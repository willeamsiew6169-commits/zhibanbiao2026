# finance_v7.py

import os
import re
import hmac
import math
import uuid
import tempfile

from copy import copy
from io import BytesIO
from decimal import Decimal
from pathlib import Path
from functools import wraps
from datetime import date, datetime, timedelta

from flask import (
    Blueprint,
    request,
    redirect,
    url_for,
    render_template_string,
    send_file,
    session,
    jsonify,
    flash,
    current_app,
    abort,
)

from psycopg2.extras import RealDictCursor

from db import (
    db_query,
    get_conn,
)

from utils import (
    normalize_member_id,
)

from finance_audit import (
    write_finance_audit,
)

from finance_common import (
    money,
    normalize_finance_date,
    get_finance_ym,
    get_month_close_record,
    is_finance_month_closed,
    get_month_lock_message,
    require_finance_month_open,
    get_current_finance_user,
)

from openpyxl import (
    Workbook,
    load_workbook,
)

from openpyxl.utils import (
    get_column_letter,
)

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

FINANCE_OPENING_DATE = date(2026, 1, 1)
# =========================================================
# Finance Web V7 Blueprint
# =========================================================

finance_v7_bp = Blueprint(
    "finance_v7",
    __name__
)



@finance_v7_bp.before_request
def require_finance_v7_login():
    """V7 只允许已通过财政负责人密码的人进入。"""
    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    if not session.get("finance_admin"):
        return redirect(
            url_for(
                "finance.finance_admin_login",
                next=request.path,
            )
        )



# =========================================================
# V7 Real-data helpers
# =========================================================

def _v7_money(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _v7_date_label(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def _v7_time_label(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    text = str(value)
    return text[11:16] if len(text) >= 16 else ""


def _v7_month_label(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m")
    return str(value)[:7]


def _v7_one(sql, params=()):
    try:
        return db_query(sql, params, fetchone=True) or {}
    except Exception as exc:
        print("Finance V7 query warning:", exc)
        return {}


def _v7_all(sql, params=()):
    try:
        return db_query(sql, params, fetchall=True) or []
    except Exception as exc:
        print("Finance V7 query warning:", exc)
        return []


def _v7_month_bounds(today=None):
    today = today or date.today()
    start = today.replace(day=1)
    if start.month == 12:
        next_start = date(start.year + 1, 1, 1)
    else:
        next_start = date(start.year, start.month + 1, 1)
    return start, next_start


def _v7_che_filter(alias="r"):
    """只按会员／收条编号判断 CHE，不用 fund_account 判断分会。"""
    return f"""(
        upper(coalesce({alias}.member_id, '')) like 'CHE-%%'
        or upper(coalesce({alias}.receipt_no, '')) like 'CHE%%'
    )"""


def _v7_stw_filter(alias="r"):
    """只按会员／收条编号判断 STW，不用 fund_account 判断分会。"""
    return f"""(
        upper(coalesce({alias}.member_id, '')) like 'STW-%%'
        or upper(coalesce({alias}.receipt_no, '')) like 'STW%%'
    )"""


def _v7_balance_summary(branch="CHE", balance_date=None):
    """Finance V7 唯一资金余额引擎。

    三个资金池严格分开：

    1. Bank（CHE 日常银行）
       = 银行期初 + CHE 银行月费 + 正式 CDM - 银行提款 ± 银行调整

    2. Cash In Hand（月费现金，等待 CDM）
       = 尚未连接到正式 CDM 明细的 CHE 现金月费
       不能用来支付任何 PV／日常支出。

    3. Petty Cash（日常备用金）
       = Petty Cash 期初 + 银行提款转入 - 现金支出 ± 现金调整

    finance_bank_deposits 是 CDM 唯一正式来源；
    finance_cash_movements.bank_in 不参与余额，避免重复计算。
    """
    branch = (branch or "CHE").strip().upper()
    balance_date = balance_date or date.today()
    opening_date = FINANCE_OPENING_DATE

    def movement_total(account_type, movement_type):
        row = _v7_one("""
            select coalesce(sum(amount), 0) as total
            from finance_cash_movements
            where upper(coalesce(branch, '')) = %s
              and account_type = %s
              and movement_type = %s
              and record_date <= %s
        """, (branch, account_type, movement_type, balance_date))
        return Decimal(str(row.get("total") or 0))

    bank_opening = movement_total("bank", "opening")
    petty_cash_opening = movement_total("cash", "opening")
    bank_cash_out = movement_total("bank", "cash_out")
    petty_cash_in = movement_total("cash", "cash_in")
    bank_adjustment = movement_total("bank", "adjustment")
    petty_cash_adjustment = movement_total("cash", "adjustment")

    # CHE 银行月费：只算月费，并且必须能由 CHE member_id／receipt_no 识别。
    bank_income_row = _v7_one(f"""
        select coalesce(sum(r.amount), 0) as total
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and coalesce(r.status, 'confirmed') <> 'cancelled'
          and r.record_date between %s and %s
          and lower(trim(coalesce(r.payment_method, ''))) in (
                '银行过账','网上转账','银行转账','duitnow',
                'online transfer','bank transfer'
          )
          and {_v7_che_filter('r')}
    """, (opening_date, balance_date))
    bank_direct_income = Decimal(str(bank_income_row.get("total") or 0))

    # CHE 现金月费总额：用于资金流统计。
    cash_income_row = _v7_one(f"""
        select coalesce(sum(r.amount), 0) as total
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and coalesce(r.status, 'confirmed') <> 'cancelled'
          and r.record_date between %s and %s
          and lower(trim(coalesce(r.payment_method, ''))) in ('现金','cash')
          and {_v7_che_filter('r')}
    """, (opening_date, balance_date))
    cash_monthly_income = Decimal(str(cash_income_row.get("total") or 0))

    # 当前 Cash In Hand：只计算尚未被任何正式 CDM 明细连接的现金月费。
    # 已经放入 finance_bank_deposit_items 的收条，代表现金已经存入银行，
    # 不再留在 Cash In Hand。这样不会用“本月现金 - 历史／本月 CDM”造成负数。
    cash_in_hand_row = _v7_one(f"""
        select coalesce(sum(r.amount), 0) as total
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and coalesce(r.status, 'confirmed') <> 'cancelled'
          and r.record_date between %s and %s
          and lower(trim(coalesce(r.payment_method, ''))) in ('现金','cash')
          and {_v7_che_filter('r')}
          and not exists (
                select 1
                from finance_bank_deposit_items i
                join finance_bank_deposits d
                  on d.id = i.deposit_id
                where i.finance_record_id = r.id
                  and upper(coalesce(d.branch, '')) = %s
          )
    """, (opening_date, balance_date, branch))
    current_cash_in_hand = Decimal(str(cash_in_hand_row.get("total") or 0))

    # 正式 CDM：Cash In Hand 减少、Bank 增加。
    cdm_row = _v7_one("""
        select coalesce(sum(amount), 0) as total
        from finance_bank_deposits
        where upper(coalesce(branch, '')) = %s
          and fund_account = '观音堂日常户口'
          and deposit_date between %s and %s
    """, (branch, opening_date, balance_date))
    cdm_total = Decimal(str(cdm_row.get("total") or 0))

    # 所有 CHE 日常现金支出只从 Petty Cash 扣除。
    expense_row = _v7_one("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'expense'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and record_date between %s and %s
          and coalesce(fund_account, '观音堂日常户口')
              in ('观音堂日常户口', 'CHE 日常户口')
    """, (opening_date, balance_date))
    petty_cash_expense = Decimal(str(expense_row.get("total") or 0))

    bank_balance = (
        bank_opening + bank_direct_income + cdm_total
        - bank_cash_out + bank_adjustment
    )
    cash_in_hand = current_cash_in_hand
    petty_cash = (
        petty_cash_opening + petty_cash_in
        - petty_cash_expense + petty_cash_adjustment
    )

    return {
        "bank_balance": float(bank_balance),
        "cash_in_hand": float(cash_in_hand),
        "petty_cash": float(petty_cash),
        "total_cash": float(cash_in_hand + petty_cash),
        "total_funds": float(bank_balance + cash_in_hand + petty_cash),
        "bank_opening": float(bank_opening),
        "bank_direct_income": float(bank_direct_income),
        "cash_income": float(cash_monthly_income),
        "cash_monthly_income": float(cash_monthly_income),
        "cdm_total": float(cdm_total),
        "bank_cash_out": float(bank_cash_out),
        "cash_opening": float(petty_cash_opening),
        "petty_cash_opening": float(petty_cash_opening),
        "petty_cash_in": float(petty_cash_in),
        "cash_expense": float(petty_cash_expense),
        "petty_cash_expense": float(petty_cash_expense),
        "bank_adjustment": float(bank_adjustment),
        "cash_adjustment": float(petty_cash_adjustment),
        "petty_cash_adjustment": float(petty_cash_adjustment),
    }


def _v7_monthly_fee_health(branch="CHE", month_count=6):
    """计算 CHE 月费是否足以维持日常支出。

    规则：
    - 月费只统计已确认、未作废的 CHE 月费。
    - 支出只统计观音堂日常户口／CHE 日常户口支出。
    - 使用最近 month_count 个月平均，避免单月波动误导。
    - 安全会员人数会根据实际月费实现率自动加入迟付／未付缓冲。
    """
    branch = (branch or "CHE").strip().upper()
    month_count = max(1, int(month_count or 6))
    monthly_fee_amount = 50.0
    today = date.today()
    current_start = today.replace(day=1)

    def add_months(value, offset):
        index = value.year * 12 + (value.month - 1) + offset
        return date(index // 12, index % 12 + 1, 1)

    periods = []
    for offset in range(-(month_count - 1), 1):
        start = add_months(current_start, offset)
        end = add_months(start, 1)
        periods.append((start, end))

    branch_filter = _v7_che_filter("r") if branch == "CHE" else _v7_stw_filter("r")
    monthly = []

    for start, end in periods:
        income_row = _v7_one(f"""
            select coalesce(sum(r.amount), 0) as total
            from finance_records r
            where r.record_type = 'income'
              and r.category = '月费'
              and coalesce(r.status, 'confirmed') <> 'cancelled'
              and r.record_date >= %s and r.record_date < %s
              and {branch_filter}
        """, (start, end))

        expense_row = _v7_one("""
            select coalesce(sum(amount), 0) as total
            from finance_records
            where record_type = 'expense'
              and coalesce(status, 'confirmed') <> 'cancelled'
              and record_date >= %s and record_date < %s
              and coalesce(fund_account, '观音堂日常户口')
                  in ('观音堂日常户口', 'CHE 日常户口')
        """, (start, end))

        income = _v7_money(income_row.get("total"))
        expense = _v7_money(expense_row.get("total"))
        monthly.append({
            "ym": start.strftime("%Y-%m"),
            "label": start.strftime("%Y年%m月"),
            "income": income,
            "expense": expense,
            "surplus": income - expense,
            "coverage": round((income / expense * 100), 1) if expense > 0 else 100.0,
        })

    avg_income = sum(row["income"] for row in monthly) / len(monthly)
    avg_expense = sum(row["expense"] for row in monthly) / len(monthly)
    avg_surplus = avg_income - avg_expense
    current = monthly[-1]

    active_row = _v7_one("""
        select count(*) as total
        from members
        where upper(coalesce(member_id, '')) like %s
          and lower(trim(coalesce(member_status, status, ''))) not in (
                '停供','停止','永久停止','往生','已往生',
                'paused','inactive','stopped','deceased'
          )
    """, (f"{branch}-%%",))
    active_members = int(active_row.get("total") or 0)

    stable_row = _v7_one("""
        select count(*) as total
        from (
            select m.member_id, max(p.end_month) as paid_until
            from members m
            left join member_payments p
              on p.member_id = m.member_id
             and coalesce(p.status, 'active') = 'active'
            where upper(coalesce(m.member_id, '')) like %s
              and lower(trim(coalesce(m.member_status, m.status, ''))) not in (
                    '停供','停止','永久停止','往生','已往生',
                    'paused','inactive','stopped','deceased'
              )
            group by m.member_id
            having max(p.end_month) >= %s
        ) x
    """, (f"{branch}-%%", current_start))
    stable_members = int(stable_row.get("total") or 0)

    late_members = max(active_members - stable_members, 0)
    theoretical_monthly = active_members * monthly_fee_amount
    realization_rate = (
        min(max(avg_income / theoretical_monthly, 0.0), 1.0)
        if theoretical_monthly > 0 else 0.0
    )

    minimum_members = math.ceil(avg_expense / monthly_fee_amount) if avg_expense > 0 else 0
    # 实现率太低时避免产生无限大目标；最低以 60% 作为可解释下限。
    safe_rate = max(realization_rate, 0.60) if active_members else 1.0
    safe_members = math.ceil(minimum_members / safe_rate) if minimum_members else 0
    need_members = max(safe_members - active_members, 0)

    current_coverage = current["coverage"]
    average_coverage = round((avg_income / avg_expense * 100), 1) if avg_expense > 0 else 100.0

    if average_coverage >= 110:
        health_level, health_icon, health_text = "healthy", "🟢", "月费平均足以维持日常支出，并有安全余量。"
    elif average_coverage >= 100:
        health_level, health_icon, health_text = "watch", "🟡", "月费目前刚好足以维持，仍需留意迟付与支出上升。"
    elif average_coverage >= 80:
        health_level, health_icon, health_text = "warning", "🟠", "月费平均不足以完全覆盖支出，建议增加稳定会员。"
    else:
        health_level, health_icon, health_text = "danger", "🔴", "月费明显不足以维持日常支出，需要优先增加稳定会员。"

    return {
        "branch": branch,
        "month_count": month_count,
        "monthly_fee_amount": monthly_fee_amount,
        "months": monthly,
        "current_month": current["ym"],
        "current_income": current["income"],
        "current_expense": current["expense"],
        "current_surplus": current["surplus"],
        "current_coverage": current_coverage,
        "avg_income": round(avg_income, 2),
        "avg_expense": round(avg_expense, 2),
        "avg_surplus": round(avg_surplus, 2),
        "average_coverage": average_coverage,
        "active_members": active_members,
        "stable_members": stable_members,
        "late_members": late_members,
        "realization_rate": round(realization_rate * 100, 1),
        "minimum_members": minimum_members,
        "safe_members": safe_members,
        "need_members": need_members,
        "health_level": health_level,
        "health_icon": health_icon,
        "health_text": health_text,
    }


def _v7_pending_summary(branch="CHE", category=None):
    clauses = ["coalesce(status, 'pending') = 'pending'"]
    params = []
    if category:
        clauses.append("category = %s")
        params.append(category)
    if branch == "CHE":
        clauses.append("(upper(coalesce(member_id,'')) like 'CHE-%%' or upper(coalesce(receipt_no,'')) like 'CHE%%')")
    elif branch == "STW":
        clauses.append("(upper(coalesce(member_id,'')) like 'STW-%%' or upper(coalesce(receipt_no,'')) like 'STW%%')")
    row = _v7_one(f"""
        select count(*) as count, coalesce(sum(amount),0) as total
        from bank_pending_records
        where {' and '.join(clauses)}
    """, tuple(params))
    return int(row.get("count") or 0), _v7_money(row.get("total"))


def _v7_che_home_data():
    today = date.today()
    start, next_start = _v7_month_bounds(today)
    balance = _v7_balance_summary("CHE", today)
    pending_count, pending_total = _v7_pending_summary("CHE")

    monthly = _v7_one(f"""
        select
            coalesce(sum(amount),0) as total,
            coalesce(sum(case when lower(coalesce(payment_method,'')) in ('现金','cash') then amount else 0 end),0) as cash_total,
            coalesce(sum(case when lower(coalesce(payment_method,'')) not in ('现金','cash') then amount else 0 end),0) as bank_total
        from finance_records r
        where r.record_type='income' and r.category='月费'
          and coalesce(r.status,'confirmed') <> 'cancelled'
          and r.record_date >= %s and r.record_date < %s
          and {_v7_che_filter('r')}
    """, (start, next_start))

    expense = _v7_one("""
        select
            coalesce(sum(amount),0) as month_total,
            coalesce(sum(case when record_date=current_date then amount else 0 end),0) as today_total,
            count(*) filter (where coalesce(payment_voucher_no,'')='') as missing_pv
        from finance_records
        where record_type='expense'
          and coalesce(status,'confirmed') <> 'cancelled'
          and record_date >= %s and record_date < %s
    """, (start, next_start))

    late = _v7_one("""
        select count(*) as count
        from (
            select m.member_id, max(p.end_month) as paid_until
            from members m
            left join member_payments p on p.member_id=m.member_id
             and coalesce(p.status,'active')='active'
            where upper(coalesce(m.member_id,'')) like 'CHE-%%'
              and coalesce(m.member_status,m.status,'') not in ('停供','停止','永久停止','往生','已往生')
            group by m.member_id
            having max(p.end_month) is not null
               and max(p.end_month) < date_trunc('month',current_date)
        ) x
    """)

    return {
        "bank_balance": balance["bank_balance"],
        "bank_pending_count": pending_count,
        "bank_pending_total": pending_total,
        "cash_balance": balance["cash_in_hand"],
        "cash_in_hand": balance["cash_in_hand"],
        "petty_cash": balance["petty_cash"],
        "total_cash": balance["total_cash"],
        "cash_month_in": _v7_money(_v7_one(f"""
            select coalesce(sum(r.amount),0) total
            from finance_records r
            where r.record_type='income' and r.category='月费'
              and coalesce(r.status,'confirmed')<>'cancelled'
              and lower(trim(coalesce(r.payment_method,''))) in ('现金','cash')
              and r.record_date >= %s and r.record_date < %s
              and {_v7_che_filter('r')}
        """, (start,next_start)).get('total')),
        "cash_month_out": _v7_money(expense.get("month_total")),
        "monthly_fee_month_total": _v7_money(monthly.get("total")),
        "monthly_fee_cash_total": _v7_money(monthly.get("cash_total")),
        "monthly_fee_bank_total": _v7_money(monthly.get("bank_total")),
        "expense_month_total": _v7_money(expense.get("month_total")),
        "expense_pending_pv_count": int(expense.get("missing_pv") or 0),
        "expense_today_total": _v7_money(expense.get("today_total")),
        "late_member_count": int(late.get("count") or 0),
        "new_late_member_count": 0,
        "care_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "month_close_period": today.strftime("%Y-%m"),
        "month_close_status": "待检查",
        "month_close_progress": 0,
    }


def _v7_monthly_data(branch="CHE"):
    today = date.today(); start, next_start = _v7_month_bounds(today)
    branch_filter = _v7_che_filter('r') if branch == 'CHE' else _v7_stw_filter('r')
    summary = _v7_one(f"""
        select
          count(distinct coalesce(member_id,name)) filter (where record_date=current_date) as today_people,
          coalesce(sum(amount) filter (where record_date=current_date),0) as today_total,
          count(distinct coalesce(member_id,name)) as month_people,
          coalesce(sum(amount),0) as month_total
        from finance_records r
        where r.record_type='income' and r.category='月费'
          and coalesce(r.status,'confirmed') <> 'cancelled'
          and r.record_date >= %s and r.record_date < %s
          and {branch_filter}
    """, (start,next_start))
    pending_count, pending_total = _v7_pending_summary(branch, '月费')
    recent = _v7_all(f"""
        select receipt_no, member_id, name, month_from, month_to, amount, payment_method, record_date, created_at
        from finance_records r
        where r.record_type='income' and r.category='月费'
          and coalesce(r.status,'confirmed') <> 'cancelled' and {branch_filter}
        order by coalesce(r.created_at, r.record_date::timestamp) desc, r.id desc
        limit 5
    """)
    rows=[]
    for r in recent:
        mf_label = _v7_month_label(r.get('month_from'))
        mt_label = _v7_month_label(r.get('month_to'))
        if mf_label and mt_label and mf_label != mt_label:
            months = f"{mf_label} ～ {mt_label}"
        else:
            months = mf_label or mt_label or '-'
        rows.append({
          'receipt_no':r.get('receipt_no') or '-', 'member_id':r.get('member_id') or '-', 'name':r.get('name') or '-',
          'months':months, 'month_label':months, 'amount':_v7_money(r.get('amount')), 'method':r.get('payment_method') or '-',
          'record_date':_v7_date_label(r.get('record_date')), 'time':_v7_time_label(r.get('created_at'))
        })
    return {
      'today':today.strftime('%Y-%m-%d'),'month':today.strftime('%Y-%m'),
      'today_people':int(summary.get('today_people') or 0),'today_total':_v7_money(summary.get('today_total')),
      'month_people':int(summary.get('month_people') or 0),'month_total':_v7_money(summary.get('month_total')),
      'bank_pending_count':pending_count,'bank_pending_total':pending_total,'pending_count':pending_count,'recent_records':rows,'recent':rows
    }


def _v7_expense_data():
    today=date.today(); start,next_start=_v7_month_bounds(today)
    summary=_v7_one("""
      select coalesce(sum(amount) filter(where record_date=current_date),0) today_total,
             count(*) filter(where record_date=current_date) today_pv_count,
             coalesce(sum(amount),0) month_total
      from finance_records
      where record_type='expense' and coalesce(status,'confirmed')<>'cancelled'
        and record_date >= %s and record_date < %s
    """,(start,next_start))
    recent=_v7_all("""
      select payment_voucher_no, record_date, category, coalesce(vendor_name,name) as vendor_name, name, amount
      from finance_records
      where record_type='expense' and coalesce(status,'confirmed')<>'cancelled'
      order by record_date desc,id desc limit 5
    """)
    return {'today_total':_v7_money(summary.get('today_total')),'today_pv_count':int(summary.get('today_pv_count') or 0),
      'month_total':_v7_money(summary.get('month_total')),'month_label':today.strftime('%Y-%m'),
      'recent_expenses':[{'pv_no':r.get('payment_voucher_no') or '未填 PV','expense_date':str(r.get('record_date') or ''),
      'category':r.get('category') or '其它支出','vendor_name':r.get('vendor_name') or r.get('name') or '未填写','amount':_v7_money(r.get('amount'))} for r in recent]}


def _v7_hq_data():
    today=date.today(); start,next_start=_v7_month_bounds(today)
    cats=['财布施','观音村','膳食结缘','观音堂纯檀香布施','813交流会财布施','临时特别布施']
    summary=_v7_one("""select coalesce(sum(amount),0) month_total, count(*) filter(where record_date=current_date) today_count, coalesce(sum(amount) filter(where record_date=current_date),0) today_total from finance_records where record_type='income' and category=any(%s) and coalesce(status,'confirmed')<>'cancelled' and record_date >= %s and record_date < %s""",(cats,start,next_start))
    pending_count,_=_v7_pending_summary('CHE')
    category_rows=_v7_all("""select category,coalesce(sum(amount),0) amount from finance_records where record_type='income' and category=any(%s) and coalesce(status,'confirmed')<>'cancelled' and record_date >= %s and record_date < %s group by category order by category""",(cats,start,next_start))
    icon_map={'财布施':'❤️','观音村':'🏠','膳食结缘':'🍚','观音堂纯檀香布施':'🪔','813交流会财布施':'🌟','临时特别布施':'🌟'}
    recent=_v7_all("""select receipt_no,name,category,amount,payment_method,created_at from finance_records where record_type='income' and category=any(%s) and coalesce(status,'confirmed')<>'cancelled' order by coalesce(created_at,record_date::timestamp) desc,id desc limit 5""",(cats,))
    return {'month_total':_v7_money(summary.get('month_total')),'today_count':int(summary.get('today_count') or 0),'today_total':_v7_money(summary.get('today_total')),'pending_count':pending_count,
      'categories':[{'name':r.get('category'),'icon':icon_map.get(r.get('category'),'🙏'),'amount':_v7_money(r.get('amount'))} for r in category_rows],
      'recent':[{'receipt_no':r.get('receipt_no') or '-','name':r.get('name') or '无名氏','category':r.get('category') or '财布施','icon':icon_map.get(r.get('category'),'🙏'),'amount':_v7_money(r.get('amount')),'method':r.get('payment_method') or '-','time':r.get('created_at').strftime('%H:%M') if r.get('created_at') else ''} for r in recent]}



def _render_v7_module_records(module, search_mode=False):
    """V7 独立记录页：记录与查找合并，并支持指定年份／月份。"""
    today = date.today()
    keyword = request.args.get("q", "").strip()
    year_text = request.args.get("year", str(today.year)).strip()
    month_text = request.args.get("month", str(today.month)).strip()
    category = request.args.get("category", "").strip()

    try:
        selected_year = int(year_text) if year_text else None
    except ValueError:
        selected_year = today.year
        year_text = str(today.year)

    try:
        selected_month = int(month_text) if month_text else None
        if selected_month is not None and not 1 <= selected_month <= 12:
            raise ValueError
    except ValueError:
        selected_month = today.month
        month_text = str(today.month)

    clauses = ["coalesce(r.status,'confirmed')<>'cancelled'"]
    params = []
    title = "记录"
    subtitle = ""
    back_endpoint = "finance_v7.finance_v7_home"

    if module == "stw":
        title = "STW 月费记录"
        clauses += ["r.record_type='income'", "r.category='月费'", _v7_stw_filter("r")]
        subtitle = "记录与查找已合并；这里只显示 STW 月费。"
        back_endpoint = "finance_v7.finance_v7_stw_home"
    elif module == "hq":
        title = "总会布施记录"
        clauses += ["r.record_type='income'", "r.category=any(%s)"]
        params.append([
            "财布施", "观音村", "膳食结缘", "观音堂纯檀香布施",
            "813交流会财布施", "临时特别布施",
        ])
        subtitle = "记录与查找已合并；这里只显示总会布施，并可选择布施类别。"
        back_endpoint = "finance_v7.finance_v7_hq_home"
        if category:
            clauses.append("r.category=%s")
            params.append(category)
    elif module == "cash":
        title = "CHE 现金活动"
        clauses += ["((r.record_type='expense') or (r.record_type='income' and lower(coalesce(r.payment_method,'')) in ('现金','cash')))"]
        subtitle = "这里只显示现金增加与现金支出。"
        back_endpoint = "finance_v7.finance_v7_cash"
    elif module == "expense":
        title = "CHE 支出记录"
        clauses += ["r.record_type='expense'"]
        subtitle = "记录与查找已合并；这里只显示 CHE 支出、PV 与供应商。"
        back_endpoint = "finance_v7.finance_v7_expense"

    if selected_year is not None:
        clauses.append("extract(year from r.record_date)=%s")
        params.append(selected_year)
    if selected_month is not None:
        clauses.append("extract(month from r.record_date)=%s")
        params.append(selected_month)

    if keyword:
        like = f"%{keyword}%"
        clauses.append("(coalesce(r.member_id,'') ilike %s or coalesce(r.name,'') ilike %s or coalesce(r.receipt_no,'') ilike %s or coalesce(r.payment_voucher_no,'') ilike %s or coalesce(r.category,'') ilike %s or coalesce(r.remarks,'') ilike %s or cast(r.amount as text) ilike %s)")
        params.extend([like] * 7)

    rows = _v7_all(f"""
        select r.id,r.record_date,r.record_type,r.category,r.receipt_no,
               r.payment_voucher_no,r.member_id,r.name,r.amount,
               r.payment_method,r.remarks
        from finance_records r
        where {' and '.join(clauses)}
        order by r.record_date desc,r.id desc
        limit 500
    """, tuple(params))

    items = []
    for row in rows:
        number = row.get("payment_voucher_no") if row.get("record_type") == "expense" else row.get("receipt_no")
        items.append({
            **row,
            "number": number or "-",
            "record_date": _v7_date_label(row.get("record_date")),
            "amount": _v7_money(row.get("amount")),
        })

    total = sum(item["amount"] for item in items)
    current_year = today.year
    years = list(range(current_year + 1, current_year - 8, -1))
    months = [(m, f"{m}月") for m in range(1, 13)]
    hq_categories = [
        "财布施", "观音村", "膳食结缘", "观音堂纯檀香布施",
        "813交流会财布施", "临时特别布施",
    ] if module == "hq" else []

    return render_template_string(r"""<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{{ title }}</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>
body{margin:0;background:#f5f7fb;color:#172033}.page{width:min(1000px,calc(100% - 24px));margin:auto;padding:22px 0 48px}.top{display:flex;justify-content:space-between;gap:14px;align-items:flex-start}.top h1{margin:0;font-size:28px}.sub{margin-top:6px;color:#748096}.back{background:#fff;border:1px solid #dce3ed;border-radius:13px;padding:10px 14px;text-decoration:none;color:#263149;font-weight:900}.search{margin-top:17px;background:#fff;border:1px solid #e0e6ef;border-radius:20px;padding:15px}.search-grid{display:grid;grid-template-columns:minmax(220px,1fr) 140px 140px {% if hq_categories %}190px {% endif %}auto auto;gap:10px}.search input,.search select{min-height:48px;border:1px solid #d4dce7;border-radius:13px;padding:0 13px;font-size:15px;background:#fff}.search button,.clear{min-height:48px;border:0;border-radius:13px;background:#3978db;color:#fff;font-weight:900;padding:0 20px;display:flex;align-items:center;justify-content:center;text-decoration:none}.clear{background:#eef2f7;color:#4d596d}.summary{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin:14px 0}.sum{background:#fff;border:1px solid #e0e6ef;border-radius:17px;padding:16px}.sum-label{color:#778194;font-size:13px}.sum-value{font-size:24px;font-weight:950;margin-top:5px}.list{display:grid;gap:10px}.row{display:grid;grid-template-columns:125px minmax(0,1fr) auto;gap:14px;background:#fff;border:1px solid #e2e7ef;border-radius:17px;padding:15px}.num{font-weight:950}.date{color:#7c8698;font-size:13px;margin-top:5px}.name{font-size:17px;font-weight:950}.meta{color:#7b8495;font-size:13px;line-height:1.55;margin-top:6px}.amount{text-align:right;font-size:19px;font-weight:950;white-space:nowrap}.method{text-align:right;color:#7b8495;font-size:12px;margin-top:5px}.empty{text-align:center;background:#fff;border:1px dashed #ccd5e2;border-radius:17px;padding:36px;color:#788397}@media(max-width:760px){.search-grid{grid-template-columns:1fr 1fr}.search-grid input{grid-column:1/-1}.search button,.clear{min-height:48px}.row{grid-template-columns:1fr auto}.row>div:nth-child(2){grid-column:1/-1;grid-row:2}.top h1{font-size:23px}}</style></head><body><main class="page"><div class="top"><div><h1>📖 {{ title }}</h1><div class="sub">{{ subtitle }}</div></div><a class="back" href="{{ url_for(back_endpoint) }}">← 返回</a></div><form class="search" method="get"><div class="search-grid"><input name="q" value="{{ keyword }}" placeholder="输入编号、姓名、类别、金额或备注"><select name="year"><option value="">全部年份</option>{% for y in years %}<option value="{{ y }}" {% if selected_year==y %}selected{% endif %}>{{ y }}年</option>{% endfor %}</select><select name="month"><option value="">全年</option>{% for value,label in months %}<option value="{{ value }}" {% if selected_month==value %}selected{% endif %}>{{ label }}</option>{% endfor %}</select>{% if hq_categories %}<select name="category"><option value="">全部布施类别</option>{% for c in hq_categories %}<option value="{{ c }}" {% if category==c %}selected{% endif %}>{{ c }}</option>{% endfor %}</select>{% endif %}<button type="submit">查询</button><a class="clear" href="?year={{ current_year }}&month={{ current_month }}">本月</a></div></form><div class="summary"><div class="sum"><div class="sum-label">记录</div><div class="sum-value">{{ items|length }} 笔</div></div><div class="sum"><div class="sum-label">合计</div><div class="sum-value">RM {{ '%.2f'|format(total) }}</div></div></div>{% if items %}<section class="list">{% for r in items %}<article class="row"><div><div class="num">{{ r.number }}</div><div class="date">{{ r.record_date }}</div></div><div><div class="name">{{ r.name or '-' }}</div><div class="meta">{{ r.category or '-' }}{% if r.member_id %} · {{ r.member_id }}{% endif %}<br>{{ r.remarks or '无备注' }}</div></div><div><div class="amount">RM {{ '%.2f'|format(r.amount) }}</div><div class="method">{{ r.payment_method or '-' }}</div></div></article>{% endfor %}</section>{% else %}<div class="empty">这个月份没有找到相关记录</div>{% endif %}</main></body></html>""",
        title=title, subtitle=subtitle, back_endpoint=back_endpoint,
        keyword=keyword, selected_year=selected_year,
        selected_month=selected_month, years=years, months=months,
        current_year=today.year, current_month=today.month,
        hq_categories=hq_categories, category=category,
        items=items, total=total,
    )

# ========================================================
# V7 单位选择首页
# =========================================================

@finance_v7_bp.route("/finance/v7")
def finance_v7_home():

    units = [
        {
            "code": "CHE",
            "name": "CHE 观音堂财政",
            "icon": "🏛️",
            "type_label": "完整财政管理",
            "description": "银行、Cash、月费、支出、月费关怀与月结。",
            "url": url_for("finance_v7.finance_v7_che_home"),
            "class_name": "unit-che",
            "features": [
                "银行与 Bank In",
                "Cash In Hand",
                "月费",
                "支出",
                "月费关怀",
                "财政月结",
            ],
        },
        {
            "code": "STW",
            "name": "STW 收款记录",
            "icon": "🧾",
            "type_label": "简单收款管理",
            "description": "记录月费、完成 Bank In，并保留清楚的收款记录。",
            "url": url_for("finance_v7.finance_v7_stw_home"),
            "class_name": "unit-stw",
            "features": [
                "记录月费",
                "Bank In",
                "收款记录查询",
            ],
        },
        {
            "code": "HQ",
            "name": "总会财布施",
            "icon": "🙏",
            "type_label": "分类财布施管理",
            "description": "按不同财布施类别分别记录、统计和 Bank In。",
            "url": url_for("finance_v7.finance_v7_hq_home"),
            "class_name": "unit-hq",
            "features": [ 
                "财布施",
                "观音村",
                "膳食结缘",
                "纯檀香布施",
                "特别布施",
            ],
        },
    ]

    return render_template_string(
        """
        <!doctype html>
        <html lang="zh">
        <head>

            <meta charset="utf-8">

            <meta
                name="viewport"
                content="width=device-width, initial-scale=1"
            >

            <title>观音堂财政 V7</title>

            <link
                rel="stylesheet"
                href="{{ url_for(
                    'static',
                    filename='css/toolbox.css'
                ) }}"
            >

            <style>

                body {
                    margin: 0;
                    background:
                        linear-gradient(
                            180deg,
                            #eef4ff 0,
                            #f6f8fc 320px,
                            #f6f8fc 100%
                        );
                    color: #172033;
                }

                .v7-unit-page {
                    width: min(1120px, calc(100% - 32px));
                    margin: 0 auto;
                    padding: 30px 0 50px;
                }

                /* =====================================
                   页面标题
                ====================================== */

                .v7-unit-header {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 20px;
                    margin-bottom: 26px;
                }

                .v7-heading-wrap {
                    display: flex;
                    align-items: center;
                    gap: 16px;
                }

                .v7-heading-icon {
                    width: 64px;
                    height: 64px;
                    display: grid;
                    place-items: center;
                    flex: 0 0 auto;
                    border-radius: 20px;
                    background: #dfeaff;
                    font-size: 33px;
                    box-shadow:
                        0 8px 22px rgba(41, 75, 140, 0.10);
                }

                .v7-heading-wrap h1 {
                    margin: 0;
                    font-size: 32px;
                    line-height: 1.2;
                }

                .v7-heading-wrap p {
                    margin: 7px 0 0;
                    color: #687286;
                    font-size: 16px;
                    line-height: 1.5;
                }

                .v7-return-button {
                    text-decoration: none;
                    color: #26324a;
                    background: white;
                    border: 1px solid #d8e0ed;
                    border-radius: 14px;
                    padding: 11px 17px;
                    font-weight: 800;
                    white-space: nowrap;
                    box-shadow:
                        0 6px 18px rgba(33, 50, 83, 0.05);
                }

                /* =====================================
                   提示区
                ====================================== */

                .v7-unit-notice {
                    display: flex;
                    align-items: flex-start;
                    gap: 12px;
                    background: rgba(255, 255, 255, 0.86);
                    border: 1px solid #dfe6f1;
                    border-radius: 17px;
                    padding: 16px 18px;
                    margin-bottom: 20px;
                    color: #556077;
                    line-height: 1.6;
                    box-shadow:
                        0 8px 24px rgba(33, 50, 83, 0.04);
                }

                .v7-unit-notice strong {
                    color: #26324a;
                }

                /* =====================================
                   单位卡片
                ====================================== */

                .v7-unit-grid {
                    display: grid;
                    grid-template-columns:
                        repeat(3, minmax(0, 1fr));
                    gap: 18px;
                }

                .v7-unit-card {
                    position: relative;
                    display: flex;
                    flex-direction: column;
                    min-height: 370px;
                    overflow: hidden;
                    text-decoration: none;
                    color: inherit;
                    background: white;
                    border: 1px solid #dee5ef;
                    border-radius: 24px;
                    padding: 23px;
                    box-shadow:
                        0 12px 32px rgba(33, 50, 83, 0.07);
                    transition:
                        transform 0.16s ease,
                        box-shadow 0.16s ease,
                        border-color 0.16s ease;
                }

                .v7-unit-card:hover {
                    transform: translateY(-3px);
                    border-color: #c5d2e7;
                    box-shadow:
                        0 18px 40px rgba(33, 50, 83, 0.12);
                }

                .v7-unit-card::before {
                    content: "";
                    position: absolute;
                    top: 0;
                    left: 0;
                    right: 0;
                    height: 7px;
                }

                .unit-che::before {
                    background: #3978db;
                }

                .unit-stw::before {
                    background: #7857c8;
                }

                .unit-hq::before {
                    background: #d59131;
                }

                .v7-unit-card-head {
                    display: flex;
                    align-items: flex-start;
                    justify-content: space-between;
                    gap: 14px;
                    margin-top: 4px;
                }

                .v7-unit-card-icon {
                    width: 58px;
                    height: 58px;
                    display: grid;
                    place-items: center;
                    border-radius: 18px;
                    font-size: 29px;
                }

                .unit-che .v7-unit-card-icon {
                    background: #e8f0ff;
                }

                .unit-stw .v7-unit-card-icon {
                    background: #f0ebff;
                }

                .unit-hq .v7-unit-card-icon {
                    background: #fff2df;
                }

                .v7-unit-code {
                    padding: 6px 10px;
                    border-radius: 99px;
                    background: #f0f3f8;
                    color: #5f697d;
                    font-size: 12px;
                    font-weight: 900;
                    letter-spacing: 0.6px;
                }

                .v7-unit-card h2 {
                    margin: 18px 0 0;
                    font-size: 23px;
                    line-height: 1.3;
                }

                .v7-unit-type {
                    margin-top: 7px;
                    color: #43506a;
                    font-size: 15px;
                    font-weight: 900;
                }

                .v7-unit-description {
                    margin-top: 10px;
                    color: #747e90;
                    font-size: 15px;
                    line-height: 1.65;
                }

                .v7-unit-items {
                    display: grid;
                    gap: 8px;
                    margin-top: 18px;
                }

                .v7-unit-item {
                    display: flex;
                    align-items: center;
                    gap: 8px;
                    color: #4d586e;
                    font-size: 14px;
                }

                .v7-unit-item::before {
                    content: "✓";
                    width: 20px;
                    height: 20px;
                    display: grid;
                    place-items: center;
                    flex: 0 0 auto;
                    border-radius: 50%;
                    background: #eef2f7;
                    font-size: 12px;
                    font-weight: 900;
                }

                .v7-unit-enter {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 12px;
                    margin-top: auto;
                    padding-top: 22px;
                    font-weight: 900;
                }

                .unit-che .v7-unit-enter {
                    color: #2f68bd;
                }

                .unit-stw .v7-unit-enter {
                    color: #6845b3;
                }

                .unit-hq .v7-unit-enter {
                    color: #a86d1f;
                }

                .v7-enter-arrow {
                    font-size: 25px;
                }

                /* =====================================
                   负责人入口
                ====================================== */

                .v7-utility-section {
                    margin-top: 26px;
                }

                .v7-section-title {
                    margin: 0 0 13px;
                    font-size: 20px;
                    font-weight: 950;
                }

                .v7-utility-grid {
                    display: grid;
                    grid-template-columns: repeat(2, minmax(0, 1fr));
                    gap: 16px;
                }

                .v7-utility-card {
                    display: flex;
                    align-items: center;
                    gap: 15px;
                    min-height: 92px;
                    padding: 18px 20px;
                    text-decoration: none;
                    color: inherit;
                    background: white;
                    border: 1px solid #dee5ef;
                    border-radius: 20px;
                    box-shadow: 0 10px 28px rgba(33, 50, 83, 0.06);
                }

                .v7-utility-card:hover {
                    border-color: #c5d2e7;
                    transform: translateY(-2px);
                }

                .v7-utility-icon {
                    width: 52px;
                    height: 52px;
                    display: grid;
                    place-items: center;
                    flex: 0 0 auto;
                    border-radius: 16px;
                    background: #edf3ff;
                    font-size: 27px;
                }

                .v7-utility-card:last-child .v7-utility-icon {
                    background: #f2edff;
                }

                .v7-utility-title {
                    font-size: 18px;
                    font-weight: 950;
                }

                .v7-utility-desc {
                    margin-top: 5px;
                    color: #747e90;
                    font-size: 14px;
                    line-height: 1.5;
                }

                /* =====================================
                   手机版
                ====================================== */

                @media (max-width: 900px) {

                    .v7-unit-grid {
                        grid-template-columns: 1fr;
                    }

                    .v7-unit-card {
                        min-height: auto;
                    }
                }

                @media (max-width: 560px) {

                    .v7-unit-page {
                        width: min(100% - 20px, 1120px);
                        padding-top: 16px;
                    }

                    .v7-unit-header {
                        align-items: flex-start;
                    }

                    .v7-heading-icon {
                        width: 50px;
                        height: 50px;
                        border-radius: 16px;
                        font-size: 26px;
                    }

                    .v7-heading-wrap h1 {
                        font-size: 25px;
                    }

                    .v7-heading-wrap p {
                        font-size: 14px;
                    }

                    .v7-return-button {
                        padding: 9px 11px;
                        font-size: 13px;
                    }

                    .v7-unit-card {
                        border-radius: 20px;
                        padding: 19px;
                    }

                    .v7-utility-grid {
                        grid-template-columns: 1fr;
                    }

                    .v7-unit-card h2 {
                        font-size: 21px;
                    }
                }

            </style>

        </head>

        <body>

        <main class="v7-unit-page">

            <header class="v7-unit-header">

                <div class="v7-heading-wrap">

                    <div class="v7-heading-icon">
                        💼
                    </div>

                    <div>
                        <h1>观音堂财政 V7</h1>

                        <p>
                            请选择今天要管理的单位
                        </p>
                    </div>

                </div>

                <a
                    class="v7-return-button"
                    href="/finance"
                >
                    ← 返回财政系统
                </a>

            </header>

            <div class="v7-unit-notice">
                <span>ℹ️</span>

                <div>
                    <strong>三个单位采用不同工作流程。</strong>
                    CHE 是完整财政管理；STW 只负责简单收款；
                    总会按不同财布施类别分别记录及 Bank In。
                </div>
            </div>

            <section class="v7-unit-grid">

                {% for unit in units %}

                    <a
                        class="v7-unit-card {{ unit.class_name }}"
                        href="{{ unit.url }}"
                    >

                        <div class="v7-unit-card-head">

                            <div class="v7-unit-card-icon">
                                {{ unit.icon }}
                            </div>

                            <div class="v7-unit-code">
                                {{ unit.code }}
                            </div>

                        </div>

                        <h2>{{ unit.name }}</h2>

                        <div class="v7-unit-type">
                            {{ unit.type_label }}
                        </div>

                        <div class="v7-unit-description">
                            {{ unit.description }}
                        </div>

                        <div class="v7-unit-items">

                            {% for feature in unit.features %}

                                <div class="v7-unit-item">
                                    {{ feature }}
                                </div>

                            {% endfor %}

                        </div>

                        <div class="v7-unit-enter">
                            <span>进入工作台</span>
                            <span class="v7-enter-arrow">›</span>
                        </div>

                    </a>

                {% endfor %}

            </section>

            <section class="v7-utility-section">
                <h2 class="v7-section-title">负责人功能</h2>

                <div class="v7-utility-grid">
                    <a class="v7-utility-card" href="{{ url_for('finance_v7.finance_v7_reports_center') }}">
                        <div class="v7-utility-icon">📊</div>
                        <div>
                            <div class="v7-utility-title">财政报表</div>
                            <div class="v7-utility-desc">Dashboard、Excel 下载及各单位报表，只保留这一个正式入口。</div>
                        </div>
                    </a>

                    <a class="v7-utility-card" href="{{ url_for('finance_v7.finance_v7_system_management') }}">
                        <div class="v7-utility-icon">⚙️</div>
                        <div>
                            <div class="v7-utility-title">系统管理</div>
                            <div class="v7-utility-desc">历史 Excel 导入、导入批次、收条簿、布施人、会员及付款对象。</div>
                        </div>
                    </a>
                </div>
            </section>

        </main>

        </body>
        </html>
        """,
        units=units,
    )



# =========================================================
# V7 Reports / System Management
# =========================================================

@finance_v7_bp.route("/finance/v7/reports/monthly-fee-health")
def finance_v7_monthly_fee_health():
    data = _v7_monthly_fee_health("CHE", 6)
    return render_template_string(r"""
<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>月费健康分析</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>
body{margin:0;background:#f5f7fb;color:#172033}.page{width:min(1100px,calc(100% - 24px));margin:auto;padding:22px 0 50px}.top{display:flex;justify-content:space-between;gap:14px;align-items:flex-start}.top h1{margin:0;font-size:30px}.sub{margin-top:6px;color:#707b8d}.back{background:#fff;border:1px solid #dce3ed;border-radius:13px;padding:10px 14px;text-decoration:none;color:#263149;font-weight:900;white-space:nowrap}.hero{margin-top:17px;border-radius:24px;padding:24px;background:#fff;border:1px solid #e0e6ef;box-shadow:0 10px 28px rgba(32,49,80,.06)}.hero.healthy{border-top:7px solid #2f9b68}.hero.watch{border-top:7px solid #d3a326}.hero.warning{border-top:7px solid #e17f2d}.hero.danger{border-top:7px solid #d94f4f}.hero-head{display:flex;justify-content:space-between;gap:18px;align-items:center}.hero-title{font-size:21px;font-weight:950}.hero-text{margin-top:7px;color:#667286;line-height:1.6}.coverage{text-align:right}.coverage strong{display:block;font-size:42px}.coverage span{color:#747f91;font-size:13px}.grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:13px;margin-top:15px}.card{background:#fff;border:1px solid #e0e6ef;border-radius:19px;padding:18px;box-shadow:0 7px 22px rgba(32,49,80,.05)}.label{color:#758093;font-size:13px}.value{font-size:25px;font-weight:950;margin-top:7px}.note{color:#8a93a3;font-size:12px;margin-top:5px;line-height:1.45}.section{margin-top:16px;background:#fff;border:1px solid #e0e6ef;border-radius:22px;padding:20px}.section h2{margin:0 0 5px;font-size:21px}.section-help{color:#778194;margin-bottom:15px}.member-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:11px}.member{background:#f6f8fb;border-radius:15px;padding:15px}.member strong{display:block;font-size:27px;margin-top:7px}.need{background:#fff4ef}.table{display:grid;gap:8px}.row{display:grid;grid-template-columns:120px repeat(4,1fr);gap:10px;align-items:center;padding:12px 13px;border-radius:14px;background:#f7f9fc}.head{font-weight:900;background:#edf2f8}.positive{color:#25734e}.negative{color:#b54343}.advice{margin-top:15px;padding:16px;border-radius:16px;background:#eef4ff;color:#34547d;line-height:1.7}.advice strong{color:#1f3f6b}@media(max-width:850px){.grid,.member-grid{grid-template-columns:1fr 1fr}.row{grid-template-columns:90px repeat(4,minmax(100px,1fr));min-width:650px}.table-wrap{overflow-x:auto}}@media(max-width:560px){.top h1{font-size:24px}.hero-head{align-items:flex-start;flex-direction:column}.coverage{text-align:left}.grid,.member-grid{grid-template-columns:1fr 1fr}.value{font-size:21px}}
</style></head><body><main class="page"><header class="top"><div><h1>❤️ 月费健康分析</h1><div class="sub">用最近 {{ data.month_count }} 个月的实际月费与支出，判断观音堂是否需要增加会员。</div></div><a class="back" href="{{ url_for('finance_v7.finance_v7_reports_center') }}">← 财政报表</a></header>
<section class="hero {{ data.health_level }}"><div class="hero-head"><div><div class="hero-title">{{ data.health_icon }} 财政判断</div><div class="hero-text">{{ data.health_text }}</div></div><div class="coverage"><strong>{{ '%.1f'|format(data.average_coverage) }}%</strong><span>最近 {{ data.month_count }} 个月平均覆盖率</span></div></div></section>
<section class="grid"><div class="card"><div class="label">本月月费</div><div class="value">RM {{ '%.2f'|format(data.current_income) }}</div><div class="note">{{ data.current_month }}</div></div><div class="card"><div class="label">本月支出</div><div class="value">RM {{ '%.2f'|format(data.current_expense) }}</div><div class="note">观音堂日常支出</div></div><div class="card"><div class="label">本月结余／不足</div><div class="value {% if data.current_surplus < 0 %}negative{% else %}positive{% endif %}">{% if data.current_surplus >= 0 %}+{% endif %}RM {{ '%.2f'|format(data.current_surplus) }}</div><div class="note">本月覆盖率 {{ '%.1f'|format(data.current_coverage) }}%</div></div><div class="card"><div class="label">6个月平均结余</div><div class="value {% if data.avg_surplus < 0 %}negative{% else %}positive{% endif %}">{% if data.avg_surplus >= 0 %}+{% endif %}RM {{ '%.2f'|format(data.avg_surplus) }}</div><div class="note">平均月费 RM {{ '%.2f'|format(data.avg_income) }} · 平均支出 RM {{ '%.2f'|format(data.avg_expense) }}</div></div></section>
<section class="section"><h2>👥 会员安全目标</h2><div class="section-help">安全人数已根据实际月费实现率自动加入迟付／未付缓冲。</div><div class="member-grid"><div class="member"><span>有效会员</span><strong>{{ data.active_members }} 位</strong></div><div class="member"><span>稳定供养</span><strong>{{ data.stable_members }} 位</strong></div><div class="member"><span>最低维持人数</span><strong>{{ data.minimum_members }} 位</strong></div><div class="member need"><span>建议安全人数</span><strong>{{ data.safe_members }} 位</strong></div></div><div class="advice">目前实际月费实现率约 <strong>{{ '%.1f'|format(data.realization_rate) }}%</strong>，有 <strong>{{ data.late_members }} 位</strong>尚未供到本月。{% if data.need_members > 0 %}按照最近支出与实际缴付情况，建议再增加约 <strong>{{ data.need_members }} 位稳定会员</strong>。{% else %}目前有效会员人数已达到安全目标，但仍需继续跟进迟付会员。{% endif %}</div></section>
<section class="section"><h2>📈 最近 {{ data.month_count }} 个月</h2><div class="section-help">逐月比较月费、支出和覆盖率。</div><div class="table-wrap"><div class="table"><div class="row head"><div>月份</div><div>月费</div><div>支出</div><div>结余</div><div>覆盖率</div></div>{% for row in data.months %}<div class="row"><div><strong>{{ row.ym }}</strong></div><div>RM {{ '%.2f'|format(row.income) }}</div><div>RM {{ '%.2f'|format(row.expense) }}</div><div class="{% if row.surplus < 0 %}negative{% else %}positive{% endif %}">{% if row.surplus >= 0 %}+{% endif %}RM {{ '%.2f'|format(row.surplus) }}</div><div>{{ '%.1f'|format(row.coverage) }}%</div></div>{% endfor %}</div></div></section>
</main></body></html>""", data=data)


@finance_v7_bp.route("/finance/v7/reports")
def finance_v7_reports_center():
    return render_template_string(r"""
    <!doctype html><html lang="zh"><head>
    <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
    <title>财政报表</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}">
    <style>
    body{margin:0;background:#f5f7fb;color:#172033}.page{max-width:920px;margin:auto;padding:24px 14px}
    .top{display:flex;justify-content:space-between;align-items:center;gap:16px}.back{text-decoration:none;font-weight:900;color:#285ea8}
    .grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px;margin-top:20px}
    .card{display:flex;gap:14px;align-items:center;text-decoration:none;color:inherit;background:#fff;border:1px solid #dfe5ee;border-radius:18px;padding:19px;min-height:82px}
    .icon{font-size:29px}.title{font-size:18px;font-weight:950}.desc{color:#70798a;margin-top:5px;line-height:1.5}
    @media(max-width:650px){.grid{grid-template-columns:1fr}.top{align-items:flex-start}}
    </style></head><body><main class="page">
    <div class="top">
        <div>
            <h1>📊 财政报表</h1>
            <p>所有报表与 Excel 下载集中在这里，月结不再重复放下载按钮。</p>
        </div>

        <a class="back" href="/finance/v7">
            ← 返回
        </a>
    </div>
    <section class="grid">
      <a class="card" href="{{ url_for('finance_v7.finance_v7_monthly_fee_health') }}"><div class="icon">❤️</div><div><div class="title">月费健康分析</div><div class="desc">判断月费能否维持支出、会员安全目标及还需增加多少会员。</div></div></a>
      <a class="card" href="/finance/dashboard"><div class="icon">📈</div><div><div class="title">Dashboard</div><div class="desc">查看收入、支出与月份统计。</div></div></a>
      <a class="card" href="/finance/reports/excel"><div class="icon">📥</div><div><div class="title">Excel 下载中心</div><div class="desc">下载 CHE、STW、布施、支出及 Petty Cash 报表。</div></div></a>
    </section></main></body></html>
    """)


@finance_v7_bp.route("/finance/v7/system")
def finance_v7_system_management():
    tools = [
        ("📥", "历史 Excel 导入", "上传旧月费、布施、支出及 Petty Cash Excel；先预览，再确认。", "/finance/import/"),
        ("🌏", "STW 银行月费导入", "上传 STW 银行月费 Excel；自动检查、预览后批量导入。", "/finance/admin/stw-bank-monthly-import"),
        ("📕", "收条簿管理", "管理 CHE、STW 与各布施类别的当前及下一张收条。", "/finance/admin/receipt-books"),
        ("📒", "收条区间检查", "检查缺号、重复、作废与区间合计。", "/finance/admin/receipt-summary"),
        ("📇", "布施人资料", "维护布施人姓名及电话号码。", "/finance/donors"),
        ("👥", "会员资料", "查看会员状态及付款历史。", "/finance/menu/member"),
        ("🏢", "付款对象", "维护供花、供油、水电等付款对象。", "/finance/vendors"),
    ]
    return render_template_string(r"""
    <!doctype html><html lang="zh"><head>
    <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
    <title>系统管理</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}">
    <style>
    body{margin:0;background:#f7f5fb;color:#202638}.page{max-width:1000px;margin:auto;padding:24px 14px}
    .top{display:flex;justify-content:space-between;align-items:center;gap:16px}.back{text-decoration:none;font-weight:900;color:#6340a7}
    .notice{background:#fff8df;border:1px solid #ead89a;border-radius:15px;padding:14px 16px;line-height:1.6;margin-top:16px}
    .grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px;margin-top:18px}
    .card{display:flex;align-items:center;gap:14px;text-decoration:none;color:inherit;background:#fff;border:1px solid #e2ddeb;border-radius:18px;padding:18px;min-height:88px}
    .icon{width:48px;height:48px;display:grid;place-items:center;flex:0 0 auto;background:#f0ebfb;border-radius:15px;font-size:25px}
    .title{font-size:17px;font-weight:950}.desc{color:#746d7e;margin-top:5px;line-height:1.5;font-size:14px}
    @media(max-width:680px){.grid{grid-template-columns:1fr}.top{align-items:flex-start}}
    </style></head><body><main class="page">
    <div class="top"><div><h1>⚙️ 系统管理</h1><p>后台工具集中在这里；所有正式入口统一由 V7 管理。</p></div><a class="back" href="{{ url_for('finance_v7.finance_v7_home') }}">← 返回</a></div>
    <div class="notice">CHE、STW、总会三个单位按钮继续保留。历史上传、导入批次、收条簿、会员、布施人及付款对象统一集中在系统管理。</div>
    <section class="grid">{% for icon,title,desc,url in tools %}<a class="card" href="{{ url }}"><div class="icon">{{ icon }}</div><div><div class="title">{{ title }}</div><div class="desc">{{ desc }}</div></div></a>{% endfor %}</section>
    </main></body></html>
    """, tools=tools)


@finance_v7_bp.route("/finance/v7/CHE")
def finance_v7_che_home():
    data = _v7_che_home_data()

    return render_template_string(
        """
        <!doctype html>
        <html lang="zh">
        <head>

            <meta charset="utf-8">

            <meta
                name="viewport"
                content="width=device-width, initial-scale=1"
            >

            <title>CHE 观音堂财政</title>

            <link
                rel="stylesheet"
                href="{{ url_for(
                    'static',
                    filename='css/toolbox.css'
                ) }}"
            >

            <style>

                body {
                    margin: 0;
                    background: #f4f7fb;
                    color: #172033;
                }

                .v7-page {
                    width: min(1180px, calc(100% - 32px));
                    margin: 0 auto;
                    padding: 24px 0 46px;
                }

                /* =========================================
                   页面顶部
                ========================================= */

                .v7-topbar {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 18px;
                    margin-bottom: 20px;
                }

                .v7-title-wrap {
                    display: flex;
                    align-items: center;
                    gap: 14px;
                }

                .v7-title-icon {
                    width: 58px;
                    height: 58px;
                    border-radius: 18px;
                    display: grid;
                    place-items: center;
                    background: #e7efff;
                    font-size: 30px;
                    flex: 0 0 auto;
                }

                .v7-title-wrap h1 {
                    margin: 0;
                    font-size: 31px;
                    line-height: 1.2;
                }

                .v7-title-wrap p {
                    margin: 6px 0 0;
                    color: #6d7687;
                    font-size: 16px;
                }

                .v7-back-button {
                    text-decoration: none;
                    background: white;
                    color: #253047;
                    border: 1px solid #d9dfeb;
                    border-radius: 13px;
                    padding: 11px 17px;
                    font-weight: 800;
                    white-space: nowrap;
                }

                /* =========================================
                   顶部状态
                ========================================= */

                .v7-status-grid {
                    display: grid;
                    grid-template-columns:
                        repeat(5, minmax(0, 1fr));
                    gap: 14px;
                    margin-bottom: 19px;
                }

                .v7-status-box {
                    background: white;
                    border: 1px solid #e1e7f0;
                    border-radius: 17px;
                    padding: 16px 18px;
                    box-shadow:
                        0 8px 24px rgba(33, 50, 83, 0.05);
                }

                .v7-status-label {
                    color: #788192;
                    font-size: 14px;
                    margin-bottom: 7px;
                }

                .v7-status-value {
                    font-size: 20px;
                    font-weight: 900;
                }

                /* =========================================
                   主工作台卡片
                ========================================= */

                .v7-card-grid {
                    display: grid;
                    grid-template-columns:
                        repeat(2, minmax(0, 1fr));
                    gap: 18px;
                }

                .v7-card {
                    display: block;
                    text-decoration: none;
                    color: inherit;
                    background: white;
                    border: 1px solid #e1e7f0;
                    border-radius: 22px;
                    padding: 22px;
                    box-shadow:
                        0 10px 30px rgba(33, 50, 83, 0.07);
                    transition:
                        transform 0.16s ease,
                        box-shadow 0.16s ease,
                        border-color 0.16s ease;
                }

                .v7-card:hover {
                    transform: translateY(-2px);
                    border-color: #c9d5e9;
                    box-shadow:
                        0 15px 36px rgba(33, 50, 83, 0.11);
                }

                .v7-card-head {
                    display: flex;
                    align-items: flex-start;
                    justify-content: space-between;
                    gap: 16px;
                }

                .v7-card-title-wrap {
                    display: flex;
                    align-items: center;
                    gap: 14px;
                }

                .v7-card-icon {
                    width: 53px;
                    height: 53px;
                    border-radius: 16px;
                    display: grid;
                    place-items: center;
                    font-size: 27px;
                    flex: 0 0 auto;
                }

                .icon-bank {
                    background: #e6efff;
                }

                .icon-cash {
                    background: #e7f8ef;
                }

                .icon-monthly-fee {
                    background: #e7f7f8;
                }

                .icon-expense {
                    background: #fff2df;
                }

                .icon-care {
                    background: #fff0f4;
                }

                .icon-close {
                    background: #efeaff;
                }

                .v7-card h2 {
                    margin: 0;
                    font-size: 22px;
                }

                .v7-card-desc {
                    margin-top: 5px;
                    color: #747d8e;
                    font-size: 15px;
                    line-height: 1.4;
                }

                .v7-arrow {
                    color: #7a8498;
                    font-size: 26px;
                    line-height: 1;
                }

                .v7-main-value {
                    margin-top: 22px;
                    font-size: 32px;
                    line-height: 1.1;
                    font-weight: 900;
                }

                .v7-main-label {
                    margin-top: 7px;
                    color: #687286;
                    font-size: 15px;
                }

                .v7-sub-grid {
                    display: grid;
                    grid-template-columns:
                        repeat(2, minmax(0, 1fr));
                    gap: 10px;
                    margin-top: 18px;
                }

                .v7-sub-box {
                    background: #f5f7fa;
                    border-radius: 14px;
                    padding: 13px 14px;
                }

                .v7-sub-label {
                    color: #747e90;
                    font-size: 13px;
                    margin-bottom: 5px;
                }

                .v7-sub-value {
                    font-size: 18px;
                    font-weight: 900;
                    overflow-wrap: anywhere;
                }

                /* =========================================
                   月结进度
                ========================================= */

                .v7-progress {
                    height: 10px;
                    margin-top: 18px;
                    background: #eceff5;
                    border-radius: 99px;
                    overflow: hidden;
                }

                .v7-progress-bar {
                    height: 100%;
                    width: {{ data.month_close_progress }}%;
                    background: #8065d8;
                    border-radius: 99px;
                }

                /* =========================================
                   手机版
                ========================================= */

                @media (max-width: 850px) {

                    .v7-status-grid {
                        grid-template-columns:
                            repeat(2, minmax(0, 1fr));
                    }

                    .v7-card-grid {
                        grid-template-columns: 1fr;
                    }
                }

                @media (max-width: 560px) {

                    .v7-page {
                        width: min(100% - 20px, 1180px);
                        padding-top: 14px;
                    }

                    .v7-topbar {
                        align-items: flex-start;
                    }

                    .v7-title-icon {
                        width: 48px;
                        height: 48px;
                        border-radius: 15px;
                        font-size: 25px;
                    }

                    .v7-title-wrap h1 {
                        font-size: 25px;
                    }

                    .v7-title-wrap p {
                        font-size: 14px;
                    }

                    .v7-back-button {
                        padding: 9px 12px;
                        font-size: 14px;
                    }

                    .v7-status-grid {
                        gap: 9px;
                    }

                    .v7-status-box {
                        padding: 13px;
                    }

                    .v7-status-value {
                        font-size: 17px;
                    }

                    .v7-card {
                        padding: 18px;
                    }

                    .v7-card h2 {
                        font-size: 20px;
                    }

                    .v7-main-value {
                        font-size: 28px;
                    }

                    .v7-sub-value {
                        font-size: 16px;
                    }
                }

            </style>

        </head>

        <body>

        <main class="v7-page">

            <!-- =========================================
                 页面标题
            ========================================== -->

            <header class="v7-topbar">

                <div class="v7-title-wrap">

                    <div class="v7-title-icon">
                        🛡️
                    </div>

                    <div>
                        <h1>CHE 观音堂财政</h1>

                        <p>
                            完整财政工作台 · 今日 {{ today }}
                        </p>
                    </div>

                </div>

                <a
                    class="v7-back-button"
                    href="/finance/v7"
                >
                    ← 返回单位选择
                </a>

            </header>

            <!-- =========================================
                 顶部总览
            ========================================== -->

            <section class="v7-status-grid">

                <div class="v7-status-box">
                    <div class="v7-status-label">
                        银行余额
                    </div>

                    <div class="v7-status-value">
                        RM {{ "%.2f"|format(
                            data.bank_balance
                        ) }}
                    </div>
                </div>

                <div class="v7-status-box">
                    <div class="v7-status-label">
                        Cash In Hand
                    </div>

                    <div class="v7-status-value">
                        RM {{ "%.2f"|format(
                            data.cash_balance
                        ) }}
                    </div>
                </div>

                <div class="v7-status-box">
                    <div class="v7-status-label">
                        Petty Cash
                    </div>

                    <div class="v7-status-value">
                        RM {{ "%.2f"|format(data.petty_cash) }}
                    </div>
                </div>

                <div class="v7-status-box">
                    <div class="v7-status-label">
                        银行待确认
                    </div>

                    <div class="v7-status-value">
                        {{ data.bank_pending_count }} 笔
                    </div>
                </div>

                <div class="v7-status-box">
                    <div class="v7-status-label">
                        月费迟付关怀
                    </div>

                    <div class="v7-status-value">
                        {{ data.late_member_count }} 人
                    </div>
                </div>

            </section>

            <!-- =========================================
                 六个最终工作台
            ========================================== -->

            <section class="v7-card-grid">

                <!-- 银行 -->

                <a
                    class="v7-card"
                    href="{{ url_for(
                        'finance_v7.finance_v7_bank'
                    ) }}"
                >

                    <div class="v7-card-head">

                        <div class="v7-card-title-wrap">

                            <div class="v7-card-icon icon-bank">
                                🏦
                            </div>

                            <div>
                                <h2>银行专区</h2>

                                <div class="v7-card-desc">
                                    CDM、银行对账、Cash In Hand 与数据维修
                                </div>
                            </div>

                        </div>

                        <div class="v7-arrow">›</div>

                    </div>

                    <div class="v7-main-value">
                        RM {{ "%.2f"|format(
                            data.bank_balance
                        ) }}
                    </div>

                    <div class="v7-main-label">
                        当前银行余额
                    </div>

                    <div class="v7-sub-grid">

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                待确认记录
                            </div>

                            <div class="v7-sub-value">
                                {{ data.bank_pending_count }} 笔
                            </div>
                        </div>

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                待确认金额
                            </div>

                            <div class="v7-sub-value">
                                RM {{ "%.2f"|format(
                                    data.bank_pending_total
                                ) }}
                            </div>
                        </div>

                    </div>

                </a>

                <!-- Cash -->

                <a
                    class="v7-card"
                    href="{{ url_for(
                        'finance_v7.finance_v7_cash'
                    ) }}"
                >

                    <div class="v7-card-head">

                        <div class="v7-card-title-wrap">

                            <div class="v7-card-icon icon-cash">
                                💵
                            </div>

                            <div>
                                <h2>Cash</h2>

                                <div class="v7-card-desc">
                                    月费现金与日常备用金分开管理
                                </div>
                            </div>

                        </div>

                        <div class="v7-arrow">›</div>

                    </div>

                    <div class="v7-main-value">
                        RM {{ "%.2f"|format(data.total_cash) }}
                    </div>

                    <div class="v7-main-label">
                        两种现金合计
                    </div>

                    <div class="v7-sub-grid">

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                Cash In Hand
                            </div>

                            <div class="v7-sub-value">
                                RM {{ "%.2f"|format(data.cash_in_hand) }}
                            </div>
                        </div>

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                Petty Cash
                            </div>

                            <div class="v7-sub-value">
                                RM {{ "%.2f"|format(data.petty_cash) }}
                            </div>
                        </div>

                    </div>

                </a>

                <!-- CHE 月费 -->

                <a
                    class="v7-card"
                    href="{{ url_for(
                        'finance_v7.finance_v7_che_monthly_fee'
                    ) }}"
                >

                    <div class="v7-card-head">

                        <div class="v7-card-title-wrap">

                            <div class="v7-card-icon icon-monthly-fee">
                                ❤️
                            </div>

                            <div>
                                <h2>CHE 月费</h2>

                                <div class="v7-card-desc">
                                    月费收款、收条及银行转账确认
                                </div>
                            </div>

                        </div>

                        <div class="v7-arrow">›</div>

                    </div>

                    <div class="v7-main-value">
                        RM {{ "%.2f"|format(
                            data.monthly_fee_month_total
                        ) }}
                    </div>

                    <div class="v7-main-label">
                        本月 CHE 月费
                    </div>

                    <div class="v7-sub-grid">

                        <div class="v7-sub-box">

                            <div class="v7-sub-label">
                                💵 现金月费
                            </div>

                            <div class="v7-sub-value">
                                RM {{ "%.2f"|format(
                                    data.monthly_fee_cash_total
                                ) }}
                            </div>

                        </div>

                        <div class="v7-sub-box">

                            <div class="v7-sub-label">
                                🏦 银行月费
                            </div>

                            <div class="v7-sub-value">
                                RM {{ "%.2f"|format(
                                    data.monthly_fee_bank_total
                                ) }}
                            </div>

                        </div>

                    </div>

                </a>

                <!-- CHE 支出 -->

                <a
                    class="v7-card"
                    href="{{ url_for(
                        'finance_v7.finance_v7_expense'
                    ) }}"
                >

                    <div class="v7-card-head">

                        <div class="v7-card-title-wrap">

                            <div class="v7-card-icon icon-expense">
                                📤
                            </div>

                            <div>
                                <h2>CHE 支出</h2>

                                <div class="v7-card-desc">
                                    支出录入、PV 与待补资料
                                </div>
                            </div>

                        </div>

                        <div class="v7-arrow">›</div>

                    </div>

                    <div class="v7-main-value">
                        RM {{ "%.2f"|format(
                            data.expense_month_total
                        ) }}
                    </div>

                    <div class="v7-main-label">
                        本月 CHE 支出
                    </div>

                    <div class="v7-sub-grid">

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                待补 PV
                            </div>

                            <div class="v7-sub-value">
                                {{ data.expense_pending_pv_count }} 笔
                            </div>
                        </div>

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                今日支出
                            </div>

                            <div class="v7-sub-value">
                                RM {{ "%.2f"|format(
                                    data.expense_today_total
                                ) }}
                            </div>
                        </div>

                    </div>

                </a>

                <!-- 月费关怀 -->

                <a
                    class="v7-card"
                    href="{{ url_for(
                        'finance_v7.finance_v7_member_care'
                    ) }}"
                >

                    <div class="v7-card-head">

                        <div class="v7-card-title-wrap">

                            <div class="v7-card-icon icon-care">
                                ❤️
                            </div>

                            <div>
                                <h2>月费关怀</h2>

                                <div class="v7-card-desc">
                                    查看迟付会员与后续关怀情况
                                </div>
                            </div>

                        </div>

                        <div class="v7-arrow">›</div>

                    </div>

                    <div class="v7-main-value">
                        {{ data.late_member_count }} 人
                    </div>

                    <div class="v7-main-label">
                        当前需要关怀
                    </div>

                    <div class="v7-sub-grid">

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                今日新增迟付
                            </div>

                            <div class="v7-sub-value">
                                {{ data.new_late_member_count }} 人
                            </div>
                        </div>

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                最后更新
                            </div>

                            <div class="v7-sub-value">
                                {{ data.care_updated_at }}
                            </div>
                        </div>

                    </div>

                </a>

                <!-- 月结 -->

                <a
                    class="v7-card"
                    href="{{ url_for(
                        'finance_v7.finance_v7_month_close'
                    ) }}"
                >

                    <div class="v7-card-head">

                        <div class="v7-card-title-wrap">

                            <div class="v7-card-icon icon-close">
                                📅
                            </div>

                            <div>
                                <h2>财政月结</h2>

                                <div class="v7-card-desc">
                                    核对、锁账与下载财政月报
                                </div>
                            </div>

                        </div>

                        <div class="v7-arrow">›</div>

                    </div>

                    <div class="v7-main-value">
                        {{ data.month_close_status }}
                    </div>

                    <div class="v7-main-label">
                        {{ data.month_close_period }} 月结状态
                    </div>

                    <div class="v7-progress">
                        <div class="v7-progress-bar"></div>
                    </div>

                    <div class="v7-sub-grid">

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                完成进度
                            </div>

                            <div class="v7-sub-value">
                                {{ data.month_close_progress }}%
                            </div>
                        </div>

                        <div class="v7-sub-box">
                            <div class="v7-sub-label">
                                月结月份
                            </div>

                            <div class="v7-sub-value">
                                {{ data.month_close_period }}
                            </div>
                        </div>

                    </div>

                </a>

            </section>

        </main>

        </body>
        </html>
        """,
        data=data,
        today=date.today().strftime("%Y-%m-%d"),
    )

@finance_v7_bp.route("/finance/v7/STW")
def finance_v7_stw_home():
    data = _v7_monthly_data("STW")

    return render_template_string(
        r"""
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>STW 收款工作台</title>
            <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
            <style>
                body{margin:0;background:#f5f3fa;color:#201b2c;}
                .stw-page{width:min(900px,calc(100% - 24px));margin:0 auto;padding:22px 0 48px;}
                .topbar{display:flex;align-items:center;justify-content:space-between;gap:14px;margin-bottom:16px;}
                .topbar h1{margin:0;font-size:30px;line-height:1.2;}
                .topbar p{margin:6px 0 0;color:#776f86;font-size:14px;}
                .back{display:inline-flex;align-items:center;justify-content:center;padding:10px 14px;border:1px solid #ded7e8;border-radius:14px;background:#fff;color:#3a3148;text-decoration:none;font-weight:900;white-space:nowrap;}
                .hero{background:linear-gradient(145deg,#5b46a5,#7b67c6);color:#fff;border-radius:26px;padding:28px;box-shadow:0 18px 42px rgba(75,55,145,.22);}
                .hero-label{font-size:14px;font-weight:800;opacity:.86;}
                .hero-amount{font-size:46px;line-height:1.08;font-weight:950;margin:7px 0 10px;letter-spacing:-1px;}
                .hero-sub{font-size:15px;font-weight:800;opacity:.88;}
                .hero-status{margin-top:15px;display:inline-flex;align-items:center;gap:8px;background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.2);border-radius:999px;padding:9px 13px;font-weight:800;}
                .stats{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:14px;}
                .stat{background:#fff;border:1px solid #e2deec;border-radius:19px;padding:18px;box-shadow:0 8px 24px rgba(54,38,89,.06);}
                .stat-label{color:#786f87;font-size:13px;font-weight:900;}
                .stat-value{margin-top:7px;font-size:27px;font-weight:950;color:#30283f;}
                .section{margin-top:18px;background:#fff;border:1px solid #e2deec;border-radius:22px;padding:20px;box-shadow:0 8px 26px rgba(54,38,89,.06);}
                .section h2{margin:0 0 14px;font-size:21px;}
                .actions{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;}
                .action{display:flex;align-items:center;gap:13px;min-height:72px;padding:15px 17px;border:1px solid #ded8ea;border-radius:18px;background:#faf9fd;color:#2d253a;text-decoration:none;font-weight:950;font-size:17px;}
                .action:hover{background:#f2eef9;border-color:#cfc5e2;}
                .action-icon{width:42px;height:42px;border-radius:14px;display:grid;place-items:center;background:#ece7f7;font-size:22px;flex:0 0 auto;}
                .recent-list{display:grid;gap:10px;}
                .recent-item{display:grid;grid-template-columns:1fr auto;gap:14px;align-items:center;border:1px solid #e6e1ee;border-radius:17px;padding:15px 16px;background:#fcfbfe;}
                .recent-top{display:flex;flex-wrap:wrap;gap:7px 12px;align-items:center;}
                .member{font-weight:950;font-size:17px;}
                .name{font-weight:850;color:#4d435a;}
                .meta{margin-top:6px;color:#7a7187;font-size:13px;line-height:1.55;}
                .amount{font-size:22px;font-weight:950;color:#5b46a5;white-space:nowrap;text-align:right;}
                .time{font-size:12px;color:#92899e;margin-top:4px;text-align:right;}
                .summary{display:flex;align-items:flex-start;gap:12px;background:#f1ecfb;border:1px solid #ddd3f0;border-radius:18px;padding:16px 17px;line-height:1.65;color:#4f4263;}
                .summary strong{color:#302440;}
                @media(max-width:680px){
                    .stw-page{width:min(100% - 16px,900px);padding-top:12px;}
                    .topbar{align-items:flex-start;}.topbar h1{font-size:25px;}.back{padding:9px 11px;font-size:13px;}
                    .hero{padding:22px 20px;border-radius:22px;}.hero-amount{font-size:38px;}
                    .stats{grid-template-columns:1fr 1fr;}.stat:last-child{grid-column:1/-1;}
                    .actions{grid-template-columns:1fr;}
                    .recent-item{grid-template-columns:1fr;}.amount,.time{text-align:left;}
                }
            </style>
        </head>
        <body>
        <main class="stw-page">
            <header class="topbar">
                <div>
                    <h1>🧾 STW 收款</h1>
                    <p>这里只处理 STW 月费与 Bank In，不显示 CHE、HQ、Cash 或支出。</p>
                </div>
                <a class="back" href="{{ url_for('finance_v7.finance_v7_home') }}">← 单位选择</a>
            </header>

            <section class="hero">
                <div class="hero-label">本月月费</div>
                <div class="hero-amount">RM {{ '%.2f'|format(data.month_total) }}</div>
                <div class="hero-sub">本月共 {{ data.month_people }} 位佛友</div>
                <div class="hero-status">💬 今日收到 {{ data.today_people }} 位佛友，共 RM {{ '%.2f'|format(data.today_total) }}</div>
            </section>

            <section class="stats">
                <div class="stat"><div class="stat-label">今日人数</div><div class="stat-value">{{ data.today_people }} 位</div></div>
                <div class="stat"><div class="stat-label">今日金额</div><div class="stat-value">RM {{ '%.2f'|format(data.today_total) }}</div></div>
                <div class="stat"><div class="stat-label">待确认银行月费</div><div class="stat-value">{{ data.pending_count }} 笔</div></div>
            </section>

            <section class="section">
                <h2>今天要做什么？</h2>
                <div class="actions">
                    <a class="action" href="{{ url_for('finance_v7.finance_v7_stw_collect') }}"><span class="action-icon">❤️</span><span>收取月费</span></a>
                    <a class="action" href="{{ url_for('finance_v7.finance_v7_stw_bank_pending') }}"><span class="action-icon">🏦</span><span>确认银行月费</span></a>
                    <a class="action" href="{{ url_for('finance_v7.finance_v7_stw_records') }}"><span class="action-icon">📖</span><span>收款记录</span></a>
                </div>
            </section>

            <section class="section">
                <h2>最近收款</h2>
                <div class="recent-list">
                    {% for row in data.recent %}
                    <div class="recent-item">
                        <div>
                            <div class="recent-top"><span class="member">{{ row.member_id }}</span><span class="name">{{ row.name }}</span></div>
                            <div class="meta">{{ row.receipt_no }} · {{ row.month_label }} · {{ row.method }}</div>
                        </div>
                        <div><div class="amount">RM {{ '%.2f'|format(row.amount) }}</div><div class="time">{{ row.time }}</div></div>
                    </div>
                    {% endfor %}
                </div>
            </section>

            <section class="section">
                <h2>今日摘要</h2>
                <div class="summary">
                    <div>✅</div>
                    <div>
                        今日已收 <strong>{{ data.today_people }} 位佛友</strong>，合计 <strong>RM {{ '%.2f'|format(data.today_total) }}</strong>。
                        {% if data.pending_count %}还有 <strong>{{ data.pending_count }} 笔银行月费</strong>待确认。{% else %}今天收款工作已完成。{% endif %}
                    </div>
                </div>
            </section>
        </main>
        </body>
        </html>
        """,
        data=data,
    )


def _render_stw_placeholder(title, icon, description):
    return render_template_string(
        """
        <!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
        <title>{{ title }}</title><link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
        <style>body{margin:0;background:#f5f3fa;color:#201b2c}.page{max-width:720px;margin:0 auto;padding:26px 14px}.box{background:#fff;border:1px solid #e2deec;border-radius:22px;padding:24px;box-shadow:0 10px 28px rgba(54,38,89,.07)}h1{margin:0 0 10px}.desc{color:#70677d;font-size:16px;line-height:1.7}.back{display:inline-flex;margin-top:18px;text-decoration:none;font-weight:900;color:#fff;background:#6550ad;padding:12px 16px;border-radius:14px}</style>
        </head><body><main class="page"><section class="box"><h1>{{ icon }} {{ title }}</h1><div class="desc">{{ description }}</div><a class="back" href="{{ url_for('finance_v7.finance_v7_stw_home') }}">← 返回 STW 收款</a></section></main></body></html>
        """,
        title=title,
        icon=icon,
        description=description,
    )


@finance_v7_bp.route("/finance/v7/STW/collect")
def finance_v7_stw_collect():
    return redirect(
        url_for(
            "finance.monthly_fee_batch",
            branch="STW"
        )
    )


@finance_v7_bp.route("/finance/v7/STW/bank-pending")
def finance_v7_stw_bank_pending():
    return redirect("/finance/bank_pending?branch=STW")



@finance_v7_bp.route("/finance/v7/STW/records")
def finance_v7_stw_records():
    return _render_v7_module_records('stw',False)


@finance_v7_bp.route("/finance/v7/STW/search")
def finance_v7_stw_search():
    return _render_v7_module_records('stw',True)


@finance_v7_bp.route("/finance/v7/HQ")
def finance_v7_hq_home():
    data = _v7_hq_data()

    return render_template_string(
        r"""
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>总会财布施中心</title>
            <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
            <style>
                body{margin:0;background:#f5f3fa;color:#201b2c;}
                .hq-page{width:min(920px,calc(100% - 24px));margin:0 auto;padding:22px 0 48px;}
                .topbar{display:flex;align-items:center;justify-content:space-between;gap:14px;margin-bottom:16px;}
                .topbar h1{margin:0;font-size:30px;line-height:1.2;}
                .topbar p{margin:6px 0 0;color:#776f86;font-size:14px;}
                .back{display:inline-flex;align-items:center;justify-content:center;padding:10px 14px;border:1px solid #ded7e8;border-radius:14px;background:#fff;color:#3a3148;text-decoration:none;font-weight:900;white-space:nowrap;}
                .hero{background:linear-gradient(145deg,#5d3f89,#7d5aae);color:#fff;border-radius:26px;padding:28px;box-shadow:0 18px 42px rgba(80,53,122,.22);}
                .hero-label{font-size:14px;font-weight:800;opacity:.86;}
                .hero-amount{font-size:46px;line-height:1.08;font-weight:950;margin:7px 0 12px;letter-spacing:-1px;}
                .hero-status{display:inline-flex;align-items:center;gap:8px;background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.2);border-radius:999px;padding:9px 13px;font-weight:800;}
                .stats{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:14px;}
                .stat{background:#fff;border:1px solid #e3ddeb;border-radius:18px;padding:17px;box-shadow:0 8px 24px rgba(55,38,78,.05);}
                .stat-label{color:#857d90;font-size:13px;font-weight:800;}
                .stat-value{margin-top:7px;font-size:24px;font-weight:950;}
                .work-title,.section-title{margin:24px 2px 11px;font-size:18px;font-weight:950;}
                .actions{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;}
                .action{display:flex;align-items:center;gap:14px;min-height:82px;padding:15px 17px;background:#fff;border:1px solid #e2dce9;border-radius:19px;text-decoration:none;color:#2b2436;box-shadow:0 8px 25px rgba(55,38,78,.05);transition:.16s ease;}
                .action:hover{transform:translateY(-2px);box-shadow:0 13px 30px rgba(55,38,78,.1);}
                .action-icon{width:46px;height:46px;display:grid;place-items:center;border-radius:15px;background:#f1ebf8;font-size:23px;flex:0 0 auto;}
                .action-name{font-size:17px;font-weight:950;}
                .action-help{margin-top:4px;color:#867e91;font-size:12px;line-height:1.35;}
                .primary-action{background:#fff9ef;border-color:#eed7a8;}
                .primary-action .action-icon{background:#ffe9bd;}
                .pending-badge{display:inline-block;margin-left:6px;padding:2px 8px;border-radius:999px;background:#fff0c9;color:#855a00;font-size:12px;font-weight:950;}
                .layout{display:grid;grid-template-columns:minmax(0,1.25fr) minmax(260px,.75fr);gap:14px;align-items:start;}
                .panel{background:#fff;border:1px solid #e2dce9;border-radius:21px;padding:18px;box-shadow:0 8px 25px rgba(55,38,78,.05);}
                .recent-list{display:grid;gap:10px;}
                .recent-row{display:grid;grid-template-columns:minmax(0,1fr) auto;gap:12px;align-items:center;padding:14px;border:1px solid #ebe6f0;border-radius:16px;background:#fcfbfd;}
                .receipt{font-size:12px;color:#8b8395;font-weight:850;}
                .donor{font-size:16px;font-weight:950;margin:3px 0;}
                .meta{color:#776f83;font-size:13px;display:flex;gap:8px;flex-wrap:wrap;}
                .amount{text-align:right;font-size:17px;font-weight:950;color:#654292;white-space:nowrap;}
                .time{margin-top:4px;color:#9a93a3;font-size:11px;}
                .category-list{display:grid;gap:2px;}
                .category-row{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:14px 3px;border-bottom:1px solid #eee9f2;}
                .category-row:last-child{border-bottom:0;}
                .category-name{font-weight:900;}
                .category-amount{font-weight:950;color:#594075;white-space:nowrap;}
                .foot-note{margin-top:14px;padding:13px 15px;border-radius:15px;background:#f0ecf6;color:#655a72;font-size:13px;line-height:1.5;}
                @media(max-width:700px){
                    .topbar{align-items:flex-start}.topbar h1{font-size:24px}.back{font-size:13px;padding:9px 11px}
                    .hero{padding:23px 20px}.hero-amount{font-size:37px}
                    .stats{grid-template-columns:1fr 1fr}.stat:last-child{grid-column:1/-1}
                    .actions{grid-template-columns:1fr}
                    .layout{grid-template-columns:1fr}
                }
            </style>
        </head>
        <body>
        <main class="hq-page">
            <div class="topbar">
                <div>
                    <h1>🙏 总会财布施中心</h1>
                    <p>这里只处理总会布施，不显示 CHE 月费、Cash、PV 或 STW。</p>
                </div>
                <a class="back" href="{{ url_for('finance_v7.finance_v7_home') }}">← 单位选择</a>
            </div>

            <section class="hero">
                <div class="hero-label">本月累计布施</div>
                <div class="hero-amount">RM {{ "{:,.2f}".format(data.month_total) }}</div>
                <div class="hero-status">💬 今日收到 {{ data.today_count }} 笔布施</div>
            </section>

            <section class="stats">
                <div class="stat">
                    <div class="stat-label">今日笔数</div>
                    <div class="stat-value">{{ data.today_count }} 笔</div>
                </div>
                <div class="stat">
                    <div class="stat-label">今日金额</div>
                    <div class="stat-value">RM {{ "{:,.2f}".format(data.today_total) }}</div>
                </div>
                <div class="stat">
                    <div class="stat-label">待确认 Bank In</div>
                    <div class="stat-value">{{ data.pending_count }} 笔</div>
                </div>
            </section>

            <div class="work-title">今天要做什么？</div>
            <section class="actions">
                <a class="action primary-action" href="{{ url_for('finance_v7.finance_v7_hq_donation_new') }}">
                    <div class="action-icon">🙏</div>
                    <div><div class="action-name">登记布施</div><div class="action-help">先选择布施类别，再进入录入。</div></div>
                </a>
                <a class="action" href="{{ url_for('finance_v7.finance_v7_hq_bank_pending') }}">
                    <div class="action-icon">🏦</div>
                    <div><div class="action-name">确认 Bank In <span class="pending-badge">{{ data.pending_count }}</span></div><div class="action-help">只处理总会布施的银行待确认。</div></div>
                </a>
                <a class="action" href="{{ url_for('finance_v7.finance_v7_hq_records') }}">
                    <div class="action-icon">📖</div>
                    <div><div class="action-name">布施记录</div><div class="action-help">查看总会最近和历史布施。</div></div>
                </a>
            </section>

            <section class="layout">
                <div>
                    <div class="section-title">最近布施</div>
                    <div class="panel recent-list">
                        {% for row in data.recent %}
                        <div class="recent-row">
                            <div>
                                <div class="receipt">{{ row.receipt_no }}</div>
                                <div class="donor">{{ row.name }}</div>
                                <div class="meta"><span>{{ row.icon }} {{ row.category }}</span><span>·</span><span>{{ row.method }}</span></div>
                            </div>
                            <div>
                                <div class="amount">RM {{ "{:,.2f}".format(row.amount) }}</div>
                                <div class="time">{{ row.time }}</div>
                            </div>
                        </div>
                        {% endfor %}
                    </div>
                </div>

                <div>
                    <div class="section-title">本月分类</div>
                    <div class="panel category-list">
                        {% for item in data.categories %}
                        <div class="category-row">
                            <div class="category-name">{{ item.icon }} {{ item.name }}</div>
                            <div class="category-amount">RM {{ "{:,.2f}".format(item.amount) }}</div>
                        </div>
                        {% endfor %}
                        <div class="foot-note">分类统计只属于总会财布施，不会混入 CHE 月费或 STW 收款。</div>
                    </div>
                </div>
            </section>
        </main>
        </body>
        </html>
        """,
        data=data,
    )


@finance_v7_bp.route("/finance/v7/HQ/donation/new")
def finance_v7_hq_donation_new():

    categories = [
        {"name": "财布施", "icon": "❤️", "help": "一般财布施记录"},
        {"name": "观音村", "icon": "🏠", "help": "观音村相关布施"},
        {"name": "膳食结缘", "icon": "🍚", "help": "初一十五及膳食结缘"},
        {"name": "特别布施", "icon": "🌟", "help": "特别活动或指定用途"},
        {"name": "纯檀香布施", "icon": "🪔", "help": "纯檀香相关布施"},
    ]

    return render_template_string(
        r"""
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>选择布施类别</title>
            <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
            <style>
                body{margin:0;background:#f5f3fa;color:#201b2c}.page{width:min(760px,calc(100% - 24px));margin:0 auto;padding:24px 0 48px}
                .head{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;margin-bottom:18px}.head h1{margin:0;font-size:28px}.head p{margin:6px 0 0;color:#7a7287}
                .back{padding:10px 14px;border:1px solid #ded7e8;border-radius:13px;background:#fff;color:#392f48;text-decoration:none;font-weight:900;white-space:nowrap}
                .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}.item{display:flex;align-items:center;gap:15px;min-height:92px;padding:18px;background:#fff;border:1px solid #e1dae9;border-radius:20px;text-decoration:none;color:#2b2436;box-shadow:0 9px 25px rgba(55,38,78,.05)}
                .icon{width:50px;height:50px;display:grid;place-items:center;border-radius:16px;background:#f1ebf8;font-size:25px}.name{font-size:18px;font-weight:950}.help{margin-top:5px;color:#847b8f;font-size:13px}
                .note{margin-top:16px;padding:14px 16px;border-radius:16px;background:#eee8f5;color:#665b73;font-size:13px;line-height:1.55}
                @media(max-width:580px){.head h1{font-size:23px}.grid{grid-template-columns:1fr}}
            </style>
        </head>
        <body><main class="page">
            <div class="head"><div><h1>🙏 请选择布施类别</h1><p>先选类别，再进入对应的布施录入。</p></div><a class="back" href="{{ url_for('finance_v7.finance_v7_hq_home') }}">← 返回 HQ</a></div>
            <section class="grid">
                {% for item in categories %}
                <a class="item" href="{{ url_for('finance_v7.finance_v7_hq_donation_form', category=item.name) }}">
                    <div class="icon">{{ item.icon }}</div><div><div class="name">{{ item.name }}</div><div class="help">{{ item.help }}</div></div>
                </a>
                {% endfor %}
            </section>
            <div class="note">这一页只负责选择类别。真正的姓名搜索、收条编号、金额和付款方式，将在后续统一重做输入功能时接上。</div>
        </main></body></html>
        """,
        categories=categories,
    )


@finance_v7_bp.route("/finance/v7/HQ/donation/form")
def finance_v7_hq_donation_form():
    """V7 负责人共用主页成熟的批量布施录入表单。"""
    raw_category = request.args.get("category", "财布施").strip() or "财布施"

    category_map = {
        "财布施": "财布施",
        "观音村": "观音村",
        "膳食结缘": "膳食结缘",
        "特别布施": "临时特别布施",
        "临时特别布施": "临时特别布施",
        "纯檀香布施": "观音堂纯檀香布施",
        "观音堂纯檀香布施": "观音堂纯檀香布施",
        "813交流会财布施": "813交流会财布施",
    }

    category = category_map.get(raw_category, "财布施")

    return redirect(
        url_for(
            "finance.income_batch",
            category=category,
            source="v7",
        )
    )



@finance_v7_bp.route("/finance/v7/HQ/bank-pending")
def finance_v7_hq_bank_pending():
    return redirect("/finance/bank_pending")



@finance_v7_bp.route("/finance/v7/HQ/records")
def finance_v7_hq_records():
    return _render_v7_module_records('hq',False)


@finance_v7_bp.route("/finance/v7/HQ/search")
def finance_v7_hq_search():
    return _render_v7_module_records('hq',True)


def render_v7_placeholder(title, description):
    return render_template_string(
        """
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta
                name="viewport"
                content="width=device-width, initial-scale=1"
            >

            <title>{{ title }}</title>

            <link
                rel="stylesheet"
                href="{{ url_for('static', filename='css/toolbox.css') }}"
            >

            <style>
                body {
                    margin: 0;
                    background: #f4f7fb;
                    color: #172033;
                }

                .page {
                    max-width: 820px;
                    margin: 0 auto;
                    padding: 24px 16px;
                }

                .card {
                    background: white;
                    border: 1px solid #e1e7f0;
                    border-radius: 22px;
                    padding: 25px;
                    box-shadow:
                        0 12px 32px rgba(34, 52, 88, 0.08);
                }

                h1 {
                    margin-top: 0;
                }

                p {
                    color: #6c7485;
                    font-size: 17px;
                    line-height: 1.6;
                }

                .notice {
                    margin-top: 20px;
                    padding: 16px;
                    background: #eef4ff;
                    border-radius: 14px;
                    color: #2450a7;
                    font-weight: 700;
                }

                .back {
                    display: inline-block;
                    margin-top: 22px;
                    text-decoration: none;
                    background: #2459d8;
                    color: white;
                    padding: 12px 18px;
                    border-radius: 12px;
                    font-weight: 700;
                }
            </style>
        </head>

        <body>

        <div class="page">
            <div class="card">

                <h1>{{ title }}</h1>

                <p>{{ description }}</p>

                <div class="notice">
                    V7 工作台正在设计中，目前没有连接旧页面。
                </div>

                <a
                    class="back"
                    href="{{ url_for('finance_v7.finance_v7_che_home') }}"
                >
                    ← 返回 CHE 财政
                </a>

            </div>
        </div>

        </body>
        </html>
        """,
        title=title,
        description=description,
    )


# =========================================================
# 旧网址兼容：避免浏览器书签或旧按钮出现 404
# =========================================================

@finance_v7_bp.route("/finance/v7/CHE/income")
def finance_v7_legacy_che_income():
    return redirect(
        url_for("finance_v7.finance_v7_che_monthly_fee")
    )


@finance_v7_bp.route("/finance/v7/monthly-fee")
def finance_v7_legacy_monthly_fee():
    return redirect(
        url_for("finance_v7.finance_v7_che_monthly_fee")
    )


@finance_v7_bp.route("/finance/v7/bank")
def finance_v7_legacy_bank():
    return redirect(
        url_for("finance_v7.finance_v7_bank")
    )


@finance_v7_bp.route("/finance/v7/cash")
def finance_v7_legacy_cash():
    return redirect(
        url_for("finance_v7.finance_v7_cash")
    )


@finance_v7_bp.route("/finance/v7/expense")
def finance_v7_legacy_expense():
    return redirect(
        url_for("finance_v7.finance_v7_expense")
    )


@finance_v7_bp.route("/finance/v7/month-close")
def finance_v7_legacy_month_close():
    return redirect(
        url_for("finance_v7.finance_v7_month_close")
    )


@finance_v7_bp.route("/finance/v7/CHE/bank")
@finance_v7_bp.route("/finance/v7/CHE/treasury")
def finance_v7_bank():
    today = date.today()
    start, next_start = _v7_month_bounds(today)
    balance = _v7_balance_summary("CHE", today)
    pending_count, pending_total = _v7_pending_summary("CHE")
    movements = _v7_all("""
        select record_date, movement_type, amount, remarks, transfer_ref, created_at
        from finance_cash_movements
        where branch='CHE' and account_type='bank'
        order by record_date desc, id desc limit 5
    """)
    recent=[]
    for r in movements:
        mt=r.get('movement_type') or ''
        direction='out' if mt=='cash_out' else ('neutral' if mt=='adjustment' else 'in')
        recent.append({'time':r.get('created_at').strftime('%H:%M') if r.get('created_at') else str(r.get('record_date') or ''),
          'type':mt,'title':r.get('remarks') or mt,'reference':r.get('transfer_ref') or '-',
          'amount':_v7_money(r.get('amount')),'direction':direction})
    month_move=_v7_one("""select
      coalesce(sum(amount) filter(where movement_type='bank_in'),0) bank_in,
      coalesce(sum(amount) filter(where movement_type='cash_out'),0) withdrawal,
      coalesce(sum(amount) filter(where movement_type='adjustment'),0) adjustment
      from finance_cash_movements where branch='CHE' and account_type='bank' and record_date >= %s and record_date < %s""",(start,next_start))
    data={'today':today.strftime('%Y-%m-%d'),'month':today.strftime('%Y-%m'),'bank_balance':balance['bank_balance'],
      'today_income':_v7_money(_v7_one("""select coalesce(sum(amount),0) total from finance_cash_movements where branch='CHE' and account_type='bank' and record_date=current_date and movement_type in ('bank_in','opening')""").get('total')),
      'today_withdrawal':_v7_money(_v7_one("""select coalesce(sum(amount),0) total from finance_cash_movements where branch='CHE' and account_type='bank' and record_date=current_date and movement_type='cash_out'""").get('total')),
      'pending_count':pending_count,'pending_total':pending_total,'month_bank_in':_v7_money(month_move.get('bank_in')),
      'month_withdrawal':_v7_money(month_move.get('withdrawal')),'month_income':_v7_money(month_move.get('bank_in')),
      'month_adjustment':_v7_money(month_move.get('adjustment')),'recent_transactions':recent}

    return render_template_string(
        """
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>CHE 银行专区</title>

            <link
                rel="stylesheet"
                href="{{ url_for('static', filename='css/toolbox.css') }}"
            >

            <style>
                body {
                    margin: 0;
                    background: #f4f7fb;
                    color: #172033;
                }

                .bank-page {
                    width: min(1140px, calc(100% - 32px));
                    margin: 0 auto;
                    padding: 24px 0 48px;
                }

                .bank-topbar {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 18px;
                    margin-bottom: 20px;
                }

                .bank-heading {
                    display: flex;
                    align-items: center;
                    gap: 14px;
                }

                .bank-heading-icon {
                    width: 58px;
                    height: 58px;
                    display: grid;
                    place-items: center;
                    flex: 0 0 auto;
                    border-radius: 18px;
                    background: #e7efff;
                    font-size: 30px;
                }

                .bank-heading h1 {
                    margin: 0;
                    font-size: 30px;
                }

                .bank-heading p {
                    margin: 6px 0 0;
                    color: #6e788a;
                    font-size: 15px;
                }

                .bank-back {
                    padding: 11px 16px;
                    border: 1px solid #d9e0eb;
                    border-radius: 13px;
                    background: white;
                    color: #263149;
                    text-decoration: none;
                    font-weight: 900;
                    white-space: nowrap;
                }

                .summary-grid {
                    display: grid;
                    grid-template-columns: repeat(4, minmax(0, 1fr));
                    gap: 13px;
                    margin-bottom: 20px;
                }

                .summary-card {
                    background: white;
                    border: 1px solid #e0e6ef;
                    border-radius: 17px;
                    padding: 16px 17px;
                    box-shadow: 0 8px 22px rgba(32, 49, 80, 0.05);
                }

                .summary-label {
                    color: #758093;
                    font-size: 14px;
                    margin-bottom: 7px;
                }

                .summary-value {
                    font-size: 21px;
                    font-weight: 900;
                }

                .summary-note {
                    margin-top: 5px;
                    color: #929aaa;
                    font-size: 12px;
                }

                .bank-layout {
                    display: grid;
                    grid-template-columns: minmax(0, 1.35fr) minmax(300px, 0.75fr);
                    gap: 19px;
                    align-items: start;
                }

                .section-card {
                    background: white;
                    border: 1px solid #dfe6ef;
                    border-radius: 22px;
                    padding: 21px;
                    box-shadow: 0 10px 28px rgba(32, 49, 80, 0.06);
                }

                .section-card + .section-card {
                    margin-top: 18px;
                }

                .section-title {
                    margin: 0;
                    font-size: 21px;
                }

                .section-help {
                    margin-top: 5px;
                    color: #7a8496;
                    font-size: 14px;
                }

                .action-grid {
                    display: grid;
                    grid-template-columns: repeat(2, minmax(0, 1fr));
                    gap: 12px;
                    margin-top: 17px;
                }

                .action-card {
                    display: flex;
                    align-items: center;
                    gap: 12px;
                    min-height: 82px;
                    padding: 15px;
                    border: 1px solid #dde5ef;
                    border-radius: 16px;
                    background: #fafbfd;
                    color: #2e3c56;
                    text-decoration: none;
                    transition: transform .15s ease, border-color .15s ease;
                }

                .action-card:hover {
                    transform: translateY(-2px);
                    border-color: #bfd0ec;
                }

                .action-primary {
                    background: #edf4ff;
                    border-color: #cbdcf7;
                }

                .action-warning {
                    background: #fff9ed;
                    border-color: #efd79f;
                }

                .action-icon {
                    width: 46px;
                    height: 46px;
                    display: grid;
                    place-items: center;
                    flex: 0 0 auto;
                    border-radius: 14px;
                    background: white;
                    font-size: 23px;
                }

                .action-title {
                    font-size: 16px;
                    font-weight: 900;
                }

                .action-desc {
                    margin-top: 4px;
                    color: #748096;
                    font-size: 12px;
                    line-height: 1.4;
                }

                .transaction-list {
                    display: grid;
                    gap: 10px;
                    margin-top: 17px;
                }

                .transaction-row {
                    display: grid;
                    grid-template-columns: 72px minmax(0, 1fr) auto;
                    gap: 13px;
                    align-items: center;
                    padding: 14px;
                    border: 1px solid #e5e9f0;
                    border-radius: 15px;
                    background: #fafbfd;
                }

                .transaction-time {
                    color: #768196;
                    font-size: 13px;
                    font-weight: 800;
                }

                .transaction-type {
                    font-size: 12px;
                    color: #66758c;
                    margin-bottom: 4px;
                }

                .transaction-title {
                    font-size: 15px;
                    font-weight: 900;
                }

                .transaction-ref {
                    margin-top: 4px;
                    color: #7d8799;
                    font-size: 12px;
                }

                .transaction-amount {
                    text-align: right;
                    font-size: 16px;
                    font-weight: 900;
                    white-space: nowrap;
                }

                .amount-in { color: #1f7a4d; }
                .amount-out { color: #b04b3f; }
                .amount-neutral { color: #687286; }

                .view-all {
                    display: block;
                    margin-top: 13px;
                    padding: 12px;
                    border-radius: 13px;
                    background: #eef3f9;
                    color: #33425d;
                    text-align: center;
                    text-decoration: none;
                    font-weight: 900;
                }

                .month-total {
                    margin-top: 14px;
                    font-size: 32px;
                    font-weight: 900;
                }

                .month-label {
                    margin-top: 6px;
                    color: #737e91;
                    font-size: 14px;
                }

                .breakdown-list {
                    display: grid;
                    gap: 10px;
                    margin-top: 18px;
                }

                .breakdown-row {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 15px;
                    padding: 13px 14px;
                    border-radius: 14px;
                    background: #f5f7fa;
                }

                .breakdown-label {
                    color: #667187;
                    font-size: 14px;
                }

                .breakdown-value {
                    font-weight: 900;
                }

                .pending-box {
                    margin-top: 17px;
                    padding: 17px;
                    border-radius: 16px;
                    background: #fff9ed;
                    border: 1px solid #efd79f;
                }

                .pending-number {
                    font-size: 28px;
                    font-weight: 900;
                }

                .pending-note {
                    margin-top: 5px;
                    color: #796947;
                    font-size: 13px;
                    line-height: 1.5;
                }

                @media (max-width: 900px) {
                    .summary-grid {
                        grid-template-columns: repeat(2, minmax(0, 1fr));
                    }

                    .bank-layout {
                        grid-template-columns: 1fr;
                    }
                }

                @media (max-width: 580px) {
                    .bank-page {
                        width: min(100% - 20px, 1140px);
                        padding-top: 14px;
                    }

                    .bank-topbar {
                        align-items: flex-start;
                    }

                    .bank-heading-icon {
                        width: 49px;
                        height: 49px;
                        border-radius: 15px;
                        font-size: 25px;
                    }

                    .bank-heading h1 {
                        font-size: 24px;
                    }

                    .bank-heading p,
                    .bank-back {
                        font-size: 13px;
                    }

                    .bank-back {
                        padding: 9px 11px;
                    }

                    .summary-grid,
                    .action-grid {
                        gap: 9px;
                    }

                    .summary-card,
                    .section-card {
                        padding: 14px;
                    }

                    .transaction-row {
                        grid-template-columns: 1fr auto;
                    }

                    .transaction-time {
                        grid-column: 1 / -1;
                    }
                }
            </style>
        </head>

        <body>
        <main class="bank-page">

            <header class="bank-topbar">
                <div class="bank-heading">
                    <div class="bank-heading-icon">🏦</div>
                    <div>
                        <h1>CHE 银行专区</h1>
                        <p>CDM、银行待确认、银行对账、Cash In Hand 与数据维修 · {{ data.today }}</p>
                    </div>
                </div>

                <a
                    class="bank-back"
                    href="{{ url_for('finance_v7.finance_v7_che_home') }}"
                >
                    ← 返回 CHE 财政
                </a>
            </header>

            <section class="summary-grid">
                <div class="summary-card">
                    <div class="summary-label">当前银行余额</div>
                    <div class="summary-value">RM {{ "%.2f"|format(data.bank_balance) }}</div>
                    <div class="summary-note">CHE 银行账户</div>
                </div>

                <div class="summary-card">
                    <div class="summary-label">今日银行收入</div>
                    <div class="summary-value">RM {{ "%.2f"|format(data.today_income) }}</div>
                    <div class="summary-note">已确认记录</div>
                </div>

                <div class="summary-card">
                    <div class="summary-label">今日提款</div>
                    <div class="summary-value">RM {{ "%.2f"|format(data.today_withdrawal) }}</div>
                    <div class="summary-note">转入 Petty Cash</div>
                </div>

                <div class="summary-card">
                    <div class="summary-label">银行待确认</div>
                    <div class="summary-value">{{ data.pending_count }} 笔</div>
                    <div class="summary-note">RM {{ "%.2f"|format(data.pending_total) }}</div>
                </div>
            </section>

            <div class="bank-layout">
                <div>
                    <section class="section-card">
                        <h2 class="section-title">快速操作</h2>
                        <div class="section-help">银行工作集中在这里完成</div>

                        <div class="action-grid">
                            <a class="action-card action-primary" href="{{ url_for('finance_v7.finance_v7_cdm_management') }}">
                                <div class="action-icon">➕</div>
                                <div>
                                    <div class="action-title">新增 Bank In</div>
                                    <div class="action-desc">记录存款、CDM 或转账入账</div>
                                </div>
                            </a>

                            <a class="action-card action-warning" href="/finance/bank_pending">
                                <div class="action-icon">🟡</div>
                                <div>
                                    <div class="action-title">银行待确认</div>
                                    <div class="action-desc">{{ data.pending_count }} 笔待处理</div>
                                </div>
                            </a>

                            <a class="action-card" href="/finance/bank_pending">
                                <div class="action-icon">📜</div>
                                <div>
                                    <div class="action-title">银行流水</div>
                                    <div class="action-desc">查看全部银行交易</div>
                                </div>
                            </a>

                            <a class="action-card" href="{{ url_for('finance_v7.finance_v7_bank_search') }}">
                                <div class="action-icon">🔍</div>
                                <div>
                                    <div class="action-title">搜索记录</div>
                                    <div class="action-desc">按编号、姓名、金额或备注查询</div>
                                </div>
                            </a>
                        </div>
                    </section>

                    <section class="section-card">
                        <h2 class="section-title">最近银行交易</h2>
                        <div class="section-help">最新 5 笔银行活动</div>

                        <div class="transaction-list">
                            {% for tx in data.recent_transactions %}
                                <div class="transaction-row">
                                    <div class="transaction-time">{{ tx.time }}</div>

                                    <div>
                                        <div class="transaction-type">{{ tx.type }}</div>
                                        <div class="transaction-title">{{ tx.title }}</div>
                                        <div class="transaction-ref">{{ tx.reference }}</div>
                                    </div>

                                    <div class="transaction-amount amount-{{ tx.direction }}">
                                        {% if tx.direction == 'in' %}+{% elif tx.direction == 'out' %}-{% endif %}
                                        RM {{ "%.2f"|format(tx.amount) }}
                                    </div>
                                </div>
                            {% endfor %}
                        </div>

                        <a class="view-all" href="{{ url_for('finance_v7.finance_v7_bank_search') }}">查看银行记录</a>
                    </section>
                </div>

                <aside>
                    <section class="section-card">
                        <h2 class="section-title">本月银行统计</h2>
                        <div class="section-help">{{ data.month }}</div>

                        <div class="month-total">RM {{ "%.2f"|format(data.month_income) }}</div>
                        <div class="month-label">本月银行总收入</div>

                        <div class="breakdown-list">
                            <div class="breakdown-row">
                                <div class="breakdown-label">Bank In</div>
                                <div class="breakdown-value">RM {{ "%.2f"|format(data.month_bank_in) }}</div>
                            </div>

                            <div class="breakdown-row">
                                <div class="breakdown-label">现金提款</div>
                                <div class="breakdown-value">RM {{ "%.2f"|format(data.month_withdrawal) }}</div>
                            </div>

                            <div class="breakdown-row">
                                <div class="breakdown-label">余额调整</div>
                                <div class="breakdown-value">RM {{ "%.2f"|format(data.month_adjustment) }}</div>
                            </div>
                        </div>
                    </section>

                    <section class="section-card">
                        <h2 class="section-title">待确认提醒</h2>
                        <div class="section-help">需要财政处理的银行记录</div>

                        <div class="pending-box">
                            <div class="pending-number">{{ data.pending_count }} 笔</div>
                            <div class="pending-note">
                                合计 RM {{ "%.2f"|format(data.pending_total) }}，
                                请核对付款人、类别及收条编号。
                            </div>
                        </div>

                        <a class="view-all" href="/finance/bank_pending">立即处理待确认</a>
                    </section>
                </aside>
            </div>

        </main>
        </body>
        </html>
        """,
        data=data,
    )


@finance_v7_bp.route("/finance/v7/CHE/cash")
def finance_v7_cash():
    """CHE 现金 Dashboard：Cash In Hand 与 Petty Cash 严格分开。"""
    today = date.today()
    start, next_start = _v7_month_bounds(today)
    balance = _v7_balance_summary("CHE", today)

    month_cash_fee = _v7_money(_v7_one(f"""
        select coalesce(sum(r.amount),0) total
        from finance_records r
        where r.record_type='income' and r.category='月费'
          and coalesce(r.status,'confirmed')<>'cancelled'
          and lower(trim(coalesce(r.payment_method,''))) in ('现金','cash')
          and r.record_date >= %s and r.record_date < %s
          and {_v7_che_filter('r')}
    """, (start, next_start)).get("total"))

    month_cdm = _v7_money(_v7_one("""
        select coalesce(sum(amount),0) total
        from finance_bank_deposits
        where upper(coalesce(branch,''))='CHE'
          and fund_account='观音堂日常户口'
          and deposit_date >= %s and deposit_date < %s
    """, (start, next_start)).get("total"))

    month_petty_in = _v7_money(_v7_one("""
        select coalesce(sum(amount),0) total
        from finance_cash_movements
        where branch='CHE' and account_type='cash'
          and movement_type='cash_in'
          and record_date >= %s and record_date < %s
    """, (start, next_start)).get("total"))

    month_expense = _v7_money(_v7_one("""
        select coalesce(sum(amount),0) total
        from finance_records
        where record_type='expense'
          and coalesce(status,'confirmed')<>'cancelled'
          and record_date >= %s and record_date < %s
          and coalesce(fund_account,'观音堂日常户口')
              in ('观音堂日常户口','CHE 日常户口')
    """, (start, next_start)).get("total"))

    data = {
        "today": today.strftime("%Y-%m-%d"),
        "month": today.strftime("%Y-%m"),
        "cash_in_hand": balance["cash_in_hand"],
        "petty_cash": balance["petty_cash"],
        "total_cash": balance["total_cash"],
        "month_cash_fee": month_cash_fee,
        "month_cdm": month_cdm,
        "month_petty_in": month_petty_in,
        "month_expense": month_expense,
    }

    return render_template_string(r"""
<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>CHE 现金 Dashboard</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>
body{margin:0;background:#f4f7fb;color:#182235}.page{width:min(1080px,calc(100% - 24px));margin:auto;padding:22px 0 48px}.top{display:flex;justify-content:space-between;align-items:flex-start;gap:14px}.top h1{margin:0;font-size:29px}.sub{margin-top:6px;color:#6f7889}.back{background:#fff;border:1px solid #dce3ed;border-radius:13px;padding:10px 14px;text-decoration:none;color:#344158;font-weight:900}.notice{margin-top:16px;background:#eef4ff;border:1px solid #d4e1f5;border-radius:17px;padding:15px 17px;line-height:1.65;color:#41536f}.grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:16px}.cash-card{background:#fff;border:1px solid #dfe6ef;border-radius:23px;padding:22px;box-shadow:0 10px 28px rgba(32,49,80,.06)}.cash-card.cih{border-top:6px solid #2c8c64}.cash-card.petty{border-top:6px solid #d68b2b}.label{font-size:20px;font-weight:950}.help{margin-top:5px;color:#737d8e;line-height:1.55}.amount{font-size:40px;font-weight:950;margin-top:18px}.flow{margin-top:18px;display:grid;gap:9px}.row{display:flex;justify-content:space-between;gap:12px;background:#f6f8fb;border-radius:13px;padding:12px 13px}.row span{color:#697489}.row strong{white-space:nowrap}.rule{margin-top:15px;border-radius:14px;padding:13px;font-weight:850;line-height:1.55}.cih .rule{background:#eaf8f1;color:#276a4d}.petty .rule{background:#fff5e6;color:#80551d}.actions{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:18px}.action{background:#fff;border:1px solid #dfe6ef;border-radius:17px;padding:17px;text-decoration:none;color:#29364d;font-weight:950}.action small{display:block;margin-top:5px;color:#778194;font-weight:700;line-height:1.45}.total{margin-top:16px;background:#172f4f;color:#fff;border-radius:20px;padding:20px;display:flex;justify-content:space-between;align-items:center;gap:15px}.total strong{font-size:29px}@media(max-width:720px){.grid,.actions{grid-template-columns:1fr}.top h1{font-size:24px}.amount{font-size:34px}.total{align-items:flex-start;flex-direction:column}}
</style></head><body><main class="page"><header class="top"><div><h1>💵 CHE 现金 Dashboard</h1><div class="sub">{{ data.today }} · 两种现金独立管理，不互相混用</div></div><a class="back" href="{{ url_for('finance_v7.finance_v7_che_home') }}">← CHE 财政</a></header>
<div class="notice"><strong>资金规则：</strong>Cash In Hand 只放收到的现金月费，完成 CDM 后转入银行；Petty Cash 来自银行提款，只用于支付 PV 与日常开支。</div>
<section class="grid"><article class="cash-card cih"><div class="label">💵 Cash In Hand</div><div class="help">CHE 现金月费 · 等待 CDM，不可支付日常支出</div><div class="amount">RM {{ '%.2f'|format(data.cash_in_hand) }}</div><div class="flow"><div class="row"><span>本月现金月费</span><strong>+ RM {{ '%.2f'|format(data.month_cash_fee) }}</strong></div><div class="row"><span>本月正式 CDM</span><strong>- RM {{ '%.2f'|format(data.month_cdm) }}</strong></div></div><div class="rule">现金月费 → Cash In Hand → CDM → Bank</div></article>
<article class="cash-card petty"><div class="label">💰 Petty Cash</div><div class="help">日常备用金 · 银行提款补充，用于所有现金 PV</div><div class="amount">RM {{ '%.2f'|format(data.petty_cash) }}</div><div class="flow"><div class="row"><span>本月银行提款补充</span><strong>+ RM {{ '%.2f'|format(data.month_petty_in) }}</strong></div><div class="row"><span>本月现金支出</span><strong>- RM {{ '%.2f'|format(data.month_expense) }}</strong></div></div><div class="rule">Bank 提款 → Petty Cash → PV／日常支出</div></article></section>
<div class="total"><span>两种现金合计（仅供总览，不代表可互相使用）</span><strong>RM {{ '%.2f'|format(data.total_cash) }}</strong></div>
<nav class="actions"><a class="action" href="{{ url_for('finance_v7.finance_v7_cdm_management') }}">🏦 管理 CDM<small>把 Cash In Hand 存入银行</small></a><a class="action" href="{{ url_for('finance_v7.finance_v7_cash_transfer') }}">➕ 银行提款到 Petty Cash<small>一次录入，同时减少 Bank、增加 Petty Cash</small></a><a class="action" href="{{ url_for('finance_v7.finance_v7_expense_new') }}">📤 开立现金 PV<small>支出只从 Petty Cash 扣除</small></a></nav>
</main></body></html>""", data=data)


@finance_v7_bp.route("/finance/v7/CHE/cash/withdrawal")
def finance_v7_cash_withdrawal():
    return redirect(url_for("finance_v7.finance_v7_cash_transfer"))



@finance_v7_bp.route("/finance/v7/CHE/cash/activity")
def finance_v7_cash_activity():
    return _render_v7_module_records('cash',False)


@finance_v7_bp.route("/finance/v7/CHE/cash/search")
def finance_v7_cash_search():
    return _render_v7_module_records('cash',True)


@finance_v7_bp.route("/finance/v7/CHE/monthly-fee")
def finance_v7_che_monthly_fee():
    data = _v7_monthly_data("CHE")
    if data["bank_pending_count"]:
        data["today_status"] = f"今日已收 {data['today_people']} 位佛友，还有 {data['bank_pending_count']} 笔银行月费待确认。"
        data["status_icon"] = "⏳"
        data["status_class"] = "status-pending"
    else:
        data["today_status"] = "今天月费工作已完成。"
        data["status_icon"] = "✅"
        data["status_class"] = "status-done"

    return render_template_string(
        """
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>CHE 月费工作台</title>
            <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
            <style>
                body{margin:0;background:#f5f7fb;color:#172033}.page{width:min(1080px,calc(100% - 30px));margin:0 auto;padding:24px 0 48px}
                .topbar{display:flex;align-items:center;justify-content:space-between;gap:16px;margin-bottom:18px}.heading{display:flex;align-items:center;gap:14px}
                .heading-icon{width:58px;height:58px;display:grid;place-items:center;border-radius:18px;background:#fff0f4;font-size:30px}.heading h1{margin:0;font-size:30px}.heading p{margin:5px 0 0;color:#758094;font-size:15px}
                .back{padding:11px 16px;border:1px solid #d9e0eb;border-radius:13px;background:white;color:#263149;text-decoration:none;font-weight:900;white-space:nowrap}
                .hero{background:linear-gradient(135deg,#fff5f7,#fff);border:1px solid #f0dce2;border-radius:24px;padding:24px;box-shadow:0 10px 28px rgba(32,49,80,.06);margin-bottom:16px}
                .hero-label{color:#7f6670;font-size:15px;font-weight:800}.hero-total{margin-top:7px;font-size:43px;font-weight:950;letter-spacing:-1px}.hero-meta{margin-top:7px;color:#6f7889;font-size:15px}
                .status-box{margin-top:18px;display:flex;align-items:center;gap:12px;padding:14px 16px;border-radius:15px;font-weight:850;line-height:1.5}.status-pending{background:#fff8e8;border:1px solid #f1d99f;color:#715c28}.status-done{background:#edf9f1;border:1px solid #cde9d6;color:#346544}
                .stats{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px;margin-bottom:16px}.stat{background:white;border:1px solid #e0e6ef;border-radius:18px;padding:18px;box-shadow:0 8px 22px rgba(32,49,80,.05)}.stat-label{color:#778194;font-size:14px}.stat-value{margin-top:7px;font-size:24px;font-weight:950}
                .actions{
                    display:grid;
                    grid-template-columns:360px;
                    gap:12px;
                    margin-bottom:18px;
                }

                .actions-single{
                    grid-template-columns:minmax(0,360px);
                }
                .action{
                    min-height:116px;
                    padding:17px 14px;

                    display:flex;
                    flex-direction:column;
                    align-items:flex-start;
                    justify-content:space-between;

                    border:1px solid #dfe5ee;
                    border-radius:18px;

                    background:#ffffff;
                    color:#263149;
                    text-decoration:none;

                    box-shadow:0 8px 22px rgba(32,49,80,.05);

                    transition:
                        transform .18s ease,
                        box-shadow .18s ease,
                        border-color .18s ease;
                }

                .action:hover{
                    transform:translateY(-2px);
                    border-color:#c7d2e2;
                    box-shadow:0 12px 28px rgba(32,49,80,.10);
                }

                .action-icon{
                    font-size:28px;
                }

                .action-title{
                    font-size:17px;
                    font-weight:950;
                }

                .action-note{
                    margin-top:5px;
                    color:#778194;
                    font-size:12px;
                    line-height:1.45;
                }
                .content-grid{display:grid;grid-template-columns:minmax(0,1fr) 300px;gap:17px;align-items:start}.card{background:white;border:1px solid #e0e6ef;border-radius:21px;padding:20px;box-shadow:0 9px 25px rgba(32,49,80,.05)}.card h2{margin:0;font-size:21px}.card-help{margin:5px 0 0;color:#7c8698;font-size:14px}
                .records{display:grid;gap:10px;margin-top:16px}.record{display:grid;grid-template-columns:minmax(0,1fr) auto;gap:14px;padding:14px;border:1px solid #e6eaf0;border-radius:15px;background:#fafbfd}.record-name{font-weight:950;font-size:16px}.record-id{margin-top:3px;color:#5f6f86;font-weight:800}.record-meta{margin-top:6px;color:#7d8798;font-size:13px;line-height:1.5}.record-amount{text-align:right;font-size:18px;font-weight:950;white-space:nowrap}.record-method{margin-top:5px;color:#7d8798;font-size:12px;text-align:right}
                .side-value{margin-top:15px;font-size:32px;font-weight:950}.side-label{margin-top:4px;color:#7c8698;font-size:14px}.pending-summary{margin-top:16px;padding:15px;border-radius:15px;background:#fff9ec;border:1px solid #efdca8}.pending-summary strong{display:block;font-size:19px;margin-top:4px}
                @media(max-width:820px){

                .actions{
                    grid-template-columns:1fr;
                }

                .actions-single{
                    grid-template-columns:1fr;
                }

                .content-grid{
                    grid-template-columns:1fr;
                }
            }
                @media(max-width:560px){.page{width:min(100% - 20px,1080px);padding-top:14px}.topbar{align-items:flex-start}.heading-icon{width:50px;height:50px;font-size:26px}.heading h1{font-size:24px}.heading p{font-size:13px}.back{padding:9px 11px;font-size:13px}.hero{padding:20px}.hero-total{font-size:36px}.stats{gap:10px}.stat{padding:15px}.stat-value{font-size:20px}.actions{gap:10px}.action{min-height:108px}.record{grid-template-columns:1fr}.record-amount,.record-method{text-align:left}}
            </style>
        </head>
        <body><main class="page">
            <div class="topbar">
                <div class="heading"><div class="heading-icon">❤️</div><div><h1>CHE 月费</h1><p>只处理 CHE 月费，不混入其它财政记录</p></div></div>
                <a class="back" href="{{ url_for('finance_v7.finance_v7_che_home') }}">← CHE 财政</a>
            </div>

            <section class="hero">
                <div class="hero-label">{{ data.month }} 本月已收月费</div>
                <div class="hero-total">RM {{ "%.2f"|format(data.month_total) }}</div>
                <div class="hero-meta">共 {{ data.month_people }} 位佛友</div>
                <div class="status-box {{ data.status_class }}"><span style="font-size:22px">{{ data.status_icon }}</span><span>{{ data.today_status }}</span></div>
            </section>

            <section class="stats">
                <div class="stat"><div class="stat-label">👥 今日收到</div><div class="stat-value">{{ data.today_people }} 位</div></div>
                <div class="stat"><div class="stat-label">💰 今日金额</div><div class="stat-value">RM {{ "%.2f"|format(data.today_total) }}</div></div>
            </section>

            <section class="actions actions-single">

                <a class="action"
                href="{{ url_for(
                    'finance_v7.finance_v7_monthly_fee_records'
                ) }}">

                    <div class="action-icon">🔎</div>

                    <div>
                        <div class="action-title">
                            查找月费记录
                        </div>

                        <div class="action-note">
                            按会员编号、姓名、收条或日期查询 CHE 月费
                        </div>
                    </div>

                </a>

            </section>

            <div class="content-grid">
                <section class="card"><h2>最近月费</h2><p class="card-help">最近 5 笔 CHE 月费</p><div class="records">
                    {% for record in data.recent_records %}
                    <div class="record"><div><div class="record-name">{{ record.name }}</div><div class="record-id">{{ record.member_id }}</div><div class="record-meta">供养月份：{{ record.months }}<br>收条：{{ record.receipt_no }} · {{ record.record_date }}</div></div><div><div class="record-amount">RM {{ "%.2f"|format(record.amount) }}</div><div class="record-method">{{ record.method }}</div></div></div>
                    {% else %}<div style="padding:22px;text-align:center;color:#818a9a">暂时没有月费记录</div>{% endfor %}
                </div></section>

                <aside class="card"><h2>今日摘要</h2><p class="card-help">{{ data.today }}</p><div class="side-value">{{ data.today_people }} 位</div><div class="side-label">今日已收月费佛友</div><div class="side-value" style="font-size:26px">RM {{ "%.2f"|format(data.today_total) }}</div><div class="side-label">今日月费金额</div><div class="pending-summary">银行待确认<strong>{{ data.bank_pending_count }} 笔</strong>RM {{ "%.2f"|format(data.bank_pending_total) }}</div></aside>
            </div>
        </main></body></html>
        """,
        data=data,
    )


@finance_v7_bp.route("/finance/v7/CHE/monthly-fee/records")
def finance_v7_monthly_fee_records():
    return _render_v7_monthly_records_page(search_mode=False)


@finance_v7_bp.route("/finance/v7/CHE/monthly-fee/search")
def finance_v7_monthly_fee_search():
    query = request.query_string.decode("utf-8")
    target = url_for("finance_v7.finance_v7_monthly_fee_records")
    return redirect(f"{target}?{query}" if query else target)


def _render_v7_monthly_records_page(search_mode=False):
    """CHE 月费记录与查找合并；默认本月，并可指定任意年份／月份。"""
    today = date.today()
    keyword = request.args.get("q", "").strip()
    method = request.args.get("method", "").strip()
    year_text = request.args.get("year", str(today.year)).strip()
    month_text = request.args.get("month", str(today.month)).strip()

    try:
        selected_year = int(year_text) if year_text else None
    except ValueError:
        selected_year = today.year
    try:
        selected_month = int(month_text) if month_text else None
        if selected_month is not None and not 1 <= selected_month <= 12:
            raise ValueError
    except ValueError:
        selected_month = today.month

    clauses = [
        "r.record_type='income'",
        "r.category='月费'",
        "coalesce(r.status,'confirmed') <> 'cancelled'",
        _v7_che_filter("r"),
    ]
    params = []

    if selected_year is not None:
        clauses.append("extract(year from r.record_date)=%s")
        params.append(selected_year)
    if selected_month is not None:
        clauses.append("extract(month from r.record_date)=%s")
        params.append(selected_month)

    if method == "cash":
        clauses.append("lower(coalesce(r.payment_method,'')) in ('现金','cash')")
    elif method == "bank":
        clauses.append("lower(coalesce(r.payment_method,'')) not in ('现金','cash')")

    if keyword:
        like = f"%{keyword}%"
        clauses.append("(coalesce(r.member_id,'') ilike %s or coalesce(r.name,'') ilike %s or coalesce(r.receipt_no,'') ilike %s or coalesce(r.phone,'') ilike %s)")
        params.extend([like, like, like, like])

    rows = _v7_all(f"""
        select r.id,r.record_date,r.receipt_no,r.member_id,r.name,r.phone,
               r.month_from,r.month_to,r.amount,r.payment_method,r.remarks
        from finance_records r
        where {' and '.join(clauses)}
        order by r.record_date desc,r.id desc
        limit 500
    """, tuple(params))

    items = []
    for row in rows:
        mf = _v7_month_label(row.get("month_from"))
        mt = _v7_month_label(row.get("month_to"))
        months_label = f"{mf} ～ {mt}" if mf and mt and mf != mt else (mf or mt or "-")
        items.append({
            **row,
            "record_date": _v7_date_label(row.get("record_date")),
            "months": months_label,
            "amount": _v7_money(row.get("amount")),
        })

    total = sum(item["amount"] for item in items)
    years = list(range(today.year + 1, today.year - 8, -1))
    months = [(m, f"{m}月") for m in range(1, 13)]

    return render_template_string(r"""
<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>CHE 月费记录</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>
body{margin:0;background:#f5f7fb;color:#172033}.page{width:min(1040px,calc(100% - 24px));margin:auto;padding:22px 0 48px}.top{display:flex;justify-content:space-between;gap:14px;align-items:flex-start}.top h1{margin:0;font-size:28px}.sub{margin-top:6px;color:#748096}.back{background:#fff;border:1px solid #dce3ed;border-radius:13px;padding:10px 14px;text-decoration:none;color:#263149;font-weight:900;white-space:nowrap}.search{margin-top:17px;background:#fff;border:1px solid #e0e6ef;border-radius:20px;padding:16px;box-shadow:0 8px 24px rgba(32,49,80,.05)}.search-grid{display:grid;grid-template-columns:minmax(220px,1fr) 145px 135px 160px auto auto;gap:10px}.search input,.search select{min-height:48px;border:1px solid #d4dce7;border-radius:13px;padding:0 13px;font-size:15px;background:#fff}.search button,.clear{min-height:48px;border:0;border-radius:13px;background:#d94e72;color:#fff;font-weight:900;padding:0 20px;display:flex;align-items:center;justify-content:center;text-decoration:none}.clear{background:#f1f4f8;color:#4d596d}.summary{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin:14px 0}.sum{background:#fff;border:1px solid #e0e6ef;border-radius:17px;padding:16px}.sum-label{color:#778194;font-size:13px}.sum-value{font-size:24px;font-weight:950;margin-top:5px}.list{display:grid;gap:10px}.row{display:grid;grid-template-columns:120px minmax(0,1fr) auto;gap:14px;background:#fff;border:1px solid #e2e7ef;border-radius:17px;padding:15px}.receipt{font-weight:950}.date{color:#7c8698;font-size:13px;margin-top:5px}.name{font-size:17px;font-weight:950}.member{color:#5d6c82;font-weight:850;margin-top:3px}.meta{color:#7b8495;font-size:13px;line-height:1.55;margin-top:6px}.amount{text-align:right;font-size:19px;font-weight:950;white-space:nowrap}.method{text-align:right;color:#7b8495;font-size:12px;margin-top:5px}.empty{text-align:center;background:#fff;border:1px dashed #ccd5e2;border-radius:17px;padding:36px;color:#788397}@media(max-width:820px){.search-grid{grid-template-columns:1fr 1fr}.search-grid input{grid-column:1/-1}.search button,.clear{min-height:48px}.row{grid-template-columns:1fr auto}.row>div:nth-child(2){grid-column:1/-1;grid-row:2}.top h1{font-size:23px}}</style></head><body><main class="page">
<div class="top"><div><h1>❤️ CHE 月费记录</h1><div class="sub">月费记录与查找已经合并；这里只显示 CHE 月费。</div></div><a class="back" href="{{ url_for('finance_v7.finance_v7_che_monthly_fee') }}">← CHE 月费</a></div>
<form class="search" method="get"><div class="search-grid"><input name="q" value="{{ keyword }}" placeholder="会员编号、姓名、电话或收条编号"><select name="year"><option value="">全部年份</option>{% for y in years %}<option value="{{ y }}" {% if selected_year==y %}selected{% endif %}>{{ y }}年</option>{% endfor %}</select><select name="month"><option value="">全年</option>{% for value,label in months %}<option value="{{ value }}" {% if selected_month==value %}selected{% endif %}>{{ label }}</option>{% endfor %}</select><select name="method"><option value="">全部付款方式</option><option value="cash" {% if method=='cash' %}selected{% endif %}>现金</option><option value="bank" {% if method=='bank' %}selected{% endif %}>银行</option></select><button type="submit">查询</button><a class="clear" href="?year={{ current_year }}&month={{ current_month }}">本月</a></div></form>
<div class="summary"><div class="sum"><div class="sum-label">记录</div><div class="sum-value">{{ items|length }} 笔</div></div><div class="sum"><div class="sum-label">合计</div><div class="sum-value">RM {{ '%.2f'|format(total) }}</div></div></div>
{% if items %}<section class="list">{% for row in items %}<article class="row"><div><div class="receipt">{{ row.receipt_no or '-' }}</div><div class="date">{{ row.record_date }}</div></div><div><div class="name">{{ row.name or '-' }}</div><div class="member">{{ row.member_id or '-' }}</div><div class="meta">供养月份：{{ row.months }}<br>{{ row.phone or '无电话' }}{% if row.remarks %} · {{ row.remarks }}{% endif %}</div></div><div><div class="amount">RM {{ '%.2f'|format(row.amount) }}</div><div class="method">{{ row.payment_method or '-' }}</div></div></article>{% endfor %}</section>{% else %}<div class="empty">这个月份没有找到 CHE 月费记录</div>{% endif %}
</main></body></html>""",
        keyword=keyword, method=method, selected_year=selected_year,
        selected_month=selected_month, years=years, months=months,
        current_year=today.year, current_month=today.month,
        items=items, total=total,
    )

@finance_v7_bp.route("/finance/v7/CHE/expense")
def finance_v7_expense():
    data = _v7_expense_data()

    return render_template_string(
        """
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta
                name="viewport"
                content="width=device-width, initial-scale=1"
            >

            <title>CHE 支出</title>

            <link
                rel="stylesheet"
                href="{{ url_for('static', filename='css/toolbox.css') }}"
            >

            <style>
                body {
                    margin: 0;
                    background: #f4f7fb;
                    color: #172033;
                }

                .expense-page {
                    width: min(1040px, calc(100% - 28px));
                    margin: 0 auto;
                    padding: 22px 0 46px;
                }

                .topbar {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 16px;
                    margin-bottom: 18px;
                }

                .title-wrap {
                    display: flex;
                    align-items: center;
                    gap: 14px;
                }

                .title-icon {
                    width: 58px;
                    height: 58px;
                    display: grid;
                    place-items: center;
                    border-radius: 18px;
                    background: #ffe8e5;
                    font-size: 30px;
                }

                .title-wrap h1 {
                    margin: 0;
                    font-size: 30px;
                    line-height: 1.2;
                }

                .title-wrap p {
                    margin: 5px 0 0;
                    color: #6c7586;
                    font-size: 15px;
                }

                .back-btn {
                    text-decoration: none;
                    color: #29344b;
                    background: white;
                    border: 1px solid #dce3ed;
                    border-radius: 13px;
                    padding: 10px 15px;
                    font-weight: 800;
                }

                .summary-grid {
                    display: grid;
                    grid-template-columns: repeat(3, minmax(0, 1fr));
                    gap: 14px;
                    margin-bottom: 16px;
                }

                .summary-card {
                    background: white;
                    border: 1px solid #e0e6ef;
                    border-radius: 20px;
                    padding: 19px;
                    box-shadow: 0 8px 24px rgba(31, 46, 78, .05);
                }

                .summary-label {
                    color: #6c7586;
                    font-size: 15px;
                    font-weight: 700;
                }

                .summary-value {
                    margin-top: 8px;
                    font-size: 28px;
                    font-weight: 900;
                    letter-spacing: -.5px;
                }

                .summary-note {
                    margin-top: 5px;
                    color: #8a92a2;
                    font-size: 13px;
                }

                .action-panel,
                .recent-panel {
                    background: white;
                    border: 1px solid #e0e6ef;
                    border-radius: 22px;
                    padding: 20px;
                    box-shadow: 0 8px 24px rgba(31, 46, 78, .05);
                    margin-bottom: 16px;
                }

                .section-title {
                    margin: 0 0 14px;
                    font-size: 20px;
                }

                .primary-action {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 14px;
                    text-decoration: none;
                    color: white;
                    background: linear-gradient(135deg, #d75248, #b83c34);
                    border-radius: 18px;
                    padding: 19px 20px;
                    margin-bottom: 13px;
                    font-size: 20px;
                    font-weight: 900;
                }

                .primary-action small {
                    display: block;
                    margin-top: 4px;
                    color: rgba(255,255,255,.82);
                    font-size: 13px;
                    font-weight: 600;
                }

                .action-grid {
                    display: grid;
                    grid-template-columns: repeat(3, minmax(0, 1fr));
                    gap: 11px;
                }

                .action-btn {
                    display: flex;
                    align-items: center;
                    justify-content: space-between;
                    gap: 8px;
                    min-height: 66px;
                    padding: 13px 15px;
                    text-decoration: none;
                    color: #27334a;
                    background: #f7f9fc;
                    border: 1px solid #e1e7f0;
                    border-radius: 15px;
                    font-weight: 850;
                }

                .expense-list {
                    display: grid;
                    gap: 10px;
                }

                .expense-row {
                    display: grid;
                    grid-template-columns: 125px minmax(0, 1fr) auto;
                    align-items: center;
                    gap: 16px;
                    padding: 15px 16px;
                    background: #fafbfd;
                    border: 1px solid #e7ebf2;
                    border-radius: 16px;
                }

                .pv-no {
                    font-size: 16px;
                    font-weight: 900;
                }

                .expense-date {
                    margin-top: 4px;
                    color: #8991a0;
                    font-size: 13px;
                }

                .expense-category {
                    font-weight: 900;
                    margin-bottom: 5px;
                }

                .vendor-name {
                    color: #586377;
                    font-size: 14px;
                    line-height: 1.45;
                    overflow-wrap: anywhere;
                }

                .expense-amount {
                    font-size: 19px;
                    font-weight: 900;
                    white-space: nowrap;
                }

                .view-all {
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    margin-top: 13px;
                    min-height: 47px;
                    text-decoration: none;
                    color: #34415a;
                    background: #f5f7fb;
                    border: 1px solid #e0e6ef;
                    border-radius: 14px;
                    font-weight: 850;
                }

                @media (max-width: 720px) {
                    .expense-page {
                        width: min(100% - 20px, 1040px);
                        padding-top: 14px;
                    }

                    .topbar {
                        align-items: flex-start;
                    }

                    .title-icon {
                        width: 50px;
                        height: 50px;
                        border-radius: 15px;
                        font-size: 26px;
                    }

                    .title-wrap h1 {
                        font-size: 25px;
                    }

                    .title-wrap p {
                        display: none;
                    }

                    .back-btn {
                        padding: 9px 11px;
                        font-size: 14px;
                    }

                    .summary-grid {
                        grid-template-columns: repeat(2, minmax(0, 1fr));
                    }

                    .summary-card:last-child {
                        grid-column: 1 / -1;
                    }

                    .summary-value {
                        font-size: 24px;
                    }

                    .action-grid {
                        grid-template-columns: 1fr;
                    }

                    .expense-row {
                        grid-template-columns: 1fr auto;
                        gap: 10px;
                    }

                    .expense-row > div:nth-child(2) {
                        grid-column: 1 / -1;
                        grid-row: 2;
                    }

                    .expense-amount {
                        grid-column: 2;
                        grid-row: 1;
                    }
                }
            </style>
        </head>

        <body>
        <main class="expense-page">

            <header class="topbar">
                <div class="title-wrap">
                    <div class="title-icon">📤</div>
                    <div>
                        <h1>CHE 支出</h1>
                        <p>开立 PV、查看支出与分类统计</p>
                    </div>
                </div>

                <a
                    class="back-btn"
                    href="{{ url_for('finance_v7.finance_v7_che_home') }}"
                >
                    ← CHE 财政
                </a>
            </header>

            <section class="summary-grid">
                <div class="summary-card">
                    <div class="summary-label">今日支出</div>
                    <div class="summary-value">
                        RM {{ "%.2f"|format(data.today_total) }}
                    </div>
                    <div class="summary-note">今天已登记的现金支出</div>
                </div>

                <div class="summary-card">
                    <div class="summary-label">今日 PV</div>
                    <div class="summary-value">
                        {{ data.today_pv_count }} 张
                    </div>
                    <div class="summary-note">今天开立的付款凭单</div>
                </div>

                <div class="summary-card">
                    <div class="summary-label">本月支出</div>
                    <div class="summary-value">
                        RM {{ "%.2f"|format(data.month_total) }}
                    </div>
                    <div class="summary-note">{{ data.month_label }}</div>
                </div>
            </section>

            <section class="action-panel">
                <h2 class="section-title">今天要做什么？</h2>

                <a
                    class="primary-action"
                    href="{{ url_for('finance_v7.finance_v7_expense_new') }}"
                >
                    <span>
                        ➕ 开立付款 PV
                        <small>登记一笔新的 CHE 支出</small>
                    </span>
                    <span>›</span>
                </a>

                <div class="action-grid">
                    <a
                        class="action-btn"
                        href="{{ url_for('finance_v7.finance_v7_expense_history') }}"
                    >
                        <span>📖 支出记录</span>
                        <span>›</span>
                    </a>

                    <a
                        class="action-btn"
                        href="{{ url_for('finance_v7.finance_v7_expense_search') }}"
                    >
                        <span>🔍 搜索支出</span>
                        <span>›</span>
                    </a>

                    <a
                        class="action-btn"
                        href="{{ url_for('finance_v7.finance_v7_expense_categories') }}"
                    >
                        <span>📊 分类统计</span>
                        <span>›</span>
                    </a>
                    <a class="action-btn" href="{{ url_for('finance_v7.finance_v7_vendor_manage') }}"><span>🏢 付款对象管理</span><span>›</span></a>
                </div>
            </section>

            <section class="recent-panel">
                <h2 class="section-title">最近支出</h2>

                <div class="expense-list">
                    {% for row in data.recent_expenses %}
                    <div class="expense-row">
                        <div>
                            <div class="pv-no">{{ row.pv_no }}</div>
                            <div class="expense-date">{{ row.expense_date }}</div>
                        </div>

                        <div>
                            <div class="expense-category">
                                {{ row.category }}
                            </div>
                            <div class="vendor-name">
                                {{ row.vendor_name }}
                            </div>
                        </div>

                        <div class="expense-amount">
                            RM {{ "%.2f"|format(row.amount) }}
                        </div>
                    </div>
                    {% endfor %}
                </div>

                <a
                    class="view-all"
                    href="{{ url_for('finance_v7.finance_v7_expense_history') }}"
                >
                    查看全部支出记录
                </a>
            </section>

        </main>
        </body>
        </html>
        """,
        data=data,
    )


def _v7_ensure_vendor_system():
    try:
        db_query("""create table if not exists finance_vendors (
            id bigserial primary key, branch varchar(10) not null default 'CHE',
            company_name varchar(220) not null, contact_person varchar(160), phone varchar(80),
            email varchar(180), expense_category varchar(100) not null, remarks text,
            is_active boolean not null default true, sort_order integer not null default 100,
            created_at timestamp not null default now(), updated_at timestamp not null default now())""")
        db_query("create unique index if not exists uq_finance_vendors_name_cat on finance_vendors(branch, lower(company_name), expense_category)")
        db_query("alter table finance_records add column if not exists vendor_id bigint")
        db_query("alter table finance_records add column if not exists vendor_name varchar(220)")
        for company, category, order_no in [
            ('Pudu Ria Florist Trading Sdn Bhd','供花',10), ('Laohongka Sdn Bhd','供油',10),
            ('Tenaga Nasional Berhad','电费',10), ('Air Selangor','水费',10),
            ('Indah Water Konsortium Sdn Bhd','水费',20), ('TM Unifi','电话及网络费',10),
            ('CelcomDigi','电话及网络费',20), ('Maxis','电话及网络费',30)]:
            db_query("""insert into finance_vendors(branch,company_name,expense_category,sort_order)
                      values('CHE',%s,%s,%s) on conflict do nothing""", (company,category,order_no))
    except Exception as exc:
        print('Finance V7 vendor system warning:', exc)


def _v7_expense_categories_list():
    return ['供花','供果','供油','电费','水费','电话及网络费','维修保养','日常采购','装修工程','其它支出']


def _v7_vendor_rows(active_only=True):
    _v7_ensure_vendor_system()
    extra = " and is_active=true" if active_only else ""
    return _v7_all(f"""select id,company_name,contact_person,phone,email,expense_category,remarks,is_active,sort_order
        from finance_vendors where branch='CHE' {extra}
        order by expense_category,sort_order,company_name""")


@finance_v7_bp.route('/finance/v7/CHE/expense/new', methods=['GET','POST'])
def finance_v7_expense_new():
    _v7_ensure_vendor_system()
    categories = _v7_expense_categories_list()
    message, error = '', ''
    form = {
        'record_date': request.form.get('record_date') or date.today().isoformat(),
        'payment_voucher_no': request.form.get('payment_voucher_no','').strip().upper(),
        'category': request.form.get('category','').strip(),
        'vendor_id': request.form.get('vendor_id','').strip(),
        'amount': request.form.get('amount','').strip(),
        'remarks': request.form.get('remarks','').strip(),
    }
    if request.method == 'POST':
        action = request.form.get('action','save_expense')
        if action == 'add_vendor':
            company = request.form.get('new_company_name','').strip()
            cat = request.form.get('new_vendor_category','').strip() or form['category']
            if not company or not cat:
                error = '请填写公司名称和支出类别。'
            else:
                try:
                    db_query("""insert into finance_vendors(branch,company_name,contact_person,phone,email,expense_category,remarks)
                        values('CHE',%s,%s,%s,%s,%s,%s) on conflict do nothing""",(
                        company,request.form.get('new_contact_person') or None,request.form.get('new_phone') or None,
                        request.form.get('new_email') or None,cat,request.form.get('new_vendor_remarks') or None))
                    row=_v7_one("select id from finance_vendors where branch='CHE' and lower(company_name)=lower(%s) and expense_category=%s order by id desc limit 1",(company,cat))
                    form['category'],form['vendor_id']=cat,str(row.get('id') or '')
                    message=f'已添加并选择：{company}'
                except Exception as exc: error=str(exc)
        else:
            try:
                if not form['payment_voucher_no']: raise ValueError('请填写 PV 编号。')
                if not form['category']: raise ValueError('请选择支出类别。')
                if not form['vendor_id']: raise ValueError('请选择付款对象／公司名称。')
                amount=Decimal(form['amount'])
                if amount<=0: raise ValueError('金额必须大于 0。')
                vendor=_v7_one("select id,company_name from finance_vendors where id=%s and branch='CHE' and is_active=true",(int(form['vendor_id']),))
                if not vendor: raise ValueError('找不到这个付款对象。')
                if _v7_one("select id from finance_records where upper(coalesce(payment_voucher_no,''))=%s and coalesce(status,'confirmed')<>'cancelled' limit 1",(form['payment_voucher_no'],)):
                    raise ValueError('这个 PV 编号已经使用。')
                db_query("""insert into finance_records(record_type,fund_account,record_date,category,payment_voucher_no,name,
                    vendor_id,vendor_name,amount,payment_method,remarks,status)
                    values('expense','观音堂日常户口',%s,%s,%s,%s,%s,%s,%s,'现金',%s,'confirmed')""",(
                    form['record_date'],form['category'],form['payment_voucher_no'],vendor['company_name'],vendor['id'],vendor['company_name'],amount,form['remarks'] or None))
                return redirect(url_for('finance_v7.finance_v7_expense',saved='1'))
            except Exception as exc: error=str(exc)
    vendors=_v7_vendor_rows(True)
    return render_template_string(r"""<!doctype html><html lang='zh'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>开立付款 PV</title><link rel='stylesheet' href="{{ url_for('static',filename='css/toolbox.css') }}"><style>
body{margin:0;background:#f5f7fb;color:#172033}.page{max-width:880px;margin:auto;padding:22px 14px 50px}.top{display:flex;justify-content:space-between;gap:12px}.back{text-decoration:none;background:#fff;border:1px solid #dce3ed;border-radius:13px;padding:10px 14px;color:#29344b;font-weight:900}.card,.vendor-box{margin-top:16px;background:#fff;border:1px solid #e0e6ef;border-radius:20px;padding:20px}.grid{display:grid;grid-template-columns:1fr 1fr;gap:15px}.field{display:grid;gap:7px}.full{grid-column:1/-1}label{font-weight:900}.field input,.field select,.field textarea{box-sizing:border-box;width:100%;border:1px solid #ccd6e3;border-radius:13px;padding:12px;font-size:16px}.field textarea{min-height:90px}.cash{grid-column:1/-1;background:#eff8f2;border-radius:14px;padding:13px;color:#28643b;font-weight:850}.actions{display:flex;gap:10px;margin-top:18px}.btn{border:0;border-radius:13px;padding:13px 18px;font-weight:950;text-decoration:none;cursor:pointer}.primary{background:#d9544d;color:#fff}.soft{background:#eef2f7;color:#34415a}.vendor-box{display:none;background:#fff9ed}.vendor-box.open{display:block}.msg{padding:12px;border-radius:12px;margin-top:12px}.ok{background:#e9f8ee;color:#216b39}.err{background:#fff0ef;color:#a12f28}@media(max-width:650px){.grid{grid-template-columns:1fr}.full,.cash{grid-column:auto}.actions{flex-direction:column}}</style></head><body><main class='page'>
<div class='top'><div><h1>➕ 开立付款 PV</h1><div>所有支出固定为现金，并从 Petty Cash 扣除。</div></div><a class='back' href="{{ url_for('finance_v7.finance_v7_expense') }}">← CHE 支出</a></div>
{% if message %}<div class='msg ok'>{{ message }}</div>{% endif %}{% if error %}<div class='msg err'>{{ error }}</div>{% endif %}
<section class='card'><form method='post'><input type='hidden' name='action' value='save_expense'><div class='grid'>
<div class='field'><label>PV 日期 *</label><input type='date' name='record_date' value='{{ form.record_date }}' required></div><div class='field'><label>PV 编号 *</label><input name='payment_voucher_no' value='{{ form.payment_voucher_no }}' required></div>
<div class='field'><label>支出类别 *</label><select name='category' id='category' required><option value=''>请选择</option>{% for c in categories %}<option value='{{ c }}' {% if form.category==c %}selected{% endif %}>{{ c }}</option>{% endfor %}</select></div>
<div class='field'><label>付款对象／公司名称 *</label><select name='vendor_id' id='vendor' required><option value=''>请先选择类别</option>{% for v in vendors %}<option value='{{ v.id }}' data-category='{{ v.expense_category }}' {% if form.vendor_id==v.id|string %}selected{% endif %}>{{ v.company_name }}</option>{% endfor %}</select><button class='btn soft' type='button' onclick='toggleVendor()'>＋ 添加新付款对象</button></div>
<div class='field'><label>金额 RM *</label><input type='number' min='0.01' step='0.01' name='amount' value='{{ form.amount }}' required></div><div class='field'><label>付款方式</label><input value='现金（固定）' readonly></div>
<div class='field full'><label>备注</label><textarea name='remarks'>{{ form.remarks }}</textarea></div><div class='cash'>💵 保存后只计入 Petty Cash 支出；不会减少 Cash In Hand，也不会直接减少银行余额。</div></div><div class='actions'><button class='btn primary'>确认并保存 PV</button><a class='btn soft' href="{{ url_for('finance_v7.finance_v7_vendor_manage') }}">管理付款对象</a></div></form></section>
<section class='vendor-box' id='vendorBox'><h2>🏢 新增付款对象</h2><form method='post'><input type='hidden' name='action' value='add_vendor'><input type='hidden' name='record_date' value='{{ form.record_date }}'><input type='hidden' name='payment_voucher_no' value='{{ form.payment_voucher_no }}'><input type='hidden' name='amount' value='{{ form.amount }}'><input type='hidden' name='remarks' value='{{ form.remarks }}'><div class='grid'><div class='field full'><label>完整公司名称 *</label><input name='new_company_name' required></div><div class='field'><label>适用类别 *</label><select name='new_vendor_category'>{% for c in categories %}<option {% if form.category==c %}selected{% endif %}>{{ c }}</option>{% endfor %}</select></div><div class='field'><label>联络人</label><input name='new_contact_person'></div><div class='field'><label>电话</label><input name='new_phone'></div><div class='field'><label>Email</label><input name='new_email'></div><div class='field full'><label>备注</label><input name='new_vendor_remarks'></div></div><div class='actions'><button class='btn primary'>保存并选择</button><button class='btn soft' type='button' onclick='toggleVendor()'>取消</button></div></form></section>
</main><script>const c=document.getElementById('category'),v=document.getElementById('vendor');function filterV(){let n=0;[...v.options].forEach((o,i)=>{if(!i)return;let show=o.dataset.category===c.value;o.hidden=!show;o.disabled=!show;if(show)n++});if(v.selectedOptions[0]&&v.selectedOptions[0].disabled)v.value='';v.options[0].text=c.value?(n?'请选择付款对象／公司名称':'这个类别还没有付款对象，请新增'):'请先选择类别'}function toggleVendor(){document.getElementById('vendorBox').classList.toggle('open')}c.addEventListener('change',filterV);filterV();</script></body></html>""",categories=categories,vendors=vendors,form=form,message=message,error=error)


@finance_v7_bp.route('/finance/v7/CHE/expense/history')
def finance_v7_expense_history():
    return _render_v7_module_records('expense',True)


@finance_v7_bp.route('/finance/v7/CHE/expense/search')
def finance_v7_expense_search():
    return redirect(url_for('finance_v7.finance_v7_expense_history', **request.args))


@finance_v7_bp.route('/finance/v7/CHE/expense/categories')
def finance_v7_expense_categories():
    rows=_v7_all("""select coalesce(category,'其它支出') category,count(*) count,coalesce(sum(amount),0) total from finance_records where record_type='expense' and coalesce(status,'confirmed')<>'cancelled' and date_trunc('month',record_date)=date_trunc('month',current_date) group by coalesce(category,'其它支出') order by total desc""")
    return render_template_string("""<!doctype html><html lang='zh'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>支出分类</title></head><body><main style='max-width:800px;margin:auto;padding:24px'><a href='{{ url_for("finance_v7.finance_v7_expense") }}'>← 返回</a><h1>📊 本月支出分类</h1>{% for r in rows %}<div style='background:white;border:1px solid #ddd;padding:14px;margin:8px;border-radius:12px'><b>{{ r.category }}</b> · {{ r.count }} 笔 · RM {{ '%.2f'|format(r.total) }}</div>{% else %}<p>本月没有支出。</p>{% endfor %}</main></body></html>""",rows=rows)


@finance_v7_bp.route('/finance/v7/CHE/expense/vendors', methods=['GET','POST'])
def finance_v7_vendor_manage():
    _v7_ensure_vendor_system(); message=''; error=''
    if request.method=='POST':
        try:
            action=request.form.get('action')
            if action=='add':
                db_query("""insert into finance_vendors(branch,company_name,contact_person,phone,email,expense_category,remarks,sort_order) values('CHE',%s,%s,%s,%s,%s,%s,%s)""",(request.form.get('company_name').strip(),request.form.get('contact_person') or None,request.form.get('phone') or None,request.form.get('email') or None,request.form.get('expense_category'),request.form.get('remarks') or None,int(request.form.get('sort_order') or 100)));message='已新增付款对象。'
            elif action=='toggle': db_query("update finance_vendors set is_active=not is_active,updated_at=now() where id=%s",(int(request.form.get('vendor_id')),));message='状态已更新。'
            elif action=='edit': db_query("update finance_vendors set company_name=%s,expense_category=%s,phone=%s,contact_person=%s,sort_order=%s,updated_at=now() where id=%s",(request.form.get('company_name'),request.form.get('expense_category'),request.form.get('phone') or None,request.form.get('contact_person') or None,int(request.form.get('sort_order') or 100),int(request.form.get('vendor_id'))));message='资料已保存。'
        except Exception as exc:error=str(exc)
    rows=_v7_vendor_rows(False);categories=_v7_expense_categories_list()
    return render_template_string("""<!doctype html><html lang='zh'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>付款对象管理</title><style>body{background:#f5f7fb;color:#172033}.page{max-width:950px;margin:auto;padding:22px}.card,.row{background:#fff;border:1px solid #dfe5ee;border-radius:16px;padding:16px;margin:10px 0}.grid,.row form{display:grid;grid-template-columns:2fr 1.2fr 1fr 1fr auto auto;gap:8px}.grid input,.grid select,.row input,.row select{padding:10px;border:1px solid #ccd6e3;border-radius:10px;min-width:0}.btn{border:0;border-radius:10px;padding:10px;font-weight:bold}.save{background:#d9544d;color:white}.inactive{opacity:.55}@media(max-width:700px){.grid,.row form{grid-template-columns:1fr}}</style></head><body><main class='page'><a href='{{ url_for("finance_v7.finance_v7_expense") }}'>← CHE 支出</a><h1>🏢 付款对象管理</h1>{% if message %}<p>{{ message }}</p>{% endif %}{% if error %}<p>{{ error }}</p>{% endif %}<section class='card'><h2>新增公司</h2><form method='post'><input type='hidden' name='action' value='add'><div class='grid'><input name='company_name' placeholder='完整公司名称' required><select name='expense_category'>{% for c in categories %}<option>{{ c }}</option>{% endfor %}</select><input name='phone' placeholder='电话'><input name='contact_person' placeholder='联络人'><input name='sort_order' type='number' value='100'><button class='btn save'>新增</button><input type='hidden' name='email'><input type='hidden' name='remarks'></div></form></section>{% for v in rows %}<section class='row {% if not v.is_active %}inactive{% endif %}'><form method='post'><input type='hidden' name='vendor_id' value='{{ v.id }}'><input name='company_name' value='{{ v.company_name }}'><select name='expense_category'>{% for c in categories %}<option {% if c==v.expense_category %}selected{% endif %}>{{ c }}</option>{% endfor %}</select><input name='phone' value='{{ v.phone or "" }}' placeholder='电话'><input name='contact_person' value='{{ v.contact_person or "" }}' placeholder='联络人'><input name='sort_order' type='number' value='{{ v.sort_order }}'><button class='btn' name='action' value='edit'>保存</button><button class='btn save' name='action' value='toggle'>{% if v.is_active %}停用{% else %}启用{% endif %}</button></form></section>{% endfor %}</main></body></html>""",rows=rows,categories=categories,message=message,error=error)


# =========================================================
# CHE Member Care CRM
# =========================================================

def _v7_ensure_member_care_table():
    try:
        db_query("""
            create table if not exists finance_member_care_logs (
                id bigserial primary key,
                branch varchar(10) not null default 'CHE',
                member_id varchar(50) not null,
                contact_date date not null default current_date,
                action_type varchar(40) not null,
                note text,
                follow_up_date date,
                status varchar(30) not null default 'completed',
                created_by varchar(120),
                created_at timestamp not null default now()
            )
        """)
        db_query("""
            create index if not exists idx_finance_member_care_logs_member
            on finance_member_care_logs(member_id, contact_date desc, id desc)
        """)
    except Exception as exc:
        print("Finance V7 member care table warning:", exc)


def _v7_phone_digits(value):
    digits = ''.join(ch for ch in str(value or '') if ch.isdigit())
    if digits.startswith('0'):
        digits = '6' + digits
    return digits


def _v7_care_rows(keyword='', mode='all'):
    keyword = str(keyword or '').strip()
    where = ["upper(coalesce(m.member_id,'')) like 'CHE-%%'"]
    params = []
    if keyword:
        like = f"%{keyword}%"
        where.append("(m.member_id ilike %s or m.name ilike %s or coalesce(m.english_name,'') ilike %s or coalesce(m.phone,'') ilike %s)")
        params.extend([like, like, like, like])

    rows = _v7_all(f"""
        select
            m.member_id,
            m.name,
            m.english_name,
            m.phone,
            coalesce(m.member_status, m.status, '') as member_status,
            m.remark,
            max(p.end_month) as paid_until,
            max(p.payment_date) as last_payment_date,
            coalesce(sum(p.amount),0) as total_paid,
            count(p.id) as payment_count,
            max(l.contact_date) as last_contact_date,
            max(l.follow_up_date) filter(where l.status='pending') as next_follow_up
        from members m
        left join member_payments p
          on p.member_id=m.member_id
         and coalesce(p.status,'active')='active'
        left join finance_member_care_logs l
          on l.member_id=m.member_id
         and l.branch='CHE'
        where {' and '.join(where)}
        group by m.member_id,m.name,m.english_name,m.phone,m.member_status,m.status,m.remark
        order by max(p.end_month) nulls first,m.member_id
    """, tuple(params))

    today_month = date.today().replace(day=1)
    out = []
    paused_values = {'暂停','paused','停供','停止','永久停止','往生','已往生'}

    for r in rows:
        paid = r.get('paid_until')
        paid_date = paid.date() if hasattr(paid, 'date') else paid
        if isinstance(paid_date, str):
            try:
                paid_date = datetime.strptime(paid_date[:10], '%Y-%m-%d').date()
            except Exception:
                paid_date = None

        if not paid_date:
            months_late = 999
        else:
            months_late = (today_month.year-paid_date.year)*12 + (today_month.month-paid_date.month)

        member_status = str(r.get('member_status') or '').strip()
        is_paused = member_status in paused_values
        phone = str(r.get('phone') or '').strip()
        phone_digits = _v7_phone_digits(phone)

        if is_paused:
            care_status, cls, icon, priority = '暂停供养', 'paused', '⏸️', 2
        elif months_late <= 0:
            care_status, cls, icon, priority = '正常供养', 'ok', '✅', 0
        elif months_late == 1:
            care_status, cls, icon, priority = '迟供 1 个月', 'one', '🟡', 5
        elif months_late < 999:
            care_status, cls, icon, priority = f'迟供 {months_late} 个月', 'two', '🟠', min(5, 3 + months_late)
        else:
            care_status, cls, icon, priority = '尚无供养记录', 'two', '🟠', 3

        if not phone:
            priority = max(1, priority - 1)

        item = {
            **r,
            'paid_until': _v7_month_label(paid) or '-',
            'last_payment_date': _v7_date_label(r.get('last_payment_date')) or '-',
            'last_contact_date': _v7_date_label(r.get('last_contact_date')) or '尚未联系',
            'next_follow_up': _v7_date_label(r.get('next_follow_up')) or '',
            'months_late': months_late,
            'care_status': care_status,
            'status_class': cls,
            'status_icon': icon,
            'priority': priority,
            'stars': '★' * priority + '☆' * (5-priority),
            'total_paid': _v7_money(r.get('total_paid')),
            'payment_count': int(r.get('payment_count') or 0),
            'phone_digits': phone_digits,
            'has_phone': bool(phone),
            'is_paused': is_paused,
        }

        if mode == 'today':
            due_follow_up = item['next_follow_up'] and item['next_follow_up'] <= date.today().isoformat()
            if not due_follow_up and not (1 <= months_late <= 3 and not is_paused):
                continue
        out.append(item)

    out.sort(key=lambda x: (-x['priority'], x['months_late'] if x['months_late'] < 999 else 999, x['member_id']))
    return out


@finance_v7_bp.route('/finance/v7/CHE/member-care')
def finance_v7_member_care():
    _v7_ensure_member_care_table()
    rows = _v7_care_rows()
    today_rows = _v7_care_rows(mode='today')[:8]
    start, end = _v7_month_bounds()
    recovered = _v7_one("""
        select count(distinct member_id) as total
        from member_payments
        where payment_date >= %s and payment_date < %s
          and upper(coalesce(member_id,'')) like 'CHE-%%'
          and coalesce(status,'active')='active'
    """, (start, end))
    contacted = _v7_one("""
        select count(distinct member_id) as total
        from finance_member_care_logs
        where branch='CHE' and contact_date >= %s and contact_date < %s
    """, (start, end))
    data = {
        'active_members': len(rows),
        'suggested_today': len(_v7_care_rows(mode='today')),
        'late_one_month': sum(1 for r in rows if r['months_late']==1 and not r['is_paused']),
        'late_two_plus': sum(1 for r in rows if r['months_late']>=2 and r['months_late']<999 and not r['is_paused']),
        'paused': sum(1 for r in rows if r['is_paused']),
        'recovered_this_month': int(recovered.get('total') or 0),
        'contacted_this_month': int(contacted.get('total') or 0),
        'missing_phone': sum(1 for r in rows if not r['has_phone']),
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'today_rows': today_rows,
    }
    return render_template_string(r'''<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>CHE 月费关怀中心</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>
body{margin:0;background:linear-gradient(180deg,#f2ecff 0,#f7f8fc 360px,#f7f8fc 100%);color:#202638}.page{width:min(1120px,calc(100% - 24px));margin:auto;padding:24px 0 50px}.top{display:flex;justify-content:space-between;gap:14px;align-items:center;margin-bottom:16px}.title{display:flex;gap:14px;align-items:center}.icon{width:58px;height:58px;border-radius:18px;display:grid;place-items:center;background:#e5d8ff;font-size:30px}.title h1{margin:0;font-size:30px}.sub{color:#70697c;margin-top:5px}.back{background:#fff;border:1px solid #ddd5ea;border-radius:13px;padding:11px 15px;text-decoration:none;color:#443456;font-weight:900}.hero{background:linear-gradient(135deg,#6840bc,#8d66d8);color:#fff;border-radius:26px;padding:25px;box-shadow:0 15px 36px rgba(84,50,145,.2)}.hero-grid{display:grid;grid-template-columns:1.3fr 1fr;gap:18px}.hero-num{font-size:54px;font-weight:950;line-height:1}.hero-label{opacity:.88;font-weight:850}.hero-box{background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.22);border-radius:18px;padding:17px}.hero-box strong{display:block;font-size:28px;margin-top:7px}.metrics{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));gap:12px;margin:16px 0}.metric{background:#fff;border:1px solid #e4dfeb;border-radius:18px;padding:15px;box-shadow:0 6px 18px rgba(45,35,62,.05)}.metric span{display:block;color:#756e80;font-size:13px;font-weight:850}.metric strong{display:block;font-size:26px;margin-top:7px}.actions{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-bottom:16px}.action{background:#fff;border:1px solid #dfd6ed;border-radius:18px;min-height:105px;padding:15px;text-align:center;text-decoration:none;color:#50327d;font-weight:950;display:grid;place-items:center}.action.primary{background:#6f42c1;color:#fff;border-color:#6f42c1}.action b{font-size:27px;display:block}.panel{background:#fff;border:1px solid #e4dfeb;border-radius:22px;padding:20px;box-shadow:0 7px 21px rgba(45,35,62,.05)}.panel-head{display:flex;justify-content:space-between;gap:12px;align-items:center;margin-bottom:14px}.panel-head h2{margin:0;font-size:22px}.updated{color:#7a7285;font-size:13px}.list{display:grid;gap:11px}.person{display:grid;grid-template-columns:1fr auto;gap:15px;border:1px solid #ebe6f1;border-radius:17px;padding:16px}.member-id{color:#6f6680;font-size:13px;font-weight:900}.name{font-size:20px;font-weight:950;margin:4px 0}.stars{color:#e0a400;font-weight:950;letter-spacing:1px}.pill{display:inline-flex;padding:6px 10px;border-radius:999px;font-size:13px;font-weight:900;margin:8px 0}.one{background:#fff1b8;color:#765600}.two{background:#ffe1cc;color:#934000}.paused{background:#e9edf2;color:#4b5563}.ok{background:#e3f7eb;color:#23734d}.meta{color:#666d79;font-size:14px;line-height:1.65}.person-actions{display:flex;gap:8px;align-items:center;flex-wrap:wrap;justify-content:flex-end}.btn{border:0;border-radius:11px;padding:9px 12px;text-decoration:none;font-weight:900;cursor:pointer;font-size:14px}.btn.view{background:#efe7fb;color:#58348c}.btn.wa{background:#e2f7e9;color:#1f7a48}.empty{text-align:center;padding:32px;color:#786f84}@media(max-width:920px){.metrics{grid-template-columns:repeat(3,1fr)}.actions{grid-template-columns:repeat(2,1fr)}}@media(max-width:620px){.top{align-items:flex-start}.title h1{font-size:24px}.hero-grid{grid-template-columns:1fr}.metrics{grid-template-columns:repeat(2,1fr)}.person{grid-template-columns:1fr}.person-actions{justify-content:flex-start}.actions{grid-template-columns:1fr 1fr}}
</style></head><body><main class="page"><header class="top"><div class="title"><div class="icon">❤️</div><div><h1>CHE 月费关怀中心</h1><div class="sub">先关心佛友近况，再处理供养安排。</div></div></div><a class="back" href="{{ url_for('finance_v7.finance_v7_che_home') }}">← CHE 财政</a></header><section class="hero"><div class="hero-grid"><div><div class="hero-label">今日建议处理</div><div class="hero-num">{{ data.suggested_today }}<small style="font-size:20px"> 位</small></div><div style="margin-top:9px;opacity:.9">系统按照迟供月份、暂停状态和跟进日期排序。</div></div><div class="hero-box">本月已联系<strong>{{ data.contacted_this_month }} 位</strong></div></div></section><section class="metrics"><div class="metric"><span>迟供 1 个月</span><strong>{{ data.late_one_month }}</strong></div><div class="metric"><span>迟供 2 个月以上</span><strong>{{ data.late_two_plus }}</strong></div><div class="metric"><span>暂停供养</span><strong>{{ data.paused }}</strong></div><div class="metric"><span>本月有缴付</span><strong>{{ data.recovered_this_month }}</strong></div><div class="metric"><span>没有电话</span><strong>{{ data.missing_phone }}</strong></div><div class="metric"><span>有效会员</span><strong>{{ data.active_members }}</strong></div></section><nav class="actions"><a class="action primary" href="{{ url_for('finance_v7.finance_v7_member_care_today') }}"><div><b>📞</b>开始处理今天名单</div></a><a class="action" href="{{ url_for('finance_v7.finance_v7_member_care_list') }}"><div><b>📋</b>全部关怀名单</div></a><a class="action" href="{{ url_for('finance_v7.finance_v7_member_care_search') }}"><div><b>🔍</b>查询佛友</div></a><a class="action" href="{{ url_for('finance_v7.finance_v7_member_care_stats') }}"><div><b>📊</b>月费分析</div></a></nav><section class="panel"><div class="panel-head"><h2>📋 今日优先名单</h2><span class="updated">更新：{{ data.updated_at }}</span></div>{% if data.today_rows %}<div class="list">{% for r in data.today_rows %}<article class="person"><div><div class="member-id">{{ r.member_id }}</div><div class="name">{{ r.name or '-' }}</div><div class="stars">{{ r.stars }}</div><div class="pill {{ r.status_class }}">{{ r.status_icon }} {{ r.care_status }}</div><div class="meta">最后供养：<strong>{{ r.paid_until }}</strong> · 最近联系：{{ r.last_contact_date }}{% if r.next_follow_up %}<br>下次跟进：<strong>{{ r.next_follow_up }}</strong>{% endif %}</div></div><div class="person-actions">{% if r.phone_digits %}<a class="btn wa" href="https://wa.me/{{ r.phone_digits }}" target="_blank">WhatsApp</a>{% endif %}<a class="btn view" href="{{ url_for('finance_v7.finance_v7_member_care_member',member_id=r.member_id) }}">处理／查看</a></div></article>{% endfor %}</div>{% else %}<div class="empty">今天暂时没有需要优先处理的佛友。</div>{% endif %}</section></main></body></html>''', data=data)


@finance_v7_bp.route('/finance/v7/CHE/member-care/today')
def finance_v7_member_care_today():
    return _render_v7_member_care_list('today')


@finance_v7_bp.route('/finance/v7/CHE/member-care/list')
def finance_v7_member_care_list():
    return _render_v7_member_care_list('all')


@finance_v7_bp.route('/finance/v7/CHE/member-care/search')
def finance_v7_member_care_search():
    return _render_v7_member_care_list('search')


def _render_v7_member_care_list(mode):
    _v7_ensure_member_care_table()
    keyword=request.args.get('q','').strip(); status=request.args.get('status','').strip()
    rows=_v7_care_rows(keyword,mode)
    if status=='one': rows=[r for r in rows if r['months_late']==1 and not r['is_paused']]
    elif status=='two': rows=[r for r in rows if r['months_late']>=2 and not r['is_paused']]
    elif status=='paused': rows=[r for r in rows if r['is_paused']]
    elif status=='ok': rows=[r for r in rows if r['months_late']<=0 and not r['is_paused']]
    title={'today':'今日建议联络','search':'查询月费佛友'}.get(mode,'全部关怀名单')
    return render_template_string(r'''<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{{ title }}</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>body{margin:0;background:#f7f5fb;color:#202638}.page{width:min(1020px,calc(100% - 24px));margin:auto;padding:22px 0 48px}.top{display:flex;justify-content:space-between;gap:12px}.top h1{margin:0}.back{text-decoration:none;font-weight:900;color:#5a348e}.search{margin:17px 0;background:#fff;border:1px solid #e3deeb;border-radius:19px;padding:15px}.grid{display:grid;grid-template-columns:1fr 210px auto;gap:10px}.grid input,.grid select{min-height:48px;border:1px solid #d9d2e4;border-radius:13px;padding:0 13px;font-size:15px;background:#fff}.grid button{border:0;border-radius:13px;background:#6f42c1;color:#fff;font-weight:900;padding:0 22px}.count{margin:14px 2px;color:#6f687a;font-weight:850}.list{display:grid;gap:11px}.person{display:grid;grid-template-columns:1fr auto;gap:15px;background:#fff;border:1px solid #e5e0ec;border-radius:18px;padding:16px}.name{font-size:20px;font-weight:950}.id{color:#6c6280;font-weight:850}.stars{color:#dda400;font-weight:950;margin-top:6px}.meta{color:#707685;font-size:14px;line-height:1.65;margin-top:8px}.pill{display:inline-flex;padding:7px 10px;border-radius:999px;font-weight:900;font-size:13px;margin-top:8px}.ok{background:#e7f7ee;color:#24724c}.one{background:#fff4c8;color:#7a5a00}.two{background:#ffe5d2;color:#944000}.paused{background:#e9edf2;color:#4b5563}.view{align-self:center;text-decoration:none;background:#f0e8fb;color:#5a348e;border-radius:12px;padding:10px 13px;font-weight:900}.empty{text-align:center;background:#fff;border:1px dashed #d5cde0;border-radius:17px;padding:36px;color:#7d7589}@media(max-width:650px){.grid{grid-template-columns:1fr}.grid button{min-height:48px}.person{grid-template-columns:1fr}.view{text-align:center}}</style></head><body><main class="page"><div class="top"><div><h1>👥 {{ title }}</h1><div style="color:#777184;margin-top:6px">CHE 月费会员专用关怀名单</div></div><a class="back" href="{{ url_for('finance_v7.finance_v7_member_care') }}">← 月费关怀</a></div><form class="search"><div class="grid"><input name="q" value="{{ keyword }}" placeholder="编号、姓名、英文名或电话"><select name="status"><option value="">全部状态</option><option value="one" {% if status=='one' %}selected{% endif %}>迟供 1 个月</option><option value="two" {% if status=='two' %}selected{% endif %}>迟供 2 个月以上</option><option value="paused" {% if status=='paused' %}selected{% endif %}>暂停供养</option><option value="ok" {% if status=='ok' %}selected{% endif %}>正常供养</option></select><button>查找</button></div></form><div class="count">共 {{ rows|length }} 位佛友</div>{% if rows %}<section class="list">{% for r in rows %}<article class="person"><div><div class="id">{{ r.member_id }}</div><div class="name">{{ r.name or '-' }} {% if r.english_name %}<small>({{ r.english_name }})</small>{% endif %}</div><div class="stars">{{ r.stars }}</div><div class="pill {{ r.status_class }}">{{ r.status_icon }} {{ r.care_status }}</div><div class="meta">已供到：<strong>{{ r.paid_until }}</strong><br>电话：{{ r.phone or '未填写' }} · 最近联系：{{ r.last_contact_date }}</div></div><a class="view" href="{{ url_for('finance_v7.finance_v7_member_care_member',member_id=r.member_id) }}">处理／查看</a></article>{% endfor %}</section>{% else %}<div class="empty">没有找到符合条件的会员</div>{% endif %}</main></body></html>''',title=title,keyword=keyword,status=status,rows=rows)


@finance_v7_bp.route('/finance/v7/CHE/member-care/action/<member_id>', methods=['POST'])
def finance_v7_member_care_action(member_id):
    _v7_ensure_member_care_table()
    action_type = request.form.get('action_type','已联系').strip()
    note = request.form.get('note','').strip()
    follow_up_date = request.form.get('follow_up_date') or None
    status = 'pending' if follow_up_date else 'completed'
    created_by = str(session.get('finance_user') or session.get('username') or '财政负责人')
    db_query("""
        insert into finance_member_care_logs
        (branch,member_id,contact_date,action_type,note,follow_up_date,status,created_by)
        values('CHE',%s,current_date,%s,%s,%s,%s,%s)
    """, (member_id,action_type,note,follow_up_date,status,created_by))
    return redirect(url_for('finance_v7.finance_v7_member_care_member',member_id=member_id,saved='1'))


@finance_v7_bp.route('/finance/v7/CHE/member-care/stats')
def finance_v7_member_care_stats():
    _v7_ensure_member_care_table()
    rows=_v7_care_rows(); start,end=_v7_month_bounds()
    contact=_v7_one("select count(*) total,count(distinct member_id) people from finance_member_care_logs where branch='CHE' and contact_date >= %s and contact_date < %s",(start,end))
    data={'total':len(rows),'ok':sum(1 for r in rows if r['months_late']<=0 and not r['is_paused']),'one':sum(1 for r in rows if r['months_late']==1 and not r['is_paused']),'two':sum(1 for r in rows if r['months_late']>=2 and not r['is_paused']),'paused':sum(1 for r in rows if r['is_paused']),'contacts':int(contact.get('total') or 0),'people':int(contact.get('people') or 0)}
    return render_template_string(r'''<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>月费关怀分析</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>body{margin:0;background:#f7f5fb;color:#202638}.page{max-width:900px;margin:auto;padding:24px 14px}.top{display:flex;justify-content:space-between}.top h1{margin:0}.back{text-decoration:none;font-weight:900;color:#5a348e}.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-top:20px}.card{background:#fff;border:1px solid #e4deeb;border-radius:20px;padding:22px}.label{color:#736a80}.value{font-size:34px;font-weight:950;margin-top:8px}@media(max-width:620px){.grid{grid-template-columns:1fr 1fr}}</style></head><body><main class="page"><div class="top"><h1>📊 月费关怀分析</h1><a class="back" href="{{ url_for('finance_v7.finance_v7_member_care') }}">← 返回</a></div><section class="grid"><div class="card"><div class="label">有效会员</div><div class="value">{{ data.total }}</div></div><div class="card"><div class="label">正常供养</div><div class="value">{{ data.ok }}</div></div><div class="card"><div class="label">迟供 1 个月</div><div class="value">{{ data.one }}</div></div><div class="card"><div class="label">迟供 2 个月以上</div><div class="value">{{ data.two }}</div></div><div class="card"><div class="label">暂停供养</div><div class="value">{{ data.paused }}</div></div><div class="card"><div class="label">本月已联系佛友</div><div class="value">{{ data.people }}</div></div></section></main></body></html>''',data=data)


@finance_v7_bp.route('/finance/v7/CHE/member-care/member/<member_id>')
def finance_v7_member_care_member(member_id):
    _v7_ensure_member_care_table()
    people=_v7_care_rows(member_id,'search'); member=next((r for r in people if r.get('member_id')==member_id),None)
    payments=_v7_all("""select payment_date,receipt_no,start_month as month_from,end_month as month_to,amount,'' as payment_method from member_payments where member_id=%s and coalesce(status,'active')='active' order by payment_date desc,id desc limit 50""",(member_id,))
    logs=_v7_all("""select contact_date,action_type,note,follow_up_date,status,created_by,created_at from finance_member_care_logs where member_id=%s and branch='CHE' order by contact_date desc,id desc limit 50""",(member_id,))
    items=[]
    for p in payments:
        mf=_v7_month_label(p.get('month_from')); mt=_v7_month_label(p.get('month_to')); months=f"{mf} ～ {mt}" if mf and mt and mf!=mt else (mf or mt or '-')
        items.append({**p,'payment_date':_v7_date_label(p.get('payment_date')),'months':months,'amount':_v7_money(p.get('amount'))})
    for l in logs:
        l['contact_date']=_v7_date_label(l.get('contact_date')); l['follow_up_date']=_v7_date_label(l.get('follow_up_date'))
    return render_template_string(r'''<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{{ member_id }} 佛友资料</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>body{margin:0;background:#f7f5fb;color:#202638}.page{max-width:960px;margin:auto;padding:24px 14px}.back{text-decoration:none;font-weight:900;color:#5a348e}.profile,.panel{background:#fff;border:1px solid #e3deea;border-radius:20px;padding:20px;margin-top:15px}.profile-grid{display:grid;grid-template-columns:1fr auto;gap:15px}.name{font-size:26px;font-weight:950}.meta{line-height:1.75;color:#6d6678;margin-top:8px}.stars{color:#dda400;font-weight:950}.quick{display:flex;gap:8px;flex-wrap:wrap;align-content:flex-start}.quick a{padding:10px 13px;border-radius:11px;text-decoration:none;font-weight:900}.tel{background:#e8f1ff;color:#235ca5}.wa{background:#e2f7e9;color:#1f7a48}.form-grid{display:grid;grid-template-columns:180px 1fr 170px;gap:10px}.form-grid input,.form-grid select,.form-grid textarea{border:1px solid #d9d2e4;border-radius:12px;padding:11px;font-size:15px}.form-grid textarea{grid-column:1/-1;min-height:90px}.save{grid-column:1/-1;border:0;border-radius:12px;background:#6f42c1;color:#fff;min-height:48px;font-weight:950}.row{display:grid;grid-template-columns:120px 1fr auto;gap:12px;padding:13px 0;border-bottom:1px solid #eeeaf2}.row:last-child{border:0}.amount{font-weight:950;white-space:nowrap}.log-note{color:#6d6678;margin-top:5px;line-height:1.6}.saved{background:#e4f7eb;color:#23734d;border-radius:12px;padding:12px 14px;margin-top:14px;font-weight:900}@media(max-width:650px){.profile-grid,.form-grid{grid-template-columns:1fr}.form-grid textarea,.save{grid-column:auto}.row{grid-template-columns:1fr auto}.row>div:nth-child(2){grid-column:1/-1;grid-row:2}}</style></head><body><main class="page"><a class="back" href="{{ url_for('finance_v7.finance_v7_member_care_list') }}">← 返回关怀名单</a>{% if request.args.get('saved') %}<div class="saved">✅ 联系纪录已经保存</div>{% endif %}{% if member %}<section class="profile"><div class="profile-grid"><div><div class="name">{{ member.name or '-' }} <small>{{ member.english_name or '' }}</small></div><div class="stars">{{ member.stars }}</div><div class="meta">{{ member.member_id }}<br>电话：{{ member.phone or '未填写' }}<br>状态：{{ member.care_status }}<br>已供到：{{ member.paid_until }}<br>累计月费：RM {{ '%.2f'|format(member.total_paid) }}<br>最近联系：{{ member.last_contact_date }}</div></div><div class="quick">{% if member.phone %}<a class="tel" href="tel:{{ member.phone }}">📞 拨电话</a>{% endif %}{% if member.phone_digits %}<a class="wa" href="https://wa.me/{{ member.phone_digits }}" target="_blank">💬 WhatsApp</a>{% endif %}</div></div></section><section class="panel"><h2>📝 新增关怀纪录</h2><form method="post" action="{{ url_for('finance_v7.finance_v7_member_care_action',member_id=member.member_id) }}" class="form-grid"><select name="action_type"><option>电话联系</option><option>WhatsApp</option><option>当面关怀</option><option>已恢复供养</option><option>暂停提醒</option><option>其它</option></select><input type="date" name="follow_up_date" title="下次跟进日期"><input value="今日" disabled><textarea name="note" placeholder="简单记录佛友近况或负责人备注"></textarea><button class="save">保存关怀纪录</button></form></section>{% endif %}<section class="panel"><h2>📞 联系历史</h2>{% for l in logs %}<div class="row"><div>{{ l.contact_date }}</div><div><strong>{{ l.action_type }}</strong>{% if l.note %}<div class="log-note">{{ l.note }}</div>{% endif %}{% if l.follow_up_date %}<div class="log-note">下次跟进：{{ l.follow_up_date }}</div>{% endif %}</div><div>{{ l.created_by or '-' }}</div></div>{% else %}<p>还没有联系纪录</p>{% endfor %}</section><section class="panel"><h2>💳 月费历史</h2>{% for p in payments %}<div class="row"><div>{{ p.payment_date }}</div><div>{{ p.months }}<br><small>{{ p.receipt_no or '-' }}</small></div><div class="amount">RM {{ '%.2f'|format(p.amount) }}</div></div>{% else %}<p>没有月费记录</p>{% endfor %}</section></main></body></html>''',member_id=member_id,member=member,payments=items,logs=logs)


@finance_v7_bp.route("/finance/v7/CHE/month-close")
def finance_v7_month_close():
    ym = request.args.get("ym") or date.today().strftime("%Y-%m")
    try:
        start = datetime.strptime(ym + "-01", "%Y-%m-%d").date()
    except ValueError:
        start = date.today().replace(day=1)
        ym = start.strftime("%Y-%m")
    _, end = _v7_month_bounds(start)

    fee_row = _v7_one(f"""
        select count(*) as count, coalesce(sum(amount),0) as total,
               count(*) - count(distinct receipt_no) filter (where coalesce(receipt_no,'')<>'') as duplicate_receipts
        from finance_records r
        where r.record_type='income' and r.category='月费'
          and coalesce(r.status,'confirmed')<>'cancelled'
          and r.record_date >= %s and r.record_date < %s
          and {_v7_che_filter('r')}
    """, (start, end))
    donation_row = _v7_one("""
        select count(*) as count, coalesce(sum(amount),0) as total
        from finance_records
        where record_type='income'
          and category=any(%s)
          and coalesce(status,'confirmed')<>'cancelled'
          and record_date >= %s and record_date < %s
    """, (["财布施","观音村","膳食结缘","观音堂纯檀香布施","813交流会财布施","临时特别布施"], start, end))
    expense_row = _v7_one("""
        select count(*) as count, coalesce(sum(amount),0) as total,
               count(*) filter(where coalesce(payment_voucher_no,'')='') as missing_pv
        from finance_records
        where record_type='expense'
          and coalesce(status,'confirmed')<>'cancelled'
          and record_date >= %s and record_date < %s
    """, (start, end))
    pending_row = _v7_one("""
        select count(*) as count, coalesce(sum(amount),0) as total
        from bank_pending_records
        where coalesce(status,'pending')='pending'
    """)
    cash_balance = _v7_balance_summary('CHE', end - __import__('datetime').timedelta(days=1)).get('cash_in_hand', 0)

    pending_count = int(pending_row.get('count') or 0)
    missing_pv = int(expense_row.get('missing_pv') or 0)
    duplicate_receipts = max(0, int(fee_row.get('duplicate_receipts') or 0))
    problem_count = sum([pending_count > 0, missing_pv > 0, duplicate_receipts > 0])
    completed_checks = 4 - problem_count
    progress = int(completed_checks / 4 * 100)
    can_close = problem_count == 0

    history_rows = _v7_all("""
        select to_char(record_date,'YYYY-MM') period,
               coalesce(sum(amount) filter(where record_type='income' and category='月费'),0) monthly_fee,
               coalesce(sum(amount) filter(where record_type='expense'),0) expense
        from finance_records
        where coalesce(status,'confirmed')<>'cancelled'
        group by to_char(record_date,'YYYY-MM')
        order by period desc limit 3
    """)
    data = {
        "period": ym,
        "period_label": f"{start.year} 年 {start.month} 月",
        "status": "可以进入正式月结" if can_close else "尚有项目需要处理",
        "status_class": "ready" if can_close else "pending",
        "progress": progress,
        "score": max(0, 100 - problem_count * 20),
        "score_stars": "★★★★★" if can_close else ("★★★★☆" if problem_count == 1 else "★★★☆☆"),
        "health_level": "good" if can_close else "warning",
        "health_title": "自动检查已通过" if can_close else f"发现 {problem_count} 项需要处理",
        "health_message": "可以进入现有正式锁账流程。" if can_close else "请处理警告项目后再完成月结。",
        "can_close": can_close,
        "summary": {
            "monthly_fee": _v7_money(fee_row.get('total')),
            "donation": _v7_money(donation_row.get('total')),
            "expense": _v7_money(expense_row.get('total')),
            "cash_balance": _v7_money(cash_balance),
        },
        "checks": [
            {"label":"CHE 月费记录已读取","detail":f"本月 {int(fee_row.get('count') or 0)} 笔，合计 RM{_v7_money(fee_row.get('total')):,.2f}。","status":"done","icon":"✓"},
            {"label":"Bank Pending 已清空","detail":"没有待确认记录。" if pending_count==0 else f"仍有 {pending_count} 笔等待确认，合计 RM{_v7_money(pending_row.get('total')):,.2f}。","status":"done" if pending_count==0 else "warning","icon":"✓" if pending_count==0 else "!"},
            {"label":"本月支出与 PV 已检查","detail":f"本月 {int(expense_row.get('count') or 0)} 笔支出。" if missing_pv==0 else f"发现 {missing_pv} 笔支出没有 PV 编号。","status":"done" if missing_pv==0 else "warning","icon":"✓" if missing_pv==0 else "!"},
            {"label":"月费收条重复检查","detail":"没有发现重复收条。" if duplicate_receipts==0 else f"发现 {duplicate_receipts} 个重复收条编号。","status":"done" if duplicate_receipts==0 else "warning","icon":"✓" if duplicate_receipts==0 else "!"},
            {"label":"Cash In Hand 系统余额","detail":f"截至月底系统余额 RM{_v7_money(cash_balance):,.2f}；实际现金点算在正式锁账页确认。","status":"waiting","icon":"…"},
        ],
        "history": [
            {"period":r.get('period'),"status":"已有账务记录","completed_at":f"月费 RM{_v7_money(r.get('monthly_fee')):,.2f}","by":f"支出 RM{_v7_money(r.get('expense')):,.2f}"}
            for r in history_rows
        ],
    }

    return render_template_string(
        r"""
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>CHE 财政月结中心</title>
            <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
            <style>
                body{
                    margin:0;
                    color:#1f2937;
                    background:linear-gradient(180deg,#eef5ff 0,#f7f8fc 330px,#f7f8fc 100%);
                }
                .close-page{
                    width:min(1100px,calc(100% - 28px));
                    margin:0 auto;
                    padding:26px 0 48px;
                }
                .topbar{
                    display:flex;justify-content:space-between;align-items:center;
                    gap:16px;margin-bottom:18px;
                }
                .title-wrap{display:flex;align-items:center;gap:14px;}
                .title-icon{
                    width:58px;height:58px;border-radius:18px;display:grid;place-items:center;
                    font-size:30px;background:#dceaff;box-shadow:0 8px 24px rgba(39,93,168,.12);
                }
                .page-title{margin:0;font-size:30px;line-height:1.15;}
                .subtitle{margin:6px 0 0;color:#6b7280;font-size:16px;}
                .back-link{
                    text-decoration:none;color:#374151;background:#fff;border:1px solid #d9e2ef;
                    border-radius:14px;padding:11px 16px;font-weight:800;white-space:nowrap;
                }
                .hero{
                    border-radius:25px;padding:25px;color:#fff;
                    background:linear-gradient(135deg,#2059a8,#4384d7);
                    box-shadow:0 16px 40px rgba(32,89,168,.20);margin-bottom:18px;
                }
                .hero-grid{display:grid;grid-template-columns:1.25fr 1fr;gap:20px;align-items:center;}
                .period-label{font-size:17px;font-weight:800;opacity:.9;}
                .period-value{font-size:43px;font-weight:900;margin:8px 0 10px;}
                .status-box{
                    background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.24);
                    border-radius:19px;padding:18px;
                }
                .status-box strong{display:block;font-size:24px;margin:7px 0 12px;}
                .progress-track{height:11px;background:rgba(255,255,255,.22);border-radius:999px;overflow:hidden;}
                .progress-bar{height:100%;background:#fff;border-radius:999px;}
                .progress-text{margin-top:8px;font-weight:800;font-size:14px;opacity:.92;}
                .summary-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin-bottom:18px;}
                .summary-card{
                    background:#fff;border:1px solid #dfe6ef;border-radius:19px;padding:18px;
                    box-shadow:0 8px 22px rgba(42,55,75,.05);
                }
                .summary-label{color:#6b7280;font-size:15px;font-weight:800;}
                .summary-value{font-size:27px;font-weight:900;margin-top:8px;}
                .summary-card.fee{border-top:5px solid #e25578;}
                .summary-card.donation{border-top:5px solid #8b5cf6;}
                .summary-card.expense{border-top:5px solid #ef7b45;}
                .summary-card.cash{border-top:5px solid #22a06b;}
                .main-grid{display:grid;grid-template-columns:minmax(0,1.55fr) minmax(290px,.8fr);gap:18px;align-items:start;}
                .section-card{
                    background:#fff;border:1px solid #dfe6ef;border-radius:22px;padding:21px;
                    box-shadow:0 8px 24px rgba(42,55,75,.05);margin-bottom:18px;
                }
                .section-head{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:15px;}
                .section-head h2{margin:0;font-size:22px;}
                .section-head small{color:#7b8190;}
                .check-list{display:grid;gap:11px;}
                .check-row{
                    display:grid;grid-template-columns:42px minmax(0,1fr);gap:13px;align-items:center;
                    border:1px solid #e7ebf1;border-radius:16px;padding:14px 15px;
                }
                .check-icon{
                    width:38px;height:38px;border-radius:50%;display:grid;place-items:center;
                    font-size:20px;font-weight:900;
                }
                .check-row.done .check-icon{background:#dcf7e8;color:#178351;}
                .check-row.warning .check-icon{background:#fff1d6;color:#b86600;}
                .check-row.waiting .check-icon{background:#edf1f6;color:#64748b;}
                .check-title{font-weight:900;font-size:17px;}
                .check-detail{color:#667085;font-size:14px;line-height:1.55;margin-top:4px;}
                .health-card.warning{background:#fff9e8;border-color:#f3d78d;}
                .health-card.good{background:#edf9f2;border-color:#b8e4c9;}
                .health-title{font-size:21px;font-weight:900;margin:7px 0;}
                .health-message{color:#667085;line-height:1.6;}
                .score{
                    text-align:center;background:#f5f8fd;border:1px solid #dce5f2;border-radius:18px;
                    padding:18px;margin-top:15px;
                }
                .score-stars{font-size:25px;letter-spacing:3px;}
                .score-number{font-size:42px;font-weight:900;color:#2059a8;line-height:1.1;margin-top:7px;}
                .score-label{color:#6b7280;font-weight:800;margin-top:4px;}
                .action-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;}
                .action-card{
                    display:flex;flex-direction:column;justify-content:center;text-align:center;
                    min-height:112px;padding:16px;border-radius:18px;text-decoration:none;
                    font-size:17px;font-weight:900;line-height:1.4;border:1px solid #d9e3f0;
                    background:#fff;color:#23456e;
                }
                .action-card span{font-size:29px;margin-bottom:7px;}
                .action-primary{background:#2059a8;color:#fff;border-color:#2059a8;}
                .action-disabled{background:#edf1f6;color:#8a94a3;border-color:#e0e5eb;cursor:not-allowed;}
                .history-list{display:grid;gap:10px;}
                .history-row{
                    display:grid;grid-template-columns:85px minmax(0,1fr);gap:12px;
                    border-bottom:1px solid #edf0f4;padding:10px 0;
                }
                .history-row:last-child{border-bottom:0;}
                .history-period{font-weight:900;}
                .history-meta{font-size:14px;color:#687181;line-height:1.55;}
                .foot-note{
                    background:#eef5ff;border:1px solid #d6e5fb;border-radius:16px;
                    padding:15px 17px;color:#4d6078;line-height:1.7;
                }
                @media(max-width:860px){
                    .summary-grid{grid-template-columns:repeat(2,minmax(0,1fr));}
                    .main-grid{grid-template-columns:1fr;}
                    .hero-grid{grid-template-columns:1fr;}
                }
                @media(max-width:560px){
                    .close-page{width:min(100% - 18px,1100px);padding-top:14px;}
                    .topbar{align-items:flex-start;}
                    .page-title{font-size:25px;}
                    .title-icon{width:50px;height:50px;font-size:26px;}
                    .back-link{padding:9px 11px;font-size:14px;}
                    .hero{padding:20px;border-radius:21px;}
                    .period-value{font-size:35px;}
                    .summary-grid{gap:10px;}
                    .summary-card{padding:15px;}
                    .summary-value{font-size:23px;}
                    .action-grid{grid-template-columns:1fr 1fr;gap:10px;}
                    .action-card{min-height:100px;padding:12px;font-size:15px;}
                }
            </style>
        </head>
        <body>
            <main class="close-page">
                <header class="topbar">
                    <div class="title-wrap">
                        <div class="title-icon">📅</div>
                        <div>
                            <h1 class="page-title">CHE 财政月结中心</h1>
                            <p class="subtitle">先完成自动检查，再正式锁定本月账目。</p>
                        </div>
                    </div>
                    <a class="back-link" href="{{ url_for('finance_v7.finance_v7_che_home') }}">← CHE 财政</a>
                </header>

                <section class="hero">
                    <div class="hero-grid">
                        <div>
                            <div class="period-label">月结月份</div>
                            <div class="period-value">{{ data.period_label }}</div>
                            <div>系统会逐项核对银行、Cash、月费、支出、收条和 PV。</div>
                        </div>
                        <div class="status-box">
                            <div>目前状态</div>
                            <strong>🟡 {{ data.status }}</strong>
                            <div class="progress-track"><div class="progress-bar" style="width:{{ data.progress }}%"></div></div>
                            <div class="progress-text">月结检查完成 {{ data.progress }}%</div>
                        </div>
                    </div>
                </section>

                <section class="summary-grid">
                    <div class="summary-card fee">
                        <div class="summary-label">❤️ CHE 月费</div>
                        <div class="summary-value">RM{{ "{:,.2f}".format(data.summary.monthly_fee) }}</div>
                    </div>
                    <div class="summary-card donation">
                        <div class="summary-label">🙏 HQ 布施</div>
                        <div class="summary-value">RM{{ "{:,.2f}".format(data.summary.donation) }}</div>
                    </div>
                    <div class="summary-card expense">
                        <div class="summary-label">📤 CHE 支出</div>
                        <div class="summary-value">RM{{ "{:,.2f}".format(data.summary.expense) }}</div>
                    </div>
                    <div class="summary-card cash">
                        <div class="summary-label">💵 Cash In Hand</div>
                        <div class="summary-value">RM{{ "{:,.2f}".format(data.summary.cash_balance) }}</div>
                    </div>
                </section>

                <div class="main-grid">
                    <div>
                        <section class="section-card">
                            <div class="section-head">
                                <h2>✅ 月结检查清单</h2>
                                <small>{{ data.period }}</small>
                            </div>
                            <div class="check-list">
                                {% for item in data.checks %}
                                <div class="check-row {{ item.status }}">
                                    <div class="check-icon">{{ item.icon }}</div>
                                    <div>
                                        <div class="check-title">{{ item.label }}</div>
                                        <div class="check-detail">{{ item.detail }}</div>
                                    </div>
                                </div>
                                {% endfor %}
                            </div>
                        </section>

                        <section class="section-card health-card {{ data.health_level }}">
                            <div>🩺 月结健康检查</div>
                            <div class="health-title">{{ data.health_title }}</div>
                            <div class="health-message">{{ data.health_message }}</div>
                        </section>

                        <div class="foot-note">
                            <strong>安全原则：</strong>只有 Bank Pending、Cash 实际点算、收条号码、PV 编号和支出审核全部通过后，系统才开放“完成月结”。
                        </div>
                    </div>

                    <aside>
                        <section class="section-card">
                            <div class="section-head"><h2>月结操作</h2></div>
                            <div class="action-grid">
                                {% if data.can_close %}
                                <a class="action-card action-primary" href="{{ url_for('finance_v7.finance_v7_month_close_start', ym=data.period) }}"><span>✅</span>完成月结</a>
                                {% else %}
                                <div class="action-card action-disabled"><span>🔒</span>完成月结</div>
                                {% endif %}
                                <a class="action-card" href="{{ url_for('finance_v7.finance_v7_month_close_report', ym=data.period) }}"><span>📄</span>月结报告</a>
                                <a class="action-card" href="{{ url_for('finance_v7.finance_v7_month_close_history') }}"><span>📜</span>月结历史</a>
                            </div>

                            <div class="score">
                                <div class="score-stars">{{ data.score_stars }}</div>
                                <div class="score-number">{{ data.score }}</div>
                                <div class="score-label">本月账务健康评分</div>
                            </div>
                        </section>

                        <section class="section-card">
                            <div class="section-head"><h2>📜 最近月结</h2></div>
                            <div class="history-list">
                                {% for row in data.history %}
                                <div class="history-row">
                                    <div class="history-period">{{ row.period }}</div>
                                    <div>
                                        <strong>{{ row.status }}</strong>
                                        <div class="history-meta">{{ row.completed_at }} · {{ row.by }}</div>
                                    </div>
                                </div>
                                {% endfor %}
                            </div>
                        </section>
                    </aside>
                </div>
            </main>
        </body>
        </html>
        """,
        data=data,
    )


@finance_v7_bp.route("/finance/v7/CHE/month-close/start", methods=["GET", "POST"])
def finance_v7_month_close_start():
    ym=request.values.get('ym') or date.today().strftime('%Y-%m')
    if request.method=='POST':
        return redirect(f"/finance/month-close?ym={ym}")
    return render_template_string(r"""<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>确认财政月结</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>body{margin:0;background:#f4f7fb;color:#172033}.page{max-width:720px;margin:auto;padding:28px 14px}.box{background:#fff;border:1px solid #dfe6ef;border-radius:22px;padding:24px}.warn{background:#fff8e8;border:1px solid #efd99f;border-radius:16px;padding:16px;line-height:1.65}.actions{display:flex;gap:10px;margin-top:18px}.btn{border:0;border-radius:13px;padding:12px 17px;font-weight:900;text-decoration:none}.primary{background:#2059a8;color:#fff}.back{background:#eef2f7;color:#344158}</style></head><body><main class="page"><section class="box"><h1>✅ 确认 {{ ym }} 月结</h1><div class="warn">系统将进入现有稳定的正式锁账处理。完成前请确认 Bank Pending、Cash 点算、收条与 PV 检查全部通过。</div><form method="post"><input type="hidden" name="ym" value="{{ ym }}"><div class="actions"><button class="btn primary">进入正式锁账</button><a class="btn back" href="{{ url_for('finance_v7.finance_v7_month_close') }}">取消</a></div></form></section></main></body></html>""",ym=ym)


@finance_v7_bp.route("/finance/v7/CHE/month-close/report")
def finance_v7_month_close_report():
    ym=request.args.get('ym') or date.today().strftime('%Y-%m'); start=datetime.strptime(ym+'-01','%Y-%m-%d').date(); _,end=_v7_month_bounds(start)
    fee=_v7_money(_v7_one(f"select coalesce(sum(amount),0) total from finance_records r where record_type='income' and category='月费' and coalesce(status,'confirmed')<>'cancelled' and record_date>=%s and record_date<%s and {_v7_che_filter('r')}",(start,end)).get('total'))
    donation=_v7_money(_v7_one("select coalesce(sum(amount),0) total from finance_records where record_type='income' and category<>'月费' and coalesce(status,'confirmed')<>'cancelled' and record_date>=%s and record_date<%s",(start,end)).get('total'))
    expense=_v7_money(_v7_one("select coalesce(sum(amount),0) total from finance_records where record_type='expense' and coalesce(status,'confirmed')<>'cancelled' and record_date>=%s and record_date<%s",(start,end)).get('total'))
    pending_count,pending_total=_v7_pending_summary('CHE')
    data={'ym':ym,'fee':fee,'donation':donation,'expense':expense,'net':fee-expense,'pending_count':pending_count,'pending_total':pending_total,'balance':_v7_balance_summary('CHE',end.replace(day=1)-__import__('datetime').timedelta(days=1))}
    return render_template_string(r"""<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>月结报告</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>body{margin:0;background:#f4f7fb;color:#172033}.page{max-width:900px;margin:auto;padding:24px 14px}.top{display:flex;justify-content:space-between}.back{text-decoration:none;font-weight:900;color:#2059a8}.grid{display:grid;grid-template-columns:repeat(2,1fr);gap:13px;margin-top:18px}.card{background:#fff;border:1px solid #dfe6ef;border-radius:19px;padding:20px}.label{color:#6f7889}.value{font-size:29px;font-weight:950;margin-top:7px}.warn{margin-top:15px;background:#fff8e8;border:1px solid #efd99f;border-radius:16px;padding:16px}@media(max-width:560px){.grid{grid-template-columns:1fr 1fr}.value{font-size:22px}}</style></head><body><main class="page"><div class="top"><h1>📄 {{ data.ym }} 月结报告</h1><a class="back" href="{{ url_for('finance_v7.finance_v7_month_close') }}">← 返回月结</a></div><section class="grid"><div class="card"><div class="label">CHE 月费</div><div class="value">RM {{ '%.2f'|format(data.fee) }}</div></div><div class="card"><div class="label">HQ 布施</div><div class="value">RM {{ '%.2f'|format(data.donation) }}</div></div><div class="card"><div class="label">CHE 支出</div><div class="value">RM {{ '%.2f'|format(data.expense) }}</div></div><div class="card"><div class="label">月费减支出</div><div class="value">RM {{ '%.2f'|format(data.net) }}</div></div><div class="card"><div class="label">银行余额</div><div class="value">RM {{ '%.2f'|format(data.balance.bank_balance) }}</div></div><div class="card"><div class="label">Cash In Hand</div><div class="value">RM {{ '%.2f'|format(data.balance.cash_in_hand) }}</div></div></section>{% if data.pending_count %}<div class="warn">仍有 {{ data.pending_count }} 笔 Bank Pending，合计 RM {{ '%.2f'|format(data.pending_total) }}。</div>{% endif %}</main></body></html>""",data=data)


@finance_v7_bp.route("/finance/v7/CHE/month-close/excel")
def finance_v7_month_close_excel():
    ym=request.args.get('ym') or date.today().strftime('%Y-%m')
    return redirect(f"/finance/reports/excel?ym={ym}&branch=CHE")


@finance_v7_bp.route("/finance/v7/CHE/month-close/history")
def finance_v7_month_close_history():
    rows=_v7_all("""select to_char(record_date,'YYYY-MM') period,coalesce(sum(amount) filter(where record_type='income' and category='月费'),0) monthly_fee,coalesce(sum(amount) filter(where record_type='expense'),0) expense from finance_records where coalesce(status,'confirmed')<>'cancelled' group by to_char(record_date,'YYYY-MM') order by period desc limit 24""")
    items=[{'period':r.get('period'),'monthly_fee':_v7_money(r.get('monthly_fee')),'expense':_v7_money(r.get('expense'))} for r in rows]
    return render_template_string(r"""<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>月结历史</title><link rel="stylesheet" href="{{ url_for('static',filename='css/toolbox.css') }}"><style>body{margin:0;background:#f4f7fb;color:#172033}.page{max-width:850px;margin:auto;padding:24px 14px}.top{display:flex;justify-content:space-between}.back{text-decoration:none;font-weight:900;color:#2059a8}.list{display:grid;gap:10px;margin-top:18px}.row{display:grid;grid-template-columns:110px 1fr 1fr auto;gap:12px;background:#fff;border:1px solid #dfe6ef;border-radius:16px;padding:15px;align-items:center}.period{font-weight:950}.view{text-decoration:none;font-weight:900;color:#2059a8}@media(max-width:600px){.row{grid-template-columns:1fr 1fr}.view{text-align:right}}</style></head><body><main class="page"><div class="top"><h1>📜 月结历史</h1><a class="back" href="{{ url_for('finance_v7.finance_v7_month_close') }}">← 返回</a></div><section class="list">{% for r in items %}<div class="row"><div class="period">{{ r.period }}</div><div>月费 RM {{ '%.2f'|format(r.monthly_fee) }}</div><div>支出 RM {{ '%.2f'|format(r.expense) }}</div><a class="view" href="{{ url_for('finance_v7.finance_v7_month_close_report',ym=r.period) }}">查看</a></div>{% else %}<p>没有记录</p>{% endfor %}</section></main></body></html>""",items=items)


@finance_v7_bp.route("/finance/v7/CHE/bank/search")
def finance_v7_bank_search():
    """
    V7 CHE 银行总账正式版。

    真实资料来源：
    1. finance_records：
       CHE 银行过账／支票收入
    2. finance_cash_movements：
       Bank In、银行提款、期初余额及银行调整
    3. finance_bank_deposits：
       只补回尚未写入 finance_cash_movements 的旧 Bank In 记录

    不再使用任何测试资料。
    """

    branch = "CHE"
    today = date.today()

    keyword = (request.args.get("q") or "").strip()
    selected_type = (request.args.get("type") or "").strip()
    selected_direction = (request.args.get("direction") or "").strip()
    all_mode = request.args.get("all") == "1"

    selected_year = (request.args.get("year") or "").strip()
    selected_month = (request.args.get("month") or "").strip()

    # 第一次进入页面默认显示本月。
    # 点击“清除筛选／全部记录”时使用 ?all=1，不再自动套用本月。
    if not request.args and not all_mode:
        selected_year = str(today.year)
        selected_month = str(today.month)

    # 防止错误参数进入筛选逻辑。
    try:
        selected_year_int = int(selected_year) if selected_year else None
    except (TypeError, ValueError):
        selected_year = ""
        selected_year_int = None

    try:
        selected_month_int = int(selected_month) if selected_month else None
        if selected_month_int and not 1 <= selected_month_int <= 12:
            raise ValueError
    except (TypeError, ValueError):
        selected_month = ""
        selected_month_int = None

    # =========================================================
    # 1. 读取真实 CHE 银行收入
    # =========================================================
    income_rows = db_query(
        """
        select
            id,
            record_date,
            category,
            name,
            member_id,
            receipt_no,
            bank_ref,
            payment_method,
            amount,
            remarks
        from finance_records
        where record_type = 'income'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and payment_method in (
                '银行过账',
                '银行转账',
                'Online Transfer',
                'DuitNow',
                '支票'
          )
          and (
                upper(coalesce(member_id, '')) like 'CHE-%%'
             or upper(coalesce(receipt_no, '')) like 'CHE%%'
             or fund_account in (
                    '观音堂日常户口',
                    'CHE 日常户口'
                )
          )
        order by record_date desc, id desc
        """,
        fetchall=True,
    ) or []

    # =========================================================
    # 2. 读取银行资金移动
    # =========================================================
    movement_rows = db_query(
        """
        select
            id,
            record_date,
            movement_type,
            amount,
            remarks,
            transfer_ref
        from finance_cash_movements
        where branch = %s
          and account_type = 'bank'
        order by record_date desc, id desc
        """,
        (branch,),
        fetchall=True,
    ) or []

    # =========================================================
    # 3. 补回旧 Bank In
    #    只有在 finance_cash_movements 找不到对应流水时才加入，
    #    避免同一笔 CDM 显示两次。
    # =========================================================
    deposit_rows = db_query(
        """
        select
            d.id,
            d.deposit_date,
            d.reference_no,
            d.amount,
            d.remarks
        from finance_bank_deposits d
        where coalesce(d.branch, 'CHE') = %s
          and not exists (
                select 1
                from finance_cash_movements m
                where m.branch = %s
                  and m.account_type = 'bank'
                  and m.movement_type = 'bank_in'
                  and m.record_date = d.deposit_date
                  and abs(coalesce(m.amount, 0) - coalesce(d.amount, 0)) < 0.005
                  and (
                        lower(trim(coalesce(m.transfer_ref, '')))
                            = lower(trim(coalesce(d.reference_no, '')))
                        or (
                            coalesce(trim(d.reference_no), '') = ''
                            and coalesce(trim(m.transfer_ref), '') = ''
                        )
                  )
          )
        order by d.deposit_date desc, d.id desc
        """,
        (branch, branch),
        fetchall=True,
    ) or []

    # =========================================================
    # 4. 转换成统一的银行总账格式
    # =========================================================
    all_rows = []

    for row in income_rows:
        payment_method = str(row.get("payment_method") or "").strip()

        if payment_method == "支票":
            row_type = "Cheque"
        elif payment_method == "DuitNow":
            row_type = "DuitNow"
        elif payment_method == "Online Transfer":
            row_type = "Online Transfer"
        else:
            row_type = "Bank Transfer"

        category = str(row.get("category") or "银行收入").strip()
        person_name = str(row.get("name") or "").strip()

        reference_parts = [
            str(row.get("bank_ref") or "").strip(),
            str(row.get("receipt_no") or "").strip(),
            str(row.get("member_id") or "").strip(),
        ]
        reference = " · ".join(part for part in reference_parts if part)

        if not reference:
            reference = person_name or str(row.get("remarks") or "").strip() or "-"

        all_rows.append({
            "sort_date": row.get("record_date"),
            "sort_id": int(row.get("id") or 0),
            "date": (
                row["record_date"].isoformat()
                if hasattr(row.get("record_date"), "isoformat")
                else str(row.get("record_date") or "")
            ),
            "time": "",
            "type": row_type,
            "title": f"{category}" + (f" · {person_name}" if person_name else ""),
            "reference": reference,
            "amount": float(row.get("amount") or 0),
            "direction": "in",
            "source": "finance_records",
        })

    movement_type_labels = {
        "bank_in": "Bank In",
        "cash_out": "Cash Withdrawal",
        "withdrawal": "Cash Withdrawal",
        "bank_out": "Bank Out",
        "opening": "Opening Balance",
        "adjustment": "Bank Adjustment",
        "adjustment_in": "Bank Adjustment",
        "adjustment_out": "Bank Adjustment",
    }

    for row in movement_rows:
        movement_type = str(row.get("movement_type") or "").strip()
        transfer_ref = str(row.get("transfer_ref") or "").strip()
        remarks = str(row.get("remarks") or "").strip()

        row_type = movement_type_labels.get(
            movement_type,
            movement_type.replace("_", " ").title() or "Bank Movement",
        )

        if movement_type == "bank_in" and transfer_ref.upper().startswith("CDM"):
            row_type = "CDM"

        is_out = movement_type in {
            "cash_out",
            "withdrawal",
            "bank_out",
            "adjustment_out",
        }

        is_neutral = movement_type == "opening" and float(row.get("amount") or 0) == 0

        if row_type == "CDM":
            title = "现金存入银行"
        elif row_type == "Cash Withdrawal":
            title = "银行提款转入 Petty Cash"
        elif row_type == "Opening Balance":
            title = "银行期初余额"
        elif row_type == "Bank Adjustment":
            title = "银行余额调整"
        else:
            title = remarks or row_type

        all_rows.append({
            "sort_date": row.get("record_date"),
            "sort_id": 10_000_000 + int(row.get("id") or 0),
            "date": (
                row["record_date"].isoformat()
                if hasattr(row.get("record_date"), "isoformat")
                else str(row.get("record_date") or "")
            ),
            "time": "",
            "type": row_type,
            "title": title,
            "reference": transfer_ref or remarks or "-",
            "amount": abs(float(row.get("amount") or 0)),
            "direction": "neutral" if is_neutral else ("out" if is_out else "in"),
            "source": "finance_cash_movements",
        })

    for row in deposit_rows:
        reference_no = str(row.get("reference_no") or "").strip()
        remarks = str(row.get("remarks") or "").strip()
        row_type = "CDM" if reference_no.upper().startswith("CDM") else "Bank In"

        all_rows.append({
            "sort_date": row.get("deposit_date"),
            "sort_id": 20_000_000 + int(row.get("id") or 0),
            "date": (
                row["deposit_date"].isoformat()
                if hasattr(row.get("deposit_date"), "isoformat")
                else str(row.get("deposit_date") or "")
            ),
            "time": "",
            "type": row_type,
            "title": "现金存入银行",
            "reference": reference_no or remarks or "-",
            "amount": abs(float(row.get("amount") or 0)),
            "direction": "in",
            "source": "finance_bank_deposits",
        })

    all_rows.sort(
        key=lambda item: (
            item.get("sort_date") or date.min,
            item.get("sort_id") or 0,
        ),
        reverse=True,
    )

    # =========================================================
    # 5. 下拉选项从真实资料产生
    # =========================================================
    available_years = {
        int(str(row.get("date") or "")[:4])
        for row in all_rows
        if len(str(row.get("date") or "")) >= 4
        and str(row.get("date") or "")[:4].isdigit()
    }
    available_years.add(today.year)
    year_options = sorted(available_years, reverse=True)

    type_options = sorted({
        str(row.get("type") or "").strip()
        for row in all_rows
        if str(row.get("type") or "").strip()
    })

    month_options = list(range(1, 13))

    # =========================================================
    # 6. 正式筛选
    # =========================================================
    needle = keyword.casefold()

    rows = []

    for row in all_rows:
        row_date = str(row.get("date") or "")

        if selected_year_int:
            if not row_date.startswith(f"{selected_year_int:04d}-"):
                continue

        if selected_month_int:
            if len(row_date) < 7 or row_date[5:7] != f"{selected_month_int:02d}":
                continue

        if selected_type and row.get("type") != selected_type:
            continue

        if selected_direction and row.get("direction") != selected_direction:
            continue

        if needle:
            searchable = " ".join([
                row_date,
                str(row.get("time") or ""),
                str(row.get("type") or ""),
                str(row.get("title") or ""),
                str(row.get("reference") or ""),
                f'{float(row.get("amount") or 0):.2f}',
            ]).casefold()

            if needle not in searchable:
                continue

        rows.append(row)

    # 防止一次显示过多记录拖慢页面。
    rows = rows[:500]

    total_in = sum(
        float(row.get("amount") or 0)
        for row in rows
        if row.get("direction") == "in"
    )

    total_out = sum(
        float(row.get("amount") or 0)
        for row in rows
        if row.get("direction") == "out"
    )

    net_total = total_in - total_out

    return render_template_string(
        """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">

<title>CHE 银行总账</title>

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">

<style>
*{box-sizing:border-box}

body{
    margin:0;
    background:#f3f7fc;
    color:#172033;
}

.page{
    width:min(1120px,calc(100% - 28px));
    margin:0 auto;
    padding:24px 0 52px;
}

.top{
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:14px;
    margin-bottom:18px;
}

.heading{
    display:flex;
    align-items:center;
    gap:14px;
}

.heading-icon{
    width:58px;
    height:58px;
    display:grid;
    place-items:center;
    border-radius:18px;
    background:linear-gradient(135deg,#2563eb,#0891b2);
    color:#fff;
    font-size:29px;
    box-shadow:0 12px 26px rgba(37,99,235,.22);
}

h1{
    margin:0;
    font-size:31px;
}

.sub{
    margin-top:5px;
    color:#718096;
    font-size:15px;
}

.back{
    padding:11px 15px;
    border:1px solid #d8e1ec;
    border-radius:14px;
    background:#fff;
    color:#263149;
    text-decoration:none;
    font-weight:900;
    white-space:nowrap;
}

.card{
    background:#fff;
    border:1px solid #dde6f0;
    border-radius:22px;
    padding:20px;
    box-shadow:0 12px 30px rgba(32,49,80,.06);
}

.filter-form{
    display:flex;
    flex-direction:column;
    gap:17px;
}

.filter-grid{
    display:grid;
    grid-template-columns:2.2fr repeat(4,minmax(120px,1fr));
    gap:13px;
    align-items:end;
}

.field{
    display:flex;
    flex-direction:column;
    gap:7px;
}

.field label{
    color:#64748b;
    font-size:14px;
    font-weight:800;
}

.field input,
.field select{
    width:100%;
    min-height:50px;
    padding:10px 13px;
    border:1px solid #cbd5e1;
    border-radius:13px;
    background:#fff;
    color:#172033;
    font-size:15px;
    outline:none;
}

.field input:focus,
.field select:focus{
    border-color:#3b82f6;
    box-shadow:0 0 0 4px rgba(59,130,246,.12);
}

.quick-row{
    display:flex;
    flex-wrap:wrap;
    gap:10px;
    align-items:center;
}

.quick-btn,
.clear-btn,
.search-btn{
    min-height:46px;
    padding:11px 17px;
    border-radius:13px;
    font-size:15px;
    font-weight:900;
    text-decoration:none;
    display:inline-flex;
    align-items:center;
    justify-content:center;
}

.quick-btn{
    color:#1d4ed8;
    background:#eff6ff;
    border:1px solid #bfdbfe;
}

.quick-btn.income{
    color:#166534;
    background:#f0fdf4;
    border-color:#bbf7d0;
}

.quick-btn.outgoing{
    color:#b42318;
    background:#fff5f4;
    border-color:#fecaca;
}

.clear-btn{
    color:#64748b;
    background:#f8fafc;
    border:1px solid #cbd5e1;
}

.search-btn{
    margin-left:auto;
    border:0;
    color:#fff;
    background:linear-gradient(135deg,#2563eb,#3b82f6);
    cursor:pointer;
    box-shadow:0 9px 20px rgba(37,99,235,.22);
}

.hint{
    margin-top:11px;
    color:#8a94a5;
    font-size:13px;
}

.summary-grid{
    display:grid;
    grid-template-columns:repeat(4,1fr);
    gap:14px;
    margin:20px 0 25px;
}

.summary-card{
    background:#fff;
    padding:19px;
    border-radius:18px;
    border:1px solid #e2e8f0;
    box-shadow:0 8px 24px rgba(15,23,42,.05);
}

.summary-label{
    color:#64748b;
    font-size:14px;
    margin-bottom:7px;
}

.summary-value{
    color:#0f172a;
    font-size:25px;
    font-weight:950;
}

.amount-in{color:#15803d}
.amount-out{color:#c2413b}

.result-head{
    display:flex;
    justify-content:space-between;
    align-items:end;
    margin:25px 3px 11px;
}

.result-title{
    font-size:26px;
    font-weight:950;
}

.result-count{
    color:#7b8597;
    font-size:14px;
}

.list{
    display:grid;
    gap:10px;
}

.row{
    position:relative;
    display:grid;
    grid-template-columns:115px minmax(0,1fr) auto;
    gap:16px;
    align-items:center;
    background:#fff;
    border:1px solid #e1e7ef;
    border-radius:17px;
    padding:15px 17px 15px 20px;
    overflow:hidden;
    box-shadow:0 6px 18px rgba(15,23,42,.035);
}

.row::before{
    content:"";
    position:absolute;
    left:0;
    top:0;
    bottom:0;
    width:6px;
    background:#94a3b8;
}

.row.direction-in::before{background:#22a06b}
.row.direction-out::before{background:#e05252}

.date{
    color:#64748b;
    font-size:13px;
    font-weight:850;
}

.type{
    display:inline-flex;
    width:max-content;
    padding:4px 9px;
    border-radius:999px;
    background:#eff6ff;
    color:#1d4ed8;
    font-size:12px;
    font-weight:900;
    margin-bottom:5px;
}

.title{
    font-size:17px;
    font-weight:950;
    line-height:1.35;
}

.ref{
    margin-top:5px;
    color:#718096;
    font-size:13px;
    overflow-wrap:anywhere;
}

.amount{
    text-align:right;
    font-size:20px;
    font-weight:950;
    white-space:nowrap;
}

.in{color:#16834f}
.out{color:#c33d34}
.neutral{color:#64748b}

.empty{
    text-align:center;
    padding:44px 18px;
    background:#fff;
    border:1px dashed #ccd6e3;
    border-radius:18px;
    color:#788397;
}

.limit-note{
    margin-top:12px;
    color:#94a3b8;
    text-align:right;
    font-size:12px;
}

@media(max-width:900px){
    .filter-grid{
        grid-template-columns:1fr 1fr;
    }

    .keyword-field{
        grid-column:1/-1;
    }

    .summary-grid{
        grid-template-columns:1fr 1fr;
    }

    .search-btn{
        margin-left:0;
    }
}

@media(max-width:620px){
    .page{
        width:min(100% - 18px,1120px);
        padding-top:14px;
    }

    .top{
        align-items:flex-start;
    }

    .heading-icon{
        width:50px;
        height:50px;
        font-size:25px;
    }

    h1{
        font-size:24px;
    }

    .sub,
    .back{
        font-size:13px;
    }

    .filter-grid,
    .summary-grid{
        grid-template-columns:1fr;
    }

    .keyword-field{
        grid-column:auto;
    }

    .quick-row > *{
        width:100%;
    }

    .row{
        grid-template-columns:1fr auto;
        padding:14px 14px 14px 18px;
    }

    .date{
        grid-column:1/-1;
    }

    .title{
        font-size:16px;
    }

    .amount{
        font-size:18px;
    }
}
</style>
</head>

<body>
<main class="page">

<header class="top">
    <div class="heading">
        <div class="heading-icon">🏦</div>
        <div>
            <h1>CHE 银行总账</h1>
            <div class="sub">
                查询银行收入、CDM Bank In、提款、期初余额及银行调整
            </div>
        </div>
    </div>

    <a class="back"
       href="{{ url_for('finance_v7.finance_v7_bank') }}">
        ← 返回银行
    </a>
</header>

<section class="card">
    <form method="get" class="filter-form">

        <div class="filter-grid">

            <div class="field keyword-field">
                <label>关键字</label>
                <input
                    type="search"
                    name="q"
                    value="{{ keyword }}"
                    placeholder="Bank Ref、收条、会员、CDM、说明或金额"
                >
            </div>

            <div class="field">
                <label>年份</label>
                <select name="year">
                    <option value="">全部年份</option>
                    {% for y in year_options %}
                        <option
                            value="{{ y }}"
                            {% if selected_year|string == y|string %}selected{% endif %}
                        >
                            {{ y }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="field">
                <label>月份</label>
                <select name="month">
                    <option value="">全部月份</option>
                    {% for m in month_options %}
                        <option
                            value="{{ m }}"
                            {% if selected_month|string == m|string %}selected{% endif %}
                        >
                            {{ "%02d"|format(m) }} 月
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="field">
                <label>记录类型</label>
                <select name="type">
                    <option value="">全部类型</option>
                    {% for type_name in type_options %}
                        <option
                            value="{{ type_name }}"
                            {% if selected_type == type_name %}selected{% endif %}
                        >
                            {{ type_name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="field">
                <label>资金方向</label>
                <select name="direction">
                    <option value="">全部方向</option>
                    <option value="in"
                        {% if selected_direction == "in" %}selected{% endif %}>
                        银行入账
                    </option>
                    <option value="out"
                        {% if selected_direction == "out" %}selected{% endif %}>
                        银行出账
                    </option>
                    <option value="neutral"
                        {% if selected_direction == "neutral" %}selected{% endif %}>
                        期初／中性记录
                    </option>
                </select>
            </div>

        </div>

        <div class="quick-row">

            <a class="quick-btn"
               href="{{ url_for(
                    'finance_v7.finance_v7_bank_search',
                    year=current_year,
                    month=current_month
               ) }}">
                📅 本月
            </a>

            <a class="quick-btn income"
               href="{{ url_for(
                    'finance_v7.finance_v7_bank_search',
                    year=selected_year,
                    month=selected_month,
                    direction='in'
               ) }}">
                ＋ 银行入账
            </a>

            <a class="quick-btn outgoing"
               href="{{ url_for(
                    'finance_v7.finance_v7_bank_search',
                    year=selected_year,
                    month=selected_month,
                    direction='out'
               ) }}">
                － 银行出账
            </a>

            <a class="clear-btn"
               href="{{ url_for(
                    'finance_v7.finance_v7_bank_search',
                    all=1
               ) }}">
                全部记录
            </a>

            <button class="search-btn" type="submit">
                🔎 查询银行总账
            </button>

        </div>
    </form>

    <div class="hint">
        首次进入默认显示本月；“全部记录”会移除年月及其它筛选。
    </div>
</section>

<section class="summary-grid">

    <div class="summary-card">
        <div class="summary-label">符合条件记录</div>
        <div class="summary-value">{{ rows|length }} 笔</div>
    </div>

    <div class="summary-card">
        <div class="summary-label">银行入账</div>
        <div class="summary-value amount-in">
            RM {{ "{:,.2f}".format(total_in) }}
        </div>
    </div>

    <div class="summary-card">
        <div class="summary-label">银行出账</div>
        <div class="summary-value amount-out">
            RM {{ "{:,.2f}".format(total_out) }}
        </div>
    </div>

    <div class="summary-card">
        <div class="summary-label">净变动</div>
        <div class="summary-value
             {% if net_total >= 0 %}amount-in{% else %}amount-out{% endif %}">
            {% if net_total < 0 %}-{% endif %}
            RM {{ "{:,.2f}".format(net_total|abs) }}
        </div>
    </div>

</section>

<div class="result-head">
    <div class="result-title">
        {% if keyword or selected_year or selected_month
              or selected_type or selected_direction %}
            查询结果
        {% else %}
            全部银行记录
        {% endif %}
    </div>

    <div class="result-count">
        {{ rows|length }} 笔
    </div>
</div>

{% if rows %}
    <section class="list">
        {% for row in rows %}
            <article class="row direction-{{ row.direction }}">

                <div class="date">
                    {{ row.date }}
                    {% if row.time %}
                        <br>{{ row.time }}
                    {% endif %}
                </div>

                <div>
                    <div class="type">{{ row.type }}</div>
                    <div class="title">{{ row.title }}</div>
                    <div class="ref">{{ row.reference }}</div>
                </div>

                <div class="amount {{ row.direction }}">
                    {% if row.direction == "in" %}
                        +
                    {% elif row.direction == "out" %}
                        -
                    {% endif %}
                    RM {{ "{:,.2f}".format(row.amount) }}
                </div>

            </article>
        {% endfor %}
    </section>

    {% if rows|length >= 500 %}
        <div class="limit-note">
            为保持页面速度，本次最多显示 500 笔记录。
        </div>
    {% endif %}
{% else %}
    <div class="empty">
        找不到符合筛选条件的银行记录
    </div>
{% endif %}

</main>
</body>
</html>
        """,
        keyword=keyword,
        rows=rows,
        selected_year=selected_year,
        selected_month=selected_month,
        selected_type=selected_type,
        selected_direction=selected_direction,
        year_options=year_options,
        month_options=month_options,
        type_options=type_options,
        current_year=today.year,
        current_month=today.month,
        total_in=total_in,
        total_out=total_out,
        net_total=net_total,
    )

# V7 compatibility helpers for migrated bank pages.
def get_finance_balance_summary(branch="CHE", balance_date=None):
    return _v7_balance_summary(branch, balance_date or date.today())

FINANCE_DATE_COMPONENT = ""

# =========================================================
# Finance V7 Bank Zone - migrated from Legacy on 2026-07-22
# =========================================================

def normalize_cdm_reference(value):
    """
    支持输入：
    1
    5
    CDM1
    CDM-1
    cdm 1

    统一转换为：
    CDM-1
    """
    text = str(value or "").strip().upper()
    text = re.sub(r"\s+", "", text)

    if not text:
        return ""

    if text.isdigit():
        return f"CDM-{int(text)}"

    match = re.fullmatch(r"CDM-?(\d+)", text)

    if match:
        return f"CDM-{int(match.group(1))}"

    return ""

def get_next_cdm_reference(deposit_date=None):
    """
    根据 Bank In 日期所在月份，建议下一个 CDM 编号。

    例如该月份已有：
    CDM-1、CDM-2、CDM-4

    系统建议：
    CDM-5
    """
    deposit_date = deposit_date or date.today()

    if isinstance(deposit_date, str):
        try:
            deposit_date = datetime.strptime(
                deposit_date,
                "%Y-%m-%d"
            ).date()
        except ValueError:
            deposit_date = date.today()

    ym = deposit_date.strftime("%Y-%m")

    rows = db_query(
        """
        select
            reference_no,
            cdm_sequence
        from finance_bank_deposits
        where branch = 'CHE'
          and (
                ym = %s
                or to_char(deposit_date, 'YYYY-MM') = %s
          )
          and (
                upper(coalesce(reference_no, '')) like 'CDM-%%'
                or cdm_sequence is not null
          )
        """,
        (ym, ym),
        fetchall=True,
    ) or []

    largest = 0

    for row in rows:
        sequence = row.get("cdm_sequence")

        if sequence is not None:
            match = re.search(r"\d+", str(sequence))

            if match:
                largest = max(largest, int(match.group()))
                continue

        reference = normalize_cdm_reference(
            row.get("reference_no")
        )

        match = re.fullmatch(r"CDM-(\d+)", reference)

        if match:
            largest = max(largest, int(match.group(1)))

    return f"CDM-{largest + 1}"

def ensure_finance_bank_in_tables():
    """建立月费现金 Bank In 所需资料表；可重复安全执行。"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                create table if not exists finance_bank_deposits (
                    id bigserial primary key,
                    branch varchar(10) not null default 'CHE',
                    deposit_date date not null,
                    reference_no varchar(120),
                    amount numeric(14,2) not null default 0,
                    remarks text,
                    created_at timestamp not null default now()
                )
            """)
            cur.execute("""
                create table if not exists finance_bank_deposit_items (
                    id bigserial primary key,
                    deposit_id bigint not null references finance_bank_deposits(id) on delete cascade,
                    finance_record_id bigint not null,
                    amount numeric(14,2) not null default 0,
                    created_at timestamp not null default now(),
                    unique(finance_record_id)
                )
            """)
        conn.commit()

OFFICIAL_CHE_CDM_ROWS = [
    # ym, CDM编号, Bank In日期, 金额, 收条起号, 收条尾号
    (
        "2026-01",
        1,
        "2026-01-06",
        Decimal("1800.00"),
        "CHE0001292",
        "CHE0001300",
    ),
    (
        "2026-01",
        2,
        "2026-01-06",
        Decimal("2700.00"),
        "CHE0001401",
        "CHE0001407",
    ),
    (
        "2026-01",
        3,
        "2026-01-14",
        Decimal("2000.00"),
        "CHE0001408",
        "CHE0001412",
    ),
    (
        "2026-01",
        4,
        "2026-01-26",
        Decimal("1150.00"),
        "CHE0001413",
        "CHE0001417",
    ),

    (
        "2026-02",
        1,
        "2026-02-09",
        Decimal("1900.00"),
        "CHE0001425",
        "CHE0001431",
    ),

    (
        "2026-03",
        1,
        "2026-03-02",
        Decimal("2250.00"),
        "CHE0001432",
        "CHE0001440",
    ),
    (
        "2026-03",
        2,
        "2026-03-11",
        Decimal("2000.00"),
        "CHE0001441",
        "CHE0001448",
    ),
    (
        "2026-03",
        3,
        "2026-03-22",
        Decimal("1250.00"),
        "CHE0001449",
        "CHE0001453",
    ),

    (
        "2026-04",
        1,
        "2026-04-12",
        Decimal("2250.00"),
        "CHE0001454",
        "CHE0001462",
    ),
    (
        "2026-04",
        2,
        "2026-04-20",
        Decimal("1100.00"),
        "CHE0001463",
        "CHE0001468",
    ),
    (
        "2026-04",
        3,
        "2026-04-29",
        Decimal("450.00"),
        "CHE0001469",
        "CHE0001471",
    ),

    (
        "2026-05",
        1,
        "2026-05-04",
        Decimal("1900.00"),
        "CHE0001472",
        "CHE0001478",
    ),
    (
        "2026-05",
        2,
        "2026-05-18",
        Decimal("300.00"),
        "CHE0001479",
        "CHE0001480",
    ),
    (
        "2026-05",
        3,
        "2026-05-26",
        Decimal("800.00"),
        "CHE0001481",
        "CHE0001484",
    ),
    (
        "2026-05",
        4,
        "2026-05-29",
        Decimal("400.00"),
        "CHE0001485",
        "CHE0001487",
    ),

    (
        "2026-06",
        1,
        "2026-06-04",
        Decimal("900.00"),
        "CHE0001488",
        "CHE0001492",
    ),
    (
        "2026-06",
        2,
        "2026-06-18",
        Decimal("200.00"),
        "CHE0001493",
        "CHE0001494",
    ),
    (
        "2026-06",
        3,
        "2026-06-26",
        Decimal("300.00"),
        "CHE0001495",
        "CHE0001500",
    ),
    (
        "2026-06",
        3,
        "2026-06-26",
        Decimal("300.00"),
        "CHE0001495",
        "CHE0001500",
    ),
        # 历史遗漏：2026年1月 CDM-5
    (
        "2026-01",
        5,
        "2026-01-31",
        Decimal("1200.00"),
        "CHE0001418",
        "CHE0001424",
    ),

    # 历史遗漏：2026年6月 CDM-4
    (
        "2026-06",
        4,
        "2026-06-30",
        Decimal("2200.00"),
        "CHE0001551",
        "CHE0001563",
    ),
]



def _bank_in_receipt_number(receipt_no):
    """CHE0001454 / CHE1454 -> 1454。"""
    match = re.search(r"(\d+)$", str(receipt_no or "").strip())
    return int(match.group(1)) if match else None

def _bank_in_decimal(value):
    try:
        return Decimal(str(value or 0)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")

def _find_che_cash_monthly_receipts(cur, receipt_from, receipt_to):
    """
    按 CHE 收条号码范围寻找现金月费。
    finance_records 没有 branch，因此以 receipt_no 的 CHE 前缀判断。
    """
    start_no = _bank_in_receipt_number(receipt_from)
    end_no = _bank_in_receipt_number(receipt_to)

    if start_no is None or end_no is None:
        raise ValueError(f"收条范围格式错误：{receipt_from} 至 {receipt_to}")

    cur.execute(
        """
        select
            r.id,
            r.record_date,
            r.receipt_date,
            r.receipt_no,
            r.member_id,
            r.name,
            r.amount,
            r.status,
            r.remarks
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and r.payment_method = '现金'
          and coalesce(r.status, 'confirmed') <> 'cancelled'
          and upper(coalesce(r.receipt_no, '')) like 'CHE%%'
          and nullif(
                regexp_replace(
                    coalesce(r.receipt_no, ''),
                    '[^0-9]',
                    '',
                    'g'
                ),
                ''
              )::bigint between %s and %s
        order by
            nullif(
                regexp_replace(
                    coalesce(r.receipt_no, ''),
                    '[^0-9]',
                    '',
                    'g'
                ),
                ''
            )::bigint,
            r.id
        """,
        (start_no, end_no),
    )
    return cur.fetchall() or []

def _get_bank_in_master_rows(cur):
    cur.execute(
        """
        select
            d.id,
            d.ym,
            d.cdm_sequence,
            d.deposit_date,
            d.amount,
            d.receipt_from,
            d.receipt_to,
            d.reference_no,
            d.branch
        from finance_bank_deposits d
        where d.branch = 'CHE'
          and d.ym between '2026-01' and '2026-06'
        order by d.ym, d.cdm_sequence, d.id
        """
    )
    return cur.fetchall() or []

def _scan_che_bank_in():
    """只扫描，不修改数据库。"""
    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            master_rows = _get_bank_in_master_rows(cur)

            master_by_key = {}
            duplicate_keys = set()

            for row in master_rows:
                key = (
                    str(row.get("ym") or ""),
                    int(row.get("cdm_sequence") or 0),
                )
                if key in master_by_key:
                    duplicate_keys.add(key)
                else:
                    master_by_key[key] = row

            official_keys = {
                (ym, sequence)
                for ym, sequence, *_ in OFFICIAL_CHE_CDM_ROWS
            }

            extra_master_rows = [
                row for row in master_rows
                if (
                    str(row.get("ym") or ""),
                    int(row.get("cdm_sequence") or 0),
                ) not in official_keys
            ]

            batches = []
            master_ok_count = 0
            validation_ok_count = 0

            for (
                ym,
                sequence,
                deposit_date_text,
                expected_amount,
                expected_from,
                expected_to,
            ) in OFFICIAL_CHE_CDM_ROWS:

                key = (ym, sequence)
                master = master_by_key.get(key)
                master_issues = []

                if not master:
                    master_issues.append("缺少 CDM 主记录")
                else:
                    actual_date = str(master.get("deposit_date") or "")
                    actual_amount = _bank_in_decimal(master.get("amount"))
                    actual_from = str(master.get("receipt_from") or "")
                    actual_to = str(master.get("receipt_to") or "")

                    if actual_date != deposit_date_text:
                        master_issues.append(
                            f"Bank In 日期应为 {deposit_date_text}"
                        )
                    if actual_amount != expected_amount:
                        master_issues.append(
                            f"CDM 金额应为 RM {expected_amount:,.2f}"
                        )
                    if (
                        actual_from != expected_from
                        or actual_to != expected_to
                    ):
                        master_issues.append(
                            f"收条范围应为 {expected_from} 至 {expected_to}"
                        )

                if key in duplicate_keys:
                    master_issues.append("同月份相同 CDM 编号有重复主记录")

                master_ok = not master_issues
                if master_ok:
                    master_ok_count += 1

                receipts = _find_che_cash_monthly_receipts(
                    cur,
                    expected_from,
                    expected_to,
                )

                receipt_total = sum(
                    (_bank_in_decimal(row.get("amount")) for row in receipts),
                    Decimal("0.00"),
                )

                validation_issues = []

                if not receipts:
                    validation_issues.append(
                        f"找不到收条 {expected_from} 至 {expected_to}"
                    )
                elif receipt_total != expected_amount:
                    validation_issues.append(
                        "官方金额："
                        f"RM {expected_amount:,.2f}；"
                        f"实际收条：RM {receipt_total:,.2f}。"
                        "（仅提示，不阻止重建）"
                    )

                validation_ok = bool(receipts)
                if validation_ok:
                    validation_ok_count += 1

                batches.append({
                    "ym": ym,
                    "sequence": sequence,
                    "deposit_date": deposit_date_text,
                    "expected_amount": expected_amount,
                    "expected_from": expected_from,
                    "expected_to": expected_to,
                    "master": master,
                    "master_issues": master_issues,
                    "master_ok": master_ok,
                    "receipts": receipts,
                    "receipt_total": receipt_total,
                    "validation_issues": validation_issues,
                    "validation_ok": validation_ok,
                })

            cur.execute(
                """
                select count(*) as count
                from finance_bank_deposit_items
                """
            )
            item_count = int((cur.fetchone() or {}).get("count") or 0)

            cur.execute(
                """
                select
                    count(*) as count,
                    coalesce(sum(r.amount), 0) as total
                from finance_records r
                where r.record_type = 'income'
                  and r.category = '月费'
                  and r.payment_method = '现金'
                  and coalesce(r.status, 'confirmed') <> 'cancelled'
                  and r.record_date >= date '2026-01-01'
                  and (
                        upper(coalesce(r.member_id, '')) like 'CHE-%%'
                     or upper(coalesce(r.receipt_no, '')) like 'CHE%%'
                  )
                  and not exists (
                        select 1
                        from finance_bank_deposit_items i
                        where i.finance_record_id = r.id
                  )
                """
            )
            waiting = cur.fetchone() or {}

    all_master_ok = master_ok_count == len(OFFICIAL_CHE_CDM_ROWS)
    all_validation_ok = validation_ok_count == len(OFFICIAL_CHE_CDM_ROWS)

    return {
        "batches": batches,
        "master_count": len(master_rows),
        "master_ok_count": master_ok_count,
        "validation_ok_count": validation_ok_count,
        "official_count": len(OFFICIAL_CHE_CDM_ROWS),
        "extra_master_rows": extra_master_rows,
        "all_master_ok": all_master_ok,
        "all_validation_ok": all_validation_ok,
        "can_rebuild": all_master_ok and all_validation_ok,
        "item_count": item_count,
        "waiting_count": int(waiting.get("count") or 0),
        "waiting_total": _bank_in_decimal(waiting.get("total")),
    }

def _repair_che_bank_in_master():
    """只修正 finance_bank_deposits，不碰明细和收条。"""
    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select *
                    from finance_bank_deposits
                    where branch = 'CHE'
                      and ym between '2026-01' and '2026-06'
                    order by ym, cdm_sequence, id
                    for update
                    """
                )
                rows = cur.fetchall() or []

                official_keys = {
                    (ym, sequence)
                    for ym, sequence, *_ in OFFICIAL_CHE_CDM_ROWS
                }

                keep_by_key = {}
                remove_ids = []

                for row in rows:
                    key = (
                        str(row.get("ym") or ""),
                        int(row.get("cdm_sequence") or 0),
                    )

                    if key not in official_keys or key in keep_by_key:
                        remove_ids.append(row["id"])
                    else:
                        keep_by_key[key] = row

                remove_ids = sorted(set(remove_ids))

                if remove_ids:
                    cur.execute(
                        """
                        delete from finance_bank_deposit_items
                        where deposit_id = any(%s)
                        """,
                        (remove_ids,),
                    )
                    cur.execute(
                        """
                        delete from finance_bank_deposits
                        where id = any(%s)
                        """,
                        (remove_ids,),
                    )

                repaired = 0

                for (
                    ym,
                    sequence,
                    deposit_date_text,
                    expected_amount,
                    receipt_from,
                    receipt_to,
                ) in OFFICIAL_CHE_CDM_ROWS:
                    
                    receipts = _find_che_cash_monthly_receipts(
                        cur,
                        receipt_from,
                        receipt_to,
                    )

                    actual_amount = sum(
                        (_bank_in_decimal(r["amount"]) for r in receipts),
                        Decimal("0.00"),
                    )

                    current = keep_by_key.get((ym, sequence))

                    if current and current["id"] not in remove_ids:
                        cur.execute(
                            """
                            update finance_bank_deposits
                            set
                                branch = 'CHE',
                                deposit_date = %s,
                                ym = %s,
                                fund_account = '观音堂日常户口',
                                bank_name = 'Hong Leong Bank',
                                reference_no = %s,
                                amount = %s,
                                remarks = '系统修复：CHE 月费现金 CDM',
                                receipt_from = %s,
                                receipt_to = %s,
                                cdm_sequence = %s
                            where id = %s
                            """,
                            (
                                deposit_date_text,
                                ym,
                                f"CDM-{sequence}",
                                actual_amount,
                                receipt_from,
                                receipt_to,
                                sequence,
                                current["id"],
                            ),
                        )
                    else:
                        cur.execute(
                            """
                            insert into finance_bank_deposits
                            (
                                branch,
                                deposit_date,
                                ym,
                                fund_account,
                                bank_name,
                                reference_no,
                                amount,
                                remarks,
                                receipt_from,
                                receipt_to,
                                cdm_sequence
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                '观音堂日常户口',
                                'Hong Leong Bank',
                                %s,
                                %s,
                                '系统修复：CHE 月费现金 CDM',
                                %s,
                                %s,
                                %s
                            )
                            """,
                            (
                                deposit_date_text,
                                ym,
                                f"CDM-{sequence}",
                                actual_amount,
                                receipt_from,
                                receipt_to,
                                sequence,
                            ),
                        )

                    repaired += 1

            conn.commit()
            return repaired

        except Exception:
            conn.rollback()
            raise

def _update_bank_in_receipt_amount(record_id, new_amount):
    """
    只允许修改 CHE 现金月费记录金额。
    这里不会猜测哪一张错误，必须由负责人看原始收条后手动输入。
    """
    amount = _bank_in_decimal(new_amount)

    if amount <= 0:
        raise ValueError("金额必须大于 0。")

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select
                        id,
                        receipt_no,
                        member_id,
                        name,
                        amount
                    from finance_records
                    where id = %s
                      and record_type = 'income'
                      and category = '月费'
                      and payment_method = '现金'
                      and coalesce(status, 'confirmed') <> 'cancelled'
                      and upper(coalesce(receipt_no, '')) like 'CHE%%'
                    for update
                    """,
                    (record_id,),
                )
                row = cur.fetchone()

                if not row:
                    raise ValueError("找不到这笔可修改的 CHE 现金月费记录。")

                old_amount = _bank_in_decimal(row.get("amount"))

                cur.execute(
                    """
                    update finance_records
                    set
                        amount = %s,
                        remarks = concat_ws(
                            ' | ',
                            nullif(remarks, ''),
                            %s
                        )
                    where id = %s
                    """,
                    (
                        amount,
                        (
                            "Bank In 修复中心调整金额："
                            f"RM {old_amount:,.2f} → RM {amount:,.2f}"
                        ),
                        record_id,
                    ),
                )

            conn.commit()

            return {
                "receipt_no": row.get("receipt_no") or "",
                "name": row.get("name") or "",
                "old_amount": old_amount,
                "new_amount": amount,
            }

        except Exception:
            conn.rollback()
            raise

def _find_missing_che_cdm_candidates():
    """
    自动寻找 2026-01 至 2026-06：
    尚未连接 Bank In 的连续 CHE 现金月费收条。

    这里只负责找候选范围，不修改数据库。
    2026-07 起不属于历史维修范围，继续留在正常 CDM 页面。
    """
    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            cur.execute(
                """
                select
                    r.id,
                    r.record_date,
                    r.receipt_no,
                    r.member_id,
                    r.name,
                    r.amount
                from finance_records r
                where r.record_type = 'income'
                  and r.category = '月费'
                  and lower(
                        trim(coalesce(r.payment_method, ''))
                      ) in ('现金', 'cash')
                  and coalesce(
                        r.status,
                        'confirmed'
                      ) <> 'cancelled'
                  and r.record_date >= date '2026-01-01'
                  and r.record_date < date '2026-07-01'
                  and upper(
                        coalesce(r.receipt_no, '')
                      ) like 'CHE%%'
                  and not exists (
                        select 1
                        from finance_bank_deposit_items i
                        where i.finance_record_id = r.id
                  )
                order by
                    r.record_date,
                    r.id
                """
            )

            rows = cur.fetchall() or []

    valid_rows = []

    for row in rows:
        receipt_number = _bank_in_receipt_number(
            row.get("receipt_no")
        )

        if receipt_number is None:
            continue

        row["receipt_number"] = receipt_number
        row["amount"] = _bank_in_decimal(
            row.get("amount")
        )

        valid_rows.append(row)

    # 必须按收条号码排列，不能只按日期。
    valid_rows.sort(
        key=lambda row: (
            row["receipt_number"],
            row["id"],
        )
    )

    groups = []
    current_group = []

    for row in valid_rows:

        if not current_group:
            current_group = [row]
            continue

        previous_number = current_group[-1][
            "receipt_number"
        ]

        if row["receipt_number"] == previous_number + 1:
            current_group.append(row)
        else:
            groups.append(current_group)
            current_group = [row]

    if current_group:
        groups.append(current_group)

    candidates = []

    for index, group in enumerate(groups, start=1):

        receipt_total = sum(
            (
                _bank_in_decimal(row.get("amount"))
                for row in group
            ),
            Decimal("0.00"),
        )

        first_row = group[0]
        last_row = group[-1]

        candidates.append({
            "candidate_id": index,
            "receipt_from": first_row["receipt_no"],
            "receipt_to": last_row["receipt_no"],
            "receipt_count": len(group),
            "receipt_total": receipt_total,
            "first_record_date": str(
                first_row.get("record_date") or ""
            ),
            "last_record_date": str(
                last_row.get("record_date") or ""
            ),
            "rows": group,
        })

    return candidates


def _create_missing_che_cdm(
    receipt_from,
    receipt_to,
    deposit_date_text,
    cdm_sequence,
):
    """
    根据负责人确认的日期和 CDM 编号，
    为一个遗漏收条范围建立：
    1. finance_bank_deposits
    2. finance_bank_deposit_items

    金额只取实际收条 SUM(amount)，不接受人工金额。
    """
    ensure_finance_bank_in_tables()

    receipt_from = str(
        receipt_from or ""
    ).strip().upper()

    receipt_to = str(
        receipt_to or ""
    ).strip().upper()

    deposit_date_text = str(
        deposit_date_text or ""
    ).strip()

    try:
        deposit_date_value = datetime.strptime(
            deposit_date_text,
            "%Y-%m-%d",
        ).date()
    except ValueError:
        raise ValueError(
            "Bank In 日期格式不正确。"
        )

    try:
        cdm_sequence = int(cdm_sequence)
    except (TypeError, ValueError):
        raise ValueError(
            "CDM 编号必须是数字。"
        )

    if cdm_sequence <= 0:
        raise ValueError(
            "CDM 编号必须大于 0。"
        )

    receipt_from_number = _bank_in_receipt_number(
        receipt_from
    )

    receipt_to_number = _bank_in_receipt_number(
        receipt_to
    )

    if receipt_from_number is None:
        raise ValueError(
            f"无法识别起始收条：{receipt_from}"
        )

    if receipt_to_number is None:
        raise ValueError(
            f"无法识别结束收条：{receipt_to}"
        )

    if receipt_from_number > receipt_to_number:
        raise ValueError(
            "起始收条不可大于结束收条。"
        )

    ym = deposit_date_value.strftime("%Y-%m")
    reference_no = f"CDM-{cdm_sequence}"

    with get_conn() as conn:
        try:
            with conn.cursor(
                cursor_factory=RealDictCursor
            ) as cur:

                # 防止相同月份出现重复 CDM 编号。
                cur.execute(
                    """
                    select id
                    from finance_bank_deposits
                    where branch = 'CHE'
                      and ym = %s
                      and cdm_sequence = %s
                    limit 1
                    for update
                    """,
                    (
                        ym,
                        cdm_sequence,
                    ),
                )

                if cur.fetchone():
                    raise ValueError(
                        f"{ym} 的 {reference_no} 已经存在。"
                    )

                receipts = (
                    _find_che_cash_monthly_receipts(
                        cur,
                        receipt_from,
                        receipt_to,
                    )
                )

                if not receipts:
                    raise ValueError(
                        "这个收条范围找不到有效的 "
                        "CHE 现金月费记录。"
                    )

                actual_numbers = {
                    _bank_in_receipt_number(
                        row.get("receipt_no")
                    )
                    for row in receipts
                }

                missing_receipts = []

                for number in range(
                    receipt_from_number,
                    receipt_to_number + 1,
                ):
                    if number not in actual_numbers:
                        missing_receipts.append(
                            f"CHE{number:07d}"
                        )

                if missing_receipts:
                    raise ValueError(
                        "这个范围有缺少的收条："
                        + "、".join(missing_receipts[:10])
                        + (
                            "……"
                            if len(missing_receipts) > 10
                            else ""
                        )
                    )

                receipt_ids = [
                    row["id"]
                    for row in receipts
                ]

                # 确认这些收条还没有属于其他 CDM。
                cur.execute(
                    """
                    select
                        i.finance_record_id,
                        d.ym,
                        d.reference_no
                    from finance_bank_deposit_items i
                    join finance_bank_deposits d
                      on d.id = i.deposit_id
                    where i.finance_record_id = any(%s)
                    limit 1
                    """,
                    (receipt_ids,),
                )

                existing_item = cur.fetchone()

                if existing_item:
                    raise ValueError(
                        "这个范围内已有收条连接至 "
                        f"{existing_item.get('ym')} "
                        f"{existing_item.get('reference_no')}。"
                    )

                actual_amount = sum(
                    (
                        _bank_in_decimal(
                            row.get("amount")
                        )
                        for row in receipts
                    ),
                    Decimal("0.00"),
                )

                cur.execute(
                    """
                    insert into finance_bank_deposits
                    (
                        branch,
                        deposit_date,
                        ym,
                        fund_account,
                        bank_name,
                        reference_no,
                        amount,
                        remarks,
                        receipt_from,
                        receipt_to,
                        cdm_sequence
                    )
                    values
                    (
                        'CHE',
                        %s,
                        %s,
                        '观音堂日常户口',
                        'Hong Leong Bank',
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s
                    )
                    returning id
                    """,
                    (
                        deposit_date_value,
                        ym,
                        reference_no,
                        actual_amount,
                        (
                            "系统维修：补回遗漏 CHE "
                            "现金月费 CDM"
                        ),
                        receipt_from,
                        receipt_to,
                        cdm_sequence,
                    ),
                )

                deposit_row = cur.fetchone() or {}
                deposit_id = deposit_row.get("id")

                if not deposit_id:
                    raise ValueError(
                        "建立 CDM 主记录失败。"
                    )

                for receipt in receipts:
                    cur.execute(
                        """
                        insert into
                            finance_bank_deposit_items
                        (
                            deposit_id,
                            finance_record_id,
                            amount
                        )
                        values (%s, %s, %s)
                        """,
                        (
                            deposit_id,
                            receipt["id"],
                            _bank_in_decimal(
                                receipt.get("amount")
                            ),
                        ),
                    )

            conn.commit()

            return {
                "ym": ym,
                "reference_no": reference_no,
                "receipt_from": receipt_from,
                "receipt_to": receipt_to,
                "receipt_count": len(receipts),
                "amount": actual_amount,
            }

        except Exception:
            conn.rollback()
            raise

def _rebuild_che_bank_in_items():
    """
    只有 18/18 主记录正确且 18/18 金额验证通过时才重建。
    """
    scan = _scan_che_bank_in()

    if not scan["all_master_ok"]:
        raise ValueError("CDM 主记录尚未全部修正，不能重建。")

    if not scan["all_validation_ok"]:
        raise ValueError("仍有收条金额不符，不能重建。")

    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select
                        id,
                        ym,
                        cdm_sequence
                    from finance_bank_deposits
                    where branch = 'CHE'
                      and ym between '2026-01' and '2026-06'
                    for update
                    """
                )
                deposit_rows = cur.fetchall() or []
                deposit_by_key = {
                    (
                        str(row.get("ym") or ""),
                        int(row.get("cdm_sequence") or 0),
                    ): row["id"]
                    for row in deposit_rows
                }

                deposit_ids = list(deposit_by_key.values())

                if deposit_ids:
                    cur.execute(
                        """
                        delete from finance_bank_deposit_items
                        where deposit_id = any(%s)
                        """,
                        (deposit_ids,),
                    )

                rebuilt_count = 0

                for batch in scan["batches"]:
                    key = (batch["ym"], batch["sequence"])
                    deposit_id = deposit_by_key.get(key)

                    if not deposit_id:
                        raise ValueError(
                            f"{batch['ym']} CDM-{batch['sequence']} "
                            "缺少主记录。"
                        )

                    for receipt in batch["receipts"]:
                        # 避免依赖 ON CONFLICT 约束名称。
                        cur.execute(
                            """
                            delete from finance_bank_deposit_items
                            where finance_record_id = %s
                            """,
                            (receipt["id"],),
                        )

                        cur.execute(
                            """
                            insert into finance_bank_deposit_items
                            (
                                deposit_id,
                                finance_record_id,
                                amount
                            )
                            values (%s, %s, %s)
                            """,
                            (
                                deposit_id,
                                receipt["id"],
                                receipt["amount"],
                            ),
                        )
                        rebuilt_count += 1

            conn.commit()
            return rebuilt_count

        except Exception:
            conn.rollback()
            raise

@finance_v7_bp.route("/finance/v7/CHE/bank/cdm", methods=["GET", "POST"])
def finance_v7_cdm_management():
    """
    月费现金 CDM 管理。

    功能：
    1. 显示尚未存入银行的 CHE 现金月费收条
    2. 按收条建立 CDM Bank Deposit
    3. 自动建议本月下一个 CDM 编号
    4. CDM 编号写入 reference_no 及 cdm_sequence
    5. 不重复增加财政收入
    """
    ensure_finance_bank_in_tables()

    message = ""
    error = ""

    today = date.today()
    default_deposit_date = today.isoformat()
    default_cdm_reference = get_next_cdm_reference(today)

    form_deposit_date = (
        request.form.get("deposit_date")
        or default_deposit_date
    ).strip()

    form_cdm_reference = (
        request.form.get("cdm_reference")
        or default_cdm_reference
    ).strip()

    form_remarks = (
        request.form.get("remarks")
        or ""
    ).strip()

    if request.method == "POST":
        selected_ids = request.form.getlist("record_ids")

        deposit_date_text = form_deposit_date
        cdm_reference = normalize_cdm_reference(
            form_cdm_reference
        )
        remarks = form_remarks

        deposit_date_value = None

        try:
            deposit_date_value = datetime.strptime(
                deposit_date_text,
                "%Y-%m-%d"
            ).date()
        except ValueError:
            error = "Bank In 日期格式不正确。"

        if not error and not selected_ids:
            error = "请至少选择一张现金月费收条。"

        if not error and not cdm_reference:
            error = "CDM 编号格式不正确，请输入例如 CDM-1。"

        if not error:
            cdm_sequence_match = re.fullmatch(
                r"CDM-(\d+)",
                cdm_reference
            )

            cdm_sequence = int(
                cdm_sequence_match.group(1)
            )

            deposit_ym = deposit_date_value.strftime("%Y-%m")

            # CDM编号每个月可以重新从1开始。
            # 只拦截同月份、同分会、相同CDM编号。
            existing_reference = db_query(
                """
                select
                    id,
                    deposit_date,
                    reference_no
                from finance_bank_deposits
                where branch = 'CHE'
                  and (
                        ym = %s
                        or to_char(deposit_date, 'YYYY-MM') = %s
                  )
                  and upper(
                        replace(
                            coalesce(reference_no, ''),
                            ' ',
                            ''
                        )
                      ) = upper(%s)
                limit 1
                """,
                (
                    deposit_ym,
                    deposit_ym,
                    cdm_reference,
                ),
                fetchone=True,
            )

            if existing_reference:
                error = (
                    f"{deposit_ym} 的 {cdm_reference} "
                    "已经存在，请检查是否重复。"
                )

        if not error:
            try:
                ids = list({
                    int(value)
                    for value in selected_ids
                })

                with get_conn() as conn:
                    with conn.cursor(
                        cursor_factory=RealDictCursor
                    ) as cur:

                        # 再次检查这些记录尚未连接到任何CDM。
                        # 这一步可避免两个人同时打开页面后重复操作。
                        cur.execute(
                            """
                            select
                                r.id,
                                r.record_date,
                                r.receipt_no,
                                r.member_id,
                                r.name,
                                r.amount
                            from finance_records r
                            where r.id = any(%s)
                              and r.record_type = 'income'
                              and r.category = '月费'
                              and r.payment_method = '现金'
                              and (
                                    upper(
                                        coalesce(r.member_id, '')
                                    ) like 'CHE-%%'
                                    or upper(
                                        coalesce(r.receipt_no, '')
                                    ) like 'CHE%%'
                              )
                              and coalesce(
                                    r.status,
                                    'confirmed'
                                  ) <> 'cancelled'
                              and not exists (
                                    select 1
                                    from finance_bank_deposit_items i
                                    where i.finance_record_id = r.id
                              )
                            order by
                                r.record_date,
                                r.receipt_no,
                                r.id
                            for update
                            """,
                            (ids,),
                        )

                        rows = cur.fetchall() or []

                        if len(rows) != len(ids):
                            raise ValueError(
                                "部分收条已完成 CDM、已经作废，"
                                "或不是 CHE 现金月费。请刷新页面后重试。"
                            )

                        total = sum(
                            (
                                Decimal(
                                    str(row["amount"] or 0)
                                )
                                for row in rows
                            ),
                            Decimal("0"),
                        )

                        receipt_numbers = [
                            str(row.get("receipt_no") or "").strip()
                            for row in rows
                            if str(
                                row.get("receipt_no") or ""
                            ).strip()
                        ]

                        receipt_from = (
                            min(receipt_numbers)
                            if receipt_numbers
                            else None
                        )

                        receipt_to = (
                            max(receipt_numbers)
                            if receipt_numbers
                            else None
                        )

                        deposit_ym = (
                            deposit_date_value.strftime("%Y-%m")
                        )

                        cur.execute(
                            """
                            insert into finance_bank_deposits
                            (
                                branch,
                                deposit_date,
                                ym,
                                fund_account,
                                reference_no,
                                cdm_sequence,
                                receipt_from,
                                receipt_to,
                                amount,
                                remarks
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                '观音堂日常户口',
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s
                            )
                            returning id
                            """,
                            (
                                deposit_date_value,
                                deposit_ym,
                                cdm_reference,
                                cdm_sequence,
                                receipt_from,
                                receipt_to,
                                total,
                                remarks or "月费现金 CDM",
                            ),
                        )

                        deposit_id = cur.fetchone()["id"]

                        for row in rows:
                            cur.execute(
                                """
                                insert into finance_bank_deposit_items
                                (
                                    deposit_id,
                                    finance_record_id,
                                    amount
                                )
                                values (%s, %s, %s)
                                on conflict (finance_record_id)
                                do update set
                                    deposit_id = excluded.deposit_id,
                                    amount = excluded.amount
                                """,
                                (
                                    deposit_id,
                                    row["id"],
                                    row["amount"],
                                ),
                            )

                        # 现金存入银行只改变钱的位置。
                        # 不重复增加收入。
                        cur.execute(
                            """
                            insert into finance_cash_movements
                            (
                                branch,
                                account_type,
                                record_date,
                                movement_type,
                                amount,
                                remarks,
                                transfer_ref
                            )
                            values
                            (
                                'CHE',
                                'bank',
                                %s,
                                'bank_in',
                                %s,
                                %s,
                                %s
                            )
                            """,
                            (
                                deposit_date_value,
                                total,
                                (
                                    f"月费现金 {cdm_reference}"
                                    + (
                                        f"：{remarks}"
                                        if remarks
                                        else ""
                                    )
                                ),
                                cdm_reference,
                            ),
                        )

                    conn.commit()

                message = (
                    f"{cdm_reference} 完成："
                    f"{len(rows)} 张收条，"
                    f"共 RM {total:,.2f}。"
                )

                # 成功后建议同月份的下一编号
                default_cdm_reference = (
                    get_next_cdm_reference(
                        deposit_date_value
                    )
                )
                form_cdm_reference = default_cdm_reference
                form_remarks = ""

            except Exception as exc:
                error = str(exc)

    waiting_rows = db_query(
        """
        select
            r.id,
            r.record_date,
            r.receipt_no,
            r.member_id,
            r.name,
            r.amount
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and r.payment_method = '现金'
          and coalesce(
                r.status,
                'confirmed'
              ) <> 'cancelled'
          and r.record_date >= %s
          and (
                upper(
                    coalesce(r.member_id, '')
                ) like 'CHE-%%'
                or upper(
                    coalesce(r.receipt_no, '')
                ) like 'CHE%%'
          )
          and not exists (
                select 1
                from finance_bank_deposit_items i
                where i.finance_record_id = r.id
          )
        order by
            r.record_date,
            r.receipt_no,
            r.id
        """,
        (date(2026, 1, 1),),
        fetchall=True,
    ) or []

    waiting_total = sum(
        (
            Decimal(str(row["amount"] or 0))
            for row in waiting_rows
        ),
        Decimal("0"),
    )

    recent_deposits = db_query(
        """
        select
            d.id,
            d.deposit_date,
            d.reference_no,
            d.cdm_sequence,
            d.amount,
            d.remarks,
            d.created_at,
            count(i.id) as receipt_count,
            coalesce(
                sum(i.amount),
                0
            ) as linked_amount
        from finance_bank_deposits d
        left join finance_bank_deposit_items i
            on i.deposit_id = d.id
        where coalesce(d.branch, 'CHE') = 'CHE'
        group by
            d.id,
            d.deposit_date,
            d.reference_no,
            d.cdm_sequence,
            d.amount,
            d.remarks,
            d.created_at
        order by
            d.deposit_date desc,
            d.created_at desc
        limit 10
        """,
        fetchall=True,
    ) or []

    return render_template_string(
        """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport"
      content="width=device-width,initial-scale=1">

<title>CDM 管理</title>

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">

<style>
body{
    background:#f4f7fb;
}

.cdm-page{
    max-width:1100px;
}

.hero{
    padding:28px;
    border-radius:24px;
    background:linear-gradient(135deg,#1d4ed8,#0f766e);
    color:#fff;
    margin-bottom:20px;
}

.hero h1{
    margin:0 0 8px;
}

.summary{
    display:grid;
    grid-template-columns:repeat(2,1fr);
    gap:14px;
    margin-bottom:18px;
}

.summary .card{
    margin:0;
}

.big{
    font-size:30px;
    font-weight:900;
    margin-top:7px;
}

.receipt-list{
    display:grid;
    gap:10px;
    margin:14px 0;
}

.receipt-row{
    display:grid;
    grid-template-columns:auto 120px 150px 1fr 120px;
    gap:12px;
    align-items:center;
    padding:14px;
    border:1px solid #e5e7eb;
    border-radius:15px;
    background:#fff;
}

.amount{
    text-align:right;
    font-weight:900;
}

.form-grid{
    display:grid;
    grid-template-columns:repeat(2,1fr);
    gap:14px;
}

.full{
    grid-column:1/-1;
}

.topbar{
    display:flex;
    justify-content:space-between;
    gap:10px;
    flex-wrap:wrap;
    margin-bottom:16px;
}

.cdm-reference-wrap{
    position:relative;
}

.cdm-hint{
    margin-top:7px;
    color:#64748b;
    font-size:14px;
    line-height:1.5;
}

.deposit-ok{
    color:#166534;
    font-weight:800;
}

.deposit-diff{
    color:#b91c1c;
    font-weight:800;
}

@media(max-width:760px){
    .summary,
    .form-grid{
        grid-template-columns:1fr;
    }

    .receipt-row{
        grid-template-columns:auto 1fr;
    }

    .receipt-row > *:not(input){
        grid-column:2;
    }

    .amount{
        text-align:left;
    }

    .cdm-page{
        padding:14px;
    }
}
</style>
</head>

<body>
<div class="page cdm-page">

    <div class="topbar">
        <a class="btn-tool btn-secondary"
           href="{{ url_for('finance_v7.finance_v7_bank') }}">
            ← 返回资金中心
        </a>

        <a class="btn-tool btn-danger"
           href="{{ url_for('finance.finance_admin_logout') }}">
            退出负责人
        </a>
    </div>

    <section class="hero">
        <h1>💰 CDM 管理</h1>

        <p>
            选择已经存入银行的现金月费收条，
            登记 CDM 编号及存款日期。
            此操作只改变资金位置，不会重复增加收入。
        </p>
    </section>

    {% if message %}
    <div class="alert alert-success">
        {{ message }}
    </div>
    {% endif %}

    {% if error %}
    <div class="alert alert-danger">
        {{ error }}
    </div>
    {% endif %}

    <div class="summary">
        <div class="card">
            <div>等待 CDM 收条</div>
            <div class="big">
                {{ waiting_rows|length }} 张
            </div>
        </div>

        <div class="card">
            <div>等待存入金额</div>
            <div class="big">
                RM {{ '{:,.2f}'.format(waiting_total) }}
            </div>
        </div>
    </div>

    <form method="post" class="card">

        <div class="section-title">
            📋 选择现金月费收条
        </div>

        {% if waiting_rows %}

        <label style="
            display:inline-flex;
            gap:8px;
            align-items:center;
            font-weight:800;
        ">
            <input type="checkbox" id="selectAll">
            全选
        </label>

        <div class="receipt-list">
            {% for r in waiting_rows %}
            <label class="receipt-row">

                <input class="receipt-check"
                       type="checkbox"
                       name="record_ids"
                       value="{{ r.id }}">

                <span>{{ r.record_date }}</span>

                <span>{{ r.receipt_no or '-' }}</span>

                <span>
                    {{ r.member_id or '-' }}
                    ·
                    {{ r.name or '-' }}
                </span>

                <span class="amount">
                    RM {{ '%.2f'|format(r.amount or 0) }}
                </span>

            </label>
            {% endfor %}
        </div>

        <div class="form-grid">

            <div class="form-group">
                <label class="form-label">
                    CDM 编号
                </label>

                <div class="cdm-reference-wrap">
                    <input
                        class="form-input"
                        id="cdmReference"
                        name="cdm_reference"
                        value="{{ form_cdm_reference }}"
                        placeholder="例如：CDM-1"
                        autocomplete="off"
                        required
                    >
                </div>

                <div class="cdm-hint">
                    系统已建议本月下一个编号。
                    输入 5、CDM5 或 CDM-5 都可以。
                </div>
            </div>

            <div class="form-group">
                <label class="form-label">
                    Bank In 日期
                </label>

                <input
                    class="form-input"
                    type="date"
                    name="deposit_date"
                    value="{{ form_deposit_date }}"
                    required
                >
            </div>

            <div class="form-group full">
                <label class="form-label">
                    备注
                </label>

                <input
                    class="form-input"
                    name="remarks"
                    value="{{ form_remarks }}"
                    placeholder="可留空"
                >
            </div>

            <div class="full">
                <button
                    class="btn-tool btn-success"
                    style="width:100%"
                    type="submit"
                    onclick="
                        return confirm(
                            '确认所选收条已经存入银行吗？'
                        )
                    "
                >
                    ✅ 确认 CDM
                </button>
            </div>

        </div>

        {% else %}

        <div class="empty-state">
            目前没有等待 CDM 的现金月费收条。
        </div>

        {% endif %}

    </form>

    <section class="card">

        <div class="section-title">
            最近 CDM
        </div>

        <div class="table-responsive">
            <table class="record-table">

                <thead>
                <tr>
                    <th>日期</th>
                    <th>CDM编号</th>
                    <th>收条</th>
                    <th>存款金额</th>
                    <th>已连接金额</th>
                    <th>状态</th>
                    <th>备注</th>
                </tr>
                </thead>

                <tbody>
                {% for d in recent_deposits %}

                {% set difference =
                    (d.amount or 0)
                    -
                    (d.linked_amount or 0)
                %}

                <tr>
                    <td>{{ d.deposit_date }}</td>

                    <td>
                        <b>{{ d.reference_no or '-' }}</b>
                    </td>

                    <td>
                        {{ d.receipt_count or 0 }} 张
                    </td>

                    <td>
                        RM {{ '%.2f'|format(d.amount or 0) }}
                    </td>

                    <td>
                        RM {{ '%.2f'|format(d.linked_amount or 0) }}
                    </td>

                    <td>
                        {% if difference|abs < 0.005 %}
                        <span class="deposit-ok">
                            ✓ 已对账
                        </span>
                        {% else %}
                        <span class="deposit-diff">
                            差 RM {{ '%.2f'|format(difference) }}
                        </span>
                        {% endif %}
                    </td>

                    <td>{{ d.remarks or '-' }}</td>
                </tr>

                {% else %}

                <tr>
                    <td colspan="7">
                        尚未有 CDM 记录
                    </td>
                </tr>

                {% endfor %}
                </tbody>

            </table>
        </div>

    </section>

</div>

<script>
const selectAll =
    document.getElementById("selectAll");

if (selectAll) {
    selectAll.addEventListener(
        "change",
        function () {
            document
                .querySelectorAll(".receipt-check")
                .forEach(function (checkbox) {
                    checkbox.checked =
                        selectAll.checked;
                });
        }
    );
}

const cdmReference =
    document.getElementById("cdmReference");

if (cdmReference) {
    cdmReference.addEventListener(
        "blur",
        function () {
            let value =
                cdmReference.value
                    .trim()
                    .toUpperCase()
                    .replace(/\\s+/g, "");

            if (/^\\d+$/.test(value)) {
                value =
                    "CDM-" +
                    parseInt(value, 10);
            } else {
                const match =
                    value.match(/^CDM-?(\\d+)$/);

                if (match) {
                    value =
                        "CDM-" +
                        parseInt(
                            match[1],
                            10
                        );
                }
            }

            cdmReference.value = value;
        }
    );
}
</script>

</body>
</html>
        """,
        waiting_rows=waiting_rows,
        waiting_total=waiting_total,
        recent_deposits=recent_deposits,
        form_deposit_date=form_deposit_date,
        form_cdm_reference=form_cdm_reference,
        form_remarks=form_remarks,
        message=message,
        error=error,
    )

@finance_v7_bp.route("/finance/v7/CHE/bank/repair", methods=["GET", "POST"])
def finance_v7_repair_center():
    message = ""
    error = ""

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        try:
            if action == "repair_master":
                confirmation = (
                    request.form.get("confirmation") or ""
                ).strip().upper()

                if confirmation != "REPAIR":
                    raise ValueError("请输入 REPAIR 才能修正 CDM 主记录。")

                repaired = _repair_che_bank_in_master()
                message = f"已修正 {repaired} 笔正式 CHE CDM 主记录。"

            elif action == "update_amount":
                record_id = int(request.form.get("record_id") or 0)
                new_amount = request.form.get("new_amount")

                result = _update_bank_in_receipt_amount(
                    record_id,
                    new_amount,
                )

                message = (
                    f"已修改 {result['receipt_no']} "
                    f"（{result['name']}）："
                    f"RM {result['old_amount']:,.2f} → "
                    f"RM {result['new_amount']:,.2f}。"
                )

            elif action == "create_missing_cdm":

                confirmation = (
                    request.form.get("confirmation")
                    or ""
                ).strip().upper()

                if confirmation != "CREATE":
                    raise ValueError(
                        "请输入 CREATE 才能建立遗漏 CDM。"
                    )

                result = _create_missing_che_cdm(
                    receipt_from=request.form.get(
                        "receipt_from"
                    ),
                    receipt_to=request.form.get(
                        "receipt_to"
                    ),
                    deposit_date_text=request.form.get(
                        "deposit_date"
                    ),
                    cdm_sequence=request.form.get(
                        "cdm_sequence"
                    ),
                )

                message = (
                    f"已建立 {result['ym']} "
                    f"{result['reference_no']}："
                    f"{result['receipt_from']} 至 "
                    f"{result['receipt_to']}，"
                    f"共 {result['receipt_count']} 张，"
                    f"RM {result['amount']:,.2f}。"
                )

            elif action == "rebuild_items":
                confirmation = (
                    request.form.get("confirmation") or ""
                ).strip().upper()

                if confirmation != "REBUILD":
                    raise ValueError("请输入 REBUILD 才能重建 Bank In 明细。")

                rebuilt = _rebuild_che_bank_in_items()
                message = f"重建完成，共建立 {rebuilt} 条 Bank In 明细。"

            else:
                raise ValueError("无法识别这次操作。")

        except Exception as exc:
            error = str(exc)

    status = _scan_che_bank_in()

    missing_candidates = (
        _find_missing_che_cdm_candidates()
    )

    return render_template_string(
        """
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta
                name="viewport"
                content="width=device-width, initial-scale=1"
            >
            <title>Bank In 数据维修中心</title>

            <link
                rel="stylesheet"
                href="{{ url_for('static', filename='css/toolbox.css') }}"
            >

            <style>
                body{
                    background:#f4f7fb;
                }

                .repair-page{
                    max-width:1180px;
                    margin:0 auto;
                    padding:22px;
                }

                .repair-hero,
                .repair-card{
                    background:#fff;
                    border:1px solid #e2e8f0;
                    border-radius:22px;
                    padding:22px;
                    margin-bottom:18px;
                    box-shadow:0 6px 20px rgba(15,23,42,.06);
                }

                .repair-hero{
                    background:linear-gradient(135deg,#fff8e8,#fff);
                }

                .repair-title{
                    margin:0 0 8px;
                    font-size:30px;
                }

                .repair-subtitle{
                    color:#64748b;
                    line-height:1.65;
                }

                .status-grid{
                    display:grid;
                    grid-template-columns:repeat(4,minmax(0,1fr));
                    gap:12px;
                    margin-top:18px;
                }

                .status-box{
                    padding:15px;
                    border:1px solid #e2e8f0;
                    border-radius:16px;
                    background:#fff;
                }

                .status-label{
                    color:#64748b;
                    font-size:14px;
                }

                .status-value{
                    margin-top:5px;
                    font-size:24px;
                    font-weight:900;
                }

                .steps{
                    display:grid;
                    grid-template-columns:repeat(3,minmax(0,1fr));
                    gap:14px;
                }

                .step-card{
                    border:1px solid #e2e8f0;
                    border-radius:17px;
                    padding:17px;
                    background:#fff;
                }

                .step-no{
                    display:inline-flex;
                    width:34px;
                    height:34px;
                    align-items:center;
                    justify-content:center;
                    border-radius:50%;
                    background:#eef2ff;
                    font-weight:900;
                    margin-right:6px;
                }

                .notice{
                    padding:14px 16px;
                    border-radius:14px;
                    margin-bottom:16px;
                    line-height:1.6;
                }

                .notice-success{
                    background:#ecfdf3;
                    border:1px solid #a7e3b6;
                    color:#166534;
                }

                .notice-error{
                    background:#fff1f2;
                    border:1px solid #fecdd3;
                    color:#9f1239;
                }

                .batch-card{
                    border:1px solid #e2e8f0;
                    border-radius:17px;
                    margin-bottom:14px;
                    overflow:hidden;
                }

                .batch-head{
                    padding:15px 17px;
                    background:#f8fafc;
                    display:flex;
                    justify-content:space-between;
                    gap:12px;
                    flex-wrap:wrap;
                    align-items:center;
                }

                .batch-title{
                    font-size:18px;
                    font-weight:900;
                }

                .badge{
                    display:inline-block;
                    padding:5px 9px;
                    border-radius:999px;
                    font-size:13px;
                    font-weight:800;
                    margin-left:5px;
                }

                .badge-ok{
                    background:#dcfce7;
                    color:#166534;
                }

                .badge-bad{
                    background:#fee2e2;
                    color:#991b1b;
                }

                .batch-body{
                    padding:16px;
                }

                .issue{
                    color:#b42318;
                    margin:4px 0;
                    font-weight:700;
                }

                .receipt-table-wrap{
                    overflow:auto;
                    margin-top:12px;
                }

                .receipt-table{
                    width:100%;
                    min-width:760px;
                    border-collapse:collapse;
                }

                .receipt-table th,
                .receipt-table td{
                    padding:10px 8px;
                    border-bottom:1px solid #edf0f3;
                    text-align:left;
                    vertical-align:middle;
                }

                .amount-form{
                    display:flex;
                    gap:7px;
                    align-items:center;
                }

                .amount-input,
                .confirm-input{
                    min-height:42px;
                    padding:8px 11px;
                    border:1px solid #cbd5e1;
                    border-radius:10px;
                    font-size:16px;
                }

                .amount-input{
                    width:115px;
                }

                .btn-small,
                .btn-action,
                .btn-danger{
                    border:0;
                    border-radius:10px;
                    padding:10px 14px;
                    font-weight:900;
                    cursor:pointer;
                }

                .btn-small{
                    background:#e0e7ff;
                    color:#3730a3;
                    white-space:nowrap;
                }

                .btn-action{
                    background:#2563eb;
                    color:#fff;
                }

                .btn-danger{
                    background:#b42318;
                    color:#fff;
                }

                .btn-disabled{
                    background:#cbd5e1;
                    color:#64748b;
                    cursor:not-allowed;
                }

                .form-row{
                    display:flex;
                    gap:10px;
                    flex-wrap:wrap;
                    align-items:center;
                    margin-top:12px;
                }

                .back-link{
                    display:inline-block;
                    margin-bottom:14px;
                }

                .summary-line{
                    font-size:17px;
                    font-weight:800;
                    margin-top:8px;
                }

                @media(max-width:820px){
                    .status-grid{
                        grid-template-columns:repeat(2,minmax(0,1fr));
                    }

                    .steps{
                        grid-template-columns:1fr;
                    }
                }
            </style>
        </head>

        <body>
        <div class="repair-page">

            <a
                class="back-link"
                href="{{ url_for('finance_v7.finance_v7_bank') }}"
            >
                ← 返回银行专区
            </a>

            <section class="repair-hero">
                <h1 class="repair-title">
                    🔧 Bank In 数据维修中心
                </h1>

                <div class="repair-subtitle">
                    分三个阶段处理：先修正 CDM 主记录，
                    再逐张验证收条，最后才重建 Bank In 明细。
                    扫描本身不会修改数据库。
                </div>

                <div class="status-grid">
                    <div class="status-box">
                        <div class="status-label">CDM 主记录正确</div>
                        <div class="status-value">
                            {{ status.master_ok_count }}
                            / {{ status.official_count }}
                        </div>
                    </div>

                    <div class="status-box">
                        <div class="status-label">收条验证通过</div>
                        <div class="status-value">
                            {{ status.validation_ok_count }}
                            / {{ status.official_count }}
                        </div>
                    </div>

                    <div class="status-box">
                        <div class="status-label">Bank In 明细</div>
                        <div class="status-value">
                            {{ status.item_count }}
                        </div>
                    </div>

                    <div class="status-box">
                        <div class="status-label">等待 Bank In</div>
                        <div class="status-value">
                            {{ status.waiting_count }} 张
                            <div style="font-size:14px;color:#64748b;">
                                RM {{
                                    "{:,.2f}".format(
                                        status.waiting_total
                                    )
                                }}
                            </div>
                        </div>
                    </div>
                </div>
            </section>

            {% if message %}
                <div class="notice notice-success">
                    {{ message }}
                </div>
            {% endif %}

            {% if error %}
                <div class="notice notice-error">
                    {{ error }}
                </div>
            {% endif %}

            <section class="repair-card">
                <div class="steps">
                    <div class="step-card">
                        <h3>
                            <span class="step-no">1</span>
                            修正 CDM 主记录
                        </h3>
                        <p>
                            只修正 finance_bank_deposits，
                            不修改收条金额，也不建立明细。
                        </p>

                        <form
                            method="post"
                            onsubmit="
                                return confirm(
                                    '确定修正 18 笔正式 CHE CDM 吗？'
                                );
                            "
                        >
                            <input
                                type="hidden"
                                name="action"
                                value="repair_master"
                            >

                            <div class="form-row">
                                <input
                                    class="confirm-input"
                                    name="confirmation"
                                    placeholder="输入 REPAIR"
                                    autocomplete="off"
                                >

                                <button class="btn-action" type="submit">
                                    修正 CDM 主记录
                                </button>
                            </div>
                        </form>
                    </div>

                    <div class="step-card">
                        <h3>
                            <span class="step-no">2</span>
                            验证及修改收条
                        </h3>
                        <p>
                            下方逐张显示收条金额。
                            必须根据纸本收条确认后再修改，
                            系统不会自动猜测哪一张错误。
                        </p>

                        <div class="summary-line">
                            已通过：
                            {{ status.validation_ok_count }}
                            / {{ status.official_count }}
                        </div>
                    </div>

                    <div class="step-card">
                        <h3>
                            <span class="step-no">3</span>
                            重建 Bank In 明细
                        </h3>

                        {% if status.can_rebuild %}
                            <p class="notice notice-success">
                                18 / 18 主记录和收条金额全部通过，
                                现在可以安全重建。
                            </p>

                            <form
                                method="post"
                                onsubmit="
                                    return confirm(
                                        '确定重建 Bank In 明细吗？'
                                    );
                                "
                            >
                                <input
                                    type="hidden"
                                    name="action"
                                    value="rebuild_items"
                                >

                                <div class="form-row">
                                    <input
                                        class="confirm-input"
                                        name="confirmation"
                                        placeholder="输入 REBUILD"
                                        autocomplete="off"
                                    >

                                    <button
                                        class="btn-danger"
                                        type="submit"
                                    >
                                        重建 Bank In 明细
                                    </button>
                                </div>
                            </form>
                        {% else %}
                            <p class="notice notice-error">
                                尚未达到 18 / 18。
                                请先修正 CDM 和错误收条。
                            </p>

                            <button
                                class="btn-danger btn-disabled"
                                type="button"
                                disabled
                            >
                                暂时不能重建
                            </button>
                        {% endif %}
                    </div>
                </div>
            </section>

            {% if status.extra_master_rows %}
                <div class="notice notice-error">
                    检测到 {{ status.extra_master_rows|length }}
                    笔不属于正式清单的 CHE CDM。
                    执行第一步时会移除这些错误主记录。
                </div>
            {% endif %}

            <section class="repair-card">

                <h2>🔎 自动找到的遗漏 CDM 候选</h2>

                <p class="repair-subtitle">
                    系统只负责寻找尚未连接 Bank In
                    的连续收条。Bank In 日期及 CDM 编号
                    必须根据纸本或银行记录确认，系统不会猜测。
                </p>

                {% if missing_candidates %}

                    {% for candidate in missing_candidates %}

                        <div class="batch-card">

                            <div class="batch-head">

                                <div>
                                    <span class="batch-title">
                                        {{ candidate.receipt_from }}
                                        至
                                        {{ candidate.receipt_to }}
                                    </span>

                                    <span class="badge badge-bad">
                                        尚未建立 CDM
                                    </span>
                                </div>

                                <div>
                                    {{ candidate.receipt_count }} 张
                                    ·
                                    RM {{
                                        "{:,.2f}".format(
                                            candidate.receipt_total
                                        )
                                    }}
                                </div>

                            </div>

                            <div class="batch-body">

                                <div class="summary-line">
                                    收条日期：
                                    {{ candidate.first_record_date }}
                                    至
                                    {{ candidate.last_record_date }}
                                </div>

                                <form
                                    method="post"
                                    onsubmit="
                                        return confirm(
                                            '确定根据填写的日期和编号建立这笔 CDM 吗？'
                                        );
                                    "
                                >
                                    <input
                                        type="hidden"
                                        name="action"
                                        value="create_missing_cdm"
                                    >

                                    <input
                                        type="hidden"
                                        name="receipt_from"
                                        value="{{ candidate.receipt_from }}"
                                    >

                                    <input
                                        type="hidden"
                                        name="receipt_to"
                                        value="{{ candidate.receipt_to }}"
                                    >

                                    <div class="form-row">

                                        <input
                                            class="confirm-input"
                                            type="date"
                                            name="deposit_date"
                                            required
                                        >

                                        <input
                                            class="confirm-input"
                                            type="number"
                                            name="cdm_sequence"
                                            min="1"
                                            placeholder="CDM 编号，例如 5"
                                            required
                                        >

                                        <input
                                            class="confirm-input"
                                            name="confirmation"
                                            placeholder="输入 CREATE"
                                            autocomplete="off"
                                            required
                                        >

                                        <button
                                            class="btn-action"
                                            type="submit"
                                        >
                                            建立这笔遗漏 CDM
                                        </button>

                                    </div>

                                </form>

                            </div>

                        </div>

                    {% endfor %}

                {% else %}

                    <p class="notice notice-success">
                        2026 年 1 月至 6 月没有发现遗漏的
                        CHE 现金月费 CDM。
                    </p>

                {% endif %}

            </section>

            <section class="repair-card">
                <h2>逐批扫描与收条明细</h2>

                {% for batch in status.batches %}
                    <div class="batch-card">
                        <div class="batch-head">
                            <div>
                                <span class="batch-title">
                                    {{ batch.ym }}
                                    CDM-{{ batch.sequence }}
                                </span>

                                {% if batch.master_ok %}
                                    <span class="badge badge-ok">
                                        主记录正确
                                    </span>
                                {% else %}
                                    <span class="badge badge-bad">
                                        主记录待修
                                    </span>
                                {% endif %}

                                {% if batch.validation_ok %}
                                    <span class="badge badge-ok">
                                        金额通过
                                    </span>
                                {% else %}
                                    <span class="badge badge-bad">
                                        金额不符
                                    </span>
                                {% endif %}
                            </div>

                            <div>
                                正式 RM {{
                                    "{:,.2f}".format(
                                        batch.expected_amount
                                    )
                                }}
                                ／
                                收条 RM {{
                                    "{:,.2f}".format(
                                        batch.receipt_total
                                    )
                                }}
                            </div>
                        </div>

                        <div class="batch-body">
                            <div>
                                正式范围：
                                <strong>
                                    {{ batch.expected_from }}
                                    至
                                    {{ batch.expected_to }}
                                </strong>
                                ｜ Bank In：
                                {{ batch.deposit_date }}
                            </div>

                            {% for issue in batch.master_issues %}
                                <div class="issue">
                                    主记录：{{ issue }}
                                </div>
                            {% endfor %}

                            {% for issue in batch.validation_issues %}
                                <div class="issue">
                                    收条：{{ issue }}
                                </div>
                            {% endfor %}

                            <div class="receipt-table-wrap">
                                <table class="receipt-table">
                                    <thead>
                                        <tr>
                                            <th>收条编号</th>
                                            <th>日期</th>
                                            <th>月费编号</th>
                                            <th>姓名</th>
                                            <th>目前金额</th>
                                            <th>根据纸本修正</th>
                                        </tr>
                                    </thead>

                                    <tbody>
                                    {% if batch.receipts %}
                                        {% for receipt in batch.receipts %}
                                            <tr>
                                                <td>
                                                    <strong>
                                                        {{ receipt.receipt_no }}
                                                    </strong>
                                                </td>
                                                <td>
                                                    {{
                                                        receipt.receipt_date
                                                        or receipt.record_date
                                                        or "-"
                                                    }}
                                                </td>
                                                <td>
                                                    {{ receipt.member_id or "-" }}
                                                </td>
                                                <td>
                                                    {{ receipt.name or "-" }}
                                                </td>
                                                <td>
                                                    RM {{
                                                        "{:,.2f}".format(
                                                            receipt.amount
                                                        )
                                                    }}
                                                </td>
                                                <td>
                                                    <form
                                                        class="amount-form"
                                                        method="post"
                                                        onsubmit="
                                                            return confirm(
                                                                '确定根据纸本收条修改这笔金额吗？'
                                                            );
                                                        "
                                                    >
                                                        <input
                                                            type="hidden"
                                                            name="action"
                                                            value="update_amount"
                                                        >
                                                        <input
                                                            type="hidden"
                                                            name="record_id"
                                                            value="{{ receipt.id }}"
                                                        >
                                                        <input
                                                            class="amount-input"
                                                            type="number"
                                                            name="new_amount"
                                                            min="0.01"
                                                            step="0.01"
                                                            value="{{ receipt.amount }}"
                                                            required
                                                        >
                                                        <button
                                                            class="btn-small"
                                                            type="submit"
                                                        >
                                                            保存
                                                        </button>
                                                    </form>
                                                </td>
                                            </tr>
                                        {% endfor %}
                                    {% else %}
                                        <tr>
                                            <td colspan="6">
                                                找不到这个正式范围内的 CHE 现金月费收条。
                                            </td>
                                        </tr>
                                    {% endif %}
                                    </tbody>
                                </table>
                            </div>
                        </div>
                    </div>
                {% endfor %}
            </section>

        </div>
        </body>
        </html>
        """,
        status=status,
        missing_candidates=missing_candidates,
        message=message,
        error=error,
    )

@finance_v7_bp.route("/finance/v7/CHE/bank/cash-transfer", methods=["GET", "POST"])
def finance_v7_cash_transfer():

    message = ""
    error = ""

    record_date = request.form.get(
        "record_date",
        date.today().isoformat()
    ).strip()

    amount_text = request.form.get(
        "amount",
        "10000.00"
    ).strip()

    remarks = request.form.get(
        "remarks",
        "银行提款转入 Petty Cash"
    ).strip()

    if request.method == "POST":

        try:
            amount = Decimal(
                amount_text.replace(",", "")
            )

            if amount <= 0:
                raise ValueError(
                    "提款金额必须大于零"
                )

            transfer_ref = (
                "CASH-"
                + uuid.uuid4().hex[:12].upper()
            )

            with get_conn() as conn:
                with conn.cursor() as cur:

                    # 银行减少
                    cur.execute("""
                        insert into finance_cash_movements
                        (
                            branch,
                            account_type,
                            record_date,
                            movement_type,
                            amount,
                            remarks,
                            transfer_ref
                        )
                        values
                        (
                            %s,
                            'bank',
                            %s,
                            'cash_out',
                            %s,
                            %s,
                            %s
                        )
                    """, (
                        "CHE",
                        record_date,
                        amount,
                        remarks,
                        transfer_ref,
                    ))

                    # 现金增加
                    cur.execute("""
                        insert into finance_cash_movements
                        (
                            branch,
                            account_type,
                            record_date,
                            movement_type,
                            amount,
                            remarks,
                            transfer_ref
                        )
                        values
                        (
                            %s,
                            'cash',
                            %s,
                            'cash_in',
                            %s,
                            %s,
                            %s
                        )
                    """, (
                        "CHE",
                        record_date,
                        amount,
                        "GYT Cash In：" + remarks,
                        transfer_ref,
                    ))

                conn.commit()

            message = (
                f"提款记录完成：RM {amount:,.2f}。"
                "银行余额已减少，Petty Cash 已增加。"
            )

            amount_text = "10000.00"

        except Exception as e:
            error = str(e)

    balance = get_finance_balance_summary(
        branch="CHE"
    )

    return render_template_string(
        FINANCE_DATE_COMPONENT + """
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">
    <meta
        name="viewport"
        content="width=device-width, initial-scale=1"
    >
    <title>银行提款到 Petty Cash</title>

    <link
        rel="stylesheet"
        href="{{ url_for('static', filename='css/toolbox.css') }}"
    >

    <style>
        .cash-page{
            max-width:900px;
            margin:0 auto;
            padding:24px;
        }

        .cash-hero{
            padding:28px;
            border-radius:24px;
            background:
                linear-gradient(
                    135deg,
                    #0f766e,
                    #0891b2
                );
            color:white;
            margin-bottom:22px;
            box-shadow:
                0 18px 40px rgba(15,118,110,.20);
        }

        .cash-hero h1{
            margin:0 0 8px;
            font-size:34px;
        }

        .cash-hero p{
            margin:0;
            opacity:.9;
            font-size:18px;
        }

        .balance-grid{
            display:grid;
            grid-template-columns:
                repeat(2, minmax(0, 1fr));
            gap:16px;
            margin-bottom:22px;
        }

        .balance-card{
            background:white;
            border-radius:22px;
            padding:24px;
            border:1px solid #e5e7eb;
            box-shadow:
                0 10px 28px rgba(15,23,42,.07);
        }

        .balance-label{
            font-size:17px;
            color:#64748b;
            font-weight:700;
        }

        .balance-value{
            margin-top:10px;
            font-size:32px;
            font-weight:900;
            color:#0f172a;
        }

        .cash-value{
            color:#d97706;
        }

        .form-card{
            background:white;
            border-radius:22px;
            padding:26px;
            border:1px solid #e5e7eb;
            box-shadow:
                0 10px 28px rgba(15,23,42,.07);
        }

        .flow-box{
            display:grid;
            grid-template-columns:
                1fr auto 1fr;
            align-items:center;
            gap:14px;
            padding:18px;
            background:#f8fafc;
            border-radius:18px;
            margin-bottom:22px;
            text-align:center;
        }

        .flow-account{
            padding:15px;
            border-radius:16px;
            background:white;
            font-weight:800;
        }

        .flow-arrow{
            font-size:30px;
        }

        .cash-form-grid{
            display:grid;
            grid-template-columns:
                repeat(2, minmax(0, 1fr));
            gap:16px;
        }

        .full-row{
            grid-column:1 / -1;
        }

        .submit-btn{
            width:100%;
            min-height:62px;
            border:none;
            border-radius:17px;
            font-size:21px;
            font-weight:900;
            cursor:pointer;
            background:
                linear-gradient(
                    135deg,
                    #f59e0b,
                    #ea580c
                );
            color:white;
        }

        @media(max-width:700px){
            .balance-grid,
            .cash-form-grid{
                grid-template-columns:1fr;
            }

            .flow-box{
                grid-template-columns:1fr;
            }

            .flow-arrow{
                transform:rotate(90deg);
            }

            .cash-page{
                padding:14px;
            }
        }
    </style>
</head>

<body>
<div class="cash-page">

    <div class="cash-hero">
        <h1>💵 银行提款到 Petty Cash</h1>
        <p>
            财政只需输入一次，系统自动完成银行扣款与现金增加。
        </p>
    </div>

    {% if message %}
        <div class="alert alert-success">
            {{ message }}
        </div>
    {% endif %}

    {% if error %}
        <div class="alert alert-danger">
            {{ error }}
        </div>
    {% endif %}

    <div class="balance-grid">

        <div class="balance-card">
            <div class="balance-label">
                🏦 CHE 银行余额
            </div>

            <div class="balance-value">
                RM {{ "{:,.2f}".format(balance.bank_balance) }}
            </div>
        </div>

        <div class="balance-card">
            <div class="balance-label">
                💵 Cash In Hand
            </div>

            <div class="balance-value cash-value">
                RM {{ "{:,.2f}".format(balance.cash_in_hand) }}
            </div>
        </div>

    </div>

    <div class="form-card">

        <div class="flow-box">
            <div class="flow-account">
                🏦 CHE 银行户口
            </div>

            <div class="flow-arrow">
                ➜
            </div>

            <div class="flow-account">
                💵 Petty Cash
            </div>
        </div>

        <form method="post">

            <div class="cash-form-grid">

                <div class="form-group">
                    <label class="form-label">
                        提款日期
                    </label>

                    <input
                        type="date"
                        name="record_date"
                        class="form-input"
                        value="{{ record_date }}"
                        required
                    >
                </div>

                <div class="form-group">
                    <label class="form-label">
                        提款金额
                    </label>

                    <input
                        type="number"
                        name="amount"
                        class="form-input"
                        step="0.01"
                        min="0.01"
                        value="{{ amount_text }}"
                        required
                    >
                </div>

                <div class="form-group full-row">
                    <label class="form-label">
                        备注
                    </label>

                    <input
                        type="text"
                        name="remarks"
                        class="form-input"
                        value="{{ remarks }}"
                    >
                </div>

                <div class="full-row">
                    <button
                        type="submit"
                        class="submit-btn"
                        onclick="
                            return confirm(
                                '确认这笔银行提款吗？'
                            );
                        "
                    >
                        💵 确认提款并转入 Petty Cash
                    </button>
                </div>

            </div>

        </form>

    </div>

    <div style="margin-top:18px;">
        <a
            href="{{ url_for('finance_v7.finance_v7_che_home') }}"
            class="btn-tool btn-secondary"
        >
            ← 返回财政负责人中心
        </a>
    </div>

</div>
</body>
</html>
        """,
        message=message,
        error=error,
        record_date=record_date,
        amount_text=amount_text,
        remarks=remarks,
        balance=balance,
    )

@finance_v7_bp.route("/finance/v7/CHE/bank/cash-in-hand")
def finance_v7_cash_in_hand():

    branch = "CHE"

    balance = get_finance_balance_summary(
        branch=branch
    )

    # 1. 读取期初、提款及调整
    movement_rows = db_query("""
        select
            id,
            record_date,
            movement_type,
            amount,
            remarks,
            account_type,
            transfer_ref
        from finance_cash_movements
        where branch = %s
          and account_type = 'cash'
        order by record_date, id
    """, (
        branch,
    ), fetchall=True)

    ledger_rows = []

    for row in movement_rows:

        movement_type = row["movement_type"]
        amount = Decimal(str(row["amount"] or 0))

        if movement_type == "opening":
            money_in = amount
            money_out = Decimal("0")
            itemised = row["remarks"] or "期初现金"

        elif movement_type == "cash_in":
            money_in = amount
            money_out = Decimal("0")
            itemised = row["remarks"] or "GYT Cash In"

        elif movement_type == "adjustment":

            if amount >= 0:
                money_in = amount
                money_out = Decimal("0")
            else:
                money_in = Decimal("0")
                money_out = abs(amount)

            itemised = row["remarks"] or "现金调整"

        else:
            continue

        ledger_rows.append({
            "sort_date": row["record_date"],
            "sort_order": 1,
            "source": "movement",
            "source_id": row["id"],
            "record_date": row["record_date"],
            "money_in": money_in,
            "money_out": money_out,
            "itemised": itemised,
            "reference_no": row.get("transfer_ref") or "",
        })

    # 2. 所有支出都是 Petty Cash 支出
    expense_rows = db_query("""
        select
            id,
            record_date,
            payment_voucher_no,
            category,
            sub_category,
            name,
            amount,
            remarks
        from finance_records
        where record_type = 'expense'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and record_date >= %s
        order by record_date, id
    """, (
        FINANCE_OPENING_DATE,
    ), fetchall=True)

    for row in expense_rows:

        itemised_parts = []

        if row.get("category"):
            itemised_parts.append(str(row["category"]))

        if (
            row.get("sub_category")
            and row["sub_category"] != row.get("category")
        ):
            itemised_parts.append(str(row["sub_category"]))

        if row.get("name"):
            itemised_parts.append(str(row["name"]))

        if row.get("remarks"):
            itemised_parts.append(str(row["remarks"]))

        itemised = " / ".join(itemised_parts)

        if not itemised:
            itemised = "现金支出"

        ledger_rows.append({
            "sort_date": row["record_date"],
            "sort_order": 2,
            "source": "expense",
            "source_id": row["id"],
            "record_date": row["record_date"],
            "money_in": Decimal("0"),
            "money_out": Decimal(str(row["amount"] or 0)),
            "itemised": itemised,
            "reference_no": (
                row.get("payment_voucher_no") or ""
            ),
        })

    # 3. 按日期排列并计算滚动余额
    ledger_rows.sort(
        key=lambda r: (
            r["sort_date"],
            r["sort_order"],
            r["source_id"],
        )
    )

    running_balance = Decimal("0")

    for row in ledger_rows:
        running_balance += row["money_in"]
        running_balance -= row["money_out"]
        row["running_balance"] = running_balance

    # 页面显示最近记录在上面
    display_rows = list(reversed(ledger_rows))

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">

    <meta
        name="viewport"
        content="width=device-width, initial-scale=1"
    >

    <title>Cash In Hand</title>

    <link
        rel="stylesheet"
        href="{{ url_for('static', filename='css/toolbox.css') }}"
    >

    <style>
        body{
            background:#f3f6fb;
        }

        .cash-page{
            max-width:1400px;
            margin:0 auto;
            padding:24px;
        }

        .cash-hero{
            border-radius:26px;
            padding:30px;
            color:white;
            background:
                linear-gradient(
                    135deg,
                    #b45309,
                    #f59e0b
                );
            box-shadow:
                0 18px 42px rgba(180,83,9,.20);
            margin-bottom:22px;
        }

        .cash-hero h1{
            margin:0;
            font-size:38px;
            font-weight:900;
        }

        .cash-hero p{
            margin:10px 0 0;
            font-size:18px;
            opacity:.92;
        }

        .balance-grid{
            display:grid;
            grid-template-columns:
                repeat(4, minmax(0, 1fr));
            gap:16px;
            margin-bottom:22px;
        }

        .balance-card{
            background:white;
            border-radius:22px;
            padding:24px;
            border:1px solid #e2e8f0;
            box-shadow:
                0 10px 28px rgba(15,23,42,.07);
        }

        .balance-label{
            color:#64748b;
            font-size:17px;
            font-weight:800;
        }

        .balance-value{
            margin-top:10px;
            font-size:30px;
            font-weight:900;
            color:#0f172a;
        }

        .balance-value.cash{
            color:#c2410c;
        }

        .action-row{
            display:flex;
            gap:14px;
            flex-wrap:wrap;
            margin-bottom:22px;
        }

        .action-row .btn-tool{
            min-height:58px;
            font-size:19px;
        }

        .ledger-card{
            background:white;
            border-radius:24px;
            padding:26px;
            border:1px solid #e2e8f0;
            box-shadow:
                0 12px 32px rgba(15,23,42,.07);
        }

        .ledger-head{
            display:flex;
            justify-content:space-between;
            align-items:center;
            gap:16px;
            margin-bottom:18px;
        }

        .ledger-title{
            margin:0;
            font-size:30px;
            font-weight:900;
        }

        .cash-table{
            width:100%;
            border-collapse:collapse;
            min-width:950px;
        }

        .cash-table th{
            padding:15px 12px;
            background:#f8fafc;
            color:#334155;
            border-bottom:2px solid #e2e8f0;
            font-size:16px;
            text-align:left;
        }

        .cash-table td{
            padding:15px 12px;
            border-bottom:1px solid #e5e7eb;
            font-size:16px;
            vertical-align:top;
        }

        .money-in{
            color:#15803d;
            font-weight:900;
            white-space:nowrap;
        }

        .money-out{
            color:#dc2626;
            font-weight:900;
            white-space:nowrap;
        }

        .running{
            color:#1d4ed8;
            font-weight:900;
            white-space:nowrap;
        }

        .reference{
            color:#64748b;
            font-size:14px;
            margin-top:5px;
        }

        .empty-box{
            padding:40px;
            text-align:center;
            color:#64748b;
            font-size:18px;
        }

        @media(max-width:900px){
            .balance-grid{
                grid-template-columns:
                    repeat(2, minmax(0, 1fr));
            }
        }

        @media(max-width:600px){
            .cash-page{
                padding:14px;
            }

            .balance-grid{
                grid-template-columns:1fr;
            }

            .cash-hero h1{
                font-size:30px;
            }
        }
    </style>
</head>

<body>
<div class="cash-page">

    <section class="cash-hero">
        <h1>💵 Cash In Hand</h1>

        <p>
            银行提款进入 Petty Cash，所有观音堂支出只从 Petty Cash 扣除。
        </p>
    </section>

    <section class="balance-grid">

        <div class="balance-card">
            <div class="balance-label">
                💵 当前 Cash In Hand
            </div>

            <div class="balance-value cash">
                RM {{ "{:,.2f}".format(balance.cash_in_hand) }}
            </div>
        </div>

        <div class="balance-card">
            <div class="balance-label">
                期初现金
            </div>

            <div class="balance-value">
                RM {{ "{:,.2f}".format(balance.cash_opening) }}
            </div>
        </div>

        <div class="balance-card">
            <div class="balance-label">
                GYT Cash In
            </div>

            <div class="balance-value">
                RM {{ "{:,.2f}".format(balance.petty_cash_in) }}
            </div>
        </div>

        <div class="balance-card">
            <div class="balance-label">
                现金支出
            </div>

            <div class="balance-value">
                RM {{ "{:,.2f}".format(balance.cash_expense) }}
            </div>
        </div>

    </section>

    <div class="action-row">

        <a
            href="{{ url_for('finance.finance_v7_cash_transfer') }}"
            class="btn-tool btn-warning"
        >
            ＋ GYT Cash In／银行提款
        </a>

        <a
            href="{{ url_for('finance.records') }}"
            class="btn-tool btn-primary"
        >
            📋 查看财政记录
        </a>

        <a
            href="{{ url_for('finance_v7.finance_v7_che_home') }}"
            class="btn-tool btn-secondary"
        >
            ← 返回负责人中心
        </a>

    </div>

    <section class="ledger-card">

        <div class="ledger-head">
            <h2 class="ledger-title">
                📒 Petty Cash Ledger
            </h2>

            <div>
                共 {{ display_rows|length }} 笔
            </div>
        </div>

        <div class="table-responsive">

            {% if display_rows %}

            <table class="cash-table">
                <thead>
                    <tr>
                        <th>日期</th>
                        <th>Cash In</th>
                        <th>Cash Out</th>
                        <th>余额</th>
                        <th>项目</th>
                    </tr>
                </thead>

                <tbody>
                {% for row in display_rows %}
                    <tr>
                        <td>
                            {{ row.record_date.strftime("%d/%m/%Y") }}
                        </td>

                        <td class="money-in">
                            {% if row.money_in %}
                                RM {{ "{:,.2f}".format(row.money_in) }}
                            {% else %}
                                -
                            {% endif %}
                        </td>

                        <td class="money-out">
                            {% if row.money_out %}
                                RM {{ "{:,.2f}".format(row.money_out) }}
                            {% else %}
                                -
                            {% endif %}
                        </td>

                        <td class="running">
                            RM {{
                                "{:,.2f}".format(
                                    row.running_balance
                                )
                            }}
                        </td>

                        <td>
                            <strong>
                                {{ row.itemised }}
                            </strong>

                            {% if row.reference_no %}
                                <div class="reference">
                                    {{ row.reference_no }}
                                </div>
                            {% endif %}
                        </td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>

            {% else %}

                <div class="empty-box">
                    暂时没有 Petty Cash 记录。
                </div>

            {% endif %}

        </div>

    </section>

</div>
</body>
</html>
    """,
        balance=balance,
        display_rows=display_rows,
    )
