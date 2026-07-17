"""Finance V6 history Excel import centre.

Supported source formats (first production version):
- excel format.xlsx
  * monthly fee sheets
  * 初一十五膳食结缘
  * 观音村善款
  * 善款收纳表
  * CDM / cash bank-in groups
- 00-PETTY CASH GYT.xlsx
  * 2026 Petty Cash Manual Book expenses

Upload is always staged first. Nothing enters finance_records until Confirm.
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from flask import (
    Blueprint,
    flash,
    redirect,
    render_template_string,
    request,
    session,
    url_for,
)
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from db import db_query, get_conn
from finance_audit import write_finance_audit
from finance_common import (
    get_current_finance_branch,
    get_current_finance_user,
    malaysia_today,
    money,
    require_finance_month_open,
)

finance_import_bp = Blueprint(
    "finance_import",
    __name__,
    url_prefix="/finance/import",
)

ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}
MAX_IMPORT_BYTES = 12 * 1024 * 1024

STATUS_NOTE_KEYWORDS = (
    "cancel", "cancelled", "canceled", "stop from", "stopped",
    "停供", "暂停", "暫停", "取消", "终止", "終止",
)


def _contains_status_note(*values: Any) -> bool:
    text = " ".join(_text(v).lower() for v in values if _text(v))
    return any(keyword in text for keyword in STATUS_NOTE_KEYWORDS)


def _is_status_note_row(
    *,
    receipt_date: date | None,
    amount: float,
    month_from: str,
    month_to: str,
    values: tuple[Any, ...],
) -> bool:
    """Recognise old member-status notes that are not accounting transactions."""
    return (
        receipt_date is None
        and amount <= 0
        and not month_from
        and not month_to
        and _contains_status_note(*values)
    )


def _require_login():
    return bool(session.get("finance_login"))


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"#N/A", "#VALUE!", "None", "nan"}:
        return ""
    return text


def _amount(value: Any) -> float:
    text = _text(value).replace("RM", "").replace("-", "").strip()
    return money(text)


def _excel_date(value: Any) -> date | None:
    if value in (None, "", "-"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    text = _text(value)
    if not text:
        return None

    formats = (
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _month_ym(value: Any) -> str:
    """
    把 Excel 月份转换成 YYYY-MM。

    支持：
    - Excel 日期
    - 2025-01
    - Jan-25
    - January-2025
    - Jan 25
    """
    if value in (None, "", "-"):
        return ""

    # Excel 日期或真正的 date / datetime
    d = _excel_date(value)
    if d:
        return d.strftime("%Y-%m")

    text = _text(value).strip()

    if not text:
        return ""

    # 已经是 YYYY-MM
    match = re.fullmatch(r"(\d{4})[-/](\d{1,2})", text)

    if match:
        year = int(match.group(1))
        month = int(match.group(2))

        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}"

    month_names = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }

    match = re.fullmatch(
        r"([A-Za-z]+)[\s./-]*(\d{2,4})",
        text
    )

    if match:
        month_name = match.group(1).lower()
        year = int(match.group(2))

        if year < 100:
            year += 2000

        month = month_names.get(month_name)

        if month:
            return f"{year:04d}-{month:02d}"

    return ""


def _normalize_receipt(value: Any) -> str:
    text = _text(value).upper()
    text = re.sub(r"\s+", "", text)
    return text


def _normalize_member_id(value: Any, prefix: str) -> str:
    text = _text(value).upper().replace(" ", "")
    if not text:
        return ""
    if re.fullmatch(r"\d+(?:\.0)?", text):
        return f"{prefix}-{int(float(text))}"
    m = re.fullmatch(r"([A-Z]+)-?(\d+)", text)
    if m:
        return f"{m.group(1)}-{int(m.group(2))}"
    return text


def _payment_method(cash: float, cheque: float, bank: float) -> tuple[str, str]:
    used = [
        name
        for name, amount in (
            ("现金", cash),
            ("支票", cheque),
            ("银行过账", bank),
        )
        if amount > 0
    ]
    if not used:
        return "", "没有付款金额"
    if len(used) == 1:
        return used[0], ""
    return "混合付款", "同一行同时包含多种付款方式，请确认"


def _bank_name_from_sheet(sheet) -> str:
    sample = " ".join(
        _text(sheet.cell(r, c).value)
        for r in range(1, min(sheet.max_row, 7) + 1)
        for c in range(1, min(sheet.max_column, 16) + 1)
    ).upper()
    if "HLB" in sample:
        return "Hong Leong Bank"
    if "MBB" in sample or "MAYBANK" in sample:
        return "Maybank"
    if re.search(r"\bAM\b", sample):
        return "AmBank"
    return ""


def _fund_account(category: str, record_type: str = "income") -> str:
    if record_type == "expense" or category == "月费":
        return "观音堂日常户口"
    return "总会户口"


def _expense_category(item: str) -> str:
    """
    根据 Petty Cash 的 Itemised 自动判断大分类。

    原始 Itemised 会另外保存到 sub_category，
    所以这里仅负责判断大分类。
    """

    raw_text = str(item or "").strip()
    text = raw_text.lower()

    # =========================
    # 供品
    # =========================
    if any(keyword in raw_text for keyword in (
        "供花",
        "供果",
        "供油",
    )):
        return "供品"

    # =========================
    # 佛台与佛具
    # =========================
    if any(keyword in raw_text for keyword in (
        "佛具",
        "佛台用品",
        "佛台用具",
        "佛堂用品",
    )):
        return "佛台用品"

    # =========================
    # 电费
    # 保留 TNB 20-1、TNB 20-2 在 sub_category
    # =========================
    if any(keyword in text for keyword in (
        "tnb",
        "electric",
        "electricity",
        "电费",
    )):
        return "电费"

    # =========================
    # 水费
    # =========================
    if any(keyword in text for keyword in (
        "air selangor",
        "syabas",
        "indah water",
        "water bill",
        "水费",
    )):
        return "水费"

    # =========================
    # 电话与网络
    # =========================
    if any(keyword in text for keyword in (
        "unifi",
        "celcom",
        "digi",
        "maxis",
        "umobile",
        "u mobile",
        "reload",
        "internet",
        "phone",
        "telephone",
        "电话",
        "网络",
    )):
        return "电话及网络费"

    # =========================
    # 装修
    # 包括 GYT装修、装修工程等
    # =========================
    if any(keyword in text for keyword in (
        "gyt renovation",
        "renovation",
        "装修",
        "装潢",
        "改造工程",
    )):
        return "装修"

    # =========================
    # 维修与保养
    # =========================
    if any(keyword in text for keyword in (
        "air-con service",
        "aircon service",
        "air cond service",
        "air conditioner service",
        "repair",
        "maintenance",
        "维修",
        "保养",
    )):
        return "维修与保养"

    # =========================
    # 执照与政府费用
    # =========================
    if any(keyword in text for keyword in (
        "license",
        "licence",
        "lesen",
        "执照",
        "牌照",
    )):
        return "执照及行政费"

    # =========================
    # 现金转入
    # 注意：这不是支出
    # Parser 应另外处理为 cash_in
    # =========================
    if any(keyword in text for keyword in (
        "gyt cash in",
        "cash in",
        "petty cash in",
        "现金转入",
        "现金存入",
    )):
        return "现金转入"

    return "其它支出"

def _expense_sub_category(item: str) -> str:
    """
    完整保留 Excel 原本的 Itemised，
    例如 TNB 20-1、Celcom Reload、GYT装修。
    """
    return str(item or "").strip()


def _db_receipt_exists(receipt_no: str) -> bool:
    if not receipt_no:
        return False
    row = db_query(
        """
        select id from finance_records
        where receipt_no = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
        limit 1
        """,
        (receipt_no,),
        fetchone=True,
    )
    return bool(row)


def _db_pv_exists(pv_no: str) -> bool:
    row = db_query(
        """
        select id from finance_records
        where payment_voucher_no = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
        limit 1
        """,
        (pv_no,),
        fetchone=True,
    )
    return bool(row)


def _member_exists(member_id: str) -> bool:
    if not member_id:
        return False
    return bool(
        db_query(
            "select member_id from members where member_id = %s limit 1",
            (member_id,),
            fetchone=True,
        )
    )


def _ym_text(value: Any) -> str:
    text = _text(value)
    return text[:7] if len(text) >= 7 else text


def _same_amount(left: Any, right: Any) -> bool:
    return abs(float(left or 0) - float(right or 0)) < 0.005


def _same_months(row: dict[str, Any], data: dict[str, Any], *, member_table: bool = False) -> bool:
    left_from = row.get("start_month") if member_table else row.get("month_from")
    left_to = row.get("end_month") if member_table else row.get("month_to")
    return (
        _ym_text(left_from) == _text(data.get("month_from"))
        and _ym_text(left_to) == _text(data.get("month_to"))
    )


def _same_person(row: dict[str, Any], data: dict[str, Any]) -> bool:
    row_id = _text(row.get("member_id")).upper()
    data_id = _text(data.get("member_id")).upper()
    if row_id and data_id:
        return row_id == data_id
    row_name = re.sub(r"\s+", "", _text(row.get("name")).lower())
    data_name = re.sub(r"\s+", "", _text(data.get("name")).lower())
    return bool(row_name and data_name and row_name == data_name)


def _finance_match(data: dict[str, Any]) -> tuple[str, dict[str, Any] | None, str]:
    """Return exact / exact_warning / conflict / none for finance_records.

    Historical workbooks and the old member total sheet often use slightly
    different transaction dates. Date-only differences are warnings, not
    conflicts, when member, amount and covered months are identical.
    """
    receipt_no = data.get("receipt_no") or ""
    if receipt_no:
        row = db_query(
            """
            select id, receipt_no, member_id, name, amount, record_date,
                   month_from, month_to, status
            from finance_records
            where receipt_no = %s
              and coalesce(status, 'confirmed') <> 'cancelled'
            order by id desc limit 1
            """,
            (receipt_no,), fetchone=True,
        )
        if row:
            if _same_amount(row.get("amount"), data.get("amount")) and _same_months(row, data):
                if _same_person(row, data):
                    return "exact", row, "收条编号、会员、金额及供养月份一致"
                return "exact_warning", row, "收条编号、金额及供养月份一致，但会员编号／姓名不同，请留意旧资料差异"
            return "conflict", row, "相同收条编号的金额或供养月份不同"

    person_clause = "member_id = %s" if data.get("member_id") else "lower(replace(coalesce(name,''),' ','')) = lower(replace(%s,' ',''))"
    person_value = data.get("member_id") or data.get("name") or ""
    if person_value:
        row = db_query(
            f"""
            select id, receipt_no, member_id, name, amount, record_date,
                   month_from, month_to, status
            from finance_records
            where record_type='income' and category='月费'
              and coalesce(status, 'confirmed') <> 'cancelled'
              and {person_clause}
              and abs(amount-%s)<0.005
              and left(cast(month_from as text),7)=%s
              and left(cast(month_to as text),7)=%s
            order by id desc limit 1
            """,
            (person_value, data.get("amount") or 0, data.get("month_from"), data.get("month_to")),
            fetchone=True,
        )
        if row:
            incoming_date = _text(data.get("record_date"))
            stored_date = _text(row.get("record_date"))[:10]
            if incoming_date and stored_date and incoming_date != stored_date:
                return "exact_warning", row, f"会员、金额及供养月份一致；日期不同（系统 {stored_date}／Excel {incoming_date}）"
            return "exact", row, "会员、金额及供养月份一致"

        # 只把“同一供养期间但金额不同”或“同日同额但月份不同”
        # 视为真正冲突。不能只因同一会员曾付过相同金额便送人工检查，
        # 否则每月 RM50 的正常历史付款都会被误判。
        similar = db_query(
            f"""
            select id, receipt_no, member_id, name, amount, record_date,
                   month_from, month_to, status
            from finance_records
            where record_type='income' and category='月费'
              and coalesce(status, 'confirmed') <> 'cancelled'
              and {person_clause}
              and (
                    (
                        left(cast(month_from as text),7)=%s
                        and left(cast(month_to as text),7)=%s
                    )
                    or (
                        record_date=%s
                        and abs(amount-%s)<0.005
                    )
              )
            order by id desc limit 1
            """,
            (
                person_value,
                data.get("month_from"), data.get("month_to"),
                data.get("record_date") or None,
                data.get("amount") or 0,
            ),
            fetchone=True,
        )
        if similar:
            return "conflict", similar, "同一会员的供养期间或同日同额记录不同，请核对"
    return "none", None, "财政记录未找到"


def _member_payment_match(data: dict[str, Any]) -> tuple[str, dict[str, Any] | None, str]:
    """Return exact / exact_warning / conflict / none for member_payments."""
    receipt_no = data.get("receipt_no") or ""
    if receipt_no:
        row = db_query(
            """
            select id, receipt_no, member_id, name, amount, payment_date, receipt_date,
                   start_month, end_month, month_count, status
            from member_payments
            where receipt_no=%s and coalesce(status,'active')='active'
            order by id desc limit 1
            """,
            (receipt_no,), fetchone=True,
        )
        if row:
            if _same_amount(row.get("amount"), data.get("amount")) and _same_months(row, data, member_table=True):
                if _same_person(row, data):
                    return "exact", row, "收条编号、会员、金额及供养月份一致"
                return "exact_warning", row, "收条编号、金额及供养月份一致，但会员编号／姓名不同，请留意旧总表差异"
            return "conflict", row, "相同收条编号的金额或供养月份不同"

    person_clause = "member_id = %s" if data.get("member_id") else "lower(replace(coalesce(name,''),' ','')) = lower(replace(%s,' ',''))"
    person_value = data.get("member_id") or data.get("name") or ""
    if person_value:
        row = db_query(
            f"""
            select id, receipt_no, member_id, name, amount, payment_date, receipt_date,
                   start_month, end_month, month_count, status
            from member_payments
            where coalesce(status,'active')='active'
              and {person_clause}
              and abs(amount-%s)<0.005
              and left(cast(start_month as text),7)=%s
              and left(cast(end_month as text),7)=%s
            order by id desc limit 1
            """,
            (person_value, data.get("amount") or 0, data.get("month_from"), data.get("month_to")),
            fetchone=True,
        )
        if row:
            incoming_receipt = _text(data.get("receipt_no"))
            stored_receipt = _text(row.get("receipt_no"))

            # 财政明细已有完整收条编号，而旧查询总表的收条编号为空：
            # 这是安全的“补编号”场景，只允许补空白，绝不覆盖已有编号。
            if incoming_receipt and not stored_receipt:
                return (
                    "exact_fill_receipt",
                    row,
                    f"会员、金额及供养月份一致；查询记录 ID {row.get('id')} 缺少收条编号，可补为 {incoming_receipt}",
                )

            # 两边都有编号但不同，不能自动覆盖。
            if incoming_receipt and stored_receipt and incoming_receipt != stored_receipt:
                return "conflict", row, "会员、金额及供养月份一致，但已有收条编号不同"

            incoming_date = _text(data.get("receipt_date") or data.get("record_date"))
            stored_date = _text(row.get("receipt_date") or row.get("payment_date"))[:10]
            if incoming_date and stored_date and incoming_date != stored_date:
                return "exact_warning", row, f"会员、金额及供养月份一致；日期不同（查询系统 {stored_date}／Excel {incoming_date}）"
            return "exact", row, "会员、金额及供养月份一致"

        similar = db_query(
            f"""
            select id, receipt_no, member_id, name, amount, payment_date, receipt_date,
                   start_month, end_month, month_count, status
            from member_payments
            where coalesce(status,'active')='active'
              and {person_clause}
              and (
                    (
                        left(cast(start_month as text),7)=%s
                        and left(cast(end_month as text),7)=%s
                    )
                    or (
                        coalesce(receipt_date, payment_date)=%s
                        and abs(amount-%s)<0.005
                    )
              )
            order by id desc limit 1
            """,
            (
                person_value,
                data.get("month_from"), data.get("month_to"),
                data.get("receipt_date") or data.get("record_date") or None,
                data.get("amount") or 0,
            ),
            fetchone=True,
        )
        if similar:
            return "conflict", similar, "同一会员的供养期间或同日同额查询记录不同，请核对"
    return "none", None, "月费查询记录未找到"




def _review_value(value: Any, *, month: bool = False, date_only: bool = False) -> str:
    text = _text(value)
    if month:
        return _ym_text(value) or "-"
    if date_only:
        return text[:10] if text else "-"
    return text or "-"


def _build_review_analysis(
    source: str,
    row: dict[str, Any] | None,
    data: dict[str, Any],
    *,
    member_table: bool = False,
) -> dict[str, Any]:
    """Explain exactly why an imported monthly-fee row needs review."""
    row = row or {}
    db_from_key = "start_month" if member_table else "month_from"
    db_to_key = "end_month" if member_table else "month_to"
    db_date = (
        row.get("receipt_date") or row.get("payment_date")
        if member_table
        else row.get("record_date")
    )
    excel_date = data.get("receipt_date") or data.get("record_date")

    differences: list[dict[str, str]] = []

    def add(code: str, label: str, excel: Any, database: Any, suggestion: str):
        differences.append({
            "code": code,
            "label": label,
            "excel": _review_value(excel),
            "database": _review_value(database),
            "suggestion": suggestion,
        })

    incoming_receipt = _text(data.get("receipt_no"))
    stored_receipt = _text(row.get("receipt_no"))
    if incoming_receipt != stored_receipt:
        add(
            "receipt", "收条编号不同", incoming_receipt, stored_receipt,
            "核对纸本收条；若查询总表没有收条编号，可保留财政Excel编号。",
        )

    incoming_member = _text(data.get("member_id") or data.get("name"))
    stored_member = _text(row.get("member_id") or row.get("name"))
    if incoming_member and stored_member and incoming_member.upper() != stored_member.upper():
        add(
            "member", "会员资料不同", incoming_member, stored_member,
            "先核对会员编号与姓名，避免把付款记到错误会员。",
        )

    if not _same_amount(row.get("amount"), data.get("amount")):
        add(
            "amount", "金额不同", f"RM {float(data.get('amount') or 0):.2f}",
            f"RM {float(row.get('amount') or 0):.2f}",
            "金额会影响总账与供养月数，必须核对原收条或银行记录。",
        )

    incoming_from = _ym_text(data.get("month_from"))
    stored_from = _ym_text(row.get(db_from_key))
    incoming_to = _ym_text(data.get("month_to"))
    stored_to = _ym_text(row.get(db_to_key))
    if incoming_from != stored_from or incoming_to != stored_to:
        differences.append({
            "code": "months",
            "label": "供养月份不同",
            "excel": f"{incoming_from or '-'} 至 {incoming_to or '-'}",
            "database": f"{stored_from or '-'} 至 {stored_to or '-'}",
            "suggestion": "检查收条上注明的开始月与结束月；不要自动覆盖。",
        })

    excel_date_text = _review_value(excel_date, date_only=True)
    db_date_text = _review_value(db_date, date_only=True)
    if excel_date_text != db_date_text:
        differences.append({
            "code": "date",
            "label": "日期不同",
            "excel": excel_date_text,
            "database": db_date_text,
            "suggestion": "若只差入账日与收条日，可采用财政Excel日期；若跨月则需核对。",
        })

    if not differences:
        differences.append({
            "code": "other",
            "label": "资料组合需要确认",
            "excel": "财政Excel记录",
            "database": "现有查询记录",
            "suggestion": "请打开原Excel与现有记录核对后再决定。",
        })

    codes = [d["code"] for d in differences]
    material = [c for c in codes if c in {"amount", "months", "member", "receipt"}]
    if len(set(material)) > 1:
        primary_code = "multiple"
        primary_label = "多项关键资料不同"
    elif material:
        primary_code = material[0]
        primary_label = next(d["label"] for d in differences if d["code"] == primary_code)
    elif "date" in codes:
        primary_code = "date"
        primary_label = "日期不同"
    else:
        primary_code = "other"
        primary_label = differences[0]["label"]

    if primary_code in {"amount", "months", "multiple", "member"}:
        overall = "必须人工核对，不建议自动覆盖。"
    elif primary_code == "receipt":
        overall = "优先核对纸本收条编号；确认后再决定。"
    elif primary_code == "date":
        overall = "若金额和供养月份一致，通常可采用财政Excel日期，但跨月时应核对。"
    else:
        overall = "请核对原始资料后再决定。"

    return {
        "source": source,
        "candidate_id": row.get("id"),
        "primary_code": primary_code,
        "primary_label": primary_label,
        "differences": differences,
        "suggestion": overall,
        "excel_snapshot": {
            "receipt_no": incoming_receipt or "-",
            "member": incoming_member or "-",
            "amount": f"RM {float(data.get('amount') or 0):.2f}",
            "months": f"{incoming_from or '-'} 至 {incoming_to or '-'}",
            "date": excel_date_text,
        },
        "database_snapshot": {
            "receipt_no": stored_receipt or "-",
            "member": stored_member or "-",
            "amount": f"RM {float(row.get('amount') or 0):.2f}",
            "months": f"{stored_from or '-'} 至 {stored_to or '-'}",
            "date": db_date_text,
        },
    }

def _monthly_import_resolution(data: dict[str, Any]) -> dict[str, Any]:
    finance_state, finance_row, finance_note = _finance_match(data)
    member_state, member_row, member_note = _member_payment_match(data)

    reviews: list[dict[str, Any]] = []
    details: list[str] = []
    if finance_state == "conflict":
        reviews.append(_build_review_analysis("finance_records", finance_row, data))
        details.append(
            f"财政记录待核对 ID {finance_row.get('id') if finance_row else '-'}：{finance_note}"
        )
    if member_state == "conflict":
        reviews.append(
            _build_review_analysis(
                "member_payments", member_row, data, member_table=True
            )
        )
        details.append(
            f"查询记录待核对 ID {member_row.get('id') if member_row else '-'}：{member_note}"
        )

    if reviews:
        # 若财政表与查询表都冲突，优先显示关键差异较多的一项。
        priority = {"multiple": 6, "amount": 5, "months": 4, "member": 3, "receipt": 2, "date": 1, "other": 0}
        primary = max(reviews, key=lambda x: priority.get(x.get("primary_code"), 0))
        return {
            "action": "manual_review",
            "status": "review",
            "note": "；".join(details),
            "review_analysis": primary,
            "review_candidates": reviews,
        }

    warnings = [
        note
        for state, note in (
            (finance_state, finance_note),
            (member_state, member_note),
        )
        if state == "exact_warning"
    ]

    # 查询系统已有完全相同的供养记录，但收条编号为空。
    # 只允许把财政Excel中的编号补入空白字段，绝不覆盖现有编号。
    if member_state == "exact_fill_receipt":
        target_id = member_row.get("id") if member_row else None
        target_receipt = data.get("receipt_no") or ""
        if finance_state in ("exact", "exact_warning"):
            return {
                "action": "fill_receipt_only",
                "status": "ready",
                "note": f"财政记录已存在；只补查询记录收条编号为 {target_receipt}",
                "member_payment_target_id": target_id,
                "receipt_to_fill": target_receipt,
            }
        if finance_state == "none":
            return {
                "action": "finance_and_fill_receipt",
                "status": "ready",
                "note": f"新增财政记录，并补查询记录收条编号为 {target_receipt}",
                "member_payment_target_id": target_id,
                "receipt_to_fill": target_receipt,
            }

    if finance_state in ("exact", "exact_warning") and member_state in ("exact", "exact_warning"):
        note = "财政记录和月费查询都已存在"
        if warnings:
            note += "；" + "；".join(warnings)
        return {"action": "skip_duplicate", "status": "duplicate", "note": note}
    if finance_state in ("exact", "exact_warning") and member_state == "none":
        note = "财政记录已存在，本次只补月费查询记录"
        if warnings:
            note += "；" + "；".join(warnings)
        return {"action": "member_only", "status": "warning" if warnings else "ready", "note": note}
    if finance_state == "none" and member_state in ("exact", "exact_warning"):
        note = "月费查询记录已存在，本次只补财政记录"
        if warnings:
            note += "；" + "；".join(warnings)
        return {"action": "finance_only", "status": "warning" if warnings else "ready", "note": note}
    return {
        "action": "finance_and_member",
        "status": "ready",
        "note": "新增财政记录和月费查询记录",
    }


def _monthly_import_action(data: dict[str, Any]) -> tuple[str, str, str]:
    result = _monthly_import_resolution(data)
    return result["action"], result["status"], result["note"]

def _expense_db_exists(data: dict[str, Any]) -> bool:
    row = db_query(
        """
        select id from finance_records
        where record_type='expense'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and record_date=%s
          and abs(amount-%s)<0.005
          and lower(trim(coalesce(remarks,'')))=lower(trim(%s))
        limit 1
        """,
        (data.get("record_date") or None, data.get("amount") or 0, data.get("remarks") or ""),
        fetchone=True,
    )
    return bool(row)


def _bank_deposit_db_exists(data: dict[str, Any]) -> bool:
    row = db_query(
        """
        select id from finance_bank_deposits
        where deposit_date=%s
          and abs(amount-%s)<0.005
          and coalesce(reference_no,'')=coalesce(%s,'')
        limit 1
        """,
        (data.get("deposit_date") or None, data.get("amount") or 0, data.get("reference_no") or ""),
        fetchone=True,
    )
    return bool(row)


def _stage_row(
    batch_id: int,
    *,
    sheet_name: str,
    source_row: int,
    row_type: str,
    status: str,
    normalized: dict[str, Any],
    raw: dict[str, Any],
    error: str = "",
    warning: str = "",
):
    db_query(
        """
        insert into finance_import_rows
        (
            batch_id, sheet_name, source_row, row_type, status,
            error_message, warning_message, normalized_data, raw_data
        )
        values (%s,%s,%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb)
        """,
        (
            batch_id,
            sheet_name,
            source_row,
            row_type,
            status,
            error,
            warning,
            _json(normalized),
            _json(raw),
        ),
    )


def _status(error: list[str], warning: list[str], duplicate: bool = False) -> str:
    if duplicate:
        return "duplicate"
    if error:
        return "error"
    if warning:
        return "warning"
    return "ready"


def _monthly_header_row(sheet) -> int | None:
    """Find the official monthly-fee header row in real GYT/STW workbooks."""
    for r in range(1, min(sheet.max_row, 15) + 1):
        values = [
            _text(sheet.cell(r, c).value).lower()
            for c in range(1, min(sheet.max_column, 18) + 1)
        ]
        joined = " | ".join(values)
        has_receipt = (
            "official receipt" in joined
            or "收据编号" in joined
            or "收據編號" in joined
        )
        has_month = (
            "start month" in joined
            and "end month" in joined
        )
        has_name = "name" in joined or "姓名" in joined
        if has_receipt and has_month and has_name:
            return r
    return None


def _sheet_month_from_title(title: str) -> str:
    """Convert Jan-26 / Jan 26 / Sept-24 sheet names into YYYY-MM."""
    text = title.strip().lower().replace("(new)", "").replace("new", "")
    text = re.sub(r"\s+", " ", text)
    months = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4,
        "may": 5, "jun": 6, "jul": 7, "aug": 8,
        "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    m = re.search(
        r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[\s-]*(\d{2,4})\b",
        text,
    )
    if not m:
        return ""
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return f"{year:04d}-{months[m.group(1)]:02d}"


def _month_in_range(ym: str, month_from: str, month_to: str) -> bool:
    if not ym:
        return True
    if month_from and ym < month_from:
        return False
    if month_to and ym > month_to:
        return False
    return True


def _infer_monthly_prefix(sheet, file_name: str, header_row: int) -> str:
    file_upper = file_name.upper()
    if "SITIAWAN" in file_upper or "STW" in file_upper:
        return "STW"
    for r in range(header_row + 1, min(sheet.max_row, header_row + 25) + 1):
        receipt = _normalize_receipt(sheet.cell(r, 2).value)
        if receipt.startswith("STW"):
            return "STW"
        if receipt.startswith("CHE"):
            return "CHE"
    return "CHE"




def _monthly_amount_columns(sheet, header_row: int) -> dict[str, int | None]:
    """Detect monthly-fee amount columns across old and new GYT templates.

    Old template:
        J = total amount, K = cash, L = cheque, M = online bank
    New template:
        J = cash, K = cheque, L = online bank, M = CDM sequence
    Some transitional sheets only contain J = total amount.
    """
    found: dict[str, int | None] = {
        "total": None,
        "cash": None,
        "cheque": None,
        "bank": None,
        "cdm_seq": None,
        "bank_in_amount": None,
        "bank_in_date": None,
        "remarks": None,
    }

    for r in range(max(1, header_row - 2), min(sheet.max_row, header_row + 1) + 1):
        for c in range(1, min(sheet.max_column, 20) + 1):
            text = _text(sheet.cell(r, c).value).lower().replace("\n", " ")
            if not text:
                continue
            if ("cash" in text or "现款" in text or "現款" in text) and "bank-in" not in text:
                found["cash"] = c
            elif "cheque" in text or "支票" in text:
                found["cheque"] = c
            elif "online bank" in text or "网上银行" in text or "網上銀行" in text:
                found["bank"] = c
            elif "cdm seq" in text:
                found["cdm_seq"] = c
            elif "bank-in amount" in text or "bank in amount" in text:
                found["bank_in_amount"] = c
            elif "bank-in date" in text or "bank in date" in text:
                found["bank_in_date"] = c
            elif "remark" in text or "备注" in text or "備註" in text:
                found["remarks"] = c

    cash_col = found["cash"]
    # In the old workbook, J is the calculated total and cash starts at K.
    if cash_col == 11:
        found["total"] = 10
    # Transitional old sheets have only J as the transaction amount.
    elif cash_col is None:
        found["total"] = 10

    # Safe fallbacks for the known GYT workbook layout.
    found["bank_in_amount"] = found["bank_in_amount"] or 14
    found["bank_in_date"] = found["bank_in_date"] or 15
    found["remarks"] = found["remarks"] or 16
    return found


def _parse_monthly_sheet(
    sheet,
    batch_id: int,
    file_name: str,
    month_from: str = "",
    month_to: str = "",
):
    header_row = _monthly_header_row(sheet)
    if not header_row:
        return False

    sheet_ym = _sheet_month_from_title(sheet.title)
    if not _month_in_range(sheet_ym, month_from, month_to):
        return False

    prefix = _infer_monthly_prefix(sheet, file_name, header_row)
    amount_cols = _monthly_amount_columns(sheet, header_row)
    bank_name = _bank_name_from_sheet(sheet)
    deposit_groups: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"receipts": [], "amount": 0.0, "date": None}
    )
    seen_receipts: set[str] = set()

    for r in range(header_row + 2, sheet.max_row + 1):
        raw = {chr(64 + c): sheet.cell(r, c).value for c in range(1, 17)}
        receipt_no = _normalize_receipt(sheet.cell(r, 2).value)
        name = _text(sheet.cell(r, 4).value)
        if not receipt_no and not name:
            continue

        receipt_date = _excel_date(sheet.cell(r, 1).value)
        member_id = _normalize_member_id(sheet.cell(r, 3).value, prefix)
        payment_month_from = _month_ym(sheet.cell(r, 5).value)
        payment_month_to = _month_ym(sheet.cell(r, 6).value)
        month_count = int(_amount(sheet.cell(r, 8).value) or 0)
        # 某些旧Excel的月份格是文字格式。
        # 若E/F仍然无法读取，则尝试从G栏REF补回。
        if not payment_month_from or not payment_month_to:

            ref_text = _text(
                sheet.cell(r, 7).value
            )

            ref_match = re.search(
                r"""
                (?P<start>[A-Za-z]{3,9})
                [\s-]*
                (?:
                    (?P<start_year>\d{2,4})
                    [\s-]*
                )?
                (?:to|至|-)
                [\s-]*
                (?P<end>[A-Za-z]{3,9})
                [\s-]*
                (?P<end_year>\d{2,4})
                """,
                ref_text,
                re.IGNORECASE | re.VERBOSE,
            )

            if ref_match:

                end_year = int(
                    ref_match.group("end_year")
                )

                if end_year < 100:
                    end_year += 2000

                start_year_raw = (
                    ref_match.group("start_year")
                )

                if start_year_raw:
                    start_year = int(start_year_raw)

                    if start_year < 100:
                        start_year += 2000
                else:
                    start_year = end_year

                payment_month_from = (
                    payment_month_from
                    or _month_ym(
                        f"{ref_match.group('start')}-{start_year}"
                    )
                )

                payment_month_to = (
                    payment_month_to
                    or _month_ym(
                        f"{ref_match.group('end')}-{end_year}"
                    )
                )

        total_amount = (
            _amount(sheet.cell(r, amount_cols["total"]).value)
            if amount_cols["total"]
            else 0.0
        )
        cash = (
            _amount(sheet.cell(r, amount_cols["cash"]).value)
            if amount_cols["cash"]
            else 0.0
        )
        cheque = (
            _amount(sheet.cell(r, amount_cols["cheque"]).value)
            if amount_cols["cheque"]
            else 0.0
        )
        bank = (
            _amount(sheet.cell(r, amount_cols["bank"]).value)
            if amount_cols["bank"]
            else 0.0
        )
        split_amount = cash + cheque + bank

        # Critical rule: when the workbook already has a calculated total,
        # use it directly. Do not add the payment-method breakdown again.
        amount = total_amount if total_amount > 0 else split_amount

        # Transitional sheets only store the amount in J and do not have
        # payment-method columns. Historically these rows are cash records.
        if amount > 0 and split_amount <= 0:
            cash = amount

        # Real workbooks contain future blank template rows with receipt numbers
        # and #N/A formulas. They are not accounting records.
        if (
            not receipt_date
            and not name
            and not member_id
            and amount <= 0
        ):
            continue
        if receipt_no in {"BANKIN", "TOTAL", "SUMMARY"}:
            continue

        method, method_warning = _payment_method(cash, cheque, bank)
        cdm_seq = (
            _text(sheet.cell(r, amount_cols["cdm_seq"]).value)
            if amount_cols["cdm_seq"]
            else ""
        )
        bank_in_amount = _amount(sheet.cell(r, amount_cols["bank_in_amount"]).value)
        bank_in_date = _excel_date(sheet.cell(r, amount_cols["bank_in_date"]).value)
        remarks = _text(sheet.cell(r, amount_cols["remarks"]).value)
        bank_holder = _text(sheet.cell(r, 17).value)
        bank_note = _text(sheet.cell(r, 18).value)
        extra_notes = []
        if bank_holder:
            extra_notes.append(f"Bank holder: {bank_holder}")
        if bank_note:
            extra_notes.append(bank_note)
        if extra_notes:
            remarks = "；".join([x for x in [remarks, *extra_notes] if x])

        if _is_status_note_row(
            receipt_date=receipt_date,
            amount=amount,
            month_from=payment_month_from,
            month_to=payment_month_to,
            values=(receipt_no, member_id, name, remarks, bank_holder, bank_note),
        ):
            normalized = {
                "record_type": "note",
                "record_date": "",
                "receipt_no": receipt_no,
                "member_id": member_id,
                "name": name,
                "amount": 0,
                "category": "会员状态备注",
                "payment_method": "",
                "import_action": "skip_status_note",
                "match_note": "会员停供／取消备注，不属于财政交易，系统自动略过",
                "import_source": file_name,
                "legacy_record": True,
            }
            _stage_row(
                batch_id, sheet_name=sheet.title, source_row=r,
                row_type="status_note", status="skipped",
                normalized=normalized, raw=raw,
                warning="会员停供／取消备注行，未写入财政或月费查询",
            )
            continue

        errors: list[str] = []
        warnings: list[str] = []
        duplicate = False
        if total_amount > 0 and split_amount > 0 and abs(total_amount - split_amount) >= 0.005:
            warnings.append(
                f"总金额RM{total_amount:.2f}与付款分项RM{split_amount:.2f}不同，系统采用总金额"
            )
        if not receipt_date:
            errors.append("收条日期无效")
        if not receipt_no:
            if method == "银行过账":
                warnings.append("银行月费尚未填写纸本收条编号，可先作为历史记录导入")
            else:
                errors.append("缺少收条编号")
        if not name:
            errors.append("缺少姓名")
        if amount <= 0:
            errors.append("金额必须大过0")
        if amount > 0 and amount % 50 != 0:
            errors.append("月费金额不是RM50倍数")
        if not payment_month_from or not payment_month_to:
            warnings.append("缺少开始或结束月份")
        if month_count <= 0 and amount > 0:
            month_count = int(amount / 50)
            warnings.append("月数为空，系统按金额自动计算")
        if method_warning:
            warnings.append(method_warning)
        if receipt_date:
            lock = require_finance_month_open(receipt_date, _fund_account("月费"))
            if lock:
                errors.append(lock)
        # 同一上传文件内部重复仍直接拦截；数据库已有资料交给
        # Smart Matching 判断，才能执行“只补查询／只补收条编号”等动作。
        if receipt_no and receipt_no in seen_receipts:
            duplicate = True
        if receipt_no:
            seen_receipts.add(receipt_no)
        if member_id and not _member_exists(member_id):
            warnings.append(f"会员表找不到 {member_id}，仍可作为历史资料导入")

        normalized = {
            "record_type": "income",
            "fund_account": _fund_account("月费"),
            "record_date": receipt_date.isoformat() if receipt_date else "",
            "receipt_date": receipt_date.isoformat() if receipt_date else "",
            "category": "月费",
            "sub_category": "",
            "receipt_no": receipt_no,
            "member_id": member_id,
            "name": name,
            "phone": "",
            "amount": amount,
            "payment_method": method,
            "month_from": payment_month_from,
            "month_to": payment_month_to,
            "month_count": month_count,
            "remarks": remarks or "历史Excel导入",
            "import_source": file_name,
            "legacy_record": True,
        }

        action = "finance_and_member"
        action_note = "新增财政记录和月费查询记录"
        action_status = ""
        if not errors and not duplicate:
            resolution = _monthly_import_resolution(normalized)
            action = resolution["action"]
            action_status = resolution["status"]
            action_note = resolution["note"]
            if resolution.get("member_payment_target_id"):
                normalized["member_payment_target_id"] = resolution["member_payment_target_id"]
            if resolution.get("receipt_to_fill"):
                normalized["receipt_to_fill"] = resolution["receipt_to_fill"]
            if resolution.get("review_analysis"):
                normalized["review_analysis"] = resolution["review_analysis"]
                normalized["review_candidates"] = resolution.get("review_candidates", [])
            if action_status == "review":
                errors.append(action_note)
            elif action_status == "warning":
                warnings.append(action_note)
            elif action_status == "duplicate":
                duplicate = True

        normalized["import_action"] = action
        normalized["match_note"] = action_note
        final_status = _status(errors, warnings, duplicate)
        if action_status == "review":
            final_status = "review"

        _stage_row(
            batch_id,
            sheet_name=sheet.title,
            source_row=r,
            row_type="monthly_fee",
            status=final_status,
            normalized=normalized,
            raw=raw,
            error="；".join(errors),
            warning="；".join(warnings),
        )

        if cdm_seq and receipt_no:
            g = deposit_groups[cdm_seq]
            g["receipts"].append(receipt_no)
            if bank_in_amount > 0:
                g["amount"] = bank_in_amount
            if bank_in_date:
                g["date"] = bank_in_date

    _stage_deposits(
        batch_id,
        sheet.title,
        file_name,
        deposit_groups,
        bank_name,
        _fund_account("月费"),
    )
    return True


def _parse_donation_sheet(sheet, batch_id: int, file_name: str, kind: str):
    if kind == "meal":
        category, sub_category = "膳食结缘", "初一十五"
        cash_col, bank_col, cheque_col = 9, 10, 11
        seq_col, deposit_amount_col, deposit_date_col, remarks_col = 12, 13, 14, 15
    elif kind == "village":
        category, sub_category = "观音村", ""
        cash_col, cheque_col, bank_col = 8, 9, 10
        seq_col, deposit_amount_col, deposit_date_col, remarks_col = 11, 12, 13, 14
    else:
        category, sub_category = "财布施", ""
        cash_col, cheque_col, bank_col = 7, 8, 9
        seq_col, deposit_amount_col, deposit_date_col, remarks_col = 10, 11, 12, 13

    bank_name = _bank_name_from_sheet(sheet)
    deposit_groups: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"receipts": [], "amount": 0.0, "date": None}
    )
    seen_receipts: set[str] = set()

    for r in range(9, sheet.max_row + 1):
        raw = {chr(64 + c): sheet.cell(r, c).value for c in range(1, min(sheet.max_column, 16) + 1)}
        receipt_date = _excel_date(sheet.cell(r, 1).value)
        receipt_no = _normalize_receipt(sheet.cell(r, 2).value)
        name = _text(sheet.cell(r, 3).value)
        if not receipt_no and not name:
            continue

        bank_holder = _text(sheet.cell(r, 4).value)
        phone = _text(sheet.cell(r, 5).value)
        cash = _amount(sheet.cell(r, cash_col).value)
        cheque = _amount(sheet.cell(r, cheque_col).value)
        bank = _amount(sheet.cell(r, bank_col).value)
        amount = cash + cheque + bank
        method, method_warning = _payment_method(cash, cheque, bank)
        cdm_seq = _text(sheet.cell(r, seq_col).value)
        bank_in_amount = _amount(sheet.cell(r, deposit_amount_col).value)
        bank_in_date = _excel_date(sheet.cell(r, deposit_date_col).value)
        remarks = _text(sheet.cell(r, remarks_col).value)

        row_sub_category = sub_category
        if kind == "village":
            printing = _amount(sheet.cell(r, 6).value)
            release = _amount(sheet.cell(r, 7).value)
            if printing > 0 and release <= 0:
                row_sub_category = "印刷／法会"
            elif release > 0 and printing <= 0:
                row_sub_category = "放生"
            elif printing > 0 and release > 0:
                row_sub_category = "印刷／法会＋放生"

        errors: list[str] = []
        warnings: list[str] = []
        duplicate = False
        if not receipt_date:
            errors.append("收条日期无效")
        if not receipt_no:
            errors.append("缺少收条编号")
        if not name:
            errors.append("缺少姓名")
        if amount <= 0:
            errors.append("金额必须大过0")
        if method_warning:
            warnings.append(method_warning)
        if bank_holder and bank_holder != name:
            warnings.append(f"银行户名：{bank_holder}")
        if receipt_date:
            lock = require_finance_month_open(receipt_date, _fund_account(category))
            if lock:
                errors.append(lock)
        if receipt_no in seen_receipts or _db_receipt_exists(receipt_no):
            duplicate = True
        seen_receipts.add(receipt_no)

        normalized = {
            "record_type": "income",
            "fund_account": _fund_account(category),
            "record_date": receipt_date.isoformat() if receipt_date else "",
            "receipt_date": receipt_date.isoformat() if receipt_date else "",
            "category": category,
            "sub_category": row_sub_category,
            "receipt_no": receipt_no,
            "member_id": "",
            "name": name,
            "phone": phone,
            "amount": amount,
            "payment_method": method,
            "month_from": "",
            "month_to": "",
            "month_count": 0,
            "remarks": remarks or "历史Excel导入",
            "bank_holder_name": bank_holder,
            "import_source": file_name,
            "legacy_record": True,
            "import_action": "finance_only",
            "match_note": "新增财政收入记录",
        }
        _stage_row(
            batch_id,
            sheet_name=sheet.title,
            source_row=r,
            row_type="income",
            status=_status(errors, warnings, duplicate),
            normalized=normalized,
            raw=raw,
            error="；".join(errors),
            warning="；".join(warnings),
        )

        if cdm_seq and receipt_no:
            g = deposit_groups[cdm_seq]
            g["receipts"].append(receipt_no)
            if bank_in_amount > 0:
                g["amount"] = bank_in_amount
            if bank_in_date:
                g["date"] = bank_in_date

    _stage_deposits(
        batch_id,
        sheet.title,
        file_name,
        deposit_groups,
        bank_name,
        _fund_account(category),
    )


def _stage_deposits(
    batch_id: int,
    sheet_name: str,
    file_name: str,
    groups: dict[str, dict[str, Any]],
    bank_name: str,
    fund_account: str,
):
    for seq, g in groups.items():
        receipts = g["receipts"]
        amount = float(g["amount"] or 0)
        deposit_date = g["date"]
        errors: list[str] = []
        warnings: list[str] = []
        if amount <= 0:
            warnings.append("CDM批次没有Bank-in金额，暂不导入Bank In")
        if not deposit_date:
            warnings.append("CDM批次没有Bank-in日期")
        if deposit_date:
            lock = require_finance_month_open(deposit_date, fund_account)
            if lock:
                errors.append(lock)

        normalized = {
            "deposit_date": deposit_date.isoformat() if deposit_date else "",
            "ym": deposit_date.strftime("%Y-%m") if deposit_date else "",
            "fund_account": fund_account,
            "bank_name": bank_name,
            "reference_no": f"CDM-{seq}",
            "receipt_from": receipts[0] if receipts else "",
            "receipt_to": receipts[-1] if receipts else "",
            "amount": amount,
            "remarks": f"历史Excel CDM批次 {seq}",
            "cdm_sequence": seq,
            "import_source": file_name,
            "import_action": "bank_only",
            "match_note": "新增Cash Bank In记录",
        }
        duplicate = _bank_deposit_db_exists(normalized) if deposit_date and amount > 0 else False
        if duplicate:
            normalized["import_action"] = "skip_duplicate"
            normalized["match_note"] = "相同日期、金额及Reference的Bank In已存在"
        _stage_row(
            batch_id,
            sheet_name=sheet_name,
            source_row=0,
            row_type="bank_deposit",
            status=_status(errors, warnings, duplicate),
            normalized=normalized,
            raw={"sequence": seq, "receipts": receipts},
            error="；".join(errors),
            warning="；".join(warnings),
        )


def _parse_petty_cash_2026(sheet, batch_id: int, file_name: str):
    # Six month blocks: Date / In / Out / Balance / Itemised, with one spacer column.
    starts = [1, 7, 13, 19, 25, 31]
    
    for start_col in starts:
        block_date = _excel_date(sheet.cell(1, start_col).value)
        year = block_date.year if block_date else 2026
        for r in range(4, sheet.max_row + 1):
            tx_date = _excel_date(sheet.cell(r, start_col).value)
            out_amount = _amount(sheet.cell(r, start_col + 2).value)
            item = _text(sheet.cell(r, start_col + 4).value)
            if out_amount <= 0 or not item:
                continue
            if not tx_date:
                tx_date = block_date

            category = _expense_category(item)
            sub_category = _expense_sub_category(item)
            errors: list[str] = []
            warnings: list[str] = ["历史Petty Cash没有原始PV编号及Vendor"]
            if not tx_date:
                errors.append("支出日期无效")
            if tx_date:
                lock = require_finance_month_open(tx_date, _fund_account(category, "expense"))
                if lock:
                    errors.append(lock)
            duplicate = False

            normalized = {
                "record_type": "expense",
                "fund_account": _fund_account(category, "expense"),
                "record_date": tx_date.isoformat() if tx_date else "",
                "receipt_date": "",
                "category": category,
                "sub_category": sub_category,
                "payment_voucher_no": "",  # generated after staging row receives an id
                "vendor": "历史资料",
                "name": "历史资料",
                "amount": out_amount,
                "payment_method": "现金",
                "remarks": item,
                "import_source": file_name,
                "legacy_record": True,
                "legacy_year": year,
                "import_action": "finance_only",
                "match_note": "新增历史支出记录",
            }

            # Excel 内部出现相同日期、金额及用途
            if duplicate:
                normalized["import_action"] = "skip_duplicate"
                normalized["match_note"] = "Excel内部重复，系统将自动略过"

            elif _expense_db_exists(normalized):
                duplicate = True
                normalized["import_action"] = "skip_duplicate"
                normalized["match_note"] = "相同日期、金额及用途的支出已存在"
            
            _stage_row(
                batch_id,
                sheet_name=sheet.title,
                source_row=r,
                row_type="expense",
                status=_status(errors, warnings, duplicate),
                normalized=normalized,
                raw={
                    "date": sheet.cell(r, start_col).value,
                    "out": sheet.cell(r, start_col + 2).value,
                    "itemised": sheet.cell(r, start_col + 4).value,
                    "block_start_column": start_col,
                },
                error="；".join(errors),
                warning="；".join(warnings),
            )


def _parse_workbook(
    content: bytes,
    file_name: str,
    batch_id: int,
    month_from: str = "",
    month_to: str = "",
) -> list[str]:
    wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
    recognized: list[str] = []
    skip_names = {
        "summary", "blank_master", "blank master", "master name list",
        "master data", "for hq", "sheet6",
    }

    for sheet in wb.worksheets:
        title = sheet.title.strip()
        normalized_title = title.lower().replace("-", " ").replace("_", " ").strip()
        if normalized_title in skip_names:
            continue

        # The real 2025/2026 monthly-fee workbooks use generic names such
        # as Jan-26, Feb-26. Detect them by their header, not filename/title.
        if _monthly_header_row(sheet):
            if _parse_monthly_sheet(
                sheet, batch_id, file_name, month_from, month_to
            ):
                recognized.append(title)
            continue

        if "初一十五" in title:
            _parse_donation_sheet(sheet, batch_id, file_name, "meal")
            recognized.append(title)
        elif "观音村" in title:
            _parse_donation_sheet(sheet, batch_id, file_name, "village")
            recognized.append(title)
        elif "善款收纳表" in title and "观音村" not in title:
            _parse_donation_sheet(sheet, batch_id, file_name, "donation")
            recognized.append(title)
        elif title == "2026 Petty Cash Manual Book":
            _parse_petty_cash_2026(sheet, batch_id, file_name)
            recognized.append(title)
    return recognized


def _refresh_batch(batch_id: int):
    counts = db_query(
        """
        select
            count(*) as total_rows,
            count(*) filter (where status = 'ready') as ready_rows,
            count(*) filter (where status = 'warning') as warning_rows,
            count(*) filter (where status = 'duplicate') as duplicate_rows,
            count(*) filter (where status = 'review') as review_rows,
            count(*) filter (where status = 'error') as error_rows,
            count(*) filter (where status = 'skipped') as skipped_status_rows,
            count(*) filter (where normalized_data->>'import_action' = 'finance_and_member') as finance_and_member_rows,
            count(*) filter (where normalized_data->>'import_action' = 'finance_only') as finance_only_rows,
            count(*) filter (where normalized_data->>'import_action' = 'member_only') as member_only_rows,
            count(*) filter (where normalized_data->>'import_action' = 'fill_receipt_only') as fill_receipt_only_rows,
            count(*) filter (where normalized_data->>'import_action' = 'finance_and_fill_receipt') as finance_and_fill_receipt_rows,
            count(*) filter (where normalized_data->>'import_action' = 'bank_only') as bank_only_rows,
            count(*) filter (where normalized_data->>'import_action' = 'skip_duplicate') as duplicate_skip_rows,
            count(*) filter (where normalized_data->>'import_action' = 'skip_status_note') as status_note_rows,
            count(*) filter (where normalized_data->'review_analysis'->>'primary_code' = 'amount') as review_amount_rows,
            count(*) filter (where normalized_data->'review_analysis'->>'primary_code' = 'months') as review_months_rows,
            count(*) filter (where normalized_data->'review_analysis'->>'primary_code' = 'receipt') as review_receipt_rows,
            count(*) filter (where normalized_data->'review_analysis'->>'primary_code' = 'member') as review_member_rows,
            count(*) filter (where normalized_data->'review_analysis'->>'primary_code' = 'date') as review_date_rows,
            count(*) filter (where normalized_data->'review_analysis'->>'primary_code' = 'multiple') as review_multiple_rows,
            count(*) filter (where normalized_data->'review_analysis'->>'primary_code' = 'other') as review_other_rows
        from finance_import_rows where batch_id = %s
        """,
        (batch_id,),
        fetchone=True,
    ) or {}
    db_query(
        """
        update finance_import_batches
        set total_rows=%s, ready_rows=%s, warning_rows=%s,
            duplicate_rows=%s, review_rows=%s, error_rows=%s, skipped_rows=%s,
            summary=%s::jsonb
        where id=%s
        """,
        (
            counts.get("total_rows") or 0,
            counts.get("ready_rows") or 0,
            counts.get("warning_rows") or 0,
            counts.get("duplicate_rows") or 0,
            counts.get("review_rows") or 0,
            counts.get("error_rows") or 0,
            counts.get("skipped_status_rows") or 0,
            _json(counts),
            batch_id,
        ),
    )


@finance_import_bp.route("/", methods=["GET", "POST"])
def import_home():
    if not _require_login():
        return redirect(url_for("finance.finance_login"))

    if request.method == "POST":
        uploaded = request.files.get("excel_file")
        if not uploaded or not uploaded.filename:
            flash("请选择Excel文件。", "danger")
            return redirect(url_for("finance_import.import_home"))

        ext = uploaded.filename.rsplit(".", 1)[-1].lower() if "." in uploaded.filename else ""
        if ext not in ALLOWED_EXTENSIONS:
            flash("只支持 .xlsx 或 .xlsm 文件。", "danger")
            return redirect(url_for("finance_import.import_home"))

        month_from = request.form.get("month_from", "").strip()
        month_to = request.form.get("month_to", "").strip()
        if month_from and month_to and month_from > month_to:
            flash("开始月份不能迟过结束月份。", "danger")
            return redirect(url_for("finance_import.import_home"))

        content = uploaded.read()
        if not content or len(content) > MAX_IMPORT_BYTES:
            flash("文件为空或超过12MB。", "danger")
            return redirect(url_for("finance_import.import_home"))

        batch = db_query(
            """
            insert into finance_import_batches
            (branch_code, file_name, import_type, status, created_by)
            values (%s,%s,%s,'preview',%s)
            returning id
            """,
            (
                get_current_finance_branch(),
                uploaded.filename,
                f"auto_excel:{month_from or 'all'}:{month_to or 'all'}",
                get_current_finance_user(),
            ),
            fetchone=True,
        )
        batch_id = batch["id"]
        try:
            recognized = _parse_workbook(
                content, uploaded.filename, batch_id, month_from, month_to
            )
            if not recognized:
                raise ValueError("没有识别到支持的Sheet")
            _refresh_batch(batch_id)
            db_query(
                "update finance_import_batches set recognized_sheets=%s::jsonb where id=%s",
                (_json(recognized), batch_id),
            )
        except Exception as exc:
            db_query(
                "update finance_import_batches set status='failed', error_message=%s where id=%s",
                (str(exc), batch_id),
            )
            flash(f"Excel解析失败：{exc}", "danger")
            return redirect(url_for("finance_import.import_home"))

        return redirect(url_for("finance_import.import_preview", batch_id=batch_id))

    history = db_query(
        """
        select * from finance_import_batches
        where branch_code=%s order by id desc limit 30
        """,
        (get_current_finance_branch(),),
        fetchall=True,
    ) or []
    return render_template_string(IMPORT_HOME_HTML, history=history)


@finance_import_bp.route("/<int:batch_id>")
def import_preview(batch_id: int):
    if not _require_login():
        return redirect(url_for("finance.finance_login"))
    batch = db_query(
        "select * from finance_import_batches where id=%s and branch_code=%s",
        (batch_id, get_current_finance_branch()),
        fetchone=True,
    )
    if not batch:
        flash("找不到导入批次。", "danger")
        return redirect(url_for("finance_import.import_home"))
    status_filter = request.args.get("status", "").strip()
    action_filter = request.args.get("action", "").strip()
    review_code = request.args.get("review_code", "").strip()
    clauses = ["batch_id=%s"]
    params: list[Any] = [batch_id]
    if status_filter:
        clauses.append("status=%s")
        params.append(status_filter)
    if action_filter:
        clauses.append("normalized_data->>'import_action'=%s")
        params.append(action_filter)
    if review_code:
        clauses.append("normalized_data->'review_analysis'->>'primary_code'=%s")
        params.append(review_code)
    rows = db_query(
        f"""
        select id, sheet_name, source_row, row_type, status,
               error_message, warning_message, normalized_data
        from finance_import_rows
        where {' and '.join(clauses)}
        order by sheet_name, source_row, id
        limit 1500
        """,
        tuple(params),
        fetchall=True,
    ) or []
    return render_template_string(
        IMPORT_PREVIEW_HTML,
        batch=batch,
        rows=rows,
        status_filter=status_filter,
        action_filter=action_filter,
        review_code=review_code,
    )


@finance_import_bp.route("/<int:batch_id>/confirm", methods=["POST"])
def import_confirm(batch_id: int):
    if not _require_login():
        return redirect(url_for("finance.finance_login"))

    batch = db_query(
        "select * from finance_import_batches where id=%s and branch_code=%s",
        (batch_id, get_current_finance_branch()),
        fetchone=True,
    )
    if not batch or batch.get("status") != "preview":
        flash("这个导入批次不能确认。", "danger")
        return redirect(url_for("finance_import.import_home"))
    if int(batch.get("error_rows") or 0) > 0 or int(batch.get("review_rows") or 0) > 0:
        flash("仍有错误或疑似冲突资料，不能确认导入。", "danger")
        return redirect(url_for("finance_import.import_preview", batch_id=batch_id))

    rows = db_query(
        """
        select * from finance_import_rows
        where batch_id=%s and status in ('ready','warning')
        order by id
        """,
        (batch_id,),
        fetchall=True,
    ) or []

    # Confirm 前再次核对，防止 Preview 后数据库才新增了相同资料。
    conflict_found = False
    for row in rows:
        if row["row_type"] != "monthly_fee":
            continue
        data = row["normalized_data"] or {}
        if isinstance(data, str):
            data = json.loads(data)
        resolution = _monthly_import_resolution(data)
        latest_action = resolution["action"]
        latest_status = resolution["status"]
        latest_note = resolution["note"]
        data["import_action"] = latest_action
        data["match_note"] = latest_note
        if resolution.get("member_payment_target_id"):
            data["member_payment_target_id"] = resolution["member_payment_target_id"]
        else:
            data.pop("member_payment_target_id", None)
        if resolution.get("receipt_to_fill"):
            data["receipt_to_fill"] = resolution["receipt_to_fill"]
        else:
            data.pop("receipt_to_fill", None)
        if resolution.get("review_analysis"):
            data["review_analysis"] = resolution["review_analysis"]
            data["review_candidates"] = resolution.get("review_candidates", [])
        else:
            data.pop("review_analysis", None)
            data.pop("review_candidates", None)
        if latest_status == "review":
            db_query(
                """
                update finance_import_rows
                set status='review', error_message=%s, normalized_data=%s::jsonb
                where id=%s
                """,
                (latest_note, _json(data), row["id"]),
            )
            conflict_found = True
        elif latest_action != row["normalized_data"].get("import_action"):
            db_query(
                """
                update finance_import_rows set normalized_data=%s::jsonb where id=%s
                """,
                (_json(data), row["id"]),
            )
    if conflict_found:
        _refresh_batch(batch_id)
        flash("确认前发现新的疑似冲突，请检查橙色“待检查”资料。", "danger")
        return redirect(url_for("finance_import.import_preview", batch_id=batch_id))

    # 重新读取，因为系统动作可能在上面的二次核对中改变。
    rows = db_query(
        """
        select * from finance_import_rows
        where batch_id=%s and status in ('ready','warning')
        order by id
        """,
        (batch_id,),
        fetchall=True,
    ) or []

    processed = 0
    finance_added = 0
    member_added = 0
    receipt_filled = 0
    bank_added = 0
    skipped = 0

    with get_conn() as conn:
        with conn.cursor() as cur:
            for row in rows:
                data = row["normalized_data"] or {}
                if isinstance(data, str):
                    data = json.loads(data)
                row_type = row["row_type"]
                action = data.get("import_action") or (
                    "finance_and_member" if row_type == "monthly_fee" else
                    "bank_only" if row_type == "bank_deposit" else
                    "finance_only"
                )
                finance_record_id = None
                member_payment_id = None
                bank_deposit_id = None


                if action == "skip_duplicate":
                    cur.execute(
                        "update finance_import_rows set status='duplicate' where id=%s",
                        (row["id"],),
                    )
                    skipped += 1
                    continue

                if row_type in ("monthly_fee", "income") and action in (
                    "finance_and_member", "finance_only", "finance_and_fill_receipt"
                ):
                    cur.execute(
                        """
                        insert into finance_records
                        (
                            record_type, fund_account, record_date, receipt_date,
                            category, sub_category, receipt_no, member_id, name,
                            phone, amount, payment_method, month_from, month_to,
                            remarks, status, import_batch_id, import_source,
                            legacy_record
                        )
                        values
                        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                         'confirmed',%s,%s,true)
                        returning id
                        """,
                        (
                            data.get("record_type"), data.get("fund_account"),
                            data.get("record_date") or None, data.get("receipt_date") or None,
                            data.get("category"), data.get("sub_category") or None,
                            data.get("receipt_no") or None, data.get("member_id") or None,
                            data.get("name") or None, data.get("phone") or None,
                            data.get("amount") or 0, data.get("payment_method") or None,
                            data.get("month_from") or None, data.get("month_to") or None,
                            data.get("remarks") or "历史Excel导入", batch_id,
                            data.get("import_source") or batch.get("file_name"),
                        ),
                    )
                    finance_record_id = cur.fetchone()["id"]
                    finance_added += 1

                if row_type == "monthly_fee" and action in (
                    "finance_and_member", "member_only"
                ):
                    month_from_db = f"{data.get('month_from')}-01" if data.get("month_from") else None
                    month_to_db = f"{data.get('month_to')}-01" if data.get("month_to") else None
                    cur.execute(
                        """
                        insert into member_payments
                        (
                            payment_date, receipt_date, member_id, name,
                            receipt_no, amount, start_month, end_month,
                            month_count, status, import_batch_id, import_source
                        )
                        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,'active',%s,%s)
                        returning id
                        """,
                        (
                            data.get("record_date") or None,
                            data.get("receipt_date") or None,
                            data.get("member_id") or None,
                            data.get("name") or None,
                            data.get("receipt_no") or None,
                            data.get("amount") or 0,
                            month_from_db, month_to_db,
                            data.get("month_count") or 0,
                            batch_id,
                            data.get("import_source") or batch.get("file_name"),
                        ),
                    )
                    member_payment_id = cur.fetchone()["id"]
                    member_added += 1

                if row_type == "monthly_fee" and action in (
                    "fill_receipt_only", "finance_and_fill_receipt"
                ):
                    target_id = data.get("member_payment_target_id")
                    receipt_to_fill = _text(data.get("receipt_to_fill") or data.get("receipt_no"))
                    if not target_id or not receipt_to_fill:
                        raise ValueError(
                            f"导入行 {row['id']} 缺少补收条编号所需的目标资料"
                        )
                    cur.execute(
                        """
                        update member_payments
                        set receipt_no = %s,
                            import_batch_id = coalesce(import_batch_id, %s),
                            import_source = coalesce(import_source, %s)
                        where id = %s
                          and coalesce(status, 'active') = 'active'
                          and (receipt_no is null or trim(receipt_no) = '')
                        returning id
                        """,
                        (
                            receipt_to_fill,
                            batch_id,
                            data.get("import_source") or batch.get("file_name"),
                            target_id,
                        ),
                    )
                    filled = cur.fetchone()
                    if not filled:
                        raise ValueError(
                            f"查询记录 ID {target_id} 的收条编号已被填写或资料已改变，请重新预览"
                        )
                    member_payment_id = filled["id"]
                    receipt_filled += 1

                elif row_type == "expense" and action == "finance_only":
                    if _expense_db_exists(data):
                        cur.execute(
                            "update finance_import_rows set status='duplicate' where id=%s",
                            (row["id"],),
                        )
                        skipped += 1
                        continue
                    pv_no = f"LEGACY-PV-{batch_id:04d}-{row['id']:06d}"
                    cur.execute(
                        """
                        insert into finance_records
                        (
                            record_type, fund_account, record_date, category,
                            sub_category, payment_voucher_no, vendor, name,
                            amount, payment_method, remarks, status,
                            import_batch_id, import_source, legacy_record
                        )
                        values
                        ('expense',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                         'confirmed',%s,%s,true)
                        returning id
                        """,
                        (
                            data.get("fund_account"), data.get("record_date") or None,
                            data.get("category"), data.get("sub_category") or None,
                            pv_no, data.get("vendor") or "历史资料",
                            data.get("name") or "历史资料", data.get("amount") or 0,
                            data.get("payment_method") or "现金",
                            data.get("remarks") or "历史Petty Cash导入",
                            batch_id, data.get("import_source") or batch.get("file_name"),
                        ),
                    )
                    finance_record_id = cur.fetchone()["id"]
                    finance_added += 1

                elif row_type == "bank_deposit" and action == "bank_only":
                    if not data.get("deposit_date") or float(data.get("amount") or 0) <= 0:
                        continue
                    # 先检查 Reference No 是否已经存在
                    if data.get("reference_no"):

                        cur.execute("""
                            select id
                            from finance_bank_deposits
                            where lower(reference_no)=lower(%s)
                            limit 1
                        """, (
                            data["reference_no"],
                        ))

                        if cur.fetchone():

                            cur.execute(
                                """
                                update finance_import_rows
                                set status='duplicate'
                                where id=%s
                                """,
                                (row["id"],),
                            )

                            skipped += 1
                            continue

                    # 再检查其它条件
                    if _bank_deposit_db_exists(data):

                        cur.execute(
                            """
                            update finance_import_rows
                            set status='duplicate'
                            where id=%s
                            """,
                            (row["id"],),
                        )

                        skipped += 1
                        continue
                    cur.execute(
                        """
                        insert into finance_bank_deposits
                        (
                            deposit_date, ym, fund_account, bank_name,
                            reference_no, receipt_from, receipt_to, amount,
                            remarks, import_batch_id, import_source, cdm_sequence
                        )
                        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        returning id
                        """,
                        (
                            data.get("deposit_date"), data.get("ym"),
                            data.get("fund_account"), data.get("bank_name") or None,
                            data.get("reference_no") or None,
                            data.get("receipt_from") or None,
                            data.get("receipt_to") or None,
                            data.get("amount") or 0,
                            data.get("remarks") or "历史Excel CDM导入",
                            batch_id, data.get("import_source") or batch.get("file_name"),
                            data.get("cdm_sequence") or None,
                        ),
                    )
                    bank_deposit_id = cur.fetchone()["id"]
                    bank_added += 1

                cur.execute(
                    """
                    update finance_import_rows
                    set status='imported', finance_record_id=%s,
                        member_payment_id=%s, bank_deposit_id=%s
                    where id=%s
                    """,
                    (finance_record_id, member_payment_id, bank_deposit_id, row["id"]),
                )
                processed += 1

            result_summary = {
                "processed_rows": processed,
                "finance_records_added": finance_added,
                "member_payments_added": member_added,
                "missing_receipts_filled": receipt_filled,
                "bank_deposits_added": bank_added,
                "duplicates_skipped": skipped,
            }
            cur.execute(
                """
                update finance_import_batches
                set status='confirmed', confirmed_by=%s,
                    confirmed_at=now(), imported_rows=%s,
                    result_summary=%s::jsonb
                where id=%s
                """,
                (
                    get_current_finance_user(), processed,
                    _json(result_summary), batch_id,
                ),
            )

    write_finance_audit(
        module="finance_import",
        action="confirm",
        record_id=batch_id,
        new_value={
            "file_name": batch.get("file_name"),
            "processed_rows": processed,
            "finance_records_added": finance_added,
            "member_payments_added": member_added,
            "missing_receipts_filled": receipt_filled,
            "bank_deposits_added": bank_added,
            "duplicates_skipped": skipped,
        },
        reason="历史Excel智能同步导入确认",
    )
    flash(
        f"导入完成：财政记录 {finance_added}，月费查询新增 {member_added}，"
        f"补回收条编号 {receipt_filled}，Bank In {bank_added}，重复跳过 {skipped}。",
        "success",
    )
    return redirect(url_for("finance_import.import_preview", batch_id=batch_id))


IMPORT_HOME_HTML = """
<!doctype html><html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>历史财政资料导入</title>
<link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
<style>
.import-page{max-width:1100px}.upload-card{border:2px dashed #93c5fd;background:#eff6ff}
.import-note{line-height:1.8;color:#475569}.status-pill{padding:5px 10px;border-radius:999px;font-weight:800}
.status-preview{background:#fef3c7;color:#92400e}.status-confirmed{background:#dcfce7;color:#166534}
.status-failed{background:#fee2e2;color:#991b1b}
</style></head><body><div class="page import-page">
<div class="card"><h1 class="page-title">📥 历史财政资料导入</h1>
<p class="page-subtitle">上传财政原有Excel，系统会先预览、检查，再由财政确认写入。</p></div>
<div class="card upload-card"><div class="section-title">① 选择Excel</div>
<form method="post" enctype="multipart/form-data">
<div class="form-group"><label class="form-label">Excel文件</label>
<input class="form-input" type="file" name="excel_file" accept=".xlsx,.xlsm" required></div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:14px">
<div class="form-group"><label class="form-label">开始月份（可选）</label>
<input class="form-input" type="month" name="month_from" value=""></div>
<div class="form-group"><label class="form-label">结束月份（可选）</label>
<input class="form-input" type="month" name="month_to" value=""></div>
</div>
<div class="alert alert-warning import-note">不会上传后立即入账。系统先识别Sheet、检查日期、金额、重复收条、月结锁定及CDM批次。</div>
<button class="btn-tool btn-primary" type="submit">🔍 上传并预览</button></form></div>
<div class="card"><div class="section-title">支持的格式</div>
<p class="import-note">支持真实的 Jan-26／Feb-26 等月费Sheet、SITIAWAN月费、初一十五膳食结缘、观音村善款、善款收纳表及2026 Petty Cash。系统会自动略过 Summary、Master、Blank_Master 和未来空白模板行。</p></div>
<div class="card"><div class="section-title">导入历史</div><div class="table-responsive"><table class="record-table">
<thead><tr><th>时间</th><th>文件</th><th>状态</th><th>总行</th><th>可导入</th><th>警告</th><th>重复</th><th>略过</th><th>待检查</th><th>错误</th><th></th></tr></thead>
<tbody>{% for b in history %}<tr><td>{{ b.created_at }}</td><td>{{ b.file_name }}</td>
<td><span class="status-pill status-{{ b.status }}">{{ b.status }}</span></td><td>{{ b.total_rows or 0 }}</td>
<td>{{ b.ready_rows or 0 }}</td><td>{{ b.warning_rows or 0 }}</td><td>{{ b.duplicate_rows or 0 }}</td><td>{{ b.skipped_rows or 0 }}</td><td>{{ b.review_rows or 0 }}</td><td>{{ b.error_rows or 0 }}</td>
<td><a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_preview', batch_id=b.id) }}">查看</a></td></tr>
{% else %}<tr><td colspan="11">还没有导入记录</td></tr>{% endfor %}</tbody></table></div></div>
<div class="btn-row"><a class="btn-tool btn-secondary" href="{{ url_for('finance.finance_home') }}">← 返回财政首页</a></div>
</div></body></html>
"""

IMPORT_PREVIEW_HTML = """
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">

    <title>历史资料预览</title>

    <link rel="stylesheet"
          href="{{ url_for('static', filename='css/toolbox.css') }}">

    <style>
        .preview-page{max-width:1500px}
        .kpi-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px}
        .kpi{padding:18px;text-align:center}
        .kpi strong{font-size:28px;display:block}
        .row-error{background:#fee2e2}
        .row-warning{background:#fef3c7}
        .row-duplicate{background:#f1f5f9}
        .row-skipped{background:#eef2ff;color:#64748b}
        .row-review{background:#fed7aa}
        .small{font-size:13px;color:#64748b}
        .import-table{min-width:2000px}
        .review-box{background:#fff7ed;border:1px solid #fdba74;border-radius:12px;padding:10px;min-width:280px}
        .diff-line{margin:5px 0}
        .diff-label{font-weight:800;color:#9a3412}
        .suggestion{margin-top:7px;color:#7c2d12;font-weight:700}
        .review-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-top:14px}
        .action-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-top:14px}
        @media(max-width:900px){.kpi-grid,.action-grid{grid-template-columns:1fr 1fr}}
        @media(max-width:700px){.kpi-grid,.action-grid,.review-grid{grid-template-columns:1fr}}
    </style>
</head>
<body>
<div class="page preview-page">

    <div class="card">
        <h1 class="page-title">📋 导入预览</h1>
        <p><b>文件：</b>{{ batch.file_name }}</p>
        <p><b>状态：</b>{{ batch.status }}</p>
    </div>

    {% set sm = batch.summary or {} %}

    <div class="kpi-grid">
        <div class="card kpi"><span>总行</span><strong>{{ batch.total_rows or 0 }}</strong></div>
        <div class="card kpi"><span>可自动处理</span><strong>{{ (batch.ready_rows or 0) + (batch.warning_rows or 0) }}</strong></div>
        <div class="card kpi"><span>提醒</span><strong>{{ batch.warning_rows or 0 }}</strong></div>
        <div class="card kpi"><span>待检查</span><strong>{{ batch.review_rows or 0 }}</strong></div>
        <div class="card kpi"><span>错误</span><strong>{{ batch.error_rows or 0 }}</strong></div>
    </div>

    <div class="card">
        <div class="section-title">系统准备执行的动作</div>

        <div class="action-grid">
            <a class="btn-tool btn-primary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, action='finance_and_member') }}">新增两边<br>{{ sm.finance_and_member_rows or 0 }}</a>
            <a class="btn-tool btn-success" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, action='finance_only') }}">只补财政<br>{{ sm.finance_only_rows or 0 }}</a>
            <a class="btn-tool btn-purple" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, action='member_only') }}">只补查询<br>{{ sm.member_only_rows or 0 }}</a>
            <a class="btn-tool btn-warning" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, action='fill_receipt_only') }}">只补收条编号<br>{{ sm.fill_receipt_only_rows or 0 }}</a>
            <a class="btn-tool btn-primary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, action='finance_and_fill_receipt') }}">新增财政＋补编号<br>{{ sm.finance_and_fill_receipt_rows or 0 }}</a>
            <a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, action='skip_duplicate') }}">完全重复<br>{{ sm.duplicate_skip_rows or 0 }}</a>
            <a class="btn-tool btn-warning" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, action='skip_status_note') }}">状态备注略过<br>{{ sm.status_note_rows or 0 }}</a>
        </div>

        {% if batch.review_rows %}
        <div class="section-title" style="margin-top:24px">待检查原因分析</div>
        <div class="review-grid">
            <a class="btn-tool btn-danger" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, review_code='amount') }}">金额不同<br>{{ sm.review_amount_rows or 0 }}</a>
            <a class="btn-tool btn-warning" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, review_code='months') }}">供养月份不同<br>{{ sm.review_months_rows or 0 }}</a>
            <a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, review_code='receipt') }}">收条编号不同<br>{{ sm.review_receipt_rows or 0 }}</a>
            <a class="btn-tool btn-purple" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, review_code='member') }}">会员资料不同<br>{{ sm.review_member_rows or 0 }}</a>
            <a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, review_code='date') }}">日期不同<br>{{ sm.review_date_rows or 0 }}</a>
            <a class="btn-tool btn-danger" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, review_code='multiple') }}">多项不同<br>{{ sm.review_multiple_rows or 0 }}</a>
            <a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, review_code='other') }}">其它<br>{{ sm.review_other_rows or 0 }}</a>
        </div>
        {% endif %}

        <div class="btn-row" style="margin-top:16px">
            <a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id) }}">全部</a>
            {% for s,label in [('ready','正常'),('warning','提醒'),('duplicate','重复'),('skipped','略过'),('review','待检查'),('error','错误'),('imported','已导入')] %}
            <a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_preview', batch_id=batch.id, status=s) }}">{{ label }}</a>
            {% endfor %}
        </div>
    </div>

    <div class="card">
        <div class="table-responsive">
            <table class="record-table import-table">
                <thead>
                <tr>
                    <th>Sheet</th>
                    <th>行</th>
                    <th>类型</th>
                    <th>状态</th>
                    <th>系统动作</th>
                    <th>日期</th>
                    <th>编号</th>
                    <th>姓名／对象</th>
                    <th>分类</th>
                    <th>用途</th>
                    <th>金额</th>
                    <th>方式</th>
                    <th>差异分析</th>
                    <th>系统建议</th>
                    <th>错误</th>
                    <th>警告</th>
                </tr>
                </thead>

                <tbody>
                {% for r in rows %}
                {% set d = r.normalized_data or {} %}
                {% set ra = d.review_analysis or {} %}

                <tr class="row-{{ r.status }}">
                    <td>{{ r.sheet_name }}</td>
                    <td>{{ r.source_row or '-' }}</td>
                    <td>{{ r.row_type }}</td>
                    <td>{{ r.status }}</td>
                    <td>
                        <b>{{ d.import_action or '-' }}</b>
                        {% if d.match_note %}<div class="small">{{ d.match_note }}</div>{% endif %}
                    </td>
                    <td>{{ d.record_date or d.deposit_date or '-' }}</td>
                    <td>{{ d.receipt_no or d.payment_voucher_no or d.reference_no or '-' }}</td>
                    <td>{{ d.name or d.vendor or '-' }}</td>
                    <td>{{ d.category or d.fund_account or '-' }}</td>
                    <td>{{ d.sub_category or d.remarks or '-' }}</td>
                    <td>RM {{ '%.2f'|format(d.amount or 0) }}</td>
                    <td>{{ d.payment_method or '-' }}</td>

                    <td>
                        {% if ra %}
                        <div class="review-box">
                            <div class="diff-label">{{ ra.primary_label }}</div>
                            {% for x in ra.differences %}
                            <div class="diff-line">
                                <b>{{ x.label }}</b><br>
                                <span class="small">Excel：{{ x.excel }}<br>系统：{{ x.database }}</span>
                            </div>
                            {% endfor %}
                        </div>
                        {% else %}-{% endif %}
                    </td>

                    <td>
                        {% if ra %}
                        <div class="suggestion">{{ ra.suggestion }}</div>
                        <div class="small">候选来源：{{ ra.source }} · ID {{ ra.candidate_id or '-' }}</div>
                        {% else %}-{% endif %}
                    </td>

                    <td>{{ r.error_message or '-' }}</td>
                    <td>{{ r.warning_message or '-' }}</td>
                </tr>

                {% else %}
                <tr><td colspan="16">没有资料</td></tr>
                {% endfor %}
                </tbody>
            </table>
        </div>
    </div>

    <div class="btn-row">
        <a class="btn-tool btn-secondary" href="{{ url_for('finance_import.import_home') }}">← 返回导入中心</a>

        {% if batch.status == 'preview' %}
        <form method="post"
              action="{{ url_for('finance_import.import_confirm', batch_id=batch.id) }}"
              onsubmit="return confirm('确定把所有正常及警告资料正式写入财政系统？重复资料会跳过。');">
            <button class="btn-tool btn-success"
                    type="submit"
                    {% if batch.error_rows or batch.review_rows %}disabled{% endif %}>
                ✅ 确认导入
            </button>
        </form>
        {% endif %}
    </div>

</div>
</body>
</html>
"""
