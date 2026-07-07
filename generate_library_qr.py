# generate_library_qr.py

import os
import qrcode
from collections import defaultdict
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

EXCEL_FILE = "library_inventory.xlsx"

OUTPUT_DIR = "library_qr_output"
QR_DIR = os.path.join(OUTPUT_DIR, "qr")
PDF_FILE = os.path.join(OUTPUT_DIR, "Library_QR_Catalogue.pdf")

FONT_FILE = "C:/Windows/Fonts/msyh.ttc"

os.makedirs(QR_DIR, exist_ok=True)
pdfmetrics.registerFont(TTFont("MY", FONT_FILE))


def read_items():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    items = []

    for r in range(1, ws.max_row + 1):
        raw_code = ws.cell(r, 1).value
        raw_name = ws.cell(r, 2).value
        raw_balance = ws.cell(r, 3).value
        raw_category = ws.cell(r, 4).value

        if not raw_code or not raw_name:
            continue

        try:
            number = int(raw_code)
        except:
            continue

        item_code = f"BOOK{number:04d}"

        category = "未分类"
        if raw_category:
            category = str(raw_category).strip()

        items.append({
            "item_code": item_code,
            "name": str(raw_name).strip(),
            "balance": raw_balance,
            "category": category,
        })

    if not items:
        raise Exception("没有读取到法宝。请确认 A栏=编号，B栏=名称，C栏=数量，D栏=分类。")

    return items


def make_qr_image(item_code):
    BASE_URL = "https://gyt-checkin.onrender.com"
    qr_text = f"{BASE_URL}/library/scan/{item_code}"

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )

    qr.add_data(qr_text)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    path = os.path.join(QR_DIR, f"{item_code}.png")
    img.save(path)

    return path


def short_text(text, max_len):
    text = str(text)

    if len(text) <= max_len:
        return text

    return text[:max_len] + "..."


def draw_top_title(c, page_no, title="藏经阁法宝 QR Code 目录"):
    width, height = A4

    c.setFont("MY", 17)
    c.drawString(16 * mm, height - 15 * mm, title)

    c.setFont("MY", 8)
    c.drawRightString(width - 16 * mm, height - 15 * mm, f"Page {page_no}")

    c.line(16 * mm, height - 22 * mm, width - 16 * mm, height - 22 * mm)


def draw_index_page(c, grouped_items, page_no):
    width, height = A4

    draw_top_title(c, page_no, "藏经阁法宝 QR Code 目录")

    c.setFont("MY", 11)
    c.drawString(
        16 * mm,
        height - 32 * mm,
        "扫一扫后可进入法宝操作页面。"
    )
    c.drawString(
        16 * mm,
        height - 39 * mm,
        "可选择：登记领取、登记入库、查看法宝资料。"
    )

    c.setFont("MY", 15)
    c.drawString(16 * mm, height - 55 * mm, "📚 分类索引")

    y = height - 70 * mm

    for category, items in grouped_items.items():
        c.setFont("MY", 12)
        c.drawString(22 * mm, y, f"📖 {category}")

        c.setFont("MY", 10)
        c.drawRightString(width - 25 * mm, y, f"{len(items)} 项")

        y -= 10 * mm

        if y < 25 * mm:
            c.showPage()
            page_no += 1
            draw_top_title(c, page_no, "藏经阁法宝 QR Code 目录")
            y = height - 35 * mm

    return page_no


def draw_category_header(c, category, count, page_no):
    width, height = A4

    draw_top_title(c, page_no)

    c.setFont("MY", 15)
    c.drawString(16 * mm, height - 33 * mm, f"📖 {category}（{count} 项）")

    c.line(16 * mm, height - 38 * mm, width - 16 * mm, height - 38 * mm)


def draw_card(c, item, x, y):
    card_w = 86 * mm
    card_h = 58 * mm
    qr_size = 32 * mm

    c.roundRect(
        x,
        y - card_h,
        card_w,
        card_h,
        3 * mm
    )

    c.setFont("MY", 10)
    c.drawString(
        x + 4 * mm,
        y - 8 * mm,
        short_text(item["name"], 25)
    )

    c.setFont("MY", 8)
    c.drawString(
        x + 4 * mm,
        y - 15 * mm,
        item["item_code"]
    )

    qr_path = make_qr_image(item["item_code"])

    c.drawImage(
        qr_path,
        x + 4 * mm,
        y - 48 * mm,
        qr_size,
        qr_size
    )

    c.setFont("MY", 8)
    c.drawString(
        x + 38 * mm,
        y - 30 * mm,
        "扫码登记"
    )

    c.drawString(
        x + 38 * mm,
        y - 37 * mm,
        "领取法宝"
    )

    c.setFont("MY", 7)
    c.drawString(
        x + 38 * mm,
        y - 47 * mm,
        item["item_code"]
    )


def make_pdf(items):
    grouped = defaultdict(list)

    for item in items:
        grouped[item["category"]].append(item)

    grouped_items = dict(sorted(grouped.items(), key=lambda x: x[0]))

    for category in grouped_items:
        grouped_items[category].sort(key=lambda x: x["name"])

    c = canvas.Canvas(PDF_FILE, pagesize=A4)

    width, height = A4
    page_no = 1

    page_no = draw_index_page(c, grouped_items, page_no)

    x_positions = [
        16 * mm,
        108 * mm
    ]

    y_positions = [
        height - 48 * mm,
        height - 108 * mm,
        height - 168 * mm,
        height - 228 * mm,
    ]

    for category, category_items in grouped_items.items():

        c.showPage()
        page_no += 1

        draw_category_header(c, category, len(category_items), page_no)

        index = 0

        for item in category_items:

            pos = index % 8

            if index > 0 and pos == 0:
                c.showPage()
                page_no += 1
                draw_category_header(c, category, len(category_items), page_no)

            x = x_positions[pos % 2]
            y = y_positions[pos // 2]

            draw_card(c, item, x, y)

            index += 1

    c.save()


def main():
    items = read_items()

    print(f"读取到 {len(items)} 项法宝")

    make_pdf(items)

    print("完成！")
    print(f"PDF 目录：{PDF_FILE}")
    print(f"QR 图片：{QR_DIR}")


if __name__ == "__main__":
    main()