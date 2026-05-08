# reading_web.py

import os
import pandas as pd

from db import db_query
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from flask import (
    Blueprint,
    request,
    redirect,
    url_for,
    render_template_string
)

from datetime import datetime
from utils import get_text, now_date_str, MY_TZ

reading_bp = Blueprint("reading", __name__)
READING_FILE = "reading.xlsx"

def ensure_reading_file():
    required_cols = ["日期", "姓名", "主题", "场次", "时间"]

    if not os.path.exists(READING_FILE):
        df = pd.DataFrame(columns=required_cols)
        df.to_excel(READING_FILE, index=False)
        return

    df = pd.read_excel(READING_FILE)

    for col in required_cols:
        if col not in df.columns:
            df[col] = ""

    df = df[required_cols]
    df.to_excel(READING_FILE, index=False)
    beautify_reading_excel()

def beautify_reading_excel():
    if not os.path.exists(READING_FILE):
        return
    
    print("🔥 正在执行 beautify")

    wb = load_workbook(READING_FILE)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(name="Microsoft YaHei", size=16, bold=True)
    body_font = Font(name="Microsoft YaHei", size=15)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = header_font if cell.row == 1 else body_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill

    # ✅ 行高：重点
    ws.row_dimensions[1].height = 32

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 30

    # ✅ 列宽：放大
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    wb.save(READING_FILE)

def get_today_reading_rows():
    rows = db_query("""
        select *
        from reading
        where date = %s
        order by id desc
    """, (now_date_str(),), fetchall=True)

    return rows or []



def add_reading_record(name, identity, topic, session, time_text):
    db_query("""
        insert into reading
        (date, name, identity, topic, session, time)
        values (%s, %s, %s, %s, %s, %s)
    """, (
        now_date_str(),
        name,
        identity,
        topic,
        session,
        time_text
    ))

def get_reading_topics():
    rows = db_query("""
        select distinct topic
        from reading
        where topic is not null and topic <> ''
        order by topic
    """, fetchall=True)

    return [r["topic"] for r in rows if r.get("topic")]

def get_today_attendees():
    rows = db_query("""
        select distinct name
        from attendance
        where date = %s and signin = 1
    """, (now_date_str(),), fetchall=True)

    return sorted([r["name"] for r in rows if r.get("name")])

@reading_bp.route("/reading", methods=["GET", "POST"])
def reading():
    t = get_text()
    attendees = get_today_attendees()
    today = now_date_str()

    if request.method == "POST":
        topic = request.form.get("topic", "").strip()
        names = request.form.getlist("names")
        session = request.form.get("session", "").strip()
        extra_text = request.form.get("extra_names", "").strip()

        extra_names = []
        if extra_text:
            for sep in ["，", "、", ",", "\n"]:
                extra_text = extra_text.replace(sep, " ")
            extra_names = [x.strip() for x in extra_text.split(" ") if x.strip()]

        if len(names) + len(extra_names) < 2:
            return f"❌ {t['need_two_people']}<br><a href='/reading'>{t['back_home']}</a>"

        if not topic:
            return f"❌ {t['enter_topic']}<br><a href='/reading'>{t['back_home']}</a>"

        now_time = datetime.now(MY_TZ).strftime("%I:%M %p").lstrip("0")

        for name in names:
            add_reading_record(
                name=name,
                identity=t["volunteer_identity"],
                topic=topic,
                session=session,
                time_text=now_time
            )

        for name in extra_names:
            add_reading_record(
                name=name,
                identity=t["friend_identity"],
                topic=topic,
                session=session,
                time_text=now_time
            )

        return redirect(url_for("reading.reading"))

    rows = get_today_reading_rows()

    today_records = []
    for r in rows:
        today_records.append({
            "id": r.get("id"),
            "日期": r.get("date"),
            "姓名": r.get("name"),
            "身份": r.get("identity"),
            "主题": r.get("topic"),
            "场次": r.get("session"),
            "时间": r.get("time"),
        })

    summary_map = {}
    for r in today_records:
        key = (r["姓名"], r["身份"])
        summary_map[key] = summary_map.get(key, 0) + 1

    today_summary_records = [
        {
            "姓名": name,
            "身份": identity,
            "共修次数": count
        }
        for (name, identity), count in summary_map.items()
    ]

    topic_options = get_reading_topics()

    html = """
    <!doctype html>
    <html lang="{{ t.html_lang }}">
    <head>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <style>
            body {
                font-family: "Microsoft YaHei", Arial;
                background: #f6f7f2;
                padding: 18px;
                font-size: 24px;
            }
            .card {
                background: white;
                padding: 28px;
                border-radius: 16px;
                margin-bottom: 18px;
                box-shadow: 0 2px 8px #ccc;
            }
            input[type=text] {
                width: 95%;
                font-size: 32px;
                padding: 20px;
                border-radius: 10px;
                border: 1px solid #aaa;
            }
            label {
                display: block;
                padding: 8px;
                font-size: 21px;
            }
            button {
                font-size: 32px;
                padding: 12px 22px;
                border-radius: 18px;
                border: none;
                background: #4CAF50;
                color: white;
                margin: 5px;
            }
            .delete {
                background: #d9534f;
            }
            .edit {
                background: #f0ad4e;
            }
            a {
                text-decoration: none;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                background: white;
                font-size: 18px;
            }
            th, td {
                border: 1px solid #ccc;
                padding: 8px;
                text-align: center;
            }
            th {
                background: #d9ead3;
            }
        </style>
    </head>
    <body>

    <a href="/"><button>⬅ {{ t.back_home }}</button></a>

    <a href="/download_reading">
        <button style="background:#2196F3;">{{ t.download_report }}</button>
    </a>

    <div class="card">
        <h2>{{ t.bhff_record_title }}</h2>

        <form method="post">
            <p>{{ t.today_topic }}：</p>
            <input type="text" name="topic" list="topicList" placeholder="{{ t.topic_placeholder }}">

            <datalist id="topicList">
            {% for topic in topic_options %}
                <option value="{{ topic }}">
            {% endfor %}
            </datalist>

            <p>{{ t.session_remark }}：</p>
            <input type="text" name="session" placeholder="{{ t.session_placeholder }}">

            <p>{{ t.volunteer_list }}：</p>

            <button type="button" onclick="selectAllNames()" style="margin-bottom:10px;">
                {{ t.select_all_volunteers }}
            </button>

            {% for name in attendees %}
                <label>
                    <input type="checkbox" name="names" value="{{name}}">
                    {{name}}
                </label>
            {% endfor %}

            <br>
            <button type="submit">{{ t.record_study }}</button>

            <p>{{ t.extra_friend_name }}：</p>
            <input type="text" name="extra_names" placeholder="{{ t.extra_friend_placeholder }}"
                style="font-size:20px;width:420px;padding:8px;">
            <p style="color:#777;">{{ t.extra_friend_tip }}</p>
        </form>
    </div>

    <div class="card">
        <h2>{{ t.today_recorded }}</h2>

        <table>
            <tr>
                <th>{{ t.name }}</th>
                <th>{{ t.identity }}</th>
                <th>{{ t.topic }}</th>
                <th>{{ t.time }}</th>
                <th>{{ t.operation }}</th>
            </tr>

            {% for r in today_records %}
            <tr>
                <td>{{ r["姓名"] }}</td>
                <td>{{ r["身份"] }}</td>
                <td>{{ r["主题"] }}</td>
                <td>{{ r["时间"] }}</td>
                <td>
                    <a href="/reading_edit/{{ r['id'] }}"><button class="edit">{{ t.edit }}</button></a>
                    <a href="/reading_delete/{{ r['id'] }}" onclick="return confirm('{{ t.delete_confirm_simple }}')">
                        <button class="delete">{{ t.delete_record }}</button>
                    </a>
                </td>
            </tr>
            {% endfor %}
        </table>

        <h3 style="margin-top:20px;">{{ t.today_study_count }}</h3>

        <table>
            <tr>
                <th>{{ t.name }}</th>
                <th>{{ t.identity }}</th>
                <th>{{ t.count }}</th>
            </tr>

            {% for r in today_summary %}
            <tr>
                <td>{{ r["姓名"] }}</td>
                <td>{{ r["身份"] }}</td>
                <td>{{ r["共修次数"] }}</td>
            </tr>
            {% endfor %}
        </table>
    </div>

    <script>
    function selectAllNames() {
        const boxes = document.querySelectorAll('input[name="names"]');
        boxes.forEach(b => b.checked = true);
    }
    </script>

    </body>
    </html>
    """

    return render_template_string(
        html,
        t=t,
        attendees=attendees,
        today_records=today_records,
        today_summary=today_summary_records,
        topic_options=topic_options
    )

@reading_bp.route("/reading_delete/<int:record_id>")
def reading_delete(record_id):
    db_query("""
        delete from reading
        where id = %s
    """, (record_id,))

    return redirect(url_for("reading.reading"))

@reading_bp.route("/reading_edit/<int:record_id>", methods=["GET", "POST"])
def reading_edit(record_id):
    row = db_query("""
        select *
        from reading
        where id = %s
    """, (record_id,), fetchone=True)

    if not row:
        return "找不到这笔记录<br><a href='/reading'>返回</a>"

    if request.method == "POST":
        new_topic = request.form.get("topic", "").strip()
        new_session = request.form.get("session", "").strip()

        db_query("""
            update reading
            set topic = %s,
                session = %s
            where id = %s
        """, (new_topic, new_session, record_id))

        return redirect(url_for("reading"))

    html = """
    <h2>修改白话佛法记录</h2>
    <form method="post">
        <p>姓名：{{ name }}</p>
        <p>身份：{{ identity }}</p>

        <p>主题：</p>
        <input name="topic" value="{{ topic }}" style="font-size:22px;width:320px;">

        <p>场次 / 备注：</p>
        <input name="session" value="{{ session }}" style="font-size:22px;width:320px;">

        <br><br>
        <button type="submit" style="font-size:22px;">保存修改</button>
    </form>
    <br>
    <a href="/reading">返回</a>
    """

    return render_template_string(
        html,
        name=row.get("name"),
        identity=row.get("identity"),
        topic=row.get("topic") or "",
        session=row.get("session") or ""
    )