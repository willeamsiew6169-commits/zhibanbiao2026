# dharma_class_web.py


from io import BytesIO
from db import get_conn
from flask import send_file
from zoneinfo import ZoneInfo
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from psycopg2.extras import RealDictCursor
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from flask import Blueprint, render_template_string, request, redirect, url_for, flash, send_file


dharma_class_bp = Blueprint(
    "dharma_class",
    __name__,
    url_prefix="/class"
)


STATUS_LABELS = {
    "present": "出席",
    "absent": "缺席",
    "late": "迟到",
    "farm": "农舍",
}

BAIHUA_STATUS_LABELS = {
    "submitted_done": "有交",
    "missing": "没交",
    "submitted_no_answer": "有交，没做题",
    "absent": "缺席",
}

SCRIPTURE_STATUS_LABELS = {
    "submitted_recited": "有交，有念",
    "submitted_not_recited": "有交，没念",
    "missing": "没交",
    "absent": "缺席",
}

def beautify_simple_excel(ws):

    header_fill = PatternFill("solid", fgColor="5B9BD5")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            value = str(value) if value is not None else ""
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 24

@dharma_class_bp.route("/")
def class_home():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    malaysia_now = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    )

    today = malaysia_now.date()
    today_str = today.isoformat()

    weekday_names = [
        "星期一",
        "星期二",
        "星期三",
        "星期四",
        "星期五",
        "星期六",
        "星期日"
    ]

    today_display = (
        f"{today.year}年"
        f"{today.month}月"
        f"{today.day}日 "
        f"{weekday_names[today.weekday()]}"
    )

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 每个组别今日统计
            cur.execute("""
                select
                    g.id,
                    g.name,
                    g.sort_order,

                    count(distinct s.id) as total_students,

                    count(
                        distinct case
                            when a.id is not null
                            then s.id
                        end
                    ) as marked_count,

                    count(
                        distinct case
                            when a.status in ('present', 'late', 'farm')
                            then s.id
                        end
                    ) as attended_count,

                    count(
                        distinct case
                            when a.status = 'present'
                            then s.id
                        end
                    ) as present_count,

                    count(
                        distinct case
                            when a.status = 'late'
                            then s.id
                        end
                    ) as late_count,

                    count(
                        distinct case
                            when a.status = 'farm'
                            then s.id
                        end
                    ) as farm_count,

                    count(
                        distinct case
                            when a.status in ('absent', 'leave')
                            then s.id
                        end
                    ) as absent_count,

                    count(distinct l.id) as lesson_count

                from dharma_class_groups g

                left join dharma_students s
                    on s.group_id = g.id
                   and s.status = 'active'
                   and s.branch = 'CHE'

                left join dharma_attendance a
                    on a.student_id = s.id
                   and a.class_date = %s
                   and a.branch = 'CHE'

                left join dharma_class_lessons l
                    on l.group_id = g.id
                   and l.lesson_date = %s
                   and l.branch = 'CHE'

                where g.is_active = true

                group by
                    g.id,
                    g.name,
                    g.sort_order

                order by
                    g.sort_order,
                    g.id
            """, (
                today_str,
                today_str
            ))

            group_stats = cur.fetchall()

            # 首页总统计
            total_students = sum(
                int(g["total_students"] or 0)
                for g in group_stats
            )

            total_marked = sum(
                int(g["marked_count"] or 0)
                for g in group_stats
            )

            total_attended = sum(
                int(g["attended_count"] or 0)
                for g in group_stats
            )

            total_lessons = sum(
                int(g["lesson_count"] or 0)
                for g in group_stats
            )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>蕉赖佛学班系统</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>
.class-dashboard-note {
    background: linear-gradient(135deg, #fff8e8, #fffdf6);
    border: 1px solid #f3d98c;
    border-radius: 16px;
    padding: 16px;
    margin: 16px 0 20px;
    text-align: center;
    font-size: 18px;
    line-height: 1.6;
}

.overall-summary {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin: 18px 0 22px;
}

.overall-box {
    background: #ffffff;
    border-radius: 16px;
    padding: 16px 10px;
    text-align: center;
    border: 1px solid #edf0f3;
    position: relative;
    overflow: hidden;
}

.overall-box::before {
    content: "";
    position: absolute;
    left: 0;
    right: 0;
    top: 0;
    height: 5px;
    background: #60a5fa;
}

.overall-box:nth-child(2)::before { background: #a78bfa; }
.overall-box:nth-child(3)::before { background: #34d399; }
.overall-box:nth-child(4)::before { background: #fbbf24; }

.overall-label {
    color: #666;
    font-size: 16px;
}

.overall-number {
    font-size: 30px;
    font-weight: 800;
    margin-top: 6px;
}

.group-dashboard-grid{
    display:grid;
    grid-template-columns:repeat(3, 1fr);
    gap:18px;
    margin-top:16px;
}

@media (max-width:900px){
    .group-dashboard-grid{
        grid-template-columns:repeat(2,1fr);
    }
}

@media (max-width:600px){
    .group-dashboard-grid{
        grid-template-columns:1fr;
    }
}

.group-dashboard-card{
    border:1px solid #e5e7eb;
    border-top:5px solid #60a5fa;
    border-radius:18px;
    padding:18px;
    background:#fff;

    display:flex;
    flex-direction:column;
    justify-content:space-between;

    min-height:260px;
}

.group-dashboard-card:nth-child(2) { border-top-color:#f59e0b; }
.group-dashboard-card:nth-child(3) { border-top-color:#a78bfa; }

.group-dashboard-name {
    display:inline-block;
    width:fit-content;
    padding:7px 12px;
    border-radius:10px;
    background:#eef6ff;
    color:#1d4ed8;
    font-size: 22px;
    font-weight: 800;
    margin-bottom: 12px;
}

.group-main-number {
    font-size: 30px;
    font-weight: 800;
    margin-bottom: 8px;
}

.group-detail {
    color: #666;
    font-size: 16px;
    line-height: 1.7;
}

.group-status {
    display: inline-block;
    border-radius: 999px;
    padding: 6px 12px;
    margin-top: 12px;
    font-size: 15px;
    font-weight: 700;
}

.status-complete {
    background: #dcfce7;
    color: #166534;
}

.status-pending {
    background: #fef3c7;
    color: #92400e;
}

.status-no-class {
    background: #f3f4f6;
    color: #555;
}

.home-section-title {
    font-size: 21px;
    font-weight: 800;
    margin: 24px 0 10px;
}

@media (max-width: 700px) {
    .overall-summary {
        grid-template-columns: repeat(2, 1fr);
    }

    .group-dashboard-grid {
        grid-template-columns: 1fr;
    }
}
</style>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            📘 蕉赖佛学班系统
        </h1>

        <p class="page-subtitle">
            学生资料、课程记录、点名与出席统计。<br>
            <span style="color:#8a6d3b;font-weight:700;">
                🌸 学习佛法・增长智慧・培养慈悲 🌸
            </span>
        </p>

        <div class="class-dashboard-note">
            📅 今日：{{ today_display }}
        </div>

        <div class="overall-summary">

            <div class="overall-box">
                <div class="overall-label">
                    在读学生
                </div>

                <div class="overall-number">
                    {{ total_students }}
                </div>
            </div>

            <div class="overall-box">
                <div class="overall-label">
                    今日已点名
                </div>

                <div class="overall-number">
                    {{ total_marked }}
                </div>
            </div>

            <div class="overall-box">
                <div class="overall-label">
                    今日出席
                </div>

                <div class="overall-number">
                    {{ total_attended }}
                </div>
            </div>

            <div class="overall-box">
                <div class="overall-label">
                    今日课程
                </div>

                <div class="overall-number">
                    {{ total_lessons }}
                </div>
            </div>

        </div>

        <h2 class="section-title">
            📊 今日各组情况
        </h2>

        {% if group_stats %}

            <div class="group-dashboard-grid">

                {% for g in group_stats %}

                    <div class="group-dashboard-card">

                        <div class="group-dashboard-name">
                            {{ g.name }}
                        </div>

                        <div class="group-main-number">
                            {{ g.attended_count or 0 }}
                            /
                            {{ g.total_students or 0 }}
                        </div>

                        <div class="group-detail">
                            已点名：{{ g.marked_count or 0 }} 位<br>
                            出席：{{ g.present_count or 0 }} 位<br>
                            迟到：{{ g.late_count or 0 }} 位<br>
                            农舍：{{ g.farm_count or 0 }} 位<br>
                            缺席：{{ g.absent_count or 0 }} 位
                        </div>

                        {% if
                            g.lesson_count
                            and g.total_students
                            and g.marked_count == g.total_students
                        %}
                            <div class="group-status status-complete">
                                ✅ 课程与点名已完成
                            </div>

                        {% elif g.lesson_count or g.marked_count %}
                            <div class="group-status status-pending">
                                ⏳ 记录尚未完成
                            </div>

                        {% else %}
                            <div class="group-status status-no-class">
                                尚未建立今日记录
                            </div>
                        {% endif %}

                    </div>

                {% endfor %}

            </div>

        {% else %}

            <div class="empty-state">
                暂时没有启用中的佛学班组别。
            </div>

        {% endif %}

        <div class="home-section-title">
            每周主要功能
        </div>

        <div class="btn-row">
            <a
                class="btn-tool btn-primary"
                href="{{ url_for(
                    'dharma_class.class_attendance'
                ) }}"
            >
                📚 今日上课
            </a>
        </div>

        <div class="btn-row">

            <a
                class="btn-tool btn-purple"
                href="{{ url_for(
                    'dharma_class.class_lessons'
                ) }}"
            >
                📖 课程记录
            </a>

            <a
                class="btn-tool btn-purple"
                href="/class/records"
            >
                📅 点名记录
            </a>

        </div>

        <div class="home-section-title">
            学生与统计管理
        </div>

        <div class="btn-row">

            <a
                class="btn-tool btn-warning"
                href="{{ url_for(
                    'dharma_class.class_students'
                ) }}"
            >
                👧 学生名单
            </a>

            <a
                class="btn-tool btn-warning"
                href="/class/reports"
            >
                📊 出席统计
            </a>

        </div>

        <div class="btn-row">
            <a
                class="btn-tool btn-warning"
                href="/class/promote"
            >
                🎓 年度升班
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        today_str=today_str,
        today_display=today_display,
        group_stats=group_stats,
        total_students=total_students,
        total_marked=total_marked,
        total_attended=total_attended,
        total_lessons=total_lessons
    )


@dharma_class_bp.route("/attendance", methods=["GET", "POST"])
def class_attendance():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    malaysia_today = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).date()

    selected_date = (
        request.values.get("class_date")
        or malaysia_today.isoformat()
    )

    selected_group_id = (
        request.values.get("group_id")
        or ""
    ).strip()

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # =========================================================
            # 读取启用组别
            # =========================================================
            cur.execute("""
                select
                    id,
                    name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)

            groups = cur.fetchall()

            if not selected_group_id and groups:
                selected_group_id = str(groups[0]["id"])

            lesson = {
                "teacher_name": "",
                "topic": "",
                "material": "",
                "content": "",
                "remark": ""
            }

            # =========================================================
            # 保存课程、点名与功课
            # =========================================================
            if request.method == "POST":

                teacher_name = request.form.get(
                    "teacher_name",
                    ""
                ).strip()

                topic = request.form.get(
                    "topic",
                    ""
                ).strip()

                material = request.form.get(
                    "material",
                    ""
                ).strip()

                content = request.form.get(
                    "content",
                    ""
                ).strip()

                lesson_remark = request.form.get(
                    "lesson_remark",
                    ""
                ).strip()

                student_ids = request.form.getlist(
                    "student_id"
                )

                lesson = {
                    "teacher_name": teacher_name,
                    "topic": topic,
                    "material": material,
                    "content": content,
                    "remark": lesson_remark
                }

                if not selected_group_id:
                    flash("请选择组别。", "bad")

                elif not student_ids:
                    flash("这个组别没有可保存的学生。", "bad")

                else:

                    has_lesson_data = any([
                        teacher_name,
                        topic,
                        material,
                        content,
                        lesson_remark
                    ])

                    # -------------------------------------------------
                    # 有填写课程资料才保存课程
                    # 补录旧点名时可以全部留空
                    # -------------------------------------------------
                    if has_lesson_data:

                        cur.execute("""
                            insert into dharma_class_lessons
                            (
                                branch,
                                lesson_date,
                                group_id,
                                teacher_name,
                                topic,
                                material,
                                content,
                                remark
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s
                            )
                            on conflict
                            (
                                branch,
                                lesson_date,
                                group_id
                            )
                            do update set
                                teacher_name = excluded.teacher_name,
                                topic = excluded.topic,
                                material = excluded.material,
                                content = excluded.content,
                                remark = excluded.remark
                        """, (
                            selected_date,
                            selected_group_id,
                            teacher_name or None,
                            topic or None,
                            material or None,
                            content or None,
                            lesson_remark or None
                        ))

                    marked_by = teacher_name or "老师"

                    # -------------------------------------------------
                    # 保存每位学生
                    # -------------------------------------------------
                    for sid in student_ids:

                        attendance_status = request.form.get(
                            f"status_{sid}",
                            "present"
                        ).strip()

                        # 防止异常状态写入
                        if attendance_status not in (
                            "present",
                            "absent",
                            "late",
                            "farm"
                        ):
                            attendance_status = "present"

                        baihua_status = request.form.get(
                            f"baihua_{sid}",
                            ""
                        ).strip()

                        scripture_status = request.form.get(
                            f"scripture_{sid}",
                            ""
                        ).strip()

                        # 缺席时，两项功课都记录为缺席。
                        if attendance_status == "absent":
                            baihua_status = "absent"
                            scripture_status = "absent"

                        if baihua_status not in (
                            "",
                            "submitted_done",
                            "submitted_no_answer",
                            "missing",
                            "absent"
                        ):
                            baihua_status = ""

                        if scripture_status not in (
                            "",
                            "submitted_recited",
                            "submitted_not_recited",
                            "missing",
                            "absent"
                        ):
                            scripture_status = ""

                        # ---------------------------------------------
                        # 保存出席：present / absent / late / farm
                        # ---------------------------------------------
                        cur.execute("""
                            insert into dharma_attendance
                            (
                                branch,
                                class_date,
                                student_id,
                                group_id,
                                status,
                                remark,
                                marked_by
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                %s,
                                %s,
                                null,
                                %s
                            )
                            on conflict
                            (
                                class_date,
                                student_id
                            )
                            do update set
                                group_id = excluded.group_id,
                                status = excluded.status,
                                marked_by = excluded.marked_by,
                                marked_at = now()
                        """, (
                            selected_date,
                            sid,
                            selected_group_id,
                            attendance_status,
                            marked_by
                        ))

                        # ---------------------------------------------
                        # 至少记录了一项功课，才写入功课表
                        # ---------------------------------------------
                        if baihua_status or scripture_status:

                            cur.execute("""
                                insert into dharma_homework
                                (
                                    branch,
                                    class_date,
                                    student_id,
                                    group_id,
                                    baihua_status,
                                    scripture_status,
                                    recorded_by,
                                    updated_at
                                )
                                values
                                (
                                    'CHE',
                                    %s,
                                    %s,
                                    %s,
                                    %s,
                                    %s,
                                    %s,
                                    now()
                                )
                                on conflict
                                (
                                    branch,
                                    class_date,
                                    student_id
                                )
                                do update set
                                    group_id = excluded.group_id,
                                    baihua_status =
                                        excluded.baihua_status,
                                    scripture_status =
                                        excluded.scripture_status,
                                    recorded_by =
                                        excluded.recorded_by,
                                    updated_at = now()
                            """, (
                                selected_date,
                                sid,
                                selected_group_id,
                                baihua_status or None,
                                scripture_status or None,
                                marked_by
                            ))

                    conn.commit()

                    if has_lesson_data:
                        flash(
                            "课程、点名和学生功课已保存。",
                            "good"
                        )
                    else:
                        flash(
                            "学生点名和功课已保存。",
                            "good"
                        )

                    return redirect(
                        url_for(
                            "dharma_class.class_attendance",
                            class_date=selected_date,
                            group_id=selected_group_id
                        )
                    )

            # =========================================================
            # 读取已有课程
            # =========================================================
            cur.execute("""
                select
                    teacher_name,
                    topic,
                    material,
                    content,
                    remark
                from dharma_class_lessons
                where branch = 'CHE'
                  and lesson_date = %s
                  and group_id = %s
                limit 1
            """, (
                selected_date,
                selected_group_id
            ))

            saved_lesson = cur.fetchone()

            if saved_lesson:
                lesson = saved_lesson

            # =========================================================
            # 读取学生和点名状态
            # =========================================================
            cur.execute("""
                select
                    s.id,
                    s.student_no,
                    s.name,
                    s.english_name,

                    coalesce(
                        a.status,
                        'present'
                    ) as attendance_status

                from dharma_students s

                left join dharma_attendance a
                    on a.student_id = s.id
                   and a.class_date = %s
                   and a.branch = 'CHE'

                where s.status = 'active'
                  and s.branch = 'CHE'
                  and s.group_id = %s

                order by
                    s.student_no nulls last,
                    s.name
            """, (
                selected_date,
                selected_group_id
            ))

            students = cur.fetchall()

            # =========================================================
            # 读取学生功课
            # =========================================================
            cur.execute("""
                select
                    student_id,
                    baihua_status,
                    scripture_status
                from dharma_homework
                where branch = 'CHE'
                  and class_date = %s
                  and group_id = %s
            """, (
                selected_date,
                selected_group_id
            ))

            homework_rows = cur.fetchall()

            homework_map = {
                row["student_id"]: row
                for row in homework_rows
            }

            for student in students:

                homework = homework_map.get(
                    student["id"]
                )

                # 兼容旧资料：旧 submitted 自动转为新的完整状态。
                if student["attendance_status"] == "leave":
                    student["attendance_status"] = "absent"

                if homework:
                    baihua_value = homework["baihua_status"]
                    scripture_value = homework["scripture_status"]

                    if baihua_value == "submitted":
                        baihua_value = "submitted_done"

                    if scripture_value == "submitted":
                        scripture_value = "submitted_recited"

                    student["baihua_status"] = baihua_value
                    student["scripture_status"] = scripture_value

                else:
                    if student["attendance_status"] == "absent":
                        student["baihua_status"] = "absent"
                        student["scripture_status"] = "absent"
                    else:
                        student["baihua_status"] = None
                        student["scripture_status"] = None

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>

<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>今日上课与点名</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>

/* =========================================================
   页面资料提示
   ========================================================= */

.current-selection {
    background: #eef6ff;
    border: 1px solid #cfe3ff;
    border-radius: 16px;
    padding: 14px 16px;
    margin-top: 16px;
    text-align: center;
    font-size: 18px;
    font-weight: 800;
}

/* =========================================================
   课程记录
   ========================================================= */

.lesson-note {
    background: #f6f7f9;
    border-radius: 14px;
    padding: 14px 16px;
    margin-bottom: 18px;
    line-height: 1.6;
}

/* =========================================================
   快速操作
   ========================================================= */

.quick-actions {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 10px;
    margin-bottom: 22px;
}

.quick-actions .btn-tool {
    min-height: 58px;
    font-size: 18px;
}

/* =========================================================
   学生卡片
   ========================================================= */

.student-list {
    display: grid;
    gap: 18px;
}

.student-card {
    border: 1px solid rgba(0, 0, 0, 0.08);
    border-radius: 20px;
    padding: 20px;
    box-shadow: 0 3px 10px rgba(0, 0, 0, 0.04);
}

/* 四种柔和底色循环 */

.student-card:nth-child(4n+1) {
    background: #f0f7ff;
}

.student-card:nth-child(4n+2) {
    background: #fff2f6;
}

.student-card:nth-child(4n+3) {
    background: #f4fbf1;
}

.student-card:nth-child(4n+4) {
    background: #fffbed;
}

.student-header {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 16px;
}

.student-number {
    display: flex;
    align-items: center;
    justify-content: center;
    width: 38px;
    height: 38px;
    border-radius: 50%;
    background: rgba(255, 255, 255, 0.85);
    font-size: 18px;
    font-weight: 900;
    flex: 0 0 38px;
}

.student-name {
    font-size: 24px;
    font-weight: 900;
}

.student-english-name {
    color: #777;
    font-size: 17px;
    font-weight: 400;
    margin-left: 6px;
}

/* =========================================================
   区块标题
   ========================================================= */

.student-section {
    border-top: 1px solid rgba(0, 0, 0, 0.09);
    padding-top: 15px;
    margin-top: 15px;
}

.student-section-title {
    font-size: 18px;
    font-weight: 900;
    margin-bottom: 10px;
}

/* =========================================================
   出席按钮
   ========================================================= */

.attendance-options {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 9px;
}

.homework-options {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 9px;
}

.choice-option {
    position: relative;
    display: flex;
    align-items: center;
    justify-content: center;
    min-height: 52px;
    border: 1px solid #d7dce2;
    border-radius: 14px;
    padding: 10px 8px;
    background: rgba(255, 255, 255, 0.82);
    cursor: pointer;
    font-size: 18px;
    font-weight: 700;
    text-align: center;
}

.choice-option input {
    position: absolute;
    opacity: 0;
    pointer-events: none;
}

.choice-option:has(input:checked) {
    border: 2px solid #2563eb;
    background: #eaf2ff;
    color: #174ea6;
    box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.08);
}

/* 不同状态选中颜色 */

.choice-present:has(input:checked) {
    border-color: #16a34a;
    background: #dcfce7;
    color: #166534;
}

.choice-late:has(input:checked) {
    border-color: #d97706;
    background: #fef3c7;
    color: #92400e;
}

.choice-farm:has(input:checked) {
    border-color: #65a30d;
    background: #ecfccb;
    color: #3f6212;
}

.choice-absent:has(input:checked) {
    border-color: #dc2626;
    background: #fee2e2;
    color: #991b1b;
}

.choice-submitted:has(input:checked) {
    border-color: #16a34a;
    background: #dcfce7;
    color: #166534;
}

.choice-missing:has(input:checked) {
    border-color: #dc2626;
    background: #fee2e2;
    color: #991b1b;
}

.choice-partial:has(input:checked) {
    border-color: #d97706;
    background: #fff7d6;
    color: #92400e;
}

.choice-homework-absent:has(input:checked) {
    border-color: #64748b;
    background: #e2e8f0;
    color: #334155;
}
                                  
.floating-toolbar{
    position:fixed;
    left:0;
    right:0;
    bottom:0;

    display:grid;
    grid-template-columns:repeat(3,1fr);
    gap:8px;

    padding:8px 10px;

    background:#fff;
    border-top:1px solid #ddd;
    box-shadow:0 -3px 10px rgba(0,0,0,.08);

    z-index:999;
}

.floating-toolbar .btn-tool{

    min-height:46px;

    font-size:18px;

    border-radius:14px;

    padding:8px 10px;

}

.page{

    padding-bottom:72px;

}

@media (max-width: 480px) {
    .floating-toolbar {
        gap: 7px;
        padding-left: 8px;
        padding-right: 8px;
    }

    .floating-toolbar .btn-tool {
        font-size: 15px;
        padding-left: 5px;
        padding-right: 5px;
    }
}

/* =========================================================
   手机
   ========================================================= */

@media (max-width: 800px) {

    .quick-actions {
        grid-template-columns: 1fr;
    }

    .attendance-options {
        grid-template-columns: repeat(2, 1fr);
    }

    .student-card {
        padding: 16px;
    }
}

@media (max-width: 480px) {

    .homework-options {
        grid-template-columns: 1fr 1fr;
    }

    .student-name {
        font-size: 21px;
    }

    .choice-option {
        font-size: 16px;
    }
}

</style>
</head>

<body>

<div class="page">

    <!-- =====================================================
         日期与组别
         ===================================================== -->

    <div class="card">

        <h1 class="page-title">
            📚 今日上课
        </h1>

        <p class="page-subtitle">
            课程、点名、白话佛法功课及经文功课。
        </p>

        {% with messages =
            get_flashed_messages(with_categories=true)
        %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form
            method="get"
            id="attendanceFilterForm"
        >

            <div class="form-group">
                <label class="form-label">
                    日期
                </label>

                <input
                    class="form-input"
                    type="date"
                    name="class_date"
                    value="{{ selected_date }}"
                    onchange="
                        document
                            .getElementById(
                                'attendanceFilterForm'
                            )
                            .submit()
                    "
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                    onchange="
                        document
                            .getElementById(
                                'attendanceFilterForm'
                            )
                            .submit()
                    "
                    required
                >
                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                selected_group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

        </form>

        <div class="current-selection">
            当前记录日期：{{ selected_date }}
        </div>

    </div>

    {% if students %}

    <form method="post" id="attendanceForm">

        <input
            type="hidden"
            name="class_date"
            value="{{ selected_date }}"
        >

        <input
            type="hidden"
            name="group_id"
            value="{{ selected_group_id }}"
        >

        <!-- =================================================
             课程资料
             ================================================= -->

        <div class="card">

            <h2 class="section-title">
                📖 课程记录
            </h2>

            <div class="lesson-note">
                补录以前的点名时，老师和课题可以留空。
                有填写课程资料时，系统才会建立课程记录。
            </div>

            <div class="form-group">
                <label class="form-label">
                    负责老师
                </label>

                <input
                    class="form-input"
                    name="teacher_name"
                    value="{{ lesson.teacher_name or '' }}"
                    placeholder="补录旧点名时可不填"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    今日课题
                </label>

                <input
                    class="form-input"
                    name="topic"
                    value="{{ lesson.topic or '' }}"
                    placeholder="补录旧点名时可不填"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    教材／教学资源
                </label>

                <input
                    class="form-input"
                    name="material"
                    value="{{ lesson.material or '' }}"
                    placeholder="例如：故事、影片、白话佛法"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    教学内容
                </label>

                <textarea
                    class="form-input"
                    name="content"
                    rows="3"
                    placeholder="可不填"
                >{{ lesson.content or '' }}</textarea>
            </div>

            <div class="form-group">
                <label class="form-label">
                    课程备注
                </label>

                <textarea
                    class="form-input"
                    name="lesson_remark"
                    rows="2"
                    placeholder="可不填"
                >{{ lesson.remark or '' }}</textarea>
            </div>

        </div>

        <!-- =================================================
             学生点名与功课
             ================================================= -->

        <div class="card">

            <h2 class="section-title">
                📝 学生点名与功课
            </h2>

            <div class="quick-actions">

                <button
                    class="btn-tool btn-success"
                    type="button"
                    onclick="markAllPresent()"
                >
                    ✅ 全部出席
                </button>

                <button
                    class="btn-tool btn-purple"
                    type="button"
                    onclick="markAllBaihuaSubmitted()"
                >
                    📚 白话全部有交
                </button>

                <button
                    class="btn-tool btn-primary"
                    type="button"
                    onclick="markAllScriptureSubmitted()"
                >
                    📿 经文全部有交有念
                </button>

            </div>

            <div class="student-list">

                {% for s in students %}

                    <div class="student-card">

                        <input
                            type="hidden"
                            name="student_id"
                            value="{{ s.id }}"
                        >

                        <div class="student-header">

                            <div class="student-number">
                                {{ loop.index }}
                            </div>

                            <div class="student-name">

                                {{ s.name }}

                                {% if s.english_name %}
                                    <span class="student-english-name">
                                        {{ s.english_name }}
                                    </span>
                                {% endif %}

                            </div>

                        </div>

                        <!-- 出席 -->

                        <div class="student-section">

                            <div class="student-section-title">
                                📝 出席情况
                            </div>

                            <div class="attendance-options">

                                {% for key, label in status_labels.items() %}

                                    <label class="
                                        choice-option
                                        {% if key == 'present' %}
                                            choice-present
                                        {% elif key == 'late' %}
                                            choice-late
                                        {% elif key == 'farm' %}
                                            choice-farm
                                        {% elif key == 'absent' %}
                                            choice-absent
                                        {% endif %}
                                    ">

                                        <input
                                            type="radio"
                                            name="status_{{ s.id }}"
                                            value="{{ key }}"
                                            {% if
                                                s.attendance_status
                                                == key
                                            %}
                                                checked
                                            {% endif %}
                                        >

                                        {% if key == "present" %}
                                            ✅
                                        {% elif key == "late" %}
                                            ⏰
                                        {% elif key == "farm" %}
                                            🌱
                                        {% elif key == "absent" %}
                                            ❌
                                        {% endif %}

                                        {{ label }}

                                    </label>

                                {% endfor %}

                            </div>

                        </div>

                        <!-- 白话佛法 -->

                        <div class="student-section">

                            <div class="student-section-title">
                                📚 白话佛法功课
                            </div>

                            <div class="homework-options">

                                <label class="choice-option choice-submitted">
                                    <input type="radio" name="baihua_{{ s.id }}"
                                           value="submitted_done"
                                           {% if s.baihua_status == "submitted_done" %}checked{% endif %}>
                                    ✓ 有交
                                </label>

                                <label class="choice-option choice-missing">
                                    <input type="radio" name="baihua_{{ s.id }}"
                                           value="missing"
                                           {% if s.baihua_status == "missing" %}checked{% endif %}>
                                    ✗ 没交
                                </label>

                                <label class="choice-option choice-partial">
                                    <input type="radio" name="baihua_{{ s.id }}"
                                           value="submitted_no_answer"
                                           {% if s.baihua_status == "submitted_no_answer" %}checked{% endif %}>
                                    ○ 有交，没做题
                                </label>

                                <label class="choice-option choice-homework-absent">
                                    <input type="radio" name="baihua_{{ s.id }}"
                                           value="absent"
                                           {% if s.baihua_status == "absent" %}checked{% endif %}>
                                    缺 缺席
                                </label>

                            </div>

                        </div>

                        <!-- 念经文 -->

                        <div class="student-section">

                            <div class="student-section-title">
                                📿 念经文功课
                            </div>

                            <div class="homework-options">

                                <label class="choice-option choice-submitted">
                                    <input type="radio" name="scripture_{{ s.id }}"
                                           value="submitted_recited"
                                           {% if s.scripture_status == "submitted_recited" %}checked{% endif %}>
                                    ✓ 有交，有念
                                </label>

                                <label class="choice-option choice-partial">
                                    <input type="radio" name="scripture_{{ s.id }}"
                                           value="submitted_not_recited"
                                           {% if s.scripture_status == "submitted_not_recited" %}checked{% endif %}>
                                    ○ 有交，没念
                                </label>

                                <label class="choice-option choice-missing">
                                    <input type="radio" name="scripture_{{ s.id }}"
                                           value="missing"
                                           {% if s.scripture_status == "missing" %}checked{% endif %}>
                                    ✗ 没交
                                </label>

                                <label class="choice-option choice-homework-absent">
                                    <input type="radio" name="scripture_{{ s.id }}"
                                           value="absent"
                                           {% if s.scripture_status == "absent" %}checked{% endif %}>
                                    缺 缺席
                                </label>

                            </div>

                        </div>

                    </div>

                {% endfor %}

            </div>

            <div class="btn-row">

                <button
                    class="btn-tool btn-success"
                    type="submit"
                >
                    💾 保存课程、点名与功课
                </button>

            </div>

        </div>

    </form>

    {% else %}

        <div class="card">

            <div class="empty-state">
                这个组别目前没有在读学生。
            </div>

        </div>

    {% endif %}

    <div class="card normal-bottom-actions">

        <div class="btn-row">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_home'
                ) }}"
            >
                ⬅ 返回佛学班首页
            </a>

        </div>

    </div>

</div>

{% if students %}

<div class="floating-toolbar">

    <a
        class="btn-tool btn-secondary"
        href="{{ url_for(
            'dharma_class.class_home'
        ) }}"
    >
        ⬅ 返回
    </a>

    <button
        class="btn-tool btn-primary"
        type="button"
        onclick="goTop()"
    >
        ⬆ 顶部
    </button>

    <button
        class="btn-tool btn-success"
        type="submit"
        form="attendanceForm"
    >
        💾 保存
    </button>

</div>

{% endif %}

<script>

/* =========================================================
   全部出席
   ========================================================= */

function markAllPresent() {

    document
        .querySelectorAll(
            'input[type="radio"][value="present"]'
        )
        .forEach(function(radio) {
            radio.checked = true;
        });
}


/* =========================================================
   白话佛法全部已交
   同时自动设为出席
   ========================================================= */

function markAllBaihuaSubmitted() {
    document
        .querySelectorAll(
            'input[name^="baihua_"][value="submitted_done"]'
        )
        .forEach(function(radio) {
            radio.checked = true;
            const studentId = radio.name.replace("baihua_", "");
            markStudentPresentIfNeeded(studentId);
        });
}

function markAllScriptureSubmitted() {
    document
        .querySelectorAll(
            'input[name^="scripture_"][value="submitted_recited"]'
        )
        .forEach(function(radio) {
            radio.checked = true;
            const studentId = radio.name.replace("scripture_", "");
            markStudentPresentIfNeeded(studentId);
        });
}

function markStudentPresentIfNeeded(studentId) {
    const selected = document.querySelector(
        'input[name="status_' + studentId + '"]:checked'
    );

    // 已经是迟到或农舍时，不覆盖老师的真实记录。
    if (selected && (selected.value === "late" || selected.value === "farm")) {
        return;
    }

    const presentRadio = document.querySelector(
        'input[name="status_' + studentId + '"][value="present"]'
    );

    if (presentRadio) {
        presentRadio.checked = true;
    }
}

function markStudentHomeworkAbsent(studentId) {
    const baihuaAbsent = document.querySelector(
        'input[name="baihua_' + studentId + '"][value="absent"]'
    );

    const scriptureAbsent = document.querySelector(
        'input[name="scripture_' + studentId + '"][value="absent"]'
    );

    if (baihuaAbsent) baihuaAbsent.checked = true;
    if (scriptureAbsent) scriptureAbsent.checked = true;
}

function goTop(){

    window.scrollTo({

        top:0,

        behavior:"smooth"

    });

}


/* =========================================================
   初始化智能联动
   ========================================================= */

document.addEventListener(
    "DOMContentLoaded",
    function() {

        // 功课选择任何非缺席状态时，若学生不是迟到或农舍，自动设为出席。
        document
            .querySelectorAll(
                'input[name^="baihua_"], input[name^="scripture_"]'
            )
            .forEach(function(radio) {
                radio.addEventListener("change", function() {
                    if (!this.checked || this.value === "absent") return;

                    const studentId = this.name
                        .replace("baihua_", "")
                        .replace("scripture_", "");

                    markStudentPresentIfNeeded(studentId);
                });
            });

        // 点缺席时，两项功课自动改成缺席。
        document
            .querySelectorAll('input[name^="status_"]')
            .forEach(function(radio) {
                radio.addEventListener("change", function() {
                    if (!this.checked) return;

                    const studentId = this.name.replace("status_", "");

                    if (this.value === "absent") {
                        markStudentHomeworkAbsent(studentId);
                    }
                });
            });
    }
);

</script>

</body>
</html>
""",
        groups=groups,
        students=students,
        selected_date=selected_date,
        selected_group_id=selected_group_id,
        status_labels=STATUS_LABELS,
        baihua_status_labels=BAIHUA_STATUS_LABELS,
        scripture_status_labels=SCRIPTURE_STATUS_LABELS,
        lesson=lesson
    )

@dharma_class_bp.route("/lessons")
def class_lessons():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    selected_year = request.args.get(
        "year",
        str(current_year)
    ).strip()

    selected_month = request.args.get(
        "month",
        ""
    ).strip()

    selected_group_id = request.args.get(
        "group_id",
        ""
    ).strip()

    q = request.args.get(
        "q",
        ""
    ).strip()

    teacher = request.args.get(
        "teacher",
        ""
    ).strip()

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取组别
            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            # 读取课程记录中已有的年份
            cur.execute("""
                select distinct
                    extract(year from lesson_date)::int as lesson_year
                from dharma_class_lessons
                where branch = 'CHE'
                order by lesson_year desc
            """)
            year_rows = cur.fetchall()

            available_years = [
                row["lesson_year"]
                for row in year_rows
                if row["lesson_year"]
            ]

            if current_year not in available_years:
                available_years.insert(0, current_year)

            where_sql = """
                where l.branch = 'CHE'
            """

            params = []

            # 年份筛选
            if selected_year:
                where_sql += """
                    and extract(year from l.lesson_date) = %s
                """
                params.append(int(selected_year))

            # 月份筛选
            if selected_month:
                where_sql += """
                    and extract(month from l.lesson_date) = %s
                """
                params.append(int(selected_month))

            # 组别筛选
            if selected_group_id:
                where_sql += """
                    and l.group_id = %s
                """
                params.append(selected_group_id)

            # 老师筛选
            if teacher:
                teacher_like = f"%{teacher}%"

                where_sql += """
                    and l.teacher_name ilike %s
                """
                params.append(teacher_like)

            # 课题、教材、内容搜索
            if q:
                search_like = f"%{q}%"

                where_sql += """
                    and (
                        l.topic ilike %s
                        or l.material ilike %s
                        or l.content ilike %s
                        or l.remark ilike %s
                    )
                """

                params.extend([
                    search_like,
                    search_like,
                    search_like,
                    search_like
                ])

            cur.execute(f"""
                select
                    l.id,
                    l.lesson_date,
                    l.group_id,
                    l.teacher_name,
                    l.topic,
                    l.material,
                    l.content,
                    l.remark,
                    l.created_at,
                    g.name as group_name,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                    ) as attendance_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status = 'present'
                    ) as present_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status = 'late'
                    ) as late_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status = 'farm'
                    ) as farm_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status in ('absent', 'leave')
                    ) as absent_count

                from dharma_class_lessons l

                left join dharma_class_groups g
                    on g.id = l.group_id

                {where_sql}

                order by
                    l.lesson_date desc,
                    g.sort_order,
                    l.id desc
            """, params)

            lessons = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>课程记录</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>
.lesson-card {
    border: 1px solid #e5e7eb;
    border-radius: 18px;
    padding: 20px;
    margin-bottom: 18px;
    background: #fff;
}

.lesson-header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 14px;
    margin-bottom: 14px;
}

.lesson-date {
    font-size: 22px;
    font-weight: 800;
}

.lesson-group {
    background: #eef2ff;
    border-radius: 999px;
    padding: 7px 14px;
    font-size: 16px;
    font-weight: 700;
    white-space: nowrap;
}

.lesson-topic {
    font-size: 25px;
    font-weight: 800;
    margin: 10px 0;
}

.lesson-meta {
    line-height: 1.8;
    font-size: 18px;
    margin-bottom: 12px;
}

.lesson-content {
    background: #f6f7f9;
    border-radius: 14px;
    padding: 14px 16px;
    line-height: 1.7;
    margin-top: 12px;
    white-space: pre-wrap;
}

.attendance-summary {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 10px;
    margin-top: 16px;
}

.attendance-box {
    background: #f8f9fa;
    border-radius: 14px;
    padding: 12px;
    text-align: center;
}

.attendance-number {
    font-size: 25px;
    font-weight: 800;
    margin-top: 4px;
}

.lesson-actions {
    display: flex;
    gap: 10px;
    margin-top: 16px;
}

.lesson-actions .btn-tool {
    flex: 1;
    min-height: 48px;
    font-size: 17px;
    padding: 10px 14px;
}

.filter-summary {
    background: #f6f7f9;
    border-radius: 14px;
    padding: 14px 16px;
    margin-top: 16px;
    text-align: center;
    font-size: 18px;
    font-weight: 700;
}

@media (max-width: 650px) {
    .lesson-header {
        display: block;
    }

    .lesson-group {
        display: inline-block;
        margin-top: 8px;
    }

    .attendance-summary {
        grid-template-columns: repeat(2, 1fr);
    }

    .lesson-actions {
        display: block;
    }

    .lesson-actions .btn-tool {
        width: 100%;
        margin-bottom: 10px;
    }
}
</style>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            📖 课程记录
        </h1>

        <p class="page-subtitle">
            查询各组历次课程、负责老师及学生出席情况。
        </p>

        <form method="get">

            <div class="form-group">
                <label class="form-label">
                    年份
                </label>

                <select
                    class="form-input"
                    name="year"
                >
                    <option value="">
                        全部年份
                    </option>

                    {% for y in available_years %}
                        <option
                            value="{{ y }}"
                            {% if selected_year|string == y|string %}
                                selected
                            {% endif %}
                        >
                            {{ y }}年
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    月份
                </label>

                <select
                    class="form-input"
                    name="month"
                >
                    <option value="">
                        全部月份
                    </option>

                    {% for m in range(1, 13) %}
                        <option
                            value="{{ m }}"
                            {% if selected_month|string == m|string %}
                                selected
                            {% endif %}
                        >
                            {{ m }}月
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                >
                    <option value="">
                        全部组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                selected_group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    负责老师
                </label>

                <input
                    class="form-input"
                    name="teacher"
                    value="{{ teacher or '' }}"
                    placeholder="例如：陈老师"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    搜索课程
                </label>

                <input
                    class="form-input"
                    name="q"
                    value="{{ q or '' }}"
                    placeholder="课题／教材／教学内容／备注"
                >
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    🔍 搜索课程
                </button>

                <a
                    class="btn-tool btn-secondary"
                    href="{{ url_for(
                        'dharma_class.class_lessons'
                    ) }}"
                >
                    ↺ 清除筛选
                </a>
            </div>

        </form>

        <div class="filter-summary">
            找到 {{ lessons|length }} 笔课程记录
        </div>

    </div>

    <div class="card">

        <h2 class="section-title">
            📚 课程列表
        </h2>

        {% if lessons %}

            {% for lesson in lessons %}

                <div class="lesson-card">

                    <div class="lesson-header">

                        <div class="lesson-date">
                            📅 {{ lesson.lesson_date.strftime("%Y年%m月%d日") }}
                        </div>

                        <div class="lesson-group">
                            {{ lesson.group_name or "未设组别" }}
                        </div>

                    </div>

                    <div class="lesson-topic">
                        📖 {{ lesson.topic }}
                    </div>

                    <div class="lesson-meta">

                        <div>
                            👨‍🏫 负责老师：
                            {{ lesson.teacher_name or "—" }}
                        </div>

                        <div>
                            📚 教材／资源：
                            {{ lesson.material or "—" }}
                        </div>

                    </div>

                    {% if lesson.content %}
                        <div class="lesson-content">
                            <strong>教学内容</strong><br>
                            {{ lesson.content }}
                        </div>
                    {% endif %}

                    {% if lesson.remark %}
                        <div class="lesson-content">
                            <strong>课程备注</strong><br>
                            {{ lesson.remark }}
                        </div>
                    {% endif %}

                    <div class="attendance-summary">

                        <div class="attendance-box">
                            <div>点名人数</div>
                            <div class="attendance-number">
                                {{ lesson.attendance_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>出席</div>
                            <div class="attendance-number">
                                {{ lesson.present_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>迟到</div>
                            <div class="attendance-number">
                                {{ lesson.late_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>农舍</div>
                            <div class="attendance-number">
                                {{ lesson.farm_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>缺席</div>
                            <div class="attendance-number">
                                {{ lesson.absent_count or 0 }}
                            </div>
                        </div>

                    </div>

                    <div class="lesson-actions">

                        <a
                            class="btn-tool btn-primary"
                            href="{{ url_for(
                                'dharma_class.class_attendance',
                                class_date=lesson.lesson_date.isoformat(),
                                group_id=lesson.group_id
                            ) }}"
                        >
                            👁 查看课程与点名
                        </a>

                        <a
                            class="btn-tool btn-warning"
                            href="{{ url_for(
                                'dharma_class.class_attendance',
                                class_date=lesson.lesson_date.isoformat(),
                                group_id=lesson.group_id
                            ) }}"
                        >
                            ✏ 编辑记录
                        </a>

                    </div>

                </div>

            {% endfor %}

        {% else %}

            <div class="empty-state">
                暂时没有符合条件的课程记录。
            </div>

        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="/class"
            >
                ⬅ 返回佛学班首页
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        groups=groups,
        lessons=lessons,
        available_years=available_years,
        selected_year=selected_year,
        selected_month=selected_month,
        selected_group_id=selected_group_id,
        teacher=teacher,
        q=q
    )

@dharma_class_bp.route("/students")
def class_students():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    group_id = request.args.get(
        "group_id",
        ""
    ).strip()

    q = request.args.get(
        "q",
        ""
    ).strip()

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取组别
            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)

            groups = cur.fetchall()

            where_sql = """
                where s.status = 'active'
            """

            params = []

            # 组别筛选
            if group_id:
                where_sql += """
                    and s.group_id = %s
                """
                params.append(group_id)

            # 搜索
            if q:
                like = f"%{q}%"

                where_sql += """
                    and (
                        s.name ilike %s
                        or s.english_name ilike %s
                        or s.parent_name ilike %s
                        or s.parent_phone ilike %s
                        or s.student_no ilike %s
                    )
                """

                params.extend([
                    like,
                    like,
                    like,
                    like,
                    like
                ])

            cur.execute(f"""
                select
                    s.id,
                    s.student_no,
                    s.name,
                    s.english_name,
                    s.gender,
                    s.birth_year,
                    s.parent_name,
                    s.parent_phone,
                    s.group_id,
                    s.status,
                    s.remark,
                    g.name as group_name,
                    g.sort_order as group_sort_order
                from dharma_students s
                left join dharma_class_groups g
                    on g.id = s.group_id
                {where_sql}
                order by
                    g.sort_order,
                    s.student_no nulls last,
                    s.name
            """, params)

            students = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>{{ current_year }}年佛学班学生名单</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>
.student-summary {
    margin-top: 12px;
    padding: 14px 16px;
    border-radius: 14px;
    background: #f6f7f9;
    font-size: 18px;
    font-weight: 700;
    text-align: center;
}

.student-name {
    font-weight: 800;
    white-space: nowrap;
}

.student-age {
    white-space: nowrap;
}

.student-contact {
    min-width: 160px;
}

.student-remark {
    min-width: 150px;
}

.student-table th,
.student-table td {
    vertical-align: middle;
}

.small-action-btn {
    font-size: 16px;
    min-height: 40px;
    padding: 8px 12px;
    white-space: nowrap;
}

@media (max-width: 700px) {
    .student-table {
        min-width: 1200px;
    }
}
</style>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            👧 {{ current_year }}年佛学班学生名单
        </h1>

        <p class="page-subtitle">
            管理佛学班学生资料、组别及监护人联系方式。
        </p>

        {% with messages =
            get_flashed_messages(
                with_categories=true
            )
        %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="get">

            <div class="form-group">
                <label class="form-label">
                    搜索学生
                </label>

                <input
                    class="form-input"
                    name="q"
                    value="{{ q or '' }}"
                    placeholder="姓名／英文名／监护人／电话／编号"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别筛选
                </label>

                <select
                    class="form-input"
                    name="group_id"
                >
                    <option value="">
                        全部组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="btn-row">

                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    🔍 搜索
                </button>

                {% if q or group_id %}
                    <a
                        class="btn-tool btn-secondary"
                        href="{{ url_for(
                            'dharma_class.class_students'
                        ) }}"
                    >
                        ↺ 清除筛选
                    </a>
                {% endif %}

            </div>

        </form>

        <div class="student-summary">
            当前显示：{{ students|length }} 位学生
        </div>

        <div class="btn-row">
            <a
                class="btn-tool btn-success"
                href="{{ url_for(
                    'dharma_class.class_students_add'
                ) }}"
            >
                ➕ 新增学生
            </a>
        </div>

        <div class="btn-row">

            <a
                class="btn-tool btn-purple"
                href="/class/students/import"
            >
                📥 导入学生 Excel
            </a>

            <a
                class="btn-tool btn-secondary"
                href="/class/students/template"
            >
                📄 下载模板
            </a>

        </div>

        <div class="btn-row">
            <a
                class="btn-tool btn-warning"
                href="/class/students/export"
            >
                📤 导出学生名单
            </a>
        </div>

    </div>

    <div class="card">

        <h2 class="section-title">
            学生列表
        </h2>

        {% if students %}

            <div class="table-responsive">

                <table class="record-table student-table">

                    <thead>
                        <tr>
                            <th>编号</th>
                            <th>姓名</th>
                            <th>英文名</th>
                            <th>性别</th>
                            <th>出生年份</th>
                            <th>年龄</th>
                            <th>组别</th>
                            <th>父／母／监护人</th>
                            <th>联系电话</th>
                            <th>备注</th>
                            <th>操作</th>
                        </tr>
                    </thead>

                    <tbody>

                        {% for s in students %}
                        <tr>

                            <td>
                                {{ s.student_no or "—" }}
                            </td>

                            <td class="student-name">
                                {{ s.name }}
                            </td>

                            <td>
                                {{ s.english_name or "—" }}
                            </td>

                            <td>
                                {{ s.gender or "—" }}
                            </td>

                            <td>
                                {{ s.birth_year or "—" }}
                            </td>

                            <td class="student-age">
                                {% if s.birth_year %}
                                    {{ current_year - s.birth_year }}岁
                                {% else %}
                                    —
                                {% endif %}
                            </td>

                            <td>
                                {{ s.group_name or "—" }}
                            </td>

                            <td class="student-contact">
                                {{ s.parent_name or "—" }}
                            </td>

                            <td class="student-contact">
                                {{ s.parent_phone or "—" }}
                            </td>

                            <td class="student-remark">
                                {{ s.remark or "—" }}
                            </td>

                            <td>
                                <a
                                    class="btn-tool btn-warning small-action-btn"
                                    href="{{ url_for(
                                        'dharma_class.class_students_edit',
                                        student_id=s.id
                                    ) }}"
                                >
                                    ✏ 编辑
                                </a>
                            </td>

                        </tr>
                        {% endfor %}

                    </tbody>

                </table>

            </div>

        {% else %}

            <div class="empty-state">
                没有找到符合条件的学生资料。
            </div>

        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="/class"
            >
                ⬅ 返回佛学班首页
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        groups=groups,
        students=students,
        group_id=group_id,
        q=q,
        current_year=current_year
    )

@dharma_class_bp.route("/students/add", methods=["GET", "POST"])
def class_students_add():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    form_data = {
        "student_no": "",
        "name": "",
        "english_name": "",
        "gender": "",
        "birth_year": "",
        "parent_name": "",
        "parent_phone": "",
        "group_id": "",
        "remark": ""
    }

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            if request.method == "POST":

                student_no = request.form.get(
                    "student_no", ""
                ).strip()

                name = request.form.get(
                    "name", ""
                ).strip()

                english_name = request.form.get(
                    "english_name", ""
                ).strip()

                gender = request.form.get(
                    "gender", ""
                ).strip()

                birth_year_text = request.form.get(
                    "birth_year", ""
                ).strip()

                parent_name = request.form.get(
                    "parent_name", ""
                ).strip()

                parent_phone = request.form.get(
                    "parent_phone", ""
                ).strip()

                group_id_text = request.form.get(
                    "group_id", ""
                ).strip()

                remark = request.form.get(
                    "remark", ""
                ).strip()

                birth_year = (
                    int(birth_year_text)
                    if birth_year_text.isdigit()
                    else None
                )

                group_id = (
                    int(group_id_text)
                    if group_id_text.isdigit()
                    else None
                )

                form_data = {
                    "student_no": student_no,
                    "name": name,
                    "english_name": english_name,
                    "gender": gender,
                    "birth_year": birth_year_text,
                    "parent_name": parent_name,
                    "parent_phone": parent_phone,
                    "group_id": group_id_text,
                    "remark": remark
                }

                if not name:
                    flash("学生姓名必须填写。", "bad")

                elif gender not in ("男", "女"):
                    flash("请选择学生性别。", "bad")

                elif birth_year is None:
                    flash("请输入正确的出生年份。", "bad")

                elif birth_year < 2000 or birth_year > current_year:
                    flash(
                        f"出生年份必须介于 2000 至 {current_year} 年。",
                        "bad"
                    )

                elif group_id is None:
                    flash("请选择学生组别。", "bad")

                else:
                    cur.execute("""
                        insert into dharma_students
                        (
                            branch,
                            student_no,
                            name,
                            english_name,
                            gender,
                            birth_year,
                            parent_name,
                            parent_phone,
                            group_id,
                            status,
                            remark
                        )
                        values
                        (
                            'CHE',
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            'active',
                            %s
                        )
                    """, (
                        student_no or None,
                        name,
                        english_name or None,
                        gender,
                        birth_year,
                        parent_name or None,
                        parent_phone or None,
                        group_id,
                        remark or None
                    ))

                    conn.commit()

                    flash("学生已新增。", "good")

                    return redirect(
                        url_for(
                            "dharma_class.class_students"
                        )
                    )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">

<title>新增学生</title>

<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">➕ 新增学生</h1>

        <p class="page-subtitle">
            加入佛学班学生资料。
        </p>

        {% with messages = get_flashed_messages(
            with_categories=true
        ) %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="post">

            <div class="form-group">
                <label class="form-label">
                    学生编号
                </label>

                <input
                    class="form-input"
                    name="student_no"
                    value="{{ form_data.student_no }}"
                    placeholder="可不填，例如：S001"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    中文姓名
                </label>

                <input
                    class="form-input"
                    name="name"
                    value="{{ form_data.name }}"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    英文名
                </label>

                <input
                    class="form-input"
                    name="english_name"
                    value="{{ form_data.english_name }}"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    性别
                </label>

                <select
                    name="gender"
                    class="form-input"
                    required
                >
                    <option value="">
                        请选择性别
                    </option>

                    <option
                        value="男"
                        {% if form_data.gender == "男" %}
                            selected
                        {% endif %}
                    >
                        男
                    </option>

                    <option
                        value="女"
                        {% if form_data.gender == "女" %}
                            selected
                        {% endif %}
                    >
                        女
                    </option>
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    出生年份
                </label>

                <input
                    type="number"
                    name="birth_year"
                    class="form-input"
                    value="{{ form_data.birth_year }}"
                    min="2000"
                    max="{{ current_year }}"
                    placeholder="例如：2016"
                    inputmode="numeric"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人姓名
                </label>

                <input
                    class="form-input"
                    name="parent_name"
                    value="{{ form_data.parent_name }}"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人电话
                </label>

                <input
                    class="form-input"
                    name="parent_phone"
                    value="{{ form_data.parent_phone }}"
                    inputmode="tel"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                    required
                >
                    <option value="">
                        请选择组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                form_data.group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    备注
                </label>

                <textarea
                    class="form-input"
                    name="remark"
                    rows="3"
                >{{ form_data.remark }}</textarea>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-success"
                    type="submit"
                >
                    ✅ 保存学生
                </button>
            </div>

        </form>

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_students'
                ) }}"
            >
                ⬅ 返回学生名单
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        groups=groups,
        form_data=form_data,
        current_year=current_year
    )

@dharma_class_bp.route(
    "/students/edit/<int:student_id>",
    methods=["GET", "POST"]
)
def class_students_edit(student_id):
    
    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取可用组别
            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            # 读取学生资料
            cur.execute("""
                select *
                from dharma_students
                where id = %s
            """, (student_id,))

            student = cur.fetchone()

            if not student:
                flash("找不到这个学生。", "bad")
                return redirect(
                    url_for(
                        "dharma_class.class_students"
                    )
                )

            if request.method == "POST":

                action = request.form.get(
                    "action",
                    "save"
                ).strip()

                # 暂停学生
                if action == "inactive":

                    cur.execute("""
                        update dharma_students
                        set status = 'inactive'
                        where id = %s
                    """, (student_id,))

                    conn.commit()

                    flash("学生已暂停。", "good")

                    return redirect(
                        url_for(
                            "dharma_class.class_students"
                        )
                    )

                # 读取表单资料
                student_no = request.form.get(
                    "student_no",
                    ""
                ).strip()

                name = request.form.get(
                    "name",
                    ""
                ).strip()

                english_name = request.form.get(
                    "english_name",
                    ""
                ).strip()

                gender = request.form.get(
                    "gender",
                    ""
                ).strip()

                birth_year_text = request.form.get(
                    "birth_year",
                    ""
                ).strip()

                parent_name = request.form.get(
                    "parent_name",
                    ""
                ).strip()

                parent_phone = request.form.get(
                    "parent_phone",
                    ""
                ).strip()

                group_id_text = request.form.get(
                    "group_id",
                    ""
                ).strip()

                remark = request.form.get(
                    "remark",
                    ""
                ).strip()

                birth_year = (
                    int(birth_year_text)
                    if birth_year_text.isdigit()
                    else None
                )

                group_id = (
                    int(group_id_text)
                    if group_id_text.isdigit()
                    else None
                )

                # 将刚提交的资料放回 student
                # 验证失败时，页面会保留用户刚输入的内容
                student["student_no"] = student_no
                student["name"] = name
                student["english_name"] = english_name
                student["gender"] = gender
                student["birth_year"] = birth_year_text
                student["parent_name"] = parent_name
                student["parent_phone"] = parent_phone
                student["group_id"] = group_id_text
                student["remark"] = remark

                # 验证资料
                if not name:
                    flash(
                        "学生姓名必须填写。",
                        "bad"
                    )

                elif gender not in ("男", "女"):
                    flash(
                        "请选择学生性别。",
                        "bad"
                    )

                elif birth_year is None:
                    flash(
                        "请输入正确的出生年份。",
                        "bad"
                    )

                elif (
                    birth_year < 2000
                    or birth_year > current_year
                ):
                    flash(
                        f"出生年份必须介于 "
                        f"2000 至 {current_year} 年。",
                        "bad"
                    )

                elif group_id is None:
                    flash(
                        "请选择学生组别。",
                        "bad"
                    )

                else:
                    cur.execute("""
                        update dharma_students
                        set
                            student_no = %s,
                            name = %s,
                            english_name = %s,
                            gender = %s,
                            birth_year = %s,
                            parent_name = %s,
                            parent_phone = %s,
                            group_id = %s,
                            remark = %s
                        where id = %s
                    """, (
                        student_no or None,
                        name,
                        english_name or None,
                        gender,
                        birth_year,
                        parent_name or None,
                        parent_phone or None,
                        group_id,
                        remark or None,
                        student_id
                    ))

                    conn.commit()

                    flash(
                        "学生资料已更新。",
                        "good"
                    )

                    return redirect(
                        url_for(
                            "dharma_class.class_students"
                        )
                    )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>编辑学生</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            ✏ 编辑学生
        </h1>

        <p class="page-subtitle">
            修改学生资料或暂停学生。
        </p>

        {% with messages =
            get_flashed_messages(
                with_categories=true
            )
        %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="post">

            <div class="form-group">
                <label class="form-label">
                    学生编号
                </label>

                <input
                    class="form-input"
                    name="student_no"
                    value="{{ student.student_no or '' }}"
                    placeholder="例如：S001"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    中文姓名
                </label>

                <input
                    class="form-input"
                    name="name"
                    value="{{ student.name or '' }}"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    英文名
                </label>

                <input
                    class="form-input"
                    name="english_name"
                    value="{{ student.english_name or '' }}"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    性别
                </label>

                <select
                    name="gender"
                    class="form-input"
                    required
                >
                    <option value="">
                        请选择性别
                    </option>

                    <option
                        value="男"
                        {% if student.gender == "男" %}
                            selected
                        {% endif %}
                    >
                        男
                    </option>

                    <option
                        value="女"
                        {% if student.gender == "女" %}
                            selected
                        {% endif %}
                    >
                        女
                    </option>
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    出生年份
                </label>

                <input
                    type="number"
                    name="birth_year"
                    class="form-input"
                    value="{{ student.birth_year or '' }}"
                    min="2000"
                    max="{{ current_year }}"
                    placeholder="例如：2016"
                    inputmode="numeric"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人姓名
                </label>

                <input
                    class="form-input"
                    name="parent_name"
                    value="{{ student.parent_name or '' }}"
                    placeholder="父亲、母亲、亲戚或监护人姓名"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人电话
                </label>

                <input
                    class="form-input"
                    name="parent_phone"
                    value="{{ student.parent_phone or '' }}"
                    placeholder="例如：0123456789"
                    inputmode="tel"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                    required
                >
                    <option value="">
                        请选择组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                student.group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    备注
                </label>

                <textarea
                    class="form-input"
                    name="remark"
                    rows="3"
                >{{ student.remark or '' }}</textarea>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-success"
                    type="submit"
                    name="action"
                    value="save"
                >
                    ✅ 保存修改
                </button>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-danger"
                    type="submit"
                    name="action"
                    value="inactive"
                    onclick="
                        return confirm(
                            '确定要暂停这个学生吗？暂停后点名不会再显示。'
                        );
                    "
                >
                    ⏸ 暂停学生
                </button>
            </div>

        </form>

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_students'
                ) }}"
            >
                ⬅ 返回学生名单
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        student=student,
        groups=groups,
        current_year=current_year
    )

@dharma_class_bp.route("/reports")
def class_reports():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    malaysia_today = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).date()

    # month = 月报，year = 年报
    mode = request.args.get(
        "mode",
        "month"
    ).strip().lower()

    if mode not in ("month", "year"):
        mode = "month"

    ym = request.args.get(
        "ym",
        malaysia_today.strftime("%Y-%m")
    ).strip()

    year_text = request.args.get(
        "year",
        ym[:4] if len(ym) >= 4 else str(malaysia_today.year)
    ).strip()

    group_id = request.args.get(
        "group_id",
        ""
    ).strip()

    try:
        selected_year = int(year_text)
    except (TypeError, ValueError):
        selected_year = malaysia_today.year

    selected_month = malaysia_today.month

    # 根据月报或年报建立日期范围
    if mode == "month":

        try:
            selected_month_date = datetime.strptime(
                ym,
                "%Y-%m"
            ).date()

            selected_year = selected_month_date.year
            selected_month = selected_month_date.month

        except ValueError:
            selected_year = malaysia_today.year
            selected_month = malaysia_today.month
            ym = malaysia_today.strftime("%Y-%m")

        if selected_month == 12:
            next_year = selected_year + 1
            next_month = 1
        else:
            next_year = selected_year
            next_month = selected_month + 1

        start_date = f"{selected_year:04d}-{selected_month:02d}-01"
        end_date = f"{next_year:04d}-{next_month:02d}-01"

        report_title = (
            f"{selected_year}年"
            f"{selected_month}月出席统计"
        )

    else:

        start_date = f"{selected_year:04d}-01-01"
        end_date = f"{selected_year + 1:04d}-01-01"

        report_title = f"{selected_year}年全年出席统计"

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取组别
            cur.execute("""
                select
                    id,
                    name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)

            groups = cur.fetchall()

            # 查询条件
            student_where = """
                where s.status = 'active'
                  and s.branch = 'CHE'
            """

            student_params = [
                start_date,
                end_date
            ]

            if group_id:
                student_where += """
                    and s.group_id = %s
                """
                student_params.append(group_id)

            # 每位学生出席统计
            cur.execute(f"""
                select
                    g.id as group_id,
                    g.name as group_name,
                    g.sort_order,

                    s.id as student_id,
                    s.student_no,
                    s.name,
                    s.english_name,

                    count(a.id) as total_records,

                    count(a.id)
                        filter (
                            where a.status = 'present'
                        ) as present_count,

                    count(a.id)
                        filter (
                            where a.status = 'late'
                        ) as late_count,

                    count(a.id)
                        filter (
                            where a.status = 'farm'
                        ) as farm_count,

                    count(a.id)
                        filter (
                            where a.status in ('absent', 'leave')
                        ) as absent_count

                from dharma_students s

                left join dharma_class_groups g
                    on g.id = s.group_id

                left join dharma_attendance a
                    on a.student_id = s.id
                   and a.branch = 'CHE'
                   and a.class_date >= %s
                   and a.class_date < %s

                {student_where}

                group by
                    g.id,
                    g.name,
                    g.sort_order,
                    s.id,
                    s.student_no,
                    s.name,
                    s.english_name

                order by
                    g.sort_order,
                    s.student_no nulls last,
                    s.name
            """, student_params)

            rows = cur.fetchall()

            # 课程堂数
            lesson_where = """
                where l.branch = 'CHE'
                  and l.lesson_date >= %s
                  and l.lesson_date < %s
            """

            lesson_params = [
                start_date,
                end_date
            ]

            if group_id:
                lesson_where += """
                    and l.group_id = %s
                """
                lesson_params.append(group_id)

            cur.execute(f"""
                select
                    count(*) as lesson_count,
                    count(distinct l.group_id) as group_count,
                    count(distinct l.teacher_name)
                        filter (
                            where coalesce(
                                l.teacher_name,
                                ''
                            ) <> ''
                        ) as teacher_count
                from dharma_class_lessons l
                {lesson_where}
            """, lesson_params)

            lesson_summary = cur.fetchone() or {}

            # 各组课程堂数
            cur.execute(f"""
                select
                    g.id as group_id,
                    g.name as group_name,
                    g.sort_order,
                    count(l.id) as lesson_count
                from dharma_class_groups g

                left join dharma_class_lessons l
                    on l.group_id = g.id
                   and l.branch = 'CHE'
                   and l.lesson_date >= %s
                   and l.lesson_date < %s

                where g.is_active = true
                {"and g.id = %s" if group_id else ""}

                group by
                    g.id,
                    g.name,
                    g.sort_order

                order by
                    g.sort_order,
                    g.id
            """, lesson_params)

            group_lessons = cur.fetchall()

    # 计算每位学生出席率
    total_present = 0
    total_late = 0
    total_farm = 0
    total_absent = 0
    total_marked_records = 0

    for row in rows:

        total_records = int(
            row["total_records"] or 0
        )

        present_count = int(
            row["present_count"] or 0
        )

        late_count = int(
            row["late_count"] or 0
        )

        farm_count = int(
            row["farm_count"] or 0
        )

        absent_count = int(
            row["absent_count"] or 0
        )

        attended_count = (
            present_count
            + late_count
            + farm_count
        )

        effective_total = attended_count + absent_count

        if effective_total > 0:
            row["rate"] = round(
                attended_count * 100 / effective_total,
                1
            )
        else:
            row["rate"] = None

        total_present += present_count
        total_late += late_count
        total_farm += farm_count
        total_absent += absent_count
        total_marked_records += effective_total

    student_count = len(rows)

    total_attended = (
        total_present
        + total_late
        + total_farm
    )

    if total_marked_records > 0:
        overall_rate = round(
            total_attended
            * 100
            / total_marked_records,
            1
        )
    else:
        overall_rate = None

    lesson_count = int(
        lesson_summary.get("lesson_count") or 0
    )


    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>佛学班报表中心</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>
.report-mode-row {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 12px;
    margin-bottom: 20px;
}

.report-mode-btn {
    display: flex;
    align-items: center;
    justify-content: center;
    min-height: 58px;
    border-radius: 15px;
    text-decoration: none;
    font-size: 20px;
    font-weight: 800;
    background: #f1f3f5;
    color: #444;
}

.report-mode-btn.active {
    background: #2563eb;
    color: white;
}

.report-summary {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin-top: 20px;
}

.report-summary-box {
    background: #f6f7f9;
    border-radius: 16px;
    padding: 16px 10px;
    text-align: center;
}

.report-summary-label {
    color: #666;
    font-size: 16px;
}

.report-summary-number {
    font-size: 29px;
    font-weight: 800;
    margin-top: 6px;
}

.group-lesson-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 12px;
    margin: 16px 0 20px;
}

.group-lesson-box {
    border: 1px solid #e5e7eb;
    border-radius: 15px;
    padding: 15px;
    text-align: center;
    background: white;
}

.group-lesson-name {
    font-size: 18px;
    font-weight: 800;
}

.group-lesson-count {
    font-size: 27px;
    font-weight: 800;
    margin-top: 6px;
}

.rate-good {
    color: #15803d;
    font-weight: 800;
}

.rate-warning {
    color: #b45309;
    font-weight: 800;
}

.rate-low {
    color: #b91c1c;
    font-weight: 800;
}

@media (max-width: 750px) {
    .report-summary {
        grid-template-columns: repeat(2, 1fr);
    }

    .group-lesson-grid {
        grid-template-columns: 1fr;
    }
}
</style>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            📊 佛学班报表中心
        </h1>

        <p class="page-subtitle">
            查看学生月度及年度出席统计。
        </p>

        <div class="report-mode-row">

            <a
                class="report-mode-btn
                    {% if mode == 'month' %}active{% endif %}"
                href="{{ url_for(
                    'dharma_class.class_reports',
                    mode='month',
                    ym=ym,
                    group_id=group_id
                ) }}"
            >
                📅 月报
            </a>

            <a
                class="report-mode-btn
                    {% if mode == 'year' %}active{% endif %}"
                href="{{ url_for(
                    'dharma_class.class_reports',
                    mode='year',
                    year=selected_year,
                    group_id=group_id
                ) }}"
            >
                📆 年报
            </a>

        </div>

        <form method="get">

            <input
                type="hidden"
                name="mode"
                value="{{ mode }}"
            >

            {% if mode == "month" %}

                <div class="form-group">
                    <label class="form-label">
                        月份
                    </label>

                    <input
                        class="form-input"
                        type="month"
                        name="ym"
                        value="{{ ym }}"
                        required
                    >
                </div>

            {% else %}

                <div class="form-group">
                    <label class="form-label">
                        年份
                    </label>

                    <input
                        class="form-input"
                        type="number"
                        name="year"
                        value="{{ selected_year }}"
                        min="2020"
                        max="2100"
                        required
                    >
                </div>

            {% endif %}

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                >
                    <option value="">
                        全部组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    🔍 查看报表
                </button>
            </div>

        </form>

        <div class="report-summary">

            <div class="report-summary-box">
                <div class="report-summary-label">
                    学生人数
                </div>

                <div class="report-summary-number">
                    {{ student_count }}
                </div>
            </div>

            <div class="report-summary-box">
                <div class="report-summary-label">
                    课程堂数
                </div>

                <div class="report-summary-number">
                    {{ lesson_count }}
                </div>
            </div>

            <div class="report-summary-box">
                <div class="report-summary-label">
                    出席次数
                </div>

                <div class="report-summary-number">
                    {{ total_attended }}
                </div>
            </div>

            <div class="report-summary-box">
                <div class="report-summary-label">
                    平均出席率
                </div>

                <div class="report-summary-number">
                    {% if overall_rate is not none %}
                        {{ overall_rate }}%
                    {% else %}
                        —
                    {% endif %}
                </div>
            </div>

        </div>

    </div>

    <div class="card">

        <h2 class="section-title">
            📚 各组课程堂数
        </h2>

        <div class="group-lesson-grid">

            {% for g in group_lessons %}
                <div class="group-lesson-box">

                    <div class="group-lesson-name">
                        {{ g.group_name }}
                    </div>

                    <div class="group-lesson-count">
                        {{ g.lesson_count or 0 }} 堂
                    </div>

                </div>
            {% endfor %}

        </div>

        <h2 class="section-title">
            {{ report_title }}
        </h2>

        {% if rows %}

            <div class="table-responsive">

                <table class="record-table">

                    <thead>
                        <tr>
                            <th>组别</th>
                            <th>编号</th>
                            <th>姓名</th>
                            <th>已记录</th>
                            <th>出席</th>
                            <th>迟到</th>
                            <th>农舍</th>
                            <th>缺席</th>
                            <th>出席率</th>
                        </tr>
                    </thead>

                    <tbody>

                        {% for r in rows %}

                        <tr>
                            <td>
                                {{ r.group_name or "—" }}
                            </td>

                            <td>
                                {{ r.student_no or "—" }}
                            </td>

                            <td>
                                <strong>{{ r.name }}</strong>

                                {% if r.english_name %}
                                    <br>
                                    <span style="color:#777;">
                                        {{ r.english_name }}
                                    </span>
                                {% endif %}
                            </td>

                            <td>
                                {{ r.total_records or 0 }}
                            </td>

                            <td>
                                {{ r.present_count or 0 }}
                            </td>

                            <td>
                                {{ r.late_count or 0 }}
                            </td>

                            <td>
                                {{ r.farm_count or 0 }}
                            </td>

                            <td>
                                {{ r.absent_count or 0 }}
                            </td>

                            <td>
                                {% if r.rate is not none %}

                                    <span class="
                                        {% if r.rate >= 80 %}
                                            rate-good
                                        {% elif r.rate >= 60 %}
                                            rate-warning
                                        {% else %}
                                            rate-low
                                        {% endif %}
                                    ">
                                        {{ r.rate }}%
                                    </span>

                                {% else %}
                                    —
                                {% endif %}
                            </td>
                        </tr>

                        {% endfor %}

                    </tbody>

                </table>

            </div>

        {% else %}

            <div class="empty-state">
                暂时没有符合条件的学生资料。
            </div>

        {% endif %}

        {% if mode == "month" %}

            <div class="btn-row">

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_monthly_report',
                        ym=ym,
                        group_id=group_id
                    ) }}"
                >
                    📥 下载月报 Excel
                </a>

            </div>

            <h2 class="section-title">
                📚 功课追踪报表
            </h2>

            <div class="btn-row">

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_baihua_homework',
                        year=selected_year,
                        month=selected_month
                    ) }}"
                >
                    📚 本月白话佛法功课
                </a>

                <a
                    class="btn-tool btn-warning"
                    href="{{ url_for(
                        'dharma_class.class_export_scripture_homework',
                        year=selected_year,
                        month=selected_month
                    ) }}"
                >
                    📿 本月经文功课
                </a>

            </div>

        {% else %}

            <div class="btn-row">

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_yearly_report',
                        year=selected_year,
                        group_id=group_id
                    ) }}"
                >
                    📥 下载年报 Excel
                </a>

            </div>

            <h2 class="section-title">
                📚 全年功课追踪报表
            </h2>

            <div class="btn-row">
                               
                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_baihua_homework',
                        year=selected_year
                    ) }}"
                >
                    📚 全年白话佛法功课
                </a>

                <a
                    class="btn-tool btn-warning"
                    href="{{ url_for(
                        'dharma_class.class_export_scripture_homework',
                        year=selected_year
                    ) }}"
                >
                    📿 全年经文功课
                </a>

            </div>

        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_home'
                ) }}"
            >
                ⬅ 返回佛学班首页
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        mode=mode,
        ym=ym,
        selected_year=selected_year,
        selected_month=selected_month,
        group_id=group_id,
        groups=groups,
        rows=rows,
        report_title=report_title,
        student_count=student_count,
        lesson_count=lesson_count,
        total_attended=total_attended,
        overall_rate=overall_rate,
        group_lessons=group_lessons
    )

@dharma_class_bp.route("/records")
def class_records():

    selected_date = request.args.get("class_date") or date.today().isoformat()
    group_id = request.args.get("group_id")

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            params = [selected_date]
            where_group = ""

            if group_id:
                where_group = "and s.group_id = %s"
                params.append(group_id)

            cur.execute(f"""
                select
                    s.student_no,
                    s.name,
                    s.english_name,
                    g.name as group_name,
                    a.status,
                    a.remark,
                    a.marked_by,
                    a.marked_at
                from dharma_attendance a
                join dharma_students s on s.id = a.student_id
                left join dharma_class_groups g on g.id = s.group_id
                where a.class_date = %s
                {where_group}
                order by g.sort_order, s.student_no nulls last, s.name
            """, params)

            records = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>点名记录</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">📅 点名记录</h1>
        <p class="page-subtitle">查看某一天佛学班点名情况。</p>

        <form method="get">
            <div class="form-group">
                <label class="form-label">日期</label>
                <input class="form-input" type="date" name="class_date" value="{{ selected_date }}">
            </div>

            <div class="form-group">
                <label class="form-label">组别</label>
                <select class="form-input" name="group_id">
                    <option value="">全部组别</option>
                    {% for g in groups %}
                        <option value="{{ g.id }}"
                            {% if group_id|string == g.id|string %}selected{% endif %}>
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-primary" type="submit">
                    🔍 查看记录
                </button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2 class="section-title">{{ selected_date }} 点名记录</h2>

        {% if records %}
            <div class="table-responsive">
                <table class="record-table">
                    <thead>
                        <tr>
                            <th>组别</th>
                            <th>编号</th>
                            <th>姓名</th>
                            <th>状态</th>
                            <th>备注</th>
                            <th>老师</th>
                            <th>保存时间</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for r in records %}
                        <tr>
                            <td>{{ r.group_name or "-" }}</td>
                            <td>{{ r.student_no or "-" }}</td>
                            <td>
                                {{ r.name }}
                                {% if r.english_name %}
                                    <span style="color:#777;">{{ r.english_name }}</span>
                                {% endif %}
                            </td>
                            <td>
                                {% if r.status == "present" %}
                                    ✅ 出席
                                {% elif r.status == "late" %}
                                    ⏰ 迟到
                                {% elif r.status == "farm" %}
                                    🌱 农舍
                                {% elif r.status in ("absent", "leave") %}
                                    ❌ 缺席
                                {% else %}
                                    {{ r.status }}
                                {% endif %}
                            </td>
                            <td>{{ r.remark or "-" }}</td>
                            <td>{{ r.marked_by or "-" }}</td>
                            <td>{{ r.marked_at or "-" }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        {% else %}
            <div class="empty-state">
                这一天还没有点名记录。
            </div>
        {% endif %}

        <div class="btn-row">
            <a class="btn-tool btn-secondary" href="/class">
                ⬅ 返回佛学班首页
            </a>
        </div>
    </div>

</div>
</body>
</html>
""",
        selected_date=selected_date,
        group_id=group_id,
        groups=groups,
        records=records
    )

@dharma_class_bp.route("/reports/monthly/export")
def class_export_monthly_report():

    ym = request.args.get("ym") or date.today().strftime("%Y-%m")

    # =========================================================
    # 出席状态显示
    # =========================================================
    status_symbol = {
        "present": "✓",
        "late": "迟",
        "farm": "农",
        "leave": "✗",
        "absent": "✗",
    }

    status_fill = {
        "✓": PatternFill("solid", fgColor="C6EFCE"),
        "迟": PatternFill("solid", fgColor="FCE4D6"),
        "农": PatternFill("solid", fgColor="D9EAD3"),
        "✗": PatternFill("solid", fgColor="F4CCCC"),
    }

    # =========================================================
    # 读取资料
    # =========================================================
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            # 组别
            cur.execute("""
                select
                    id,
                    name,
                    sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            # 在籍学生
            cur.execute("""
                select
                    s.id,
                    s.student_no,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.name as group_name,
                    g.sort_order
                from dharma_students s
                join dharma_class_groups g
                  on g.id = s.group_id
                where s.status = 'active'
                  and g.is_active = true
                order by
                    g.sort_order,
                    s.student_no nulls last,
                    s.name
            """)
            students = cur.fetchall()

            # 本月出席记录
            cur.execute("""
                select
                    student_id,
                    group_id,
                    class_date,
                    status
                from dharma_attendance
                where to_char(class_date, 'YYYY-MM') = %s
                order by class_date
            """, (ym,))
            att_rows = cur.fetchall()

    # =========================================================
    # 整理出席资料
    # =========================================================
    attendance_map = {}

    for r in att_rows:
        attendance_map[
            (
                r["student_id"],
                r["class_date"]
            )
        ] = r["status"]

    # 每组实际有点名记录的日期
    group_dates_map = {}

    for g in groups:
        group_dates_map[g["id"]] = []

    for r in att_rows:
        group_id = r["group_id"]
        class_date = r["class_date"]

        if group_id not in group_dates_map:
            group_dates_map[group_id] = []

        if class_date not in group_dates_map[group_id]:
            group_dates_map[group_id].append(class_date)

    for group_id in group_dates_map:
        group_dates_map[group_id].sort()

    # 全部课程日期，用于总览显示
    all_class_dates = sorted({
        r["class_date"]
        for r in att_rows
    })

    # =========================================================
    # 建立 Excel
    # =========================================================
    wb = Workbook()

    # =========================================================
    # 样式
    # =========================================================
    title_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="5B9BD5"
    )

    summary_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_font = Font(
        size=18,
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    thin = Side(
        style="thin",
        color="CCCCCC"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    left = Alignment(
        horizontal="left",
        vertical="center"
    )

    # =========================================================
    # 总览 Sheet
    # =========================================================
    ws = wb.active
    ws.title = "总览"

    ws.merge_cells("A1:H1")

    ws["A1"] = f"蕉赖佛学班出席月报  {ym}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center

    ws.row_dimensions[1].height = 34

    ws["A3"] = "本月课程日期"
    ws["A3"].font = bold_font
    ws["A3"].alignment = left

    if all_class_dates:
        date_text = "、".join(
            f"{d.month}/{d.day}"
            for d in all_class_dates
        )
    else:
        date_text = "本月还没有点名记录"

    ws["B3"] = date_text
    ws["B3"].alignment = left

    ws.append([])

    ws.append([
        "组别",
        "总学生人数",
        "课程次数",
        "出席人数",
        "迟到人数",
        "农舍人数",
        "缺席人数",
        "平均出席率",
    ])

    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    for g in groups:

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        group_dates = group_dates_map.get(
            g["id"],
            []
        )

        total_students = len(group_students)

        group_present = 0
        group_late = 0
        group_farm = 0
        group_absent = 0

        student_rates = []

        for s in group_students:

            present_count = 0
            late_count = 0
            farm_count = 0
            absent_count = 0

            for class_date in group_dates:

                status = attendance_map.get(
                    (
                        s["id"],
                        class_date
                    )
                )

                if status == "present":
                    present_count += 1
                    group_present += 1

                elif status == "late":
                    late_count += 1
                    group_late += 1

                elif status == "farm":
                    farm_count += 1
                    group_farm += 1

                elif status in ("absent", "leave"):
                    absent_count += 1
                    group_absent += 1

            attended_count = (
                present_count +
                late_count +
                farm_count
            )

            effective_total = attended_count + absent_count

            if effective_total > 0:
                student_rates.append(
                    attended_count / effective_total
                )

        avg_rate = (
            sum(student_rates) / len(student_rates)
            if student_rates
            else None
        )

        ws.append([
            g["name"],
            total_students,
            len(group_dates),
            group_present,
            group_late,
            group_farm,
            group_absent,
            avg_rate,
        ])

    for row in ws.iter_rows(
        min_row=6,
        max_row=ws.max_row
    ):
        for cell in row:
            cell.border = border
            cell.alignment = center

        rate_cell = row[7]

        if (
            rate_cell.value is not None
            and isinstance(
                rate_cell.value,
                (int, float)
            )
        ):
            rate_cell.number_format = "0%"

            if rate_cell.value >= 0.9:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="C6EFCE"
                )

            elif rate_cell.value >= 0.75:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="FFEB9C"
                )

            else:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="F4CCCC"
                )

    ws.freeze_panes = "A6"

    if ws.max_row >= 6:
        ws.auto_filter.ref = (
            f"A5:H{ws.max_row}"
        )

    # =========================================================
    # 各组 Sheet
    # =========================================================
    for g in groups:

        sheet_name = g["name"][:31]
        ws_g = wb.create_sheet(sheet_name)

        group_dates = group_dates_map.get(
            g["id"],
            []
        )

        headers = [
            "学生编号",
            "学生姓名",
        ]

        for class_date in group_dates:
            headers.append(
                f"{class_date.month}/{class_date.day}"
            )

        headers += [
            "出席",
            "迟到",
            "农舍",
            "缺席",
            "出席率",
        ]

        end_col = max(
            7,
            len(headers)
        )

        ws_g.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=end_col
        )

        ws_g["A1"] = (
            f"{g['name']} 出席月报  {ym}"
        )

        ws_g["A1"].font = title_font
        ws_g["A1"].fill = title_fill
        ws_g["A1"].alignment = center

        ws_g.row_dimensions[1].height = 34

        ws_g.append([])
        ws_g.append(headers)

        for cell in ws_g[3]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        if not group_students:
            ws_g.append([
                "",
                "目前没有学生资料"
            ])

        for s in group_students:

            present_count = 0
            late_count = 0
            farm_count = 0
            absent_count = 0

            row_data = [
                s["student_no"] or "",
                s["name"] or "",
            ]

            for class_date in group_dates:

                status = attendance_map.get(
                    (
                        s["id"],
                        class_date
                    )
                )

                if status == "present":
                    present_count += 1

                elif status == "late":
                    late_count += 1

                elif status == "farm":
                    farm_count += 1

                elif status in ("absent", "leave"):
                    absent_count += 1

                row_data.append(
                    status_symbol.get(
                        status,
                        ""
                    )
                )

            attended_count = (
                present_count +
                late_count +
                farm_count
            )

            effective_total = attended_count + absent_count

            if effective_total > 0:
                attendance_rate = (
                    attended_count /
                    effective_total
                )
            else:
                attendance_rate = None

            row_data += [
                present_count,
                late_count,
                farm_count,
                absent_count,
                attendance_rate,
            ]

            ws_g.append(row_data)

        # =====================================================
        # 学生资料样式
        # =====================================================
        for row in ws_g.iter_rows(
            min_row=4,
            max_row=ws_g.max_row
        ):
            for cell in row:
                cell.border = border
                cell.alignment = center

                if cell.value in status_fill:
                    cell.fill = status_fill[cell.value]
                    cell.font = Font(
                        bold=True
                    )

            rate_cell = row[-1]

            if (
                rate_cell.value is not None
                and isinstance(
                    rate_cell.value,
                    (int, float)
                )
            ):
                rate_cell.number_format = "0%"

                if rate_cell.value >= 0.9:
                    rate_cell.fill = PatternFill(
                        "solid",
                        fgColor="C6EFCE"
                    )

                elif rate_cell.value >= 0.75:
                    rate_cell.fill = PatternFill(
                        "solid",
                        fgColor="FFEB9C"
                    )

                else:
                    rate_cell.fill = PatternFill(
                        "solid",
                        fgColor="F4CCCC"
                    )

        ws_g.freeze_panes = "C4"

        if ws_g.max_row >= 4:
            ws_g.auto_filter.ref = (
                f"A3:"
                f"{get_column_letter(ws_g.max_column)}"
                f"{ws_g.max_row}"
            )

        ws_g.column_dimensions["A"].width = 14
        ws_g.column_dimensions["B"].width = 22

    # =========================================================
    # 全部 Sheet 自动美化
    # =========================================================
    for sheet in wb.worksheets:

        for col_idx in range(
            1,
            sheet.max_column + 1
        ):
            col_letter = get_column_letter(
                col_idx
            )

            max_len = 0

            for row_idx in range(
                1,
                sheet.max_row + 1
            ):
                cell = sheet.cell(
                    row=row_idx,
                    column=col_idx
                )

                value = (
                    str(cell.value)
                    if cell.value is not None
                    else ""
                )

                max_len = max(
                    max_len,
                    len(value)
                )

            width = max_len + 4

            if col_idx == 1:
                width = max(
                    width,
                    14
                )

            elif col_idx == 2:
                width = max(
                    width,
                    22
                )

            else:
                width = max(
                    width,
                    10
                )

            # 防止课程日期太多时列宽过大
            width = min(
                width,
                35
            )

            sheet.column_dimensions[
                col_letter
            ].width = width

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=(
                        cell.alignment.horizontal
                        or "center"
                    ),
                    vertical="center",
                    wrap_text=True
                )

        for row_idx in range(
            1,
            sheet.max_row + 1
        ):
            sheet.row_dimensions[
                row_idx
            ].height = 24

        sheet.sheet_view.showGridLines = False

    # =========================================================
    # 输出 Excel
    # =========================================================
    output = BytesIO()

    wb.save(output)
    output.seek(0)

    filename = (
        f"蕉赖佛学班出席月报_{ym}.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@dharma_class_bp.route("/students/import", methods=["GET", "POST"])
def class_students_import():

    result = None

    if request.method == "POST":
        file = request.files.get("excel_file")

        if not file or file.filename == "":
            flash("请选择 Excel 文件。", "bad")
            return redirect(url_for("dharma_class.class_students_import"))

        wb = load_workbook(file, data_only=True)
        ws = wb.active

        inserted = 0
        duplicate = 0
        blank = 0
        invalid = 0

        inserted_names = []
        duplicate_names = []
        error_messages = []

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                cur.execute("""
                    select id, name
                    from dharma_class_groups
                    where is_active = true
                """)
                groups = cur.fetchall()

                group_map = {
                    g["name"].strip(): g["id"]
                    for g in groups
                }

                for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

                    student_no = str(row[0]).strip() if len(row) > 0 and row[0] else ""
                    name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    english_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                    gender = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                    birth_year = None
                    if len(row) > 4 and row[4]:
                        try:
                            birth_year = int(row[4])
                        except:
                            birth_year = None

                    parent_name = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                    parent_phone = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                    group_name = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                    remark = str(row[8]).strip() if len(row) > 8 and row[8] else ""

                    if not name:
                        blank += 1
                        continue

                    group_id = group_map.get(group_name)

                    if not group_id:
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：组别「{group_name}」不存在"
                        )
                        continue

                    # 性别检查
                    if gender not in ("男", "女"):
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：性别必须填写「男」或「女」"
                        )
                        continue

                    # 出生年份检查
                    if birth_year is None:
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：请填写出生年份"
                        )
                        continue

                    if birth_year < 2000 or birth_year > 2100:
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：出生年份错误"
                        )
                        continue

                    # 重复检查：优先学生编号，其次 中文姓名 + 家长电话
                    if student_no:
                        cur.execute("""
                            select id
                            from dharma_students
                            where student_no = %s
                            limit 1
                        """, (student_no,))
                    else:
                        cur.execute("""
                            select id
                            from dharma_students
                            where name = %s
                              and coalesce(parent_phone, '') = %s
                            limit 1
                        """, (name, parent_phone or ""))

                    exists = cur.fetchone()

                    if exists:
                        duplicate += 1
                        duplicate_names.append(name)
                        continue

                    cur.execute("""
                        insert into dharma_students
                        (
                            branch,
                            student_no,
                            name,
                            english_name,
                            gender,
                            birth_year,
                            parent_name,
                            parent_phone,
                            group_id,
                            status,
                            remark
                        )
                        values
                            ('CHE', %s, %s, %s, %s, %s, %s, %s, %s, 'active', %s)
                    """, (
                        student_no or None,
                        name,
                        english_name or None,
                        gender or None,
                        birth_year,
                        parent_name or None,
                        parent_phone or None,
                        group_id,
                        remark or None
                    ))

                    inserted += 1
                    inserted_names.append(name)

                conn.commit()

        result = {
            "inserted": inserted,
            "duplicate": duplicate,
            "blank": blank,
            "invalid": invalid,
            "inserted_names": inserted_names,
            "duplicate_names": duplicate_names,
            "error_messages": error_messages,
        }

        flash("导入完成。", "good")

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>导入学生 Excel</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">📥 导入学生 Excel</h1>
        <p class="page-subtitle">上传学生名单，系统会自动新增并检查重复。</p>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert">{{ msg }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <div class="empty-state" style="text-align:left;">
            Excel 第一行标题请使用：<br><br>
            学生编号｜中文姓名｜英文名｜家长姓名｜家长电话｜组别｜备注<br><br>
            组别目前只接受：低年组、少年组、高年组
        </div>

        <div class="btn-row">
            <a class="btn-tool btn-secondary" href="/class/students/template">
                📄 下载 Excel 模板
            </a>
        </div>

        <form method="post" enctype="multipart/form-data">

            <div class="form-group">
                <label class="form-label">选择 Excel 文件</label>
                <input class="form-input" type="file" name="excel_file" accept=".xlsx" required>
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-success" type="submit">
                    ✅ 上传并导入
                </button>
            </div>

        </form>
    </div>

    {% if result %}
    <div class="card">
        <h2 class="section-title">📋 导入结果</h2>

        <div class="summary-grid">
            <div class="summary-box">
                <div>新增</div>
                <div style="font-size:30px;font-weight:800;">{{ result.inserted }}</div>
            </div>

            <div class="summary-box">
                <div>重复</div>
                <div style="font-size:30px;font-weight:800;">{{ result.duplicate }}</div>
            </div>

            <div class="summary-box">
                <div>空白</div>
                <div style="font-size:30px;font-weight:800;">{{ result.blank }}</div>
            </div>

            <div class="summary-box">
                <div>错误</div>
                <div style="font-size:30px;font-weight:800;">{{ result.invalid }}</div>
            </div>
        </div>

        {% if result.inserted_names %}
            <h3 class="section-title">✅ 新增学生</h3>
            <div class="empty-state" style="text-align:left;">
                {% for n in result.inserted_names[:30] %}
                    ✓ {{ n }}<br>
                {% endfor %}
            </div>
        {% endif %}

        {% if result.duplicate_names %}
            <h3 class="section-title">⚠️ 重复已跳过</h3>
            <div class="empty-state" style="text-align:left;">
                {% for n in result.duplicate_names[:30] %}
                    • {{ n }}<br>
                {% endfor %}
            </div>
        {% endif %}

        {% if result.error_messages %}
            <h3 class="section-title">❌ 错误资料</h3>
            <div class="empty-state" style="text-align:left;color:#b91c1c;">
                {% for e in result.error_messages[:30] %}
                    {{ e }}<br>
                {% endfor %}
            </div>
        {% endif %}
    </div>
    {% endif %}

    <div class="card">
        <div class="btn-row">
            <a class="btn-tool btn-secondary" href="/class/students">
                ⬅ 返回学生名单
            </a>
        </div>
    </div>

</div>
</body>
</html>
""", result=result)

@dharma_class_bp.route("/students/template")
def class_students_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"

    headers = [
        "学生编号",
        "中文姓名",
        "英文名",
        "性别",
        "出生年份",
        "父／母／监护人姓名",
        "父／母／监护人电话",
        "组别",
        "备注"
    ]

    ws.append(headers)

    ws.append([
        "001",
        "张小明",
        "Zhang Xiao Ming",
        "男",
        2016,
        "张爸爸",
        "0123456789",
        "低年组",
        ""
    ])

    ws.append([
        "002",
        "李小美",
        "Lee Mei Mei",
        "女",
        2014,
        "李妈妈",
        "01122334455",
        "少年组",
        ""
    ])

    beautify_simple_excel(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="佛学班学生导入模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@dharma_class_bp.route("/students/export")
def class_students_export():

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select
                    s.student_no,
                    s.name,
                    s.english_name,
                    s.parent_name,
                    s.parent_phone,
                    g.name as group_name,
                    s.status,
                    s.remark
                from dharma_students s
                left join dharma_class_groups g on g.id = s.group_id
                order by g.sort_order, s.student_no nulls last, s.name
            """)
            rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "学生名单"

    ws.append([
        "学生编号",
        "中文姓名",
        "英文名",
        "家长姓名",
        "家长电话",
        "组别",
        "状态",
        "备注"
    ])

    for r in rows:
        ws.append([
            r["student_no"] or "",
            r["name"] or "",
            r["english_name"] or "",
            r["parent_name"] or "",
            r["parent_phone"] or "",
            r["group_name"] or "",
            r["status"] or "",
            r["remark"] or ""
        ])

    beautify_simple_excel(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="佛学班学生名单.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@dharma_class_bp.route("/reports/yearly/export")
def class_export_yearly_report():

    year = request.args.get("year") or date.today().strftime("%Y")

    # =========================================================
    # 读取资料
    # =========================================================
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    id,
                    name,
                    sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            cur.execute("""
                select
                    s.id,
                    s.student_no,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.name as group_name,
                    g.sort_order
                from dharma_students s
                join dharma_class_groups g
                  on g.id = s.group_id
                where s.status = 'active'
                  and g.is_active = true
                order by
                    g.sort_order,
                    s.student_no nulls last,
                    s.name
            """)
            students = cur.fetchall()

            cur.execute("""
                select
                    student_id,
                    to_char(class_date, 'MM') as month_no,
                    status
                from dharma_attendance
                where to_char(class_date, 'YYYY') = %s
                order by class_date
            """, (year,))
            att_rows = cur.fetchall()

    # =========================================================
    # 整理统计资料
    #
    # student_id
    #     -> month
    #         -> present / late / farm / absent
    # =========================================================
    stat = {}

    for r in att_rows:

        student_id = r["student_id"]
        month_no = int(r["month_no"])
        status = r["status"]

        stat.setdefault(student_id, {})

        stat[student_id].setdefault(month_no, {
            "records": 0,
            "present": 0,
            "late": 0,
            "farm": 0,
            "absent": 0,
            "attended": 0,
            "effective_total": 0,
        })

        month_data = stat[student_id][month_no]

        # 所有点名记录，包括请假
        month_data["records"] += 1

        if status == "present":
            month_data["present"] += 1
            month_data["attended"] += 1
            month_data["effective_total"] += 1

        elif status == "late":
            month_data["late"] += 1
            month_data["attended"] += 1
            month_data["effective_total"] += 1

        elif status == "farm":
            month_data["farm"] += 1
            month_data["attended"] += 1
            month_data["effective_total"] += 1

        elif status in ("absent", "leave"):
            month_data["absent"] += 1
            month_data["effective_total"] += 1

    # =========================================================
    # 建立 Excel
    # =========================================================
    wb = Workbook()

    # =========================================================
    # 样式
    # =========================================================
    title_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="5B9BD5"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_font = Font(
        size=18,
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    thin = Side(
        style="thin",
        color="CCCCCC"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    # =========================================================
    # 总览 Sheet
    # =========================================================
    ws = wb.active
    ws.title = "总览"

    ws.merge_cells("A1:H1")

    ws["A1"] = f"蕉赖佛学班出席年报  {year}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center

    ws.row_dimensions[1].height = 34

    ws.append([])

    ws.append([
        "组别",
        "学生人数",
        "全年记录",
        "出席",
        "迟到",
        "农舍",
        "缺席",
        "平均出席率",
    ])

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    for g in groups:

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        student_rates = []

        total_records = 0
        total_present = 0
        total_late = 0
        total_farm = 0
        total_absent = 0

        for s in group_students:

            student_attended = 0
            student_effective_total = 0

            for month_no in range(1, 13):

                data = stat.get(
                    s["id"],
                    {}
                ).get(month_no)

                if not data:
                    continue

                total_records += data["records"]
                total_present += data["present"]
                total_late += data["late"]
                total_farm += data["farm"]
                total_absent += data["absent"]

                student_attended += data["attended"]
                student_effective_total += data["effective_total"]

            if student_effective_total > 0:
                student_rates.append(
                    student_attended /
                    student_effective_total
                )

        avg_rate = (
            sum(student_rates) / len(student_rates)
            if student_rates
            else None
        )

        ws.append([
            g["name"],
            len(group_students),
            total_records,
            total_present,
            total_late,
            total_farm,
            total_absent,
            avg_rate,
        ])

    for row in ws.iter_rows(
        min_row=4,
        max_row=ws.max_row
    ):
        for cell in row:
            cell.border = border
            cell.alignment = center

        rate_cell = row[7]

        if (
            rate_cell.value is not None
            and isinstance(
                rate_cell.value,
                (int, float)
            )
        ):
            rate_cell.number_format = "0%"

            if rate_cell.value >= 0.9:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="C6EFCE"
                )

            elif rate_cell.value >= 0.75:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="FFEB9C"
                )

            else:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="F4CCCC"
                )

    ws.freeze_panes = "A4"

    if ws.max_row >= 4:
        ws.auto_filter.ref = (
            f"A3:H{ws.max_row}"
        )

    # =========================================================
    # 各组 Sheet
    # =========================================================
    for g in groups:

        ws_g = wb.create_sheet(
            g["name"][:31]
        )

        headers = [
            "学生编号",
            "学生姓名",
        ]

        headers += [
            f"{month_no}月"
            for month_no in range(1, 13)
        ]

        headers += [
            "全年出席率",
            "出席",
            "迟到",
            "农舍",
            "缺席",
        ]

        end_col = len(headers)

        ws_g.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=end_col
        )

        ws_g["A1"] = (
            f"{g['name']} 出席年报  {year}"
        )

        ws_g["A1"].font = title_font
        ws_g["A1"].fill = title_fill
        ws_g["A1"].alignment = center

        ws_g.row_dimensions[1].height = 34

        ws_g.append([])
        ws_g.append(headers)

        for cell in ws_g[3]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        if not group_students:
            ws_g.append([
                "",
                "目前没有学生资料"
            ])

        for s in group_students:

            row_data = [
                s["student_no"] or "",
                s["name"] or "",
            ]

            yearly_attended = 0
            yearly_effective_total = 0

            yearly_present = 0
            yearly_late = 0
            yearly_farm = 0
            yearly_absent = 0

            for month_no in range(1, 13):

                data = stat.get(
                    s["id"],
                    {}
                ).get(month_no)

                if data:

                    if data["effective_total"] > 0:
                        month_rate = (
                            data["attended"] /
                            data["effective_total"]
                        )
                    else:
                        # 当月只有请假，没有可计算记录
                        month_rate = None

                    row_data.append(month_rate)

                    yearly_attended += data["attended"]
                    yearly_effective_total += data["effective_total"]

                    yearly_present += data["present"]
                    yearly_late += data["late"]
                    yearly_farm += data["farm"]
                    yearly_absent += data["absent"]

                else:
                    row_data.append(None)

            if yearly_effective_total > 0:
                yearly_rate = (
                    yearly_attended /
                    yearly_effective_total
                )
            else:
                yearly_rate = None

            row_data += [
                yearly_rate,
                yearly_present,
                yearly_late,
                yearly_farm,
                yearly_absent,
            ]

            ws_g.append(row_data)

        # =====================================================
        # 内容样式
        # =====================================================
        for row in ws_g.iter_rows(
            min_row=4,
            max_row=ws_g.max_row
        ):
            for cell in row:
                cell.border = border
                cell.alignment = center

            # C 至 N：1月至12月
            # O：全年出席率
            for cell in row[2:15]:

                if isinstance(
                    cell.value,
                    (int, float)
                ):
                    cell.number_format = "0%"

                    if cell.value >= 0.9:
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="C6EFCE"
                        )

                    elif cell.value >= 0.75:
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="FFEB9C"
                        )

                    else:
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="F4CCCC"
                        )

        ws_g.freeze_panes = "C4"

        if ws_g.max_row >= 4:
            ws_g.auto_filter.ref = (
                f"A3:"
                f"{get_column_letter(ws_g.max_column)}"
                f"{ws_g.max_row}"
            )

        ws_g.column_dimensions["A"].width = 14
        ws_g.column_dimensions["B"].width = 22

    # =========================================================
    # 全部 Sheet 自动美化
    # =========================================================
    for sheet in wb.worksheets:

        for col_idx in range(
            1,
            sheet.max_column + 1
        ):

            col_letter = get_column_letter(
                col_idx
            )

            max_len = 0

            for row_idx in range(
                1,
                sheet.max_row + 1
            ):

                value = sheet.cell(
                    row=row_idx,
                    column=col_idx
                ).value

                value = (
                    str(value)
                    if value is not None
                    else ""
                )

                max_len = max(
                    max_len,
                    len(value)
                )

            width = max(
                max_len + 4,
                12
            )

            if col_idx == 1:
                width = max(
                    width,
                    14
                )

            elif col_idx == 2:
                width = max(
                    width,
                    22
                )

            width = min(
                width,
                35
            )

            sheet.column_dimensions[
                col_letter
            ].width = width

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=(
                        cell.alignment.horizontal
                        or "center"
                    ),
                    vertical="center",
                    wrap_text=True
                )

        for row_idx in range(
            1,
            sheet.max_row + 1
        ):
            sheet.row_dimensions[
                row_idx
            ].height = 24

        sheet.sheet_view.showGridLines = False

    # =========================================================
    # 输出 Excel
    # =========================================================
    output = BytesIO()

    wb.save(output)
    output.seek(0)

    filename = (
        f"蕉赖佛学班出席年报_{year}.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

def build_homework_tracking_excel(
    report_type,
    year,
    month=None
):
    """生成白话佛法／念经文功课追踪 Excel。"""

    if report_type == "baihua":
        report_title = "白话佛法功课追踪表"
        status_field = "baihua_status"
        full_statuses = {"submitted_done", "submitted"}
        partial_statuses = {"submitted_no_answer"}
        full_label = "有交"
        partial_label = "有交没做题"
        legend_full = "✓ = 有交"
        legend_partial = "○ = 有交，没做题"

    elif report_type == "scripture":
        report_title = "念经文功课追踪表"
        status_field = "scripture_status"
        full_statuses = {"submitted_recited", "submitted"}
        partial_statuses = {"submitted_not_recited"}
        full_label = "有交有念"
        partial_label = "有交没念"
        legend_full = "✓ = 有交，有念"
        legend_partial = "○ = 有交，没念"

    else:
        raise ValueError("不正确的功课报表类型")

    if month:
        period_text = f"{year}-{month:02d}"
        homework_date_condition = """
            extract(year from h.class_date) = %s
            and extract(month from h.class_date) = %s
        """
        attendance_date_condition = """
            extract(year from a.class_date) = %s
            and extract(month from a.class_date) = %s
        """
        date_params = (int(year), int(month))
    else:
        period_text = str(year)
        homework_date_condition = """
            extract(year from h.class_date) = %s
        """
        attendance_date_condition = """
            extract(year from a.class_date) = %s
        """
        date_params = (int(year),)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name, sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            cur.execute("""
                select
                    s.id,
                    s.student_no,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.sort_order
                from dharma_students s
                join dharma_class_groups g on g.id = s.group_id
                where s.status = 'active'
                  and s.branch = 'CHE'
                  and g.is_active = true
                order by
                    g.sort_order,
                    s.student_no nulls last,
                    s.name
            """)
            students = cur.fetchall()

            cur.execute(f"""
                select
                    h.student_id,
                    h.group_id,
                    h.class_date,
                    h.{status_field} as homework_status
                from dharma_homework h
                where h.branch = 'CHE'
                  and {homework_date_condition}
                order by h.class_date, h.student_id
            """, date_params)
            homework_rows = cur.fetchall()

            # 以点名日期为课程日期，才能显示“未记录”。
            cur.execute(f"""
                select distinct
                    a.group_id,
                    a.class_date
                from dharma_attendance a
                where a.branch = 'CHE'
                  and {attendance_date_condition}
                order by a.class_date
            """, date_params)
            attendance_dates = cur.fetchall()

    homework_map = {
        (row["student_id"], row["class_date"]): row["homework_status"]
        for row in homework_rows
    }

    group_dates_map = {group["id"]: [] for group in groups}

    for row in attendance_dates:
        group_dates_map.setdefault(row["group_id"], [])
        if row["class_date"] not in group_dates_map[row["group_id"]]:
            group_dates_map[row["group_id"]].append(row["class_date"])

    # 若旧功课资料没有对应点名记录，也保留其日期。
    for row in homework_rows:
        group_dates_map.setdefault(row["group_id"], [])
        if row["class_date"] not in group_dates_map[row["group_id"]]:
            group_dates_map[row["group_id"]].append(row["class_date"])

    for group_id in group_dates_map:
        group_dates_map[group_id].sort()

    wb = Workbook()
    wb.remove(wb.active)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    full_fill = PatternFill("solid", fgColor="C6EFCE")
    partial_fill = PatternFill("solid", fgColor="FFF2CC")
    missing_fill = PatternFill("solid", fgColor="F4CCCC")
    absent_fill = PatternFill("solid", fgColor="D9E2F3")
    unknown_fill = PatternFill("solid", fgColor="E7E6E6")
    summary_fill = PatternFill("solid", fgColor="D9EAF7")

    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=18, color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for group in groups:
        ws = wb.create_sheet(group["name"][:31])

        group_students = [
            student for student in students
            if student["group_id"] == group["id"]
        ]
        class_dates = group_dates_map.get(group["id"], [])

        headers = ["学生编号", "学生姓名"]
        headers += [f"{d.month}/{d.day}" for d in class_dates]
        headers += [
            full_label,
            partial_label,
            "没交",
            "缺席",
            "未记录",
            "完整完成率",
        ]

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(len(headers), 8)
        )
        ws["A1"] = f"蕉赖佛学班 {group['name']} {report_title} {period_text}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 36

        ws["A2"] = "说明"
        ws["A2"].font = bold_font
        ws["B2"] = legend_full
        ws["C2"] = legend_partial
        ws["D2"] = "✗ = 没交"
        ws["E2"] = "缺 = 缺席"
        ws["F2"] = "空白 = 未记录"

        ws["B2"].fill = full_fill
        ws["C2"].fill = partial_fill
        ws["D2"].fill = missing_fill
        ws["E2"].fill = absent_fill
        ws["F2"].fill = unknown_fill

        for cell in ws[2]:
            cell.alignment = center

        ws.append(headers)
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        if not group_students:
            ws.append(["", "目前没有学生资料"])

        for student in group_students:
            row_data = [student["student_no"] or "", student["name"] or ""]

            full_count = 0
            partial_count = 0
            missing_count = 0
            absent_count = 0
            unrecorded_count = 0

            for class_date in class_dates:
                status = homework_map.get((student["id"], class_date))

                if status in full_statuses:
                    value = "✓"
                    full_count += 1
                elif status in partial_statuses:
                    value = "○"
                    partial_count += 1
                elif status == "missing":
                    value = "✗"
                    missing_count += 1
                elif status == "absent":
                    value = "缺"
                    absent_count += 1
                else:
                    value = ""
                    unrecorded_count += 1

                row_data.append(value)

            effective_total = full_count + partial_count + missing_count
            completion_rate = (
                full_count / effective_total
                if effective_total > 0
                else None
            )

            row_data += [
                full_count,
                partial_count,
                missing_count,
                absent_count,
                unrecorded_count,
                completion_rate,
            ]
            ws.append(row_data)

        date_start_col = 3
        date_end_col = 2 + len(class_dates)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = center

            for col_idx in range(date_start_col, date_end_col + 1):
                cell = ws.cell(row=row[0].row, column=col_idx)
                if cell.value == "✓":
                    cell.fill = full_fill
                    cell.font = bold_font
                elif cell.value == "○":
                    cell.fill = partial_fill
                    cell.font = bold_font
                elif cell.value == "✗":
                    cell.fill = missing_fill
                    cell.font = bold_font
                elif cell.value == "缺":
                    cell.fill = absent_fill
                    cell.font = bold_font
                else:
                    cell.fill = unknown_fill

            rate_cell = row[-1]
            if isinstance(rate_cell.value, (int, float)):
                rate_cell.number_format = "0%"
                if rate_cell.value >= 0.9:
                    rate_cell.fill = full_fill
                elif rate_cell.value >= 0.75:
                    rate_cell.fill = partial_fill
                else:
                    rate_cell.fill = missing_fill

        if group_students:
            total_row = ws.max_row + 1
            ws.merge_cells(
                start_row=total_row,
                start_column=1,
                end_row=total_row,
                end_column=2
            )
            ws.cell(total_row, 1, "合计")

            first_summary_col = 3 + len(class_dates)
            last_student_row = total_row - 1

            for col_idx in range(first_summary_col, first_summary_col + 5):
                letter = get_column_letter(col_idx)
                ws.cell(
                    total_row,
                    col_idx,
                    f"=SUM({letter}4:{letter}{last_student_row})"
                )

            full_col = first_summary_col
            partial_col = first_summary_col + 1
            missing_col = first_summary_col + 2
            rate_col = first_summary_col + 5

            full_letter = get_column_letter(full_col)
            partial_letter = get_column_letter(partial_col)
            missing_letter = get_column_letter(missing_col)

            ws.cell(
                total_row,
                rate_col,
                f'=IF({full_letter}{total_row}+{partial_letter}{total_row}+'
                f'{missing_letter}{total_row}=0,"",'
                f'{full_letter}{total_row}/('
                f'{full_letter}{total_row}+{partial_letter}{total_row}+'
                f'{missing_letter}{total_row}))'
            )
            ws.cell(total_row, rate_col).number_format = "0%"

            for cell in ws[total_row]:
                cell.border = border
                cell.alignment = center
                cell.font = bold_font
                cell.fill = summary_fill

        ws.freeze_panes = "C4"
        if ws.max_row >= 4:
            ws.auto_filter.ref = (
                f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
            )

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22

        for col_idx in range(3, 3 + len(class_dates)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 9

        for col_idx in range(3 + len(class_dates), ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 25

        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:3"
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    if not wb.sheetnames:
        ws = wb.create_sheet("提示")
        ws["A1"] = "目前没有启用中的组别"

    return wb


@dharma_class_bp.route("/reports/baihua/export")
def class_export_baihua_homework():

    year = request.args.get("year") or date.today().strftime("%Y")
    month = request.args.get("month")

    try:
        year_int = int(year)
    except (TypeError, ValueError):
        year_int = date.today().year

    try:
        month_int = int(month) if month else None

        if month_int is not None and not 1 <= month_int <= 12:
            month_int = None

    except (TypeError, ValueError):
        month_int = None

    wb = build_homework_tracking_excel(
        report_type="baihua",
        year=year_int,
        month=month_int
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if month_int:
        filename = (
            f"蕉赖佛学班白话佛法功课追踪表_"
            f"{year_int}-{month_int:02d}.xlsx"
        )
    else:
        filename = (
            f"蕉赖佛学班白话佛法功课追踪表_"
            f"{year_int}.xlsx"
        )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@dharma_class_bp.route("/reports/scripture/export")
def class_export_scripture_homework():

    year = request.args.get("year") or date.today().strftime("%Y")
    month = request.args.get("month")

    try:
        year_int = int(year)
    except (TypeError, ValueError):
        year_int = date.today().year

    try:
        month_int = int(month) if month else None

        if month_int is not None and not 1 <= month_int <= 12:
            month_int = None

    except (TypeError, ValueError):
        month_int = None

    wb = build_homework_tracking_excel(
        report_type="scripture",
        year=year_int,
        month=month_int
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if month_int:
        filename = (
            f"蕉赖佛学班经文功课追踪表_"
            f"{year_int}-{month_int:02d}.xlsx"
        )
    else:
        filename = (
            f"蕉赖佛学班经文功课追踪表_"
            f"{year_int}.xlsx"
        )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@dharma_class_bp.route("/promote", methods=["GET", "POST"])
def class_promote():

    target_year = request.values.get("target_year") or str(date.today().year + 1)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name, sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            cur.execute("""
                select
                    s.id,
                    s.student_no,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.name as group_name
                from dharma_students s
                join dharma_class_groups g on g.id = s.group_id
                where s.status = 'active'
                order by g.sort_order, s.student_no nulls last, s.name
            """)
            students = cur.fetchall()

            if request.method == "POST":

                updated = 0
                inactive = 0

                for s in students:
                    new_value = request.form.get(f"new_group_{s['id']}")

                    if not new_value:
                        continue

                    if new_value == "same":
                        continue

                    if new_value == "inactive":
                        cur.execute("""
                            update dharma_students
                            set status = 'inactive'
                            where id = %s
                        """, (s["id"],))
                        inactive += 1
                        continue

                    new_group_id = int(new_value)

                    if new_group_id != s["group_id"]:
                        cur.execute("""
                            update dharma_students
                            set group_id = %s
                            where id = %s
                        """, (new_group_id, s["id"]))
                        updated += 1

                conn.commit()

                flash(
                    f"{target_year} 年度升班完成：更新组别 {updated} 位，暂停 {inactive} 位。",
                    "good"
                )
                return redirect(url_for("dharma_class.class_students"))

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>年度升班</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">🎓 年度升班</h1>
        <p class="page-subtitle">
            老师可逐位选择新组别。没有要更动的学生，保留「保持原组」即可。
        </p>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert">{{ msg }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="get">
            <div class="form-group">
                <label class="form-label">目标学年</label>
                <input class="form-input" name="target_year" value="{{ target_year }}">
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-primary" type="submit">
                    🔍 查看学生
                </button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2 class="section-title">升班名单</h2>

        {% if students %}
        <form method="post">
            <input type="hidden" name="target_year" value="{{ target_year }}">

            <div class="table-responsive">
                <table class="record-table">
                    <thead>
                        <tr>
                            <th>编号</th>
                            <th>姓名</th>
                            <th>英文名</th>
                            <th>目前组别</th>
                            <th>新组别</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for s in students %}
                        <tr>
                            <td>{{ s.student_no or "-" }}</td>
                            <td>{{ s.name }}</td>
                            <td>{{ s.english_name or "-" }}</td>
                            <td>{{ s.group_name }}</td>
                            <td>
                                <select class="form-input" name="new_group_{{ s.id }}">
                                    <option value="same">保持原组</option>

                                    {% for g in groups %}
                                        <option value="{{ g.id }}"
                                            {% if
                                                (s.group_name == '低年组' and g.name == '少年组') or
                                                (s.group_name == '少年组' and g.name == '高年组')
                                            %}
                                                selected
                                            {% endif %}
                                        >
                                            {{ g.name }}
                                        </option>
                                    {% endfor %}

                                    <option value="inactive">暂停学生</option>
                                </select>
                            </td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-danger"
                        type="submit"
                        onclick="return confirm('确定要执行年度升班吗？请确认每位学生的新组别。');">
                    ✅ 确认执行升班
                </button>
            </div>
        </form>

        {% else %}
            <div class="empty-state">
                目前没有在册学生。
            </div>
        {% endif %}

        <div class="btn-row">
            <a class="btn-tool btn-secondary" href="/class">
                ⬅ 返回佛学班首页
            </a>
        </div>
    </div>

</div>
</body>
</html>
""",
        target_year=target_year,
        groups=groups,
        students=students
    )