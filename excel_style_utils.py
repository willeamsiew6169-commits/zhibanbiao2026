# excel_style_utils.py

import pandas as pd
import datetime

from openpyxl import load_workbook
from lunardate import LunarDate
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment  


def beautify_attendance_file(file_path, sheet_name):
       
    def get_display_width(s):
        """中文按2格，英文按1格估算显示宽度"""
        width = 0
        for ch in str(s):
            if ord(ch) > 127:
                width += 2
            else:
                width += 1
        return width

    def get_lunar_info(date_obj):
        """返回：星期、农历文字、特殊日子名称"""

        if isinstance(date_obj, datetime.datetime):
            date_obj = date_obj.date()

        elif isinstance(date_obj, datetime.date):
            pass

        else:
            date_obj = pd.to_datetime(date_obj, errors="coerce")
            if pd.isna(date_obj):
                return "", "", "", None
            date_obj = date_obj.date()

        weekday_map = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
        weekday_text = weekday_map[dt.weekday()]

        lunar = LunarDate.fromSolarDate(dt.year, dt.month, dt.day)

        month_names = {
            1: "正月", 2: "二月", 3: "三月", 4: "四月", 5: "五月", 6: "六月",
            7: "七月", 8: "八月", 9: "九月", 10: "十月", 11: "冬月", 12: "腊月"
        }
        day_names = {
            1: "初一", 2: "初二", 3: "初三", 4: "初四", 5: "初五",
            6: "初六", 7: "初七", 8: "初八", 9: "初九", 10: "初十",
            11: "十一", 12: "十二", 13: "十三", 14: "十四", 15: "十五",
            16: "十六", 17: "十七", 18: "十八", 19: "十九", 20: "二十",
            21: "廿一", 22: "廿二", 23: "廿三", 24: "廿四", 25: "廿五",
            26: "廿六", 27: "廿七", 28: "廿八", 29: "廿九", 30: "三十"
        }

        lunar_text = f"{month_names.get(lunar.month, str(lunar.month) + '月')}{day_names.get(lunar.day, str(lunar.day))}"

        # 你常用的特殊日子，可继续加
        special_days = {
            (1, 1): "农历初一",
            (1, 15): "农历十五",
            (2, 8): "释迦摩尼佛出家日",
            (2, 15): "释迦摩尼佛涅槃日",
            (2, 19): "观世音菩萨诞辰日",
            (4, 8): "释迦摩尼佛诞辰日",
            (6, 19): "观世音菩萨成道日",
            (9, 19): "观世音菩萨出家日",
            (10, 6):"恩师卢军宏涅槃日",
            (12, 8): "释迦摩尼佛成道日",
        }

        special_name = special_days.get((lunar.month, lunar.day), "")
        return weekday_text, lunar_text, special_name, lunar.day

    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        print(f"找不到工作表：{sheet_name}")
        print("现有工作表：", wb.sheetnames)
        return

    ws = wb[sheet_name]

    # ===== 样式 =====
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, name="微软雅黑", size=12)
    body_font = Font(name="微软雅黑", size=12)
    name_font = Font(name="微软雅黑", size=13, bold=True)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 日期颜色
    weekend_fill = PatternFill("solid", fgColor="DDEBF7")   # 周末：浅蓝
    lunar_fill = PatternFill("solid", fgColor="FFF2CC")     # 初一/十五：浅黄
    special_fill = PatternFill("solid", fgColor="F4CCCC")   # 观音诞/佛陀诞：浅红

    # 冻结首行
    ws.freeze_panes = "A2"

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row >= 1 and max_col >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # 找列位置
    col_map = {}
    for cell in ws[1]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    name_col = col_map.get("姓名")
    date_col = col_map.get("日期")

    # 全表样式
    for row in ws.iter_rows():
        for cell in row:
            cell.font = body_font
            cell.alignment = center_align
            cell.border = border

    # 表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 数据区样式
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 28

        if name_col:
            ws.cell(r, name_col).font = name_font
            ws.cell(r, name_col).alignment = left_align

    # 日期批注 + 颜色
    if date_col:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, date_col)
            val = cell.value

            if val is None or str(val).strip() == "":
                continue

            date_val = None

            if isinstance(val, datetime.datetime):
                dt_val = val.date()

            elif isinstance(val, datetime.date):
                dt_val = val

            else:
                parsed = pd.to_datetime(val, errors="coerce")
                if pd.isna(parsed):
                    continue
                dt = parsed.date()

            weekday_text, lunar_text, special_name, lunar_day = get_lunar_info(dt)

            # 保留原本批注内容
            existing_comment = cell.comment.text.strip() if cell.comment and cell.comment.text else ""

            comment_lines = []
            if existing_comment:
                comment_lines.append(existing_comment)

            comment_lines.append(weekday_text)
            comment_lines.append(lunar_text)

            if special_name:
                comment_lines.append(special_name)

            cell.comment = Comment("\n".join(comment_lines), "system")

            # 先清除日期格原本填色，再按优先级上色
            cell.fill = PatternFill(fill_type=None)

            if special_name:
                cell.fill = special_fill
            elif lunar_day in (1, 15):
                cell.fill = lunar_fill
            elif dt.weekday() >= 5:  # 周六=5, 周日=6
                cell.fill = weekend_fill

    # 自动列宽（支持中文）
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_length = 0

        for cell in col_cells:
            try:
                val = "" if cell.value is None else str(cell.value)
                length = get_display_width(val)
                max_length = max(max_length, length)
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    # 常用列宽微调
    if "日期" in col_map:
        ws.column_dimensions[get_column_letter(col_map["日期"])].width = 14

    if name_col:
        ws.column_dimensions[get_column_letter(name_col)].width = 16

    if "报名" in col_map:
        ws.column_dimensions[get_column_letter(col_map["报名"])].width = 8

    if "签到" in col_map:
        ws.column_dimensions[get_column_letter(col_map["签到"])].width = 8

    if "岗位" in col_map:
        ws.column_dimensions[get_column_letter(col_map["岗位"])].width = 12

    if "开始时间" in col_map:
        ws.column_dimensions[get_column_letter(col_map["开始时间"])].width = 13

    if "结束时间" in col_map:
        ws.column_dimensions[get_column_letter(col_map["结束时间"])].width = 13

    if "时数" in col_map:
        ws.column_dimensions[get_column_letter(col_map["时数"])].width = 8

    if "备注" in col_map:
        ws.column_dimensions[get_column_letter(col_map["备注"])].width = 20

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)