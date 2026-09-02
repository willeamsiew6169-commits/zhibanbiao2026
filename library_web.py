# library_web.py

import re
import os
import openpyxl

from db import get_conn
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from psycopg2.extras import RealDictCursor, Json
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from schedule.builders.time_utils import malaysia_now
from flask import Blueprint, flash, request, redirect, url_for, session, render_template_string


library_bp = Blueprint(
    "library",
    __name__,
    url_prefix="/library"
)

MALAYSIA_TZ = ZoneInfo("Asia/Kuala_Lumpur")
# 只用于第一次建立「系统开发者」账号。建立后，日常登录改用 Email + 个人 PIN。
LIBRARY_ADMIN_PIN = os.getenv("LIBRARY_ADMIN_PIN", "1234")

LIBRARY_ROLES = ("developer", "main_admin", "team_leader")
LIBRARY_ROLE_LABELS = {
    "developer": "系统开发者",
    "main_admin": "主负责人",
    "team_leader": "小组长",
}

LIBRARY_PERMISSION_FIELDS = [
    "can_batch_in",
    "can_batch_out",
    "can_view_records",
    "can_adjust_stock",
    "can_cancel_transaction",
    "can_manage_items",
    "can_stocktake",
    "can_view_audit_log",
    "can_manage_permissions",
]

def get_library_admin():
    admin_id = session.get("library_admin_id")
    if not admin_id:
        return None

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select
                    a.*,
                    v.name as volunteer_name
                from library_admin_users a
                left join volunteers v
                  on v.id::text = a.volunteer_id::text
                where a.id = %s
                  and a.is_active = true
                limit 1
            """, (admin_id,))
            return cur.fetchone()

def is_library_admin():
    return get_library_admin() is not None

def library_role(admin=None):
    admin = admin or get_library_admin()
    if not admin:
        return ""
    role = (admin.get("role") or "").strip()
    # 兼容旧资料：旧 Super Admin 自动视为系统开发者
    if not role and admin.get("is_super_admin"):
        return "developer"
    return role or "team_leader"


def library_role_label(admin=None):
    return LIBRARY_ROLE_LABELS.get(library_role(admin), "小组长")


def library_admin_can(permission):
    admin = get_library_admin()
    if not admin:
        return False

    role = library_role(admin)

    # 系统开发者拥有全部权限
    if role == "developer":
        return True

    # 只有系统开发者 / 主负责人可以管理负责人
    if permission == "can_manage_permissions" and role not in ("developer", "main_admin"):
        return False

    return bool(admin.get(permission))


def library_admin_label(admin=None):
    admin = admin or get_library_admin()
    if not admin:
        return "负责人"
    name = admin.get("volunteer_name") or admin.get("volunteer_id") or "负责人"
    vid = admin.get("volunteer_id") or ""
    return f"{vid} {name}".strip()

def write_library_audit(
    cur,
    action,
    *,
    item_code=None,
    item_name=None,
    quantity_change=None,
    balance_before=None,
    balance_after=None,
    transaction_id=None,
    detail=None,
    metadata=None,
    admin=None,
):
    admin = admin or get_library_admin()
    if not admin:
        return

    cur.execute("""
        insert into library_audit_logs (
            admin_user_id, volunteer_id, volunteer_name, user_email,
            action, item_code, item_name, quantity_change,
            balance_before, balance_after, transaction_id, detail, metadata
        )
        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        admin.get("id"),
        admin.get("volunteer_id"),
        admin.get("volunteer_name"),
        admin.get("email"),
        action, item_code, item_name, quantity_change,
        balance_before, balance_after, transaction_id, detail,
        Json(metadata or {})
    ))

def library_permission_denied():
    return render_template_string("""
    <!doctype html><html lang="zh"><head>
    <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
    <link rel="stylesheet" href="/static/css/toolbox.css">
    <title>没有权限</title></head><body><div class="page"><div class="card">
    <h1 class="page-title">🔒 没有权限</h1>
    <p class="page-subtitle">你的负责人账号没有这个功能的权限。</p>
    <div class="btn-row"><a class="btn-tool btn-secondary" href="/library/admin/home">返回负责人专区</a></div>
    </div></div></body></html>
    """), 403

def get_library_out_volunteer():
    return session.get("library_out_volunteer")

def get_library_cart():
    return session.setdefault("library_cart", [])

def now_malaysia():
    return datetime.now(MALAYSIA_TZ)

def guess_category(name):

    if "白话佛法" in name or "白话广播" in name:
        return "白话佛法"

    if "小房子" in name or "XFZ" in name:
        return "小房子"

    if "DVD" in name:
        return "DVD"

    if "MP3" in name:
        return "MP3"

    if "CD" in name:
        return "CD"

    if "宣传单" in name:
        return "宣传单"

    if (
        "经书" in name
        or "礼佛卡" in name
        or "自修" in name
        or "自存" in name
        or "大悲咒" in name
        or "心经" in name
        or "往生咒" in name
        or "准提神咒" in name
        or "消灾吉祥神咒" in name
    ):
        return "经书"

    if (
        "观 A" in name
        or "南 A" in name
        or "太 A" in name
        or "关 A" in name
        or "释 A" in name
        or "师父 A" in name
        or "背景画" in name
    ):
        return "佛像"

    if (
        "无烟香" in name
        or "吊坠" in name
        or "护身卡" in name
        or "念佛机" in name
        or "红信封" in name
        or "计数器" in name
        or "经书套" in name
        or "佛珠" in name
        or "扭蛋卡" in name
        or "法宝袋" in name
    ):
        return "法物"

    return "书籍"


def guess_unit(name, category):

    if category in ["DVD", "MP3", "CD"]:
        return "片"

    if category in ["宣传单", "佛像"]:
        return "张"

    if category == "法物":
        return "个"

    if "小房子" in name or "A4" in name or "A6" in name:
        return "张"

    return "本"


def make_item_code(index, category):

    prefix_map = {
        "白话佛法": "BHF",
        "经书": "JING",
        "宣传单": "FLYER",
        "DVD": "DVD",
        "MP3": "MP3",
        "CD": "CD",
        "佛像": "IMG",
        "法物": "MAT",
        "小房子": "XFZ",
        "书籍": "BOOK",
        "其他": "OTH",
    }

    prefix = prefix_map.get(category, "OTH")
    return f"{prefix}{index:04d}"


def make_keywords(name):

    import re

    text = name.replace(" ", "")
    keywords = {name, text}

    if "白话佛法" in name or "白话广播" in name:
        keywords.add("白话")
        keywords.add("bh")
        keywords.add("BHF")

        m = re.search(r"(\d+)", name)
        if m:
            n = m.group(1).zfill(2)
            keywords.add(f"白话{n}")
            keywords.add(f"bh{n}")
            keywords.add(f"BHF{n}")

    if "小房子" in name or "XFZ" in name:
        keywords.add("小房子")
        keywords.add("xfz")

    if "婚姻" in name:
        keywords.add("婚姻")
        keywords.add("情感")

    if "DVD" in name:
        keywords.add("dvd")

    if "MP3" in name:
        keywords.add("mp3")

    return " ".join(keywords)


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "")


def find_inventory_columns(ws):
    name_col = None
    balance_col = None
    header_row = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            text = normalize_header(cell.value)

            if not text:
                continue

            if "项目名称" in text or "itemname" in text or "name" == text:
                name_col = cell.column

            if (
                "closingbalance" in text
                or "期末余额" in text
                or "balance" in text
            ):
                balance_col = cell.column

        if name_col and balance_col:
            header_row = row[0].row
            break

    if not name_col or not balance_col:
        raise ValueError("找不到『项目名称』或『Closing Balance』栏位")

    return header_row, name_col, balance_col


def guess_category(name):

    if "白话佛法" in name or "白话广播" in name:
        return "白话佛法"

    if "小房子" in name or "XFZ" in name or "xfz" in name:
        return "小房子"

    if "DVD" in name:
        return "DVD"

    if "MP3" in name:
        return "MP3"

    if "CD" in name:
        return "CD"

    if "宣传单" in name:
        return "宣传单"

    if (
        "经书" in name
        or "礼佛卡" in name
        or "自修" in name
        or "自存" in name
        or "大悲咒" in name
        or "心经" in name
        or "往生咒" in name
        or "准提神咒" in name
        or "消灾吉祥神咒" in name
        or "礼佛大忏悔文" in name
    ):
        return "经书"

    if (
        "观 A" in name
        or "南 A" in name
        or "太 A" in name
        or "关 A" in name
        or "释 A" in name
        or "师父 A" in name
        or "背景画" in name
    ):
        return "佛像"

    if (
        "无烟香" in name
        or "吊坠" in name
        or "护身卡" in name
        or "念佛机" in name
        or "红信封" in name
        or "计数器" in name
        or "经书套" in name
        or "佛珠" in name
        or "扭蛋卡" in name
        or "法宝袋" in name
    ):
        return "法物"

    return "书籍"


def guess_unit(name, category):

    if category in ["DVD", "MP3", "CD"]:
        return "片"

    if category in ["宣传单", "佛像"]:
        return "张"

    if category == "法物":
        return "个"

    if "小房子" in name or "A4" in name or "A6" in name:
        return "张"

    return "本"


def make_keywords(name):

    text = name.replace(" ", "")
    keywords = {name, text}

    if "白话佛法" in name or "白话广播" in name:
        keywords.add("白话")
        keywords.add("bh")
        keywords.add("BHF")

        m = re.search(r"(\d+)", name)
        if m:
            n = m.group(1).zfill(2)
            keywords.add(f"白话{n}")
            keywords.add(f"bh{n}")
            keywords.add(f"BHF{n}")

    if "小房子" in name or "XFZ" in name or "xfz" in name:
        keywords.add("小房子")
        keywords.add("xfz")

    if "婚姻" in name:
        keywords.add("婚姻")
        keywords.add("情感")

    if "佛言佛语" in name:
        keywords.add("佛言佛语")

    if "DVD" in name:
        keywords.add("dvd")

    if "MP3" in name:
        keywords.add("mp3")

    return " ".join(sorted(keywords))


def make_item_code(index, category):

    prefix_map = {
        "白话佛法": "BHF",
        "经书": "JING",
        "宣传单": "FLYER",
        "DVD": "DVD",
        "MP3": "MP3",
        "CD": "CD",
        "佛像": "IMG",
        "法物": "MAT",
        "小房子": "XFZ",
        "书籍": "BOOK",
    }

    prefix = prefix_map.get(category, "BOOK")
    return f"{prefix}{index:04d}"

@library_bp.route("/upload", methods=["GET", "POST"])
def library_upload():

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_manage_items"):
        return library_permission_denied()

    message = None
    error = None

    if request.method == "POST":

        file = request.files.get("file")

        if not file or file.filename == "":
            error = "请选择 Excel 文件"

        else:
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active

                header_row, name_col, balance_col = find_inventory_columns(ws)

                rows = []

                for row_idx in range(header_row + 1, ws.max_row + 1):

                    name = ws.cell(row=row_idx, column=name_col).value
                    balance = ws.cell(row=row_idx, column=balance_col).value

                    if not name:
                        continue

                    name = str(name).strip()

                    if not name or name == "项目名称":
                        continue

                    try:
                        balance = int(balance or 0)
                    except:
                        balance = 0

                    category = guess_category(name)
                    unit = guess_unit(name, category)

                    rows.append({
                        "name": name,
                        "balance": balance,
                        "category": category,
                        "unit": unit,
                        "keywords": make_keywords(name),
                    })

                with get_conn() as conn:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:

                        for index, item in enumerate(rows, start=1):

                            item_code = make_item_code(index, item["category"])

                            cur.execute("""
                                insert into library_items (
                                    item_code,
                                    name,
                                    category,
                                    keywords,
                                    unit,
                                    balance,
                                    min_balance,
                                    is_active
                                )
                                values (%s,%s,%s,%s,%s,%s,%s,true)
                                on conflict (item_code)
                                do update set
                                    name = excluded.name,
                                    category = excluded.category,
                                    keywords = excluded.keywords,
                                    unit = excluded.unit,
                                    balance = excluded.balance,
                                    is_active = true
                            """, (
                                item_code,
                                item["name"],
                                item["category"],
                                item["keywords"],
                                item["unit"],
                                item["balance"],
                                0,
                            ))

                        conn.commit()

                message = f"上传完成：共导入 {len(rows)} 项法物"

            except Exception as e:
                error = f"上传失败：{e}"

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>上传藏经阁库存</title>
        <link rel="stylesheet" href="/static/css/toolbox.css">
    </head>
    <body>
    <div class="page">

        <div class="card">
            <h1 class="page-title">📤 上传藏经阁库存</h1>
            <p class="page-subtitle">上传 library_inventory.xlsx 更新库存资料</p>

            {% if message %}
                <div class="alert alert-success">{{ message }}</div>
            {% endif %}

            {% if error %}
                <div class="alert alert-danger">{{ error }}</div>
            {% endif %}

            <form method="post" enctype="multipart/form-data">

                <div class="form-group">
                    <label class="form-label">选择 Excel 文件</label>
                    <input class="form-input"
                           type="file"
                           name="file"
                           accept=".xlsx"
                           required>
                </div>

                <div class="btn-row">
                    <button class="btn-tool btn-success" type="submit">
                        上传并更新库存
                    </button>

                    <a class="btn-tool btn-secondary" href="/library">
                        返回
                    </a>
                </div>

            </form>
        </div>

    </div>
    </body>
    </html>
    """, message=message, error=error)


@library_bp.route("/")
def library_home():
    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">

    <title>藏经阁系统</title>

    <meta name="theme-color" content="#8B4513">

    <link rel="manifest" href="/library-manifest.json?v=5">

    <link rel="icon"
        type="image/png"
        sizes="192x192"
        href="/static/library_icon_192.png?v=5">

    <link rel="apple-touch-icon"
        href="/static/library_icon_512.png?v=5">

    <meta name="theme-color" content="#8B4513">

    <link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>

<!-- 法物查询 -->
<div class="page">

    <div class="card"
         style="
            max-width:1000px;
            margin:40px auto;
            padding:45px 55px 50px;
         ">

        <!-- 标题 -->
        <div style="text-align:center; margin-bottom:32px;">

            <h1 class="page-title"
                style="font-size:38px; margin-bottom:8px;">
                📚 藏经阁系统
            </h1>

            <p class="page-subtitle"
               style="font-size:20px; margin:0;">
                查询法物、登记领取及管理藏经阁库存
            </p>

        </div>


        <!-- 法物查询 -->
        <div style="
            max-width:760px;
            margin:0 auto;
        ">

            <form method="get" action="/library/search">

                <div class="form-group">

                    <label style="
                        font-size:25px;
                        font-weight:700;
                        display:block;
                        margin-bottom:12px;
                    ">
                        🔍 法物查询
                    </label>

                    <input
                        class="form-input"
                        type="text"
                        name="q"
                        placeholder="例如：白话06 / 婚姻 / 小房子 / DVD"
                        autocomplete="off"
                        style="
                            font-size:21px;
                            min-height:64px;
                        "
                    >

                </div>

                <button
                    class="btn-tool btn-primary"
                    type="submit"
                    style="
                        width:100%;
                        font-size:22px;
                        min-height:64px;
                        margin-top:10px;
                    ">
                    🔍 搜索
                </button>

            </form>

        </div>


        <!-- 手动登记领取 -->
        <div style="
            display:flex;
            justify-content:center;
            margin-top:28px;
        ">

            <a class="btn-tool btn-success"
               href="/library/out"
               style="
                    width:430px;
                    max-width:90%;
                    font-size:21px;
                    min-height:60px;
               ">
                📝 手动登记领取
            </a>

        </div>


        <!-- 负责人专区 -->
        <div style="
            text-align:center;
            margin-top:42px;
        ">

            <a class="btn-tool btn-warning"
               href="/library/admin"
               style="
                    display:inline-flex;
                    width:auto;
                    min-width:210px;
                    justify-content:center;
                    font-size:19px;
                    min-height:52px;
                    padding:10px 26px;
               ">
                ⚙️ 负责人专区
            </a>

        </div>

    </div>

</div>

</body>
</html>
""")

@library_bp.route("/search")
def library_search():

    keyword = request.args.get("q", "").strip()

    items = []

    if keyword:

        like = f"%{keyword}%"
        like_no_space = f"%{keyword.replace(' ', '')}%"

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                cur.execute("""
                    select
                        item_code,
                        name,
                        category,
                        keywords,
                        location,
                        unit,
                        balance,
                        min_balance
                    from library_items
                    where
                        is_active = true
                        and (
                            name ilike %s
                            or replace(name,' ','') ilike %s
                            or category ilike %s
                            or keywords ilike %s
                            or replace(coalesce(keywords,''),' ','') ilike %s
                            or item_code ilike %s
                        )
                    order by
                        category,
                        name
                    limit 100
                """, (
                    like,
                    like_no_space,
                    like,
                    like,
                    like_no_space,
                    like,
                ))

                items = cur.fetchall()

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>法物查询</title>
        <link rel="stylesheet" href="/static/css/toolbox.css">
    </head>

    <body>
    <div class="page">

        <div class="card">
            <h1 class="page-title">🔍 法物查询</h1>
            <p class="page-subtitle">输入书名、关键字、分类或编号</p>

            <form method="get" action="/library/search">
                <div class="form-group">
                    <label class="form-label">搜索</label>
                    <input
                        class="form-input"
                        name="q"
                        value="{{ keyword }}"
                        placeholder="例如：白话06 / 婚姻 / 小房子 / DVD"
                        autofocus
                    >
                </div>

                <div class="btn-row">
                    <button class="btn-tool btn-primary" type="submit">
                        🔎 搜索
                    </button>

                    <a class="btn-tool btn-secondary" href="/library">
                        返回
                    </a>
                </div>
            </form>
        </div>

        {% if keyword and not items %}
        <div class="card">
            <div class="empty-state">
                找不到相关物
            </div>
        </div>
        {% endif %}

        {% if items %}
        <div class="card">
            <h2 class="section-title">查询结果</h2>
            <p class="page-subtitle">共找到 {{ items|length }} 项法物</p>
        </div>

        {% for item in items %}
        <div class="card" style="margin-bottom:16px;">

            <h2 class="section-title" style="margin-bottom:8px;">
                <a href="/library/item/{{ item.item_code }}"
                   style="text-decoration:none; color:inherit;">
                    📚 {{ item.name }}
                </a>
            </h2>

            <p style="color:#666; margin-top:0;">
                {{ item.item_code }}
            </p>

            <div class="summary-grid">

                <div class="summary-box">
                    <div class="summary-title">分类</div>
                    <div class="summary-value" style="font-size:18px;">
                        {{ item.category or "-" }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">库存</div>
                    <div class="summary-value" style="font-size:22px;">
                        {{ item.balance }}
                        {{ item.unit or "" }}
                    </div>
                </div>

            </div>

            {% if item.min_balance and item.balance <= item.min_balance %}
            <div class="alert alert-warning" style="margin-top:12px;">
                ⚠ 库存偏低，最低库存：{{ item.min_balance }} {{ item.unit or "" }}
            </div>
            {% endif %}

            <p style="margin-top:14px;">
                📍 位置：{{ item.location or "未设置" }}
            </p>

            <div class="btn-row">
                <a class="btn-tool btn-primary"
                   href="/library/item/{{ item.item_code }}">
                    查看详情
                </a>
            </div>

        </div>
        {% endfor %}
        {% endif %}

    </div>
    </body>
    </html>
    """, keyword=keyword, items=items)

@library_bp.route("/item/<item_code>")
def library_item_detail(item_code):

    item = None
    records = []

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    item_code,
                    name,
                    category,
                    keywords,
                    location,
                    unit,
                    balance,
                    min_balance,
                    description
                from library_items
                where item_code = %s
                limit 1
            """, (item_code,))

            item = cur.fetchone()

            if item:
                cur.execute("""
                    select
                        transaction_type,
                        quantity,
                        volunteer_id,
                        volunteer_name,
                        handled_by,
                        remark,
                        to_char(
                            created_at at time zone 'Asia/Kuala_Lumpur',
                            'YYYY-MM-DD HH24:MI'
                        ) as created_at_my
                    from library_transactions
                    where item_code = %s
                    order by created_at desc
                    limit 10
                """, (item_code,))

                records = cur.fetchall()

    if not item:
        return "找不到这个法物", 404

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{{ item.name }}</title>
        <link rel="stylesheet" href="/static/css/toolbox.css">
    </head>

    <body>
    <div class="page">

        <div class="card">
            <h1 class="page-title">📘 {{ item.name }}</h1>
            <p class="page-subtitle">
                {{ item.category or "未分类" }}
            </p>

            <div class="summary-grid">

                <div class="summary-box">
                    <div class="summary-title">当前库存</div>
                    <div class="summary-value">
                        {{ item.balance }} {{ item.unit or "" }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">存放位置</div>
                    <div class="summary-value">
                        {{ item.location or "未设置" }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">最低库存</div>
                    <div class="summary-value">
                        {{ item.min_balance or 0 }} {{ item.unit or "" }}
                    </div>
                </div>

            </div>

            {% if item.description %}
                <div class="alert alert-info">
                    {{ item.description }}
                </div>
            {% endif %}

            {% if item.min_balance and item.balance <= item.min_balance %}
                <div class="alert alert-danger">
                    ⚠ 这个法物库存偏低，请留意补货。
                </div>
            {% endif %}

            <div class="btn-row">
                <a class="btn-tool btn-success"
                   href="/library/out?item_code={{ item.item_code }}">
                    📤 出库
                </a>

                <a class="btn-tool btn-warning"
                   href="/library/in?item_code={{ item.item_code }}">
                    📥 入库
                </a>

                <a class="btn-tool btn-secondary"
                   href="/library/search">
                    返回查询
                </a>
            </div>
        </div>

        <div class="card">
            <h2 class="section-title">最近记录</h2>

            {% if records %}
                <div class="table-responsive">
                    <table class="record-table">
                        <thead>
                            <tr>
                                <th>时间</th>
                                <th>类型</th>
                                <th>数量</th>
                                <th>义工/负责人</th>
                                <th>备注</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for r in records %}
                            <tr>
                                <td>{{ r.created_at_my }}</td>
                                <td>
                                    {% if r.transaction_type == "out" %}
                                        出库
                                    {% elif r.transaction_type == "in" %}
                                        入库
                                    {% elif r.transaction_type == "adjust" %}
                                        调整
                                    {% else %}
                                        {{ r.transaction_type }}
                                    {% endif %}
                                </td>
                                <td>{{ r.quantity }}</td>
                                <td>
                                    {{ r.volunteer_id or r.handled_by or "-" }}
                                    {% if r.volunteer_name %}
                                        <br>{{ r.volunteer_name }}
                                    {% endif %}
                                </td>
                                <td>{{ r.remark or "-" }}</td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% else %}
                <div class="empty-state">
                    暂时还没有进出记录
                </div>
            {% endif %}
        </div>

    </div>
    </body>
    </html>
    """, item=item, records=records)

@library_bp.route("/out", methods=["GET", "POST"])
def library_out():

    volunteer = get_library_out_volunteer()

    if volunteer:
        return redirect(url_for("library.library_out_search"))

    error = None

    if request.method == "POST":

        keyword = request.form.get("volunteer_id", "").strip()
        branch = request.form.get("branch", "CHE").strip().upper()

        if keyword.isdigit():
            volunteer_id = f"{branch}-{int(keyword)}"
        else:
            volunteer_id = keyword.upper()

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                cur.execute("""
                    select
                        id,
                        name,
                        status
                    from volunteers
                    where id=%s
                    limit 1
                """, (volunteer_id,))

                vol = cur.fetchone()

        if not vol:
            error = "找不到這位義工"

        else:

            session["library_out_volunteer"] = {
                "id": str(vol["id"]),
                "name": vol["name"]
            }

            # 如果是掃 QR Code 進來，登入後回到 QR 法寶頁
            next_url = session.pop("library_after_login_url", None)

            if next_url:
                return redirect(next_url)

            # 一般登入，照舊到搜尋頁
            return redirect(
                url_for("library.library_out_search")
            )

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">
    <link rel="stylesheet" href="/static/css/toolbox.css">
    <title>义工领取法物</title>
    </head>

    <body>

    <div class="page">

    <div class="card">

    <h1 class="page-title">📝 登记领取法物</h1>

    <p class="page-subtitle">
    会使用系统的义工可直接登记；不会使用的义工可照旧写在纸上，由负责人之后补录。
    </p>

    {% if error %}
    <div class="alert alert-danger">
    {{ error }}
    </div>
    {% endif %}

    <form method="post">

    <div class="form-group">
    <label class="form-label">分会</label>

    <div class="btn-row">
        <button
            type="button"
            id="branchBtn"
            class="btn-tool btn-success"
            onclick="toggleBranch()">
            CHE
        </button>
    </div>

    <input type="hidden" id="branch" name="branch" value="CHE">

    </div>

    <div class="form-group">

    <label class="form-label">义工编号</label>

    <input
    name="volunteer_id"
    class="form-input"
    placeholder="例如：108"
    inputmode="numeric"
    autocomplete="off"
    required>

    </div>

    <div class="btn-row" style="margin-top:25px;">

    <button
    type="submit"
    class="btn-tool btn-success">
    继续领取
    </button>

    <a
    href="/library"
    class="btn-tool btn-secondary">
    返回首页
    </a>

    </div>

    </form>

    </div>

    </div>

    <script>
function toggleBranch() {
    const branchInput = document.getElementById("branch");
    const branchBtn = document.getElementById("branchBtn");

    if (branchInput.value === "CHE") {
        branchInput.value = "STW";
        branchBtn.innerText = "STW";

        branchBtn.className = "btn-tool btn-danger";
    } else {
        branchInput.value = "CHE";
        branchBtn.innerText = "CHE";

        branchBtn.className = "btn-tool btn-success";
    }
}
</script>

    </body>
    </html>
    """, error=error)

@library_bp.route("/admin", methods=["GET", "POST"])
def library_admin():
    if is_library_admin():
        return redirect(url_for("library.library_admin_home"))

    error = None
    bootstrap = False

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("select count(*) as c from library_admin_users")
            bootstrap = int(cur.fetchone()["c"] or 0) == 0

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        pin = request.form.get("pin", "").strip()

        if bootstrap:
            volunteer_id = request.form.get("volunteer_id", "").strip().upper()
            master_pin = request.form.get("master_pin", "").strip()

            if master_pin != LIBRARY_ADMIN_PIN:
                error = "首次设置 PIN 不正确"
            elif not volunteer_id or not email or len(pin) < 4:
                error = "请完整填写义工编号、Email 和至少 4 位个人 PIN"
            else:
                with get_conn() as conn:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute("select id, name from volunteers where id::text=%s limit 1", (volunteer_id,))
                        vol = cur.fetchone()
                        if not vol:
                            error = "找不到这位义工"
                        else:
                            cur.execute("""
                                insert into library_admin_users (
                                    volunteer_id, email, pin_hash,
                                    role, is_active, is_super_admin,
                                    can_batch_in, can_batch_out, can_view_records,
                                    can_adjust_stock, can_cancel_transaction,
                                    can_manage_items, can_stocktake,
                                    can_view_audit_log, can_manage_permissions
                                ) values (
                                    %s,%s,%s,
                                    'developer',true,true,
                                    true,true,true,true,true,true,true,true,true
                                )
                                returning id
                            """, (volunteer_id, email, generate_password_hash(pin)))
                            admin_id = cur.fetchone()["id"]
                            conn.commit()
                            session["library_admin_id"] = admin_id
                            session.modified = True

                            admin = get_library_admin()
                            with get_conn() as log_conn:
                                with log_conn.cursor(cursor_factory=RealDictCursor) as log_cur:
                                    write_library_audit(log_cur, "ADMIN_BOOTSTRAP", detail="建立系统开发者账号", admin=admin)
                                    log_conn.commit()
                            return redirect(url_for("library.library_admin_home"))
        else:
            if not email or not pin:
                error = "请输入 Email 和个人 PIN"
            else:
                with get_conn() as conn:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute("""
                            select id, pin_hash
                            from library_admin_users
                            where lower(email)=lower(%s)
                              and is_active=true
                            limit 1
                        """, (email,))
                        row = cur.fetchone()

                if not row or not check_password_hash(row["pin_hash"], pin):
                    error = "Email 或个人 PIN 不正确"
                else:
                    session["library_admin_id"] = row["id"]
                    session.modified = True
                    admin = get_library_admin()
                    with get_conn() as log_conn:
                        with log_conn.cursor(cursor_factory=RealDictCursor) as log_cur:
                            write_library_audit(log_cur, "LOGIN", detail="登入负责人专区", admin=admin)
                            log_conn.commit()
                    return redirect(url_for("library.library_admin_home"))

    return render_template_string("""
<!doctype html>
<html lang="zh"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>藏经阁负责人专区</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head><body><div class="page"><div class="card">
<h1 class="page-title">⚙️ 负责人专区</h1>

{% if bootstrap %}
<p class="page-subtitle">第一次使用：建立系统开发者账号。以后主负责人和小组长都使用自己的 Email + 个人 PIN 登录。</p>
{% else %}
<p class="page-subtitle">请输入自己的 Email 与个人 PIN</p>
{% endif %}

{% if error %}<div class="alert alert-danger">{{ error }}</div>{% endif %}

<form method="post">
{% if bootstrap %}
<div class="form-group"><label class="form-label">义工编号</label>
<input class="form-input" name="volunteer_id" placeholder="例如 CHE-208" required></div>
{% endif %}

<div class="form-group"><label class="form-label">Email</label>
<input class="form-input" type="email" name="email" placeholder="例如 name@gmail.com" required autocomplete="username"></div>

<div class="form-group"><label class="form-label">个人 PIN</label>
<input class="form-input" type="password" name="pin" placeholder="至少 4 位" required autocomplete="current-password"></div>

{% if bootstrap %}
<div class="form-group"><label class="form-label">原负责人 PIN（只用于第一次设置）</label>
<input class="form-input" type="password" name="master_pin" required></div>
{% endif %}

<div class="btn-row">
<button class="btn-tool btn-warning" type="submit">登入</button>
<a class="btn-tool btn-secondary" href="/library">返回</a>
</div></form>
</div></div></body></html>
""", error=error, bootstrap=bootstrap)

@library_bp.route("/admin/home")
def library_admin_home():
    admin = get_library_admin()
    if not admin:
        return redirect(url_for("library.library_admin"))

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""select count(*) as total_items, coalesce(sum(balance),0) as total_balance
                           from library_items where is_active=true""")
            summary = cur.fetchone()
            cur.execute("""select count(*) as low_stock_count from library_items
                           where is_active=true and balance <= min_balance""")
            low_stock = cur.fetchone()
            cur.execute("""select coalesce(sum(quantity),0) as today_out from library_transactions
                           where transaction_type='out'
                           and (created_at at time zone 'Asia/Kuala_Lumpur')::date=(now() at time zone 'Asia/Kuala_Lumpur')::date""")
            today_out = cur.fetchone()
            cur.execute("""select coalesce(sum(quantity),0) as today_in from library_transactions
                           where transaction_type='in'
                           and (created_at at time zone 'Asia/Kuala_Lumpur')::date=(now() at time zone 'Asia/Kuala_Lumpur')::date""")
            today_in = cur.fetchone()
            cur.execute("""select item_code,item_name,transaction_type,quantity,volunteer_name,handled_by,
                           to_char(created_at at time zone 'Asia/Kuala_Lumpur','YYYY-MM-DD HH24:MI') as created_at_my
                           from library_transactions order by created_at desc limit 5""")
            recent_rows = cur.fetchall()

    return render_template_string("""
<!doctype html><html lang="zh"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>藏经阁负责人</title><link rel="stylesheet" href="/static/css/toolbox.css"></head><body>
<div class="page"><div class="card">
<h1 class="page-title">📚 藏经阁负责人</h1>
<p class="page-subtitle">{{ admin.volunteer_name or admin.volunteer_id }} · {{ role_label }} · {{ admin.email }}</p>

<div class="card"><h2 class="section-title">📊 今日概况</h2><div class="summary-grid">
<div class="summary-box"><div class="summary-title">法物种类</div><div class="summary-value">{{ summary.total_items }}</div></div>
<div class="summary-box"><div class="summary-title">总库存</div><div class="summary-value">{{ summary.total_balance }}</div></div>
<div class="summary-box"><div class="summary-title">今日领取</div><div class="summary-value">{{ today_out.today_out }}</div></div>
<div class="summary-box"><div class="summary-title">今日入库</div><div class="summary-value">{{ today_in.today_in }}</div></div>
<div class="summary-box"><div class="summary-title">库存不足</div><div class="summary-value">{{ low_stock.low_stock_count }}</div></div>
</div></div>

<div class="card"><h2 class="section-title">🕒 最近记录</h2>
{% if not recent_rows %}<div class="empty-state">暂时没有进出记录。</div>{% else %}
{% for row in recent_rows %}<div style="border-bottom:1px solid #eee;padding:12px 0;">
{% if row.transaction_type == 'out' %}<b style="color:#dc3545;">📤 出库</b>
{% elif row.transaction_type == 'in' %}<b style="color:#28a745;">📥 入库</b>
{% else %}<b style="color:#ff9800;">⚙️ 调整</b>{% endif %}<br>
📚 {{ row.item_name }}<br>数量：<b>{{ row.quantity }}</b>
{% if row.volunteer_name %}<br>👤 {{ row.volunteer_name }}{% endif %}
{% if row.handled_by %}<br>👨‍💼 {{ row.handled_by }}{% endif %}<br>
<small style="color:#666;">{{ row.created_at_my }}</small></div>{% endfor %}{% endif %}</div>

<div class="btn-row">
{% if can_batch_in %}<a class="btn-tool btn-success" href="/library/batch-in">📦 批量入库</a><a class="btn-tool btn-success" href="/library/in">📥 单项入库</a>{% endif %}
{% if can_view_records %}<a class="btn-tool btn-primary" href="/library/records">📋 进出记录</a>{% endif %}
<a class="btn-tool btn-warning" href="/library/low-stock">📊 库存不足</a>
{% if can_stocktake or can_adjust_stock %}<a class="btn-tool btn-purple" href="{{ url_for('library.library_stocktake') }}">📤 上传盘点 Excel</a>{% endif %}
{% if can_manage_items %}<a class="btn-tool btn-primary" href="/library/items">⚙️ 法物管理</a>{% endif %}
{% if can_view_audit_log %}<a class="btn-tool" style="background:#5b6472;color:white;" href="/library/admin/audit">🧾 操作日志</a>{% endif %}
{% if can_manage_permissions %}<a class="btn-tool" style="background:#0d9488;color:white;" href="/library/admin/users">👥 负责人权限</a>{% endif %}
</div>
<hr style="margin:35px 0;"><div class="btn-row">
<a class="btn-tool btn-secondary" href="/library">返回藏经阁首页</a>
<a class="btn-tool btn-danger" href="/library/admin/logout">退出负责人专区</a>
</div></div></div></body></html>
""",
        admin=admin,
        role_label=library_role_label(admin),
        can_batch_in=library_admin_can("can_batch_in"),
        can_view_records=library_admin_can("can_view_records"),
        can_stocktake=library_admin_can("can_stocktake"),
        can_adjust_stock=library_admin_can("can_adjust_stock"),
        can_manage_items=library_admin_can("can_manage_items"),
        can_view_audit_log=library_admin_can("can_view_audit_log"),
        can_manage_permissions=library_admin_can("can_manage_permissions"),
        summary=summary,
        low_stock=low_stock,
        today_out=today_out,
        today_in=today_in,
        recent_rows=recent_rows
    )

@library_bp.route("/admin/users", methods=["GET", "POST"])
def library_admin_users():
    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_manage_permissions"):
        return library_permission_denied()

    admin = get_library_admin()
    current_role = library_role(admin)
    error = None
    success = None

    # 系统开发者可以建立主负责人 / 小组长；
    # 主负责人只能建立和管理小组长。
    allowed_roles = ["main_admin", "team_leader"] if current_role == "developer" else ["team_leader"]

    if request.method == "POST":
        user_id = request.form.get("user_id", "").strip()
        volunteer_id = request.form.get("volunteer_id", "").strip().upper()
        email = request.form.get("email", "").strip().lower()
        pin = request.form.get("pin", "").strip()
        is_active = request.form.get("is_active") == "1"
        target_role = request.form.get("role", "team_leader").strip()

        if target_role not in LIBRARY_ROLES:
            target_role = "team_leader"

        perms = {p: request.form.get(p) == "1" for p in LIBRARY_PERMISSION_FIELDS}

        if not volunteer_id or not email:
            error = "义工编号和 Email 必须填写"
        elif target_role == "developer":
            error = "系统开发者账号不能在这里新增或升级"
        elif target_role not in allowed_roles:
            error = "你没有权限建立或修改这个级别的负责人"
        else:
            with get_conn() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("select id,name from volunteers where id::text=%s limit 1", (volunteer_id,))
                    vol = cur.fetchone()

                    if not vol:
                        error = "找不到这位义工"
                    elif not user_id and len(pin) < 4:
                        error = "新增负责人必须设置至少 4 位个人 PIN"
                    else:
                        try:
                            before = None

                            if user_id:
                                cur.execute("select * from library_admin_users where id=%s limit 1", (user_id,))
                                before = cur.fetchone()

                                if not before:
                                    error = "找不到这位负责人"
                                else:
                                    before_role = (before.get("role") or "").strip()
                                    if not before_role and before.get("is_super_admin"):
                                        before_role = "developer"
                                    before_role = before_role or "team_leader"

                                    # 开发者账号不能在一般权限页面被修改/停用。
                                    if before_role == "developer":
                                        error = "系统开发者账号不能在负责人权限页面修改"
                                    elif current_role == "main_admin" and before_role != "team_leader":
                                        error = "主负责人只能管理小组长"
                                    elif current_role == "main_admin" and target_role != "team_leader":
                                        error = "主负责人不能把小组长升级为主负责人"
                                    elif str(before["id"]) == str(admin["id"]):
                                        error = "不能在这里停用或修改自己的账号"
                                    else:
                                        sets = [
                                            "volunteer_id=%s",
                                            "email=%s",
                                            "role=%s",
                                            "is_active=%s",
                                            "is_super_admin=false",
                                        ]
                                        vals = [volunteer_id, email, target_role, is_active]

                                        for p in LIBRARY_PERMISSION_FIELDS:
                                            sets.append(f"{p}=%s")
                                            vals.append(perms[p])

                                        if pin:
                                            sets.append("pin_hash=%s")
                                            vals.append(generate_password_hash(pin))

                                        sets.append("updated_at=now()")
                                        vals.append(user_id)

                                        cur.execute(
                                            f"update library_admin_users set {', '.join(sets)} where id=%s",
                                            vals
                                        )

                                        action = "ADMIN_EDIT"
                                        detail = (
                                            f"修改负责人：{volunteer_id} "
                                            f"({LIBRARY_ROLE_LABELS.get(before_role, before_role)}"
                                            f" → {LIBRARY_ROLE_LABELS.get(target_role, target_role)})"
                                        )
                            else:
                                cols = [
                                    "volunteer_id", "email", "pin_hash", "role",
                                    "is_active", "is_super_admin"
                                ] + LIBRARY_PERMISSION_FIELDS

                                vals = [
                                    volunteer_id,
                                    email,
                                    generate_password_hash(pin),
                                    target_role,
                                    is_active,
                                    False,
                                ] + [perms[p] for p in LIBRARY_PERMISSION_FIELDS]

                                placeholders = ",".join(["%s"] * len(vals))
                                cur.execute(
                                    f"insert into library_admin_users ({','.join(cols)}) "
                                    f"values ({placeholders}) returning id",
                                    vals
                                )
                                user_id = str(cur.fetchone()["id"])
                                action = "ADMIN_CREATE"
                                detail = (
                                    f"新增{LIBRARY_ROLE_LABELS.get(target_role, target_role)}："
                                    f"{volunteer_id}"
                                )

                            if not error:
                                write_library_audit(
                                    cur,
                                    action,
                                    detail=detail,
                                    metadata={
                                        "target_admin_id": user_id,
                                        "volunteer_id": volunteer_id,
                                        "email": email,
                                        "role": target_role,
                                        "is_active": is_active,
                                        "permissions": perms,
                                    },
                                    admin=admin,
                                )
                                conn.commit()
                                success = "负责人资料已保存"

                        except Exception as e:
                            conn.rollback()
                            error = f"保存失败：{e}"

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if current_role == "developer":
                cur.execute("""
                    select a.*, v.name as volunteer_name
                    from library_admin_users a
                    left join volunteers v on v.id::text=a.volunteer_id::text
                    order by
                        case coalesce(a.role,'team_leader')
                            when 'developer' then 1
                            when 'main_admin' then 2
                            else 3
                        end,
                        a.is_active desc,
                        coalesce(v.name,a.volunteer_id)
                """)
            else:
                cur.execute("""
                    select a.*, v.name as volunteer_name
                    from library_admin_users a
                    left join volunteers v on v.id::text=a.volunteer_id::text
                    where coalesce(a.role,'team_leader')='team_leader'
                    order by a.is_active desc, coalesce(v.name,a.volunteer_id)
                """)
            rows = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>负责人权限</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>
<body>
<div class="page">

<div class="card">
    <h1 class="page-title">👥 负责人权限</h1>
    <p class="page-subtitle">
        当前身份：{{ current_role_label }}。
        {% if current_role == 'developer' %}
            你可以建立主负责人和小组长。
        {% else %}
            主负责人只管理小组长。
        {% endif %}
    </p>

    {% if error %}<div class="alert alert-danger">{{ error }}</div>{% endif %}
    {% if success %}<div class="alert alert-success">{{ success }}</div>{% endif %}

    <a class="btn-tool btn-secondary" href="/library/admin/home">← 返回负责人专区</a>
</div>

<div class="card">
    <h2 class="section-title">➕ 新增负责人</h2>

    <form method="post">
        <input type="hidden" name="is_active" value="1">

        <div class="form-group">
            <label class="form-label">义工编号</label>
            <input class="form-input" name="volunteer_id" placeholder="CHE-208 / STW-160" required>
        </div>

        <div class="form-group">
            <label class="form-label">Email</label>
            <input class="form-input" type="email" name="email" required>
        </div>

        <div class="form-group">
            <label class="form-label">个人 PIN</label>
            <input class="form-input" type="password" name="pin" minlength="4" required>
        </div>

        <div class="form-group">
            <label class="form-label">身份级别</label>
            <select class="form-input" name="role" required>
                {% for value,label in allowed_role_options %}
                <option value="{{ value }}">{{ label }}</option>
                {% endfor %}
            </select>
        </div>

        <h3 class="section-title" style="margin-top:22px;">功能权限</h3>
        <div class="summary-grid">
            {% for p,label in permission_labels %}
            <label class="summary-box" style="text-align:left;">
                <input type="checkbox" name="{{ p }}" value="1"> {{ label }}
            </label>
            {% endfor %}
        </div>

        <div class="btn-row" style="margin-top:18px;">
            <button class="btn-tool btn-success" type="submit">保存负责人</button>
        </div>
    </form>
</div>

{% for u in rows %}
{% set urole = u.role or ('developer' if u.is_super_admin else 'team_leader') %}

<div class="card">
    <h2 class="section-title">
        {{ u.volunteer_name or u.volunteer_id }}
        {% if urole == 'developer' %} 🛠️
        {% elif urole == 'main_admin' %} ⭐
        {% else %} 👤
        {% endif %}
    </h2>

    <p>
        {{ u.volunteer_id }} · {{ u.email }} ·
        <b>{{ role_labels.get(urole, '小组长') }}</b> ·
        {% if u.is_active %}启用{% else %}已停用{% endif %}
    </p>

    {% if urole == 'developer' %}
        <div class="alert alert-info">
            🛠️ 系统开发者为最高级账号，拥有全部权限；不能在这里被停用或降级。
        </div>
    {% else %}
        <form method="post">
            <input type="hidden" name="user_id" value="{{ u.id }}">
            <input type="hidden" name="volunteer_id" value="{{ u.volunteer_id }}">

            <div class="form-group">
                <label class="form-label">Email</label>
                <input class="form-input" type="email" name="email" value="{{ u.email }}" required>
            </div>

            <div class="form-group">
                <label class="form-label">新 PIN（不修改可留空）</label>
                <input class="form-input" type="password" name="pin">
            </div>

            <div class="form-group">
                <label class="form-label">身份级别</label>
                <select class="form-input" name="role">
                    {% for value,label in allowed_role_options %}
                    <option value="{{ value }}" {% if urole == value %}selected{% endif %}>{{ label }}</option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label>
                    <input type="checkbox" name="is_active" value="1"
                           {% if u.is_active %}checked{% endif %}>
                    启用账号
                </label>
            </div>

            <div class="summary-grid">
                {% for p,label in permission_labels %}
                <label class="summary-box" style="text-align:left;">
                    <input type="checkbox" name="{{ p }}" value="1"
                           {% if u[p] %}checked{% endif %}>
                    {{ label }}
                </label>
                {% endfor %}
            </div>

            <div class="btn-row" style="margin-top:18px;">
                <button class="btn-tool btn-warning" type="submit">保存修改</button>
            </div>
        </form>
    {% endif %}
</div>
{% endfor %}

</div>
</body>
</html>
""",
        rows=rows,
        error=error,
        success=success,
        current_role=current_role,
        current_role_label=library_role_label(admin),
        role_labels=LIBRARY_ROLE_LABELS,
        allowed_role_options=[(r, LIBRARY_ROLE_LABELS[r]) for r in allowed_roles],
        permission_labels=[
            ("can_batch_in","批量/单项入库"),
            ("can_batch_out","批量出库"),
            ("can_view_records","查看进出记录"),
            ("can_adjust_stock","调整库存"),
            ("can_cancel_transaction","撤销记录"),
            ("can_manage_items","法物管理"),
            ("can_stocktake","库存盘点"),
            ("can_view_audit_log","查看操作日志"),
            ("can_manage_permissions","管理负责人权限"),
        ],
    )


@library_bp.route("/admin/audit")
def library_admin_audit():
    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_view_audit_log"):
        return library_permission_denied()

    keyword = request.args.get("q", "").strip()
    action = request.args.get("action", "").strip()
    like = f"%{keyword}%"
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select *, to_char(created_at at time zone 'Asia/Kuala_Lumpur','DD/MM/YYYY HH24:MI:SS') as created_at_my
                from library_audit_logs
                where (%s='' or action=%s)
                  and (%s='' or coalesce(volunteer_name,'') ilike %s or coalesce(user_email,'') ilike %s
                       or coalesce(item_name,'') ilike %s or coalesce(item_code,'') ilike %s or coalesce(detail,'') ilike %s)
                order by created_at desc limit 500
            """, (action, action, keyword, like, like, like, like, like))
            rows = cur.fetchall()

    return render_template_string("""
<!doctype html><html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>操作日志</title><link rel="stylesheet" href="/static/css/toolbox.css"></head><body><div class="page">
<div class="card"><h1 class="page-title">🧾 操作日志</h1><p class="page-subtitle">负责人做过的入库、调整、撤销、资料修改和权限修改都会留下记录。</p>
<form method="get"><div class="form-group"><label class="form-label">搜索</label><input class="form-input" name="q" value="{{ q }}" placeholder="姓名 / Email / 法物 / 编号 / 备注"></div>
<div class="form-group"><label class="form-label">操作类型</label><select class="form-input" name="action"><option value="">全部</option>{% for a in actions %}<option value="{{ a }}" {% if a==action %}selected{% endif %}>{{ a }}</option>{% endfor %}</select></div>
<div class="btn-row"><button class="btn-tool btn-primary">搜索</button><a class="btn-tool btn-secondary" href="/library/admin/home">返回负责人专区</a></div></form></div>
<div class="card"><div class="table-responsive"><table class="record-table"><thead><tr><th>时间</th><th>负责人</th><th>操作</th><th>法物</th><th>变化</th><th>库存</th><th>详情</th></tr></thead><tbody>
{% for r in rows %}<tr><td>{{ r.created_at_my }}</td><td>{{ r.volunteer_name or r.volunteer_id or '-' }}<br><small>{{ r.user_email or '' }}</small></td><td>{{ r.action }}</td>
<td>{{ r.item_name or '-' }}{% if r.item_code %}<br><small>{{ r.item_code }}</small>{% endif %}</td><td>{% if r.quantity_change is not none %}{{ '%+d'|format(r.quantity_change) }}{% else %}-{% endif %}</td>
<td>{% if r.balance_before is not none %}{{ r.balance_before }} → {{ r.balance_after }}{% else %}-{% endif %}</td><td>{{ r.detail or '-' }}</td></tr>{% endfor %}
</tbody></table></div>{% if not rows %}<div class="empty-state">没有操作日志</div>{% endif %}</div></div></body></html>
""", rows=rows, q=keyword, action=action, actions=["LOGIN","LOGOUT","ADMIN_BOOTSTRAP","ADMIN_CREATE","ADMIN_EDIT","BATCH_IN","SINGLE_IN","STOCKTAKE_ADJUST","CANCEL_TRANSACTION","ITEM_EDIT"])

@library_bp.route("/items")
def library_items():

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_manage_items"):
        return library_permission_denied()

    keyword = request.args.get("q", "").strip()
    items = []

    if keyword:
        like = f"%{keyword}%"
        like_no_space = f"%{keyword.replace(' ', '')}%"

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    select
                        item_code,
                        name,
                        category,
                        location,
                        unit,
                        balance,
                        min_balance
                    from library_items
                    where is_active = true
                    and (
                        name ilike %s
                        or replace(name,' ','') ilike %s
                        or category ilike %s
                        or keywords ilike %s
                        or replace(coalesce(keywords,''),' ','') ilike %s
                        or item_code ilike %s
                        or location ilike %s
                    )
                    order by category, name
                    limit 100
                """, (
                    like,
                    like_no_space,
                    like,
                    like,
                    like_no_space,
                    like,
                    like
                ))

                items = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>法物管理</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">
<div class="card">

<h1 class="page-title">⚙️ 法物管理</h1>

<p class="page-subtitle">
修改法物位置、最低库存及说明
</p>

<form method="get">

<div class="form-group">
<label class="form-label">搜索法物</label>
<input
name="q"
class="form-input"
placeholder="例如：白话、心经、A-01"
value="{{ request.args.get('q','') }}">
</div>

<div class="btn-row">
<button class="btn-tool btn-primary" type="submit">
搜索
</button>

<a class="btn-tool btn-secondary" href="/library/admin/home">
返回负责人专区
</a>
</div>

</form>

{% for item in items %}

<div class="card" style="margin-top:18px;">

<h3>{{ item.name }}</h3>

<p>编号：{{ item.item_code }}</p>
<p>分类：{{ item.category or "-" }}</p>
<p>位置：{{ item.location or "-" }}</p>
<p>库存：{{ item.balance }} {{ item.unit }}</p>
<p>最低库存：{{ item.min_balance }}</p>

<a
class="btn-tool btn-warning"
href="/library/items/edit/{{ item.item_code }}">
编辑
</a>

</div>

{% endfor %}

{% if keyword and not items %}
<div class="empty-state">
找不到相关法物
</div>
{% endif %}

</div>
</div>
</body>
</html>
""",
items=items,
keyword=keyword
)

@library_bp.route("/records")
def library_records():

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_view_records"):
        return library_permission_denied()

    keyword = request.args.get("q", "").strip()

    records = []

    like = f"%{keyword}%"

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            if keyword:
                cur.execute("""
                    select
                        id,
                        item_code,
                        item_name,
                        transaction_type,
                        quantity,
                        status,
                        volunteer_id,
                        volunteer_name,
                        handled_by,
                        remark,
                        to_char(
                            created_at at time zone 'Asia/Kuala_Lumpur',
                            'YYYY-MM-DD HH24:MI'
                        ) as created_at_my
                    from library_transactions
                    where
                        item_name ilike %s
                        or item_code ilike %s
                        or volunteer_id ilike %s
                        or volunteer_name ilike %s
                        or handled_by ilike %s
                        or remark ilike %s
                    order by created_at desc
                    limit 200
                """, (
                    like,
                    like,
                    like,
                    like,
                    like,
                    like
                ))
            else:
                cur.execute("""
                    select
                        id,
                        item_code,
                        item_name,
                        transaction_type,
                        quantity,
                        status,
                        volunteer_id,
                        volunteer_name,
                        handled_by,
                        remark,
                        to_char(
                            created_at at time zone 'Asia/Kuala_Lumpur',
                            'YYYY-MM-DD HH24:MI'
                        ) as created_at_my
                    from library_transactions
                    order by created_at desc
                    limit 200
                """)

            records = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>进出记录</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">
<div class="card">

<h1 class="page-title">📋 进出记录</h1>

<p class="page-subtitle">
查看最近的法物入库、出库及调整记录
</p>

<form method="get">

<div class="form-group">
<label class="form-label">搜索记录</label>

<input
name="q"
class="form-input"
placeholder="例如：白话、CHE-208、姓名"
value="{{ request.args.get('q','') }}">
</div>

<div class="btn-row">

<button class="btn-tool btn-primary" type="submit">
搜索
</button>

<a class="btn-tool btn-secondary" href="/library/admin/home">
返回负责人专区
</a>

</div>

</form>

{% if records %}

<div class="table-responsive" style="margin-top:20px;">

<table class="record-table">

<thead>
<tr>
    <th>时间</th>
    <th>类型</th>
    <th>法物</th>
    <th>数量</th>
    <th>领取人</th>
    <th>处理人</th>
    <th>备注</th>
    <th>操作</th>
</tr>
</thead>

<tbody>

{% for r in records %}

<tr
{% if r.status == "cancelled" %}
style="color:#999;text-decoration:line-through;background:#f8f8f8;"
{% endif %}
>
    <td>{{ r.created_at_my }}</td>

    <td>

        {% if r.transaction_type == "out" %}
            📤 出库
        {% elif r.transaction_type == "in" %}
            📥 入库
        {% elif r.transaction_type == "adjust" %}
            🔧 调整
        {% else %}
            {{ r.transaction_type }}
        {% endif %}

        {% if r.status == "cancelled" %}
            <br>
            <small style="color:#999;">
                ❌ 已撤销
            </small>
        {% endif %}

    </td>

    <td>
        <b>{{ r.item_name }}</b><br>
        <small>{{ r.item_code }}</small>
    </td>

    <td>
        {{ r.quantity }}
    </td>

    <td>
        {{ r.volunteer_id or "-" }}<br>
        {{ r.volunteer_name or "" }}
    </td>

    <td>
        {{ r.handled_by or "-" }}
    </td>

    <td>
        {{ r.remark or "-" }}
    </td>

    <td>

    {% if r.status=="active" %}

    <form
    method="post"
    action="/library/records/cancel/{{ r.id }}"
    onsubmit="return confirm('确定撤销这笔记录？');">

    <button
    class="btn-tool btn-danger"
    style="padding:6px 12px;font-size:14px;">

    撤销

    </button>

    </form>

    {% else %}

    <span style="color:#999;">
    已撤销
    </span>

    {% endif %}

    </td>
</tr>

{% endfor %}

</tbody>

</table>

</div>

{% else %}

<div class="empty-state" style="margin-top:20px;">
暂无进出记录
</div>

{% endif %}

</div>
</div>
</body>
</html>
""",
records=records
)

@library_bp.route("/low-stock")
def library_low_stock():

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select
                    item_code,
                    name,
                    category,
                    location,
                    unit,
                    balance,
                    min_balance
                from library_items
                where is_active = true
                  and balance <= min_balance
                order by
                    category,
                    name
            """)

            items = cur.fetchall()

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>库存不足</title>
        <link rel="stylesheet" href="/static/css/toolbox.css">
    </head>

    <body>
    <div class="page">

        <div class="card">
            <h1 class="page-title">📊 库存不足</h1>
            <p class="page-subtitle">
                以下法物库存已低于或等于最低库存，需要留意补货。
            </p>

            <div class="btn-row">
                <a class="btn-tool btn-secondary" href="{{ url_for('library.library_admin_home') }}">
                    ← 返回负责人专区
                </a>
            </div>
        </div>

        <div class="card">
            <h2 class="section-title">库存不足法物</h2>

            {% if not items %}
                <div class="empty-state">
                    ✅ 目前没有库存不足的法物。
                </div>
            {% else %}

                <div class="summary-grid">
                    <div class="summary-box">
                        <div class="summary-title">不足项目</div>
                        <div class="summary-value">{{ items|length }}</div>
                    </div>
                </div>

                <div class="table-responsive">
                    <table class="record-table">
                        <thead>
                            <tr>
                                <th>编号</th>
                                <th>名称</th>
                                <th>分类</th>
                                <th>库存</th>
                                <th>最低</th>
                                <th>位置</th>
                                <th>操作</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for item in items %}
                            <tr>
                                <td>{{ item.item_code }}</td>
                                <td>
                                    <b>{{ item.name }}</b>
                                </td>
                                <td>{{ item.category or "-" }}</td>
                                <td>
                                    <b style="color:#dc3545;">
                                        {{ item.balance }}
                                    </b>
                                    {{ item.unit or "" }}
                                </td>
                                <td>
                                    {{ item.min_balance }}
                                    {{ item.unit or "" }}
                                </td>
                                <td>{{ item.location or "-" }}</td>
                                <td>
                                    <a class="btn-tool btn-primary"
                                       href="{{ url_for('library.library_item_detail', item_code=item.item_code) }}">
                                        详情
                                    </a>
                                    <a class="btn-tool btn-success"
                                       href="{{ url_for('library.library_in_item', item_code=item.item_code) }}">
                                        入库
                                    </a>
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>

            {% endif %}
        </div>

    </div>
    </body>
    </html>
    """, items=items)

@library_bp.route("/in")
def library_in():

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_batch_in"):
        return library_permission_denied()

    keyword = request.args.get("q", "").strip()

    items = []

    if keyword:

        like = f"%{keyword}%"
        like_no_space = f"%{keyword.replace(' ', '')}%"

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                cur.execute("""
                    select
                        item_code,
                        name,
                        category,
                        location,
                        unit,
                        balance
                    from library_items
                    where is_active = true
                    and (
                        name ilike %s
                        or replace(name,' ','') ilike %s
                        or category ilike %s
                        or keywords ilike %s
                        or replace(coalesce(keywords,''),' ','') ilike %s
                        or item_code ilike %s
                    )
                    order by
                        category,
                        name
                    limit 100
                """,(
                    like,
                    like_no_space,
                    like,
                    like,
                    like_no_space,
                    like
                ))

                items = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>

<meta charset="utf-8">
<meta name="viewport"
content="width=device-width,initial-scale=1">

<link rel="stylesheet"
href="/static/css/toolbox.css">

<title>法物入库</title>

</head>

<body>

<div class="page">

<div class="card">

<h1 class="page-title">
📥 法物入库
</h1>

<p class="page-subtitle">
搜索需要增加库存的法物
</p>

<form method="get">

<div class="form-group">

<label class="form-label">
搜索法物
</label>

<input
name="q"
class="form-input"
placeholder="例如：白话、心经、DVD"
value="{{ request.args.get('q','') }}">

</div>

<div class="btn-row">

<button
class="btn-tool btn-primary">

搜索

</button>

<a
class="btn-tool btn-secondary"
href="/library/admin/home">

返回负责人专区

</a>

</div>

</form>

{% for item in items %}

<div class="card"
style="margin-top:18px;">

<h3>

{{ item.name }}

</h3>

<p>

分类：
{{ item.category }}

</p>

<p>

库存：

<b>

{{ item.balance }}

{{ item.unit }}

</b>

</p>

<p>

位置：

{{ item.location or "-" }}

</p>

<a
class="btn-tool btn-success"
href="/library/in/{{ item.item_code }}">

📥 入库

</a>

</div>

{% endfor %}

{% if keyword and not items %}

<div class="empty-state"
style="margin-top:20px;">

找不到相关法物

</div>

{% endif %}

</div>

</div>

</body>

</html>

""",
items=items,
keyword=keyword
)

@library_bp.route("/in/<item_code>", methods=["GET", "POST"])
def library_in_item(item_code):

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_batch_in"):
        return library_permission_denied()

    success = None
    error = None
    admin = get_library_admin()
    handled_by = library_admin_label(admin)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    item_code,
                    name,
                    category,
                    location,
                    unit,
                    balance
                from library_items
                where item_code=%s
                limit 1
            """, (item_code,))

            item = cur.fetchone()

            if not item:
                return "找不到法物", 404

            if request.method == "POST":

                try:
                    qty = int(request.form.get("qty", "1"))
                except:
                    qty = 1

                if qty <= 0:
                    error = "数量必须大于 0"

                else:

                    remark = request.form.get("remark", "").strip()

                    cur.execute("""
                        update library_items
                        set balance = balance + %s
                        where item_code=%s
                    """, (
                        qty,
                        item_code
                    ))

                    cur.execute("""
                        insert into library_transactions(

                            item_code,
                            item_name,
                            transaction_type,
                            quantity,
                            handled_by,
                            remark

                        )
                        values(

                            %s,
                            %s,
                            'in',
                            %s,
                            %s,
                            %s

                        )
                    """, (

                        item["item_code"],
                        item["name"],
                        qty,
                        handled_by,
                        remark

                    ))

                    before = int(item["balance"] or 0)
                    after = before + qty
                    write_library_audit(
                        cur, "SINGLE_IN", item_code=item["item_code"], item_name=item["name"],
                        quantity_change=qty, balance_before=before, balance_after=after,
                        detail=remark or "单项入库", admin=admin
                    )
                    conn.commit()

                    item["balance"] = after

                    success = "入库成功！"

    return render_template_string("""

<!doctype html>

<html lang="zh">

<head>

<meta charset="utf-8">

<meta name="viewport"
content="width=device-width,initial-scale=1">

<link
rel="stylesheet"
href="/static/css/toolbox.css">

<title>法物入库</title>

</head>

<body>

<div class="page">

<div class="card">

<h1 class="page-title">

📥 法物入库

</h1>

<p class="page-subtitle">

{{ item.name }}

</p>

{% if error %}

<div class="alert alert-danger">

{{ error }}

</div>

{% endif %}

{% if success %}

<div class="alert alert-success">

{{ success }}

</div>

{% endif %}

<div class="summary-grid">

<div class="summary-box">

<div class="summary-title">

目前库存

</div>

<div class="summary-value">

{{ item.balance }}

{{ item.unit }}

</div>

</div>

<div class="summary-box">

<div class="summary-title">

位置

</div>

<div class="summary-value">

{{ item.location or "-" }}

</div>

</div>

</div>

<form method="post">

<div class="form-group">

<label class="form-label">

增加数量

</label>

<input
type="number"
name="qty"
value="1"
min="1"
class="form-input"
required>

</div>

<div class="form-group">

<label class="form-label">

备注（可选）

</label>

<input
name="remark"
class="form-input"
placeholder="例如：新到货 / 委员捐赠">

</div>

<div class="btn-row">

<button
class="btn-tool btn-success">

📥 确认入库

</button>

<a
class="btn-tool btn-secondary"
href="/library/in">

返回搜索

</a>

</div>

</form>

</div>

</div>

</body>

</html>

""",
item=item,
success=success,
error=error
)

@library_bp.route("/items/edit/<item_code>", methods=["GET", "POST"])
def library_item_edit(item_code):

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_manage_items"):
        return library_permission_denied()

    error = None
    success = None
    admin = get_library_admin()

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("select * from library_items where item_code=%s limit 1", (item_code,))
            before_item = cur.fetchone()

            if request.method == "POST":

                name = request.form.get("name", "").strip()
                category = request.form.get("category", "").strip()
                location = request.form.get("location", "").strip()
                unit = request.form.get("unit", "本").strip()
                min_balance = request.form.get("min_balance", "0").strip()
                description = request.form.get("description", "").strip()

                if not name:
                    error = "名称不能为空"
                else:
                    try:
                        min_balance = int(min_balance)
                    except:
                        min_balance = 0

                    cur.execute("""
                        update library_items
                        set
                            name=%s,
                            category=%s,
                            location=%s,
                            unit=%s,
                            min_balance=%s,
                            description=%s
                        where item_code=%s
                    """, (
                        name,
                        category,
                        location,
                        unit,
                        min_balance,
                        description,
                        item_code
                    ))

                    write_library_audit(
                        cur, "ITEM_EDIT", item_code=item_code, item_name=name,
                        detail="修改法物资料",
                        metadata={
                            "before": dict(before_item) if before_item else {},
                            "after": {"name": name, "category": category, "location": location,
                                      "unit": unit, "min_balance": min_balance, "description": description}
                        }, admin=admin
                    )
                    conn.commit()
                    success = "法物资料已保存"

            cur.execute("""
                select
                    item_code,
                    name,
                    category,
                    keywords,
                    location,
                    unit,
                    balance,
                    min_balance,
                    description
                from library_items
                where item_code=%s
                limit 1
            """, (item_code,))

            item = cur.fetchone()

    if not item:
        return "找不到这个法物", 404

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>编辑法物</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">
<div class="card">

<h1 class="page-title">✏️ 编辑法物</h1>

<p class="page-subtitle">
{{ item.item_code }}
</p>

{% if error %}
<div class="alert alert-danger">{{ error }}</div>
{% endif %}

{% if success %}
<div class="alert alert-success">{{ success }}</div>
{% endif %}

<form method="post">

<div class="form-group">
<label class="form-label">名称</label>
<input name="name" class="form-input" value="{{ item.name }}" required>
</div>

<div class="form-group">
<label class="form-label">分类</label>
<input name="category" class="form-input" value="{{ item.category or '' }}">
</div>

<div class="form-group">
<label class="form-label">位置</label>
<input name="location" class="form-input" placeholder="例如：A-01" value="{{ item.location or '' }}">
</div>

<div class="form-group">
<label class="form-label">单位</label>
<input name="unit" class="form-input" value="{{ item.unit or '本' }}">
</div>

<div class="form-group">
<label class="form-label">最低库存</label>
<input type="number" name="min_balance" class="form-input" value="{{ item.min_balance or 0 }}">
</div>

<div class="form-group">
<label class="form-label">说明</label>
<textarea name="description" class="form-input" rows="4">{{ item.description or '' }}</textarea>
</div>

<div class="card" style="margin-top:18px;">
<p>当前库存：<b>{{ item.balance }} {{ item.unit }}</b></p>
<p>关键词：{{ item.keywords or "-" }}</p>
</div>

<div class="btn-row" style="margin-top:25px;">

<button class="btn-tool btn-success" type="submit">
保存
</button>

<a class="btn-tool btn-secondary" href="/library/items">
返回法物管理
</a>

<a class="btn-tool btn-primary" href="/library/item/{{ item.item_code }}">
查看详情
</a>

</div>

</form>

</div>
</div>
</body>
</html>
""",
item=item,
error=error,
success=success
)

@library_bp.route("/admin/logout")
def library_admin_logout():
    admin = get_library_admin()
    if admin:
        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                write_library_audit(cur, "LOGOUT", detail="退出负责人专区", admin=admin)
                conn.commit()
    session.pop("library_admin_id", None)
    session.modified = True
    return redirect(url_for("library.library_home"))

@library_bp.route("/out/search")
def library_out_search():

    volunteer = get_library_out_volunteer()

    if not volunteer:
        return redirect(url_for("library.library_out"))

    keyword = request.args.get("q", "").strip()
    cart = session.get("library_cart", [])

    total_qty = sum(
        int(c.get("qty", 0))
        for c in cart
    )

    print("SEARCH PAGE CART =", cart)

    items = []

    if keyword:

        like = f"%{keyword}%"
        like_no_space = f"%{keyword.replace(' ', '')}%"

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                cur.execute("""
                    select
                        item_code,
                        name,
                        category,
                        location,
                        unit,
                        balance,
                        min_balance
                    from library_items
                    where is_active = true
                    and (
                        name ilike %s
                        or replace(name,' ','') ilike %s
                        or category ilike %s
                        or keywords ilike %s
                        or replace(coalesce(keywords,''),' ','') ilike %s
                        or item_code ilike %s
                    )
                    order by category, name
                    limit 100
                """, (
                    like,
                    like_no_space,
                    like,
                    like,
                    like_no_space,
                    like
                ))

                items = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="stylesheet" href="/static/css/toolbox.css">
<title>领取法物</title>
</head>

<body>

<div class="page">

<div class="card">

<h1 class="page-title">
📤 领取法物
</h1>

<p class="page-subtitle">

{{ volunteer.id }}
　
{{ volunteer.name }}

</p>

{% with messages = get_flashed_messages(with_categories=true) %}
{% if messages %}

{% for category, message in messages %}

<div class="alert alert-danger">
{{ message }}
</div>

{% endfor %}

{% endif %}
{% endwith %}

{% if cart %}

<div class="card" style="margin-top:18px;">

<h2>🛒 本次领取</h2>

<p>
共 {{ cart|length }} 种，
共 {{ total_qty }} {{ cart[0].unit if cart else "" }}
</p>

{% for c in cart %}

<div class="card" style="margin-top:14px;">

<h3>{{ c.name }}</h3>

<p>
位置：{{ c.location or "-" }}
</p>

<form method="post" action="/library/out/cart/update">

<input type="hidden" name="item_code" value="{{ c.item_code }}">
<input type="hidden" name="q" value="{{ request.args.get('q','') }}">

<div style="display:flex;gap:10px;align-items:center;">

<input
type="number"
name="qty"
value="{{ c.qty }}"
min="1"
class="form-input"
style="width:100px;">

<span>{{ c.unit }}</span>

<button class="btn-tool btn-primary" type="submit">
更新
</button>

</div>

</form>

<form method="post" action="/library/out/cart/delete" style="margin-top:10px;">

<input type="hidden" name="item_code" value="{{ c.item_code }}">
<input type="hidden" name="q" value="{{ request.args.get('q','') }}">

<button class="btn-tool btn-danger" type="submit">
删除
</button>

</form>

</div>

{% endfor %}
                                  
<form
method="post"
action="/library/out/confirm"
style="margin-top:20px;"
onsubmit="return confirm(
'确认领取以下法物？\\n\\n确认后库存会立即扣减。'
);">

</div>

{% endif %}

<form method="get">

<div class="form-group">
<label class="form-label">搜索法物</label>

<input
name="q"
class="form-input"
placeholder="例如：白话、心经、DVD"
value="{{ request.args.get('q','') }}">
</div>

<div class="btn-row">

<button class="btn-tool btn-primary" type="submit">
搜索
</button>

<a href="/library/out/logout" class="btn-tool btn-secondary">
退出
</a>

</div>

</form>

{% for item in items %}

<div class="card" style="margin-top:18px;">

<h3>{{ item.name }}</h3>

<p>分类：{{ item.category }}</p>

{% if item.balance <= 0 %}

<p style="color:#d32f2f;font-weight:bold;">
🔴 已领完
</p>

{% elif item.balance <= item.min_balance %}

<p style="color:#f57c00;font-weight:bold;">
🟡 库存偏低：{{ item.balance }} {{ item.unit }}
</p>

{% else %}

<p style="color:#2e7d32;font-weight:bold;">
🟢 库存：{{ item.balance }} {{ item.unit }}
</p>

{% endif %}

<p>位置：{{ item.location or "-" }}</p>

<form method="post" action="/library/out/add">

<input type="hidden" name="item_code" value="{{ item.item_code }}">
<input type="hidden" name="q" value="{{ request.args.get('q','') }}">

<div style="display:flex;gap:10px;align-items:center;">

<input
type="number"
name="qty"
value="1"
min="1"
class="form-input"
style="width:100px;">

{% if item.balance <= 0 %}

<button
class="btn-tool btn-secondary"
disabled>
已领完
</button>

{% else %}

<button
class="btn-tool btn-success"
type="submit">
加入领取
</button>

{% endif %}

</div>

</form>

</div>

{% endfor %}

</div>

</div>

</body>
</html>
""",
volunteer=volunteer,
items=items,
cart=cart,
total_qty=total_qty
)

@library_bp.route("/out/add", methods=["POST"])
def library_out_add():

    print("=" * 50)
    print("进入 /out/add")
    print("form =", request.form)
    print("session before =", dict(session))

    volunteer = get_library_out_volunteer()

    if not volunteer:
        return redirect(url_for("library.library_out"))

    item_code = request.form.get("item_code", "").strip()
    qty = int(request.form.get("qty", 1))

    if qty <= 0:
        qty = 1

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    item_code,
                    name,
                    category,
                    location,
                    unit,
                    balance,
                    min_balance
                from library_items
                where item_code=%s
                limit 1
            """, (item_code,))

            item = cur.fetchone()

    if not item:
        return redirect(url_for("library.library_out_search"))
    
    if qty > item["balance"]:

        flash(
            f"❌ {item['name']} 库存只有 {item['balance']} {item['unit']}",
            "danger"
        )

        return redirect(
            url_for(
                "library.library_out_search",
                q=request.form.get("q", "")
            )
        )

    cart = get_library_cart()

    found = False

    for c in cart:

        if c["item_code"] == item_code:

            new_qty = c["qty"] + qty

            if new_qty > item["balance"]:

                flash(
                    f"❌ {item['name']} 库存只有 {item['balance']} {item['unit']}",
                    "danger"
                )

                return redirect(
                    url_for(
                        "library.library_out_search",
                        q=request.form.get("q", "")
                    )
                )

            c["qty"] = new_qty
            found = True
            break

    if not found:

        cart.append({

            "item_code": item["item_code"],
            "name": item["name"],
            "category": item["category"],
            "location": item["location"],
            "unit": item["unit"],
            "balance": item["balance"],
            "qty": qty

        })

    session["library_cart"] = cart
    session.modified = True

    print("cart after =", session.get("library_cart"))
    print("session after =", dict(session))
    print("=" * 50)

    return redirect(
        url_for(
            "library.library_out_search",
            q=request.form.get("q", "")
        )
    )

@library_bp.route("/out/logout")
def library_out_logout():

    session.pop("library_out_volunteer", None)
    session.pop("library_cart", None)

    return redirect(url_for("library.library_out"))

@library_bp.route("/out/cart/update", methods=["POST"])
def library_out_cart_update():

    volunteer = get_library_out_volunteer()

    if not volunteer:
        return redirect(url_for("library.library_out"))

    item_code = request.form.get("item_code", "").strip()
    qty = int(request.form.get("qty", 1))
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    balance,
                    name,
                    unit
                from library_items
                where item_code=%s
                limit 1
            """, (item_code,))

            item = cur.fetchone()

    if item and qty > item["balance"]:

        flash(
            f"❌ {item['name']} 库存只有 {item['balance']} {item['unit']}",
            "danger"
        )

        return redirect(
            url_for(
                "library.library_out_search",
                q=q
            )
        )
    q = request.form.get("q", "")

    cart = session.get("library_cart", [])

    new_cart = []

    for c in cart:
        if c["item_code"] == item_code:
            if qty > 0:
                c["qty"] = qty
                new_cart.append(c)
        else:
            new_cart.append(c)

    session["library_cart"] = new_cart
    session.modified = True

    return redirect(url_for("library.library_out_search", q=q))


@library_bp.route("/out/cart/delete", methods=["POST"])
def library_out_cart_delete():

    volunteer = get_library_out_volunteer()

    if not volunteer:
        return redirect(url_for("library.library_out"))

    item_code = request.form.get("item_code", "").strip()
    q = request.form.get("q", "")

    cart = session.get("library_cart", [])

    cart = [
        c for c in cart
        if c["item_code"] != item_code
    ]

    session["library_cart"] = cart
    session.modified = True

    return redirect(url_for("library.library_out_search", q=q))

@library_bp.route("/out/confirm", methods=["POST"])
def library_out_confirm():

    volunteer = get_library_out_volunteer()

    if not volunteer:
        return redirect(url_for("library.library_out"))

    cart = session.get("library_cart", [])

    if not cart:
        return redirect(url_for("library.library_out_search"))

    error = None
    success_items = []

    try:
        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                # 先检查每一样库存
                for c in cart:

                    cur.execute("""
                        select
                            item_code,
                            name,
                            balance
                        from library_items
                        where item_code=%s
                        for update
                    """, (c["item_code"],))

                    item = cur.fetchone()

                    if not item:
                        error = f"找不到法物：{c['name']}"
                        raise Exception(error)

                    if item["balance"] < c["qty"]:
                        error = (
                            f"库存不足：{item['name']}，"
                            f"目前只有 {item['balance']}，"
                            f"需要 {c['qty']}"
                        )
                        raise Exception(error)

                # 全部库存足够，才正式扣库存
                for c in cart:

                    cur.execute("""
                        update library_items
                        set balance = balance - %s
                        where item_code=%s
                    """, (
                        c["qty"],
                        c["item_code"]
                    ))

                    cur.execute("""
                        insert into library_transactions
                        (
                            item_code,
                            item_name,
                            transaction_type,
                            quantity,
                            volunteer_id,
                            volunteer_name,
                            handled_by,
                            remark
                        )
                        values
                        (%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        c["item_code"],
                        c["name"],
                        "out",
                        c["qty"],
                        volunteer["id"],
                        volunteer["name"],
                        volunteer["name"],
                        "义工领取"
                    ))

                    success_items.append(c)

            conn.commit()

    except Exception as e:

        print("library_out_confirm error:", e)

        return render_template_string("""
<link rel="stylesheet" href="/static/css/toolbox.css">

<div class="page">
<div class="card">

<h1 class="page-title">❌ 领取失败</h1>

<div class="alert alert-danger">
{{ error or "系统错误，请稍后再试" }}
</div>

<a class="btn-tool btn-secondary" href="/library/out/search">
返回领取页面
</a>

</div>
</div>
""", error=error)
    
    session["library_last_receipt"] = {
        "volunteer": volunteer,
        "items": success_items
    }

    session.pop("library_cart", None)
    session.modified = True

    return redirect(
        url_for("library.library_out_success")
    )


@library_bp.route("/out/success")
def library_out_success():

    now = malaysia_now()

    volunteer = get_library_out_volunteer()
    receipt = session.get("library_last_receipt")

    if not volunteer or not receipt:
        return redirect(url_for("library.library_out"))

    success_items = receipt.get("items", [])

    return render_template_string("""
<link rel="stylesheet" href="/static/css/toolbox.css">

<div class="page">
<div class="card">

<h1 class="page-title">✅ 领取成功</h1>

<p class="page-subtitle">
{{ volunteer.id }}　{{ volunteer.name }}
</p>

<h2>本次领取</h2>

{% for c in success_items %}

<div class="card" style="margin-top:12px;">
<h3>{{ c.name }}</h3>
<p>数量：{{ c.qty }} {{ c.unit }}</p>
</div>

{% endfor %}

<div class="alert alert-success" style="margin-top:25px;">
谢谢发心弘法 🙏
</div>

<div class="btn-row">

<a class="btn-tool btn-success" href="/library/out/search">
继续领取
</a>

<a class="btn-tool btn-secondary" href="/library">
返回首页
</a>

</div>

</div>
</div>
""",
volunteer=volunteer,
success_items=success_items
)

@library_bp.route("/stocktake", methods=["GET", "POST"])
def library_stocktake():

    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not (library_admin_can("can_stocktake") or library_admin_can("can_adjust_stock")):
        return library_permission_denied()

    admin = get_library_admin()
    handled_by = library_admin_label(admin)

    preview_rows = []
    error = None

    if request.method == "POST":

        action = request.form.get("action")

        # 第二步：确认调整
        if action == "confirm":

            rows = session.get("library_stocktake_preview", [])

            with get_conn() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:

                    for row in rows:

                        diff = int(row["diff"])

                        if diff == 0:
                            continue

                        cur.execute("""
                            update library_items
                            set balance = %s
                            where item_code = %s
                        """, (
                            row["actual_balance"],
                            row["item_code"],
                        ))

                        cur.execute("""
                            insert into library_transactions (
                                item_code,
                                item_name,
                                transaction_type,
                                quantity,
                                handled_by,
                                remark
                            )
                            values (%s, %s, %s, %s, %s, %s)
                        """, (
                            row["item_code"],
                            row["name"],
                            "adjust",
                            diff,
                            handled_by,
                            "盘点调整"
                        ))

                        write_library_audit(
                            cur, "STOCKTAKE_ADJUST",
                            item_code=row["item_code"], item_name=row["name"],
                            quantity_change=diff, balance_before=row["system_balance"],
                            balance_after=row["actual_balance"], detail="盘点调整", admin=admin
                        )

                    conn.commit()

            session.pop("library_stocktake_preview", None)

            return redirect(url_for("library.library_records"))

        # 第一步：上传 Excel 预览
        file = request.files.get("file")

        if not file or file.filename == "":
            error = "请先选择 Excel 文件。"

        else:
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            headers = {}
            header_row = None

            for r in range(1, min(ws.max_row, 10) + 1):
                for c in range(1, ws.max_column + 1):
                    value = ws.cell(r, c).value
                    if not value:
                        continue

                    text = str(value).strip()

                    if text in ["编号", "item_code", "法物编号"]:
                        headers["item_code"] = c
                        header_row = r

                    if text in ["实盘数量", "盘点数量", "数量", "balance"]:
                        headers["actual_balance"] = c
                        header_row = r

                if "item_code" in headers and "actual_balance" in headers:
                    break

            if "item_code" not in headers or "actual_balance" not in headers:
                error = "Excel 需要有：编号 / 实盘数量 两个栏目。"

            else:
                excel_rows = []

                for r in range(header_row + 1, ws.max_row + 1):

                    item_code = ws.cell(r, headers["item_code"]).value
                    actual_balance = ws.cell(r, headers["actual_balance"]).value

                    if not item_code or actual_balance is None:
                        continue

                    try:
                        actual_balance = int(actual_balance)
                    except:
                        continue

                    excel_rows.append({
                        "item_code": str(item_code).strip(),
                        "actual_balance": actual_balance
                    })

                with get_conn() as conn:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:

                        for row in excel_rows:

                            cur.execute("""
                                select
                                    item_code,
                                    name,
                                    balance
                                from library_items
                                where item_code = %s
                                  and is_active = true
                                limit 1
                            """, (row["item_code"],))

                            item = cur.fetchone()

                            if not item:
                                continue

                            diff = row["actual_balance"] - item["balance"]

                            preview_rows.append({
                                "item_code": item["item_code"],
                                "name": item["name"],
                                "system_balance": item["balance"],
                                "actual_balance": row["actual_balance"],
                                "diff": diff
                            })

                session["library_stocktake_preview"] = preview_rows

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>上传盘点 Excel</title>
        <link rel="stylesheet" href="/static/css/toolbox.css">
    </head>

    <body>
    <div class="page">

        <div class="card">
            <h1 class="page-title">📤 上传盘点 Excel</h1>
            <p class="page-subtitle">
                上传实际盘点数量，系统会先显示差异，确认后才调整库存。
            </p>

            <div class="btn-row">
                <a class="btn-tool btn-secondary" href="{{ url_for('library.library_admin_home') }}">
                    ← 返回负责人专区
                </a>
            </div>
        </div>

        {% if error %}
        <div class="card">
            <div class="alert alert-danger">{{ error }}</div>
        </div>
        {% endif %}

        <div class="card">
            <h2 class="section-title">上传 Excel</h2>

            <form method="post" enctype="multipart/form-data">
                <input type="hidden" name="action" value="preview">

                <div class="form-group">
                    <label class="form-label">盘点 Excel</label>
                    <input class="form-input" type="file" name="file" accept=".xlsx" required>
                </div>

                <button class="btn-tool btn-primary" type="submit">
                    上传并预览
                </button>
            </form>
        </div>

        {% if preview_rows %}
        <div class="card">
            <h2 class="section-title">盘点差异预览</h2>

            <div class="summary-grid">
                <div class="summary-box">
                    <div class="summary-title">读取项目</div>
                    <div class="summary-value">{{ preview_rows|length }}</div>
                </div>
            </div>
        </div>

        {% for row in preview_rows %}
        <div class="card" style="margin-bottom:16px;">

            <h2 class="section-title">📚 {{ row.name }}</h2>
            <p style="color:#666;">{{ row.item_code }}</p>

            <div class="summary-grid">
                <div class="summary-box">
                    <div class="summary-title">系统库存</div>
                    <div class="summary-value">{{ row.system_balance }}</div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">实盘数量</div>
                    <div class="summary-value">{{ row.actual_balance }}</div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">差异</div>
                    <div class="summary-value">
                        {% if row.diff > 0 %}
                            +{{ row.diff }}
                        {% else %}
                            {{ row.diff }}
                        {% endif %}
                    </div>
                </div>
            </div>

        </div>
        {% endfor %}

        <div class="card">
            <form method="post">
                <input type="hidden" name="action" value="confirm">

                <button class="btn-tool btn-success" type="submit">
                    ✅ 确认调整库存
                </button>
            </form>
        </div>
        {% endif %}

    </div>
    </body>
    </html>
    """, preview_rows=preview_rows, error=error)

@library_bp.route("/quick-out/<item_code>", methods=["GET", "POST"])
def library_quick_out(item_code):

    volunteer = session.get("library_out_volunteer")

    if not volunteer:
        session["library_after_login_url"] = url_for(
            "library.library_quick_out",
            item_code=item_code
        )
        return redirect(url_for("library.library_out"))

    item_code = item_code.strip().upper()

    error = None
    success = None

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    item_code,
                    name,
                    category,
                    unit,
                    balance,
                    location
                from library_items
                where item_code = %s
                  and is_active = true
                limit 1
            """, (item_code,))

            item = cur.fetchone()

            if not item:
                return render_template_string("""
                <!doctype html>
                <html lang="zh">
                <head>
                    <meta charset="utf-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1">
                    <title>找不到法物</title>
                    <link rel="stylesheet" href="/static/css/toolbox.css">
                </head>
                <body>
                <div class="page">
                    <div class="card">
                        <h1 class="page-title">找不到法物</h1>
                        <p class="page-subtitle">这个 QR Code 对应的法物不存在。</p>

                        <div class="btn-row">
                            <a class="btn-tool btn-secondary" href="/library/out">
                                返回登记领取
                            </a>
                        </div>
                    </div>
                </div>
                </body>
                </html>
                """)

            if request.method == "POST":

                qty_text = request.form.get("quantity", "").strip()
                remark = request.form.get("remark", "").strip()

                try:
                    qty = int(qty_text)
                except:
                    qty = 0

                if qty <= 0:
                    error = "数量必须大过 0。"

                elif qty > item["balance"]:
                    error = f"库存不足，目前只有 {item['balance']} {item['unit'] or ''}。"

                else:
                    new_balance = item["balance"] - qty

                    cur.execute("""
                        update library_items
                        set balance = %s
                        where item_code = %s
                    """, (
                        new_balance,
                        item_code
                    ))

                    cur.execute("""
                        insert into library_transactions (
                            item_code,
                            item_name,
                            transaction_type,
                            quantity,
                            volunteer_id,
                            volunteer_name,
                            handled_by,
                            remark
                        )
                        values (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        item["item_code"],
                        item["name"],
                        "out",
                        qty,
                        volunteer.get("id"),
                        volunteer.get("name"),
                        volunteer.get("name"),
                        remark
                    ))

                    conn.commit()

                    success = f"已登记领取：{item['name']} x {qty} {item['unit'] or ''}"

                    item["balance"] = new_balance

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>快速领取</title>
        <link rel="stylesheet" href="/static/css/toolbox.css">
    </head>

    <body>
    <div class="page">

        <div class="card">
            <h1 class="page-title">📷 快速领取</h1>
            <p class="page-subtitle">
                扫 QR Code 后，直接输入数量登记领取。
            </p>

            <div class="btn-row">
                <a class="btn-tool btn-secondary" href="/library/out">
                    返回登记领取
                </a>
            </div>
        </div>

        {% if error %}
        <div class="card">
            <div class="alert alert-danger">
                {{ error }}
            </div>
        </div>
        {% endif %}

        {% if success %}
        <div class="card">
            <div class="alert alert-success">
                {{ success }}
            </div>

            <div class="btn-row">
                <a class="btn-tool btn-primary" href="/library/out">
                    继续领取其它法物
                </a>
            </div>
        </div>
        {% endif %}

        <div class="card">
            <h2 class="section-title">📚 {{ item.name }}</h2>

            <p style="color:#666;">
                {{ item.item_code }}
            </p>

            <div class="summary-grid">
                <div class="summary-box">
                    <div class="summary-title">分类</div>
                    <div class="summary-value" style="font-size:18px;">
                        {{ item.category or "-" }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">当前库存</div>
                    <div class="summary-value">
                        {{ item.balance }}
                        {{ item.unit or "" }}
                    </div>
                </div>
            </div>

            <p style="margin-top:14px;">
                📍 位置：{{ item.location or "未设置" }}
            </p>
        </div>

        <div class="card">
            <h2 class="section-title">登记数量</h2>

            <form method="post">

                <div class="form-group">
                    <label class="form-label">领取数量</label>
                    <input
                        class="form-input"
                        type="number"
                        name="quantity"
                        min="1"
                        value="1"
                        required
                        autofocus
                    >
                </div>

                <div class="form-group">
                    <label class="form-label">备注（可选）</label>
                    <input
                        class="form-input"
                        name="remark"
                        placeholder="例如：共修 / 结缘 / 法会"
                    >
                </div>

                <button class="btn-tool btn-success" type="submit">
                    ✅ 确认领取
                </button>

            </form>
        </div>

    </div>
    </body>
    </html>
    """, item=item, error=error, success=success)

@library_bp.route("/scan/<item_code>")
def library_scan(item_code):

    item_code = item_code.strip().upper()

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    item_code,
                    name,
                    category,
                    balance,
                    unit,
                    location
                from library_items
                where item_code = %s
                  and is_active = true
                limit 1
            """, (item_code,))

            item = cur.fetchone()

    if not item:
        return render_template_string("""
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <link rel="stylesheet" href="/static/css/toolbox.css">
            <title>找不到法物</title>
        </head>

        <body>
        <div class="page">
            <div class="card">
                <h1 class="page-title">📚 找不到法物</h1>
                <p class="page-subtitle">
                    此 QR Code 对应的法物不存在，或已经停用。
                </p>

                <div class="btn-row">
                    <a class="btn-tool btn-secondary" href="/library">
                        返回首页
                    </a>
                </div>
            </div>
        </div>
        </body>
        </html>
        """)

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <link rel="stylesheet" href="/static/css/toolbox.css">
        <title>{{ item.name }}</title>
    </head>

    <body>
    <div class="page">

        <div class="card">
            <h1 class="page-title">📚 {{ item.name }}</h1>
            <p class="page-subtitle">{{ item.item_code }}</p>

            <div class="summary-grid">
                <div class="summary-box">
                    <div class="summary-title">分类</div>
                    <div class="summary-value" style="font-size:18px;">
                        {{ item.category or "-" }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">库存</div>
                    <div class="summary-value">
                        {{ item.balance }} {{ item.unit or "" }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-title">位置</div>
                    <div class="summary-value" style="font-size:18px;">
                        {{ item.location or "未设置" }}
                    </div>
                </div>
            </div>
        </div>

        <div class="card">
            <h2 class="section-title">请选择操作</h2>

            <div class="btn-row">
                <a class="btn-tool btn-danger"
                   href="{{ url_for('library.library_quick_out', item_code=item.item_code) }}">
                    📤 登记领取
                </a>

                <a class="btn-tool btn-success"
                   href="{{ url_for('library.library_in_item', item_code=item.item_code) }}">
                    📥 登记入库
                </a>

                <a class="btn-tool btn-primary"
                   href="{{ url_for('library.library_item_detail', item_code=item.item_code) }}">
                    📄 法物资料
                </a>

                <a class="btn-tool btn-secondary"
                   href="{{ url_for('library.library_search') }}">
                    🔍 搜索法物
                </a>
            </div>
        </div>

    </div>
    </body>
    </html>
    """, item=item)

@library_bp.route("/scan")
def library_scan_camera():

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>扫描法物</title>
<link rel="stylesheet" href="/static/css/toolbox.css">

<style>
video {
    width:100%;
    border-radius:16px;
    background:black;
}

.scan-box {
    margin-top:20px;
}

.tip {
    text-align:center;
    color:#666;
    margin-top:12px;
}

.manual-input {
    font-size:28px;
    text-align:center;
    letter-spacing:1px;
}
</style>
</head>

<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">📷 扫描法物</h1>
        <p class="page-subtitle">
            手机可用相机扫描；电脑可直接使用扫码枪。
        </p>

        <div class="btn-row">
            <a class="btn-tool btn-secondary" href="/library">
                返回首页
            </a>
        </div>
    </div>

    <div class="card scan-box">
        <h2 class="section-title">扫码枪 / 手动输入</h2>

        <p class="page-subtitle">
            请扫描 QR Code，或输入 BOOK 编号。
        </p>

        <input
            id="scanInput"
            class="form-input manual-input"
            placeholder="BOOK0001"
            autofocus
            autocomplete="off"
        >

        <div class="btn-row" style="margin-top:14px;">
            <button class="btn-tool btn-primary" onclick="goScanInput()">
                查询
            </button>
        </div>

        <div class="tip" id="inputTip">
            等待扫码枪输入...
        </div>
    </div>

    <div class="card scan-box">
        <h2 class="section-title">手机相机扫描</h2>

        <video id="video" playsinline autoplay muted></video>

        <div class="tip" id="cameraTip">
            正在尝试开启相机...
        </div>
    </div>

</div>

<script>
const video = document.getElementById("video");
const input = document.getElementById("scanInput");
const cameraTip = document.getElementById("cameraTip");

input.focus();

function extractItemCode(value) {
    value = (value || "").trim();

    if (!value) {
        return "";
    }

    // 如果扫码出来是完整网址
    if (value.includes("/library/scan/")) {
        let parts = value.split("/library/scan/");
        return parts[1].split(/[?#]/)[0].trim().toUpperCase();
    }

    // 如果是 GYTLIB:BOOK0001
    if (value.toUpperCase().startsWith("GYTLIB:")) {
        return value.substring(7).trim().toUpperCase();
    }

    // 如果是 BOOK0001
    return value.toUpperCase();
}

function goToItem(value) {
    let itemCode = extractItemCode(value);

    if (!itemCode) {
        return;
    }

    window.location.href = "/library/scan/" + encodeURIComponent(itemCode);
}

function goScanInput() {
    goToItem(input.value);
}

// 扫码枪通常会自动送 Enter
input.addEventListener("keydown", function(e) {
    if (e.key === "Enter") {
        e.preventDefault();
        goScanInput();
    }
});

// 有些扫码枪不会送 Enter，输入完成后稍等自动跳转
let scanTimer = null;

input.addEventListener("input", function() {
    clearTimeout(scanTimer);

    scanTimer = setTimeout(function() {
        let value = input.value.trim();

        if (
            value.toUpperCase().startsWith("BOOK") ||
            value.toUpperCase().startsWith("GYTLIB:") ||
            value.includes("/library/scan/")
        ) {
            goToItem(value);
        }
    }, 500);
});

async function startCamera() {
    if (!("BarcodeDetector" in window)) {
        cameraTip.innerText = "此浏览器不支持相机扫码，可使用扫码枪或手动输入。";
        return;
    }

    try {
        const detector = new BarcodeDetector({
            formats: ["qr_code"]
        });

        const stream = await navigator.mediaDevices.getUserMedia({
            video: {
                facingMode: "environment"
            }
        });

        video.srcObject = stream;
        video.play();

        cameraTip.innerText = "请把 QR Code 放在画面中央。";

        setInterval(async () => {
            const codes = await detector.detect(video);

            if (codes.length > 0) {
                stream.getTracks().forEach(track => track.stop());

                let value = codes[0].rawValue;
                goToItem(value);
            }
        }, 300);

    } catch (err) {
        cameraTip.innerText = "无法开启相机，可使用扫码枪或手动输入。";
    }
}

startCamera();
</script>

</body>
</html>
""")

@library_bp.route("/records/cancel/<int:id>", methods=["POST"])
def library_cancel_record(id):
    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_cancel_transaction"):
        return library_permission_denied()

    admin = get_library_admin()
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""select id,item_code,item_name,transaction_type,quantity,status
                           from library_transactions where id=%s limit 1""", (id,))
            row = cur.fetchone()
            if not row or row["status"] == "cancelled":
                return redirect(url_for("library.library_records"))

            cur.execute("select balance from library_items where item_code=%s", (row["item_code"],))
            bal_row = cur.fetchone()
            before = int(bal_row["balance"] or 0) if bal_row else None
            qty = int(row["quantity"])
            delta = -qty if row["transaction_type"] == "in" else qty if row["transaction_type"] == "out" else 0

            if delta != 0:
                cur.execute("update library_items set balance=balance+%s where item_code=%s", (delta, row["item_code"]))
            cur.execute("update library_transactions set status='cancelled' where id=%s", (id,))
            after = before + delta if before is not None else None
            write_library_audit(cur, "CANCEL_TRANSACTION", item_code=row["item_code"], item_name=row["item_name"],
                                quantity_change=delta, balance_before=before, balance_after=after,
                                transaction_id=id, detail=f"撤销 {row['transaction_type']} 记录", admin=admin)
            conn.commit()
    return redirect(url_for("library.library_records"))

@library_bp.route("/batch-in", methods=["GET", "POST"])
def library_batch_in():
    if not is_library_admin():
        return redirect(url_for("library.library_admin"))
    if not library_admin_can("can_batch_in"):
        return library_permission_denied()

    admin = get_library_admin()
    handled_by = library_admin_label(admin)

    if request.method == "POST":
        remark = request.form.get("remark", "").strip()
        codes = request.form.getlist("item_code[]")
        quantities = request.form.getlist("quantity[]")
        batch_time = now_malaysia()
        items = []

        for code, qty in zip(codes, quantities):
            code = code.strip().upper()
            try:
                qty = int(qty)
            except Exception:
                qty = 0
            if code and qty > 0:
                items.append((code, qty))

        if not items:
            flash("还没有扫描任何法物", "bad")
            return redirect(url_for("library.library_batch_in"))

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                for item_code, qty in items:
                    cur.execute("""select item_code,name,balance from library_items
                                   where item_code=%s and is_active=true for update""", (item_code,))
                    item = cur.fetchone()
                    if not item:
                        conn.rollback()
                        flash(f"找不到法物编号：{item_code}", "bad")
                        return redirect(url_for("library.library_batch_in"))

                    before = int(item["balance"] or 0)
                    after = before + qty
                    cur.execute("update library_items set balance=%s where item_code=%s", (after, item_code))
                    cur.execute("""insert into library_transactions
                                   (item_code,item_name,transaction_type,quantity,handled_by,remark,created_at)
                                   values (%s,%s,'in',%s,%s,%s,%s) returning id""",
                                (item["item_code"], item["name"], qty, handled_by, remark, batch_time))
                    transaction_id = cur.fetchone()["id"]
                    write_library_audit(cur, "BATCH_IN", item_code=item["item_code"], item_name=item["name"],
                                        quantity_change=qty, balance_before=before, balance_after=after,
                                        transaction_id=transaction_id, detail=remark or "批量入库", admin=admin)
                conn.commit()

        flash(f"批量入库完成：{len(items)} 种法物", "good")
        if library_admin_can("can_view_records"):
            return redirect(url_for("library.library_records"))
        return redirect(url_for("library.library_admin_home"))

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>批量入库</title>
    <link rel="stylesheet" href="/static/css/toolbox.css">
</head>
<body>

<div class="page">

    <div class="card">
        <h1 class="page-title">📦 批量入库</h1>
        <div class="page-subtitle">
            扫码枪或手机相机连续扫描。同一本自动累计，也可以直接修改数量。
        </div>

        <div class="alert alert-info">
            👤 当前负责人：<b>{{ admin_label }}</b>
        </div>

        <form method="post" onsubmit="return beforeSubmit();">
<div class="form-group">
                <label>备注</label>
                <input class="form-input" name="remark" placeholder="例如：新书到货 / 批量入库">
            </div>

            <div class="btn-row">
                <button type="button"
                        class="btn-tool btn-purple"
                        onclick="startCameraScan()">
                    📷 手机连续扫码
                </button>

                <button type="button"
                        class="btn-tool btn-secondary"
                        onclick="stopCameraScan()">
                    ⏹ 停止相机
                </button>
            </div>

            <video
                id="cameraPreview"
                style="width:100%;border-radius:18px;margin-top:15px;display:none;"
                autoplay
                muted
                playsinline>
            </video>

            <div id="scanStatus"
                 class="empty-state"
                 style="margin-top:12px;">
                扫码枪可直接扫描；手机请按“手机连续扫码”
            </div>

            <div class="form-group" style="margin-top:20px;">
                <label>扫码枪输入</label>
                <input
                    id="scanInput"
                    class="form-input"
                    placeholder="扫码枪扫描 BOOK0001，或手动输入编号后按 Enter"
                    autocomplete="off"
                    autofocus
                >
            </div>

            <div class="summary-grid">
                <div class="summary-box">
                    <div>种类</div>
                    <strong id="typeCount">0</strong>
                </div>
                <div class="summary-box">
                    <div>总数量</div>
                    <strong id="totalCount">0</strong>
                </div>
            </div>

            <div class="table-responsive" style="margin-top:18px;">
                <table class="record-table">
                    <thead>
                        <tr>
                            <th>编号</th>
                            <th>名称</th>
                            <th>现有库存</th>
                            <th>入库数量</th>
                            <th>操作</th>
                        </tr>
                    </thead>
                    <tbody id="batchBody">
                        <tr>
                            <td colspan="5" class="empty-state">
                                还没有扫描任何法物
                            </td>
                        </tr>
                    </tbody>
                </table>
            </div>

            <div id="hiddenInputs"></div>

            <div style="height:18px;"></div>

            <button class="btn-tool btn-success" type="submit">
                ✅ 确认全部入库
            </button>

            <a class="btn-tool btn-secondary" href="/library/admin/home">
                ⬅ 返回负责人专区
            </a>

        </form>
    </div>

</div>

<script>
let batch = {};
let itemNames = {};
let itemBalances = {};

let cameraStream = null;
let cameraTimer = null;
let lastCameraCode = "";
let lastCameraTime = 0;

const scanInput = document.getElementById("scanInput");
const batchBody = document.getElementById("batchBody");
const hiddenInputs = document.getElementById("hiddenInputs");
const cameraPreview = document.getElementById("cameraPreview");
const scanStatus = document.getElementById("scanStatus");

function normalizeCode(raw) {
    let code = (raw || "").trim().toUpperCase();

    if (code.includes("/LIBRARY/SCAN/")) {
        code = code.split("/LIBRARY/SCAN/").pop();
    }

    code = code.replace(/[^A-Z0-9]/g, "");

    let match = code.match(/^BOOK(\\d+)$/);
    if (match) {
        code = "BOOK" + match[1].padStart(4, "0");
    }

    return code;
}

async function addCode(code) {
    code = normalizeCode(code);

    if (!code) return;

    try {
        let res = await fetch(`/library/api/item/${code}`);

        if (!res.ok) {
            showStatus("❌ 找不到法物编号：" + code);
            return;
        }

        let data = await res.json();
        let item = data.item;

        itemNames[code] = item.name;
        itemBalances[code] = item.balance;

        if (!batch[code]) {
            batch[code] = 1;
        } else {
            batch[code] += 1;
        }

        renderBatch();
        showStatus("✅ " + item.name + "，目前数量：" + batch[code]);

    } catch (err) {
        showStatus("❌ 查询法物失败：" + code);
    }
}

function showStatus(text) {
    scanStatus.innerText = text;
}

function removeCode(code) {
    delete batch[code];
    delete itemNames[code];
    delete itemBalances[code];
    renderBatch();
    scanInput.focus();
}

function changeQty(code, value) {
    let qty = parseInt(value || "0");

    if (qty <= 0) {
        delete batch[code];
    } else {
        batch[code] = qty;
    }

    renderBatch();
    scanInput.focus();
}

function renderBatch() {
    batchBody.innerHTML = "";
    hiddenInputs.innerHTML = "";

    let codes = Object.keys(batch).sort();
    let total = 0;

    if (codes.length === 0) {
        batchBody.innerHTML = `
            <tr>
                <td colspan="5" class="empty-state">
                    还没有扫描任何法物
                </td>
            </tr>
        `;
    }

    codes.forEach(code => {
        let qty = batch[code];
        total += qty;

        let tr = document.createElement("tr");

        tr.innerHTML = `
            <td><strong>${code}</strong></td>
            <td>${itemNames[code] || ""}</td>
            <td>${itemBalances[code] ?? ""}</td>
            <td>
                <input
                    class="form-input"
                    type="number"
                    min="1"
                    value="${qty}"
                    onchange="changeQty('${code}', this.value)"
                    style="max-width:130px;"
                >
            </td>
            <td>
                <button
                    type="button"
                    class="btn-tool btn-danger"
                    onclick="removeCode('${code}')"
                    style="padding:8px 14px;"
                >
                    删除
                </button>
            </td>
        `;

        batchBody.appendChild(tr);

        hiddenInputs.innerHTML += `
            <input type="hidden" name="item_code[]" value="${code}">
            <input type="hidden" name="quantity[]" value="${qty}">
        `;
    });

    document.getElementById("typeCount").innerText = codes.length;
    document.getElementById("totalCount").innerText = total;
}

scanInput.addEventListener("keydown", function(e) {
    if (e.key === "Enter") {
        e.preventDefault();

        let code = normalizeCode(scanInput.value);
        scanInput.value = "";

        addCode(code);
    }
});

async function startCameraScan() {
    if (!("BarcodeDetector" in window)) {
        showStatus("⚠️ 这个手机浏览器不支持连续扫码，请用 Chrome 或直接用扫码枪。");
        return;
    }

    try {
        cameraStream = await navigator.mediaDevices.getUserMedia({
            video: {
                facingMode: "environment"
            }
        });

        cameraPreview.srcObject = cameraStream;
        cameraPreview.style.display = "block";

        const detector = new BarcodeDetector({
            formats: ["qr_code", "code_128"]
        });

        showStatus("📷 相机已开启，请对准 QR / Barcode");

        cameraTimer = setInterval(async function() {
            if (cameraPreview.readyState < 2) return;

            try {
                let barcodes = await detector.detect(cameraPreview);

                if (barcodes.length > 0) {
                    let raw = barcodes[0].rawValue;
                    let code = normalizeCode(raw);

                    let now = Date.now();

                    // 避免手机同一本在一秒内重复扫很多次
                    if (code === lastCameraCode && now - lastCameraTime < 1200) {
                        return;
                    }

                    lastCameraCode = code;
                    lastCameraTime = now;

                    await addCode(code);
                }

            } catch (err) {
                console.log(err);
            }

        }, 500);

    } catch (err) {
        showStatus("❌ 无法开启相机，请检查手机浏览器权限。");
    }
}

function stopCameraScan() {
    if (cameraTimer) {
        clearInterval(cameraTimer);
        cameraTimer = null;
    }

    if (cameraStream) {
        cameraStream.getTracks().forEach(track => track.stop());
        cameraStream = null;
    }

    cameraPreview.style.display = "none";
    cameraPreview.srcObject = null;

    showStatus("⏹ 相机已停止。扫码枪仍然可以继续扫描。");
    scanInput.focus();
}

document.addEventListener("click", function() {
    scanInput.focus();
});

function beforeSubmit() {
    if (Object.keys(batch).length === 0) {
        alert("还没有扫描任何法物");
        scanInput.focus();
        return false;
    }

    return confirm("确定要批量入库吗？");
}

window.onload = function() {
    scanInput.focus();
};
</script>

</body>
</html>
""", admin_label=library_admin_label(admin))

@library_bp.route("/api/item/<item_code>")
def library_api_item(item_code):
    item_code = item_code.strip().upper()

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select item_code, name, balance, category
                from library_items
                where item_code = %s
                  and is_active = true
            """, (item_code,))

            item = cur.fetchone()

    if not item:
        return {
            "ok": False,
            "message": "找不到法物"
        }, 404

    return {
        "ok": True,
        "item": {
            "item_code": item["item_code"],
            "name": item["name"],
            "balance": item["balance"],
            "category": item["category"],
        }
    }