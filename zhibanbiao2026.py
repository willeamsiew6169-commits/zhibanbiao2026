import os
import re
import shutil
import calendar
import datetime
import pandas as pd

from opencc import OpenCC
from openpyxl import Workbook
from lunardate import LunarDate
from pypinyin import lazy_pinyin
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl.comments import Comment
from difflib import get_close_matches
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from excel_style_utils import beautify_attendance_file
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

cc = OpenCC('t2s')



REPORT_MONTH = os.getenv("REPORT_MONTH")  # 例如 2026-04
REPORT_YEAR = os.getenv("REPORT_YEAR")    # 例如 2026


DEBUG = False
REBUILD_ALL_MONTHS = False   # True=重建全年所有有资料月份；False=只生成目标月份

def log(msg):
    print(msg)

def debug(msg):
    if DEBUG:
        print(msg)

# =========================
# 配置区
# =========================
VOLUNTEERS_FILE = "volunteers.xlsx"
VOLUNTEERS_SHEET = "volunteers"

ATTENDANCE_FILE = "attendance.xlsx"
ATTENDANCE_SHEET = "records"

SCHEDULE_FILE = "schedule_messages.xlsx"

PROJECT_NAME = "Cheras义工"
OUTPUT_DIR = "reports"

ROLE_ORDER = ["值班", "卫生", "佛台","供台", "供花",  "供果", "膳食", "佛学班"]

ROLE_COLORS = {
    "值班": "D9EAF7",
    "卫生": "FFF2CC",
    "佛台": "E2F0D9",
    "供花": "FCE4D6",
    "供台": "CCE5FF",
    "供果": "D9D2E9",
    "膳食": "EAD1DC",
    "佛学班": "F4CCCC",
}

ANOMALY_COLORS = {
    "报名没签到": "FF6666",
    "签到没报名": "C00000",
}

MULTI_ROLE_COLOR = "C9DAF8"
TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
NAME_FILL = PatternFill("solid", fgColor="FCE5CD")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

THIN = Side(style="thin", color="000000")

MANUAL_FILE = r"C:\gyt_system\值班报表2026\manual_adjustments.xlsx"
BACKUP_DIR = r"C:\gyt_system\值班报表2026\backups"

# ===== 白话佛法年报准备 =====
READING_FILE = "reading.xlsx"
import datetime
target_year = datetime.datetime.now().year

try:
    if os.path.exists(READING_FILE):
        reading_all = pd.read_excel(READING_FILE)

        if "日期" in reading_all.columns:
            reading_all["日期"] = pd.to_datetime(reading_all["日期"], errors="coerce")

            reading_year = reading_all[
                reading_all["日期"].dt.year == target_year
            ].copy()
        else:
            print("⚠️ reading.xlsx 没有【日期】栏位")

    else:
        print("⚠️ 找不到 reading.xlsx")

except Exception as e:
    print("⚠️ 读取 reading.xlsx 失败：", e)
    reading_year = pd.DataFrame()

def detect_target_year_month():
    # ========= 1) 优先：schedule_messages 里还没处理的资料 =========
    try:
        sch_df = pd.read_excel(SCHEDULE_FILE)

        if "日期" in sch_df.columns:
            sch_df["日期"] = pd.to_datetime(sch_df["日期"], errors="coerce")

            if "已处理" not in sch_df.columns:
                sch_df["已处理"] = ""

            if "原始信息" not in sch_df.columns:
                if "报名原文" in sch_df.columns:
                    sch_df["原始信息"] = sch_df["报名原文"]
                elif "whatsapp_raw" in sch_df.columns:
                    sch_df["原始信息"] = sch_df["whatsapp_raw"]
                else:
                    sch_df["原始信息"] = ""

            pending_df = sch_df[
                sch_df["日期"].notna() &
                sch_df["原始信息"].astype(str).str.strip().ne("") &
                (sch_df["已处理"].astype(str).str.strip().str.upper() != "YES")
            ].copy()

            print("schedule 未处理行数：", len(pending_df))

            if not pending_df.empty:
                latest_date = pending_df["日期"].max()
                print(f"[自动判断] 使用 schedule 未处理月份：{latest_date.year}-{latest_date.month:02d}")
                return latest_date.year, latest_date.month

    except Exception as e:
        print("schedule 自动判断失败：", e)

    # ========= 2) 如果已经进入新月份，自动开新月份 =========
    today = datetime.datetime.today()
    today_ym = (today.year, today.month)

    try:
        att_df = pd.read_excel(ATTENDANCE_FILE, sheet_name=ATTENDANCE_SHEET)

        if "日期" in att_df.columns:
            att_df["日期"] = pd.to_datetime(att_df["日期"], errors="coerce")
            att_dates = att_df["日期"].dropna()

            if not att_dates.empty:
                latest_date = att_dates.max()
                latest_ym = (latest_date.year, latest_date.month)

                # ✅ 关键：今天月份已经比 attendance 最新月份新
                if today_ym > latest_ym:
                    print(f"[自动判断] 发现新月份，自动使用今天月份：{today.year}-{today.month:02d}")
                    return today.year, today.month

                print(f"[自动判断] 使用 attendance 最新月份：{latest_date.year}-{latest_date.month:02d}")
                return latest_date.year, latest_date.month

    except Exception as e:
        print("attendance 自动判断失败：", e)

    # ========= 3) 最后 fallback =========
    print(f"[自动判断] fallback 使用今天：{today.year}-{today.month:02d}")
    return today.year, today.month

def get_available_months(att_df, year):
    df = att_df.copy()
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
    df = df[df["日期"].dt.year == year].copy()

    month_list = sorted(df["日期"].dt.month.dropna().unique().tolist())
    return month_list

def backup_file_if_exists(filepath, backup_dir):
    if not os.path.exists(filepath):
        return

    os.makedirs(backup_dir, exist_ok=True)

    base = os.path.basename(filepath)
    name, ext = os.path.splitext(base)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"{name}_{ts}{ext}")

    shutil.copy2(filepath, backup_path)
    print(f"[备份] 已备份: {backup_path}")

def load_manual_adjustments():
    cols = ["日期", "姓名", "报名", "签到", "岗位", "开始时间", "结束时间", "时数", "备注"]

    if not os.path.exists(MANUAL_FILE):
        return pd.DataFrame(columns=cols)

    try:
        df = pd.read_excel(MANUAL_FILE)
    except Exception:
        return pd.DataFrame(columns=cols)

    if df.empty:
        return pd.DataFrame(columns=cols)

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    df = df[cols].copy()

    df["姓名"] = df["姓名"].astype(str).str.strip()
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")

    df["报名"] = pd.to_numeric(df["报名"], errors="coerce").fillna(0).astype(int)
    df["签到"] = pd.to_numeric(df["签到"], errors="coerce").fillna(0).astype(int)

    for c in ["岗位", "开始时间", "结束时间", "时数", "备注"]:
        df[c] = df[c].fillna("").astype(str).str.strip()

    return df

def normalize_flag(x):
    if pd.isna(x):
        return 0

    s = str(x).strip().lower()

    if s in {"1", "1.0", "true", "yes", "y", "是", "有", "已签到", "签到", "√", "✓"}:
        return 1

    if s in {"0", "0.0", "false", "no", "n", "否", "无", "", "nan", "none"}:
        return 0

    # 数字型再兜底一次
    try:
        return 1 if float(s) >= 1 else 0
    except:
        return 0

def normalize_role(x):
    if pd.isna(x) or str(x).strip() == "":
        return "未填写"

    x = str(x).strip()

    if "佛学班" in x:
        return "佛学班"
    if "佛台" in x:
        return "佛台"
    if "卫生" in x:
        return "卫生"
    if "供花" in x:
        return "供花"
    if "供果" in x:
        return "供果"
    if "膳食" in x:
        return "膳食"
    if "值班" in x:
        return "值班"

    return x

def clean_attendance_df(df):
    cols = ["日期", "姓名", "报名", "签到", "岗位", "开始时间", "结束时间", "时数", "备注"]

    if df is None or df.empty:
        return pd.DataFrame(columns=cols)

    df = df.copy()

    for c in cols:
        if c not in df.columns:
            df[c] = ""

    df = df[cols].copy()

    df["姓名"] = df["姓名"].astype(str).str.strip()
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")

    df["报名"] = df["报名"].apply(normalize_flag)
    df["签到"] = df["签到"].apply(normalize_flag)

    for c in ["岗位", "开始时间", "结束时间", "时数", "备注"]:
        df[c] = df[c].fillna("").astype(str).str.strip()

    df["岗位"] = df["岗位"].apply(normalize_role)

    return df

def build_final_attendance(att_df, schedule_df):
    cols = ["日期", "姓名", "报名", "签到", "岗位", "开始时间", "结束时间", "时数", "备注"]

    if att_df is None or att_df.empty:
        att_df = pd.DataFrame(columns=cols)
    if schedule_df is None or schedule_df.empty:
        schedule_df = pd.DataFrame(columns=cols)

    att_df = clean_attendance_df(att_df)
    schedule_df = clean_attendance_df(schedule_df)
    manual_df = load_manual_adjustments()
    manual_df = clean_attendance_df(manual_df)

    dfs = []
    for df in [att_df, schedule_df, manual_df]:
        if df is not None and not df.empty:
            dfs.append(df[cols].copy())

    if dfs:
        final_df = pd.concat(dfs, ignore_index=True)
    else:
        final_df = pd.DataFrame(columns=cols)

    final_df = final_df[
        final_df["姓名"].astype(str).str.strip().ne("") &
        final_df["日期"].notna()
    ].copy()

    final_df = final_df.drop_duplicates(
        subset=["日期", "姓名", "岗位", "开始时间", "结束时间",],
        keep="first"
    ).reset_index(drop=True)

    return final_df

def save_attendance(att_df, attendance_file):
    export_df = att_df.copy()

    if "日期" in export_df.columns:
        export_df["日期"] = pd.to_datetime(export_df["日期"], errors="coerce")

    export_df = calc_hours(export_df)

    export_df["日期"] = export_df["日期"].dt.strftime("%Y-%m-%d")

    with pd.ExcelWriter(attendance_file, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name=ATTENDANCE_SHEET, index=False)

# ===== 固定义工名单 =====
GROUPS = {
    "膳食组": [
        {"姓名": "叶鏸郧", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "叶碧燕", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "陈金清", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "王碧琳", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "余丽苹", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "刘玉云", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "陈柔霓", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "颜美荭", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "黎嘉祺", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "葉荔銖", "报名": 1, "签到": 1, "岗位": "膳食"},
        {"姓名": "张凤", "报名": 1, "签到": 1, "岗位": "膳食"},

    ],
 
 
    "供花组": [
        {"姓名": "陈仪伶", "报名": 1, "签到": 1, "岗位": "供花"},
        {"姓名": "余丽苹", "报名": 1, "签到": 1, "岗位": "供花"},
    ],
    "供果组": [
        {"姓名": "王康芬", "报名": 1, "签到": 1, "岗位": "供果"},
        {"姓名": "郭丽诗", "报名": 1, "签到": 1, "岗位": "供果"},
        {"姓名": "蕭志倫", "报名": 1, "签到": 1, "岗位": "供果"},
        {"姓名": "程秋萍", "报名": 1, "签到": 1, "岗位": "供果"},
        {"姓名": "陈金清", "报名": 1, "签到": 1, "岗位": "供果"},
        {"姓名": "伍蔚枋", "报名": 1, "签到": 1, "岗位": "供果"},
        {"姓名": "黎嘉祺", "报名": 1, "签到": 1, "岗位": "供果"},
        {"姓名": "石秀萍", "报名": 1, "签到": 1, "岗位": "供果"},
    ],
    "佛学班": [
        {"姓名": "黄莉珍", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "王康芬", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "林臣顺", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "刘永耀", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "许愫芩", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "许银铃", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "方玉芬", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "黄丽萍", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "伍蔚枋", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "林凤美", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "陈映如", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "刘铧忆", "报名": 1, "签到": 1, "岗位": "佛学班"},
        {"姓名": "吴文杰", "报名": 1, "签到": 1, "岗位": "佛学班"},
    ]
}

def groups_to_df(groups):
    rows = []

    for group_name, people in groups.items():
        for p in people:
            row = p.copy()
            row["分组"] = group_name
            rows.append(row)

    df = pd.DataFrame(rows)
    return df

# =========================
# 工具函数
# =========================
def ensure_output_dir(path: str):
    if not os.path.exists(path):
        os.makedirs(path)


def normalize_yes_no(value):
    if pd.isna(value):
        return 0

    s = str(value).strip().lower()

    if s in {"1", "1.0", "true", "yes", "y", "是", "有", "已签到", "签到", "√", "✓"}:
        return 1

    if s in {"0", "0.0", "false", "no", "n", "否", "无", "", "nan", "none"}:
        return 0

    try:
        return 1 if float(s) >= 1 else 0
    except:
        return 0

def safe_to_datetime(v):
    if pd.isna(v):
        return None

    # pandas 时间
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()

    # datetime
    if isinstance(v, datetime.datetime):
        return v

    # date → 转 datetime
    if isinstance(v, datetime.date):
        return datetime.datetime.combine(v, datetime.time.min)

    # time → 补日期
    if isinstance(v, datetime.time):
        return datetime.datetime.combine(datetime.date(2000, 1, 1), v)

    s = str(v).strip()
    if not s:
        return None

    # 常见格式
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%H:%M:%S",
        "%H:%M",
    ]:
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue

    # pandas兜底
    try:
        x = pd.to_datetime(s, errors="coerce")
        if pd.isna(x):
            return None
        return x.to_pydatetime()
    except Exception:
        return None

def safe_to_date(v):
    if pd.isna(v):
        return None
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return None


def compute_hours(start_val, end_val):
    start_dt = safe_to_datetime(start_val)
    end_dt = safe_to_datetime(end_val)

    if not start_dt or not end_dt:
        return 0.0

    start_minutes = start_dt.hour * 60 + start_dt.minute
    end_minutes = end_dt.hour * 60 + end_dt.minute

    if end_minutes < start_minutes:
        return 0.0

    return round((end_minutes - start_minutes) / 60.0, 2)


def merge_intervals(intervals):
    if not intervals:
        return 0.0

    intervals = sorted(intervals, key=lambda x: x[0])
    merged = [intervals[0]]

    for curr_start, curr_end in intervals[1:]:
        last_start, last_end = merged[-1]
        if curr_start <= last_end:
            merged[-1] = (last_start, max(last_end, curr_end))
        else:
            merged.append((curr_start, curr_end))

    total_minutes = sum(end - start for start, end in merged)
    return round(total_minutes / 60.0, 2)


def auto_fit_columns(ws, extra=2, min_width=8, max_width=24):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        width = min(max(max_len + extra, min_width), max_width)
        ws.column_dimensions[col_letter].width = width

def calc_hours(df):
    def _calc(row):
        try:
            start_raw = row.get("开始时间")
            end_raw = row.get("结束时间")

            if pd.isna(start_raw) or pd.isna(end_raw) or start_raw == "" or end_raw == "":
                return 0

            start = pd.to_datetime(to_ampm(start_raw), errors="coerce")
            end = pd.to_datetime(to_ampm(end_raw), errors="coerce")

            if pd.isna(start) or pd.isna(end):
                return 0

            hours = (end - start).total_seconds() / 3600

            if hours < 0:
                return 0

            return round(hours, 2)

        except Exception:
            return 0

    df = df.copy()
    df["时数"] = df.apply(_calc, axis=1)
    return df

def set_cell(cell, value=None, fill=None, bold=False, size=10, center=True):
    if value is not None:
        cell.value = value
    if fill:
        cell.fill = fill
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    cell.font = Font(bold=bold, size=size)
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")


def role_combo_label(roles):
    unique_roles = [r for r in dict.fromkeys(roles) if r]
    if not unique_roles:
        return ""
    if len(unique_roles) == 1:
        return unique_roles[0]
    return " & ".join(unique_roles)


# =========================
# 读取主名单
# =========================
SORT_MODE = "pinyin"

def load_volunteers():
    if not os.path.exists(VOLUNTEERS_FILE):
        raise FileNotFoundError(f"找不到文件：{VOLUNTEERS_FILE}")

    df = pd.read_excel(VOLUNTEERS_FILE, sheet_name=VOLUNTEERS_SHEET)
    df = df.copy()

    # =========================
    # 必要栏位检查
    # =========================
    required_cols = ["姓名"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"volunteers.xlsx 缺少这些栏位：{missing}")

    # =========================
    # 自动补新栏位
    # =========================
    if "编号" not in df.columns:
        df["编号"] = ""

    if "状态" not in df.columns:
        df["状态"] = "在册"

    if "是否义工" not in df.columns:
        df["是否义工"] = "是"

    if "是否月费" not in df.columns:
        df["是否月费"] = "否"

    if "月费编号" not in df.columns:
        df["月费编号"] = ""

    if "备注" not in df.columns:
        df["备注"] = ""

    # =========================
    # 基本清理
    # =========================
    df["姓名"] = df["姓名"].astype(str).str.strip()
    df = df[df["姓名"] != ""].copy()

    df["状态"] = df["状态"].astype(str).str.strip()
    df["是否义工"] = df["是否义工"].astype(str).str.strip()
    df["是否月费"] = df["是否月费"].astype(str).str.strip()
    df["月费编号"] = df["月费编号"].astype(str).str.strip()
    df["备注"] = df["备注"].astype(str).str.strip()

    # 统一是否义工 / 是否月费
    def normalize_yes_no(x, default="否"):
        x = str(x).strip().lower()
        if x in ["是", "y", "yes", "1", "true"]:
            return "是"
        if x in ["否", "n", "no", "0", "false", "", "nan", "none"]:
            return "否"
        return default

    df["是否义工"] = df["是否义工"].apply(lambda x: normalize_yes_no(x, default="是"))
    df["是否月费"] = df["是否月费"].apply(lambda x: normalize_yes_no(x, default="否"))

    # 状态空白时默认
    df.loc[df["状态"].isin(["", "nan", "None"]), "状态"] = "在册"

    # =========================
    # ✅ 统一电话号码
    # =========================
    if "电话号码" in df.columns and "联络号码" in df.columns:
        df["电话号码"] = df["电话号码"].fillna(df["联络号码"])
    elif "电话号码" not in df.columns and "联络号码" in df.columns:
        df["电话号码"] = df["联络号码"]
    elif "电话号码" not in df.columns:
        df["电话号码"] = ""

    df["电话号码"] = df["电话号码"].apply(normalize_phone)

    if "联络号码" in df.columns:
        df = df.drop(columns=["联络号码"])

    # =========================
    # 义工编号处理
    # 规则：
    # 1. 只有 是否义工=是 才分配编号
    # 2. 非义工编号留空
    # =========================
    def clean_number(x):
        if pd.isna(x):
            return None
        s = str(x).strip()
        if s in ["", "nan", "None"]:
            return None
        try:
            return int(float(s))
        except:
            return None

    df["编号"] = df["编号"].apply(clean_number)

    volunteer_mask = df["是否义工"] == "是"
    non_volunteer_mask = df["是否义工"] != "是"

    # 非义工不保留义工编号
    df.loc[non_volunteer_mask, "编号"] = None

    existing_ids = df.loc[volunteer_mask, "编号"].dropna().tolist()
    max_id = int(max(existing_ids)) if existing_ids else 0

    missing_id_mask = volunteer_mask & df["编号"].isna()
    if missing_id_mask.any():
        for i in df[missing_id_mask].index:
            max_id += 1
            df.at[i, "编号"] = max_id

    # 编号输出格式：义工为整数，非义工留空
    df["编号"] = df["编号"].apply(lambda x: int(x) if pd.notna(x) else "")

    # =========================
    # 月费编号清理
    # 规则：
    # 1. 是否月费=否 -> 月费编号清空
    # 2. 是否月费=是 -> 保留原值
    # =========================
    df.loc[df["是否月费"] != "是", "月费编号"] = ""
    df["月费编号"] = df["月费编号"].replace(["nan", "None"], "").fillna("").astype(str).str.strip()

    # =========================
    # 排序辅助列
    # =========================
    df["姓名拼音"] = df["姓名"].apply(lambda x: "".join(lazy_pinyin(str(x))).lower())
    df["姓氏"] = df["姓名"].apply(lambda x: str(x)[0] if str(x) else "")

    # =========================
    # 排序逻辑
    # 义工在前，非义工在后
    # =========================
    df["_义工排序"] = df["是否义工"].apply(lambda x: 0 if x == "是" else 1)

    if SORT_MODE == "pinyin":
        df = df.sort_values(
            ["_义工排序", "姓名拼音", "编号"],
            kind="stable"
        )

    elif SORT_MODE == "surname":
        df = df.sort_values(
            ["_义工排序", "姓氏", "姓名拼音", "编号"],
            kind="stable"
        )

    else:
        # 编号排序时，非义工会排后面
        df["_编号排序"] = df["编号"].apply(lambda x: 999999 if x == "" else int(x))
        df = df.sort_values(
            ["_义工排序", "_编号排序"],
            kind="stable"
        )
        df = df.drop(columns=["_编号排序"])

    df = df.drop(columns=["_义工排序"])
    df = df.reset_index(drop=True)

    # =========================
    # 栏位顺序统一
    # =========================
    final_cols = [
        "编号", "姓名", "状态", "电话号码",
        "姓名拼音", "姓氏",
        "是否义工", "是否月费", "月费编号", "备注"
    ]

    # 自动加上其他新栏（例如 PIN）
    for col in df.columns:
        if col not in final_cols:
            final_cols.append(col)

    df = df[final_cols]

    return df


def normalize_phone(x):
    if pd.isna(x):
        return ""

    s = str(x).strip()

    # 去掉 Excel 的 .0
    if s.endswith(".0"):
        s = s[:-2]

    # 只保留数字
    s = re.sub(r"\D", "", s)

    # =========================
    # 马来西亚电话修复（关键）
    # =========================

    # 情况1：少了0（9位）
    # 104391956 → 0104391956
    if len(s) == 9 and not s.startswith("0"):
        s = "0" + s

    # 情况2：少了0（10位）
    # 1110280429 → 01110280429
    elif len(s) == 10 and not s.startswith("0"):
        s = "0" + s

    return s

def beautify_volunteers_file(file_path, sheet_name):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def get_display_width(s):
        """中文按2格，英文按1格"""
        width = 0
        for ch in str(s):
            if ord(ch) > 127:
                width += 2
            else:
                width += 1
        return width

    wb = load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        print(f"找不到工作表：{sheet_name}")
        print("现有工作表：", wb.sheetnames)
        return

    ws = wb[sheet_name]

    # ===== 样式 =====
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, name="微软雅黑", size=12)

    body_font = Font(name="微软雅黑", size=12)
    name_font = Font(name="微软雅黑", size=13, bold=True)
    phone_font = Font(name="微软雅黑", size=12)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 冻结窗格
    ws.freeze_panes = "A2"

    # 自动筛选
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 1 and max_col >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # 找列
    col_map = {}
    for cell in ws[1]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    name_col = col_map.get("姓名")
    phone_col = col_map.get("电话号码")
    status_col = col_map.get("状态")

    # 全表样式
    for row in ws.iter_rows():
        for cell in row:
            cell.font = body_font
            cell.alignment = center_align
            cell.border = border

    # 表头
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 数据区
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 28

        if name_col:
            ws.cell(r, name_col).font = name_font
            ws.cell(r, name_col).alignment = left_align

        if phone_col:
            ws.cell(r, phone_col).font = phone_font

    # 自动列宽（支持中文）
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_length = 0

        for cell in col_cells:
            try:
                val = "" if cell.value is None else str(cell.value)
                length = get_display_width(val)
                max_length = max(max_length, length)
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    # 常用列优化
    if "编号" in col_map:
        ws.column_dimensions[get_column_letter(col_map["编号"])].width = 8

    if name_col:
        ws.column_dimensions[get_column_letter(name_col)].width = 20

    if phone_col:
        ws.column_dimensions[get_column_letter(phone_col)].width = 16

    if status_col:
        ws.column_dimensions[get_column_letter(status_col)].width = 10

    wb.save(file_path)
    print(f"volunteers 已美化：{file_path} / {sheet_name}")

# =========================
# 读取出席记录
# =========================
def load_attendance():
    if not os.path.exists(ATTENDANCE_FILE):
        raise FileNotFoundError(f"找不到文件：{ATTENDANCE_FILE}")

    df = pd.read_excel(ATTENDANCE_FILE, sheet_name=ATTENDANCE_SHEET)

    required_cols = ["日期", "姓名", "报名", "签到", "岗位", "开始时间", "结束时间"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"attendance.xlsx 缺少这些栏位：{missing}")

    df = df.copy()
    df["日期"] = df["日期"].apply(safe_to_date)
    df["姓名"] = df["姓名"].astype(str).str.strip()
    df["报名"] = df["报名"].apply(normalize_flag)
    df["签到"] = df["签到"].apply(normalize_flag)

    print("清洗后签到唯一值：", df["签到"].unique())

    df["岗位"] = df["岗位"].apply(normalize_role)

    if "备注" not in df.columns:
        df["备注"] = ""

    if "时数" not in df.columns:
        df["时数"] = ""

    df["自动时数"] = df.apply(
        lambda row: compute_hours(row["开始时间"], row["结束时间"]), axis=1
    )

    def final_hours(row):
        val = row.get("时数", "")
        if pd.isna(val) or str(val).strip() == "":
            return row["自动时数"]
        try:
            return float(val)
        except Exception:
            return row["自动时数"]

    df["时数"] = df.apply(final_hours, axis=1)
    df = df[df["日期"].notna() & (df["姓名"] != "")].reset_index(drop=True)

    # 按日期 + 姓名 + 时间排序
    df = df.sort_values(
        by=["日期", "姓名", "开始时间", "结束时间"],
        ascending=[True, True, True, True]
    ).reset_index(drop=True)

    return df

def get_lunar_info(date_obj):
    """
    输入 datetime/date，返回：
    - 星期
    - 农历月日文字
    - 节日名称（如有）
    """

    if isinstance(date_obj, datetime.datetime):
        date_obj = date_obj.date()

    elif isinstance(date_obj, datetime.date):
        pass

    else:
        date_obj = pd.to_datetime(date_obj, errors="coerce")
        if pd.isna(date_obj):
            return "", "", "", None
        date_obj = date_obj.date()

    weekday_map = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
    weekday_text = weekday_map[date_obj.weekday()]

    lunar = LunarDate.fromSolarDate(date_obj.year, date_obj.month, date_obj.day)

    month_names = {
        1: "正月", 2: "二月", 3: "三月", 4: "四月", 5: "五月", 6: "六月",
        7: "七月", 8: "八月", 9: "九月", 10: "十月", 11: "冬月", 12: "腊月"
    }
    day_names = {
        1: "初一", 2: "初二", 3: "初三", 4: "初四", 5: "初五",
        6: "初六", 7: "初七", 8: "初八", 9: "初九", 10: "初十",
        11: "十一", 12: "十二", 13: "十三", 14: "十四", 15: "十五",
        16: "十六", 17: "十七", 18: "十八", 19: "十九", 20: "二十",
        21: "廿一", 22: "廿二", 23: "廿三", 24: "廿四", 25: "廿五",
        26: "廿六", 27: "廿七", 28: "廿八", 29: "廿九", 30: "三十"
    }

    lunar_text = f"{month_names.get(lunar.month, str(lunar.month)+'月')}{day_names.get(lunar.day, str(lunar.day))}"

    # 你要的重点日子
    special_days = {
        (1, 1): "农历初一",
        (1, 15): "农历十五",
        (2, 8): "释迦摩尼佛出家日",
        (2, 15): "释迦摩尼佛涅槃日",
        (2, 19): "观世音菩萨诞辰日",
        (4, 8): "释迦摩尼佛诞辰日",
        (6, 19): "观世音菩萨成道日",
        (9, 19): "观世音菩萨出家日",
        (10, 6):"恩师卢军宏涅槃日",
        (12, 8): "释迦摩尼佛成道日",
    }

    special_name = special_days.get((lunar.month, lunar.day), "")

    return weekday_text, lunar_text, special_name, lunar



# =========================
# 月报统计
# =========================
def build_monthly_summary(vol_df, att_df, year, month):
    vol_df = vol_df.copy()
    att_df = att_df.copy()

    # ========= 基础清洗 =========
    vol_df["姓名"] = vol_df["姓名"].astype(str).str.strip()
    att_df["姓名"] = att_df["姓名"].astype(str).str.strip()

    att_df["日期"] = pd.to_datetime(att_df["日期"], errors="coerce")

    if "报名" in att_df.columns:
        att_df["报名"] = pd.to_numeric(att_df["报名"], errors="coerce").fillna(0).astype(int)
    else:
        att_df["报名"] = 0

    if "签到" in att_df.columns:
        att_df["签到"] = pd.to_numeric(att_df["签到"], errors="coerce").fillna(0).astype(int)
    else:
        att_df["签到"] = 0

    if "岗位" not in att_df.columns:
        att_df["岗位"] = ""

    if "备注" not in att_df.columns:
        att_df["备注"] = ""

    if "开始时间" not in att_df.columns:
        att_df["开始时间"] = ""

    if "结束时间" not in att_df.columns:
        att_df["结束时间"] = ""

    def normalize_role(x):
        if pd.isna(x) or str(x).strip() == "":
            return "未填写"

        x = str(x).strip()

        if "佛学班" in x:
            return "佛学班"
        if "佛台" in x:
            return "佛台"
        if "卫生" in x:
            return "卫生"
        if "供花" in x:
            return "供花"
        if "供果" in x :
            return "供果"
        if "膳食" in x:
            return "膳食"
        if "值班" in x:
            return "值班"

        return x

    att_df["岗位"] = att_df["岗位"].apply(normalize_role)

    # ========= 只取当月 =========
    month_df = att_df[
        (att_df["日期"].dt.year == year) &
        (att_df["日期"].dt.month == month)
    ].copy()

    days_in_month = calendar.monthrange(year, month)[1]

    # 主名单
    all_names = vol_df["姓名"].dropna().astype(str).str.strip().tolist()

    if "编号" in vol_df.columns:
        no_map = dict(zip(vol_df["姓名"], vol_df["编号"]))
    else:
        no_map = dict(zip(vol_df["姓名"], [""] * len(vol_df)))

    # ========= 把“有出席记录但不在主名单的人”也补进来 =========
    extra_names = [
        n for n in month_df["姓名"].dropna().astype(str).str.strip().unique().tolist()
        if n not in no_map
    ]
    for n in extra_names:
        no_map[n] = ""
        all_names.append(n)

    # 去重并排序，避免重复名字
    all_names = list(dict.fromkeys(all_names))

    daily_info = defaultdict(lambda: defaultdict(lambda: {
        "roles": [],
        "signup": 0,
        "checkin": 0,
        "notes": [],
    }))

    # ========= 汇总每天资料 =========
    for _, row in month_df.iterrows():
        name = str(row["姓名"]).strip()

        if not name:
            continue

        if pd.isna(row["日期"]):
            continue

        day = row["日期"].day

        if name not in no_map:
            no_map[name] = ""
            if name not in all_names:
                all_names.append(name)

        role = str(row["岗位"]).strip() if pd.notna(row["岗位"]) else "未填写"
        note = str(row["备注"]).strip() if pd.notna(row["备注"]) else ""

        # 岗位去重
        if role and role not in daily_info[name][day]["roles"]:
            daily_info[name][day]["roles"].append(role)

        daily_info[name][day]["signup"] = max(
            daily_info[name][day]["signup"],
            int(row["报名"]) if pd.notna(row["报名"]) else 0
        )
        daily_info[name][day]["checkin"] = max(
            daily_info[name][day]["checkin"],
            int(row["签到"]) if pd.notna(row["签到"]) else 0
        )

        if note and note not in daily_info[name][day]["notes"]:
            daily_info[name][day]["notes"].append(note)

    # ========= 异常表 =========
    anomaly_rows = []
    for _, row in month_df.iterrows():
        anomaly_type = None

        signup = int(row["报名"]) if pd.notna(row["报名"]) else 0
        checkin = int(row["签到"]) if pd.notna(row["签到"]) else 0

        if signup != checkin:
            anomaly_type = "数据异常（报名≠签到）"

        start_raw = row.get("开始时间")
        end_raw = row.get("结束时间")

        anomaly_rows.append({
            "日期": row["日期"].strftime("%Y-%m-%d") if pd.notna(row.get("日期")) else "",
            "姓名": str(row.get("姓名", "")).strip(),
            "岗位": str(row.get("岗位", "")).strip() if pd.notna(row.get("岗位")) else "",
            "开始时间": to_ampm(start_raw).strip() if pd.notna(start_raw) else "",
            "结束时间": to_ampm(end_raw).strip() if pd.notna(end_raw) else "",
            "异常": anomaly_type,
            "备注": str(row.get("备注", "")).strip() if pd.notna(row.get("备注")) else "",
        })

    anomaly_df = pd.DataFrame(anomaly_rows)

    # ========= 底部每天总人数（按签到人数） =========
    daily_totals = []
    for day in range(1, days_in_month + 1):
        cnt = 0
        for name in all_names:
            if daily_info[name][day]["checkin"] == 1:
                cnt += 1
        daily_totals.append(cnt)

    month_grand_total = sum(daily_totals)

    return month_df, daily_info, days_in_month, all_names, no_map, anomaly_df, daily_totals, month_grand_total

# =========================
# 年报统计
# =========================
def build_yearly_summary(vol_df, att_df, year):
    att_df = att_df.copy()
    att_df["日期"] = pd.to_datetime(att_df["日期"], errors="coerce")
    att_df["姓名"] = att_df["姓名"].astype(str).str.strip()
    att_df["岗位"] = att_df["岗位"].astype(str).str.strip()

    # 关键：不要再只用 to_numeric
    att_df["签到"] = att_df["签到"].apply(normalize_flag)
    att_df["报名"] = att_df["报名"].apply(normalize_flag)

    year_df = att_df[att_df["日期"].dt.year == year].copy()

    debug(f"year_df 行数: {len(year_df)}")
    print("签到所有值:", year_df["签到"].unique())

    if year_df.empty:
        empty = vol_df[["编号", "姓名"]].copy()
        empty["全年出席天数"] = 0
        empty["岗位记录次数"] = 0
        empty["全年总时数"] = 0.0
        empty["总服务时数"] = 0.0
        empty["出席率"] = "0.00%"
        return empty, pd.DataFrame(), pd.DataFrame()

    signed_df = year_df[year_df["签到"] == 1].copy()

    print("signed_df 行数:", len(signed_df))

    if year_df.empty:
        empty = vol_df[["编号", "姓名"]].copy()
        empty["全年出席天数"] = 0
        empty["岗位记录次数"] = 0
        empty["全年总时数"] = 0.0
        empty["总服务时数"] = 0.0
        empty["出席率"] = "0.00%"
        return empty, pd.DataFrame(), pd.DataFrame()

    signed_df = year_df[year_df["签到"] == 1].copy()

    # 防止没有时数字段
    if "时数" not in signed_df.columns:
        signed_df["时数"] = 0

    signed_df["时数"] = pd.to_numeric(signed_df["时数"], errors="coerce").fillna(0)

    # 总服务时数（直接按每条记录的时数加总）
    hours_df = (
        signed_df.groupby("姓名")["时数"]
        .sum()
        .reset_index()
        .rename(columns={"时数": "总服务时数"})
    )

    all_activity_days = signed_df["日期"].nunique()
    if all_activity_days == 0:
        all_activity_days = 1

    # 全年出席天数：同一个人同一天算1天
    attend_days = (
        signed_df.groupby("姓名")["日期"]
        .nunique()
        .rename("全年出席天数")
    )

    # 岗位记录次数：按签到记录行数
    shift_counts = (
        signed_df.groupby("姓名")
        .size()
        .rename("岗位记录次数")
    )

    # 全年总时数：同一天多个岗位如果时间重叠，自动去重
    person_day_intervals = defaultdict(list)
    for _, row in signed_df.iterrows():
        start_dt = safe_to_datetime(row["开始时间"])
        end_dt = safe_to_datetime(row["结束时间"])
        if not start_dt or not end_dt:
            continue

        start_min = start_dt.hour * 60 + start_dt.minute
        end_min = end_dt.hour * 60 + end_dt.minute

        if end_min < start_min:
            continue

        key = (row["姓名"], row["日期"])
        person_day_intervals[key].append((start_min, end_min))

    person_total_hours = defaultdict(float)
    for (name, _day), intervals in person_day_intervals.items():
        person_total_hours[name] += merge_intervals(intervals)

    total_hours = pd.Series(person_total_hours, name="全年总时数")

    # 岗位统计（按天数，不按记录次数）
    role_stat_df = signed_df.copy()
    role_stat_df["日期"] = pd.to_datetime(role_stat_df["日期"], errors="coerce").dt.strftime("%Y-%m-%d")

    # 同一天、同一人、同一岗位，只算一次
    role_stat_df = role_stat_df.drop_duplicates(
        subset=["日期", "姓名", "岗位"],
        keep="first"
    ).copy()

    role_pivot = pd.pivot_table(
        role_stat_df,
        index="姓名",
        columns="岗位",
        values="日期",
        aggfunc="count",
        fill_value=0
    )

    # 强制所有岗位都出现（包括供台）
    for role in ROLE_ORDER:
        if role not in role_pivot.columns:
            role_pivot[role] = 0

    # 按固定顺序排列
    role_pivot = role_pivot[ROLE_ORDER]

    role_pivot = role_pivot.rename(columns={"值班": "值班天数"})

    # 每月出席天数
    signed_df["月份"] = signed_df["日期"].apply(lambda d: d.month)
    month_unique = (
        signed_df.groupby(["姓名", "月份"])["日期"]
        .nunique()
        .reset_index()
    )

    monthly_pivot = month_unique.pivot_table(
        index="姓名",
        columns="月份",
        values="日期",
        aggfunc="sum",
        fill_value=0
    )

    for m in range(1, 13):
        if m not in monthly_pivot.columns:
            monthly_pivot[m] = 0

    monthly_pivot = monthly_pivot[sorted(monthly_pivot.columns)]
    monthly_pivot.columns = [f"{m}月" for m in monthly_pivot.columns]

    # 转成 DataFrame 方便 merge
    attend_days_df = attend_days.reset_index()
    shift_counts_df = shift_counts.reset_index()
    total_hours_df = total_hours.reset_index()
    total_hours_df.columns = ["姓名", "全年总时数"]

    # 主汇总表
    summary = vol_df.copy()   # ⭐ 改这里（关键）

    # 如果没有状态列，自动补
    if "状态" not in summary.columns:
        summary["状态"] = "在册"
    summary = summary.merge(attend_days_df, on="姓名", how="left")
    summary = summary.merge(shift_counts_df, on="姓名", how="left")
    summary = summary.merge(total_hours_df, on="姓名", how="left")
    summary = summary.merge(hours_df, on="姓名", how="left")

    if not role_pivot.empty:
        role_reset = role_pivot.reset_index()
        summary = summary.merge(role_reset, on="姓名", how="left")

    if not monthly_pivot.empty:
        month_reset = monthly_pivot.reset_index()
        summary = summary.merge(month_reset, on="姓名", how="left")

    # 补空值
    if "全年出席天数" not in summary.columns:
        summary["全年出席天数"] = 0
    if "岗位记录次数" not in summary.columns:
        summary["岗位记录次数"] = 0
    if "全年总时数" not in summary.columns:
        summary["全年总时数"] = 0.0
    if "总服务时数" not in summary.columns:
        summary["总服务时数"] = 0.0

    # 👉 ⭐ 加这一段（关键）
    numeric_cols = ["全年出席天数", "岗位记录次数", "全年总时数", "总服务时数"]

    for col in numeric_cols:
        if col in summary.columns:
            summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0)

    summary = summary.copy()
    # 👉 再做格式
    summary["全年总时数"] = summary["全年总时数"].round(2)
    summary["总服务时数"] = summary["总服务时数"].round(2)

    summary["全年出席天数"] = pd.to_numeric(summary["全年出席天数"], errors="coerce").fillna(0).astype(int)
    summary["岗位记录次数"] = pd.to_numeric(summary["岗位记录次数"], errors="coerce").fillna(0).astype(int)
    summary["全年总时数"] = summary["全年总时数"].fillna(0).round(2)
    summary["总服务时数"] = summary["总服务时数"].fillna(0).round(2)

    protected_cols = ["编号", "姓名", "全年总时数", "总服务时数", "出席率"]

    for col in summary.columns:
        if col not in protected_cols:
            summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0).astype(int)
        # 出席率
        summary["出席率"] = summary["全年出席天数"].apply(
            lambda x: f"{(x / all_activity_days) * 100:.2f}%"
        )

    summary = summary.sort_values(by=["编号"]).reset_index(drop=True)

    # 明细
    detail = signed_df[["日期", "姓名", "岗位", "开始时间", "结束时间", "时数", "备注"]].copy()
    detail["日期"] = detail["日期"].astype(str)

    return summary, role_pivot.reset_index() if not role_pivot.empty else pd.DataFrame(), detail

def build_year_summary_table(vol_df, att_df, year):
    att_df = att_df.copy()
    att_df["日期"] = pd.to_datetime(att_df["日期"], errors="coerce")

    year_df = att_df[att_df["日期"].dt.year == year].copy()
    signed_df = year_df[year_df["签到"] == 1].copy()

    # 每人每月出席天数（同一天算1次）
    if not signed_df.empty:
        signed_df["月份"] = signed_df["日期"].dt.month
        month_unique = (
            signed_df.groupby(["姓名", "月份"])["日期"]
            .nunique()
            .reset_index()
        )

        monthly_pivot = month_unique.pivot_table(
            index="姓名",
            columns="月份",
            values="日期",
            aggfunc="sum",
            fill_value=0
        )
    else:
        monthly_pivot = pd.DataFrame()

    # 补足 1~12 月
    for m in range(1, 13):
        if m not in monthly_pivot.columns:
            monthly_pivot[m] = 0

    if not monthly_pivot.empty:
        monthly_pivot = monthly_pivot[sorted(monthly_pivot.columns)]
        monthly_pivot.columns = [f"{m}月" for m in monthly_pivot.columns]
        monthly_pivot = monthly_pivot.reset_index()
    else:
        monthly_pivot = vol_df[["姓名"]].copy()
        for m in range(1, 13):
            monthly_pivot[f"{m}月"] = 0

    month_cols = [f"{m}月" for m in range(1, 13)]
    monthly_pivot["Total"] = monthly_pivot[month_cols].sum(axis=1)

    # 开始值班日期（最早签到日期）
    if not signed_df.empty:
        first_dates = (
            signed_df.groupby("姓名")["日期"]
            .min()
            .reset_index()
            .rename(columns={"日期": "开始值班日期"})
        )
        first_dates["开始值班日期"] = first_dates["开始值班日期"].dt.strftime("%Y-%m-%d")
    else:
        first_dates = pd.DataFrame(columns=["姓名", "开始值班日期"])

    # 合并主名单
    summary = vol_df.copy()

    # 统一电话号码列
    if "电话号码" not in summary.columns and "联络号码" in summary.columns:
        summary["电话号码"] = summary["联络号码"]
    elif "电话号码" not in summary.columns:
        summary["电话号码"] = ""

    summary = summary.merge(monthly_pivot, on="姓名", how="left")
    summary = summary.merge(first_dates, on="姓名", how="left")

    for col in month_cols + ["Total"]:
        if col not in summary.columns:
            summary[col] = 0

        summary[col] = pd.to_numeric(summary[col], errors="coerce").fillna(0).astype(int)

    summary["开始值班日期"] = summary["开始值班日期"].fillna("")

    # 保持主名单顺序
    final_cols = ["姓名", "电话号码", "状态"] + month_cols + ["Total", "开始值班日期"]
    summary = summary[final_cols]

    return summary

def add_date_comments_and_colors(ws, date_col_idx=1, start_row=2):
    weekday_map = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

    month_names = {
        1: "正月", 2: "二月", 3: "三月", 4: "四月", 5: "五月", 6: "六月",
        7: "七月", 8: "八月", 9: "九月", 10: "十月", 11: "冬月", 12: "腊月"
    }
    day_names = {
        1: "初一", 2: "初二", 3: "初三", 4: "初四", 5: "初五",
        6: "初六", 7: "初七", 8: "初八", 9: "初九", 10: "初十",
        11: "十一", 12: "十二", 13: "十三", 14: "十四", 15: "十五",
        16: "十六", 17: "十七", 18: "十八", 19: "十九", 20: "二十",
        21: "廿一", 22: "廿二", 23: "廿三", 24: "廿四", 25: "廿五",
        26: "廿六", 27: "廿七", 28: "廿八", 29: "廿九", 30: "三十"
    }

    # 你要的大日子，可继续加
    special_days = {
        (1, 1): "农历初一",
        (1, 15): "农历十五",
        (2, 8): "释迦摩尼佛出家日",
        (2, 15): "释迦摩尼佛涅槃日",
        (2, 19): "观世音菩萨诞辰日",
        (4, 8): "释迦摩尼佛诞辰日",
        (6, 19): "观世音菩萨成道日",
        (9, 19): "观世音菩萨出家日",
        (10, 6): "恩师卢军宏涅槃日",
        (12, 8): "释迦摩尼佛成道日",
    }

    weekend_fill = PatternFill("solid", fgColor="DDEBF7")   # 周末：浅蓝
    lunar_fill = PatternFill("solid", fgColor="FFF2CC")     # 初一/十五：浅黄
    special_fill = PatternFill("solid", fgColor="F4CCCC")   # 大日子：浅红

    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(r, date_col_idx)
        val = cell.value

        if val is None or str(val).strip() in ("", "Total"):
            continue

        try:
            parsed = pd.to_datetime(val, errors="coerce")
            if pd.isna(parsed):
                continue
            d = parsed.date()
        except Exception:
            continue

        weekday_text = weekday_map[d.weekday()]
        lunar = LunarDate.fromSolarDate(d.year, d.month, d.day)
        lunar_text = f"{month_names.get(lunar.month, str(lunar.month)+'月')}{day_names.get(lunar.day, str(lunar.day))}"
        special_name = special_days.get((lunar.month, lunar.day), "")

        # 保留原有批注
        existing_comment = cell.comment.text.strip() if cell.comment and cell.comment.text else ""
        lines = []
        if existing_comment:
            lines.append(existing_comment)
        lines.append(weekday_text)
        lines.append(lunar_text)
        if special_name:
            lines.append(special_name)

        cell.comment = Comment("\n".join(lines), "system")

        # 日期格上色
        cell.fill = PatternFill(fill_type=None)
        if special_name:
            cell.fill = special_fill
        elif lunar.day in (1, 15):
            cell.fill = lunar_fill
        elif d.weekday() >= 5:
            cell.fill = weekend_fill

def add_month_table_date_marks(ws, header_row, start_day_col, year, month, days_in_month):
    weekday_map = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

    month_names = {
        1: "正月", 2: "二月", 3: "三月", 4: "四月", 5: "五月", 6: "六月",
        7: "七月", 8: "八月", 9: "九月", 10: "十月", 11: "冬月", 12: "腊月"
    }
    day_names = {
        1: "初一", 2: "初二", 3: "初三", 4: "初四", 5: "初五",
        6: "初六", 7: "初七", 8: "初八", 9: "初九", 10: "初十",
        11: "十一", 12: "十二", 13: "十三", 14: "十四", 15: "十五",
        16: "十六", 17: "十七", 18: "十八", 19: "十九", 20: "二十",
        21: "廿一", 22: "廿二", 23: "廿三", 24: "廿四", 25: "廿五",
        26: "廿六", 27: "廿七", 28: "廿八", 29: "廿九", 30: "三十"
    }

    # 你常用的大日子，可以继续加
    special_days = {
        # ===== 常规 =====
        (1, 1): "农历初一",
        (1, 15): "农历十五",

        # ===== 二月 =====
        (2, 8): "释迦摩尼佛出家日",
        (2, 15): "释迦摩尼佛涅槃日",
        (2, 19): "观世音菩萨诞辰日",

        # ===== 四月 =====
        (4, 8): "释迦摩尼佛诞辰日",

        # ===== 六月 =====
        (6, 19): "观世音菩萨成道日",

        # ===== 九月 =====
        (9, 19): "观世音菩萨出家日",

        # ===== 十月 =====
        (10, 6): "恩师卢军宏涅槃日",
        
        # ===== 十二月 =====
        (12, 8): "释迦摩尼佛成道日",
    }

    # 颜色
    # 🟦 周末（改成淡灰更安全）
    weekend_fill = PatternFill("solid", fgColor="F2F2F2")

    # 🟨 初一 / 十五
    lunar_fill = PatternFill("solid", fgColor="FFF2CC")

    # 🟥 佛陀相关（用淡红）
    buddha_fill = PatternFill("solid", fgColor="F4CCCC")

    # 🟪 观音相关（用淡紫）
    guanyin_fill = PatternFill("solid", fgColor="EAD1DC")

    master_fill = PatternFill("solid", fgColor="D9D2E9")

    for day in range(1, days_in_month + 1):
        col = start_day_col + day - 1
        cell = ws.cell(header_row, col)

        try:
            d = datetime.date(year, month, day)
        except Exception:
            continue

        weekday_text = weekday_map[d.weekday()]
        lunar = LunarDate.fromSolarDate(d.year, d.month, d.day)
        lunar_text = f"{month_names.get(lunar.month, str(lunar.month) + '月')}{day_names.get(lunar.day, str(lunar.day))}"
        special_name = special_days.get((lunar.month, lunar.day), "")

        # 先决定这一天用什么颜色
        fill = None
        if "恩师" in special_name or "卢军宏" in special_name:
            fill = master_fill
        elif "观世音菩萨" in special_name or "观世音" in special_name or "观音" in special_name:
            fill = guanyin_fill
        elif "释迦摩尼佛" in special_name or "释迦牟尼佛" in special_name or "释迦" in special_name or "佛" in special_name:
            fill = buddha_fill
        elif lunar.day in (1, 15):
            fill = lunar_fill
        elif d.weekday() >= 5:
            fill = weekend_fill

        # 如果这个表头日期格原本就有批注，保留再追加
        existing_comment = cell.comment.text.strip() if cell.comment and cell.comment.text else ""

        comment_lines = []
        if existing_comment:
            comment_lines.append(existing_comment)

        comment_lines.append(d.strftime("%Y-%m-%d"))
        comment_lines.append(weekday_text)
        comment_lines.append(lunar_text)

        if special_name:
            comment_lines.append(special_name)

        cell.comment = Comment("\n".join(comment_lines), "system")

        # 整列上色（从表头到最后一行）
        from openpyxl.styles import Border, Side, Font

        thick = Side(style="medium", color="000000")
        thick_border = Border(top=thick, bottom=thick)

        for r in range(header_row, ws.max_row + 1):
            target_cell = ws.cell(r, col)

            # 有数据的格子不盖掉原本岗位颜色
            if target_cell.value not in (None, "", 0):
                continue

            if fill is not None:
                target_cell.fill = fill

        # 表头强化（只针对日期那一格）
        header_cell = ws.cell(header_row, col)

        if "释迦摩尼佛" in special_name:
            header_cell.fill = buddha_fill
            header_cell.font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
            header_cell.border = thick_border

        elif "观世音菩萨" in special_name:
            header_cell.fill = guanyin_fill
            header_cell.font = Font(name="微软雅黑", size=12, bold=True, color="7030A0")
            header_cell.border = thick_border

        elif lunar.day in (1, 15):
            header_cell.fill = lunar_fill
            header_cell.font = Font(name="微软雅黑", size=12, bold=True)

        elif d.weekday() >= 5:
            header_cell.fill = weekend_fill
            header_cell.font = Font(name="微软雅黑", size=12, bold=True)

# =========================
# 写月报
# =========================
def write_monthly_report(month_df, daily_info, days_in_month, all_names, no_map,
                         anomaly_df, daily_totals, month_grand_total, year, month):
    ensure_output_dir(OUTPUT_DIR)
    out_file = os.path.join(OUTPUT_DIR, f"{year}_{month:02d}_月报.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月月表"
    
    total_col = days_in_month + 3
    thin_black = Side(style="thin", color="000000")
    thin_gray = Side(style="thin", color="999999")
    border_black = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
    border_gray = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    # ==============================
    # 1) 标题区
    # ==============================
    ws.merge_cells("B1:Q1")
    ws.merge_cells("B2:Q2")

    set_cell(ws.cell(1, 2), "马来西亚卢台长心灵法门共修会（蕉赖分会- Chears）", bold=True, size=18)
    set_cell(ws.cell(2, 2), f"{year}年义工值班表", bold=True, size=14)

    ws.cell(1, 2).font = Font(name="微软雅黑", size=18, bold=True)
    ws.cell(2, 2).font = Font(name="微软雅黑", size=14, bold=True)
    ws.cell(1, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(2, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    # 清标题区边框，只留底线
    for r in range(1, 3):
        for c in range(2, 18):  # B~Q
            ws.cell(r, c).border = Border()
    for c in range(2, 18):
        ws.cell(1, c).border = Border(bottom=thin_black)
        ws.cell(2, c).border = Border(bottom=thin_black)

    # ==============================
    # 2) 图例区（统一顶部对齐）
    # ==============================
    legend_top = 2

    # 异常图例（放月表右边）
    anomaly_col = total_col + 2
    ws.merge_cells(start_row=legend_top + 2, start_column=anomaly_col, end_row=legend_top + 2, end_column=anomaly_col + 1)
    ws.merge_cells(start_row=legend_top + 3, start_column=anomaly_col, end_row=legend_top + 3, end_column=anomaly_col + 1)

    set_cell(
        ws.cell(legend_top + 2, anomaly_col),
        "报名没签到",
        fill=PatternFill("solid", fgColor=ANOMALY_COLORS["报名没签到"]),
        bold=True
    )
    set_cell(
        ws.cell(legend_top + 3, anomaly_col),
        "签到没报名",
        fill=PatternFill("solid", fgColor=ANOMALY_COLORS["签到没报名"]),
        bold=True
    )
    ws.cell(legend_top + 2, anomaly_col).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(legend_top + 3, anomaly_col).alignment = Alignment(horizontal="center", vertical="center")

    # 岗位图例
    legend_col = total_col + 6
    combo_start = legend_col + 2
    anomaly_col = legend_col - 4

    # 隐藏中间空列
    for col in range(total_col + 1, legend_col):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    # ==============================
    # 2B) 右侧图例区：岗位 + 组合 + 本月特别日
    # 不使用 merge，避免边框断线
    # ==============================

    role_order = ROLE_ORDER[:]
    combo_labels = [
        "值班 & 卫生",
        "值班 & 供花",
        "值班 & 供果",
        "值班 & 佛台",
        "值班 & 膳食",
        "多岗位同日"
    ]

    # 图例颜色
    weekend_fill = PatternFill("solid", fgColor="F2F2F2")
    lunar_fill = PatternFill("solid", fgColor="FFF2CC")
    buddha_fill = PatternFill("solid", fgColor="F4CCCC")
    guanyin_fill = PatternFill("solid", fgColor="EAD1DC")
    special_title_fill = PatternFill("solid", fgColor="D9EAD3")
    master_fill = PatternFill("solid", fgColor="D9D2E9")  # 淡紫蓝，可换

    def draw_legend_row(row, col, text, fill, font_color="000000", font_size=11):
        cell = ws.cell(row, col)
        cell.value = text
        cell.fill = fill
        cell.border = border_black
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="微软雅黑", size=font_size, bold=True, color=font_color)

        main_cell = ws.cell(row, col)
        main_cell.value = text
        main_cell.font = Font(
            name="微软雅黑",
            size=font_size,
            bold=True,
            color=font_color
        )

    # ==============================
    # 右侧图例区：岗位 + 组合岗位 + 本月特别日
    # ==============================

    role_order = ROLE_ORDER[:]
    combo_labels = [
        "值班 & 卫生",
        "值班 & 供花",
        "值班 & 供果",
        "值班 & 佛台",
        "值班 & 膳食",
        "多岗位同日"
    ]

    # ===== 左边岗位图例：不 merge =====
    for i, role in enumerate(role_order):
        r = legend_top + i
        cell = ws.cell(r, legend_col)

        set_cell(
            cell,
            role,
            fill=PatternFill("solid", fgColor=ROLE_COLORS[role]),
            bold=True
        )

        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="微软雅黑", size=11, bold=True)
        cell.border = border_black


    # ===== 右边组合岗位图例：不 merge =====
    for i, label in enumerate(combo_labels):
        r = legend_top + i
        cell = ws.cell(r, combo_start)

        set_cell(
            cell,
            label,
            fill=PatternFill("solid", fgColor=MULTI_ROLE_COLOR),
            bold=True
        )

        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="微软雅黑", size=11, bold=True)
        cell.border = border_black


    # =====================
    # 本月特别日：merge 两格，比较好看
    # =====================
    special_start_row = legend_top + len(role_order) + 2

    weekend_fill = PatternFill("solid", fgColor="F2F2F2")
    lunar_fill = PatternFill("solid", fgColor="FFF2CC")
    buddha_fill = PatternFill("solid", fgColor="F4CCCC")
    guanyin_fill = PatternFill("solid", fgColor="EAD1DC")
    special_title_fill = PatternFill("solid", fgColor="D9EAD3")


    def draw_special_row(row, text, fill_color, font_color="000000", font_size=11):
        ws.merge_cells(
            start_row=row,
            start_column=legend_col,
            end_row=row,
            end_column=legend_col + 1
        )

        cell = ws.cell(row, legend_col)

        set_cell(
            cell,
            text,
            fill=fill_color,
            bold=True
        )

        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(
            name="微软雅黑",
            size=font_size,
            bold=True,
            color=font_color
        )

        # 🔥 关键：补整行边框
        for c in range(legend_col, legend_col + 2):
            ws.cell(row, c).border = border_black


    # 标题
    draw_special_row(
        special_start_row,
        "本月特别日",
        special_title_fill,
        font_color="000000",
        font_size=11
    )

    # 普通特别日
    normal_items = [
        ("星期六", weekend_fill, "000000"),
        ("星期日", weekend_fill, "000000"),
        ("初一", lunar_fill, "000000"),
        ("十五", lunar_fill, "000000"),
    ]

    for i, (label, fill_color, font_color) in enumerate(normal_items, start=1):
        draw_special_row(
            special_start_row + i,
            label,
            fill_color,
            font_color=font_color,
            font_size=11
        )


    # ===== 动态大日子 =====
    highlight_items = []

    for day in range(1, days_in_month + 1):
        date_obj = datetime.date(year, month, day)

        _, _, special_name, _ = get_lunar_info(date_obj)

        if special_name:
            if "恩师" in special_name or "卢军宏" in special_name:
                highlight_items.append((special_name, master_fill, "351C75"))

            elif "观世音" in special_name or "观音" in special_name:
                highlight_items.append((special_name, guanyin_fill, "800080"))

            elif "释迦" in special_name or "释迦牟尼" in special_name or "释迦摩尼" in special_name or "佛" in special_name:
                highlight_items.append((special_name, buddha_fill, "FF0000"))
                    # 去重
    highlight_items = list(dict.fromkeys(highlight_items))


    highlight_start_row = special_start_row + len(normal_items) + 1

    for i, (label, fill_color, font_color) in enumerate(highlight_items):
        r = highlight_start_row + i

        draw_special_row(
            r,
            label,
            fill_color,
            font_color=font_color,
            font_size=12
        )

        ws.row_dimensions[r].height = 24


    # ===== 图例列宽 =====
    ws.column_dimensions[get_column_letter(legend_col)].width = 18
    ws.column_dimensions[get_column_letter(legend_col + 1)].width = 3

    ws.column_dimensions[get_column_letter(combo_start)].width = 22
    ws.column_dimensions[get_column_letter(combo_start + 1)].width = 3

    # ==============================
    # 3) 项目栏
    # ==============================
    title_row = 9
    for col in range(1, total_col + 1):
        cell = ws.cell(title_row, col)
        cell.fill = HEADER_FILL
        cell.border = border_black
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=title_row, start_column=3, end_row=title_row, end_column=5)
    set_cell(ws.cell(title_row, 3), "项目：", bold=True, size=16)
    ws.cell(title_row, 3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(title_row, 3).font = Font(name="微软雅黑", size=16, bold=True)

    ws.merge_cells(start_row=title_row, start_column=6, end_row=title_row, end_column=total_col - 1)
    set_cell(ws.cell(title_row, 6), f"{year}年{month}月份", bold=True, size=18)
    ws.cell(title_row, 6).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(title_row, 6).font = Font(name="微软雅黑", size=18, bold=True)

    ws.row_dimensions[title_row].height = 26

    # ==============================
    # 4) 表头
    # ==============================
    header_row = 10
    start_row = 11
    headers = ["No.", PROJECT_NAME] + list(range(1, days_in_month + 1)) + ["Total"]

    for c, h in enumerate(headers, start=1):
        fill = TOTAL_FILL if h == "Total" else HEADER_FILL
        set_cell(ws.cell(header_row, c), h, fill=fill, bold=True, size=13)
        ws.cell(header_row, c).font = Font(name="微软雅黑", size=13, bold=True)
        ws.cell(header_row, c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(header_row, c).border = border_black

    # ==============================
    # 5) 数据区
    # ==============================
    for idx, name in enumerate(all_names, start=1):
        r = start_row + idx - 1
        set_cell(ws.cell(r, 1), no_map[name])
        set_cell(ws.cell(r, 2), name, fill=NAME_FILL, center=False)

        ws.cell(r, 1).border = border_gray
        ws.cell(r, 2).border = border_gray
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")

        total_days = 0

        for day in range(1, days_in_month + 1):
            c = day + 2
            info = daily_info[name].get(day, {})
            roles = info.get("roles", [])
            signup = info.get("signup", 0)
            checkin = info.get("checkin", 0)
            notes = info.get("notes", [])

            cell = ws.cell(r, c)
            unique_roles = [x for x in dict.fromkeys(roles) if x]
            comment_text = ""

            if signup == 1 and checkin == 0:
                set_cell(cell, "", fill=PatternFill("solid", fgColor=ANOMALY_COLORS["报名没签到"]))
                comment_text = f"报名没签到\n岗位：{role_combo_label(unique_roles)}"

            elif signup == 0 and checkin == 1:
                set_cell(cell, 1, fill=PatternFill("solid", fgColor=ANOMALY_COLORS["签到没报名"]), bold=True)
                total_days += 1
                comment_text = f"签到没报名\n岗位：{role_combo_label(unique_roles)}"

            elif checkin == 1:
                total_days += 1
                if len(unique_roles) == 1:
                    fill = PatternFill("solid", fgColor=ROLE_COLORS.get(unique_roles[0], "FFFFFF"))
                elif len(unique_roles) > 1:
                    fill = PatternFill("solid", fgColor=MULTI_ROLE_COLOR)
                else:
                    fill = WHITE_FILL

                set_cell(cell, 1, fill=fill, bold=True)
                comment_text = f"签到\n岗位：{role_combo_label(unique_roles)}"

            else:
                set_cell(cell, "", fill=WHITE_FILL)

            if notes:
                comment_text += "\n备注：" + "；".join(notes)

            if comment_text:
                cell.comment = Comment(comment_text, "ChatGPT")

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_gray

        set_cell(ws.cell(r, total_col), total_days, fill=TOTAL_FILL, bold=True)
        ws.cell(r, total_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, total_col).border = border_gray

    # ==============================
    # 6) Total 行
    # ==============================
    total_row = start_row + len(all_names)
    set_cell(ws.cell(total_row, 1), "")
    set_cell(ws.cell(total_row, 2), "Total", fill=HEADER_FILL, bold=True, center=False)

    for day in range(1, days_in_month + 1):
        set_cell(ws.cell(total_row, day + 2), daily_totals[day - 1], fill=HEADER_FILL, bold=True)

    set_cell(ws.cell(total_row, total_col), month_grand_total, fill=TOTAL_FILL, bold=True)

    for col in range(1, total_col + 1):
        ws.cell(total_row, col).font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(total_row, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(total_row, col).border = border_black

    for r in range(header_row, total_row + 1):
        ws.row_dimensions[r].height = 20

    # ==============================
    # 7) 冻结与列宽
    # ==============================
    ws.freeze_panes = "C11"

    for col in range(1, total_col + 1):
        if col == 1:
            width = 6
        elif col == 2:
            width = 16
        elif col == total_col:
            width = 8
        else:
            width = 4.5
        ws.column_dimensions[get_column_letter(col)].width = width

    # 图例专属列宽
    ws.column_dimensions[get_column_letter(anomaly_col)].width = 12
    ws.column_dimensions[get_column_letter(anomaly_col + 1)].width = 5
    ws.column_dimensions[get_column_letter(legend_col)].width = 12
    ws.column_dimensions[get_column_letter(legend_col + 1)].width = 4
    ws.column_dimensions[get_column_letter(combo_start)].width = 22
    ws.column_dimensions[get_column_letter(combo_start + 1)].width = 4

    # 月表日期格：星期 / 农历 / 大日子批注 + 上色
    add_month_table_date_marks(ws, header_row=10, start_day_col=3, year=year, month=month, days_in_month=days_in_month)

    # ==============================
    # 8) 月明细
    # ==============================
    ws2 = wb.create_sheet("月明细")
    detail_headers = ["日期", "姓名", "报名", "签到", "岗位", "开始时间", "结束时间", "时数", "备注"]

    # 表头
    for c, h in enumerate(detail_headers, start=1):
        fill = TOTAL_FILL if h == "时数" else HEADER_FILL
        set_cell(ws2.cell(1, c), h, fill=fill, bold=True)
        ws2.cell(1, c).font = Font(name="微软雅黑", size=13, bold=True)
        ws2.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(1, c).border = border_black

    # 内容
    for r, (_, row) in enumerate(month_df.iterrows(), start=2):
        values = [
            row["日期"].strftime("%Y-%m-%d") if pd.notna(row.get("日期")) else "",
            row.get("姓名", ""),
            row.get("报名", ""),
            row.get("签到", ""),
            row.get("岗位", ""),
            to_ampm(row.get("开始时间", "")),
            to_ampm(row.get("结束时间", "")),
            row.get("时数", 0),
            row.get("备注", ""),
        ]
        for c, v in enumerate(values, start=1):
            set_cell(ws2.cell(r, c), v)
            ws2.cell(r, c).font = Font(name="微软雅黑", size=12)
            ws2.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
            ws2.cell(r, c).border = border_gray

    # Total 行
    total_row_ws2 = ws2.max_row + 1
    set_cell(ws2.cell(total_row_ws2, 1), "Total", fill=HEADER_FILL, bold=True)
    ws2.cell(total_row_ws2, 1).font = Font(name="微软雅黑", size=12, bold=True)
    ws2.cell(total_row_ws2, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(total_row_ws2, 1).border = border_black

    set_cell(ws2.cell(total_row_ws2, 2), f"共 {total_row_ws2 - 2} 条", fill=HEADER_FILL, bold=True)
    ws2.cell(total_row_ws2, 2).font = Font(name="微软雅黑", size=12, bold=True)
    ws2.cell(total_row_ws2, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(total_row_ws2, 2).border = border_black

    # 中间空白列也补边框
    for c in range(3, 8):
        ws2.cell(total_row_ws2, c).fill = HEADER_FILL
        ws2.cell(total_row_ws2, c).font = Font(name="微软雅黑", size=12, bold=True)
        ws2.cell(total_row_ws2, c).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(total_row_ws2, c).border = border_black

    # H列=第8列，时数总和
    set_cell(ws2.cell(total_row_ws2, 8), f"=SUM(H2:H{total_row_ws2 - 1})", fill=TOTAL_FILL, bold=True)
    ws2.cell(total_row_ws2, 8).font = Font(name="微软雅黑", size=12, bold=True)
    ws2.cell(total_row_ws2, 8).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(total_row_ws2, 8).border = border_black

    # 备注列
    ws2.cell(total_row_ws2, 9).fill = HEADER_FILL
    ws2.cell(total_row_ws2, 9).font = Font(name="微软雅黑", size=12, bold=True)
    ws2.cell(total_row_ws2, 9).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(total_row_ws2, 9).border = border_black

    # 行高放大
    ws2.row_dimensions[1].height = 24
    for r in range(2, ws2.max_row + 1):
        ws2.row_dimensions[r].height = 22

    # 列宽放大
    ws2.column_dimensions["A"].width = 14   # 日期
    ws2.column_dimensions["B"].width = 16   # 姓名
    ws2.column_dimensions["C"].width = 8    # 报名
    ws2.column_dimensions["D"].width = 8    # 签到
    ws2.column_dimensions["E"].width = 14   # 岗位
    ws2.column_dimensions["F"].width = 12   # 开始时间
    ws2.column_dimensions["G"].width = 12   # 结束时间
    ws2.column_dimensions["H"].width = 10   # 时数
    ws2.column_dimensions["I"].width = 22   # 备注

    # 冻结表头
    ws2.freeze_panes = "A2"

    add_date_comments_and_colors(ws2, date_col_idx=1, start_row=2)
    
    # ==============================
    # 9) 当月岗位统计
    # ==============================
    ws4 = wb.create_sheet("当月岗位统计")
    role_order_month = ["值班天数", "卫生", "佛台", "供花", "供台", "供果", "膳食", "佛学班"]

    if month_df.empty:
        set_cell(ws4.cell(1, 1), "本月没有岗位统计数据", fill=HEADER_FILL, bold=True)
        ws4.cell(1, 1).font = Font(name="微软雅黑", size=13, bold=True)
        ws4.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws4.row_dimensions[1].height = 24

    else:
        stat_df = month_df.copy()
        stat_df = stat_df[stat_df["签到"] == 1].copy()
        stat_df["岗位"] = stat_df["岗位"].apply(normalize_role)
        stat_df = stat_df[stat_df["岗位"].notna()].copy()
        stat_df["日期"] = pd.to_datetime(stat_df["日期"], errors="coerce").dt.strftime("%Y-%m-%d")

        stat_df = stat_df.drop_duplicates(
            subset=["日期", "姓名", "岗位"],
            keep="first"
        ).copy()

        if stat_df.empty:
            set_cell(ws4.cell(1, 1), "本月没有可统计的岗位数据", fill=HEADER_FILL, bold=True)
            ws4.cell(1, 1).font = Font(name="微软雅黑", size=13, bold=True)
            ws4.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws4.row_dimensions[1].height = 24

        else:
            pivot_df = pd.pivot_table(
                stat_df,
                index="姓名",
                columns="岗位",
                values="签到",
                aggfunc="sum",
                fill_value=0
            ).reset_index()

            pivot_df.rename(columns={"值班": "值班天数"}, inplace=True)

            for role in role_order_month:
                if role not in pivot_df.columns:
                    pivot_df[role] = 0

            pivot_df = pivot_df[["姓名"] + role_order_month]
            pivot_df["合计"] = pivot_df[role_order_month].sum(axis=1)
            pivot_df = pivot_df.sort_values(["合计", "姓名"], ascending=[False, True]).reset_index(drop=True)

            # 标题
            ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(pivot_df.columns))
            set_cell(ws4.cell(1, 1), f"{year}年{month}月岗位统计", fill=HEADER_FILL, bold=True, size=15)
            ws4.cell(1, 1).font = Font(name="微软雅黑", size=15, bold=True)
            ws4.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws4.cell(1, 1).border = border_black
            ws4.row_dimensions[1].height = 26

            # 表头
            for c, h in enumerate(pivot_df.columns.tolist(), start=1):
                fill = TOTAL_FILL if h == "合计" else HEADER_FILL
                set_cell(ws4.cell(3, c), h, fill=fill, bold=True)
                ws4.cell(3, c).font = Font(name="微软雅黑", size=13, bold=True)
                ws4.cell(3, c).alignment = Alignment(horizontal="center", vertical="center")
                ws4.cell(3, c).border = border_black

            # 内容
            for r, (_, row) in enumerate(pivot_df.iterrows(), start=4):
                for c, h in enumerate(pivot_df.columns.tolist(), start=1):
                    fill = NAME_FILL if h == "姓名" else (TOTAL_FILL if h == "合计" else None)
                    set_cell(ws4.cell(r, c), row[h], fill=fill)
                    ws4.cell(r, c).font = Font(name="微软雅黑", size=12)
                    ws4.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                    ws4.cell(r, c).border = border_gray

            # Total 行
            total_row_ws4 = ws4.max_row + 1
            set_cell(ws4.cell(total_row_ws4, 1), "Total", fill=HEADER_FILL, bold=True)
            ws4.cell(total_row_ws4, 1).font = Font(name="微软雅黑", size=12, bold=True)
            ws4.cell(total_row_ws4, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws4.cell(total_row_ws4, 1).border = border_black

            for col in range(2, ws4.max_column + 1):
                col_letter = get_column_letter(col)
                fill = TOTAL_FILL if ws4.cell(3, col).value == "合计" else HEADER_FILL
                set_cell(
                    ws4.cell(total_row_ws4, col),
                    f"=SUM({col_letter}4:{col_letter}{total_row_ws4 - 1})",
                    fill=fill,
                    bold=True
                )
                ws4.cell(total_row_ws4, col).font = Font(name="微软雅黑", size=12, bold=True)
                ws4.cell(total_row_ws4, col).alignment = Alignment(horizontal="center", vertical="center")
                ws4.cell(total_row_ws4, col).border = border_black

            # 冻结
            ws4.freeze_panes = "B4"

            # 列宽放大
            ws4.column_dimensions["A"].width = 18   # 姓名
            for col in range(2, ws4.max_column):
                ws4.column_dimensions[get_column_letter(col)].width = 12
            ws4.column_dimensions[get_column_letter(ws4.max_column)].width = 12  # 合计

            # 行高放大
            ws4.row_dimensions[3].height = 24
            for r in range(4, ws4.max_row + 1):
                ws4.row_dimensions[r].height = 22

        # ←←← 这里要退出来（对齐 ws4 = wb.create_sheet）
        file_name = os.path.basename(out_file)
        
        # 🔥 最后才保存
        file_name = os.path.basename(out_file)

        # ==============================
        # 📖 白话佛法共修统计（升级版）
        # ==============================
        try:
            reading_df = pd.read_excel("reading.xlsx")

            required_cols = ["日期", "姓名", "主题", "场次", "时间"]
            for col in required_cols:
                if col not in reading_df.columns:
                    reading_df[col] = ""

            reading_df["日期"] = pd.to_datetime(reading_df["日期"], errors="coerce")

            reading_month = reading_df[
                (reading_df["日期"].dt.year == year) &
                (reading_df["日期"].dt.month == month)
            ].copy()

            if not reading_month.empty:
                reading_month = reading_month.sort_values(["姓名", "日期", "时间"])

                # 每个人最后一次记录
                last_rows = (
                    reading_month.sort_values(["姓名", "日期", "时间"])
                    .groupby("姓名", as_index=False)
                    .tail(1)
                )

                # 每人次数
                count_df = (
                    reading_month.groupby("姓名")
                    .size()
                    .reset_index(name="本月共修次数")
                )

                person_stats = count_df.merge(
                    last_rows[["姓名", "日期", "主题"]],
                    on="姓名",
                    how="left"
                )

                person_stats.rename(
                    columns={
                        "日期": "最近一次共修",
                        "主题": "最后主题"
                    },
                    inplace=True
                )

                person_stats["最近一次共修"] = person_stats["最近一次共修"].dt.strftime("%Y-%m-%d")

                # 排序：次数多的在上面
                person_stats = person_stats.sort_values(
                    ["本月共修次数", "最近一次共修", "姓名"],
                    ascending=[False, False, True]
                ).reset_index(drop=True)

                ws5 = wb.create_sheet("白话佛法统计")

                headers = ["姓名", "本月共修次数", "最近一次共修", "最后主题"]

                # 标题
                ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                set_cell(ws5.cell(1, 1), f"{year}年{month}月白话佛法共修统计", fill=HEADER_FILL, bold=True, size=15)
                ws5.cell(1, 1).font = Font(name="微软雅黑", size=15, bold=True)
                ws5.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
                ws5.cell(1, 1).border = border_black
                ws5.row_dimensions[1].height = 28

                # 表头
                for c, h in enumerate(headers, start=1):
                    fill = TOTAL_FILL if h == "本月共修次数" else HEADER_FILL
                    set_cell(ws5.cell(3, c), h, fill=fill, bold=True)
                    ws5.cell(3, c).font = Font(name="微软雅黑", size=13, bold=True)
                    ws5.cell(3, c).alignment = Alignment(horizontal="center", vertical="center")
                    ws5.cell(3, c).border = border_black

                # 内容
                for r, (_, row) in enumerate(person_stats.iterrows(), start=4):
                    values = [
                        row["姓名"],
                        row["本月共修次数"],
                        row["最近一次共修"],
                        row["最后主题"],
                    ]

                    for c, v in enumerate(values, start=1):
                        fill = None

                        # 温和提醒：本月只共修1次，浅红
                        if headers[c - 1] == "本月共修次数" and int(v) <= 1:
                            fill = PatternFill("solid", fgColor="F4CCCC")

                        # 次数较多，浅绿鼓励
                        if headers[c - 1] == "本月共修次数" and int(v) >= 5:
                            fill = PatternFill("solid", fgColor="D9EAD3")

                        set_cell(ws5.cell(r, c), v, fill=fill)
                        ws5.cell(r, c).font = Font(name="微软雅黑", size=12)
                        ws5.cell(r, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws5.cell(r, c).border = border_gray

                # Total 行
                total_row = ws5.max_row + 1
                set_cell(ws5.cell(total_row, 1), "Total", fill=HEADER_FILL, bold=True)
                ws5.cell(total_row, 1).font = Font(name="微软雅黑", size=12, bold=True)
                ws5.cell(total_row, 1).alignment = Alignment(horizontal="center", vertical="center")
                ws5.cell(total_row, 1).border = border_black

                set_cell(ws5.cell(total_row, 2), f"=SUM(B4:B{total_row - 1})", fill=TOTAL_FILL, bold=True)
                ws5.cell(total_row, 2).font = Font(name="微软雅黑", size=12, bold=True)
                ws5.cell(total_row, 2).alignment = Alignment(horizontal="center", vertical="center")
                ws5.cell(total_row, 2).border = border_black

                for c in range(3, 5):
                    ws5.cell(total_row, c).fill = HEADER_FILL
                    ws5.cell(total_row, c).border = border_black

                # 列宽
                ws5.column_dimensions["A"].width = 20
                ws5.column_dimensions["B"].width = 18
                ws5.column_dimensions["C"].width = 20
                ws5.column_dimensions["D"].width = 45

                # 行高
                ws5.row_dimensions[3].height = 24
                for r in range(4, ws5.max_row + 1):
                    ws5.row_dimensions[r].height = 26

                # 冻结 + 筛选功能
                ws5.freeze_panes = "A4"
                ws5.auto_filter.ref = f"A3:D{ws5.max_row}"

        except Exception as e:
            print("❌ 白话佛法统计失败：", e)

        debug(f"准备保存月报：{out_file}")
        wb.save(out_file)
        log(f"   ✔ 已保存：{file_name}")

        return out_file
                
def write_year_summary_sheet(wb, summary_table, year):
    ws = wb.create_sheet("全年值班表")

    summary_table = summary_table.copy()

    # ========= 状态列处理 =========
    if "状态" not in summary_table.columns:
        summary_table["状态"] = "在册"

    # 状态排序：在册 -> 暂停/失联/退出 -> 往生
    status_order = {
        "在册": 1,
        "暂停": 2,
        "失联": 3,
        "退出": 4,
        "往生": 5,
    }
    summary_table["_status_order"] = summary_table["状态"].map(status_order).fillna(9)

    # 如果有 Total，就先按状态，再按 Total 高到低，再按姓名
    sort_cols = ["_status_order"]
    ascending_list = [True]

    if "Total" in summary_table.columns:
        sort_cols.append("Total")
        ascending_list.append(False)

    if "姓名" in summary_table.columns:
        sort_cols.append("姓名")
        ascending_list.append(True)

    summary_table = summary_table.sort_values(
        by=sort_cols,
        ascending=ascending_list
    ).reset_index(drop=True)

    headers = ["Cheras义工", "电话号码"] + [f"{m}月" for m in range(1, 13)] + ["Total", "开始值班日期"]

    # ===== 通用样式 =====
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    body_font = Font(name="微软雅黑", size=11)
    gray_font = Font(name="微软雅黑", size=11, color="999999")
    passed_font = Font(name="微软雅黑", size=11, color="BBBBBB")

    header_font = Font(name="微软雅黑", size=12, bold=True)
    title_font = Font(name="微软雅黑", size=16, bold=True)

    name_font = Font(name="微软雅黑", size=13, bold=True)
    name_gray_font = Font(name="微软雅黑", size=13, bold=True, color="999999")
    name_passed_font = Font(name="微软雅黑", size=13, bold=True, color="BBBBBB")

    phone_font = Font(name="微软雅黑", size=11)
    phone_gray_font = Font(name="微软雅黑", size=11, color="999999")
    phone_passed_font = Font(name="微软雅黑", size=11, color="BBBBBB")

    total_font = Font(name="微软雅黑", size=11, bold=True)
    total_gray_font = Font(name="微软雅黑", size=11, bold=True, color="999999")
    total_passed_font = Font(name="微软雅黑", size=11, bold=True, color="BBBBBB")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ===== 标题 =====
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1)
    set_cell(title_cell, f"{year}年 全年(义工值班表)", fill=HEADER_FILL, bold=True, size=16)
    title_cell.font = title_font
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # ===== 表头 =====
    for c, h in enumerate(headers, start=1):
        fill = TOTAL_FILL if h == "Total" else HEADER_FILL
        set_cell(ws.cell(3, c), h, fill=fill, bold=True, size=12)

    # ===== 数据 =====
    for r, (_, row) in enumerate(summary_table.iterrows(), start=4):
        ws.row_dimensions[r].height = 28

        for c, h in enumerate(headers, start=1):
            source_col = "姓名" if h == "Cheras义工" else h

            fill = None
            if h in ["Cheras义工", "电话号码", "开始值班日期"]:
                fill = NAME_FILL
            elif h == "Total":
                fill = TOTAL_FILL

            set_cell(ws.cell(r, c), row.get(source_col, ""), fill=fill)

    max_row = ws.max_row
    max_col = ws.max_column

    # ===== 全表基础样式 =====
    for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = center_align

    # ===== 表头强化 =====
    for c in range(1, max_col + 1):
        cell = ws.cell(3, c)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # ===== 特定列样式 + 状态分色 =====
    for r in range(4, max_row + 1):
        row_status = str(summary_table.iloc[r - 4].get("状态", "在册")).strip()

        # 默认
        row_body_font = body_font
        row_name_font = name_font
        row_phone_font = phone_font
        row_total_font = total_font

        # 非在册灰色
        if row_status in ["暂停", "失联", "退出"]:
            row_body_font = gray_font
            row_name_font = name_gray_font
            row_phone_font = phone_gray_font
            row_total_font = total_gray_font

        # 往生更淡
        elif row_status == "往生":
            row_body_font = passed_font
            row_name_font = name_passed_font
            row_phone_font = phone_passed_font
            row_total_font = total_passed_font

        # 整列字体
        for c in range(1, max_col + 1):
            ws.cell(r, c).font = row_body_font
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = center_align

        # A列：Cheras义工
        ws.cell(r, 1).alignment = left_align
        ws.cell(r, 1).font = row_name_font

        # B列：电话号码
        ws.cell(r, 2).alignment = left_align
        ws.cell(r, 2).font = row_phone_font

        # O列：Total
        ws.cell(r, 15).font = row_total_font

        # P列：开始值班日期
        ws.cell(r, 16).alignment = left_align
        ws.cell(r, 16).font = row_body_font

    # ===== 冻结窗格 =====
    ws.freeze_panes = "C4"

    # ===== 自动筛选（你要的下拉箭头）=====
    if max_row >= 3 and max_col >= 1:
        ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{max_row}"

    # ===== 列宽 =====
    ws.column_dimensions["A"].width = 20   # Cheras义工
    ws.column_dimensions["B"].width = 18   # 电话号码
    for col in range(3, 15):  # C ~ N（1月~12月）
        ws.column_dimensions[get_column_letter(col)].width = 8
    ws.column_dimensions["O"].width = 10   # Total
    ws.column_dimensions["P"].width = 16   # 开始值班日期

    # ===== 自动筛色：月份区 C:N =====
    if max_row >= 4:
        ws.conditional_formatting.add(
            f"C4:N{max_row}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FFFFFF",
                mid_type="num", mid_value=1, mid_color="FFF2CC",
                end_type="max", end_color="63BE7B"
            )
        )

        # ===== 自动筛色：Total 列 O =====
        ws.conditional_formatting.add(
            f"O4:O{max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )
        )

    # 清理临时列
    if "_status_order" in summary_table.columns:
        summary_table.drop(columns=["_status_order"], inplace=True, errors="ignore")
# =========================
# 写年报
# =========================
def write_yearly_report(summary_df, role_df, detail_df, year, summary_table=None):
    ensure_output_dir(OUTPUT_DIR)
    out_file = os.path.join(OUTPUT_DIR, f"{year}_年报.xlsx")

    wb = Workbook()

    # ===== 通用样式 =====
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    body_font = Font(name="微软雅黑", size=11)
    body_font_gray = Font(name="微软雅黑", size=11, color="999999")
    header_font = Font(name="微软雅黑", size=12, bold=True)
    title_font = Font(name="微软雅黑", size=16, bold=True)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    def get_display_width(s):
        """中文按2格，英文按1格估算宽度"""
        width = 0
        for ch in str(s):
            if ord(ch) > 127:
                width += 2
            else:
                width += 1
        return width

    def auto_fit_columns_better(ws, min_width=8, max_width=30):
        for col_cells in ws.columns:
            col_idx = col_cells[0].column
            col_letter = get_column_letter(col_idx)
            max_len = 0

            for cell in col_cells:
                try:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, get_display_width(val))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max(min(max_len + 4, max_width), min_width)

    def beautify_sheet(
        ws,
        header_row,
        data_start_row,
        freeze_cell=None,
        name_col_idx=None,
        gray_body=False
    ):
        max_row = ws.max_row
        max_col = ws.max_column

        # 表头
        for c in range(1, max_col + 1):
            cell = ws.cell(header_row, c)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # 内容
        for r in range(data_start_row, max_row + 1):
            ws.row_dimensions[r].height = 26
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.font = body_font_gray if gray_body else body_font
                cell.border = border
                cell.alignment = center_align

        # 姓名列靠左并更突出
        if name_col_idx is not None:
            for r in range(data_start_row, max_row + 1):
                ws.cell(r, name_col_idx).alignment = left_align
                if gray_body:
                    ws.cell(r, name_col_idx).font = Font(name="微软雅黑", size=12, bold=True, color="999999")
                else:
                    ws.cell(r, name_col_idx).font = Font(name="微软雅黑", size=12, bold=True)

        # 冻结
        if freeze_cell:
            ws.freeze_panes = freeze_cell

        if max_row >= header_row and max_col >= 1:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"

        auto_fit_columns_better(ws)

        if name_col_idx is not None:
            ws.column_dimensions[get_column_letter(name_col_idx)].width = 20

        ws.row_dimensions[1].height = 30

    def write_summary_sheet(ws, df, title_text, gray_body=False):
        end_col = max(1, len(df.columns))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

        title_cell = ws.cell(1, 1)
        set_cell(title_cell, title_text, fill=HEADER_FILL, bold=True, size=16)
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 32

        # 空表处理
        if df.empty:
            for c, h in enumerate(df.columns.tolist(), start=1):
                set_cell(ws.cell(3, c), h, fill=HEADER_FILL, bold=True)
            set_cell(ws.cell(4, 1), "没有数据")
            beautify_sheet(
                ws,
                header_row=3,
                data_start_row=4,
                freeze_cell="C4" if len(df.columns) >= 3 else "A4",
                name_col_idx=(df.columns.get_loc("Cheras义工") + 1) if "Cheras义工" in df.columns else None,
                gray_body=gray_body
            )
            return

        # 表头
        for c, h in enumerate(df.columns.tolist(), start=1):
            set_cell(ws.cell(3, c), h, fill=HEADER_FILL, bold=True)

        # 内容
        for r, (_, row) in enumerate(df.iterrows(), start=4):
            for c, h in enumerate(df.columns.tolist(), start=1):
                val = row[h]
                fill = NAME_FILL if h == "Cheras义工" else None
                set_cell(ws.cell(r, c), val, fill=fill)

        name_col_idx = df.columns.get_loc("Cheras义工") + 1 if "Cheras义工" in df.columns else None
        freeze_cell = "C4" if len(df.columns) >= 3 else "A4"

        beautify_sheet(
            ws,
            header_row=3,
            data_start_row=4,
            freeze_cell=freeze_cell,
            name_col_idx=name_col_idx,
            gray_body=gray_body
        )

        # 出席率颜色渐变
        if "出席率" in df.columns and not df.empty and not gray_body:
            rate_col_idx = df.columns.get_loc("出席率") + 1
            rate_col_letter = get_column_letter(rate_col_idx)
            ws.conditional_formatting.add(
                f"{rate_col_letter}4:{rate_col_letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B"
                )
            )

    # =========================
    # 先复制，避免改到外部 DataFrame
    # =========================
    summary_df = summary_df.copy()
    role_df = role_df.copy()
    detail_df = detail_df.copy()

    # 如果没有状态列，自动补
    if "状态" not in summary_df.columns:
        summary_df["状态"] = "在册"

    # 先分类（要在 rename 前做）
    # 主表（只在册）
    main_df = summary_df[
        summary_df["状态"] == "在册"
    ].copy()

    # 未活跃（全部非在册 + 非往生）
    inactive_df = summary_df[
        (summary_df["状态"] != "在册") &
        (summary_df["状态"] != "往生")
    ].copy()

    # 往生
    passed_df = summary_df[
        summary_df["状态"] == "往生"
    ].copy()

    # 排序
    if "全年出席天数" in main_df.columns:
        if "全年总时数" in main_df.columns:
            main_df = main_df.sort_values(by=["全年出席天数", "全年总时数"], ascending=[False, False])
        else:
            main_df = main_df.sort_values(by="全年出席天数", ascending=False)

    if not inactive_df.empty:
        sort_cols = [c for c in ["状态", "编号"] if c in inactive_df.columns]
        if sort_cols:
            inactive_df = inactive_df.sort_values(by=sort_cols)

    if not passed_df.empty and "编号" in passed_df.columns:
        passed_df = passed_df.sort_values(by="编号")

    # 再 rename（只影响显示）
    def rename_summary_columns(df):
        df = df.copy()
        return df.rename(columns={
            "姓名": "Cheras义工",
            "电话号码": "联络号码"
        })

    summary_df = rename_summary_columns(summary_df)
    main_df = rename_summary_columns(main_df)
    inactive_df = rename_summary_columns(inactive_df)
    passed_df = rename_summary_columns(passed_df)

    # ===== 1) 在册义工 =====
    ws = wb.active
    ws.title = "在册义工"
    write_summary_sheet(ws, main_df, f"{year}年在册义工年度报表", gray_body=False)

    # ===== 2) 未活跃义工 =====
    ws_inactive = wb.create_sheet("未活跃义工")
    write_summary_sheet(ws_inactive, inactive_df, f"{year}年未活跃义工名单", gray_body=True)

    # ===== 3) 往生义工 =====
    ws_passed = wb.create_sheet("往生义工")
    write_summary_sheet(ws_passed, passed_df, f"{year}年往生义工名单", gray_body=False)

    # ===== 4) 岗位统计 =====
    ws2 = wb.create_sheet("岗位统计")

    if role_df.empty:
        set_cell(ws2.cell(1, 1), "没有岗位统计数据", fill=HEADER_FILL, bold=True)
        ws2.cell(1, 1).font = title_font
        ws2.cell(1, 1).alignment = center_align
        ws2.row_dimensions[1].height = 28
        ws2.column_dimensions["A"].width = 22
    else:
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(role_df.columns))
        t2 = ws2.cell(1, 1)
        set_cell(t2, f"{year}年岗位统计", fill=HEADER_FILL, bold=True, size=15)
        t2.font = Font(name="微软雅黑", size=15, bold=True)
        t2.alignment = center_align
        ws2.row_dimensions[1].height = 30

        for c, h in enumerate(role_df.columns.tolist(), start=1):
            set_cell(ws2.cell(3, c), h, fill=HEADER_FILL, bold=True)

        for r, (_, row) in enumerate(role_df.iterrows(), start=4):
            for c, h in enumerate(role_df.columns.tolist(), start=1):
                fill = NAME_FILL if h == "姓名" else None
                set_cell(ws2.cell(r, c), row[h], fill=fill)

        name_col_idx = role_df.columns.get_loc("姓名") + 1 if "姓名" in role_df.columns else None
        beautify_sheet(ws2, header_row=3, data_start_row=4, freeze_cell="C4", name_col_idx=name_col_idx)

    # ===== 5) 签到明细 =====
    ws3 = wb.create_sheet("签到明细")

    if detail_df.empty:
        set_cell(ws3.cell(1, 1), "没有签到明细", fill=HEADER_FILL, bold=True)
        ws3.cell(1, 1).font = title_font
        ws3.cell(1, 1).alignment = center_align
        ws3.row_dimensions[1].height = 28
        ws3.column_dimensions["A"].width = 20
    else:
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(detail_df.columns))
        t3 = ws3.cell(1, 1)
        set_cell(t3, f"{year}年签到明细", fill=HEADER_FILL, bold=True, size=15)
        t3.font = Font(name="微软雅黑", size=15, bold=True)
        t3.alignment = center_align
        ws3.row_dimensions[1].height = 30

        for c, h in enumerate(detail_df.columns.tolist(), start=1):
            set_cell(ws3.cell(3, c), h, fill=HEADER_FILL, bold=True)

        for r, (_, row) in enumerate(detail_df.iterrows(), start=4):
            for c, h in enumerate(detail_df.columns.tolist(), start=1):
                val = row[h]
                fill = NAME_FILL if h == "姓名" else None
                set_cell(ws3.cell(r, c), val, fill=fill)

        name_col_idx = detail_df.columns.get_loc("姓名") + 1 if "姓名" in detail_df.columns else None
        beautify_sheet(ws3, header_row=3, data_start_row=4, freeze_cell="C4", name_col_idx=name_col_idx)

    # ===== 6) 全年值班表 =====
    if summary_table is not None and not summary_table.empty:
        write_year_summary_sheet(wb, summary_table, year)

    # ⭐ 一定先初始化（关键）
    reading_year = pd.DataFrame()

    try:
        if os.path.exists(READING_FILE):
            reading_all = pd.read_excel(READING_FILE)

            # 确保有日期栏
            if "日期" in reading_all.columns:
                reading_all["日期"] = pd.to_datetime(reading_all["日期"], errors="coerce")

                reading_year = reading_all[
                    reading_all["日期"].dt.year == target_year
                ].copy()
            else:
                print("⚠️ reading.xlsx 没有【日期】栏位")

        else:
            print("⚠️ 找不到 reading.xlsx")

    except Exception as e:
        print("⚠️ 读取 reading.xlsx 失败：", e)
        reading_year = pd.DataFrame()

    try:
        print("📝 写入年报白话佛法...")

        # 如果已经有旧 sheet，先删除
        if "白话佛法年统计" in wb.sheetnames:
            del wb["白话佛法年统计"]

        ws6 = wb.create_sheet("白话佛法年统计")
        ws6.freeze_panes = "A2"

        headers = ["姓名", "身份", "全年共修次数", "最近一次共修", "最后主题"]

        for c, h in enumerate(headers, start=1):
            ws6.cell(1, c, h)

        # 保证栏位存在
        for col in ["日期", "姓名", "身份", "主题", "时间"]:
            if col not in reading_year.columns:
                reading_year[col] = ""

        if reading_year.empty:
            ws6.cell(2, 1, "本年暂无共修记录")

        else:
            reading_year = reading_year.copy()
            reading_year["日期"] = pd.to_datetime(reading_year["日期"], errors="coerce")
            reading_year = reading_year.dropna(subset=["日期"])

            reading_year["姓名"] = reading_year["姓名"].astype(str).str.strip()
            reading_year["身份"] = reading_year["身份"].astype(str).str.strip()
            reading_year["主题"] = reading_year["主题"].astype(str).str.strip()
            reading_year["时间"] = reading_year["时间"].astype(str).str.strip()

            reading_year = reading_year[reading_year["姓名"] != ""]

            if reading_year.empty:
                ws6.cell(2, 1, "本年暂无共修记录")
            else:
                reading_year = reading_year.sort_values(["姓名", "日期", "时间"])

                # 每人最后一次
                last_rows = (
                    reading_year
                    .groupby(["姓名", "身份"], as_index=False)
                    .tail(1)
                )

                # 次数：每一行 = 1次
                count_df = (
                    reading_year
                    .groupby(["姓名", "身份"])
                    .size()
                    .reset_index(name="全年共修次数")
                )

                year_stats = count_df.merge(
                    last_rows[["姓名", "身份", "日期", "主题"]],
                    on=["姓名", "身份"],
                    how="left"
                )

                year_stats.rename(
                    columns={
                        "日期": "最近一次共修",
                        "主题": "最后主题"
                    },
                    inplace=True
                )

                year_stats["最近一次共修"] = pd.to_datetime(
                    year_stats["最近一次共修"],
                    errors="coerce"
                ).dt.strftime("%Y-%m-%d")

                year_stats = year_stats.sort_values(
                    ["全年共修次数", "最近一次共修", "姓名"],
                    ascending=[False, False, True]
                ).reset_index(drop=True)

                for r, (_, row) in enumerate(year_stats.iterrows(), start=2):
                    ws6.cell(r, 1, row["姓名"])
                    ws6.cell(r, 2, row["身份"])
                    ws6.cell(r, 3, row["全年共修次数"])
                    ws6.cell(r, 4, row["最近一次共修"])
                    ws6.cell(r, 5, row["最后主题"])
                
                # ✅ 美化白话佛法年统计
                header_fill = PatternFill("solid", fgColor="D9EAD3")
                header_font = Font(name="Microsoft YaHei", size=14, bold=True)
                body_font = Font(name="Microsoft YaHei", size=12)
                thin = Side(style="thin", color="999999")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # 表头
                for cell in ws6[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                # 内容
                for row in ws6.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = body_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border

                # 最后主题靠左
                for row in range(2, ws6.max_row + 1):
                    ws6.cell(row, 5).alignment = Alignment(horizontal="left", vertical="center")

                # 栏宽
                widths = {
                    1: 18,   # 姓名
                    2: 12,   # 身份
                    3: 16,   # 全年共修次数
                    4: 16,   # 最近一次共修
                    5: 38,   # 最后主题
                }

                for col, width in widths.items():
                    ws6.column_dimensions[get_column_letter(col)].width = width

                # 行高
                ws6.row_dimensions[1].height = 28
                for r in range(2, ws6.max_row + 1):
                    ws6.row_dimensions[r].height = 24

                # 冻结表头
                ws6.freeze_panes = "A2"

                # 自动筛选
                ws6.auto_filter.ref = ws6.dimensions

    except Exception as e:
        print("❌ 年报白话佛法失败：", e)

    wb.save(out_file)
    return out_file

def to_ampm(t):
    import pandas as pd

    if pd.isna(t):
        return ""

    s = str(t).strip().lower()

    # 已经是 am/pm
    if "am" in s or "pm" in s:
        return s

    try:
        dt = pd.to_datetime(s)
        return dt.strftime("%I:%M%p").lstrip("0").lower()
    except:
        return s

def to_simplified(text):
    return cc.convert(str(text).strip())

def normalize_name_by_map(name, name_map):
    key = to_simplified(name)
    return name_map.get(key, str(name).strip())

def import_schedule_to_attendance(att_df, schedule_df):
    if schedule_df is None or schedule_df.empty:
        log("⚠️ 没有 schedule 数据需要导入")
        return att_df

    log(f"📥 导入 schedule 数据：{len(schedule_df)} 条")

    # 补齐字段（防止缺列）
    required_cols = ["日期", "姓名", "报名", "签到", "岗位", "开始时间", "结束时间", "时数", "备注"]
    for col in required_cols:
        if col not in schedule_df.columns:
            schedule_df[col] = ""

    schedule_df = schedule_df[required_cols].copy()

    # 合并
    new_att = pd.concat([att_df, schedule_df], ignore_index=True)

    # 去重
    new_att = new_att.drop_duplicates(
        subset=["日期", "姓名", "岗位", "开始时间", "结束时间"],
        keep="last"
    ).reset_index(drop=True)

    log(f"✔ 导入后 attendance 共 {len(new_att)} 条")

    return new_att
# =========================
# 主程序
# =========================
def main():
    # 🔥 优先用网页传进来的参数
    if REPORT_MONTH:
        year_str, month_str = REPORT_MONTH.split("-")
        target_year = int(year_str)
        target_month = int(month_str)

    elif REPORT_YEAR:
        target_year = int(REPORT_YEAR)
        target_month = None  # 年报

    else:
        # 👉 没传参数，就用今天日期（备用）
        from datetime import datetime
        today = datetime.today()
        target_year = today.year
        target_month = today.month

    # 👉 下面继续你原本逻辑
    log(f"\n📊 年度处理：{target_year}")
    log(f"\n📊 年度处理：{target_year}")

    if target_month is not None:
        log(f"📅 目标月份：{target_month:02d}\n")
    else:
        log("📅 目标月份：全年\n")

    # =====================
    # 1) 主名单
    # =====================
    log("👥 读取主名单...")
    vol_df = load_volunteers()
    vol_df = vol_df.copy()
    vol_df["姓名"] = vol_df["姓名"].astype(str).str.strip()

    # =====================
    # ✅ 只保留义工进入报表（关键）
    # =====================
    if "是否义工" in vol_df.columns:
        vol_df = vol_df[vol_df["是否义工"] == "是"].copy()

    name_map = {
        to_simplified(name): name
        for name in vol_df["姓名"]
    }

    # 写回并美化 volunteers.xlsx
    vol_df.to_excel(
        VOLUNTEERS_FILE,
        index=False,
        sheet_name=VOLUNTEERS_SHEET
    )
    beautify_volunteers_file(VOLUNTEERS_FILE, VOLUNTEERS_SHEET)

    log(f"✔ 共 {len(vol_df)} 位义工\n")

    # =====================
    # 2) 出席记录
    # =====================
    log("📖 读取出席记录...")
    att_df = load_attendance()
    att_df = clean_attendance_df(att_df)
    att_df = calc_hours(att_df)

    debug(f"清洗后签到唯一值：{att_df['签到'].dropna().unique().tolist() if '签到' in att_df.columns else []}")
    log(f"✔ 共 {len(att_df)} 条记录\n")

    if "姓名" in att_df.columns:
        att_df["姓名"] = att_df["姓名"].astype(str).str.strip().apply(lambda x: normalize_name_by_map(x, name_map))

    # =====================
    # 3) WhatsApp 排班
    # =====================
    log("📩 读取 WhatsApp 排班...")

    try:
        schedule_df = pd.read_excel(SCHEDULE_FILE)
    except Exception:
        schedule_df = pd.DataFrame(
            columns=["日期", "姓名", "报名", "签到", "岗位", "开始时间", "结束时间", "时数", "备注"]
        )
    log("⚠️ 没有 schedule 数据需要导入")

    schedule_df = clean_attendance_df(schedule_df)

    if "姓名" in schedule_df.columns:
        schedule_df["姓名"] = schedule_df["姓名"].astype(str).str.strip().apply(lambda x: normalize_name_by_map(x, name_map))

    if not schedule_df.empty:
        debug(f"schedule_df 报名值: {schedule_df['报名'].unique().tolist()}")
        debug(f"schedule_df 签到值: {schedule_df['签到'].unique().tolist()}")

    log(f"✔ 导入 {len(schedule_df)} 条排班\n")

    if not schedule_df.empty:
        unique_dates = sorted(schedule_df["日期"].dropna().dt.strftime("%Y-%m-%d").unique())
        debug(f"[排班日期] 共 {len(unique_dates)} 天：")
        for d in unique_dates:
            debug(f" - {d}")

    # =====================
    # 4) 合并 attendance + 排班 + 手动补录
    # =====================
    log("📦 合并 attendance + 排班 + 手动补录...")

    # =====================
    # 名字检查（非常重要）
    # =====================
    log("🔍 检查异常名字...")

    known_names = set(vol_df["姓名"].dropna().astype(str).str.strip())

    att_names = att_df["姓名"] if "姓名" in att_df.columns else pd.Series(dtype=object)
    sch_names = schedule_df["姓名"] if "姓名" in schedule_df.columns else pd.Series(dtype=object)

    all_names = pd.concat([att_names, sch_names], ignore_index=True)
    all_names = all_names.dropna().astype(str).str.strip()
    all_names = all_names[(all_names != "") & (all_names.str.lower() != "nan")]

    unknown_names = sorted(set(all_names) - known_names)
    
    if unknown_names:
        log(f"⚠️ 发现 {len(unknown_names)} 个不在主名单的名字：")
        for name in unknown_names:
            suggestion = get_close_matches(name, list(known_names), n=1)
            if suggestion:
                log(f"   - {name}（可能是：{suggestion[0]}）")
            else:
                log(f"   - {name}")
        log("")
    else:
        log("✔ 没有异常名字\n")

    # 过滤异常名字（不进入合并/报表）
    if "姓名" in att_df.columns:
        att_df["姓名"] = att_df["姓名"].where(att_df["姓名"].notna(), "")
        att_df["姓名"] = att_df["姓名"].astype(str).str.strip()
        att_df = att_df[
            (att_df["姓名"] != "") &
            (att_df["姓名"].str.lower() != "nan") &
            (att_df["姓名"].isin(known_names))
        ].copy()

    if "姓名" in schedule_df.columns:
        schedule_df["姓名"] = schedule_df["姓名"].where(schedule_df["姓名"].notna(), "")
        schedule_df["姓名"] = schedule_df["姓名"].astype(str).str.strip()
        schedule_df = schedule_df[
            (schedule_df["姓名"] != "") &
            (schedule_df["姓名"].str.lower() != "nan") &
            (schedule_df["姓名"].isin(known_names))
        ].copy()

    final_att_df = import_schedule_to_attendance(att_df, schedule_df)

    att_export = final_att_df.copy()
    att_export = calc_hours(att_export)

    att_export["日期"] = pd.to_datetime(att_export["日期"], errors="coerce")

    att_export = att_export.sort_values(
        by=["日期", "姓名", "开始时间", "结束时间"],
        ascending=[True, True, True, True]
    ).reset_index(drop=True)

    # 保存前备份
    backup_file_if_exists(ATTENDANCE_FILE, BACKUP_DIR)

    att_export_save = att_export.copy()
    att_export_save["日期"] = pd.to_datetime(
        att_export_save["日期"], errors="coerce"
    ).dt.strftime("%Y-%m-%d")

    att_export_save["开始时间"] = att_export_save["开始时间"].apply(to_ampm)
    att_export_save["结束时间"] = att_export_save["结束时间"].apply(to_ampm)

    att_export_save["时数"] = pd.to_numeric(
        att_export_save["时数"], errors="coerce"
    ).fillna(0).round(2)

    log("💾 写回 attendance.xlsx...")
    with pd.ExcelWriter(ATTENDANCE_FILE, engine="openpyxl") as writer:
        att_export_save.to_excel(writer, sheet_name=ATTENDANCE_SHEET, index=False)

        df_groups = groups_to_df(GROUPS)
        df_groups.to_excel(writer, sheet_name="sheet1", index=False)

    beautify_attendance_file(ATTENDANCE_FILE, ATTENDANCE_SHEET)
    beautify_attendance_file(ATTENDANCE_FILE, "sheet1")

    log(f"✔ 已写入 attendance.xlsx")
    log(f"✔ 合并后共 {len(att_export)} 条记录\n")

    # =====================
    # 🔥 清空 schedule_messages（放这里）
    # =====================
    if schedule_df is not None and not schedule_df.empty:
        try:
            empty_df = pd.DataFrame(columns=schedule_df.columns)
            empty_df.to_excel(SCHEDULE_FILE, index=False)
            log("🧹 已清空 schedule_messages.xlsx")
        except Exception as e:
            log(f"⚠️ 清空 schedule 失败: {e}")

    # =====================
    # 5) 年报数据
    # =====================
    log("📘 生成年报数据...")
    summary_df, role_df, detail_df = build_yearly_summary(vol_df, att_export, target_year)
    summary_table = build_year_summary_table(vol_df, att_export, target_year)

    debug(f"year_df 行数: {len(att_export)}")
    if "签到" in att_export.columns:
        debug(f"签到所有值: {att_export['签到'].dropna().unique().tolist()}")

        # =====================
        # 6) 生成月报
        # =====================
        
        att_export["日期"] = pd.to_datetime(att_export["日期"], errors="coerce")

        if REBUILD_ALL_MONTHS:
            log("📅 生成所有有资料月份的月报...")
            available_months = sorted(
                att_export.loc[
                    att_export["日期"].notna() & (att_export["日期"].dt.year == target_year),
                    "日期"
                ].dt.month.unique().tolist()
            )
        else:
            if target_month is not None:
                log(f"📅 只生成目标月份月报：{target_year}-{target_month:02d}")
            else:
                log(f"📅 年报模式：不生成指定月报，只生成年报")

            has_target_data = (
                att_export["日期"].notna()
                & (att_export["日期"].dt.year == target_year)
                & (att_export["日期"].dt.month == target_month)
            ).any()

            available_months = [target_month] if has_target_data else []

        month_files = []
        month_done = []

        for m in available_months:
            log(f"   ➜ 生成 {target_year}-{m:02d} 月报...")

            month_df, daily_info, days_in_month, all_names, no_map, anomaly_df, daily_totals, month_grand_total = build_monthly_summary(
                vol_df, att_export, target_year, m
            )

            month_file = write_monthly_report(
                month_df, daily_info, days_in_month, all_names, no_map, anomaly_df,
                daily_totals, month_grand_total, target_year, m
            )
            month_files.append(month_file)
            month_done.append(f"{m:02d}")

        if month_done:
            if REBUILD_ALL_MONTHS:
                log("✔ 已生成月份：" + "、".join(month_done) + "\n")
            else:
                log(f"✔ 已生成目标月份月报：{target_year}-{month_done[0]}\n")
        else:
            if REBUILD_ALL_MONTHS:
                log("⚠ 没有可生成的月报数据\n")
            else:
                log(f"⚠ 目标月份 {target_year}-{target_month:02d} 没有数据，未生成月报\n")

    # =====================
    # 7) 写入年报
    # =====================
    log("📝 写入年报...")
    year_file = write_yearly_report(
        summary_df, role_df, detail_df, target_year, summary_table
    )

    # =====================
    # 8) 完成
    # =====================
    log("🎉 完成！")
    for mf in month_files:
        log(f"月报：{mf}")
    log(f"年报：{year_file}\n")

if __name__ == "__main__":
    main()