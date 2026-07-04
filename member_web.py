# member_web.py

import os
import json
import psycopg2
import pandas as pd

from io import BytesIO
from db import db_query
from opencc import OpenCC
from flask import send_file
from psycopg2.extras import RealDictCursor
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from schedule.builders.time_utils import malaysia_now
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Blueprint, request, render_template_string, redirect, url_for, flash, session, send_file


cc = OpenCC('t2s')  # 繁 → 简

member_bp = Blueprint("member", __name__, url_prefix="/member")

DATABASE_URL = os.environ.get("DATABASE_URL")

MEMBER_ADMIN_PIN = os.environ.get("MEMBER_ADMIN_PIN", "123789")
FINANCE_PIN = os.environ.get("FINANCE_PIN", "1234")

PAYMENT_YEAR = 2026
MONTHS = [f"{PAYMENT_YEAR}-{m:02d}" for m in range(1, 13)]

def to_simple(text):
    if not text:
        return ""
    return cc.convert(str(text).strip())


def search_volunteer_names(keyword, volunteers):
    keyword_simple = to_simple(keyword)

    if not keyword_simple:
        return []

    results = []

    for v in volunteers:
        name = str(v.get("name") or v.get("姓名") or "").strip()
        phone = str(v.get("phone") or v.get("电话") or v.get("电话号码") or "").strip()

        name_simple = to_simple(name)

        if keyword_simple in name_simple or keyword_simple in phone:
            results.append(v)

    return results

def check_missing_months(paid_months):
    if not paid_months:
        return []

    paid_set = set(paid_months)
    sorted_months = sorted(paid_months)

    first = sorted_months[0]
    last = sorted_months[-1]

    expected = []
    start_m = int(first.split("-")[1])
    end_m = int(last.split("-")[1])
    year = first.split("-")[0]

    for m in range(start_m, end_m + 1):
        month = f"{year}-{m:02d}"
        if month not in paid_set:
            expected.append(month)

    return expected

def get_conn():
    return psycopg2.connect(DATABASE_URL)

def normalize_volunteer_id(keyword):
    keyword = str(keyword).strip().upper()

    if keyword.startswith("CHE-") or keyword.startswith("STW-"):
        return keyword

    if keyword.isdigit():
        return f"CHE-{int(keyword)}"

    return keyword

def get_member_query_volunteer():
    return session.get("member_query_volunteer")


def normalize_member_id(raw):
    s = str(raw or "").strip().upper()

    if not s:
        return ""

    # 208 -> CHE-208
    if s.isdigit():
        if s.startswith("0"):
            # 0208 / 0160 这种先保留给 STW 逻辑
            return "STW-" + s.lstrip("0")
        return "CHE-" + s

    return s

def find_volunteer_for_member(cur, member):
    member_id = str(member.get("member_id") or "").strip()
    name = str(member.get("name") or "").strip()
    phone = str(member.get("phone") or "").strip()

    # 1. 先用编号找
    cur.execute("""
        select *
        from volunteers
        where id = %s
        limit 1
    """, (member_id,))
    vol = cur.fetchone()
    if vol:
        return vol

    # 2. 再用电话找
    if phone:
        cur.execute("""
            select *
            from volunteers
            where phone = %s
            limit 1
        """, (phone,))
        vol = cur.fetchone()
        if vol:
            return vol

    # 3. 最后用姓名找
    if name:
        cur.execute("""
            select *
            from volunteers
            where name = %s
            limit 1
        """, (name,))
        vol = cur.fetchone()
        if vol:
            return vol

    return None


def verify_member_pin(member, pin):
    pin = str(pin or "").strip()

    db_pin = str(member.get("pin") or "").strip()
    phone = str(member.get("phone") or "").strip()
    default_pin = phone[-4:] if len(phone) >= 4 else ""

    if db_pin:
        return pin == db_pin

    return pin == default_pin


from datetime import datetime
import pandas as pd

def parse_month(value):

    if value is None or pd.isna(value):
        return None

    # Excel 日期
    if isinstance(value, datetime):
        return value.date().replace(day=1)

    text = str(value).strip()

    # Nov-24
    for fmt in (
        "%b-%y",
        "%B-%y",
        "%Y-%m",
        "%Y/%m",
    ):
        try:
            return datetime.strptime(text, fmt).date().replace(day=1)
        except:
            pass

    # 最后再交给 pandas
    return pd.to_datetime(text).date().replace(day=1)


@member_bp.route("/query-logout")
def member_query_logout():
    session.pop("member_query_volunteer", None)
    return redirect(url_for("member.member_query_login"))


@member_bp.route("/query-login", methods=["GET", "POST"])
def member_query_login():

    error = None

    if request.method == "POST":
        keyword = request.form.get("volunteer_id", "").strip()
        branch = request.form.get("branch", "CHE").strip().upper()

        if keyword.isdigit():
            volunteer_id = f"{branch}-{int(keyword)}"
        else:
            volunteer_id = normalize_volunteer_id(keyword)

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    select id, name, status
                    from volunteers
                    where id = %s
                    limit 1
                """, (volunteer_id,))

                vol = cur.fetchone()

        if not vol:
            error = "找不到这个义工编号"
        elif vol.get("status") != "在册":
            error = "这个义工不是在册状态"
        else:
            session["member_query_volunteer"] = {
                "id": str(vol["id"]),
                "name": vol["name"]
            }
            return redirect(url_for("member.member_home"))

    return render_template_string("""
    <!doctype html>
    <html>
    <head>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>月费查询系统</title>

    <style>
    body {
        margin: 0;
        font-family: Arial, "Microsoft YaHei", sans-serif;
        background: #f4f6f8;
    }

    .card {
        max-width: 520px;
        margin: 60px auto;
        background: white;
        padding: 35px;
        border-radius: 18px;
        box-shadow: 0 8px 24px rgba(0,0,0,0.12);
    }

    h1 {
        margin-top: 0;
        color: #7a0000;
        font-size: 34px;
    }

    h2 {
        font-size: 22px;
        color: #333;
    }

    .branch-row {
        display: flex;
        gap: 12px;
        margin: 20px 0;
    }

    .branch-btn {
        flex: 1;
        padding: 18px;
        font-size: 26px;
        font-weight: bold;
        border: none;
        border-radius: 14px;
        cursor: pointer;
        background: #ddd;
        color: #333;
    }

    .branch-btn.active {
        background: #7a0000;
        color: white;
    }

    input {
        width: 100%;
        box-sizing: border-box;
        font-size: 26px;
        padding: 16px;
        border-radius: 12px;
        border: 1px solid #aaa;
        margin-bottom: 18px;
    }

    button.submit-btn {
        width: 100%;
        font-size: 26px;
        padding: 18px;
        border: none;
        border-radius: 14px;
        background: #198754;
        color: white;
        font-weight: bold;
        cursor: pointer;
    }

    .error {
        background: #ffe5e5;
        color: #b00020;
        padding: 14px;
        border-radius: 10px;
        font-size: 20px;
        margin-bottom: 15px;
    }

    .note {
        color: #666;
        font-size: 16px;
    }
    </style>
    </head>

    <body>

    <div class="card">
        <h1>月费查询系统</h1>
        <h2>请先输入值班义工编号</h2>

        {% if error %}
            <div class="error">{{ error }}</div>
        {% endif %}

        <form method="post">
            <input type="hidden" id="branch" name="branch" value="CHE">

            <div class="branch-row">
                <button type="button" id="btnCHE" class="branch-btn active" onclick="setBranch('CHE')">
                    CHE
                </button>
                <button type="button" id="btnSTW" class="branch-btn" onclick="setBranch('STW')">
                    STW
                </button>
            </div>

            <input name="volunteer_id"
                placeholder="例如：108"
                required
                autofocus>

            <button class="submit-btn" type="submit">
                进入查询
            </button>

            <p class="note">
                系统会记录查询义工与查询次数，方便财政查看。
            </p>
        </form>
    </div>

    <script>
    function setBranch(branch) {
        document.getElementById("branch").value = branch;

        document.getElementById("btnCHE").classList.remove("active");
        document.getElementById("btnSTW").classList.remove("active");

        document.getElementById("btn" + branch).classList.add("active");
    }
    </script>

    </body>
    </html>
    """, error=error)


@member_bp.route("/query-logs")
def member_query_logs():

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    mode = request.args.get("mode", "today")

    where_sql = ""
    params = []

    if mode == "today":
        where_sql = "where queried_at::date = current_date"
    elif mode == "month":
        where_sql = """
            where date_trunc('month', queried_at)
            = date_trunc('month', current_date)
        """

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(f"""
                select *
                from member_query_logs
                {where_sql}
                order by queried_at desc
                limit 300
            """, params)

            logs = cur.fetchall()

            cur.execute(f"""
                select
                    volunteer_id,
                    volunteer_name,
                    count(*) as total
                from member_query_logs
                {where_sql}
                group by volunteer_id, volunteer_name
                order by total desc
            """, params)

            summary = cur.fetchall()

        return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
    <meta charset="utf-8">
    <title>月费查询记录</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">

    <link rel="stylesheet"
        href="{{ url_for('static', filename='css/toolbox.css') }}">

    <style>
    .log-filter-row{
        display:grid;
        grid-template-columns:repeat(4, 1fr);
        gap:12px;
    }

    .query-log-page .section-title{
        font-size:26px;
        margin-bottom:16px;
    }

    .query-log-page .card{
        padding:22px;
    }

    .query-log-page .btn-tool{
        min-height:54px;
        font-size:21px;
    }

    @media(max-width:700px){
        .log-filter-row{
            grid-template-columns:1fr;
        }
    }
    </style>
    </head>

    <body>

    <div class="page query-log-page">

        <h1 class="page-title">🔍 月费查询记录</h1>
        <p class="page-subtitle">查看义工查询佛友月费资料的记录</p>

        <div class="card">
            <div class="section-title">📅 筛选</div>

            <div class="log-filter-row">
                <a class="btn-tool {% if mode == 'today' %}btn-primary{% else %}btn-secondary{% endif %}"
                href="/member/query-logs?mode=today">
                    今天
                </a>

                <a class="btn-tool {% if mode == 'month' %}btn-primary{% else %}btn-secondary{% endif %}"
                href="/member/query-logs?mode=month">
                    本月
                </a>

                <a class="btn-tool {% if mode == 'all' %}btn-primary{% else %}btn-secondary{% endif %}"
                href="/member/query-logs?mode=all">
                    全部
                </a>

                <a class="btn-tool btn-warning"
                href="/member/finance-upload">
                    ← 返回管理中心
                </a>
            </div>
        </div>

        <div class="card">
            <div class="section-title">👥 义工查询统计</div>

            {% if summary %}
            <div class="table-responsive">
                <table class="record-table">
                    <tr>
                        <th>义工编号</th>
                        <th>义工姓名</th>
                        <th>查询次数</th>
                    </tr>

                    {% for s in summary %}
                    <tr>
                        <td>{{ s.volunteer_id }}</td>
                        <td>{{ s.volunteer_name }}</td>
                        <td>{{ s.total }}</td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% else %}
            <div class="empty-state">
                <div class="empty-title">暂无查询统计</div>
            </div>
            {% endif %}
        </div>

        <div class="card">
            <div class="section-title">📋 详细记录</div>

            {% if logs %}
            <div class="table-responsive">
                <table class="record-table">
                    <tr>
                        <th>时间</th>
                        <th>查询义工</th>
                        <th>佛友编号</th>
                        <th>佛友姓名</th>
                        <th>输入关键词</th>
                        <th>IP</th>
                    </tr>

                    {% for r in logs %}
                    <tr>
                        <td>
                        {{ (r.queried_at + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S") if r.queried_at else "-" }}
                        </td>
                        <td>{{ r.volunteer_id }} {{ r.volunteer_name }}</td>
                        <td>{{ r.member_id }}</td>
                        <td>{{ r.member_name }}</td>
                        <td>{{ r.keyword }}</td>
                        <td>{{ r.ip_address }}</td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% else %}
            <div class="empty-state">
                <div class="empty-title">暂无查询记录</div>
            </div>
            {% endif %}
        </div>

    </div>

    </body>
    </html>
    """, logs=logs, summary=summary, mode=mode, timedelta=timedelta)


@member_bp.route("/", methods=["GET", "POST"])
def member_home():

    query_volunteer = get_member_query_volunteer()

    if not query_volunteer:
        return redirect(url_for("member.member_query_login"))
    
    member = None
    error = None
    paid_until = None
    summary = None
    payments = []
    candidates = []

    def load_member_payment_data(cur, real_member_id):
        cur.execute("""
            select
                coalesce(sum(amount), 0) as total_payment,
                coalesce(sum(month_count), 0) as total_months,
                max(end_month) as paid_until,
                max(payment_date) as last_payment_date
            from member_payments
            where member_id = %s
        """, (real_member_id,))
        summary_data = cur.fetchone()

        paid_until_text = None
        if summary_data and summary_data["paid_until"]:
            paid_until_text = summary_data["paid_until"].strftime("%Y年%m月")

        cur.execute("""
            select amount
            from member_payments
            where member_id = %s
            order by payment_date desc, id desc
            limit 1
        """, (real_member_id,))
        last_payment = cur.fetchone()

        if summary_data:
            summary_data["last_payment_amount"] = (
                last_payment["amount"] if last_payment else 0
            )

        cur.execute("""
            select
                payment_date,
                receipt_no,
                start_month,
                end_month,
                month_count,
                amount
            from member_payments
            where member_id = %s
            order by payment_date desc, receipt_no desc
        """, (real_member_id,))
        payment_rows = cur.fetchall()

        return summary_data, paid_until_text, payment_rows

    if request.method == "POST":
        raw_member_id = request.form.get("member_id", "").strip()
        selected_member_id = request.form.get("selected_member_id", "").strip()
        branch = request.form.get("branch", "CHE").strip().upper()
        

        try:
            with get_conn() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:

                    # 第二次提交：用户已经从候选名单选择会员
                    if selected_member_id:
                        cur.execute("""
                            select *
                            from members
                            where member_id = %s
                            limit 1
                        """, (selected_member_id,))
                        member = cur.fetchone()

                    else:
                        keyword = raw_member_id.strip()
                        digits_only = "".join(ch for ch in keyword if ch.isdigit())

                        # 情况 1：纯数字，而且 6 位或以下，当会员编号
                        if keyword.isdigit() and len(keyword) <= 6:
                            if branch == "STW":
                                member_id = normalize_member_id(f"STW-{keyword}")
                            else:
                                member_id = normalize_member_id(keyword)

                            cur.execute("""
                                select *
                                from members
                                where member_id = %s
                                limit 1
                            """, (member_id,))
                            member = cur.fetchone()

                        # 情况 2：纯数字，而且超过 6 位，当电话号码
                        elif keyword.isdigit() and len(keyword) > 6:
                            cur.execute("""
                                select *
                                from members
                                where replace(replace(coalesce(phone,''), '-', ''), ' ', '') = %s
                                order by member_id
                            """, (digits_only,))
                            candidates = cur.fetchall()

                            if len(candidates) == 1:
                                member = candidates[0]
                                candidates = []

                            elif len(candidates) > 1:
                                member = None

                            else:
                                member = None

                        # 情况 3：非纯数字，当会员编号 / 中文名 / 英文名 搜索
                        else:
                            member_id = normalize_member_id(keyword)

                            cur.execute("""
                                select *
                                from members
                                where member_id = %s
                                   or name ilike %s
                                   or english_name ilike %s
                                order by member_id
                            """, (
                                member_id,
                                f"%{keyword}%",
                                f"%{keyword}%"
                            ))
                            candidates = cur.fetchall()

                            if len(candidates) == 1:
                                member = candidates[0]
                                candidates = []

                            elif len(candidates) > 1:
                                member = None

                            else:
                                member = None

                    if candidates:
                        error = None

                    elif not member:
                        error = "找不到这个会员，请检查编号 / 姓名 / 电话"

                    else:
                        real_member_id = member["member_id"]

                        # ✅ 成功找到会员后，记录是谁查询了这个会员
                        query_volunteer = get_member_query_volunteer()

                        if query_volunteer:
                            cur.execute("""
                                insert into member_query_logs (
                                    queried_at,
                                    volunteer_id,
                                    volunteer_name,
                                    member_id,
                                    member_name,
                                    keyword,
                                    ip_address,
                                    user_agent
                                )
                                values (%s, %s, %s, %s, %s, %s, %s, %s)
                            """, (
                                malaysia_now(),
                                query_volunteer["id"],
                                query_volunteer["name"],
                                member["member_id"],
                                member["name"],
                                raw_member_id or selected_member_id,
                                request.remote_addr,
                                request.headers.get("User-Agent", "")
                            ))
                            conn.commit()

                        summary, paid_until, payments = load_member_payment_data(
                            cur,
                            real_member_id
                        )

                        if summary and summary["paid_until"]:

                            today = date.today()
                            paid_until_date = summary["paid_until"]

                            late_months = (
                                (today.year - paid_until_date.year) * 12
                                + (today.month - paid_until_date.month)
                            )

                            summary["late_months"] = late_months

                            if late_months <= 0:
                                summary["payment_status"] = "🟢 正常"

                            elif late_months <= 3:
                                summary["payment_status"] = f"🟡 已过期 {late_months} 个月"

                            else:
                                summary["payment_status"] = f"🔴 停供参考 {late_months} 个月"

                        else:
                            if summary:
                                summary["late_months"] = None
                                summary["payment_status"] = "⚪ 暂无缴费记录"

        except Exception as e:
            error = f"系统错误：{e}"

    return render_template_string(
        MEMBER_HTML,
        member=member,
        error=error,
        paid_until=paid_until,
        summary=summary,
        payments=payments,
        candidates=candidates,
        query_volunteer=query_volunteer
    )

@member_bp.route("/admin", methods=["GET", "POST"])
def member_admin():
    error = None
    member = None
    payments = []
    summary = None
    warnings = []

    page = int(request.args.get("page", 1) or 1)
    per_page = 15
    offset = (page - 1) * per_page
    total_pages = 0

    admin_pin = request.form.get("admin_pin", "").strip()
    raw_member_id = request.values.get("member_id", "").strip()
    branch = request.values.get("branch", "CHE").strip().upper()

    if request.method == "POST":

        if not session.get("member_admin"):

            admin_pin = request.form.get("admin_pin", "")

            if admin_pin == MEMBER_ADMIN_PIN:
                session["member_admin"] = True
                return redirect("/member/admin")

            error = "管理员 PIN 不正确"
            

    if raw_member_id and not error:
        keyword = raw_member_id.strip()

        if keyword.isdigit():
            if branch == "STW":
                member_id = normalize_member_id(f"STW-{keyword}")
            else:
                member_id = normalize_member_id(keyword)
        else:
            member_id = normalize_member_id(keyword)

        try:
            with get_conn() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:

                    # ✅ 数字只查编号，避免 208 查到电话 0123992081
                    if keyword.isdigit():
                        cur.execute("""
                            select *
                            from members
                            where member_id = %s
                            limit 1
                        """, (member_id,))
                    else:
                        cur.execute("""
                            select *
                            from members
                            where member_id = %s
                               or name ilike %s
                               or english_name ilike %s
                               or phone ilike %s
                            limit 1
                        """, (
                            member_id,
                            f"%{keyword}%",
                            f"%{keyword}%",
                            f"%{keyword}%"
                        ))

                    member = cur.fetchone()

                    if not member:
                        error = "找不到这个月费编号"
                    else:
                        real_member_id = member["member_id"]

                        # ✅ 汇总
                        cur.execute("""
                            select
                                coalesce(sum(amount), 0) as total_payment,
                                coalesce(sum(month_count), 0) as total_months,
                                max(end_month) as paid_until,
                                max(payment_date) as last_payment_date
                            from member_payments
                            where member_id = %s
                        """, (real_member_id,))
                        summary = cur.fetchone()

                        # ✅ 最后付款金额
                        if summary:
                            summary["last_payment_amount"] = None

                            if summary.get("last_payment_date"):
                                cur.execute("""
                                    select amount
                                    from member_payments
                                    where member_id = %s
                                      and payment_date = %s
                                    order by id desc
                                    limit 1
                                """, (
                                    real_member_id,
                                    summary["last_payment_date"]
                                ))
                                last_row = cur.fetchone()

                                if last_row:
                                    summary["last_payment_amount"] = last_row["amount"]

                        # ✅ 总页数
                        cur.execute("""
                            select count(*) as cnt
                            from member_payments
                            where member_id = %s
                        """, (real_member_id,))
                        total_rows = cur.fetchone()["cnt"] or 0
                        total_pages = (total_rows + per_page - 1) // per_page

                        # ✅ 当前页，只显示 15 行
                        cur.execute("""
                            select
                                id,
                                payment_date,
                                receipt_no,
                                member_id,
                                start_month,
                                end_month,
                                month_count,
                                amount,
                                name
                            from member_payments
                            where member_id = %s
                            order by payment_date desc, id desc
                            limit %s offset %s
                        """, (
                            real_member_id,
                            per_page,
                            offset
                        ))
                        payments = cur.fetchall()

                        # ✅ 另外查全部记录，用来检查错误
                        cur.execute("""
                            select
                                id,
                                payment_date,
                                receipt_no,
                                start_month,
                                end_month,
                                month_count,
                                amount,
                                name
                            from member_payments
                            where member_id = %s
                            order by payment_date asc, id asc
                        """, (real_member_id,))
                        all_payments = cur.fetchall()

                        seen_months = {}
                        seen_receipts = {}

                        for r in all_payments:
                            receipt_no = r.get("receipt_no") or "-"
                            amount = float(r.get("amount") or 0)
                            month_count = int(r.get("month_count") or 0)

                            start_month = r.get("start_month")
                            end_month = r.get("end_month")
                            payment_date = r.get("payment_date")

                            # 1. 收据重复
                            if receipt_no != "-":
                                if receipt_no in seen_receipts:
                                    warnings.append(
                                        f"收据 {receipt_no}：收据编号重复，之前已经出现过"
                                    )
                                else:
                                    seen_receipts[receipt_no] = True

                            # 2. 金额不是 RM50 倍数
                            if amount % 50 != 0:
                                warnings.append(
                                    f"收据 {receipt_no}：金额 RM {amount:.2f} 不是 RM50 的倍数"
                                )

                            # 3. 月数不正确
                            if month_count <= 0:
                                warnings.append(
                                    f"收据 {receipt_no}：月数不正确"
                                )

                            # 4. 金额和月数不符合
                            if amount > 0 and month_count > 0:
                                expected_amount = month_count * 50
                                if amount != expected_amount:
                                    warnings.append(
                                        f"收据 {receipt_no}：金额 RM {amount:.2f} 和月数 {month_count} 不符合，正常应是 RM {expected_amount:.2f}"
                                    )

                            # 5. 付款日期未来
                            if payment_date and payment_date > date.today():
                                warnings.append(
                                    f"收据 {receipt_no}：付款日期是未来日期"
                                )

                            # 6. 开始结束月份检查
                            if start_month and end_month:
                                if start_month > end_month:
                                    warnings.append(
                                        f"收据 {receipt_no}：开始月份大过结束月份"
                                    )
                                else:
                                    real_month_count = (
                                        (end_month.year - start_month.year) * 12
                                        + (end_month.month - start_month.month)
                                        + 1
                                    )

                                    if real_month_count != month_count:
                                        warnings.append(
                                            f"收据 {receipt_no}：开始月份到结束月份是 {real_month_count} 个月，但记录写 {month_count} 个月"
                                        )

                                    # 7. 重复月份检查
                                    current = start_month

                                    for i in range(real_month_count):
                                        ym = current.strftime("%Y-%m")

                                        if ym in seen_months:
                                            warnings.append(
                                                f"月份重复：{ym} 已在收据 {seen_months[ym]} 记录过，现在又出现在收据 {receipt_no}"
                                            )
                                        else:
                                            seen_months[ym] = receipt_no

                                        if current.month == 12:
                                            current = current.replace(
                                                year=current.year + 1,
                                                month=1
                                            )
                                        else:
                                            current = current.replace(
                                                month=current.month + 1
                                            )

        except Exception as e:
            error = f"系统错误：{e}"

    return render_template_string(
        MEMBER_ADMIN_HTML,
        error=error,
        member=member,
        payments=payments,
        summary=summary,
        warnings=warnings,
        raw_member_id=raw_member_id,
        page=page,
        total_pages=total_pages
    )

@member_bp.route("/admin/late_members")
def admin_late_members():

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    rows = db_query("""
        select
            m.member_id,
            m.name,
            m.phone,
            m.remark,
            max(p.end_month) as paid_until,
            max(p.payment_date) as last_payment_date
        from members m
        left join member_payments p
            on p.member_id = m.member_id
        where coalesce(m.member_status, m.status, '') not in ('停供', '往生', '已往生')
        group by
            m.member_id,
            m.name,
            m.phone,
            m.remark
        having
            max(p.end_month) is not null
            and max(p.end_month) < date_trunc('month', current_date)
        order by
            paid_until asc,
            m.member_id
    """, fetchall=True)

    today = date.today()
    current_month_index = today.year * 12 + today.month

    for r in rows:
        paid_until = r["paid_until"]
        paid_index = paid_until.year * 12 + paid_until.month
        late_months = current_month_index - paid_index

        r["late_months"] = late_months
        r["reference_amount"] = late_months * 50

        if late_months <= 2:
            r["level"] = "green"
        elif late_months <= 6:
            r["level"] = "yellow"
        else:
            r["level"] = "red"

        phone = (r["phone"] or "").strip()
        phone_digits = "".join(ch for ch in phone if ch.isdigit())

        if phone_digits.startswith("0"):
            phone_digits = "6" + phone_digits

        r["wa_link"] = "https://wa.me/" + phone_digits if phone_digits else ""

    green_count = sum(1 for r in rows if r["level"] == "green")
    yellow_count = sum(1 for r in rows if r["level"] == "yellow")
    red_count = sum(1 for r in rows if r["level"] == "red")
    total_amount = sum(r["reference_amount"] for r in rows)

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
    <meta charset="utf-8">
    <title>月费迟付名单</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">

    <link rel="stylesheet"
        href="{{ url_for('static', filename='css/toolbox.css') }}">

    <style>
    .late-page .section-title{
        font-size:26px;
        margin-bottom:16px;
    }

    .late-page .card{
        padding:22px;
    }

    .late-page .btn-tool{
        min-height:54px;
        font-size:21px;
    }

    .late-row-green{
        background:#ecfdf5;
    }

    .late-row-yellow{
        background:#fffbeb;
    }

    .late-row-red{
        background:#fef2f2;
    }

    .floating-back{
        position:fixed;
        left:24px;
        bottom:24px;
        z-index:999;
        padding:14px 22px;
        border-radius:14px;
        background:#4b5563;
        color:white;
        text-decoration:none;
        font-size:18px;
        font-weight:700;
        box-shadow:0 8px 20px rgba(0,0,0,.18);
    }

    .floating-back:hover{
        background:#374151;
    }
    </style>
    </head>

    <body>

    <div class="page late-page">

        <h1 class="page-title">🌿 月费迟付名单</h1>
        <p class="page-subtitle">查看已过期月费会员，方便后续关怀联系</p>

        <div class="card">
            <div class="info-summary-grid">

                <div class="info-summary-box">
                    <div class="info-summary-title">总人数</div>
                    <div class="info-summary-value">{{ rows|length }} 人</div>
                </div>

                <div class="info-summary-box">
                    <div class="info-summary-title">参考金额</div>
                    <div class="info-summary-value">RM {{ "{:,.2f}".format(total_amount) }}</div>
                </div>

                <div class="info-summary-box">
                    <div class="info-summary-title">🟢 最近缴费</div>
                    <div class="info-summary-value">{{ green_count }} 人</div>
                </div>

                <div class="info-summary-box">
                    <div class="info-summary-title">🟡 一段时间未缴费</div>
                    <div class="info-summary-value">{{ yellow_count }} 人</div>
                </div>

                <div class="info-summary-box">
                    <div class="info-summary-title">🔴 较久未缴费</div>
                    <div class="info-summary-value">{{ red_count }} 人</div>
                </div>

            </div>
        </div>

        <div class="card">
            <div class="section-title">📋 名单</div>

            {% if rows %}
            <div class="table-responsive">
                <table class="record-table">
                    <tr>
                        <th>会员编号</th>
                        <th>姓名</th>
                        <th>电话</th>
                        <th>WhatsApp</th>
                        <th>已缴至</th>
                        <th>缴费间隔</th>
                        <th>参考金额</th>
                        <th>状态</th>
                        <th>最后付款日期</th>
                        <th>备注</th>
                    </tr>

                    {% for r in rows %}
                    <tr class="
                        {% if r.level == 'green' %}
                            late-row-green
                        {% elif r.level == 'yellow' %}
                            late-row-yellow
                        {% elif r.level == 'red' %}
                            late-row-red
                        {% endif %}
                    ">
                        <td>{{ r.member_id }}</td>
                        <td>{{ r.name }}</td>
                        <td>{{ r.phone or "-" }}</td>
                        <td>
                            {% if r.wa_link %}
                                <a class="btn-tool btn-success mini-btn"
                                href="{{ r.wa_link }}"
                                target="_blank">
                                    打开
                                </a>
                            {% else %}
                                -
                            {% endif %}
                        </td>
                        <td>{{ r.paid_until.strftime("%Y-%m") }}</td>
                        <td>{{ r.late_months }} 个月</td>
                        <td>RM {{ "%.2f"|format(r.reference_amount) }}</td>
                        <td>
                            {% if r.level == "green" %}
                                🟢 最近缴费
                            {% elif r.level == "yellow" %}
                                🟡 一段时间未缴费
                            {% else %}
                                🔴 较久未缴费
                            {% endif %}
                        </td>
                        <td>{{ r.last_payment_date or "-" }}</td>
                        <td>{{ r.remark or "-" }}</td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% else %}
            <div class="empty-state">
                <div class="empty-title">暂无迟付名单</div>
                <div class="empty-text">目前没有需要列入迟付关怀的会员。</div>
            </div>
            {% endif %}
        </div>

    </div>

    <a class="floating-back"
    href="{{ url_for('member.member_admin') }}">
        ⬅ 返回月费管理员
    </a>

    </body>
    </html>
    """,
    rows=rows,
    green_count=green_count,
    yellow_count=yellow_count,
    red_count=red_count,
    total_amount=total_amount
    )

@member_bp.route("/payment/edit/<int:payment_id>", methods=["GET", "POST"])
def edit_member_payment(payment_id):
    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    error = None

    try:
        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                cur.execute("""
                    select *
                    from member_payments
                    where id = %s
                """, (payment_id,))
                payment = cur.fetchone()

                if not payment:
                    return "找不到这笔缴费记录", 404

                if request.method == "POST":
                    payment_date = request.form.get("payment_date")
                    receipt_no = request.form.get("receipt_no", "").strip()
                    name = request.form.get("name", "").strip()
                    start_month = request.form.get("start_month")
                    end_month = request.form.get("end_month")
                    month_count = int(request.form.get("month_count") or 0)
                    amount = float(request.form.get("amount") or 0)

                    if month_count <= 0:
                        error = "月数不能小过 1"
                    elif amount <= 0:
                        error = "金额不能小过 0"
                    else:
                        old_data = dict(payment)

                        new_data = {
                            "payment_date": payment_date,
                            "receipt_no": receipt_no,
                            "name": name,
                            "start_month": start_month + "-01",
                            "end_month": end_month + "-01",
                            "month_count": month_count,
                            "amount": amount
                        }

                        cur.execute("""
                            update member_payments
                            set
                                payment_date = %s,
                                receipt_no = %s,
                                name = %s,
                                start_month = %s,
                                end_month = %s,
                                month_count = %s,
                                amount = %s
                            where id = %s
                        """, (
                            payment_date,
                            receipt_no,
                            name,
                            start_month + "-01",
                            end_month + "-01",
                            month_count,
                            amount,
                            payment_id
                        ))

                        cur.execute("""
                            insert into member_payment_history
                            (
                                payment_id,
                                member_id,
                                receipt_no,
                                action,
                                old_data,
                                new_data,
                                changed_by
                            )
                            values (%s, %s, %s, %s, %s::jsonb, %s::jsonb, %s)
                        """, (
                            payment_id,
                            payment["member_id"],
                            receipt_no,
                            "edit",
                            json.dumps(old_data, default=str),
                            json.dumps(new_data, default=str),
                            "member_admin"
                        ))

                        conn.commit()

                        return redirect(url_for("member.member_admin", member_id=payment["member_id"]))

        return render_template_string(
            PAYMENT_EDIT_HTML,
            payment=payment,
            error=error
        )

    except Exception as e:
        return f"系统错误：{e}", 500
    
@member_bp.route("/payment/history/<int:history_id>")
def member_payment_history_detail(history_id):
    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    try:
        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    select *
                    from member_payment_history
                    where id = %s
                """, (history_id,))
                h = cur.fetchone()

                if not h:
                    return "找不到这笔修改历史", 404

        changes = format_history_changes(h.get("old_data"), h.get("new_data"))

        return render_template_string(
            PAYMENT_HISTORY_DETAIL_HTML,
            h=h,
            changes=changes,
            timedelta=timedelta
        )

    except Exception as e:
        return f"系统错误：{e}", 500

@member_bp.route("/change-pin", methods=["GET", "POST"])
def member_change_pin():
    error = None
    ok = None

    if request.method == "POST":
        raw_member_id = request.form.get("member_id", "")
        old_pin = request.form.get("old_pin", "")
        new_pin = request.form.get("new_pin", "")
        confirm_pin = request.form.get("confirm_pin", "")

        member_id = normalize_member_id(raw_member_id)

        if not new_pin or len(new_pin) < 4:
            error = "新密码至少 4 位"
        elif new_pin != confirm_pin:
            error = "两次新密码不一样"
        else:
            try:
                with get_conn() as conn:
                    with conn.cursor(cursor_factory=RealDictCursor) as cur:
                        cur.execute("""
                            select *
                            from members
                            where member_id = %s
                            limit 1
                        """, (member_id,))
                        member = cur.fetchone()

                        if not member:
                            error = "找不到这个月费编号"
                        elif not verify_member_pin(member, old_pin):
                            error = "旧密码不正确"
                        else:
                            cur.execute("""
                                update members
                                set pin = %s
                                where member_id = %s
                            """, (new_pin, member_id))
                            conn.commit()
                            ok = "月费密码已更改"

            except Exception as e:
                error = f"系统错误：{e}"

    return render_template_string(CHANGE_PIN_HTML, error=error, ok=ok)

@member_bp.route("/admin/logout")
def member_admin_logout():
    session.pop("member_admin", None)
    return redirect(url_for("member.member_admin"))

@member_bp.route("/member-management")
def member_management():

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>会员资料管理</title>
<meta name="viewport" content="width=device-width, initial-scale=1">

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">

<style>
.member-menu-grid{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(260px,1fr));
    gap:18px;
}

.member-menu-card{
    display:block;
    text-decoration:none;
    border-radius:16px;
    padding:24px;
    color:#111827;
    transition:.2s;
}

.member-menu-card:hover{
    transform:translateY(-2px);
}

.member-menu-title{
    font-size:24px;
    font-weight:bold;
    margin-bottom:10px;
}

.member-menu-desc{
    font-size:16px;
    color:#6b7280;
    line-height:1.6;
}

.card-blue{
    background:#eff6ff;
}

.card-green{
    background:#ecfdf5;
}

.card-yellow{
    background:#fffbeb;
}

.card-gray{
    background:#f9fafb;
}
</style>

</head>
<body>

<div class="page">

    <h1 class="page-title">
        👤 会员资料管理
    </h1>

    <p class="page-subtitle">
        管理会员资料、下载资料、查看迟付名单
    </p>

    <div class="card">

        <div class="member-menu-grid">

            <a class="member-menu-card card-blue"
               href="/member/member-management/search">

                <div class="member-menu-title">
                    🔍 查找 / 编辑会员
                </div>

                <div class="member-menu-desc">
                    修改电话、PIN、备注、会员资料
                </div>

            </a>

            <a class="member-menu-card card-green"
               href="/member/download-members">

                <div class="member-menu-title">
                    📥 下载会员资料 Excel
                </div>

                <div class="member-menu-desc">
                    下载最新 members 主档
                </div>

            </a>

            <a class="member-menu-card card-yellow"
               href="/member/admin/late_members">

                <div class="member-menu-title">
                    ⚠️ 月费迟付名单
                </div>

                <div class="member-menu-desc">
                    查看已过期会员
                </div>

            </a>

            <a class="member-menu-card card-gray"
               href="/member/member-management/add">

                <div class="member-menu-title">
                    ➕ 新增会员
                </div>

                <div class="member-menu-desc">
                    新增 CHE / STW 月费会员
                </div>

            </a>

        </div>

    </div>

    <div class="card">
        <a class="btn-tool btn-secondary"
           href="{{ url_for('member.member_admin') }}">
            ← 返回月费管理员
        </a>
    </div>

</div>

</body>
</html>
""")

@member_bp.route("/download-members")
def download_members():

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    df = pd.read_sql_query("""
        select
            member_id,
            branch,
            name,
            english_name,
            ic_number,
            phone,
            pin,
            status,
            member_status,
            remark
        from members
        order by branch, member_id
    """, get_conn())

    df.columns = [
        "会员编号",
        "分会",
        "姓名",
        "英文名",
        "身份证号码",
        "电话",
        "PIN",
        "状态",
        "月费状态",
        "备注"
    ]

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="会员总表",
            startrow=2
        )

        ws = writer.sheets["会员总表"]

        ws.merge_cells("A1:J1")
        ws["A1"] = "观音堂会员资料总表"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="4472C4")
        ws["A1"].alignment = Alignment(horizontal="center")

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        blue_fill = PatternFill("solid", fgColor="DDEBF7")
        white_fill = PatternFill("solid", fgColor="FFFFFF")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_num in range(4, ws.max_row + 1):
            fill = blue_fill if row_num % 2 == 0 else white_fill

            for col in range(1, 11):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:J{ws.max_row}"

        widths = {
            "A": 14,
            "B": 10,
            "C": 18,
            "D": 22,
            "E": 20,
            "F": 18,
            "G": 10,
            "H": 12,
            "I": 12,
            "J": 28,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 24

    output.seek(0)

    filename = f"members_{date.today().strftime('%Y-%m-%d')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@member_bp.route("/member-management/search")
def member_management_search():

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    q = request.args.get("q", "").strip()
    rows = []

    if q:
        keyword = q
        member_id = normalize_member_id(keyword)

        if keyword.isdigit():
            rows = db_query("""
                select
                    member_id,
                    name,
                    english_name,
                    phone,
                    pin,
                    branch,
                    status,
                    remark,
                    member_status
                from members
                where member_id = %s
                order by member_id
                limit 50
            """, (member_id,), fetchall=True)

        else:
            rows = db_query("""
                select
                    member_id,
                    name,
                    english_name,
                    phone,
                    pin,
                    branch,
                    status,
                    remark,
                    member_status
                from members
                where member_id = %s
                or name ilike %s
                or english_name ilike %s
                or phone ilike %s
                order by member_id
                limit 50
            """, (
                member_id,
                f"%{keyword}%",
                f"%{keyword}%",
                f"%{keyword}%"
            ), fetchall=True)

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>查找 / 编辑会员</title>
<style>
body{
    font-family:Arial,"Microsoft YaHei",sans-serif;
    background:#f3f4f6;
    padding:20px;
}
.box{
    max-width:1000px;
    margin:auto;
    background:white;
    padding:24px;
    border-radius:18px;
    box-shadow:0 3px 12px rgba(0,0,0,.08);
}
h1{
    text-align:center;
    font-size:36px;
}
input{
    width:100%;
    font-size:22px;
    padding:12px;
    border:1px solid #ccc;
    border-radius:10px;
}
button,.btn{
    display:inline-block;
    background:#2d7ff9;
    color:white;
    padding:12px 18px;
    border-radius:10px;
    border:0;
    font-size:20px;
    font-weight:bold;
    text-decoration:none;
    margin-top:12px;
}
table{
    width:100%;
    border-collapse:collapse;
    margin-top:20px;
}
th,td{
    border:1px solid #ddd;
    padding:10px;
    text-align:center;
}
th{
    background:#eaf2ff;
}
.edit{
    background:#16a34a;
    color:white;
    padding:7px 12px;
    border-radius:8px;
    text-decoration:none;
}
</style>
</head>
<body>
<div class="box">

    <a href="/member/member-management">← 返回会员资料管理</a>

    <h1>🔍 查找 / 编辑会员</h1>

    <form method="get">
        <input name="q" value="{{ q }}" placeholder="输入编号 / 姓名 / 英文名 / 电话">
        <button type="submit">查找会员</button>
    </form>

    {% if q %}
        <h3>搜索结果：{{ rows|length }} 位</h3>

        {% if rows %}
        <table>
            <tr>
                <th>编号</th>
                <th>姓名</th>
                <th>英文名</th>
                <th>电话</th>
                <th>PIN</th>
                <th>分会</th>
                <th>状态</th>
                <th>备注</th>
                <th>操作</th>
            </tr>

            {% for r in rows %}
            <tr>
                <td>{{ r.member_id }}</td>
                <td>{{ r.name }}</td>
                <td>{{ r.english_name or "-" }}</td>
                <td>{{ r.phone or "-" }}</td>
                <td>{{ r.pin or "-" }}</td>
                <td>{{ r.branch or "-" }}</td>
                <td>{{ r.member_status or r.status or "-" }}</td>
                <td>{{ r.remark or "-" }}</td>
                <td>
                    <a class="edit" href="/member/member-management/edit/{{ r.member_id }}">
                        ✏ 编辑
                    </a>
                </td>
            </tr>
            {% endfor %}
        </table>
        {% else %}
            <p>找不到会员。</p>
        {% endif %}
    {% endif %}

</div>
</body>
</html>
""", q=q, rows=rows)

@member_bp.route("/member-management/edit/<member_id>", methods=["GET", "POST"])
def edit_member(member_id):

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    member = db_query("""
        select *
        from members
        where member_id = %s
    """, (member_id,), fetchone=True)

    if not member:
        return "会员不存在"

    msg = None
    error = None

    if request.method == "POST":

        old_data = dict(member)

        name = request.form.get("name", "").strip()
        english_name = request.form.get("english_name", "").strip()
        ic_number = request.form.get("ic_number", "").strip()
        phone = request.form.get("phone", "").strip()
        pin = request.form.get("pin", "").strip()
        branch = request.form.get("branch", "").strip().upper()
        member_status = request.form.get("member_status", "").strip()
        remark = request.form.get("remark", "").strip()

        if not name:
            error = "姓名不能为空"

        if not error and branch not in ["CHE", "STW"]:
            error = "分会只能是 CHE 或 STW"

        if not error and ic_number:
            exist = db_query("""
                select member_id, name
                from members
                where ic_number = %s
                  and member_id <> %s
                limit 1
            """, (ic_number, member_id), fetchone=True)

            if exist:
                error = f"身份证号码已存在：{exist['member_id']} / {exist['name']}"

        if not error:
            db_query("""
                update members
                set
                    name = %s,
                    english_name = %s,
                    ic_number = %s,
                    phone = %s,
                    pin = %s,
                    branch = %s,
                    status = %s,
                    member_status = %s,
                    remark = %s
                where member_id = %s
            """, (
                name,
                english_name,
                ic_number,
                phone,
                pin,
                branch,
                member_status,
                member_status,
                remark or member_status,
                member_id
            ))

            new_member = db_query("""
                select *
                from members
                where member_id = %s
            """, (member_id,), fetchone=True)

            new_data = dict(new_member)

            if json.dumps(old_data, default=str, sort_keys=True) != json.dumps(new_data, default=str, sort_keys=True):
                db_query("""
                    insert into member_profile_history
                    (
                        member_id,
                        action,
                        old_data,
                        new_data,
                        changed_by
                    )
                    values (%s,%s,%s,%s,%s)
                """, (
                    member_id,
                    "edit",
                    json.dumps(old_data, default=str),
                    json.dumps(new_data, default=str),
                    "member_admin"
                ))

                msg = "保存成功，已记录修改历史"
            else:
                msg = "没有实际修改"

            member = new_member

    history_rows = db_query("""
        select *
        from member_profile_history
        where member_id = %s
        order by changed_at desc
        limit 10
    """, (member_id,), fetchall=True)

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>编辑会员</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
body{
    font-family:Arial,"Microsoft YaHei",sans-serif;
    background:#f3f4f6;
    margin:0;
    padding:20px;
}
.box{
    max-width:900px;
    margin:auto;
    background:white;
    padding:24px;
    border-radius:18px;
    box-shadow:0 3px 12px rgba(0,0,0,.08);
}
h1{
    text-align:center;
    font-size:38px;
}
label{
    font-size:20px;
    font-weight:bold;
    display:block;
    margin-top:16px;
}
input,select,textarea{
    width:100%;
    padding:12px;
    font-size:20px;
    border:1px solid #ccc;
    border-radius:10px;
    box-sizing:border-box;
    margin-top:6px;
}
textarea{
    height:120px;
}
button{
    background:#16a34a;
    color:white;
    border:0;
    padding:14px 22px;
    border-radius:10px;
    font-size:22px;
    font-weight:bold;
    margin-top:20px;
    width:100%;
}
.ok{
    background:#e8f5e9;
    color:#176b2c;
    padding:12px;
    border-radius:8px;
    margin-bottom:15px;
}
.error{
    background:#ffe5e5;
    color:#a10000;
    padding:12px;
    border-radius:8px;
    margin-bottom:15px;
}
.member-id{
    background:#eef2ff;
    padding:14px;
    border-radius:10px;
    font-size:22px;
    font-weight:bold;
    margin-bottom:18px;
}
table{
    width:100%;
    border-collapse:collapse;
    margin-top:15px;
}
th,td{
    border:1px solid #ddd;
    padding:10px;
    text-align:center;
}
th{
    background:#eeeeee;
}
.note{
    background:#fff7d6;
    padding:12px;
    border-radius:10px;
    margin-top:15px;
}
</style>
</head>
<body>

<div class="box">

    <p>
        <a href="/member/member-management/search">← 返回搜索</a>
        &nbsp; | &nbsp;
        <a href="/member/member-management">返回会员资料管理</a>
    </p>

    <h1>✏ 编辑会员</h1>

    {% if error %}
    <div class="error">❌ {{ error }}</div>
    {% endif %}

    {% if msg %}
    <div class="ok">✅ {{ msg }}</div>
    {% endif %}

    <div class="member-id">
        会员编号：{{ member.member_id }}
    </div>

    <form method="post">

        <label>姓名</label>
        <input name="name" value="{{ member.name or '' }}" required>

        <label>英文名</label>
        <input name="english_name" value="{{ member.english_name or '' }}">

        <label>身份证号码</label>
        <input name="ic_number"
               value="{{ member.ic_number or '' }}"
               placeholder="例如：800101-14-1234">

        <label>电话</label>
        <input name="phone" value="{{ member.phone or '' }}">

        <label>PIN</label>
        <input name="pin" value="{{ member.pin or '' }}">

        <label>分会</label>
        <select name="branch">
            <option value="CHE" {% if member.branch=="CHE" %}selected{% endif %}>CHE</option>
            <option value="STW" {% if member.branch=="STW" %}selected{% endif %}>STW</option>
        </select>

        <label>状态</label>
        <select name="member_status">
            <option value="在供" {% if member.member_status=="在供" or member.status=="在供" %}selected{% endif %}>在供</option>
            <option value="停供" {% if member.member_status=="停供" or member.status=="停供" %}selected{% endif %}>停供</option>
            <option value="暂停" {% if member.member_status=="暂停" or member.status=="暂停" %}selected{% endif %}>暂停</option>
            <option value="已往生" {% if member.member_status=="已往生" or member.status=="已往生" %}selected{% endif %}>已往生</option>
        </select>

        <label>备注</label>
        <textarea name="remark">{{ member.remark or '' }}</textarea>

        <button type="submit">
            💾 保存修改
        </button>

    </form>

    <div class="note">
        说明：会员编号不可修改；若需要转分会，之后会另外做「转分会」功能。
    </div>

    <h2>📜 最近会员资料修改历史</h2>

    {% if history_rows %}
    <table>
        <tr>
            <th>时间</th>
            <th>动作</th>
            <th>修改者</th>
        </tr>
        {% for h in history_rows %}
        <tr>
            <td>{{ (h.changed_at + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M") if h.changed_at else "-" }}</td>
            <td>{{ h.action }}</td>
            <td>{{ h.changed_by }}</td>
        </tr>
        {% endfor %}
    </table>
    {% else %}
        <div class="note">暂时没有会员资料修改历史。</div>
    {% endif %}

</div>

</body>
</html>
""",
member=member,
msg=msg,
error=error,
history_rows=history_rows,
timedelta=timedelta
)


@member_bp.route("/member-management/add", methods=["GET", "POST"])
def add_member():

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))

    msg = None
    error = None

    if request.method == "POST":
        branch = request.form.get("branch", "CHE").strip().upper()
        name = request.form.get("name", "").strip()
        english_name = request.form.get("english_name", "").strip()
        phone = request.form.get("phone", "").strip()
        remark = request.form.get("remark", "").strip()
        ic_number = request.form.get("ic_number", "").strip()

        if branch not in ["CHE", "STW"]:
            error = "分会不正确"
        elif not name:
            error = "姓名不能为空"
        else:
            last_no = db_query("""
                select max(
                    cast(regexp_replace(member_id, '^(CHE-|STW-)', '') as integer)
                ) as last_no
                from members
                where member_id like %s
            """, (branch + "-%",), fetchone=True)

            next_no = int(last_no["last_no"] or 0) + 1
            member_id = f"{branch}-{next_no}"

            phone_digits = "".join(ch for ch in phone if ch.isdigit())
            pin = phone_digits[-4:] if len(phone_digits) >= 4 else "0000"

            exist = db_query("""
                select member_id,name
                from members
                where ic_number=%s
            """, (ic_number,), fetchone=True)

            if exist:
                error = f"""
                身份证号码已存在：
                {exist['member_id']}
                {exist['name']}
                """

            db_query("""
                insert into members
                (
                    member_id,
                    name,
                    english_name,
                    ic_number,
                    phone,
                    pin,
                    branch,
                    status,
                    remark,
                    member_status
                )
                values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                member_id,
                name,
                english_name,
                ic_number,
                phone,
                pin,
                branch,
                "在供",
                remark or "在供",
                "在供"
            ))

            msg = f"新增成功：{member_id} / {name} / PIN：{pin}"

    return render_template_string("""
<div style="max-width:900px;margin:auto;padding:20px;">

    <p>
        <a href="/member/member-management">← 返回会员资料管理</a>
    </p>

    <h1>➕ 新增会员</h1>

    {% if error %}
    <div style="background:#ffe5e5;color:#a10000;padding:12px;border-radius:8px;">
        ❌ {{ error }}
    </div>
    {% endif %}

    {% if msg %}
    <div style="background:#e8f5e9;color:#176b2c;padding:12px;border-radius:8px;">
        ✅ {{ msg }}
    </div>
    {% endif %}

    <form method="post">

        <label>分会</label><br>
        <select name="branch" style="width:100%;padding:12px;font-size:20px;">
            <option value="CHE">CHE</option>
            <option value="STW">STW</option>
        </select><br><br>

        <label>姓名</label><br>
        <input name="name" required style="width:100%;padding:12px;font-size:20px;"><br><br>

        <label>英文名</label><br>
        <input name="english_name" style="width:100%;padding:12px;font-size:20px;"><br><br>
                                  
        <label>身份证号码</label>
        <input
            name="ic_number"
            placeholder="例如：800101-14-1234"
            style="width:100%;padding:12px;font-size:20px;">

        <br><br>

        <label>电话</label><br>
        <input name="phone" style="width:100%;padding:12px;font-size:20px;"><br><br>

        <label>备注</label><br>
        <textarea name="remark" style="width:100%;height:100px;padding:12px;font-size:20px;"></textarea><br><br>

        <button type="submit"
            style="background:#2d7ff9;color:white;border:0;padding:14px 22px;border-radius:10px;font-size:22px;font-weight:bold;">
            新增会员
        </button>

    </form>
</div>
""", msg=msg, error=error)


@member_bp.route("/finance-upload", methods=["GET", "POST"])
def finance_upload():

    if not session.get("member_admin"):
        return redirect(url_for("member.member_admin"))
    
    error = None
    msg = None
    rows = []
    history_rows = []
    safety_issues = []

    q = request.args.get("q", "").strip()
    branch = request.args.get("branch", "CHE").strip().upper()
    year = request.args.get("year", str(date.today().year)).strip()
    months = request.args.getlist("months")
    
    # 下载 Excel
    if request.args.get("download") == "1":
        try:
            params = []
            where = "where 1=1"

            if year:
                where += " and to_char(payment_date, 'YYYY') = %s"
                params.append(year)

            if months:
                where += " and to_char(payment_date, 'MM') = any(%s)"
                params.append(months)

            df = pd.read_sql_query(f"""
                select
                    payment_date as "日期\nDate",
                    receipt_no as "收据编号 \nOfficial Receipt No",
                    regexp_replace(member_id, '^(CHE-|STW-)', '') as "编号 No/",
                    name as "捐款人\n姓名\nName",
                    start_month as "START MONTH",
                    end_month as "END MONTH",
                    month_count as "No/ of Mth",
                    amount as "Total Amt"
                from member_payments
                {where}
                order by payment_date asc, id asc
            """, get_conn(), params=params)

            output = BytesIO()

            with pd.ExcelWriter(output, engine="openpyxl") as writer:

                df.to_excel(
                    writer,
                    index=False,
                    sheet_name="月费记录",
                    startrow=5
                )

                ws = writer.book["月费记录"]

                # ======================
                # 标题
                # ======================

                title = f"{year}年{'全部月份' if months == 'all' else '、'.join(months) + '月'}月费记录"

                ws.merge_cells("A1:H1")
                ws["A1"] = title

                ws["A1"].font = Font(
                    size=18,
                    bold=True,
                    color="FFFFFF"
                )

                ws["A1"].fill = PatternFill(
                    "solid",
                    fgColor="4472C4"
                )

                ws["A1"].alignment = Alignment(horizontal="center")

                # ======================
                # 摘要
                # ======================

                total_people = len(df)
                total_months = int(df["No/ of Mth"].sum())
                total_amount = float(df["Total Amt"].sum())

                # ======================
                # 摘要卡片
                # ======================

                summary_fill = PatternFill("solid", fgColor="EAF4FF")
                summary_border = Border(
                    left=Side(style="thin", color="9EADCC"),
                    right=Side(style="thin", color="9EADCC"),
                    top=Side(style="thin", color="9EADCC"),
                    bottom=Side(style="thin", color="9EADCC")
                )

                summary_items = [
                    ("A3:B4", "总人数", f"{total_people} 人"),
                    ("D3:E4", "总月数", f"{total_months}"),
                    ("G3:H4", "总金额", f"RM {total_amount:,.2f}"),
                ]

                for cell_range, label, value in summary_items:
                    ws.merge_cells(cell_range)
                    cell = ws[cell_range.split(":")[0]]
                    cell.value = f"{label}\n{value}"
                    cell.fill = summary_fill
                    cell.border = summary_border
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )

                ws.row_dimensions[3].height = 28
                ws.row_dimensions[4].height = 28

                # ======================
                # 表头美化
                # ======================

                header_fill = PatternFill("solid", fgColor="0F5F76")
                header_font = Font(color="FFFFFF", bold=True)

                row_fill = PatternFill("solid", fgColor="BFE8F5")

                for cell in ws[6]:

                    cell.font = Font(
                        bold=True
                    )

                    cell.fill = header_fill

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )

                # ======================
                # 边框样式
                # ======================

                thin = Side(
                    border_style="thin",
                    color="CCCCCC"
                )

                border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin
                )

                # ======================
                # 自动栏宽
                # ======================

                for col in ws.columns:

                    max_len = 0

                    letter = get_column_letter(col[0].column)

                    for cell in col:

                        try:
                            max_len = max(
                                max_len,
                                len(str(cell.value))
                            )
                        except:
                            pass

                    ws.column_dimensions[letter].width = min(max_len + 4, 30)

                # ======================
                # 冻结标题
                # ======================

                ws.freeze_panes = "A7"

                # ======================
                # 自动筛选
                # ======================

                # ======================
                # 表格边框
                # ======================

                blue_fill = PatternFill(
                    "solid",
                    fgColor="D8EEF8"
                )

                white_fill = PatternFill(
                    "solid",
                    fgColor="FFFFFF"
                )

                for row_num in range(7, ws.max_row + 1):

                    fill = blue_fill if row_num % 2 == 0 else white_fill

                    for col in range(1, 9):

                        cell = ws.cell(row=row_num, column=col)

                        cell.fill = fill
                        cell.border = border

                ws.auto_filter.ref = f"A6:H{ws.max_row}"

                # ======================
                # 金额格式
                # ======================

                for row in range(7, ws.max_row + 1):

                    ws[f"A{row}"].number_format = "dd/mm/yyyy"

                    ws[f"H{row}"].number_format = '#,##0.00'

                    ws[f"H{row}"].alignment = Alignment(
                        horizontal="right"
                    )

            output.seek(0)

            month_text = "全部" if not months else "_".join(months)
            filename = f"{year}_{month_text}_月费记录.xlsx"

            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            error = f"下载失败：{e}"

    # 上传 Excel
    if request.method == "POST":

        file = request.files.get("file")

        if not file or file.filename == "":
            error = "请选择 Excel 文件"
        else:
            try:
                df = pd.read_excel(file)

                df = df[[
                    "日期\nDate",
                    "收据编号 \nOfficial Receipt No",
                    "编号 No/",
                    "捐款人\n姓名\nName",
                    "START MONTH",
                    "END MONTH",
                    "No/ of Mth",
                    "Total Amt"
                ]]

                df = df.dropna(subset=["编号 No/"])

                inserted = 0
                skipped = 0

                for _, row in df.iterrows():
                    try:
                        receipt_raw = row["收据编号 \nOfficial Receipt No"]

                        if pd.isna(receipt_raw):
                            receipt_no = None
                        else:
                            receipt_no = str(receipt_raw).strip().replace(" ", "")
                            if receipt_no == "":
                                receipt_no = None

                        member_no = int(row["编号 No/"])
                        member_id = f"CHE-{member_no}"

                        member_info = db_query("""
                            select name
                            from members
                            where member_id = %s
                        """, (member_id,), fetchone=True)

                        if not member_info:
                            print("会员不存在：", member_id)
                            skipped += 1
                            continue
                                                
                        system_name = member_info["name"]

                        excel_name = str(
                            row["捐款人\n姓名\nName"]
                        ).strip()

                        name_warnings = {}

                        if excel_name != system_name:
                            name_warnings[member_id] = (
                                f"{member_id}：Excel={excel_name}，Members={system_name}"
                            )
                                             
                        name = system_name

                        payment_date = pd.to_datetime(
                            row["日期\nDate"]
                        ).date()

                        start_month = parse_month(
                            row["START MONTH"]
                        )

                        end_month = parse_month(
                            row["END MONTH"]
                        )

                        month_count = int(row["No/ of Mth"])
                        amount = float(row["Total Amt"])

                        result = db_query("""
                            insert into member_payments
                            (
                                receipt_no,
                                member_id,
                                name,
                                payment_date,
                                start_month,
                                end_month,
                                month_count,
                                amount
                            )
                            values
                            (
                                %s, %s, %s, %s,
                                %s, %s, %s, %s
                            )
                            on conflict (receipt_no) do nothing
                            returning id
                        """, (
                            receipt_no,
                            member_id,
                            name,
                            payment_date,
                            start_month,
                            end_month,
                            month_count,
                            amount
                        ))

                        if result:
                            inserted += 1
                        else:
                            skipped += 1

                    except Exception as e:
                        print("导入失败:", row.to_dict(), e)
                        skipped += 1

                msg = f"上传完成：读取 {len(df)} 行，新增 {inserted} 行，跳过 {skipped} 行。"

                if name_warnings:

                    print("=" * 60)
                    print(f"发现 {len(name_warnings)} 位会员姓名不一致：")

                    for w in sorted(name_warnings.values()):
                        print(w)

                    print("=" * 60)

            except Exception as e:
                print("上传失败:", e)
                error = f"上传失败：{e}"

    # 搜索记录
    try:
        if q:
            keyword = q.strip()

            if keyword.isdigit():
                if branch == "STW":
                    member_id = normalize_member_id(f"STW-{keyword}")
                else:
                    member_id = normalize_member_id(keyword)
            else:
                member_id = normalize_member_id(keyword)

            with get_conn() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    if keyword.isdigit() and len(keyword) <= 6:
                        cur.execute("""
                            select *
                            from member_payments
                            where member_id = %s
                            order by payment_date desc, id desc
                            limit 200
                        """, (member_id,))
                    else:
                        cur.execute("""
                            select *
                            from member_payments
                            where member_id ilike %s
                               or name ilike %s
                               or receipt_no ilike %s
                            order by payment_date desc, id desc
                            limit 200
                        """, (
                            f"%{keyword}%",
                            f"%{keyword}%",
                            f"%{keyword}%"
                        ))

                    rows = cur.fetchall()

    except Exception as e:
        error = f"搜索失败：{e}"

    try:
        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    select *
                    from member_payment_history
                    order by changed_at desc
                    limit 30
                """)
                history_rows = cur.fetchall()

                # =========================
                # 财政安全检查
                # =========================

                # 1. 重复收据
                cur.execute("""
                    select
                        receipt_no,
                        count(*) as cnt
                    from member_payments
                    where receipt_no is not null
                    and receipt_no <> ''
                    group by receipt_no
                    having count(*) > 1
                """)

                for r in cur.fetchall():
                    safety_issues.append({
                        "type": "重复收据",
                        "detail": f"收据 {r['receipt_no']} 出现 {r['cnt']} 次"
                    })

                # 2. 金额和月数不符
                cur.execute("""
                    select
                        id,
                        member_id,
                        receipt_no,
                        amount,
                        month_count
                    from member_payments
                """)

                for r in cur.fetchall():

                    amount = float(r["amount"] or 0)
                    months = int(r["month_count"] or 0)

                    if months > 0:

                        expected = months * 50

                        if amount != expected:

                            safety_issues.append({
                                "type": "金额不符",
                                "detail":
                                f"{r['member_id']} / {r['receipt_no']}：RM {amount:.2f}，应为 RM {expected:.2f}",
                                "payment_id": r["id"]
                            })

                # 3. 未来日期
                cur.execute("""
                    select
                        id,
                        member_id,
                        receipt_no,
                        payment_date
                    from member_payments
                    where payment_date > current_date
                """)

                for r in cur.fetchall():

                    safety_issues.append({
                        "type": "未来日期",
                        "detail":
                        f"{r['member_id']} / {r['receipt_no']}：{r['payment_date']}",
                        "payment_id": r["id"]
                    })

                # 4. 会员编号不存在
                cur.execute("""
                    select
                        p.id,
                        p.member_id,
                        p.receipt_no
                    from member_payments p
                    left join members m
                        on p.member_id = m.member_id
                    where m.member_id is null
                """)

                for r in cur.fetchall():

                    safety_issues.append({
                        "type": "会员不存在",
                        "detail":
                        f"{r['member_id']} / {r['receipt_no']}",
                        "payment_id": r["id"]
                    })

    except Exception as e:
        print("读取修改历史失败:", e)

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
    <meta charset="utf-8">
    <title>月费管理中心</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">

    <link rel="stylesheet"
        href="{{ url_for('static', filename='css/toolbox.css') }}">

    <style>
    .member-center .section-title{
        font-size:26px;
        margin-bottom:16px;
    }
    .member-center .card{
        padding:22px;
    }
    .member-center .btn-tool{
        min-height:54px;
        font-size:21px;
    }
    .member-grid{
        display:grid;
        grid-template-columns:1fr 1fr;
        gap:16px;
    }
    .month-grid{
        display:grid;
        grid-template-columns:repeat(4,1fr);
        gap:10px;
    }
    .month-option{
        display:flex;
        align-items:center;
        gap:8px;
        font-size:20px;
        font-weight:bold;
        cursor:pointer;
        padding:10px;
        border:1px solid #e5e7eb;
        border-radius:12px;
        background:white;
    }
    .month-option input{
        width:22px;
        height:22px;
    }
    .search-row{
        display:flex;
        gap:10px;
        align-items:center;
    }
    .branch-small-btn{
        width:95px;
        height:54px;
        font-size:20px;
        font-weight:bold;
        background:#16a34a;
        color:white;
        border:none;
        border-radius:12px;
        cursor:pointer;
        flex-shrink:0;
    }
    @media(max-width:700px){
        .member-grid{
            grid-template-columns:1fr;
        }
        .month-grid{
            grid-template-columns:repeat(2,1fr);
        }
        .search-row{
            align-items:stretch;
        }
    }
    </style>
    </head>

    <body>

    <div class="page member-center">

        <h1 class="page-title">📂 月费管理中心</h1>
        <p class="page-subtitle">上传、下载、搜索、检查月费资料</p>

        <div class="card">
            <a class="btn-tool btn-secondary" href="/member/admin">
                ← 返回月费管理员
            </a>
        </div>

        {% if error %}
        <div class="alert alert-danger">❌ {{ error }}</div>
        {% endif %}

        {% if msg %}
        <div class="alert alert-success">✅ {{ msg }}</div>
        {% endif %}

        <div class="member-grid">

            <div class="card">
                <div class="section-title">① 上传单月 Excel</div>

                <form method="post" enctype="multipart/form-data">
                    <div class="form-group">
                        <label>选择 Excel 文件</label>
                        <input class="form-input" name="file" type="file" accept=".xlsx,.xls" required>
                    </div>

                    <button class="btn-tool btn-primary" type="submit" style="width:100%;">
                        ⬆️ 上传 Excel
                    </button>
                </form>

                <div class="alert alert-warning">
                    系统会保留旧数据，只导入新记录。重复收据编号会自动跳过。
                </div>
            </div>

            <div class="card">
                <div class="section-title">② 下载月费 Excel</div>

                <form method="get">
                    <input type="hidden" name="download" value="1">

                    <div class="form-group">
                        <label>年份</label>
                        <input class="form-input" name="year" value="{{ year }}">
                    </div>

                    <div class="form-group">
                        <label>月份</label>
                        <div class="small-text">不打勾 = 下载全年，可同时选择多个月</div>
                    </div>

                    <div class="month-grid">
                        {% for m in range(1,13) %}
                        <label class="month-option">
                            <input
                                type="checkbox"
                                name="months"
                                value="{{ '%02d'|format(m) }}"
                            >
                            {{ m }}月
                        </label>
                        {% endfor %}
                    </div>

                    <button class="btn-tool btn-primary" type="submit" style="width:100%;">
                        ⬇️ 下载 Excel
                    </button>
                </form>
            </div>

        </div>

        <div class="card">
            <div class="section-title">③ 搜索月费记录</div>

            <form method="get">
                <div class="form-group">
                    <label>寻找编号 / 姓名 / 收据编号</label>

                    <div class="search-row">
                        <button
                            type="button"
                            id="search_branch_btn"
                            onclick="toggleSearchBranch()"
                            class="branch-small-btn">
                            {{ branch }}
                        </button>

                        <input
                            type="hidden"
                            id="search_branch"
                            name="branch"
                            value="{{ branch }}">

                        <input
                            class="form-input"
                            name="q"
                            value="{{ q }}"
                            placeholder="例如：108 / 张三 / CHE0001493"
                            style="flex:1;">
                    </div>
                </div>

                <button class="btn-tool btn-primary" type="submit" style="width:100%;">
                    🔍 搜索记录
                </button>
            </form>

            {% if q %}
            <div class="section-title" style="margin-top:22px;">
                搜索结果：{{ rows|length }} 笔
            </div>

            {% if rows %}
            <div class="table-responsive">
                <table class="record-table">
                    <tr>
                        <th>付款日期</th>
                        <th>月费编号</th>
                        <th>姓名</th>
                        <th>收据编号</th>
                        <th>开始月份</th>
                        <th>结束月份</th>
                        <th>月数</th>
                        <th>金额</th>
                        <th>操作</th>
                    </tr>

                    {% for r in rows %}
                    <tr>
                        <td>{{ r.payment_date.strftime("%Y-%m-%d") if r.payment_date else "-" }}</td>
                        <td>{{ r.member_id }}</td>
                        <td>{{ r.name }}</td>
                        <td>{{ r.receipt_no }}</td>
                        <td>{{ r.start_month.strftime("%Y-%m") if r.start_month else "-" }}</td>
                        <td>{{ r.end_month.strftime("%Y-%m") if r.end_month else "-" }}</td>
                        <td>{{ r.month_count }}</td>
                        <td>RM {{ "%.2f"|format(r.amount or 0) }}</td>
                        <td>
                            <a class="btn-tool btn-primary mini-btn"
                            href="/member/payment/edit/{{ r.id }}">
                                ✏ 编辑
                            </a>
                        </td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% else %}
            <div class="empty-state">
                <div class="empty-title">找不到相关记录</div>
            </div>
            {% endif %}
            {% endif %}
        </div>

        <div class="card">
            <div class="section-title">④ 修改历史</div>

            {% if history_rows %}
            <div class="table-responsive">
                <table class="record-table">
                    <tr>
                        <th>时间</th>
                        <th>会员编号</th>
                        <th>收据编号</th>
                        <th>动作</th>
                        <th>修改者</th>
                        <th>详情</th>
                    </tr>

                    {% for h in history_rows %}
                    <tr>
                        <td>{{ (h.changed_at + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M") if h.changed_at else "-" }}</td>
                        <td>{{ h.member_id }}</td>
                        <td>{{ h.receipt_no }}</td>
                        <td>{{ h.action }}</td>
                        <td>{{ h.changed_by }}</td>
                        <td>
                            <a class="btn-tool btn-primary mini-btn"
                            href="/member/payment/history/{{ h.id }}">
                                查看详情
                            </a>
                        </td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% else %}
            <div class="empty-state">
                <div class="empty-title">暂时没有修改记录</div>
            </div>
            {% endif %}
        </div>

        <div class="card">
            <div class="section-title">⑤ 月费安全检查</div>

            {% if safety_issues %}
            <div class="alert alert-danger">
                ⚠️ 发现 {{ safety_issues|length }} 个问题
            </div>

            <div class="table-responsive">
                <table class="record-table">
                    <tr>
                        <th>类型</th>
                        <th>问题</th>
                        <th>操作</th>
                    </tr>

                    {% for s in safety_issues %}
                    <tr>
                        <td>{{ s.type }}</td>
                        <td>{{ s.detail }}</td>
                        <td>
                            {% if s.payment_id %}
                            <a class="btn-tool btn-primary mini-btn"
                            href="/member/payment/edit/{{ s.payment_id }}">
                                ✏ 编辑
                            </a>
                            {% else %}
                            -
                            {% endif %}
                        </td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% else %}
            <div class="alert alert-success">
                ✅ 没有发现问题
            </div>
            {% endif %}
        </div>

        <div class="card">
            <div class="section-title">⑥ 其他管理</div>

            <div class="member-grid">
                <a class="btn-tool btn-primary" href="/member/member-management">
                    👤 会员资料管理
                </a>

                <a class="btn-tool btn-purple" href="/member/query-logs">
                    🔍 月费查询记录
                </a>
            </div>
        </div>

    </div>

    <script>
    function toggleSearchBranch(){

        const btn = document.getElementById("search_branch_btn");
        const branch = document.getElementById("search_branch");

        if(branch.value==="CHE"){
            branch.value="STW";
            btn.innerText="STW";
            btn.style.background="#dc2626";
        }else{
            branch.value="CHE";
            btn.innerText="CHE";
            btn.style.background="#16a34a";
        }
    }
    </script>

    </body>
    </html>
    """,
    error=error,
    msg=msg,
    rows=rows,
    q=q,
    year=year,
    branch=branch,
    months=months,
    timedelta=timedelta,
    safety_issues=safety_issues,
    history_rows=history_rows
    )

MEMBER_HTML = """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>月费查询</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="manifest" href="/member-manifest.json?v=2">
<link rel="icon" href="/static/member_icon.png?v=2">

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">

<style>
.member-search-row{
    display:flex;
    gap:10px;
    align-items:center;
}
.branch-btn{
    width:100px;
    min-height:64px;
    font-size:24px;
    font-weight:800;
    border:0;
    border-radius:16px;
    color:white;
    background:#16a34a;
    flex-shrink:0;
    cursor:pointer;
}
.member-title-line{
    font-size:30px;
    font-weight:800;
}
.member-sub{
    color:#666;
    font-size:16px;
    margin-top:6px;
}
.record-table{
    width:100%;
    border-collapse:collapse;
    background:white;
    margin-top:14px;
}
.record-table th,
.record-table td{
    border:1px solid #e5e7eb;
    padding:10px;
    text-align:center;
    font-size:15px;
}
.record-table th{
    background:#f3f4f6;
}
.candidate-card{
    background:white;
    border:1px solid #e5e7eb;
    border-radius:16px;
    padding:16px;
    margin-top:12px;
}

.member-summary-grid{
    grid-template-columns:repeat(3, 1fr);
    margin-top:18px;
}

.member-summary-box{
    text-align:left;
    box-shadow:none;
    border:1px solid #e5e7eb;
    padding:16px 18px;
}

.member-summary-value{
    font-size:26px;
    color:#111827;
}

.member-last-payment{
    margin-top:14px;
}

@media(max-width:700px){
    .member-summary-grid{
        grid-template-columns:1fr;
    }

    .member-summary-value{
        font-size:24px;
    }
}

@media(max-width:700px){
    .member-search-row{
        align-items:stretch;
    }
    .branch-btn{
        width:82px;
        font-size:21px;
    }
    .record-table th,
    .record-table td{
        font-size:12px;
        padding:6px;
    }
    .member-title-line{
        font-size:24px;
    }
}
</style>
</head>

<body>

<div class="page">

    <div class="btn-row" style="justify-content:space-between;">
        <a class="btn-tool btn-light" href="/member/admin">⚙ 管理员入口</a>
    </div>

    <h1 class="page-title">🙏 月费查询</h1>
    <p class="page-subtitle">请输入月费编号、姓名、英文名或电话查询供养记录</p>

    {% if query_volunteer %}
    <div class="alert alert-warning">
        当前查询义工：
        <b>{{ query_volunteer.id }} {{ query_volunteer.name }}</b>

        <div style="margin-top:12px;">
            <a class="btn-tool btn-danger" href="/member/query-logout">
                🔒 退出查询
            </a>
        </div>
    </div>
    {% endif %}

    {% if error %}
    <div class="alert alert-danger">
        ❌ {{ error }}
    </div>
    {% endif %}

    <div class="card">
        <div class="section-title">🔍 查询资料</div>

        <form method="post">
            <div class="form-group">
                <label>月费编号 / 姓名 / 英文名 / 电话</label>

                <div class="member-search-row">
                    <button
                        type="button"
                        id="branch_btn"
                        class="branch-btn"
                        onclick="toggleBranch()"
                    >
                        CHE
                    </button>

                    <input type="hidden" id="branch" name="branch" value="CHE">

                    <input
                        class="form-input"
                        id="member_id"
                        name="member_id"
                        placeholder="例如：108 / 张三 / 0123456789"
                        autocomplete="off"
                        required
                    >
                </div>
            </div>

            <button class="btn-tool btn-primary" type="submit" style="width:100%;">
                🔍 查询
            </button>
        </form>
    </div>

    {% if candidates %}
    <div class="card">
        <div class="section-title">👥 找到多位会员，请选择</div>

        <form method="post">
            {% for m in candidates %}
            <div class="candidate-card">
                <div class="member-title-line">{{ m.name }}</div>
                <div class="member-sub">
                    编号：<b>{{ m.member_id }}</b><br>
                    {% if m.english_name %}
                    英文名：{{ m.english_name }}<br>
                    {% endif %}
                    电话：{{ m.phone or "-" }}
                </div>

                <button
                    class="btn-tool btn-primary"
                    type="submit"
                    name="selected_member_id"
                    value="{{ m.member_id }}"
                    style="width:100%;margin-top:12px;"
                >
                    查看这个会员
                </button>
            </div>
            {% endfor %}
        </form>
    </div>
    {% endif %}

    {% if member %}
    <div class="card">
        <div class="section-title">👤 会员资料</div>

        <div class="member-title-line">
            {{ member.name }}
        </div>

        <div class="member-sub">
            月费编号：<b>{{ member.member_id }}</b><br>
            {% if member.english_name %}
            英文名：{{ member.english_name }}<br>
            {% endif %}
            电话：{{ member.phone or "-" }}
        </div>

        {% if summary %}
        <div class="summary-grid member-summary-grid">

            <div class="summary-box member-summary-box">
                <div class="summary-title">已付月费总额</div>
                <div class="summary-value member-summary-value">
                    RM {{ "%.2f"|format(summary.total_payment or 0) }}
                </div>
            </div>

            <div class="summary-box member-summary-box">
                <div class="summary-title">累计已付月数</div>
                <div class="summary-value member-summary-value">
                    {{ summary.total_months or 0 }} 个月
                </div>
            </div>

            <div class="summary-box member-summary-box">
                <div class="summary-title">月费已缴至</div>
                <div class="summary-value member-summary-value">
                    {% if summary.paid_until %}
                        {{ summary.paid_until.strftime("%Y-%m") }}
                    {% else %}
                        暂无记录
                    {% endif %}
                </div>
            </div>

        </div>

        <div class="card-soft member-last-payment">
            <div class="summary-title">最后付款记录</div>
            <div class="summary-value member-summary-value">
                {% if summary.last_payment_date %}
                    {{ summary.last_payment_date.strftime("%Y-%m-%d") }}
                    ｜RM {{ "%.2f"|format(summary.last_payment_amount or 0) }}
                {% else %}
                    暂无记录
                {% endif %}
            </div>
        </div>
        {% endif %}
    </div>

    <div class="card">
        <div class="section-title">📋 缴费记录</div>

        {% if payments %}
        <div style="overflow-x:auto;">
            <table class="record-table">
                <tr>
                    <th>付款日期</th>
                    <th>收据编号</th>
                    <th>开始月份</th>
                    <th>结束月份</th>
                    <th>月数</th>
                    <th>金额</th>
                </tr>

                {% for p in payments %}
                <tr>
                    <td>{{ p.payment_date.strftime("%Y-%m-%d") if p.payment_date else "-" }}</td>
                    <td>{{ p.receipt_no }}</td>
                    <td>{{ p.start_month.strftime("%Y-%m") if p.start_month else "-" }}</td>
                    <td>{{ p.end_month.strftime("%Y-%m") if p.end_month else "-" }}</td>
                    <td>{{ p.month_count }}</td>
                    <td>RM {{ "%.2f"|format(p.amount or 0) }}</td>
                </tr>
                {% endfor %}
            </table>
        </div>
        {% else %}
        <div class="empty-state">
            这个会员目前没有缴费记录。
        </div>
        {% endif %}
    </div>
    {% endif %}

</div>

<script>
function toggleBranch() {
    const btn = document.getElementById("branch_btn");
    const branch = document.getElementById("branch");

    if (branch.value === "CHE") {
        branch.value = "STW";
        btn.innerText = "STW";
        btn.style.background = "#dc2626";
    } else {
        branch.value = "CHE";
        btn.innerText = "CHE";
        btn.style.background = "#16a34a";
    }
}
</script>

</body>
</html>
"""

MEMBER_ADMIN_HTML = """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>月费管理员</title>
<meta name="viewport" content="width=device-width, initial-scale=1">

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">

<style>
.admin-link-grid{
    display:grid;
    grid-template-columns:repeat(3, 1fr);
    gap:12px;
}

.member-search-row{
    display:flex;
    gap:10px;
    align-items:center;
}

.branch-btn{
    width:100px;
    min-height:64px;
    font-size:24px;
    font-weight:800;
    border:0;
    border-radius:16px;
    color:white;
    background:#16a34a;
    flex-shrink:0;
    cursor:pointer;
}

.pagination{
    margin-top:20px;
    text-align:center;
}

.pagination a,
.pagination span{
    display:inline-block;
    padding:8px 13px;
    margin:3px;
    border-radius:8px;
    text-decoration:none;
    font-size:16px;
}

.pagination a{
    background:#e5e7eb;
    color:#333;
}

.pagination span{
    background:#2563eb;
    color:white;
    font-weight:bold;
}

@media(max-width:700px){
    .admin-link-grid{
        grid-template-columns:1fr;
    }

    .member-search-row{
        align-items:stretch;
    }

    .branch-btn{
        width:82px;
        font-size:21px;
    }
}
</style>
</head>

<body>

<div class="page">

    <h1 class="page-title">🙏 月费管理员</h1>
    <p class="page-subtitle">管理员查询佛友月费记录、检查缴费资料与编辑记录</p>

    <div class="card">
        <div class="section-title">📂 功能入口</div>

        <div class="admin-link-grid">
            <a class="btn-tool btn-secondary" href="/member">
                ← 月费查询
            </a>

            <a class="btn-tool btn-primary" href="/member/finance-upload">
                📂 资料管理中心
            </a>

            <a class="btn-tool btn-warning" href="/member/admin/late_members">
                ⚠️ 迟付名单
            </a>
        </div>
    </div>

    {% if session.get("member_admin") %}
    <div class="alert alert-success">
        ✅ 管理员已登录

        <div style="margin-top:12px;">
            <a class="btn-tool btn-danger" href="/member/admin/logout">
                🚪 退出管理员
            </a>
        </div>
    </div>
    {% endif %}

    {% if error %}
    <div class="alert alert-danger">
        ❌ {{ error }}
    </div>
    {% endif %}

    <div class="card">
        {% if not session.get("member_admin") %}

        <div class="section-title">🔐 管理员登录</div>

        <form method="post" action="/member/admin">
            <div class="form-group">
                <label>管理员 PIN</label>

                <input
                    class="form-input"
                    id="member_admin_pin"
                    name="admin_pin"
                    type="password"
                    inputmode="numeric"
                    autocomplete="new-password"
                    readonly
                    onfocus="this.removeAttribute('readonly');"
                    required
                >
            </div>

            <button class="btn-tool btn-primary" type="submit" style="width:100%;">
                登录管理员
            </button>
        </form>

        {% else %}

        <div class="section-title">🔍 查询会员</div>

        <form method="post" action="/member/admin">
            <div class="form-group">
                <label>月费编号 / 姓名 / 英文名</label>

                <div class="member-search-row">
                    <button
                        type="button"
                        id="branch_btn"
                        class="branch-btn"
                        onclick="toggleBranch()"
                    >
                        CHE
                    </button>

                    <input type="hidden" id="branch" name="branch" value="CHE">

                    <input
                        class="form-input"
                        name="member_id"
                        placeholder="例如：108 / 张三 / zhangsan"
                        autocomplete="off"
                    >
                </div>
            </div>

            <button class="btn-tool btn-primary" type="submit" style="width:100%;">
                查找会员
            </button>
        </form>

        {% endif %}
    </div>

    {% if member %}
    <div class="info-card">
        <div class="section-title">👤 会员资料</div>

        <div class="person-name">
            {{ member.name }}
        </div>

        <div class="person-meta">
            月费编号：<b>{{ member.member_id }}</b><br>
            {% if member.english_name %}
            英文名：{{ member.english_name }}<br>
            {% endif %}
            电话：{{ member.phone or "-" }}
        </div>
    </div>

    {% if summary %}
    <div class="card">
        <div class="section-title">📊 月费统计</div>

        <div class="info-summary-grid">

            <div class="info-summary-box">
                <div class="info-summary-title">已付月费总额</div>
                <div class="info-summary-value">
                    RM {{ "%.2f"|format(summary.total_payment or 0) }}
                </div>
            </div>

            <div class="info-summary-box">
                <div class="info-summary-title">累计已付月数</div>
                <div class="info-summary-value">
                    {{ summary.total_months or 0 }} 个月
                </div>
            </div>

            <div class="info-summary-box">
                <div class="info-summary-title">月费已缴至</div>
                <div class="info-summary-value">
                    {% if summary.paid_until %}
                        {{ summary.paid_until.strftime("%Y-%m") }}
                    {% else %}
                        暂无记录
                    {% endif %}
                </div>
            </div>

        </div>

        <div class="info-summary-box" style="margin-top:14px;">
            <div class="info-summary-title">最后付款记录</div>
            <div class="info-summary-value">
                {% if summary.last_payment_date %}
                    {{ summary.last_payment_date.strftime("%Y-%m-%d") }}
                    ｜RM {{ "%.2f"|format(summary.last_payment_amount or 0) }}
                {% else %}
                    暂无记录
                {% endif %}
            </div>
        </div>
    </div>
    {% endif %}

    {% if warnings %}
    <div class="alert alert-warning">
        <b>⚠️ 系统发现可能有错误：</b>
        <ul>
            {% for w in warnings %}
                <li>{{ w }}</li>
            {% endfor %}
        </ul>
    </div>
    {% endif %}

    <div class="card">
        <div class="section-title">📋 缴费记录</div>

        {% if payments %}
        <div class="table-responsive">
            <table class="record-table">
                <tr>
                    <th>日期<br>Date</th>
                    <th>收据编号<br>Official Receipt No</th>
                    <th>编号 No.</th>
                    <th>捐款人<br>姓名</th>
                    <th>START MONTH</th>
                    <th>END MONTH</th>
                    <th>No. of Mth</th>
                    <th>Total Amt</th>
                    <th>操作</th>
                </tr>

                {% for p in payments %}
                <tr>
                    <td>{{ p.payment_date.strftime("%d/%m/%Y") if p.payment_date else "-" }}</td>
                    <td>{{ p.receipt_no }}</td>
                    <td>{{ p.member_id.replace("CHE-", "").replace("STW-", "") if p.member_id else "-" }}</td>
                    <td>{{ p.name or member.name }}</td>
                    <td>{{ p.start_month.strftime("%b-%y") if p.start_month else "-" }}</td>
                    <td>{{ p.end_month.strftime("%b-%y") if p.end_month else "-" }}</td>
                    <td>{{ p.month_count }}</td>
                    <td>{{ "%.2f"|format(p.amount or 0) }}</td>
                    <td>
                        <a class="btn-tool btn-primary mini-btn"
                           href="/member/payment/edit/{{ p.id }}">
                            ✏ 编辑
                        </a>
                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>

        {% if total_pages and total_pages > 1 %}
        <div class="pagination">
            {% for p in range(1, total_pages + 1) %}
                {% if p == page %}
                    <span>{{ p }}</span>
                {% else %}
                    <a href="/member/admin?member_id={{ raw_member_id }}&page={{ p }}">{{ p }}</a>
                {% endif %}
            {% endfor %}
        </div>
        {% endif %}

        {% else %}
        <div class="empty-state">
            <div class="empty-icon">📭</div>
            <div class="empty-title">暂无缴费记录</div>
            <div class="empty-text">这个会员目前没有缴费记录。</div>
        </div>
        {% endif %}
    </div>
    {% endif %}

</div>

<script>
function toggleBranch() {
    const btn = document.getElementById("branch_btn");
    const branch = document.getElementById("branch");

    if (branch.value === "CHE") {
        branch.value = "STW";
        btn.innerText = "STW";
        btn.style.background = "#dc2626";
    } else {
        branch.value = "CHE";
        btn.innerText = "CHE";
        btn.style.background = "#16a34a";
    }
}
</script>

</body>
</html>
"""

PAYMENT_EDIT_HTML = """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>修改缴费记录</title>
<meta name="viewport" content="width=device-width, initial-scale=1">

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">
</head>

<body>

<div class="page">

    <h1 class="page-title">✏ 修改缴费记录</h1>
    <p class="page-subtitle">请确认日期、收据编号、月份与金额是否正确</p>

    <div class="card">
        <a class="btn-tool btn-secondary"
           href="/member/admin?member_id={{ payment.member_id }}">
            ← 返回会员记录
        </a>
    </div>

    {% if error %}
    <div class="alert alert-danger">
        ❌ {{ error }}
    </div>
    {% endif %}

    <div class="card">
        <div class="section-title">📋 缴费资料</div>

        <form method="post">

            <div class="form-group">
                <label>日期 Date</label>
                <input
                    class="form-input"
                    type="date"
                    name="payment_date"
                    value="{{ payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '' }}"
                    required
                >
            </div>

            <div class="form-group">
                <label>收据编号 Official Receipt No</label>
                <input
                    class="form-input"
                    name="receipt_no"
                    value="{{ payment.receipt_no or '' }}"
                    required
                >
            </div>

            <div class="form-group">
                <label>编号 No.</label>
                <input
                    class="form-input"
                    value="{{ payment.member_id.replace('CHE-', '').replace('STW-', '') }}"
                    disabled
                >
            </div>

            <div class="form-group">
                <label>捐款人 姓名 Name</label>
                <input
                    class="form-input"
                    name="name"
                    value="{{ payment.name or '' }}"
                    required
                >
            </div>

            <div class="form-grid">
                <div class="form-group">
                    <label>START MONTH</label>
                    <input
                        class="form-input"
                        type="month"
                        name="start_month"
                        value="{{ payment.start_month.strftime('%Y-%m') if payment.start_month else '' }}"
                        required
                    >
                </div>

                <div class="form-group">
                    <label>END MONTH</label>
                    <input
                        class="form-input"
                        type="month"
                        name="end_month"
                        value="{{ payment.end_month.strftime('%Y-%m') if payment.end_month else '' }}"
                        required
                    >
                </div>
            </div>

            <div class="form-grid">
                <div class="form-group">
                    <label>No. of Mth</label>
                    <input
                        class="form-input"
                        type="number"
                        name="month_count"
                        value="{{ payment.month_count or 1 }}"
                        required
                    >
                </div>

                <div class="form-group">
                    <label>Total Amt</label>
                    <input
                        class="form-input"
                        type="number"
                        step="0.01"
                        name="amount"
                        value="{{ payment.amount or 0 }}"
                        required
                    >
                </div>
            </div>

            <button class="btn-tool btn-primary" type="submit" style="width:100%;">
                💾 保存修改
            </button>

        </form>
    </div>

</div>

</body>
</html>
"""

CHANGE_PIN_HTML = """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>更改月费密码</title>
<meta name="viewport" content="width=device-width, initial-scale=1">

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">
</head>

<body>

<div class="page">

    <h1 class="page-title">🔐 更改月费密码</h1>
    <p class="page-subtitle">请输入会员编号、旧密码及新密码</p>

    <div class="card">
        <a class="btn-tool btn-secondary" href="/member">
            ← 返回月费查询
        </a>
    </div>

    {% if error %}
    <div class="alert alert-danger">
        ❌ {{ error }}
    </div>
    {% endif %}

    {% if ok %}
    <div class="alert alert-success">
        ✅ {{ ok }}
    </div>
    {% endif %}

    <div class="card">

        <div class="section-title">
            🔑 修改密码
        </div>

        <form method="post">

            <div class="form-group">
                <label>月费编号</label>

                <input
                    class="form-input"
                    name="member_id"
                    placeholder="例如：108 / CHE-108 / STW-108"
                    autocomplete="off"
                    required
                >
            </div>

            <div class="form-group">
                <label>旧密码</label>

                <input
                    class="form-input"
                    name="old_pin"
                    type="password"
                    inputmode="numeric"
                    autocomplete="new-password"
                    autocorrect="off"
                    autocapitalize="off"
                    spellcheck="false"
                    readonly
                    onfocus="this.removeAttribute('readonly');"
                    required
                >
            </div>

            <div class="form-group">
                <label>新密码</label>

                <input
                    class="form-input"
                    name="new_pin"
                    type="password"
                    inputmode="numeric"
                    autocomplete="new-password"
                    autocorrect="off"
                    autocapitalize="off"
                    spellcheck="false"
                    readonly
                    onfocus="this.removeAttribute('readonly');"
                    required
                >
            </div>

            <div class="form-group">
                <label>确认新密码</label>

                <input
                    class="form-input"
                    name="confirm_pin"
                    type="password"
                    inputmode="numeric"
                    autocomplete="new-password"
                    autocorrect="off"
                    autocapitalize="off"
                    spellcheck="false"
                    readonly
                    onfocus="this.removeAttribute('readonly');"
                    required
                >
            </div>

            <button
                class="btn-tool btn-primary"
                style="width:100%;"
                type="submit">
                💾 确认更改密码
            </button>

        </form>

    </div>

</div>

</body>
</html>
"""

def format_history_changes(old_data, new_data):
    old = old_data or {}
    new = new_data or {}

    labels = {
        "payment_date": "付款日期",
        "receipt_no": "收据编号",
        "name": "姓名",
        "start_month": "开始月份",
        "end_month": "结束月份",
        "month_count": "月数",
        "amount": "金额",
    }

    changes = []

    for key, label in labels.items():
        old_value = old.get(key, "")
        new_value = new.get(key, "")

        # 新资料没有这个字段
        if key not in new:
            continue

        # 金额格式统一
        if key == "amount":
            try:
                old_value = float(old_value)
                new_value = float(new_value)
            except:
                pass

        if str(old_value) != str(new_value):

            changes.append({
                "label": label,
                "old": old_value or "-",
                "new": new_value or "-"
            })

    return changes


PAYMENT_HISTORY_DETAIL_HTML = """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<title>修改历史详情</title>
<meta name="viewport" content="width=device-width, initial-scale=1">

<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">
</head>

<body>

<div class="page">

    <h1 class="page-title">📝 修改历史详情</h1>
    <p class="page-subtitle">查看这笔缴费记录曾经修改过的内容</p>

    <div class="card">
        <a class="btn-tool btn-secondary"
           href="/member/finance-upload">
            ← 返回月费资料管理中心
        </a>
    </div>

    <div class="info-card">

        <div class="section-title">
            📄 基本资料
        </div>

        <div class="person-meta">
            <b>修改时间：</b>
            {{ (h.changed_at + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M") if h.changed_at else "-" }}
            <br>

            <b>会员编号：</b>{{ h.member_id }}
            <br>

            <b>收据编号：</b>{{ h.receipt_no }}
            <br>

            <b>修改者：</b>{{ h.changed_by }}
        </div>

    </div>

    <div class="card">

        <div class="section-title">
            🔍 修改内容
        </div>

        {% if changes %}

        <div class="table-responsive">

            <table class="record-table">

                <tr>
                    <th>项目</th>
                    <th>修改前</th>
                    <th>修改后</th>
                </tr>

                {% for c in changes %}
                <tr>

                    <td>
                        {{ c.label }}
                    </td>

                    <td style="color:#dc2626;font-weight:bold;">
                        {{ c.old }}
                    </td>

                    <td style="color:#16a34a;font-weight:bold;">
                        {{ c.new }}
                    </td>

                </tr>
                {% endfor %}

            </table>

        </div>

        {% else %}

        <div class="empty-state">

            <div class="empty-icon">
                📭
            </div>

            <div class="empty-title">
                没有发现明显变化
            </div>

            <div class="empty-text">
                这次修改没有记录到任何字段差异。
            </div>

        </div>

        {% endif %}

    </div>

</div>

</body>
</html>
"""