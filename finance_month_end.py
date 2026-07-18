# finance_month_end.py

import re

from io import BytesIO
from db import db_query
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from finance_web_old import (
    FINANCE_V5_STYLE,
    FINANCE_DATE_COMPONENT,
)
from finance_common import money, require_finance_month_open, get_current_finance_user
from finance_audit import write_finance_audit
from flask import (
    Blueprint, request, render_template_string, redirect,
    url_for, session, send_file, flash
)

finance_month_end_bp = Blueprint(
    "finance_month_end",
    __name__,
    url_prefix="/finance/month_end"
)


def get_month_cash_income(ym, fund_account):
    row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'income'
          and payment_method = '现金'
          and fund_account = %s
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
    """, (fund_account, ym), fetchone=True)

    return float((row or {}).get("total") or 0)


def get_month_cash_expense(ym, fund_account):
    row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'expense'
          and payment_method = '现金'
          and fund_account = %s
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
    """, (fund_account, ym), fetchone=True)

    return float((row or {}).get("total") or 0)


def get_month_bank_in_total(ym, fund_account):
    row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_bank_deposits
        where ym = %s
          and fund_account = %s
    """, (ym, fund_account), fetchone=True)

    return float((row or {}).get("total") or 0)


def get_cash_receipt_range(ym, fund_account):
    row = db_query("""
        select
            min(receipt_no) as receipt_from,
            max(receipt_no) as receipt_to
        from finance_records
        where record_type = 'income'
          and payment_method = '现金'
          and fund_account = %s
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and coalesce(receipt_no, '') <> ''
    """, (fund_account, ym), fetchone=True)

    return {
        "receipt_from": (row or {}).get("receipt_from") or "",
        "receipt_to": (row or {}).get("receipt_to") or "",
    }


@finance_month_end_bp.route("/")
def month_end_home():
    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    ym = request.args.get(
        "ym",
        date.today().strftime("%Y-%m")
    )

    return render_template_string(
        FINANCE_V5_STYLE + FINANCE_DATE_COMPONENT + """
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta
                name="viewport"
                content="width=device-width, initial-scale=1"
            >
            <title>财政月结</title>
            <link
                rel="stylesheet"
                href="{{ url_for(
                    'static',
                    filename='css/toolbox.css'
                ) }}"
            >
        </head>

        <body>
        <div class="finance-v5">

            <div class="v5-topbar">
                <a
                    class="v5-back"
                    href="{{ url_for('finance.finance_home') }}"
                >
                    ← 返回财政首页
                </a>
            </div>

            <div class="v5-header">
                <h1>📒 财政月结</h1>
                <p>
                    Cash Bank In、现金对账、
                    银行对账与总会报表
                </p>
            </div>

            <div class="card" style="margin-bottom:20px;">
                <form method="get">
                    <div class="form-group">
                        <label class="form-label">
                            月结月份
                        </label>

                        <input
                            class="form-input"
                            name="ym"
                            type="month"
                            value="{{ ym }}"
                            required
                        >
                    </div>

                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                    >
                        查看月份
                    </button>
                </form>
            </div>

            <div class="v5-menu-grid">

                <a
                    class="v5-menu-btn v5-income"
                    href="{{ url_for(
                        'finance_month_end.cash_bank_in',
                        ym=ym
                    ) }}"
                >
                    <div class="v5-icon">💰</div>
                    <div class="v5-menu-text">
                        <div class="v5-menu-title">
                            Cash Bank In
                        </div>
                        <div class="v5-menu-desc">
                            登记现金收款存入银行
                        </div>
                    </div>
                </a>

                <a
                    class="v5-menu-btn"
                    href="{{ url_for(
                        'finance_month_end.cash_reconciliation',
                        ym=ym
                    ) }}"
                >
                    <div class="v5-icon">💵</div>

                    <div class="v5-menu-text">
                        <div class="v5-menu-title">
                            Cash Reconciliation
                        </div>

                        <div class="v5-menu-desc">
                            现金在手核对与差异检查
                        </div>
                    </div>
                </a>

                <a
                    class="v5-menu-btn"
                    href="{{ url_for(
                        'finance_month_end.bank_reconciliation',
                        ym=ym
                    ) }}"
                >
                    <div class="v5-icon">🏦</div>

                    <div class="v5-menu-text">
                        <div class="v5-menu-title">
                            Bank Reconciliation
                        </div>

                        <div class="v5-menu-desc">
                            银行月结余额与 Statement 核对
                        </div>
                    </div>
                </a>

                <a
                    class="v5-menu-btn"
                    href="{{ url_for(
                        'finance_month_end.month_close',
                        ym=ym
                    ) }}"
                >
                    <div class="v5-icon">🔒</div>

                    <div class="v5-menu-text">
                        <div class="v5-menu-title">
                            Month Close
                        </div>

                        <div class="v5-menu-desc">
                            确认现金与银行对账后完成月结
                        </div>
                    </div>
                </a>

                <div class="v5-menu-btn">
                    <div class="v5-icon">📄</div>
                    <div class="v5-menu-text">
                        <div class="v5-menu-title">
                            HQ Report Package
                        </div>
                        <div class="v5-menu-desc">
                            下一阶段：总会七项报表
                        </div>
                    </div>
                </div>

            </div>
        </div>
        </body>
        </html>
        """,
        ym=ym
    )


@finance_month_end_bp.route(
    "/bank_in",
    methods=["GET", "POST"]
)
def cash_bank_in():
    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    fund_accounts = [
        "观音堂日常户口",
        "总会户口",
    ]

    ym = (
        request.form.get("ym")
        or request.args.get("ym")
        or date.today().strftime("%Y-%m")
    ).strip()

    fund_account = (
        request.form.get("fund_account")
        or request.args.get("fund_account")
        or "观音堂日常户口"
    ).strip()

    if fund_account not in fund_accounts:
        fund_account = "观音堂日常户口"

    cash_income = get_month_cash_income(
        ym,
        fund_account
    )

    cash_expense = get_month_cash_expense(
        ym,
        fund_account
    )

    banked_in = get_month_bank_in_total(
        ym,
        fund_account
    )

    unbanked = max(cash_income - banked_in, 0)
    over_banked = max(banked_in - cash_income, 0)

    receipt_range = get_cash_receipt_range(
        ym,
        fund_account
    )

    message = ""
    message_type = "danger"

    form_data = {
        "deposit_date": date.today().isoformat(),
        "bank_name": "",
        "reference_no": "",
        "receipt_from": receipt_range["receipt_from"],
        "receipt_to": receipt_range["receipt_to"],
        "amount": f"{unbanked:.2f}" if unbanked > 0 else "",
        "remarks": "",
    }

    if request.method == "POST":
        deposit_date = (
            request.form.get("deposit_date")
            or date.today().isoformat()
        )

        bank_name = request.form.get(
            "bank_name",
            ""
        ).strip()

        reference_no = request.form.get(
            "reference_no",
            ""
        ).strip()

        receipt_from = request.form.get(
            "receipt_from",
            ""
        ).strip().upper()

        receipt_to = request.form.get(
            "receipt_to",
            ""
        ).strip().upper()

        amount_raw = request.form.get(
            "amount",
            ""
        ).strip()

        amount = money(amount_raw)

        remarks = request.form.get(
            "remarks",
            ""
        ).strip()

        form_data = {
            "deposit_date": str(deposit_date),
            "bank_name": bank_name,
            "reference_no": reference_no,
            "receipt_from": receipt_from,
            "receipt_to": receipt_to,
            "amount": amount_raw,
            "remarks": remarks,
        }

        month_lock_error = require_finance_month_open(
            ym,
            fund_account
        )

        existing_reference = None

        if reference_no:
            existing_reference = db_query("""
                select id
                from finance_bank_deposits
                where lower(reference_no) = lower(%s)
                limit 1
            """, (reference_no,), fetchone=True)

        if month_lock_error:
            message = month_lock_error

        elif not re.match(r"^\d{4}-\d{2}$", ym):
            message = "月份格式错误。"

        elif amount <= 0:
            message = "Bank In 金额必须大过 RM0。"

        elif not bank_name:
            message = "请选择存入的银行。"

        elif existing_reference:
            message = (
                "这个 Bank In Reference 已经存在，"
                "请检查是否重复登记。"
            )

        elif unbanked <= 0:
            message = (
                "这个月份目前没有尚未存入银行的现金。"
            )

        elif amount > unbanked + 0.005:
            message = (
                "Bank In 金额大过系统计算的尚未存入金额。"
                f"目前尚未存入 RM{unbanked:.2f}。"
            )

        else:
            db_query("""
                insert into finance_bank_deposits
                (
                    deposit_date,
                    ym,
                    fund_account,
                    bank_name,
                    reference_no,
                    receipt_from,
                    receipt_to,
                    amount,
                    remarks
                )
                values
                (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, (
                deposit_date,
                ym,
                fund_account,
                bank_name,
                reference_no,
                receipt_from,
                receipt_to,
                amount,
                remarks
            ))

            return redirect(
                url_for(
                    "finance_month_end.cash_bank_in",
                    ym=ym,
                    fund_account=fund_account,
                    saved="1"
                )
            )

    if request.args.get("saved") == "1":
        message = "Bank In 记录已保存。"
        message_type = "success"

    if request.args.get("deleted") == "1":
        message = "Bank In 记录已删除。"
        message_type = "success"

    deposits = db_query("""
        select
            id,
            deposit_date,
            bank_name,
            reference_no,
            receipt_from,
            receipt_to,
            amount,
            remarks
        from finance_bank_deposits
        where ym = %s
          and fund_account = %s
        order by deposit_date desc, id desc
    """, (ym, fund_account), fetchall=True)

    cash_income = get_month_cash_income(
        ym,
        fund_account
    )

    cash_expense = get_month_cash_expense(
        ym,
        fund_account
    )

    banked_in = get_month_bank_in_total(
        ym,
        fund_account
    )

    unbanked = max(cash_income - banked_in, 0)
    over_banked = max(banked_in - cash_income, 0)

    return render_template_string(
        FINANCE_DATE_COMPONENT + """
        <!doctype html>
        <html lang="zh">
        <head>
            <meta charset="utf-8">
            <meta
                name="viewport"
                content="width=device-width, initial-scale=1"
            >

            <title>Cash Bank In</title>

            <link
                rel="stylesheet"
                href="{{ url_for(
                    'static',
                    filename='css/toolbox.css'
                ) }}"
            >

            <style>
                .bank-in-page{
                    max-width:1180px;
                }

                .bank-in-header{
                    background:linear-gradient(
                        135deg,
                        #15803d,
                        #166534
                    );
                    color:#fff;
                    padding:28px;
                    border-radius:22px;
                    margin-bottom:20px;
                }

                .bank-in-header h1{
                    margin:0 0 8px;
                }

                .bank-in-header p{
                    margin:0;
                    opacity:.92;
                }

                .filter-grid{
                    display:grid;
                    grid-template-columns:
                        minmax(180px,1fr)
                        minmax(240px,1fr)
                        auto;
                    gap:14px;
                    align-items:end;
                }

                .summary-grid-bank{
                    display:grid;
                    grid-template-columns:
                        repeat(4,minmax(0,1fr));
                    gap:14px;
                    margin:20px 0;
                }

                .summary-grid-bank .summary-box{
                    min-height:120px;
                    display:flex;
                    flex-direction:column;
                    justify-content:center;
                    text-align:center;
                }

                .summary-label-bank{
                    color:#64748b;
                    margin-bottom:7px;
                }

                .summary-value-bank{
                    font-size:25px;
                    font-weight:900;
                }

                .money-green{color:#15803d;}
                .money-blue{color:#1d4ed8;}
                .money-orange{color:#c2410c;}
                .money-red{color:#b91c1c;}

                .entry-grid{
                    display:grid;
                    grid-template-columns:
                        repeat(2,minmax(0,1fr));
                    gap:16px;
                }

                .entry-full{
                    grid-column:1 / -1;
                }

                .form-help{
                    color:#64748b;
                    font-size:14px;
                    margin-top:7px;
                }

                .bank-in-table{
                    min-width:1050px;
                }

                .money-cell{
                    text-align:right;
                    white-space:nowrap;
                    font-weight:900;
                    color:#15803d;
                }

                .reference-cell,
                .receipt-cell{
                    white-space:nowrap;
                }

                .reference-cell{
                    font-family:Consolas,monospace;
                }

                .bank-in-actions{
                    display:flex;
                    gap:10px;
                    flex-wrap:wrap;
                    margin-top:20px;
                }

                .bank-query-form{
                    display:grid;
                    grid-template-columns:
                        minmax(260px, 1fr)
                        minmax(320px, 1fr)
                        auto;
                    gap:16px;
                    align-items:end;
                }

                .query-button-wrap{
                    display:flex;
                    align-items:flex-end;
                }

                .query-button-wrap .btn-tool{
                    min-height:76px;
                    padding:0 28px;
                    white-space:nowrap;
                }

                @media(max-width:800px){
                    .bank-query-form{
                        grid-template-columns:1fr;
                    }

                    .query-button-wrap .btn-tool{
                        width:100%;
                    }
                }

                @media(max-width:850px){
                    .summary-grid-bank{
                        grid-template-columns:
                            repeat(2,minmax(0,1fr));
                    }

                    .filter-grid,
                    .entry-grid{
                        grid-template-columns:1fr;
                    }

                    .entry-full{
                        grid-column:auto;
                    }
                }

                @media(max-width:600px){
                    .summary-grid-bank{
                        grid-template-columns:1fr;
                    }

                    .filter-grid .btn-tool,
                    .bank-in-actions .btn-tool{
                        width:100%;
                    }
                }
            </style>
        </head>

        <body>
        <div class="page bank-in-page">

            <div class="bank-in-header">
                <h1>💰 Cash Collections Banked In</h1>
                <p>
                    系统自动统计现金收入，
                    财政只登记实际存入银行的资料。
                </p>
            </div>

            {% if message %}
                <div class="
                    alert
                    {% if message_type == 'success' %}
                        alert-success
                    {% else %}
                        alert-danger
                    {% endif %}
                ">
                    {{ message }}
                </div>
            {% endif %}

            <div class="card">
                <div class="section-title">
                    🔎 查询月份与户口
                </div>

                <form method="get"
                    class="bank-query-form">

                    <div class="form-group">
                        <label class="form-label">
                            月份
                        </label>

                        <input class="form-input"
                            type="month"
                            name="ym"
                            value="{{ ym }}">
                    </div>

                    <div class="form-group">
                        <label class="form-label">
                            基金户口
                        </label>

                        <select class="form-input"
                                name="fund_account">

                            <option value="观音堂日常户口"
                                {% if fund_account == '观音堂日常户口' %}
                                    selected
                                {% endif %}>
                                观音堂日常户口
                            </option>

                            <option value="总会户口"
                                {% if fund_account == '总会户口' %}
                                    selected
                                {% endif %}>
                                总会户口
                            </option>

                        </select>
                    </div>

                    <div class="query-button-wrap">
                        <button class="btn-tool btn-primary"
                                type="submit">
                            查看
                        </button>
                    </div>

                </form>
            </div>

            <div class="summary-grid-bank">

                <div class="summary-box">
                    <div class="summary-label-bank">
                        本月现金收入
                    </div>
                    <div class="
                        summary-value-bank
                        money-green
                    ">
                        RM {{ "%.2f"|format(cash_income) }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-label-bank">
                        本月现金支出
                    </div>
                    <div class="
                        summary-value-bank
                        money-red
                    ">
                        RM {{ "%.2f"|format(cash_expense) }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-label-bank">
                        已存入银行
                    </div>
                    <div class="
                        summary-value-bank
                        money-blue
                    ">
                        RM {{ "%.2f"|format(banked_in) }}
                    </div>
                </div>

                <div class="summary-box">
                    <div class="summary-label-bank">
                        尚未存入银行
                    </div>
                    <div class="
                        summary-value-bank
                        {% if unbanked > 0 %}
                            money-orange
                        {% else %}
                            money-green
                        {% endif %}
                    ">
                        RM {{ "%.2f"|format(unbanked) }}
                    </div>

                    {% if over_banked > 0 %}
                        <div style="color:#b91c1c;font-weight:800;">
                            多存 RM
                            {{ "%.2f"|format(over_banked) }}
                        </div>
                    {% endif %}
                </div>

            </div>

            <div class="card">
                <div class="section-title">
                    ➕ 登记 Cash Bank In
                </div>

                {% if unbanked > 0 %}
                    <form method="post">

                        <input
                            type="hidden"
                            name="ym"
                            value="{{ ym }}"
                        >

                        <input
                            type="hidden"
                            name="fund_account"
                            value="{{ fund_account }}"
                        >

                        <div class="entry-grid">

                            <div class="form-group">
                                <label class="form-label">
                                    Bank In 日期
                                </label>

                                <input
                                    class="form-input"
                                    name="deposit_date"
                                    type="date"
                                    value="{{ form_data.deposit_date }}"
                                    required
                                >
                            </div>

                            <div class="form-group">
                                <label class="form-label">
                                    存入银行
                                </label>

                                <select
                                    class="form-input"
                                    name="bank_name"
                                    required
                                >
                                    <option value="">
                                        请选择银行
                                    </option>

                                    {% for bank in bank_names %}
                                        <option
                                            value="{{ bank }}"
                                            {% if
                                                form_data.bank_name
                                                == bank
                                            %}
                                                selected
                                            {% endif %}
                                        >
                                            {{ bank }}
                                        </option>
                                    {% endfor %}
                                </select>
                            </div>

                            <div class="form-group">
                                <label class="form-label">
                                    Bank In Reference
                                </label>

                                <input
                                    class="form-input"
                                    name="reference_no"
                                    value="{{ form_data.reference_no }}"
                                    autocomplete="off"
                                >
                            </div>

                            <div class="form-group">
                                <label class="form-label">
                                    Bank In 金额 RM
                                </label>

                                <input
                                    class="form-input"
                                    name="amount"
                                    type="number"
                                    step="0.01"
                                    min="0.01"
                                    max="{{ unbanked }}"
                                    value="{{ form_data.amount }}"
                                    required
                                >

                                <div class="form-help">
                                    尚未存入：
                                    RM {{ "%.2f"|format(unbanked) }}
                                </div>
                            </div>

                            <div class="form-group">
                                <label class="form-label">
                                    Receipt From
                                </label>

                                <input
                                    class="form-input"
                                    name="receipt_from"
                                    value="{{ form_data.receipt_from }}"
                                >
                            </div>

                            <div class="form-group">
                                <label class="form-label">
                                    Receipt To
                                </label>

                                <input
                                    class="form-input"
                                    name="receipt_to"
                                    value="{{ form_data.receipt_to }}"
                                >
                            </div>

                            <div class="
                                form-group
                                entry-full
                            ">
                                <label class="form-label">
                                    备注
                                </label>

                                <textarea
                                    class="form-input"
                                    name="remarks"
                                    rows="4"
                                >{{ form_data.remarks }}</textarea>
                            </div>

                        </div>

                        <button
                            class="btn-tool btn-success"
                            type="submit"
                            style="width:100%;margin-top:18px;"
                            onclick="
                                return confirm(
                                    '确定保存这笔 Cash Bank In？'
                                );
                            "
                        >
                            ✅ 确认 Bank In
                        </button>
                    </form>

                {% else %}
                    <div class="empty-state">
                        ✅ 此月份的现金收入已经全部登记存入银行。
                    </div>
                {% endif %}
            </div>

            <div class="card">
                <div class="section-title">
                    📋 Bank In 历史
                </div>

                {% if deposits %}
                    <div class="table-responsive">
                        <table class="
                            record-table
                            bank-in-table
                        ">
                            <thead>
                                <tr>
                                    <th>日期</th>
                                    <th>银行</th>
                                    <th>Reference</th>
                                    <th>Receipt From</th>
                                    <th>Receipt To</th>
                                    <th>金额</th>
                                    <th>备注</th>
                                    <th>操作</th>
                                </tr>
                            </thead>

                            <tbody>
                                {% for row in deposits %}
                                    <tr>
                                        <td>
                                            {{ row.deposit_date }}
                                        </td>

                                        <td>
                                            {{ row.bank_name or "-" }}
                                        </td>

                                        <td class="reference-cell">
                                            {{ row.reference_no or "-" }}
                                        </td>

                                        <td class="receipt-cell">
                                            {{ row.receipt_from or "-" }}
                                        </td>

                                        <td class="receipt-cell">
                                            {{ row.receipt_to or "-" }}
                                        </td>

                                        <td class="money-cell">
                                            RM
                                            {{ "%.2f"|format(
                                                row.amount or 0
                                            ) }}
                                        </td>

                                        <td>
                                            {{ row.remarks or "-" }}
                                        </td>

                                        <td>
                                            <form
                                                method="post"
                                                action="{{ url_for(
                                                    'finance_month_end.delete_cash_bank_in',
                                                    deposit_id=row.id
                                                ) }}"
                                                onsubmit="
                                                    return confirm(
                                                        '确定删除这笔 Bank In 记录？'
                                                    );
                                                "
                                            >
                                                <button
                                                    class="
                                                        btn-tool
                                                        btn-danger
                                                    "
                                                    type="submit"
                                                >
                                                    删除
                                                </button>
                                            </form>
                                        </td>
                                    </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>

                {% else %}
                    <div class="empty-state">
                        这个月份还没有 Bank In 记录。
                    </div>
                {% endif %}
            </div>

            <div class="bank-in-actions">
                <a
                    class="btn-tool btn-secondary"
                    href="{{ url_for(
                        'finance_month_end.month_end_home',
                        ym=ym
                    ) }}"
                >
                    ← 返回财政月结
                </a>
            </div>

        </div>
        </body>
        </html>
        """,
        ym=ym,
        fund_account=fund_account,
        fund_accounts=fund_accounts,
        cash_income=cash_income,
        cash_expense=cash_expense,
        banked_in=banked_in,
        unbanked=unbanked,
        over_banked=over_banked,
        deposits=deposits,
        form_data=form_data,
        message=message,
        message_type=message_type,
        bank_names=[
            "Maybank",
            "Public Bank",
            "CIMB",
            "Hong Leong",
            "RHB",
            "AmBank",
            "Bank Islam",
            "BSN",
        ]
    )

def get_previous_month_ym(ym):
    year, month = map(int, ym.split("-"))
    month -= 1
    if month == 0:
        year -= 1
        month = 12
    return f"{year:04d}-{month:02d}"


def get_previous_closing_cash(ym, fund_account):
    previous_ym = get_previous_month_ym(ym)

    row = db_query("""
        select actual_cash
        from finance_cash_reconciliation
        where ym = %s
          and fund_account = %s
          and status = 'confirmed'
        limit 1
    """, (previous_ym, fund_account), fetchone=True)

    return float((row or {}).get("actual_cash") or 0)


def get_cash_reconciliation_summary(
    ym,
    fund_account,
    opening_cash=None
):
    if opening_cash is None:
        opening_cash = get_previous_closing_cash(
            ym,
            fund_account
        )

    cash_income = get_month_cash_income(
        ym,
        fund_account
    )

    cash_expense = get_month_cash_expense(
        ym,
        fund_account
    )

    cash_banked_in = get_month_bank_in_total(
        ym,
        fund_account
    )

    expected_cash = (
        float(opening_cash or 0)
        + float(cash_income or 0)
        - float(cash_expense or 0)
        - float(cash_banked_in or 0)
    )

    return {
        "opening_cash": float(opening_cash or 0),
        "cash_income": float(cash_income or 0),
        "cash_expense": float(cash_expense or 0),
        "cash_banked_in": float(cash_banked_in or 0),
        "expected_cash": float(expected_cash or 0),
    }


@finance_month_end_bp.route(
    "/cash_reconciliation",
    methods=["GET", "POST"]
)
def cash_reconciliation():

    if not session.get("finance_login"):
        return redirect(
            url_for("finance.finance_login")
        )

    fund_accounts = [
        "观音堂日常户口",
        "总会户口",
    ]

    ym = (
        request.form.get("ym")
        or request.args.get("ym")
        or date.today().strftime("%Y-%m")
    ).strip()

    fund_account = (
        request.form.get("fund_account")
        or request.args.get("fund_account")
        or "观音堂日常户口"
    ).strip()

    if fund_account not in fund_accounts:
        fund_account = "观音堂日常户口"

    existing = db_query("""
        select *
        from finance_cash_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    opening_cash = (
        float(existing["opening_cash"] or 0)
        if existing
        else get_previous_closing_cash(
            ym,
            fund_account
        )
    )

    summary = get_cash_reconciliation_summary(
        ym,
        fund_account,
        opening_cash
    )

    message = ""
    message_type = "danger"

    actual_cash = (
        float(existing["actual_cash"] or 0)
        if existing
        else None
    )

    difference = (
        actual_cash - summary["expected_cash"]
        if actual_cash is not None
        else None
    )

    form_data = {
        "opening_cash": f"{opening_cash:.2f}",
        "actual_cash": (
            f"{actual_cash:.2f}"
            if actual_cash is not None
            else ""
        ),
        "checked_by": (
            existing.get("checked_by") or ""
            if existing
            else ""
        ),
        "checked_date": (
            str(existing.get("checked_date"))
            if existing and existing.get("checked_date")
            else date.today().isoformat()
        ),
        "remarks": (
            existing.get("remarks") or ""
            if existing
            else ""
        ),
    }

    if request.method == "POST":

        month_lock_error = require_finance_month_open(
            ym,
            fund_account
        )

        opening_cash = money(
            request.form.get("opening_cash")
        )

        actual_cash = money(
            request.form.get("actual_cash")
        )

        checked_by = request.form.get(
            "checked_by",
            ""
        ).strip()

        checked_date = (
            request.form.get("checked_date")
            or date.today().isoformat()
        )

        remarks = request.form.get(
            "remarks",
            ""
        ).strip()

        action = request.form.get(
            "action",
            "save"
        )

        summary = get_cash_reconciliation_summary(
            ym,
            fund_account,
            opening_cash
        )

        difference = (
            actual_cash
            - summary["expected_cash"]
        )

        if month_lock_error:
            message = month_lock_error

        elif opening_cash < 0:
            message = "上月结转现金不能少于 RM0。"

        elif actual_cash < 0:
            message = "实际点算现金不能少于 RM0。"

        elif action == "confirm" and not checked_by:
            message = "确认对账时必须填写核对人。"

        else:
            status = (
                "confirmed"
                if action == "confirm"
                else "draft"
            )

            db_query("""
                insert into finance_cash_reconciliation
                (
                    ym,
                    fund_account,
                    opening_cash,
                    cash_income,
                    cash_expense,
                    cash_banked_in,
                    expected_cash,
                    actual_cash,
                    difference,
                    checked_by,
                    checked_date,
                    status,
                    remarks,
                    updated_at
                )
                values
                (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, now()
                )
                on conflict (ym, fund_account)
                do update set
                    opening_cash = excluded.opening_cash,
                    cash_income = excluded.cash_income,
                    cash_expense = excluded.cash_expense,
                    cash_banked_in = excluded.cash_banked_in,
                    expected_cash = excluded.expected_cash,
                    actual_cash = excluded.actual_cash,
                    difference = excluded.difference,
                    checked_by = excluded.checked_by,
                    checked_date = excluded.checked_date,
                    status = excluded.status,
                    remarks = excluded.remarks,
                    updated_at = now()
            """, (
                ym,
                fund_account,
                summary["opening_cash"],
                summary["cash_income"],
                summary["cash_expense"],
                summary["cash_banked_in"],
                summary["expected_cash"],
                actual_cash,
                difference,
                checked_by,
                checked_date,
                status,
                remarks
            ))

            return redirect(
                url_for(
                    "finance_month_end.cash_reconciliation",
                    ym=ym,
                    fund_account=fund_account,
                    saved="1",
                    status=status
                )
            )

    if request.args.get("saved") == "1":
        message = (
            "Cash Reconciliation 已确认完成。"
            if request.args.get("status") == "confirmed"
            else "Cash Reconciliation 草稿已保存。"
        )
        message_type = "success"

    history = db_query("""
        select *
        from finance_cash_reconciliation
        where fund_account = %s
        order by ym desc
        limit 24
    """, (fund_account,), fetchall=True)

    return render_template_string(
        FINANCE_DATE_COMPONENT + """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>
<title>Cash Reconciliation</title>

<link
    rel="stylesheet"
    href="{{ url_for(
        'static',
        filename='css/toolbox.css'
    ) }}"
>

<style>
.cash-recon-page{
    max-width:1100px;
}

.cash-recon-header{
    background:linear-gradient(
        135deg,
        #7c3aed,
        #5b21b6
    );
    color:white;
    padding:28px;
    border-radius:22px;
    margin-bottom:20px;
}

.cash-recon-header h1{
    margin:0 0 8px;
}

.cash-recon-header p{
    margin:0;
    opacity:.92;
}

.filter-grid,
.form-grid{
    display:grid;
    grid-template-columns:
        repeat(2,minmax(0,1fr));
    gap:16px;
}

.filter-actions{
    grid-column:1 / -1;
}

.recon-list{
    display:grid;
    gap:10px;
}

.recon-row{
    display:grid;
    grid-template-columns:
        60px 1fr 220px;
    gap:12px;
    align-items:center;
    padding:15px;
    background:#f8fafc;
    border:1px solid #e2e8f0;
    border-radius:14px;
}

.recon-symbol{
    font-size:27px;
    text-align:center;
    font-weight:900;
}

.recon-label{
    font-size:18px;
    font-weight:800;
}

.recon-value{
    text-align:right;
    font-size:22px;
    font-weight:900;
}

.expected-row{
    background:#eff6ff;
    border-color:#bfdbfe;
}

.expected-row .recon-value{
    color:#1d4ed8;
    font-size:27px;
}

.full-width{
    grid-column:1 / -1;
}

.difference-box{
    margin-top:18px;
    padding:20px;
    border-radius:16px;
    text-align:center;
}

.difference-ok{
    background:#f0fdf4;
    border:1px solid #86efac;
    color:#166534;
}

.difference-bad{
    background:#fef2f2;
    border:1px solid #fca5a5;
    color:#991b1b;
}

.difference-value{
    font-size:32px;
    font-weight:900;
}

.action-grid{
    display:grid;
    grid-template-columns:1fr 1fr;
    gap:12px;
    margin-top:18px;
}

.history-table{
    min-width:1150px;
}

.money-cell{
    text-align:right;
    white-space:nowrap;
}

@media(max-width:700px){
    .filter-grid,
    .form-grid,
    .action-grid{
        grid-template-columns:1fr;
    }

    .full-width,
    .filter-actions{
        grid-column:auto;
    }

    .recon-row{
        grid-template-columns:45px 1fr;
    }

    .recon-value{
        grid-column:1 / -1;
        text-align:left;
        padding-left:57px;
    }
}
</style>
</head>

<body>
<div class="page cash-recon-page">

    <div class="cash-recon-header">
        <h1>💵 Cash in Hand Reconciliation</h1>
        <p>
            系统自动计算应有现金，
            财政只需填写实际点算金额。
        </p>
    </div>

    {% if message %}
        <div class="
            alert
            {% if message_type == 'success' %}
                alert-success
            {% else %}
                alert-danger
            {% endif %}
        ">
            {{ message }}
        </div>
    {% endif %}

    <div class="card">
        <div class="section-title">
            🔎 查询月份与户口
        </div>

        <form method="get">
            <div class="filter-grid">

                <div class="form-group">
                    <label class="form-label">
                        月份
                    </label>

                    <input
                        class="form-input"
                        name="ym"
                        type="month"
                        value="{{ ym }}"
                        required
                    >
                </div>

                <div class="form-group">
                    <label class="form-label">
                        基金户口
                    </label>

                    <select
                        class="form-input"
                        name="fund_account"
                    >
                        {% for account in fund_accounts %}
                            <option
                                value="{{ account }}"
                                {% if
                                    fund_account == account
                                %}
                                    selected
                                {% endif %}
                            >
                                {{ account }}
                            </option>
                        {% endfor %}
                    </select>
                </div>

                <div class="filter-actions">
                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                    >
                        查看
                    </button>
                </div>

            </div>
        </form>
    </div>

    <div class="card">
        <div class="section-title">
            🧮 系统现金计算
        </div>

        <div class="recon-list">

            <div class="recon-row">
                <div class="recon-symbol">🧾</div>
                <div class="recon-label">
                    上月结转现金
                </div>
                <div class="recon-value">
                    RM {{ "%.2f"|format(
                        summary.opening_cash
                    ) }}
                </div>
            </div>

            <div class="recon-row">
                <div class="recon-symbol">＋</div>
                <div class="recon-label">
                    本月现金收入
                </div>
                <div class="recon-value">
                    RM {{ "%.2f"|format(
                        summary.cash_income
                    ) }}
                </div>
            </div>

            <div class="recon-row">
                <div class="recon-symbol">−</div>
                <div class="recon-label">
                    本月现金支出
                </div>
                <div class="recon-value">
                    RM {{ "%.2f"|format(
                        summary.cash_expense
                    ) }}
                </div>
            </div>

            <div class="recon-row">
                <div class="recon-symbol">−</div>
                <div class="recon-label">
                    已存入银行
                </div>
                <div class="recon-value">
                    RM {{ "%.2f"|format(
                        summary.cash_banked_in
                    ) }}
                </div>
            </div>

            <div class="recon-row expected-row">
                <div class="recon-symbol">＝</div>
                <div class="recon-label">
                    系统应有现金
                </div>
                <div class="recon-value">
                    RM {{ "%.2f"|format(
                        summary.expected_cash
                    ) }}
                </div>
            </div>

        </div>
    </div>

    <div class="card">
        <div class="section-title">
            ✍️ 实际点算与确认
        </div>

        <form method="post">

            <input
                type="hidden"
                name="ym"
                value="{{ ym }}"
            >

            <input
                type="hidden"
                name="fund_account"
                value="{{ fund_account }}"
            >

            <div class="form-grid">

                <div class="form-group">
                    <label class="form-label">
                        上月结转现金 RM
                    </label>

                    <input
                        class="form-input"
                        name="opening_cash"
                        type="number"
                        step="0.01"
                        min="0"
                        value="{{ form_data.opening_cash }}"
                        required
                    >
                </div>

                <div class="form-group">
                    <label class="form-label">
                        实际点算现金 RM
                    </label>

                    <input
                        class="form-input"
                        name="actual_cash"
                        type="number"
                        step="0.01"
                        min="0"
                        value="{{ form_data.actual_cash }}"
                        required
                    >
                </div>

                <div class="form-group">
                    <label class="form-label">
                        核对人
                    </label>

                    <input
                        class="form-input"
                        name="checked_by"
                        value="{{ form_data.checked_by }}"
                    >
                </div>

                <div class="form-group">
                    <label class="form-label">
                        核对日期
                    </label>

                    <input
                        class="form-input"
                        name="checked_date"
                        type="date"
                        value="{{ form_data.checked_date }}"
                        required
                    >
                </div>

                <div class="form-group full-width">
                    <label class="form-label">
                        备注
                    </label>

                    <textarea
                        class="form-input"
                        name="remarks"
                        rows="4"
                    >{{ form_data.remarks }}</textarea>
                </div>

            </div>

            {% if difference is not none %}
                <div class="
                    difference-box
                    {% if difference|abs < 0.005 %}
                        difference-ok
                    {% else %}
                        difference-bad
                    {% endif %}
                ">
                    <div>
                        {% if difference|abs < 0.005 %}
                            ✅ 对账一致
                        {% else %}
                            ⚠️ 现金差异
                        {% endif %}
                    </div>

                    <div class="difference-value">
                        RM {{ "%.2f"|format(difference) }}
                    </div>
                </div>
            {% endif %}

            <div class="action-grid">
                <button
                    class="btn-tool btn-secondary"
                    type="submit"
                    name="action"
                    value="save"
                >
                    💾 保存草稿
                </button>

                <button
                    class="btn-tool btn-success"
                    type="submit"
                    name="action"
                    value="confirm"
                    onclick="
                        return confirm(
                            '确定完成本月现金对账？'
                        );
                    "
                >
                    ✅ 确认完成
                </button>
            </div>

        </form>
    </div>

    <div class="card">
        <div class="section-title">
            📋 对账历史
        </div>

        {% if history %}
            <div class="table-responsive">
                <table class="record-table history-table">
                    <thead>
                        <tr>
                            <th>月份</th>
                            <th>户口</th>
                            <th>期初现金</th>
                            <th>现金收入</th>
                            <th>现金支出</th>
                            <th>Bank In</th>
                            <th>应有现金</th>
                            <th>实际现金</th>
                            <th>差异</th>
                            <th>状态</th>
                            <th>核对人</th>
                        </tr>
                    </thead>

                    <tbody>
                        {% for row in history %}
                            <tr>
                                <td>{{ row.ym }}</td>
                                <td>{{ row.fund_account }}</td>

                                <td class="money-cell">
                                    RM {{ "%.2f"|format(
                                        row.opening_cash or 0
                                    ) }}
                                </td>

                                <td class="money-cell">
                                    RM {{ "%.2f"|format(
                                        row.cash_income or 0
                                    ) }}
                                </td>

                                <td class="money-cell">
                                    RM {{ "%.2f"|format(
                                        row.cash_expense or 0
                                    ) }}
                                </td>

                                <td class="money-cell">
                                    RM {{ "%.2f"|format(
                                        row.cash_banked_in or 0
                                    ) }}
                                </td>

                                <td class="money-cell">
                                    RM {{ "%.2f"|format(
                                        row.expected_cash or 0
                                    ) }}
                                </td>

                                <td class="money-cell">
                                    RM {{ "%.2f"|format(
                                        row.actual_cash or 0
                                    ) }}
                                </td>

                                <td class="money-cell">
                                    RM {{ "%.2f"|format(
                                        row.difference or 0
                                    ) }}
                                </td>

                                <td>
                                    {{ row.status }}
                                </td>

                                <td>
                                    {{ row.checked_by or "-" }}
                                </td>
                            </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>

        {% else %}
            <div class="empty-state">
                目前还没有现金对账记录。
            </div>
        {% endif %}
    </div>

    <div class="btn-row">
        <a
            class="btn-tool btn-secondary"
            href="{{ url_for(
                'finance_month_end.month_end_home',
                ym=ym
            ) }}"
        >
            ← 返回财政月结
        </a>

        <a
            class="btn-tool btn-primary"
            href="{{ url_for(
                'finance_month_end.cash_bank_in',
                ym=ym,
                fund_account=fund_account
            ) }}"
        >
            查看 Cash Bank In
        </a>
    </div>

</div>
</body>
</html>
        """,
        ym=ym,
        fund_account=fund_account,
        fund_accounts=fund_accounts,
        summary=summary,
        form_data=form_data,
        difference=difference,
        history=history,
        message=message,
        message_type=message_type
    )


@finance_month_end_bp.route(
    "/bank_in/<int:deposit_id>/delete",
    methods=["POST"]
)
def delete_cash_bank_in(deposit_id):
    if not session.get("finance_login"):
        return redirect(url_for("finance_month_end.cash_bank_in"))

    row = db_query("""
        select
            ym,
            fund_account
        from finance_bank_deposits
        where id = %s
        limit 1
    """, (deposit_id,), fetchone=True)

    if row:
        month_lock_error = require_finance_month_open(
            row["ym"],
            row["fund_account"]
        )

        if month_lock_error:
            flash(month_lock_error, "danger")
            return redirect(
                url_for(
                    "finance_month_end.cash_bank_in",
                    ym=row["ym"],
                    fund_account=row["fund_account"]
                )
            )

        db_query("""
            delete from finance_bank_deposits
            where id = %s
        """, (deposit_id,))

        return redirect(
            url_for(
                "finance_month_end.cash_bank_in",
                ym=row["ym"],
                fund_account=row["fund_account"],
                deleted="1"
            )
        )

    return redirect(
        url_for("finance_month_end.month_end_home")
    )




def get_month_bank_income(ym, fund_account):
    row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'income'
          and fund_account = %s
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and payment_method in (
              '银行过账',
              '支票'
          )
    """, (fund_account, ym), fetchone=True)

    return float((row or {}).get("total") or 0)


def get_month_bank_expense(ym, fund_account):
    row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'expense'
          and fund_account = %s
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and payment_method in (
              '银行过账',
              '支票'
          )
    """, (fund_account, ym), fetchone=True)

    return float((row or {}).get("total") or 0)


def get_previous_bank_closing(ym, fund_account):
    previous_ym = get_previous_month_ym(ym)

    row = db_query("""
        select statement_balance
        from finance_bank_reconciliation
        where ym = %s
          and fund_account = %s
          and status = 'confirmed'
        limit 1
    """, (previous_ym, fund_account), fetchone=True)

    return float((row or {}).get("statement_balance") or 0)


def get_bank_reconciliation_summary(
    ym,
    fund_account,
    opening_bank=None
):
    if opening_bank is None:
        opening_bank = get_previous_bank_closing(
            ym,
            fund_account
        )

    bank_income = get_month_bank_income(
        ym,
        fund_account
    )

    bank_expense = get_month_bank_expense(
        ym,
        fund_account
    )

    cash_bank_in = get_month_bank_in_total(
        ym,
        fund_account
    )

    expected_bank = (
        float(opening_bank or 0)
        + bank_income
        + cash_bank_in
        - bank_expense
    )

    return {
        "opening_bank": float(opening_bank or 0),
        "bank_income": bank_income,
        "bank_expense": bank_expense,
        "cash_bank_in": cash_bank_in,
        "expected_bank": expected_bank,
    }


@finance_month_end_bp.route(
    "/bank_reconciliation",
    methods=["GET", "POST"]
)
def bank_reconciliation():

    if not session.get("finance_login"):
        return redirect(
            url_for("finance.finance_login")
        )

    fund_accounts = [
        "观音堂日常户口",
        "总会户口",
    ]

    ym = (
        request.form.get("ym")
        or request.args.get("ym")
        or date.today().strftime("%Y-%m")
    ).strip()

    fund_account = (
        request.form.get("fund_account")
        or request.args.get("fund_account")
        or "观音堂日常户口"
    ).strip()

    if fund_account not in fund_accounts:
        fund_account = "观音堂日常户口"

    existing = db_query("""
        select *
        from finance_bank_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    opening_bank = (
        float(existing.get("opening_bank") or 0)
        if existing
        else get_previous_bank_closing(
            ym,
            fund_account
        )
    )

    summary = get_bank_reconciliation_summary(
        ym,
        fund_account,
        opening_bank
    )

    statement_balance = (
        float(existing.get("statement_balance") or 0)
        if existing
        else None
    )

    difference = (
        statement_balance - summary["expected_bank"]
        if statement_balance is not None
        else None
    )

    message = ""
    message_type = "danger"

    form_data = {
        "opening_bank": f"{opening_bank:.2f}",
        "statement_balance": (
            f"{statement_balance:.2f}"
            if statement_balance is not None
            else ""
        ),
        "checked_by": (
            existing.get("checked_by") or ""
            if existing else ""
        ),
        "checked_date": (
            str(existing.get("checked_date"))
            if existing and existing.get("checked_date")
            else date.today().isoformat()
        ),
        "remarks": (
            existing.get("remarks") or ""
            if existing else ""
        ),
    }

    if request.method == "POST":

        month_lock_error = require_finance_month_open(
            ym,
            fund_account
        )

        opening_bank = money(
            request.form.get("opening_bank")
        )

        statement_balance = money(
            request.form.get("statement_balance")
        )

        checked_by = request.form.get(
            "checked_by",
            ""
        ).strip()

        checked_date = (
            request.form.get("checked_date")
            or date.today().isoformat()
        )

        remarks = request.form.get(
            "remarks",
            ""
        ).strip()

        action = request.form.get(
            "action",
            "save"
        )

        summary = get_bank_reconciliation_summary(
            ym,
            fund_account,
            opening_bank
        )

        difference = (
            statement_balance
            - summary["expected_bank"]
        )

        form_data = {
            "opening_bank": request.form.get(
                "opening_bank",
                "0"
            ),
            "statement_balance": request.form.get(
                "statement_balance",
                "0"
            ),
            "checked_by": checked_by,
            "checked_date": str(checked_date),
            "remarks": remarks,
        }

        if month_lock_error:
            message = month_lock_error

        elif opening_bank < 0:
            message = "期初银行余额不能少于 RM0。"

        elif statement_balance < 0:
            message = "Statement Ending Balance 不能少于 RM0。"

        elif action == "confirm" and not checked_by:
            message = "确认银行对账时必须填写核对人。"

        else:
            status = (
                "confirmed"
                if action == "confirm"
                else "draft"
            )

            db_query("""
                insert into finance_bank_reconciliation
                (
                    ym,
                    fund_account,
                    opening_bank,
                    bank_income,
                    bank_expense,
                    cash_bank_in,
                    expected_bank,
                    statement_balance,
                    difference,
                    checked_by,
                    checked_date,
                    status,
                    remarks,
                    updated_at
                )
                values
                (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, now()
                )
                on conflict (ym, fund_account)
                do update set
                    opening_bank = excluded.opening_bank,
                    bank_income = excluded.bank_income,
                    bank_expense = excluded.bank_expense,
                    cash_bank_in = excluded.cash_bank_in,
                    expected_bank = excluded.expected_bank,
                    statement_balance = excluded.statement_balance,
                    difference = excluded.difference,
                    checked_by = excluded.checked_by,
                    checked_date = excluded.checked_date,
                    status = excluded.status,
                    remarks = excluded.remarks,
                    updated_at = now()
            """, (
                ym,
                fund_account,
                summary["opening_bank"],
                summary["bank_income"],
                summary["bank_expense"],
                summary["cash_bank_in"],
                summary["expected_bank"],
                statement_balance,
                difference,
                checked_by,
                checked_date,
                status,
                remarks
            ))

            return redirect(
                url_for(
                    "finance_month_end.bank_reconciliation",
                    ym=ym,
                    fund_account=fund_account,
                    saved="1",
                    status=status
                )
            )

    if request.args.get("saved") == "1":
        message = (
            "Bank Reconciliation 已确认完成。"
            if request.args.get("status") == "confirmed"
            else "Bank Reconciliation 草稿已保存。"
        )
        message_type = "success"

    history = db_query("""
        select *
        from finance_bank_reconciliation
        where fund_account = %s
        order by ym desc
        limit 24
    """, (fund_account,), fetchall=True)

    return render_template_string(
        FINANCE_DATE_COMPONENT + """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Bank Reconciliation</title>
<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">

<style>
.recon-page{max-width:1100px}
.recon-header{
    background:linear-gradient(135deg,#1d4ed8,#1e3a8a);
    color:#fff;padding:28px;border-radius:22px;margin-bottom:20px
}
.recon-header h1{margin:0 0 8px}
.recon-header p{margin:0;opacity:.92}
.two-grid{
    display:grid;
    grid-template-columns:repeat(2,minmax(0,1fr));
    gap:16px
}
.full{grid-column:1/-1}
.calc-list{display:grid;gap:10px}
.calc-row{
    display:grid;
    grid-template-columns:60px 1fr 220px;
    align-items:center;gap:12px;
    padding:15px;border:1px solid #e2e8f0;
    border-radius:14px;background:#f8fafc
}
.calc-symbol{font-size:27px;font-weight:900;text-align:center}
.calc-label{font-size:18px;font-weight:800}
.calc-value{text-align:right;font-size:22px;font-weight:900}
.expected{background:#eff6ff;border-color:#93c5fd}
.expected .calc-value{color:#1d4ed8;font-size:27px}
.diff{
    padding:20px;border-radius:16px;text-align:center;margin-top:18px
}
.diff-ok{background:#f0fdf4;border:1px solid #86efac;color:#166534}
.diff-bad{background:#fef2f2;border:1px solid #fca5a5;color:#991b1b}
.diff-value{font-size:32px;font-weight:900}
.action-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:18px}
.history-table{min-width:1200px}
.money{text-align:right;white-space:nowrap}
@media(max-width:700px){
    .two-grid,.action-grid{grid-template-columns:1fr}
    .full{grid-column:auto}
    .calc-row{grid-template-columns:45px 1fr}
    .calc-value{grid-column:1/-1;text-align:left;padding-left:57px}
}
</style>
</head>

<body>
<div class="page recon-page">

    <div class="recon-header">
        <h1>🏦 Bank Reconciliation</h1>
        <p>系统自动计算银行账面余额，财政填写银行月结单余额。</p>
    </div>

    {% if message %}
        <div class="alert
            {% if message_type == 'success' %}
                alert-success
            {% else %}
                alert-danger
            {% endif %}
        ">
            {{ message }}
        </div>
    {% endif %}

    <div class="card">
        <div class="section-title">🔎 查询月份与户口</div>

        <form method="get">
            <div class="two-grid">
                <div class="form-group">
                    <label class="form-label">月份</label>
                    <input class="form-input"
                           name="ym"
                           type="month"
                           value="{{ ym }}"
                           required>
                </div>

                <div class="form-group">
                    <label class="form-label">基金户口</label>
                    <select class="form-input" name="fund_account">
                        {% for account in fund_accounts %}
                            <option value="{{ account }}"
                                {% if fund_account == account %}selected{% endif %}
                            >
                                {{ account }}
                            </option>
                        {% endfor %}
                    </select>
                </div>

                <div class="full">
                    <button class="btn-tool btn-primary" type="submit">
                        查看
                    </button>
                </div>
            </div>
        </form>
    </div>

    <div class="card">
        <div class="section-title">🧮 系统银行余额</div>

        <div class="calc-list">
            <div class="calc-row">
                <div class="calc-symbol">🏦</div>
                <div class="calc-label">上月银行结余</div>
                <div class="calc-value">
                    RM {{ "%.2f"|format(summary.opening_bank) }}
                </div>
            </div>

            <div class="calc-row">
                <div class="calc-symbol">＋</div>
                <div class="calc-label">本月银行收入</div>
                <div class="calc-value">
                    RM {{ "%.2f"|format(summary.bank_income) }}
                </div>
            </div>

            <div class="calc-row">
                <div class="calc-symbol">＋</div>
                <div class="calc-label">现金存入银行</div>
                <div class="calc-value">
                    RM {{ "%.2f"|format(summary.cash_bank_in) }}
                </div>
            </div>

            <div class="calc-row">
                <div class="calc-symbol">−</div>
                <div class="calc-label">本月银行支出</div>
                <div class="calc-value">
                    RM {{ "%.2f"|format(summary.bank_expense) }}
                </div>
            </div>

            <div class="calc-row expected">
                <div class="calc-symbol">＝</div>
                <div class="calc-label">系统银行余额</div>
                <div class="calc-value">
                    RM {{ "%.2f"|format(summary.expected_bank) }}
                </div>
            </div>
        </div>
    </div>

    <div class="card">
        <div class="section-title">✍️ 银行 Statement 核对</div>

        <form method="post">
            <input type="hidden" name="ym" value="{{ ym }}">
            <input type="hidden" name="fund_account" value="{{ fund_account }}">

            <div class="two-grid">
                <div class="form-group">
                    <label class="form-label">上月银行结余 RM</label>
                    <input class="form-input"
                           name="opening_bank"
                           type="number"
                           step="0.01"
                           min="0"
                           value="{{ form_data.opening_bank }}"
                           required>
                </div>

                <div class="form-group">
                    <label class="form-label">
                        Statement Ending Balance RM
                    </label>
                    <input class="form-input"
                           name="statement_balance"
                           type="number"
                           step="0.01"
                           min="0"
                           value="{{ form_data.statement_balance }}"
                           required>
                </div>

                <div class="form-group">
                    <label class="form-label">核对人</label>
                    <input class="form-input"
                           name="checked_by"
                           value="{{ form_data.checked_by }}">
                </div>

                <div class="form-group">
                    <label class="form-label">核对日期</label>
                    <input class="form-input"
                           name="checked_date"
                           type="date"
                           value="{{ form_data.checked_date }}"
                           required>
                </div>

                <div class="form-group full">
                    <label class="form-label">备注／差异说明</label>
                    <textarea class="form-input"
                              name="remarks"
                              rows="4">{{ form_data.remarks }}</textarea>
                </div>
            </div>

            {% if difference is not none %}
                <div class="diff
                    {% if difference|abs < 0.005 %}
                        diff-ok
                    {% else %}
                        diff-bad
                    {% endif %}
                ">
                    <div>
                        {% if difference|abs < 0.005 %}
                            ✅ 银行对账一致
                        {% else %}
                            ⚠️ 银行差异
                        {% endif %}
                    </div>
                    <div class="diff-value">
                        RM {{ "%.2f"|format(difference) }}
                    </div>
                </div>
            {% endif %}

            <div class="action-grid">
                <button class="btn-tool btn-secondary"
                        type="submit"
                        name="action"
                        value="save">
                    💾 保存草稿
                </button>

                <button class="btn-tool btn-success"
                        type="submit"
                        name="action"
                        value="confirm"
                        onclick="return confirm('确定完成银行对账？');">
                    ✅ 确认完成
                </button>
            </div>
        </form>
    </div>

    <div class="card">
        <div class="section-title">📋 银行对账历史</div>

        {% if history %}
            <div class="table-responsive">
                <table class="record-table history-table">
                    <thead>
                        <tr>
                            <th>月份</th>
                            <th>户口</th>
                            <th>期初</th>
                            <th>银行收入</th>
                            <th>Cash Bank In</th>
                            <th>银行支出</th>
                            <th>系统余额</th>
                            <th>Statement</th>
                            <th>差异</th>
                            <th>状态</th>
                            <th>核对人</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for row in history %}
                            <tr>
                                <td>{{ row.ym }}</td>
                                <td>{{ row.fund_account }}</td>
                                <td class="money">RM {{ "%.2f"|format(row.opening_bank or 0) }}</td>
                                <td class="money">RM {{ "%.2f"|format(row.bank_income or 0) }}</td>
                                <td class="money">RM {{ "%.2f"|format(row.cash_bank_in or 0) }}</td>
                                <td class="money">RM {{ "%.2f"|format(row.bank_expense or 0) }}</td>
                                <td class="money">RM {{ "%.2f"|format(row.expected_bank or 0) }}</td>
                                <td class="money">RM {{ "%.2f"|format(row.statement_balance or 0) }}</td>
                                <td class="money">RM {{ "%.2f"|format(row.difference or 0) }}</td>
                                <td>{{ row.status }}</td>
                                <td>{{ row.checked_by or "-" }}</td>
                            </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        {% else %}
            <div class="empty-state">目前还没有银行对账记录。</div>
        {% endif %}
    </div>

    <div class="btn-row">
        <a class="btn-tool btn-secondary"
           href="{{ url_for(
               'finance_month_end.month_end_home',
               ym=ym
           ) }}">
            ← 返回财政月结
        </a>
    </div>

</div>
</body>
</html>
        """,
        ym=ym,
        fund_account=fund_account,
        fund_accounts=fund_accounts,
        summary=summary,
        form_data=form_data,
        difference=difference,
        history=history,
        message=message,
        message_type=message_type
    )


def _hq_style_sheet(ws, title, headers):
    dark = "1F4E78"
    light = "D9EAF7"
    thin = Side(style="thin", color="B7C9D6")
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    ws.sheet_view.showGridLines = False
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=2,
        end_column=len(headers)
    )

    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )
    ws.cell(1, 1).fill = PatternFill(
        "solid",
        fgColor=dark
    )
    ws.cell(1, 1).alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor=light
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = (
        f"A4:{get_column_letter(len(headers))}4"
    )

    return border


def _hq_write_rows(ws, rows, start_row, border):
    for r_index, row_values in enumerate(rows, start_row):
        for c_index, value in enumerate(row_values, 1):
            cell = ws.cell(r_index, c_index)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )


def _hq_auto_width(ws, max_width=35):
    for column_cells in ws.columns:
        letter = get_column_letter(
            column_cells[0].column
        )
        max_len = 0

        for cell in column_cells:
            if cell.value is not None:
                max_len = max(
                    max_len,
                    len(str(cell.value))
                )

        ws.column_dimensions[letter].width = min(
            max_len + 4,
            max_width
        )


@finance_month_end_bp.route("/hq_report")
def hq_report_package():

    if not session.get("finance_login"):
        return redirect(
            url_for("finance.finance_login")
        )

    ym = request.args.get(
        "ym",
        date.today().strftime("%Y-%m")
    )

    fund_account = request.args.get(
        "fund_account",
        "观音堂日常户口"
    )

    cash_recon = db_query("""
        select *
        from finance_cash_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    bank_recon = db_query("""
        select *
        from finance_bank_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    deposit_summary = db_query("""
        select
            count(*) as count,
            coalesce(sum(amount), 0) as total
        from finance_bank_deposits
        where ym = %s
          and fund_account = %s
    """, (ym, fund_account), fetchone=True)

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>HQ Report Package</title>
<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">
<style>
.hq-page{max-width:900px}
.hq-header{
    background:linear-gradient(135deg,#b45309,#92400e);
    color:#fff;padding:28px;border-radius:22px;margin-bottom:20px
}
.hq-header h1{margin:0 0 8px}
.hq-header p{margin:0;opacity:.92}
.status-grid{
    display:grid;
    grid-template-columns:repeat(2,minmax(0,1fr));
    gap:14px
}
.status-card{
    border:1px solid #e2e8f0;
    border-radius:15px;padding:17px;background:#f8fafc
}
.status-title{font-weight:900;margin-bottom:7px}
.ok{color:#15803d}
.wait{color:#b45309}
@media(max-width:650px){
    .status-grid{grid-template-columns:1fr}
}
</style>
</head>
<body>
<div class="page hq-page">

    <div class="hq-header">
        <h1>📄 HQ Report Package</h1>
        <p>一次生成总会需要的七项财政资料。</p>
    </div>

    <div class="card">
        <form method="get">
            <div class="form-group">
                <label class="form-label">月份</label>
                <input class="form-input"
                       name="ym"
                       type="month"
                       value="{{ ym }}">
            </div>

            <div class="form-group">
                <label class="form-label">基金户口</label>
                <select class="form-input" name="fund_account">
                    {% for account in [
                        '观音堂日常户口',
                        '总会户口'
                    ] %}
                        <option value="{{ account }}"
                            {% if fund_account == account %}selected{% endif %}
                        >
                            {{ account }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <button class="btn-tool btn-primary" type="submit">
                查看状态
            </button>
        </form>
    </div>

    <div class="card">
        <div class="section-title">📋 月结资料状态</div>

        <div class="status-grid">
            <div class="status-card">
                <div class="status-title">Cash Bank In</div>
                <div class="{% if deposit_summary.count %}ok{% else %}wait{% endif %}">
                    {{ deposit_summary.count or 0 }} 笔，
                    RM {{ "%.2f"|format(deposit_summary.total or 0) }}
                </div>
            </div>

            <div class="status-card">
                <div class="status-title">Cash Reconciliation</div>
                <div class="{% if cash_recon and cash_recon.status == 'confirmed' %}ok{% else %}wait{% endif %}">
                    {% if cash_recon %}
                        {{ cash_recon.status }}
                    {% else %}
                        尚未建立
                    {% endif %}
                </div>
            </div>

            <div class="status-card">
                <div class="status-title">Bank Reconciliation</div>
                <div class="{% if bank_recon and bank_recon.status == 'confirmed' %}ok{% else %}wait{% endif %}">
                    {% if bank_recon %}
                        {{ bank_recon.status }}
                    {% else %}
                        尚未建立
                    {% endif %}
                </div>
            </div>

            <div class="status-card">
                <div class="status-title">HQ Excel Package</div>
                <div class="ok">可以生成</div>
            </div>
        </div>
    </div>

    <div class="card">
        <a class="btn-tool btn-success"
           style="width:100%;"
           href="{{ url_for(
               'finance_month_end.export_hq_report',
               ym=ym,
               fund_account=fund_account
           ) }}">
            📥 生成总会七项 Excel
        </a>
    </div>

    <div class="btn-row">
        <a class="btn-tool btn-secondary"
           href="{{ url_for(
               'finance_month_end.month_end_home',
               ym=ym
           ) }}">
            ← 返回财政月结
        </a>
    </div>

</div>
</body>
</html>
    """,
        ym=ym,
        fund_account=fund_account,
        cash_recon=cash_recon,
        bank_recon=bank_recon,
        deposit_summary=deposit_summary
    )


@finance_month_end_bp.route("/hq_report/export")
def export_hq_report():

    if not session.get("finance_login"):
        return redirect(
            url_for("finance.finance_login")
        )

    ym = request.args.get(
        "ym",
        date.today().strftime("%Y-%m")
    )

    fund_account = request.args.get(
        "fund_account",
        "观音堂日常户口"
    )

    records = db_query("""
        select *
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and fund_account = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
        order by record_date, id
    """, (ym, fund_account), fetchall=True)

    deposits = db_query("""
        select *
        from finance_bank_deposits
        where ym = %s
          and fund_account = %s
        order by deposit_date, id
    """, (ym, fund_account), fetchall=True)

    cash_recon = db_query("""
        select *
        from finance_cash_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    bank_recon = db_query("""
        select *
        from finance_bank_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    wb = Workbook()
    wb.remove(wb.active)

    # 01 Bank Reconciliation
    ws = wb.create_sheet("01 Bank Reconciliation")
    headers = ["项目", "金额 RM", "状态／说明"]
    border = _hq_style_sheet(
        ws,
        f"Bank Reconciliation - {ym}",
        headers
    )

    br = bank_recon or {}
    rows = [
        ["Opening Bank", float(br.get("opening_bank") or 0), ""],
        ["Bank Income", float(br.get("bank_income") or 0), ""],
        ["Cash Bank In", float(br.get("cash_bank_in") or 0), ""],
        ["Bank Expense", float(br.get("bank_expense") or 0), ""],
        ["System Balance", float(br.get("expected_bank") or 0), ""],
        ["Statement Balance", float(br.get("statement_balance") or 0), ""],
        ["Difference", float(br.get("difference") or 0), br.get("status") or "未对账"],
    ]
    _hq_write_rows(ws, rows, 5, border)

    # 02 Invoices
    ws = wb.create_sheet("02 Invoices")
    headers = [
        "日期", "Payment Voucher", "付款对象",
        "类别", "金额 RM", "付款方式", "备注"
    ]
    border = _hq_style_sheet(
        ws,
        f"Invoices / Expense Supporting List - {ym}",
        headers
    )

    invoice_rows = []
    for r in records:
        if r.get("record_type") == "expense":
            invoice_rows.append([
                r.get("record_date"),
                r.get("payment_voucher_no") or "-",
                r.get("vendor") or r.get("name") or "-",
                r.get("category") or "-",
                float(r.get("amount") or 0),
                r.get("payment_method") or "-",
                r.get("remarks") or "-",
            ])
    _hq_write_rows(ws, invoice_rows, 5, border)

    # 03 Cash Receipts
    ws = wb.create_sheet("03 Cash Receipts")
    headers = [
        "付款日期", "收条日期", "Receipt No.",
        "类别", "编号", "姓名", "金额 RM", "备注"
    ]
    border = _hq_style_sheet(
        ws,
        f"Cash Receipts - {ym}",
        headers
    )

    cash_rows = []
    for r in records:
        if (
            r.get("record_type") == "income"
            and r.get("payment_method") == "现金"
        ):
            cash_rows.append([
                r.get("record_date"),
                r.get("receipt_date"),
                r.get("receipt_no") or "-",
                r.get("category") or "-",
                r.get("member_id") or "-",
                r.get("name") or "-",
                float(r.get("amount") or 0),
                r.get("remarks") or "-",
            ])
    _hq_write_rows(ws, cash_rows, 5, border)

    # 04 Bank Receipts
    ws = wb.create_sheet("04 Bank Receipts")
    headers = [
        "付款日期", "Receipt No.", "类别",
        "编号", "姓名", "金额 RM",
        "Bank Reference", "备注"
    ]
    border = _hq_style_sheet(
        ws,
        f"Bank Receipts - {ym}",
        headers
    )

    bank_rows = []
    for r in records:
        if (
            r.get("record_type") == "income"
            and r.get("payment_method") in (
                "银行过账",
                "支票"
            )
        ):
            bank_rows.append([
                r.get("record_date"),
                r.get("receipt_no") or "-",
                r.get("category") or "-",
                r.get("member_id") or "-",
                r.get("name") or "-",
                float(r.get("amount") or 0),
                r.get("bank_ref") or "-",
                r.get("remarks") or "-",
            ])
    _hq_write_rows(ws, bank_rows, 5, border)

    # 05 Cash Collections Banked In
    ws = wb.create_sheet("05 Cash Banked In")
    headers = [
        "Bank In 日期", "银行", "Reference",
        "Receipt From", "Receipt To", "金额 RM", "备注"
    ]
    border = _hq_style_sheet(
        ws,
        f"Cash Collections Banked In - {ym}",
        headers
    )

    deposit_rows = []
    for r in deposits:
        deposit_rows.append([
            r.get("deposit_date"),
            r.get("bank_name") or "-",
            r.get("reference_no") or "-",
            r.get("receipt_from") or "-",
            r.get("receipt_to") or "-",
            float(r.get("amount") or 0),
            r.get("remarks") or "-",
        ])
    _hq_write_rows(ws, deposit_rows, 5, border)

    # 06 Petty Cash Vouchers
    ws = wb.create_sheet("06 Petty Cash Vouchers")
    headers = [
        "日期", "Payment Voucher", "类别",
        "付款对象", "金额 RM", "付款方式", "用途"
    ]
    border = _hq_style_sheet(
        ws,
        f"Petty Cash Vouchers - {ym}",
        headers
    )

    petty_rows = []
    for r in records:
        if (
            r.get("record_type") == "expense"
            and r.get("payment_method") == "现金"
        ):
            petty_rows.append([
                r.get("record_date"),
                r.get("payment_voucher_no") or "-",
                r.get("category") or "-",
                r.get("vendor") or r.get("name") or "-",
                float(r.get("amount") or 0),
                r.get("payment_method") or "-",
                r.get("remarks") or "-",
            ])
    _hq_write_rows(ws, petty_rows, 5, border)

    # 07 Cash in Hand Reconciliation
    ws = wb.create_sheet("07 Cash Reconciliation")
    headers = ["项目", "金额 RM", "状态／说明"]
    border = _hq_style_sheet(
        ws,
        f"Cash in Hand Reconciliation - {ym}",
        headers
    )

    cr = cash_recon or {}
    cash_recon_rows = [
        ["Opening Cash", float(cr.get("opening_cash") or 0), ""],
        ["Cash Income", float(cr.get("cash_income") or 0), ""],
        ["Cash Expense", float(cr.get("cash_expense") or 0), ""],
        ["Cash Banked In", float(cr.get("cash_banked_in") or 0), ""],
        ["Expected Cash", float(cr.get("expected_cash") or 0), ""],
        ["Actual Cash", float(cr.get("actual_cash") or 0), ""],
        ["Difference", float(cr.get("difference") or 0), cr.get("status") or "未对账"],
    ]
    _hq_write_rows(ws, cash_recon_rows, 5, border)

    for ws in wb.worksheets:
        _hq_auto_width(ws)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        for row in ws.iter_rows(min_row=5):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '"RM"#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=(
            f"CHE_HQ_Finance_Package_{ym}.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


@finance_month_end_bp.route(
    "/month_close",
    methods=["GET", "POST"]
)
def month_close():

    if not session.get("finance_login"):
        return redirect(
            url_for("finance.finance_login")
        )

    ym = (
        request.form.get("ym")
        or request.args.get("ym")
        or date.today().strftime("%Y-%m")
    )

    fund_account = (
        request.form.get("fund_account")
        or request.args.get("fund_account")
        or "观音堂日常户口"
    )

    message = ""

    cash_recon = db_query("""
        select status
        from finance_cash_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    bank_recon = db_query("""
        select status
        from finance_bank_reconciliation
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    closed = db_query("""
        select *
        from finance_month_close
        where ym = %s
          and fund_account = %s
        limit 1
    """, (ym, fund_account), fetchone=True)

    if request.method == "POST":

        closed_by = request.form.get(
            "closed_by",
            ""
        ).strip()

        remarks = request.form.get(
            "remarks",
            ""
        ).strip()

        if closed and closed.get("status") == "closed":
            message = "这个月份已经完成月结，无需重复关闭。"

        elif not closed_by:
            message = "请填写月结负责人。"

        elif not cash_recon or cash_recon.get("status") != "confirmed":
            message = "Cash Reconciliation 尚未确认。"

        elif not bank_recon or bank_recon.get("status") != "confirmed":
            message = "Bank Reconciliation 尚未确认。"

        else:
            db_query("""
                insert into finance_month_close
                (
                    ym,
                    fund_account,
                    cash_reconciliation_status,
                    bank_reconciliation_status,
                    closed_by,
                    closed_date,
                    status,
                    remarks
                )
                values
                (
                    %s, %s, %s, %s,
                    %s, %s, 'closed', %s
                )
                on conflict (ym, fund_account)
                do update set
                    cash_reconciliation_status =
                        excluded.cash_reconciliation_status,
                    bank_reconciliation_status =
                        excluded.bank_reconciliation_status,
                    closed_by = excluded.closed_by,
                    closed_date = excluded.closed_date,
                    status = 'closed',
                    remarks = excluded.remarks
            """, (
                ym,
                fund_account,
                cash_recon.get("status"),
                bank_recon.get("status"),
                closed_by,
                date.today().isoformat(),
                remarks
            ))

            write_finance_audit(
                module="month_close",
                action="close",
                record_id=f"{ym}:{fund_account}",
                new_value={
                    "ym": ym,
                    "fund_account": fund_account,
                    "cash_reconciliation_status": cash_recon.get("status"),
                    "bank_reconciliation_status": bank_recon.get("status"),
                    "closed_by": closed_by,
                    "status": "closed",
                    "remarks": remarks,
                },
                reason=remarks,
                actor=closed_by or get_current_finance_user(),
            )

            return redirect(
                url_for(
                    "finance_month_end.month_close",
                    ym=ym,
                    fund_account=fund_account,
                    saved="1"
                )
            )

    if request.args.get("saved") == "1":
        message = "本月财政月结已完成。"

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Month Close</title>
<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">
<style>
.close-page{max-width:820px}
.close-header{
    background:linear-gradient(135deg,#334155,#0f172a);
    color:#fff;padding:28px;border-radius:22px;margin-bottom:20px
}
.close-header h1{margin:0 0 8px}
.close-header p{margin:0;opacity:.92}
.status-line{
    display:flex;justify-content:space-between;gap:12px;
    padding:15px;border:1px solid #e2e8f0;
    border-radius:14px;margin-bottom:10px
}
.ok{color:#15803d;font-weight:900}
.wait{color:#b45309;font-weight:900}
</style>
</head>
<body>
<div class="page close-page">

    <div class="close-header">
        <h1>🔒 Month Close</h1>
        <p>确认现金与银行对账完成后，登记本月结账。</p>
    </div>

    {% if message %}
        <div class="alert alert-success">{{ message }}</div>
    {% endif %}

    <div class="card">
        <form method="get">
            <div class="form-group">
                <label class="form-label">月份</label>
                <input class="form-input"
                       name="ym"
                       type="month"
                       value="{{ ym }}">
            </div>

            <div class="form-group">
                <label class="form-label">基金户口</label>
                <select class="form-input" name="fund_account">
                    {% for account in [
                        '观音堂日常户口',
                        '总会户口'
                    ] %}
                        <option value="{{ account }}"
                            {% if fund_account == account %}selected{% endif %}
                        >
                            {{ account }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <button class="btn-tool btn-primary" type="submit">
                查看
            </button>
        </form>
    </div>

    <div class="card">
        <div class="status-line">
            <span>Cash Reconciliation</span>
            <span class="
                {% if cash_recon and cash_recon.status == 'confirmed' %}
                    ok
                {% else %}
                    wait
                {% endif %}
            ">
                {{ cash_recon.status if cash_recon else "未完成" }}
            </span>
        </div>

        <div class="status-line">
            <span>Bank Reconciliation</span>
            <span class="
                {% if bank_recon and bank_recon.status == 'confirmed' %}
                    ok
                {% else %}
                    wait
                {% endif %}
            ">
                {{ bank_recon.status if bank_recon else "未完成" }}
            </span>
        </div>

        <div class="status-line">
            <span>Month Close</span>
            <span class="{% if closed %}ok{% else %}wait{% endif %}">
                {{ closed.status if closed else "未结账" }}
            </span>
        </div>
    </div>

    <div class="card">
        <form method="post">
            <input type="hidden" name="ym" value="{{ ym }}">
            <input type="hidden" name="fund_account" value="{{ fund_account }}">

            <div class="form-group">
                <label class="form-label">月结负责人</label>
                <input class="form-input"
                       name="closed_by"
                       value="{{ closed.closed_by if closed else '' }}"
                       required>
            </div>

            <div class="form-group">
                <label class="form-label">备注</label>
                <textarea class="form-input"
                          name="remarks"
                          rows="4">{{ closed.remarks if closed else '' }}</textarea>
            </div>

            <button class="btn-tool btn-success"
                    type="submit"
                    style="width:100%;"
                    onclick="return confirm('确定完成本月财政月结？');">
                ✅ 完成本月结账
            </button>
        </form>
    </div>

    <div class="btn-row">
        <a class="btn-tool btn-secondary"
           href="{{ url_for(
               'finance_month_end.month_end_home',
               ym=ym
           ) }}">
            ← 返回财政月结
        </a>
    </div>

</div>
</body>
</html>
    """,
        ym=ym,
        fund_account=fund_account,
        cash_recon=cash_recon,
        bank_recon=bank_recon,
        closed=closed,
        message=message
    )