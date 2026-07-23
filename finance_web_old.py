# finance_web.py

import os
import re
import tempfile
import hmac
import uuid

from copy import copy
from io import BytesIO
from functools import wraps
from decimal import Decimal
from pathlib import Path
from db import db_query, get_conn
from utils import normalize_member_id
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from finance_audit import write_finance_audit
from datetime import date, datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import (
    Blueprint,
    request,
    redirect,
    url_for,
    render_template_string,
    send_file,
    session,
    jsonify,
    flash,
    current_app,
    abort,
)

# Payment Voucher PDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

from finance_common import (
    money,
    normalize_finance_date,
    get_finance_ym,
    get_month_close_record,
    is_finance_month_closed,
    get_month_lock_message,
    require_finance_month_open,
    get_current_finance_user,
)


EXPENSE_SUB_CATEGORY_OPTIONS = {
    "供花": ["鲜花"],
    "供果": ["水果", "供果用品", "其它"],
    "供油": ["灯油"],
    "佛台用品": ["佛具", "佛台用品", "佛堂用品", "其它"],
    "电费": ["TNB 20-1", "TNB 20-2", "其它"],
    "水费": [
        "Air Selangor 20-1",
        "Air Selangor 20-2",
        "Indah Water 20-1",
        "Indah Water 20-2",
        "其它",
    ],
    "电话及网络费": [
        "Digi Reload",
        "Celcom Reload",
        "Unifi Internet",
        "Maxis",
        "其它",
    ],
    "维修保养": [
        "Air-con Service",
        "电器维修",
        "水管维修",
        "建筑维修",
        "其它",
    ],
    "装修": ["GYT装修", "观音堂装修", "活动中心装修", "其它"],
    "日常采购": ["文具", "厨房用品", "清洁用品", "日常用品", "其它"],
    "执照及行政费": ["License Fee", "政府费用", "行政费用", "其它"],
    "其它支出": ["印刷", "交通", "杂项", "其它"],
}

EXPENSE_VENDOR_BY_SUB_CATEGORY = {
    # 供花
    "鲜花": "Pudu Ria Florist Trading Sdn Bhd",
    "供花用品": "Pudu Ria Florist Trading Sdn Bhd",

    # 供油
    "灯油": "Laohongka Sdn Bhd",
    "油灯用品": "Laohongka Sdn Bhd",

    # 电费
    "TNB 20-1": "Tenaga Nasional",
    "TNB 20-2": "Tenaga Nasional",

    # 水费
    "Air Selangor 20-1": "Air Selangor",
    "Air Selangor 20-2": "Air Selangor",
    "Indah Water 20-1": "Indah Water",
    "Indah Water 20-2": "Indah Water",

    # 电话及网络
    "Digi Reload": "Digi",
    "Celcom Reload": "Celcom",
    "Unifi Internet": "TM Unifi",
    "Maxis": "Maxis",
}

AUTO_VENDOR_CATEGORIES = {
    "供花",
    "供油",
    "电费",
    "水费",
    "电话及网络费",
}
# =========================================================
# Finance Web local helpers
# These remain here because they are domain-specific to
# daily finance input rather than generic shared utilities.
# =========================================================

FINANCE_OPENING_DATE = date(2026, 1, 1)


def get_finance_balance_summary(
    branch="CHE",
    balance_date=None
):
    """
    计算指定分会截至某一天的资金余额。

    CHE 银行余额：
    银行期初余额
    + CHE 月费收入
    - 银行提款到 Petty Cash
    ± 银行调整

    Cash In Hand：
    现金期初余额
    + 银行提款转入现金
    - 所有现金支出
    ± 现金调整
    """

    branch = (branch or "CHE").strip().upper()
    balance_date = balance_date or date.today()

    # 1. 银行期初余额
    bank_opening_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_cash_movements
        where branch = %s
          and account_type = 'bank'
          and movement_type = 'opening'
          and record_date <= %s
    """, (
        branch,
        balance_date,
    ), fetchone=True)

    bank_opening = Decimal(
        str(bank_opening_row["total"] or 0)
    )

    # 2. 现金期初余额
    cash_opening_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_cash_movements
        where branch = %s
          and account_type = 'cash'
          and movement_type = 'opening'
          and record_date <= %s
    """, (
        branch,
        balance_date,
    ), fetchone=True)

    cash_opening = Decimal(
        str(cash_opening_row["total"] or 0)
    )

    # 3. 银行提款总额
    cash_out_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_cash_movements
        where branch = %s
          and account_type = 'bank'
          and movement_type = 'cash_out'
          and record_date between %s and %s
    """, (
        branch,
        FINANCE_OPENING_DATE,
        balance_date,
    ), fetchone=True)

    bank_cash_out = Decimal(
        str(cash_out_row["total"] or 0)
    )

    # 4. 转入 Petty Cash 的金额
    cash_in_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_cash_movements
        where branch = %s
          and account_type = 'cash'
          and movement_type = 'cash_in'
          and record_date between %s and %s
    """, (
        branch,
        FINANCE_OPENING_DATE,
        balance_date,
    ), fetchone=True)

    petty_cash_in = Decimal(
        str(cash_in_row["total"] or 0)
    )

    # 5. 银行调整
    bank_adjustment_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_cash_movements
        where branch = %s
          and account_type = 'bank'
          and movement_type = 'adjustment'
          and record_date between %s and %s
    """, (
        branch,
        FINANCE_OPENING_DATE,
        balance_date,
    ), fetchone=True)

    bank_adjustment = Decimal(
        str(bank_adjustment_row["total"] or 0)
    )

    # 6. 现金调整
    cash_adjustment_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_cash_movements
        where branch = %s
          and account_type = 'cash'
          and movement_type = 'adjustment'
          and record_date between %s and %s
    """, (
        branch,
        FINANCE_OPENING_DATE,
        balance_date,
    ), fetchone=True)

    cash_adjustment = Decimal(
        str(cash_adjustment_row["total"] or 0)
    )

    # 7. CHE 月费收入
    monthly_income_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'income'
          and category = '月费'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and record_date between %s and %s
          and (
                coalesce(member_id, '') ilike %s
             or coalesce(receipt_no, '') ilike %s
             or fund_account in (
                    '观音堂日常户口',
                    'CHE 日常户口'
                )
          )
    """, (
        FINANCE_OPENING_DATE,
        balance_date,
        "CHE%",
        "CHE%",
    ), fetchone=True)

    monthly_income = Decimal(
        str(monthly_income_row["total"] or 0)
    )

    # 8. 所有有效支出
    expense_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'expense'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and record_date between %s and %s
    """, (
        FINANCE_OPENING_DATE,
        balance_date,
    ), fetchone=True)

    cash_expense = Decimal(
        str(expense_row["total"] or 0)
    )

    bank_balance = (
        bank_opening
        + monthly_income
        - bank_cash_out
        + bank_adjustment
    )

    cash_in_hand = (
        cash_opening
        + petty_cash_in
        - cash_expense
        + cash_adjustment
    )

    return {
        "balance_date": balance_date,
        "bank_opening": bank_opening,
        "cash_opening": cash_opening,
        "monthly_income": monthly_income,
        "bank_cash_out": bank_cash_out,
        "petty_cash_in": petty_cash_in,
        "cash_expense": cash_expense,
        "bank_adjustment": bank_adjustment,
        "cash_adjustment": cash_adjustment,
        "bank_balance": bank_balance,
        "cash_in_hand": cash_in_hand,
        "total_funds": bank_balance + cash_in_hand,
    }

def normalize_cdm_reference(value):
    """
    支持输入：
    1
    5
    CDM1
    CDM-1
    cdm 1

    统一转换为：
    CDM-1
    """
    text = str(value or "").strip().upper()
    text = re.sub(r"\s+", "", text)

    if not text:
        return ""

    if text.isdigit():
        return f"CDM-{int(text)}"

    match = re.fullmatch(r"CDM-?(\d+)", text)

    if match:
        return f"CDM-{int(match.group(1))}"

    return ""


def get_next_cdm_reference(deposit_date=None):
    """
    根据 Bank In 日期所在月份，建议下一个 CDM 编号。

    例如该月份已有：
    CDM-1、CDM-2、CDM-4

    系统建议：
    CDM-5
    """
    deposit_date = deposit_date or date.today()

    if isinstance(deposit_date, str):
        try:
            deposit_date = datetime.strptime(
                deposit_date,
                "%Y-%m-%d"
            ).date()
        except ValueError:
            deposit_date = date.today()

    ym = deposit_date.strftime("%Y-%m")

    rows = db_query(
        """
        select
            reference_no,
            cdm_sequence
        from finance_bank_deposits
        where branch = 'CHE'
          and (
                ym = %s
                or to_char(deposit_date, 'YYYY-MM') = %s
          )
          and (
                upper(coalesce(reference_no, '')) like 'CDM-%%'
                or cdm_sequence is not null
          )
        """,
        (ym, ym),
        fetchall=True,
    ) or []

    largest = 0

    for row in rows:
        sequence = row.get("cdm_sequence")

        if sequence is not None:
            match = re.search(r"\d+", str(sequence))

            if match:
                largest = max(largest, int(match.group()))
                continue

        reference = normalize_cdm_reference(
            row.get("reference_no")
        )

        match = re.fullmatch(r"CDM-(\d+)", reference)

        if match:
            largest = max(largest, int(match.group(1)))

    return f"CDM-{largest + 1}"

def get_fund_account(
    category,
    branch="CHE",
    record_type="income"
):

    branch = (branch or "CHE").strip().upper()

    # 支出
    if record_type == "expense":

        if branch == "STW":
            return "STW 日常户口"

        return "观音堂日常户口"

    # 月费
    if category == "月费":

        if branch == "STW":
            return "STW 日常户口"

        return "观音堂日常户口"

    # 其它全部
    return "总会户口"


def add_months_ym(ym, months):
    y, m = map(int, ym.split("-"))
    m += months
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return f"{y:04d}-{m:02d}"


def normalize_phone(phone):
    """统一电话号码格式，方便历史资料和新资料互相搜索。"""
    digits = re.sub(r"\D", "", str(phone or ""))

    if digits.startswith("0060"):
        digits = digits[4:]
    elif digits.startswith("60"):
        digits = digits[2:]

    if digits and not digits.startswith("0"):
        digits = "0" + digits

    return digits


def phone_search_variants(phone):
    local_phone = normalize_phone(phone)

    if not local_phone:
        return "", ""

    international_phone = (
        "60" + local_phone[1:]
        if local_phone.startswith("0")
        else local_phone
    )

    return local_phone, international_phone


def next_month_ym(d):
    if not d:
        return date.today().strftime("%Y-%m")

    return add_months_ym(
        d.strftime("%Y-%m"),
        1
    )


def normalize_receipt_category(category: str) -> str:

    category = str(category or "").strip()

    mapping = {
        "月费": "月费",
        "财布施": "财布施",
        "观音村": "观音村",

        "膳食": "膳食结缘",
        "膳食结缘": "膳食结缘",
        "初一十五": "膳食结缘",
        "初一十五膳食结缘": "膳食结缘",
    }

    return mapping.get(category, category)


def get_next_receipt_no(
    branch: str,
    category: str,
) -> str:

    branch = str(branch or "CHE").strip().upper()
    category = normalize_receipt_category(category)

    row = db_query(
        """
        select
            prefix,
            current_number,
            number_width
        from finance_receipt_books
        where branch = %s
          and category = %s
        limit 1
        """,
        (
            branch,
            category,
        ),
        fetchone=True,
    )

    if not row:
        raise ValueError(
            f"找不到收条簿设置：{branch} / {category}"
        )

    next_number = int(
        row.get("current_number") or 0
    ) + 1

    prefix = str(
        row.get("prefix") or ""
    )

    width = int(
        row.get("number_width") or 6
    )

    return f"{prefix}{next_number:0{width}d}"

def get_current_receipt_book(
    branch,
    category,
):

    branch = str(branch or "CHE").strip().upper()
    category = normalize_receipt_category(category)

    row = db_query(
        """
        select current_book_no
        from finance_receipt_books
        where branch=%s
          and category=%s
        limit 1
        """,
        (
            branch,
            category,
        ),
        fetchone=True,
    )

    if not row:
        return None

    return row["current_book_no"]

def update_receipt_book_number(
    branch: str,
    category: str,
    receipt_no: str,
):

    branch = str(branch or "CHE").strip().upper()
    category = normalize_receipt_category(category)
    receipt_no = str(receipt_no or "").strip().upper()

    digits = "".join(
        ch for ch in receipt_no
        if ch.isdigit()
    )

    if not digits:
        return

    number = int(digits)

    db_query(
        """
        update finance_receipt_books
        set
            current_number = greatest(
                current_number,
                %s
            ),
            updated_at = now()
        where branch = %s
          and category = %s
        """,
        (
            number,
            branch,
            category,
        ),
    )

def update_receipt_book_no(
    branch,
    category,
    book_no,
):

    branch = str(branch or "CHE").strip().upper()
    category = normalize_receipt_category(category)

    db_query(
        """
        update finance_receipt_books
        set
            current_book_no=%s,
            updated_at=now()
        where branch=%s
          and category=%s
        """,
        (
            int(book_no),
            branch,
            category,
        ),
    )

def get_receipt_book_info(
    branch,
    category,
):

    branch = str(branch or "CHE").strip().upper()
    category = normalize_receipt_category(category)

    return db_query(
        """
        select
            prefix,
            current_number,
            current_book_no,
            number_width
        from finance_receipt_books
        where branch=%s
          and category=%s
        limit 1
        """,
        (
            branch,
            category,
        ),
        fetchone=True,
    )


def get_active_finance_vendors():
    return db_query("""
        select
            id,
            vendor_name,
            sort_order
        from finance_vendors
        where is_active = true
        order by
            sort_order,
            vendor_name
    """, fetchall=True)


finance_bp = Blueprint("finance", __name__, url_prefix="/finance")

# Legacy boundary: V7 banking pages live in finance_v7.py.
# These old URLs only redirect for compatibility; do not add new V7 modules here.

FINANCE_PIN = "123456"

# 财政负责人密码必须放在 Render Environment：
# FINANCE_ADMIN_PASSWORD=你的负责人密码
FINANCE_ADMIN_PASSWORD = "123789"
      
def finance_admin_required(view_func):
    """只允许已经通过负责人密码验证的财政负责人进入。"""
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("finance_login"):
            return redirect(url_for("finance.finance_login"))

        if not session.get("finance_admin"):
            return redirect(
                url_for(
                    "finance.finance_admin_login",
                    next=request.path,
                )
            )

        return view_func(*args, **kwargs)

    return wrapped


def get_receipt_book_status(branch, category):
    """返回收条簿最后号码与下一张号码；菜单显示用。"""
    branch = str(branch or "CHE").strip().upper()
    category = normalize_receipt_category(category)

    row = db_query(
        """
        select prefix, current_number, number_width
        from finance_receipt_books
        where branch = %s and category = %s
        limit 1
        """,
        (branch, category),
        fetchone=True,
    )

    if not row:
        return {
            "branch": branch,
            "category": category,
            "last": "尚未设置",
            "next": "尚未设置",
            "configured": False,
        }

    prefix = str(row.get("prefix") or "")
    current_number = int(row.get("current_number") or 0)
    width = int(row.get("number_width") or 6)

    last_no = (
        f"{prefix}{current_number:0{width}d}"
        if current_number > 0
        else "尚未使用"
    )

    return {
        "branch": branch,
        "category": category,
        "last": last_no,
        "next": f"{prefix}{current_number + 1:0{width}d}",
        "configured": True,
    }


SPECIAL_DONATION_TITLE = "813交流会财布施"

INCOME_CATEGORIES = [
    "月费",
    "财布施",
    "观音村",
    "膳食结缘",
    "观音堂纯檀香布施",
    SPECIAL_DONATION_TITLE,
    "临时特别布施"
]

FINANCE_STYLE = """
<style>
body{
    font-family: Microsoft JhengHei, Arial;
    font-size:22px;
    padding:25px;
}

h1{
    font-size:42px;
    margin-bottom:25px;
}

label{
    font-size:24px;
    font-weight:bold;
}

input,
select,
textarea{
    font-size:22px;
    padding:10px;
    min-height:50px;
    min-width:260px;
}

button,
input[type="submit"]{
    font-size:22px;
    padding:14px 28px;
    border-radius:10px;
    cursor:pointer;
}

a{
    font-size:22px;
}

table{
    font-size:20px;
    border-collapse:collapse;
    margin-top:20px;
}

th,td{
    padding:12px;
}

.card{
    padding:25px;
    border-radius:12px;
    box-shadow:0 2px 8px rgba(0,0,0,.15);
    margin-bottom:25px;
}

form{
    line-height:2.2;
}
</style>
"""

FINANCE_V5_STYLE = """
<style>
*{
    box-sizing:border-box;
}

body{
    margin:0;
    background:#f4f7fb;
    color:#1f2937;
    font-family:"Microsoft JhengHei","Microsoft YaHei",Arial,sans-serif;
}

.finance-v5{
    max-width:900px;
    margin:0 auto;
    padding:28px 20px 50px;
}

.v5-header{
    background:linear-gradient(135deg,#1769aa,#2387c9);
    color:#fff;
    padding:28px;
    border-radius:22px;
    margin-bottom:24px;
    box-shadow:0 8px 24px rgba(23,105,170,.20);
}

.v5-header h1{
    margin:0;
    font-size:36px;
}

.v5-header p{
    margin:8px 0 0;
    font-size:17px;
    opacity:.9;
}

.v5-topbar{
    display:flex;
    justify-content:space-between;
    align-items:center;
    gap:12px;
    margin-bottom:18px;
}

.v5-back,
.v5-logout{
    display:inline-flex;
    align-items:center;
    justify-content:center;
    min-height:44px;
    padding:10px 18px;
    border-radius:12px;
    text-decoration:none;
    font-weight:bold;
    font-size:17px;
}

.v5-back{
    color:#1769aa;
    background:#eaf4ff;
}

.v5-logout{
    color:#fff;
    background:#d94343;
}

.v5-menu-grid{
    display:grid;
    grid-template-columns:repeat(2,minmax(0,1fr));
    gap:16px;
}

.v5-menu-btn{
    display:flex;
    align-items:center;
    gap:16px;
    min-height:94px;
    padding:20px;
    background:#fff;
    border:1px solid #e5e7eb;
    border-radius:18px;
    text-decoration:none;
    color:#1f2937;
    box-shadow:0 4px 16px rgba(0,0,0,.06);
    transition:.18s ease;
}

.v5-menu-btn:hover{
    transform:translateY(-2px);
    box-shadow:0 8px 22px rgba(0,0,0,.10);
}

.v5-icon{
    width:54px;
    height:54px;
    flex:0 0 54px;
    display:flex;
    align-items:center;
    justify-content:center;
    border-radius:15px;
    font-size:30px;
    background:#eef6ff;
}

.v5-menu-text{
    min-width:0;
}

.v5-menu-title{
    font-size:22px;
    font-weight:800;
    margin-bottom:5px;
}

.v5-menu-desc{
    color:#667085;
    font-size:15px;
    line-height:1.45;
}

.v5-section-title{
    margin:26px 0 12px;
    font-size:18px;
    font-weight:800;
    color:#475467;
}

.v5-alert{
    background:#fff6e5;
    border-left:6px solid #f59e0b;
}

.v5-income{
    background:#edf9f1;
}

.v5-expense{
    background:#fff1f2;
}

.v5-member{
    background:#eef5ff;
}

.v5-report{
    background:#f7f2ff;
}

@media(max-width:680px){
    .finance-v5{
        padding:16px 12px 35px;
    }

    .v5-header{
        padding:22px 18px;
        border-radius:17px;
    }

    .v5-header h1{
        font-size:29px;
    }

    .v5-menu-grid{
        grid-template-columns:1fr;
        gap:12px;
    }

    .v5-menu-btn{
        min-height:82px;
        padding:16px;
    }

    .v5-menu-title{
        font-size:20px;
    }
}
</style>
"""

FINANCE_DATE_COMPONENT = """
<style>
.finance-date-control{
    display:grid;
    gap:9px;
    width:100%;
}

.finance-date-main{
    display:grid;
    grid-template-columns:52px minmax(0,1fr) 52px;
    gap:9px;
    align-items:stretch;
}

.finance-date-main .form-input,
.finance-date-main input[type="date"]{
    width:100%;
    min-width:0;
    margin:0;
    text-align:center;
    font-weight:800;
}

.finance-date-step,
.finance-date-quick{
    border:1px solid #cbd5e1;
    background:#f8fafc;
    color:#334155;
    border-radius:10px;
    cursor:pointer;
    font-weight:800;
    transition:.15s ease;
}

.finance-date-step{
    min-height:52px;
    font-size:25px;
}

.finance-date-step:hover,
.finance-date-quick:hover{
    background:#eaf4ff;
    border-color:#8fbce0;
    color:#1769aa;
}

.finance-date-step:active,
.finance-date-quick:active{
    transform:scale(.97);
}

.finance-date-shortcuts{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:8px;
}

.finance-date-quick{
    min-height:42px;
    padding:8px 10px;
    font-size:15px;
}

.finance-date-status{
    min-height:22px;
    font-size:14px;
    font-weight:700;
    line-height:1.45;
}

.finance-date-today{
    border-color:#86efac !important;
    background:#f0fdf4 !important;
    color:#166534 !important;
}

.finance-date-yesterday{
    border-color:#fde68a !important;
    background:#fffbeb !important;
    color:#92400e !important;
}

.finance-date-old{
    border-color:#fdba74 !important;
    background:#fff7ed !important;
    color:#9a3412 !important;
}

.finance-date-future{
    border-color:#fca5a5 !important;
    background:#fef2f2 !important;
    color:#b91c1c !important;
}

@media(max-width:560px){
    .finance-date-main{
        grid-template-columns:48px minmax(0,1fr) 48px;
    }

    .finance-date-shortcuts{
        grid-template-columns:1fr;
    }
}
</style>

<script>
(function(){
    function localISO(dateValue){
        const year = dateValue.getFullYear();
        const month = String(dateValue.getMonth() + 1).padStart(2, "0");
        const day = String(dateValue.getDate()).padStart(2, "0");
        return `${year}-${month}-${day}`;
    }

    function parseISO(value){
        if(!value){
            return null;
        }

        const parts = value.split("-").map(Number);

        if(parts.length !== 3 || parts.some(Number.isNaN)){
            return null;
        }

        return new Date(parts[0], parts[1] - 1, parts[2]);
    }

    function startOfDay(value){
        return new Date(
            value.getFullYear(),
            value.getMonth(),
            value.getDate()
        );
    }

    function changeDate(input, days){
        let current = parseISO(input.value);

        if(!current){
            current = new Date();
        }

        current.setDate(current.getDate() + days);
        input.value = localISO(current);
        input.dispatchEvent(new Event("change", {bubbles:true}));
    }

    function setRelativeDate(input, days){
        const selected = new Date();
        selected.setDate(selected.getDate() + days);
        input.value = localISO(selected);
        input.dispatchEvent(new Event("change", {bubbles:true}));
    }

    function openCalendar(input){
        if(typeof input.showPicker === "function"){
            try{
                input.showPicker();
                return;
            }catch(error){}
        }

        input.focus();
        input.click();
    }

    function updateStatus(input, status){
        input.classList.remove(
            "finance-date-today",
            "finance-date-yesterday",
            "finance-date-old",
            "finance-date-future"
        );

        const selected = parseISO(input.value);

        if(!selected){
            status.textContent = "";
            return;
        }

        const today = startOfDay(new Date());
        const chosen = startOfDay(selected);
        const difference = Math.round(
            (chosen - today) / 86400000
        );

        if(difference === 0){
            input.classList.add("finance-date-today");
            status.textContent = "🟢 今天";
        }else if(difference === -1){
            input.classList.add("finance-date-yesterday");
            status.textContent = "🟡 昨天";
        }else if(difference > 0){
            input.classList.add("finance-date-future");
            status.textContent = `🔴 未来日期（${difference} 天后），请检查`;
        }else if(difference < -30){
            input.classList.add("finance-date-old");
            status.textContent = `🟠 ${Math.abs(difference)} 天前，可能是旧账补录`;
        }else{
            status.textContent = `📅 ${Math.abs(difference)} 天前`;
        }
    }

    function enhanceDateInput(input){
        if(input.dataset.financeDateReady === "1"){
            return;
        }

        input.dataset.financeDateReady = "1";

        const control = document.createElement("div");
        control.className = "finance-date-control";

        const main = document.createElement("div");
        main.className = "finance-date-main";

        const previous = document.createElement("button");
        previous.type = "button";
        previous.className = "finance-date-step";
        previous.textContent = "◀";
        previous.title = "前一天";

        const next = document.createElement("button");
        next.type = "button";
        next.className = "finance-date-step";
        next.textContent = "▶";
        next.title = "后一天";

        const shortcuts = document.createElement("div");
        shortcuts.className = "finance-date-shortcuts";

        const yesterday = document.createElement("button");
        yesterday.type = "button";
        yesterday.className = "finance-date-quick";
        yesterday.textContent = "昨天";

        const today = document.createElement("button");
        today.type = "button";
        today.className = "finance-date-quick";
        today.textContent = "今天";

        const calendar = document.createElement("button");
        calendar.type = "button";
        calendar.className = "finance-date-quick";
        calendar.textContent = "📅 选日期";

        const status = document.createElement("div");
        status.className = "finance-date-status";

        const parent = input.parentNode;
        parent.insertBefore(control, input);

        main.appendChild(previous);
        main.appendChild(input);
        main.appendChild(next);

        shortcuts.appendChild(yesterday);
        shortcuts.appendChild(today);
        shortcuts.appendChild(calendar);

        control.appendChild(main);
        control.appendChild(shortcuts);
        control.appendChild(status);

        previous.addEventListener("click", function(){
            changeDate(input, -1);
        });

        next.addEventListener("click", function(){
            changeDate(input, 1);
        });

        yesterday.addEventListener("click", function(){
            setRelativeDate(input, -1);
        });

        today.addEventListener("click", function(){
            setRelativeDate(input, 0);
        });

        calendar.addEventListener("click", function(){
            openCalendar(input);
        });

        input.addEventListener("change", function(){
            updateStatus(input, status);
        });

        input.addEventListener("keydown", function(event){
            if(event.altKey || event.ctrlKey || event.metaKey){
                return;
            }

            if(event.key === "ArrowLeft"){
                event.preventDefault();
                changeDate(input, -1);
            }else if(event.key === "ArrowRight"){
                event.preventDefault();
                changeDate(input, 1);
            }
        });

        updateStatus(input, status);
    }

    function initFinanceDates(root){
        const selector = [
            'input[type="date"][name="receipt_date"]',
            'input[type="date"][name="record_date"]',
            'input[type="date"][name="payment_date"]',
            'input[type="date"][name="voucher_date"]',
            'input[type="date"][name="payment_voucher_date"]'
        ].join(",");

        (root || document).querySelectorAll(selector).forEach(
            enhanceDateInput
        );
    }

    document.addEventListener("DOMContentLoaded", function(){
        initFinanceDates(document);
    });

    window.initFinanceDates = initFinanceDates;
})();
</script>
"""


# Finance shared helpers are imported from finance_common.py.

def get_next_receipt_no_by_category(category):
    row = db_query("""
        select receipt_no
        from finance_records
        where category = %s
          and receipt_no is not null
          and receipt_no <> ''
        order by id desc
        limit 1
    """, (category,), fetchone=True)

    if not row or not row["receipt_no"]:
        return ""

    last_no = row["receipt_no"].strip()

    match = re.match(r"^([A-Za-z]+)(\d+)$", last_no)

    if not match:
        return ""

    prefix = match.group(1)
    number = match.group(2)

    next_number = str(int(number) + 1).zfill(len(number))

    return prefix + next_number


@finance_bp.route("/login", methods=["GET", "POST"])
def finance_login():

    error = ""

    if request.method == "POST":

        pin = request.form.get("pin", "").strip()

        if pin == FINANCE_PIN:
            session["finance_login"] = True
            return redirect(url_for("finance.finance_home"))

        error = "财政 PIN 不正确，请重新输入。"

    return render_template_string(FINANCE_DATE_COMPONENT + """
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">

    <meta
        name="viewport"
        content="width=device-width, initial-scale=1"
    >

    <title>财政系统登入</title>

    <link
        rel="stylesheet"
        href="{{ url_for('static', filename='css/toolbox.css') }}"
    >

    <style>
        .finance-login-page{
            max-width:680px;
        }

        .finance-login-card{
            margin-top:22px;
        }

        .finance-login-icon{
            font-size:52px;
            text-align:center;
            margin-bottom:10px;
        }

        .finance-login-title{
            text-align:center;
            margin-bottom:8px;
        }

        .finance-login-subtitle{
            text-align:center;
            color:#667085;
            margin-bottom:26px;
        }

        .finance-pin-input{
            text-align:center;
            font-size:28px;
            letter-spacing:8px;
        }

        .finance-login-actions{
            display:grid;
            gap:12px;
            margin-top:22px;
        }

        .finance-login-actions .btn-tool{
            width:100%;
        }

        @media(max-width:600px){
            .finance-login-page{
                padding:16px;
            }

            .finance-login-icon{
                font-size:44px;
            }
        }
    </style>
</head>

<body>

<div class="page finance-login-page">

    <div class="card finance-login-card">

        <div class="finance-login-icon">
            🏦
        </div>

        <h1 class="page-title finance-login-title">
            财政系统
        </h1>

        <p class="page-subtitle finance-login-subtitle">
            Finance Management V5
        </p>

        {% if error %}
            <div class="alert alert-danger">
                {{ error }}
            </div>
        {% endif %}

        <form method="post">

            <div class="form-group">

                <label class="form-label">
                    财政 PIN
                </label>

                <input
                    class="form-input finance-pin-input"
                    name="pin"
                    type="password"
                    inputmode="numeric"
                    autocomplete="new-password"
                    maxlength="8"
                    autofocus
                    required
                >

            </div>

            <div class="finance-login-actions">

                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    进入财政系统
                </button>

                <a
                    class="btn-tool btn-secondary"
                    href="/admin-home"
                >
                    返回管理员首页
                </a>

            </div>

        </form>

    </div>

</div>

</body>
</html>
""",
    error=error
    )

@finance_bp.route("/logout")
def finance_logout():

    session.pop("finance_login", None)
    session.pop("finance_admin", None)

    return redirect(url_for("finance.finance_login"))

@finance_bp.route("/late_members")
@finance_admin_required
def late_members():

    branch = request.args.get("branch", "CHE").strip().upper()

    if branch not in ("CHE", "STW"):
        branch = "CHE"

    rows = db_query("""
        select
            m.member_id,
            m.name,
            m.phone,
            m.remark,
            max(p.end_month) as paid_until,
            max(p.payment_date) as last_payment_date
        from members m
        left join member_payments p
            on p.member_id = m.member_id
           and coalesce(p.status, 'active') = 'active'
        where coalesce(m.member_status, m.status, '') not in (
            '停供',
            '停止',
            '永久停止',
            '往生',
            '已往生'
        )
          and upper(coalesce(m.member_id, '')) like %s
        group by
            m.member_id,
            m.name,
            m.phone,
            m.remark
        having
            max(p.end_month) is not null
            and max(p.end_month) < date_trunc(
                'month',
                current_date
            )
        order by
            paid_until asc,
            m.member_id
    """, (f"{branch}-%",), fetchall=True)

    today = date.today()

    current_month_index = (
        today.year * 12
        + today.month
    )

    green_count = 0
    yellow_count = 0
    red_count = 0
    total_amount = 0

    for r in rows:

        paid_until = r["paid_until"]

        paid_index = (
            paid_until.year * 12
            + paid_until.month
        )

        late_months = max(
            current_month_index - paid_index,
            0
        )

        reference_amount = late_months * 50

        r["late_months"] = late_months
        r["reference_amount"] = reference_amount

        if late_months <= 2:
            r["level"] = "green"
            r["level_text"] = "最近停止"
            green_count += 1

        elif late_months <= 6:
            r["level"] = "yellow"
            r["level_text"] = "一段时间未缴费"
            yellow_count += 1

        else:
            r["level"] = "red"
            r["level_text"] = "较久未缴费"
            red_count += 1

        total_amount += reference_amount

        phone = (
            r["phone"]
            or ""
        ).strip()

        phone_digits = "".join(
            ch
            for ch in phone
            if ch.isdigit()
        )

        if phone_digits.startswith("0"):
            phone_digits = "6" + phone_digits

        if phone_digits:
            r["wa_link"] = (
                "https://wa.me/"
                + phone_digits
            )
        else:
            r["wa_link"] = ""

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>

        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>{{ branch }} 月费关怀名单</title>

        <link
            rel="stylesheet"
            href="{{ url_for(
                'static',
                filename='css/toolbox.css'
            ) }}"
        >

        <style>

            .care-page {
                max-width: 1450px;
            }

            .care-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;

                box-shadow:
                    0 12px 30px
                    rgba(37, 99, 235, 0.18);
            }

            .care-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .care-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .care-note {
                background: #eff6ff;
                border: 1px solid #bfdbfe;
                color: #1e40af;
                border-radius: 14px;
                padding: 14px 16px;
                margin-bottom: 20px;
                line-height: 1.65;
            }

            .care-summary {
                display: grid;
                grid-template-columns:
                    repeat(5, minmax(0, 1fr));
                gap: 16px;
                margin-bottom: 20px;
            }

            .care-summary .summary-box {
                min-height: 120px;
                text-align: center;

                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
            }

            .summary-icon {
                font-size: 29px;
                margin-bottom: 6px;
            }

            .summary-label {
                color: #64748b;
                font-size: 15px;
                margin-bottom: 6px;
            }

            .summary-value {
                color: #0f172a;
                font-size: 26px;
                font-weight: 800;
            }

            .summary-green {
                background: #f0fdf4;
                border-color: #bbf7d0;
            }

            .summary-yellow {
                background: #fffbeb;
                border-color: #fde68a;
            }

            .summary-red {
                background: #fef2f2;
                border-color: #fecaca;
            }

            .summary-green .summary-value {
                color: #15803d;
            }

            .summary-yellow .summary-value {
                color: #a16207;
            }

            .summary-red .summary-value {
                color: #b91c1c;
            }

            .table-topbar {
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 12px;
                flex-wrap: wrap;
                margin-bottom: 14px;
            }

            .result-text {
                color: #64748b;
                font-size: 16px;
            }

            .care-table {
                min-width: 1280px;
            }

            .care-table td {
                vertical-align: middle;
            }

            .member-id {
                color: #1d4ed8;
                font-weight: 800;
                white-space: nowrap;
            }

            .member-name {
                color: #1e293b;
                font-weight: 700;
                min-width: 90px;
            }

            .phone-cell {
                white-space: nowrap;
            }

            .money-cell {
                color: #15803d;
                font-weight: 800;
                white-space: nowrap;
                text-align: right;
            }

            .date-cell {
                white-space: nowrap;
            }

            .remark-cell {
                min-width: 180px;
                max-width: 320px;
                white-space: normal;
                line-height: 1.5;
                overflow-wrap: anywhere;
            }

            .level-row-green {
                background: #f0fdf4;
            }

            .level-row-yellow {
                background: #fffbeb;
            }

            .level-row-red {
                background: #fef2f2;
            }

            .level-badge {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                padding: 7px 11px;
                border-radius: 999px;
                font-size: 14px;
                font-weight: 800;
                white-space: nowrap;
            }

            .level-green {
                background: #dcfce7;
                color: #166534;
            }

            .level-yellow {
                background: #fef3c7;
                color: #92400e;
            }

            .level-red {
                background: #fee2e2;
                color: #991b1b;
            }

            .wa-button {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                gap: 5px;
                padding: 8px 12px;
                border-radius: 10px;
                background: #dcfce7;
                color: #166534;
                font-weight: 800;
                text-decoration: none;
                white-space: nowrap;
                border: 1px solid #bbf7d0;
            }

            .wa-button:hover {
                background: #bbf7d0;
            }

            .empty-care {
                text-align: center;
                padding: 55px 20px;
                color: #64748b;
            }

            .empty-care-icon {
                font-size: 48px;
                margin-bottom: 10px;
            }

            .empty-care h3 {
                color: #334155;
                margin: 0 0 8px;
            }

            .bottom-actions {
                margin-top: 20px;
            }

            @media (max-width: 1100px) {

                .care-summary {
                    grid-template-columns:
                        repeat(2, minmax(0, 1fr));
                }

            }

            @media (max-width: 700px) {

                .care-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .care-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .care-header h1 {
                    font-size: 26px;
                }

                .care-summary {
                    grid-template-columns: 1fr;
                }

                .care-summary .summary-box {
                    min-height: 100px;
                }

                .bottom-actions .btn-tool {
                    width: 100%;
                }

            }

        </style>

    </head>

    <body>

    <div class="page care-page">

        <div class="care-header">

            <h1>🌿 {{ branch }} 月费关怀名单</h1>

            <p>
                协助负责人了解会员最近缴费情况，
                方便适时联络与关怀。
            </p>

        </div>

        <div class="branch-switch" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:18px;">
            <a class="btn-tool {{ 'btn-primary' if branch == 'CHE' else 'btn-secondary' }}"
               href="{{ url_for('finance.late_members', branch='CHE') }}">
                CHE 蕉赖
            </a>
            <a class="btn-tool {{ 'btn-primary' if branch == 'STW' else 'btn-secondary' }}"
               href="{{ url_for('finance.late_members', branch='STW') }}">
                STW 实兆远
            </a>
        </div>

        <div class="care-note">

            ℹ️ 此名单不是追缴名单。
            “参考金额”只按照每月 RM50 计算，
            方便负责人了解大概月份，
            不代表会员必须补缴该金额。

        </div>

        <div class="care-summary">

            <div class="summary-box">

                <div class="summary-icon">
                    👥
                </div>

                <div class="summary-label">
                    总人数
                </div>

                <div class="summary-value">
                    {{ rows|length }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    💰
                </div>

                <div class="summary-label">
                    参考金额
                </div>

                <div class="summary-value">
                    RM {{ "{:,.2f}".format(total_amount) }}
                </div>

            </div>

            <div class="summary-box summary-green">

                <div class="summary-icon">
                    🟢
                </div>

                <div class="summary-label">
                    最近停止
                </div>

                <div class="summary-value">
                    {{ green_count }}
                </div>

            </div>

            <div class="summary-box summary-yellow">

                <div class="summary-icon">
                    🟡
                </div>

                <div class="summary-label">
                    一段时间未缴费
                </div>

                <div class="summary-value">
                    {{ yellow_count }}
                </div>

            </div>

            <div class="summary-box summary-red">

                <div class="summary-icon">
                    🔴
                </div>

                <div class="summary-label">
                    较久未缴费
                </div>

                <div class="summary-value">
                    {{ red_count }}
                </div>

            </div>

        </div>

        <div class="card">

            <div class="table-topbar">

                <div
                    class="section-title"
                    style="margin-bottom:0;"
                >
                    📋 会员关怀资料
                </div>

                <div class="result-text">
                    共
                    <strong>{{ rows|length }}</strong>
                    位会员
                </div>

            </div>

            {% if rows %}

                <div class="table-responsive">

                    <table class="record-table care-table">

                        <thead>

                            <tr>
                                <th>会员编号</th>
                                <th>姓名</th>
                                <th>电话</th>
                                <th>WhatsApp</th>
                                <th>已缴至</th>
                                <th>间隔月份</th>
                                <th>参考金额</th>
                                <th>关怀状态</th>
                                <th>最后付款日期</th>
                                <th>备注</th>
                            </tr>

                        </thead>

                        <tbody>

                            {% for r in rows %}

                                <tr class="
                                    {% if r.level == 'green' %}
                                        level-row-green
                                    {% elif r.level == 'yellow' %}
                                        level-row-yellow
                                    {% else %}
                                        level-row-red
                                    {% endif %}
                                ">

                                    <td>
                                        <span class="member-id">
                                            {{ r.member_id }}
                                        </span>
                                    </td>

                                    <td>
                                        <span class="member-name">
                                            {{ r.name }}
                                        </span>
                                    </td>

                                    <td class="phone-cell">
                                        {{ r.phone or "-" }}
                                    </td>

                                    <td>

                                        {% if r.wa_link %}

                                            <a
                                                class="wa-button"
                                                href="{{ r.wa_link }}"
                                                target="_blank"
                                                rel="noopener noreferrer"
                                            >
                                                💬 打开
                                            </a>

                                        {% else %}

                                            -

                                        {% endif %}

                                    </td>

                                    <td class="date-cell">
                                        {{ r.paid_until.strftime("%Y-%m") }}
                                    </td>

                                    <td>
                                        {{ r.late_months }} 个月
                                    </td>

                                    <td class="money-cell">
                                        RM
                                        {{ "%.2f"|format(
                                            r.reference_amount
                                        ) }}
                                    </td>

                                    <td>

                                        <span class="
                                            level-badge
                                            {% if r.level == 'green' %}
                                                level-green
                                            {% elif r.level == 'yellow' %}
                                                level-yellow
                                            {% else %}
                                                level-red
                                            {% endif %}
                                        ">

                                            {% if r.level == "green" %}
                                                🟢
                                            {% elif r.level == "yellow" %}
                                                🟡
                                            {% else %}
                                                🔴
                                            {% endif %}

                                            {{ r.level_text }}

                                        </span>

                                    </td>

                                    <td class="date-cell">

                                        {% if r.last_payment_date %}
                                            {{ r.last_payment_date }}
                                        {% else %}
                                            -
                                        {% endif %}

                                    </td>

                                    <td class="remark-cell">
                                        {{ r.remark or "-" }}
                                    </td>

                                </tr>

                            {% endfor %}

                        </tbody>

                    </table>

                </div>

            {% else %}

                <div class="empty-care">

                    <div class="empty-care-icon">
                        🌿
                    </div>

                    <h3>目前没有需要关怀的会员</h3>

                    <p>
                        所有在册会员的月费记录都在当前月份内。
                    </p>

                </div>

            {% endif %}

        </div>

        <div class="bottom-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'finance.finance_admin_home'
                ) }}"
            >
                ← 返回负责人中心
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        green_count=green_count,
        yellow_count=yellow_count,
        red_count=red_count,
        total_amount=total_amount,
        branch=branch
    )
    
    
@finance_bp.route("/member/<member_id>/edit",
                  methods=["GET","POST"])
@finance_admin_required
def edit_member(member_id):

    member = db_query("""
        select *
        from members
        where member_id = %s
        limit 1
    """, (member_id,), fetchone=True)

    if not member:
        return "Member not found"

    if request.method == "POST":

        status = request.form.get("status")

        db_query("""
            update members
            set status = %s
            where member_id = %s
        """, (
            status,
            member_id
        ))

        return redirect(
            url_for("finance.member_management")
        )

    return render_template_string(FINANCE_STYLE + FINANCE_DATE_COMPONENT + """

    <h1>{{ member.member_id }}</h1>

    <p>姓名：{{ member.name }}</p>

    <p>电话：{{ member.phone }}</p>

    <form method="post">

        <p>

            状态：

            <select name="status">

                <option
                {% if member.status == '在册' %}
                selected
                {% endif %}
                >
                在册
                </option>

                <option
                {% if member.status == '停供' %}
                selected
                {% endif %}
                >
                停供
                </option>

                <option
                {% if member.status == '往生' %}
                selected
                {% endif %}
                >
                往生
                </option>

            </select>

        </p>

        <button type="submit">
            保存
        </button>

    </form>

    <p>
        <a href="{{ url_for('finance.member_management') }}">
            返回
        </a>
    </p>

    """, member=member)


@finance_bp.route("/")
def finance_home():

    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    today_text = date.today().strftime("%Y-%m-%d")

    return render_template_string(
        FINANCE_DATE_COMPONENT + """
<!doctype html>
<html lang="zh">
<head>

    <meta charset="utf-8">
    <meta name="viewport"
          content="width=device-width, initial-scale=1">

    <title>财政录入工作台</title>

    <link rel="manifest"
          href="/finance-manifest.json">

    <link rel="icon"
          href="/static/finance_icon.png?v=1">

    <link rel="stylesheet"
          href="{{ url_for('static', filename='css/toolbox.css') }}">

    <style>
        body{
            background:#f4f7fb;
        }

        .entry-page{
            max-width:940px;
        }

        .entry-topbar{
            display:flex;
            justify-content:space-between;
            align-items:center;
            gap:12px;
            margin-bottom:18px;
        }

        .entry-topbar .btn-tool{
            min-height:46px;
            padding:10px 17px;
            font-size:16px;
        }

        .entry-hero{
            padding:28px 30px;
            border-radius:24px;
            color:#fff;
            background:linear-gradient(135deg,#1769aa,#2387c9);
            box-shadow:0 12px 28px rgba(23,105,170,.20);
            margin-bottom:22px;
        }

        .entry-hero-row{
            display:flex;
            align-items:center;
            gap:18px;
        }

        .entry-hero-icon{
            width:72px;
            height:72px;
            border-radius:20px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:rgba(255,255,255,.16);
            border:1px solid rgba(255,255,255,.25);
            font-size:39px;
            flex:0 0 auto;
        }

        .entry-hero h1{
            margin:0;
            font-size:34px;
            line-height:1.2;
        }

        .entry-hero p{
            margin:8px 0 0;
            font-size:17px;
            opacity:.92;
            line-height:1.6;
        }

        .entry-date{
            display:inline-flex;
            margin-top:14px;
            padding:7px 12px;
            border-radius:999px;
            background:rgba(255,255,255,.16);
            font-size:15px;
            font-weight:700;
        }

        .entry-section-title{
            margin:23px 0 12px;
            color:#475467;
            font-size:18px;
            font-weight:800;
        }

        .entry-grid{
            display:grid;
            grid-template-columns:repeat(2,minmax(0,1fr));
            gap:15px;
        }

        .entry-action{
            display:flex;
            align-items:center;
            gap:16px;
            min-height:102px;
            padding:20px;
            border:1px solid #e3e8ef;
            border-radius:19px;
            background:#fff;
            color:#1f2937;
            text-decoration:none;
            box-shadow:0 5px 18px rgba(15,23,42,.06);
            transition:.18s ease;
        }

        .entry-action:hover{
            transform:translateY(-2px);
            box-shadow:0 10px 25px rgba(15,23,42,.10);
        }

        .entry-action-icon{
            width:58px;
            height:58px;
            border-radius:17px;
            display:flex;
            align-items:center;
            justify-content:center;
            flex:0 0 58px;
            font-size:31px;
            background:#eef6ff;
        }

        .entry-action.income .entry-action-icon{
            background:#eaf8ef;
        }

        .entry-action.expense .entry-action-icon{
            background:#fff0f1;
        }

        .entry-action.bank .entry-action-icon{
            background:#edf4ff;
        }

        .entry-action.records .entry-action-icon{
            background:#f6f0ff;
        }

        .entry-action-title{
            font-size:22px;
            font-weight:850;
            line-height:1.3;
        }

        .entry-action-desc{
            margin-top:5px;
            color:#667085;
            font-size:15px;
            line-height:1.45;
        }

        .entry-admin-card{
            margin-top:24px;
            padding:18px;
            border:1px dashed #b8c7d6;
            border-radius:17px;
            background:#f8fafc;
            text-align:center;
        }

        .entry-admin-card p{
            margin:0 0 12px;
            color:#667085;
            font-size:15px;
        }

        @media(max-width:680px){
            .entry-page{
                padding:14px 12px 34px;
            }

            .entry-hero{
                padding:23px 19px;
                border-radius:19px;
            }

            .entry-hero-icon{
                width:60px;
                height:60px;
                font-size:33px;
            }

            .entry-hero h1{
                font-size:27px;
            }

            .entry-grid{
                grid-template-columns:1fr;
                gap:12px;
            }

            .entry-action{
                min-height:88px;
                padding:16px;
            }

            .entry-action-title{
                font-size:20px;
            }

            .entry-topbar .btn-tool{
                padding:9px 12px;
                font-size:15px;
            }
        }
    </style>
</head>
<body>

<div class="page entry-page">

    <div class="entry-topbar">
        <a class="btn-tool btn-secondary" href="/admin-home">
            ← 管理员首页
        </a>

        <a class="btn-tool btn-danger"
           href="{{ url_for('finance.finance_logout') }}">
            退出
        </a>
    </div>

    <section class="entry-hero">
        <div class="entry-hero-row">
            <div class="entry-hero-icon">💰</div>
            <div>
                <h1>财政录入工作台</h1>
                <p>只需选择要输入的项目，然后依照收条或 PV 录入。</p>
                <div class="entry-date">📅 {{ today_text }}</div>
            </div>
        </div>
    </section>

    <div class="entry-section-title">今天要输入什么？</div>

    <div class="entry-grid">

        <a class="entry-action income"
           href="{{ url_for('finance.finance_income_menu') }}">
            <div class="entry-action-icon">💵</div>
            <div>
                <div class="entry-action-title">收入录入</div>
                <div class="entry-action-desc">
                    CHE／STW 月费、财布施、观音村、膳食结缘及其它收入
                </div>
            </div>
        </a>

        <a class="entry-action expense"
           href="{{ url_for('finance.finance_expense_menu') }}">
            <div class="entry-action-icon">🧾</div>
            <div>
                <div class="entry-action-title">支出录入</div>
                <div class="entry-action-desc">
                    根据纸本 PV 选择类别并输入支出资料
                </div>
            </div>
        </a>

        <a class="entry-action bank"
           href="{{ url_for('finance.bank_pending') }}">
            <div class="entry-action-icon">🏦</div>
            <div>
                <div class="entry-action-title">银行资料</div>
                <div class="entry-action-desc">
                    输入或处理银行过账待确认资料
                </div>
            </div>
        </a>

        <a class="entry-action records"
           href="{{ url_for('finance.finance_basic_records') }}">
            <div class="entry-action-icon">🔍</div>
            <div>
                <div class="entry-action-title">最近记录</div>
                <div class="entry-action-desc">
                    核对刚才录入的收条、PV、姓名及金额
                </div>
            </div>
        </a>

    </div>

    <div class="entry-admin-card">
        <p>Dashboard、月结、报表、完整记录及系统设置仅供负责人使用。</p>
        <a class="btn-tool btn-purple"
           href="{{ url_for('finance.finance_admin_login') }}">
            🔐 进入财政负责人中心
        </a>
    </div>

</div>

</body>
</html>
""",
        today_text=today_text,
    )
@finance_bp.route("/admin/login", methods=["GET", "POST"])
def finance_admin_login():

    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    error = ""
    next_url = request.values.get("next", "").strip()

    if request.method == "POST":

        password = request.form.get("password", "").strip()

        if not FINANCE_ADMIN_PASSWORD:

            error = (
                "尚未设置 FINANCE_ADMIN_PASSWORD，"
                "请先在 Render Environment 设置。"
            )

        elif hmac.compare_digest(password, FINANCE_ADMIN_PASSWORD):

            session["finance_admin"] = True

            return redirect(
                next_url
                or url_for("finance_v7.finance_v7_home")
            )

        else:
            error = "负责人密码不正确，请重新输入。"

    return render_template_string(
        FINANCE_V5_STYLE + """
        <style>

        .finance-admin-login-page{
            max-width:980px;
            margin:0 auto;
            padding:34px 20px 60px;
        }

        .finance-admin-login-page .admin-topbar{
            margin-bottom:22px;
        }

        .finance-admin-login-page .admin-back{
            display:inline-flex;
            align-items:center;
            gap:8px;
            min-height:50px;
            padding:0 22px;
            border-radius:16px;
            background:#eaf4ff;
            color:#1769aa;
            text-decoration:none;
            font-size:19px;
            font-weight:700;
            transition:all .18s ease;
        }

        .finance-admin-login-page .admin-back:hover{
            background:#dceeff;
            transform:translateY(-1px);
        }

        .finance-admin-card{
            overflow:hidden;
            border-radius:26px;
            background:#fff;
            box-shadow:
                0 18px 50px rgba(25, 83, 139, .14),
                0 3px 12px rgba(0, 0, 0, .05);
        }

        .finance-admin-hero{
            position:relative;
            padding:42px 48px 38px;
            color:#fff;
            background:
                radial-gradient(
                    circle at 86% 20%,
                    rgba(255,255,255,.18),
                    transparent 24%
                ),
                linear-gradient(
                    135deg,
                    #1268bd 0%,
                    #268fd0 100%
                );
        }

        .finance-admin-hero::after{
            content:"🔒";
            position:absolute;
            right:46px;
            top:30px;
            font-size:96px;
            opacity:.16;
            filter:grayscale(1) brightness(4);
        }

        .finance-admin-title-row{
            display:flex;
            align-items:center;
            gap:18px;
            position:relative;
            z-index:1;
        }

        .finance-admin-lock{
            width:76px;
            height:76px;
            display:flex;
            align-items:center;
            justify-content:center;
            flex:0 0 auto;
            border-radius:22px;
            background:rgba(255,255,255,.16);
            border:1px solid rgba(255,255,255,.24);
            font-size:42px;
            box-shadow:inset 0 1px 0 rgba(255,255,255,.25);
        }

        .finance-admin-title{
            margin:0;
            font-size:40px;
            line-height:1.2;
            font-weight:800;
            letter-spacing:1px;
        }

        .finance-admin-subtitle{
            margin:10px 0 0;
            font-size:20px;
            line-height:1.7;
            color:rgba(255,255,255,.9);
        }

        .finance-admin-features{
            position:relative;
            z-index:1;
            display:grid;
            grid-template-columns:repeat(4, 1fr);
            gap:12px;
            margin-top:28px;
            padding-top:22px;
            border-top:1px solid rgba(255,255,255,.25);
        }

        .finance-admin-feature{
            display:flex;
            align-items:center;
            justify-content:center;
            gap:8px;
            min-height:42px;
            font-size:17px;
            font-weight:700;
            color:rgba(255,255,255,.96);
        }

        .finance-admin-feature-icon{
            width:24px;
            height:24px;
            display:flex;
            align-items:center;
            justify-content:center;
            border:2px solid rgba(255,255,255,.9);
            border-radius:50%;
            font-size:13px;
        }

        .finance-admin-body{
            padding:42px 48px 46px;
        }

        .finance-admin-form{
            max-width:620px;
            margin:0 auto;
        }

        .finance-admin-label{
            display:flex;
            align-items:center;
            gap:10px;
            margin-bottom:12px;
            color:#24364a;
            font-size:21px;
            font-weight:800;
        }

        .finance-admin-input-wrap{
            position:relative;
        }

        .finance-admin-password{
            width:100%;
            height:66px;
            padding:0 62px 0 54px;
            border:2px solid #d7e1eb;
            border-radius:16px;
            background:#fff;
            color:#24364a;
            font-size:21px;
            outline:none;
            box-sizing:border-box;
            transition:
                border-color .18s ease,
                box-shadow .18s ease;
        }

        .finance-admin-password:focus{
            border-color:#2384ce;
            box-shadow:0 0 0 5px rgba(35,132,206,.12);
        }

        .finance-admin-password::placeholder{
            color:#9aa9b9;
        }

        .finance-admin-input-icon{
            position:absolute;
            left:20px;
            top:50%;
            transform:translateY(-50%);
            font-size:24px;
            pointer-events:none;
        }

        .finance-admin-eye{
            position:absolute;
            right:14px;
            top:50%;
            transform:translateY(-50%);
            width:42px;
            height:42px;
            border:0;
            border-radius:12px;
            background:transparent;
            cursor:pointer;
            font-size:22px;
            color:#60758a;
        }

        .finance-admin-eye:hover{
            background:#eef5fb;
        }

        .finance-admin-submit{
            width:100%;
            min-height:66px;
            margin-top:22px;
            border:0;
            border-radius:16px;
            background:linear-gradient(
                135deg,
                #176fc5,
                #258ed1
            );
            color:#fff;
            font-size:22px;
            font-weight:800;
            cursor:pointer;
            box-shadow:0 10px 24px rgba(26,114,190,.22);
            transition:
                transform .18s ease,
                box-shadow .18s ease;
        }

        .finance-admin-submit:hover{
            transform:translateY(-2px);
            box-shadow:0 14px 28px rgba(26,114,190,.28);
        }

        .finance-admin-error{
            max-width:620px;
            margin:0 auto 20px;
            padding:17px 20px;
            border:1px solid #f1c1c1;
            border-left:6px solid #d94343;
            border-radius:14px;
            background:#fff4f4;
            color:#a72c2c;
            font-size:18px;
            font-weight:700;
            line-height:1.6;
        }

        .finance-admin-security{
            max-width:620px;
            margin:24px auto 0;
            display:flex;
            gap:14px;
            padding:18px 20px;
            border:1px solid #d7e7f5;
            border-radius:15px;
            background:#f3f8fd;
        }

        .finance-admin-security-icon{
            flex:0 0 auto;
            width:34px;
            height:34px;
            display:flex;
            align-items:center;
            justify-content:center;
            border-radius:50%;
            background:#1976c9;
            color:#fff;
            font-weight:800;
        }

        .finance-admin-security-title{
            margin:0 0 4px;
            color:#1769aa;
            font-size:17px;
            font-weight:800;
        }

        .finance-admin-security-text{
            margin:0;
            color:#647588;
            font-size:15px;
            line-height:1.6;
        }

        @media (max-width:760px){

            .finance-admin-login-page{
                padding:20px 12px 40px;
            }

            .finance-admin-hero{
                padding:30px 24px;
            }

            .finance-admin-title{
                font-size:29px;
            }

            .finance-admin-subtitle{
                font-size:17px;
            }

            .finance-admin-lock{
                width:60px;
                height:60px;
                font-size:34px;
            }

            .finance-admin-hero::after{
                display:none;
            }

            .finance-admin-features{
                grid-template-columns:repeat(2, 1fr);
            }

            .finance-admin-feature{
                justify-content:flex-start;
                font-size:15px;
            }

            .finance-admin-body{
                padding:30px 22px 34px;
            }

            .finance-admin-password{
                height:60px;
                font-size:18px;
            }

            .finance-admin-submit{
                min-height:60px;
                font-size:20px;
            }

        }

        </style>

        <div class="finance-admin-login-page">

            <div class="admin-topbar">
                <a
                    class="admin-back"
                    href="{{ url_for('finance.finance_home') }}"
                >
                    ← 返回录入工作台
                </a>
            </div>

            <div class="finance-admin-card">

                <div class="finance-admin-hero">

                    <div class="finance-admin-title-row">

                        <div class="finance-admin-lock">
                            🔐
                        </div>

                        <div>
                            <h1 class="finance-admin-title">
                                财政负责人登入
                            </h1>

                            <p class="finance-admin-subtitle">
                                仅限财政负责人使用
                            </p>
                        </div>

                    </div>

                    <div class="finance-admin-features">

                        <div class="finance-admin-feature">
                            <span class="finance-admin-feature-icon">✓</span>
                            审核月结
                        </div>

                        <div class="finance-admin-feature">
                            <span class="finance-admin-feature-icon">✓</span>
                            收条簿管理
                        </div>

                        <div class="finance-admin-feature">
                            <span class="finance-admin-feature-icon">✓</span>
                            敏感资料
                        </div>

                        <div class="finance-admin-feature">
                            <span class="finance-admin-feature-icon">✓</span>
                            系统设置
                        </div>

                    </div>

                </div>

                <div class="finance-admin-body">

                    {% if error %}

                    <div class="finance-admin-error">
                        ⚠️ {{ error }}
                    </div>

                    {% endif %}

                    <form
                        method="post"
                        class="finance-admin-form"
                    >

                        <input
                            type="hidden"
                            name="next"
                            value="{{ next_url }}"
                        >

                        <label
                            class="finance-admin-label"
                            for="admin-password"
                        >
                            👤 负责人密码
                        </label>

                        <div class="finance-admin-input-wrap">

                            <span class="finance-admin-input-icon">
                                🔒
                            </span>

                            <input
                                id="admin-password"
                                class="finance-admin-password"
                                type="password"
                                name="password"
                                placeholder="请输入负责人密码"
                                autocomplete="current-password"
                                autofocus
                                required
                            >

                            <button
                                type="button"
                                class="finance-admin-eye"
                                onclick="toggleAdminPassword()"
                                aria-label="显示或隐藏密码"
                                title="显示或隐藏密码"
                            >
                                👁️
                            </button>

                        </div>

                        <button
                            class="finance-admin-submit"
                            type="submit"
                        >
                            🛡️ 进入负责人中心
                        </button>

                    </form>

                    <div class="finance-admin-security">

                        <div class="finance-admin-security-icon">
                            i
                        </div>

                        <div>
                            <p class="finance-admin-security-title">
                                安全提示
                            </p>

                            <p class="finance-admin-security-text">
                                请妥善保管负责人密码，
                                不要将密码透露给其他人员。
                            </p>
                        </div>

                    </div>

                </div>

            </div>

        </div>

        <script>

        function toggleAdminPassword(){

            const input =
                document.getElementById("admin-password");

            const button =
                document.querySelector(".finance-admin-eye");

            if(input.type === "password"){
                input.type = "text";
                button.textContent = "🙈";
            }else{
                input.type = "password";
                button.textContent = "👁️";
            }

            input.focus();
        }

        </script>
        """,
        error=error,
        next_url=next_url,
    )


@finance_bp.route("/admin/logout")
def finance_admin_logout():
    session.pop("finance_admin", None)
    return redirect(url_for("finance.finance_home"))


@finance_bp.route("/admin")
@finance_admin_required
def finance_admin_home():

    ensure_finance_bank_in_tables()

    today = date.today()
    current_ym = today.strftime("%Y-%m")
    branch = "CHE"

    # =========================================================
    # 1. CHE 银行与 Petty Cash 真实余额
    #    直接使用资金中心现有的同一套计算，避免两边数字不同
    # =========================================================

    balance = get_finance_balance_summary(branch=branch)

    bank_balance = Decimal(str(
        getattr(balance, "bank_balance", 0) or 0
    ))

    cash_in_hand = Decimal(str(
        getattr(balance, "cash_in_hand", 0) or 0
    ))


    # =========================================================
    # 2. CHE 现金月费等待 Bank In
    # =========================================================

    waiting_row = db_query("""
        select
            coalesce(sum(r.amount), 0) as total,
            count(*) as count
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and r.payment_method = '现金'
          and coalesce(r.status, 'confirmed') <> 'cancelled'
          and (
                r.member_id ilike 'CHE%%'
                or r.receipt_no ilike 'CHE%%'
          )
          and not exists (
              select 1
              from finance_bank_deposit_items i
              where i.finance_record_id = r.id
          )
    """, fetchone=True) or {
        "total": 0,
        "count": 0
    }

    waiting_bank_in = Decimal(str(
        waiting_row.get("total") or 0
    ))

    waiting_bank_in_count = int(
        waiting_row.get("count") or 0
    )


    # =========================================================
    # 3. 银行待确认
    # =========================================================

    pending_bank_row = db_query("""
        select count(*) as total
        from bank_pending_records
    """, fetchone=True) or {
        "total": 0
    }

    pending_bank_count = int(
        pending_bank_row.get("total") or 0
    )


    # =========================================================
    # 4. CHE 本月收入
    #
    # 这里只统计 CHE 日常户口收入。
    # 总会布施不会混进 CHE。
    # =========================================================

    income_row = db_query("""
        select
            coalesce(sum(amount), 0) as total,
            count(*) as count
        from finance_records
        where record_type = 'income'
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and (
                member_id ilike 'CHE%%'
                or receipt_no ilike 'CHE%%'
                or (
                    coalesce(member_id, '') = ''
                    and coalesce(receipt_no, '') = ''
                )
          )
    """, (current_ym,), fetchone=True) or {
        "total": 0,
        "count": 0
    }

    che_month_income = Decimal(str(
        income_row.get("total") or 0
    ))

    che_month_income_count = int(
        income_row.get("count") or 0
    )


    # =========================================================
    # 5. CHE 今日收入
    # =========================================================

    today_income_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'income'
          and record_date = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and (
                member_id ilike 'CHE%%'
                or receipt_no ilike 'CHE%%'
                or (
                    coalesce(member_id, '') = ''
                    and coalesce(receipt_no, '') = ''
                )
          )
    """, (today,), fetchone=True) or {
        "total": 0
    }

    che_today_income = Decimal(str(
        today_income_row.get("total") or 0
    ))


    # =========================================================
    # 6. CHE 本月支出
    # =========================================================

    expense_row = db_query("""
        select
            coalesce(sum(amount), 0) as total,
            count(*) as count
        from finance_records
        where record_type = 'expense'
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
    """, (current_ym,), fetchone=True) or {
        "total": 0,
        "count": 0
    }

    che_month_expense = Decimal(str(
        expense_row.get("total") or 0
    ))

    che_month_expense_count = int(
        expense_row.get("count") or 0
    )


    # =========================================================
    # 7. CHE 今日支出
    # =========================================================

    today_expense_row = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where record_type = 'expense'
          and record_date = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
    """, (today,), fetchone=True) or {
        "total": 0
    }

    che_today_expense = Decimal(str(
        today_expense_row.get("total") or 0
    ))


    # =========================================================
    # 8. 支出待补 PV
    # =========================================================

    pending_pv_row = db_query("""
        select count(*) as total
        from finance_records
        where record_type = 'expense'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and (
                payment_voucher_no is null
                or trim(payment_voucher_no) = ''
          )
    """, fetchone=True) or {
        "total": 0
    }

    pending_pv_count = int(
        pending_pv_row.get("total") or 0
    )


    # =========================================================
    # 9. CHE 本月月费
    # =========================================================

    che_monthly_row = db_query("""
        select
            coalesce(sum(amount), 0) as total,
            count(*) as record_count,
            count(distinct member_id) as member_count
        from finance_records
        where record_type = 'income'
          and category = '月费'
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and (
                member_id ilike 'CHE%%'
                or receipt_no ilike 'CHE%%'
          )
    """, (current_ym,), fetchone=True) or {
        "total": 0,
        "record_count": 0,
        "member_count": 0
    }

    che_monthly_total = Decimal(str(
        che_monthly_row.get("total") or 0
    ))

    che_monthly_record_count = int(
        che_monthly_row.get("record_count") or 0
    )

    che_monthly_member_count = int(
        che_monthly_row.get("member_count") or 0
    )


    # =========================================================
    # 10. STW 本月月费
    # =========================================================

    stw_row = db_query("""
        select
            coalesce(sum(amount), 0) as total,
            count(*) as record_count
        from finance_records
        where record_type = 'income'
          and category = '月费'
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and (
                member_id ilike 'STW%%'
                or receipt_no ilike 'STW%%'
          )
    """, (current_ym,), fetchone=True) or {
        "total": 0,
        "record_count": 0
    }

    stw_month_income = Decimal(str(
        stw_row.get("total") or 0
    ))

    stw_month_count = int(
        stw_row.get("record_count") or 0
    )


    # =========================================================
    # 11. 总会本月布施
    # =========================================================

    hq_row = db_query("""
        select
            coalesce(sum(amount), 0) as total,
            count(*) as record_count
        from finance_records
        where record_type = 'income'
          and to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '总会户口'
    """, (current_ym,), fetchone=True) or {
        "total": 0,
        "record_count": 0
    }

    hq_month_income = Decimal(str(
        hq_row.get("total") or 0
    ))

    hq_month_count = int(
        hq_row.get("record_count") or 0
    )


    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">

    <meta name="viewport"
          content="width=device-width, initial-scale=1">

    <title>CHE 财政工作台</title>

    <link rel="stylesheet"
          href="{{ url_for('static', filename='css/toolbox.css') }}">

    <style>

        body{
            background:#f4f7fb;
        }

        .finance-admin-page{
            max-width:1160px;
        }

        /* =====================================================
           顶部
        ===================================================== */

        .finance-topbar{
            display:flex;
            justify-content:space-between;
            align-items:center;
            gap:12px;
            margin-bottom:18px;
        }

        .finance-hero{
            position:relative;
            overflow:hidden;
            padding:30px;
            border-radius:26px;
            color:#fff;
            background:linear-gradient(
                135deg,
                #2851a3 0%,
                #4d55b4 52%,
                #7150bd 100%
            );
            box-shadow:0 14px 34px rgba(58,67,157,.22);
        }

        .finance-hero::after{
            content:"";
            position:absolute;
            right:-70px;
            top:-95px;
            width:230px;
            height:230px;
            border-radius:50%;
            background:rgba(255,255,255,.10);
        }

        .hero-small{
            position:relative;
            z-index:1;
            font-size:15px;
            font-weight:850;
            opacity:.88;
        }

        .finance-hero h1{
            position:relative;
            z-index:1;
            margin:8px 0 0;
            font-size:37px;
            font-weight:950;
        }

        .finance-hero p{
            position:relative;
            z-index:1;
            margin:10px 0 0;
            font-size:17px;
            line-height:1.6;
            opacity:.92;
        }

        .hero-meta{
            position:relative;
            z-index:1;
            display:flex;
            flex-wrap:wrap;
            gap:9px;
            margin-top:17px;
        }

        .hero-chip{
            padding:8px 13px;
            border-radius:999px;
            background:rgba(255,255,255,.15);
            font-size:13px;
            font-weight:850;
        }

        /* =====================================================
           Section
        ===================================================== */

        .finance-section{
            margin-top:28px;
        }

        .section-heading{
            display:flex;
            justify-content:space-between;
            align-items:flex-end;
            gap:12px;
            margin-bottom:14px;
        }

        .section-heading-title{
            color:#26344c;
            font-size:23px;
            font-weight:950;
        }

        .section-heading-note{
            margin-top:3px;
            color:#8290a3;
            font-size:14px;
        }

        /* =====================================================
           六张主要工作卡
        ===================================================== */

        .main-work-grid{
            display:grid;
            grid-template-columns:repeat(2,minmax(0,1fr));
            gap:17px;
        }

        .main-work-card{
            position:relative;
            display:block;
            min-height:238px;
            padding:23px;
            overflow:hidden;
            border:1px solid #e1e7ef;
            border-radius:23px;
            background:#fff;
            color:#172033;
            text-decoration:none;
            box-shadow:0 7px 22px rgba(15,23,42,.065);
            transition:
                transform .18s ease,
                box-shadow .18s ease;
        }

        .main-work-card:hover{
            transform:translateY(-3px);
            box-shadow:0 13px 30px rgba(15,23,42,.11);
        }

        .card-top{
            display:flex;
            align-items:center;
            gap:14px;
        }

        .card-icon{
            width:58px;
            height:58px;
            flex:0 0 58px;
            display:flex;
            align-items:center;
            justify-content:center;
            border-radius:18px;
            background:#edf3ff;
            font-size:31px;
        }

        .card-name{
            color:#1d2a40;
            font-size:23px;
            font-weight:950;
        }

        .card-description{
            margin-top:3px;
            color:#778398;
            font-size:13px;
            line-height:1.4;
        }

        .card-balance{
            margin-top:21px;
            color:#101828;
            font-size:34px;
            font-weight:950;
            letter-spacing:-.8px;
        }

        .card-balance-label{
            margin-top:4px;
            color:#667085;
            font-size:14px;
            font-weight:750;
        }

        .card-stats{
            display:grid;
            grid-template-columns:repeat(2,minmax(0,1fr));
            gap:10px;
            margin-top:18px;
        }

        .card-stat{
            padding:12px 13px;
            border-radius:15px;
            background:#f6f8fb;
        }

        .card-stat.warning{
            background:#fff6e8;
        }

        .card-stat.danger{
            background:#fff0f0;
        }

        .card-stat.success{
            background:#ecf8f1;
        }

        .stat-label{
            color:#7a8699;
            font-size:12px;
            font-weight:800;
        }

        .stat-value{
            margin-top:5px;
            color:#27364d;
            font-size:18px;
            font-weight:950;
        }

        .card-enter{
            display:flex;
            justify-content:space-between;
            align-items:center;
            margin-top:18px;
            padding-top:15px;
            border-top:1px solid #edf0f4;
            color:#5066b4;
            font-size:14px;
            font-weight:900;
        }

        /* 各卡不同色 */

        .bank-card{
            border-top:5px solid #4772c5;
        }

        .cash-card{
            border-top:5px solid #35a276;
        }

        .income-card{
            border-top:5px solid #4f8bdb;
        }

        .expense-card{
            border-top:5px solid #dc7a60;
        }

        .monthly-card{
            border-top:5px solid #b16ca6;
        }

        .close-card{
            border-top:5px solid #db9a38;
        }

        .cash-card .card-icon{
            background:#eaf8f1;
        }

        .expense-card .card-icon{
            background:#fff0ea;
        }

        .monthly-card .card-icon{
            background:#faeef8;
        }

        .close-card .card-icon{
            background:#fff4df;
        }

        /* =====================================================
           STW / HQ
        ===================================================== */

        .other-account-grid{
            display:grid;
            grid-template-columns:repeat(2,minmax(0,1fr));
            gap:16px;
        }

        .account-card{
            display:block;
            padding:22px;
            border:1px solid #e2e7ef;
            border-radius:21px;
            background:#fff;
            color:#172033;
            text-decoration:none;
            box-shadow:0 6px 19px rgba(15,23,42,.055);
            transition:.18s ease;
        }

        .account-card:hover{
            transform:translateY(-2px);
            box-shadow:0 10px 26px rgba(15,23,42,.09);
        }

        .account-card.stw{
            border-left:6px solid #6579c8;
        }

        .account-card.hq{
            border-left:6px solid #b96993;
        }

        .account-title{
            font-size:22px;
            font-weight:950;
        }

        .account-desc{
            margin-top:5px;
            color:#758195;
            font-size:13px;
        }

        .account-number{
            margin-top:18px;
            color:#172033;
            font-size:28px;
            font-weight:950;
        }

        .account-meta{
            display:flex;
            flex-wrap:wrap;
            gap:8px;
            margin-top:12px;
        }

        .account-chip{
            padding:7px 10px;
            border-radius:999px;
            background:#f2f4f7;
            color:#536174;
            font-size:12px;
            font-weight:850;
        }

        .account-enter{
            margin-top:17px;
            color:#5368b4;
            font-size:14px;
            font-weight:900;
        }

        /* =====================================================
           辅助功能
        ===================================================== */

        .utility-grid{
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:13px;
        }

        .utility-card{
            display:flex;
            align-items:center;
            gap:12px;
            min-height:84px;
            padding:16px;
            border:1px solid #e2e7ef;
            border-radius:17px;
            background:#fff;
            color:#27364d;
            text-decoration:none;
            box-shadow:0 4px 14px rgba(15,23,42,.05);
        }

        .utility-icon{
            width:45px;
            height:45px;
            flex:0 0 45px;
            display:flex;
            align-items:center;
            justify-content:center;
            border-radius:14px;
            background:#f0f3f8;
            font-size:24px;
        }

        .utility-title{
            font-size:16px;
            font-weight:900;
        }

        .utility-desc{
            margin-top:3px;
            color:#8a95a5;
            font-size:11px;
            line-height:1.35;
        }

        /* =====================================================
           Mobile
        ===================================================== */

        @media(max-width:950px){

            .utility-grid{
                grid-template-columns:repeat(2,minmax(0,1fr));
            }
        }

        @media(max-width:720px){

            .finance-admin-page{
                padding:14px 12px 38px;
            }

            .finance-topbar{
                flex-direction:column;
                align-items:stretch;
            }

            .finance-topbar .btn-tool{
                width:100%;
            }

            .finance-hero{
                padding:25px 21px;
                border-radius:21px;
            }

            .finance-hero h1{
                font-size:29px;
            }

            .main-work-grid,
            .other-account-grid,
            .utility-grid{
                grid-template-columns:1fr;
            }

            .main-work-card{
                min-height:auto;
            }

            .card-balance{
                font-size:30px;
            }

            .section-heading{
                flex-direction:column;
                align-items:flex-start;
                gap:4px;
            }
        }

    </style>
</head>

<body>

<div class="page finance-admin-page">

    <div class="finance-topbar">

        <a class="btn-tool btn-secondary"
           href="{{ url_for('finance.finance_home') }}">
            ← 返回录入工作台
        </a>

        <a class="btn-tool btn-danger"
           href="{{ url_for('finance.finance_admin_logout') }}">
            退出负责人
        </a>

    </div>


    <!-- =====================================================
         HERO
    ====================================================== -->

    <section class="finance-hero">

        <div class="hero-small">
            财政负责人专用
        </div>

        <h1>🏠 CHE 财政工作台</h1>

        <p>
            每项工作只有一个入口。
            银行、现金、收入、支出、月费与月结各自独立处理。
        </p>

        <div class="hero-meta">

            <div class="hero-chip">
                📅 {{ today.strftime("%d/%m/%Y") }}
            </div>

            {% if pending_bank_count > 0
                  or waiting_bank_in_count > 0
                  or pending_pv_count > 0 %}

                <div class="hero-chip">
                    ⚠️ 目前有待处理工作
                </div>

            {% else %}

                <div class="hero-chip">
                    ✅ 暂无紧急待处理
                </div>

            {% endif %}

        </div>

    </section>


    <!-- =====================================================
         六张主要工作卡
    ====================================================== -->

    <section class="finance-section">

        <div class="section-heading">

            <div>
                <div class="section-heading-title">
                    CHE 日常工作
                </div>

                <div class="section-heading-note">
                    数字与操作放在同一张卡，不再重复显示
                </div>
            </div>

        </div>


        <div class="main-work-grid">


            <!-- 银行 -->

            <a class="main-work-card bank-card"
               href="{{ url_for('finance.finance_treasury') }}">

                <div class="card-top">

                    <div class="card-icon">
                        🏦
                    </div>

                    <div>
                        <div class="card-name">
                            银行
                        </div>

                        <div class="card-description">
                            Bank In、银行确认、提款与流水
                        </div>
                    </div>

                </div>

                <div class="card-balance">
                    RM {{ '{:,.2f}'.format(bank_balance) }}
                </div>

                <div class="card-balance-label">
                    当前 CHE 银行余额
                </div>

                <div class="card-stats">

                    <div class="card-stat
                        {{ 'warning' if waiting_bank_in_count > 0 else 'success' }}">

                        <div class="stat-label">
                            等待 Bank In
                        </div>

                        <div class="stat-value">
                            RM {{ '{:,.2f}'.format(waiting_bank_in) }}
                        </div>

                    </div>

                    <div class="card-stat
                        {{ 'warning' if pending_bank_count > 0 else 'success' }}">

                        <div class="stat-label">
                            银行待确认
                        </div>

                        <div class="stat-value">
                            {{ pending_bank_count }} 笔
                        </div>

                    </div>

                </div>

                <div class="card-enter">
                    <span>进入银行与现金工作台</span>
                    <span>→</span>
                </div>

            </a>


            <!-- Petty Cash -->

            <a class="main-work-card cash-card"
               href="{{ url_for('finance.finance_cash_in_hand') }}">

                <div class="card-top">

                    <div class="card-icon">
                        💵
                    </div>

                    <div>
                        <div class="card-name">
                            Petty Cash
                        </div>

                        <div class="card-description">
                            Cash In Hand、提款转入与现金流水
                        </div>
                    </div>

                </div>

                <div class="card-balance">
                    RM {{ '{:,.2f}'.format(cash_in_hand) }}
                </div>

                <div class="card-balance-label">
                    当前现金在手
                </div>

                <div class="card-stats">

                    <div class="card-stat">

                        <div class="stat-label">
                            本月现金支出
                        </div>

                        <div class="stat-value">
                            RM {{ '{:,.2f}'.format(che_month_expense) }}
                        </div>

                    </div>

                    <div class="card-stat">

                        <div class="stat-label">
                            今日现金支出
                        </div>

                        <div class="stat-value">
                            RM {{ '{:,.2f}'.format(che_today_expense) }}
                        </div>

                    </div>

                </div>

                <div class="card-enter">
                    <span>进入 Cash Ledger</span>
                    <span>→</span>
                </div>

            </a>


            <!-- 收入 -->

            <a class="main-work-card income-card"
               href="{{ url_for('finance.records', record_type='income') }}">

                <div class="card-top">

                    <div class="card-icon">
                        📥
                    </div>

                    <div>
                        <div class="card-name">
                            CHE 收入
                        </div>

                        <div class="card-description">
                            月费、收条、现金收入与收入查询
                        </div>
                    </div>

                </div>

                <div class="card-balance">
                    RM {{ '{:,.2f}'.format(che_month_income) }}
                </div>

                <div class="card-balance-label">
                    本月 CHE 日常户口收入
                </div>

                <div class="card-stats">

                    <div class="card-stat">

                        <div class="stat-label">
                            今日收入
                        </div>

                        <div class="stat-value">
                            RM {{ '{:,.2f}'.format(che_today_income) }}
                        </div>

                    </div>

                    <div class="card-stat">

                        <div class="stat-label">
                            本月记录
                        </div>

                        <div class="stat-value">
                            {{ che_month_income_count }} 笔
                        </div>

                    </div>

                </div>

                <div class="card-enter">
                    <span>进入收入工作台</span>
                    <span>→</span>
                </div>

            </a>


            <!-- 支出 -->

            <a class="main-work-card expense-card"
               href="{{ url_for('finance.records', record_type='expense') }}">

                <div class="card-top">

                    <div class="card-icon">
                        📤
                    </div>

                    <div>
                        <div class="card-name">
                            CHE 支出
                        </div>

                        <div class="card-description">
                            支出记录、PV、修改与作废
                        </div>
                    </div>

                </div>

                <div class="card-balance">
                    RM {{ '{:,.2f}'.format(che_month_expense) }}
                </div>

                <div class="card-balance-label">
                    本月 CHE 支出
                </div>

                <div class="card-stats">

                    <div class="card-stat
                        {{ 'danger' if pending_pv_count > 0 else 'success' }}">

                        <div class="stat-label">
                            待补 PV
                        </div>

                        <div class="stat-value">
                            {{ pending_pv_count }} 笔
                        </div>

                    </div>

                    <div class="card-stat">

                        <div class="stat-label">
                            今日支出
                        </div>

                        <div class="stat-value">
                            RM {{ '{:,.2f}'.format(che_today_expense) }}
                        </div>

                    </div>

                </div>

                <div class="card-enter">
                    <span>进入支出工作台</span>
                    <span>→</span>
                </div>

            </a>


            <!-- 月费 -->

            <a class="main-work-card monthly-card"
               href="{{ url_for('finance.late_members', branch='CHE') }}">

                <div class="card-top">

                    <div class="card-icon">
                        ❤️
                    </div>

                    <div>
                        <div class="card-name">
                            CHE 月费
                        </div>

                        <div class="card-description">
                            月费缴付、会员记录及关怀名单
                        </div>
                    </div>

                </div>

                <div class="card-balance">
                    RM {{ '{:,.2f}'.format(che_monthly_total) }}
                </div>

                <div class="card-balance-label">
                    本月收到的 CHE 月费
                </div>

                <div class="card-stats">

                    <div class="card-stat">

                        <div class="stat-label">
                            本月缴付会员
                        </div>

                        <div class="stat-value">
                            {{ che_monthly_member_count }} 人
                        </div>

                    </div>

                    <div class="card-stat">

                        <div class="stat-label">
                            本月月费记录
                        </div>

                        <div class="stat-value">
                            {{ che_monthly_record_count }} 笔
                        </div>

                    </div>

                </div>

                <div class="card-enter">
                    <span>进入月费工作台</span>
                    <span>→</span>
                </div>

            </a>


            <!-- 月结 -->

            <a class="main-work-card close-card"
               href="{{ url_for('finance_month_end.month_close') }}">

                <div class="card-top">

                    <div class="card-icon">
                        📒
                    </div>

                    <div>
                        <div class="card-name">
                            财政月结
                        </div>

                        <div class="card-description">
                            核对现金、银行、收条与 PV
                        </div>
                    </div>

                </div>

                <div class="card-balance">
                    {{ today.strftime("%Y年%m月") }}
                </div>

                <div class="card-balance-label">
                    当前处理月份
                </div>

                <div class="card-stats">

                    <div class="card-stat">

                        <div class="stat-label">
                            银行待确认
                        </div>

                        <div class="stat-value">
                            {{ pending_bank_count }} 笔
                        </div>

                    </div>

                    <div class="card-stat">

                        <div class="stat-label">
                            待补 PV
                        </div>

                        <div class="stat-value">
                            {{ pending_pv_count }} 笔
                        </div>

                    </div>

                </div>

                <div class="card-enter">
                    <span>开始本月月结</span>
                    <span>→</span>
                </div>

            </a>

        </div>

    </section>


    <!-- =====================================================
         STW 与总会
    ====================================================== -->

    <section class="finance-section">

        <div class="section-heading">

            <div>
                <div class="section-heading-title">
                    其他户口
                </div>

                <div class="section-heading-note">
                    STW 与总会独立显示，不混入 CHE
                </div>
            </div>

        </div>

        <div class="other-account-grid">


            <a class="account-card stw"
               href="{{ url_for('finance.late_members', branch='STW') }}">

                <div class="account-title">
                    🌏 STW 月费
                </div>

                <div class="account-desc">
                    只管理 STW 月费及相关存款记录
                </div>

                <div class="account-number">
                    RM {{ '{:,.2f}'.format(stw_month_income) }}
                </div>

                <div class="account-meta">

                    <span class="account-chip">
                        本月 {{ stw_month_count }} 笔
                    </span>

                    <span class="account-chip">
                        独立户口
                    </span>

                </div>

                <div class="account-enter">
                    进入 STW 工作区 →
                </div>

            </a>


            <a class="account-card hq"
               href="{{ url_for(
                    'finance.records',
                    fund_account='总会户口'
               ) }}">

                <div class="account-title">
                    🌸 总会布施
                </div>

                <div class="account-desc">
                    财布施、观音村及膳食结缘等总会款项
                </div>

                <div class="account-number">
                    RM {{ '{:,.2f}'.format(hq_month_income) }}
                </div>

                <div class="account-meta">

                    <span class="account-chip">
                        本月 {{ hq_month_count }} 笔
                    </span>

                    <span class="account-chip">
                        总会户口
                    </span>

                </div>

                <div class="account-enter">
                    进入总会工作区 →
                </div>

            </a>

        </div>

    </section>


    <!-- =====================================================
         辅助入口
    ====================================================== -->

    <section class="finance-section">

        <div class="section-heading">

            <div>
                <div class="section-heading-title">
                    查询与工具
                </div>

                <div class="section-heading-note">
                    不常用功能集中放在最后
                </div>
            </div>

        </div>

        <div class="utility-grid">


            <a class="utility-card"
               href="{{ url_for('finance.records') }}">

                <div class="utility-icon">
                    🔎
                </div>

                <div>
                    <div class="utility-title">
                        财务查询
                    </div>

                    <div class="utility-desc">
                        搜索收条、PV、姓名及历史记录
                    </div>
                </div>

            </a>


            <a class="utility-card"
               href="{{ url_for('finance.dashboard') }}">

                <div class="utility-icon">
                    📊
                </div>

                <div>
                    <div class="utility-title">
                        财政报表
                    </div>

                    <div class="utility-desc">
                        收入、支出及月度统计
                    </div>
                </div>

            </a>


            <a class="utility-card"
               href="/finance/reports/excel">

                <div class="utility-icon">
                    📥
                </div>

                <div>
                    <div class="utility-title">
                        Excel 下载
                    </div>

                    <div class="utility-desc">
                        下载月费、布施与支出报表
                    </div>
                </div>

            </a>


            <a class="utility-card"
               href="{{ url_for(
                    'finance.receipt_range_summary'
               ) }}">

                <div class="utility-icon">
                    📒
                </div>

                <div>
                    <div class="utility-title">
                        收条检查
                    </div>

                    <div class="utility-desc">
                        检查跳号、作废及金额
                    </div>
                </div>

            </a>


            <a class="utility-card"
               href="{{ url_for(
                    'finance.finance_member_menu'
               ) }}">

                <div class="utility-icon">
                    👥
                </div>

                <div>
                    <div class="utility-title">
                        会员资料
                    </div>

                    <div class="utility-desc">
                        会员状态及付款历史
                    </div>
                </div>

            </a>


            <a class="utility-card"
               href="{{ url_for(
                    'finance.donor_management'
               ) }}">

                <div class="utility-icon">
                    📇
                </div>

                <div>
                    <div class="utility-title">
                        布施人资料
                    </div>

                    <div class="utility-desc">
                        维护姓名及电话号码
                    </div>
                </div>

            </a>


            <a class="utility-card"
               href="{{ url_for(
                    'finance.receipt_book_management'
               ) }}">

                <div class="utility-icon">
                    📕
                </div>

                <div>
                    <div class="utility-title">
                        收条簿
                    </div>

                    <div class="utility-desc">
                        管理当前及下一张收条
                    </div>
                </div>

            </a>


            <a class="utility-card"
               href="{{ url_for(
                    'finance.finance_vendors'
               ) }}">

                <div class="utility-icon">
                    ⚙️
                </div>

                <div>
                    <div class="utility-title">
                        财政设置
                    </div>

                    <div class="utility-desc">
                        付款对象及历史导入
                    </div>
                </div>

            </a>

        </div>

    </section>

</div>

</body>
</html>
""",
        today=today,

        bank_balance=bank_balance,
        cash_in_hand=cash_in_hand,

        waiting_bank_in=waiting_bank_in,
        waiting_bank_in_count=waiting_bank_in_count,
        pending_bank_count=pending_bank_count,
        pending_pv_count=pending_pv_count,

        che_month_income=che_month_income,
        che_month_income_count=che_month_income_count,
        che_today_income=che_today_income,

        che_month_expense=che_month_expense,
        che_month_expense_count=che_month_expense_count,
        che_today_expense=che_today_expense,

        che_monthly_total=che_monthly_total,
        che_monthly_record_count=che_monthly_record_count,
        che_monthly_member_count=che_monthly_member_count,

        stw_month_income=stw_month_income,
        stw_month_count=stw_month_count,

        hq_month_income=hq_month_income,
        hq_month_count=hq_month_count,
    )

@finance_bp.route("/records/basic")
def finance_basic_records():
    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    q = request.args.get("q", "").strip()
    record_type = request.args.get("type", "").strip()
    year = request.args.get("year", "").strip()
    month = request.args.get("month", "").strip()
    category = request.args.get("category", "").strip()
    payment_method = request.args.get("method", "").strip()
    status = request.args.get("status", "").strip()

    clauses = ["1=1"]
    params = []

    if q:
        keyword = f"%{q}%"
        clauses.append("""
            (
                coalesce(receipt_no, '') ilike %s
                or coalesce(payment_voucher_no, '') ilike %s
                or coalesce(name, '') ilike %s
                or coalesce(vendor_name, '') ilike %s
                or coalesce(category, '') ilike %s
                or coalesce(remarks, '') ilike %s
            )
        """)
        params.extend([keyword] * 6)

    if record_type in ("income", "expense"):
        clauses.append("record_type = %s")
        params.append(record_type)

    if year.isdigit() and len(year) == 4:
        clauses.append("extract(year from record_date)::int = %s")
        params.append(int(year))

    if month.isdigit() and 1 <= int(month) <= 12:
        clauses.append("extract(month from record_date)::int = %s")
        params.append(int(month))

    if category:
        clauses.append("category = %s")
        params.append(category)

    if payment_method:
        clauses.append("payment_method = %s")
        params.append(payment_method)

    if status:
        clauses.append("coalesce(status, 'confirmed') = %s")
        params.append(status)

    where_sql = " and ".join(clauses)

    rows = db_query(f"""
        select
            id,
            record_date,
            record_type,
            receipt_no,
            payment_voucher_no,
            name,
            vendor_name,
            category,
            amount,
            payment_method,
            status,
            remarks
        from finance_records
        where {where_sql}
        order by record_date desc, id desc
        limit 200
    """, tuple(params), fetchall=True)

    summary = db_query(f"""
        select
            count(*) as record_count,
            coalesce(sum(case when record_type = 'income' and coalesce(status, 'confirmed') <> 'cancelled' then amount else 0 end), 0) as income_total,
            coalesce(sum(case when record_type = 'expense' and coalesce(status, 'confirmed') <> 'cancelled' then amount else 0 end), 0) as expense_total,
            count(*) filter (where coalesce(status, 'confirmed') = 'cancelled') as cancelled_count
        from finance_records
        where {where_sql}
    """, tuple(params), fetchone=True) or {}

    categories = db_query("""
        select distinct category
        from finance_records
        where category is not null and trim(category) <> ''
        order by category
    """, fetchall=True)

    methods = db_query("""
        select distinct payment_method
        from finance_records
        where payment_method is not null and trim(payment_method) <> ''
        order by payment_method
    """, fetchall=True)

    years = db_query("""
        select distinct extract(year from record_date)::int as year
        from finance_records
        where record_date is not null
        order by year desc
    """, fetchall=True)

    return render_template_string(FINANCE_V5_STYLE + r"""
    <style>
        .ledger-page{max-width:1220px;margin:0 auto;padding-bottom:40px;}
        .ledger-hero{background:linear-gradient(135deg,#1769aa,#2e9ad4);border-radius:26px;padding:30px;color:white;box-shadow:0 14px 34px rgba(23,105,170,.18);}
        .ledger-hero h1{margin:0 0 8px;font-size:34px;}
        .ledger-hero p{margin:0;opacity:.92;font-size:17px;}
        .ledger-summary{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin:18px 0;}
        .ledger-stat{background:white;border-radius:18px;padding:18px;box-shadow:0 8px 24px rgba(15,23,42,.07);}
        .ledger-stat-label{font-size:14px;color:#64748b;margin-bottom:6px;}
        .ledger-stat-value{font-size:25px;font-weight:800;color:#0f172a;}
        .ledger-stat.income .ledger-stat-value{color:#15803d;}
        .ledger-stat.expense .ledger-stat-value{color:#b91c1c;}
        .ledger-stat.cancelled .ledger-stat-value{color:#475569;}
        .filter-card{background:white;border-radius:20px;padding:18px;margin-bottom:18px;box-shadow:0 8px 24px rgba(15,23,42,.06);}
        .quick-row{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px;}
        .quick-btn{border:1px solid #dbe4ee;background:#f8fafc;color:#334155;border-radius:999px;padding:9px 15px;text-decoration:none;font-weight:700;}
        .quick-btn.active{background:#1769aa;color:white;border-color:#1769aa;}
        .filter-grid{display:grid;grid-template-columns:2fr repeat(5,minmax(120px,1fr));gap:10px;align-items:end;}
        .filter-grid label{display:block;font-size:13px;color:#64748b;margin:0 0 5px 4px;font-weight:700;}
        .filter-actions{display:flex;gap:12px;margin-top:16px;justify-content:flex-end;align-items:center;}
        .filter-actions .btn-tool{min-width:150px;min-height:52px;border-radius:14px;font-size:17px;font-weight:800;display:inline-flex;align-items:center;justify-content:center;text-decoration:none;}
        .filter-actions .btn-primary{min-width:190px;background:linear-gradient(135deg,#1769aa,#238fcb);border:none;box-shadow:0 10px 22px rgba(23,105,170,.24);}
        .filter-actions .btn-primary:hover{transform:translateY(-1px);box-shadow:0 13px 26px rgba(23,105,170,.30);}
        .ledger-list{display:grid;gap:10px;}
        .ledger-item{background:white;border-radius:18px;padding:15px 18px;display:grid;grid-template-columns:minmax(210px,1.2fr) minmax(220px,1.5fr) minmax(120px,.7fr) 150px;gap:18px;align-items:center;box-shadow:0 7px 20px rgba(15,23,42,.06);border-left:7px solid #94a3b8;transition:.18s ease;}
        .ledger-item:hover{transform:translateY(-1px);box-shadow:0 10px 25px rgba(15,23,42,.09);}
        .ledger-item.income{border-left-color:#22a447;}
        .ledger-item.expense{border-left-color:#dc4c4c;}
        .ledger-item.cancelled{opacity:.7;border-left-color:#64748b;}
        .record-no{font-size:18px;font-weight:850;color:#0f172a;word-break:break-word;margin-bottom:4px;}
        .record-party{font-size:20px;font-weight:850;color:#172033;margin-bottom:4px;}
        .record-sub{font-size:14px;color:#64748b;line-height:1.45;}
        .record-meta{display:flex;flex-wrap:wrap;gap:7px;margin-top:6px;}
        .meta-chip{display:inline-flex;align-items:center;padding:4px 8px;border-radius:999px;background:#f1f5f9;color:#475569;font-size:12px;font-weight:750;}
        .record-amount{font-size:24px;font-weight:900;text-align:right;white-space:nowrap;}
        .income .record-amount{color:#15803d;}.expense .record-amount{color:#b91c1c;}
        .status-pill{display:inline-block;padding:5px 10px;border-radius:999px;font-size:12px;font-weight:800;background:#e2e8f0;color:#475569;margin-top:6px;}
        .status-pill.confirmed{background:#dcfce7;color:#166534;}.status-pill.cancelled{background:#e2e8f0;color:#475569;}
        .empty-box{background:white;border-radius:20px;padding:44px;text-align:center;color:#64748b;}
        @media(max-width:900px){
            .ledger-summary{grid-template-columns:repeat(2,1fr)}
            .filter-grid{grid-template-columns:1fr 1fr}
            .ledger-item{grid-template-columns:1fr 1fr;gap:12px}
            .record-amount{text-align:left}
        }
        @media(max-width:600px){
            .ledger-summary{grid-template-columns:1fr 1fr;gap:9px}
            .ledger-stat{padding:14px}.ledger-stat-value{font-size:19px}
            .filter-grid{grid-template-columns:1fr}
            .filter-actions{display:grid;grid-template-columns:1fr;margin-top:14px}
            .filter-actions .btn-tool,.filter-actions .btn-primary{width:100%;min-width:0}
            .ledger-item{grid-template-columns:1fr;padding:15px}
            .record-party{font-size:19px}.record-amount{font-size:22px}
            .ledger-hero{padding:24px 20px}.ledger-hero h1{font-size:28px}
        }
    </style>

    <div class="finance-v5 ledger-page">
        <div class="v5-topbar">
            <a class="v5-back" href="{{ url_for('finance.finance_home') }}">← 返回录入工作台</a>
        </div>

        <div class="ledger-hero">
            <h1>📒 财政总账</h1>
            <p>供 Key In 组员核对最近录入资料；只读查看，不可修改或作废</p>
        </div>

        <div class="ledger-summary">
            <div class="ledger-stat"><div class="ledger-stat-label">符合条件记录</div><div class="ledger-stat-value">{{ summary.record_count or 0 }} 笔</div></div>
            <div class="ledger-stat income"><div class="ledger-stat-label">收入合计</div><div class="ledger-stat-value">RM {{ '%.2f'|format(summary.income_total or 0) }}</div></div>
            <div class="ledger-stat expense"><div class="ledger-stat-label">支出合计</div><div class="ledger-stat-value">RM {{ '%.2f'|format(summary.expense_total or 0) }}</div></div>
            <div class="ledger-stat cancelled"><div class="ledger-stat-label">作废记录</div><div class="ledger-stat-value">{{ summary.cancelled_count or 0 }} 笔</div></div>
        </div>

        <div class="filter-card">
            <div class="quick-row">
                <a class="quick-btn {{ 'active' if not record_type else '' }}" href="{{ url_for('finance.finance_basic_records', q=q, year=year, month=month, category=category, method=payment_method, status=status) }}">全部</a>
                <a class="quick-btn {{ 'active' if record_type == 'income' else '' }}" href="{{ url_for('finance.finance_basic_records', q=q, type='income', year=year, month=month, category=category, method=payment_method, status=status) }}">🟢 收入</a>
                <a class="quick-btn {{ 'active' if record_type == 'expense' else '' }}" href="{{ url_for('finance.finance_basic_records', q=q, type='expense', year=year, month=month, category=category, method=payment_method, status=status) }}">🔴 支出</a>
                <a class="quick-btn {{ 'active' if status == 'cancelled' else '' }}" href="{{ url_for('finance.finance_basic_records', q=q, type=record_type, year=year, month=month, category=category, method=payment_method, status='cancelled') }}">⚫ 作废</a>
            </div>

            <form method="get">
                <div class="filter-grid">
                    <div><label>关键字</label><input class="form-input" name="q" value="{{ q }}" placeholder="收条、PV、姓名、公司或类别"></div>
                    <div><label>类型</label><select class="form-input" name="type"><option value="">全部</option><option value="income" {{ 'selected' if record_type == 'income' else '' }}>收入</option><option value="expense" {{ 'selected' if record_type == 'expense' else '' }}>支出</option></select></div>
                    <div><label>年份</label><select class="form-input" name="year"><option value="">全部</option>{% for y in years %}<option value="{{ y.year }}" {{ 'selected' if year == (y.year|string) else '' }}>{{ y.year }}</option>{% endfor %}</select></div>
                    <div><label>月份</label><select class="form-input" name="month"><option value="">全部</option>{% for m in range(1,13) %}<option value="{{ m }}" {{ 'selected' if month == (m|string) else '' }}>{{ '%02d'|format(m) }} 月</option>{% endfor %}</select></div>
                    <div><label>类别</label><select class="form-input" name="category"><option value="">全部</option>{% for c in categories %}<option value="{{ c.category }}" {{ 'selected' if category == c.category else '' }}>{{ c.category }}</option>{% endfor %}</select></div>
                    <div><label>付款方式</label><select class="form-input" name="method"><option value="">全部</option>{% for m in methods %}<option value="{{ m.payment_method }}" {{ 'selected' if payment_method == m.payment_method else '' }}>{{ m.payment_method }}</option>{% endfor %}</select></div>
                </div>
                <div class="filter-actions">
                    <a class="btn-tool" href="{{ url_for('finance.finance_basic_records') }}">↺ 清除筛选</a>
                    <button class="btn-tool btn-primary" type="submit">🔎 查询总账</button>
                </div>
            </form>
        </div>

        <div class="ledger-list">
            {% for r in rows %}
                {% set display_status = r.status or 'confirmed' %}
                <div class="ledger-item {{ r.record_type }} {{ 'cancelled' if display_status == 'cancelled' else '' }}">
                    <div>
                        <div class="record-no">{{ r.receipt_no or r.payment_voucher_no or '未填写编号' }}</div>
                        <div class="record-sub">{{ r.record_date }}</div>
                    </div>
                    <div>
                        <div class="record-party">
                            {% if r.record_type == 'expense' %}
                                {{ r.vendor_name or r.name or '未填写付款对象' }}
                            {% else %}
                                {{ r.name or '未填写姓名' }}
                            {% endif %}
                        </div>
                        <div class="record-meta">
                            <span class="meta-chip">{{ '收入' if r.record_type == 'income' else '支出' }}</span>
                            <span class="meta-chip">{{ r.category or '-' }}</span>
                            <span class="meta-chip">{{ r.payment_method or '-' }}</span>
                        </div>
                    </div>
                    <div>
                        <span class="status-pill {{ display_status }}">{{ '已确认' if display_status == 'confirmed' else ('已作废' if display_status == 'cancelled' else display_status) }}</span>
                    </div>
                    <div class="record-amount">{{ '+' if r.record_type == 'income' else '-' }} RM {{ '%.2f'|format(r.amount or 0) }}</div>
                </div>
            {% else %}
                <div class="empty-box">没有找到符合条件的财政记录</div>
            {% endfor %}
        </div>
    </div>
    """,
        rows=rows,
        q=q,
        record_type=record_type,
        year=year,
        month=month,
        category=category,
        payment_method=payment_method,
        status=status,
        summary=summary,
        categories=categories,
        methods=methods,
        years=years,
    )


@finance_bp.route("/menu/income")
def finance_income_menu():
    items = [
        ("C", "CHE 月费", "CHE", "月费", url_for("finance.monthly_fee_batch", branch="CHE")),
        ("S", "STW 月费", "STW", "月费", url_for("finance.monthly_fee_batch", branch="STW")),
        ("🙏", "财布施", "CHE", "财布施", url_for("finance.income_batch", category="财布施")),
        ("🏡", "观音村", "CHE", "观音村", url_for("finance.income_batch", category="观音村")),
        ("🥗", "膳食结缘", "CHE", "膳食结缘", url_for("finance.income_batch", category="膳食结缘")),
        ("🪵", "观音堂纯檀香布施", "CHE", "观音堂纯檀香布施", url_for("finance.income_batch", category="观音堂纯檀香布施")),
        ("🎊", SPECIAL_DONATION_TITLE, "CHE", SPECIAL_DONATION_TITLE, url_for("finance.income_batch", category=SPECIAL_DONATION_TITLE)),
        ("➕", "临时特别布施", "CHE", "临时特别布施", url_for("finance.income_batch", category="临时特别布施")),
    ]

    menu_items = []
    for icon, title, branch, category, href in items:
        menu_items.append({
            "icon": icon,
            "title": title,
            "href": href,
            "status": get_receipt_book_status(branch, category),
        })

    return render_template_string(FINANCE_V5_STYLE + """
    <style>
        .receipt-mini{margin-top:9px;padding-top:9px;border-top:1px dashed #cbd5e1;display:grid;gap:3px;font-size:14px;}
        .receipt-mini strong{color:#1769aa;font-size:15px;}
    </style>
    <div class="finance-v5">
        <div class="v5-topbar">
            <a class="v5-back"
            href="{{ url_for('finance.finance_home') }}">
                ← 返回财政首页
            </a>
        </div>
        <div class="v5-header"><h1>💵 收入录入</h1><p>请选择收入项目；每个项目会显示最后及下一张收条</p></div>
        <div class="v5-menu-grid">
            {% for item in menu_items %}
            <a class="v5-menu-btn v5-income" href="{{ item.href }}">
                <div class="v5-icon">{{ item.icon }}</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">{{ item.title }}</div>
                    <div class="receipt-mini">
                        <span>上次最后一张：<strong>{{ item.status.last }}</strong></span>
                        <span>下一张建议：<strong>{{ item.status.next }}</strong></span>
                    </div>
                </div>
            </a>
            {% endfor %}
        </div>
    </div>
    """, menu_items=menu_items)


@finance_bp.route("/admin/receipt-books")
@finance_admin_required
def receipt_book_management():
    rows = db_query("""
        select branch, category, prefix, current_number, number_width
        from finance_receipt_books
        order by branch, category
    """, fetchall=True)

    for row in rows:
        width = int(row.get("number_width") or 6)
        current = int(row.get("current_number") or 0)
        prefix = str(row.get("prefix") or "")
        row["last_no"] = f"{prefix}{current:0{width}d}" if current else "尚未使用"
        row["next_no"] = f"{prefix}{current + 1:0{width}d}"

    return render_template_string(FINANCE_V5_STYLE + """
    <div class="finance-v5" style="max-width:1100px;">
        <div class="v5-topbar"><a class="v5-back" href="{{ url_for('finance.finance_admin_home') }}">← 负责人中心</a></div>
        <div class="v5-header"><h1>📕 收条簿管理</h1><p>查看每个分会及收入项目目前使用到的号码</p></div>
        <div class="card" style="overflow:auto;">
            <table class="record-table" style="width:100%;min-width:760px;">
                <thead><tr><th>分会</th><th>收入项目</th><th>前缀</th><th>上次最后一张</th><th>下一张建议</th></tr></thead>
                <tbody>{% for r in rows %}<tr><td>{{ r.branch }}</td><td>{{ r.category }}</td><td>{{ r.prefix }}</td><td><strong>{{ r.last_no }}</strong></td><td><strong style="color:#1769aa;">{{ r.next_no }}</strong></td></tr>{% else %}<tr><td colspan="5">尚未建立 finance_receipt_books 设置</td></tr>{% endfor %}</tbody>
            </table>
        </div>
    </div>
    """, rows=rows)


@finance_bp.route(
    "/admin/receipt-summary",
    methods=["GET", "POST"]
)
@finance_admin_required
def receipt_range_summary():

    branch = request.values.get(
        "branch",
        "CHE"
    ).strip().upper()

    category = request.values.get(
        "category",
        "月费"
    ).strip()

    receipt_from_input = request.values.get(
        "receipt_from",
        ""
    ).strip().upper()

    receipt_to_input = request.values.get(
        "receipt_to",
        ""
    ).strip().upper()

    receipt_from = ""
    receipt_to = ""

    result = None
    records = []
    error = ""

    if branch not in ["CHE", "STW"]:
        branch = "CHE"

    if request.method == "POST":

        def parse_receipt_number(raw_value, selected_branch):

            raw_value = str(raw_value or "").strip().upper()

            if not raw_value:
                return None

            # 允许只输入数字，例如 1555
            if raw_value.isdigit():
                return int(raw_value)

            # 也兼容完整编号，例如 CHE0001555
            match = re.match(
                r"^([A-Z]+)[\s\-]?(\d+)$",
                raw_value
            )

            if not match:
                return None

            prefix = match.group(1)
            number = int(match.group(2))

            if prefix != selected_branch:
                return "wrong_branch"

            return number


        start_result = parse_receipt_number(
            receipt_from_input,
            branch
        )

        end_result = parse_receipt_number(
            receipt_to_input,
            branch
        )

        if start_result == "wrong_branch" or end_result == "wrong_branch":

            error = (
                f"目前选择的是 {branch} 分会，"
                f"请输入 {branch} 收条号码。"
            )

        elif start_result is None or end_result is None:

            error = (
                "请输入收条号码。"
                "可以只输入数字，例如 1555。"
            )

        else:

            start_no = start_result
            end_no = end_result

            receipt_width = 7

            receipt_from = (
                f"{branch}{start_no:0{receipt_width}d}"
            )

            receipt_to = (
                f"{branch}{end_no:0{receipt_width}d}"
            )

            if start_no > end_no:

                error = (
                    "开始号码不可大于结束号码。"
                )

            elif end_no - start_no > 1000:

                error = (
                    "一次最多检查 1001 张收条，"
                    "请缩小范围。"
                )

            else:

                records = db_query(
                    """
                    select
                        id,
                        receipt_no,
                        record_date,
                        receipt_date,
                        name,
                        category,
                        amount,
                        payment_method,
                        status
                    from finance_records
                    where record_type = 'income'
                      and receipt_no is not null
                      and receipt_no <> ''
                      and receipt_no >= %s
                      and receipt_no <= %s
                      and receipt_no ilike %s
                      and category = %s
                    order by receipt_no, id
                    """,
                    (
                        receipt_from,
                        receipt_to,
                        f"{branch}%",
                        category,
                    ),
                    fetchall=True,
                )

                by_number = {}

                for row in records:

                    receipt_value = str(
                        row.get("receipt_no") or ""
                    ).strip().upper()

                    match = re.match(
                        r"^([A-Z]+)(\d+)$",
                        receipt_value
                    )

                    if (
                        match
                        and match.group(1) == branch
                    ):
                        by_number[
                            int(match.group(2))
                        ] = row

                valid = [
                    row
                    for row in records
                    if str(
                        row.get("status")
                        or "confirmed"
                    ) != "cancelled"
                ]

                cancelled = [
                    row
                    for row in records
                    if str(
                        row.get("status")
                        or "confirmed"
                    ) == "cancelled"
                ]

                cash_total = sum(
                    float(row.get("amount") or 0)
                    for row in valid
                    if str(
                        row.get("payment_method")
                        or ""
                    ).strip() == "现金"
                )

                non_cash_total = sum(
                    float(row.get("amount") or 0)
                    for row in valid
                    if str(
                        row.get("payment_method")
                        or ""
                    ).strip() != "现金"
                )

                all_total = sum(
                    float(row.get("amount") or 0)
                    for row in valid
                )

                missing_numbers = [
                    number
                    for number in range(
                        start_no,
                        end_no + 1
                    )
                    if number not in by_number
                ]

                width = receipt_width
                prefix = branch

                result = {
                    "span_count":
                        end_no - start_no + 1,

                    "used_count":
                        len(records),

                    "valid_count":
                        len(valid),

                    "cancelled_count":
                        len(cancelled),

                    "missing_count":
                        len(missing_numbers),

                    "missing_text":
                        ", ".join(
                            f"{prefix}{number:0{width}d}"
                            for number
                            in missing_numbers[:30]
                        ),

                    "cash_total":
                        cash_total,

                    "non_cash_total":
                        non_cash_total,

                    "all_total":
                        all_total,
                }

    categories = INCOME_CATEGORIES

    return render_template_string(
        FINANCE_V5_STYLE + """
        <style>

        .receipt-summary-page{
            max-width:1280px;
            margin:0 auto;
            padding:30px 20px 70px;
        }

        .receipt-summary-page .receipt-topbar{
            margin-bottom:22px;
        }

        .receipt-summary-page .receipt-back{
            display:inline-flex;
            align-items:center;
            justify-content:center;
            gap:8px;
            min-height:50px;
            padding:0 22px;
            border-radius:16px;
            background:#eaf4ff;
            color:#1769aa;
            text-decoration:none;
            font-size:18px;
            font-weight:800;
            transition:
                transform .18s ease,
                background .18s ease;
        }

        .receipt-summary-page .receipt-back:hover{
            background:#dceeff;
            transform:translateY(-1px);
        }

        .receipt-hero{
            position:relative;
            overflow:hidden;
            padding:38px 42px;
            border-radius:26px;
            color:#fff;
            background:
                radial-gradient(
                    circle at 88% 15%,
                    rgba(255,255,255,.18),
                    transparent 25%
                ),
                linear-gradient(
                    135deg,
                    #166ebf 0%,
                    #2693d2 100%
                );
            box-shadow:
                0 16px 38px rgba(27,105,170,.18);
        }

        .receipt-hero::after{
            content:"📒";
            position:absolute;
            right:48px;
            top:19px;
            font-size:120px;
            opacity:.14;
            transform:rotate(-7deg);
        }

        .receipt-hero-title{
            position:relative;
            z-index:1;
            display:flex;
            align-items:center;
            gap:18px;
        }

        .receipt-hero-icon{
            width:74px;
            height:74px;
            display:flex;
            align-items:center;
            justify-content:center;
            flex:0 0 auto;
            border-radius:22px;
            background:rgba(255,255,255,.17);
            border:1px solid rgba(255,255,255,.25);
            font-size:40px;
        }

        .receipt-hero h1{
            margin:0;
            font-size:38px;
            line-height:1.25;
            font-weight:850;
            letter-spacing:.5px;
        }

        .receipt-hero p{
            margin:9px 0 0;
            font-size:19px;
            line-height:1.6;
            color:rgba(255,255,255,.92);
        }

        .receipt-form-card{
            margin-top:24px;
            padding:30px;
            border:1px solid #e1eaf2;
            border-radius:22px;
            background:#fff;
            box-shadow:
                0 10px 30px rgba(30,75,115,.08);
        }

        .receipt-card-title{
            display:flex;
            align-items:center;
            gap:10px;
            margin:0 0 22px;
            color:#21384d;
            font-size:23px;
            font-weight:850;
        }

        .receipt-form-grid{
            display:grid;
            grid-template-columns:
                repeat(2, minmax(0,1fr));
            gap:20px;
        }

        .receipt-field{
            min-width:0;
        }

        .receipt-field label{
            display:block;
            margin-bottom:9px;
            color:#34495e;
            font-size:18px;
            font-weight:800;
        }

        .receipt-field .form-input{
            width:100%;
            height:58px;
            padding:0 16px;
            border:2px solid #d9e3ec;
            border-radius:14px;
            background:#fff;
            color:#24384c;
            font-size:18px;
            box-sizing:border-box;
            outline:none;
            transition:
                border-color .18s ease,
                box-shadow .18s ease;
        }

        .receipt-field .form-input:focus{
            border-color:#2485cb;
            box-shadow:
                0 0 0 5px rgba(36,133,203,.11);
        }

        .receipt-field-help{
            margin-top:7px;
            color:#7b8b9b;
            font-size:14px;
            line-height:1.5;
        }

        .receipt-submit{
            width:100%;
            min-height:62px;
            margin-top:24px;
            border:0;
            border-radius:15px;
            background:
                linear-gradient(
                    135deg,
                    #176fc5,
                    #258fd2
                );
            color:#fff;
            font-size:21px;
            font-weight:850;
            cursor:pointer;
            box-shadow:
                0 10px 25px rgba(27,112,188,.22);
            transition:
                transform .18s ease,
                box-shadow .18s ease;
        }

        .receipt-submit:hover{
            transform:translateY(-2px);
            box-shadow:
                0 14px 30px rgba(27,112,188,.28);
        }

        .receipt-error{
            display:flex;
            gap:13px;
            margin-top:22px;
            padding:18px 20px;
            border:1px solid #efc0c0;
            border-left:6px solid #d94343;
            border-radius:15px;
            background:#fff4f4;
            color:#a82e2e;
            font-size:17px;
            font-weight:750;
            line-height:1.6;
        }

        .receipt-summary-grid{
            display:grid;
            grid-template-columns:
                repeat(3, minmax(0,1fr));
            gap:18px;
            margin-top:24px;
        }

        .receipt-summary-box{
            position:relative;
            overflow:hidden;
            min-height:130px;
            padding:24px;
            border:1px solid #e1e9f0;
            border-radius:20px;
            background:#fff;
            box-shadow:
                0 8px 24px rgba(35,72,105,.07);
        }

        .receipt-summary-box::after{
            content:"";
            position:absolute;
            left:0;
            top:0;
            bottom:0;
            width:6px;
            background:#2589ce;
        }

        .receipt-summary-box.warning::after{
            background:#e0a320;
        }

        .receipt-summary-box.danger::after{
            background:#d9534f;
        }

        .receipt-summary-box.success::after{
            background:#2aa56f;
        }

        .receipt-summary-box.money::after{
            background:#7357c8;
        }

        .receipt-summary-icon{
            margin-bottom:12px;
            font-size:25px;
        }

        .receipt-summary-value{
            display:block;
            color:#183b59;
            font-size:31px;
            font-weight:900;
            line-height:1.2;
            word-break:break-word;
        }

        .receipt-summary-label{
            display:block;
            margin-top:8px;
            color:#748596;
            font-size:16px;
            font-weight:700;
        }

        .receipt-missing-card{
            margin-top:22px;
            padding:22px 24px;
            border:1px solid #f0d49b;
            border-radius:18px;
            background:#fffaf0;
        }

        .receipt-missing-title{
            display:flex;
            align-items:center;
            gap:9px;
            margin-bottom:10px;
            color:#945f00;
            font-size:19px;
            font-weight:850;
        }

        .receipt-missing-list{
            color:#755524;
            font-size:16px;
            line-height:1.8;
            overflow-wrap:anywhere;
        }

        .receipt-table-card{
            margin-top:24px;
            overflow:hidden;
            border:1px solid #e1e8ef;
            border-radius:20px;
            background:#fff;
            box-shadow:
                0 8px 24px rgba(35,72,105,.07);
        }

        .receipt-table-header{
            display:flex;
            align-items:center;
            justify-content:space-between;
            gap:15px;
            padding:20px 24px;
            border-bottom:1px solid #e5edf4;
            background:#f7fafc;
        }

        .receipt-table-title{
            margin:0;
            color:#21384d;
            font-size:21px;
            font-weight:850;
        }

        .receipt-table-count{
            padding:7px 12px;
            border-radius:999px;
            background:#e6f2fc;
            color:#1769aa;
            font-size:14px;
            font-weight:800;
        }

        .receipt-table-wrap{
            overflow-x:auto;
        }

        .receipt-record-table{
            width:100%;
            min-width:930px;
            border-collapse:collapse;
        }

        .receipt-record-table th{
            padding:15px 16px;
            background:#eef5fb;
            color:#355067;
            font-size:15px;
            font-weight:850;
            text-align:left;
            white-space:nowrap;
        }

        .receipt-record-table td{
            padding:15px 16px;
            border-top:1px solid #edf1f5;
            color:#34495e;
            font-size:15px;
            vertical-align:middle;
        }

        .receipt-record-table tbody tr:hover{
            background:#f8fbfe;
        }

        .receipt-no-badge{
            display:inline-block;
            padding:7px 10px;
            border-radius:10px;
            background:#eaf4fd;
            color:#1769aa;
            font-weight:850;
            white-space:nowrap;
        }

        .receipt-amount{
            color:#166d4a;
            font-weight:850;
            white-space:nowrap;
        }

        .receipt-status{
            display:inline-flex;
            align-items:center;
            justify-content:center;
            min-width:78px;
            padding:6px 11px;
            border-radius:999px;
            font-size:13px;
            font-weight:850;
        }

        .receipt-status.valid{
            background:#e8f7ef;
            color:#23774f;
        }

        .receipt-status.cancelled{
            background:#fdeaea;
            color:#b63f3f;
        }

        .receipt-empty{
            padding:45px 20px;
            color:#7b8b9b;
            font-size:17px;
            text-align:center;
        }

        @media (max-width:800px){

            .receipt-summary-page{
                padding:20px 12px 50px;
            }

            .receipt-hero{
                padding:28px 23px;
            }

            .receipt-hero::after{
                display:none;
            }

            .receipt-hero-icon{
                width:60px;
                height:60px;
                border-radius:18px;
                font-size:32px;
            }

            .receipt-hero h1{
                font-size:29px;
            }

            .receipt-hero p{
                font-size:16px;
            }

            .receipt-form-card{
                padding:23px 18px;
            }

            .receipt-form-grid{
                grid-template-columns:1fr;
                gap:16px;
            }

            .receipt-summary-grid{
                grid-template-columns:
                    repeat(2, minmax(0,1fr));
            }

            .receipt-summary-box{
                min-height:115px;
                padding:20px;
            }

            .receipt-summary-value{
                font-size:25px;
            }

        }

        @media (max-width:520px){

            .receipt-summary-grid{
                grid-template-columns:1fr;
            }

            .receipt-hero-title{
                align-items:flex-start;
            }

            .receipt-hero h1{
                font-size:25px;
            }

        }

        </style>

        <div class="receipt-summary-page">

            <div class="receipt-topbar">

                <a
                    class="receipt-back"
                    href="{{ url_for(
                        'finance.finance_admin_home'
                    ) }}"
                >
                    ← 返回负责人中心
                </a>

            </div>

            <section class="receipt-hero">

                <div class="receipt-hero-title">

                    <div class="receipt-hero-icon">
                        📒
                    </div>

                    <div>

                        <h1>
                            收条范围汇总
                        </h1>

                        <p>
                            核对收条张数、现金金额、
                            作废记录及未录入号码
                        </p>

                    </div>

                </div>

            </section>

            {% if error %}

            <div class="receipt-error">
                <span>⚠️</span>
                <span>{{ error }}</span>
            </div>

            {% endif %}

            <section class="receipt-form-card">

                <h2 class="receipt-card-title">
                    🔎 选择要检查的收条范围
                </h2>

                <form method="post">

                    <div class="receipt-form-grid">

                        <div class="receipt-field">

                            <label for="receipt-branch">
                                分会
                            </label>

                            <select
                                id="receipt-branch"
                                class="form-input"
                                name="branch"
                                onchange="updateReceiptPrefix()"
                            >

                                <option
                                    value="CHE"
                                    {% if branch == 'CHE' %}
                                    selected
                                    {% endif %}
                                >
                                    CHE — 蕉赖分会
                                </option>

                                <option
                                    value="STW"
                                    {% if branch == 'STW' %}
                                    selected
                                    {% endif %}
                                >
                                    STW — 实达阿南分会
                                </option>

                            </select>

                        </div>

                        <div class="receipt-field">

                            <label for="receipt-category">
                                收入项目
                            </label>

                            <select
                                id="receipt-category"
                                class="form-input"
                                name="category"
                            >

                                {% for item in categories %}

                                <option
                                    value="{{ item }}"
                                    {% if item == category %}
                                    selected
                                    {% endif %}
                                >
                                    {{ item }}
                                </option>

                                {% endfor %}

                            </select>

                        </div>

                        <div class="receipt-field">

                            <label for="receipt-from">
                                开始收条
                            </label>

                            <<input
                                id="receipt-from"
                                class="form-input"
                                name="receipt_from"
                                value="{{ receipt_from_input }}"
                                placeholder="例如 1555"
                                inputmode="numeric"
                                autocomplete="off"
                                required
                            >

                            <div class="receipt-field-help">
                                只需输入数字，系统会自动补成
                                {{ branch }}0001555
                            </div>

                        </div>

                        <div class="receipt-field">

                            <label for="receipt-to">
                                结束收条
                            </label>

                            <input
                                id="receipt-to"
                                class="form-input"
                                name="receipt_to"
                                value="{{ receipt_to_input }}"
                                placeholder="例如 1560"
                                inputmode="numeric"
                                autocomplete="off"
                                required
                            >

                            <div class="receipt-field-help">
                                只需输入数字，系统会自动补上分会及前导零
                            </div>

                        </div>

                    </div>

                    <button
                        class="receipt-submit"
                        type="submit"
                    >
                        🧮 计算收条范围
                    </button>

                </form>

            </section>

            {% if result %}

            <section class="receipt-summary-grid">

                <div class="receipt-summary-box">

                    <div class="receipt-summary-icon">
                        🔢
                    </div>

                    <strong class="receipt-summary-value">
                        {{ result.span_count }}
                    </strong>

                    <span class="receipt-summary-label">
                        号码跨度
                    </span>

                </div>

                <div class="receipt-summary-box success">

                    <div class="receipt-summary-icon">
                        ✅
                    </div>

                    <strong class="receipt-summary-value">
                        {{ result.valid_count }}
                    </strong>

                    <span class="receipt-summary-label">
                        有效收条
                    </span>

                </div>

                <div class="receipt-summary-box danger">

                    <div class="receipt-summary-icon">
                        ❌
                    </div>

                    <strong class="receipt-summary-value">
                        {{ result.cancelled_count }}
                    </strong>

                    <span class="receipt-summary-label">
                        已作废
                    </span>

                </div>

                <div class="receipt-summary-box warning">

                    <div class="receipt-summary-icon">
                        ⚠️
                    </div>

                    <strong class="receipt-summary-value">
                        {{ result.missing_count }}
                    </strong>

                    <span class="receipt-summary-label">
                        未录入／跳号
                    </span>

                </div>

                <div class="receipt-summary-box money">

                    <div class="receipt-summary-icon">
                        💵
                    </div>

                    <strong class="receipt-summary-value">
                        RM {{
                            "%.2f"|format(
                                result.cash_total
                            )
                        }}
                    </strong>

                    <span class="receipt-summary-label">
                        现金应存银行
                    </span>

                </div>

                <div class="receipt-summary-box success">

                    <div class="receipt-summary-icon">
                        💰
                    </div>

                    <strong class="receipt-summary-value">
                        RM {{
                            "%.2f"|format(
                                result.all_total
                            )
                        }}
                    </strong>

                    <span class="receipt-summary-label">
                        全部有效金额
                    </span>

                </div>

            </section>

            {% if result.missing_count %}

            <section class="receipt-missing-card">

                <div class="receipt-missing-title">
                    ⚠️ 未找到以下收条号码
                </div>

                <div class="receipt-missing-list">

                    {{ result.missing_text }}

                    {% if result.missing_count > 30 %}
                        ……
                    {% endif %}

                </div>

            </section>

            {% endif %}

            <section class="receipt-table-card">

                <div class="receipt-table-header">

                    <h2 class="receipt-table-title">
                        📋 收条明细
                    </h2>

                    <span class="receipt-table-count">
                        共 {{ records|length }} 笔
                    </span>

                </div>

                {% if records %}

                <div class="receipt-table-wrap">

                    <table class="receipt-record-table">

                        <thead>

                            <tr>
                                <th>收条编号</th>
                                <th>日期</th>
                                <th>姓名</th>
                                <th>类别</th>
                                <th>付款方式</th>
                                <th>金额</th>
                                <th>状态</th>
                            </tr>

                        </thead>

                        <tbody>

                            {% for row in records %}

                            {% set row_status =
                                row.status
                                or 'confirmed'
                            %}

                            <tr>

                                <td>
                                    <span class="receipt-no-badge">
                                        {{ row.receipt_no }}
                                    </span>
                                </td>

                                <td>
                                    {{
                                        row.receipt_date
                                        or row.record_date
                                    }}
                                </td>

                                <td>
                                    {{ row.name or "—" }}
                                </td>

                                <td>
                                    {{ row.category or "—" }}
                                </td>

                                <td>
                                    {{
                                        row.payment_method
                                        or "—"
                                    }}
                                </td>

                                <td class="receipt-amount">
                                    RM {{
                                        "%.2f"|format(
                                            row.amount or 0
                                        )
                                    }}
                                </td>

                                <td>

                                    {% if row_status ==
                                          'cancelled' %}

                                    <span
                                        class="
                                            receipt-status
                                            cancelled
                                        "
                                    >
                                        已作废
                                    </span>

                                    {% else %}

                                    <span
                                        class="
                                            receipt-status
                                            valid
                                        "
                                    >
                                        有效
                                    </span>

                                    {% endif %}

                                </td>

                            </tr>

                            {% endfor %}

                        </tbody>

                    </table>

                </div>

                {% else %}

                <div class="receipt-empty">
                    此范围没有找到符合条件的收条记录。
                </div>

                {% endif %}

            </section>

            {% endif %}

        </div>

        <script>

        function updateReceiptPrefix(){

            const branch =
                document.getElementById(
                    "receipt-branch"
                ).value;

            const fromInput =
                document.getElementById(
                    "receipt-from"
                );

            const toInput =
                document.getElementById(
                    "receipt-to"
                );

            fromInput.placeholder =
                branch + "0001201";

            toInput.placeholder =
                branch + "0001250";
        }

        </script>
        """,
        branch=branch,
        category=category,
        categories=categories,
        receipt_from_input=receipt_from_input,
        receipt_to_input=receipt_to_input,
        result=result,
        records=records,
        error=error,
    )

@finance_bp.route("/menu/expense")
def finance_expense_menu():

    expense_items = [

        ("🌸", "供花", "鲜花及供花相关支出"),

        ("🍎", "供果", "水果及供品相关支出"),

        ("🪔", "供油", "灯油及油品相关支出"),

        ("🛕", "佛台用品", "佛具、佛台及佛堂用品"),

        ("⚡", "电费", "TNB 20-1、TNB 20-2"),

        ("💧", "水费", "Air Selangor、Indah Water"),

        ("📶", "电话及网络费", "Celcom、Digi、Unifi"),

        ("🛠️", "维修保养", "维修及保养费用"),

        ("🏗️", "装修", "GYT、观音堂及活动中心装修"),

        ("🛒", "日常采购", "文具、厨房及日常用品"),

        ("📄", "执照及行政费", "License、政府及行政费用"),

        ("🧾", "其它支出", "印刷、交通及其它杂项费用"),

    ]

    return render_template_string(
        FINANCE_V5_STYLE
        + FINANCE_DATE_COMPONENT
        + """
<div class="finance-v5">

    <div class="v5-topbar">

        <a class="v5-back"
        href="{{ url_for('finance.finance_home') }}">
            ← 返回财政首页
        </a>

    </div>

    <div class="v5-header">

        <h1>🧾 支出录入</h1>

        <p>请选择支出项目</p>

    </div>

    <div class="v5-menu-grid">

        {% for icon, category, desc in expense_items %}

        <a class="v5-menu-btn v5-expense"
           href="{{ url_for('finance.expense', category=category) }}">

            <div class="v5-icon">
                {{ icon }}
            </div>

            <div class="v5-menu-text">

                <div class="v5-menu-title">
                    {{ category }}
                </div>

                <div class="v5-menu-desc">
                    {{ desc }}
                </div>

            </div>

        </a>

        {% endfor %}

    </div>

</div>
""",
        expense_items=expense_items,
    )

@finance_bp.route("/menu/member")
@finance_admin_required
def finance_member_menu():

    return render_template_string(FINANCE_V5_STYLE + FINANCE_DATE_COMPONENT + """
    <div class="finance-v5">

        <div class="v5-topbar">
            <a class="v5-back"
               href="{{ url_for('finance.finance_admin_home') }}">
                ← 返回负责人中心
            </a>
        </div>

        <div class="v5-header">
            <h1>👥 会员与月费</h1>
            <p>会员资料与月费情况</p>
        </div>

        <div class="v5-menu-grid">

            <a class="v5-menu-btn v5-member"
               href="{{ url_for('finance.member_management') }}">
                <div class="v5-icon">👤</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">会员管理</div>
                    <div class="v5-menu-desc">
                        搜索、查看和维护会员资料
                    </div>
                </div>
            </a>
            
            <a class="v5-menu-btn v5-member"
               href="{{ url_for('finance.add_member') }}">
                <div class="v5-icon">➕</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">新增会员</div>
                    <div class="v5-menu-desc">
                        建立新月费会员资料
                    </div>
                </div>
            </a>

        </div>

    </div>
    """)

@finance_bp.route("/menu/report")
@finance_admin_required
def finance_report_menu():

    today_ym = date.today().strftime("%Y-%m")

    return render_template_string(
        FINANCE_V5_STYLE + FINANCE_DATE_COMPONENT + """
        <style>
            .v5-export-card{
                display:flex;
                align-items:center;
                gap:16px;
                padding:22px;
                border-radius:18px;
                background:#ffffff;
                box-shadow:0 4px 14px rgba(0,0,0,.08);
            }

            .v5-export-icon{
                font-size:42px;
                flex:0 0 auto;
            }

            .v5-export-content{
                flex:1;
                min-width:0;
            }

            .v5-export-title{
                font-size:24px;
                font-weight:700;
                margin-bottom:6px;
            }

            .v5-export-desc{
                font-size:17px;
                color:#666;
                margin-bottom:14px;
                line-height:1.5;
            }

            .v5-export-form{
                display:flex;
                align-items:center;
                gap:10px;
                flex-wrap:wrap;
            }

            .v5-month-input{
                min-height:48px;
                padding:8px 12px;
                border:2px solid #d6dbe3;
                border-radius:10px;
                font-size:19px;
                background:#fff;
            }

            .v5-download-btn{
                min-height:48px;
                padding:9px 18px;
                border:0;
                border-radius:10px;
                font-size:18px;
                font-weight:700;
                cursor:pointer;
                background:#198754;
                color:#fff;
            }

            .v5-download-btn:hover{
                filter:brightness(.95);
            }

            .v5-download-btn.report{
                background:#246bfd;
            }

            @media (max-width:600px){
                .v5-export-card{
                    align-items:flex-start;
                }

                .v5-export-form{
                    display:grid;
                    grid-template-columns:1fr;
                }

                .v5-month-input,
                .v5-download-btn{
                    width:100%;
                    box-sizing:border-box;
                }
            }
        </style>

        <div class="finance-v5">

            <div class="v5-topbar">
                <a class="v5-back"
                href="{{ url_for('finance.finance_admin_home') }}">
                    ← 返回负责人中心
                </a>
            </div>

            <div class="v5-header">
                <h1>📊 报表与查询</h1>
                <p>统计、记录与 Excel 月报</p>
            </div>

            <div class="v5-menu-grid">

                <a class="v5-menu-btn v5-report"
                href="{{ url_for('finance.dashboard') }}">
                    <div class="v5-icon">📈</div>

                    <div class="v5-menu-text">
                        <div class="v5-menu-title">
                            财政 Dashboard
                        </div>

                        <div class="v5-menu-desc">
                            查看每月收入、支出和户口统计
                        </div>
                    </div>
                </a>

                <a class="v5-menu-btn v5-report"
                href="{{ url_for('finance.records') }}">
                    <div class="v5-icon">🔎</div>

                    <div class="v5-menu-text">
                        <div class="v5-menu-title">
                            财政记录搜索
                        </div>

                        <div class="v5-menu-desc">
                            搜索收条、会员编号、姓名与银行 Reference
                        </div>
                    </div>
                </a>
                
                <!-- 专业版月报 -->
                <div class="v5-export-card">

                    <div class="v5-export-icon">
                        📊
                    </div>

                    <div class="v5-export-content">

                        <div class="v5-export-title">
                            下载专业版月报
                        </div>

                        <div class="v5-export-desc">
                            选择月份后，下载收入、支出及户口统计月报
                        </div>

                        <form
                            method="get"
                            action="{{ url_for('finance.export_monthly_report') }}"
                            class="v5-export-form"
                        >
                            <input
                                type="month"
                                name="ym"
                                value="{{ today_ym }}"
                                class="v5-month-input"
                                required
                            >

                            <button
                                type="submit"
                                class="v5-download-btn report"
                            >
                                📥 下载专业版月报
                            </button>
                        </form>

                    </div>
                </div>

            </div>
            
        </div>
        """,
        today_ym=today_ym,
    )

@finance_bp.route("/member_management")
@finance_admin_required
def member_management():

    q = request.args.get("q", "").strip()

    if q:
        keyword = f"%{q}%"

        rows = db_query("""
            select
                member_id,
                name,
                english_name,
                phone,
                coalesce(status, '在册') as status,
                remark
            from members
            where
                member_id ilike %s
                or name ilike %s
                or english_name ilike %s
                or phone ilike %s
            order by member_id
            limit 300
        """, (
            keyword,
            keyword,
            keyword,
            keyword
        ), fetchall=True)

    else:
        rows = db_query("""
            select
                member_id,
                name,
                english_name,
                phone,
                coalesce(status, '在册') as status,
                remark
            from members
            order by member_id
            limit 300
        """, fetchall=True)

    total_members = len(rows)

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>月费会员管理</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .member-page {
                max-width: 1100px;
            }

            .member-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                border-radius: 22px;
                padding: 26px;
                margin-bottom: 20px;
                box-shadow: 0 12px 30px rgba(37, 99, 235, 0.18);
            }

            .member-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .member-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .member-toolbar {
                display: grid;
                grid-template-columns: 1fr auto;
                gap: 14px;
                align-items: end;
            }

            .member-search-row {
                display: grid;
                grid-template-columns: 1fr auto auto;
                gap: 10px;
                align-items: center;
            }

            .member-search-row .form-input {
                margin: 0;
            }

            .member-count {
                font-size: 17px;
                color: #475569;
                margin-top: 10px;
            }

            .member-name {
                font-weight: 700;
                color: #1e293b;
            }

            .member-id {
                font-weight: 700;
                color: #1d4ed8;
                white-space: nowrap;
            }

            .member-phone {
                white-space: nowrap;
            }

            .status-badge {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                min-width: 72px;
                padding: 6px 10px;
                border-radius: 999px;
                font-size: 15px;
                font-weight: 700;
                white-space: nowrap;
            }

            .status-active {
                background: #dcfce7;
                color: #166534;
            }

            .status-paused {
                background: #fef3c7;
                color: #92400e;
            }

            .status-stopped {
                background: #fee2e2;
                color: #991b1b;
            }

            .status-default {
                background: #e2e8f0;
                color: #334155;
            }

            .table-actions {
                display: flex;
                gap: 8px;
                flex-wrap: wrap;
            }

            .table-actions .btn-tool {
                min-height: auto;
                padding: 8px 13px;
                font-size: 15px;
            }

            .member-empty {
                text-align: center;
                padding: 45px 20px;
                color: #64748b;
            }

            @media (max-width: 760px) {

                .member-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .member-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .member-header h1 {
                    font-size: 26px;
                }

                .member-toolbar {
                    grid-template-columns: 1fr;
                }

                .member-search-row {
                    grid-template-columns: 1fr;
                }

                .member-toolbar .btn-tool,
                .member-search-row .btn-tool {
                    width: 100%;
                }

                .record-table {
                    min-width: 880px;
                }
            }
        </style>
    </head>

    <body>

    <div class="page member-page">

        <div class="member-header">
            <h1>👥 月费会员管理</h1>

            <p>
                查询、新增及编辑 CHE 与 STW 月费会员资料。
            </p>
        </div>

        <div class="card">

            <div class="member-toolbar">

                <div>
                    <div class="section-title">
                        🔍 搜索会员
                    </div>

                    <form method="get">

                        <div class="member-search-row">

                            <input
                                class="form-input"
                                name="q"
                                value="{{ q }}"
                                placeholder="输入编号、姓名、英文名或电话"
                                autocomplete="off"
                            >

                            <button
                                class="btn-tool btn-primary"
                                type="submit"
                            >
                                🔍 搜索
                            </button>

                            {% if q %}
                            <a
                                class="btn-tool btn-secondary"
                                href="{{ url_for(
                                    'finance.member_management'
                                ) }}"
                            >
                                ✕ 清除
                            </a>
                            {% endif %}

                        </div>

                    </form>
                </div>

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for('finance.add_member') }}"
                >
                    ➕ 新增会员
                </a>

            </div>

            <div class="member-count">
                {% if q %}
                    搜索结果：
                    <strong>{{ total_members }}</strong>
                    位会员
                {% else %}
                    当前显示：
                    <strong>{{ total_members }}</strong>
                    位会员
                {% endif %}
            </div>

        </div>

        <div class="card">

            <div class="section-title">
                📋 会员名单
            </div>

            {% if rows %}

            <div class="table-responsive">

                <table class="record-table">

                    <thead>
                        <tr>
                            <th>会员编号</th>
                            <th>姓名</th>
                            <th>英文名</th>
                            <th>电话</th>
                            <th>会员状态</th>
                            <th>备注</th>
                            <th>操作</th>
                        </tr>
                    </thead>

                    <tbody>

                        {% for r in rows %}

                        {% set status_text = r.status or '在册' %}

                        <tr>

                            <td>
                                <span class="member-id">
                                    {{ r.member_id }}
                                </span>
                            </td>

                            <td>
                                <span class="member-name">
                                    {{ r.name }}
                                </span>
                            </td>

                            <td>
                                {{ r.english_name or "-" }}
                            </td>

                            <td class="member-phone">
                                {% if r.phone %}
                                    <a href="tel:{{ r.phone }}">
                                        {{ r.phone }}
                                    </a>
                                {% else %}
                                    -
                                {% endif %}
                            </td>

                            <td>

                                {% if status_text in [
                                    '在册',
                                    'active',
                                    '正常',
                                    '供养中'
                                ] %}

                                    <span
                                        class="
                                            status-badge
                                            status-active
                                        "
                                    >
                                        在册
                                    </span>

                                {% elif status_text in [
                                    '暂停',
                                    'paused',
                                    '停供'
                                ] %}

                                    <span
                                        class="
                                            status-badge
                                            status-paused
                                        "
                                    >
                                        暂停
                                    </span>

                                {% elif status_text in [
                                    '停止',
                                    'stopped',
                                    'inactive',
                                    '永久停止'
                                ] %}

                                    <span
                                        class="
                                            status-badge
                                            status-stopped
                                        "
                                    >
                                        停止
                                    </span>

                                {% else %}

                                    <span
                                        class="
                                            status-badge
                                            status-default
                                        "
                                    >
                                        {{ status_text }}
                                    </span>

                                {% endif %}

                            </td>

                            <td>
                                {{ r.remark or "-" }}
                            </td>

                            <td>

                                <div class="table-actions">

                                    <a
                                        class="
                                            btn-tool
                                            btn-primary
                                        "
                                        href="{{ url_for(
                                            'finance.edit_member',
                                            member_id=r.member_id
                                        ) }}"
                                    >
                                        ✏️ 编辑
                                    </a>

                                </div>

                            </td>

                        </tr>

                        {% endfor %}

                    </tbody>

                </table>

            </div>

            {% else %}

            <div class="member-empty">
                <div style="font-size:44px;">
                    🔎
                </div>

                <h3>找不到相关会员</h3>

                <p>
                    请检查会员编号、姓名、英文名或电话号码。
                </p>
            </div>

            {% endif %}

        </div>

        <div class="btn-row">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_admin_home') }}"
            >
                ← 返回负责人中心
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        q=q,
        total_members=total_members
    )

@finance_bp.route("/member/<member_id>/status", methods=["POST"])
@finance_admin_required
def update_member_status(member_id):

    status = request.form.get("status", "").strip()

    allowed_status = ["在册", "停供", "往生"]

    if status not in allowed_status:
        return "Invalid status", 400

    db_query("""
        update members
        set status = %s
        where member_id = %s
    """, (status, member_id))

    return redirect(url_for("finance.member_management"))

@finance_bp.route("/add_member", methods=["GET","POST"])
@finance_admin_required
def add_member():
    return "Add Member Coming Soon"

@finance_bp.route("/monthly_fee_batch/<branch>", methods=["GET", "POST"])
def monthly_fee_batch(branch):

    branch = branch.upper()

    if branch not in ["CHE", "STW"]:
        return "Invalid branch", 400

    message = ""
    preview_rows = []

    raw_text = request.form.get("raw_text", "").strip()

    next_receipt_no = get_next_receipt_no(
        branch,
        "月费",
    )

    current_book_no = get_current_receipt_book(
        branch,
        "月费",
    )
    next_receipt_raw = next_receipt_no.replace(branch, "", 1)

    receipt_start_raw = request.form.get(
        "receipt_start",
        ""
    ).strip().upper()

    if not receipt_start_raw:
        receipt_start_raw = next_receipt_raw

    if receipt_start_raw.isdigit():
        receipt_start = (
            branch
            + str(int(receipt_start_raw)).zfill(7)
        )
    else:
        receipt_start = receipt_start_raw

    default_receipt_date = (
        request.form.get("receipt_date")
        or date.today().isoformat()
    )

    payment_method = request.form.get(
        "payment_method",
        "现金"
    )

    receipt_book_no = (
        request.form.get("receipt_book_no")
        or current_book_no
    )

    try:
        receipt_book_no = int(receipt_book_no)
    except (TypeError, ValueError):
        receipt_book_no = current_book_no

    bank_payment_date = (
        request.form.get("payment_date")
        or default_receipt_date
    )

    action = request.form.get("action", "")

    default_amount = money(
        request.form.get("default_amount")
        or 350
    )

    def make_receipt_no(start_no, index):
        match = re.match(r"^([A-Z]+)(\d+)$", start_no)

        if not match:
            return ""

        prefix = match.group(1)
        number_text = match.group(2)

        return (
            prefix
            + str(int(number_text) + index).zfill(
                len(number_text)
            )
        )

    def parse_date_header(line):
        """
        支持：
        @2026-07-10
        @ 2026-07-10
        """
        match = re.match(
            r"^@\s*(\d{4}-\d{2}-\d{2})$",
            line
        )

        if not match:
            return None

        date_text = match.group(1)

        try:
            return date.fromisoformat(date_text).isoformat()
        except ValueError:
            return "INVALID"

    def build_preview():
        rows = []
        current_receipt_date = default_receipt_date
        receipt_index = 0
        today_date = date.today()

        for line_no, original_line in enumerate(
            raw_text.splitlines(),
            start=1
        ):
            line = original_line.strip()

            if not line:
                continue

            if line.startswith("@"):
                parsed_date = parse_date_header(line)

                if parsed_date == "INVALID":
                    rows.append({
                        "error": (
                            f"第 {line_no} 行日期无效：{line}"
                        ),
                        "warning": "",
                        "raw": line,
                        "receipt_date": "",
                    })

                elif not parsed_date:
                    rows.append({
                        "error": (
                            f"第 {line_no} 行日期格式错误，"
                            "请使用 @YYYY-MM-DD"
                        ),
                        "warning": "",
                        "raw": line,
                        "receipt_date": "",
                    })

                else:
                    current_receipt_date = parsed_date

                continue

            parts = line.split()

            # 支持：
            # 108
            # 108 100
            # 张三
            # 张三 100
            if (
                len(parts) >= 2
                and re.fullmatch(
                    r"\d+(?:\.\d+)?",
                    parts[-1]
                )
            ):
                amount = money(parts[-1])
                raw_member_keyword = (
                    " ".join(parts[:-1]).strip()
                )

            else:
                amount = default_amount
                raw_member_keyword = line

            receipt_no = make_receipt_no(
                receipt_start,
                receipt_index
            )
            receipt_index += 1

            error_messages = []
            warning_messages = []

            # A. 金额基本检查
            if amount <= 0:
                error_messages.append(
                    f"第 {line_no} 行金额无效：{line}"
                )

            # 月费必须是 RM50 的倍数
            elif amount % 50 != 0:
                error_messages.append(
                    f"月费 RM{amount:.2f} "
                    "不是 RM50 的倍数"
                )

            # B. 日期检查
            try:
                receipt_date_obj = date.fromisoformat(
                    current_receipt_date
                )

                month_lock_error = require_finance_month_open(
                    current_receipt_date,
                    get_fund_account(
                        "月费",
                        branch=branch
                    )
                )

                if month_lock_error:
                    error_messages.append(
                        month_lock_error
                    )

                if receipt_date_obj > today_date:
                    error_messages.append(
                        "收条日期是未来日期"
                    )

                else:
                    days_old = (
                        today_date - receipt_date_obj
                    ).days

                    if days_old > 30:
                        warning_messages.append(
                            f"收条日期距今已有 "
                            f"{days_old} 天，"
                            "请确认是否为旧账补录"
                        )

            except (TypeError, ValueError):
                error_messages.append(
                    "收条日期格式无效"
                )

            if error_messages:
                rows.append({
                    "error": "；".join(error_messages),
                    "warning": "；".join(warning_messages),
                    "raw": line,
                    "receipt_no": receipt_no,
                    "receipt_date": current_receipt_date,
                })
                continue

            # C. 会员查找
            if raw_member_keyword.isdigit():
                member_id = (
                    f"{branch}-{int(raw_member_keyword)}"
                )

            else:
                member_id = normalize_member_id(
                    raw_member_keyword,
                    default_branch=branch
                )

            # 先按会员编号查找
            member = db_query("""
                select *
                from members
                where member_id = %s
                limit 1
            """, (
                member_id,
            ), fetchone=True)

            # 编号找不到时，再按中文名或英文名查找
            if not member:
                name_matches = db_query("""
                    select *
                    from members
                    where branch = %s
                      and (
                            lower(trim(name))
                                = lower(trim(%s))
                         or lower(
                                trim(
                                    coalesce(
                                        english_name,
                                        ''
                                    )
                                )
                            )
                                = lower(trim(%s))
                         or regexp_replace(coalesce(phone, ''), '[^0-9]', '', 'g')
                                in (%s, %s)
                      )
                    order by member_id
                    limit 2
                """, (
                    branch,
                    raw_member_keyword,
                    raw_member_keyword,
                    phone_search_variants(raw_member_keyword)[0],
                    phone_search_variants(raw_member_keyword)[1],
                ), fetchall=True) or []

                if len(name_matches) == 1:
                    member = name_matches[0]

                elif len(name_matches) > 1:
                    rows.append({
                        "error": (
                            f"第 {line_no} 行姓名有重复，"
                            "请改用会员编号："
                            f"{raw_member_keyword}"
                        ),
                        "warning": "",
                        "raw": line,
                        "receipt_no": receipt_no,
                        "receipt_date": current_receipt_date,
                    })
                    continue

            if not member:
                rows.append({
                    "error": (
                        f"第 {line_no} 行找不到会员："
                        f"{raw_member_keyword}"
                    ),
                    "warning": "",
                    "raw": line,
                    "receipt_no": receipt_no,
                    "receipt_date": current_receipt_date,
                })
                continue

            # 姓名查找成功后，必须改用数据库里的正式编号
            member_id = member["member_id"]

            # D. 计算供养月份
            paid = db_query("""
                select max(end_month) as paid_until
                from member_payments
                where member_id = %s
                  and coalesce(status, 'active') = 'active'
            """, (
                member_id,
            ), fetchone=True)

            paid_until_date = (
                paid["paid_until"]
                if paid
                else None
            )

            month_from = next_month_ym(
                paid_until_date
            )

            month_count = int(amount / 50)

            month_to = add_months_ym(
                month_from,
                month_count - 1
            )

            # E. 收条重复检查
            existing_finance = db_query("""
                select id
                from finance_records
                where receipt_no = %s
                limit 1
            """, (
                receipt_no,
            ), fetchone=True)

            existing_payment = db_query("""
                select id
                from member_payments
                where receipt_no = %s
                  and coalesce(status, 'active') = 'active'
                limit 1
            """, (
                receipt_no,
            ), fetchone=True)

            if existing_finance and existing_payment:
                error_messages.append(
                    "收条已存在于财政记录和月费查询"
                )

            elif existing_finance:
                error_messages.append(
                    "收条已存在于财政记录"
                )

            elif existing_payment:
                error_messages.append(
                    "收条已存在于月费查询"
                )

            # F. 同一会员、同一天重复月费提醒
            same_day_payment = db_query("""
                select
                    receipt_no,
                    amount
                from member_payments
                where member_id = %s
                  and receipt_date = %s
                  and coalesce(status, 'active') = 'active'
                order by id desc
                limit 1
            """, (
                member_id,
                current_receipt_date,
            ), fetchone=True)

            if same_day_payment:
                old_receipt = (
                    same_day_payment.get("receipt_no")
                    or "无收条编号"
                )

                old_amount = float(
                    same_day_payment.get("amount")
                    or 0
                )

                warning_messages.append(
                    "此会员同一天已有月费记录："
                    f"{old_receipt}，"
                    f"RM{old_amount:.2f}"
                )

            # G. 银行付款日期检查
            if payment_method == "银行过账":
                row_payment_date = bank_payment_date

                try:
                    payment_date_obj = date.fromisoformat(
                        row_payment_date
                    )

                    if payment_date_obj > today_date:
                        error_messages.append(
                            "银行付款日期是未来日期"
                        )

                except (TypeError, ValueError):
                    error_messages.append(
                        "银行付款日期格式无效"
                    )

            else:
                row_payment_date = current_receipt_date

            rows.append({
                "error": "；".join(error_messages),
                "warning": "；".join(warning_messages),
                "receipt_no": receipt_no,
                "receipt_date": current_receipt_date,
                "payment_date": row_payment_date,
                "member_id": member_id,
                "name": (
                    member.get("姓名")
                    or member.get("name")
                ),
                "phone": (
                    member.get("电话号码")
                    or member.get("phone")
                ),
                "amount": amount,
                "month_from": month_from,
                "month_to": month_to,
                "month_count": month_count,
            })

        return rows

    if request.method == "POST":

        if receipt_book_no != current_book_no:

            update_receipt_book_no(
                branch,
                "月费",
                receipt_book_no,
            )

            current_book_no = receipt_book_no

        if not receipt_start:
            message = "请填写收条开始号码"

        elif not re.match(
            rf"^{branch}\d+$",
            receipt_start
        ):
            message = (
                "收条号码格式错误，例如："
                f"{branch}0001501，或只输入 1501"
            )

        elif not raw_text:
            message = "请先加入或贴上月费资料"

        else:
            preview_rows = build_preview()

            data_rows = [
                row
                for row in preview_rows
                if row.get("member_id")
            ]

            has_error = any(
                row.get("error")
                for row in preview_rows
            )

            if not data_rows and not has_error:
                message = "没有可预览的会员资料"

            elif action == "confirm" and not has_error:

                confirm_lock_error = None

                # 保存前重新检查所有月份
                # 防止预览后，该月份才被别人完成月结
                for row in data_rows:

                    confirm_lock_error = (
                        require_finance_month_open(
                            row["receipt_date"],
                            get_fund_account(
                                "月费",
                                branch=branch
                            )
                        )
                    )

                    if confirm_lock_error:
                        break

                if confirm_lock_error:

                    message = confirm_lock_error

                else:

                    last_receipt_no = ""

                    for row in data_rows:
                        member_id = row["member_id"]
                        branch = (
                            "STW"
                            if str(member_id).startswith("STW-")
                            else "CHE"
                        )
                        month_from_db = (
                            row["month_from"] + "-01"
                        )

                        month_to_db = (
                            row["month_to"] + "-01"
                        )

                        db_query("""
                            insert into finance_records
                            (
                                record_type,
                                fund_account,
                                record_date,
                                receipt_date,
                                category,
                                receipt_no,
                                receipt_book_no,
                                member_id,
                                name,
                                phone,
                                amount,
                                payment_method,
                                month_from,
                                month_to,
                                remarks
                            )
                            values
                            (
                                %s, %s, %s, %s,
                                '月费',
                                %s, %s, %s, %s,
                                %s, %s, %s, %s,
                                %s, %s
                            )
                        """, (
                            "income",
                            get_fund_account(
                                "月费",
                                branch=branch
                            ),
                            row["payment_date"],
                            row["receipt_date"],
                            row["receipt_no"],
                            receipt_book_no,
                            row["member_id"],
                            row["name"],
                            row["phone"],
                            row["amount"],
                            payment_method,
                            month_from_db,
                            month_to_db,
                            "批量月费录入",
                        ))

                        if row.get("phone"):
                            db_query("""
                                update members
                                set phone = %s
                                where member_id = %s
                                  and coalesce(trim(phone), '') = ''
                            """, (
                                normalize_phone(row["phone"]),
                                row["member_id"],
                            ))

                        db_query("""
                            insert into member_payments
                            (
                                payment_date,
                                receipt_date,
                                member_id,
                                name,
                                receipt_no,
                                amount,
                                start_month,
                                end_month,
                                month_count
                            )
                            values
                            (
                                %s, %s, %s, %s, %s,
                                %s, %s, %s, %s
                            )
                        """, (
                            row["payment_date"],
                            row["receipt_date"],
                            row["member_id"],
                            row["name"],
                            row["receipt_no"],
                            row["amount"],
                            month_from_db,
                            month_to_db,
                            row["month_count"],
                        ))

                        last_receipt_no = row["receipt_no"]

                    if last_receipt_no:
                        update_receipt_book_number(
                            branch,
                            "月费",
                            last_receipt_no,
                        )

                    return redirect(
                        url_for("finance.records")
                    )

# ============================================================
# 页面预览区还要做以下两处修改
# ============================================================

# 1. 在 <style> 中加入：
#
# .preview-warning{
#     color:#b77900;
#     font-weight:700;
# }
#
# 2. 把“检查结果”单元格替换成：
#
# {% if r.error %}
#     <span class="preview-error">
#         ❌ {{ r.error }}
#     </span>
#
# {% elif r.warning %}
#     <span class="preview-warning">
#         ⚠️ {{ r.warning }}
#     </span>
#
# {% else %}
#     <span class="preview-success">
#         ✅ 可以入账
#     </span>
# {% endif %}
#
# 注意：
# - error 会阻止“确认全部入账”
# - warning 只提醒，仍然允许确认入账

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">

        <title>{{ branch }} 月费录入</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .finance-form-page{
                max-width:920px;
            }

            .finance-header-card{
                margin-bottom:18px;
                background:linear-gradient(135deg,#1769aa,#2589c5);
                color:#fff;
            }

            .finance-header-card .page-title,
            .finance-header-card .page-subtitle{
                color:#fff;
            }

            .finance-back-row{
                margin-bottom:16px;
            }

            .step-title{
                display:flex;
                align-items:center;
                gap:10px;
                margin-bottom:18px;
            }

            .step-title .section-title{
                margin:0;
            }

            .step-badge{
                flex:0 0 auto;
                display:inline-flex;
                align-items:center;
                justify-content:center;
                width:36px;
                height:36px;
                border-radius:50%;
                background:#1769aa;
                color:#fff;
                font-size:17px;
                font-weight:800;
            }

            .settings-grid{
                display:grid;
                grid-template-columns:repeat(2,minmax(0,1fr));
                gap:16px;
            }

            .settings-grid .form-group{
                margin:0;
            }

            .finance-full{
                grid-column:1 / -1;
            }

            .receipt-input-row{
                display:flex;
                align-items:stretch;
                gap:10px;
            }

            .receipt-prefix{
                min-width:78px;
                display:flex;
                align-items:center;
                justify-content:center;
                padding:0 15px;
                border-radius:10px;
                background:#eaf4ff;
                color:#1769aa;
                font-weight:800;
                font-size:20px;
            }

            .receipt-input-row .form-input{
                flex:1;
                min-width:0;
            }

            .field-help{
                display:block;
                color:#667085;
                font-size:15px;
                line-height:1.55;
                margin-top:7px;
            }

            .quick-member-row{
                display:grid;
                grid-template-columns:minmax(0,1fr) 280px;
                gap:14px;
                align-items:end;
            }

            .amount-panel{
                display:grid;
                grid-template-columns:54px 1fr 54px;
                gap:8px;
                align-items:stretch;
            }

            .amount-step-btn{
                border:1px solid #cbd5e1;
                border-radius:10px;
                background:#f8fafc;
                color:#172033;
                font-size:25px;
                font-weight:800;
                cursor:pointer;
            }

            .amount-step-btn:hover{
                background:#eaf4ff;
                border-color:#8fbce0;
            }

            .amount-step-btn:active{
                transform:scale(.97);
            }

            #quick_amount{
                text-align:center;
                font-size:25px;
                font-weight:800;
            }

            .amount-shortcuts{
                display:flex;
                flex-wrap:wrap;
                gap:8px;
                margin-top:12px;
            }

            .amount-chip{
                min-width:72px;
                border:1px solid #cbd5e1;
                border-radius:999px;
                background:#fff;
                padding:9px 14px;
                color:#1769aa;
                font-size:16px;
                font-weight:800;
                cursor:pointer;
            }

            .amount-chip:hover{
                background:#eaf4ff;
                border-color:#8fbce0;
            }

            .quick-add-btn{
                width:100%;
                margin-top:16px;
            }

            .batch-textarea{
                width:100%;
                min-height:250px;
                resize:vertical;
                line-height:1.75;
                font-family:Consolas,"Microsoft YaHei",monospace;
                font-size:18px;
            }

            .batch-help-grid{
                display:grid;
                grid-template-columns:repeat(5,minmax(0,1fr));
                gap:10px;
                margin-top:12px;
            }

            .batch-help-item{
                padding:12px 14px;
                border:1px solid #dbe3ed;
                border-radius:12px;
                background:#f8fafc;
                color:#475467;
                font-size:14px;
                line-height:1.6;
            }

            .batch-help-item strong{
                display:block;
                margin-bottom:4px;
                color:#172033;
            }

            .batch-help-item code{
                color:#1769aa;
                font-weight:800;
            }

            .action-row{
                display:grid;
                grid-template-columns:1fr 1fr;
                gap:12px;
                margin-top:18px;
            }

            .action-row.single{
                grid-template-columns:1fr;
            }

            .action-row .btn-tool{
                width:100%;
            }

            .payment-date-box{
                display:none;
            }

            .preview-success{
                color:#16863a;
                font-weight:700;
            }

            .preview-error{
                color:#c62828;
                font-weight:700;
            }

            .preview-date{
                white-space:nowrap;
                font-weight:700;
                color:#1769aa;
            }
                                  
            .preview-warning{
                color:#b77900;
                font-weight:700;
            }

            @media(max-width:760px){
                .settings-grid,
                .quick-member-row,
                .batch-help-grid,
                .action-row{
                    grid-template-columns:1fr;
                }

                .finance-full{
                    grid-column:auto;
                }

                .receipt-input-row{
                    align-items:stretch;
                }
            }
        </style>
    </head>

    <body>
    <div class="page finance-form-page">

        <div class="finance-back-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_income_menu') }}"
            >
                ← 返回收入录入
            </a>
        </div>

        <div class="card finance-header-card">
            <h1 class="page-title">💳 {{ branch }} 月费录入</h1>
            <p class="page-subtitle">
                先设定收条资料，再加入会员；单笔与批量共用同一份待录入清单。
            </p>
        </div>

        {% if message %}
            <div class="alert alert-danger">{{ message }}</div>
        {% endif %}

        <form method="post">

            <div class="card">
                <div class="step-title">
                    <span class="step-badge">1</span>
                    <h2 class="section-title">收条资料与单笔加入</h2>
                </div>
                                  
                <div class="card" style="margin-bottom:18px;background:#f8fbff;border:1px solid #d8e7f7;">

                    <div class="form-group" style="margin-bottom:0;">

                        <label class="form-label">
                            📒 当前 OR Book
                        </label>

                        <div style="
                            display:flex;
                            gap:12px;
                            align-items:center;
                        ">

                            <input
                                class="form-input"
                                style="max-width:150px;"
                                type="number"
                                min="1"
                                name="receipt_book_no"
                                value="{{ current_book_no }}"
                            >

                            <span class="field-help" style="margin-top:0;">
                                开始使用新的实体收条簿时才需要修改。
                            </span>

                        </div>

                    </div>

                </div>

                <div class="settings-grid">
                    <div class="form-group">
                        <label class="form-label">收条开始号码</label>

                        <div class="receipt-input-row">
                            <div class="receipt-prefix">{{ branch }}</div>

                            <input
                                class="form-input"
                                name="receipt_start"
                                value="{{ receipt_start_raw }}"
                                inputmode="numeric"
                                placeholder="例如 1501"
                                required
                            >
                        </div>

                        <span class="field-help">
                            系统建议下一张：<strong>{{ next_receipt_no }}</strong>
                        </span>
                    </div>

                    <div class="form-group">
                        <label class="form-label">默认开收条日期</label>

                        <input
                            class="form-input"
                            name="receipt_date"
                            type="date"
                            value="{{ default_receipt_date }}"
                            required
                        >

                        <span class="field-help">
                            没有写 @日期 的会员会使用这个日期。
                        </span>
                    </div>
                </div>

                <div style="border-top:1px solid #e2e8f0;margin:22px 0 18px;"></div>

                <h3 class="entry-method-title" style="margin-bottom:12px;">
                    👤 单笔快速加入
                </h3>

                <div class="quick-member-row">
                    <div class="form-group">
                        <label class="form-label">会员编号或姓名</label>
                        <input
                            class="form-input"
                            id="quick_member_keyword"
                            type="text"
                            placeholder="例如：108、{{ branch }}-108、张三"
                            autocomplete="off"
                        >
                    </div>

                    <div class="form-group">
                        <label class="form-label">本笔金额 RM</label>
                        <div class="amount-panel">
                            <button
                                type="button"
                                class="amount-step-btn"
                                onclick="changeQuickAmount(-50)"
                            >−</button>

                            <input
                                class="form-input"
                                id="quick_amount"
                                type="number"
                                min="50"
                                step="50"
                                value="{{ default_amount }}"
                            >

                            <button
                                type="button"
                                class="amount-step-btn"
                                onclick="changeQuickAmount(50)"
                            >＋</button>
                        </div>
                    </div>
                </div>

                <div class="amount-shortcuts">
                    {% for amount in [50, 100, 150, 200, 250, 300] %}
                        <button
                            type="button"
                            class="amount-chip"
                            onclick="setQuickAmount({{ amount }})"
                        >
                            RM{{ amount }}
                        </button>
                    {% endfor %}
                </div>

                <button
                    class="btn-tool btn-success quick-add-btn"
                    type="button"
                    onclick="addQuickMember()"
                >
                    ➕ 加入待录入清单
                </button>
            </div>

            <div class="card">
                <div class="step-title">
                    <span class="step-badge">2</span>
                    <h2 class="section-title">待录入清单与付款资料</h2>
                </div>

                <p class="page-subtitle" style="margin-top:-6px;">
                    单笔加入会自动出现在这里，也可以直接贴上多位会员资料。
                </p>

                <textarea
                    class="form-input batch-textarea"
                    id="raw_text"
                    name="raw_text"
                    placeholder="例如：
    @2026-07-10
    108
    张三
    188 100

    @2026-07-11
    69
    205 300"
                >{{ raw_text }}</textarea>

                <div class="batch-help-grid">
                    <div class="batch-help-item">
                        <strong>普通月费</strong>
                        <code>108</code> 或 <code>张三</code>
                    </div>

                    <div class="batch-help-item">
                        <strong>指定金额</strong>
                        <code>188 100</code>
                    </div>

                    <div class="batch-help-item">
                        <strong>切换收条日期</strong>
                        <code>@2026-07-11</code>
                    </div>
                </div>

                <div style="border-top:1px solid #e2e8f0;margin:22px 0 18px;"></div>

                <div class="settings-grid">
                    <div class="form-group">
                        <label class="form-label">付款方式</label>

                        <select
                            class="form-input"
                            name="payment_method"
                            id="payment_method"
                            onchange="togglePaymentDate()"
                        >
                            {% for method in ['现金', '银行过账', '支票'] %}
                                <option
                                    value="{{ method }}"
                                    {% if payment_method == method %}selected{% endif %}
                                >
                                    {{ method }}
                                </option>
                            {% endfor %}
                        </select>
                    </div>

                    <div
                        class="form-group payment-date-box"
                        id="payment_date_box"
                    >
                        <label class="form-label">银行付款日期</label>

                        <input
                            class="form-input"
                            name="payment_date"
                            type="date"
                            value="{{ bank_payment_date }}"
                        >
                    </div>

                    <div class="form-group finance-full">
                        <label class="form-label">批量清单默认月费 RM</label>

                        <input
                            class="form-input"
                            name="default_amount"
                            type="number"
                            step="50"
                            min="50"
                            value="{{ default_amount }}"
                            required
                        >

                        <span class="field-help">
                            清单只写编号或姓名时使用；行内有金额时，以行内金额为准。
                        </span>
                    </div>
                </div>

                <div class="action-row {% if not (preview_rows and not has_preview_error) %}single{% endif %}">
                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                        name="action"
                        value="preview"
                    >
                        👁️ 预览资料
                    </button>

                    {% if preview_rows and not has_preview_error %}
                        <button
                            class="btn-tool btn-success"
                            type="submit"
                            name="action"
                            value="confirm"
                            onclick="return confirm('确定全部入账？');"
                        >
                            ✅ 确认全部入账
                        </button>
                    {% endif %}
                </div>
            </div>

            {% if preview_rows %}
                <div class="card">
                    <div class="step-title">
                        <span class="step-badge">3</span>
                        <h2 class="section-title">月费录入预览</h2>
                    </div>

                    <div class="table-responsive">
                        <table class="record-table">
                            <thead>
                                <tr>
                                    <th>收条</th>
                                    <th>收条日期</th>
                                    <th>会员编号</th>
                                    <th>姓名</th>
                                    <th>电话</th>
                                    <th>金额</th>
                                    <th>开始月份</th>
                                    <th>缴费至</th>
                                    <th>月数</th>
                                    <th>检查结果</th>
                                </tr>
                            </thead>

                            <tbody>
                                {% for r in preview_rows %}
                                    <tr>
                                        <td>{{ r.receipt_no or '-' }}</td>
                                        <td class="preview-date">{{ r.receipt_date or '-' }}</td>
                                        <td><strong>{{ r.member_id or '-' }}</strong></td>
                                        <td>{{ r.name or '-' }}</td>
                                        <td>{{ r.phone or '-' }}</td>
                                        <td>
                                            {% if r.amount %}
                                                RM {{ '%.2f'|format(r.amount) }}
                                            {% else %}
                                                -
                                            {% endif %}
                                        </td>
                                        <td>{{ r.month_from or '-' }}</td>
                                        <td>{{ r.month_to or '-' }}</td>
                                        <td>
                                            {% if r.month_count %}
                                                {{ r.month_count }} 个月
                                            {% else %}
                                                -
                                            {% endif %}
                                        </td>
                                        <td>
                                            {% if r.error %}
                                                <span class="preview-error">
                                                    ❌ {{ r.error }}
                                                </span>

                                            {% elif r.warning %}
                                                <span class="preview-warning">
                                                    ⚠️ {{ r.warning }}
                                                </span>

                                            {% else %}
                                                <span class="preview-success">
                                                    ✅ 可以入账
                                                </span>
                                            {% endif %}
                                        </td>
                                    </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                </div>
            {% endif %}

        </form>
    </div>

    <script>
    function changeQuickAmount(change){
        const input = document.getElementById("quick_amount");

        if(!input){
            return;
        }

        let current = Number(input.value);

        if(!Number.isFinite(current)){
            current = 50;
        }

        current += change;

        if(current < 50){
            current = 50;
        }

        input.value = current;
    }

    function setQuickAmount(amount){
        const input = document.getElementById("quick_amount");

        if(input){
            input.value = amount;
        }
    }

    function addQuickMember(){
        const keywordInput = document.getElementById("quick_member_keyword");
        const amountInput = document.getElementById("quick_amount");
        const textarea = document.getElementById("raw_text");

        if(!keywordInput || !amountInput || !textarea){
            return;
        }

        const keyword = keywordInput.value.trim();
        const amount = amountInput.value.trim();

        if(!keyword){
            alert("请输入会员编号或姓名");
            keywordInput.focus();
            return;
        }

        const newLine = amount
            ? keyword + " " + amount
            : keyword;

        const current = textarea.value.trim();

        textarea.value = current
            ? current + "\\n" + newLine
            : newLine;

        keywordInput.value = "";
        keywordInput.focus();
        textarea.scrollTop = textarea.scrollHeight;
    }

    function togglePaymentDate(){
        const method = document.getElementById("payment_method");
        const box = document.getElementById("payment_date_box");

        if(!method || !box){
            return;
        }

        box.style.display = method.value === "银行过账"
            ? "block"
            : "none";
    }

    document.addEventListener("DOMContentLoaded", function(){
        const keywordInput = document.getElementById("quick_member_keyword");

        if(keywordInput){
            keywordInput.addEventListener("keydown", function(event){
                if(event.key === "Enter"){
                    event.preventDefault();
                    addQuickMember();
                }
            });
        }

        togglePaymentDate();
    });
    </script>

    </body>
    </html>
    """,
        branch=branch,
        message=message,
        raw_text=raw_text,
        receipt_start_raw=receipt_start_raw,
        next_receipt_no=next_receipt_no,
        default_receipt_date=default_receipt_date,
        bank_payment_date=bank_payment_date,
        current_book_no=current_book_no,
        payment_method=payment_method,
        default_amount=default_amount,
        preview_rows=preview_rows,
        has_preview_error=any(
            row.get("error")
            for row in preview_rows
        ),
    )

def get_recent_donors(category=None):
    if category:
        return db_query("""
            select
                name,
                max(phone) as phone,
                count(*) as times,
                max(record_date) as last_date
            from finance_records
            where record_type = 'income'
              and category = %s
              and coalesce(status, 'confirmed') <> 'cancelled'
              and coalesce(name, '') <> ''
              and category <> '月费'
            group by name
            order by max(record_date) desc, count(*) desc
            limit 100
        """, (category,), fetchall=True)

    return db_query("""
        select
            name,
            max(phone) as phone,
            count(*) as times,
            max(record_date) as last_date
        from finance_records
        where record_type = 'income'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and coalesce(name, '') <> ''
          and category <> '月费'
        group by name
        order by max(record_date) desc, count(*) desc
        limit 100
    """, fetchall=True)
    

@finance_bp.route("/income/<category>", methods=["GET", "POST"])
def normal_income(category):

    allowed_categories = ["财布施", "观音村", "膳食结缘"]

    if category not in allowed_categories:
        return "Invalid category", 400

    receipt_prefix = "CHE"

    last_receipt = db_query("""
        select receipt_no
        from finance_records
        where receipt_no like %s
        order by receipt_no desc
        limit 1
    """, (receipt_prefix + "%",), fetchone=True)

    if last_receipt and last_receipt["receipt_no"]:
        old_no = last_receipt["receipt_no"]
        number = int(old_no[3:])
        next_receipt_no = receipt_prefix + str(number + 1).zfill(len(old_no) - 3)
    else:
        next_receipt_no = "CHE0000001"

    message = ""

    if request.method == "POST":

        receipt_no = request.form.get("receipt_no", "").strip().upper()
        receipt_date = request.form.get("receipt_date") or date.today()
        record_date = request.form.get("record_date") or date.today()

        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        amount = money(request.form.get("amount"))
        payment_method = request.form.get("payment_method", "现金")
        remarks = request.form.get("remarks", "").strip()

        if not receipt_no.startswith("CHE"):
            message = "收条号码必须以 CHE 开头"

        elif amount <= 0:
            message = "金额必须大过 0"

        else:
            existing = db_query("""
                select id
                from finance_records
                where receipt_no = %s
                limit 1
            """, (receipt_no,), fetchone=True)

            if existing:
                message = "这个收条号码已经记录过了，请检查是否重复输入"
            else:
                db_query("""
                    insert into finance_records
                    (
                        record_type,
                        fund_account,
                        record_date,
                        receipt_date,
                        category,
                        receipt_no,
                        name,
                        phone,
                        amount,
                        payment_method,
                        remarks
                    )
                    values
                    (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s
                    )
                """, (
                    "income",
                    get_fund_account(category),
                    record_date,
                    receipt_date,
                    category,
                    receipt_no,
                    name,
                    phone,
                    amount,
                    payment_method,
                    remarks
                ))

                return redirect(url_for("finance.records"))

    return render_template_string(FINANCE_STYLE + FINANCE_DATE_COMPONENT + """
    <h1>{{ category }}</h1>

    {% if message %}
        <p style="color:red;">{{ message }}</p>
    {% endif %}

    <form method="post">

        <p>
            收条号码：
            <input
                name="receipt_no"
                value="{{ next_receipt_no }}"
                required
            >
            <br>
            <small style="color:#666;">
                系统会自动建议 CHE 编号；如果换新收条簿，可以手动修改。
            </small>
        </p>

        <p>
            开收条日期：
            <input name="receipt_date" type="date" value="{{ today }}" required>
        </p>

        <p>
            付款日期：
            <input name="record_date" type="date" value="{{ today }}" required>
        </p>

        <p>
            姓名：
            <input name="name">
        </p>

        <p>
            电话：
            <input name="phone">
        </p>

        <p>
            金额 RM：
            <input
                name="amount"
                type="number"
                step="1.00"
                min="0"
                value="50"
                required
            >
        </p>

        <p>
            付款方式：
            <select name="payment_method">
                <option>现金</option>
                <option>银行过账</option>
                <option>支票</option>
            </select>
        </p>

        <p>
            备注：
            <input name="remarks">
        </p>

        <button type="submit">
            保存
        </button>

    </form>

    <p>
        <a href="{{ url_for('finance.finance_admin_home') }}">
            返回负责人中心
        </a>
    </p>
                                  
    <script>

    function addDonor(name){

        let textarea = document.querySelector(
            'textarea[name="raw_text"]'
        );

        if(textarea.value.trim() === ""){
            textarea.value = name;
        }else{
            textarea.value += "\\n" + name;
        }

        textarea.focus();
    }

    </script>

    """,
    category=category,
    next_receipt_no=next_receipt_no,
    message=message,
    today=date.today().isoformat()
    )

def search_finance_donors(keyword: str, branch: str = "CHE", limit_per_source: int = 20):
    """统一搜索月费会员、非会员义工及历史布施人；每个来源保留独立名额。"""

    keyword = str(keyword or "").strip()
    branch = str(branch or "CHE").strip().upper()
    limit_per_source = max(1, min(int(limit_per_source or 20), 50))

    if not keyword:
        return []

    like_keyword = f"%{keyword}%"
    compact_keyword = "".join(keyword.split())
    phone_local, phone_international = phone_search_variants(keyword)
    phone_search = phone_local or compact_keyword
    like_compact = f"%{phone_search}%"

    results = []
    seen = set()

    def add_result(item):
        name = str(item.get("name") or "").strip()
        phone = str(item.get("phone") or "").strip()
        member_id = str(item.get("member_id") or "").strip()
        volunteer_id = str(item.get("volunteer_id") or "").strip()

        if not name:
            return

        identity = (
            member_id.upper()
            or volunteer_id.upper()
            or (name.casefold(), "".join(ch for ch in phone if ch.isdigit()))
        )

        if identity in seen:
            return

        seen.add(identity)
        results.append(item)

    member_rows = db_query("""
        select
            member_id,
            name,
            coalesce(english_name, '') as english_name,
            coalesce(phone, '') as phone
        from members
        where (%s = '' or upper(coalesce(branch, 'CHE')) = %s)
          and (
                member_id ilike %s
             or name ilike %s
             or coalesce(english_name, '') ilike %s
             or regexp_replace(coalesce(phone, ''), '[^0-9]', '', 'g') ilike %s
          )
        order by
            case
                when upper(member_id) = upper(%s) then 0
                when name = %s then 1
                when coalesce(english_name, '') = %s then 2
                else 3
            end,
            name
        limit %s
    """, (
        branch,
        branch,
        like_keyword,
        like_keyword,
        like_keyword,
        like_compact,
        keyword,
        keyword,
        keyword,
        limit_per_source,
    ), fetchall=True) or []

    for row in member_rows:
        add_result({
            "source": "member",
            "source_label": "月费会员",
            "name": row.get("name") or "",
            "english_name": row.get("english_name") or "",
            "member_id": row.get("member_id") or "",
            "volunteer_id": "",
            "phone": row.get("phone") or "",
            "last_date": "",
            "times": 0,
        })

    volunteer_rows = db_query("""
        select
            id as volunteer_id,
            name,
            coalesce(phone, '') as phone
        from volunteers
        where (%s = '' or upper(coalesce(branch, 'CHE')) = %s)
          and not exists (
                select 1
                from members m
                where upper(coalesce(m.branch, 'CHE')) = upper(coalesce(volunteers.branch, 'CHE'))
                  and (
                        trim(m.name) = trim(volunteers.name)
                     or (
                            regexp_replace(coalesce(m.phone, ''), '[^0-9]', '', 'g') <> ''
                        and regexp_replace(coalesce(m.phone, ''), '[^0-9]', '', 'g')
                            = regexp_replace(coalesce(volunteers.phone, ''), '[^0-9]', '', 'g')
                     )
                  )
          )
          and (
                cast(id as text) ilike %s
             or name ilike %s
             or regexp_replace(coalesce(phone, ''), '[^0-9]', '', 'g') ilike %s
          )
        order by
            case
                when upper(cast(id as text)) = upper(%s) then 0
                when name = %s then 1
                else 2
            end,
            name
        limit %s
    """, (
        branch,
        branch,
        like_keyword,
        like_keyword,
        like_compact,
        keyword,
        keyword,
        limit_per_source,
    ), fetchall=True) or []

    for row in volunteer_rows:
        volunteer_id = str(row.get("volunteer_id") or "").strip()
        name = str(row.get("name") or "").strip()
        phone = str(row.get("phone") or "").strip()

        duplicate = any(
            (name and item.get("name") == name)
            and (
                not phone
                or not item.get("phone")
                or "".join(ch for ch in item.get("phone", "") if ch.isdigit())
                   == "".join(ch for ch in phone if ch.isdigit())
            )
            for item in results
        )
        if duplicate:
            continue

        add_result({
            "source": "volunteer",
            "source_label": "义工",
            "name": name,
            "english_name": "",
            "member_id": "",
            "volunteer_id": volunteer_id,
            "phone": phone,
            "last_date": "",
            "times": 0,
        })

    history_rows = db_query("""
        select
            name,
            (array_agg(coalesce(phone, '') order by record_date desc, id desc))[1] as phone,
            max(record_date) as last_date,
            count(*) as times
        from finance_records
        where category <> '月费'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and coalesce(name, '') <> ''
          and (
                name ilike %s
             or regexp_replace(coalesce(phone, ''), '[^0-9]', '', 'g') ilike %s
          )
        group by name
        order by
            case when name = %s then 0 else 1 end,
            max(record_date) desc,
            count(*) desc
        limit %s
    """, (
        like_keyword,
        like_compact,
        keyword,
        limit_per_source,
    ), fetchall=True) or []

    for row in history_rows:
        name = str(row.get("name") or "").strip()
        phone = str(row.get("phone") or "").strip()

        duplicate = any(
            item.get("name") == name
            and (
                not phone
                or not item.get("phone")
                or "".join(ch for ch in item.get("phone", "") if ch.isdigit())
                   == "".join(ch for ch in phone if ch.isdigit())
            )
            for item in results
        )
        if duplicate:
            continue

        last_date = row.get("last_date")
        add_result({
            "source": "history",
            "source_label": "曾布施佛友",
            "name": name,
            "english_name": "",
            "member_id": "",
            "volunteer_id": "",
            "phone": phone,
            "last_date": last_date.isoformat() if last_date else "",
            "times": int(row.get("times") or 0),
        })

    # 不再让月费会员先占满总名额。
    # 三个来源各自最多 limit_per_source 笔，确保义工和历史布施人都会出现。
    return results


@finance_bp.route("/api/donor-search")
def finance_donor_search():
    keyword = request.args.get("q", "").strip()
    branch = request.args.get("branch", "CHE").strip().upper()

    if len(keyword) < 1:
        return jsonify({"ok": True, "results": []})

    try:
        results = search_finance_donors(keyword, branch, 20)
        return jsonify({"ok": True, "results": results})
    except Exception as exc:
        print("donor search error:", exc)
        return jsonify({
            "ok": False,
            "message": "搜索布施人失败",
            "results": [],
        }), 500


@finance_bp.route("/income_batch/<category>", methods=["GET", "POST"])
def income_batch(category):

    # 普通 Key In 与 V7 负责人共用同一套成熟录入表单。
    # 只有已通过负责人密码时，source=v7 才会生效。
    source = (
        request.args.get("source")
        or request.form.get("source")
        or ""
    ).strip().lower()

    is_v7_source = (
        source == "v7"
        and bool(session.get("finance_admin"))
    )

    if is_v7_source:
        return_url = url_for("finance_v7.finance_v7_hq_home")
        return_label = "返回 V7 负责人中心"
        success_url = url_for("finance_v7.finance_v7_hq_home")
    else:
        return_url = url_for("finance.finance_income_menu")
        return_label = "返回收入录入"
        success_url = url_for("finance.records")

    allowed_categories = [
        "财布施",
        "观音村",
        "膳食结缘",
        "观音堂纯檀香布施",
        SPECIAL_DONATION_TITLE,
        "临时特别布施"
    ]

    if category not in allowed_categories:
        return "Invalid category", 400

    message = ""
    preview_rows = []

    raw_text = request.form.get("raw_text", "").strip()

    receipt_category = normalize_receipt_category(category)

    next_receipt_no = get_next_receipt_no(
        "CHE",
        receipt_category,
    )

    current_book_no = get_current_receipt_book(
        "CHE",
        receipt_category,
    )

    next_receipt_raw = next_receipt_no.replace("CHE", "", 1)

    receipt_start_raw = (
        request.form.get("receipt_start", "")
        .strip()
        .upper()
    )

    if not receipt_start_raw:
        receipt_start_raw = next_receipt_raw

    if receipt_start_raw.isdigit():
        receipt_start = (
            "CHE"
            + str(int(receipt_start_raw)).zfill(7)
        )
    else:
        receipt_start = receipt_start_raw

    receipt_date = request.form.get("receipt_date") or date.today().isoformat()
    record_date = receipt_date
    payment_method = request.form.get("payment_method", "现金")
    default_amount = money(request.form.get("default_amount") or 50)
    remarks = request.form.get("remarks", "").strip()
    action = request.form.get("action", "")

    def make_receipt_no(start_no, index):
        m = re.match(r"^([A-Z]+)(\d+)$", start_no)
        if not m:
            return ""

        prefix = m.group(1)
        num = m.group(2)

        return prefix + str(int(num) + index).zfill(len(num))

    def find_donor_info(keyword):
        keyword = str(keyword or "").strip()
        matches = search_finance_donors(keyword, "CHE", 10)

        if matches:
            selected = matches[0]
            return {
                "name": selected.get("name") or keyword,
                "phone": selected.get("phone") or "",
                "source": selected.get("source_label") or "资料库",
            }

        return {
            "name": keyword,
            "phone": "",
            "source": "手动输入",
        }

    def build_preview():
        rows = []
        valid_index = 0
        current_receipt_date = receipt_date

        for line in raw_text.splitlines():
            line = line.strip()

            if not line:
                continue

            # 日期分组格式：
            # @2026-07-10
            # 后续记录都会使用这个收条日期，
            # 直到遇到下一个 @日期。
            if line.startswith("@") or line.startswith("＠"):
                date_text = line[1:].strip()

                try:
                    current_receipt_date = date.fromisoformat(
                        date_text
                    ).isoformat()
                except ValueError:
                    rows.append({
                        "error": (
                            f"日期格式错误：{date_text}，"
                            "请使用 @YYYY-MM-DD"
                        ),
                        "receipt_no": "",
                        "receipt_date": date_text,
                        "name": "",
                        "phone": "",
                        "amount": 0,
                        "source": "日期设置",
                    })

                continue

            parts = line.split()

            if len(parts) >= 2:
                amount = money(parts[-1])
                name = " ".join(parts[:-1]).strip()
            else:
                amount = default_amount
                name = parts[0].strip()

            receipt_no = make_receipt_no(
                receipt_start,
                valid_index
            )
            valid_index += 1

            donor = find_donor_info(name)

            name = donor["name"]
            phone = donor["phone"]
            source = donor["source"]

            existing = db_query("""
                select id
                from finance_records
                where receipt_no = %s
                limit 1
            """, (receipt_no,), fetchone=True)

            error = ""

            month_lock_error = require_finance_month_open(
                current_receipt_date,
                get_fund_account(category)
            )

            if month_lock_error:
                error = month_lock_error
            elif not name:
                error = "姓名不能为空"
            elif amount <= 0:
                error = "金额必须大过 0"
            elif existing:
                error = "收条已存在"

            rows.append({
                "error": error,
                "receipt_no": receipt_no,
                "receipt_date": current_receipt_date,
                "name": name,
                "phone": phone,
                "amount": amount,
                "source": source,
            })

        return rows

    recent_donors = db_query("""
        select
            name,
            max(phone) as phone,
            count(*) as times,
            max(record_date) as last_date
        from finance_records
        where category <> '月费'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and coalesce(name, '') <> ''
        group by name
        order by max(record_date) desc, count(*) desc
        limit 50
    """, fetchall=True)

    if request.method == "POST":

        receipt_book_no = (
            request.form.get("receipt_book_no")
            or current_book_no
        )

        try:
            receipt_book_no = int(receipt_book_no)
        except (TypeError, ValueError):
            receipt_book_no = current_book_no

        if receipt_book_no != current_book_no:
            update_receipt_book_no(
                "CHE",
                receipt_category,
                receipt_book_no,
            )
            current_book_no = receipt_book_no

        if category == "临时特别布施" and not remarks:
            message = "临时特别布施请填写活动名称或用途"

        elif not receipt_start:
            message = "请填写收条开始号码"

        elif not re.match(r"^CHE\d+$", receipt_start):
            message = "收条号码格式错误，例如：CHE0001501，或只输入 1501"

        elif not raw_text:
            message = "请粘贴批量资料"

        else:
            preview_rows = build_preview()
            has_error = any(r.get("error") for r in preview_rows)

            if action == "confirm" and not has_error:

                confirm_lock_error = None

                for r in preview_rows:
                    confirm_lock_error = require_finance_month_open(
                        r["receipt_date"],
                        get_fund_account(category)
                    )

                    if confirm_lock_error:
                        break

                if confirm_lock_error:
                    message = confirm_lock_error

                else:
                    last_receipt_no = ""

                    for r in preview_rows:

                        db_query("""
                            insert into finance_records
                            (
                                record_type,
                                fund_account,
                                record_date,
                                receipt_date,
                                category,
                                receipt_no,
                                receipt_book_no,
                                name,
                                phone,
                                amount,
                                payment_method,
                                remarks
                            )
                            values
                            (
                                %s, %s, %s, %s, %s,
                                %s, %s, %s, %s, %s, %s, %s
                            )
                        """, (
                            "income",
                            get_fund_account(category),
                            r["receipt_date"],
                            r["receipt_date"],
                            category,
                            r["receipt_no"],
                            receipt_book_no,
                            r["name"],
                            r["phone"],
                            r["amount"],
                            payment_method,
                            remarks or "批量布施录入"
                        ))

                        last_receipt_no = r["receipt_no"]

                    if last_receipt_no:
                        update_receipt_book_number(
                            "CHE",
                            receipt_category,
                            last_receipt_no,
                        )

                    return redirect(success_url)

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>{{ category }}录入</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .finance-form-page{
                max-width:900px;
            }

            .finance-header-card{
                margin-bottom:18px;
                background:linear-gradient(135deg,#1877b8,#2695cf);
                color:white;
            }

            .finance-header-card .page-title,
            .finance-header-card .page-subtitle{
                color:white;
            }

            .finance-back-row{
                margin-bottom:16px;
            }

            .finance-form-grid{
                display:grid;
                grid-template-columns:repeat(2,minmax(0,1fr));
                gap:16px;
            }

            .finance-form-grid .form-group{
                margin:0;
            }

            .finance-full{
                grid-column:1 / -1;
            }

            .receipt-input-row{
                display:flex;
                align-items:center;
                gap:10px;
            }

            .receipt-prefix{
                min-height:52px;
                display:flex;
                align-items:center;
                justify-content:center;
                padding:0 16px;
                border-radius:10px;
                background:#eaf4ff;
                color:#1769aa;
                font-weight:800;
                font-size:20px;
            }

            .receipt-input-row .form-input{
                flex:1;
                min-width:0;
            }

            .field-help{
                display:block;
                color:#667085;
                font-size:15px;
                line-height:1.5;
                margin-top:7px;
            }

            .batch-textarea{
                width:100%;
                min-height:260px;
                resize:vertical;
                line-height:1.7;
                font-family:inherit;
            }

            .batch-example{
                background:#f8fafc;
                border:1px dashed #cbd5e1;
                border-radius:12px;
                padding:14px 16px;
                margin-top:12px;
                color:#475467;
                font-size:15px;
                line-height:1.75;
            }

            .batch-example code{
                font-family:Consolas,monospace;
                font-weight:700;
                color:#1769aa;
            }

            .action-row{
                display:flex;
                flex-wrap:wrap;
                gap:12px;
                margin-top:20px;
            }

            .action-row .btn-tool{
                min-width:160px;
            }

            .preview-success{
                color:#16863a;
                font-weight:700;
            }

            .preview-error{
                color:#c62828;
                font-weight:700;
            }

            .donor-buttons{
                display:flex;
                flex-wrap:wrap;
                gap:9px;
                margin:14px 0;
            }

            .donor-chip{
                border:1px solid #cbd5e1;
                background:#fff;
                border-radius:999px;
                padding:9px 14px;
                font-size:16px;
                cursor:pointer;
            }

            .donor-chip:hover{
                background:#eef6ff;
                border-color:#69a8df;
            }

            .recent-donor-details{
                margin-top:18px;
            }

            .recent-donor-details summary{
                cursor:pointer;
                font-weight:800;
                font-size:18px;
                padding:12px 0;
            }

            .alert{
                margin-bottom:16px;
            }
                                  
            .quick-amount-row{
                display:grid;
                grid-template-columns:58px 1fr 58px;
                gap:10px;
                align-items:stretch;
            }

            .amount-step-btn{
                border:0;
                border-radius:10px;
                background:#e8eef7;
                color:#172033;
                font-size:30px;
                font-weight:800;
                cursor:pointer;
            }

            .amount-step-btn:hover{
                background:#d8e4f4;
            }

            .amount-step-btn:active{
                transform:scale(.96);
            }

            .quick-amount-row .form-input{
                text-align:center;
                font-size:30px;
                font-weight:800;
            }

            .donor-search-box{
                position:relative;
                margin-bottom:14px;
            }

            .donor-search-results{
                display:none;
                position:absolute;
                left:0;
                right:0;
                top:calc(100% + 6px);
                z-index:50;
                max-height:360px;
                overflow:auto;
                background:#fff;
                border:1px solid #cbd5e1;
                border-radius:12px;
                box-shadow:0 12px 30px rgba(15,23,42,.16);
            }

            .donor-search-results.show{
                display:block;
            }

            .donor-result-item{
                width:100%;
                border:0;
                border-bottom:1px solid #eef2f6;
                background:#fff;
                padding:13px 15px;
                text-align:left;
                cursor:pointer;
            }

            .donor-result-item:last-child{
                border-bottom:0;
            }

            .donor-result-item:hover,
            .donor-result-item.active{
                background:#eef6ff;
            }

            .donor-result-main{
                display:flex;
                gap:9px;
                align-items:center;
                flex-wrap:wrap;
                font-size:18px;
                font-weight:800;
                color:#172033;
            }

            .donor-source-badge{
                display:inline-flex;
                align-items:center;
                border-radius:999px;
                padding:3px 9px;
                font-size:13px;
                background:#eaf4ff;
                color:#1769aa;
            }

            .donor-result-meta{
                margin-top:5px;
                color:#667085;
                font-size:14px;
                line-height:1.5;
            }

            .donor-search-empty{
                padding:15px;
                color:#667085;
                text-align:center;
            }

            .smart-entry-grid{
                display:grid;
                grid-template-columns:minmax(0,1.15fr) minmax(300px,.85fr);
                gap:18px;
                align-items:start;
            }

            .selected-donor-card{
                min-height:220px;
                border:1px solid #d8e2ee;
                border-radius:16px;
                background:linear-gradient(180deg,#fbfdff,#f4f8fc);
                padding:18px;
            }

            .selected-donor-empty{
                min-height:180px;
                display:flex;
                align-items:center;
                justify-content:center;
                text-align:center;
                color:#667085;
                line-height:1.7;
            }

            .selected-donor-name{
                font-size:25px;
                font-weight:900;
                color:#172033;
                margin-bottom:4px;
            }

            .selected-donor-english{
                color:#667085;
                margin-bottom:12px;
            }

            .selected-donor-badges{
                display:flex;
                flex-wrap:wrap;
                gap:8px;
                margin-bottom:14px;
            }

            .selected-donor-badge{
                display:inline-flex;
                align-items:center;
                border-radius:999px;
                padding:5px 10px;
                background:#eaf4ff;
                color:#1769aa;
                font-size:14px;
                font-weight:800;
            }

            .selected-donor-info{
                display:grid;
                grid-template-columns:110px 1fr;
                gap:8px 12px;
                font-size:15px;
                line-height:1.5;
                margin-bottom:16px;
            }

            .selected-donor-info .label{
                color:#667085;
            }

            .selected-donor-info .value{
                color:#172033;
                font-weight:700;
                word-break:break-word;
            }

            .selected-donor-actions{
                display:grid;
                gap:10px;
            }

            .selected-donor-actions .btn-tool{
                width:100%;
            }

            @media(max-width:700px){
                .finance-form-grid,
                .smart-entry-grid{
                    grid-template-columns:1fr;
                }

                .finance-full{
                    grid-column:auto;
                }

                .receipt-input-row{
                    align-items:stretch;
                }

                .action-row{
                    display:grid;
                }

                .action-row .btn-tool{
                    width:100%;
                }
            }
        </style>
    </head>

    <body>

    <div class="page finance-form-page">

        <div class="finance-back-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ return_url }}"
            >
                ← {{ return_label }}
            </a>
        </div>

        <div class="card finance-header-card">

            <h1 class="page-title">
                💵 {{ category }}
            </h1>

            <p class="page-subtitle">
                可批量录入多笔，并在同一批中使用不同收条日期
            </p>

        </div>

        {% if message %}
            <div class="alert alert-danger">
                {{ message }}
            </div>
        {% endif %}

        <form method="post">

            <input
                type="hidden"
                name="source"
                value="{{ entry_source }}"
            >

            <div class="card">

                <h2 class="section-title">
                    收条与付款资料
                </h2>

                <div class="form-group finance-full">

                    <label class="form-label">
                        📒 收条簿编号
                    </label>

                    <div style="display:flex;gap:12px;align-items:center;">

                        <input
                            class="form-input"
                            type="number"
                            min="1"
                            name="receipt_book_no"
                            value="{{ current_book_no }}"
                            style="max-width:140px;"
                        >

                        <span class="field-help" style="margin-top:0;">
                            开始使用新的实体收条簿时才需要修改。
                        </span>

                    </div>

                </div>

                <div class="finance-form-grid">

                    <div class="form-group">

                        <label class="form-label">
                            收条开始号码
                        </label>

                        <div class="receipt-input-row">

                            <div class="receipt-prefix">
                                CHE
                            </div>

                            <input
                                class="form-input"
                                name="receipt_start"
                                value="{{ receipt_start_raw }}"
                                inputmode="numeric"
                                placeholder="例如 1501"
                                required
                            >

                        </div>

                        <span class="field-help">
                            系统建议下一张收条：
                            <strong>{{ next_receipt_no }}</strong>
                        </span>

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            默认开收条日期
                        </label>

                        <input
                            class="form-input"
                            name="receipt_date"
                            type="date"
                            value="{{ receipt_date }}"
                            required
                        >

                        <span class="field-help">
                            没有写 @日期 的记录，会使用这个日期。
                        </span>

                    </div>

                    <div class="quick-amount-row">

                        <button
                            type="button"
                            class="amount-step-btn"
                            onclick="changeQuickAmount(-50)"
                        >
                            −
                        </button>

                        <input
                            class="form-input"
                            id="quick_amount"
                            name="default_amount"
                            type="number"
                            step="50"
                            min="50"
                            value="{{ default_amount }}"
                            placeholder="金额"
                        >

                        <button
                            type="button"
                            class="amount-step-btn"
                            onclick="changeQuickAmount(50)"
                        >
                            ＋
                        </button>

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            付款方式
                        </label>

                        <select
                            class="form-input"
                            name="payment_method"
                        >
                            <option
                                {% if payment_method == '现金' %}
                                    selected
                                {% endif %}
                            >
                                现金
                            </option>

                            <option
                                {% if payment_method == '银行过账' %}
                                    selected
                                {% endif %}
                            >
                                银行过账
                            </option>

                            <option
                                {% if payment_method == '支票' %}
                                    selected
                                {% endif %}
                            >
                                支票
                            </option>
                        </select>

                    </div>

                    <div class="form-group finance-full">

                        <label class="form-label">
                            备注／活动名称
                        </label>

                        <input
                            class="form-input"
                            name="remarks"
                            value="{{ remarks }}"
                            placeholder="{% if category == '临时特别布施' %}临时特别布施必须填写活动名称{% else %}例如：观音诞、法会、特别活动{% endif %}"
                        >

                        {% if category == "临时特别布施" %}
                            <span class="field-help">
                                此项目必须填写活动名称或用途。
                            </span>
                        {% endif %}

                    </div>

                </div>

            </div>

            <div class="card">

                <h2 class="section-title">
                    批量输入
                </h2>

                <p class="page-subtitle">
                    可用 @日期 分组；只输入姓名会使用默认金额，
                    姓名后面加金额可单独修改。
                </p>

                <div class="smart-entry-grid">

                    <div class="donor-search-box">

                        <label class="form-label">
                            🔎 快速找布施人
                        </label>

                        <input
                            class="form-input"
                            id="donor_search_input"
                            type="search"
                            autocomplete="off"
                            placeholder="输入姓名、英文名、会员／义工编号或电话号码"
                        >

                        <span class="field-help">
                            点选搜索结果后，先检查右边资料，再加入批量输入。
                        </span>

                        <div
                            class="donor-search-results"
                            id="donor_search_results"
                        ></div>

                    </div>

                    <div
                        class="selected-donor-card"
                        id="selected_donor_card"
                    >
                        <div class="selected-donor-empty">
                            先搜索并点选布施人。<br>
                            系统会在这里显示电话、编号和最近布施资料。
                        </div>
                    </div>

                </div>

                <textarea
                    class="form-input batch-textarea"
                    name="raw_text"
                    placeholder="例如：
@2026-07-10
王小明
李大华
陈美玲 100

@2026-07-11
郑依颖 30
林美珍 50"
                    required
                >{{ raw_text }}</textarea>

                <div class="batch-example">

                    <strong>输入规则：</strong><br>

                    <code>@2026-07-10</code>
                    → 后续记录使用 2026-07-10<br>

                    <code>王小明</code>
                    → 使用默认金额<br>

                    <code>陈美玲 100</code>
                    → 金额 RM100<br>

                    没有写 <code>@日期</code> 的记录，
                    会使用上面的默认开收条日期。

                </div>

                <div class="action-row">

                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                        name="action"
                        value="preview"
                    >
                        👁️ 预览资料
                    </button>

                    {% if preview_rows %}
                        <button
                            class="btn-tool btn-success"
                            type="submit"
                            name="action"
                            value="confirm"
                            onclick="return confirm('确定全部入账？');"
                        >
                            ✅ 确认全部入账
                        </button>
                    {% endif %}

                </div>

            </div>

            {% if preview_rows %}

                <div class="card">

                    <h2 class="section-title">
                        录入预览
                    </h2>

                    <div class="table-responsive">

                        <table class="record-table">

                            <thead>
                                <tr>
                                    <th>收条</th>
                                    <th>收条日期</th>
                                    <th>姓名</th>
                                    <th>电话</th>
                                    <th>来源</th>
                                    <th>金额</th>
                                    <th>检查结果</th>
                                </tr>
                            </thead>

                            <tbody>

                                {% for r in preview_rows %}
                                <tr>

                                    <td>{{ r.receipt_no or "-" }}</td>

                                    <td>
                                        {{ r.receipt_date or "-" }}
                                    </td>

                                    <td>
                                        <strong>{{ r.name or "-" }}</strong>
                                    </td>

                                    <td>{{ r.phone or "-" }}</td>

                                    <td>{{ r.source or "-" }}</td>

                                    <td>
                                        RM {{ "%.2f"|format(r.amount or 0) }}
                                    </td>

                                    <td>
                                        {% if r.error %}
                                            <span class="preview-error">
                                                ❌ {{ r.error }}
                                            </span>
                                        {% else %}
                                            <span class="preview-success">
                                                ✅ 可以入账
                                            </span>
                                        {% endif %}
                                    </td>

                                </tr>
                                {% endfor %}

                            </tbody>

                        </table>

                    </div>

                </div>

            {% endif %}

        </form>

        {% if recent_donors %}

            <div class="card">

                <h2 class="section-title">
                    常用捐赠者
                </h2>

                <p class="page-subtitle">
                    点一下姓名，会自动加入批量输入框。
                </p>

                <div class="donor-buttons">

                    {% for d in recent_donors %}

                        <button
                            class="donor-chip"
                            type="button"
                            data-donor-name="{{ d.name }}"
                        >
                            {{ d.name }}
                        </button>

                    {% endfor %}

                </div>

                <details class="recent-donor-details">

                    <summary>
                        查看捐赠者详细记录
                    </summary>

                    <div class="table-responsive">

                        <table class="record-table">

                            <thead>
                                <tr>
                                    <th>姓名</th>
                                    <th>电话</th>
                                    <th>次数</th>
                                    <th>最后日期</th>
                                </tr>
                            </thead>

                            <tbody>

                                {% for d in recent_donors %}
                                <tr>
                                    <td>{{ d.name }}</td>
                                    <td>{{ d.phone or "-" }}</td>
                                    <td>{{ d.times }}</td>
                                    <td>{{ d.last_date }}</td>
                                </tr>
                                {% endfor %}

                            </tbody>

                        </table>

                    </div>

                </details>

            </div>

        {% endif %}

    </div>

    <script>
const donorSearchInput =
    document.getElementById("donor_search_input");

const donorSearchResults =
    document.getElementById("donor_search_results");

const selectedDonorCard =
    document.getElementById("selected_donor_card");

let selectedDonor = null;
let donorSearchTimer = null;
let donorSearchController = null;


function addDonor(name){

    const textarea = document.querySelector(
        'textarea[name="raw_text"]'
    );

    if(!textarea){
        console.error("找不到 raw_text 输入框");
        return;
    }

    const cleanName = String(name || "").trim();

    if(!cleanName){
        return;
    }

    if(textarea.value.trim() === ""){
        textarea.value = cleanName;
    }else{
        textarea.value += "\\n" + cleanName;
    }

    textarea.focus();
    textarea.scrollTop = textarea.scrollHeight;
}


document.querySelectorAll(".donor-chip").forEach(function(button){

    button.addEventListener("click", function(){

        addDonor(
            button.dataset.donorName || ""
        );

    });

});


function escapeHtml(value){

    return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
}


function donorIcon(source){

    if(source === "member"){
        return "👤";
    }

    if(source === "volunteer"){
        return "👷";
    }

    return "💖";
}


function closeDonorResults(){

    if(donorSearchResults){

        donorSearchResults.classList.remove("show");
        donorSearchResults.innerHTML = "";

    }
}


function renderSelectedDonor(result){

    if(!selectedDonorCard){
        return;
    }

    if(!result){
        selectedDonorCard.innerHTML = `
            <div class="selected-donor-empty">
                先搜索并点选布施人。<br>
                系统会在这里显示电话、编号和最近布施资料。
            </div>
        `;
        return;
    }

    const memberId = result.member_id || "";
    const volunteerId = result.volunteer_id || "";
    const phone = result.phone || "-";
    const lastCategory = result.last_category || "-";
    const lastAmount = Number(result.last_amount || 0);
    const lastDate = result.last_record_date || "-";

    const badges = [];

    if(memberId){
        badges.push(`<span class="selected-donor-badge">👤 月费会员</span>`);
    }

    if(volunteerId){
        badges.push(`<span class="selected-donor-badge">🤝 义工</span>`);
    }

    if(!memberId && !volunteerId){
        badges.push(`<span class="selected-donor-badge">💖 历史布施人</span>`);
    }

    selectedDonorCard.innerHTML = `
        <div class="selected-donor-name">
            ${escapeHtml(result.name || "-")}
        </div>

        <div class="selected-donor-english">
            ${escapeHtml(result.english_name || "")}
        </div>

        <div class="selected-donor-badges">
            ${badges.join("")}
        </div>

        <div class="selected-donor-info">
            <div class="label">电话</div>
            <div class="value">${escapeHtml(phone)}</div>

            <div class="label">会员编号</div>
            <div class="value">${escapeHtml(memberId || "-")}</div>

            <div class="label">义工编号</div>
            <div class="value">${escapeHtml(volunteerId || "-")}</div>

            <div class="label">最近布施</div>
            <div class="value">
                ${escapeHtml(lastCategory)}
                ${lastAmount > 0 ? ` · RM ${lastAmount.toFixed(2)}` : ""}
                ${lastDate !== "-" ? ` · ${escapeHtml(lastDate)}` : ""}
            </div>
        </div>

        <div class="selected-donor-actions">
            <button
                type="button"
                class="btn-tool btn-success"
                id="add_selected_donor_btn"
            >
                ➕ 加入批量输入
            </button>

            ${lastAmount > 0 ? `
                <button
                    type="button"
                    class="btn-tool btn-secondary"
                    id="use_last_amount_btn"
                >
                    使用上次金额 RM ${lastAmount.toFixed(2)}
                </button>
            ` : ""}
        </div>
    `;

    const addButton = document.getElementById("add_selected_donor_btn");

    if(addButton){
        addButton.addEventListener("click", function(){
            addDonor(result.name || "");
            selectedDonor = null;
            renderSelectedDonor(null);

            if(donorSearchInput){
                donorSearchInput.value = "";
                donorSearchInput.focus();
            }
        });
    }

    const amountButton = document.getElementById("use_last_amount_btn");

    if(amountButton){
        amountButton.addEventListener("click", function(){
            const amountInput = document.getElementById("quick_amount");
            if(amountInput){
                amountInput.value = lastAmount;
            }
        });
    }
}


function selectDonor(result){

    selectedDonor = result || null;
    renderSelectedDonor(selectedDonor);
    closeDonorResults();

    if(donorSearchInput){
        donorSearchInput.focus();
    }
}


function renderDonorResults(results){

    if(!donorSearchResults){
        return;
    }

    if(!Array.isArray(results) || results.length === 0){

        donorSearchResults.innerHTML = `
            <div class="donor-search-empty">
                找不到记录，可直接在下面手动输入新姓名。
            </div>
        `;

        donorSearchResults.classList.add("show");
        return;
    }

    donorSearchResults.innerHTML = results.map((r, index) => {

        const code =
            r.member_id
            || r.volunteer_id
            || "";

        const englishName = r.english_name
            ? ` · ${escapeHtml(r.english_name)}`
            : "";

        const phone = r.phone
            ? `电话：${escapeHtml(r.phone)}`
            : "电话：-";

        const lastAmount = Number(r.last_amount || 0);

        const history = r.last_category
            ? ` · 最近：${escapeHtml(r.last_category)}${lastAmount > 0 ? ` RM ${lastAmount.toFixed(2)}` : ""}${r.last_record_date ? ` · ${escapeHtml(r.last_record_date)}` : ""}`
            : "";

        return `
            <button
                type="button"
                class="donor-result-item"
                data-index="${index}"
            >
                <div class="donor-result-main">

                    <span>
                        ${donorIcon(r.source)}
                    </span>

                    <span>
                        ${escapeHtml(r.name)}
                    </span>

                    ${
                        code
                        ? `<span>${escapeHtml(code)}</span>`
                        : ""
                    }

                    <span class="donor-source-badge">
                        ${escapeHtml(
                            r.source_label || "资料库"
                        )}
                    </span>

                </div>

                <div class="donor-result-meta">
                    ${phone}${englishName}${history}
                </div>

            </button>
        `;

    }).join("");

    donorSearchResults
        .querySelectorAll(".donor-result-item")
        .forEach(function(button){

            button.addEventListener("click", function(){

                const index = Number(
                    button.dataset.index
                );

                selectDonor(
                    results[index]
                );

            });

        });

    donorSearchResults.classList.add("show");
}


async function runDonorSearch(keyword){

    if(!keyword){

        closeDonorResults();
        return;

    }

    if(donorSearchController){

        donorSearchController.abort();

    }

    donorSearchController = new AbortController();

    try{

        const url = new URL(
            "/finance/api/donor-search",
            window.location.origin
        );

        url.searchParams.set("q", keyword);
        url.searchParams.set("branch", "CHE");

        const response = await fetch(url, {

            signal: donorSearchController.signal,
            headers: {
                "Accept": "application/json"
            }

        });

        const data = await response.json();

        if(!response.ok || !data.ok){

            throw new Error(
                data.message || "搜索失败"
            );

        }

        renderDonorResults(
            data.results || []
        );

    }catch(error){

        if(error.name === "AbortError"){
            return;
        }

        console.error(
            "布施人搜索失败：",
            error
        );

        if(donorSearchResults){

            donorSearchResults.innerHTML = `
                <div class="donor-search-empty">
                    搜索暂时失败，请直接输入姓名。
                </div>
            `;

            donorSearchResults.classList.add("show");

        }
    }
}


if(donorSearchInput){

    donorSearchInput.addEventListener(
        "input",
        function(){

            const keyword =
                donorSearchInput.value.trim();

            clearTimeout(
                donorSearchTimer
            );

            if(!keyword){

                closeDonorResults();
                return;

            }

            donorSearchTimer = setTimeout(
                function(){

                    runDonorSearch(keyword);

                },
                220
            );

        }
    );
}


document.addEventListener(
    "click",
    function(event){

        if(
            donorSearchInput
            && donorSearchResults
            && !donorSearchInput.contains(event.target)
            && !donorSearchResults.contains(event.target)
        ){
            closeDonorResults();
        }

    }
);


function changeQuickAmount(change){

    const input = document.getElementById(
        "quick_amount"
    );

    if(!input){
        return;
    }

    let current = Number(
        input.value
    );

    if(!Number.isFinite(current)){
        current = 50;
    }

    current += change;

    if(current < 50){
        current = 50;
    }

    input.value = current;
}
</script>

    </body>
    </html>
    """,
        category=category,
        message=message,
        raw_text=raw_text,
        receipt_start_raw=receipt_start_raw,
        next_receipt_no=next_receipt_no,
        receipt_date=receipt_date,
        payment_method=payment_method,
        default_amount=default_amount,
        current_book_no=current_book_no,
        remarks=remarks,
        preview_rows=preview_rows,
        recent_donors=recent_donors,
        entry_source="v7" if is_v7_source else "",
        return_url=return_url,
        return_label=return_label
        )

@finance_bp.route("/donors")
@finance_admin_required
def donor_management():

    q = request.args.get("q", "").strip()
    missing_only = request.args.get("missing", "") == "1"

    where_parts = [
        "record_type = 'income'",
        "category <> '月费'",
        "coalesce(status, 'confirmed') <> 'cancelled'",
        "coalesce(trim(name), '') <> ''",
    ]
    params = []

    if q:
        where_parts.append(
            "(name ilike %s or regexp_replace(coalesce(phone, ''), '[^0-9]', '', 'g') ilike %s)"
        )
        params.extend([
            f"%{q}%",
            f"%{normalize_phone(q) or q}%",
        ])

    if missing_only:
        where_parts.append("coalesce(trim(phone), '') = ''")

    rows = db_query(f"""
        select
            name,
            (array_agg(nullif(trim(phone), '') order by record_date desc, id desc)
                filter (where coalesce(trim(phone), '') <> ''))[1] as phone,
            count(*) as times,
            min(record_date) as first_date,
            max(record_date) as last_date,
            coalesce(sum(amount), 0) as total_amount
        from finance_records
        where {' and '.join(where_parts)}
        group by name
        order by
            case
                when (
                    array_agg(nullif(trim(phone), '') order by record_date desc, id desc)
                    filter (where coalesce(trim(phone), '') <> '')
                )[1] is null then 0
                else 1
            end,
            max(record_date) desc,
            name
        limit 500
    """, tuple(params), fetchall=True) or []

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>布施人管理</title>
        <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
        <style>
            .donor-page{max-width:1250px}
            .donor-header{background:linear-gradient(135deg,#7c3aed,#5b21b6);color:#fff;padding:28px;border-radius:22px;margin-bottom:20px}
            .donor-header h1{margin:0 0 8px}
            .donor-filter{display:grid;grid-template-columns:1fr auto auto;gap:10px;align-items:end}
            .donor-table{min-width:1050px}
            .phone-form{display:flex;gap:8px;align-items:center}
            .phone-form .form-input{min-width:180px;margin:0}
            .missing-phone{background:#fff7ed}
            @media(max-width:700px){.donor-filter{grid-template-columns:1fr}.donor-filter .btn-tool{width:100%}}
        </style>
    </head>
    <body>
    <div class="page donor-page">
        <div class="donor-header">
            <h1>👥 布施人管理</h1>
            <p>补充电话号码后，系统会更新这个姓名的全部历史布施记录。</p>
        </div>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% for category_name, text in messages %}
                <div class="alert alert-{{ 'danger' if category_name == 'danger' else 'success' }}">{{ text }}</div>
            {% endfor %}
        {% endwith %}

        <div class="card">
            <form method="get" class="donor-filter">
                <div class="form-group">
                    <label class="form-label">搜索姓名或电话</label>
                    <input class="form-input" name="q" value="{{ q }}" placeholder="输入姓名或电话号码">
                </div>
                <label style="display:flex;align-items:center;gap:8px;min-height:50px">
                    <input type="checkbox" name="missing" value="1" {% if missing_only %}checked{% endif %}>
                    只看缺电话
                </label>
                <button class="btn-tool btn-primary" type="submit">🔍 查询</button>
            </form>
        </div>

        <div class="card">
            <div class="section-title">📋 布施人名单（{{ rows|length }} 位）</div>
            <div class="table-responsive">
                <table class="record-table donor-table">
                    <thead>
                        <tr><th>姓名</th><th>电话</th><th>次数</th><th>累计</th><th>第一次</th><th>最后一次</th><th>保存</th></tr>
                    </thead>
                    <tbody>
                    {% for r in rows %}
                        <tr {% if not r.phone %}class="missing-phone"{% endif %}>
                            <td><strong>{{ r.name }}</strong></td>
                            <td>{{ r.phone or '-' }}</td>
                            <td>{{ r.times }}</td>
                            <td>RM {{ '%.2f'|format(r.total_amount or 0) }}</td>
                            <td>{{ r.first_date or '-' }}</td>
                            <td>{{ r.last_date or '-' }}</td>
                            <td>
                                <form class="phone-form" method="post" action="{{ url_for('finance.update_donor_phone') }}">
                                    <input type="hidden" name="name" value="{{ r.name }}">
                                    <input class="form-input" name="phone" value="{{ r.phone or '' }}" inputmode="tel" placeholder="0123456789" required>
                                    <button class="btn-tool btn-success" type="submit">💾 保存</button>
                                </form>
                            </td>
                        </tr>
                    {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>

        <a class="btn-tool btn-secondary" href="{{ url_for('finance.finance_admin_home') }}">← 返回负责人中心</a>
    </div>
    </body>
    </html>
    """, rows=rows, q=q, missing_only=missing_only)


@finance_bp.route("/donors/update-phone", methods=["POST"])
def update_donor_phone():

    name = request.form.get("name", "").strip()
    phone = normalize_phone(request.form.get("phone", ""))

    if not name or not phone:
        flash("姓名和电话号码不能为空。", "danger")
        return redirect(url_for("finance.donor_management"))

    db_query("""
        update finance_records
        set phone = %s
        where record_type = 'income'
          and trim(name) = trim(%s)
          and coalesce(status, 'confirmed') <> 'cancelled'
    """, (phone, name))

    db_query("""
        update members
        set phone = %s
        where trim(name) = trim(%s)
          and coalesce(trim(phone), '') = ''
    """, (phone, name))

    db_query("""
        update volunteers
        set phone = %s
        where trim(name) = trim(%s)
          and coalesce(trim(phone), '') = ''
    """, (phone, name))

    flash(f"已更新 {name} 的电话号码。", "success")
    return redirect(url_for("finance.donor_management", q=name))


@finance_bp.route("/bank_pending/<int:pending_id>/delete", methods=["POST"])
@finance_admin_required
def delete_bank_pending(pending_id):

    db_query("""
        delete from bank_pending_records
        where id = %s
        and status = 'pending'
    """, (pending_id,))

    return redirect(url_for("finance.bank_pending"))


def ensure_finance_bank_in_tables():
    """建立月费现金 Bank In 所需资料表；可重复安全执行。"""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                create table if not exists finance_bank_deposits (
                    id bigserial primary key,
                    branch varchar(10) not null default 'CHE',
                    deposit_date date not null,
                    reference_no varchar(120),
                    amount numeric(14,2) not null default 0,
                    remarks text,
                    created_at timestamp not null default now()
                )
            """)
            cur.execute("""
                create table if not exists finance_bank_deposit_items (
                    id bigserial primary key,
                    deposit_id bigint not null references finance_bank_deposits(id) on delete cascade,
                    finance_record_id bigint not null,
                    amount numeric(14,2) not null default 0,
                    created_at timestamp not null default now(),
                    unique(finance_record_id)
                )
            """)
        conn.commit()


@finance_bp.route(
    "/admin/monthly-fee-bank-in",
    methods=["GET", "POST"]
)
@finance_admin_required
def finance_monthly_fee_bank_in():
    """Legacy compatibility redirect. Formal page moved to finance_v7.py."""
    return redirect(url_for("finance_v7.finance_v7_cdm_management"))

# ============================================================
# Finance V7：CHE 月费 Bank In 修复中心（正式版）
# 适配当前结构：
# - finance_records 没有 branch 栏位
# - CHE / STW 依 member_id、receipt_no 判断
# - finance_bank_deposit_items 使用 finance_record_id 唯一
# ============================================================

OFFICIAL_CHE_CDM_ROWS = [
    # ym, CDM编号, Bank In日期, 金额, 收条起号, 收条尾号
    ("2026-01", 1, "2026-01-06", Decimal("1800.00"), "CHE0001292", "CHE0001300"),
    ("2026-01", 2, "2026-01-06", Decimal("2700.00"), "CHE0001401", "CHE0001407"),
    ("2026-01", 3, "2026-01-14", Decimal("2000.00"), "CHE0001408", "CHE0001412"),
    ("2026-01", 4, "2026-01-26", Decimal("1150.00"), "CHE0001413", "CHE0001417"),

    ("2026-02", 1, "2026-02-09", Decimal("1900.00"), "CHE0001425", "CHE0001431"),

    ("2026-03", 1, "2026-03-02", Decimal("2250.00"), "CHE0001432", "CHE0001440"),
    ("2026-03", 2, "2026-03-11", Decimal("2000.00"), "CHE0001441", "CHE0001448"),
    ("2026-03", 3, "2026-03-22", Decimal("1250.00"), "CHE0001449", "CHE0001453"),

    ("2026-04", 1, "2026-04-12", Decimal("2250.00"), "CHE0001454", "CHE0001462"),
    ("2026-04", 2, "2026-04-20", Decimal("1100.00"), "CHE0001463", "CHE0001468"),
    ("2026-04", 3, "2026-04-29", Decimal("450.00"), "CHE0001469", "CHE0001471"),

    ("2026-05", 1, "2026-05-04", Decimal("1900.00"), "CHE0001472", "CHE0001478"),
    ("2026-05", 2, "2026-05-18", Decimal("300.00"), "CHE0001479", "CHE0001480"),
    ("2026-05", 3, "2026-05-26", Decimal("800.00"), "CHE0001481", "CHE0001484"),
    ("2026-05", 4, "2026-05-29", Decimal("400.00"), "CHE0001485", "CHE0001487"),

    ("2026-06", 1, "2026-06-04", Decimal("900.00"), "CHE0001488", "CHE0001492"),
    ("2026-06", 2, "2026-06-18", Decimal("200.00"), "CHE0001493", "CHE0001494"),
    ("2026-06", 3, "2026-06-26", Decimal("300.00"), "CHE0001495", "CHE0001500"),
]


def _bank_in_receipt_number(receipt_no):
    """CHE0001454 / CHE1454 -> 1454。"""
    match = re.search(r"(\d+)$", str(receipt_no or "").strip())
    return int(match.group(1)) if match else None


def _bank_in_decimal(value):
    try:
        return Decimal(str(value or 0)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _find_che_cash_monthly_receipts(cur, receipt_from, receipt_to):
    """
    按 CHE 收条号码范围寻找现金月费。
    finance_records 没有 branch，因此以 receipt_no 的 CHE 前缀判断。
    """
    start_no = _bank_in_receipt_number(receipt_from)
    end_no = _bank_in_receipt_number(receipt_to)

    if start_no is None or end_no is None:
        raise ValueError(f"收条范围格式错误：{receipt_from} 至 {receipt_to}")

    cur.execute(
        """
        select
            r.id,
            r.record_date,
            r.receipt_date,
            r.receipt_no,
            r.member_id,
            r.name,
            r.amount,
            r.status,
            r.remarks
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and r.payment_method = '现金'
          and coalesce(r.status, 'confirmed') <> 'cancelled'
          and upper(coalesce(r.receipt_no, '')) like 'CHE%%'
          and nullif(
                regexp_replace(
                    coalesce(r.receipt_no, ''),
                    '[^0-9]',
                    '',
                    'g'
                ),
                ''
              )::bigint between %s and %s
        order by
            nullif(
                regexp_replace(
                    coalesce(r.receipt_no, ''),
                    '[^0-9]',
                    '',
                    'g'
                ),
                ''
            )::bigint,
            r.id
        """,
        (start_no, end_no),
    )
    return cur.fetchall() or []


def _get_bank_in_master_rows(cur):
    cur.execute(
        """
        select
            d.id,
            d.ym,
            d.cdm_sequence,
            d.deposit_date,
            d.amount,
            d.receipt_from,
            d.receipt_to,
            d.reference_no,
            d.branch
        from finance_bank_deposits d
        where d.branch = 'CHE'
          and d.ym between '2026-01' and '2026-06'
        order by d.ym, d.cdm_sequence, d.id
        """
    )
    return cur.fetchall() or []


def _scan_che_bank_in():
    """只扫描，不修改数据库。"""
    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            master_rows = _get_bank_in_master_rows(cur)

            master_by_key = {}
            duplicate_keys = set()

            for row in master_rows:
                key = (
                    str(row.get("ym") or ""),
                    int(row.get("cdm_sequence") or 0),
                )
                if key in master_by_key:
                    duplicate_keys.add(key)
                else:
                    master_by_key[key] = row

            official_keys = {
                (ym, sequence)
                for ym, sequence, *_ in OFFICIAL_CHE_CDM_ROWS
            }

            extra_master_rows = [
                row for row in master_rows
                if (
                    str(row.get("ym") or ""),
                    int(row.get("cdm_sequence") or 0),
                ) not in official_keys
            ]

            batches = []
            master_ok_count = 0
            validation_ok_count = 0

            for (
                ym,
                sequence,
                deposit_date_text,
                expected_amount,
                expected_from,
                expected_to,
            ) in OFFICIAL_CHE_CDM_ROWS:

                key = (ym, sequence)
                master = master_by_key.get(key)
                master_issues = []

                if not master:
                    master_issues.append("缺少 CDM 主记录")
                else:
                    actual_date = str(master.get("deposit_date") or "")
                    actual_amount = _bank_in_decimal(master.get("amount"))
                    actual_from = str(master.get("receipt_from") or "")
                    actual_to = str(master.get("receipt_to") or "")

                    if actual_date != deposit_date_text:
                        master_issues.append(
                            f"Bank In 日期应为 {deposit_date_text}"
                        )
                    if actual_amount != expected_amount:
                        master_issues.append(
                            f"CDM 金额应为 RM {expected_amount:,.2f}"
                        )
                    if (
                        actual_from != expected_from
                        or actual_to != expected_to
                    ):
                        master_issues.append(
                            f"收条范围应为 {expected_from} 至 {expected_to}"
                        )

                if key in duplicate_keys:
                    master_issues.append("同月份相同 CDM 编号有重复主记录")

                master_ok = not master_issues
                if master_ok:
                    master_ok_count += 1

                receipts = _find_che_cash_monthly_receipts(
                    cur,
                    expected_from,
                    expected_to,
                )

                receipt_total = sum(
                    (_bank_in_decimal(row.get("amount")) for row in receipts),
                    Decimal("0.00"),
                )

                validation_issues = []

                if not receipts:
                    validation_issues.append(
                        f"找不到收条 {expected_from} 至 {expected_to}"
                    )
                elif receipt_total != expected_amount:
                    validation_issues.append(
                        "金额不符："
                        f"CDM RM {expected_amount:,.2f}，"
                        f"收条合计 RM {receipt_total:,.2f}，"
                        f"共 {len(receipts)} 张"
                    )

                validation_ok = not validation_issues
                if validation_ok:
                    validation_ok_count += 1

                batches.append({
                    "ym": ym,
                    "sequence": sequence,
                    "deposit_date": deposit_date_text,
                    "expected_amount": expected_amount,
                    "expected_from": expected_from,
                    "expected_to": expected_to,
                    "master": master,
                    "master_issues": master_issues,
                    "master_ok": master_ok,
                    "receipts": receipts,
                    "receipt_total": receipt_total,
                    "validation_issues": validation_issues,
                    "validation_ok": validation_ok,
                })

            cur.execute(
                """
                select count(*) as count
                from finance_bank_deposit_items
                """
            )
            item_count = int((cur.fetchone() or {}).get("count") or 0)

            cur.execute(
                """
                select
                    count(*) as count,
                    coalesce(sum(r.amount), 0) as total
                from finance_records r
                where r.record_type = 'income'
                  and r.category = '月费'
                  and r.payment_method = '现金'
                  and coalesce(r.status, 'confirmed') <> 'cancelled'
                  and r.record_date >= date '2026-01-01'
                  and (
                        upper(coalesce(r.member_id, '')) like 'CHE-%%'
                     or upper(coalesce(r.receipt_no, '')) like 'CHE%%'
                  )
                  and not exists (
                        select 1
                        from finance_bank_deposit_items i
                        where i.finance_record_id = r.id
                  )
                """
            )
            waiting = cur.fetchone() or {}

    all_master_ok = master_ok_count == len(OFFICIAL_CHE_CDM_ROWS)
    all_validation_ok = validation_ok_count == len(OFFICIAL_CHE_CDM_ROWS)

    return {
        "batches": batches,
        "master_count": len(master_rows),
        "master_ok_count": master_ok_count,
        "validation_ok_count": validation_ok_count,
        "official_count": len(OFFICIAL_CHE_CDM_ROWS),
        "extra_master_rows": extra_master_rows,
        "all_master_ok": all_master_ok,
        "all_validation_ok": all_validation_ok,
        "can_rebuild": all_master_ok and all_validation_ok,
        "item_count": item_count,
        "waiting_count": int(waiting.get("count") or 0),
        "waiting_total": _bank_in_decimal(waiting.get("total")),
    }


def _repair_che_bank_in_master():
    """只修正 finance_bank_deposits，不碰明细和收条。"""
    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select *
                    from finance_bank_deposits
                    where branch = 'CHE'
                      and ym between '2026-01' and '2026-06'
                    order by ym, cdm_sequence, id
                    for update
                    """
                )
                rows = cur.fetchall() or []

                official_keys = {
                    (ym, sequence)
                    for ym, sequence, *_ in OFFICIAL_CHE_CDM_ROWS
                }

                keep_by_key = {}
                remove_ids = []

                for row in rows:
                    key = (
                        str(row.get("ym") or ""),
                        int(row.get("cdm_sequence") or 0),
                    )

                    if key not in official_keys or key in keep_by_key:
                        remove_ids.append(row["id"])
                    else:
                        keep_by_key[key] = row

                remove_ids = sorted(set(remove_ids))

                if remove_ids:
                    cur.execute(
                        """
                        delete from finance_bank_deposit_items
                        where deposit_id = any(%s)
                        """,
                        (remove_ids,),
                    )
                    cur.execute(
                        """
                        delete from finance_bank_deposits
                        where id = any(%s)
                        """,
                        (remove_ids,),
                    )

                repaired = 0

                for (
                    ym,
                    sequence,
                    deposit_date_text,
                    expected_amount,
                    receipt_from,
                    receipt_to,
                ) in OFFICIAL_CHE_CDM_ROWS:

                    current = keep_by_key.get((ym, sequence))

                    if current and current["id"] not in remove_ids:
                        cur.execute(
                            """
                            update finance_bank_deposits
                            set
                                branch = 'CHE',
                                deposit_date = %s,
                                ym = %s,
                                fund_account = '观音堂日常户口',
                                bank_name = 'Hong Leong Bank',
                                reference_no = %s,
                                amount = %s,
                                remarks = '系统修复：CHE 月费现金 CDM',
                                receipt_from = %s,
                                receipt_to = %s,
                                cdm_sequence = %s
                            where id = %s
                            """,
                            (
                                deposit_date_text,
                                ym,
                                f"CDM-{sequence}",
                                expected_amount,
                                receipt_from,
                                receipt_to,
                                sequence,
                                current["id"],
                            ),
                        )
                    else:
                        cur.execute(
                            """
                            insert into finance_bank_deposits
                            (
                                branch,
                                deposit_date,
                                ym,
                                fund_account,
                                bank_name,
                                reference_no,
                                amount,
                                remarks,
                                receipt_from,
                                receipt_to,
                                cdm_sequence
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                '观音堂日常户口',
                                'Hong Leong Bank',
                                %s,
                                %s,
                                '系统修复：CHE 月费现金 CDM',
                                %s,
                                %s,
                                %s
                            )
                            """,
                            (
                                deposit_date_text,
                                ym,
                                f"CDM-{sequence}",
                                expected_amount,
                                receipt_from,
                                receipt_to,
                                sequence,
                            ),
                        )

                    repaired += 1

            conn.commit()
            return repaired

        except Exception:
            conn.rollback()
            raise


def _update_bank_in_receipt_amount(record_id, new_amount):
    """
    只允许修改 CHE 现金月费记录金额。
    这里不会猜测哪一张错误，必须由负责人看原始收条后手动输入。
    """
    amount = _bank_in_decimal(new_amount)

    if amount <= 0:
        raise ValueError("金额必须大于 0。")

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select
                        id,
                        receipt_no,
                        member_id,
                        name,
                        amount
                    from finance_records
                    where id = %s
                      and record_type = 'income'
                      and category = '月费'
                      and payment_method = '现金'
                      and coalesce(status, 'confirmed') <> 'cancelled'
                      and upper(coalesce(receipt_no, '')) like 'CHE%%'
                    for update
                    """,
                    (record_id,),
                )
                row = cur.fetchone()

                if not row:
                    raise ValueError("找不到这笔可修改的 CHE 现金月费记录。")

                old_amount = _bank_in_decimal(row.get("amount"))

                cur.execute(
                    """
                    update finance_records
                    set
                        amount = %s,
                        remarks = concat_ws(
                            ' | ',
                            nullif(remarks, ''),
                            %s
                        )
                    where id = %s
                    """,
                    (
                        amount,
                        (
                            "Bank In 修复中心调整金额："
                            f"RM {old_amount:,.2f} → RM {amount:,.2f}"
                        ),
                        record_id,
                    ),
                )

            conn.commit()

            return {
                "receipt_no": row.get("receipt_no") or "",
                "name": row.get("name") or "",
                "old_amount": old_amount,
                "new_amount": amount,
            }

        except Exception:
            conn.rollback()
            raise


def _rebuild_che_bank_in_items():
    """
    只有 18/18 主记录正确且 18/18 金额验证通过时才重建。
    """
    scan = _scan_che_bank_in()

    if not scan["all_master_ok"]:
        raise ValueError("CDM 主记录尚未全部修正，不能重建。")

    if not scan["all_validation_ok"]:
        raise ValueError("仍有收条金额不符，不能重建。")

    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select
                        id,
                        ym,
                        cdm_sequence
                    from finance_bank_deposits
                    where branch = 'CHE'
                      and ym between '2026-01' and '2026-06'
                    for update
                    """
                )
                deposit_rows = cur.fetchall() or []
                deposit_by_key = {
                    (
                        str(row.get("ym") or ""),
                        int(row.get("cdm_sequence") or 0),
                    ): row["id"]
                    for row in deposit_rows
                }

                deposit_ids = list(deposit_by_key.values())

                if deposit_ids:
                    cur.execute(
                        """
                        delete from finance_bank_deposit_items
                        where deposit_id = any(%s)
                        """,
                        (deposit_ids,),
                    )

                rebuilt_count = 0

                for batch in scan["batches"]:
                    key = (batch["ym"], batch["sequence"])
                    deposit_id = deposit_by_key.get(key)

                    if not deposit_id:
                        raise ValueError(
                            f"{batch['ym']} CDM-{batch['sequence']} "
                            "缺少主记录。"
                        )

                    for receipt in batch["receipts"]:
                        # 避免依赖 ON CONFLICT 约束名称。
                        cur.execute(
                            """
                            delete from finance_bank_deposit_items
                            where finance_record_id = %s
                            """,
                            (receipt["id"],),
                        )

                        cur.execute(
                            """
                            insert into finance_bank_deposit_items
                            (
                                deposit_id,
                                finance_record_id,
                                amount
                            )
                            values (%s, %s, %s)
                            """,
                            (
                                deposit_id,
                                receipt["id"],
                                receipt["amount"],
                            ),
                        )
                        rebuilt_count += 1

            conn.commit()
            return rebuilt_count

        except Exception:
            conn.rollback()
            raise


OFFICIAL_CHE_CDM_ROWS = [
    # ym, CDM编号, Bank In日期, 金额, 收条起号, 收条尾号
    ("2026-01", 1, "2026-01-06", Decimal("1800.00"), "CHE0001292", "CHE0001300"),
    ("2026-01", 2, "2026-01-06", Decimal("2700.00"), "CHE0001401", "CHE0001407"),
    ("2026-01", 3, "2026-01-14", Decimal("2000.00"), "CHE0001408", "CHE0001412"),
    ("2026-01", 4, "2026-01-26", Decimal("1150.00"), "CHE0001413", "CHE0001417"),

    ("2026-02", 1, "2026-02-09", Decimal("1900.00"), "CHE0001425", "CHE0001431"),

    ("2026-03", 1, "2026-03-02", Decimal("2250.00"), "CHE0001432", "CHE0001440"),
    ("2026-03", 2, "2026-03-11", Decimal("2000.00"), "CHE0001441", "CHE0001448"),
    ("2026-03", 3, "2026-03-22", Decimal("1250.00"), "CHE0001449", "CHE0001453"),

    ("2026-04", 1, "2026-04-12", Decimal("2250.00"), "CHE0001454", "CHE0001462"),
    ("2026-04", 2, "2026-04-20", Decimal("1100.00"), "CHE0001463", "CHE0001468"),
    ("2026-04", 3, "2026-04-29", Decimal("450.00"), "CHE0001469", "CHE0001471"),

    ("2026-05", 1, "2026-05-04", Decimal("1900.00"), "CHE0001472", "CHE0001478"),
    ("2026-05", 2, "2026-05-18", Decimal("300.00"), "CHE0001479", "CHE0001480"),
    ("2026-05", 3, "2026-05-26", Decimal("800.00"), "CHE0001481", "CHE0001484"),
    ("2026-05", 4, "2026-05-29", Decimal("400.00"), "CHE0001485", "CHE0001487"),

    ("2026-06", 1, "2026-06-04", Decimal("900.00"), "CHE0001488", "CHE0001492"),
    ("2026-06", 2, "2026-06-18", Decimal("200.00"), "CHE0001493", "CHE0001494"),
    ("2026-06", 3, "2026-06-26", Decimal("300.00"), "CHE0001495", "CHE0001500"),
]


def _bank_in_receipt_number(receipt_no):
    """CHE0001454 / CHE1454 -> 1454。"""
    match = re.search(r"(\d+)$", str(receipt_no or "").strip())
    return int(match.group(1)) if match else None


def _bank_in_decimal(value):
    try:
        return Decimal(str(value or 0)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _find_che_cash_monthly_receipts(cur, receipt_from, receipt_to):
    """
    按 CHE 收条号码范围寻找现金月费。
    finance_records 没有 branch，因此以 receipt_no 的 CHE 前缀判断。
    """
    start_no = _bank_in_receipt_number(receipt_from)
    end_no = _bank_in_receipt_number(receipt_to)

    if start_no is None or end_no is None:
        raise ValueError(f"收条范围格式错误：{receipt_from} 至 {receipt_to}")

    cur.execute(
        """
        select
            r.id,
            r.record_date,
            r.receipt_date,
            r.receipt_no,
            r.member_id,
            r.name,
            r.amount,
            r.status,
            r.remarks
        from finance_records r
        where r.record_type = 'income'
          and r.category = '月费'
          and r.payment_method = '现金'
          and coalesce(r.status, 'confirmed') <> 'cancelled'
          and upper(coalesce(r.receipt_no, '')) like 'CHE%%'
          and nullif(
                regexp_replace(
                    coalesce(r.receipt_no, ''),
                    '[^0-9]',
                    '',
                    'g'
                ),
                ''
              )::bigint between %s and %s
        order by
            nullif(
                regexp_replace(
                    coalesce(r.receipt_no, ''),
                    '[^0-9]',
                    '',
                    'g'
                ),
                ''
            )::bigint,
            r.id
        """,
        (start_no, end_no),
    )
    return cur.fetchall() or []


def _get_bank_in_master_rows(cur):
    cur.execute(
        """
        select
            d.id,
            d.ym,
            d.cdm_sequence,
            d.deposit_date,
            d.amount,
            d.receipt_from,
            d.receipt_to,
            d.reference_no,
            d.branch
        from finance_bank_deposits d
        where d.branch = 'CHE'
          and d.ym between '2026-01' and '2026-06'
        order by d.ym, d.cdm_sequence, d.id
        """
    )
    return cur.fetchall() or []


def _scan_che_bank_in():
    """只扫描，不修改数据库。"""
    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            master_rows = _get_bank_in_master_rows(cur)

            master_by_key = {}
            duplicate_keys = set()

            for row in master_rows:
                key = (
                    str(row.get("ym") or ""),
                    int(row.get("cdm_sequence") or 0),
                )
                if key in master_by_key:
                    duplicate_keys.add(key)
                else:
                    master_by_key[key] = row

            official_keys = {
                (ym, sequence)
                for ym, sequence, *_ in OFFICIAL_CHE_CDM_ROWS
            }

            extra_master_rows = [
                row for row in master_rows
                if (
                    str(row.get("ym") or ""),
                    int(row.get("cdm_sequence") or 0),
                ) not in official_keys
            ]

            batches = []
            master_ok_count = 0
            validation_ok_count = 0

            for (
                ym,
                sequence,
                deposit_date_text,
                expected_amount,
                expected_from,
                expected_to,
            ) in OFFICIAL_CHE_CDM_ROWS:

                key = (ym, sequence)
                master = master_by_key.get(key)
                master_issues = []

                if not master:
                    master_issues.append("缺少 CDM 主记录")
                else:
                    actual_date = str(master.get("deposit_date") or "")
                    actual_amount = _bank_in_decimal(master.get("amount"))
                    actual_from = str(master.get("receipt_from") or "")
                    actual_to = str(master.get("receipt_to") or "")

                    if actual_date != deposit_date_text:
                        master_issues.append(
                            f"Bank In 日期应为 {deposit_date_text}"
                        )
                    if actual_amount != expected_amount:
                        master_issues.append(
                            f"CDM 金额应为 RM {expected_amount:,.2f}"
                        )
                    if (
                        actual_from != expected_from
                        or actual_to != expected_to
                    ):
                        master_issues.append(
                            f"收条范围应为 {expected_from} 至 {expected_to}"
                        )

                if key in duplicate_keys:
                    master_issues.append("同月份相同 CDM 编号有重复主记录")

                master_ok = not master_issues
                if master_ok:
                    master_ok_count += 1

                receipts = _find_che_cash_monthly_receipts(
                    cur,
                    expected_from,
                    expected_to,
                )

                receipt_total = sum(
                    (_bank_in_decimal(row.get("amount")) for row in receipts),
                    Decimal("0.00"),
                )

                validation_issues = []

                if not receipts:
                    validation_issues.append(
                        f"找不到收条 {expected_from} 至 {expected_to}"
                    )
                elif receipt_total != expected_amount:
                    validation_issues.append(
                        "金额不符："
                        f"CDM RM {expected_amount:,.2f}，"
                        f"收条合计 RM {receipt_total:,.2f}，"
                        f"共 {len(receipts)} 张"
                    )

                validation_ok = not validation_issues
                if validation_ok:
                    validation_ok_count += 1

                batches.append({
                    "ym": ym,
                    "sequence": sequence,
                    "deposit_date": deposit_date_text,
                    "expected_amount": expected_amount,
                    "expected_from": expected_from,
                    "expected_to": expected_to,
                    "master": master,
                    "master_issues": master_issues,
                    "master_ok": master_ok,
                    "receipts": receipts,
                    "receipt_total": receipt_total,
                    "validation_issues": validation_issues,
                    "validation_ok": validation_ok,
                })

            cur.execute(
                """
                select count(*) as count
                from finance_bank_deposit_items
                """
            )
            item_count = int((cur.fetchone() or {}).get("count") or 0)

            cur.execute(
                """
                select
                    count(*) as count,
                    coalesce(sum(r.amount), 0) as total
                from finance_records r
                where r.record_type = 'income'
                  and r.category = '月费'
                  and r.payment_method = '现金'
                  and coalesce(r.status, 'confirmed') <> 'cancelled'
                  and r.record_date >= date '2026-01-01'
                  and (
                        upper(coalesce(r.member_id, '')) like 'CHE-%%'
                     or upper(coalesce(r.receipt_no, '')) like 'CHE%%'
                  )
                  and not exists (
                        select 1
                        from finance_bank_deposit_items i
                        where i.finance_record_id = r.id
                  )
                """
            )
            waiting = cur.fetchone() or {}

    all_master_ok = master_ok_count == len(OFFICIAL_CHE_CDM_ROWS)
    all_validation_ok = validation_ok_count == len(OFFICIAL_CHE_CDM_ROWS)

    return {
        "batches": batches,
        "master_count": len(master_rows),
        "master_ok_count": master_ok_count,
        "validation_ok_count": validation_ok_count,
        "official_count": len(OFFICIAL_CHE_CDM_ROWS),
        "extra_master_rows": extra_master_rows,
        "all_master_ok": all_master_ok,
        "all_validation_ok": all_validation_ok,
        "can_rebuild": all_master_ok and all_validation_ok,
        "item_count": item_count,
        "waiting_count": int(waiting.get("count") or 0),
        "waiting_total": _bank_in_decimal(waiting.get("total")),
    }


def _repair_che_bank_in_master():
    """只修正 finance_bank_deposits，不碰明细和收条。"""
    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select *
                    from finance_bank_deposits
                    where branch = 'CHE'
                      and ym between '2026-01' and '2026-06'
                    order by ym, cdm_sequence, id
                    for update
                    """
                )
                rows = cur.fetchall() or []

                official_keys = {
                    (ym, sequence)
                    for ym, sequence, *_ in OFFICIAL_CHE_CDM_ROWS
                }

                keep_by_key = {}
                remove_ids = []

                for row in rows:
                    key = (
                        str(row.get("ym") or ""),
                        int(row.get("cdm_sequence") or 0),
                    )

                    if key not in official_keys or key in keep_by_key:
                        remove_ids.append(row["id"])
                    else:
                        keep_by_key[key] = row

                remove_ids = sorted(set(remove_ids))

                if remove_ids:
                    cur.execute(
                        """
                        delete from finance_bank_deposit_items
                        where deposit_id = any(%s)
                        """,
                        (remove_ids,),
                    )
                    cur.execute(
                        """
                        delete from finance_bank_deposits
                        where id = any(%s)
                        """,
                        (remove_ids,),
                    )

                repaired = 0

                for (
                    ym,
                    sequence,
                    deposit_date_text,
                    expected_amount,
                    receipt_from,
                    receipt_to,
                ) in OFFICIAL_CHE_CDM_ROWS:

                    current = keep_by_key.get((ym, sequence))

                    if current and current["id"] not in remove_ids:
                        cur.execute(
                            """
                            update finance_bank_deposits
                            set
                                branch = 'CHE',
                                deposit_date = %s,
                                ym = %s,
                                fund_account = '观音堂日常户口',
                                bank_name = 'Hong Leong Bank',
                                reference_no = %s,
                                amount = %s,
                                remarks = '系统修复：CHE 月费现金 CDM',
                                receipt_from = %s,
                                receipt_to = %s,
                                cdm_sequence = %s
                            where id = %s
                            """,
                            (
                                deposit_date_text,
                                ym,
                                f"CDM-{sequence}",
                                expected_amount,
                                receipt_from,
                                receipt_to,
                                sequence,
                                current["id"],
                            ),
                        )
                    else:
                        cur.execute(
                            """
                            insert into finance_bank_deposits
                            (
                                branch,
                                deposit_date,
                                ym,
                                fund_account,
                                bank_name,
                                reference_no,
                                amount,
                                remarks,
                                receipt_from,
                                receipt_to,
                                cdm_sequence
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                '观音堂日常户口',
                                'Hong Leong Bank',
                                %s,
                                %s,
                                '系统修复：CHE 月费现金 CDM',
                                %s,
                                %s,
                                %s
                            )
                            """,
                            (
                                deposit_date_text,
                                ym,
                                f"CDM-{sequence}",
                                expected_amount,
                                receipt_from,
                                receipt_to,
                                sequence,
                            ),
                        )

                    repaired += 1

            conn.commit()
            return repaired

        except Exception:
            conn.rollback()
            raise


def _update_bank_in_receipt_amount(record_id, new_amount):
    """
    只允许修改 CHE 现金月费记录金额。
    这里不会猜测哪一张错误，必须由负责人看原始收条后手动输入。
    """
    amount = _bank_in_decimal(new_amount)

    if amount <= 0:
        raise ValueError("金额必须大于 0。")

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select
                        id,
                        receipt_no,
                        member_id,
                        name,
                        amount
                    from finance_records
                    where id = %s
                      and record_type = 'income'
                      and category = '月费'
                      and payment_method = '现金'
                      and coalesce(status, 'confirmed') <> 'cancelled'
                      and upper(coalesce(receipt_no, '')) like 'CHE%%'
                    for update
                    """,
                    (record_id,),
                )
                row = cur.fetchone()

                if not row:
                    raise ValueError("找不到这笔可修改的 CHE 现金月费记录。")

                old_amount = _bank_in_decimal(row.get("amount"))

                cur.execute(
                    """
                    update finance_records
                    set
                        amount = %s,
                        remarks = concat_ws(
                            ' | ',
                            nullif(remarks, ''),
                            %s
                        )
                    where id = %s
                    """,
                    (
                        amount,
                        (
                            "Bank In 修复中心调整金额："
                            f"RM {old_amount:,.2f} → RM {amount:,.2f}"
                        ),
                        record_id,
                    ),
                )

            conn.commit()

            return {
                "receipt_no": row.get("receipt_no") or "",
                "name": row.get("name") or "",
                "old_amount": old_amount,
                "new_amount": amount,
            }

        except Exception:
            conn.rollback()
            raise


def _rebuild_che_bank_in_items():
    """
    只有 18/18 主记录正确且 18/18 金额验证通过时才重建。
    """
    scan = _scan_che_bank_in()

    if not scan["all_master_ok"]:
        raise ValueError("CDM 主记录尚未全部修正，不能重建。")

    if not scan["all_validation_ok"]:
        raise ValueError("仍有收条金额不符，不能重建。")

    ensure_finance_bank_in_tables()

    with get_conn() as conn:
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    select
                        id,
                        ym,
                        cdm_sequence
                    from finance_bank_deposits
                    where branch = 'CHE'
                      and ym between '2026-01' and '2026-06'
                    for update
                    """
                )
                deposit_rows = cur.fetchall() or []
                deposit_by_key = {
                    (
                        str(row.get("ym") or ""),
                        int(row.get("cdm_sequence") or 0),
                    ): row["id"]
                    for row in deposit_rows
                }

                deposit_ids = list(deposit_by_key.values())

                if deposit_ids:
                    cur.execute(
                        """
                        delete from finance_bank_deposit_items
                        where deposit_id = any(%s)
                        """,
                        (deposit_ids,),
                    )

                rebuilt_count = 0

                for batch in scan["batches"]:
                    key = (batch["ym"], batch["sequence"])
                    deposit_id = deposit_by_key.get(key)

                    if not deposit_id:
                        raise ValueError(
                            f"{batch['ym']} CDM-{batch['sequence']} "
                            "缺少主记录。"
                        )

                    for receipt in batch["receipts"]:
                        # 避免依赖 ON CONFLICT 约束名称。
                        cur.execute(
                            """
                            delete from finance_bank_deposit_items
                            where finance_record_id = %s
                            """,
                            (receipt["id"],),
                        )

                        cur.execute(
                            """
                            insert into finance_bank_deposit_items
                            (
                                deposit_id,
                                finance_record_id,
                                amount
                            )
                            values (%s, %s, %s)
                            """,
                            (
                                deposit_id,
                                receipt["id"],
                                receipt["amount"],
                            ),
                        )
                        rebuilt_count += 1

            conn.commit()
            return rebuilt_count

        except Exception:
            conn.rollback()
            raise


@finance_bp.route(
    "/admin/treasury/rebuild-bank-in-items",
    methods=["GET", "POST"],
)
@finance_admin_required
def finance_rebuild_bank_in_items():
    """Legacy compatibility redirect. Formal page moved to finance_v7.py."""
    return redirect(url_for("finance_v7.finance_v7_repair_center"))

@finance_bp.route("/admin/treasury")
@finance_admin_required
def finance_treasury():
    """Legacy compatibility redirect. Formal page moved to finance_v7.py."""
    return redirect(url_for("finance_v7.finance_v7_bank"))


@finance_bp.route(
    "/admin/cash-transfer",
    methods=["GET", "POST"]
)
@finance_admin_required
def finance_cash_transfer():
    """Legacy compatibility redirect. Formal page moved to finance_v7.py."""
    return redirect(url_for("finance_v7.finance_v7_cash_transfer"))

@finance_bp.route("/bank_pending", methods=["GET", "POST"])
def bank_pending():

    message = ""
    message_type = "bad"

    form_data = {
        "raw_text": "",
        "member_id": "",
        "name": "",
        "amount": "",
        "payment_date": date.today().isoformat(),
        "bank_ref": "",
        "bank_name": "",
        "category": "月费",
        "remarks": "",
    }

    if request.method == "POST":

        raw_text = request.form.get("raw_text", "").strip()

        member_id_raw = request.form.get(
            "member_id",
            ""
        ).strip()

        name = request.form.get(
            "name",
            ""
        ).strip()

        amount_raw = request.form.get(
            "amount",
            ""
        ).strip()

        payment_date = (
            request.form.get("payment_date")
            or date.today().isoformat()
        )

        bank_ref = request.form.get(
            "bank_ref",
            ""
        ).strip()

        bank_name = request.form.get(
            "bank_name",
            ""
        ).strip()

        category = request.form.get(
            "category",
            "月费"
        ).strip()

        remarks = request.form.get(
            "remarks",
            ""
        ).strip()

        if raw_text:

            member_match = re.search(
                r"(CHE|STW)[-\s]?\d+",
                raw_text,
                re.IGNORECASE
            )

            if member_match and not member_id_raw:

                member_id_raw = (
                    member_match
                    .group(0)
                    .replace(" ", "-")
                    .upper()
                )

            amount_match = re.search(
                r"(RM|MYR)\s*([0-9]+(?:\.[0-9]{1,2})?)",
                raw_text,
                re.IGNORECASE
            )

            if amount_match and not amount_raw:
                amount_raw = amount_match.group(2)

            ref_match = re.search(
                r"(Reference|Ref|Transaction|DuitNow)"
                r"\s*(No|ID|Number)?[:\s#-]*"
                r"([A-Za-z0-9\-]+)",
                raw_text,
                re.IGNORECASE
            )

            if ref_match and not bank_ref:
                bank_ref = ref_match.group(3)

            bank_options = [
                "Maybank",
                "Public Bank",
                "CIMB",
                "Hong Leong",
                "RHB",
                "AmBank",
                "Bank Islam",
                "BSN",
            ]

            for bank in bank_options:

                if (
                    bank.lower() in raw_text.lower()
                    and not bank_name
                ):
                    bank_name = bank
                    break

            if not remarks:
                remarks = raw_text[:1000]

        member_id = (
            normalize_member_id(member_id_raw)
            if member_id_raw
            else ""
        )

        amount = money(amount_raw)

        if category == "月费" and member_id:

            member = db_query("""
                select *
                from members
                where member_id = %s
                limit 1
            """, (
                member_id,
            ), fetchone=True)

            if member:
                name = (
                    member.get("name")
                    or member.get("姓名")
                    or name
                )

        form_data = {
            "raw_text": raw_text,
            "member_id": member_id_raw,
            "name": name,
            "amount": amount_raw,
            "payment_date": str(payment_date),
            "bank_ref": bank_ref,
            "bank_name": bank_name,
            "category": category,
            "remarks": remarks,
        }

        existing_ref = None

        if bank_ref:

            existing_ref = db_query("""
                select id
                from bank_pending_records
                where bank_ref = %s
                  and coalesce(bank_ref, '') <> ''
                limit 1
            """, (
                bank_ref,
            ), fetchone=True)

        if existing_ref:

            message = (
                "这个 Bank Reference 已经存在，"
                "请检查是否重复导入。"
            )

        elif amount <= 0:

            message = "请输入正确的银行过账金额。"

        else:

            db_query("""
                insert into bank_pending_records
                (
                    member_id,
                    name,
                    amount,
                    payment_date,
                    bank_ref,
                    bank_name,
                    category,
                    remarks
                )
                values
                (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, (
                member_id,
                name,
                amount,
                payment_date,
                bank_ref,
                bank_name,
                category,
                remarks
            ))

            return redirect(
                url_for("finance.bank_pending")
            )

    summary = db_query("""
        select
            count(*) as cnt,
            coalesce(sum(amount), 0) as total
        from bank_pending_records
        where status = 'pending'
    """, fetchone=True)

    rows = db_query("""
        select *
        from bank_pending_records
        where status = 'pending'
        order by upload_date desc
    """, fetchall=True)

    pending_count = int(
        summary["cnt"] or 0
    )

    pending_total = float(
        summary["total"] or 0
    )

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>

        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>银行过账中心</title>

        <link
            rel="stylesheet"
            href="{{ url_for(
                'static',
                filename='css/toolbox.css'
            ) }}"
        >

        <style>

            .bank-page {
                max-width: 1450px;
            }

            .bank-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;

                box-shadow:
                    0 12px 30px
                    rgba(37, 99, 235, 0.18);
            }

            .bank-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .bank-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .bank-summary {
                display: grid;
                grid-template-columns:
                    repeat(2, minmax(0, 1fr));
                gap: 16px;
                margin-bottom: 20px;
            }

            .bank-summary .summary-box {
                min-height: 125px;
                text-align: center;

                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
            }

            .summary-icon {
                font-size: 30px;
                margin-bottom: 6px;
            }

            .summary-label {
                color: #64748b;
                font-size: 16px;
                margin-bottom: 6px;
            }

            .summary-value {
                font-size: 28px;
                font-weight: 800;
                color: #0f172a;
            }

            .summary-value.money-value {
                color: #15803d;
            }

            .entry-grid {
                display: grid;
                grid-template-columns:
                    minmax(0, 1.2fr)
                    minmax(360px, 0.8fr);
                gap: 20px;
                align-items: start;
            }

            .form-grid {
                display: grid;
                grid-template-columns:
                    repeat(2, minmax(0, 1fr));
                gap: 16px;
            }

            .form-grid .full-width {
                grid-column: 1 / -1;
            }

            .receipt-textarea {
                width: 100%;
                min-height: 245px;
                resize: vertical;
                line-height: 1.6;
            }

            .form-help {
                margin-top: 7px;
                color: #64748b;
                font-size: 14px;
                line-height: 1.5;
            }

            .smart-note {
                background: #eff6ff;
                border: 1px solid #bfdbfe;
                color: #1e40af;
                border-radius: 14px;
                padding: 14px 16px;
                margin-bottom: 16px;
                line-height: 1.6;
            }

            .records-card {
                margin-top: 20px;
            }

            .table-topbar {
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 12px;
                flex-wrap: wrap;
                margin-bottom: 14px;
            }

            .record-count {
                color: #64748b;
                font-size: 16px;
            }

            .pending-table {
                min-width: 1400px;
            }

            .pending-table td {
                vertical-align: middle;
            }

            .member-id {
                font-weight: 800;
                color: #1d4ed8;
                white-space: nowrap;
            }

            .member-name {
                font-weight: 700;
                color: #1e293b;
                min-width: 90px;
            }

            .money-cell {
                color: #15803d;
                font-weight: 800;
                white-space: nowrap;
                text-align: right;
            }

            .reference-cell {
                font-family: Consolas, monospace;
                font-size: 14px;
                white-space: nowrap;
            }

            .remarks-cell {
                min-width: 180px;
                max-width: 320px;
                white-space: normal;
                line-height: 1.5;
                overflow-wrap: anywhere;
            }

            .confirm-box {
                min-width: 285px;
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 14px;
                padding: 12px;
            }

            .confirm-grid {
                display: grid;
                gap: 10px;
            }

            .confirm-box .form-group {
                margin: 0;
            }

            .confirm-box .form-label {
                font-size: 14px;
            }

            .confirm-box .form-input {
                min-height: 41px;
                padding: 8px 10px;
                font-size: 14px;
                margin: 0;
            }

            .confirm-box .btn-tool {
                width: 100%;
                min-height: 42px;
                font-size: 14px;
                padding: 9px 12px;
            }

            .delete-form .btn-tool {
                min-height: 42px;
                padding: 9px 13px;
                font-size: 14px;
                white-space: nowrap;
            }

            .empty-bank {
                text-align: center;
                padding: 55px 20px;
                color: #64748b;
            }

            .empty-bank-icon {
                font-size: 48px;
                margin-bottom: 10px;
            }

            .empty-bank h3 {
                margin: 0 0 8px;
                color: #334155;
            }

            .bottom-actions {
                margin-top: 20px;
            }

            @media (max-width: 950px) {

                .entry-grid {
                    grid-template-columns: 1fr;
                }

            }

            @media (max-width: 700px) {

                .bank-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .bank-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .bank-header h1 {
                    font-size: 26px;
                }

                .bank-summary {
                    grid-template-columns: 1fr;
                }

                .bank-summary .summary-box {
                    min-height: 105px;
                }

                .form-grid {
                    grid-template-columns: 1fr;
                }

                .form-grid .full-width {
                    grid-column: auto;
                }

                .entry-grid {
                    display: block;
                }

                .entry-grid .card + .card {
                    margin-top: 20px;
                }

                .bottom-actions .btn-tool {
                    width: 100%;
                }

            }

        </style>

    </head>

    <body>

    <div class="page bank-page">

        <div class="bank-header">

            <h1>🏦 银行过账中心</h1>

            <p>
                先登记银行转账资料，财政确认后再填写收条号码并正式入账。
            </p>

        </div>

        {% if message %}

            <div class="alert alert-danger">
                ⚠️ {{ message }}
            </div>

        {% endif %}

        <div class="bank-summary">

            <div class="summary-box">

                <div class="summary-icon">
                    ⏳
                </div>

                <div class="summary-label">
                    待确认笔数
                </div>

                <div class="summary-value">
                    {{ pending_count }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    💰
                </div>

                <div class="summary-label">
                    待确认总额
                </div>

                <div class="summary-value money-value">
                    RM {{ "%.2f"|format(pending_total) }}
                </div>

            </div>

        </div>

        <form method="post">

            <div class="entry-grid">

                <div class="card">

                    <div class="section-title">
                        📄 银行资料智能识别
                    </div>

                    <div class="smart-note">
                        可以直接粘贴 WhatsApp 信息或银行转账
                        Receipt 文字。系统会尝试识别会员编号、
                        金额、银行及 Reference。
                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            项目
                        </label>

                        <select
                            class="form-input"
                            name="category"
                        >

                            {% set categories = [
                                '月费',
                                '财布施',
                                '观音村',
                                '膳食结缘',
                                '观音堂纯檀香布施',
                                special_donation_title,
                                '临时特别布施'
                            ] %}

                            {% for category_item in categories %}

                                <option
                                    value="{{ category_item }}"
                                    {% if
                                        form_data.category
                                        == category_item
                                    %}
                                        selected
                                    {% endif %}
                                >
                                    {{ category_item }}
                                </option>

                            {% endfor %}

                        </select>

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            粘贴 WhatsApp／银行 Receipt 文字
                        </label>

                        <textarea
                            class="form-input receipt-textarea"
                            name="raw_text"
                            placeholder="例如：

CHE-108
RM50.00
Maybank
Reference：ABC123456"
                        >{{ form_data.raw_text }}</textarea>

                        <div class="form-help">
                            系统识别后，仍可在右边检查和修改资料。
                        </div>

                    </div>

                </div>

                <div class="card">

                    <div class="section-title">
                        ✏️ 手动补充／修正资料
                    </div>

                    <div class="form-grid">

                        <div class="form-group">

                            <label class="form-label">
                                会员编号
                            </label>

                            <input
                                class="form-input"
                                name="member_id"
                                value="{{ form_data.member_id }}"
                                placeholder="CHE-108 / STW-108 / 108"
                                autocomplete="off"
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                姓名
                            </label>

                            <input
                                class="form-input"
                                name="name"
                                value="{{ form_data.name }}"
                                placeholder="非月费或找不到会员时填写"
                                autocomplete="off"
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                金额 RM
                            </label>

                            <input
                                class="form-input"
                                name="amount"
                                type="number"
                                step="0.01"
                                min="0.01"
                                value="{{ form_data.amount }}"
                                placeholder="例如 50.00"
                                required
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                付款日期
                            </label>

                            <input
                                class="form-input"
                                name="payment_date"
                                type="date"
                                value="{{ form_data.payment_date }}"
                                required
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                银行 Reference
                            </label>

                            <input
                                class="form-input"
                                name="bank_ref"
                                value="{{ form_data.bank_ref }}"
                                placeholder="Transaction Reference"
                                autocomplete="off"
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                银行
                            </label>

                            <select
                                class="form-input"
                                name="bank_name"
                            >

                                <option value="">
                                    请选择银行
                                </option>

                                {% for bank in bank_names %}

                                    <option
                                        value="{{ bank }}"
                                        {% if
                                            form_data.bank_name == bank
                                        %}
                                            selected
                                        {% endif %}
                                    >
                                        {{ bank }}
                                    </option>

                                {% endfor %}

                            </select>

                        </div>

                        <div class="form-group full-width">

                            <label class="form-label">
                                备注
                            </label>

                            <textarea
                                class="form-input"
                                name="remarks"
                                rows="4"
                                placeholder="填写补充说明或保留银行转账文字"
                            >{{ form_data.remarks }}</textarea>

                        </div>

                    </div>

                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                        style="width:100%;"
                    >
                        ➕ 加入待确认
                    </button>

                </div>

            </div>

        </form>

        <div class="card records-card">

            <div class="table-topbar">

                <div
                    class="section-title"
                    style="margin-bottom:0;"
                >
                    📋 待确认列表
                </div>

                <div class="record-count">
                    共
                    <strong>{{ pending_count }}</strong>
                    笔待确认记录
                </div>

            </div>

            {% if rows %}

                <div class="table-responsive">

                    <table class="record-table pending-table">

                        <thead>

                            <tr>
                                <th>付款日期</th>
                                <th>编号</th>
                                <th>姓名</th>
                                <th>项目</th>
                                <th>金额</th>
                                <th>Reference</th>
                                <th>银行</th>
                                <th>备注</th>
                                <th>确认入账</th>
                                <th>删除</th>
                            </tr>

                        </thead>

                        <tbody>

                            {% for r in rows %}

                                <tr>

                                    <td style="white-space:nowrap;">
                                        {{ r.payment_date }}
                                    </td>

                                    <td>
                                        <span class="member-id">
                                            {{ r.member_id or "-" }}
                                        </span>
                                    </td>

                                    <td>
                                        <span class="member-name">
                                            {{ r.name or "-" }}
                                        </span>
                                    </td>

                                    <td>
                                        {{ r.category or "-" }}
                                    </td>

                                    <td class="money-cell">
                                        RM
                                        {{ "%.2f"|format(
                                            r.amount or 0
                                        ) }}
                                    </td>

                                    <td class="reference-cell">
                                        {{ r.bank_ref or "-" }}
                                    </td>

                                    <td style="white-space:nowrap;">
                                        {{ r.bank_name or "-" }}
                                    </td>

                                    <td class="remarks-cell">
                                        {{ r.remarks or "-" }}
                                    </td>

                                    <td>

                                        <div class="confirm-box">

                                            <form
                                                method="post"
                                                action="{{ url_for(
                                                    'finance.confirm_bank',
                                                    pending_id=r.id
                                                ) }}"
                                                onsubmit="
                                                    return confirm(
                                                        '确定确认入账？'
                                                    );
                                                "
                                            >

                                                <div class="confirm-grid">

                                                    <div class="form-group">

                                                        <label
                                                            class="form-label"
                                                        >
                                                            收条号码
                                                        </label>

                                                        <input
                                                            class="form-input"
                                                            name="receipt_no"
                                                            placeholder="CHE0000001"
                                                            autocomplete="off"
                                                            required
                                                        >

                                                    </div>

                                                    <div class="form-group">

                                                        <label
                                                            class="form-label"
                                                        >
                                                            开收条日期
                                                        </label>

                                                        <input
                                                            class="form-input"
                                                            name="receipt_date"
                                                            type="date"
                                                            value="{{ today }}"
                                                            required
                                                        >

                                                    </div>

                                                    <button
                                                        class="
                                                            btn-tool
                                                            btn-success
                                                        "
                                                        type="submit"
                                                    >
                                                        ✅ 确认入账
                                                    </button>

                                                </div>

                                            </form>

                                        </div>

                                    </td>

                                    <td>

                                        <form
                                            class="delete-form"
                                            method="post"
                                            action="{{ url_for(
                                                'finance.delete_bank_pending',
                                                pending_id=r.id
                                            ) }}"
                                            onsubmit="
                                                return confirm(
                                                    '确定删除这笔待确认记录？'
                                                );
                                            "
                                        >

                                            <button
                                                class="
                                                    btn-tool
                                                    btn-danger
                                                "
                                                type="submit"
                                            >
                                                🗑️ 删除
                                            </button>

                                        </form>

                                    </td>

                                </tr>

                            {% endfor %}

                        </tbody>

                    </table>

                </div>

            {% else %}

                <div class="empty-bank">

                    <div class="empty-bank-icon">
                        ✅
                    </div>

                    <h3>目前没有待确认记录</h3>

                    <p>
                        所有银行转账记录都已经处理完成。
                    </p>

                </div>

            {% endif %}

        </div>

        <div class="bottom-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'finance.finance_home'
                ) }}"
            >
                ← 返回财政首页
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        summary=summary,
        message=message,
        today=date.today().isoformat(),
        pending_count=pending_count,
        pending_total=pending_total,
        form_data=form_data,
        special_donation_title=SPECIAL_DONATION_TITLE,
        bank_names=[
            "Maybank",
            "Public Bank",
            "CIMB",
            "Hong Leong",
            "RHB",
            "AmBank",
            "Bank Islam",
            "BSN",
        ]
    )

@finance_bp.route("/bank_pending/<int:pending_id>/confirm", methods=["POST"])
def confirm_bank(pending_id):

    receipt_no = request.form.get("receipt_no", "").strip().upper()
    receipt_date = request.form.get("receipt_date") or date.today()

    if not receipt_no:
        return "必须填写收条号码", 400

    existing = db_query("""
        select id
        from finance_records
        where receipt_no = %s
        limit 1
    """, (receipt_no,), fetchone=True)

    if existing:
        return "这个收条号码已经用过了，请检查是否重复输入", 400

    p = db_query("""
        select *
        from bank_pending_records
        where id = %s and status = 'pending'
    """, (pending_id,), fetchone=True)

    if not p:
        return redirect(url_for("finance.bank_pending"))

    member = None
    paid_until_date = None
    month_from = ""
    month_to = ""
    month_from_db = None
    month_to_db = None
    month_count = None

    if p["category"] == "月费" and p["member_id"]:

        member = db_query("""
            select *
            from members
            where member_id = %s
            limit 1
        """, (p["member_id"],), fetchone=True)

        paid = db_query("""
            select max(end_month) as paid_until
            from member_payments
            where member_id = %s
              and coalesce(status, 'active') = 'active'
        """, (p["member_id"],), fetchone=True)

        paid_until_date = paid["paid_until"] if paid else None

        month_from = next_month_ym(paid_until_date)

        month_count = max(1, round(float(p["amount"] or 0) / 50))
        month_to = add_months_ym(month_from, month_count - 1)

        month_from_db = month_from + "-01"
        month_to_db = month_to + "-01"

    pending_fund_account = get_fund_account(
        p["category"],
        "income"
    )

    month_lock_error = require_finance_month_open(
        receipt_date,
        pending_fund_account
    )

    if month_lock_error:
        flash(month_lock_error, "danger")
        return redirect(
            url_for("finance.bank_pending")
        )

    db_query("""
        insert into finance_records
        (
            record_type,
            fund_account,
            record_date,
            receipt_date,
            category,
            receipt_no,
            member_id,
            name,
            amount,
            payment_method,
            bank_name,
            bank_ref,
            month_from,
            month_to,
            remarks
        )
        values
        (
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            '银行过账',
            %s, %s, %s, %s, %s
        )
    """, (
        "income",
        pending_fund_account,
        p["payment_date"],
        receipt_date,
        p["category"],
        receipt_no,
        p["member_id"],
        p["name"],
        p["amount"],
        p["bank_name"],
        p["bank_ref"],
        month_from,
        month_to,
        p["remarks"]
    ))

    if p["category"] == "月费" and p["member_id"]:

        db_query("""
            insert into member_payments
            (
                payment_date,
                receipt_date,
                member_id,
                name,
                receipt_no,
                amount,
                start_month,
                end_month,
                month_count
            )
            values
            (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s
            )
        """, (
            p["payment_date"],
            receipt_date,
            p["member_id"],
            p["name"],
            receipt_no,
            p["amount"],
            month_from_db,
            month_to_db,
            month_count
        ))

    db_query("""
        update bank_pending_records
        set status = 'confirmed'
        where id = %s
    """, (pending_id,))

    return redirect(url_for("finance.records"))

@finance_bp.route(
    "/records/<int:record_id>/cancel",
    methods=["POST"]
)
@finance_admin_required
def cancel_record(record_id):

    # 先读取完整资料
    record = db_query("""
        select
            id,
            receipt_no,
            record_date,
            fund_account,
            status
        from finance_records
        where id = %s
        limit 1
    """, (
        record_id,
    ), fetchone=True)

    if not record:

        flash(
            "找不到这笔财政记录。",
            "danger"
        )

        return redirect(
            url_for("finance.records")
        )

    # 已经作废，不可重复操作
    if record.get("status") == "cancelled":

        flash(
            "这笔财政记录已经作废。",
            "warning"
        )

        return redirect(
            url_for("finance.records")
        )

    # =====================================================
    # V6：Month Close Lock
    # 已月结月份禁止作废
    # =====================================================
    month_lock_error = require_finance_month_open(
        record["record_date"],
        record.get("fund_account")
    )

    if month_lock_error:

        flash(
            month_lock_error,
            "danger"
        )

        return redirect(
            url_for("finance.records")
        )

    cancel_reason = request.form.get(
        "cancel_reason",
        ""
    ).strip()

    if not cancel_reason:

        flash(
            "请填写作废原因。",
            "danger"
        )

        return redirect(
            url_for("finance.records")
        )

    # 暂时使用 session 中的财政用户
    # 如果没有登录名字，则显示 Finance User
    cancelled_by = get_current_finance_user()

    # =====================================================
    # 作废 finance_records
    # 不再 DELETE
    # =====================================================
    db_query("""
        update finance_records
        set
            status = 'cancelled',
            cancel_reason = %s,
            cancelled_by = %s,
            cancelled_at = now()
        where id = %s
          and coalesce(status, 'confirmed') != 'cancelled'
    """, (
        cancel_reason,
        cancelled_by,
        record_id,
    ))

    # =====================================================
    # 若这笔记录有收条编号，
    # 同步作废 member_payments
    # 不再永久删除月费记录
    # =====================================================
    if record.get("receipt_no"):

        db_query("""
            update member_payments
            set
                status = 'cancelled',
                cancel_reason = %s,
                cancelled_by = %s,
                cancelled_at = now()
            where receipt_no = %s
              and coalesce(status, 'active') != 'cancelled'
        """, (
            cancel_reason,
            cancelled_by,
            record["receipt_no"],
        ))

    updated_record = db_query("""
        select *
        from finance_records
        where id = %s
        limit 1
    """, (record_id,), fetchone=True)

    write_finance_audit(
        module="finance_records",
        action="cancel",
        record_id=record_id,
        old_value=dict(record),
        new_value=dict(updated_record) if updated_record else None,
        reason=cancel_reason,
        actor=cancelled_by,
    )

    flash(
        "财政记录已成功作废，原始资料仍保留在系统中。",
        "success"
    )

    return redirect(
        url_for("finance.records")
    )

@finance_bp.route("/admin/cash-in-hand")
@finance_admin_required
def finance_cash_in_hand():
    """Legacy compatibility redirect. Formal page moved to finance_v7.py."""
    return redirect(url_for("finance_v7.finance_v7_cash_in_hand"))

@finance_bp.route("/records")
def records():

    q = request.args.get("q", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    record_type = request.args.get("record_type", "").strip()
    category = request.args.get("category", "").strip()
    number_type = request.args.get("number_type", "").strip().upper()
    status_filter = request.args.get("status", "").strip()

    where_parts = []
    params = []

    if q:
        keyword = f"%{q}%"
        where_parts.append("""
            (
                coalesce(receipt_no, '') ilike %s
                or coalesce(payment_voucher_no, '') ilike %s
                or coalesce(member_id, '') ilike %s
                or coalesce(name, '') ilike %s
                or coalesce(phone, '') ilike %s
                or coalesce(category, '') ilike %s
                or coalesce(payment_method, '') ilike %s
                or coalesce(bank_ref, '') ilike %s
                or coalesce(remarks, '') ilike %s
                or coalesce(fund_account, '') ilike %s
            )
        """)
        params.extend([keyword] * 10)

    if date_from:
        where_parts.append("record_date >= %s")
        params.append(date_from)

    if date_to:
        where_parts.append("record_date <= %s")
        params.append(date_to)

    if record_type in ["income", "expense"]:
        where_parts.append("record_type = %s")
        params.append(record_type)

    if category:
        where_parts.append("category = %s")
        params.append(category)

    if number_type == "CHE":
        where_parts.append(
            "coalesce(receipt_no, '') ilike %s"
        )
        params.append("CHE%")

    elif number_type == "STW":
        where_parts.append(
            "coalesce(receipt_no, '') ilike %s"
        )
        params.append("STW%")

    elif number_type == "PV":
        where_parts.append("""
            coalesce(payment_voucher_no, '') <> ''
        """)

    elif number_type == "CDM":
        where_parts.append("""
            coalesce(bank_ref, '') ilike %s
        """)
        params.append("CDM-%")

    if status_filter == "active":
        where_parts.append("""
            coalesce(status, 'confirmed') <> 'cancelled'
        """)

    elif status_filter == "cancelled":
        where_parts.append("status = 'cancelled'")

    where_sql = ""

    if where_parts:
        where_sql = " where " + " and ".join(where_parts)

    rows = db_query(
        f"""
            select *
            from finance_records
            {where_sql}
            order by record_date desc, id desc
            limit 500
        """,
        tuple(params),
        fetchall=True
    )

    summary = db_query(
        f"""
            select
                count(*) as shown_count,

                /*
                 * 观音堂日常户口收入：
                 * 只计算 CHE 月费。
                 *
                 * receipt_no / member_id 可覆盖正常录入与部分历史导入；
                 * fund_account 是额外保险条件，但明确排除 STW 编号。
                 */
                coalesce(sum(
                    case
                        when record_type = 'income'
                        and category = '月费'
                        and coalesce(status, 'confirmed') <> 'cancelled'
                        and (
                                coalesce(receipt_no, '') ilike 'CHE%%'
                            or coalesce(member_id, '') ilike 'CHE%%'
                        )
                        then amount
                        else 0
                    end
                ), 0) as che_monthly_income_total,

                coalesce(sum(
                    case
                        when record_type = 'income'
                        and category = '月费'
                        and coalesce(status, 'confirmed') <> 'cancelled'
                        and (
                                coalesce(receipt_no, '') ilike 'STW%%'
                            or coalesce(member_id, '') ilike 'STW%%'
                        )
                        then amount
                        else 0
                    end
                ), 0) as stw_monthly_income_total,

                /*
                 * 总会户口收入：
                 * 财布施、观音村、膳食结缘、纯檀香布施及临时特别布施。
                 * 不把月费放进总会户口。
                 */
                coalesce(sum(
                    case
                        when record_type = 'income'
                         and coalesce(status, 'confirmed') <> 'cancelled'
                         and category <> '月费'
                         and (
                                fund_account = '总会户口'
                             or category in (
                                    '财布施',
                                    '观音村',
                                    '膳食结缘',
                                    '初一十五',
                                    '观音堂纯檀香布施',
                                    '临时特别布施'
                                )
                         )
                        then amount
                        else 0
                    end
                ), 0) as head_office_income_total,

                /*
                 * 观音堂日常支出：
                 * 所有有效支出。付款来源以后再区分银行或 Petty Cash。
                 */
                coalesce(sum(
                    case
                        when record_type = 'expense'
                         and coalesce(status, 'confirmed') <> 'cancelled'
                        then amount
                        else 0
                    end
                ), 0) as expense_total,

                coalesce(sum(
                    case
                        when coalesce(status, 'confirmed') <> 'cancelled'
                        then 1
                        else 0
                    end
                ), 0) as active_count,

                coalesce(sum(
                    case
                        when status = 'cancelled'
                        then 1
                        else 0
                    end
                ), 0) as cancelled_count
            from finance_records
            {where_sql}
        """,
        tuple(params),
        fetchone=True
    )

    categories = db_query("""
        select distinct category
        from finance_records
        where category is not null
          and trim(category) <> ''
        order by category
    """, fetchall=True)

    shown_count = int(summary["shown_count"] or 0)
    active_count = int(summary["active_count"] or 0)
    cancelled_count = int(summary["cancelled_count"] or 0)
    
    head_office_income_total = float(
        summary["head_office_income_total"] or 0
    )

    che_monthly_income_total = float(
        summary["che_monthly_income_total"] or 0
    )

    stw_monthly_income_total = float(
        summary["stw_monthly_income_total"] or 0
    )

    expense_total = float(
        summary["expense_total"] or 0
    )

    che_daily_balance_total = (
        che_monthly_income_total - expense_total
    )

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>财政总账</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .records-page{
                max-width:1500px;
            }

            .records-header{
                background:linear-gradient(135deg,#2563eb,#1d4ed8);
                color:#fff;
                padding:28px;
                border-radius:22px;
                margin-bottom:20px;
                box-shadow:0 12px 30px rgba(37,99,235,.18);
            }

            .records-header h1{
                margin:0 0 8px;
                font-size:32px;
            }

            .records-header p{
                margin:0;
                opacity:.92;
                line-height:1.6;
            }

            .quick-filter-row{
                display:flex;
                flex-wrap:wrap;
                gap:10px;
                margin-top:18px;
            }

            .quick-filter-btn{
                display:inline-flex;
                align-items:center;
                justify-content:center;
                min-height:42px;
                padding:8px 14px;
                border-radius:999px;
                background:rgba(255,255,255,.16);
                color:#fff;
                text-decoration:none;
                font-weight:800;
                border:1px solid rgba(255,255,255,.28);
            }

            .quick-filter-btn:hover{
                background:rgba(255,255,255,.25);
            }

            .filter-grid{
                display:grid;
                grid-template-columns:repeat(4,minmax(0,1fr));
                gap:14px;
            }

            .filter-grid .form-group{
                margin:0;
            }

            .filter-full{
                grid-column:1 / -1;
            }

            .filter-actions{
                display:flex;
                flex-wrap:wrap;
                gap:10px;
                margin-top:18px;
            }

            .records-summary{
                display:grid;
                grid-template-columns:repeat(5,minmax(0,1fr));
                gap:16px;
                margin:20px 0;
            }

            .records-summary .summary-box{
                position:relative;
                overflow:hidden;
                min-height:150px;
                padding:22px 22px 20px;
                border:1px solid #e2e8f0;
                border-radius:20px;
                background:#fff;
                box-shadow:0 8px 24px rgba(15,23,42,.07);
                display:flex;
                flex-direction:column;
                align-items:flex-start;
                justify-content:center;
                text-align:left;
            }

            .records-summary .summary-box::before{
                content:"";
                position:absolute;
                top:0;
                left:0;
                bottom:0;
                width:6px;
                background:#2563eb;
            }

            .summary-daily::before{
                background:#16a34a !important;
            }

            .summary-head-office::before{
                background:#7c3aed !important;
            }

            .summary-expense-card::before{
                background:#dc2626 !important;
            }

            .summary-balance-card::before{
                background:#2563eb !important;
            }

            .summary-cash-card::before{
                background:#d97706 !important;
            }

            .summary-record-card::before{
                background:#64748b !important;
            }

            .summary-icon{
                font-size:25px;
                margin-bottom:10px;
            }

            .summary-label{
                color:#64748b;
                font-size:16px;
                font-weight:800;
                margin-bottom:8px;
            }

            .summary-help{
                color:#94a3b8;
                font-size:13px;
                line-height:1.5;
                margin-top:8px;
            }

            .summary-value{
                color:#0f172a;
                font-size:28px;
                font-weight:900;
                line-height:1.2;
                word-break:break-word;
            }

            .summary-income{
                color:#15803d;
            }

            .summary-head-office-value{
                color:#7c3aed;
            }

            .summary-expense{
                color:#b91c1c;
            }

            .summary-balance{
                color:#1d4ed8;
            }

            .cash-in-hand-button{
                display:inline-flex;
                align-items:center;
                justify-content:center;
                width:100%;
                min-height:46px;
                margin-top:13px;
                padding:8px 14px;
                border:0;
                border-radius:13px;
                background:linear-gradient(135deg,#f59e0b,#d97706);
                color:#fff;
                font-size:16px;
                font-weight:900;
                cursor:pointer;
                box-shadow:0 8px 18px rgba(217,119,6,.2);
            }

            .cash-in-hand-button:hover{
                transform:translateY(-1px);
            }

            .table-topbar{
                display:flex;
                align-items:center;
                justify-content:space-between;
                flex-wrap:wrap;
                gap:12px;
                margin-bottom:14px;
            }

            .result-text{
                color:#64748b;
                font-size:16px;
            }

            .record-table{
                min-width:1320px;
            }

            .record-table th{
                white-space:nowrap;
            }

            .record-table td{
                vertical-align:middle;
            }

            .record-link{
                color:#1d4ed8;
                font-weight:900;
                text-decoration:none;
                white-space:nowrap;
            }

            .record-link:hover{
                text-decoration:underline;
            }

            .type-badge,
            .number-badge,
            .status-badge{
                display:inline-flex;
                align-items:center;
                justify-content:center;
                padding:7px 11px;
                border-radius:999px;
                font-size:14px;
                font-weight:900;
                white-space:nowrap;
            }

            .type-income{
                background:#dcfce7;
                color:#166534;
            }

            .type-expense{
                background:#fee2e2;
                color:#991b1b;
            }

            .number-che{
                background:#dbeafe;
                color:#1d4ed8;
            }

            .number-stw{
                background:#fee2e2;
                color:#b91c1c;
            }

            .number-pv{
                background:#f3e8ff;
                color:#7e22ce;
            }

            .status-confirmed{
                background:#dcfce7;
                color:#166534;
            }

            .status-cancelled{
                background:#fee2e2;
                color:#991b1b;
            }

            .status-pending{
                background:#fef3c7;
                color:#92400e;
            }

            .member-name{
                font-weight:800;
                color:#1e293b;
            }

            .member-id{
                color:#64748b;
                font-size:14px;
                margin-top:3px;
            }

            .money{
                font-weight:900;
                white-space:nowrap;
                text-align:right;
            }

            .money-income{
                color:#15803d;
            }

            .money-expense{
                color:#b91c1c;
            }

            .month-purpose{
                min-width:180px;
                line-height:1.5;
            }

            .remarks-cell{
                min-width:150px;
                max-width:250px;
                white-space:normal;
                line-height:1.5;
            }

            .cancelled-row{
                background:#f8fafc;
                color:#94a3b8;
            }

            .cancelled-row td:not(.action-cell){
                text-decoration:line-through;
            }

            .cancelled-row .status-badge,
            .cancelled-row .record-link,
            .cancelled-row .type-badge,
            .cancelled-row .number-badge{
                text-decoration:none;
            }

            .row-actions{
                display:flex;
                align-items:center;
                gap:8px;
                white-space:nowrap;
            }

            .mini-action{
                min-height:38px;
                padding:8px 12px;
                font-size:14px;
            }

            .cancel-form{
                display:flex;
                gap:8px;
                align-items:center;
            }

            .cancel-form .form-input{
                min-width:150px;
                min-height:38px;
                padding:7px 9px;
                font-size:14px;
                margin:0;
            }

            .empty-state-custom{
                text-align:center;
                padding:55px 20px;
                color:#64748b;
            }

            .bottom-actions{
                margin-top:20px;
            }

            @media(max-width:900px){
                .filter-grid{
                    grid-template-columns:1fr 1fr;
                }

                .records-summary{
                    grid-template-columns:1fr 1fr;
                }
            }

            @media(max-width:620px){
                .records-page{
                    padding-left:12px;
                    padding-right:12px;
                }

                .records-header{
                    padding:22px 18px;
                    border-radius:18px;
                }

                .records-header h1{
                    font-size:27px;
                }

                .filter-grid,
                .records-summary{
                    grid-template-columns:1fr;
                }

                .filter-full{
                    grid-column:auto;
                }

                .filter-actions{
                    display:grid;
                }

                .filter-actions .btn-tool,
                .bottom-actions .btn-tool{
                    width:100%;
                }
            }
        </style>
    </head>

    <body>

    <div class="page records-page">

        <div class="records-header">
            <h1>📚 财政总账</h1>

            <p>
                收条、Payment Voucher、收入、支出、月费月份和各类布施，都可在这里查询。
            </p>

            <div class="quick-filter-row">

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records') }}">
                    全部
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', record_type='income') }}">
                    收入
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', record_type='expense') }}">
                    支出
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', category='月费') }}">
                    月费
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', category='财布施') }}">
                    财布施
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', category='观音村') }}">
                    观音村
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', category='膳食结缘') }}">
                    初一十五
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', number_type='PV') }}">
                    PV
                </a>

                <a class="quick-filter-btn"
                href="{{ url_for('finance.records', number_type='CDM') }}">
                    CDM
                </a>

            </div>
        </div>

        <div class="card">
            <div class="section-title">
                🔎 查询条件
            </div>

            <form method="get">
                <div class="filter-grid">

                    <div class="form-group">
                        <label class="form-label">开始日期</label>

                        <input
                            class="form-input"
                            type="date"
                            name="date_from"
                            value="{{ date_from }}"
                        >
                    </div>

                    <div class="form-group">
                        <label class="form-label">结束日期</label>

                        <input
                            class="form-input"
                            type="date"
                            name="date_to"
                            value="{{ date_to }}"
                        >
                    </div>

                    <div class="form-group">
                        <label class="form-label">记录类型</label>

                        <select class="form-input" name="record_type">
                            <option value="">全部</option>
                            <option value="income"
                                {% if record_type == 'income' %}selected{% endif %}>
                                收入
                            </option>
                            <option value="expense"
                                {% if record_type == 'expense' %}selected{% endif %}>
                                支出
                            </option>
                        </select>
                    </div>

                    <div class="form-group">
                        <label class="form-label">编号类型</label>

                        <select class="form-input" name="number_type">
                            <option value="">全部</option>
                            <option value="CHE"
                                {% if number_type == 'CHE' %}selected{% endif %}>
                                CHE 收条
                            </option>
                            <option value="STW"
                                {% if number_type == 'STW' %}selected{% endif %}>
                                STW 收条
                            </option>
                            <option value="PV"
                                {% if number_type == 'PV' %}selected{% endif %}>
                                Payment Voucher
                            </option>
                        </select>
                    </div>
                                  
                    
                    <div class="form-group">
                        <label class="form-label">类别</label>

                        <select class="form-input" name="category">
                            <option value="">全部类别</option>

                            {% for c in categories %}
                                <option
                                    value="{{ c.category }}"
                                    {% if category == c.category %}selected{% endif %}
                                >
                                    {{ c.category }}
                                </option>
                            {% endfor %}
                        </select>
                    </div>

                    <div class="form-group">
                        <label class="form-label">状态</label>

                        <select class="form-input" name="status">
                            <option value="">全部状态</option>
                            <option value="active"
                                {% if status_filter == 'active' %}selected{% endif %}>
                                正常记录
                            </option>
                            <option value="cancelled"
                                {% if status_filter == 'cancelled' %}selected{% endif %}>
                                已作废
                            </option>
                        </select>
                    </div>

                    <div class="form-group filter-full">
                        <label class="form-label">关键字</label>

                        <input
                            class="form-input"
                            name="q"
                            value="{{ q }}"
                            placeholder="收条、PV、会员编号、姓名、电话、类别、付款方式、Reference、备注"
                            autocomplete="off"
                        >
                    </div>
                </div>

                <div class="filter-actions">
                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                    >
                        🔍 查询记录
                    </button>

                    <a
                        class="btn-tool btn-secondary"
                        href="{{ url_for('finance.records') }}"
                    >
                        ✕ 清除条件
                    </a>
                </div>
            </form>
        </div>

        <div class="records-summary">

            <div class="summary-box summary-daily">
                <div class="summary-icon">🟢</div>
                <div class="summary-label">
                    CHE 观音堂月费收入
                </div>
                <div class="summary-value summary-income">
                    RM {{ "%.2f"|format(che_monthly_income_total) }}
                </div>
                <div class="summary-help">
                    只计算 CHE 月费
                </div>
            </div>

            <div class="summary-box summary-stw">
                <div class="summary-icon">🔵</div>
                <div class="summary-label">
                    STW 分会月费收入
                </div>
                <div class="summary-value" style="color:#2563eb;">
                    RM {{ "%.2f"|format(stw_monthly_income_total) }}
                </div>
                <div class="summary-help">
                    只计算 STW 月费，不进入 CHE 户口
                </div>
            </div>

            <div class="summary-box summary-head-office">
                <div class="summary-icon">🟣</div>
                <div class="summary-label">
                    总会户口布施
                </div>
                <div class="summary-value summary-head-office-value">
                    RM {{ "%.2f"|format(head_office_income_total) }}
                </div>
                <div class="summary-help">
                    财布施、观音村、膳食结缘及其他布施
                </div>
            </div>

            <div class="summary-box summary-expense-card">
                <div class="summary-icon">🔴</div>
                <div class="summary-label">
                    观音堂日常支出
                </div>
                <div class="summary-value summary-expense">
                    RM {{ "%.2f"|format(expense_total) }}
                </div>
                <div class="summary-help">
                    包含银行付款与现金付款
                </div>
            </div>
            
            <div class="summary-box summary-cash-card">
                <div class="summary-icon">💵</div>
                <div class="summary-label">
                    Cash In Hand
                </div>
                <div class="summary-value" style="color:#b45309;">
                    待对账
                </div>
                <a
                    href="{{ url_for('finance.finance_cash_in_hand') }}"
                    class="cash-in-hand-button"
                >
                    💵 查看 Cash In Hand
                </a>
            </div>
                                  
            <div class="summary-box summary-balance-card">
                <div class="summary-icon">🏦</div>
                <div class="summary-label">
                    CHE 日常户口账面差额
                </div>
                <div class="summary-value summary-balance">
                    RM {{ "%.2f"|format(che_daily_balance_total) }}
                </div>
                <div class="summary-help">
                    CHE 月费减去 CHE 日常支出
                </div>
            </div>

            <div class="summary-box summary-record-card">
                <div class="summary-icon">📄</div>
                <div class="summary-label">
                    记录状态
                </div>
                <div class="summary-value">
                    {{ active_count }} 笔
                </div>
                <div class="summary-help">
                    正常 {{ active_count }} 笔 · 已作废 {{ cancelled_count }} 笔
                </div>
            </div>

        </div>

        <div class="card">
            <div class="table-topbar">
                <div class="section-title" style="margin-bottom:0;">
                    🧾 记录明细
                </div>

                <div class="result-text">
                    当前找到
                    <strong>{{ shown_count }}</strong>
                    笔记录
                </div>
            </div>

            {% if rows %}
                <div class="table-responsive">
                    <table class="record-table">
                        <thead>
                            <tr>
                                <th>日期</th>
                                <th>编号</th>
                                <th>类型</th>
                                <th>类别</th>
                                <th>会员／对象</th>
                                <th>月份／用途</th>
                                <th>金额</th>
                                <th>方式</th>
                                <th>状态</th>
                                <th>操作</th>
                            </tr>
                        </thead>

                        <tbody>
                            {% for r in rows %}
                                {% if r.record_type == 'expense' %}
                                    {% set display_no =
                                        r.payment_voucher_no
                                        or '-'
                                    %}
                                {% else %}
                                    {% set display_no =
                                        r.receipt_no
                                        or '-'
                                    %}
                                {% endif %}

                                <tr
                                    {% if r.status == 'cancelled' %}
                                        class="cancelled-row"
                                    {% endif %}
                                >
                                    <td style="white-space:nowrap;">
                                        {{ r.record_date }}
                                    </td>

                                    <td>
                                        <a
                                            class="record-link"
                                            href="{{ url_for(
                                                'finance.record_detail',
                                                record_id=r.id
                                            ) }}"
                                        >
                                            {{ display_no }}
                                        </a>

                                        <div style="margin-top:5px;">
                                            {% if r.payment_voucher_no %}
                                                <span class="number-badge number-pv">
                                                    PV
                                                </span>
                                            {% elif r.receipt_no and r.receipt_no.startswith('STW') %}
                                                <span class="number-badge number-stw">
                                                    STW
                                                </span>
                                            {% elif r.receipt_no and r.receipt_no.startswith('CHE') %}
                                                <span class="number-badge number-che">
                                                    CHE
                                                </span>
                                            {% endif %}
                                        </div>
                                    </td>

                                    <td>
                                        {% if r.record_type == 'expense' %}
                                            <span class="type-badge type-expense">
                                                支出
                                            </span>
                                        {% else %}
                                            <span class="type-badge type-income">
                                                收入
                                            </span>
                                        {% endif %}
                                    </td>

                                    <td>
                                        {{ r.category or '-' }}
                                    </td>

                                    <td>
                                        <div class="member-name">
                                            {{ r.name or '-' }}
                                        </div>

                                        {% if r.member_id %}
                                            <div class="member-id">
                                                {{ r.member_id }}
                                            </div>
                                        {% endif %}
                                    </td>

                                    <td class="month-purpose">
                                        {% if r.month_from or r.month_to %}
                                            {{ r.month_from or '-' }}
                                            至
                                            {{ r.month_to or '-' }}
                                        {% else %}
                                            {{ r.remarks or '-' }}
                                        {% endif %}
                                    </td>

                                    <td class="
                                        money
                                        {% if r.record_type == 'expense' %}
                                            money-expense
                                        {% else %}
                                            money-income
                                        {% endif %}
                                    ">
                                        {% if r.record_type == 'expense' %}
                                            −
                                        {% endif %}
                                        RM {{ "%.2f"|format(r.amount or 0) }}
                                    </td>

                                    <td>
                                        {{ r.payment_method or '-' }}
                                    </td>

                                    <td>
                                        {% if r.status == 'cancelled' %}
                                            <span class="status-badge status-cancelled">
                                                已作废
                                            </span>
                                        {% elif r.status == 'confirmed' or not r.status %}
                                            <span class="status-badge status-confirmed">
                                                已入账
                                            </span>
                                        {% else %}
                                            <span class="status-badge status-pending">
                                                待确认
                                            </span>
                                        {% endif %}
                                    </td>

                                    <td class="action-cell">
                                        <div class="row-actions">
                                            <a
                                                class="btn-tool btn-secondary mini-action"
                                                href="{{ url_for(
                                                    'finance.record_detail',
                                                    record_id=r.id
                                                ) }}"
                                            >
                                                查看
                                            </a>

                                            {% if r.status != 'cancelled' %}
                                                <form
                                                    class="cancel-form"
                                                    method="post"
                                                    action="{{ url_for(
                                                        'finance.cancel_record',
                                                        record_id=r.id
                                                    ) }}"
                                                    onsubmit="return confirm(
                                                        '确定要作废这笔记录吗？作废后不会计入财政统计。'
                                                    );"
                                                >
                                                    <input
                                                        class="form-input"
                                                        name="cancel_reason"
                                                        placeholder="作废原因"
                                                        required
                                                    >

                                                    <button
                                                        class="btn-tool btn-warning mini-action"
                                                        type="submit"
                                                    >
                                                        作废
                                                    </button>
                                                </form>
                                            {% endif %}
                                        </div>
                                    </td>
                                </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>

            {% else %}
                <div class="empty-state-custom">
                    <div style="font-size:48px;margin-bottom:10px;">
                        🔎
                    </div>

                    <h3>没有找到财政记录</h3>

                    <p>
                        请调整日期、类型、类别、编号或关键字。
                    </p>
                </div>
            {% endif %}
        </div>

        <div class="bottom-actions">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_admin_home') }}"
            >
                ← 返回负责人中心
            </a>
        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        categories=categories,
        q=q,
        date_from=date_from,
        date_to=date_to,
        record_type=record_type,
        category=category,
        number_type=number_type,
        status_filter=status_filter,
        shown_count=shown_count,
        active_count=active_count,
        cancelled_count=cancelled_count,
        che_monthly_income_total=che_monthly_income_total,
        stw_monthly_income_total=stw_monthly_income_total,
        head_office_income_total=head_office_income_total,
        expense_total=expense_total,
        che_daily_balance_total=che_daily_balance_total,
    )


@finance_bp.route("/records/<int:record_id>")
def record_detail(record_id):

    record = db_query("""
        select *
        from finance_records
        where id = %s
        limit 1
    """, (
        record_id,
    ), fetchone=True)

    if not record:
        return "找不到这笔财政记录", 404

    if record["record_type"] == "expense":
        display_no = (
            record.get("payment_voucher_no")
            or "-"
        )
        document_label = "Payment Voucher No."
        person_label = "付款对象"
    else:
        display_no = (
            record.get("receipt_no")
            or "-"
        )
        document_label = "Receipt No."
        person_label = "姓名"

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>财政记录详情</title>

        <link
            rel="stylesheet"
            href="{{ url_for(
                'static',
                filename='css/toolbox.css'
            ) }}"
        >

        <style>
            .detail-page{
                max-width:900px;
            }

            .detail-header{
                background:linear-gradient(
                    135deg,
                    #1d4ed8,
                    #2563eb
                );
                color:#fff;
                padding:28px;
                border-radius:22px;
                margin-bottom:20px;
                box-shadow:0 12px 30px rgba(37,99,235,.18);
            }

            .detail-document-label{
                font-size:14px;
                font-weight:800;
                opacity:.86;
                margin-bottom:6px;
                text-transform:uppercase;
                letter-spacing:.5px;
            }

            .detail-number{
                font-size:34px;
                font-weight:900;
                margin:0 0 10px;
                overflow-wrap:anywhere;
            }

            .detail-header-money{
                font-size:36px;
                font-weight:900;
                margin:4px 0 10px;
                line-height:1.2;
            }

            .detail-header-income{
                color:#bbf7d0;
            }

            .detail-header-expense{
                color:#fecaca;
            }

            .detail-subtitle{
                margin:0;
                opacity:.92;
                line-height:1.6;
            }

            .detail-grid{
                display:grid;
                grid-template-columns:repeat(2,minmax(0,1fr));
                gap:14px;
            }

            .detail-item{
                background:#f8fafc;
                border:1px solid #e2e8f0;
                border-radius:14px;
                padding:16px;
            }

            .detail-full{
                grid-column:1 / -1;
            }

            .detail-label{
                color:#64748b;
                font-size:14px;
                margin-bottom:6px;
            }

            .detail-value{
                color:#0f172a;
                font-size:19px;
                font-weight:800;
                line-height:1.5;
                overflow-wrap:anywhere;
            }

            .detail-money{
                color:#15803d;
                font-size:30px;
            }

            .detail-money.expense{
                color:#b91c1c;
            }

            .status-badge{
                display:inline-flex;
                align-items:center;
                justify-content:center;
                padding:8px 13px;
                border-radius:999px;
                font-size:15px;
                font-weight:900;
                white-space:nowrap;
            }

            .status-confirmed{
                background:#dcfce7;
                color:#166534;
            }

            .status-cancelled{
                background:#fee2e2;
                color:#991b1b;
            }

            .status-pending{
                background:#fef3c7;
                color:#92400e;
            }

            .cancel-box{
                background:#fef2f2;
                border-color:#fecaca;
            }

            .detail-actions{
                display:flex;
                flex-wrap:wrap;
                gap:10px;
                margin-top:20px;
            }

            @media print{

                body{
                    background:#fff;
                }

                .detail-actions{
                    display:none;
                }

                .detail-page{
                    max-width:none;
                    padding:0;
                }

                .detail-header,
                .card{
                    box-shadow:none;
                }
            }

            @media(max-width:650px){

                .detail-page{
                    padding-left:12px;
                    padding-right:12px;
                }

                .detail-header{
                    padding:22px 18px;
                    border-radius:18px;
                }

                .detail-number{
                    font-size:28px;
                }

                .detail-header-money{
                    font-size:30px;
                }

                .detail-grid{
                    grid-template-columns:1fr;
                }

                .detail-full{
                    grid-column:auto;
                }

                .detail-actions{
                    display:grid;
                }

                .detail-actions .btn-tool{
                    width:100%;
                }
            }
        </style>
    </head>

    <body>

    <div class="page detail-page">

        <div class="detail-header">

            <div class="detail-document-label">
                {{ document_label }}
            </div>

            <div class="detail-number">
                {{ display_no }}
            </div>

            <div class="
                detail-header-money
                {% if record.record_type == 'expense' %}
                    detail-header-expense
                {% else %}
                    detail-header-income
                {% endif %}
            ">
                {% if record.record_type == "expense" %}
                    −
                {% endif %}
                RM {{ "%.2f"|format(record.amount or 0) }}
            </div>

            <p class="detail-subtitle">
                {{ "支出记录" if record.record_type == "expense" else "收入记录" }}
                ·
                {{ record.category or "未分类" }}
            </p>

        </div>

        <div class="card">

            <h2 class="section-title">
                📄 完整记录
            </h2>

            <div class="detail-grid">

                <div class="detail-item">
                    <div class="detail-label">
                        记录类型
                    </div>

                    <div class="detail-value">
                        {{ "支出" if record.record_type == "expense" else "收入" }}
                    </div>
                </div>

                <div class="detail-item">
                    <div class="detail-label">
                        类别
                    </div>

                    <div class="detail-value">
                        {{ record.category or "-" }}
                    </div>
                </div>

                <div class="detail-item">
                    <div class="detail-label">
                        {{ "支出日期" if record.record_type == "expense" else "付款日期" }}
                    </div>

                    <div class="detail-value">
                        {{ record.record_date or "-" }}
                    </div>
                </div>

                {% if record.record_type != "expense" %}

                    <div class="detail-item">
                        <div class="detail-label">
                            开收条日期
                        </div>

                        <div class="detail-value">
                            {{ record.receipt_date or "-" }}
                        </div>
                    </div>

                {% endif %}

                <div class="detail-item">
                    <div class="detail-label">
                        {{ document_label }}
                    </div>

                    <div class="detail-value">
                        {{ display_no }}
                    </div>
                </div>

                {% if record.member_id %}

                    <div class="detail-item">
                        <div class="detail-label">
                            会员编号
                        </div>

                        <div class="detail-value">
                            {{ record.member_id }}
                        </div>
                    </div>

                {% endif %}

                <div class="detail-item">
                    <div class="detail-label">
                        {{ person_label }}
                    </div>

                    <div class="detail-value">
                        {{ record.name or "-" }}
                    </div>
                </div>

                {% if record.phone %}

                    <div class="detail-item">
                        <div class="detail-label">
                            电话号码
                        </div>

                        <div class="detail-value">
                            {{ record.phone }}
                        </div>
                    </div>

                {% endif %}

                <div class="detail-item">
                    <div class="detail-label">
                        付款方式
                    </div>

                    <div class="detail-value">
                        {{ record.payment_method or "-" }}
                    </div>
                </div>

                {% if record.bank_ref %}

                    <div class="detail-item">
                        <div class="detail-label">
                            银行 Reference
                        </div>

                        <div class="detail-value">
                            {{ record.bank_ref }}
                        </div>
                    </div>

                {% endif %}

                <div class="detail-item">
                    <div class="detail-label">
                        基金户口
                    </div>

                    <div class="detail-value">
                        {{ record.fund_account or "-" }}
                    </div>
                </div>

                {% if record.month_from or record.month_to %}

                    <div class="detail-item">
                        <div class="detail-label">
                            开始月份
                        </div>

                        <div class="detail-value">
                            {{ record.month_from or "-" }}
                        </div>
                    </div>

                    <div class="detail-item">
                        <div class="detail-label">
                            缴费至
                        </div>

                        <div class="detail-value">
                            {{ record.month_to or "-" }}
                        </div>
                    </div>

                {% endif %}

                <div class="detail-item detail-full">
                    <div class="detail-label">
                        金额
                    </div>

                    <div class="
                        detail-value
                        detail-money
                        {% if record.record_type == 'expense' %}
                            expense
                        {% endif %}
                    ">
                        {% if record.record_type == "expense" %}
                            −
                        {% endif %}
                        RM {{ "%.2f"|format(record.amount or 0) }}
                    </div>
                </div>

                <div class="detail-item detail-full">
                    <div class="detail-label">
                        备注／用途
                    </div>

                    <div class="detail-value">
                        {{ record.remarks or "-" }}
                    </div>
                </div>

                <div class="detail-item detail-full">
                    <div class="detail-label">
                        状态
                    </div>

                    <div class="detail-value">

                        {% if record.status == "cancelled" %}

                            <span class="status-badge status-cancelled">
                                ❌ 已作废
                            </span>

                        {% elif record.status == "confirmed"
                            or not record.status %}

                            <span class="status-badge status-confirmed">
                                ✅ 已入账
                            </span>

                        {% else %}

                            <span class="status-badge status-pending">
                                ⏳ {{ record.status }}
                            </span>

                        {% endif %}

                    </div>
                </div>

                {% if record.cancel_reason %}

                    <div class="
                        detail-item
                        detail-full
                        cancel-box
                    ">
                        <div class="detail-label">
                            作废原因
                        </div>

                        <div class="detail-value">
                            {{ record.cancel_reason }}
                        </div>
                    </div>

                {% endif %}

                {% if record.created_at %}

                    <div class="detail-item detail-full">
                        <div class="detail-label">
                            建立时间
                        </div>

                        <div class="detail-value">
                            {{ record.created_at }}
                        </div>
                    </div>

                {% endif %}

            </div>

        </div>

        <div class="detail-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.records') }}"
            >
                ← 返回财政总账
            </a>

            <a
                class="btn-tool btn-primary"
                href="{{ url_for(
                    'finance.records',
                    q=display_no
                ) }}"
            >
                🔍 查询相同编号
            </a>

            <button
                class="btn-tool btn-success"
                type="button"
                onclick="window.print()"
            >
                🖨️ 打印
            </button>

        </div>

    </div>

    </body>
    </html>
    """,
        record=record,
        display_no=display_no,
        document_label=document_label,
        person_label=person_label
    )

@finance_bp.route("/dashboard")
@finance_admin_required
def dashboard():

    ym = request.args.get(
        "ym",
        date.today().strftime("%Y-%m")
    )

    daily_income = db_query("""
        select
            category,
            coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'income'
        group by category
        order by category
    """, (ym,), fetchall=True)

    daily_expense = db_query("""
        select
            category,
            coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'expense'
        group by category
        order by category
    """, (ym,), fetchall=True)

    hq_income = db_query("""
        select
            category,
            coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '总会户口'
          and record_type = 'income'
        group by category
        order by category
    """, (ym,), fetchall=True)

    daily_income_total = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'income'
    """, (ym,), fetchone=True)

    daily_expense_total = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'expense'
    """, (ym,), fetchone=True)

    hq_income_total = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '总会户口'
          and record_type = 'income'
    """, (ym,), fetchone=True)

    selected_month = datetime.strptime(
        ym,
        "%Y-%m"
    ).date()

    previous_month_last_day = (
        selected_month.replace(day=1)
        - timedelta(days=1)
    )

    previous_ym = previous_month_last_day.strftime("%Y-%m")

    previous_close = get_month_close_record(
        previous_ym,
        "观音堂日常户口"
    )

    opening_balance = float(
        (previous_close or {}).get("closing_balance") or 0
    )

    daily_income_value = float(
        daily_income_total["total"] or 0
    )

    daily_expense_value = float(
        daily_expense_total["total"] or 0
    )

    hq_income_value = float(
        hq_income_total["total"] or 0
    )

    daily_balance = (
        opening_balance
        + daily_income_value
        - daily_expense_value
    )

    monthly_net = (
        daily_income_value
        - daily_expense_value
    )

    return render_template_string(FINANCE_DATE_COMPONENT + """
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>财政统计 Dashboard</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .dashboard-page {
                max-width: 1180px;
            }

            .dashboard-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;
                box-shadow:
                    0 12px 30px
                    rgba(37, 99, 235, 0.18);
            }

            .dashboard-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .dashboard-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .filter-grid {
                display:grid;
                grid-template-columns:minmax(260px,1fr) auto;
                gap:14px;
                align-items:end;
            }

            .filter-grid .form-group{
                margin:0;
            }

            .filter-submit-btn{
                min-height:52px;
                padding:10px 22px;
                white-space:nowrap;
            }

            .opening-balance-bar{
                display:flex;
                align-items:center;
                justify-content:space-between;
                gap:16px;
                margin-top:16px;
                padding:14px 16px;
                border:1px solid #dbe5ef;
                border-radius:14px;
                background:#f8fafc;
            }

            .opening-balance-label{
                color:#64748b;
                font-size:15px;
                font-weight:700;
            }

            .opening-balance-value{
                color:#0f172a;
                font-size:22px;
                font-weight:900;
                white-space:nowrap;
            }

            .opening-balance-help{
                margin-top:4px;
                color:#64748b;
                font-size:13px;
            }

            .finance-float-actions{
                position:fixed;
                right:22px;
                bottom:22px;
                z-index:9999;
                display:flex;
                flex-direction:column;
                gap:10px;
            }

            .finance-float-btn{
                width:54px;
                height:54px;
                display:flex;
                align-items:center;
                justify-content:center;
                border:0;
                border-radius:16px;
                box-shadow:0 8px 24px rgba(15,23,42,.18);
                cursor:pointer;
                text-decoration:none;
                font-size:23px;
                font-weight:900;
            }

            .finance-float-top{
                background:#2563eb;
                color:#fff;
            }

            .finance-float-back{
                background:#fff;
                color:#3157a4;
            }

            .finance-float-top.is-hidden{
                opacity:0;
                pointer-events:none;
                transform:translateY(8px);
            }

            .dashboard-summary {
                display: grid;
                grid-template-columns:
                    repeat(4, minmax(0, 1fr));
                gap: 16px;
                margin-bottom: 20px;
            }

            .dashboard-summary .summary-box {
                min-height: 130px;
                display: flex;
                flex-direction: column;
                justify-content: center;
                text-align: center;
            }

            .summary-icon {
                font-size: 30px;
                margin-bottom: 6px;
            }

            .summary-label {
                color: #64748b;
                font-size: 16px;
                margin-bottom: 7px;
            }

            .summary-value {
                color: #0f172a;
                font-size: 25px;
                font-weight: 800;
                white-space: nowrap;
            }

            .summary-positive {
                color: #15803d;
            }

            .summary-negative {
                color: #b91c1c;
            }

            .account-grid {
                display: grid;
                grid-template-columns:
                    minmax(0, 1fr)
                    minmax(0, 1fr);
                gap: 20px;
                align-items: start;
            }

            .account-title {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 12px;
                margin-bottom: 14px;
            }

            .account-title h2 {
                margin: 0;
                font-size: 23px;
                color: #1e293b;
            }

            .account-badge {
                display: inline-flex;
                align-items: center;
                padding: 7px 12px;
                border-radius: 999px;
                background: #dbeafe;
                color: #1d4ed8;
                font-size: 14px;
                font-weight: 700;
                white-space: nowrap;
            }

            .account-badge.hq {
                background: #ede9fe;
                color: #6d28d9;
            }

            .money {
                text-align: right;
                font-weight: 700;
                white-space: nowrap;
            }

            .money-income {
                color: #15803d;
            }

            .money-expense {
                color: #b91c1c;
            }

            .total-row {
                background: #f8fafc;
                font-weight: 800;
            }

            .balance-row {
                background: #eff6ff;
                font-weight: 800;
            }

            .dashboard-actions {
                display: flex;
                justify-content: space-between;
                gap: 12px;
                flex-wrap: wrap;
                margin-top: 20px;
            }

            .empty-row {
                text-align: center;
                padding: 28px 12px;
                color: #64748b;
            }

            .section-gap {
                margin-top: 20px;
            }

            @media (max-width: 900px) {

                .dashboard-summary {
                    grid-template-columns:
                        repeat(2, minmax(0, 1fr));
                }

                .account-grid {
                    grid-template-columns: 1fr;
                }
            }

            @media (max-width: 650px) {

                .dashboard-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .dashboard-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .dashboard-header h1 {
                    font-size: 26px;
                }

                .filter-grid {
                    grid-template-columns:1fr;
                }

                .filter-grid .btn-tool{
                    width:100%;
                }

                .opening-balance-bar{
                    align-items:flex-start;
                    flex-direction:column;
                    gap:6px;
                }

                .finance-float-actions{
                    right:14px;
                    bottom:14px;
                }

                .finance-float-btn{
                    width:48px;
                    height:48px;
                    border-radius:14px;
                    font-size:20px;
                }

                .dashboard-summary {
                    grid-template-columns: 1fr;
                }

                .dashboard-summary .summary-box {
                    min-height: 110px;
                }

                .dashboard-actions {
                    flex-direction: column;
                }

                .dashboard-actions .btn-tool {
                    width: 100%;
                }

                .record-table {
                    min-width: 620px;
                }
            }
        </style>
    </head>

    <body>

    <div class="page dashboard-page">

        <div class="dashboard-header">

            <h1>📊 财政统计 Dashboard</h1>

            <p>
                查看每月收入、支出、户口结余及总会户口收入。
            </p>

        </div>

        <div class="card">

            <div class="section-title">
                🔎 查询条件
            </div>

            <form method="get">

                <div class="filter-grid">

                    <div class="form-group">
                        <label class="form-label">月份</label>
                        <input
                            class="form-input"
                            type="month"
                            name="ym"
                            value="{{ ym }}"
                            required
                        >
                    </div>

                    <button
                        class="btn-tool btn-primary filter-submit-btn"
                        type="submit"
                    >
                        🔍 查看
                    </button>

                </div>

                <div class="opening-balance-bar">
                    <div>
                        <div class="opening-balance-label">
                            上月结余（自动）
                        </div>
                        <div class="opening-balance-help">
                            读取 {{ previous_ym }} 的月结余额
                        </div>
                    </div>

                    <div class="opening-balance-value">
                        RM {{ "%.2f"|format(opening_balance) }}
                    </div>
                </div>

            </form>

        </div>

        <div class="dashboard-summary">

            <div class="summary-box">

                <div class="summary-icon">
                    💰
                </div>

                <div class="summary-label">
                    日常户口收入
                </div>

                <div class="summary-value summary-positive">
                    RM {{ "%.2f"|format(daily_income_value) }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    💸
                </div>

                <div class="summary-label">
                    日常户口支出
                </div>

                <div class="summary-value summary-negative">
                    RM {{ "%.2f"|format(daily_expense_value) }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    📈
                </div>

                <div class="summary-label">
                    本月收支差额
                </div>

                <div class="
                    summary-value
                    {% if monthly_net >= 0 %}
                        summary-positive
                    {% else %}
                        summary-negative
                    {% endif %}
                ">
                    RM {{ "%.2f"|format(monthly_net) }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    🏦
                </div>

                <div class="summary-label">
                    日常户口结余
                </div>

                <div class="
                    summary-value
                    {% if daily_balance >= 0 %}
                        summary-positive
                    {% else %}
                        summary-negative
                    {% endif %}
                ">
                    RM {{ "%.2f"|format(daily_balance) }}
                </div>

            </div>

        </div>

        <div class="account-grid">

            <div class="card">

                <div class="account-title">

                    <h2>
                        🏠 观音堂日常户口
                    </h2>

                    <span class="account-badge">
                        收入与支出
                    </span>

                </div>

                <div class="section-title">
                    📥 收入明细
                </div>

                <div class="table-responsive">

                    <table class="record-table">

                        <thead>
                            <tr>
                                <th>收入项目</th>
                                <th>金额 RM</th>
                            </tr>
                        </thead>

                        <tbody>

                            {% for r in daily_income %}

                            <tr>
                                <td>
                                    {{ r.category }}
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(r.total) }}
                                </td>
                            </tr>

                            {% else %}

                            <tr>
                                <td
                                    colspan="2"
                                    class="empty-row"
                                >
                                    本月暂无收入记录
                                </td>
                            </tr>

                            {% endfor %}

                            <tr class="total-row">
                                <td>
                                    日常户口总收入
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(daily_income_value) }}
                                </td>
                            </tr>

                        </tbody>

                    </table>

                </div>

                <div class="section-title section-gap">
                    📤 支出明细
                </div>

                <div class="table-responsive">

                    <table class="record-table">

                        <thead>
                            <tr>
                                <th>支出项目</th>
                                <th>金额 RM</th>
                            </tr>
                        </thead>

                        <tbody>

                            {% for r in daily_expense %}

                            <tr>
                                <td>
                                    {{ r.category }}
                                </td>

                                <td class="money money-expense">
                                    {{ "%.2f"|format(r.total) }}
                                </td>
                            </tr>

                            {% else %}

                            <tr>
                                <td
                                    colspan="2"
                                    class="empty-row"
                                >
                                    本月暂无支出记录
                                </td>
                            </tr>

                            {% endfor %}

                            <tr class="total-row">
                                <td>
                                    日常户口总支出
                                </td>

                                <td class="money money-expense">
                                    {{ "%.2f"|format(daily_expense_value) }}
                                </td>
                            </tr>

                            <tr class="balance-row">
                                <td>
                                    上月结余
                                </td>

                                <td class="money">
                                    {{ "%.2f"|format(opening_balance) }}
                                </td>
                            </tr>

                            <tr class="balance-row">
                                <td>
                                    本月结余
                                </td>

                                <td class="
                                    money
                                    {% if daily_balance >= 0 %}
                                        money-income
                                    {% else %}
                                        money-expense
                                    {% endif %}
                                ">
                                    {{ "%.2f"|format(daily_balance) }}
                                </td>
                            </tr>

                        </tbody>

                    </table>

                </div>

            </div>

            <div class="card">

                <div class="account-title">

                    <h2>
                        🏛️ 总会户口
                    </h2>

                    <span class="account-badge hq">
                        收入统计
                    </span>

                </div>

                <div class="section-title">
                    📥 收入明细
                </div>

                <div class="table-responsive">

                    <table class="record-table">

                        <thead>
                            <tr>
                                <th>收入项目</th>
                                <th>金额 RM</th>
                            </tr>
                        </thead>

                        <tbody>

                            {% for r in hq_income %}

                            <tr>
                                <td>
                                    {{ r.category }}
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(r.total) }}
                                </td>
                            </tr>

                            {% else %}

                            <tr>
                                <td
                                    colspan="2"
                                    class="empty-row"
                                >
                                    本月暂无总会户口收入
                                </td>
                            </tr>

                            {% endfor %}

                            <tr class="total-row">
                                <td>
                                    总会户口总收入
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(hq_income_value) }}
                                </td>
                            </tr>

                        </tbody>

                    </table>

                </div>

            </div>

        </div>

        <div class="dashboard-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_admin_home') }}"
            >
                ← 返回负责人中心
            </a>

            <a
                class="btn-tool btn-success"
                href="{{ url_for(
                    'finance.export_monthly_report',
                    ym=ym
                ) }}"
            >
                📥 下载专业版 Excel 月报
            </a>

        </div>

    </div>

    <div class="finance-float-actions">
        <button
            type="button"
            class="finance-float-btn finance-float-top is-hidden"
            id="dashboardBackToTop"
            title="回到顶部"
            aria-label="回到顶部"
        >↑</button>

        <a
            class="finance-float-btn finance-float-back"
            href="{{ url_for('finance.finance_admin_home') }}"
            title="返回负责人中心"
            aria-label="返回负责人中心"
        >←</a>
    </div>

    <script>
    (function(){
        const topButton = document.getElementById("dashboardBackToTop");

        if(!topButton){
            return;
        }

        function updateTopButton(){
            topButton.classList.toggle(
                "is-hidden",
                window.scrollY <= 320
            );
        }

        window.addEventListener("scroll", updateTopButton, {passive:true});

        topButton.addEventListener("click", function(){
            window.scrollTo({top:0, behavior:"smooth"});
        });

        updateTopButton();
    })();
    </script>

    </body>
    </html>
    """,
        ym=ym,
        previous_ym=previous_ym,
        opening_balance=opening_balance,
        daily_income=daily_income,
        daily_expense=daily_expense,
        hq_income=hq_income,
        daily_income_value=daily_income_value,
        daily_expense_value=daily_expense_value,
        hq_income_value=hq_income_value,
        daily_balance=daily_balance,
        monthly_net=monthly_net
    )


@finance_bp.route("/export_monthly_report")
def export_monthly_report():

    ym = request.args.get(
        "ym",
        date.today().strftime("%Y-%m")
    )

    rows = db_query("""
        select
            record_date,
            receipt_date,
            receipt_no,
            payment_voucher_no,
            category,
            record_type,
            fund_account,
            member_id,
            name,
            phone,
            amount,
            payment_method,
            bank_ref,
            month_from,
            month_to,
            remarks as note
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
        order by
            record_date,
            case
                when record_type = 'expense'
                then payment_voucher_no
                else receipt_no
            end,
            id
    """, (
        ym,
    ), fetchall=True)

    wb = Workbook()

    ws = wb.active
    ws.title = "财政摘要"

    ws_income = wb.create_sheet(
        "收入明细"
    )

    ws_expense = wb.create_sheet(
        "支出明细"
    )

    # =========================
    # 样式
    # =========================
    green = "1F5E20"
    light_green = "EAF4E4"

    blue = "0B3A78"
    light_blue = "EAF2FF"

    gold = "F8E7B8"
    light_gold = "FFF8E8"

    dark_text = "1F2937"
    red = "D00000"
    gray = "64748B"

    thin_green = Side(
        style="thin",
        color="8DBA7C"
    )

    thin_blue = Side(
        style="thin",
        color="7FA6D9"
    )

    thin_gold = Side(
        style="thin",
        color="C9962C"
    )

    thin_gray = Side(
        style="thin",
        color="CCCCCC"
    )

    border_green = Border(
        left=thin_green,
        right=thin_green,
        top=thin_green,
        bottom=thin_green
    )

    border_blue = Border(
        left=thin_blue,
        right=thin_blue,
        top=thin_blue,
        bottom=thin_blue
    )

    border_gold = Border(
        left=thin_gold,
        right=thin_gold,
        top=thin_gold,
        bottom=thin_gold
    )

    border_gray = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray
    )

    money_fmt = '"RM"#,##0.00'

    def money(value):
        return float(value or 0)

    def set_range_border(
        sheet,
        cell_range,
        border
    ):
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = border

    def fill_range(
        sheet,
        cell_range,
        fill
    ):
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = fill

    def auto_width(
        sheet,
        max_width=32
    ):
        for column_cells in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is None:
                    continue

                cell_text = str(cell.value)

                for line in cell_text.splitlines():
                    max_length = max(
                        max_length,
                        len(line)
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 4,
                max_width
            )

    def set_print_layout(
        sheet,
        orientation="landscape"
    ):
        sheet.page_setup.orientation = orientation
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.4
        sheet.page_margins.bottom = 0.4
        sheet.page_margins.header = 0.2
        sheet.page_margins.footer = 0.2

    # =========================
    # 数据整理
    # =========================
    summary = {
        "观音堂日常户口": {
            "income": 0,
            "expense": 0
        },
        "总会户口": {
            "income": 0,
            "expense": 0
        },
    }

    income_rows = []
    expense_rows = []

    income_by_category = {}
    expense_by_category = {}

    for row in rows:

        fund_account = (
            row["fund_account"]
            or "未分类"
        )

        record_type = (
            row["record_type"]
            or "income"
        )

        amount = money(
            row["amount"]
        )

        if fund_account not in summary:

            summary[fund_account] = {
                "income": 0,
                "expense": 0
            }

        summary[fund_account][
            record_type
        ] += amount

        category_name = (
            row["category"]
            or "未分类"
        )

        if record_type == "income":

            income_rows.append(row)

            income_by_category[
                category_name
            ] = (
                income_by_category.get(
                    category_name,
                    0
                )
                + amount
            )

        else:

            expense_rows.append(row)

            expense_by_category[
                category_name
            ] = (
                expense_by_category.get(
                    category_name,
                    0
                )
                + amount
            )

    daily_income = summary.get(
        "观音堂日常户口",
        {}
    ).get(
        "income",
        0
    )

    daily_expense = summary.get(
        "观音堂日常户口",
        {}
    ).get(
        "expense",
        0
    )

    daily_balance = (
        daily_income
        - daily_expense
    )

    hq_income = summary.get(
        "总会户口",
        {}
    ).get(
        "income",
        0
    )

    hq_expense = summary.get(
        "总会户口",
        {}
    ).get(
        "expense",
        0
    )

    hq_balance = (
        hq_income
        - hq_expense
    )

    total_income = sum(
        item["income"]
        for item in summary.values()
    )

    total_expense = sum(
        item["expense"]
        for item in summary.values()
    )

    total_balance = (
        total_income
        - total_expense
    )

    # =========================
    # Sheet 1：财政摘要
    # =========================
    ws.sheet_view.showGridLines = False

    for column in range(1, 16):
        ws.column_dimensions[
            get_column_letter(column)
        ].width = 14

    for row_no in range(1, 40):
        ws.row_dimensions[row_no].height = 24

    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 12

    ws.merge_cells("A1:O4")

    ws["A1"] = (
        "观音堂财政月报\n"
        f"{ym}"
    )

    ws["A1"].font = Font(
        bold=True,
        size=24,
        color=dark_text
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=light_gold
    )

    set_range_border(
        ws,
        "A1:O4",
        border_gold
    )

    # 观音堂日常户口
    ws.merge_cells("A5:G5")

    ws["A5"] = "观音堂日常户口"

    ws["A5"].fill = PatternFill(
        "solid",
        fgColor=green
    )

    ws["A5"].font = Font(
        bold=True,
        size=14,
        color="FFFFFF"
    )

    ws["A5"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    daily_table = [
        [
            "项目",
            "收入（RM）",
            "支出（RM）",
            "本月结余（RM）"
        ],
        [
            "收入",
            daily_income,
            "-",
            daily_income
        ],
        [
            "支出",
            "-",
            daily_expense,
            -daily_expense
        ],
        [
            "本月结余",
            "-",
            "-",
            daily_balance
        ],
    ]

    start_row = 6

    for row_no, row_data in enumerate(
        daily_table,
        start_row
    ):

        ws.merge_cells(
            start_row=row_no,
            start_column=1,
            end_row=row_no,
            end_column=2
        )

        ws.merge_cells(
            start_row=row_no,
            start_column=3,
            end_row=row_no,
            end_column=4
        )

        ws.merge_cells(
            start_row=row_no,
            start_column=5,
            end_row=row_no,
            end_column=5
        )

        ws.merge_cells(
            start_row=row_no,
            start_column=6,
            end_row=row_no,
            end_column=7
        )

        ws.cell(
            row_no,
            1
        ).value = row_data[0]

        ws.cell(
            row_no,
            3
        ).value = row_data[1]

        ws.cell(
            row_no,
            5
        ).value = row_data[2]

        ws.cell(
            row_no,
            6
        ).value = row_data[3]

        for column in [
            1,
            3,
            5,
            6
        ]:

            cell = ws.cell(
                row_no,
                column
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = border_green

        if row_no == 6:

            fill_range(
                ws,
                f"A{row_no}:G{row_no}",
                PatternFill(
                    "solid",
                    fgColor=light_green
                )
            )

            for column in [
                1,
                3,
                5,
                6
            ]:

                ws.cell(
                    row_no,
                    column
                ).font = Font(
                    bold=True
                )

        elif row_no == 9:

            fill_range(
                ws,
                f"A{row_no}:G{row_no}",
                PatternFill(
                    "solid",
                    fgColor="DDEED2"
                )
            )

            ws.cell(
                row_no,
                6
            ).font = Font(
                bold=True,
                size=14,
                color=green
            )

    for cell_ref in [
        "C7",
        "E8",
        "F7",
        "F8",
        "F9"
    ]:
        ws[cell_ref].number_format = money_fmt

    # 总会户口
    ws.merge_cells("I5:O5")

    ws["I5"] = "总会户口"

    ws["I5"].fill = PatternFill(
        "solid",
        fgColor=blue
    )

    ws["I5"].font = Font(
        bold=True,
        size=14,
        color="FFFFFF"
    )

    ws["I5"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    hq_table = [
        [
            "项目",
            "收入（RM）",
            "支出（RM）",
            "本月结余（RM）"
        ],
        [
            "收入",
            hq_income,
            "-",
            hq_income
        ],
        [
            "支出",
            "-",
            hq_expense,
            -hq_expense
        ],
        [
            "本月结余",
            "-",
            "-",
            hq_balance
        ],
    ]

    for row_no, row_data in enumerate(
        hq_table,
        start_row
    ):

        ws.merge_cells(
            start_row=row_no,
            start_column=9,
            end_row=row_no,
            end_column=10
        )

        ws.merge_cells(
            start_row=row_no,
            start_column=11,
            end_row=row_no,
            end_column=12
        )

        ws.merge_cells(
            start_row=row_no,
            start_column=13,
            end_row=row_no,
            end_column=13
        )

        ws.merge_cells(
            start_row=row_no,
            start_column=14,
            end_row=row_no,
            end_column=15
        )

        ws.cell(
            row_no,
            9
        ).value = row_data[0]

        ws.cell(
            row_no,
            11
        ).value = row_data[1]

        ws.cell(
            row_no,
            13
        ).value = row_data[2]

        ws.cell(
            row_no,
            14
        ).value = row_data[3]

        for column in [
            9,
            11,
            13,
            14
        ]:

            cell = ws.cell(
                row_no,
                column
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = border_blue

        if row_no == 6:

            fill_range(
                ws,
                f"I{row_no}:O{row_no}",
                PatternFill(
                    "solid",
                    fgColor=light_blue
                )
            )

            for column in [
                9,
                11,
                13,
                14
            ]:

                ws.cell(
                    row_no,
                    column
                ).font = Font(
                    bold=True,
                    color=blue
                )

        elif row_no == 9:

            fill_range(
                ws,
                f"I{row_no}:O{row_no}",
                PatternFill(
                    "solid",
                    fgColor="DCEAFF"
                )
            )

            ws.cell(
                row_no,
                14
            ).font = Font(
                bold=True,
                size=14,
                color=blue
            )

    for cell_ref in [
        "K7",
        "M8",
        "N7",
        "N8",
        "N9"
    ]:
        ws[cell_ref].number_format = money_fmt

    # 总计横条
    ws.merge_cells("A11:C12")

    ws["A11"] = "总计"

    ws["A11"].font = Font(
        bold=True,
        size=16,
        color="7A5200"
    )

    ws["A11"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws["A11"].fill = PatternFill(
        "solid",
        fgColor="FFF2CC"
    )

    ws.merge_cells("D11:F12")

    ws["D11"] = (
        "总收入（RM）\n"
        f"{total_income:,.2f}"
    )

    ws["D11"].font = Font(
        bold=True,
        size=12,
        color=green
    )

    ws["D11"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    ws["D11"].fill = PatternFill(
        "solid",
        fgColor=light_gold
    )

    ws.merge_cells("G11:I12")

    ws["G11"] = (
        "总支出（RM）\n"
        f"{total_expense:,.2f}"
    )

    ws["G11"].font = Font(
        bold=True,
        size=12,
        color=red
    )

    ws["G11"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    ws["G11"].fill = PatternFill(
        "solid",
        fgColor=light_gold
    )

    ws.merge_cells("J11:O12")

    ws["J11"] = (
        "总体结余（RM）\n"
        f"{total_balance:,.2f}"
    )

    ws["J11"].font = Font(
        bold=True,
        size=12,
        color=blue
    )

    ws["J11"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    ws["J11"].fill = PatternFill(
        "solid",
        fgColor=light_gold
    )

    set_range_border(
        ws,
        "A11:O12",
        border_gold
    )

    # 收入分类汇总
    ws.merge_cells("A14:G14")

    ws["A14"] = "收入分类汇总"

    ws["A14"].fill = PatternFill(
        "solid",
        fgColor=green
    )

    ws["A14"].font = Font(
        bold=True,
        size=14,
        color="FFFFFF"
    )

    ws["A14"].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells("A15:E15")
    ws.merge_cells("F15:G15")

    ws["A15"] = "收入项目"
    ws["F15"] = "金额（RM）"

    for cell_ref in [
        "A15",
        "F15"
    ]:

        ws[cell_ref].fill = PatternFill(
            "solid",
            fgColor=light_green
        )

        ws[cell_ref].font = Font(
            bold=True
        )

        ws[cell_ref].alignment = Alignment(
            horizontal="center"
        )

    income_summary_start = 16
    income_summary_row = income_summary_start

    for category_name, amount in sorted(
        income_by_category.items(),
        key=lambda item: item[1],
        reverse=True
    ):

        ws.merge_cells(
            start_row=income_summary_row,
            start_column=1,
            end_row=income_summary_row,
            end_column=5
        )

        ws.merge_cells(
            start_row=income_summary_row,
            start_column=6,
            end_row=income_summary_row,
            end_column=7
        )

        ws.cell(
            income_summary_row,
            1
        ).value = category_name

        ws.cell(
            income_summary_row,
            6
        ).value = amount

        ws.cell(
            income_summary_row,
            6
        ).number_format = money_fmt

        ws.cell(
            income_summary_row,
            1
        ).alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        ws.cell(
            income_summary_row,
            6
        ).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        income_summary_row += 1

    if not income_by_category:

        ws.merge_cells("A16:G16")

        ws["A16"] = "本月没有收入记录"

        ws["A16"].alignment = Alignment(
            horizontal="center"
        )

        income_summary_row = 17

    income_total_row = income_summary_row

    ws.merge_cells(
        start_row=income_total_row,
        start_column=1,
        end_row=income_total_row,
        end_column=5
    )

    ws.merge_cells(
        start_row=income_total_row,
        start_column=6,
        end_row=income_total_row,
        end_column=7
    )

    ws.cell(
        income_total_row,
        1
    ).value = "收入总计"

    ws.cell(
        income_total_row,
        6
    ).value = total_income

    ws.cell(
        income_total_row,
        6
    ).number_format = money_fmt

    fill_range(
        ws,
        f"A{income_total_row}:G{income_total_row}",
        PatternFill(
            "solid",
            fgColor="DDEED2"
        )
    )

    ws.cell(
        income_total_row,
        1
    ).font = Font(
        bold=True,
        color=green
    )

    ws.cell(
        income_total_row,
        6
    ).font = Font(
        bold=True,
        color=green
    )

    ws.cell(
        income_total_row,
        1
    ).alignment = Alignment(
        horizontal="right"
    )

    ws.cell(
        income_total_row,
        6
    ).alignment = Alignment(
        horizontal="center"
    )

    set_range_border(
        ws,
        f"A15:G{income_total_row}",
        border_green
    )

    # 支出分类汇总
    ws.merge_cells("I14:O14")

    ws["I14"] = "支出分类汇总"

    ws["I14"].fill = PatternFill(
        "solid",
        fgColor=blue
    )

    ws["I14"].font = Font(
        bold=True,
        size=14,
        color="FFFFFF"
    )

    ws["I14"].alignment = Alignment(
        horizontal="center"
    )

    ws.merge_cells("I15:M15")
    ws.merge_cells("N15:O15")

    ws["I15"] = "支出项目"
    ws["N15"] = "金额（RM）"

    for cell_ref in [
        "I15",
        "N15"
    ]:

        ws[cell_ref].fill = PatternFill(
            "solid",
            fgColor=light_blue
        )

        ws[cell_ref].font = Font(
            bold=True,
            color=blue
        )

        ws[cell_ref].alignment = Alignment(
            horizontal="center"
        )

    expense_summary_start = 16
    expense_summary_row = expense_summary_start

    for category_name, amount in sorted(
        expense_by_category.items(),
        key=lambda item: item[1],
        reverse=True
    ):

        ws.merge_cells(
            start_row=expense_summary_row,
            start_column=9,
            end_row=expense_summary_row,
            end_column=13
        )

        ws.merge_cells(
            start_row=expense_summary_row,
            start_column=14,
            end_row=expense_summary_row,
            end_column=15
        )

        ws.cell(
            expense_summary_row,
            9
        ).value = category_name

        ws.cell(
            expense_summary_row,
            14
        ).value = amount

        ws.cell(
            expense_summary_row,
            14
        ).number_format = money_fmt

        ws.cell(
            expense_summary_row,
            9
        ).alignment = Alignment(
            horizontal="left",
            vertical="center"
        )

        ws.cell(
            expense_summary_row,
            14
        ).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        expense_summary_row += 1

    if not expense_by_category:

        ws.merge_cells("I16:O16")

        ws["I16"] = "本月没有支出记录"

        ws["I16"].alignment = Alignment(
            horizontal="center"
        )

        expense_summary_row = 17

    expense_total_row = expense_summary_row

    ws.merge_cells(
        start_row=expense_total_row,
        start_column=9,
        end_row=expense_total_row,
        end_column=13
    )

    ws.merge_cells(
        start_row=expense_total_row,
        start_column=14,
        end_row=expense_total_row,
        end_column=15
    )

    ws.cell(
        expense_total_row,
        9
    ).value = "支出总计"

    ws.cell(
        expense_total_row,
        14
    ).value = total_expense

    ws.cell(
        expense_total_row,
        14
    ).number_format = money_fmt

    fill_range(
        ws,
        f"I{expense_total_row}:O{expense_total_row}",
        PatternFill(
            "solid",
            fgColor="DCEAFF"
        )
    )

    ws.cell(
        expense_total_row,
        9
    ).font = Font(
        bold=True,
        color=blue
    )

    ws.cell(
        expense_total_row,
        14
    ).font = Font(
        bold=True,
        color=red
    )

    ws.cell(
        expense_total_row,
        9
    ).alignment = Alignment(
        horizontal="right"
    )

    ws.cell(
        expense_total_row,
        14
    ).alignment = Alignment(
        horizontal="center"
    )

    set_range_border(
        ws,
        f"I15:O{expense_total_row}",
        border_blue
    )

    max_summary_row = max(
        income_total_row,
        expense_total_row
    )

    ws.freeze_panes = "A14"

    ws.print_area = (
        f"A1:O{max_summary_row}"
    )

    set_print_layout(
        ws,
        orientation="landscape"
    )

    # =========================
    # Sheet 2：收入明细
    # =========================
    def write_income_sheet(
        sheet,
        data_rows
    ):

        sheet.sheet_view.showGridLines = False

        sheet.merge_cells("A1:N2")

        sheet["A1"] = (
            f"收入明细 - {ym}"
        )

        sheet["A1"].font = Font(
            bold=True,
            size=18,
            color="FFFFFF"
        )

        sheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        sheet["A1"].fill = PatternFill(
            "solid",
            fgColor=green
        )

        headers = [
            "付款日期",
            "开收条日期",
            "Receipt No.",
            "类别",
            "基金户口",
            "会员编号",
            "姓名",
            "电话",
            "金额（RM）",
            "付款方式",
            "银行 Reference",
            "开始月份",
            "缴费至",
            "备注",
        ]

        for column, header in enumerate(
            headers,
            1
        ):

            cell = sheet.cell(
                4,
                column
            )

            cell.value = header

            cell.fill = PatternFill(
                "solid",
                fgColor=light_green
            )

            cell.font = Font(
                bold=True,
                color=dark_text
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cell.border = border_gray

        row_no = 5

        for row in data_rows:

            values = [
                row["record_date"],
                row["receipt_date"],
                row["receipt_no"],
                row["category"],
                row["fund_account"],
                row["member_id"],
                row["name"],
                row["phone"],
                money(row["amount"]),
                row["payment_method"],
                row["bank_ref"],
                row["month_from"],
                row["month_to"],
                row["note"],
            ]

            for column, value in enumerate(
                values,
                1
            ):

                cell = sheet.cell(
                    row_no,
                    column
                )

                cell.value = value

                cell.border = border_gray

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=(
                        column in [
                            7,
                            14
                        ]
                    )
                )

                if column == 9:
                    cell.number_format = money_fmt

            row_no += 1

        if not data_rows:

            sheet.merge_cells(
                start_row=5,
                start_column=1,
                end_row=5,
                end_column=14
            )

            sheet.cell(
                5,
                1
            ).value = "本月没有收入记录"

            sheet.cell(
                5,
                1
            ).alignment = Alignment(
                horizontal="center"
            )

            row_no = 6

        total_row = row_no + 1

        sheet.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=8
        )

        sheet.cell(
            total_row,
            1
        ).value = "收入总计"

        sheet.cell(
            total_row,
            1
        ).font = Font(
            bold=True
        )

        sheet.cell(
            total_row,
            1
        ).alignment = Alignment(
            horizontal="right"
        )

        sheet.cell(
            total_row,
            9
        ).value = sum(
            money(row["amount"])
            for row in data_rows
        )

        sheet.cell(
            total_row,
            9
        ).number_format = money_fmt

        sheet.cell(
            total_row,
            9
        ).font = Font(
            bold=True,
            color=green
        )

        fill_range(
            sheet,
            f"A{total_row}:N{total_row}",
            PatternFill(
                "solid",
                fgColor=light_green
            )
        )

        set_range_border(
            sheet,
            f"A{total_row}:N{total_row}",
            border_gray
        )

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = (
            f"A4:N{max(4, row_no - 1)}"
        )

        sheet.print_title_rows = "1:4"
        sheet.print_area = (
            f"A1:N{total_row}"
        )

        auto_width(
            sheet,
            max_width=30
        )

        sheet.column_dimensions["N"].width = 34
        sheet.column_dimensions["G"].width = 18

        set_print_layout(
            sheet,
            orientation="landscape"
        )

    # =========================
    # Sheet 3：支出明细
    # =========================
    def write_expense_sheet(
        sheet,
        data_rows
    ):

        sheet.sheet_view.showGridLines = False

        sheet.merge_cells("A1:I2")

        sheet["A1"] = (
            f"支出明细 - {ym}"
        )

        sheet["A1"].font = Font(
            bold=True,
            size=18,
            color="FFFFFF"
        )

        sheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        sheet["A1"].fill = PatternFill(
            "solid",
            fgColor=blue
        )

        headers = [
            "支出日期",
            "Payment Voucher No.",
            "类别",
            "基金户口",
            "付款对象",
            "金额（RM）",
            "付款方式",
            "银行 Reference",
            "用途／备注",
        ]

        for column, header in enumerate(
            headers,
            1
        ):

            cell = sheet.cell(
                4,
                column
            )

            cell.value = header

            cell.fill = PatternFill(
                "solid",
                fgColor=light_blue
            )

            cell.font = Font(
                bold=True,
                color=dark_text
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cell.border = border_gray

        row_no = 5

        for row in data_rows:

            values = [
                row["record_date"],
                row["payment_voucher_no"],
                row["category"],
                row["fund_account"],
                row["name"],
                money(row["amount"]),
                row["payment_method"],
                row["bank_ref"],
                row["note"],
            ]

            for column, value in enumerate(
                values,
                1
            ):

                cell = sheet.cell(
                    row_no,
                    column
                )

                cell.value = value

                cell.border = border_gray

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=(
                        column in [
                            5,
                            9
                        ]
                    )
                )

                if column == 6:
                    cell.number_format = money_fmt

            row_no += 1

        if not data_rows:

            sheet.merge_cells(
                start_row=5,
                start_column=1,
                end_row=5,
                end_column=9
            )

            sheet.cell(
                5,
                1
            ).value = "本月没有支出记录"

            sheet.cell(
                5,
                1
            ).alignment = Alignment(
                horizontal="center"
            )

            row_no = 6

        total_row = row_no + 1

        sheet.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=5
        )

        sheet.cell(
            total_row,
            1
        ).value = "支出总计"

        sheet.cell(
            total_row,
            1
        ).font = Font(
            bold=True
        )

        sheet.cell(
            total_row,
            1
        ).alignment = Alignment(
            horizontal="right"
        )

        sheet.cell(
            total_row,
            6
        ).value = sum(
            money(row["amount"])
            for row in data_rows
        )

        sheet.cell(
            total_row,
            6
        ).number_format = money_fmt

        sheet.cell(
            total_row,
            6
        ).font = Font(
            bold=True,
            color=red
        )

        fill_range(
            sheet,
            f"A{total_row}:I{total_row}",
            PatternFill(
                "solid",
                fgColor=light_blue
            )
        )

        set_range_border(
            sheet,
            f"A{total_row}:I{total_row}",
            border_gray
        )

        sheet.freeze_panes = "A5"
        sheet.auto_filter.ref = (
            f"A4:I{max(4, row_no - 1)}"
        )

        sheet.print_title_rows = "1:4"
        sheet.print_area = (
            f"A1:I{total_row}"
        )

        auto_width(
            sheet,
            max_width=32
        )

        sheet.column_dimensions["I"].width = 36
        sheet.column_dimensions["E"].width = 22

        set_print_layout(
            sheet,
            orientation="landscape"
        )

    write_income_sheet(
        ws_income,
        income_rows
    )

    write_expense_sheet(
        ws_expense,
        expense_rows
    )

    # =========================
    # 输出
    # =========================
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    filename = (
        f"观音堂财政月报_{ym}.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

def get_next_payment_voucher():

    row = db_query("""
        select payment_voucher_no
        from finance_records
        where payment_voucher_no is not null
          and payment_voucher_no <> ''
        order by payment_voucher_no desc
        limit 1
    """, fetchone=True)

    if not row:
        return "PV000001"

    m = re.search(r"(\d+)$", row["payment_voucher_no"])

    if not m:
        return "PV000001"

    return f"PV{int(m.group(1))+1:06d}"


def ensure_finance_vendor_schema():
    """兼容旧版 finance_vendors，并补齐 V7 公司分类字段。"""
    db_query("""
        create table if not exists finance_vendors (
            id bigserial primary key,
            vendor_name varchar(220),
            is_active boolean not null default true,
            sort_order integer not null default 100,
            created_at timestamptz not null default now()
        )
    """)

    db_query("alter table finance_vendors add column if not exists company_name varchar(220)")
    db_query("alter table finance_vendors add column if not exists branch varchar(10) default 'CHE'")
    db_query("alter table finance_vendors add column if not exists expense_category varchar(120)")
    db_query("alter table finance_vendors add column if not exists contact_person varchar(160)")
    db_query("alter table finance_vendors add column if not exists phone varchar(80)")
    db_query("alter table finance_vendors add column if not exists email varchar(180)")
    db_query("alter table finance_vendors add column if not exists remarks text")
    db_query("alter table finance_vendors add column if not exists updated_at timestamptz default now()")

    db_query("""
        update finance_vendors
        set company_name = vendor_name
        where (company_name is null or trim(company_name) = '')
          and vendor_name is not null
    """)

    db_query("""
        update finance_vendors
        set vendor_name = company_name
        where (vendor_name is null or trim(vendor_name) = '')
          and company_name is not null
    """)

    db_query("alter table finance_records add column if not exists vendor_id bigint")
    db_query("alter table finance_records add column if not exists vendor_name varchar(220)")
    db_query("alter table finance_records add column if not exists prepared_by varchar(100)")
    db_query("alter table finance_records add column if not exists particular text")
    db_query("alter table finance_records add column if not exists reference_no varchar(180)")

    defaults = [
        ("Pudu Ria Florist Trading Sdn Bhd", "供花", 10),
        ("Laohongka Sdn Bhd", "供油", 10),
        ("Tenaga Nasional Berhad", "电费", 10),
        ("Air Selangor", "水费", 10),
        ("Indah Water Konsortium Sdn Bhd", "水费", 20),
        ("TM Unifi", "电话及网络费", 10),
        ("CelcomDigi", "电话及网络费", 20),
        ("Maxis", "电话及网络费", 30),
    ]

    for company_name, category, sort_order in defaults:
        exists = db_query("""
            select id
            from finance_vendors
            where branch = 'CHE'
              and lower(coalesce(company_name, vendor_name)) = lower(%s)
              and coalesce(expense_category, '') = %s
            limit 1
        """, (company_name, category), fetchone=True)

        if not exists:
            db_query("""
                insert into finance_vendors
                (
                    branch,
                    company_name,
                    vendor_name,
                    expense_category,
                    is_active,
                    sort_order
                )
                values ('CHE', %s, %s, %s, true, %s)
            """, (company_name, company_name, category, sort_order))


def get_finance_vendors_for_category(category, include_inactive=False):
    ensure_finance_vendor_schema()

    active_sql = "" if include_inactive else "and is_active = true"

    return db_query(f"""
        select
            id,
            coalesce(company_name, vendor_name) as company_name,
            coalesce(expense_category, '') as expense_category,
            contact_person,
            phone,
            email,
            remarks,
            is_active,
            sort_order,
            created_at
        from finance_vendors
        where branch = 'CHE'
          {active_sql}
          and (
                coalesce(expense_category, '') = %s
                or coalesce(expense_category, '') = ''
          )
        order by
            case when coalesce(expense_category, '') = %s then 0 else 1 end,
            sort_order,
            coalesce(company_name, vendor_name)
    """, (category, category), fetchall=True)


@finance_bp.route(
    "/vendors",
    methods=["GET", "POST"]
)
@finance_admin_required
def finance_vendors():
    ensure_finance_vendor_schema()

    message = ""
    error = ""
    categories = list(EXPENSE_SUB_CATEGORY_OPTIONS.keys())

    if request.method == "POST":
        action = request.form.get("action", "add").strip()

        try:
            if action == "add":
                company_name = request.form.get("company_name", "").strip()
                expense_category = request.form.get("expense_category", "").strip()
                phone = request.form.get("phone", "").strip()
                contact_person = request.form.get("contact_person", "").strip()
                sort_order = int(request.form.get("sort_order", "100") or 100)

                if not company_name:
                    raise ValueError("请填写公司名称。")
                if not expense_category:
                    raise ValueError("请选择适用支出类别。")

                existing = db_query("""
                    select id
                    from finance_vendors
                    where branch = 'CHE'
                      and lower(coalesce(company_name, vendor_name)) = lower(%s)
                      and coalesce(expense_category, '') = %s
                    limit 1
                """, (company_name, expense_category), fetchone=True)

                if existing:
                    raise ValueError("这个公司已经存在于该支出类别。")

                db_query("""
                    insert into finance_vendors
                    (
                        branch,
                        company_name,
                        vendor_name,
                        expense_category,
                        phone,
                        contact_person,
                        is_active,
                        sort_order
                    )
                    values ('CHE', %s, %s, %s, %s, %s, true, %s)
                """, (
                    company_name,
                    company_name,
                    expense_category,
                    phone or None,
                    contact_person or None,
                    sort_order,
                ))
                message = "已新增付款对象。"

            elif action == "edit":
                vendor_id = int(request.form.get("vendor_id"))
                company_name = request.form.get("company_name", "").strip()
                expense_category = request.form.get("expense_category", "").strip()
                phone = request.form.get("phone", "").strip()
                contact_person = request.form.get("contact_person", "").strip()
                sort_order = int(request.form.get("sort_order", "100") or 100)

                if not company_name:
                    raise ValueError("请填写公司名称。")

                db_query("""
                    update finance_vendors
                    set
                        company_name = %s,
                        vendor_name = %s,
                        expense_category = %s,
                        phone = %s,
                        contact_person = %s,
                        sort_order = %s,
                        updated_at = now()
                    where id = %s
                """, (
                    company_name,
                    company_name,
                    expense_category,
                    phone or None,
                    contact_person or None,
                    sort_order,
                    vendor_id,
                ))
                message = "付款对象资料已保存。"

            elif action == "toggle":
                vendor_id = int(request.form.get("vendor_id"))
                db_query("""
                    update finance_vendors
                    set is_active = not is_active,
                        updated_at = now()
                    where id = %s
                """, (vendor_id,))
                message = "付款对象状态已更新。"

        except (ValueError, TypeError) as exc:
            error = str(exc)

    rows = db_query("""
        select
            id,
            coalesce(company_name, vendor_name) as company_name,
            coalesce(expense_category, '') as expense_category,
            phone,
            contact_person,
            is_active,
            sort_order
        from finance_vendors
        where branch = 'CHE'
        order by
            is_active desc,
            expense_category,
            sort_order,
            coalesce(company_name, vendor_name)
    """, fetchall=True)

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>付款对象管理</title>
        <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
        <style>
            .vendor-page{max-width:1050px}
            .vendor-header{background:linear-gradient(135deg,#7c3aed,#5b21b6);color:#fff;padding:28px;border-radius:22px;margin-bottom:20px}
            .vendor-header h1{margin:0 0 8px}.vendor-header p{margin:0;opacity:.92}
            .vendor-grid{display:grid;grid-template-columns:2fr 1.3fr 1fr 1fr 90px auto;gap:10px;align-items:end}
            .vendor-row{display:grid;grid-template-columns:2fr 1.3fr 1fr 1fr 90px auto auto;gap:8px;align-items:center;padding:12px 0;border-bottom:1px solid #e5e7eb}
            .vendor-row input,.vendor-row select{width:100%;padding:10px;border:1px solid #d1d5db;border-radius:10px}
            .inactive{opacity:.55}
            @media(max-width:800px){.vendor-grid,.vendor-row{grid-template-columns:1fr}.vendor-row{padding:16px 0}}
        </style>
    </head>
    <body>
    <div class="page vendor-page">
        <div class="vendor-header">
            <h1>🏢 付款对象管理</h1>
            <p>此页供负责人维护公司资料；普通录入组员可在支出表单即时新增公司。</p>
        </div>
        {% if message %}<div class="alert alert-success">✅ {{ message }}</div>{% endif %}
        {% if error %}<div class="alert alert-danger">⚠️ {{ error }}</div>{% endif %}

        <div class="card">
            <div class="section-title">➕ 新增公司</div>
            <form method="post">
                <input type="hidden" name="action" value="add">
                <div class="vendor-grid">
                    <div><label class="form-label">完整公司名称</label><input class="form-input" name="company_name" required></div>
                    <div><label class="form-label">适用类别</label><select class="form-input" name="expense_category" required>{% for c in categories %}<option value="{{ c }}">{{ c }}</option>{% endfor %}</select></div>
                    <div><label class="form-label">电话</label><input class="form-input" name="phone"></div>
                    <div><label class="form-label">联络人</label><input class="form-input" name="contact_person"></div>
                    <div><label class="form-label">排序</label><input class="form-input" type="number" name="sort_order" value="100"></div>
                    <button class="btn-tool btn-primary" type="submit">保存</button>
                </div>
            </form>
        </div>

        <div class="card">
            <div class="section-title">📋 公司名单</div>
            {% for v in rows %}
            <form method="post" class="vendor-row {% if not v.is_active %}inactive{% endif %}">
                <input type="hidden" name="vendor_id" value="{{ v.id }}">
                <input name="company_name" value="{{ v.company_name }}" required>
                <select name="expense_category">{% for c in categories %}<option value="{{ c }}" {% if c == v.expense_category %}selected{% endif %}>{{ c }}</option>{% endfor %}</select>
                <input name="phone" value="{{ v.phone or '' }}" placeholder="电话">
                <input name="contact_person" value="{{ v.contact_person or '' }}" placeholder="联络人">
                <input name="sort_order" type="number" value="{{ v.sort_order }}">
                <button class="btn-tool btn-secondary" name="action" value="edit">保存</button>
                <button class="btn-tool btn-danger" name="action" value="toggle">{% if v.is_active %}停用{% else %}启用{% endif %}</button>
            </form>
            {% else %}<p>还没有付款对象。</p>{% endfor %}
        </div>

        <div class="btn-row"><a class="btn-tool btn-secondary" href="{{ url_for('finance.finance_expense_menu') }}">← 返回支出项目</a></div>
    </div>
    </body>
    </html>
    """, rows=rows, categories=categories, message=message, error=error)


@finance_bp.route("/vendors/<int:vendor_id>/toggle", methods=["POST"])
@finance_admin_required
def toggle_finance_vendor(vendor_id):
    ensure_finance_vendor_schema()
    db_query("""
        update finance_vendors
        set is_active = not is_active,
            updated_at = now()
        where id = %s
    """, (vendor_id,))
    return redirect(url_for("finance.finance_vendors"))




PV_PREPARED_BY_OPTIONS = [
    "陈柔霓",
    "黄薈菏",
    "",
]


def _register_pv_fonts():
    """Use ReportLab's built-in CJK font; no external font file is required."""
    try:
        pdfmetrics.getFont("STSong-Light")
    except KeyError:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))


def _find_finance_logo():
    """Find the same logo already stored in static for Excel exports."""
    static_folder = Path(current_app.static_folder or "static")

    preferred_names = [
        "finance_lotus_logo.png",
        "finance_hq_logo_cheras.png",
        "finance_hq_logo.png",
        "finance_logo.png",
        "finance_logo.jpg",
        "finance_logo.jpeg",
        "gyt_logo.png",
        "gyt_logo.jpg",
        "logo.png",
        "logo.jpg",
        "logo.jpeg",
    ]

    for name in preferred_names:
        candidate = static_folder / name
        if candidate.is_file():
            return candidate

    # Compatible fallback: search files whose names contain logo.
    for pattern in ("*logo*.png", "*logo*.jpg", "*logo*.jpeg"):
        matches = sorted(static_folder.rglob(pattern))
        if matches:
            return matches[0]

    return None


def _pv_safe_text(value):
    return str(value or "").strip()


def _pv_date_text(value):
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return _pv_safe_text(value)


def _draw_wrapped_text(pdf, text, x, y, max_width, font_name, font_size, leading, max_lines=4):
    """Draw Chinese/English text without overflowing the voucher box."""
    text = _pv_safe_text(text)
    if not text:
        return y

    lines = []
    current = ""

    for char in text:
        candidate = current + char
        if pdfmetrics.stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = char

        if len(lines) >= max_lines:
            break

    if current and len(lines) < max_lines:
        lines.append(current)

    if len(lines) == max_lines and len("".join(lines)) < len(text):
        last = lines[-1]
        while last and pdfmetrics.stringWidth(last + "...", font_name, font_size) > max_width:
            last = last[:-1]
        lines[-1] = last + "..."

    pdf.setFont(font_name, font_size)
    for line in lines:
        pdf.drawString(x, y, line)
        y -= leading

    return y


def _draw_single_payment_voucher(pdf, row, bottom_y, voucher_height, logo_path=None):
    """Draw one digital PV closely matching the organisation's original paper voucher."""
    _register_pv_fonts()

    page_width, _ = A4
    left = 7 * mm
    right = page_width - 7 * mm
    usable_width = right - left
    top = bottom_y + voucher_height - 2.5 * mm
    cjk = "STSong-Light"

    # ---------------------------- Letterhead ----------------------------
    logo_w = 29 * mm
    logo_h = 22 * mm
    logo_x = left + 2 * mm
    logo_y = top - logo_h

    if logo_path:
        try:
            pdf.drawImage(
                ImageReader(str(logo_path)),
                logo_x,
                logo_y,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                mask="auto",
                anchor="c",
            )
        except Exception:
            pass

    text_x = left + 34 * mm
    text_right = right - 2 * mm

    # English name: strong black, matching the original printed form.
    pdf.setFillColorRGB(0, 0, 0)
    pdf.setFont("Helvetica-Bold", 14.1)
    pdf.drawString(
        text_x,
        top - 3.7 * mm,
        "PERSATUAN PENGANUT GUAN YIN PH MALAYSIA",
    )

    # Full registration number, kept on its own line so it never overlaps.
    pdf.setFont("Helvetica-Bold", 7.2)
    pdf.drawRightString(
        text_right,
        top - 7.7 * mm,
        "(PPM-009-14-26062012-000025)",
    )

    # Correct traditional Chinese organisation name. Draw twice with a tiny
    # offset to reproduce the stronger weight of the original voucher.
    chinese_name = "馬來西亞盧台長心靈法門共修會"
    chinese_y = top - 12.4 * mm
    pdf.setFillColorRGB(0.55, 0.04, 0.04)
    pdf.setFont(cjk, 14.4)
    pdf.drawString(text_x, chinese_y, chinese_name)
    pdf.drawString(text_x + 0.22, chinese_y, chinese_name)

    # Cheras branch contact details.
    pdf.setFillColorRGB(0, 0, 0)
    pdf.setFont("Helvetica", 8.2)
    pdf.drawString(
        text_x,
        top - 17.3 * mm,
        "20-1, 20-2, Jalan Damai Perdana 1/8B, Bandar Damai Perdana, 56000 Kuala Lumpur.",
    )
    pdf.drawString(
        text_x,
        top - 21.1 * mm,
        "TEL: 03-91070300    Email: guanyinph.cheras@gmail.com",
    )

    # Branch block at the far right, as on the original form.
    branch_x = right - 37 * mm
    branch_y = top - 16.8 * mm
    pdf.setFont("Helvetica-Bold", 8.2)
    pdf.drawString(branch_x, branch_y, "BRANCH /")
    pdf.setFont(cjk, 8.2)
    pdf.drawString(branch_x + 15.4 * mm, branch_y, "分會：")
    pdf.setFont("Helvetica-Bold", 10.8)
    pdf.drawCentredString(right - 16 * mm, branch_y - 5.0 * mm, "CHERAS")
    pdf.setLineWidth(0.7)
    pdf.line(right - 31 * mm, branch_y - 6.7 * mm, right, branch_y - 6.7 * mm)

    header_line_y = top - 25.0 * mm
    pdf.setLineWidth(1.0)
    pdf.line(left, header_line_y, right, header_line_y)

    # ------------------------- Voucher heading --------------------------
    title_y = header_line_y - 7.0 * mm
    pdf.setFont("Helvetica-Bold", 17.2)
    pdf.drawCentredString(page_width / 2, title_y, "PAYMENT VOUCHER")

    # Paid To and No/Date rows.
    detail_y = title_y - 9.4 * mm
    pdf.setFont("Helvetica-Bold", 11.7)
    pdf.drawString(left, detail_y, "Paid to :")

    paid_line_x = left + 24 * mm
    paid_line_right = right - 61 * mm
    pdf.setLineWidth(0.75)
    pdf.line(paid_line_x, detail_y - 2.0 * mm, paid_line_right, detail_y - 2.0 * mm)
    _draw_wrapped_text(
        pdf,
        _pv_safe_text(row.get("vendor_name") or row.get("name")),
        paid_line_x + 3 * mm,
        detail_y + 0.15 * mm,
        paid_line_right - paid_line_x - 5 * mm,
        cjk,
        13.8,
        5.0 * mm,
        max_lines=1,
    )

    right_label_x = right - 56 * mm
    right_value_x = right - 34 * mm
    pdf.setFont("Helvetica-Bold", 11.0)
    pdf.drawString(right_label_x, detail_y + 3.1 * mm, "No :")
    pdf.drawString(right_label_x, detail_y - 3.8 * mm, "Date :")
    pdf.line(right_value_x, detail_y + 1.3 * mm, right, detail_y + 1.3 * mm)
    pdf.line(right_value_x, detail_y - 5.7 * mm, right, detail_y - 5.7 * mm)
    pdf.setFont("Helvetica", 12.0)
    pdf.drawString(
        right_value_x + 2.5 * mm,
        detail_y + 3.0 * mm,
        _pv_safe_text(row.get("payment_voucher_no")),
    )
    pdf.drawString(
        right_value_x + 2.5 * mm,
        detail_y - 3.9 * mm,
        _pv_date_text(row.get("record_date")),
    )

    # ---------------------- Particular / Amount box ---------------------
    table_top = detail_y - 10 * mm
    table_bottom = table_top - 33 * mm
    amount_x = right - 50 * mm
    pdf.setLineWidth(0.8)
    pdf.rect(left, table_bottom, usable_width, table_top - table_bottom)
    pdf.line(amount_x, table_bottom, amount_x, table_top)
    header_row_y = table_top - 8.5 * mm
    pdf.line(left, header_row_y, right, header_row_y)

    pdf.setFont("Helvetica-Bold", 11.6)
    pdf.drawCentredString((left + amount_x) / 2, table_top - 6.0 * mm, "Particular")
    pdf.drawCentredString((amount_x + right) / 2, table_top - 6.0 * mm, "Amount")

    particular_parts = []
    category_text = " / ".join(filter(None, [
        _pv_safe_text(row.get("category")),
        _pv_safe_text(row.get("sub_category")),
    ]))
    if category_text:
        particular_parts.append(category_text)
    if row.get("reference_no"):
        particular_parts.append("Reference No: " + _pv_safe_text(row.get("reference_no")))
    if row.get("remarks"):
        particular_parts.append("Remark: " + _pv_safe_text(row.get("remarks")))

    current_y = header_row_y - 7.0 * mm
    for paragraph in particular_parts:
        current_y = _draw_wrapped_text(
            pdf,
            paragraph,
            left + 4 * mm,
            current_y,
            amount_x - left - 8 * mm,
            cjk,
            11.8,
            5.3 * mm,
            max_lines=2,
        )
        current_y -= 0.7 * mm

    amount = Decimal(str(row.get("amount") or 0))
    pdf.setFont("Helvetica-Bold", 18.2)
    pdf.drawCentredString(
        (amount_x + right) / 2,
        header_row_y - 14.8 * mm,
        f"RM {amount:,.2f}",
    )

    # ---------------------------- Signatures ----------------------------
    sig_title_y = table_bottom - 7.2 * mm
    col_width = usable_width / 4
    columns = [
        (left + col_width * 0, "Prepared By:"),
        (left + col_width * 1, "Witness By:"),
        (left + col_width * 2, "Witness By:"),
        (left + col_width * 3, "Received By:"),
    ]

    pdf.setFont("Helvetica-Bold", 10.1)
    for x, title in columns:
        pdf.drawString(x, sig_title_y, title)
        pdf.line(x, sig_title_y - 11.3 * mm, x + col_width - 5 * mm, sig_title_y - 11.3 * mm)

    prepared_by = _pv_safe_text(row.get("prepared_by"))
    if prepared_by:
        pdf.setFont(cjk, 14.0)
        pdf.drawCentredString(
            left + (col_width - 5 * mm) / 2,
            sig_title_y - 8.1 * mm,
            prepared_by,
        )

    pdf.setFont("Helvetica", 9.0)
    for x, _ in columns:
        pdf.drawString(x, sig_title_y - 15.4 * mm, "Name :")

@finance_bp.route("/payment-voucher/<int:record_id>/pdf")
def payment_voucher_pdf(record_id):
    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    ensure_finance_vendor_schema()

    row = db_query("""
        select
            id,
            record_date,
            payment_voucher_no,
            category,
            sub_category,
            vendor_name,
            name,
            amount,
            remarks,
            prepared_by,
            particular,
            reference_no,
            status
        from finance_records
        where id = %s
          and record_type = 'expense'
        limit 1
    """, (record_id,), fetchone=True)

    if not row:
        abort(404)

    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    pdf.setTitle(_pv_safe_text(row.get("payment_voucher_no")) or "Payment Voucher")

    _, page_height = A4
    margin = 7 * mm
    gap = 6 * mm
    voucher_height = (page_height - (margin * 2) - gap) / 2
    logo_path = _find_finance_logo()

    _draw_single_payment_voucher(
        pdf,
        row,
        bottom_y=page_height - margin - voucher_height,
        voucher_height=voucher_height,
        logo_path=logo_path,
    )

    # Cut line between the two copies.
    cut_y = page_height / 2
    pdf.saveState()
    pdf.setDash(3, 3)
    pdf.setLineWidth(0.5)
    pdf.line(10 * mm, cut_y, A4[0] - 10 * mm, cut_y)
    pdf.setDash()
    pdf.setFont("Helvetica", 6.5)
    pdf.drawCentredString(A4[0] / 2, cut_y + 1.5 * mm, "CUT HERE")
    pdf.restoreState()

    _draw_single_payment_voucher(
        pdf,
        row,
        bottom_y=margin,
        voucher_height=voucher_height,
        logo_path=logo_path,
    )

    pdf.showPage()
    pdf.save()
    output.seek(0)

    download = request.args.get("download") == "1"
    filename = f"{_pv_safe_text(row.get('payment_voucher_no')) or 'payment_voucher'}.pdf"

    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=download,
        download_name=filename,
    )


@finance_bp.route("/payment-voucher/<int:record_id>")
def payment_voucher(record_id):
    """Payment Voucher 功能页：预览、打印与下载。"""
    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    ensure_finance_vendor_schema()

    row = db_query("""
        select
            id,
            record_date,
            payment_voucher_no,
            category,
            sub_category,
            vendor_name,
            name,
            amount,
            particular,
            reference_no,
            prepared_by,
            remarks,
            status
        from finance_records
        where id = %s
          and record_type = 'expense'
        limit 1
    """, (record_id,), fetchone=True)

    if not row:
        abort(404)

    pdf_url = url_for(
        "finance.payment_voucher_pdf",
        record_id=record_id,
    )

    download_url = url_for(
        "finance.payment_voucher_pdf",
        record_id=record_id,
        download=1,
    )

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{{ row.payment_voucher_no }} Payment Voucher</title>
        <link rel="stylesheet"
              href="{{ url_for('static', filename='css/toolbox.css') }}">

        <style>
            body{background:#f4f7fb;}
            .pv-page{max-width:1180px;}
            .pv-topbar{
                display:flex;
                justify-content:space-between;
                align-items:center;
                gap:12px;
                margin-bottom:18px;
            }
            .pv-hero{
                position:relative;
                overflow:hidden;
                padding:28px 30px;
                border-radius:24px;
                color:#fff;
                background:linear-gradient(135deg,#1769aa,#2387c9);
                box-shadow:0 12px 28px rgba(23,105,170,.20);
                margin-bottom:18px;
            }
            .pv-hero h1{margin:0;font-size:34px;}
            .pv-hero p{margin:8px 0 0;opacity:.92;font-size:17px;}
            .pv-number{
                display:inline-flex;
                margin-top:14px;
                padding:8px 14px;
                border-radius:999px;
                background:rgba(255,255,255,.17);
                font-size:18px;
                font-weight:900;
            }
            .pv-summary{
                display:grid;
                grid-template-columns:repeat(4,minmax(0,1fr));
                gap:12px;
                margin-bottom:18px;
            }
            .pv-summary-box{
                background:#fff;
                border:1px solid #e2e8f0;
                border-radius:16px;
                padding:15px;
                box-shadow:0 5px 16px rgba(15,23,42,.05);
            }
            .pv-summary-label{color:#64748b;font-size:13px;margin-bottom:5px;}
            .pv-summary-value{font-size:17px;font-weight:850;color:#172033;overflow-wrap:anywhere;}
            .pv-amount{color:#b91c1c;font-size:22px;}
            .pv-actions{
                display:grid;
                grid-template-columns:repeat(5,minmax(0,1fr));
                gap:12px;
                margin-bottom:18px;
            }
            .pv-actions .btn-tool{width:100%;min-height:55px;}
            .pv-preview-card{
                padding:14px;
                border-radius:20px;
                background:#fff;
                box-shadow:0 8px 25px rgba(15,23,42,.08);
            }
            .pv-frame{
                display:block;
                width:100%;
                height:88vh;min-height:820px;
                border:1px solid #dbe3ed;
                border-radius:14px;
                background:#eef2f7;
            }
            .pv-note{
                margin-top:12px;
                padding:12px 14px;
                border-radius:12px;
                background:#eff6ff;
                color:#1e40af;
                line-height:1.55;
            }
            @media(max-width:800px){
                .pv-page{padding:14px 12px 34px;}
                .pv-summary{grid-template-columns:1fr 1fr;}
                .pv-actions{grid-template-columns:1fr;}
                .pv-frame{height:650px;}
                .pv-topbar{display:grid;}
                .pv-topbar .btn-tool{width:100%;}
            }
            @media(max-width:520px){
                .pv-summary{grid-template-columns:1fr;}
                .pv-hero{padding:23px 19px;}
                .pv-hero h1{font-size:27px;}
                .pv-frame{height:560px;}
            }
        </style>
    </head>
    <body>
    <div class="page pv-page">

        <div class="pv-topbar">
            <a class="btn-tool btn-secondary"
               href="{{ url_for('finance.finance_expense_menu') }}">
                ← 返回支出录入
            </a>

            <a class="btn-tool btn-primary"
               href="{{ url_for('finance.expense', category=row.category) }}">
                ＋ 继续输入 {{ row.category }}
            </a>
        </div>

        <section class="pv-hero">
            <h1>🧾 Payment Voucher 已保存</h1>
            <p>请预览内容；确认无误后打印或下载 PDF。</p>
            <div class="pv-number">{{ row.payment_voucher_no or '-' }}</div>
        </section>

        <div class="pv-summary">
            <div class="pv-summary-box">
                <div class="pv-summary-label">日期</div>
                <div class="pv-summary-value">{{ row.record_date }}</div>
            </div>
            <div class="pv-summary-box">
                <div class="pv-summary-label">付款对象</div>
                <div class="pv-summary-value">{{ row.vendor_name or row.name or '-' }}</div>
            </div>
            <div class="pv-summary-box">
                <div class="pv-summary-label">类别</div>
                <div class="pv-summary-value">{{ row.category }}{% if row.sub_category %}／{{ row.sub_category }}{% endif %}</div>
            </div>
            <div class="pv-summary-box">
                <div class="pv-summary-label">金额</div>
                <div class="pv-summary-value pv-amount">RM {{ '%.2f'|format(row.amount or 0) }}</div>
            </div>
        </div>

        <div class="pv-actions">
            <a class="btn-tool btn-primary"
               href="{{ pdf_url }}"
               target="_blank"
               rel="noopener">
                👁 打开 PDF
            </a>

            <button class="btn-tool btn-success"
                    type="button"
                    onclick="printPaymentVoucher()">
                🖨 打印
            </button>

            <a class="btn-tool btn-purple"
               href="{{ download_url }}">
                📥 下载 PDF
            </a>

            <a class="btn-tool btn-secondary"
               href="{{ url_for('finance.record_detail', record_id=row.id) }}">
                🔎 查看记录
            </a>

            <form method="post"
                  action="{{ url_for('finance.delete_payment_voucher', record_id=row.id) }}"
                  onsubmit="return confirm('这会永久删除测试记录 {{ row.payment_voucher_no }}，确定继续？');"
                  style="margin:0;">
                <button class="btn-tool btn-danger" type="submit" style="width:100%;min-height:55px;">
                    🗑 删除测试记录
                </button>
            </form>
        </div>

        <div class="pv-preview-card">
            <iframe
                id="pv-frame"
                class="pv-frame"
                src="{{ pdf_url }}#toolbar=1&navpanes=0&view=FitH"
                title="Payment Voucher PDF Preview">
            </iframe>

            <div class="pv-note">
                PDF 内含上下两张相同的 Payment Voucher，中间有裁切线。Particular 由类别、Reference No. 与备注自动组成；Witness 与 Received By 保持空白。
            </div>
        </div>

    </div>

    <script>
        function printPaymentVoucher(){
            const frame = document.getElementById("pv-frame");

            try{
                frame.contentWindow.focus();
                frame.contentWindow.print();
            }catch(error){
                const printWindow = window.open("{{ pdf_url }}", "_blank");
                if(printWindow){
                    printWindow.addEventListener("load", function(){
                        printWindow.focus();
                        printWindow.print();
                    });
                }
            }
        }
    </script>
    </body>
    </html>
    """,
        row=row,
        pdf_url=pdf_url,
        download_url=download_url,
    )


@finance_bp.route("/payment-voucher/<int:record_id>/delete", methods=["POST"])
def delete_payment_voucher(record_id):
    """Development-stage hard delete for a test expense/PV record."""
    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    row = db_query("""
        select id, category, payment_voucher_no
        from finance_records
        where id = %s
          and record_type = 'expense'
        limit 1
    """, (record_id,), fetchone=True)

    if not row:
        abort(404)

    # Remove optional dependent rows first when those tables/columns exist.
    for sql in (
        "delete from finance_bank_deposit_items where finance_record_id = %s",
        "delete from finance_import_rows where finance_record_id = %s",
    ):
        try:
            db_query(sql, (record_id,))
        except Exception:
            pass

    db_query("delete from finance_records where id = %s", (record_id,))
    flash(f"测试记录 {row.get('payment_voucher_no') or record_id} 已删除。", "success")
    return redirect(url_for("finance.expense", category=row.get("category") or "其它支出"))


@finance_bp.route("/expense/<category>", methods=["GET", "POST"])
def expense(category):
    """普通 key-in 组员使用的独立支出录入页。"""
    ensure_finance_vendor_schema()

    allowed_categories = set(EXPENSE_SUB_CATEGORY_OPTIONS.keys())
    if category not in allowed_categories:
        return "Invalid expense category", 400

    message = ""
    success_message = request.args.get("success", "")
    saved_id = request.args.get("saved_id", "").strip()
    fund_account = "观音堂日常户口"
    suggested_payment_voucher_no = get_next_payment_voucher()
    sub_category_options = EXPENSE_SUB_CATEGORY_OPTIONS.get(category, ["其它"])

    form_data = {
        "payment_voucher_no": suggested_payment_voucher_no,
        "record_date": date.today().isoformat(),
        "sub_category": "",
        "sub_category_custom": "",
        "vendor_id": "",
        "amount": "",
        "particular": "",
        "reference_no": "",
        "prepared_by": "陈柔霓",
        "remarks": "",
    }

    if request.method == "POST":
        action = request.form.get("action", "save_expense").strip()

        form_data = {
            "payment_voucher_no": request.form.get("payment_voucher_no", "").strip().upper(),
            "record_date": request.form.get("record_date", "") or date.today().isoformat(),
            "sub_category": request.form.get("sub_category", "").strip(),
            "sub_category_custom": request.form.get("sub_category_custom", "").strip(),
            "vendor_id": request.form.get("vendor_id", "").strip(),
            "amount": request.form.get("amount", "").strip(),
            "reference_no": request.form.get("reference_no", "").strip(),
            "prepared_by": request.form.get("prepared_by", "陈柔霓").strip(),
            "remarks": request.form.get("remarks", "").strip(),
        }

        if action == "add_vendor":
            company_name = request.form.get("new_company_name", "").strip()
            phone = request.form.get("new_vendor_phone", "").strip()
            contact_person = request.form.get("new_vendor_contact_person", "").strip()

            if not company_name:
                message = "请填写新公司的完整名称。"
            else:
                existing = db_query("""
                    select id
                    from finance_vendors
                    where branch = 'CHE'
                      and lower(coalesce(company_name, vendor_name)) = lower(%s)
                      and coalesce(expense_category, '') = %s
                    limit 1
                """, (company_name, category), fetchone=True)

                if existing:
                    form_data["vendor_id"] = str(existing["id"])
                    message = "这家公司已经存在，系统已替你选中。"
                else:
                    db_query("""
                        insert into finance_vendors
                        (
                            branch,
                            company_name,
                            vendor_name,
                            expense_category,
                            phone,
                            contact_person,
                            is_active,
                            sort_order
                        )
                        values ('CHE', %s, %s, %s, %s, %s, true, 100)
                    """, (
                        company_name,
                        company_name,
                        category,
                        phone or None,
                        contact_person or None,
                    ))

                    new_vendor = db_query("""
                        select id
                        from finance_vendors
                        where branch = 'CHE'
                          and lower(coalesce(company_name, vendor_name)) = lower(%s)
                          and coalesce(expense_category, '') = %s
                        order by id desc
                        limit 1
                    """, (company_name, category), fetchone=True)

                    form_data["vendor_id"] = str(new_vendor["id"])
                    message = "新公司已保存并自动选中，可以继续填写 PV。"

        elif action == "save_expense":
            payment_voucher_no = form_data["payment_voucher_no"]
            record_date = form_data["record_date"]
            selected_sub_category = form_data["sub_category"]
            sub_category = (
                form_data["sub_category_custom"]
                if selected_sub_category == "__custom__"
                else selected_sub_category
            )
            amount = money(form_data["amount"])
            particular = None  # generated automatically in the PDF
            reference_no = form_data["reference_no"]
            prepared_by = form_data["prepared_by"]
            remarks = form_data["remarks"]

            existing_voucher = None
            if payment_voucher_no:
                existing_voucher = db_query("""
                    select id
                    from finance_records
                    where upper(payment_voucher_no) = upper(%s)
                      and coalesce(status, 'active') <> 'cancelled'
                    limit 1
                """, (payment_voucher_no,), fetchone=True)

            vendor_row = None
            if form_data["vendor_id"].isdigit():
                vendor_row = db_query("""
                    select
                        id,
                        coalesce(company_name, vendor_name) as company_name
                    from finance_vendors
                    where id = %s
                      and branch = 'CHE'
                      and is_active = true
                    limit 1
                """, (int(form_data["vendor_id"]),), fetchone=True)

            if not payment_voucher_no:
                message = "请填写 Payment Voucher 编号。"
            elif not re.match(r"^PV\d+$", payment_voucher_no):
                message = "Payment Voucher 格式错误，例如 PV000001。"
            elif existing_voucher:
                message = "这个 Payment Voucher 编号已经存在。"
            elif not sub_category:
                message = "请选择或填写费用明细／单位。"
            elif not vendor_row:
                message = "请选择付款对象／公司名称。"
            elif amount <= 0:
                message = "请输入正确的支出金额。"
            elif prepared_by not in PV_PREPARED_BY_OPTIONS:
                message = "Prepared By 选项无效。"
            else:
                month_lock_error = require_finance_month_open(record_date, fund_account)
                if month_lock_error:
                    message = month_lock_error
                else:
                    vendor_name = vendor_row["company_name"]
                    saved_row = db_query("""
                        insert into finance_records
                        (
                            record_type,
                            payment_voucher_no,
                            record_date,
                            category,
                            sub_category,
                            vendor_id,
                            vendor_name,
                            vendor,
                            name,
                            amount,
                            payment_method,
                            fund_account,
                            particular,
                            reference_no,
                            prepared_by,
                            remarks
                        )
                        values
                        (
                            'expense', %s, %s, %s, %s,
                            %s, %s, %s, %s,
                            %s, '现金', %s, %s, %s, %s, %s
                        )
                        returning id
                    """, (
                        payment_voucher_no,
                        record_date,
                        category,
                        sub_category,
                        vendor_row["id"],
                        vendor_name,
                        vendor_name,
                        vendor_name,
                        amount,
                        fund_account,
                        particular,
                        reference_no or None,
                        prepared_by or None,
                        remarks or None,
                    ), fetchone=True)

                    return redirect(url_for(
                        "finance.payment_voucher",
                        record_id=saved_row["id"],
                    ))

    vendors = get_finance_vendors_for_category(category)

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{{ category }}支出录入</title>
        <link rel="stylesheet" href="{{ url_for('static', filename='css/toolbox.css') }}">
        <style>
            .expense-page{max-width:860px}
            .expense-header{background:linear-gradient(135deg,#dc2626,#b91c1c);color:#fff;padding:28px;border-radius:22px;margin-bottom:20px}
            .expense-header h1{margin:0 0 8px}.expense-header p{margin:0;opacity:.92}
            .expense-form-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:16px}.full-width{grid-column:1/-1}
            .voucher-box{background:#eff6ff;border:1px solid #bfdbfe;border-radius:14px;padding:16px;margin-bottom:18px}
            .cash-box{background:#ecfdf5;border:1px solid #a7f3d0;border-radius:14px;padding:14px;font-weight:700;color:#166534}
            .inline-add{display:none;background:#fff7ed;border:1px solid #fed7aa;border-radius:16px;padding:16px;margin-top:12px}
            .inline-grid{display:grid;grid-template-columns:2fr 1fr 1fr;gap:10px}
            .expense-actions{display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-top:22px}
            @media(max-width:700px){.expense-form-grid,.inline-grid{grid-template-columns:1fr}.full-width{grid-column:auto}.expense-actions{display:grid}.expense-actions .btn-tool{width:100%}}
        </style>
    </head>
    <body>
    <div class="page expense-page">
        <div class="expense-header">
            <h1>💸 {{ category }}</h1>
            <p>普通财政组员可在这里选择公司、即时新增公司并保存 PV。</p>
        </div>

        {% if success_message %}
        <div class="alert alert-success">
            ✅ {{ success_message }}
            {% if saved_id %}
            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:12px;">
                <a class="btn-tool btn-primary" target="_blank"
                   href="{{ url_for('finance.payment_voucher_pdf', record_id=saved_id) }}">
                    👁️ 预览／打印 PV
                </a>
                <a class="btn-tool btn-success"
                   href="{{ url_for('finance.payment_voucher_pdf', record_id=saved_id, download=1) }}">
                    📥 下载 PV PDF
                </a>
            </div>
            {% endif %}
        </div>
        {% endif %}
        {% if message %}<div class="alert alert-danger">⚠️ {{ message }}</div>{% endif %}

        <div class="card">
            <form method="post" id="expenseForm">
                <input type="hidden" name="action" id="formAction" value="save_expense">

                <div class="voucher-box">
                    <label class="form-label">Payment Voucher No.</label>
                    <input class="form-input" name="payment_voucher_no" value="{{ form_data.payment_voucher_no }}" required>
                    <div class="form-help">系统建议：{{ suggested_payment_voucher_no }}</div>
                </div>

                <div class="expense-form-grid">
                    <div class="form-group">
                        <label class="form-label">支出日期</label>
                        <input class="form-input" name="record_date" type="date" value="{{ form_data.record_date }}" required>
                    </div>
                    <div class="form-group">
                        <label class="form-label">付款方式</label>
                        <div class="cash-box">💵 现金（固定）— 从 Petty Cash 扣除，不直接扣银行余额</div>
                    </div>

                    <div class="form-group full-width">
                        <label class="form-label">费用明细／单位</label>
                        <select class="form-input" name="sub_category" id="subCategory" onchange="toggleCustomSubCategory()" required>
                            <option value="">请选择费用明细</option>
                            {% for item in sub_category_options %}<option value="{{ item }}" {% if form_data.sub_category == item %}selected{% endif %}>{{ item }}</option>{% endfor %}
                            <option value="__custom__" {% if form_data.sub_category == '__custom__' %}selected{% endif %}>其它／手动输入</option>
                        </select>
                        <div id="customSubCategoryBox" style="display:none;margin-top:10px"><input class="form-input" name="sub_category_custom" value="{{ form_data.sub_category_custom }}" placeholder="填写费用明细／单位"></div>
                    </div>

                    <div class="form-group full-width">
                        <label class="form-label">付款对象／公司名称 *</label>
                        <select class="form-input" name="vendor_id" id="vendorSelect" required>
                            <option value="">请选择 {{ category }} 的付款对象</option>
                            {% for v in vendors %}<option value="{{ v.id }}" {% if form_data.vendor_id == v.id|string %}selected{% endif %}>{{ v.company_name }}{% if v.phone %} — {{ v.phone }}{% endif %}</option>{% endfor %}
                        </select>
                        <div class="form-help">这里只显示适用于“{{ category }}”的公司。</div>
                        <button class="btn-tool btn-secondary" type="button" onclick="toggleNewVendor()" style="margin-top:10px">＋ 添加新付款对象／公司</button>

                        <div class="inline-add" id="newVendorBox">
                            <h3 style="margin-top:0">新增 {{ category }} 付款对象</h3>
                            <div class="inline-grid">
                                <input class="form-input" name="new_company_name" placeholder="完整公司名称">
                                <input class="form-input" name="new_vendor_phone" placeholder="电话（可选）">
                                <input class="form-input" name="new_vendor_contact_person" placeholder="联络人（可选）">
                            </div>
                            <button class="btn-tool btn-primary" type="submit" onclick="document.getElementById('formAction').value='add_vendor'" style="margin-top:12px">保存公司并自动选择</button>
                        </div>
                    </div>

                    <div class="form-group full-width">
                        <label class="form-label">支出金额 RM</label>
                        <input class="form-input" name="amount" type="number" step="0.01" min="0.01" value="{{ form_data.amount }}" required>
                    </div>
                    </div>

                    <div class="form-group">
                        <label class="form-label">Reference No.（可选）</label>
                        <input class="form-input" name="reference_no" value="{{ form_data.reference_no }}" placeholder="Invoice／Receipt／Bill 编号">
                    </div>

                    <div class="form-group">
                        <label class="form-label">Prepared By</label>
                        <select class="form-input" name="prepared_by">
                            <option value="陈柔霓" {% if form_data.prepared_by == '陈柔霓' %}selected{% endif %}>陈柔霓（默认）</option>
                            <option value="黄薈菏" {% if form_data.prepared_by == '黄薈菏' %}selected{% endif %}>黄薈菏</option>
                            <option value="" {% if not form_data.prepared_by %}selected{% endif %}>（空白）</option>
                        </select>
                        <div class="form-help">选择空白时，PV 会保留签名线供现场填写。</div>
                    </div>

                    <div class="form-group full-width">
                        <label class="form-label">备注（可选）</label>
                        <textarea class="form-input" name="remarks" rows="2" placeholder="如有填写，会显示在 Payment Voucher；没有可留空">{{ form_data.remarks }}</textarea>
                    </div>
                </div>

                <div class="expense-actions">
                    <a class="btn-tool btn-secondary" href="{{ url_for('finance.finance_expense_menu') }}">← 返回支出项目</a>
                    <button class="btn-tool btn-danger" type="submit" onclick="document.getElementById('formAction').value='save_expense';return confirm('确定保存这笔现金支出？')">💾 保存支出</button>
                </div>
            </form>
        </div>
    </div>

    <script>
        function toggleCustomSubCategory(){
            const select=document.getElementById('subCategory');
            document.getElementById('customSubCategoryBox').style.display=select.value==='__custom__'?'block':'none';
        }
        function toggleNewVendor(){
            const box=document.getElementById('newVendorBox');
            box.style.display=box.style.display==='block'?'none':'block';
        }
        toggleCustomSubCategory();
        {% if message and request.form.get('action') == 'add_vendor' %}document.getElementById('newVendorBox').style.display='block';{% endif %}
    </script>
    </body>
    </html>
    """,
        category=category,
        message=message,
        success_message=success_message,
        saved_id=saved_id,
        form_data=form_data,
        vendors=vendors,
        sub_category_options=sub_category_options,
        suggested_payment_voucher_no=suggested_payment_voucher_no,
    )
