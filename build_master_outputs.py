# build_master_outputs.py

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

from tomlkit import item

BASE = Path(__file__).parent

MASTER_FILE = BASE / "master_volunteers.xlsx"
VOL_FILE = BASE / "volunteers.xlsx"
MEMBER_FILE = BASE / "members.xlsx"
PAYMENT_INPUT_FILE = BASE / "payment_input.xlsx"
PAYMENT_RECORDS_FILE = BASE / "payment_records.xlsx"

YEAR = 2026


def clean(v):
    return "" if v is None else str(v).strip()


def phone_last4(phone):
    digits = "".join(ch for ch in clean(phone) if ch.isdigit())
    return digits[-4:] if len(digits) >= 4 else ""


def beautify(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Microsoft YaHei", size=12, bold=True)
    body_font = Font(name="Microsoft YaHei", size=12)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(clean(cell.value)) for cell in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 28)

    ws.row_dimensions[1].height = 28

def beautify_master_file():
    if not MASTER_FILE.exists():
        return

    wb = load_workbook(MASTER_FILE)
    ws = wb.active

    # 如果原本有 Excel Table，先移除，避免只美化旧范围
    ws._tables.clear()

    beautify(ws)

    # 重新设定筛选范围
    ws.auto_filter.ref = ws.dimensions

    # 固定列宽
    widths = {
        "分会": 10,
        "编号": 14,
        "姓名": 18,
        "状态": 12,
        "电话号码": 18,
        "是否义工": 12,
        "是否月费": 12,
        "月费编号": 16,
        "月费姓名": 18,
        "月费英文名": 22,
        "备注": 38,
    }

    headers = [clean(c.value) for c in ws[1]]

    for idx, h in enumerate(headers, start=1):
        if h in widths:
            ws.column_dimensions[get_column_letter(idx)].width = widths[h]

    # 行高统一
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24

    wb.save(MASTER_FILE)
    print("✅ master_volunteers.xlsx 已美化")


def save_excel(path, sheet_name, headers, rows=None):
    rows = rows or []

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)

    for item in rows:
        ws.append([item.get(h, "") for h in headers])

    beautify(ws)
    # volunteers / members 固定列宽
    widths = {
        "编号": 14,
        "姓名": 18,
        "状态": 12,
        "电话号码": 18,
        "PIN": 10,
        "分会": 10,
        "备注": 38,
    }

    for idx, h in enumerate(headers, start=1):
        if h in widths:
            ws.column_dimensions[get_column_letter(idx)].width = widths[h]

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24
    wb.save(path)
    print(f"✅ 已生成：{path.name}")


def main():
    if not MASTER_FILE.exists():
        raise FileNotFoundError("找不到 master_volunteers.xlsx")

    wb = load_workbook(MASTER_FILE)
    ws = wb.active

    headers = [clean(c.value) for c in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        item = {k: clean(v) for k, v in item.items()}

        if not item.get("PIN"):
            phone = item.get("义工电话") or item.get("电话号码") or item.get("月费电话")
            item["PIN"] = phone_last4(phone) or "0000"

        data.append(item)

    volunteers = []
    for x in data:
        if x.get("是否义工") == "是" and (x.get("义工状态") or x.get("状态")) == "在册":
            volunteers.append({
                "编号": x.get("义工编号") or x.get("编号"),
                "姓名": x.get("义工姓名") or x.get("姓名") or x.get("月费姓名"),
                "状态": x.get("义工状态") or x.get("状态"),
                "电话号码": x.get("义工电话") or x.get("电话号码") or x.get("月费电话"),
                "PIN": x.get("PIN"),
                "分会": x.get("分会"),
                "备注": x.get("备注"),
            })

    month_cols = [f"{YEAR}-{m:02d}" for m in range(1, 13)]

    members = []
    for x in data:
        if x.get("是否月费") == "是":
            row = {
                "月费编号": x.get("月费编号"),
                "姓名": x.get("月费姓名") or x.get("义工姓名"),
                "英文名": x.get("月费英文名"),
                "电话号码": x.get("月费电话") or x.get("义工电话"),
                "PIN": x.get("PIN"),
                "分会": x.get("分会"),
                "状态": x.get("义工状态"),
                "身份证号码": x.get("身份证号码"),
                "备注": x.get("备注"),
            }

            for m in month_cols:
                row[m] = ""

            members.append(row)

    save_excel(
        VOL_FILE,
        "volunteers",
        ["编号", "姓名", "状态", "电话号码", "PIN", "分会", "备注"],
        volunteers,
    )

    save_excel(
        MEMBER_FILE,
        "members",
        ["月费编号", "姓名", "英文名", "电话号码", "PIN", "分会", "状态", "身份证号码", "备注"] + month_cols,
        members,
    )

    save_excel(
        PAYMENT_INPUT_FILE,
        "payment_input",
        ["日期", "月费编号", "姓名", "月份", "备注"],
        [],
    )

    save_excel(
        PAYMENT_RECORDS_FILE,
        "payment_records",
        ["日期", "月费编号", "姓名", "月份", "录入时间", "备注"],
        [],
    )

    print("\n🎉 完成！以后只维护 master_volunteers.xlsx")
    beautify_master_file()

if __name__ == "__main__":
    main()