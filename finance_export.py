# finance_export.py

from __future__ import annotations

import io
import re
import zipfile

from copy import copy
from db import db_query
from pathlib import Path
from typing import Any, Iterable
from finance_web_old import finance_bp
from openpyxl import load_workbook, Workbook
from datetime import date, datetime
from openpyxl.styles import Alignment, Font, Border, PatternFill, Side
from openpyxl.formula.translate import Translator
from flask import request, send_file, abort, render_template_string, url_for

# -----------------------------------------------------------------------------
# 基本设置
# -----------------------------------------------------------------------------

MONTHLY_FEE_TEMPLATE = (
    Path(__file__).resolve().parent
    / "finance_templates"
    / "monthly_fee_master.xlsx"
)

TEMPLATE_SHEET = "Blank_Master"
MASTER_NAME_SHEET = "MASTER Name List"

CASH_START_ROW = 9
CASH_END_ROW = 23
BANK_TITLE_ROW = 25
BANK_START_ROW = 26
BANK_END_ROW = 43
TOTAL_ROW = 47


# -----------------------------------------------------------------------------
# 分会与全部财政报表设置
# -----------------------------------------------------------------------------

BRANCH_CONFIG = {
    "CHE": {
        "label": "蕉赖 CHE",
        "monthly_account": "HLB 05500129613",
        "donation_account": "MBB 5125 5832 5341",
        "village_account": "MBB 5125 5832 5341",
        "meal_account": "HLB 10900080388",
        "petty_cash_opening": 0.0,
    },
    "STW": {
        "label": "实兆远 STW",
        "monthly_account": "AM 8881010982085",
        # STW 其他户口若与月费不同，只需在这里修改。
        "donation_account": "AM 8881010982085",
        "village_account": "AM 8881010982085",
        "meal_account": "AM 8881010982085",
        "petty_cash_opening": 0.0,
    },
}

PETTY_CASH_TEMPLATE = (
    Path(__file__).resolve().parent
    / "finance_templates"
    / "petty_cash_master.xlsx"
)

PETTY_CASH_TEMPLATE_SHEET = "2026 Petty Cash Manual Book"
PETTY_CASH_DATA_START = 4
PETTY_CASH_TEMPLATE_DATA_END = 47


def _normalise_branch(value: str | None) -> str:
    branch = (value or "CHE").strip().upper()
    if branch not in BRANCH_CONFIG:
        abort(400, description="分会必须是 CHE 或 STW")
    return branch


def _branch_sql_condition(
    branch: str,
    alias: str = "fr",
) -> tuple[str, tuple]:

    branch = _normalise_branch(branch)

    stw_condition = f"""
        (
            coalesce({alias}.member_id, '') ilike 'STW-%%'
            or coalesce({alias}.receipt_no, '') ilike 'STW%%'
        )
    """

    if branch == "STW":
        return stw_condition, ()

    return f"""
        not (
            {stw_condition}
        )
    """, ()


def _workbook_bytes(wb) -> io.BytesIO:
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------------------------------------------------------
# 通用辅助函数
# -----------------------------------------------------------------------------

def _row_value(row: Any, key: str, default: Any = None) -> Any:
    """同时支持 RealDictRow、dict 和普通对象。"""
    if row is None:
        return default

    if isinstance(row, dict):
        return row.get(key, default)

    try:
        return row[key]
    except (KeyError, TypeError, IndexError):
        return getattr(row, key, default)


def _first_value(row: Any, *keys: str, default: Any = None) -> Any:
    for key in keys:
        value = _row_value(row, key)
        if value not in (None, ""):
            return value
    return default


def _as_date(value: Any) -> date | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m",
        "%Y/%m",
    ):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date()
        except ValueError:
            continue

    return None


def _month_date(value: Any) -> date | None:
    """月份字段统一写成该月 1 日，让原模板 DATEDIF 正常计算。"""
    parsed = _as_date(value)
    if parsed is None:
        return None
    return date(parsed.year, parsed.month, 1)


def _normalize_ym(raw_ym: str | None) -> str:
    raw_ym = (raw_ym or "").strip()

    if not re.fullmatch(r"\d{4}-\d{2}", raw_ym):
        abort(400, description="月份格式必须是 YYYY-MM，例如 2026-07")

    year, month = map(int, raw_ym.split("-"))
    if month < 1 or month > 12:
        abort(400, description="月份不正确")

    return f"{year:04d}-{month:02d}"


def _excel_member_no(member_id: Any) -> int | str | None:
    """
    原财政 Master Name List 的 VLOOKUP key 位于 B 栏，以数字为主。

    CHE-73  -> 73
    STW-160 -> 160
    0160    -> 160

    若日后 Master Name List 改为完整编号，只需修改这个函数。
    """
    if member_id in (None, ""):
        return None

    text = str(member_id).strip().upper()
    match = re.search(r"(\d+)$", text)

    if not match:
        return text

    return int(match.group(1))


def _normalise_payment_method(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def _is_cash(row: Any) -> bool:
    method = _normalise_payment_method(
        _first_value(row, "payment_method", "method", default="")
    )

    cash_keywords = (
        "cash",
        "现款",
        "现金",
        "tunai",
    )

    return any(keyword in method for keyword in cash_keywords)


def _is_cheque(row: Any) -> bool:
    method = _normalise_payment_method(
        _first_value(row, "payment_method", "method", default="")
    )
    return "cheque" in method or "支票" in method or "cek" in method


def _reference_text(row: Any) -> str | date | None:
    """
    REF. 栏优先顺序：bank_ref -> reference -> remarks。
    保持财政旧表习惯，不自动拼接过长说明。
    """
    return _first_value(
        row,
        "bank_ref",
        "reference",
        "ref",
        "remarks",
        default=None,
    )


def _monthly_phone_reference_text(row: Any) -> str | date | None:
    """
    月费旧模板没有独立电话号码栏。
    为了不破坏原模板列位，把电话与原 REF. 资料合并写入 G 栏。

    例：
        0123456789 | ABC123
    """
    phone = str(_first_value(row, "phone", default="") or "").strip()
    reference = _reference_text(row)

    parts = []

    if phone:
        parts.append(phone)

    if reference not in (None, ""):
        parts.append(str(reference))

    return " | ".join(parts) or None


def _month_unit_amount(row: Any) -> float:
    """
    原模板 I 栏为每月 RM50。
    若资料真的不是 RM50/月，也尽量按 amount / month_count 还原。
    """
    amount = _first_value(row, "amount", default=0) or 0
    month_count = _first_value(row, "month_count", default=0) or 0

    try:
        amount = float(amount)
        month_count = int(month_count)
    except (TypeError, ValueError):
        return 50.0

    if month_count > 0 and amount > 0:
        unit = amount / month_count
        if abs(unit - 50.0) > 0.001:
            return unit

    return 50.0


def _copy_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def _copy_template_row(ws, source_row: int, target_row: int) -> None:
    """复制整行样式、公式及行高。"""
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden

    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)

        _copy_cell_style(source_cell, target_cell)

        if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
            try:
                target_cell.value = Translator(
                    source_cell.value,
                    origin=source_cell.coordinate,
                ).translate_formula(target_cell.coordinate)
            except Exception:
                target_cell.value = source_cell.value
        else:
            target_cell.value = source_cell.value


# -----------------------------------------------------------------------------
# 模板扩行
# -----------------------------------------------------------------------------

def _ensure_capacity(ws, cash_count: int, bank_count: int) -> tuple[int, int, int]:
    """
    保留原模板的两个区块。

    原容量：
      现金 15 行（9-23）
      BANK IN 18 行（26-43）

    超出时只在相应区块末端插入行，并从该区块最后一行复制格式与公式。
    返回：cash_end_row、bank_start_row、bank_end_row。
    """
    cash_capacity = CASH_END_ROW - CASH_START_ROW + 1
    bank_capacity = BANK_END_ROW - BANK_START_ROW + 1

    extra_cash = max(0, cash_count - cash_capacity)

    cash_end_row = CASH_END_ROW
    bank_title_row = BANK_TITLE_ROW
    bank_start_row = BANK_START_ROW
    bank_end_row = BANK_END_ROW

    if extra_cash:
        insert_at = CASH_END_ROW + 1
        ws.insert_rows(insert_at, amount=extra_cash)

        for offset in range(extra_cash):
            target_row = insert_at + offset
            _copy_template_row(ws, CASH_END_ROW, target_row)

        cash_end_row += extra_cash
        bank_title_row += extra_cash
        bank_start_row += extra_cash
        bank_end_row += extra_cash

    extra_bank = max(0, bank_count - bank_capacity)

    if extra_bank:
        insert_at = bank_end_row + 1
        ws.insert_rows(insert_at, amount=extra_bank)

        for offset in range(extra_bank):
            target_row = insert_at + offset
            _copy_template_row(ws, bank_end_row, target_row)

        bank_end_row += extra_bank

    total_row = TOTAL_ROW + extra_cash + extra_bank

    # 插入行后明确恢复合计公式，避免 Excel 的公式范围没有跟着扩展。
    ws.cell(total_row, 10).value = f"=SUM(J8:J{bank_end_row + 3})"
    ws.cell(total_row, 11).value = f"=SUM(K8:K{bank_end_row + 3})"
    ws.cell(total_row, 12).value = f"=SUM(L8:L{bank_end_row + 3})"
    ws.cell(total_row, 14).value = f"=SUM(N8:N{bank_end_row + 3})"

    return cash_end_row, bank_start_row, bank_end_row


# -----------------------------------------------------------------------------
# 写入资料
# -----------------------------------------------------------------------------

def _clear_data_row(ws, row_no: int) -> None:
    """清除旧输入资料，保留全部格式；D 栏姓名稍后由数据库写入。"""
    for col in (1, 2, 3, 4, 5, 6, 7, 13, 14, 15):
        ws.cell(row_no, col).value = None

    ws.cell(row_no, 9).value = 50


def _prepare_formula_row(ws, row_no: int, payment_column: int) -> None:
    # 姓名由数据库直接写入 D 栏，不再依赖 VLOOKUP。
    ws.cell(row_no, 8).value = (
        f'=IF(OR(E{row_no}="",F{row_no}=""),"",DATEDIF(E{row_no},F{row_no},"M")+1)'
    )

    # 清除三个付款金额栏，再根据付款方式只放一个公式。
    for col in (10, 11, 12):
        ws.cell(row_no, col).value = None

    ws.cell(row_no, payment_column).value = (
        f'=IF($C{row_no}="","-",$H{row_no}*$I{row_no})'
    )


def _write_monthly_fee_row(ws, row_no: int, record: Any, is_cash_section: bool) -> None:
    record_date = _as_date(
        _first_value(record, "record_date", "payment_date", "receipt_date")
    )
    receipt_no = _first_value(record, "receipt_no", default=None)
    member_id = _first_value(record, "member_id", default=None)
    name = _first_value(record, "name", default=None)
    month_from = _month_date(_first_value(record, "month_from", "start_month"))
    month_to = _month_date(_first_value(record, "month_to", "end_month"))

    ws.cell(row_no, 1).value = record_date
    ws.cell(row_no, 2).value = receipt_no if is_cash_section else None
    ws.cell(row_no, 3).value = _excel_member_no(member_id)
    ws.cell(row_no, 4).value = name
    ws.cell(row_no, 5).value = month_from
    ws.cell(row_no, 6).value = month_to
    ws.cell(row_no, 7).value = _monthly_phone_reference_text(record)
    ws.cell(row_no, 9).value = _month_unit_amount(record)

    payment_column = 10 if is_cash_section else (11 if _is_cheque(record) else 12)
    _prepare_formula_row(ws, row_no, payment_column)

    # 金额必须直接采用 finance_records.amount。
    # 不再依赖 H × I 的 Excel 公式，避免导出后公式尚未重算而显示空白。
    raw_amount = _first_value(record, "amount", default=0) or 0
    try:
        actual_amount = float(raw_amount)
    except (TypeError, ValueError):
        actual_amount = 0.0

    for col in (10, 11, 12):
        ws.cell(row_no, col).value = None

    ws.cell(row_no, payment_column).value = actual_amount

    ws.cell(row_no, 1).number_format = "dd/mm/yyyy"
    ws.cell(row_no, 5).number_format = "mmm-yy"
    ws.cell(row_no, 6).number_format = "mmm-yy"
    ws.cell(row_no, 9).number_format = '#,##0.00'
    ws.cell(row_no, payment_column).number_format = '#,##0.00'


def _set_report_title(
    ws,
    year: int,
    month: int,
    branch: str = "CHE",
) -> None:
    branch = _normalise_branch(branch)
    account = BRANCH_CONFIG[branch]["monthly_account"]
    ws["N4"] = f"   {year}  年 {month}  月份 - {account}"


def _sheet_name(year: int, month: int) -> str:
    month_name = (
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    )[month - 1]
    return f"{month_name}-{str(year)[-2:]}"


def build_monthly_fee_workbook(records: Iterable[Any], ym: str, branch: str = "CHE"):
    """建立与原财政格式一致的月费 Excel workbook。"""
    if not MONTHLY_FEE_TEMPLATE.exists():
        raise FileNotFoundError(
            f"找不到月费模板：{MONTHLY_FEE_TEMPLATE}"
        )

    branch = _normalise_branch(branch)
    year, month = map(int, ym.split("-"))
    records = list(records)

    # 现金和非现金分区，区内按日期、收条编号排序。
    cash_records = [row for row in records if _is_cash(row)]
    bank_records = [row for row in records if not _is_cash(row)]

    def sort_key(row: Any):
        d = _as_date(_first_value(row, "record_date", "payment_date"))
        receipt = str(_first_value(row, "receipt_no", default="") or "")
        member = str(_first_value(row, "member_id", default="") or "")
        return (d or date.min, receipt, member)

    cash_records.sort(key=sort_key)
    bank_records.sort(key=sort_key)

    wb = load_workbook(MONTHLY_FEE_TEMPLATE, data_only=False)

    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"模板缺少工作表：{TEMPLATE_SHEET}")

    if MASTER_NAME_SHEET not in wb.sheetnames:
        raise ValueError(f"模板缺少工作表：{MASTER_NAME_SHEET}")

    target_sheet_name = _sheet_name(year, month)

    # 不能使用 copy_worksheet()：openpyxl 不会复制图片、Logo 和 drawing。
    # 若模板已有该月份，直接使用该月份工作表；否则直接使用 Blank_Master。
    # 这是独立导出副本，因此可以安全地清空和改名，同时完整保留图片。
    if target_sheet_name in wb.sheetnames:
        ws = wb[target_sheet_name]
    else:
        ws = wb[TEMPLATE_SHEET]
        ws.title = target_sheet_name

    ws.sheet_view.showGridLines = False
    _set_report_title(ws, year, month, branch)

    cash_end_row, bank_start_row, bank_end_row = _ensure_capacity(
        ws,
        cash_count=len(cash_records),
        bank_count=len(bank_records),
    )

    # 先清除整个两区的旧输入资料。
    for row_no in range(CASH_START_ROW, cash_end_row + 1):
        _clear_data_row(ws, row_no)
        _prepare_formula_row(ws, row_no, 10)

    for row_no in range(bank_start_row, bank_end_row + 1):
        _clear_data_row(ws, row_no)
        _prepare_formula_row(ws, row_no, 12)

    # 写入现金记录。
    for row_no, record in enumerate(cash_records, start=CASH_START_ROW):
        _write_monthly_fee_row(ws, row_no, record, is_cash_section=True)

    # 写入银行／支票记录。
    for row_no, record in enumerate(bank_records, start=bank_start_row):
        _write_monthly_fee_row(ws, row_no, record, is_cash_section=False)

    # BANK IN 标题行是插行后 bank_start_row - 1。
    ws.cell(bank_start_row - 1, 2).value = "BANK IN"

    # 只保留本次生成月份工作表 + 原始资料表。
    # 若你希望导出文件继续包含过去所有月份，可删除这一段。
    keep_sheets = {
        ws.title,
        MASTER_NAME_SHEET,
        "Summary",
    }

    # 若 Blank_Master 不是本次生成页，可保留并隐藏；这样以后仍可检查模板。
    if TEMPLATE_SHEET in wb.sheetnames and wb[TEMPLATE_SHEET] is not ws:
        keep_sheets.add(TEMPLATE_SHEET)

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in keep_sheets:
            del wb[sheet_name]

    # 保证日期与月份打开时不会因为栏宽不足显示 ####。
    for column, minimum_width in {
        "A": 12.0,
        "E": 11.0,
        "F": 11.0,
        "J": 12.0,
        "K": 12.0,
        "L": 15.0,
    }.items():
        current_width = ws.column_dimensions[column].width or 0
        if current_width < minimum_width:
            ws.column_dimensions[column].width = minimum_width

    # 生成月份放到第一张。
    wb._sheets.remove(ws)
    wb._sheets.insert(0, ws)
    wb.active = 0

    if TEMPLATE_SHEET in wb.sheetnames and wb[TEMPLATE_SHEET] is not ws:
        wb[TEMPLATE_SHEET].sheet_state = "hidden"

    # 要求 Excel 打开时自动重算公式。
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    return wb


# -----------------------------------------------------------------------------
# Flask Route
# -----------------------------------------------------------------------------

# 注意：下面 route 假设你的文件里已经存在：
#   finance_bp
#   db_query
# 若 db_query 参数形式不同，只需调整查询调用部分。


@finance_bp.route("/export/monthly-fee")
def export_monthly_fee_excel():
    ym = _normalize_ym(request.args.get("ym"))
    branch = _normalise_branch(request.args.get("branch"))

    branch_condition, branch_params = _branch_sql_condition(branch, "fr")

    records = db_query(
        f"""
        select
            fr.id,
            fr.record_date,
            fr.receipt_no,
            fr.member_id,
            fr.name,
            fr.phone,
            fr.amount,
            fr.payment_method,
            fr.bank_name,
            fr.bank_ref,
            fr.month_from,
            fr.month_to,
            fr.remarks,
            mp.month_count
        from finance_records fr
        left join member_payments mp
          on mp.member_id = fr.member_id
         and coalesce(mp.receipt_no, '') = coalesce(fr.receipt_no, '')
         and mp.amount = fr.amount
         and mp.payment_date = fr.record_date
        where fr.record_type = 'income'
          and fr.category = '月费'
          and to_char(fr.record_date, 'YYYY-MM') = %s
          and {branch_condition}
        order by
            fr.record_date,
            fr.receipt_no nulls last,
            fr.id
        """,
        (ym, *branch_params),
        fetchall=True,
    ) or []

    wb = build_monthly_fee_workbook(records, ym, branch)
    filename = f"{branch}_月费收纳表_{ym}.xlsx"

    return send_file(
        _workbook_bytes(wb),
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# =============================================================================
# 布施／善款 Excel 导出
# =============================================================================

DONATION_TEMPLATE = (
    Path(__file__).resolve().parent
    / "finance_templates"
    / "donation_master.xlsx"
)

# 统一支持财布施、观音村及初一十五／膳食结缘。
DONATION_REPORTS = {
    "财布施": {
        "sheet": "05-善款收纳表",
        "db_categories": ("财布施",),
        "title_cell": "K5",
        "account_key": "donation_account",
        "title": "   {year}    年 {month}  月份  - {account}",
        "data_start": 9,
        "template_total_row": 77,
        "marker_cols": (6,),
        "cash_col": 7,
        "cheque_col": 8,
        "bank_col": 9,
        "cdm_seq_col": 10,
        "cdm_amount_col": 11,
        "cdm_date_col": 12,
        "remark_col": 13,
        "bank_holder_col": 4,
        "phone_col": 5,
    },
    "观音村": {
        "sheet": "04-观音村善款收纳表",
        "db_categories": ("观音村",),
        "title_cell": "K5",
        "account_key": "village_account",
        "title": "   {year}   年 {month} 月份  - {account}",
        "data_start": 9,
        "template_total_row": 153,
        # F、G 是旧表的印刷／法会、放生选项。系统目前只有观音村大分类，
        # 因此不自动勾选，避免把资料归入错误子分类。
        "marker_cols": (),
        "cash_col": 8,
        "cheque_col": 9,
        "bank_col": 10,
        "cdm_seq_col": 11,
        "cdm_amount_col": 12,
        "cdm_date_col": 13,
        "remark_col": 14,
        "bank_holder_col": 4,
        "phone_col": 5,
    },
    "初一十五": {
        "sheet": "03-初一十五膳食结缘",
        "db_categories": ("初一十五", "膳食结缘", "初一十五膳食结缘"),
        "title_cell": "M5",
        "account_key": "meal_account",
        "title": "   {year}  年  {month}  月份  -  {account}",
        "data_start": 9,
        "template_total_row": 25,
        "marker_cols": (8,),
        "cash_col": 9,
        "bank_col": 10,
        "cheque_col": 11,
        "cdm_seq_col": 12,
        "cdm_amount_col": 13,
        "cdm_date_col": 14,
        "remark_col": 15,
        # 该模板有两组 Bank Holder Name / Mobile No。
        "bank_holder_col": 6,
        "phone_col": 7,
    },
}


def _normalise_donation_category(value: str | None) -> str:
    text = (value or "财布施").strip()

    aliases = {
        "donation": "财布施",
        "财布施": "财布施",
        "善款": "财布施",
        "village": "观音村",
        "观音村": "观音村",
        "meal": "初一十五",
        "初一十五": "初一十五",
        "膳食结缘": "初一十五",
        "初一十五膳食结缘": "初一十五",
    }

    result = aliases.get(text.lower(), aliases.get(text))
    if not result:
        abort(400, description="不支持的布施报表类别")
    return result


def _amount_value(record: Any) -> float:
    raw = _first_value(record, "amount", default=0) or 0
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


def _donation_payment_column(record: Any, config: dict[str, Any]) -> int:
    if _is_cash(record):
        return int(config["cash_col"])
    if _is_cheque(record):
        return int(config["cheque_col"])
    return int(config["bank_col"])


def _clear_donation_data_row(
    ws,
    row_no: int,
    config: dict[str, Any],
) -> None:
    """
    清空布施报表资料行，但保留模板格式、边框和底色。
    """

    columns_to_clear = {
        1,  # 日期
        2,  # 收据编号
        3,  # 姓名
        int(config.get("bank_holder_col") or 4),
        int(config.get("phone_col") or 5),
        int(config["cash_col"]),
        int(config["cheque_col"]),
        int(config["bank_col"]),
        int(config["cdm_amount_col"]),
        int(config["cdm_date_col"]),
        int(config["remark_col"]),
    }

    for marker_col in config.get("marker_cols", ()):
        columns_to_clear.add(int(marker_col))

    for col_no in columns_to_clear:
        ws.cell(row_no, col_no).value = None


def _ensure_donation_capacity(
    ws,
    record_count: int,
    data_start: int,
    template_total_row: int,
) -> tuple[int, int]:
    """
    按实际记录数量调整布施报表长度。

    规则：
    - 有多少笔记录，就显示多少行资料；
    - 资料后保留一行原模板的空白间隔行；
    - 下一行放原模板 Total 样式；
    - 超过模板原容量时自动增加资料行；
    - 少于模板原容量时删除多余空白行。

    返回：
        data_end
        total_row
    """

    actual_count = max(record_count, 1)
    data_end = data_start + actual_count - 1
    spacer_row = data_end + 1
    total_row = data_end + 2

    template_spacer_row = template_total_row - 1
    template_data_end = template_total_row - 2
    source_data_row = max(data_start, template_data_end)

    def save_row_style(row_no: int):
        saved = []

        for col_no in range(1, ws.max_column + 1):
            source_cell = ws.cell(row_no, col_no)

            saved.append({
                "style": copy(source_cell._style),
                "number_format": source_cell.number_format,
                "font": copy(source_cell.font),
                "fill": copy(source_cell.fill),
                "border": copy(source_cell.border),
                "alignment": copy(source_cell.alignment),
                "protection": copy(source_cell.protection),
            })

        return saved, ws.row_dimensions[row_no].height

    def restore_row_style(row_no: int, saved_styles, row_height):
        ws.row_dimensions[row_no].height = row_height

        for col_no, saved in enumerate(saved_styles, start=1):
            target_cell = ws.cell(row_no, col_no)

            target_cell._style = copy(saved["style"])
            target_cell.number_format = saved["number_format"]
            target_cell.font = copy(saved["font"])
            target_cell.fill = copy(saved["fill"])
            target_cell.border = copy(saved["border"])
            target_cell.alignment = copy(saved["alignment"])
            target_cell.protection = copy(saved["protection"])
            target_cell.value = None

    # 保存原模板的空白间隔行与合计行格式。
    spacer_styles, spacer_height = save_row_style(template_spacer_row)
    total_styles, total_height = save_row_style(template_total_row)

    # 记录超过模板原容量时，自动增加资料行。
    if data_end > template_data_end:
        extra_rows = data_end - template_data_end

        ws.insert_rows(
            template_spacer_row,
            amount=extra_rows,
        )

        for row_no in range(
            template_spacer_row,
            template_spacer_row + extra_rows,
        ):
            ws.row_dimensions[row_no].height = (
                ws.row_dimensions[source_data_row].height
            )

            for col_no in range(1, ws.max_column + 1):
                source_cell = ws.cell(source_data_row, col_no)
                target_cell = ws.cell(row_no, col_no)

                _copy_cell_style(source_cell, target_cell)
                target_cell.value = None

    # 第 14 行这类间隔行必须使用模板真正的空白行格式，
    # 避免出现多余的 L 型角线或直线。
    restore_row_style(
        spacer_row,
        spacer_styles,
        spacer_height,
    )

    # 最后一笔资料行采用模板最后一笔资料行的边框，
    # 保证资料区底部有完整横线。
    for col_no in range(1, ws.max_column + 1):
        source_cell = ws.cell(template_data_end, col_no)
        target_cell = ws.cell(data_end, col_no)
        target_cell.border = copy(source_cell.border)

    # 将模板原本的合计行格式放到实际 Total 行。
    restore_row_style(
        total_row,
        total_styles,
        total_height,
    )

    return data_end, total_row

def _set_donation_totals(
    ws,
    config: dict[str, Any],
    data_start: int,
    data_end: int,
    total_row: int,
) -> None:
    """
    设置布施报表底部 Total Amount。

    功能：
    1. 写入 Total Amount 标题
    2. 分别计算 Cash / Cheque / Bank In / CDM
    3. 保留模板原本格式
    4. 合计范围自动跟随实际资料行
    """

    cash_col = int(config["cash_col"])
    cheque_col = int(config["cheque_col"])
    bank_col = int(config["bank_col"])
    cdm_amount_col = int(config["cdm_amount_col"])

    # 分付款方式合计。
    ws.cell(total_row, cash_col).value = (
        f"=SUM({ws.cell(data_start,cash_col).coordinate}:"
        f"{ws.cell(data_end,cash_col).coordinate})"
    )

    ws.cell(total_row, cheque_col).value = (
        f"=SUM("
        f"{ws.cell(data_start, cheque_col).coordinate}:"
        f"{ws.cell(data_end, cheque_col).coordinate}"
        f")"
    )

    ws.cell(total_row, bank_col).value = (
        f"=SUM("
        f"{ws.cell(data_start, bank_col).coordinate}:"
        f"{ws.cell(data_end, bank_col).coordinate}"
        f")"
    )

    ws.cell(total_row, cdm_amount_col).value = (
        f"=SUM("
        f"{ws.cell(data_start, cdm_amount_col).coordinate}:"
        f"{ws.cell(data_end, cdm_amount_col).coordinate}"
        f")"
    )

    # 金额格式。
    for col in (
        cash_col,
        cheque_col,
        bank_col,
        cdm_amount_col,
    ):
        ws.cell(total_row, col).number_format = '#,##0.00'



def _write_donation_row(
    ws,
    row_no: int,
    record: Any,
    config: dict[str, Any],
) -> None:
    record_date = _as_date(
        _first_value(record, "record_date", "payment_date", "receipt_date")
    )
    receipt_no = _first_value(record, "receipt_no", default=None)
    name = _first_value(record, "name", default=None)
    phone = _first_value(record, "phone", default=None)
    bank_holder_name = _first_value(
        record,
        "bank_holder_name",
        "account_holder_name",
        default=None,
    )
    remarks = _first_value(record, "remarks", default=None)
    bank_ref = _first_value(record, "bank_ref", "reference", default=None)

    ws.cell(row_no, 1).value = record_date
    ws.cell(row_no, 2).value = receipt_no
    ws.cell(row_no, 3).value = name

    # 财布施／观音村模板使用 D、E；初一十五模板使用 F、G。
    bank_holder_col = int(config.get("bank_holder_col") or 4)
    phone_col = int(config.get("phone_col") or 5)
    ws.cell(row_no, bank_holder_col).value = bank_holder_name
    ws.cell(row_no, phone_col).value = phone

    for marker_col in config.get("marker_cols", ()):
        cell = ws.cell(row_no, int(marker_col))
        cell.value = "✓"
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    payment_col = _donation_payment_column(record, config)
    for col in (
        int(config["cash_col"]),
        int(config["cheque_col"]),
        int(config["bank_col"]),
    ):
        ws.cell(row_no, col).value = None

    ws.cell(row_no, payment_col).value = _amount_value(record)

    # 银行 Reference 没有独立栏时放入备注，避免资料丢失。
    remark_parts = []
    if bank_ref and not _is_cash(record):
        remark_parts.append(f"Bank Ref: {bank_ref}")
    if remarks:
        remark_parts.append(str(remarks))
    ws.cell(row_no, int(config["remark_col"])).value = " | ".join(remark_parts) or None

    ws.cell(row_no, 1).number_format = "dd/mm/yyyy"
    ws.cell(row_no, payment_col).number_format = '#,##0.00'


def build_donation_workbook(
    records: Iterable[Any],
    ym: str,
    category: str = "财布施",
    branch: str = "CHE",
):
    """按照财政旧模板建立财布施／观音村／初一十五 Excel。"""
    category = _normalise_donation_category(category)
    branch = _normalise_branch(branch)
    config = DONATION_REPORTS[category]

    if not DONATION_TEMPLATE.exists():
        raise FileNotFoundError(
            f"找不到布施模板：{DONATION_TEMPLATE}。"
            "请把上传的 excel format(1).xlsx 改名为 donation_master.xlsx，"
            "放进 finance_templates 文件夹。"
        )

    year, month = map(int, ym.split("-"))
    records = list(records)

    def sort_key(row: Any):
        d = _as_date(_first_value(row, "record_date", "payment_date"))
        receipt = str(_first_value(row, "receipt_no", default="") or "")
        return (d or date.min, receipt, _row_value(row, "id", 0) or 0)

    records.sort(key=sort_key)

    wb = load_workbook(DONATION_TEMPLATE, data_only=False)
    sheet_name = str(config["sheet"])
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"布施模板缺少工作表：{sheet_name}")

    # 直接使用原工作表，不能 copy_worksheet()，否则 Logo／图片会丢失。
    ws = wb[sheet_name]
    ws.sheet_view.showGridLines = False

    title_cell = str(config["title_cell"])
    account_key = str(config["account_key"])
    account = BRANCH_CONFIG[branch][account_key]

    ws[title_cell] = str(config["title"]).format(
        year=year,
        month=month,
        account=account,
    )

    data_start = int(config["data_start"])

    data_end, total_row = _ensure_donation_capacity(
        ws,
        record_count=len(records),
        data_start=data_start,
        template_total_row=int(config["template_total_row"]),
    )

    # 清除实际资料区旧内容，包括模板原本的 TRUE／FALSE。
    for row_no in range(data_start, data_end + 1):
        _clear_donation_data_row(
            ws,
            row_no,
            config,
        )

    # 写入系统记录。
    for row_no, record in enumerate(records, start=data_start):
        _write_donation_row(
            ws,
            row_no,
            record,
            config,
        )

    _set_donation_totals(
        ws,
        config,
        data_start=data_start,
        data_end=data_end,
        total_row=total_row,
    )

    # 删除 Total Amount 以下的多余模板空白行。
    if ws.max_row > total_row:
        ws.delete_rows(
            total_row + 1,
            ws.max_row - total_row,
        )

    # 打印／转 PDF 时只使用实际报表范围。
    ws.print_area = (
        f"A1:{ws.cell(total_row, ws.max_column).coordinate}"
    )

    # 避免日期及金额打开时显示 ####。
    minimum_widths = {
        "A": 12.0,
        ws.cell(1, int(config["cash_col"])).column_letter: 12.0,
        ws.cell(1, int(config["cheque_col"])).column_letter: 12.0,
        ws.cell(1, int(config["bank_col"])).column_letter: 17.0,
        ws.cell(1, int(config["cdm_amount_col"])).column_letter: 15.0,
        ws.cell(1, int(config["cdm_date_col"])).column_letter: 14.0,
    }
    for column, minimum_width in minimum_widths.items():
        current_width = ws.column_dimensions[column].width or 0
        if current_width < minimum_width:
            ws.column_dimensions[column].width = minimum_width

    # 导出只保留目标报表工作表，避免财政误改其他旧样本。
    for other_name in list(wb.sheetnames):
        if other_name != sheet_name:
            del wb[other_name]

    wb.active = 0
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    return wb


@finance_bp.route("/export/donation")
def export_donation_excel():
    ym = _normalize_ym(request.args.get("ym"))
    branch = _normalise_branch(request.args.get("branch"))
    if branch != "CHE":
        abort(
            400,
            description="STW 目前只有月费报表"
        )
    category = _normalise_donation_category(request.args.get("category"))
    config = DONATION_REPORTS[category]

    db_categories = tuple(config["db_categories"])
    placeholders = ", ".join(["%s"] * len(db_categories))
    branch_condition, branch_params = _branch_sql_condition(branch, "fr")

    sql = f"""
        select
            fr.id,
            fr.record_date,
            fr.receipt_no,
            fr.member_id,
            fr.name,
            fr.phone,
            fr.amount,
            fr.payment_method,
            fr.bank_name,
            fr.bank_ref,
            fr.remarks
        from finance_records fr
        where fr.record_type = 'income'
          and fr.category in ({placeholders})
          and to_char(fr.record_date, 'YYYY-MM') = %s
          and {branch_condition}
        order by
            fr.record_date,
            fr.receipt_no nulls last,
            fr.id
    """

    params = (*db_categories, ym, *branch_params)
    records = db_query(sql, params, fetchall=True) or []

    wb = build_donation_workbook(records, ym, category, branch)
    filename = f"{branch}_{category}收纳表_{ym}.xlsx"

    return send_file(
        _workbook_bytes(wb),
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# =============================================================================
# Petty Cash Excel
# =============================================================================

def _is_petty_cash_in(record: Any) -> bool:
    record_type = str(_first_value(record, "record_type", default="") or "").lower()
    category = str(_first_value(record, "category", default="") or "").lower()

    return (
        record_type == "income"
        and (
            "petty" in category
            or "零用" in category
            or "备用金" in category
        )
    )


def _petty_cash_amount(record: Any) -> float:
    return _amount_value(record)


def _query_petty_cash_records(
    ym: str,
    branch: str,
):

    branch_condition, branch_params = (
        _branch_sql_condition(
            branch,
            "fr",
        )
    )

    return db_query(
        f"""
        select
            fr.id,
            fr.record_date,
            fr.record_type,
            fr.category,
            fr.name,
            fr.receipt_no,
            fr.amount,
            fr.payment_method,
            fr.remarks

        from finance_records fr

        where to_char(fr.record_date, 'YYYY-MM') = %s
          and {branch_condition}
          and (
                (
                    fr.record_type = 'income'
                    and (
                        lower(coalesce(fr.category, ''))
                            like '%%petty%%'

                        or coalesce(fr.category, '')
                            like '%%零用%%'

                        or coalesce(fr.category, '')
                            like '%%备用金%%'
                    )
                )

                or

                (
                    fr.record_type = 'expense'
                    and (
                        lower(coalesce(fr.payment_method, ''))
                            like '%%cash%%'

                        or coalesce(fr.payment_method, '')
                            like '%%现金%%'

                        or coalesce(fr.payment_method, '')
                            like '%%现款%%'

                        or lower(coalesce(fr.payment_method, ''))
                            like '%%petty%%'
                    )
                )
          )

        order by
            fr.record_date,
            fr.id
        """,
        (
            ym,
            *branch_params,
        ),
        fetchall=True,
    ) or []


def _query_petty_cash_before(ym: str, branch: str) -> float:

    first_day = datetime.strptime(
        f"{ym}-01",
        "%Y-%m-%d"
    ).date()

    branch_condition, branch_params = (
        _branch_sql_condition(
            branch,
            "fr",
        )
    )

    row = db_query(
        f"""
        select
            coalesce(sum(
                case
                    when fr.record_type = 'income'
                     and (
                        lower(coalesce(fr.category, '')) like '%%petty%%'
                        or coalesce(fr.category, '') like '%%零用%%'
                        or coalesce(fr.category, '') like '%%备用金%%'
                     )
                    then fr.amount

                    when fr.record_type = 'expense'
                     and (
                        lower(coalesce(fr.payment_method, '')) like '%%cash%%'
                        or coalesce(fr.payment_method, '') like '%%现金%%'
                        or coalesce(fr.payment_method, '') like '%%现款%%'
                        or lower(coalesce(fr.payment_method, '')) like '%%petty%%'
                     )
                    then -fr.amount

                    else 0
                end
            ), 0) as balance

        from finance_records fr

        where fr.record_date < %s
          and {branch_condition}
        """,
        (
            first_day,
            *branch_params,
        ),
        fetchone=True,
    )

    if not row:
        return 0.0

    return float(
        row.get("balance") or 0
    )

def set_merged_cell_value(ws, cell_ref: str, value):

    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            top_left = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            )
            top_left.value = value
            return

    ws[cell_ref] = value


def build_petty_cash_workbook(
    records: Iterable[Any],
    ym: str,
    branch: str = "CHE",
    opening_balance: float = 0.0,
):
    """
    建立 Petty Cash 月报。

    不使用 Excel 模板，直接建立新的工作簿，
    避免合并单元格、旧格式及模板损坏问题。
    """

    branch = _normalise_branch(branch)
    records = list(records)

    year, month = map(int, ym.split("-"))

    branch_label = (
        BRANCH_CONFIG.get(branch, {}).get("label")
        or branch
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{branch}-{ym}"

    ws.sheet_view.showGridLines = False

    # ==========================================
    # 颜色及样式
    # ==========================================
    title_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    opening_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    total_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    thin_side = Side(
        style="thin",
        color="B7B7B7",
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    left_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    right_alignment = Alignment(
        horizontal="right",
        vertical="center",
    )

    # ==========================================
    # 第 1 行：主标题
    # ==========================================
    ws.merge_cells("A1:E1")

    ws["A1"] = f"{branch_label} Petty Cash"

    ws["A1"].font = Font(
        name="Arial",
        size=18,
        bold=True,
        color="FFFFFF",
    )

    ws["A1"].fill = title_fill
    ws["A1"].alignment = center_alignment
    ws["A1"].border = border

    ws.row_dimensions[1].height = 34

    # ==========================================
    # 第 2 行：月份
    # ==========================================
    ws.merge_cells("A2:E2")

    ws["A2"] = date(year, month, 1)
    ws["A2"].number_format = "mmmm yyyy"

    ws["A2"].font = Font(
        name="Arial",
        size=13,
        bold=True,
    )

    ws["A2"].alignment = center_alignment
    ws["A2"].border = border

    ws.row_dimensions[2].height = 25

    # ==========================================
    # 第 3 行：表头
    # ==========================================
    headers = [
        "Date",
        "Cash In",
        "Cash Out",
        "Balance",
        "Description / Reference",
    ]

    for col_no, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=3,
            column=col_no,
        )

        cell.value = header
        cell.font = Font(
            name="Arial",
            size=11,
            bold=True,
        )
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    ws.row_dimensions[3].height = 30

    # ==========================================
    # 第 4 行：期初余额
    # ==========================================
    ws["A4"] = date(year, month, 1)
    ws["A4"].number_format = "dd/mm/yyyy"

    ws["B4"] = None
    ws["C4"] = None
    ws["D4"] = float(opening_balance)
    ws["D4"].number_format = '#,##0.00'

    ws["E4"] = "Opening Balance / 期初余额"

    for col_no in range(1, 6):

        cell = ws.cell(
            row=4,
            column=col_no,
        )

        cell.fill = opening_fill
        cell.border = border

        if col_no == 5:
            cell.alignment = left_alignment
        elif col_no in (2, 3, 4):
            cell.alignment = right_alignment
        else:
            cell.alignment = center_alignment

    ws["E4"].font = Font(
        name="Arial",
        bold=True,
    )

    # ==========================================
    # 第 5 行开始：本月记录
    # ==========================================
    data_start_row = 5
    current_row = data_start_row

    total_cash_in = 0.0
    total_cash_out = 0.0

    for record in records:

        record_date = _as_date(
            _first_value(
                record,
                "record_date",
                "payment_date",
            )
        )

        amount = float(
            _petty_cash_amount(record)
            or 0
        )

        is_cash_in = _is_petty_cash_in(record)

        if is_cash_in:
            total_cash_in += amount
        else:
            total_cash_out += amount

        ws.cell(
            row=current_row,
            column=1,
        ).value = record_date

        ws.cell(
            row=current_row,
            column=2,
        ).value = (
            amount
            if is_cash_in
            else None
        )

        ws.cell(
            row=current_row,
            column=3,
        ).value = (
            amount
            if not is_cash_in
            else None
        )

        previous_balance_row = current_row - 1

        ws.cell(
            row=current_row,
            column=4,
        ).value = (
            f"=D{previous_balance_row}"
            f"+IF(B{current_row}=\"\",0,B{current_row})"
            f"-IF(C{current_row}=\"\",0,C{current_row})"
        )

        category = str(
            _first_value(
                record,
                "category",
                default="",
            )
            or ""
        ).strip()

        name = str(
            _first_value(
                record,
                "name",
                default="",
            )
            or ""
        ).strip()

        reference = str(
            _first_value(
                record,
                "receipt_no",
                "bank_ref",
                "pv_no",
                default="",
            )
            or ""
        ).strip()

        remarks = str(
            _first_value(
                record,
                "remarks",
                default="",
            )
            or ""
        ).strip()

        description_parts = [
            value
            for value in (
                category,
                name,
                reference,
                remarks,
            )
            if value
        ]

        ws.cell(
            row=current_row,
            column=5,
        ).value = (
            " | ".join(description_parts)
            or "-"
        )

        # 日期格式
        ws.cell(
            row=current_row,
            column=1,
        ).number_format = "dd/mm/yyyy"

        # 金额格式
        for col_no in (2, 3, 4):
            ws.cell(
                row=current_row,
                column=col_no,
            ).number_format = '#,##0.00'

        # 边框和对齐
        for col_no in range(1, 6):

            cell = ws.cell(
                row=current_row,
                column=col_no,
            )

            cell.border = border

            if col_no == 5:
                cell.alignment = left_alignment
            elif col_no in (2, 3, 4):
                cell.alignment = right_alignment
            else:
                cell.alignment = center_alignment

        ws.row_dimensions[current_row].height = 24

        current_row += 1

    # ==========================================
    # 没有资料
    # ==========================================
    if not records:

        ws.merge_cells(
            start_row=data_start_row,
            start_column=1,
            end_row=data_start_row,
            end_column=5,
        )

        empty_cell = ws.cell(
            row=data_start_row,
            column=1,
        )

        empty_cell.value = "本月没有 Petty Cash 记录"
        empty_cell.alignment = center_alignment
        empty_cell.font = Font(
            italic=True,
            color="666666",
        )
        empty_cell.border = border

        current_row += 1

    # ==========================================
    # 合计行
    # ==========================================
    total_row = current_row

    ws["A" + str(total_row)] = "TOTAL / 合计"

    ws["B" + str(total_row)] = total_cash_in
    ws["C" + str(total_row)] = total_cash_out

    if records:
        last_data_row = total_row - 1

        ws["D" + str(total_row)] = (
            f"=D{last_data_row}"
        )
    else:
        ws["D" + str(total_row)] = (
            float(opening_balance)
        )

    ws["E" + str(total_row)] = (
        "Closing Balance / 期末余额"
    )

    for col_no in range(1, 6):

        cell = ws.cell(
            row=total_row,
            column=col_no,
        )

        cell.fill = total_fill
        cell.border = border
        cell.font = Font(
            name="Arial",
            bold=True,
        )

        if col_no == 5:
            cell.alignment = left_alignment
        elif col_no in (2, 3, 4):
            cell.alignment = right_alignment
        else:
            cell.alignment = center_alignment

    for col_no in (2, 3, 4):
        ws.cell(
            row=total_row,
            column=col_no,
        ).number_format = '#,##0.00'

    ws.row_dimensions[total_row].height = 26

    # ==========================================
    # 栏宽
    # ==========================================
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 52

    # ==========================================
    # 冻结及筛选
    # ==========================================
    ws.freeze_panes = "A5"

    if records:
        ws.auto_filter.ref = (
            f"A3:E{total_row - 1}"
        )

    # ==========================================
    # 列印设置
    # ==========================================
    ws.print_area = f"A1:E{total_row}"

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.print_options.horizontalCentered = True

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    ws.oddFooter.center.text = (
        f"{branch} Petty Cash · {ym}"
    )

    # ==========================================
    # Excel 打开时重新计算公式
    # ==========================================
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    wb.active = 0

    return wb


@finance_bp.route("/export/petty-cash")
def export_petty_cash_excel():
    ym = _normalize_ym(request.args.get("ym"))
    branch = _normalise_branch(request.args.get("branch"))
    if branch != "CHE":
        abort(
            400,
            description="STW 目前只有月费报表"
        )
    records = _query_petty_cash_records(ym, branch)
    opening_balance = _query_petty_cash_before(ym, branch)

    wb = build_petty_cash_workbook(
        records,
        ym,
        branch,
        opening_balance,
    )

    return send_file(
        _workbook_bytes(wb),
        as_attachment=True,
        download_name=f"{branch}_Petty_Cash_{ym}.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# =============================================================================
# 下载全部 Excel（ZIP）
# =============================================================================

def _query_monthly_records(ym: str, branch: str):
    branch_condition, branch_params = _branch_sql_condition(branch, "fr")

    return db_query(
        f"""
        select
            fr.id,
            fr.record_date,
            fr.receipt_no,
            fr.member_id,
            fr.name,
            fr.phone,
            fr.amount,
            fr.payment_method,
            fr.bank_name,
            fr.bank_ref,
            fr.month_from,
            fr.month_to,
            fr.remarks,
            mp.month_count
        from finance_records fr
        left join member_payments mp
          on mp.member_id = fr.member_id
         and coalesce(mp.receipt_no, '') = coalesce(fr.receipt_no, '')
         and mp.amount = fr.amount
         and mp.payment_date = fr.record_date
        where fr.record_type = 'income'
          and fr.category = '月费'
          and to_char(fr.record_date, 'YYYY-MM') = %s
          and {branch_condition}
        order by fr.record_date, fr.receipt_no nulls last, fr.id
        """,
        (ym, *branch_params),
        fetchall=True,
    ) or []


def _query_donation_records(ym: str, branch: str, category: str):
    config = DONATION_REPORTS[category]
    db_categories = tuple(config["db_categories"])
    placeholders = ", ".join(["%s"] * len(db_categories))
    branch_condition, branch_params = _branch_sql_condition(branch, "fr")

    return db_query(
        f"""
        select
            fr.id,
            fr.record_date,
            fr.receipt_no,
            fr.member_id,
            fr.name,
            fr.phone,
            fr.amount,
            fr.payment_method,
            fr.bank_name,
            fr.bank_ref,
            fr.remarks
        from finance_records fr
        where fr.record_type = 'income'
          and fr.category in ({placeholders})
          and to_char(fr.record_date, 'YYYY-MM') = %s
          and {branch_condition}
        order by fr.record_date, fr.receipt_no nulls last, fr.id
        """,
        (*db_categories, ym, *branch_params),
        fetchall=True,
    ) or []


def _add_workbook_to_zip(zip_file, filename: str, wb) -> None:
    workbook_file = _workbook_bytes(wb)
    zip_file.writestr(filename, workbook_file.getvalue())


@finance_bp.route("/export/all")
def export_all_finance_excel():
    ym = _normalize_ym(request.args.get("ym"))
    branch = _normalise_branch(request.args.get("branch"))
    if branch != "CHE":
        abort(
            400,
            description="STW 目前只有月费报表"
        )
    output = io.BytesIO()

    with zipfile.ZipFile(
        output,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    ) as zip_file:

        monthly_records = _query_monthly_records(ym, branch)
        _add_workbook_to_zip(
            zip_file,
            f"{branch}_月费收纳表_{ym}.xlsx",
            build_monthly_fee_workbook(monthly_records, ym, branch),
        )

        for category in ("财布施", "观音村", "初一十五"):
            donation_records = _query_donation_records(
                ym,
                branch,
                category,
            )
            _add_workbook_to_zip(
                zip_file,
                f"{branch}_{category}收纳表_{ym}.xlsx",
                build_donation_workbook(
                    donation_records,
                    ym,
                    category,
                    branch,
                ),
            )

        petty_records = _query_petty_cash_records(ym, branch)
        petty_opening = _query_petty_cash_before(ym, branch)
        _add_workbook_to_zip(
            zip_file,
            f"{branch}_Petty_Cash_{ym}.xlsx",
            build_petty_cash_workbook(
                petty_records,
                ym,
                branch,
                petty_opening,
            ),
        )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"{branch}_财政Excel_{ym}.zip",
        mimetype="application/zip",
    )


# =============================================================================
# 财政 Excel 下载中心
# =============================================================================

FINANCE_EXPORT_CENTER_HTML = """
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>财政 Excel 下载中心</title>
<link rel="stylesheet"
      href="{{ url_for('static', filename='css/toolbox.css') }}">
<style>
.export-page{
    max-width:900px;
}
.export-filter{
    display:grid;
    grid-template-columns:1fr 1fr;
    gap:16px;
    margin-bottom:20px;
}
.export-grid{
    display:grid;
    grid-template-columns:repeat(2, minmax(0, 1fr));
    gap:16px;
}
.export-card{
    display:block;
    padding:20px;
    border-radius:18px;
    text-decoration:none;
    color:#222;
    background:#fff;
    border:1px solid #e5e7eb;
    box-shadow:0 6px 18px rgba(0,0,0,.06);
}
.export-card strong{
    display:block;
    font-size:23px;
    margin-bottom:8px;
}
.export-card span{
    color:#666;
    font-size:17px;
}
.export-card-main{
    grid-column:1 / -1;
    background:#fff7d6;
    border-color:#e7bd43;
}
@media(max-width:650px){
    .export-filter,
    .export-grid{
        grid-template-columns:1fr;
    }
    .export-card-main{
        grid-column:auto;
    }
}
</style>
</head>
<body>
<div class="page export-page">
<div class="card">

<h1 class="page-title">📥 财政 Excel 下载中心</h1>
<p class="page-subtitle">
    选择月份和分会，再下载所需报表。
</p>

<form method="get" class="export-filter">
    <div class="form-group">
        <label class="form-label">月份</label>
        <input
            class="form-input"
            type="month"
            name="ym"
            value="{{ ym }}"
            onchange="this.form.submit()">
    </div>

    <div class="form-group">
        <label class="form-label">分会</label>
        <select
            class="form-input"
            name="branch"
            onchange="this.form.submit()">
            <option value="CHE" {% if branch == "CHE" %}selected{% endif %}>
                蕉赖 CHE
            </option>
            <option value="STW" {% if branch == "STW" %}selected{% endif %}>
                实兆远 STW
            </option>
        </select>
    </div>
</form>

<div class="export-grid">

    {% if branch == "CHE" %}

    <a class="export-card export-card-main"
       href="{{ url_for('finance.export_all_finance_excel',
                        ym=ym, branch=branch) }}">
        <strong>📦 下载本月全部 Excel</strong>
        <span>
            月费、财布施、观音村、初一十五及 Petty Cash
        </span>
    </a>

    {% endif %}


    <a class="export-card"
       href="{{ url_for('finance.export_monthly_fee_excel',
                        ym=ym, branch=branch) }}">
        <strong>💳 月费收纳表</strong>
        <span>{{ branch }} · {{ ym }}</span>
    </a>


    {% if branch == "CHE" %}

    <a class="export-card"
       href="{{ url_for('finance.export_donation_excel',
                        ym=ym,
                        branch=branch,
                        category='财布施') }}">
        <strong>🙏 财布施收纳表</strong>
        <span>{{ branch }} · {{ ym }}</span>
    </a>

    <a class="export-card"
       href="{{ url_for('finance.export_donation_excel',
                        ym=ym,
                        branch=branch,
                        category='观音村') }}">
        <strong>🌿 观音村善款收纳表</strong>
        <span>{{ branch }} · {{ ym }}</span>
    </a>

    <a class="export-card"
       href="{{ url_for('finance.export_donation_excel',
                        ym=ym,
                        branch=branch,
                        category='初一十五') }}">
        <strong>🍱 初一十五／膳食结缘</strong>
        <span>{{ branch }} · {{ ym }}</span>
    </a>

    <a class="export-card"
       href="{{ url_for('finance.export_petty_cash_excel',
                        ym=ym,
                        branch=branch) }}">
        <strong>💵 Petty Cash</strong>
        <span>{{ branch }} · {{ ym }}</span>
    </a>

    {% endif %}

</div>

<div class="btn-row" style="margin-top:22px;">
    <a class="btn-tool btn-secondary"
       href="/finance">
        ← 返回财政首页
    </a>
</div>

</div>
</div>
</body>
</html>
"""


@finance_bp.route("/reports/excel")
def finance_excel_export_center():
    ym = request.args.get("ym") or date.today().strftime("%Y-%m")
    ym = _normalize_ym(ym)
    branch = _normalise_branch(request.args.get("branch"))
    
    return render_template_string(
        FINANCE_EXPORT_CENTER_HTML,
        ym=ym,
        branch=branch,
    )