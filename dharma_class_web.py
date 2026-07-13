# dharma_class_web.py

from io import BytesIO
from db import get_conn
from flask import send_file
from zoneinfo import ZoneInfo
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from psycopg2.extras import RealDictCursor
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from flask import Blueprint, render_template_string, request, redirect, url_for, flash, send_file


dharma_class_bp = Blueprint(
    "dharma_class",
    __name__,
    url_prefix="/class"
)


STATUS_LABELS = {
    "present": "出席",
    "absent": "缺席",
    "late": "迟到",
    "farm": "农舍",
}

BAIHUA_STATUS_LABELS = {
    "submitted_done": "有交",
    "missing": "没交",
    "submitted_no_answer": "有交，没做题",
    "absent": "缺席",
}

SCRIPTURE_STATUS_LABELS = {
    "submitted_recited": "有交，有念",
    "submitted_not_recited": "有交，没念",
    "missing": "没交",
    "absent": "缺席",
}

AUDIT_FIELD_LABELS = {
    "status": "出席状态",
    "baihua_status": "白话佛法功课",
    "scripture_status": "经文功课",
    "teacher_name": "负责老师",
    "topic": "课题",
    "material": "教材／资源",
    "content": "教学内容",
    "remark": "备注",
    "name": "学生姓名",
    "english_name": "英文名",
    "gender": "性别",
    "birth_year": "出生年份",
    "parent_name": "父／母／监护人",
    "parent_phone": "联系电话",
    "group_id": "组别",
    "student_status": "学生状态",
}

AUDIT_VALUE_LABELS = {
    "present": "出席",
    "late": "迟到",
    "farm": "农舍",
    "absent": "缺席",
    "submitted_done": "有交",
    "submitted_no_answer": "有交，没做题",
    "submitted_recited": "有交，有念",
    "submitted_not_recited": "有交，没念",
    "submitted": "有交（旧状态）",
    "missing": "没交",
    "active": "在读",
    "inactive": "暂停",
    "paused": "暂停",
    "graduated": "毕业",
    "withdrawn": "退学",
    "transferred": "转组／转会",
}

HIGH_GROUP_TEACHERS = [
    "黄莉珍",
    "刘永耀",
    "许愫芩",
    "王康芬",
    "黄薈菏",
]

YOUTH_GROUP_TEACHERS = [
    "许银铃",
    "林臣顺",
    "许茹慧",
    "刘铧忆",
    "吴文杰",
    "王康芬",
    "方玉芬",
]

LOW_GROUP_TEACHERS = [
    "陈映如",
    "黄丽萍",
    "伍蔚枋",
]

GROUP_TEACHERS = {
    "低年组": LOW_GROUP_TEACHERS,
    "少年组": YOUTH_GROUP_TEACHERS,
    "高年组": HIGH_GROUP_TEACHERS,
}


def _audit_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return AUDIT_VALUE_LABELS.get(text, text)


def write_dharma_audit_log(
    cur,
    *,
    actor_name,
    action_type,
    entity_type,
    field_name,
    old_value=None,
    new_value=None,
    entity_id=None,
    student_id=None,
    student_name=None,
    record_date=None,
    group_id=None,
    branch="CHE",
):
    """只在值真的改变时写入审计记录。"""

    old_text = "" if old_value is None else str(old_value)
    new_text = "" if new_value is None else str(new_value)

    if old_text == new_text:
        return

    cur.execute("""
        insert into dharma_audit_logs
        (
            branch,
            actor_name,
            action_type,
            entity_type,
            entity_id,
            student_id,
            student_name,
            record_date,
            group_id,
            field_name,
            old_value,
            new_value,
            source_ip
        )
        values
        (
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s
        )
    """, (
        branch,
        actor_name or "老师",
        action_type,
        entity_type,
        entity_id,
        student_id,
        student_name,
        record_date,
        group_id,
        field_name,
        old_text or None,
        new_text or None,
        request.headers.get("X-Forwarded-For", request.remote_addr),
    ))

def beautify_simple_excel(ws):

    header_fill = PatternFill("solid", fgColor="5B9BD5")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            value = str(value) if value is not None else ""
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 24

@dharma_class_bp.route("/")
def class_home():

    malaysia_now = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    )

    today = malaysia_now.date()
    today_str = today.isoformat()

    weekday_names = [
        "星期一",
        "星期二",
        "星期三",
        "星期四",
        "星期五",
        "星期六",
        "星期日",
    ]

    today_display = (
        f"{today.year}年"
        f"{today.month}月"
        f"{today.day}日 "
        f"{weekday_names[today.weekday()]}"
    )

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            cur.execute("""
                select
                    g.id,
                    g.name,
                    g.sort_order,

                    count(distinct s.id) as total_students,

                    count(
                        distinct case
                            when a.id is not null
                            then s.id
                        end
                    ) as marked_count,

                    count(
                        distinct case
                            when a.status in (
                                'present',
                                'late',
                                'farm'
                            )
                            then s.id
                        end
                    ) as attended_count,

                    count(
                        distinct case
                            when a.status = 'present'
                            then s.id
                        end
                    ) as present_count,

                    count(
                        distinct case
                            when a.status = 'late'
                            then s.id
                        end
                    ) as late_count,

                    count(
                        distinct case
                            when a.status = 'farm'
                            then s.id
                        end
                    ) as farm_count,

                    count(
                        distinct case
                            when a.status in (
                                'absent',
                                'leave'
                            )
                            then s.id
                        end
                    ) as absent_count,

                    count(distinct l.id) as lesson_count

                from dharma_class_groups g

                left join dharma_students s
                    on s.group_id = g.id
                   and s.status = 'active'
                   and s.branch = 'CHE'

                left join dharma_attendance a
                    on a.student_id = s.id
                   and a.class_date = %s
                   and a.branch = 'CHE'

                left join dharma_class_lessons l
                    on l.group_id = g.id
                   and l.lesson_date = %s
                   and l.branch = 'CHE'

                where g.is_active = true

                group by
                    g.id,
                    g.name,
                    g.sort_order

                order by
                    g.sort_order,
                    g.id
            """, (
                today_str,
                today_str,
            ))

            group_stats = cur.fetchall()

            group_order = {
                "低年组": 1,
                "高年组": 2,
                "少年组": 3,
            }

            group_stats.sort(
                key=lambda x: group_order.get(x["name"], 99)
            )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<link
    rel="manifest"
    href="/dharma-class-manifest.json"
>

<link
    rel="icon"
    href="/static/dharma_icon.png?v=1"
>

<meta
    name="theme-color"
    content="#f6c54e"
>

<title>蕉赖佛学班系统</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>

/* =========================================================
   老师首页
   ========================================================= */

.teacher-home-card {
    position: relative;
    overflow: hidden;
    background:
        linear-gradient(
            145deg,
            #ffffff 0%,
            #fffdf7 58%,
            #fff6dc 100%
        );
}

.teacher-home-card::after {
    content: "🌸";
    position: absolute;
    right: 18px;
    bottom: 10px;
    font-size: 72px;
    opacity: 0.08;
    pointer-events: none;
}

.teacher-admin-link {
    position: absolute;
    top: 15px;
    right: 15px;

    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 5px;

    min-height: 40px;
    padding: 7px 12px;

    border: 1px solid #ead9a4;
    border-radius: 999px;

    background: rgba(255,255,255,0.9);
    color: #795d17;

    font-size: 15px;
    font-weight: 800;
    text-decoration: none;
}

.teacher-admin-link:hover {
    background: #fff7d6;
}

.teacher-welcome {
    padding-right: 112px;
}

.teacher-blessing {
    color: #8a6d3b;
    font-weight: 800;
    line-height: 1.7;
}

.today-banner {
    margin: 18px 0 22px;
    padding: 17px;

    border: 1px solid #f2d58d;
    border-radius: 18px;

    background:
        linear-gradient(
            135deg,
            #fff8e8,
            #fffdf7
        );

    text-align: center;
    font-size: 19px;
    font-weight: 900;
    color: #6f5315;
}

/* =========================================================
   今日组别
   ========================================================= */

.today-group-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 14px;
    margin-top: 14px;
}

.today-group-card {
    border: 1px solid #e6e9ed;
    border-top: 5px solid #60a5fa;
    border-radius: 18px;
    padding: 17px;
    background: #fff;
}

.today-group-card:nth-child(2) {
    border-top-color: #f59e0b;
}

.today-group-card:nth-child(3) {
    border-top-color: #a78bfa;
}

.today-group-name {
    font-size: 21px;
    font-weight: 900;
}

.today-group-main {
    margin: 10px 0 7px;
    font-size: 29px;
    font-weight: 900;
}

.today-group-detail {
    color: #666;
    font-size: 15px;
    line-height: 1.7;
}

.today-group-status {
    display: inline-block;
    margin-top: 11px;
    padding: 6px 11px;
    border-radius: 999px;
    font-size: 14px;
    font-weight: 800;
}

.group-complete {
    background: #dcfce7;
    color: #166534;
}

.group-pending {
    background: #fef3c7;
    color: #92400e;
}

.group-empty {
    background: #f1f5f9;
    color: #64748b;
}

/* =========================================================
   老师主要功能
   ========================================================= */

.teacher-main-action {
    margin-top: 22px;
}

.teacher-main-action .btn-tool {
    min-height: 76px;
    font-size: 26px;
    border-radius: 20px;
}

.student-search-box {
    margin-top: 24px;
    padding-top: 22px;
    border-top: 1px solid #ece6d6;
}

.student-search-title {
    margin-bottom: 12px;
    font-size: 21px;
    font-weight: 900;
}

.student-search-help {
    margin: -4px 0 13px;
    color: #777;
    line-height: 1.6;
}

.student-search-box .btn-tool {
    min-height: 58px;
    font-size: 20px;
}

/* =========================================================
   手机
   ========================================================= */

@media (max-width: 760px) {

    .teacher-welcome {
        padding-right: 0;
        padding-top: 48px;
    }

    .teacher-admin-link {
        top: 12px;
        right: 12px;
    }

    .today-group-grid {
        grid-template-columns: 1fr;
    }

    .today-group-card {
        min-height: auto;
    }
}

</style>
</head>

<body>
<div class="page">

    <div class="card teacher-home-card">

        <a
            class="teacher-admin-link"
            href="{{ url_for(
                'dharma_class.class_admin'
            ) }}"
        >
            ⚙️ 负责人
        </a>

        <div class="teacher-welcome">

            <h1 class="page-title">
                📘 蕉赖佛学班
            </h1>

            <p class="page-subtitle teacher-blessing">
                🌸 学习佛法・增长智慧・培养慈悲 🌸
            </p>

        </div>

        <div class="today-banner">
            📅 今日：{{ today_display }}
        </div>

        <h2 class="section-title">
            📊 今日各组情况
        </h2>

        {% if group_stats %}

            <div class="today-group-grid">

                {% for g in group_stats %}

                    <div class="today-group-card">

                        <div class="today-group-name">
                            {{ g.name }}
                        </div>

                        <div class="today-group-main">
                            {{ g.attended_count or 0 }}
                            /
                            {{ g.total_students or 0 }}
                        </div>

                        <div class="today-group-detail">
                            已点名：{{ g.marked_count or 0 }} 位<br>
                            出席：{{ g.present_count or 0 }} 位<br>
                            迟到：{{ g.late_count or 0 }} 位<br>
                            农舍：{{ g.farm_count or 0 }} 位<br>
                            缺席：{{ g.absent_count or 0 }} 位
                        </div>

                        {% if
                            g.lesson_count
                            and g.total_students
                            and g.marked_count == g.total_students
                        %}

                            <div class="
                                today-group-status
                                group-complete
                            ">
                                ✅ 已完成
                            </div>

                        {% elif
                            g.lesson_count
                            or g.marked_count
                        %}

                            <div class="
                                today-group-status
                                group-pending
                            ">
                                ⏳ 记录中
                            </div>

                        {% else %}

                            <div class="
                                today-group-status
                                group-empty
                            ">
                                尚未记录
                            </div>

                        {% endif %}

                    </div>

                {% endfor %}

            </div>

        {% else %}

            <div class="empty-state">
                暂时没有启用中的佛学班组别。
            </div>

        {% endif %}

        <div class="teacher-main-action">

            <a
                class="btn-tool btn-primary"
                href="{{ url_for(
                    'dharma_class.class_attendance'
                ) }}"
            >
                📚 点名
            </a>

        </div>

        <div class="student-search-box">

            <div class="student-search-title">
                
            </div>

            <div class="student-search-help">
                输入学生姓名、英文名或父／母／监护人电话。
            </div>

            <form
                method="get"
                action="{{ url_for(
                    'dharma_class.class_student_search'
                ) }}"
            >

                <div class="form-group">

                    <input
                        class="form-input"
                        name="q"
                        placeholder="例如：陈小明"
                        required
                    >

                </div>

                <div class="btn-row">

                    <button
                        class="btn-tool btn-success"
                        type="submit"
                    >
                        🔎 查看学生个人档案
                    </button>

                </div>

            </form>

        </div>

    </div>

</div>
</body>
</html>
""",
        today_str=today_str,
        today_display=today_display,
        group_stats=group_stats,
    )

@dharma_class_bp.route("/admin")
def class_admin():

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>佛学班负责人中心</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>

/* =========================================================
   负责人中心
   ========================================================= */

.admin-center-header {
    background:
        linear-gradient(
            135deg,
            #fff8e8,
            #ffffff
        );
    border: 1px solid #f0d89b;
}

.admin-center-note {
    margin-top: 15px;
    padding: 14px 16px;

    border-radius: 15px;
    background: #f8fafc;
    color: #666;

    font-size: 17px;
    line-height: 1.6;
}

.admin-section {
    margin-top: 25px;
}

.admin-section-title {
    margin-bottom: 12px;
    font-size: 21px;
    font-weight: 900;
}

.admin-tool-grid {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 13px;
}

.admin-tool {
    display: flex;
    align-items: center;
    gap: 14px;

    min-height: 82px;
    padding: 15px 17px;

    border: 1px solid #e4e7eb;
    border-radius: 18px;

    background: #fff;
    color: #333;
    text-decoration: none;

    transition:
        transform .15s ease,
        box-shadow .15s ease;
}

.admin-tool:hover {
    transform: translateY(-2px);
    box-shadow: 0 7px 20px rgba(0,0,0,.07);
}

.admin-tool-icon {
    display: flex;
    align-items: center;
    justify-content: center;

    width: 48px;
    height: 48px;
    flex: 0 0 48px;

    border-radius: 15px;
    background: #eef6ff;

    font-size: 25px;
}

.admin-tool-text {
    min-width: 0;
}

.admin-tool-title {
    font-size: 19px;
    font-weight: 900;
}

.admin-tool-help {
    margin-top: 4px;
    color: #777;
    font-size: 14px;
    line-height: 1.4;
}

.tool-student .admin-tool-icon {
    background: #fff2df;
}

.tool-teaching .admin-tool-icon {
    background: #f1edff;
}

.tool-report .admin-tool-icon {
    background: #e8f7ed;
}

.tool-system .admin-tool-icon {
    background: #eef2f6;
}
                                  
.admin-tool-grid-one{
    grid-template-columns:1fr;
}

.compact-tool-row{
    display:grid;
    grid-template-columns:1fr 1fr;
    gap:14px;
    margin:14px 0 24px;
}

.compact-tool-row a{
    display:flex;
    align-items:center;
    justify-content:center;

    height:50px;

    background:#fff8e8;
    border:1px solid #f0d68a;
    border-radius:14px;

    color:#795d17;
    text-decoration:none;
    font-size:18px;
    font-weight:700;

    transition:.2s;
}

.compact-tool-row a:hover{
    background:#ffe8a6;
    transform:translateY(-1px);
}

@media (max-width: 680px) {

    .admin-tool-grid {
        grid-template-columns: 1fr;
    }

}

</style>
</head>

<body>
<div class="page">

    <div class="card admin-center-header">

        <h1 class="page-title">
            ⚙️ 佛学班负责人中心
        </h1>

        <p class="page-subtitle">
            学生、课程、报表及系统管理功能。
        </p>

        <div class="admin-center-note">
            老师日常点名请使用首页的
            “今日上课”。这里主要提供负责人管理功能。
        </div>

    </div>

    <div class="card">

        <!-- 学生管理 -->

        <div class="admin-section">

            <div class="admin-section-title">
                👨‍🎓 学生管理
            </div>

            <div class="admin-tool-grid admin-tool-grid-one">

                <a
                    class="admin-tool tool-student"
                    href="{{ url_for(
                        'dharma_class.class_students'
                    ) }}"
                >
                    <div class="admin-tool-icon">
                        👧
                    </div>

                    <div class="admin-tool-text">
                        <div class="admin-tool-title">
                            学生名单
                        </div>

                        <div class="admin-tool-help">
                            搜索、查看档案、编辑及新增学生
                        </div>
                    </div>
                </a>

            </div>

            <div class="compact-tool-row">

                <a href="/class/students/import">
                    📥 导入学生 Excel
                </a>

                <a href="/class/students/export">
                    📤 导出学生名单
                </a>

            </div>

        </div>


        <!-- 教学管理 -->

        <div class="admin-section">

            <div class="admin-section-title">
                📖 教学管理
            </div>

            <div class="admin-tool-grid">

                <a
                    class="admin-tool tool-teaching"
                    href="{{ url_for(
                        'dharma_class.class_lessons'
                    ) }}"
                >
                    <div class="admin-tool-icon">
                        📖
                    </div>

                    <div class="admin-tool-text">
                        <div class="admin-tool-title">
                            课程记录
                        </div>

                        <div class="admin-tool-help">
                            查看及编辑历次课程内容
                        </div>
                    </div>
                </a>

                <a
                    class="admin-tool tool-teaching"
                    href="{{ url_for(
                        'dharma_class.class_records'
                    ) }}"
                >
                    <div class="admin-tool-icon">
                        📅
                    </div>

                    <div class="admin-tool-text">
                        <div class="admin-tool-title">
                            点名记录
                        </div>

                        <div class="admin-tool-help">
                            按日期查看及修改点名资料
                        </div>
                    </div>
                </a>

            </div>

        </div>


        <!-- 报表 -->

        <div class="admin-section">

            <div class="admin-section-title">
                📊 报表
            </div>

            <div class="admin-tool-grid admin-tool-grid-one">

                <a
                    class="admin-tool tool-report"
                    href="{{ url_for(
                        'dharma_class.class_reports'
                    ) }}"
                >
                    <div class="admin-tool-icon">
                        📊
                    </div>

                    <div class="admin-tool-text">
                        <div class="admin-tool-title">
                            报表中心
                        </div>

                        <div class="admin-tool-help">
                            月报、年报及功课追踪表
                        </div>
                    </div>
                </a>

            </div>

        </div>


        <!-- 系统管理 -->

        <div class="admin-section">

            <div class="admin-section-title">
                🛠️ 系统管理
            </div>

            <div class="admin-tool-grid">

                <a
                    class="admin-tool tool-system"
                    href="{{ url_for(
                        'dharma_class.class_audit_log'
                    ) }}"
                >
                    <div class="admin-tool-icon">
                        🛡️
                    </div>

                    <div class="admin-tool-text">
                        <div class="admin-tool-title">
                            修改记录
                        </div>

                        <div class="admin-tool-help">
                            查看哪位老师修改了资料
                        </div>
                    </div>
                </a>

                <a
                    class="admin-tool tool-system"
                    href="/class/promote"
                >
                    <div class="admin-tool-icon">
                        🎓
                    </div>

                    <div class="admin-tool-text">
                        <div class="admin-tool-title">
                            年度升班
                        </div>

                        <div class="admin-tool-help">
                            处理学生年度组别调整
                        </div>
                    </div>
                </a>

            </div>

        </div>


        <div class="btn-row">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_home'
                ) }}"
            >
                ⬅ 返回老师首页
            </a>

        </div>

    </div>

</div>
</body>
</html>
""")


@dharma_class_bp.route("/attendance", methods=["GET", "POST"])
def class_attendance():

    
    malaysia_today = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).date()

    selected_date = (
        request.values.get("class_date")
        or malaysia_today.isoformat()
    )

    selected_group_id = (
        request.values.get("group_id")
        or ""
    ).strip()
    
    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # =========================================================
            # 读取启用组别
            # =========================================================
            cur.execute("""
                select
                    id,
                    name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)

            groups = cur.fetchall()

            if not selected_group_id and groups:
                selected_group_id = str(groups[0]["id"])

            selected_group_name = ""

            for group in groups:
                if str(group["id"]) == str(selected_group_id):
                    selected_group_name = group["name"]
                    break

            teacher_list = GROUP_TEACHERS.get(
                selected_group_name,
                []
            )

            lesson = {
                "teacher_name": "",
                "record_teacher": "",
                "topic": "",
                "content": ""
            }

            # =========================================================
            # 保存课程、点名与功课
            # =========================================================
            if request.method == "POST":

                teacher_name = request.form.get(
                    "teacher_name",
                    ""
                ).strip()

                record_teacher = request.form.get(
                    "record_teacher",
                    ""
                ).strip()

                topic = request.form.get(
                    "topic",
                    ""
                ).strip()

                content = request.form.get(
                    "content",
                    ""
                ).strip()

                student_ids = request.form.getlist(
                    "student_id"
                )

                lesson = {
                    "teacher_name": teacher_name,
                    "record_teacher": record_teacher,
                    "topic": topic,
                    "content": content
                }

                if not selected_group_id:
                    flash("请选择组别。", "bad")

                elif not student_ids:
                    flash("这个组别没有可保存的学生。", "bad")

                else:

                    # 保存前先读取旧值，供 Audit Log 比较。
                    sid_list = [int(sid) for sid in student_ids if str(sid).isdigit()]

                    old_attendance_map = {}
                    old_homework_map = {}
                    student_name_map = {}

                    if sid_list:
                        cur.execute("""
                            select id, name
                            from dharma_students
                            where id = any(%s)
                        """, (sid_list,))
                        student_name_map = {
                            row["id"]: row["name"]
                            for row in cur.fetchall()
                        }

                        cur.execute("""
                            select student_id, status
                            from dharma_attendance
                            where branch = 'CHE'
                              and class_date = %s
                              and student_id = any(%s)
                        """, (selected_date, sid_list))
                        old_attendance_map = {
                            row["student_id"]: row["status"]
                            for row in cur.fetchall()
                        }

                        cur.execute("""
                            select student_id, baihua_status, scripture_status
                            from dharma_homework
                            where branch = 'CHE'
                              and class_date = %s
                              and student_id = any(%s)
                        """, (selected_date, sid_list))
                        old_homework_map = {
                            row["student_id"]: row
                            for row in cur.fetchall()
                        }

                    has_lesson_data = any([
                        teacher_name,
                        record_teacher,
                        topic,
                        content
                    ])

                    # -------------------------------------------------
                    # 有填写课程资料才保存课程
                    # 补录旧点名时可以全部留空
                    # -------------------------------------------------
                    if has_lesson_data:

                        cur.execute("""
                            select id, teacher_name, record_teacher, topic, content
                            from dharma_class_lessons
                            where branch = 'CHE'
                              and lesson_date = %s
                              and group_id = %s
                            limit 1
                        """, (selected_date, selected_group_id))
                        old_lesson = cur.fetchone() or {}

                        cur.execute("""
                            insert into dharma_class_lessons
                            (
                                branch,
                                lesson_date,
                                group_id,
                                teacher_name,
                                record_teacher,
                                topic,
                                content
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s
                            )
                            on conflict
                            (
                                branch,
                                lesson_date,
                                group_id
                            )
                            do update set
                                teacher_name = excluded.teacher_name,
                                record_teacher = excluded.record_teacher,
                                topic = excluded.topic,
                                content = excluded.content
                        """, (
                            selected_date,
                            selected_group_id,
                            teacher_name or None,
                            record_teacher or None,
                            topic or None,
                            content or None
                        ))

                        lesson_changes = {
                            "teacher_name": teacher_name or None,
                            "record_teacher": record_teacher or None,
                            "topic": topic or None,
                            "content": content or None,
                        }

                        for field_name, new_value in lesson_changes.items():
                            write_dharma_audit_log(
                                cur,
                                actor_name=record_teacher or teacher_name or "老师",
                                action_type="update" if old_lesson else "create",
                                entity_type="lesson",
                                entity_id=old_lesson.get("id"),
                                record_date=selected_date,
                                group_id=selected_group_id,
                                field_name=field_name,
                                old_value=old_lesson.get(field_name),
                                new_value=new_value,
                            )

                    marked_by = record_teacher or teacher_name or "老师"

                    # -------------------------------------------------
                    # 保存每位学生
                    # -------------------------------------------------
                    for sid in student_ids:

                        attendance_status = request.form.get(
                            f"status_{sid}",
                            "present"
                        ).strip()

                        # 防止异常状态写入
                        if attendance_status not in (
                            "present",
                            "absent",
                            "late",
                            "farm"
                        ):
                            attendance_status = "present"

                        baihua_status = request.form.get(
                            f"baihua_{sid}",
                            ""
                        ).strip()

                        scripture_status = request.form.get(
                            f"scripture_{sid}",
                            ""
                        ).strip()

                        # 缺席时，两项功课都记录为缺席。
                        if attendance_status == "absent":
                            baihua_status = "absent"
                            scripture_status = "absent"

                        if baihua_status not in (
                            "",
                            "submitted_done",
                            "submitted_no_answer",
                            "missing",
                            "absent"
                        ):
                            baihua_status = ""

                        if scripture_status not in (
                            "",
                            "submitted_recited",
                            "submitted_not_recited",
                            "missing",
                            "absent"
                        ):
                            scripture_status = ""

                        # ---------------------------------------------
                        # 保存出席：present / absent / late / farm
                        # ---------------------------------------------
                        cur.execute("""
                            insert into dharma_attendance
                            (
                                branch,
                                class_date,
                                student_id,
                                group_id,
                                status,
                                remark,
                                marked_by
                            )
                            values
                            (
                                'CHE',
                                %s,
                                %s,
                                %s,
                                %s,
                                null,
                                %s
                            )
                            on conflict
                            (
                                class_date,
                                student_id
                            )
                            do update set
                                group_id = excluded.group_id,
                                status = excluded.status,
                                marked_by = excluded.marked_by,
                                marked_at = now()
                        """, (
                            selected_date,
                            sid,
                            selected_group_id,
                            attendance_status,
                            marked_by
                        ))

                        sid_int = int(sid)
                        student_name = student_name_map.get(sid_int, "")

                        write_dharma_audit_log(
                            cur,
                            actor_name=marked_by,
                            action_type=(
                                "update"
                                if sid_int in old_attendance_map
                                else "create"
                            ),
                            entity_type="attendance",
                            student_id=sid_int,
                            student_name=student_name,
                            record_date=selected_date,
                            group_id=selected_group_id,
                            field_name="status",
                            old_value=old_attendance_map.get(sid_int),
                            new_value=attendance_status,
                        )

                        # ---------------------------------------------
                        # 至少记录了一项功课，才写入功课表
                        # ---------------------------------------------
                        if baihua_status or scripture_status:

                            cur.execute("""
                                insert into dharma_homework
                                (
                                    branch,
                                    class_date,
                                    student_id,
                                    group_id,
                                    baihua_status,
                                    scripture_status,
                                    recorded_by,
                                    updated_at
                                )
                                values
                                (
                                    'CHE',
                                    %s,
                                    %s,
                                    %s,
                                    %s,
                                    %s,
                                    %s,
                                    now()
                                )
                                on conflict
                                (
                                    branch,
                                    class_date,
                                    student_id
                                )
                                do update set
                                    group_id = excluded.group_id,
                                    baihua_status =
                                        excluded.baihua_status,
                                    scripture_status =
                                        excluded.scripture_status,
                                    recorded_by =
                                        excluded.recorded_by,
                                    updated_at = now()
                            """, (
                                selected_date,
                                sid,
                                selected_group_id,
                                baihua_status or None,
                                scripture_status or None,
                                marked_by
                            ))

                            old_hw = old_homework_map.get(sid_int, {})

                            write_dharma_audit_log(
                                cur,
                                actor_name=marked_by,
                                action_type=(
                                    "update"
                                    if sid_int in old_homework_map
                                    else "create"
                                ),
                                entity_type="homework",
                                student_id=sid_int,
                                student_name=student_name,
                                record_date=selected_date,
                                group_id=selected_group_id,
                                field_name="baihua_status",
                                old_value=old_hw.get("baihua_status"),
                                new_value=baihua_status or None,
                            )

                            write_dharma_audit_log(
                                cur,
                                actor_name=marked_by,
                                action_type=(
                                    "update"
                                    if sid_int in old_homework_map
                                    else "create"
                                ),
                                entity_type="homework",
                                student_id=sid_int,
                                student_name=student_name,
                                record_date=selected_date,
                                group_id=selected_group_id,
                                field_name="scripture_status",
                                old_value=old_hw.get("scripture_status"),
                                new_value=scripture_status or None,
                            )

                    conn.commit()

                    if has_lesson_data:
                        flash(
                            "课程、点名和学生功课已保存。",
                            "good"
                        )
                    else:
                        flash(
                            "学生点名和功课已保存。",
                            "good"
                        )

                    return redirect(
                        url_for(
                            "dharma_class.class_attendance",
                            class_date=selected_date,
                            group_id=selected_group_id
                        )
                    )

            # =========================================================
            # 读取已有课程
            # =========================================================
            cur.execute("""
                select
                    teacher_name,
                    record_teacher,
                    topic,
                    content
                from dharma_class_lessons
                where branch = 'CHE'
                  and lesson_date = %s
                  and group_id = %s
                limit 1
            """, (
                selected_date,
                selected_group_id
            ))

            saved_lesson = cur.fetchone()

            if saved_lesson:
                lesson = saved_lesson

            # =========================================================
            # 读取学生和点名状态
            # =========================================================
            cur.execute("""
                select
                    s.id,
                    s.name,
                    s.english_name,
                    s.gender,
                    s.birth_year,

                    coalesce(
                        a.status,
                        'present'
                    ) as attendance_status

                from dharma_students s

                left join dharma_attendance a
                    on a.student_id = s.id
                   and a.class_date = %s
                   and a.branch = 'CHE'

                where s.status = 'active'
                  and s.branch = 'CHE'
                  and s.group_id = %s

                order by
                    s.birth_year desc nulls last,
                    s.name
            """, (
                selected_date,
                selected_group_id
            ))

            students = cur.fetchall()

            if selected_group_name == "少年组":

                students.sort(
                    key=lambda s: (
                        0 if s["gender"] == "女"
                        else 1 if s["gender"] == "男"
                        else 2,
                        -(s["birth_year"] or 0),
                        s["name"]
                    )
                )

            # =========================================================
            # 读取学生功课
            # =========================================================
            cur.execute("""
                select
                    student_id,
                    baihua_status,
                    scripture_status
                from dharma_homework
                where branch = 'CHE'
                  and class_date = %s
                  and group_id = %s
            """, (
                selected_date,
                selected_group_id
            ))

            homework_rows = cur.fetchall()

            homework_map = {
                row["student_id"]: row
                for row in homework_rows
            }

            for student in students:

                homework = homework_map.get(
                    student["id"]
                )

                # 兼容旧资料：旧 submitted 自动转为新的完整状态。
                if student["attendance_status"] == "leave":
                    student["attendance_status"] = "absent"

                if homework:
                    baihua_value = homework["baihua_status"]
                    scripture_value = homework["scripture_status"]

                    if baihua_value == "submitted":
                        baihua_value = "submitted_done"

                    if scripture_value == "submitted":
                        scripture_value = "submitted_recited"

                    student["baihua_status"] = baihua_value
                    student["scripture_status"] = scripture_value

                else:
                    if student["attendance_status"] == "absent":
                        student["baihua_status"] = "absent"
                        student["scripture_status"] = "absent"
                    else:
                        student["baihua_status"] = "submitted_done"
                        student["scripture_status"] = "submitted_recited"

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>

<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>今日上课与点名</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>

/* =========================================================
   页面资料提示
   ========================================================= */

.current-selection {
    background: #eef6ff;
    border: 1px solid #cfe3ff;
    border-radius: 16px;
    padding: 14px 16px;
    margin-top: 16px;
    text-align: center;
    font-size: 18px;
    font-weight: 800;
}

/* =========================================================
   课程记录
   ========================================================= */

.lesson-note {
    background: #f6f7f9;
    border-radius: 14px;
    padding: 14px 16px;
    margin-bottom: 18px;
    line-height: 1.6;
}

/* =========================================================
   学生卡片
   ========================================================= */

.student-list {
    display: grid;
    gap: 18px;
}

.student-card {
    border: 1px solid rgba(0, 0, 0, 0.08);
    border-radius: 20px;
    padding: 20px;
    box-shadow: 0 3px 10px rgba(0, 0, 0, 0.04);
}

/* 统一白色学生卡，避免颜色被误解为缺席状态 */
.student-card {
    background: #ffffff;
    border-left: 6px solid #e7bd4f;
}

.student-header {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 16px;
}

.student-number{
    display:flex;
    align-items:center;
    justify-content:center;
    width:46px;
    height:46px;
    flex:0 0 46px;
    border-radius:50%;
    font-size:22px;
    font-weight:900;
}

.student-number-boy{

    background:#EAF6FF;
    border:2px solid #7EC8F8;
    color:#1565C0;

}

.student-number-girl{

    background:#FFF1F7;
    border:2px solid #F5A3C7;
    color:#C2185B;

}

.student-number-unknown{
    background:#FFF8DF;
    border:2px solid #F0C94E;
    color:#7A5A00;

}
                                  
.student-name {
    font-size: 40px;
    font-weight: 900;
}

.student-english-name {
    color: #777;
    font-size: 17px;
    font-weight: 400;
    margin-left: 6px;
}
                                  
.student-card-1{
    background:#FFF8E8;   /* 莲花米黄（比现在明显一点） */
}

.student-card-2{
    background:#E8F4FF;   /* 天空蓝 */
}

.student-card-3{
    background:#F5FFF2;   /* 莲花绿 */
}

/* =========================================================
   区块标题
   ========================================================= */

.student-section {
    border-top: 1px solid rgba(0, 0, 0, 0.09);
    padding-top: 15px;
    margin-top: 15px;
}

.student-section-title {
    font-size: 18px;
    font-weight: 900;
    margin-bottom: 10px;
}

/* =========================================================
   出席按钮
   ========================================================= */

.attendance-options{
    display:grid;
    gap:9px;
}

.attendance-options-3{
    grid-template-columns:repeat(3, 1fr);
}

.attendance-options-4{
    grid-template-columns:repeat(4, 1fr);
}

.homework-options {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 9px;
}

.choice-option {
    position: relative;
    display: flex;
    align-items: center;
    justify-content: center;
    min-height: 52px;
    border: 1px solid #d7dce2;
    border-radius: 14px;
    padding: 10px 8px;
    background: rgba(255, 255, 255, 0.82);
    cursor: pointer;
    font-size: 18px;
    font-weight: 700;
    text-align: center;
}

.choice-option input {
    position: absolute;
    opacity: 0;
    pointer-events: none;
}

.choice-option:has(input:checked) {
    border: 2px solid #2563eb;
    background: #eaf2ff;
    color: #174ea6;
    box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.08);
}

/* 不同状态选中颜色 */

.choice-present:has(input:checked) {
    border-color: #16a34a;
    background: #dcfce7;
    color: #166534;
}

.choice-late:has(input:checked) {
    border-color: #d97706;
    background: #fef3c7;
    color: #92400e;
}

.choice-farm:has(input:checked) {
    border-color: #65a30d;
    background: #ecfccb;
    color: #3f6212;
}

.choice-absent:has(input:checked) {
    border-color: #dc2626;
    background: #fee2e2;
    color: #991b1b;
}

.choice-submitted:has(input:checked) {
    border-color: #16a34a;
    background: #dcfce7;
    color: #166534;
}

.choice-missing:has(input:checked) {
    border-color: #dc2626;
    background: #fee2e2;
    color: #991b1b;
}

.choice-partial:has(input:checked) {
    border-color: #d97706;
    background: #fff7d6;
    color: #92400e;
}

.choice-homework-absent:has(input:checked) {
    border-color: #64748b;
    background: #e2e8f0;
    color: #334155;
}
                                  
.floating-toolbar{
    position:fixed;
    left:0;
    right:0;
    bottom:0;

    display:grid;
    grid-template-columns:repeat(3,1fr);
    gap:8px;

    padding:8px 10px;

    background:#fff;
    border-top:1px solid #ddd;
    box-shadow:0 -3px 10px rgba(0,0,0,.08);

    z-index:999;
}

.floating-toolbar .btn-tool{

    min-height:46px;

    font-size:18px;

    border-radius:14px;

    padding:8px 10px;

}

.page{

    padding-bottom:72px;

}

@media (max-width: 480px) {
    .floating-toolbar {
        gap: 7px;
        padding-left: 8px;
        padding-right: 8px;
    }

    .floating-toolbar .btn-tool {
        font-size: 15px;
        padding-left: 5px;
        padding-right: 5px;
    }
}
                                  
.teacher-grid{

    display:grid;

    grid-template-columns:1fr 1fr;

    gap:16px;

    margin-bottom:16px;

}

@media (max-width:700px){

    .teacher-grid{

        grid-template-columns:1fr;

    }

}

/* =========================================================
   手机
   ========================================================= */

@media (max-width: 800px){

    .attendance-options-3{
        grid-template-columns:repeat(3, 1fr);
    }

    .attendance-options-4{
        grid-template-columns:repeat(2, 1fr);
    }
}

@media (max-width: 480px) {

    .homework-options {
        grid-template-columns: 1fr 1fr;
    }

    .student-name {
        font-size: 21px;
    }

    .choice-option {
        font-size: 16px;
    }
}

</style>
</head>

<body>

<div class="page">

    <!-- =====================================================
         日期与组别
         ===================================================== -->

    <div class="card">

        <h1 class="page-title">
            📚 今日上课
        </h1>

        <p class="page-subtitle">
            课程、点名、白话佛法功课及经文功课。
        </p>

        {% with messages =
            get_flashed_messages(with_categories=true)
        %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form
            method="get"
            id="attendanceFilterForm"
        >

            <div class="form-group">
                <label class="form-label">
                    日期
                </label>

                <input
                    class="form-input"
                    type="date"
                    name="class_date"
                    value="{{ selected_date }}"
                    onchange="
                        document
                            .getElementById(
                                'attendanceFilterForm'
                            )
                            .submit()
                    "
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                    onchange="
                        document
                            .getElementById(
                                'attendanceFilterForm'
                            )
                            .submit()
                    "
                    required
                >
                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                selected_group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

        </form>

        <div class="current-selection">
            当前记录日期：{{ selected_date }}
        </div>

    </div>

    {% if students %}

    <form method="post" id="attendanceForm">

        <input
            type="hidden"
            name="class_date"
            value="{{ selected_date }}"
        >

        <input
            type="hidden"
            name="group_id"
            value="{{ selected_group_id }}"
        >

        <!-- =================================================
             课程资料
             ================================================= -->

        <div class="card">

            <h2 class="section-title">
                📖 教课记录
            </h2>

            <div class="lesson-note">
                请填写教课老师、记录老师及简单课程内容，下面即可开始点名。
            </div>

            <div class="teacher-grid">

                <div class="form-group">

                    <label class="form-label">
                        🎤 教课老师
                    </label>

                    <select
                        class="form-input"
                        name="teacher_name">

                        <option value="">
                            请选择教课老师
                        </option>

                        {% for teacher in teacher_list %}

                            <option
                                value="{{ teacher }}"
                                {% if lesson.teacher_name == teacher %}
                                    selected
                                {% endif %}
                            >
                                {{ teacher }}
                            </option>

                        {% endfor %}

                    </select>

                </div>

                <div class="form-group">

                    <label class="form-label">
                        📝 记录老师
                    </label>

                    <select
                        class="form-input"
                        name="record_teacher">

                        <option value="">
                            请选择记录老师
                        </option>

                        {% for teacher in teacher_list %}

                            <option
                                value="{{ teacher }}"
                                {% if lesson.record_teacher == teacher %}
                                    selected
                                {% endif %}
                            >
                                {{ teacher }}
                            </option>

                        {% endfor %}

                    </select>

                </div>

            </div>

            <div class="form-group">
                <label class="form-label">
                    📖 今日课题
                </label>

                <input
                    class="form-input"
                    name="topic"
                    value="{{ lesson.topic or '' }}"
                    placeholder="例如：学习慈悲"
                >
            </div>
            
        </div>

        <!-- =================================================
             学生点名与功课
             ================================================= -->

        <div class="card">

            <h2 class="section-title">
                📝 学生点名与功课
            </h2>

            <div class="student-list">

                {% for s in students %}

                    <div class="
                    student-card
                    student-card-{{ (loop.index0 % 3)+1 }}
                    ">

                        <input
                            type="hidden"
                            name="student_id"
                            value="{{ s.id }}"
                        >

                        <div class="student-header">

                            <div class="
                                student-number
                                {% if s.gender == '男' %}
                                    student-number-boy
                                {% elif s.gender == '女' %}
                                    student-number-girl
                                {% else %}
                                    student-number-unknown
                                {% endif %}
                            ">
                                {{ loop.index }}
                            </div>

                            <div class="student-name">

                                {{ s.name }}

                                {% if s.english_name %}
                                    <span class="student-english-name">
                                        {{ s.english_name }}
                                    </span>
                                {% endif %}

                            </div>

                        </div>

                        <!-- 出席 -->

                        <div class="student-section">

                            <div class="student-section-title">
                                📝 出席情况
                            </div>

                            <div class="
                                attendance-options
                                {% if selected_group_name == '少年组' %}
                                    attendance-options-4
                                {% else %}
                                    attendance-options-3
                                {% endif %}
                            ">

                                {% for key, label in status_labels.items() %}

                                    {% if key != "farm" or selected_group_name == "少年组" %}

                                        <label class="
                                            choice-option
                                            {% if key == 'present' %}
                                                choice-present
                                            {% elif key == 'late' %}
                                                choice-late
                                            {% elif key == 'farm' %}
                                                choice-farm
                                            {% elif key == 'absent' %}
                                                choice-absent
                                            {% endif %}
                                        ">

                                            <input
                                                type="radio"
                                                name="status_{{ s.id }}"
                                                value="{{ key }}"
                                                {% if s.attendance_status == key %}
                                                    checked
                                                {% endif %}
                                            >

                                            {% if key == "present" %}
                                                ✅
                                            {% elif key == "absent" %}
                                                ❌
                                            {% elif key == "late" %}
                                                ⏰
                                            {% elif key == "farm" %}
                                                🌱
                                            {% endif %}

                                            {{ label }}

                                        </label>

                                    {% endif %}

                                {% endfor %}

                            </div>

                        </div>

                        <!-- 白话佛法 -->

                        <div class="student-section">

                            <div class="student-section-title">
                                📚 白话佛法功课
                            </div>

                            <div class="homework-options">

                                <label class="choice-option choice-submitted">
                                    <input type="radio" name="baihua_{{ s.id }}"
                                           value="submitted_done"
                                           {% if s.baihua_status == "submitted_done" %}checked{% endif %}>
                                    ✓ 有交
                                </label>

                                <label class="choice-option choice-missing">
                                    <input
                                        type="radio"
                                        name="baihua_{{ s.id }}"
                                        value="missing"
                                        {% if s.baihua_status in ["missing", "absent"] %}
                                            checked
                                        {% endif %}
                                    >
                                    ✗ 没交
                                </label>

                                <label class="choice-option choice-partial">
                                    <input type="radio" name="baihua_{{ s.id }}"
                                           value="submitted_no_answer"
                                           {% if s.baihua_status == "submitted_no_answer" %}checked{% endif %}>
                                    ○ 有交，没做题
                                </label>
                                
                            </div>

                        </div>

                        <!-- 念经文 -->

                        <div class="student-section">

                            <div class="student-section-title">
                                📿 经文功课
                            </div>

                            <div class="homework-options">

                                <label class="choice-option choice-submitted">
                                    <input type="radio" name="scripture_{{ s.id }}"
                                           value="submitted_recited"
                                           {% if s.scripture_status == "submitted_recited" %}checked{% endif %}>
                                    ✓ 有交，有念
                                </label>

                                <label class="choice-option choice-partial">
                                    <input type="radio" name="scripture_{{ s.id }}"
                                           value="submitted_not_recited"
                                           {% if s.scripture_status == "submitted_not_recited" %}checked{% endif %}>
                                    ○ 有交，没念
                                </label>

                                <label class="choice-option choice-missing">
                                    <input
                                        type="radio"
                                        name="scripture_{{ s.id }}"
                                        value="missing"
                                        {% if s.scripture_status in ["missing", "absent"] %}
                                            checked
                                        {% endif %}
                                    >
                                    ✗ 没交
                                </label>
                                
                            </div>

                        </div>

                    </div>

                {% endfor %}

            </div>

            <div class="btn-row">
                
            </div>

        </div>

    </form>

    {% else %}

        <div class="card">

            <div class="empty-state">
                这个组别目前没有在读学生。
            </div>

        </div>

    {% endif %}

    <div class="card normal-bottom-actions">

        <div class="btn-row">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_home'
                ) }}"
            >
                ⬅ 返回佛学班首页
            </a>

        </div>

    </div>

</div>

{% if students %}

<div class="floating-toolbar">

    <a
        class="btn-tool btn-secondary"
        href="{{ url_for(
            'dharma_class.class_home'
        ) }}"
    >
        ⬅ 返回
    </a>

    <button
        class="btn-tool btn-primary"
        type="button"
        onclick="goTop()"
    >
        ⬆ 顶部
    </button>

    <button
        class="btn-tool btn-success"
        type="submit"
        form="attendanceForm"
    >
        💾 保存
    </button>

</div>

{% endif %}

<script>

function markStudentPresentIfNeeded(studentId) {
    const selected = document.querySelector(
        'input[name="status_' + studentId + '"]:checked'
    );

    // 已经是迟到或农舍时，不覆盖老师的真实记录。
    if (selected && (selected.value === "late" || selected.value === "farm")) {
        return;
    }

    const presentRadio = document.querySelector(
        'input[name="status_' + studentId + '"][value="present"]'
    );

    if (presentRadio) {
        presentRadio.checked = true;
    }
}

function markStudentHomeworkAbsent(studentId) {
    const baihuaAbsent = document.querySelector(
        'input[name="baihua_' + studentId + '"][value="absent"]'
    );

    const scriptureAbsent = document.querySelector(
        'input[name="scripture_' + studentId + '"][value="absent"]'
    );

    if (baihuaAbsent) baihuaAbsent.checked = true;
    if (scriptureAbsent) scriptureAbsent.checked = true;
}

function markStudentHomeworkAbsent(studentId) {

    // 页面已经取消功课“缺席”按钮，
    // 所以学生缺席时，画面自动选中“没交”。
    const baihuaMissing = document.querySelector(
        'input[name="baihua_' + studentId + '"][value="missing"]'
    );

    const scriptureMissing = document.querySelector(
        'input[name="scripture_' + studentId + '"][value="missing"]'
    );

    if (baihuaMissing) {
        baihuaMissing.checked = true;
        baihuaMissing.dispatchEvent(
            new Event("change", { bubbles: true })
        );
    }

    if (scriptureMissing) {
        scriptureMissing.checked = true;
        scriptureMissing.dispatchEvent(
            new Event("change", { bubbles: true })
        );
    }
}

function goTop(){

    window.scrollTo({

        top:0,

        behavior:"smooth"

    });

}


/* =========================================================
   初始化智能联动
   ========================================================= */

document.addEventListener(
    "DOMContentLoaded",
    function() {

        // 功课选择任何非缺席状态时，若学生不是迟到或农舍，自动设为出席。
        document
            .querySelectorAll(
                'input[name^="baihua_"], input[name^="scripture_"]'
            )
            .forEach(function(radio) {
                radio.addEventListener("change", function() {
                    if (!this.checked || this.value === "absent") return;

                    const studentId = this.name
                        .replace("baihua_", "")
                        .replace("scripture_", "");

                    markStudentPresentIfNeeded(studentId);
                });
            });

        // 点缺席时，两项功课自动改成缺席。
        document
            .querySelectorAll('input[name^="status_"]')
            .forEach(function(radio) {
                radio.addEventListener("change", function() {
                    if (!this.checked) return;

                    const studentId = this.name.replace("status_", "");

                    if (this.value === "absent") {
                        markStudentHomeworkAbsent(studentId);
                    } else {
                        const baihuaSelected = document.querySelector(
                            'input[name="baihua_' + studentId + '"]:checked'
                        );
                        const scriptureSelected = document.querySelector(
                            'input[name="scripture_' + studentId + '"]:checked'
                        );

                        if (
                            !baihuaSelected
                            || baihuaSelected.value === "absent"
                            || !scriptureSelected
                            || scriptureSelected.value === "absent"
                        ) {
                            markStudentHomeworkDefault(studentId);
                        }
                    }
                });
            });
    }
);

</script>

</body>
</html>
""",
        groups=groups,
        students=students,
        selected_date=selected_date,
        selected_group_id=selected_group_id,
        selected_group_name=selected_group_name,
        status_labels=STATUS_LABELS,
        baihua_status_labels=BAIHUA_STATUS_LABELS,
        scripture_status_labels=SCRIPTURE_STATUS_LABELS,
        lesson=lesson,
        teacher_list=teacher_list
    )

@dharma_class_bp.route("/lessons")
def class_lessons():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    selected_year = request.args.get(
        "year",
        str(current_year)
    ).strip()

    selected_month = request.args.get(
        "month",
        ""
    ).strip()

    selected_group_id = request.args.get(
        "group_id",
        ""
    ).strip()

    q = request.args.get(
        "q",
        ""
    ).strip()

    teacher = request.args.get(
        "teacher",
        ""
    ).strip()

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取组别
            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            # 读取课程记录中已有的年份
            cur.execute("""
                select distinct
                    extract(year from lesson_date)::int as lesson_year
                from dharma_class_lessons
                where branch = 'CHE'
                order by lesson_year desc
            """)
            year_rows = cur.fetchall()

            available_years = [
                row["lesson_year"]
                for row in year_rows
                if row["lesson_year"]
            ]

            if current_year not in available_years:
                available_years.insert(0, current_year)

            where_sql = """
                where l.branch = 'CHE'
            """

            params = []

            # 年份筛选
            if selected_year:
                where_sql += """
                    and extract(year from l.lesson_date) = %s
                """
                params.append(int(selected_year))

            # 月份筛选
            if selected_month:
                where_sql += """
                    and extract(month from l.lesson_date) = %s
                """
                params.append(int(selected_month))

            # 组别筛选
            if selected_group_id:
                where_sql += """
                    and l.group_id = %s
                """
                params.append(selected_group_id)

            # 老师筛选
            if teacher:
                teacher_like = f"%{teacher}%"

                where_sql += """
                    and l.teacher_name ilike %s
                """
                params.append(teacher_like)

            # 课题、教材、内容搜索
            if q:
                search_like = f"%{q}%"

                where_sql += """
                    and (
                        l.topic ilike %s
                        or l.material ilike %s
                        or l.content ilike %s
                        or l.remark ilike %s
                    )
                """

                params.extend([
                    search_like,
                    search_like,
                    search_like,
                    search_like
                ])

            cur.execute(f"""
                select
                    l.id,
                    l.lesson_date,
                    l.group_id,
                    l.teacher_name,
                    l.topic,
                    l.material,
                    l.content,
                    l.remark,
                    l.created_at,
                    g.name as group_name,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                    ) as attendance_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status = 'present'
                    ) as present_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status = 'late'
                    ) as late_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status = 'farm'
                    ) as farm_count,

                    (
                        select count(*)
                        from dharma_attendance a
                        where a.branch = l.branch
                          and a.class_date = l.lesson_date
                          and a.group_id = l.group_id
                          and a.status in ('absent', 'leave')
                    ) as absent_count

                from dharma_class_lessons l

                left join dharma_class_groups g
                    on g.id = l.group_id

                {where_sql}

                order by
                    l.lesson_date desc,
                    g.sort_order,
                    l.id desc
            """, params)

            lessons = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>课程记录</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>
.lesson-card {
    border: 1px solid #e5e7eb;
    border-radius: 18px;
    padding: 20px;
    margin-bottom: 18px;
    background: #fff;
}

.lesson-header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 14px;
    margin-bottom: 14px;
}

.lesson-date {
    font-size: 22px;
    font-weight: 800;
}

.lesson-group {
    background: #eef2ff;
    border-radius: 999px;
    padding: 7px 14px;
    font-size: 16px;
    font-weight: 700;
    white-space: nowrap;
}

.lesson-topic {
    font-size: 25px;
    font-weight: 800;
    margin: 10px 0;
}

.lesson-meta {
    line-height: 1.8;
    font-size: 18px;
    margin-bottom: 12px;
}

.lesson-content {
    background: #f6f7f9;
    border-radius: 14px;
    padding: 14px 16px;
    line-height: 1.7;
    margin-top: 12px;
    white-space: pre-wrap;
}

.attendance-summary {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 10px;
    margin-top: 16px;
}

.attendance-box {
    background: #f8f9fa;
    border-radius: 14px;
    padding: 12px;
    text-align: center;
}

.attendance-number {
    font-size: 25px;
    font-weight: 800;
    margin-top: 4px;
}

.lesson-actions {
    display: flex;
    gap: 10px;
    margin-top: 16px;
}

.lesson-actions .btn-tool {
    flex: 1;
    min-height: 48px;
    font-size: 17px;
    padding: 10px 14px;
}

.filter-summary {
    background: #f6f7f9;
    border-radius: 14px;
    padding: 14px 16px;
    margin-top: 16px;
    text-align: center;
    font-size: 18px;
    font-weight: 700;
}

@media (max-width: 650px) {
    .lesson-header {
        display: block;
    }

    .lesson-group {
        display: inline-block;
        margin-top: 8px;
    }

    .attendance-summary {
        grid-template-columns: repeat(2, 1fr);
    }

    .lesson-actions {
        display: block;
    }

    .lesson-actions .btn-tool {
        width: 100%;
        margin-bottom: 10px;
    }
}
</style>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            📖 课程记录
        </h1>

        <p class="page-subtitle">
            查询各组历次课程、负责老师及学生出席情况。
        </p>

        <form method="get">

            <div class="form-group">
                <label class="form-label">
                    年份
                </label>

                <select
                    class="form-input"
                    name="year"
                >
                    <option value="">
                        全部年份
                    </option>

                    {% for y in available_years %}
                        <option
                            value="{{ y }}"
                            {% if selected_year|string == y|string %}
                                selected
                            {% endif %}
                        >
                            {{ y }}年
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    月份
                </label>

                <select
                    class="form-input"
                    name="month"
                >
                    <option value="">
                        全部月份
                    </option>

                    {% for m in range(1, 13) %}
                        <option
                            value="{{ m }}"
                            {% if selected_month|string == m|string %}
                                selected
                            {% endif %}
                        >
                            {{ m }}月
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                >
                    <option value="">
                        全部组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                selected_group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    负责老师
                </label>

                <input
                    class="form-input"
                    name="teacher"
                    value="{{ teacher or '' }}"
                    placeholder="例如：陈老师"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    搜索课程
                </label>

                <input
                    class="form-input"
                    name="q"
                    value="{{ q or '' }}"
                    placeholder="课题／教材／教学内容／备注"
                >
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    🔍 搜索课程
                </button>

                <a
                    class="btn-tool btn-secondary"
                    href="{{ url_for(
                        'dharma_class.class_lessons'
                    ) }}"
                >
                    ↺ 清除筛选
                </a>
            </div>

        </form>

        <div class="filter-summary">
            找到 {{ lessons|length }} 笔课程记录
        </div>

    </div>

    <div class="card">

        <h2 class="section-title">
            📚 课程列表
        </h2>

        {% if lessons %}

            {% for lesson in lessons %}

                <div class="lesson-card">

                    <div class="lesson-header">

                        <div class="lesson-date">
                            📅 {{ lesson.lesson_date.strftime("%Y年%m月%d日") }}
                        </div>

                        <div class="lesson-group">
                            {{ lesson.group_name or "未设组别" }}
                        </div>

                    </div>

                    <div class="lesson-topic">
                        📖 {{ lesson.topic }}
                    </div>

                    <div class="lesson-meta">

                        <div>
                            👨‍🏫 负责老师：
                            {{ lesson.teacher_name or "—" }}
                        </div>

                        <div>
                            📚 教材／资源：
                            {{ lesson.material or "—" }}
                        </div>

                    </div>

                    {% if lesson.content %}
                        <div class="lesson-content">
                            <strong>教学内容</strong><br>
                            {{ lesson.content }}
                        </div>
                    {% endif %}

                    {% if lesson.remark %}
                        <div class="lesson-content">
                            <strong>课程备注</strong><br>
                            {{ lesson.remark }}
                        </div>
                    {% endif %}

                    <div class="attendance-summary">

                        <div class="attendance-box">
                            <div>点名人数</div>
                            <div class="attendance-number">
                                {{ lesson.attendance_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>出席</div>
                            <div class="attendance-number">
                                {{ lesson.present_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>迟到</div>
                            <div class="attendance-number">
                                {{ lesson.late_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>农舍</div>
                            <div class="attendance-number">
                                {{ lesson.farm_count or 0 }}
                            </div>
                        </div>

                        <div class="attendance-box">
                            <div>缺席</div>
                            <div class="attendance-number">
                                {{ lesson.absent_count or 0 }}
                            </div>
                        </div>

                    </div>

                    <div class="lesson-actions">

                        <a
                            class="btn-tool btn-primary"
                            href="{{ url_for(
                                'dharma_class.class_attendance',
                                class_date=lesson.lesson_date.isoformat(),
                                group_id=lesson.group_id
                            ) }}"
                        >
                            👁 查看课程与点名
                        </a>

                        <a
                            class="btn-tool btn-warning"
                            href="{{ url_for(
                                'dharma_class.class_attendance',
                                class_date=lesson.lesson_date.isoformat(),
                                group_id=lesson.group_id
                            ) }}"
                        >
                            ✏ 编辑记录
                        </a>

                    </div>

                </div>

            {% endfor %}

        {% else %}

            <div class="empty-state">
                暂时没有符合条件的课程记录。
            </div>

        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('dharma_class.class_admin') }}"
            >
                ⬅ 返回负责人中心
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        groups=groups,
        lessons=lessons,
        available_years=available_years,
        selected_year=selected_year,
        selected_month=selected_month,
        selected_group_id=selected_group_id,
        teacher=teacher,
        q=q
    )

@dharma_class_bp.route("/students")
def class_students():
   
    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    group_id = request.args.get(
        "group_id",
        ""
    ).strip()

    q = request.args.get(
        "q",
        ""
    ).strip()

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取组别
            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)

            groups = cur.fetchall()

            where_sql = """
                where s.status = 'active'
            """

            params = []

            # 组别筛选
            if group_id:
                where_sql += """
                    and s.group_id = %s
                """
                params.append(group_id)

            # 搜索
            if q:
                like = f"%{q}%"

                where_sql += """
                    and (
                        s.name ilike %s
                        or s.english_name ilike %s
                        or s.parent_name ilike %s
                        or s.parent_phone ilike %s
                        )
                """

                params.extend([
                    like,
                    like,
                    like,
                    like,
                    like
                ])

            cur.execute(f"""
                select
                    s.id,
                    s.name,
                    s.english_name,
                    s.gender,
                    s.birth_year,
                    s.parent_name,
                    s.parent_phone,
                    s.group_id,
                    s.status,
                    s.remark,
                    g.name as group_name,
                    g.sort_order as group_sort_order
                from dharma_students s
                left join dharma_class_groups g
                    on g.id = s.group_id
                {where_sql}
                order by
                    g.sort_order,
                    s.name
            """, params)

            students = cur.fetchall()

            def student_list_sort_key(student):

                group_name = student["group_name"] or ""
                birth_year = student["birth_year"]
                gender = student["gender"] or ""

                # 没有出生年份的排最后
                no_birth_year = 1 if not birth_year else 0

                # 年龄小到大：
                # 出生年份越大，年龄越小
                birth_order = -birth_year if birth_year else 0

                if group_name == "少年组":

                    if gender == "女":
                        gender_order = 0
                    elif gender == "男":
                        gender_order = 1
                    else:
                        gender_order = 2

                else:
                    gender_order = 0

                return (
                    student["group_sort_order"]
                    if student["group_sort_order"] is not None
                    else 999,

                    gender_order,
                    no_birth_year,
                    birth_order,
                    student["name"] or "",
                )


            students.sort(key=student_list_sort_key)

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>{{ current_year }}年佛学班学生名单</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>
.student-summary {
    margin-top: 12px;
    padding: 14px 16px;
    border-radius: 14px;
    background: #f6f7f9;
    font-size: 18px;
    font-weight: 700;
    text-align: center;
}

.student-name {
    font-weight: 800;
    white-space: nowrap;
}

.student-age {
    white-space: nowrap;
}

.student-contact {
    min-width: 160px;
}

.student-remark {
    min-width: 150px;
}

.student-table th,
.student-table td {
    vertical-align: middle;
}

.small-action-btn {
    font-size: 16px;
    min-height: 40px;
    padding: 8px 12px;
    white-space: nowrap;
}
                                  
.bottom-bar{

    position:fixed;
    left:0;
    right:0;
    bottom:0;

    display:flex;
    justify-content:center;
    gap:18px;

    padding:10px 14px;

    background:#fff;

    border-top:1px solid #ddd;
    box-shadow:0 -4px 12px rgba(0,0,0,.08);

    z-index:9999;

}

                                  
body{
    padding-bottom:90px;
}
                                  
@media (max-width: 700px) {
    .student-table {
        min-width: 1200px;
    }
}
</style>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            👧 {{ current_year }}年佛学班学生名单
        </h1>

        <p class="page-subtitle">
            管理佛学班学生资料、组别及监护人联系方式。
        </p>

        {% with messages =
            get_flashed_messages(
                with_categories=true
            )
        %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="get">

            <div class="form-group">
                <label class="form-label">
                    搜索学生
                </label>

                <input
                    class="form-input"
                    name="q"
                    value="{{ q or '' }}"
                    placeholder="姓名／英文名／监护人／电话"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别筛选
                </label>

                <select
                    class="form-input"
                    name="group_id"
                >
                    <option value="">
                        全部组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="btn-row">

                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    🔍 搜索
                </button>

                {% if q or group_id %}
                    <a
                        class="btn-tool btn-secondary"
                        href="{{ url_for(
                            'dharma_class.class_students'
                        ) }}"
                    >
                        ↺ 清除筛选
                    </a>
                {% endif %}

            </div>

        </form>

        <div class="student-summary">
            当前显示：{{ students|length }} 位学生
        </div>

        <div class="btn-row">
            <a
                class="btn-tool btn-success"
                href="{{ url_for(
                    'dharma_class.class_students_add'
                ) }}"
            >
                ➕ 新增学生
            </a>
        </div>

        <div class="btn-row">

            <a
                class="btn-tool btn-purple"
                href="/class/students/import"
            >
                📥 导入学生 Excel
            </a>

            <a
                class="btn-tool btn-secondary"
                href="/class/students/template"
            >
                📄 下载模板
            </a>

        </div>

        <div class="btn-row">
            <a
                class="btn-tool btn-warning"
                href="/class/students/export"
            >
                📤 导出学生名单
            </a>
        </div>

    </div>

    <div class="card">

        <h2 class="section-title">
            学生名单
        </h2>

        {% if students %}

            <div class="table-responsive">

                <table class="record-table student-table">

                    <thead>
                        <tr>
                            <th>姓名</th>
                            <th>英文名</th>
                            <th>性别</th>
                            <th>出生年份</th>
                            <th>年龄</th>
                            <th>组别</th>
                            <th>父／母／监护人</th>
                            <th>联系电话</th>
                            <th>备注</th>
                            <th>操作</th>
                        </tr>
                    </thead>

                    <tbody>

                        {% for s in students %}
                        <tr>
                            
                            <td class="student-name">
                                {{ s.name }}
                            </td>

                            <td>
                                {{ s.english_name or "—" }}
                            </td>

                            <td>
                                {{ s.gender or "—" }}
                            </td>

                            <td>
                                {{ s.birth_year or "—" }}
                            </td>

                            <td class="student-age">
                                {% if s.birth_year %}
                                    {{ current_year - s.birth_year }}岁
                                {% else %}
                                    —
                                {% endif %}
                            </td>

                            <td>
                                {{ s.group_name or "—" }}
                            </td>

                            <td class="student-contact">
                                {{ s.parent_name or "—" }}
                            </td>

                            <td class="student-contact">
                                {{ s.parent_phone or "—" }}
                            </td>

                            <td class="student-remark">
                                {{ s.remark or "—" }}
                            </td>

                            <td>
                                <a
                                    class="btn-tool btn-primary small-action-btn"
                                    href="{{ url_for(
                                        'dharma_class.class_student_profile',
                                        student_id=s.id
                                    ) }}"
                                >
                                    📊 档案
                                </a>

                                <a
                                    class="btn-tool btn-warning small-action-btn"
                                    href="{{ url_for(
                                        'dharma_class.class_students_edit',
                                        student_id=s.id
                                    ) }}"
                                >
                                    ✏ 编辑
                                </a>
                            </td>

                        </tr>
                        {% endfor %}

                    </tbody>

                </table>

                    </div>

                {% else %}

                    <div class="empty-state">
                        没有找到符合条件的学生资料。
                    </div>

                {% endif %}

            </div>

            <div class="bottom-bar">

                <a
                    class="btn-tool btn-secondary"
                    href="{{ url_for('dharma_class.class_admin') }}"
                >
                    ← 返回
                </a>

                

                <button
                    type="button"
                    class="btn-tool btn-primary"
                    onclick="goTop()"
                >
                    ↑ 顶部
                </button>

            </div>

        </div>

</div>
<script>
function goTop() {
    window.scrollTo({
        top: 0,
        left: 0,
        behavior: "smooth"
    });
}
</script>
</body>
</html>
""",
        groups=groups,
        students=students,
        group_id=group_id,
        q=q,
        current_year=current_year
    )


@dharma_class_bp.route("/students/search")
def class_student_search():

    q = request.args.get("q", "").strip()
    students = []

    if q:
        like = f"%{q}%"

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    select
                        s.id,
                        s.name,
                        s.english_name,
                        s.parent_name,
                        s.parent_phone,
                        s.status,
                        g.name as group_name
                    from dharma_students s
                    left join dharma_class_groups g
                        on g.id = s.group_id
                    where s.branch = 'CHE'
                      and (
                            s.name ilike %s
                         or coalesce(s.english_name, '') ilike %s
                         or coalesce(s.parent_name, '') ilike %s
                         or coalesce(s.parent_phone, '') ilike %s
                      )
                    order by
                        case when s.status = 'active' then 0 else 1 end,
                        g.sort_order,
                        s.name
                    limit 50
                """, (
                    like,
                    like,
                    like,
                    like
                ))

                students = cur.fetchall()

    if q and len(students) == 1:
        return redirect(
            url_for(
                "dharma_class.class_student_profile",
                student_id=students[0]["id"]
            )
        )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>搜索学生个人档案</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
<style>
.search-result-grid{
    display:grid;
    gap:14px;
    margin-top:18px;
}
.search-result-card{
    border:1px solid #e5e7eb;
    border-radius:17px;
    padding:17px;
    background:#fffdf7;
}
.search-result-name{
    font-size:23px;
    font-weight:900;
}
.search-result-meta{
    margin-top:7px;
    color:#666;
    line-height:1.7;
}
</style>
</head>
<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">🔎 搜索学生个人档案</h1>
        <p class="page-subtitle">
            可使用中文姓名、英文名、或联系电话搜索。
        </p>

        <form method="get">
            <div class="form-group">
                <label class="form-label">学生资料</label>
                <input
                    class="form-input"
                    name="q"
                    value="{{ q }}"
                    placeholder="例如：陈小明"
                    autofocus
                    required
                >
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-primary" type="submit">
                    🔎 搜索
                </button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2 class="section-title">搜索结果</h2>

        {% if q and students %}
            <div class="search-result-grid">
                {% for s in students %}
                    <div class="search-result-card">
                        <div class="search-result-name">
                            {{ s.name }}
                            {% if s.english_name %}
                                <span style="font-size:17px;color:#777;font-weight:500;">
                                    {{ s.english_name }}
                                </span>
                            {% endif %}
                        </div>

                        <div class="search-result-meta">
                            组别：{{ s.group_name or '—' }}<br>
                            父／母／监护人：{{ s.parent_name or '—' }}<br>
                            电话：{{ s.parent_phone or '—' }}<br>
                            状态：{{ s.status or '—' }}
                        </div>

                        <div class="btn-row">
                            <a
                                class="btn-tool btn-success"
                                href="{{ url_for(
                                    'dharma_class.class_student_profile',
                                    student_id=s.id
                                ) }}"
                            >
                                📊 查看个人全年档案
                            </a>
                        </div>
                    </div>
                {% endfor %}
            </div>

        {% elif q %}
            <div class="empty-state">
                找不到符合“{{ q }}”的学生。
            </div>

        {% else %}
            <div class="empty-state">
                请输入学生姓名或其他资料开始搜索。
            </div>
        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('dharma_class.class_admin') }}"
            >
                ⬅ 返回负责人中心
            </a>
        </div>
    </div>

</div>
</body>
</html>
""",
        q=q,
        students=students
    )


@dharma_class_bp.route("/student/<int:student_id>")
def class_student_profile(student_id):

    malaysia_today = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).date()

    year_text = request.args.get(
        "year",
        str(malaysia_today.year)
    ).strip()

    try:
        selected_year = int(year_text)
    except (TypeError, ValueError):
        selected_year = malaysia_today.year

    start_date = date(selected_year, 1, 1)
    end_date = date(selected_year + 1, 1, 1)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    s.id,
                    s.branch,
                    s.name,
                    s.english_name,
                    s.gender,
                    s.birth_year,
                    s.parent_name,
                    s.parent_phone,
                    s.group_id,
                    s.status,
                    s.remark,
                    g.name as group_name
                from dharma_students s
                left join dharma_class_groups g
                    on g.id = s.group_id
                where s.id = %s
                  and s.branch = 'CHE'
                limit 1
            """, (student_id,))

            student = cur.fetchone()

            if not student:
                return "找不到学生资料。", 404

            cur.execute("""
                select
                    a.class_date,
                    a.status as attendance_status,
                    a.marked_by,
                    h.baihua_status,
                    h.scripture_status,
                    l.topic,
                    l.teacher_name
                from dharma_attendance a
                left join dharma_homework h
                    on h.branch = a.branch
                   and h.class_date = a.class_date
                   and h.student_id = a.student_id
                left join dharma_class_lessons l
                    on l.branch = a.branch
                   and l.lesson_date = a.class_date
                   and l.group_id = a.group_id
                where a.branch = 'CHE'
                  and a.student_id = %s
                  and a.class_date >= %s
                  and a.class_date < %s
                order by a.class_date desc
            """, (
                student_id,
                start_date,
                end_date
            ))

            records = cur.fetchall()

            cur.execute("""
                select distinct
                    extract(year from class_date)::int as record_year
                from dharma_attendance
                where branch = 'CHE'
                  and student_id = %s
                order by record_year desc
            """, (student_id,))

            available_years = [
                row["record_year"]
                for row in cur.fetchall()
                if row["record_year"]
            ]

    if malaysia_today.year not in available_years:
        available_years.insert(0, malaysia_today.year)

    attendance_labels = {
        "present": ("✓", "出席", "good"),
        "late": ("迟", "迟到", "partial"),
        "farm": ("农", "农舍", "farm"),
        "absent": ("✗", "缺席", "bad"),
        "leave": ("缺", "请假", "leave"),
    }

    baihua_labels = {
        "submitted_done": ("✓", "有交", "good"),
        "submitted": ("✓", "有交", "good"),
        "submitted_no_answer": ("○", "有交，没做题", "partial"),
        "missing": ("✗", "没交", "bad"),
        "absent": ("缺", "缺席", "leave"),
    }

    scripture_labels = {
        "submitted_recited": ("✓", "有交，有念", "good"),
        "submitted": ("✓", "有交，有念", "good"),
        "submitted_not_recited": ("○", "有交，没念", "partial"),
        "missing": ("✗", "没交", "bad"),
        "absent": ("缺", "缺席", "leave"),
    }

    attendance_count = {
        "present": 0,
        "late": 0,
        "farm": 0,
        "absent": 0,
    }

    baihua_count = {
        "full": 0,
        "partial": 0,
        "missing": 0,
        "absent": 0,
    }

    scripture_count = {
        "full": 0,
        "partial": 0,
        "missing": 0,
        "absent": 0,
    }

    for row in records:
        attendance_status = row["attendance_status"]
        if attendance_status in attendance_count:
            attendance_count[attendance_status] += 1

        baihua_status = row["baihua_status"]
        if baihua_status in ("submitted_done", "submitted"):
            baihua_count["full"] += 1
        elif baihua_status == "submitted_no_answer":
            baihua_count["partial"] += 1
        elif baihua_status == "missing":
            baihua_count["missing"] += 1
        elif baihua_status == "absent":
            baihua_count["absent"] += 1

        scripture_status = row["scripture_status"]
        if scripture_status in ("submitted_recited", "submitted"):
            scripture_count["full"] += 1
        elif scripture_status == "submitted_not_recited":
            scripture_count["partial"] += 1
        elif scripture_status == "missing":
            scripture_count["missing"] += 1
        elif scripture_status == "absent":
            scripture_count["absent"] += 1

        row["attendance_display"] = attendance_labels.get(
            attendance_status,
            ("—", attendance_status or "未记录", "empty")
        )
        row["baihua_display"] = baihua_labels.get(
            baihua_status,
            ("—", "未记录", "empty")
        )
        row["scripture_display"] = scripture_labels.get(
            scripture_status,
            ("—", "未记录", "empty")
        )

    if student["group_name"] == "少年组":

        attended = (
            attendance_count["present"]
            + attendance_count["late"]
            + attendance_count["farm"]
        )

    else:

        attended = (
            attendance_count["present"]
            + attendance_count["late"]
        )
    attendance_denominator = attended + attendance_count["absent"]
    attendance_rate = (
        round(attended * 100 / attendance_denominator, 1)
        if attendance_denominator > 0
        else None
    )

    baihua_denominator = (
        baihua_count["full"]
        + baihua_count["partial"]
        + baihua_count["missing"]
    )
    baihua_rate = (
        round(baihua_count["full"] * 100 / baihua_denominator, 1)
        if baihua_denominator > 0
        else None
    )

    scripture_denominator = (
        scripture_count["full"]
        + scripture_count["partial"]
        + scripture_count["missing"]
    )
    scripture_rate = (
        round(scripture_count["full"] * 100 / scripture_denominator, 1)
        if scripture_denominator > 0
        else None
    )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{ student.name }}个人全年档案</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
<style>
.profile-hero{
    background:linear-gradient(135deg,#fffaf0,#f4f8ff);
    border:1px solid #f0dfb4;
}
.profile-name{
    font-size:31px;
    font-weight:900;
}
.profile-meta{
    color:#666;
    margin-top:8px;
    line-height:1.8;
}
.profile-summary-grid{
    display:grid;
    grid-template-columns:repeat(5,minmax(0,1fr));
    gap:11px;
    margin-top:16px;
}

.profile-summary-grid-4{
    grid-template-columns:repeat(4,minmax(0,1fr));
}

.profile-summary-grid-5{
    grid-template-columns:repeat(5,minmax(0,1fr));
}
.profile-summary-box{
    border-radius:15px;
    padding:14px 8px;
    text-align:center;
    background:#f7f8fa;
}
.profile-summary-label{font-size:15px;color:#666;}
.profile-summary-number{font-size:27px;font-weight:900;margin-top:5px;}
.profile-rate{
    border-radius:16px;
    padding:16px;
    margin-top:15px;
    background:#fff8e8;
    border:1px solid #f3d98c;
    font-size:20px;
    font-weight:900;
    text-align:center;
}
.status-chip{
    display:inline-flex;
    align-items:center;
    justify-content:center;
    gap:5px;
    min-width:94px;
    border-radius:999px;
    padding:6px 10px;
    font-weight:800;
    white-space:nowrap;
}
.status-good{background:#dcfce7;color:#166534;}
.status-partial{background:#fef3c7;color:#92400e;}
.status-farm{background:#ecfccb;color:#3f6212;}
.status-bad{background:#fee2e2;color:#991b1b;}
.status-leave{background:#e0e7ff;color:#3730a3;}
.status-empty{background:#f3f4f6;color:#666;}
.profile-table{min-width:1000px;}
@media(max-width:850px){
    .profile-summary-grid{
        grid-template-columns:repeat(2, 1fr) !important;
    }
}

@media(max-width:480px){
    .profile-summary-grid{
        grid-template-columns:1fr !important;
    }
}
</style>
</head>
<body>
<div class="page">

    <div class="card profile-hero">
        <div class="profile-name">
            👤 {{ student.name }}
            {% if student.english_name %}
                <span style="font-size:18px;color:#777;font-weight:500;">
                    {{ student.english_name }}
                </span>
            {% endif %}
        </div>

        <div class="profile-meta">

            组别：{{ student.group_name or '—' }}
            ｜
            {{ student.gender or '—' }}
            ｜
            {% if student.birth_year %}
                {{ selected_year - student.birth_year }}岁
            {% else %}
                —
            {% endif %}

            <br>

            父／母／监护人：{{ student.parent_name or '—' }}
            ｜
            电话：{{ student.parent_phone or '—' }}
            ｜
            状态：{{ student.status or '—' }}

            <br>

            备注：{{ student.remark or '—' }}

        </div>

        <div class="btn-row">
            <a
                class="btn-tool btn-warning"
                href="{{ url_for(
                    'dharma_class.class_students_edit',
                    student_id=student.id
                ) }}"
            >
                ✏ 编辑学生基本资料
            </a>

            <a
                class="btn-tool btn-success"
                href="{{ url_for(
                    'dharma_class.class_student_profile_export',
                    student_id=student.id,
                    year=selected_year
                ) }}"
            >
                📥 下载个人全年报告
            </a>
        </div>
    </div>

    <div class="card">
        <form method="get">
            <div class="form-group">
                <label class="form-label">查看年份</label>
                <select class="form-input" name="year">
                    {% for y in available_years %}
                        <option value="{{ y }}" {% if y == selected_year %}selected{% endif %}>
                            {{ y }}年
                        </option>
                    {% endfor %}
                </select>
            </div>
            <div class="btn-row">
                <button class="btn-tool btn-primary" type="submit">🔍 查看全年记录</button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2 class="section-title">📝 {{ selected_year }}年出席总览</h2>
        <div class="
            profile-summary-grid
            {% if student.group_name == '少年组' %}
                profile-summary-grid-5
            {% else %}
                profile-summary-grid-4
            {% endif %}
        ">

            <div class="profile-summary-box">
                <div class="profile-summary-label">课程记录</div>
                <div class="profile-summary-number">{{ records|length }}</div>
            </div>

            <div class="profile-summary-box">
                <div class="profile-summary-label">出席</div>
                <div class="profile-summary-number">{{ attendance_count.present }}</div>
            </div>

            <div class="profile-summary-box">
                <div class="profile-summary-label">迟到</div>
                <div class="profile-summary-number">{{ attendance_count.late }}</div>
            </div>

            {% if student.group_name == "少年组" %}
                <div class="profile-summary-box">
                    <div class="profile-summary-label">农舍</div>
                    <div class="profile-summary-number">{{ attendance_count.farm }}</div>
                </div>
            {% endif %}

            <div class="profile-summary-box">
                <div class="profile-summary-label">缺席</div>
                <div class="profile-summary-number">{{ attendance_count.absent }}</div>
            </div>

        </div>
        <div class="profile-rate">全年出席率：{{ attendance_rate ~ '%' if attendance_rate is not none else '—' }}</div>
    </div>

    <div class="card">
        <h2 class="section-title">📚 白话佛法功课</h2>
        <div class="profile-summary-grid">
            <div class="profile-summary-box"><div class="profile-summary-label">✓ 有交</div><div class="profile-summary-number">{{ baihua_count.full }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">○ 没做题</div><div class="profile-summary-number">{{ baihua_count.partial }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">✗ 没交</div><div class="profile-summary-number">{{ baihua_count.missing }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">缺席</div><div class="profile-summary-number">{{ baihua_count.absent }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">完整完成率</div><div class="profile-summary-number">{{ baihua_rate ~ '%' if baihua_rate is not none else '—' }}</div></div>
        </div>
    </div>

    <div class="card">
        <h2 class="section-title">📿 经文功课</h2>
        <div class="profile-summary-grid">
            <div class="profile-summary-box"><div class="profile-summary-label">✓ 有交有念</div><div class="profile-summary-number">{{ scripture_count.full }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">○ 有交没念</div><div class="profile-summary-number">{{ scripture_count.partial }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">✗ 没交</div><div class="profile-summary-number">{{ scripture_count.missing }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">缺席</div><div class="profile-summary-number">{{ scripture_count.absent }}</div></div>
            <div class="profile-summary-box"><div class="profile-summary-label">完整完成率</div><div class="profile-summary-number">{{ scripture_rate ~ '%' if scripture_rate is not none else '—' }}</div></div>
        </div>
    </div>

    <div class="card">
        <h2 class="section-title">📅 全年逐堂记录</h2>

        {% if records %}
            <div class="table-responsive">
                <table class="record-table profile-table">
                    <thead>
                        <tr>
                            <th>日期</th>
                            <th>出席</th>
                            <th>白话佛法</th>
                            <th>念经文</th>
                            <th>课程课题</th>
                            <th>老师</th>
                            <th>操作</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for r in records %}
                            <tr>
                                <td>{{ r.class_date.strftime('%Y-%m-%d') }}</td>
                                <td><span class="status-chip status-{{ r.attendance_display[2] }}">{{ r.attendance_display[0] }} {{ r.attendance_display[1] }}</span></td>
                                <td><span class="status-chip status-{{ r.baihua_display[2] }}">{{ r.baihua_display[0] }} {{ r.baihua_display[1] }}</span></td>
                                <td><span class="status-chip status-{{ r.scripture_display[2] }}">{{ r.scripture_display[0] }} {{ r.scripture_display[1] }}</span></td>
                                <td>{{ r.topic or '—' }}</td>
                                <td>{{ r.teacher_name or r.marked_by or '—' }}</td>
                                <td>
                                    <a
                                        class="btn-tool btn-warning"
                                        style="font-size:15px;min-height:38px;padding:7px 10px;"
                                        href="{{ url_for(
                                            'dharma_class.class_attendance',
                                            class_date=r.class_date.isoformat(),
                                            group_id=student.group_id
                                        ) }}"
                                    >
                                        ✏ 编辑当天记录
                                    </a>
                                </td>
                            </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        {% else %}
            <div class="empty-state">{{ selected_year }}年还没有这位学生的记录。</div>
        {% endif %}

        <div class="btn-row">
            <a class="btn-tool btn-secondary" href="{{ url_for('dharma_class.class_student_search') }}">🔎 搜索其他学生</a>
            <a class="btn-tool btn-secondary" href="{{ url_for('dharma_class.class_students') }}">👧 返回学生名单</a>
        </div>
    </div>

</div>
</body>
</html>
""",
        student=student,
        selected_year=selected_year,
        available_years=available_years,
        records=records,
        attendance_count=attendance_count,
        attendance_rate=attendance_rate,
        baihua_count=baihua_count,
        baihua_rate=baihua_rate,
        scripture_count=scripture_count,
        scripture_rate=scripture_rate
    )


@dharma_class_bp.route("/student/<int:student_id>/export")
def class_student_profile_export(student_id):

    malaysia_today = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).date()

    try:
        selected_year = int(
            request.args.get("year") or malaysia_today.year
        )
    except (TypeError, ValueError):
        selected_year = malaysia_today.year

    start_date = date(selected_year, 1, 1)
    end_date = date(selected_year + 1, 1, 1)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select
                    s.id,
                    s.name,
                    s.english_name,
                    s.gender,
                    s.birth_year,
                    s.parent_name,
                    s.parent_phone,
                    s.group_id,
                    s.status,
                    s.remark,
                    g.name as group_name
                from dharma_students s
                left join dharma_class_groups g
                    on g.id = s.group_id
                where s.id = %s
                  and s.branch = 'CHE'
                limit 1
            """, (student_id,))

            student = cur.fetchone()

            if not student:
                return "找不到学生资料。", 404

            cur.execute("""
                select
                    a.class_date,
                    a.status as attendance_status,
                    h.baihua_status,
                    h.scripture_status,
                    l.topic,
                    l.teacher_name
                from dharma_attendance a
                left join dharma_homework h
                    on h.branch = a.branch
                   and h.class_date = a.class_date
                   and h.student_id = a.student_id
                left join dharma_class_lessons l
                    on l.branch = a.branch
                   and l.lesson_date = a.class_date
                   and l.group_id = a.group_id
                where a.branch = 'CHE'
                  and a.student_id = %s
                  and a.class_date >= %s
                  and a.class_date < %s
                order by a.class_date
            """, (
                student_id,
                start_date,
                end_date
            ))

            records = cur.fetchall()

    attendance_text = {
        "present": "✓ 出席",
        "late": "迟 迟到",
        "farm": "农 农舍",
        "absent": "✗ 缺席",
        "leave": "缺 请假",
    }
    baihua_text = {
        "submitted_done": "✓ 有交",
        "submitted": "✓ 有交",
        "submitted_no_answer": "○ 有交，没做题",
        "missing": "✗ 没交",
        "absent": "缺 缺席",
    }
    scripture_text = {
        "submitted_recited": "✓ 有交，有念",
        "submitted": "✓ 有交，有念",
        "submitted_not_recited": "○ 有交，没念",
        "missing": "✗ 没交",
        "absent": "缺 缺席",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "个人全年报告"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"蕉赖佛学班 {student['name']} {selected_year}年个人学习报告"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(size=18, color="FFFFFF", bold=True)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 34

    details = [
        ("中文姓名", student["name"] or "—"),
        ("英文名", student["english_name"] or "—"),
        ("组别", student["group_name"] or "—"),
        ("性别", student["gender"] or "—"),
        ("出生年份", student["birth_year"] or "—"),
        ("父／母／监护人", student["parent_name"] or "—"),
        ("联系电话", student["parent_phone"] or "—"),
        ("学生状态", student["status"] or "—"),
        ("备注", student["remark"] or "—"),
    ]

    row_no = 3
    for label, value in details:
        ws.cell(row=row_no, column=1, value=label)
        ws.cell(row=row_no, column=2, value=value)
        ws.cell(row=row_no, column=1).font = bold_font
        ws.cell(row=row_no, column=1).fill = section_fill
        ws.cell(row=row_no, column=1).border = border
        ws.cell(row=row_no, column=2).border = border
        row_no += 1

    row_no += 1
    headers = ["日期", "出席", "白话佛法", "念经文", "课程课题", "老师", "备注"]
    for col_no, header in enumerate(headers, 1):
        cell = ws.cell(row=row_no, column=col_no, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    for record in records:
        row_no += 1
        values = [
            record["class_date"],
            attendance_text.get(record["attendance_status"], "—"),
            baihua_text.get(record["baihua_status"], "—"),
            scripture_text.get(record["scripture_status"], "—"),
            record["topic"] or "—",
            record["teacher_name"] or "—",
            "",
        ]

        for col_no, value in enumerate(values, 1):
            cell = ws.cell(row=row_no, column=col_no, value=value)
            cell.border = border
            cell.alignment = center

        ws.cell(row=row_no, column=1).number_format = "yyyy-mm-dd"

    widths = {
        "A": 14,
        "B": 14,
        "C": 22,
        "D": 22,
        "E": 28,
        "F": 18,
        "G": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = f"A{row_no - len(records) + 1}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = (student["name"] or f"学生{student_id}").replace("/", "-")
    filename = f"{safe_name}_{selected_year}年个人学习报告.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@dharma_class_bp.route("/students/add", methods=["GET", "POST"])
def class_students_add():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    form_data = {
        "name": "",
        "english_name": "",
        "gender": "",
        "birth_year": "",
        "parent_name": "",
        "parent_phone": "",
        "group_id": "",
        "remark": ""
    }

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            if request.method == "POST":
                
                name = request.form.get(
                    "name", ""
                ).strip()

                english_name = request.form.get(
                    "english_name", ""
                ).strip()

                gender = request.form.get(
                    "gender", ""
                ).strip()

                birth_year_text = request.form.get(
                    "birth_year", ""
                ).strip()

                parent_name = request.form.get(
                    "parent_name", ""
                ).strip()

                parent_phone = request.form.get(
                    "parent_phone", ""
                ).strip()

                group_id_text = request.form.get(
                    "group_id", ""
                ).strip()

                remark = request.form.get(
                    "remark", ""
                ).strip()

                birth_year = (
                    int(birth_year_text)
                    if birth_year_text.isdigit()
                    else None
                )

                group_id = (
                    int(group_id_text)
                    if group_id_text.isdigit()
                    else None
                )

                form_data = {
                    "name": name,
                    "english_name": english_name,
                    "gender": gender,
                    "birth_year": birth_year_text,
                    "parent_name": parent_name,
                    "parent_phone": parent_phone,
                    "group_id": group_id_text,
                    "remark": remark
                }

                if not name:
                    flash("学生姓名必须填写。", "bad")

                elif gender not in ("男", "女"):
                    flash("请选择学生性别。", "bad")

                elif birth_year is None:
                    flash("请输入正确的出生年份。", "bad")

                elif birth_year < 2000 or birth_year > current_year:
                    flash(
                        f"出生年份必须介于 2000 至 {current_year} 年。",
                        "bad"
                    )

                elif group_id is None:
                    flash("请选择学生组别。", "bad")

                else:
                    cur.execute("""
                        insert into dharma_students
                        (
                            branch,
                            name,
                            english_name,
                            gender,
                            birth_year,
                            parent_name,
                            parent_phone,
                            group_id,
                            status,
                            remark
                        )
                        values
                        (
                            'CHE',
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            'active',
                            %s
                        )
                    """, (
                        name,
                        english_name or None,
                        gender,
                        birth_year,
                        parent_name or None,
                        parent_phone or None,
                        group_id,
                        remark or None
                    ))

                    conn.commit()

                    flash("学生已新增。", "good")

                    return redirect(
                        url_for(
                            "dharma_class.class_students"
                        )
                    )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">

<title>新增学生</title>

<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">➕ 新增学生</h1>

        <p class="page-subtitle">
            加入佛学班学生资料。
        </p>

        {% with messages = get_flashed_messages(
            with_categories=true
        ) %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="post">

            <div class="form-group">
                <label class="form-label">
                    中文姓名
                </label>

                <input
                    class="form-input"
                    name="name"
                    value="{{ form_data.name }}"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    英文名
                </label>

                <input
                    class="form-input"
                    name="english_name"
                    value="{{ form_data.english_name }}"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    性别
                </label>

                <select
                    name="gender"
                    class="form-input"
                    required
                >
                    <option value="">
                        请选择性别
                    </option>

                    <option
                        value="男"
                        {% if form_data.gender == "男" %}
                            selected
                        {% endif %}
                    >
                        男
                    </option>

                    <option
                        value="女"
                        {% if form_data.gender == "女" %}
                            selected
                        {% endif %}
                    >
                        女
                    </option>
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    出生年份
                </label>

                <input
                    type="number"
                    name="birth_year"
                    class="form-input"
                    value="{{ form_data.birth_year }}"
                    min="2000"
                    max="{{ current_year }}"
                    placeholder="例如：2016"
                    inputmode="numeric"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人姓名
                </label>

                <input
                    class="form-input"
                    name="parent_name"
                    value="{{ form_data.parent_name }}"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人电话
                </label>

                <input
                    class="form-input"
                    name="parent_phone"
                    value="{{ form_data.parent_phone }}"
                    inputmode="tel"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                    required
                >
                    <option value="">
                        请选择组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                form_data.group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    备注
                </label>

                <textarea
                    class="form-input"
                    name="remark"
                    rows="3"
                >{{ form_data.remark }}</textarea>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-success"
                    type="submit"
                >
                    ✅ 保存学生
                </button>
            </div>

        </form>

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_students'
                ) }}"
            >
                ⬅ 返回学生名单
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        groups=groups,
        form_data=form_data,
        current_year=current_year
    )

@dharma_class_bp.route(
    "/students/edit/<int:student_id>",
    methods=["GET", "POST"]
)
def class_students_edit(student_id):
    
    current_year = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).year

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取可用组别
            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            # 读取学生资料
            cur.execute("""
                select *
                from dharma_students
                where id = %s
            """, (student_id,))

            student = cur.fetchone()
            original_student = dict(student) if student else None

            if not student:
                flash("找不到这个学生。", "bad")
                return redirect(
                    url_for(
                        "dharma_class.class_students"
                    )
                )

            if request.method == "POST":

                action = request.form.get(
                    "action",
                    "save"
                ).strip()

                # 暂停学生
                if action == "inactive":

                    cur.execute("""
                        update dharma_students
                        set status = 'inactive'
                        where id = %s
                    """, (student_id,))

                    write_dharma_audit_log(
                        cur,
                        actor_name="管理员",
                        action_type="update",
                        entity_type="student",
                        entity_id=student_id,
                        student_id=student_id,
                        student_name=original_student.get("name"),
                        group_id=original_student.get("group_id"),
                        field_name="student_status",
                        old_value=original_student.get("status"),
                        new_value="inactive",
                    )

                    conn.commit()

                    flash("学生已暂停。", "good")

                    return redirect(
                        url_for(
                            "dharma_class.class_students"
                        )
                    )

                # 读取表单资料
                
                name = request.form.get(
                    "name",
                    ""
                ).strip()

                english_name = request.form.get(
                    "english_name",
                    ""
                ).strip()

                gender = request.form.get(
                    "gender",
                    ""
                ).strip()

                birth_year_text = request.form.get(
                    "birth_year",
                    ""
                ).strip()

                parent_name = request.form.get(
                    "parent_name",
                    ""
                ).strip()

                parent_phone = request.form.get(
                    "parent_phone",
                    ""
                ).strip()

                group_id_text = request.form.get(
                    "group_id",
                    ""
                ).strip()

                remark = request.form.get(
                    "remark",
                    ""
                ).strip()

                birth_year = (
                    int(birth_year_text)
                    if birth_year_text.isdigit()
                    else None
                )

                group_id = (
                    int(group_id_text)
                    if group_id_text.isdigit()
                    else None
                )

                # 将刚提交的资料放回 student
                # 验证失败时，页面会保留用户刚输入的内容
                student["name"] = name
                student["english_name"] = english_name
                student["gender"] = gender
                student["birth_year"] = birth_year_text
                student["parent_name"] = parent_name
                student["parent_phone"] = parent_phone
                student["group_id"] = group_id_text
                student["remark"] = remark

                # 验证资料
                if not name:
                    flash(
                        "学生姓名必须填写。",
                        "bad"
                    )

                elif gender not in ("男", "女"):
                    flash(
                        "请选择学生性别。",
                        "bad"
                    )

                elif birth_year is None:
                    flash(
                        "请输入正确的出生年份。",
                        "bad"
                    )

                elif (
                    birth_year < 2000
                    or birth_year > current_year
                ):
                    flash(
                        f"出生年份必须介于 "
                        f"2000 至 {current_year} 年。",
                        "bad"
                    )

                elif group_id is None:
                    flash(
                        "请选择学生组别。",
                        "bad"
                    )

                else:
                    cur.execute("""
                        update dharma_students
                        set
                            name = %s,
                            english_name = %s,
                            gender = %s,
                            birth_year = %s,
                            parent_name = %s,
                            parent_phone = %s,
                            group_id = %s,
                            remark = %s
                        where id = %s
                    """, (
                        name,
                        english_name or None,
                        gender,
                        birth_year,
                        parent_name or None,
                        parent_phone or None,
                        group_id,
                        remark or None,
                        student_id
                    ))

                    student_changes = {
                        "name": name,
                        "english_name": english_name or None,
                        "gender": gender,
                        "birth_year": birth_year,
                        "parent_name": parent_name or None,
                        "parent_phone": parent_phone or None,
                        "group_id": group_id,
                        "remark": remark or None,
                    }

                    for field_name, new_value in student_changes.items():
                        write_dharma_audit_log(
                            cur,
                            actor_name="管理员",
                            action_type="update",
                            entity_type="student",
                            entity_id=student_id,
                            student_id=student_id,
                            student_name=name,
                            group_id=group_id,
                            field_name=field_name,
                            old_value=original_student.get(field_name),
                            new_value=new_value,
                        )

                    conn.commit()

                    flash(
                        "学生资料已更新。",
                        "good"
                    )

                    return redirect(
                        url_for(
                            "dharma_class.class_students"
                        )
                    )

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>编辑学生</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            ✏ 编辑学生
        </h1>

        <p class="page-subtitle">
            修改学生资料或暂停学生。
        </p>

        {% with messages =
            get_flashed_messages(
                with_categories=true
            )
        %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert alert-{{ category }}">
                        {{ msg }}
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="post">
            
            <div class="form-group">
                <label class="form-label">
                    中文姓名
                </label>

                <input
                    class="form-input"
                    name="name"
                    value="{{ student.name or '' }}"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    英文名
                </label>

                <input
                    class="form-input"
                    name="english_name"
                    value="{{ student.english_name or '' }}"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    性别
                </label>

                <select
                    name="gender"
                    class="form-input"
                    required
                >
                    <option value="">
                        请选择性别
                    </option>

                    <option
                        value="男"
                        {% if student.gender == "男" %}
                            selected
                        {% endif %}
                    >
                        男
                    </option>

                    <option
                        value="女"
                        {% if student.gender == "女" %}
                            selected
                        {% endif %}
                    >
                        女
                    </option>
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    出生年份
                </label>

                <input
                    type="number"
                    name="birth_year"
                    class="form-input"
                    value="{{ student.birth_year or '' }}"
                    min="2000"
                    max="{{ current_year }}"
                    placeholder="例如：2016"
                    inputmode="numeric"
                    required
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人姓名
                </label>

                <input
                    class="form-input"
                    name="parent_name"
                    value="{{ student.parent_name or '' }}"
                    placeholder="父亲、母亲、亲戚或监护人姓名"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    父／母／监护人电话
                </label>

                <input
                    class="form-input"
                    name="parent_phone"
                    value="{{ student.parent_phone or '' }}"
                    placeholder="例如：0123456789"
                    inputmode="tel"
                >
            </div>

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                    required
                >
                    <option value="">
                        请选择组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                student.group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="form-group">
                <label class="form-label">
                    备注
                </label>

                <textarea
                    class="form-input"
                    name="remark"
                    rows="3"
                >{{ student.remark or '' }}</textarea>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-success"
                    type="submit"
                    name="action"
                    value="save"
                >
                    ✅ 保存修改
                </button>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-danger"
                    type="submit"
                    name="action"
                    value="inactive"
                    onclick="
                        return confirm(
                            '确定要暂停这个学生吗？暂停后点名不会再显示。'
                        );
                    "
                >
                    ⏸ 暂停学生
                </button>
            </div>

        </form>

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'dharma_class.class_students'
                ) }}"
            >
                ⬅ 返回学生名单
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        student=student,
        groups=groups,
        current_year=current_year
    )

@dharma_class_bp.route("/audit")
def class_audit_log():

    q = request.args.get("q", "").strip()
    actor = request.args.get("actor", "").strip()
    entity_type = request.args.get("entity_type", "").strip()
    field_name = request.args.get("field_name", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()

    where_sql = "where a.branch = 'CHE'"
    params = []

    if q:
        like = f"%{q}%"
        where_sql += """
            and (
                a.student_name ilike %s
                or a.actor_name ilike %s
                or a.old_value ilike %s
                or a.new_value ilike %s
            )
        """
        params.extend([like, like, like, like])

    if actor:
        where_sql += " and a.actor_name = %s"
        params.append(actor)

    if entity_type:
        where_sql += " and a.entity_type = %s"
        params.append(entity_type)

    if field_name:
        where_sql += " and a.field_name = %s"
        params.append(field_name)

    if date_from:
        where_sql += " and a.created_at >= %s::date"
        params.append(date_from)

    if date_to:
        where_sql += " and a.created_at < (%s::date + interval '1 day')"
        params.append(date_to)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select distinct actor_name
                from dharma_audit_logs
                where branch = 'CHE'
                  and coalesce(actor_name, '') <> ''
                order by actor_name
            """)
            actors = [row["actor_name"] for row in cur.fetchall()]

            cur.execute(f"""
                select
                    a.id,
                    a.actor_name,
                    a.action_type,
                    a.entity_type,
                    a.student_id,
                    a.student_name,
                    a.record_date,
                    a.group_id,
                    g.name as group_name,
                    a.field_name,
                    a.old_value,
                    a.new_value,
                    a.source_ip,
                    a.created_at
                from dharma_audit_logs a
                left join dharma_class_groups g
                    on g.id = a.group_id
                {where_sql}
                order by a.created_at desc, a.id desc
                limit 1000
            """, params)
            logs = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>佛学班 Audit Log</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
<style>
.audit-summary{background:#fff8e8;border:1px solid #f3d98c;border-radius:16px;padding:14px 16px;margin:14px 0;text-align:center;font-weight:800}
.audit-old{color:#b91c1c;font-weight:700}.audit-new{color:#15803d;font-weight:700}.audit-time{white-space:nowrap}.audit-table{min-width:1200px}
.audit-badge{display:inline-block;border-radius:999px;padding:5px 10px;background:#eef2ff;font-weight:800;font-size:14px}
@media(max-width:700px){.audit-table{min-width:1250px}}
</style>
</head>
<body>
<div class="page">
<div class="card">
<h1 class="page-title">🛡️ 修改记录 Audit Log</h1>
<p class="page-subtitle">查看哪位老师在什么时间，把哪一项资料从什么改成什么。</p>
<div class="audit-summary">目前显示最近 {{ logs|length }} 笔修改记录（最多 1000 笔）</div>
<form method="get">
<div class="form-group"><label class="form-label">搜索</label><input class="form-input" name="q" value="{{ q }}" placeholder="学生姓名／老师／旧值／新值"></div>
<div class="form-group"><label class="form-label">老师</label><select class="form-input" name="actor"><option value="">全部老师</option>{% for name in actors %}<option value="{{ name }}" {% if actor == name %}selected{% endif %}>{{ name }}</option>{% endfor %}</select></div>
<div class="form-group"><label class="form-label">资料类型</label><select class="form-input" name="entity_type"><option value="">全部</option><option value="attendance" {% if entity_type=='attendance' %}selected{% endif %}>出席</option><option value="homework" {% if entity_type=='homework' %}selected{% endif %}>功课</option><option value="lesson" {% if entity_type=='lesson' %}selected{% endif %}>课程</option><option value="student" {% if entity_type=='student' %}selected{% endif %}>学生资料</option></select></div>
<div class="form-group"><label class="form-label">修改项目</label><select class="form-input" name="field_name"><option value="">全部项目</option>{% for key, label in field_labels.items() %}<option value="{{ key }}" {% if field_name==key %}selected{% endif %}>{{ label }}</option>{% endfor %}</select></div>
<div class="form-group"><label class="form-label">开始日期</label><input class="form-input" type="date" name="date_from" value="{{ date_from }}"></div>
<div class="form-group"><label class="form-label">结束日期</label><input class="form-input" type="date" name="date_to" value="{{ date_to }}"></div>
<div class="btn-row"><button class="btn-tool btn-primary" type="submit">🔍 查询</button><a class="btn-tool btn-secondary" href="{{ url_for('dharma_class.class_audit_log') }}">↺ 清除</a></div>
</form>
</div>
<div class="card">
{% if logs %}
<div class="table-responsive"><table class="record-table audit-table"><thead><tr><th>修改时间</th><th>老师</th><th>学生</th><th>记录日期</th><th>组别</th><th>资料</th><th>项目</th><th>原本</th><th>修改后</th><th>来源</th></tr></thead><tbody>
{% for r in logs %}<tr><td class="audit-time">{{ r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '—' }}</td><td><strong>{{ r.actor_name or '老师' }}</strong></td><td>{% if r.student_id %}<a href="{{ url_for('dharma_class.class_student_profile', student_id=r.student_id) }}">{{ r.student_name or ('学生 #' ~ r.student_id) }}</a>{% else %}—{% endif %}</td><td>{{ r.record_date or '—' }}</td><td>{{ r.group_name or '—' }}</td><td><span class="audit-badge">{% if r.entity_type=='attendance' %}出席{% elif r.entity_type=='homework' %}功课{% elif r.entity_type=='lesson' %}课程{% elif r.entity_type=='student' %}学生资料{% else %}{{ r.entity_type }}{% endif %}</span></td><td>{{ field_labels.get(r.field_name, r.field_name) }}</td><td class="audit-old">{{ value_labels.get(r.old_value, r.old_value) if r.old_value else '（空白）' }}</td><td class="audit-new">{{ value_labels.get(r.new_value, r.new_value) if r.new_value else '（空白）' }}</td><td>{{ r.source_ip or '—' }}</td></tr>{% endfor %}
</tbody></table></div>
{% else %}<div class="empty-state">没有找到符合条件的修改记录。</div>{% endif %}
<a
    class="btn-tool btn-secondary"
    href="{{ url_for('dharma_class.class_admin') }}"
>
    ⬅ 返回负责人中心
</a>
</div></div></body></html>
""",
        logs=logs,
        actors=actors,
        q=q,
        actor=actor,
        entity_type=entity_type,
        field_name=field_name,
        date_from=date_from,
        date_to=date_to,
        field_labels=AUDIT_FIELD_LABELS,
        value_labels=AUDIT_VALUE_LABELS,
    )


@dharma_class_bp.route("/reports")
def class_reports():

    from datetime import datetime
    from zoneinfo import ZoneInfo

    malaysia_today = datetime.now(
        ZoneInfo("Asia/Kuala_Lumpur")
    ).date()

    # month = 月报，year = 年报
    mode = request.args.get(
        "mode",
        "month"
    ).strip().lower()

    if mode not in ("month", "year"):
        mode = "month"

    ym = request.args.get(
        "ym",
        malaysia_today.strftime("%Y-%m")
    ).strip()

    year_text = request.args.get(
        "year",
        ym[:4] if len(ym) >= 4 else str(malaysia_today.year)
    ).strip()

    group_id = request.args.get(
        "group_id",
        ""
    ).strip()

    try:
        selected_year = int(year_text)
    except (TypeError, ValueError):
        selected_year = malaysia_today.year

    selected_month = malaysia_today.month

    # 根据月报或年报建立日期范围
    if mode == "month":

        try:
            selected_month_date = datetime.strptime(
                ym,
                "%Y-%m"
            ).date()

            selected_year = selected_month_date.year
            selected_month = selected_month_date.month

        except ValueError:
            selected_year = malaysia_today.year
            selected_month = malaysia_today.month
            ym = malaysia_today.strftime("%Y-%m")

        if selected_month == 12:
            next_year = selected_year + 1
            next_month = 1
        else:
            next_year = selected_year
            next_month = selected_month + 1

        start_date = f"{selected_year:04d}-{selected_month:02d}-01"
        end_date = f"{next_year:04d}-{next_month:02d}-01"

        report_title = (
            f"{selected_year}年"
            f"{selected_month}月出席统计"
        )

    else:

        start_date = f"{selected_year:04d}-01-01"
        end_date = f"{selected_year + 1:04d}-01-01"

        report_title = f"{selected_year}年全年出席统计"

    with get_conn() as conn:
        with conn.cursor(
            cursor_factory=RealDictCursor
        ) as cur:

            # 读取组别
            cur.execute("""
                select
                    id,
                    name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)

            groups = cur.fetchall()

            # 查询条件
            student_where = """
                where s.status = 'active'
                  and s.branch = 'CHE'
            """

            student_params = [
                start_date,
                end_date
            ]

            if group_id:
                student_where += """
                    and s.group_id = %s
                """
                student_params.append(group_id)

            # 每位学生出席统计
            cur.execute(f"""
                select
                    g.id as group_id,
                    g.name as group_name,
                    g.sort_order,

                    s.id as student_id,
                    s.name,
                    s.english_name,

                    count(a.id) as total_records,

                    count(a.id)
                        filter (
                            where a.status = 'present'
                        ) as present_count,

                    count(a.id)
                        filter (
                            where a.status = 'late'
                        ) as late_count,

                    count(a.id)
                        filter (
                            where a.status = 'farm'
                        ) as farm_count,

                    count(a.id)
                        filter (
                            where a.status in ('absent', 'leave')
                        ) as absent_count

                from dharma_students s

                left join dharma_class_groups g
                    on g.id = s.group_id

                left join dharma_attendance a
                    on a.student_id = s.id
                   and a.branch = 'CHE'
                   and a.class_date >= %s
                   and a.class_date < %s

                {student_where}

                group by
                    g.id,
                    g.name,
                    g.sort_order,
                    s.id,
                    s.name,
                    s.english_name

                order by
                    g.sort_order,
                    s.name
            """, student_params)

            rows = cur.fetchall()

            # 课程堂数
            lesson_where = """
                where l.branch = 'CHE'
                  and l.lesson_date >= %s
                  and l.lesson_date < %s
            """

            lesson_params = [
                start_date,
                end_date
            ]

            if group_id:
                lesson_where += """
                    and l.group_id = %s
                """
                lesson_params.append(group_id)

            cur.execute(f"""
                select
                    count(*) as lesson_count,
                    count(distinct l.group_id) as group_count,
                    count(distinct l.teacher_name)
                        filter (
                            where coalesce(
                                l.teacher_name,
                                ''
                            ) <> ''
                        ) as teacher_count
                from dharma_class_lessons l
                {lesson_where}
            """, lesson_params)

            lesson_summary = cur.fetchone() or {}

            # 各组课程堂数
            cur.execute(f"""
                select
                    g.id as group_id,
                    g.name as group_name,
                    g.sort_order,
                    count(l.id) as lesson_count
                from dharma_class_groups g

                left join dharma_class_lessons l
                    on l.group_id = g.id
                   and l.branch = 'CHE'
                   and l.lesson_date >= %s
                   and l.lesson_date < %s

                where g.is_active = true
                {"and g.id = %s" if group_id else ""}

                group by
                    g.id,
                    g.name,
                    g.sort_order

                order by
                    g.sort_order,
                    g.id
            """, lesson_params)

            group_lessons = cur.fetchall()

    # 计算每位学生出席率
    total_present = 0
    total_late = 0
    total_farm = 0
    total_absent = 0
    total_marked_records = 0

    for row in rows:

        total_records = int(
            row["total_records"] or 0
        )

        present_count = int(
            row["present_count"] or 0
        )

        late_count = int(
            row["late_count"] or 0
        )

        farm_count = int(
            row["farm_count"] or 0
        )

        absent_count = int(
            row["absent_count"] or 0
        )

        attended_count = (
            present_count
            + late_count
            + farm_count
        )

        effective_total = attended_count + absent_count

        if effective_total > 0:
            row["rate"] = round(
                attended_count * 100 / effective_total,
                1
            )
        else:
            row["rate"] = None

        total_present += present_count
        total_late += late_count
        total_farm += farm_count
        total_absent += absent_count
        total_marked_records += effective_total

    student_count = len(rows)

    total_attended = (
        total_present
        + total_late
        + total_farm
    )

    if total_marked_records > 0:
        overall_rate = round(
            total_attended
            * 100
            / total_marked_records,
            1
        )
    else:
        overall_rate = None

    lesson_count = int(
        lesson_summary.get("lesson_count") or 0
    )


    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">

<meta
    name="viewport"
    content="width=device-width, initial-scale=1"
>

<title>佛学班报表中心</title>

<link
    rel="stylesheet"
    href="/static/css/toolbox.css"
>

<style>
.report-mode-row {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 12px;
    margin-bottom: 20px;
}

.report-mode-btn {
    display: flex;
    align-items: center;
    justify-content: center;
    min-height: 58px;
    border-radius: 15px;
    text-decoration: none;
    font-size: 20px;
    font-weight: 800;
    background: #f1f3f5;
    color: #444;
}

.report-mode-btn.active {
    background: #2563eb;
    color: white;
}

.report-summary {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin-top: 20px;
}

.report-summary-box {
    background: #f6f7f9;
    border-radius: 16px;
    padding: 16px 10px;
    text-align: center;
}

.report-summary-label {
    color: #666;
    font-size: 16px;
}

.report-summary-number {
    font-size: 29px;
    font-weight: 800;
    margin-top: 6px;
}

.group-lesson-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 12px;
    margin: 16px 0 20px;
}

.group-lesson-box {
    border: 1px solid #e5e7eb;
    border-radius: 15px;
    padding: 15px;
    text-align: center;
    background: white;
}

.group-lesson-name {
    font-size: 18px;
    font-weight: 800;
}

.group-lesson-count {
    font-size: 27px;
    font-weight: 800;
    margin-top: 6px;
}

.rate-good {
    color: #15803d;
    font-weight: 800;
}

.rate-warning {
    color: #b45309;
    font-weight: 800;
}

.rate-low {
    color: #b91c1c;
    font-weight: 800;
}

@media (max-width: 750px) {
    .report-summary {
        grid-template-columns: repeat(2, 1fr);
    }

    .group-lesson-grid {
        grid-template-columns: 1fr;
    }
}
</style>
</head>

<body>
<div class="page">

    <div class="card">

        <h1 class="page-title">
            📊 佛学班报表中心
        </h1>

        <p class="page-subtitle">
            查看学生月度及年度出席统计。
        </p>

        <div class="report-mode-row">

            <a
                class="report-mode-btn
                    {% if mode == 'month' %}active{% endif %}"
                href="{{ url_for(
                    'dharma_class.class_reports',
                    mode='month',
                    ym=ym,
                    group_id=group_id
                ) }}"
            >
                📅 月报
            </a>

            <a
                class="report-mode-btn
                    {% if mode == 'year' %}active{% endif %}"
                href="{{ url_for(
                    'dharma_class.class_reports',
                    mode='year',
                    year=selected_year,
                    group_id=group_id
                ) }}"
            >
                📆 年报
            </a>

        </div>

        <form method="get">

            <input
                type="hidden"
                name="mode"
                value="{{ mode }}"
            >

            {% if mode == "month" %}

                <div class="form-group">
                    <label class="form-label">
                        月份
                    </label>

                    <input
                        class="form-input"
                        type="month"
                        name="ym"
                        value="{{ ym }}"
                        required
                    >
                </div>

            {% else %}

                <div class="form-group">
                    <label class="form-label">
                        年份
                    </label>

                    <input
                        class="form-input"
                        type="number"
                        name="year"
                        value="{{ selected_year }}"
                        min="2020"
                        max="2100"
                        required
                    >
                </div>

            {% endif %}

            <div class="form-group">
                <label class="form-label">
                    组别
                </label>

                <select
                    class="form-input"
                    name="group_id"
                >
                    <option value="">
                        全部组别
                    </option>

                    {% for g in groups %}
                        <option
                            value="{{ g.id }}"
                            {% if
                                group_id|string
                                == g.id|string
                            %}
                                selected
                            {% endif %}
                        >
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="btn-row">
                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    🔍 查看报表
                </button>
            </div>

        </form>

        <div class="report-summary">

            <div class="report-summary-box">
                <div class="report-summary-label">
                    学生人数
                </div>

                <div class="report-summary-number">
                    {{ student_count }}
                </div>
            </div>

            <div class="report-summary-box">
                <div class="report-summary-label">
                    课程堂数
                </div>

                <div class="report-summary-number">
                    {{ lesson_count }}
                </div>
            </div>

            <div class="report-summary-box">
                <div class="report-summary-label">
                    出席次数
                </div>

                <div class="report-summary-number">
                    {{ total_attended }}
                </div>
            </div>

            <div class="report-summary-box">
                <div class="report-summary-label">
                    平均出席率
                </div>

                <div class="report-summary-number">
                    {% if overall_rate is not none %}
                        {{ overall_rate }}%
                    {% else %}
                        —
                    {% endif %}
                </div>
            </div>

        </div>

    </div>

    <div class="card">

        <h2 class="section-title">
            📚 各组课程堂数
        </h2>

        <div class="group-lesson-grid">

            {% for g in group_lessons %}
                <div class="group-lesson-box">

                    <div class="group-lesson-name">
                        {{ g.group_name }}
                    </div>

                    <div class="group-lesson-count">
                        {{ g.lesson_count or 0 }} 堂
                    </div>

                </div>
            {% endfor %}

        </div>

        <h2 class="section-title">
            {{ report_title }}
        </h2>

        {% if rows %}

            <div class="table-responsive">

                <table class="record-table">

                    <thead>
                        <tr>
                            <th>组别</th>
                            <th>姓名</th>
                            <th>已记录</th>
                            <th>出席</th>
                            <th>迟到</th>
                            <th>农舍</th>
                            <th>缺席</th>
                            <th>出席率</th>
                        </tr>
                    </thead>

                    <tbody>

                        {% for r in rows %}

                        <tr>
                            <td>
                                {{ r.group_name or "—" }}
                            </td>
                            
                            <td>
                                <strong>{{ r.name }}</strong>

                                {% if r.english_name %}
                                    <br>
                                    <span style="color:#777;">
                                        {{ r.english_name }}
                                    </span>
                                {% endif %}
                            </td>

                            <td>
                                {{ r.total_records or 0 }}
                            </td>

                            <td>
                                {{ r.present_count or 0 }}
                            </td>

                            <td>
                                {{ r.late_count or 0 }}
                            </td>

                            <td>
                                {{ r.farm_count or 0 }}
                            </td>

                            <td>
                                {{ r.absent_count or 0 }}
                            </td>

                            <td>
                                {% if r.rate is not none %}

                                    <span class="
                                        {% if r.rate >= 80 %}
                                            rate-good
                                        {% elif r.rate >= 60 %}
                                            rate-warning
                                        {% else %}
                                            rate-low
                                        {% endif %}
                                    ">
                                        {{ r.rate }}%
                                    </span>

                                {% else %}
                                    —
                                {% endif %}
                            </td>
                        </tr>

                        {% endfor %}

                    </tbody>

                </table>

            </div>

        {% else %}

            <div class="empty-state">
                暂时没有符合条件的学生资料。
            </div>

        {% endif %}

        {% if mode == "month" %}

            <div class="btn-row">

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_monthly_report',
                        ym=ym,
                        group_id=group_id
                    ) }}"
                >
                    📥 下载月报 Excel
                </a>

            </div>

            <h2 class="section-title">
                📚 功课追踪报表
            </h2>

            <div class="btn-row">

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_baihua_homework',
                        year=selected_year,
                        month=selected_month
                    ) }}"
                >
                    📚 本月白话佛法功课
                </a>

                <a
                    class="btn-tool btn-warning"
                    href="{{ url_for(
                        'dharma_class.class_export_scripture_homework',
                        year=selected_year,
                        month=selected_month
                    ) }}"
                >
                    📿 本月经文功课
                </a>

            </div>

        {% else %}

            <div class="btn-row">

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_yearly_report',
                        year=selected_year,
                        group_id=group_id
                    ) }}"
                >
                    📥 下载年报 Excel
                </a>

            </div>

            <h2 class="section-title">
                📚 全年功课追踪报表
            </h2>

            <div class="btn-row">
                               
                <a
                    class="btn-tool btn-success"
                    href="{{ url_for(
                        'dharma_class.class_export_baihua_homework',
                        year=selected_year
                    ) }}"
                >
                    📚 全年白话佛法功课
                </a>

                <a
                    class="btn-tool btn-warning"
                    href="{{ url_for(
                        'dharma_class.class_export_scripture_homework',
                        year=selected_year
                    ) }}"
                >
                    📿 全年经文功课
                </a>

            </div>

        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('dharma_class.class_admin') }}"
            >
                ⬅ 返回负责人中心
            </a>
        </div>

    </div>

</div>
</body>
</html>
""",
        mode=mode,
        ym=ym,
        selected_year=selected_year,
        selected_month=selected_month,
        group_id=group_id,
        groups=groups,
        rows=rows,
        report_title=report_title,
        student_count=student_count,
        lesson_count=lesson_count,
        total_attended=total_attended,
        overall_rate=overall_rate,
        group_lessons=group_lessons
    )

@dharma_class_bp.route("/records")
def class_records():

    selected_date = request.args.get("class_date") or date.today().isoformat()
    group_id = request.args.get("group_id")

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            params = [selected_date]
            where_group = ""

            if group_id:
                where_group = "and s.group_id = %s"
                params.append(group_id)

            cur.execute(f"""
                select
                    s.name,
                    s.english_name,
                    g.name as group_name,
                    a.status,
                    a.remark,
                    a.marked_by,
                    a.marked_at
                from dharma_attendance a
                join dharma_students s on s.id = a.student_id
                left join dharma_class_groups g on g.id = s.group_id
                where a.class_date = %s
                {where_group}
                order by g.sort_order, s.name
            """, params)

            records = cur.fetchall()

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>点名记录</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">📅 点名记录</h1>
        <p class="page-subtitle">查看某一天佛学班点名情况。</p>

        <form method="get">
            <div class="form-group">
                <label class="form-label">日期</label>
                <input class="form-input" type="date" name="class_date" value="{{ selected_date }}">
            </div>

            <div class="form-group">
                <label class="form-label">组别</label>
                <select class="form-input" name="group_id">
                    <option value="">全部组别</option>
                    {% for g in groups %}
                        <option value="{{ g.id }}"
                            {% if group_id|string == g.id|string %}selected{% endif %}>
                            {{ g.name }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-primary" type="submit">
                    🔍 查看记录
                </button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2 class="section-title">{{ selected_date }} 点名记录</h2>

        {% if records %}
            <div class="table-responsive">
                <table class="record-table">
                    <thead>
                        <tr>
                            <th>组别</th>
                            <th>姓名</th>
                            <th>状态</th>
                            <th>备注</th>
                            <th>老师</th>
                            <th>保存时间</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for r in records %}
                        <tr>
                            <td>{{ r.group_name or "-" }}</td>
                            
                            <td>
                                {{ r.name }}
                                {% if r.english_name %}
                                    <span style="color:#777;">{{ r.english_name }}</span>
                                {% endif %}
                            </td>
                            <td>
                                {% if r.status == "present" %}
                                    ✅ 出席
                                {% elif r.status == "late" %}
                                    ⏰ 迟到
                                {% elif r.status == "farm" %}
                                    🌱 农舍
                                {% elif r.status in ("absent", "leave") %}
                                    ❌ 缺席
                                {% else %}
                                    {{ r.status }}
                                {% endif %}
                            </td>
                            <td>{{ r.remark or "-" }}</td>
                            <td>{{ r.marked_by or "-" }}</td>
                            <td>{{ r.marked_at or "-" }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        {% else %}
            <div class="empty-state">
                这一天还没有点名记录。
            </div>
        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('dharma_class.class_admin') }}"
            >
                ⬅ 返回负责人中心
            </a>
        </div>
    </div>

</div>
</body>
</html>
""",
        selected_date=selected_date,
        group_id=group_id,
        groups=groups,
        records=records
    )

@dharma_class_bp.route("/reports/monthly/export")
def class_export_monthly_report():

    ym = request.args.get("ym") or date.today().strftime("%Y-%m")

    # =========================================================
    # 出席状态显示
    # =========================================================
    status_symbol = {
        "present": "✓",
        "late": "迟",
        "farm": "农",
        "leave": "✗",
        "absent": "✗",
    }

    status_fill = {
        "✓": PatternFill("solid", fgColor="C6EFCE"),
        "迟": PatternFill("solid", fgColor="FCE4D6"),
        "农": PatternFill("solid", fgColor="D9EAD3"),
        "✗": PatternFill("solid", fgColor="F4CCCC"),
    }

    # =========================================================
    # 读取资料
    # =========================================================
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            # 组别
            cur.execute("""
                select
                    id,
                    name,
                    sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            # 在籍学生
            cur.execute("""
                select
                    s.id,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.name as group_name,
                    g.sort_order
                from dharma_students s
                join dharma_class_groups g
                  on g.id = s.group_id
                where s.status = 'active'
                  and g.is_active = true
                order by
                    g.sort_order,
                    s.name
            """)
            students = cur.fetchall()

            # 本月出席记录
            cur.execute("""
                select
                    student_id,
                    group_id,
                    class_date,
                    status
                from dharma_attendance
                where to_char(class_date, 'YYYY-MM') = %s
                order by class_date
            """, (ym,))
            att_rows = cur.fetchall()

    # =========================================================
    # 整理出席资料
    # =========================================================
    attendance_map = {}

    for r in att_rows:
        attendance_map[
            (
                r["student_id"],
                r["class_date"]
            )
        ] = r["status"]

    # 每组实际有点名记录的日期
    group_dates_map = {}

    for g in groups:
        group_dates_map[g["id"]] = []

    for r in att_rows:
        group_id = r["group_id"]
        class_date = r["class_date"]

        if group_id not in group_dates_map:
            group_dates_map[group_id] = []

        if class_date not in group_dates_map[group_id]:
            group_dates_map[group_id].append(class_date)

    for group_id in group_dates_map:
        group_dates_map[group_id].sort()

    # 全部课程日期，用于总览显示
    all_class_dates = sorted({
        r["class_date"]
        for r in att_rows
    })

    # =========================================================
    # 建立 Excel
    # =========================================================
    wb = Workbook()

    # =========================================================
    # 样式
    # =========================================================
    title_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="5B9BD5"
    )

    summary_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_font = Font(
        size=18,
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    thin = Side(
        style="thin",
        color="CCCCCC"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    left = Alignment(
        horizontal="left",
        vertical="center"
    )

    # =========================================================
    # 总览 Sheet
    # =========================================================
    ws = wb.active
    ws.title = "总览"

    ws.merge_cells("A1:H1")

    ws["A1"] = f"蕉赖佛学班出席月报  {ym}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center

    ws.row_dimensions[1].height = 34

    ws["A3"] = "本月课程日期"
    ws["A3"].font = bold_font
    ws["A3"].alignment = left

    if all_class_dates:
        date_text = "、".join(
            f"{d.month}/{d.day}"
            for d in all_class_dates
        )
    else:
        date_text = "本月还没有点名记录"

    ws["B3"] = date_text
    ws["B3"].alignment = left

    ws.append([])

    ws.append([
        "组别",
        "总学生人数",
        "课程次数",
        "出席人数",
        "迟到人数",
        "农舍人数",
        "缺席人数",
        "平均出席率",
    ])

    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    for g in groups:

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        group_dates = group_dates_map.get(
            g["id"],
            []
        )

        total_students = len(group_students)

        group_present = 0
        group_late = 0
        group_farm = 0
        group_absent = 0

        student_rates = []

        for s in group_students:

            present_count = 0
            late_count = 0
            farm_count = 0
            absent_count = 0

            for class_date in group_dates:

                status = attendance_map.get(
                    (
                        s["id"],
                        class_date
                    )
                )

                if status == "present":
                    present_count += 1
                    group_present += 1

                elif status == "late":
                    late_count += 1
                    group_late += 1

                elif status == "farm":
                    farm_count += 1
                    group_farm += 1

                elif status in ("absent", "leave"):
                    absent_count += 1
                    group_absent += 1

            attended_count = (
                present_count +
                late_count +
                farm_count
            )

            effective_total = attended_count + absent_count

            if effective_total > 0:
                student_rates.append(
                    attended_count / effective_total
                )

        avg_rate = (
            sum(student_rates) / len(student_rates)
            if student_rates
            else None
        )

        ws.append([
            g["name"],
            total_students,
            len(group_dates),
            group_present,
            group_late,
            group_farm,
            group_absent,
            avg_rate,
        ])

    for row in ws.iter_rows(
        min_row=6,
        max_row=ws.max_row
    ):
        for cell in row:
            cell.border = border
            cell.alignment = center

        rate_cell = row[7]

        if (
            rate_cell.value is not None
            and isinstance(
                rate_cell.value,
                (int, float)
            )
        ):
            rate_cell.number_format = "0%"

            if rate_cell.value >= 0.9:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="C6EFCE"
                )

            elif rate_cell.value >= 0.75:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="FFEB9C"
                )

            else:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="F4CCCC"
                )

    ws.freeze_panes = "A6"

    if ws.max_row >= 6:
        ws.auto_filter.ref = (
            f"A5:H{ws.max_row}"
        )

    # =========================================================
    # 各组 Sheet
    # =========================================================
    for g in groups:

        sheet_name = g["name"][:31]
        ws_g = wb.create_sheet(sheet_name)

        group_dates = group_dates_map.get(
            g["id"],
            []
        )

        headers = [
            "学生姓名",
        ]

        for class_date in group_dates:
            headers.append(
                f"{class_date.month}/{class_date.day}"
            )

        headers += [
            "出席",
            "迟到",
            "农舍",
            "缺席",
            "出席率",
        ]

        end_col = max(
            7,
            len(headers)
        )

        ws_g.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=end_col
        )

        ws_g["A1"] = (
            f"{g['name']} 出席月报  {ym}"
        )

        ws_g["A1"].font = title_font
        ws_g["A1"].fill = title_fill
        ws_g["A1"].alignment = center

        ws_g.row_dimensions[1].height = 34

        ws_g.append([])
        ws_g.append(headers)

        for cell in ws_g[3]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        if not group_students:
            ws_g.append([
                "",
                "目前没有学生资料"
            ])

        for s in group_students:

            present_count = 0
            late_count = 0
            farm_count = 0
            absent_count = 0

            row_data = [
                s["name"] or "",
            ]

            for class_date in group_dates:

                status = attendance_map.get(
                    (
                        s["id"],
                        class_date
                    )
                )

                if status == "present":
                    present_count += 1

                elif status == "late":
                    late_count += 1

                elif status == "farm":
                    farm_count += 1

                elif status in ("absent", "leave"):
                    absent_count += 1

                row_data.append(
                    status_symbol.get(
                        status,
                        ""
                    )
                )

            attended_count = (
                present_count +
                late_count +
                farm_count
            )

            effective_total = attended_count + absent_count

            if effective_total > 0:
                attendance_rate = (
                    attended_count /
                    effective_total
                )
            else:
                attendance_rate = None

            row_data += [
                present_count,
                late_count,
                farm_count,
                absent_count,
                attendance_rate,
            ]

            ws_g.append(row_data)

        # =====================================================
        # 学生资料样式
        # =====================================================
        for row in ws_g.iter_rows(
            min_row=4,
            max_row=ws_g.max_row
        ):
            for cell in row:
                cell.border = border
                cell.alignment = center

                if cell.value in status_fill:
                    cell.fill = status_fill[cell.value]
                    cell.font = Font(
                        bold=True
                    )

            rate_cell = row[-1]

            if (
                rate_cell.value is not None
                and isinstance(
                    rate_cell.value,
                    (int, float)
                )
            ):
                rate_cell.number_format = "0%"

                if rate_cell.value >= 0.9:
                    rate_cell.fill = PatternFill(
                        "solid",
                        fgColor="C6EFCE"
                    )

                elif rate_cell.value >= 0.75:
                    rate_cell.fill = PatternFill(
                        "solid",
                        fgColor="FFEB9C"
                    )

                else:
                    rate_cell.fill = PatternFill(
                        "solid",
                        fgColor="F4CCCC"
                    )

        ws_g.freeze_panes = "C4"

        if ws_g.max_row >= 4:
            ws_g.auto_filter.ref = (
                f"A3:"
                f"{get_column_letter(ws_g.max_column)}"
                f"{ws_g.max_row}"
            )

        ws_g.column_dimensions["A"].width = 14
        ws_g.column_dimensions["B"].width = 22

    # =========================================================
    # 全部 Sheet 自动美化
    # =========================================================
    for sheet in wb.worksheets:

        for col_idx in range(
            1,
            sheet.max_column + 1
        ):
            col_letter = get_column_letter(
                col_idx
            )

            max_len = 0

            for row_idx in range(
                1,
                sheet.max_row + 1
            ):
                cell = sheet.cell(
                    row=row_idx,
                    column=col_idx
                )

                value = (
                    str(cell.value)
                    if cell.value is not None
                    else ""
                )

                max_len = max(
                    max_len,
                    len(value)
                )

            width = max_len + 4

            if col_idx == 1:
                width = max(
                    width,
                    14
                )

            elif col_idx == 2:
                width = max(
                    width,
                    22
                )

            else:
                width = max(
                    width,
                    10
                )

            # 防止课程日期太多时列宽过大
            width = min(
                width,
                35
            )

            sheet.column_dimensions[
                col_letter
            ].width = width

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=(
                        cell.alignment.horizontal
                        or "center"
                    ),
                    vertical="center",
                    wrap_text=True
                )

        for row_idx in range(
            1,
            sheet.max_row + 1
        ):
            sheet.row_dimensions[
                row_idx
            ].height = 24

        sheet.sheet_view.showGridLines = False

    # =========================================================
    # 输出 Excel
    # =========================================================
    output = BytesIO()

    wb.save(output)
    output.seek(0)

    filename = (
        f"蕉赖佛学班出席月报_{ym}.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@dharma_class_bp.route("/students/import", methods=["GET", "POST"])
def class_students_import():

    result = None

    if request.method == "POST":
        file = request.files.get("excel_file")

        if not file or file.filename == "":
            flash("请选择 Excel 文件。", "bad")
            return redirect(url_for("dharma_class.class_students_import"))

        wb = load_workbook(file, data_only=True)
        ws = wb.active

        inserted = 0
        duplicate = 0
        blank = 0
        invalid = 0

        inserted_names = []
        duplicate_names = []
        error_messages = []

        with get_conn() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:

                cur.execute("""
                    select id, name
                    from dharma_class_groups
                    where is_active = true
                """)
                groups = cur.fetchall()

                group_map = {
                    g["name"].strip(): g["id"]
                    for g in groups
                }

                for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    
                    name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    english_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                    gender = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                    birth_year = None
                    if len(row) > 4 and row[4]:
                        try:
                            birth_year = int(row[4])
                        except:
                            birth_year = None

                    parent_name = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                    parent_phone = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                    group_name = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                    remark = str(row[8]).strip() if len(row) > 8 and row[8] else ""

                    if not name:
                        blank += 1
                        continue

                    group_id = group_map.get(group_name)

                    if not group_id:
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：组别「{group_name}」不存在"
                        )
                        continue

                    # 性别检查
                    if gender not in ("男", "女"):
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：性别必须填写「男」或「女」"
                        )
                        continue

                    # 出生年份检查
                    if birth_year is None:
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：请填写出生年份"
                        )
                        continue

                    if birth_year < 2000 or birth_year > 2100:
                        invalid += 1
                        error_messages.append(
                            f"第 {row_no} 行：出生年份错误"
                        )
                        continue

                    # 重复检查：中文姓名 + 家长电话
                    cur.execute("""
                        select id
                        from dharma_students
                        where name = %s
                        and coalesce(parent_phone, '') = %s
                        limit 1
                    """, (
                        name,
                        parent_phone or ""
                    ))

                    exists = cur.fetchone()

                    if exists:
                        duplicate += 1
                        duplicate_names.append(name)
                        continue


                    cur.execute("""
                        insert into dharma_students
                        (
                            branch,
                            name,
                            english_name,
                            gender,
                            birth_year,
                            parent_name,
                            parent_phone,
                            group_id,
                            status,
                            remark
                        )
                        values
                        (
                            'CHE',
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            %s,
                            'active',
                            %s
                        )
                    """, (
                        name,
                        english_name or None,
                        gender or None,
                        birth_year,
                        parent_name or None,
                        parent_phone or None,
                        group_id,
                        remark or None
                    ))

                    inserted += 1
                    inserted_names.append(name)

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>导入学生 Excel</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">📥 导入学生 Excel</h1>
        <p class="page-subtitle">上传学生名单，系统会自动新增并检查重复。</p>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert">{{ msg }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <div class="empty-state" style="text-align:left;">
            Excel 第一行标题请使用：<br><br>
            中文姓名｜家长姓名｜家长电话｜组别｜备注<br><br>
            组别目前只接受：低年组、少年组、高年组
        </div>

        <div class="btn-row">
            <a class="btn-tool btn-secondary" href="/class/students/template">
                📄 下载 Excel 模板
            </a>
        </div>

        <form method="post" enctype="multipart/form-data">

            <div class="form-group">
                <label class="form-label">选择 Excel 文件</label>
                <input class="form-input" type="file" name="excel_file" accept=".xlsx" required>
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-success" type="submit">
                    ✅ 上传并导入
                </button>
            </div>

        </form>
    </div>

    {% if result %}
    <div class="card">
        <h2 class="section-title">📋 导入结果</h2>

        <div class="summary-grid">
            <div class="summary-box">
                <div>新增</div>
                <div style="font-size:30px;font-weight:800;">{{ result.inserted }}</div>
            </div>

            <div class="summary-box">
                <div>重复</div>
                <div style="font-size:30px;font-weight:800;">{{ result.duplicate }}</div>
            </div>

            <div class="summary-box">
                <div>空白</div>
                <div style="font-size:30px;font-weight:800;">{{ result.blank }}</div>
            </div>

            <div class="summary-box">
                <div>错误</div>
                <div style="font-size:30px;font-weight:800;">{{ result.invalid }}</div>
            </div>
        </div>

        {% if result.inserted_names %}
            <h3 class="section-title">✅ 新增学生</h3>
            <div class="empty-state" style="text-align:left;">
                {% for n in result.inserted_names[:30] %}
                    ✓ {{ n }}<br>
                {% endfor %}
            </div>
        {% endif %}

        {% if result.duplicate_names %}
            <h3 class="section-title">⚠️ 重复已跳过</h3>
            <div class="empty-state" style="text-align:left;">
                {% for n in result.duplicate_names[:30] %}
                    • {{ n }}<br>
                {% endfor %}
            </div>
        {% endif %}

        {% if result.error_messages %}
            <h3 class="section-title">❌ 错误资料</h3>
            <div class="empty-state" style="text-align:left;color:#b91c1c;">
                {% for e in result.error_messages[:30] %}
                    {{ e }}<br>
                {% endfor %}
            </div>
        {% endif %}
    </div>
    {% endif %}

    <div class="card">
        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('dharma_class.class_admin') }}"
            >
                ⬅ 返回负责人中心
            </a>
        </div>
    </div>

</div>
</body>
</html>
""", result=result)

@dharma_class_bp.route("/students/template")
def class_students_template():

    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"

    headers = [
        "中文姓名",
        "英文名",
        "性别",
        "出生年份",
        "父／母／监护人姓名",
        "父／母／监护人电话",
        "组别",
        "备注"
    ]

    ws.append(headers)

    ws.append([
        "001",
        "张小明",
        "Zhang Xiao Ming",
        "男",
        2016,
        "张爸爸",
        "0123456789",
        "低年组",
        ""
    ])

    ws.append([
        "002",
        "李小美",
        "Lee Mei Mei",
        "女",
        2014,
        "李妈妈",
        "01122334455",
        "少年组",
        ""
    ])

    beautify_simple_excel(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="佛学班学生导入模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@dharma_class_bp.route("/students/export")
def class_students_export():

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                select
                    s.name,
                    s.english_name,
                    s.parent_name,
                    s.parent_phone,
                    g.name as group_name,
                    s.status,
                    s.remark
                from dharma_students s
                left join dharma_class_groups g on g.id = s.group_id
                order by g.sort_order, s.name
            """)
            rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "学生名单"

    ws.append([
        "中文姓名",
        "英文名",
        "家长姓名",
        "家长电话",
        "组别",
        "状态",
        "备注"
    ])

    for r in rows:
        ws.append([
            r["name"] or "",
            r["english_name"] or "",
            r["parent_name"] or "",
            r["parent_phone"] or "",
            r["group_name"] or "",
            r["status"] or "",
            r["remark"] or ""
        ])

    beautify_simple_excel(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="佛学班学生名单.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@dharma_class_bp.route("/reports/yearly/export")
def class_export_yearly_report():

    year = request.args.get("year") or date.today().strftime("%Y")

    # =========================================================
    # 读取资料
    # =========================================================
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select
                    id,
                    name,
                    sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            cur.execute("""
                select
                    s.id,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.name as group_name,
                    g.sort_order
                from dharma_students s
                join dharma_class_groups g
                  on g.id = s.group_id
                where s.status = 'active'
                  and g.is_active = true
                order by
                    g.sort_order,
                    s.name
            """)
            students = cur.fetchall()

            cur.execute("""
                select
                    student_id,
                    to_char(class_date, 'MM') as month_no,
                    status
                from dharma_attendance
                where to_char(class_date, 'YYYY') = %s
                order by class_date
            """, (year,))
            att_rows = cur.fetchall()

    # =========================================================
    # 整理统计资料
    #
    # student_id
    #     -> month
    #         -> present / late / farm / absent
    # =========================================================
    stat = {}

    for r in att_rows:

        student_id = r["student_id"]
        month_no = int(r["month_no"])
        status = r["status"]

        stat.setdefault(student_id, {})

        stat[student_id].setdefault(month_no, {
            "records": 0,
            "present": 0,
            "late": 0,
            "farm": 0,
            "absent": 0,
            "attended": 0,
            "effective_total": 0,
        })

        month_data = stat[student_id][month_no]

        # 所有点名记录，包括请假
        month_data["records"] += 1

        if status == "present":
            month_data["present"] += 1
            month_data["attended"] += 1
            month_data["effective_total"] += 1

        elif status == "late":
            month_data["late"] += 1
            month_data["attended"] += 1
            month_data["effective_total"] += 1

        elif status == "farm":
            month_data["farm"] += 1
            month_data["attended"] += 1
            month_data["effective_total"] += 1

        elif status in ("absent", "leave"):
            month_data["absent"] += 1
            month_data["effective_total"] += 1

    # =========================================================
    # 建立 Excel
    # =========================================================
    wb = Workbook()

    # =========================================================
    # 样式
    # =========================================================
    title_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="5B9BD5"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_font = Font(
        size=18,
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    thin = Side(
        style="thin",
        color="CCCCCC"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    # =========================================================
    # 总览 Sheet
    # =========================================================
    ws = wb.active
    ws.title = "总览"

    ws.merge_cells("A1:H1")

    ws["A1"] = f"蕉赖佛学班出席年报  {year}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center

    ws.row_dimensions[1].height = 34

    ws.append([])

    ws.append([
        "组别",
        "学生人数",
        "全年记录",
        "出席",
        "迟到",
        "农舍",
        "缺席",
        "平均出席率",
    ])

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    for g in groups:

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        student_rates = []

        total_records = 0
        total_present = 0
        total_late = 0
        total_farm = 0
        total_absent = 0

        for s in group_students:

            student_attended = 0
            student_effective_total = 0

            for month_no in range(1, 13):

                data = stat.get(
                    s["id"],
                    {}
                ).get(month_no)

                if not data:
                    continue

                total_records += data["records"]
                total_present += data["present"]
                total_late += data["late"]
                total_farm += data["farm"]
                total_absent += data["absent"]

                student_attended += data["attended"]
                student_effective_total += data["effective_total"]

            if student_effective_total > 0:
                student_rates.append(
                    student_attended /
                    student_effective_total
                )

        avg_rate = (
            sum(student_rates) / len(student_rates)
            if student_rates
            else None
        )

        ws.append([
            g["name"],
            len(group_students),
            total_records,
            total_present,
            total_late,
            total_farm,
            total_absent,
            avg_rate,
        ])

    for row in ws.iter_rows(
        min_row=4,
        max_row=ws.max_row
    ):
        for cell in row:
            cell.border = border
            cell.alignment = center

        rate_cell = row[7]

        if (
            rate_cell.value is not None
            and isinstance(
                rate_cell.value,
                (int, float)
            )
        ):
            rate_cell.number_format = "0%"

            if rate_cell.value >= 0.9:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="C6EFCE"
                )

            elif rate_cell.value >= 0.75:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="FFEB9C"
                )

            else:
                rate_cell.fill = PatternFill(
                    "solid",
                    fgColor="F4CCCC"
                )

    ws.freeze_panes = "A4"

    if ws.max_row >= 4:
        ws.auto_filter.ref = (
            f"A3:H{ws.max_row}"
        )

    # =========================================================
    # 各组 Sheet
    # =========================================================
    for g in groups:

        ws_g = wb.create_sheet(
            g["name"][:31]
        )

        headers = [
            
            "学生姓名",
        ]

        headers += [
            f"{month_no}月"
            for month_no in range(1, 13)
        ]

        headers += [
            "全年出席率",
            "出席",
            "迟到",
            "农舍",
            "缺席",
        ]

        end_col = len(headers)

        ws_g.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=end_col
        )

        ws_g["A1"] = (
            f"{g['name']} 出席年报  {year}"
        )

        ws_g["A1"].font = title_font
        ws_g["A1"].fill = title_fill
        ws_g["A1"].alignment = center

        ws_g.row_dimensions[1].height = 34

        ws_g.append([])
        ws_g.append(headers)

        for cell in ws_g[3]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        group_students = [
            s for s in students
            if s["group_id"] == g["id"]
        ]

        if not group_students:
            ws_g.append([
                "",
                "目前没有学生资料"
            ])

        for s in group_students:

            row_data = [
                s["name"] or "",
            ]

            yearly_attended = 0
            yearly_effective_total = 0

            yearly_present = 0
            yearly_late = 0
            yearly_farm = 0
            yearly_absent = 0

            for month_no in range(1, 13):

                data = stat.get(
                    s["id"],
                    {}
                ).get(month_no)

                if data:

                    if data["effective_total"] > 0:
                        month_rate = (
                            data["attended"] /
                            data["effective_total"]
                        )
                    else:
                        # 当月只有请假，没有可计算记录
                        month_rate = None

                    row_data.append(month_rate)

                    yearly_attended += data["attended"]
                    yearly_effective_total += data["effective_total"]

                    yearly_present += data["present"]
                    yearly_late += data["late"]
                    yearly_farm += data["farm"]
                    yearly_absent += data["absent"]

                else:
                    row_data.append(None)

            if yearly_effective_total > 0:
                yearly_rate = (
                    yearly_attended /
                    yearly_effective_total
                )
            else:
                yearly_rate = None

            row_data += [
                yearly_rate,
                yearly_present,
                yearly_late,
                yearly_farm,
                yearly_absent,
            ]

            ws_g.append(row_data)

        # =====================================================
        # 内容样式
        # =====================================================
        for row in ws_g.iter_rows(
            min_row=4,
            max_row=ws_g.max_row
        ):
            for cell in row:
                cell.border = border
                cell.alignment = center

            # C 至 N：1月至12月
            # O：全年出席率
            for cell in row[2:15]:

                if isinstance(
                    cell.value,
                    (int, float)
                ):
                    cell.number_format = "0%"

                    if cell.value >= 0.9:
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="C6EFCE"
                        )

                    elif cell.value >= 0.75:
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="FFEB9C"
                        )

                    else:
                        cell.fill = PatternFill(
                            "solid",
                            fgColor="F4CCCC"
                        )

        ws_g.freeze_panes = "C4"

        if ws_g.max_row >= 4:
            ws_g.auto_filter.ref = (
                f"A3:"
                f"{get_column_letter(ws_g.max_column)}"
                f"{ws_g.max_row}"
            )

        ws_g.column_dimensions["A"].width = 14
        ws_g.column_dimensions["B"].width = 22

    # =========================================================
    # 全部 Sheet 自动美化
    # =========================================================
    for sheet in wb.worksheets:

        for col_idx in range(
            1,
            sheet.max_column + 1
        ):

            col_letter = get_column_letter(
                col_idx
            )

            max_len = 0

            for row_idx in range(
                1,
                sheet.max_row + 1
            ):

                value = sheet.cell(
                    row=row_idx,
                    column=col_idx
                ).value

                value = (
                    str(value)
                    if value is not None
                    else ""
                )

                max_len = max(
                    max_len,
                    len(value)
                )

            width = max(
                max_len + 4,
                12
            )

            if col_idx == 1:
                width = max(
                    width,
                    14
                )

            elif col_idx == 2:
                width = max(
                    width,
                    22
                )

            width = min(
                width,
                35
            )

            sheet.column_dimensions[
                col_letter
            ].width = width

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=(
                        cell.alignment.horizontal
                        or "center"
                    ),
                    vertical="center",
                    wrap_text=True
                )

        for row_idx in range(
            1,
            sheet.max_row + 1
        ):
            sheet.row_dimensions[
                row_idx
            ].height = 24

        sheet.sheet_view.showGridLines = False

    # =========================================================
    # 输出 Excel
    # =========================================================
    output = BytesIO()

    wb.save(output)
    output.seek(0)

    filename = (
        f"蕉赖佛学班出席年报_{year}.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

def build_homework_tracking_excel(
    report_type,
    year,
    month=None
):
    """生成白话佛法／经文功课追踪 Excel。"""

    if report_type == "baihua":
        report_title = "白话佛法功课追踪表"
        status_field = "baihua_status"
        full_statuses = {"submitted_done", "submitted"}
        partial_statuses = {"submitted_no_answer"}
        full_label = "有交"
        partial_label = "有交没做题"
        legend_full = "✓ = 有交"
        legend_partial = "○ = 有交，没做题"

    elif report_type == "scripture":
        report_title = "经文功课追踪表"
        status_field = "scripture_status"
        full_statuses = {"submitted_recited", "submitted"}
        partial_statuses = {"submitted_not_recited"}
        full_label = "有交有念"
        partial_label = "有交没念"
        legend_full = "✓ = 有交，有念"
        legend_partial = "○ = 有交，没念"

    else:
        raise ValueError("不正确的功课报表类型")

    if month:
        period_text = f"{year}-{month:02d}"
        homework_date_condition = """
            extract(year from h.class_date) = %s
            and extract(month from h.class_date) = %s
        """
        attendance_date_condition = """
            extract(year from a.class_date) = %s
            and extract(month from a.class_date) = %s
        """
        date_params = (int(year), int(month))
    else:
        period_text = str(year)
        homework_date_condition = """
            extract(year from h.class_date) = %s
        """
        attendance_date_condition = """
            extract(year from a.class_date) = %s
        """
        date_params = (int(year),)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name, sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            cur.execute("""
                select
                    s.id,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.sort_order
                from dharma_students s
                join dharma_class_groups g on g.id = s.group_id
                where s.status = 'active'
                  and s.branch = 'CHE'
                  and g.is_active = true
                order by
                    g.sort_order,
                    s.name
            """)
            students = cur.fetchall()

            cur.execute(f"""
                select
                    h.student_id,
                    h.group_id,
                    h.class_date,
                    h.{status_field} as homework_status
                from dharma_homework h
                where h.branch = 'CHE'
                  and {homework_date_condition}
                order by h.class_date, h.student_id
            """, date_params)
            homework_rows = cur.fetchall()

            # 以点名日期为课程日期，才能显示“未记录”。
            cur.execute(f"""
                select distinct
                    a.group_id,
                    a.class_date
                from dharma_attendance a
                where a.branch = 'CHE'
                  and {attendance_date_condition}
                order by a.class_date
            """, date_params)
            attendance_dates = cur.fetchall()

    homework_map = {
        (row["student_id"], row["class_date"]): row["homework_status"]
        for row in homework_rows
    }

    group_dates_map = {group["id"]: [] for group in groups}

    for row in attendance_dates:
        group_dates_map.setdefault(row["group_id"], [])
        if row["class_date"] not in group_dates_map[row["group_id"]]:
            group_dates_map[row["group_id"]].append(row["class_date"])

    # 若旧功课资料没有对应点名记录，也保留其日期。
    for row in homework_rows:
        group_dates_map.setdefault(row["group_id"], [])
        if row["class_date"] not in group_dates_map[row["group_id"]]:
            group_dates_map[row["group_id"]].append(row["class_date"])

    for group_id in group_dates_map:
        group_dates_map[group_id].sort()

    wb = Workbook()
    wb.remove(wb.active)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    full_fill = PatternFill("solid", fgColor="C6EFCE")
    partial_fill = PatternFill("solid", fgColor="FFF2CC")
    missing_fill = PatternFill("solid", fgColor="F4CCCC")
    absent_fill = PatternFill("solid", fgColor="D9E2F3")
    unknown_fill = PatternFill("solid", fgColor="E7E6E6")
    summary_fill = PatternFill("solid", fgColor="D9EAF7")

    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=18, color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for group in groups:
        ws = wb.create_sheet(group["name"][:31])

        group_students = [
            student for student in students
            if student["group_id"] == group["id"]
        ]
        class_dates = group_dates_map.get(group["id"], [])

        headers = ["学生姓名"]
        headers += [f"{d.month}/{d.day}" for d in class_dates]
        headers += [
            full_label,
            partial_label,
            "没交",
            "缺席",
            "未记录",
            "完整完成率",
        ]

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(len(headers), 8)
        )
        ws["A1"] = f"蕉赖佛学班 {group['name']} {report_title} {period_text}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 36

        ws["A2"] = "说明"
        ws["A2"].font = bold_font
        ws["B2"] = legend_full
        ws["C2"] = legend_partial
        ws["D2"] = "✗ = 没交"
        ws["E2"] = "缺 = 缺席"
        ws["F2"] = "空白 = 未记录"

        ws["B2"].fill = full_fill
        ws["C2"].fill = partial_fill
        ws["D2"].fill = missing_fill
        ws["E2"].fill = absent_fill
        ws["F2"].fill = unknown_fill

        for cell in ws[2]:
            cell.alignment = center

        ws.append(headers)
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        if not group_students:
            ws.append(["", "目前没有学生资料"])

        for student in group_students:
            row_data = [student["name"] or ""]

            full_count = 0
            partial_count = 0
            missing_count = 0
            absent_count = 0
            unrecorded_count = 0

            for class_date in class_dates:
                status = homework_map.get((student["id"], class_date))

                if status in full_statuses:
                    value = "✓"
                    full_count += 1
                elif status in partial_statuses:
                    value = "○"
                    partial_count += 1
                elif status == "missing":
                    value = "✗"
                    missing_count += 1
                elif status == "absent":
                    value = "缺"
                    absent_count += 1
                else:
                    value = ""
                    unrecorded_count += 1

                row_data.append(value)

            effective_total = full_count + partial_count + missing_count
            completion_rate = (
                full_count / effective_total
                if effective_total > 0
                else None
            )

            row_data += [
                full_count,
                partial_count,
                missing_count,
                absent_count,
                unrecorded_count,
                completion_rate,
            ]
            ws.append(row_data)

        date_start_col = 3
        date_end_col = 2 + len(class_dates)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = center

            for col_idx in range(date_start_col, date_end_col + 1):
                cell = ws.cell(row=row[0].row, column=col_idx)
                if cell.value == "✓":
                    cell.fill = full_fill
                    cell.font = bold_font
                elif cell.value == "○":
                    cell.fill = partial_fill
                    cell.font = bold_font
                elif cell.value == "✗":
                    cell.fill = missing_fill
                    cell.font = bold_font
                elif cell.value == "缺":
                    cell.fill = absent_fill
                    cell.font = bold_font
                else:
                    cell.fill = unknown_fill

            rate_cell = row[-1]
            if isinstance(rate_cell.value, (int, float)):
                rate_cell.number_format = "0%"
                if rate_cell.value >= 0.9:
                    rate_cell.fill = full_fill
                elif rate_cell.value >= 0.75:
                    rate_cell.fill = partial_fill
                else:
                    rate_cell.fill = missing_fill

        if group_students:
            total_row = ws.max_row + 1
            ws.merge_cells(
                start_row=total_row,
                start_column=1,
                end_row=total_row,
                end_column=2
            )
            ws.cell(total_row, 1, "合计")

            first_summary_col = 3 + len(class_dates)
            last_student_row = total_row - 1

            for col_idx in range(first_summary_col, first_summary_col + 5):
                letter = get_column_letter(col_idx)
                ws.cell(
                    total_row,
                    col_idx,
                    f"=SUM({letter}4:{letter}{last_student_row})"
                )

            full_col = first_summary_col
            partial_col = first_summary_col + 1
            missing_col = first_summary_col + 2
            rate_col = first_summary_col + 5

            full_letter = get_column_letter(full_col)
            partial_letter = get_column_letter(partial_col)
            missing_letter = get_column_letter(missing_col)

            ws.cell(
                total_row,
                rate_col,
                f'=IF({full_letter}{total_row}+{partial_letter}{total_row}+'
                f'{missing_letter}{total_row}=0,"",'
                f'{full_letter}{total_row}/('
                f'{full_letter}{total_row}+{partial_letter}{total_row}+'
                f'{missing_letter}{total_row}))'
            )
            ws.cell(total_row, rate_col).number_format = "0%"

            for cell in ws[total_row]:
                cell.border = border
                cell.alignment = center
                cell.font = bold_font
                cell.fill = summary_fill

        ws.freeze_panes = "C4"
        if ws.max_row >= 4:
            ws.auto_filter.ref = (
                f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
            )

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22

        for col_idx in range(3, 3 + len(class_dates)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 9

        for col_idx in range(3 + len(class_dates), ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 25

        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:3"
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    if not wb.sheetnames:
        ws = wb.create_sheet("提示")
        ws["A1"] = "目前没有启用中的组别"

    return wb


@dharma_class_bp.route("/reports/baihua/export")
def class_export_baihua_homework():

    year = request.args.get("year") or date.today().strftime("%Y")
    month = request.args.get("month")

    try:
        year_int = int(year)
    except (TypeError, ValueError):
        year_int = date.today().year

    try:
        month_int = int(month) if month else None

        if month_int is not None and not 1 <= month_int <= 12:
            month_int = None

    except (TypeError, ValueError):
        month_int = None

    wb = build_homework_tracking_excel(
        report_type="baihua",
        year=year_int,
        month=month_int
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if month_int:
        filename = (
            f"蕉赖佛学班白话佛法功课追踪表_"
            f"{year_int}-{month_int:02d}.xlsx"
        )
    else:
        filename = (
            f"蕉赖佛学班白话佛法功课追踪表_"
            f"{year_int}.xlsx"
        )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@dharma_class_bp.route("/reports/scripture/export")
def class_export_scripture_homework():

    year = request.args.get("year") or date.today().strftime("%Y")
    month = request.args.get("month")

    try:
        year_int = int(year)
    except (TypeError, ValueError):
        year_int = date.today().year

    try:
        month_int = int(month) if month else None

        if month_int is not None and not 1 <= month_int <= 12:
            month_int = None

    except (TypeError, ValueError):
        month_int = None

    wb = build_homework_tracking_excel(
        report_type="scripture",
        year=year_int,
        month=month_int
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if month_int:
        filename = (
            f"蕉赖佛学班经文功课追踪表_"
            f"{year_int}-{month_int:02d}.xlsx"
        )
    else:
        filename = (
            f"蕉赖佛学班经文功课追踪表_"
            f"{year_int}.xlsx"
        )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

@dharma_class_bp.route("/promote", methods=["GET", "POST"])
def class_promote():

    target_year = request.values.get("target_year") or str(date.today().year + 1)

    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                select id, name, sort_order
                from dharma_class_groups
                where is_active = true
                order by sort_order, id
            """)
            groups = cur.fetchall()

            cur.execute("""
                select
                    s.id,
                    s.name,
                    s.english_name,
                    s.group_id,
                    g.name as group_name
                from dharma_students s
                join dharma_class_groups g on g.id = s.group_id
                where s.status = 'active'
                order by g.sort_order, s.name
            """)
            students = cur.fetchall()

            if request.method == "POST":

                updated = 0
                inactive = 0

                for s in students:
                    new_value = request.form.get(f"new_group_{s['id']}")

                    if not new_value:
                        continue

                    if new_value == "same":
                        continue

                    if new_value == "inactive":
                        cur.execute("""
                            update dharma_students
                            set status = 'inactive'
                            where id = %s
                        """, (s["id"],))
                        inactive += 1
                        continue

                    new_group_id = int(new_value)

                    if new_group_id != s["group_id"]:
                        cur.execute("""
                            update dharma_students
                            set group_id = %s
                            where id = %s
                        """, (new_group_id, s["id"]))
                        updated += 1

                conn.commit()

                flash(
                    f"{target_year} 年度升班完成：更新组别 {updated} 位，暂停 {inactive} 位。",
                    "good"
                )
                return redirect(url_for("dharma_class.class_students"))

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>年度升班</title>
<link rel="stylesheet" href="/static/css/toolbox.css">
</head>

<body>
<div class="page">

    <div class="card">
        <h1 class="page-title">🎓 年度升班</h1>
        <p class="page-subtitle">
            老师可逐位选择新组别。没有要更动的学生，保留「保持原组」即可。
        </p>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, msg in messages %}
                    <div class="alert">{{ msg }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <form method="get">
            <div class="form-group">
                <label class="form-label">目标学年</label>
                <input class="form-input" name="target_year" value="{{ target_year }}">
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-primary" type="submit">
                    🔍 查看学生
                </button>
            </div>
        </form>
    </div>

    <div class="card">
        <h2 class="section-title">升班名单</h2>

        {% if students %}
        <form method="post">
            <input type="hidden" name="target_year" value="{{ target_year }}">

            <div class="table-responsive">
                <table class="record-table">
                    <thead>
                        <tr>
                            <th>姓名</th>
                            <th>英文名</th>
                            <th>目前组别</th>
                            <th>新组别</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for s in students %}
                        <tr>
                            <td>{{ s.name }}</td>
                            <td>{{ s.english_name or "-" }}</td>
                            <td>{{ s.group_name }}</td>
                            <td>
                                <select class="form-input" name="new_group_{{ s.id }}">
                                    <option value="same">保持原组</option>

                                    {% for g in groups %}
                                        <option value="{{ g.id }}"
                                            {% if
                                                (s.group_name == '低年组' and g.name == '少年组') or
                                                (s.group_name == '少年组' and g.name == '高年组')
                                            %}
                                                selected
                                            {% endif %}
                                        >
                                            {{ g.name }}
                                        </option>
                                    {% endfor %}

                                    <option value="inactive">暂停学生</option>
                                </select>
                            </td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>

            <div class="btn-row">
                <button class="btn-tool btn-danger"
                        type="submit"
                        onclick="return confirm('确定要执行年度升班吗？请确认每位学生的新组别。');">
                    ✅ 确认执行升班
                </button>
            </div>
        </form>

        {% else %}
            <div class="empty-state">
                目前没有在册学生。
            </div>
        {% endif %}

        <div class="btn-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('dharma_class.class_admin') }}"
            >
                ⬅ 返回负责人中心
            </a>
        </div>
    </div>

</div>
</body>
</html>
""",
        target_year=target_year,
        groups=groups,
        students=students
    )
