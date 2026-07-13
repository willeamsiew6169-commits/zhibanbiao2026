# finance_web.py

import os
import re

from flask import Blueprint, request, redirect, url_for, render_template_string, send_file
from psycopg2.extras import RealDictCursor
from datetime import date
from openpyxl import Workbook
from db import db_query
from utils import normalize_member_id
from flask import send_file, request
from openpyxl import Workbook
from flask import session
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

finance_bp = Blueprint("finance", __name__, url_prefix="/finance")

FINANCE_PIN = "1234"

SPECIAL_DONATION_TITLE = "813交流会财布施"

INCOME_CATEGORIES = [
    "月费",
    "财布施",
    "观音村",
    "膳食结缘",
    "观音堂纯檀香布施",
    SPECIAL_DONATION_TITLE,
    "临时特别布施"
]

FINANCE_STYLE = """
<style>
body{
    font-family: Microsoft JhengHei, Arial;
    font-size:22px;
    padding:25px;
}

h1{
    font-size:42px;
    margin-bottom:25px;
}

label{
    font-size:24px;
    font-weight:bold;
}

input,
select,
textarea{
    font-size:22px;
    padding:10px;
    min-height:50px;
    min-width:260px;
}

button,
input[type="submit"]{
    font-size:22px;
    padding:14px 28px;
    border-radius:10px;
    cursor:pointer;
}

a{
    font-size:22px;
}

table{
    font-size:20px;
    border-collapse:collapse;
    margin-top:20px;
}

th,td{
    padding:12px;
}

.card{
    padding:25px;
    border-radius:12px;
    box-shadow:0 2px 8px rgba(0,0,0,.15);
    margin-bottom:25px;
}

form{
    line-height:2.2;
}
</style>
"""

FINANCE_V5_STYLE = """
<style>
*{
    box-sizing:border-box;
}

body{
    margin:0;
    background:#f4f7fb;
    color:#1f2937;
    font-family:"Microsoft JhengHei","Microsoft YaHei",Arial,sans-serif;
}

.finance-v5{
    max-width:900px;
    margin:0 auto;
    padding:28px 20px 50px;
}

.v5-header{
    background:linear-gradient(135deg,#1769aa,#2387c9);
    color:#fff;
    padding:28px;
    border-radius:22px;
    margin-bottom:24px;
    box-shadow:0 8px 24px rgba(23,105,170,.20);
}

.v5-header h1{
    margin:0;
    font-size:36px;
}

.v5-header p{
    margin:8px 0 0;
    font-size:17px;
    opacity:.9;
}

.v5-topbar{
    display:flex;
    justify-content:space-between;
    align-items:center;
    gap:12px;
    margin-bottom:18px;
}

.v5-back,
.v5-logout{
    display:inline-flex;
    align-items:center;
    justify-content:center;
    min-height:44px;
    padding:10px 18px;
    border-radius:12px;
    text-decoration:none;
    font-weight:bold;
    font-size:17px;
}

.v5-back{
    color:#1769aa;
    background:#eaf4ff;
}

.v5-logout{
    color:#fff;
    background:#d94343;
}

.v5-menu-grid{
    display:grid;
    grid-template-columns:repeat(2,minmax(0,1fr));
    gap:16px;
}

.v5-menu-btn{
    display:flex;
    align-items:center;
    gap:16px;
    min-height:94px;
    padding:20px;
    background:#fff;
    border:1px solid #e5e7eb;
    border-radius:18px;
    text-decoration:none;
    color:#1f2937;
    box-shadow:0 4px 16px rgba(0,0,0,.06);
    transition:.18s ease;
}

.v5-menu-btn:hover{
    transform:translateY(-2px);
    box-shadow:0 8px 22px rgba(0,0,0,.10);
}

.v5-icon{
    width:54px;
    height:54px;
    flex:0 0 54px;
    display:flex;
    align-items:center;
    justify-content:center;
    border-radius:15px;
    font-size:30px;
    background:#eef6ff;
}

.v5-menu-text{
    min-width:0;
}

.v5-menu-title{
    font-size:22px;
    font-weight:800;
    margin-bottom:5px;
}

.v5-menu-desc{
    color:#667085;
    font-size:15px;
    line-height:1.45;
}

.v5-section-title{
    margin:26px 0 12px;
    font-size:18px;
    font-weight:800;
    color:#475467;
}

.v5-alert{
    background:#fff6e5;
    border-left:6px solid #f59e0b;
}

.v5-income{
    background:#edf9f1;
}

.v5-expense{
    background:#fff1f2;
}

.v5-member{
    background:#eef5ff;
}

.v5-report{
    background:#f7f2ff;
}

@media(max-width:680px){
    .finance-v5{
        padding:16px 12px 35px;
    }

    .v5-header{
        padding:22px 18px;
        border-radius:17px;
    }

    .v5-header h1{
        font-size:29px;
    }

    .v5-menu-grid{
        grid-template-columns:1fr;
        gap:12px;
    }

    .v5-menu-btn{
        min-height:82px;
        padding:16px;
    }

    .v5-menu-title{
        font-size:20px;
    }
}
</style>
"""


def money(v):
    try:
        return float(v or 0)
    except:
        return 0
    
def get_fund_account(category, record_type="income"):
    if record_type == "expense":
        return "观音堂日常户口"

    if category == "月费":
        return "观音堂日常户口"

    return "总会户口"
    
def add_months_ym(ym, months):
    y, m = map(int, ym.split("-"))
    m += months
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return f"{y:04d}-{m:02d}"

def date_to_ym(d):
    if not d:
        return ""
    return d.strftime("%Y-%m")

def next_month_ym(d):
    if not d:
        return date.today().strftime("%Y-%m")
    return add_months_ym(d.strftime("%Y-%m"), 1)

def calc_month_count(amount):
    if amount <= 0:
        return 1
    return max(1, round(amount / 50))

def get_next_receipt_no_by_category(category):
    row = db_query("""
        select receipt_no
        from finance_records
        where category = %s
          and receipt_no is not null
          and receipt_no <> ''
        order by id desc
        limit 1
    """, (category,), fetchone=True)

    if not row or not row["receipt_no"]:
        return ""

    last_no = row["receipt_no"].strip()

    match = re.match(r"^([A-Za-z]+)(\d+)$", last_no)

    if not match:
        return ""

    prefix = match.group(1)
    number = match.group(2)

    next_number = str(int(number) + 1).zfill(len(number))

    return prefix + next_number

def get_next_receipt_no(prefix="CHE"):

    row = db_query("""
        select receipt_no
        from finance_records
        where receipt_no like %s
          and receipt_no is not null
          and receipt_no <> ''
        order by receipt_no desc
        limit 1
    """, (prefix + "%",), fetchone=True)

    if not row or not row["receipt_no"]:
        return f"{prefix}0000001"

    last_no = row["receipt_no"].strip().upper()

    m = re.match(rf"^({prefix})(\d+)$", last_no)

    if not m:
        return f"{prefix}0000001"

    number = m.group(2)

    return prefix + str(int(number) + 1).zfill(len(number))

@finance_bp.route("/login", methods=["GET", "POST"])
def finance_login():

    error = ""

    if request.method == "POST":

        pin = request.form.get("pin", "").strip()

        if pin == FINANCE_PIN:
            session["finance_login"] = True
            return redirect(url_for("finance.finance_home"))

        error = "财政 PIN 不正确，请重新输入。"

    return render_template_string("""
<!doctype html>
<html lang="zh">
<head>
    <meta charset="utf-8">

    <meta
        name="viewport"
        content="width=device-width, initial-scale=1"
    >

    <title>财政系统登入</title>

    <link
        rel="stylesheet"
        href="{{ url_for('static', filename='css/toolbox.css') }}"
    >

    <style>
        .finance-login-page{
            max-width:680px;
        }

        .finance-login-card{
            margin-top:22px;
        }

        .finance-login-icon{
            font-size:52px;
            text-align:center;
            margin-bottom:10px;
        }

        .finance-login-title{
            text-align:center;
            margin-bottom:8px;
        }

        .finance-login-subtitle{
            text-align:center;
            color:#667085;
            margin-bottom:26px;
        }

        .finance-pin-input{
            text-align:center;
            font-size:28px;
            letter-spacing:8px;
        }

        .finance-login-actions{
            display:grid;
            gap:12px;
            margin-top:22px;
        }

        .finance-login-actions .btn-tool{
            width:100%;
        }

        @media(max-width:600px){
            .finance-login-page{
                padding:16px;
            }

            .finance-login-icon{
                font-size:44px;
            }
        }
    </style>
</head>

<body>

<div class="page finance-login-page">

    <div class="card finance-login-card">

        <div class="finance-login-icon">
            🏦
        </div>

        <h1 class="page-title finance-login-title">
            财政系统
        </h1>

        <p class="page-subtitle finance-login-subtitle">
            Finance Management V5
        </p>

        {% if error %}
            <div class="alert alert-danger">
                {{ error }}
            </div>
        {% endif %}

        <form method="post">

            <div class="form-group">

                <label class="form-label">
                    财政 PIN
                </label>

                <input
                    class="form-input finance-pin-input"
                    name="pin"
                    type="password"
                    inputmode="numeric"
                    autocomplete="new-password"
                    maxlength="8"
                    autofocus
                    required
                >

            </div>

            <div class="finance-login-actions">

                <button
                    class="btn-tool btn-primary"
                    type="submit"
                >
                    进入财政系统
                </button>

                <a
                    class="btn-tool btn-secondary"
                    href="/admin-home"
                >
                    返回管理员首页
                </a>

            </div>

        </form>

    </div>

</div>

</body>
</html>
""",
    error=error
    )

@finance_bp.route("/logout")
def finance_logout():

    session.pop("finance_login", None)

    return redirect(url_for("finance.finance_login"))

@finance_bp.route("/late_members")
def late_members():

    rows = db_query("""
        select
            m.member_id,
            m.name,
            m.phone,
            m.remark,
            max(p.end_month) as paid_until,
            max(p.payment_date) as last_payment_date
        from members m
        left join member_payments p
            on p.member_id = m.member_id
        where coalesce(m.member_status, m.status, '') not in (
            '停供',
            '停止',
            '永久停止',
            '往生',
            '已往生'
        )
        group by
            m.member_id,
            m.name,
            m.phone,
            m.remark
        having
            max(p.end_month) is not null
            and max(p.end_month) < date_trunc(
                'month',
                current_date
            )
        order by
            paid_until asc,
            m.member_id
    """, fetchall=True)

    today = date.today()

    current_month_index = (
        today.year * 12
        + today.month
    )

    green_count = 0
    yellow_count = 0
    red_count = 0
    total_amount = 0

    for r in rows:

        paid_until = r["paid_until"]

        paid_index = (
            paid_until.year * 12
            + paid_until.month
        )

        late_months = max(
            current_month_index - paid_index,
            0
        )

        reference_amount = late_months * 50

        r["late_months"] = late_months
        r["reference_amount"] = reference_amount

        if late_months <= 2:
            r["level"] = "green"
            r["level_text"] = "最近停止"
            green_count += 1

        elif late_months <= 6:
            r["level"] = "yellow"
            r["level_text"] = "一段时间未缴费"
            yellow_count += 1

        else:
            r["level"] = "red"
            r["level_text"] = "较久未缴费"
            red_count += 1

        total_amount += reference_amount

        phone = (
            r["phone"]
            or ""
        ).strip()

        phone_digits = "".join(
            ch
            for ch in phone
            if ch.isdigit()
        )

        if phone_digits.startswith("0"):
            phone_digits = "6" + phone_digits

        if phone_digits:
            r["wa_link"] = (
                "https://wa.me/"
                + phone_digits
            )
        else:
            r["wa_link"] = ""

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>

        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>月费关怀名单</title>

        <link
            rel="stylesheet"
            href="{{ url_for(
                'static',
                filename='css/toolbox.css'
            ) }}"
        >

        <style>

            .care-page {
                max-width: 1450px;
            }

            .care-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;

                box-shadow:
                    0 12px 30px
                    rgba(37, 99, 235, 0.18);
            }

            .care-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .care-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .care-note {
                background: #eff6ff;
                border: 1px solid #bfdbfe;
                color: #1e40af;
                border-radius: 14px;
                padding: 14px 16px;
                margin-bottom: 20px;
                line-height: 1.65;
            }

            .care-summary {
                display: grid;
                grid-template-columns:
                    repeat(5, minmax(0, 1fr));
                gap: 16px;
                margin-bottom: 20px;
            }

            .care-summary .summary-box {
                min-height: 120px;
                text-align: center;

                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
            }

            .summary-icon {
                font-size: 29px;
                margin-bottom: 6px;
            }

            .summary-label {
                color: #64748b;
                font-size: 15px;
                margin-bottom: 6px;
            }

            .summary-value {
                color: #0f172a;
                font-size: 26px;
                font-weight: 800;
            }

            .summary-green {
                background: #f0fdf4;
                border-color: #bbf7d0;
            }

            .summary-yellow {
                background: #fffbeb;
                border-color: #fde68a;
            }

            .summary-red {
                background: #fef2f2;
                border-color: #fecaca;
            }

            .summary-green .summary-value {
                color: #15803d;
            }

            .summary-yellow .summary-value {
                color: #a16207;
            }

            .summary-red .summary-value {
                color: #b91c1c;
            }

            .table-topbar {
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 12px;
                flex-wrap: wrap;
                margin-bottom: 14px;
            }

            .result-text {
                color: #64748b;
                font-size: 16px;
            }

            .care-table {
                min-width: 1280px;
            }

            .care-table td {
                vertical-align: middle;
            }

            .member-id {
                color: #1d4ed8;
                font-weight: 800;
                white-space: nowrap;
            }

            .member-name {
                color: #1e293b;
                font-weight: 700;
                min-width: 90px;
            }

            .phone-cell {
                white-space: nowrap;
            }

            .money-cell {
                color: #15803d;
                font-weight: 800;
                white-space: nowrap;
                text-align: right;
            }

            .date-cell {
                white-space: nowrap;
            }

            .remark-cell {
                min-width: 180px;
                max-width: 320px;
                white-space: normal;
                line-height: 1.5;
                overflow-wrap: anywhere;
            }

            .level-row-green {
                background: #f0fdf4;
            }

            .level-row-yellow {
                background: #fffbeb;
            }

            .level-row-red {
                background: #fef2f2;
            }

            .level-badge {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                padding: 7px 11px;
                border-radius: 999px;
                font-size: 14px;
                font-weight: 800;
                white-space: nowrap;
            }

            .level-green {
                background: #dcfce7;
                color: #166534;
            }

            .level-yellow {
                background: #fef3c7;
                color: #92400e;
            }

            .level-red {
                background: #fee2e2;
                color: #991b1b;
            }

            .wa-button {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                gap: 5px;
                padding: 8px 12px;
                border-radius: 10px;
                background: #dcfce7;
                color: #166534;
                font-weight: 800;
                text-decoration: none;
                white-space: nowrap;
                border: 1px solid #bbf7d0;
            }

            .wa-button:hover {
                background: #bbf7d0;
            }

            .empty-care {
                text-align: center;
                padding: 55px 20px;
                color: #64748b;
            }

            .empty-care-icon {
                font-size: 48px;
                margin-bottom: 10px;
            }

            .empty-care h3 {
                color: #334155;
                margin: 0 0 8px;
            }

            .bottom-actions {
                margin-top: 20px;
            }

            @media (max-width: 1100px) {

                .care-summary {
                    grid-template-columns:
                        repeat(2, minmax(0, 1fr));
                }

            }

            @media (max-width: 700px) {

                .care-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .care-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .care-header h1 {
                    font-size: 26px;
                }

                .care-summary {
                    grid-template-columns: 1fr;
                }

                .care-summary .summary-box {
                    min-height: 100px;
                }

                .bottom-actions .btn-tool {
                    width: 100%;
                }

            }

        </style>

    </head>

    <body>

    <div class="page care-page">

        <div class="care-header">

            <h1>🌿 月费关怀名单</h1>

            <p>
                协助负责人了解会员最近缴费情况，
                方便适时联络与关怀。
            </p>

        </div>

        <div class="care-note">

            ℹ️ 此名单不是追缴名单。
            “参考金额”只按照每月 RM50 计算，
            方便负责人了解大概月份，
            不代表会员必须补缴该金额。

        </div>

        <div class="care-summary">

            <div class="summary-box">

                <div class="summary-icon">
                    👥
                </div>

                <div class="summary-label">
                    总人数
                </div>

                <div class="summary-value">
                    {{ rows|length }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    💰
                </div>

                <div class="summary-label">
                    参考金额
                </div>

                <div class="summary-value">
                    RM {{ "{:,.2f}".format(total_amount) }}
                </div>

            </div>

            <div class="summary-box summary-green">

                <div class="summary-icon">
                    🟢
                </div>

                <div class="summary-label">
                    最近停止
                </div>

                <div class="summary-value">
                    {{ green_count }}
                </div>

            </div>

            <div class="summary-box summary-yellow">

                <div class="summary-icon">
                    🟡
                </div>

                <div class="summary-label">
                    一段时间未缴费
                </div>

                <div class="summary-value">
                    {{ yellow_count }}
                </div>

            </div>

            <div class="summary-box summary-red">

                <div class="summary-icon">
                    🔴
                </div>

                <div class="summary-label">
                    较久未缴费
                </div>

                <div class="summary-value">
                    {{ red_count }}
                </div>

            </div>

        </div>

        <div class="card">

            <div class="table-topbar">

                <div
                    class="section-title"
                    style="margin-bottom:0;"
                >
                    📋 会员关怀资料
                </div>

                <div class="result-text">
                    共
                    <strong>{{ rows|length }}</strong>
                    位会员
                </div>

            </div>

            {% if rows %}

                <div class="table-responsive">

                    <table class="record-table care-table">

                        <thead>

                            <tr>
                                <th>会员编号</th>
                                <th>姓名</th>
                                <th>电话</th>
                                <th>WhatsApp</th>
                                <th>已缴至</th>
                                <th>间隔月份</th>
                                <th>参考金额</th>
                                <th>关怀状态</th>
                                <th>最后付款日期</th>
                                <th>备注</th>
                            </tr>

                        </thead>

                        <tbody>

                            {% for r in rows %}

                                <tr class="
                                    {% if r.level == 'green' %}
                                        level-row-green
                                    {% elif r.level == 'yellow' %}
                                        level-row-yellow
                                    {% else %}
                                        level-row-red
                                    {% endif %}
                                ">

                                    <td>
                                        <span class="member-id">
                                            {{ r.member_id }}
                                        </span>
                                    </td>

                                    <td>
                                        <span class="member-name">
                                            {{ r.name }}
                                        </span>
                                    </td>

                                    <td class="phone-cell">
                                        {{ r.phone or "-" }}
                                    </td>

                                    <td>

                                        {% if r.wa_link %}

                                            <a
                                                class="wa-button"
                                                href="{{ r.wa_link }}"
                                                target="_blank"
                                                rel="noopener noreferrer"
                                            >
                                                💬 打开
                                            </a>

                                        {% else %}

                                            -

                                        {% endif %}

                                    </td>

                                    <td class="date-cell">
                                        {{ r.paid_until.strftime("%Y-%m") }}
                                    </td>

                                    <td>
                                        {{ r.late_months }} 个月
                                    </td>

                                    <td class="money-cell">
                                        RM
                                        {{ "%.2f"|format(
                                            r.reference_amount
                                        ) }}
                                    </td>

                                    <td>

                                        <span class="
                                            level-badge
                                            {% if r.level == 'green' %}
                                                level-green
                                            {% elif r.level == 'yellow' %}
                                                level-yellow
                                            {% else %}
                                                level-red
                                            {% endif %}
                                        ">

                                            {% if r.level == "green" %}
                                                🟢
                                            {% elif r.level == "yellow" %}
                                                🟡
                                            {% else %}
                                                🔴
                                            {% endif %}

                                            {{ r.level_text }}

                                        </span>

                                    </td>

                                    <td class="date-cell">

                                        {% if r.last_payment_date %}
                                            {{ r.last_payment_date }}
                                        {% else %}
                                            -
                                        {% endif %}

                                    </td>

                                    <td class="remark-cell">
                                        {{ r.remark or "-" }}
                                    </td>

                                </tr>

                            {% endfor %}

                        </tbody>

                    </table>

                </div>

            {% else %}

                <div class="empty-care">

                    <div class="empty-care-icon">
                        🌿
                    </div>

                    <h3>目前没有需要关怀的会员</h3>

                    <p>
                        所有在册会员的月费记录都在当前月份内。
                    </p>

                </div>

            {% endif %}

        </div>

        <div class="bottom-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'finance.finance_home'
                ) }}"
            >
                ← 返回财政首页
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        green_count=green_count,
        yellow_count=yellow_count,
        red_count=red_count,
        total_amount=total_amount
    )
    
    
@finance_bp.route("/records/<int:record_id>/delete", methods=["POST"])
def delete_record(record_id):

    db_query("""
        delete from finance_records
        where id = %s
    """, (record_id,))

    return redirect(url_for("finance.records"))

@finance_bp.route("/member/<member_id>/edit",
                  methods=["GET","POST"])
def edit_member(member_id):

    member = db_query("""
        select *
        from members
        where member_id = %s
        limit 1
    """, (member_id,), fetchone=True)

    if not member:
        return "Member not found"

    if request.method == "POST":

        status = request.form.get("status")

        db_query("""
            update members
            set status = %s
            where member_id = %s
        """, (
            status,
            member_id
        ))

        return redirect(
            url_for("finance.member_management")
        )

    return render_template_string(FINANCE_STYLE + """

    <h1>{{ member.member_id }}</h1>

    <p>姓名：{{ member.name }}</p>

    <p>电话：{{ member.phone }}</p>

    <form method="post">

        <p>

            状态：

            <select name="status">

                <option
                {% if member.status == '在册' %}
                selected
                {% endif %}
                >
                在册
                </option>

                <option
                {% if member.status == '停供' %}
                selected
                {% endif %}
                >
                停供
                </option>

                <option
                {% if member.status == '往生' %}
                selected
                {% endif %}
                >
                往生
                </option>

            </select>

        </p>

        <button type="submit">
            保存
        </button>

    </form>

    <p>
        <a href="{{ url_for('finance.member_management') }}">
            返回
        </a>
    </p>

    """, member=member)


@finance_bp.route("/")
def finance_home():

    if not session.get("finance_login"):
        return redirect(url_for("finance.finance_login"))

    return render_template_string(FINANCE_V5_STYLE + """
    <div class="finance-v5">

        <div class="v5-topbar">
            <a class="v5-back" href="/admin-home">
                ← 管理员首页
            </a>

            <a class="v5-logout"
               href="{{ url_for('finance.finance_logout') }}">
                退出
            </a>
        </div>

        <div class="v5-header">
            <h1>💰 财政系统 V5</h1>
            <p>请选择要处理的工作</p>
        </div>

        <div class="v5-menu-grid">

            <a class="v5-menu-btn v5-income"
               href="{{ url_for('finance.finance_income_menu') }}">
                <div class="v5-icon">💵</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">收入录入</div>
                    <div class="v5-menu-desc">
                        月费、财布施、观音村、膳食结缘
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn"
               href="{{ url_for('finance.bank_pending') }}">
                <div class="v5-icon">🏦</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">银行过账</div>
                    <div class="v5-menu-desc">
                        银行转账待确认与补开收条
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-expense"
                href="{{ url_for('finance.finance_expense_menu') }}">

                <div class="v5-icon">🧾</div>

                <div class="v5-menu-text">

                    <div class="v5-menu-title">
                        支出录入
                    </div>

                    <div class="v5-menu-desc">
                        供花、供果、供油、水电、电话网络及其它支出
                    </div>

                </div>

            </a>

            <a class="v5-menu-btn v5-member"
               href="{{ url_for('finance.finance_member_menu') }}">
                <div class="v5-icon">👥</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">会员与月费</div>
                    <div class="v5-menu-desc">
                        会员资料、月费情况与关怀名单
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-report"
               href="{{ url_for('finance.finance_report_menu') }}">
                <div class="v5-icon">📊</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">报表与查询</div>
                    <div class="v5-menu-desc">
                        Dashboard、财政记录及月报
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-alert"
               href="{{ url_for('finance.late_members') }}">
                <div class="v5-icon">🌿</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">月费关怀名单</div>
                    <div class="v5-menu-desc">
                        查看供养间隔与会员月费资料
                    </div>
                </div>
            </a>

        </div>

    </div>
    """)

@finance_bp.route("/menu/income")
def finance_income_menu():

    return render_template_string(FINANCE_V5_STYLE + """
    <div class="finance-v5">

        <div class="v5-topbar">
            <a class="v5-back"
               href="{{ url_for('finance.finance_home') }}">
                ← 返回财政首页
            </a>
        </div>

        <div class="v5-header">
            <h1>💵 收入录入</h1>
            <p>请选择收入项目</p>
        </div>

        <div class="v5-menu-grid">

            <a class="v5-menu-btn v5-member"
               href="{{ url_for('finance.monthly_fee_batch', branch='CHE') }}">
                <div class="v5-icon">C</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">CHE 月费</div>
                    <div class="v5-menu-desc">
                        批量登记 CHE 月费
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-member"
               href="{{ url_for('finance.monthly_fee_batch', branch='STW') }}">
                <div class="v5-icon">S</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">STW 月费</div>
                    <div class="v5-menu-desc">
                        批量登记 STW 月费
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-income"
               href="{{ url_for('finance.income_batch', category='财布施') }}">
                <div class="v5-icon">🙏</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">财布施</div>
                    <div class="v5-menu-desc">
                        批量登记财布施
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-income"
               href="{{ url_for('finance.income_batch', category='观音村') }}">
                <div class="v5-icon">🏡</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">观音村</div>
                    <div class="v5-menu-desc">
                        批量登记观音村收入
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-income"
               href="{{ url_for('finance.income_batch', category='膳食结缘') }}">
                <div class="v5-icon">🥗</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">膳食结缘</div>
                    <div class="v5-menu-desc">
                        批量登记膳食结缘收入
                    </div>
                </div>
            </a>
                                  
            <a class="v5-menu-btn v5-income"
            href="{{ url_for(
                'finance.income_batch',
                category='观音堂纯檀香布施'
            ) }}">
                <div class="v5-icon">🪵</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">
                        观音堂纯檀香布施
                    </div>
                    <div class="v5-menu-desc">
                        批量登记纯檀香相关布施
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-income"
            href="{{ url_for(
                'finance.income_batch',
                category=special_donation_title
            ) }}">
                <div class="v5-icon">🎊</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">
                        {{ special_donation_title }}
                    </div>
                    <div class="v5-menu-desc">
                        本年度特别活动布施
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-income"
            href="{{ url_for(
                'finance.income_batch',
                category='临时特别布施'
            ) }}">
                <div class="v5-icon">➕</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">
                        临时特别布施
                    </div>
                    <div class="v5-menu-desc">
                        突发活动或临时项目使用
                    </div>
                </div>
            </a>

        </div>

    </div>
    """,
    special_donation_title=SPECIAL_DONATION_TITLE
    )

@finance_bp.route("/menu/expense")
def finance_expense_menu():

    expense_items = [
        ("🌸", "供花", "鲜花及供花相关支出"),
        ("🍎", "供果", "水果及供品相关支出"),
        ("🪔", "供油", "灯油及油品相关支出"),
        ("⚡", "电费", "TNB 电费记录"),
        ("💧", "水费", "水费记录"),
        ("📶", "电话及网络费", "电话、WiFi、Internet"),
        ("🛠️", "维修保养", "维修及保养费用"),
        ("🛒", "日常采购", "文具、厨房及日常用品"),
        ("🧾", "其它支出", "其它杂项费用"),
    ]

    return render_template_string(FINANCE_V5_STYLE + """
    <div class="finance-v5">

        <div class="v5-topbar">
            <a class="v5-back"
               href="{{ url_for('finance.finance_home') }}">
                ← 返回财政首页
            </a>
        </div>

        <div class="v5-header">
            <h1>🧾 支出录入</h1>
            <p>请选择支出项目</p>
        </div>

        <div class="v5-menu-grid">

            {% for icon, category, desc in expense_items %}
            <a class="v5-menu-btn v5-expense"
               href="{{ url_for('finance.expense', category=category) }}">

                <div class="v5-icon">{{ icon }}</div>

                <div class="v5-menu-text">
                    <div class="v5-menu-title">{{ category }}</div>
                    <div class="v5-menu-desc">{{ desc }}</div>
                </div>

            </a>
            {% endfor %}

        </div>

    </div>
    """, expense_items=expense_items)

@finance_bp.route("/menu/member")
def finance_member_menu():

    return render_template_string(FINANCE_V5_STYLE + """
    <div class="finance-v5">

        <div class="v5-topbar">
            <a class="v5-back"
               href="{{ url_for('finance.finance_home') }}">
                ← 返回财政首页
            </a>
        </div>

        <div class="v5-header">
            <h1>👥 会员与月费</h1>
            <p>会员资料与月费情况</p>
        </div>

        <div class="v5-menu-grid">

            <a class="v5-menu-btn v5-member"
               href="{{ url_for('finance.member_management') }}">
                <div class="v5-icon">👤</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">会员管理</div>
                    <div class="v5-menu-desc">
                        搜索、查看和维护会员资料
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-alert"
               href="{{ url_for('finance.late_members') }}">
                <div class="v5-icon">🌿</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">月费关怀名单</div>
                    <div class="v5-menu-desc">
                        查看已缴至、供养间隔与参考资料
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-member"
               href="{{ url_for('finance.add_member') }}">
                <div class="v5-icon">➕</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">新增会员</div>
                    <div class="v5-menu-desc">
                        建立新月费会员资料
                    </div>
                </div>
            </a>

        </div>

    </div>
    """)

@finance_bp.route("/menu/report")
def finance_report_menu():

    today_ym = date.today().strftime("%Y-%m")

    return render_template_string(FINANCE_V5_STYLE + """
    <div class="finance-v5">

        <div class="v5-topbar">
            <a class="v5-back"
               href="{{ url_for('finance.finance_home') }}">
                ← 返回财政首页
            </a>
        </div>

        <div class="v5-header">
            <h1>📊 报表与查询</h1>
            <p>统计、记录与 Excel 月报</p>
        </div>

        <div class="v5-menu-grid">

            <a class="v5-menu-btn v5-report"
               href="{{ url_for('finance.dashboard') }}">
                <div class="v5-icon">📈</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">财政 Dashboard</div>
                    <div class="v5-menu-desc">
                        查看每月收入、支出和户口统计
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-report"
               href="{{ url_for('finance.records') }}">
                <div class="v5-icon">🔎</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">财政记录搜索</div>
                    <div class="v5-menu-desc">
                        搜索收条、会员编号、姓名与银行 Reference
                    </div>
                </div>
            </a>

            <a class="v5-menu-btn v5-report"
               href="{{ url_for('finance.export_monthly_report', ym=today_ym) }}">
                <div class="v5-icon">📥</div>
                <div class="v5-menu-text">
                    <div class="v5-menu-title">下载专业版月报</div>
                    <div class="v5-menu-desc">
                        下载 {{ today_ym }} Excel 财政月报
                    </div>
                </div>
            </a>

        </div>

    </div>
    """, today_ym=today_ym)

@finance_bp.route("/member_management")
def member_management():

    q = request.args.get("q", "").strip()

    if q:
        keyword = f"%{q}%"

        rows = db_query("""
            select
                member_id,
                name,
                english_name,
                phone,
                coalesce(status, '在册') as status,
                remark
            from members
            where
                member_id ilike %s
                or name ilike %s
                or english_name ilike %s
                or phone ilike %s
            order by member_id
            limit 300
        """, (
            keyword,
            keyword,
            keyword,
            keyword
        ), fetchall=True)

    else:
        rows = db_query("""
            select
                member_id,
                name,
                english_name,
                phone,
                coalesce(status, '在册') as status,
                remark
            from members
            order by member_id
            limit 300
        """, fetchall=True)

    total_members = len(rows)

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>月费会员管理</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .member-page {
                max-width: 1100px;
            }

            .member-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                border-radius: 22px;
                padding: 26px;
                margin-bottom: 20px;
                box-shadow: 0 12px 30px rgba(37, 99, 235, 0.18);
            }

            .member-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .member-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .member-toolbar {
                display: grid;
                grid-template-columns: 1fr auto;
                gap: 14px;
                align-items: end;
            }

            .member-search-row {
                display: grid;
                grid-template-columns: 1fr auto auto;
                gap: 10px;
                align-items: center;
            }

            .member-search-row .form-input {
                margin: 0;
            }

            .member-count {
                font-size: 17px;
                color: #475569;
                margin-top: 10px;
            }

            .member-name {
                font-weight: 700;
                color: #1e293b;
            }

            .member-id {
                font-weight: 700;
                color: #1d4ed8;
                white-space: nowrap;
            }

            .member-phone {
                white-space: nowrap;
            }

            .status-badge {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                min-width: 72px;
                padding: 6px 10px;
                border-radius: 999px;
                font-size: 15px;
                font-weight: 700;
                white-space: nowrap;
            }

            .status-active {
                background: #dcfce7;
                color: #166534;
            }

            .status-paused {
                background: #fef3c7;
                color: #92400e;
            }

            .status-stopped {
                background: #fee2e2;
                color: #991b1b;
            }

            .status-default {
                background: #e2e8f0;
                color: #334155;
            }

            .table-actions {
                display: flex;
                gap: 8px;
                flex-wrap: wrap;
            }

            .table-actions .btn-tool {
                min-height: auto;
                padding: 8px 13px;
                font-size: 15px;
            }

            .member-empty {
                text-align: center;
                padding: 45px 20px;
                color: #64748b;
            }

            @media (max-width: 760px) {

                .member-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .member-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .member-header h1 {
                    font-size: 26px;
                }

                .member-toolbar {
                    grid-template-columns: 1fr;
                }

                .member-search-row {
                    grid-template-columns: 1fr;
                }

                .member-toolbar .btn-tool,
                .member-search-row .btn-tool {
                    width: 100%;
                }

                .record-table {
                    min-width: 880px;
                }
            }
        </style>
    </head>

    <body>

    <div class="page member-page">

        <div class="member-header">
            <h1>👥 月费会员管理</h1>

            <p>
                查询、新增及编辑 CHE 与 STW 月费会员资料。
            </p>
        </div>

        <div class="card">

            <div class="member-toolbar">

                <div>
                    <div class="section-title">
                        🔍 搜索会员
                    </div>

                    <form method="get">

                        <div class="member-search-row">

                            <input
                                class="form-input"
                                name="q"
                                value="{{ q }}"
                                placeholder="输入编号、姓名、英文名或电话"
                                autocomplete="off"
                            >

                            <button
                                class="btn-tool btn-primary"
                                type="submit"
                            >
                                🔍 搜索
                            </button>

                            {% if q %}
                            <a
                                class="btn-tool btn-secondary"
                                href="{{ url_for(
                                    'finance.member_management'
                                ) }}"
                            >
                                ✕ 清除
                            </a>
                            {% endif %}

                        </div>

                    </form>
                </div>

                <a
                    class="btn-tool btn-success"
                    href="{{ url_for('finance.add_member') }}"
                >
                    ➕ 新增会员
                </a>

            </div>

            <div class="member-count">
                {% if q %}
                    搜索结果：
                    <strong>{{ total_members }}</strong>
                    位会员
                {% else %}
                    当前显示：
                    <strong>{{ total_members }}</strong>
                    位会员
                {% endif %}
            </div>

        </div>

        <div class="card">

            <div class="section-title">
                📋 会员名单
            </div>

            {% if rows %}

            <div class="table-responsive">

                <table class="record-table">

                    <thead>
                        <tr>
                            <th>会员编号</th>
                            <th>姓名</th>
                            <th>英文名</th>
                            <th>电话</th>
                            <th>会员状态</th>
                            <th>备注</th>
                            <th>操作</th>
                        </tr>
                    </thead>

                    <tbody>

                        {% for r in rows %}

                        {% set status_text = r.status or '在册' %}

                        <tr>

                            <td>
                                <span class="member-id">
                                    {{ r.member_id }}
                                </span>
                            </td>

                            <td>
                                <span class="member-name">
                                    {{ r.name }}
                                </span>
                            </td>

                            <td>
                                {{ r.english_name or "-" }}
                            </td>

                            <td class="member-phone">
                                {% if r.phone %}
                                    <a href="tel:{{ r.phone }}">
                                        {{ r.phone }}
                                    </a>
                                {% else %}
                                    -
                                {% endif %}
                            </td>

                            <td>

                                {% if status_text in [
                                    '在册',
                                    'active',
                                    '正常',
                                    '供养中'
                                ] %}

                                    <span
                                        class="
                                            status-badge
                                            status-active
                                        "
                                    >
                                        在册
                                    </span>

                                {% elif status_text in [
                                    '暂停',
                                    'paused',
                                    '停供'
                                ] %}

                                    <span
                                        class="
                                            status-badge
                                            status-paused
                                        "
                                    >
                                        暂停
                                    </span>

                                {% elif status_text in [
                                    '停止',
                                    'stopped',
                                    'inactive',
                                    '永久停止'
                                ] %}

                                    <span
                                        class="
                                            status-badge
                                            status-stopped
                                        "
                                    >
                                        停止
                                    </span>

                                {% else %}

                                    <span
                                        class="
                                            status-badge
                                            status-default
                                        "
                                    >
                                        {{ status_text }}
                                    </span>

                                {% endif %}

                            </td>

                            <td>
                                {{ r.remark or "-" }}
                            </td>

                            <td>

                                <div class="table-actions">

                                    <a
                                        class="
                                            btn-tool
                                            btn-primary
                                        "
                                        href="{{ url_for(
                                            'finance.edit_member',
                                            member_id=r.member_id
                                        ) }}"
                                    >
                                        ✏️ 编辑
                                    </a>

                                </div>

                            </td>

                        </tr>

                        {% endfor %}

                    </tbody>

                </table>

            </div>

            {% else %}

            <div class="member-empty">
                <div style="font-size:44px;">
                    🔎
                </div>

                <h3>找不到相关会员</h3>

                <p>
                    请检查会员编号、姓名、英文名或电话号码。
                </p>
            </div>

            {% endif %}

        </div>

        <div class="btn-row">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_home') }}"
            >
                ← 返回财政首页
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        q=q,
        total_members=total_members
    )

@finance_bp.route("/member/<member_id>/status", methods=["POST"])
def update_member_status(member_id):

    status = request.form.get("status", "").strip()

    allowed_status = ["在册", "停供", "往生"]

    if status not in allowed_status:
        return "Invalid status", 400

    db_query("""
        update members
        set status = %s
        where member_id = %s
    """, (status, member_id))

    return redirect(url_for("finance.member_management"))

@finance_bp.route("/add_member", methods=["GET","POST"])
def add_member():
    return "Add Member Coming Soon"

@finance_bp.route("/monthly_fee_batch/<branch>", methods=["GET", "POST"])
def monthly_fee_batch(branch):

    branch = branch.upper()

    if branch not in ["CHE", "STW"]:
        return "Invalid branch", 400

    message = ""
    preview_rows = []

    raw_text = request.form.get("raw_text", "").strip()

    next_receipt_no = get_next_receipt_no(branch)
    next_receipt_raw = next_receipt_no.replace(branch, "", 1)

    receipt_start_raw = request.form.get(
        "receipt_start",
        ""
    ).strip().upper()

    if not receipt_start_raw:
        receipt_start_raw = next_receipt_raw

    if receipt_start_raw.isdigit():
        receipt_start = (
            branch
            + str(int(receipt_start_raw)).zfill(7)
        )
    else:
        receipt_start = receipt_start_raw

    default_receipt_date = (
        request.form.get("receipt_date")
        or date.today().isoformat()
    )

    payment_method = request.form.get(
        "payment_method",
        "现金"
    )

    bank_payment_date = (
        request.form.get("payment_date")
        or default_receipt_date
    )

    action = request.form.get("action", "")

    default_amount = money(
        request.form.get("default_amount")
        or 50
    )

    def make_receipt_no(start_no, index):
        match = re.match(r"^([A-Z]+)(\d+)$", start_no)

        if not match:
            return ""

        prefix = match.group(1)
        number_text = match.group(2)

        return (
            prefix
            + str(int(number_text) + index).zfill(
                len(number_text)
            )
        )

    def parse_date_header(line):
        """
        支持：
        @2026-07-10
        @ 2026-07-10
        """
        match = re.match(
            r"^@\s*(\d{4}-\d{2}-\d{2})$",
            line
        )

        if not match:
            return None

        date_text = match.group(1)

        try:
            return date.fromisoformat(date_text).isoformat()
        except ValueError:
            return "INVALID"

    def build_preview():
        rows = []
        current_receipt_date = default_receipt_date
        receipt_index = 0

        for line_no, original_line in enumerate(
            raw_text.splitlines(),
            start=1
        ):
            line = original_line.strip()

            if not line:
                continue

            if line.startswith("@"):
                parsed_date = parse_date_header(line)

                if parsed_date == "INVALID":
                    rows.append({
                        "error": (
                            f"第 {line_no} 行日期无效：{line}"
                        ),
                        "raw": line,
                        "receipt_date": "",
                    })
                elif not parsed_date:
                    rows.append({
                        "error": (
                            f"第 {line_no} 行日期格式错误，"
                            "请使用 @YYYY-MM-DD"
                        ),
                        "raw": line,
                        "receipt_date": "",
                    })
                else:
                    current_receipt_date = parsed_date

                continue

            parts = line.split()
            raw_member_id = parts[0].strip()

            if len(parts) >= 2:
                amount = money(parts[1])
            else:
                amount = default_amount

            receipt_no = make_receipt_no(
                receipt_start,
                receipt_index
            )
            receipt_index += 1

            if amount <= 0:
                rows.append({
                    "error": (
                        f"第 {line_no} 行金额无效：{line}"
                    ),
                    "raw": line,
                    "receipt_no": receipt_no,
                    "receipt_date": current_receipt_date,
                })
                continue

            if raw_member_id.isdigit():
                member_id = f"{branch}-{int(raw_member_id)}"
            else:
                member_id = normalize_member_id(
                    raw_member_id,
                    default_branch=branch
                )

            member = db_query("""
                select *
                from members
                where member_id = %s
                limit 1
            """, (
                member_id,
            ), fetchone=True)

            if not member:
                rows.append({
                    "error": (
                        f"第 {line_no} 行找不到会员："
                        f"{raw_member_id}"
                    ),
                    "raw": line,
                    "receipt_no": receipt_no,
                    "receipt_date": current_receipt_date,
                })
                continue

            paid = db_query("""
                select max(end_month) as paid_until
                from member_payments
                where member_id = %s
            """, (
                member_id,
            ), fetchone=True)

            paid_until_date = (
                paid["paid_until"]
                if paid
                else None
            )

            month_from = next_month_ym(
                paid_until_date
            )

            month_count = max(
                1,
                round(amount / 50)
            )

            month_to = add_months_ym(
                month_from,
                month_count - 1
            )

            existing = db_query("""
                select id
                from finance_records
                where receipt_no = %s
                limit 1
            """, (
                receipt_no,
            ), fetchone=True)

            if payment_method == "银行过账":
                row_payment_date = bank_payment_date
            else:
                row_payment_date = current_receipt_date

            rows.append({
                "error": (
                    "收条已存在"
                    if existing
                    else ""
                ),
                "receipt_no": receipt_no,
                "receipt_date": current_receipt_date,
                "payment_date": row_payment_date,
                "member_id": member["member_id"],
                "name": (
                    member.get("姓名")
                    or member.get("name")
                ),
                "phone": (
                    member.get("电话号码")
                    or member.get("phone")
                ),
                "amount": amount,
                "month_from": month_from,
                "month_to": month_to,
                "month_count": month_count,
            })

        return rows

    if request.method == "POST":

        if not receipt_start:
            message = "请填写收条开始号码"

        elif not re.match(
            rf"^{branch}\d+$",
            receipt_start
        ):
            message = (
                "收条号码格式错误，例如："
                f"{branch}0001501，或只输入 1501"
            )

        elif not raw_text:
            message = "请输入批量月费资料"

        else:
            preview_rows = build_preview()

            data_rows = [
                row
                for row in preview_rows
                if row.get("member_id")
            ]

            has_error = any(
                row.get("error")
                for row in preview_rows
            )

            if not data_rows and not has_error:
                message = "没有可预览的会员资料"

            elif action == "confirm" and not has_error:

                for row in data_rows:
                    month_from_db = (
                        row["month_from"] + "-01"
                    )

                    month_to_db = (
                        row["month_to"] + "-01"
                    )

                    db_query("""
                        insert into finance_records
                        (
                            record_type,
                            fund_account,
                            record_date,
                            receipt_date,
                            category,
                            receipt_no,
                            member_id,
                            name,
                            phone,
                            amount,
                            payment_method,
                            month_from,
                            month_to,
                            remarks
                        )
                        values
                        (
                            %s, %s, %s, %s,
                            '月费',
                            %s, %s, %s, %s,
                            %s, %s, %s, %s, %s
                        )
                    """, (
                        "income",
                        get_fund_account("月费"),
                        row["payment_date"],
                        row["receipt_date"],
                        row["receipt_no"],
                        row["member_id"],
                        row["name"],
                        row["phone"],
                        row["amount"],
                        payment_method,
                        row["month_from"],
                        row["month_to"],
                        "批量月费录入",
                    ))

                    db_query("""
                        insert into member_payments
                        (
                            payment_date,
                            receipt_date,
                            member_id,
                            name,
                            receipt_no,
                            amount,
                            start_month,
                            end_month,
                            month_count
                        )
                        values
                        (
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s
                        )
                    """, (
                        row["payment_date"],
                        row["receipt_date"],
                        row["member_id"],
                        row["name"],
                        row["receipt_no"],
                        row["amount"],
                        month_from_db,
                        month_to_db,
                        row["month_count"],
                    ))

                return redirect(
                    url_for("finance.records")
                )

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">
        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>{{ branch }} 批量月费录入</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .finance-form-page{
                max-width:980px;
            }

            .finance-header-card{
                margin-bottom:18px;
                background:linear-gradient(135deg,#1769aa,#2589c5);
                color:#fff;
            }

            .finance-header-card .page-title,
            .finance-header-card .page-subtitle{
                color:#fff;
            }

            .finance-back-row{
                margin-bottom:16px;
            }

            .step-badge{
                display:inline-flex;
                align-items:center;
                justify-content:center;
                min-width:34px;
                height:34px;
                padding:0 10px;
                margin-right:8px;
                border-radius:999px;
                background:#1769aa;
                color:#fff;
                font-size:16px;
                font-weight:800;
            }

            .finance-form-grid{
                display:grid;
                grid-template-columns:repeat(2,minmax(0,1fr));
                gap:16px;
            }

            .finance-form-grid .form-group{
                margin:0;
            }

            .finance-full{
                grid-column:1 / -1;
            }

            .receipt-input-row{
                display:flex;
                align-items:center;
                gap:10px;
            }

            .receipt-prefix{
                min-height:52px;
                min-width:76px;
                display:flex;
                align-items:center;
                justify-content:center;
                padding:0 15px;
                border-radius:10px;
                background:#eaf4ff;
                color:#1769aa;
                font-weight:800;
                font-size:20px;
            }

            .receipt-input-row .form-input{
                flex:1;
                min-width:0;
            }

            .field-help{
                display:block;
                color:#667085;
                font-size:15px;
                line-height:1.55;
                margin-top:7px;
            }

            .batch-textarea{
                width:100%;
                min-height:330px;
                resize:vertical;
                line-height:1.75;
                font-family:Consolas,"Microsoft YaHei",monospace;
                font-size:18px;
            }

            .batch-guide-grid{
                display:grid;
                grid-template-columns:repeat(3,minmax(0,1fr));
                gap:12px;
                margin-top:14px;
            }

            .batch-guide-item{
                background:#f8fafc;
                border:1px solid #dbe3ed;
                border-radius:12px;
                padding:13px 15px;
                color:#475467;
                font-size:14px;
                line-height:1.65;
            }

            .batch-guide-item strong{
                display:block;
                margin-bottom:4px;
                color:#172033;
            }

            .batch-guide-item code{
                font-family:Consolas,monospace;
                color:#1769aa;
                font-weight:800;
            }

            .date-note{
                margin-top:14px;
                padding:14px 16px;
                border-radius:12px;
                background:#fff8e8;
                border:1px solid #f1d39b;
                color:#795400;
                line-height:1.65;
            }

            .action-row{
                display:flex;
                flex-wrap:wrap;
                gap:12px;
                margin-top:20px;
            }

            .action-row .btn-tool{
                min-width:180px;
            }

            .preview-success{
                color:#16863a;
                font-weight:700;
            }

            .preview-error{
                color:#c62828;
                font-weight:700;
            }

            .payment-date-box{
                display:none;
            }

            .preview-date{
                white-space:nowrap;
                font-weight:700;
                color:#1769aa;
            }

            @media(max-width:760px){
                .finance-form-grid,
                .batch-guide-grid{
                    grid-template-columns:1fr;
                }

                .finance-full{
                    grid-column:auto;
                }

                .receipt-input-row{
                    align-items:stretch;
                }

                .action-row{
                    display:grid;
                }

                .action-row .btn-tool{
                    width:100%;
                }
            }
        </style>
    </head>

    <body>

    <div class="page finance-form-page">

        <div class="finance-back-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_income_menu') }}"
            >
                ← 返回收入录入
            </a>
        </div>

        <div class="card finance-header-card">
            <h1 class="page-title">
                💳 {{ branch }} 批量月费录入
            </h1>

            <p class="page-subtitle">
                可在同一批资料中切换不同收条日期，系统会自动找姓名、计算月份和排列收条。
            </p>
        </div>

        {% if message %}
            <div class="alert alert-danger">
                {{ message }}
            </div>
        {% endif %}

        <form method="post">

            <div class="card">
                <h2 class="section-title">
                    <span class="step-badge">1</span>
                    输入会员编号、金额与日期
                </h2>

                <p class="page-subtitle">
                    每行一位会员；遇到不同收条日期时，先输入一行
                    <strong>@YYYY-MM-DD</strong>。
                </p>

                <textarea
                    class="form-input batch-textarea"
                    name="raw_text"
                    autofocus
                    placeholder="例如：
@2026-07-10
208
160
188 100

@2026-07-11
69
205 300"
                    required
                >{{ raw_text }}</textarea>

                <div class="batch-guide-grid">
                    <div class="batch-guide-item">
                        <strong>普通月费</strong>
                        <code>208</code><br>
                        使用下方默认金额。
                    </div>

                    <div class="batch-guide-item">
                        <strong>特别金额</strong>
                        <code>188 100</code><br>
                        会员 188，金额 RM100。
                    </div>

                    <div class="batch-guide-item">
                        <strong>切换收条日期</strong>
                        <code>@2026-07-11</code><br>
                        之后的会员使用这个日期。
                    </div>
                </div>

                <div class="date-note">
                    没有写 <strong>@日期</strong> 的会员，会使用下方的“默认开收条日期”。
                    收条号码仍会按输入顺序连续排列。
                </div>
            </div>

            <div class="card">
                <h2 class="section-title">
                    <span class="step-badge">2</span>
                    检查收条与付款资料
                </h2>

                <div class="finance-form-grid">
                    <div class="form-group">
                        <label class="form-label">
                            收条开始号码
                        </label>

                        <div class="receipt-input-row">
                            <div class="receipt-prefix">
                                {{ branch }}
                            </div>

                            <input
                                class="form-input"
                                name="receipt_start"
                                value="{{ receipt_start_raw }}"
                                inputmode="numeric"
                                placeholder="例如 1501"
                                required
                            >
                        </div>

                        <span class="field-help">
                            系统建议下一张收条：
                            <strong>{{ next_receipt_no }}</strong>
                        </span>
                    </div>

                    <div class="form-group">
                        <label class="form-label">
                            默认开收条日期
                        </label>

                        <input
                            class="form-input"
                            name="receipt_date"
                            type="date"
                            value="{{ default_receipt_date }}"
                            required
                        >

                        <span class="field-help">
                            只用于没有写 @日期 的会员。
                        </span>
                    </div>

                    <div class="form-group">
                        <label class="form-label">
                            付款方式
                        </label>

                        <select
                            class="form-input"
                            name="payment_method"
                            id="payment_method"
                            onchange="togglePaymentDate()"
                        >
                            {% for method in ['现金', '银行过账', '支票'] %}
                                <option
                                    value="{{ method }}"
                                    {% if payment_method == method %}
                                        selected
                                    {% endif %}
                                >
                                    {{ method }}
                                </option>
                            {% endfor %}
                        </select>
                    </div>

                    <div
                        class="form-group payment-date-box"
                        id="payment_date_box"
                    >
                        <label class="form-label">
                            银行付款日期
                        </label>

                        <input
                            class="form-input"
                            name="payment_date"
                            type="date"
                            value="{{ bank_payment_date }}"
                        >

                        <span class="field-help">
                            银行过账时整批共用这个付款日期；每笔收条日期仍按 @日期 分组。
                        </span>
                    </div>

                    <div class="form-group finance-full">
                        <label class="form-label">
                            默认月费 RM
                        </label>

                        <input
                            class="form-input"
                            name="default_amount"
                            type="number"
                            step="50"
                            min="1"
                            value="{{ default_amount }}"
                            required
                        >

                        <span class="field-help">
                            只输入会员编号时使用这个金额；编号后面有金额时，以该金额为准。
                        </span>
                    </div>
                </div>

                <div class="action-row">
                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                        name="action"
                        value="preview"
                    >
                        👁️ 预览资料
                    </button>

                    {% if preview_rows and not has_preview_error %}
                        <button
                            class="btn-tool btn-success"
                            type="submit"
                            name="action"
                            value="confirm"
                            onclick="return confirm('确定全部入账？');"
                        >
                            ✅ 确认全部入账
                        </button>
                    {% endif %}
                </div>
            </div>

            {% if preview_rows %}
                <div class="card">
                    <h2 class="section-title">
                        <span class="step-badge">3</span>
                        月费录入预览
                    </h2>

                    <div class="table-responsive">
                        <table class="record-table">
                            <thead>
                                <tr>
                                    <th>收条</th>
                                    <th>收条日期</th>
                                    <th>会员编号</th>
                                    <th>姓名</th>
                                    <th>金额</th>
                                    <th>开始月份</th>
                                    <th>缴费至</th>
                                    <th>月数</th>
                                    <th>检查结果</th>
                                </tr>
                            </thead>

                            <tbody>
                                {% for r in preview_rows %}
                                    <tr>
                                        <td>{{ r.receipt_no or '-' }}</td>

                                        <td class="preview-date">
                                            {{ r.receipt_date or '-' }}
                                        </td>

                                        <td>
                                            <strong>
                                                {{ r.member_id or '-' }}
                                            </strong>
                                        </td>

                                        <td>{{ r.name or '-' }}</td>

                                        <td>
                                            {% if r.amount %}
                                                RM {{ '%.2f'|format(r.amount) }}
                                            {% else %}
                                                -
                                            {% endif %}
                                        </td>

                                        <td>{{ r.month_from or '-' }}</td>
                                        <td>{{ r.month_to or '-' }}</td>

                                        <td>
                                            {% if r.month_count %}
                                                {{ r.month_count }} 个月
                                            {% else %}
                                                -
                                            {% endif %}
                                        </td>

                                        <td>
                                            {% if r.error %}
                                                <span class="preview-error">
                                                    ❌ {{ r.error }}
                                                </span>
                                            {% else %}
                                                <span class="preview-success">
                                                    ✅ 可以入账
                                                </span>
                                            {% endif %}
                                        </td>
                                    </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                </div>
            {% endif %}

        </form>

    </div>

    <script>
    function togglePaymentDate(){
        const method = document.getElementById(
            "payment_method"
        ).value;

        const box = document.getElementById(
            "payment_date_box"
        );

        box.style.display = (
            method === "银行过账"
            ? "block"
            : "none"
        );
    }

    togglePaymentDate();
    </script>

    </body>
    </html>
    """,
        branch=branch,
        message=message,
        raw_text=raw_text,
        receipt_start_raw=receipt_start_raw,
        next_receipt_no=next_receipt_no,
        default_receipt_date=default_receipt_date,
        bank_payment_date=bank_payment_date,
        payment_method=payment_method,
        default_amount=default_amount,
        preview_rows=preview_rows,
        has_preview_error=any(
            row.get("error")
            for row in preview_rows
        ),
    )

def get_recent_donors(category=None):
    if category:
        return db_query("""
            select
                name,
                max(phone) as phone,
                count(*) as times,
                max(record_date) as last_date
            from finance_records
            where record_type = 'income'
              and category = %s
              and coalesce(status, 'confirmed') <> 'cancelled'
              and coalesce(name, '') <> ''
              and category <> '月费'
            group by name
            order by max(record_date) desc, count(*) desc
            limit 100
        """, (category,), fetchall=True)

    return db_query("""
        select
            name,
            max(phone) as phone,
            count(*) as times,
            max(record_date) as last_date
        from finance_records
        where record_type = 'income'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and coalesce(name, '') <> ''
          and category <> '月费'
        group by name
        order by max(record_date) desc, count(*) desc
        limit 100
    """, fetchall=True)
    

@finance_bp.route("/income/<category>", methods=["GET", "POST"])
def normal_income(category):

    allowed_categories = ["财布施", "观音村", "膳食结缘"]

    if category not in allowed_categories:
        return "Invalid category", 400

    receipt_prefix = "CHE"

    last_receipt = db_query("""
        select receipt_no
        from finance_records
        where receipt_no like %s
        order by receipt_no desc
        limit 1
    """, (receipt_prefix + "%",), fetchone=True)

    if last_receipt and last_receipt["receipt_no"]:
        old_no = last_receipt["receipt_no"]
        number = int(old_no[3:])
        next_receipt_no = receipt_prefix + str(number + 1).zfill(len(old_no) - 3)
    else:
        next_receipt_no = "CHE0000001"

    message = ""

    if request.method == "POST":

        receipt_no = request.form.get("receipt_no", "").strip().upper()
        receipt_date = request.form.get("receipt_date") or date.today()
        record_date = request.form.get("record_date") or date.today()

        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        amount = money(request.form.get("amount"))
        payment_method = request.form.get("payment_method", "现金")
        remarks = request.form.get("remarks", "").strip()

        if not receipt_no.startswith("CHE"):
            message = "收条号码必须以 CHE 开头"

        elif amount <= 0:
            message = "金额必须大过 0"

        else:
            existing = db_query("""
                select id
                from finance_records
                where receipt_no = %s
                limit 1
            """, (receipt_no,), fetchone=True)

            if existing:
                message = "这个收条号码已经记录过了，请检查是否重复输入"
            else:
                db_query("""
                    insert into finance_records
                    (
                        record_type,
                        fund_account,
                        record_date,
                        receipt_date,
                        category,
                        receipt_no,
                        name,
                        phone,
                        amount,
                        payment_method,
                        remarks
                    )
                    values
                    (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s
                    )
                """, (
                    "income",
                    get_fund_account(category),
                    record_date,
                    receipt_date,
                    category,
                    receipt_no,
                    name,
                    phone,
                    amount,
                    payment_method,
                    remarks
                ))

                return redirect(url_for("finance.records"))

    return render_template_string(FINANCE_STYLE + """
    <h1>{{ category }}</h1>

    {% if message %}
        <p style="color:red;">{{ message }}</p>
    {% endif %}

    <form method="post">

        <p>
            收条号码：
            <input
                name="receipt_no"
                value="{{ next_receipt_no }}"
                required
            >
            <br>
            <small style="color:#666;">
                系统会自动建议 CHE 编号；如果换新收条簿，可以手动修改。
            </small>
        </p>

        <p>
            开收条日期：
            <input name="receipt_date" type="date" value="{{ today }}" required>
        </p>

        <p>
            付款日期：
            <input name="record_date" type="date" value="{{ today }}" required>
        </p>

        <p>
            姓名：
            <input name="name">
        </p>

        <p>
            电话：
            <input name="phone">
        </p>

        <p>
            金额 RM：
            <input
                name="amount"
                type="number"
                step="1.00"
                min="0"
                value="50"
                required
            >
        </p>

        <p>
            付款方式：
            <select name="payment_method">
                <option>现金</option>
                <option>银行过账</option>
                <option>支票</option>
            </select>
        </p>

        <p>
            备注：
            <input name="remarks">
        </p>

        <button type="submit">
            保存
        </button>

    </form>

    <p>
        <a href="{{ url_for('finance.finance_home') }}">
            返回财政首页
        </a>
    </p>
                                  
    <script>

    function addDonor(name){

        let textarea = document.querySelector(
            'textarea[name="raw_text"]'
        );

        if(textarea.value.trim() === ""){
            textarea.value = name;
        }else{
            textarea.value += "\\n" + name;
        }

        textarea.focus();
    }

    </script>

    """,
    category=category,
    next_receipt_no=next_receipt_no,
    message=message,
    today=date.today().isoformat()
    )

@finance_bp.route("/income_batch/<category>", methods=["GET", "POST"])
def income_batch(category):

    allowed_categories = [
        "财布施",
        "观音村",
        "膳食结缘",
        "观音堂纯檀香布施",
        SPECIAL_DONATION_TITLE,
        "临时特别布施"
    ]

    if category not in allowed_categories:
        return "Invalid category", 400

    message = ""
    preview_rows = []

    raw_text = request.form.get("raw_text", "").strip()
    next_receipt_no = get_next_receipt_no("CHE")
    next_receipt_raw = next_receipt_no.replace("CHE", "", 1)
    receipt_start_raw = request.form.get("receipt_start", "").strip().upper()

    if not receipt_start_raw:
        receipt_start_raw = next_receipt_raw

    if receipt_start_raw.isdigit():
        receipt_start = "CHE" + str(int(receipt_start_raw)).zfill(7)
    else:
        receipt_start = receipt_start_raw

    receipt_date = request.form.get("receipt_date") or date.today().isoformat()
    record_date = receipt_date
    payment_method = request.form.get("payment_method", "现金")
    default_amount = money(request.form.get("default_amount") or 50)
    remarks = request.form.get("remarks", "").strip()
    action = request.form.get("action", "")

    def make_receipt_no(start_no, index):
        m = re.match(r"^([A-Z]+)(\d+)$", start_no)
        if not m:
            return ""

        prefix = m.group(1)
        num = m.group(2)

        return prefix + str(int(num) + index).zfill(len(num))

    def find_donor_info(keyword):

        keyword = keyword.strip()

        # 1. 如果是数字，先当会员编号找
        if keyword.isdigit():
            member_id = f"CHE-{int(keyword)}"

            row = db_query("""
                select member_id, name, phone
                from members
                where member_id = %s
                limit 1
            """, (member_id,), fetchone=True)

            if row:
                return {
                    "name": row["name"],
                    "phone": row["phone"],
                    "source": "会员"
                }

        # 2. 用姓名找 members
        row = db_query("""
            select member_id, name, phone
            from members
            where name ilike %s
            limit 1
        """, (f"%{keyword}%",), fetchone=True)

        if row:
            return {
                "name": row["name"],
                "phone": row["phone"],
                "source": "会员"
            }

        # 3. 用姓名找义工 volunteers
        row = db_query("""
            select name, phone
            from volunteers
            where name ilike %s
            limit 1
        """, (f"%{keyword}%",), fetchone=True)

        if row:
            return {
                "name": row["name"],
                "phone": row["phone"],
                "source": "义工"
            }

        # 4. 最后找历史捐赠者
        row = db_query("""
            select name, phone
            from finance_records
            where category <> '月费'
            and coalesce(status, 'confirmed') <> 'cancelled'
            and name ilike %s
            and coalesce(name, '') <> ''
            order by record_date desc, id desc
            limit 1
        """, (f"%{keyword}%",), fetchone=True)

        if row:
            return {
                "name": row["name"],
                "phone": row["phone"],
                "source": "历史捐赠"
            }

        return {
            "name": keyword,
            "phone": "",
            "source": "手动"
        }

    def build_preview():
        rows = []
        valid_index = 0
        current_receipt_date = receipt_date

        for line in raw_text.splitlines():
            line = line.strip()

            if not line:
                continue

            # 日期分组格式：
            # @2026-07-10
            # 后续记录都会使用这个收条日期，
            # 直到遇到下一个 @日期。
            if line.startswith("@") or line.startswith("＠"):
                date_text = line[1:].strip()

                try:
                    current_receipt_date = date.fromisoformat(
                        date_text
                    ).isoformat()
                except ValueError:
                    rows.append({
                        "error": (
                            f"日期格式错误：{date_text}，"
                            "请使用 @YYYY-MM-DD"
                        ),
                        "receipt_no": "",
                        "receipt_date": date_text,
                        "name": "",
                        "phone": "",
                        "amount": 0,
                        "source": "日期设置",
                    })

                continue

            parts = line.split()

            if len(parts) >= 2:
                amount = money(parts[-1])
                name = " ".join(parts[:-1]).strip()
            else:
                amount = default_amount
                name = parts[0].strip()

            receipt_no = make_receipt_no(
                receipt_start,
                valid_index
            )
            valid_index += 1

            donor = find_donor_info(name)

            name = donor["name"]
            phone = donor["phone"]
            source = donor["source"]

            existing = db_query("""
                select id
                from finance_records
                where receipt_no = %s
                limit 1
            """, (receipt_no,), fetchone=True)

            error = ""

            if not name:
                error = "姓名不能为空"
            elif amount <= 0:
                error = "金额必须大过 0"
            elif existing:
                error = "收条已存在"

            rows.append({
                "error": error,
                "receipt_no": receipt_no,
                "receipt_date": current_receipt_date,
                "name": name,
                "phone": phone,
                "amount": amount,
                "source": source,
            })

        return rows

    recent_donors = db_query("""
        select
            name,
            max(phone) as phone,
            count(*) as times,
            max(record_date) as last_date
        from finance_records
        where category <> '月费'
          and coalesce(status, 'confirmed') <> 'cancelled'
          and coalesce(name, '') <> ''
        group by name
        order by max(record_date) desc, count(*) desc
        limit 50
    """, fetchall=True)

    if request.method == "POST":

        if category == "临时特别布施" and not remarks:
            message = "临时特别布施请填写活动名称或用途"

        elif not receipt_start:
            message = "请填写收条开始号码"

        elif not re.match(r"^CHE\d+$", receipt_start):
            message = "收条号码格式错误，例如：CHE0001501，或只输入 1501"

        elif not raw_text:
            message = "请粘贴批量资料"

        else:
            preview_rows = build_preview()
            has_error = any(r.get("error") for r in preview_rows)

            if action == "confirm" and not has_error:

                for r in preview_rows:

                    db_query("""
                        insert into finance_records
                        (
                            record_type,
                            fund_account,
                            record_date,
                            receipt_date,
                            category,
                            receipt_no,
                            name,
                            phone,
                            amount,
                            payment_method,
                            remarks
                        )
                        values
                        (
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s
                        )
                    """, (
                        "income",
                        get_fund_account(category),
                        r["receipt_date"],
                        r["receipt_date"],
                        category,
                        r["receipt_no"],
                        r["name"],
                        r["phone"],
                        r["amount"],
                        payment_method,
                        remarks or "批量布施录入"
                    ))

                return redirect(url_for("finance.records"))

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>{{ category }}录入</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .finance-form-page{
                max-width:900px;
            }

            .finance-header-card{
                margin-bottom:18px;
                background:linear-gradient(135deg,#1877b8,#2695cf);
                color:white;
            }

            .finance-header-card .page-title,
            .finance-header-card .page-subtitle{
                color:white;
            }

            .finance-back-row{
                margin-bottom:16px;
            }

            .finance-form-grid{
                display:grid;
                grid-template-columns:repeat(2,minmax(0,1fr));
                gap:16px;
            }

            .finance-form-grid .form-group{
                margin:0;
            }

            .finance-full{
                grid-column:1 / -1;
            }

            .receipt-input-row{
                display:flex;
                align-items:center;
                gap:10px;
            }

            .receipt-prefix{
                min-height:52px;
                display:flex;
                align-items:center;
                justify-content:center;
                padding:0 16px;
                border-radius:10px;
                background:#eaf4ff;
                color:#1769aa;
                font-weight:800;
                font-size:20px;
            }

            .receipt-input-row .form-input{
                flex:1;
                min-width:0;
            }

            .field-help{
                display:block;
                color:#667085;
                font-size:15px;
                line-height:1.5;
                margin-top:7px;
            }

            .batch-textarea{
                width:100%;
                min-height:260px;
                resize:vertical;
                line-height:1.7;
                font-family:inherit;
            }

            .batch-example{
                background:#f8fafc;
                border:1px dashed #cbd5e1;
                border-radius:12px;
                padding:14px 16px;
                margin-top:12px;
                color:#475467;
                font-size:15px;
                line-height:1.75;
            }

            .batch-example code{
                font-family:Consolas,monospace;
                font-weight:700;
                color:#1769aa;
            }

            .action-row{
                display:flex;
                flex-wrap:wrap;
                gap:12px;
                margin-top:20px;
            }

            .action-row .btn-tool{
                min-width:160px;
            }

            .preview-success{
                color:#16863a;
                font-weight:700;
            }

            .preview-error{
                color:#c62828;
                font-weight:700;
            }

            .donor-buttons{
                display:flex;
                flex-wrap:wrap;
                gap:9px;
                margin:14px 0;
            }

            .donor-chip{
                border:1px solid #cbd5e1;
                background:#fff;
                border-radius:999px;
                padding:9px 14px;
                font-size:16px;
                cursor:pointer;
            }

            .donor-chip:hover{
                background:#eef6ff;
                border-color:#69a8df;
            }

            .recent-donor-details{
                margin-top:18px;
            }

            .recent-donor-details summary{
                cursor:pointer;
                font-weight:800;
                font-size:18px;
                padding:12px 0;
            }

            .alert{
                margin-bottom:16px;
            }

            @media(max-width:700px){
                .finance-form-grid{
                    grid-template-columns:1fr;
                }

                .finance-full{
                    grid-column:auto;
                }

                .receipt-input-row{
                    align-items:stretch;
                }

                .action-row{
                    display:grid;
                }

                .action-row .btn-tool{
                    width:100%;
                }
            }
        </style>
    </head>

    <body>

    <div class="page finance-form-page">

        <div class="finance-back-row">
            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_income_menu') }}"
            >
                ← 返回收入录入
            </a>
        </div>

        <div class="card finance-header-card">

            <h1 class="page-title">
                💵 {{ category }}
            </h1>

            <p class="page-subtitle">
                可批量录入多笔，并在同一批中使用不同收条日期
            </p>

        </div>

        {% if message %}
            <div class="alert alert-danger">
                {{ message }}
            </div>
        {% endif %}

        <form method="post">

            <div class="card">

                <h2 class="section-title">
                    收条与付款资料
                </h2>

                <div class="finance-form-grid">

                    <div class="form-group">

                        <label class="form-label">
                            收条开始号码
                        </label>

                        <div class="receipt-input-row">

                            <div class="receipt-prefix">
                                CHE
                            </div>

                            <input
                                class="form-input"
                                name="receipt_start"
                                value="{{ receipt_start_raw }}"
                                inputmode="numeric"
                                placeholder="例如 1501"
                                required
                            >

                        </div>

                        <span class="field-help">
                            系统建议下一张收条：
                            <strong>{{ next_receipt_no }}</strong>
                        </span>

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            默认开收条日期
                        </label>

                        <input
                            class="form-input"
                            name="receipt_date"
                            type="date"
                            value="{{ receipt_date }}"
                            required
                        >

                        <span class="field-help">
                            没有写 @日期 的记录，会使用这个日期。
                        </span>

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            默认金额 RM
                        </label>

                        <input
                            class="form-input"
                            name="default_amount"
                            type="number"
                            step="1"
                            min="0"
                            value="{{ default_amount }}"
                            required
                        >

                        <span class="field-help">
                            只输入姓名时，系统会自动使用这个金额。
                        </span>

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            付款方式
                        </label>

                        <select
                            class="form-input"
                            name="payment_method"
                        >
                            <option
                                {% if payment_method == '现金' %}
                                    selected
                                {% endif %}
                            >
                                现金
                            </option>

                            <option
                                {% if payment_method == '银行过账' %}
                                    selected
                                {% endif %}
                            >
                                银行过账
                            </option>

                            <option
                                {% if payment_method == '支票' %}
                                    selected
                                {% endif %}
                            >
                                支票
                            </option>
                        </select>

                    </div>

                    <div class="form-group finance-full">

                        <label class="form-label">
                            备注／活动名称
                        </label>

                        <input
                            class="form-input"
                            name="remarks"
                            value="{{ remarks }}"
                            placeholder="{% if category == '临时特别布施' %}临时特别布施必须填写活动名称{% else %}例如：观音诞、法会、特别活动{% endif %}"
                        >

                        {% if category == "临时特别布施" %}
                            <span class="field-help">
                                此项目必须填写活动名称或用途。
                            </span>
                        {% endif %}

                    </div>

                </div>

            </div>

            <div class="card">

                <h2 class="section-title">
                    批量输入
                </h2>

                <p class="page-subtitle">
                    可用 @日期 分组；只输入姓名会使用默认金额，
                    姓名后面加金额可单独修改。
                </p>

                <textarea
                    class="form-input batch-textarea"
                    name="raw_text"
                    placeholder="例如：
@2026-07-10
王小明
李大华
陈美玲 100

@2026-07-11
郑依颖 30
林美珍 50"
                    required
                >{{ raw_text }}</textarea>

                <div class="batch-example">

                    <strong>输入规则：</strong><br>

                    <code>@2026-07-10</code>
                    → 后续记录使用 2026-07-10<br>

                    <code>王小明</code>
                    → 使用默认金额<br>

                    <code>陈美玲 100</code>
                    → 金额 RM100<br>

                    没有写 <code>@日期</code> 的记录，
                    会使用上面的默认开收条日期。

                </div>

                <div class="action-row">

                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                        name="action"
                        value="preview"
                    >
                        👁️ 预览资料
                    </button>

                    {% if preview_rows %}
                        <button
                            class="btn-tool btn-success"
                            type="submit"
                            name="action"
                            value="confirm"
                            onclick="return confirm('确定全部入账？');"
                        >
                            ✅ 确认全部入账
                        </button>
                    {% endif %}

                </div>

            </div>

            {% if preview_rows %}

                <div class="card">

                    <h2 class="section-title">
                        录入预览
                    </h2>

                    <div class="table-responsive">

                        <table class="record-table">

                            <thead>
                                <tr>
                                    <th>收条</th>
                                    <th>收条日期</th>
                                    <th>姓名</th>
                                    <th>电话</th>
                                    <th>来源</th>
                                    <th>金额</th>
                                    <th>检查结果</th>
                                </tr>
                            </thead>

                            <tbody>

                                {% for r in preview_rows %}
                                <tr>

                                    <td>{{ r.receipt_no or "-" }}</td>

                                    <td>
                                        {{ r.receipt_date or "-" }}
                                    </td>

                                    <td>
                                        <strong>{{ r.name or "-" }}</strong>
                                    </td>

                                    <td>{{ r.phone or "-" }}</td>

                                    <td>{{ r.source or "-" }}</td>

                                    <td>
                                        RM {{ "%.2f"|format(r.amount or 0) }}
                                    </td>

                                    <td>
                                        {% if r.error %}
                                            <span class="preview-error">
                                                ❌ {{ r.error }}
                                            </span>
                                        {% else %}
                                            <span class="preview-success">
                                                ✅ 可以入账
                                            </span>
                                        {% endif %}
                                    </td>

                                </tr>
                                {% endfor %}

                            </tbody>

                        </table>

                    </div>

                </div>

            {% endif %}

        </form>

        {% if recent_donors %}

            <div class="card">

                <h2 class="section-title">
                    常用捐赠者
                </h2>

                <p class="page-subtitle">
                    点一下姓名，会自动加入批量输入框。
                </p>

                <div class="donor-buttons">

                    {% for d in recent_donors %}
                        <button
                            class="donor-chip"
                            type="button"
                            onclick="addDonor({{ d.name|tojson }})"
                        >
                            {{ d.name }}
                        </button>
                    {% endfor %}

                </div>

                <details class="recent-donor-details">

                    <summary>
                        查看捐赠者详细记录
                    </summary>

                    <div class="table-responsive">

                        <table class="record-table">

                            <thead>
                                <tr>
                                    <th>姓名</th>
                                    <th>电话</th>
                                    <th>次数</th>
                                    <th>最后日期</th>
                                </tr>
                            </thead>

                            <tbody>

                                {% for d in recent_donors %}
                                <tr>
                                    <td>{{ d.name }}</td>
                                    <td>{{ d.phone or "-" }}</td>
                                    <td>{{ d.times }}</td>
                                    <td>{{ d.last_date }}</td>
                                </tr>
                                {% endfor %}

                            </tbody>

                        </table>

                    </div>

                </details>

            </div>

        {% endif %}

    </div>

    <script>
    function addDonor(name){

        const textarea = document.querySelector(
            'textarea[name="raw_text"]'
        );

        if(!textarea){
            return;
        }

        const currentText = textarea.value.trim();

        if(currentText === ""){
            textarea.value = name;
        }else{
            textarea.value += "\\n" + name;
        }

        textarea.focus();
        textarea.scrollTop = textarea.scrollHeight;
    }
    </script>

    </body>
    </html>
    """,
        category=category,
        message=message,
        raw_text=raw_text,
        receipt_start_raw=receipt_start_raw,
        next_receipt_no=next_receipt_no,
        receipt_date=receipt_date,
        payment_method=payment_method,
        default_amount=default_amount,
        remarks=remarks,
        preview_rows=preview_rows,
        recent_donors=recent_donors
        )

@finance_bp.route("/bank_pending/<int:pending_id>/delete", methods=["POST"])
def delete_bank_pending(pending_id):

    db_query("""
        delete from bank_pending_records
        where id = %s
        and status = 'pending'
    """, (pending_id,))

    return redirect(url_for("finance.bank_pending"))

@finance_bp.route("/bank_pending", methods=["GET", "POST"])
def bank_pending():

    message = ""
    message_type = "bad"

    form_data = {
        "raw_text": "",
        "member_id": "",
        "name": "",
        "amount": "",
        "payment_date": date.today().isoformat(),
        "bank_ref": "",
        "bank_name": "",
        "category": "月费",
        "remarks": "",
    }

    if request.method == "POST":

        raw_text = request.form.get("raw_text", "").strip()

        member_id_raw = request.form.get(
            "member_id",
            ""
        ).strip()

        name = request.form.get(
            "name",
            ""
        ).strip()

        amount_raw = request.form.get(
            "amount",
            ""
        ).strip()

        payment_date = (
            request.form.get("payment_date")
            or date.today().isoformat()
        )

        bank_ref = request.form.get(
            "bank_ref",
            ""
        ).strip()

        bank_name = request.form.get(
            "bank_name",
            ""
        ).strip()

        category = request.form.get(
            "category",
            "月费"
        ).strip()

        remarks = request.form.get(
            "remarks",
            ""
        ).strip()

        if raw_text:

            member_match = re.search(
                r"(CHE|STW)[-\s]?\d+",
                raw_text,
                re.IGNORECASE
            )

            if member_match and not member_id_raw:

                member_id_raw = (
                    member_match
                    .group(0)
                    .replace(" ", "-")
                    .upper()
                )

            amount_match = re.search(
                r"(RM|MYR)\s*([0-9]+(?:\.[0-9]{1,2})?)",
                raw_text,
                re.IGNORECASE
            )

            if amount_match and not amount_raw:
                amount_raw = amount_match.group(2)

            ref_match = re.search(
                r"(Reference|Ref|Transaction|DuitNow)"
                r"\s*(No|ID|Number)?[:\s#-]*"
                r"([A-Za-z0-9\-]+)",
                raw_text,
                re.IGNORECASE
            )

            if ref_match and not bank_ref:
                bank_ref = ref_match.group(3)

            bank_options = [
                "Maybank",
                "Public Bank",
                "CIMB",
                "Hong Leong",
                "RHB",
                "AmBank",
                "Bank Islam",
                "BSN",
            ]

            for bank in bank_options:

                if (
                    bank.lower() in raw_text.lower()
                    and not bank_name
                ):
                    bank_name = bank
                    break

            if not remarks:
                remarks = raw_text[:1000]

        member_id = (
            normalize_member_id(member_id_raw)
            if member_id_raw
            else ""
        )

        amount = money(amount_raw)

        if category == "月费" and member_id:

            member = db_query("""
                select *
                from members
                where member_id = %s
                limit 1
            """, (
                member_id,
            ), fetchone=True)

            if member:
                name = (
                    member.get("name")
                    or member.get("姓名")
                    or name
                )

        form_data = {
            "raw_text": raw_text,
            "member_id": member_id_raw,
            "name": name,
            "amount": amount_raw,
            "payment_date": str(payment_date),
            "bank_ref": bank_ref,
            "bank_name": bank_name,
            "category": category,
            "remarks": remarks,
        }

        existing_ref = None

        if bank_ref:

            existing_ref = db_query("""
                select id
                from bank_pending_records
                where bank_ref = %s
                  and coalesce(bank_ref, '') <> ''
                limit 1
            """, (
                bank_ref,
            ), fetchone=True)

        if existing_ref:

            message = (
                "这个 Bank Reference 已经存在，"
                "请检查是否重复导入。"
            )

        elif amount <= 0:

            message = "请输入正确的银行过账金额。"

        else:

            db_query("""
                insert into bank_pending_records
                (
                    member_id,
                    name,
                    amount,
                    payment_date,
                    bank_ref,
                    bank_name,
                    category,
                    remarks
                )
                values
                (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, (
                member_id,
                name,
                amount,
                payment_date,
                bank_ref,
                bank_name,
                category,
                remarks
            ))

            return redirect(
                url_for("finance.bank_pending")
            )

    summary = db_query("""
        select
            count(*) as cnt,
            coalesce(sum(amount), 0) as total
        from bank_pending_records
        where status = 'pending'
    """, fetchone=True)

    rows = db_query("""
        select *
        from bank_pending_records
        where status = 'pending'
        order by upload_date desc
    """, fetchall=True)

    pending_count = int(
        summary["cnt"] or 0
    )

    pending_total = float(
        summary["total"] or 0
    )

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>

        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>银行过账中心</title>

        <link
            rel="stylesheet"
            href="{{ url_for(
                'static',
                filename='css/toolbox.css'
            ) }}"
        >

        <style>

            .bank-page {
                max-width: 1450px;
            }

            .bank-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;

                box-shadow:
                    0 12px 30px
                    rgba(37, 99, 235, 0.18);
            }

            .bank-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .bank-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .bank-summary {
                display: grid;
                grid-template-columns:
                    repeat(2, minmax(0, 1fr));
                gap: 16px;
                margin-bottom: 20px;
            }

            .bank-summary .summary-box {
                min-height: 125px;
                text-align: center;

                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
            }

            .summary-icon {
                font-size: 30px;
                margin-bottom: 6px;
            }

            .summary-label {
                color: #64748b;
                font-size: 16px;
                margin-bottom: 6px;
            }

            .summary-value {
                font-size: 28px;
                font-weight: 800;
                color: #0f172a;
            }

            .summary-value.money-value {
                color: #15803d;
            }

            .entry-grid {
                display: grid;
                grid-template-columns:
                    minmax(0, 1.2fr)
                    minmax(360px, 0.8fr);
                gap: 20px;
                align-items: start;
            }

            .form-grid {
                display: grid;
                grid-template-columns:
                    repeat(2, minmax(0, 1fr));
                gap: 16px;
            }

            .form-grid .full-width {
                grid-column: 1 / -1;
            }

            .receipt-textarea {
                width: 100%;
                min-height: 245px;
                resize: vertical;
                line-height: 1.6;
            }

            .form-help {
                margin-top: 7px;
                color: #64748b;
                font-size: 14px;
                line-height: 1.5;
            }

            .smart-note {
                background: #eff6ff;
                border: 1px solid #bfdbfe;
                color: #1e40af;
                border-radius: 14px;
                padding: 14px 16px;
                margin-bottom: 16px;
                line-height: 1.6;
            }

            .records-card {
                margin-top: 20px;
            }

            .table-topbar {
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 12px;
                flex-wrap: wrap;
                margin-bottom: 14px;
            }

            .record-count {
                color: #64748b;
                font-size: 16px;
            }

            .pending-table {
                min-width: 1400px;
            }

            .pending-table td {
                vertical-align: middle;
            }

            .member-id {
                font-weight: 800;
                color: #1d4ed8;
                white-space: nowrap;
            }

            .member-name {
                font-weight: 700;
                color: #1e293b;
                min-width: 90px;
            }

            .money-cell {
                color: #15803d;
                font-weight: 800;
                white-space: nowrap;
                text-align: right;
            }

            .reference-cell {
                font-family: Consolas, monospace;
                font-size: 14px;
                white-space: nowrap;
            }

            .remarks-cell {
                min-width: 180px;
                max-width: 320px;
                white-space: normal;
                line-height: 1.5;
                overflow-wrap: anywhere;
            }

            .confirm-box {
                min-width: 285px;
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 14px;
                padding: 12px;
            }

            .confirm-grid {
                display: grid;
                gap: 10px;
            }

            .confirm-box .form-group {
                margin: 0;
            }

            .confirm-box .form-label {
                font-size: 14px;
            }

            .confirm-box .form-input {
                min-height: 41px;
                padding: 8px 10px;
                font-size: 14px;
                margin: 0;
            }

            .confirm-box .btn-tool {
                width: 100%;
                min-height: 42px;
                font-size: 14px;
                padding: 9px 12px;
            }

            .delete-form .btn-tool {
                min-height: 42px;
                padding: 9px 13px;
                font-size: 14px;
                white-space: nowrap;
            }

            .empty-bank {
                text-align: center;
                padding: 55px 20px;
                color: #64748b;
            }

            .empty-bank-icon {
                font-size: 48px;
                margin-bottom: 10px;
            }

            .empty-bank h3 {
                margin: 0 0 8px;
                color: #334155;
            }

            .bottom-actions {
                margin-top: 20px;
            }

            @media (max-width: 950px) {

                .entry-grid {
                    grid-template-columns: 1fr;
                }

            }

            @media (max-width: 700px) {

                .bank-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .bank-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .bank-header h1 {
                    font-size: 26px;
                }

                .bank-summary {
                    grid-template-columns: 1fr;
                }

                .bank-summary .summary-box {
                    min-height: 105px;
                }

                .form-grid {
                    grid-template-columns: 1fr;
                }

                .form-grid .full-width {
                    grid-column: auto;
                }

                .entry-grid {
                    display: block;
                }

                .entry-grid .card + .card {
                    margin-top: 20px;
                }

                .bottom-actions .btn-tool {
                    width: 100%;
                }

            }

        </style>

    </head>

    <body>

    <div class="page bank-page">

        <div class="bank-header">

            <h1>🏦 银行过账中心</h1>

            <p>
                先登记银行转账资料，财政确认后再填写收条号码并正式入账。
            </p>

        </div>

        {% if message %}

            <div class="alert alert-danger">
                ⚠️ {{ message }}
            </div>

        {% endif %}

        <div class="bank-summary">

            <div class="summary-box">

                <div class="summary-icon">
                    ⏳
                </div>

                <div class="summary-label">
                    待确认笔数
                </div>

                <div class="summary-value">
                    {{ pending_count }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    💰
                </div>

                <div class="summary-label">
                    待确认总额
                </div>

                <div class="summary-value money-value">
                    RM {{ "%.2f"|format(pending_total) }}
                </div>

            </div>

        </div>

        <form method="post">

            <div class="entry-grid">

                <div class="card">

                    <div class="section-title">
                        📄 银行资料智能识别
                    </div>

                    <div class="smart-note">
                        可以直接粘贴 WhatsApp 信息或银行转账
                        Receipt 文字。系统会尝试识别会员编号、
                        金额、银行及 Reference。
                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            项目
                        </label>

                        <select
                            class="form-input"
                            name="category"
                        >

                            {% set categories = [
                                '月费',
                                '财布施',
                                '观音村',
                                '膳食结缘',
                                '观音堂纯檀香布施',
                                special_donation_title,
                                '临时特别布施'
                            ] %}

                            {% for category_item in categories %}

                                <option
                                    value="{{ category_item }}"
                                    {% if
                                        form_data.category
                                        == category_item
                                    %}
                                        selected
                                    {% endif %}
                                >
                                    {{ category_item }}
                                </option>

                            {% endfor %}

                        </select>

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            粘贴 WhatsApp／银行 Receipt 文字
                        </label>

                        <textarea
                            class="form-input receipt-textarea"
                            name="raw_text"
                            placeholder="例如：

CHE-108
RM50.00
Maybank
Reference：ABC123456"
                        >{{ form_data.raw_text }}</textarea>

                        <div class="form-help">
                            系统识别后，仍可在右边检查和修改资料。
                        </div>

                    </div>

                </div>

                <div class="card">

                    <div class="section-title">
                        ✏️ 手动补充／修正资料
                    </div>

                    <div class="form-grid">

                        <div class="form-group">

                            <label class="form-label">
                                会员编号
                            </label>

                            <input
                                class="form-input"
                                name="member_id"
                                value="{{ form_data.member_id }}"
                                placeholder="CHE-108 / STW-108 / 108"
                                autocomplete="off"
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                姓名
                            </label>

                            <input
                                class="form-input"
                                name="name"
                                value="{{ form_data.name }}"
                                placeholder="非月费或找不到会员时填写"
                                autocomplete="off"
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                金额 RM
                            </label>

                            <input
                                class="form-input"
                                name="amount"
                                type="number"
                                step="0.01"
                                min="0.01"
                                value="{{ form_data.amount }}"
                                placeholder="例如 50.00"
                                required
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                付款日期
                            </label>

                            <input
                                class="form-input"
                                name="payment_date"
                                type="date"
                                value="{{ form_data.payment_date }}"
                                required
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                银行 Reference
                            </label>

                            <input
                                class="form-input"
                                name="bank_ref"
                                value="{{ form_data.bank_ref }}"
                                placeholder="Transaction Reference"
                                autocomplete="off"
                            >

                        </div>

                        <div class="form-group">

                            <label class="form-label">
                                银行
                            </label>

                            <select
                                class="form-input"
                                name="bank_name"
                            >

                                <option value="">
                                    请选择银行
                                </option>

                                {% for bank in bank_names %}

                                    <option
                                        value="{{ bank }}"
                                        {% if
                                            form_data.bank_name == bank
                                        %}
                                            selected
                                        {% endif %}
                                    >
                                        {{ bank }}
                                    </option>

                                {% endfor %}

                            </select>

                        </div>

                        <div class="form-group full-width">

                            <label class="form-label">
                                备注
                            </label>

                            <textarea
                                class="form-input"
                                name="remarks"
                                rows="4"
                                placeholder="填写补充说明或保留银行转账文字"
                            >{{ form_data.remarks }}</textarea>

                        </div>

                    </div>

                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                        style="width:100%;"
                    >
                        ➕ 加入待确认
                    </button>

                </div>

            </div>

        </form>

        <div class="card records-card">

            <div class="table-topbar">

                <div
                    class="section-title"
                    style="margin-bottom:0;"
                >
                    📋 待确认列表
                </div>

                <div class="record-count">
                    共
                    <strong>{{ pending_count }}</strong>
                    笔待确认记录
                </div>

            </div>

            {% if rows %}

                <div class="table-responsive">

                    <table class="record-table pending-table">

                        <thead>

                            <tr>
                                <th>付款日期</th>
                                <th>编号</th>
                                <th>姓名</th>
                                <th>项目</th>
                                <th>金额</th>
                                <th>Reference</th>
                                <th>银行</th>
                                <th>备注</th>
                                <th>确认入账</th>
                                <th>删除</th>
                            </tr>

                        </thead>

                        <tbody>

                            {% for r in rows %}

                                <tr>

                                    <td style="white-space:nowrap;">
                                        {{ r.payment_date }}
                                    </td>

                                    <td>
                                        <span class="member-id">
                                            {{ r.member_id or "-" }}
                                        </span>
                                    </td>

                                    <td>
                                        <span class="member-name">
                                            {{ r.name or "-" }}
                                        </span>
                                    </td>

                                    <td>
                                        {{ r.category or "-" }}
                                    </td>

                                    <td class="money-cell">
                                        RM
                                        {{ "%.2f"|format(
                                            r.amount or 0
                                        ) }}
                                    </td>

                                    <td class="reference-cell">
                                        {{ r.bank_ref or "-" }}
                                    </td>

                                    <td style="white-space:nowrap;">
                                        {{ r.bank_name or "-" }}
                                    </td>

                                    <td class="remarks-cell">
                                        {{ r.remarks or "-" }}
                                    </td>

                                    <td>

                                        <div class="confirm-box">

                                            <form
                                                method="post"
                                                action="{{ url_for(
                                                    'finance.confirm_bank',
                                                    pending_id=r.id
                                                ) }}"
                                                onsubmit="
                                                    return confirm(
                                                        '确定确认入账？'
                                                    );
                                                "
                                            >

                                                <div class="confirm-grid">

                                                    <div class="form-group">

                                                        <label
                                                            class="form-label"
                                                        >
                                                            收条号码
                                                        </label>

                                                        <input
                                                            class="form-input"
                                                            name="receipt_no"
                                                            placeholder="CHE0000001"
                                                            autocomplete="off"
                                                            required
                                                        >

                                                    </div>

                                                    <div class="form-group">

                                                        <label
                                                            class="form-label"
                                                        >
                                                            开收条日期
                                                        </label>

                                                        <input
                                                            class="form-input"
                                                            name="receipt_date"
                                                            type="date"
                                                            value="{{ today }}"
                                                            required
                                                        >

                                                    </div>

                                                    <button
                                                        class="
                                                            btn-tool
                                                            btn-success
                                                        "
                                                        type="submit"
                                                    >
                                                        ✅ 确认入账
                                                    </button>

                                                </div>

                                            </form>

                                        </div>

                                    </td>

                                    <td>

                                        <form
                                            class="delete-form"
                                            method="post"
                                            action="{{ url_for(
                                                'finance.delete_bank_pending',
                                                pending_id=r.id
                                            ) }}"
                                            onsubmit="
                                                return confirm(
                                                    '确定删除这笔待确认记录？'
                                                );
                                            "
                                        >

                                            <button
                                                class="
                                                    btn-tool
                                                    btn-danger
                                                "
                                                type="submit"
                                            >
                                                🗑️ 删除
                                            </button>

                                        </form>

                                    </td>

                                </tr>

                            {% endfor %}

                        </tbody>

                    </table>

                </div>

            {% else %}

                <div class="empty-bank">

                    <div class="empty-bank-icon">
                        ✅
                    </div>

                    <h3>目前没有待确认记录</h3>

                    <p>
                        所有银行转账记录都已经处理完成。
                    </p>

                </div>

            {% endif %}

        </div>

        <div class="bottom-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for(
                    'finance.finance_home'
                ) }}"
            >
                ← 返回财政首页
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        summary=summary,
        message=message,
        today=date.today().isoformat(),
        pending_count=pending_count,
        pending_total=pending_total,
        form_data=form_data,
        special_donation_title=SPECIAL_DONATION_TITLE,
        bank_names=[
            "Maybank",
            "Public Bank",
            "CIMB",
            "Hong Leong",
            "RHB",
            "AmBank",
            "Bank Islam",
            "BSN",
        ]
    )

@finance_bp.route("/bank_pending/<int:pending_id>/confirm", methods=["POST"])
def confirm_bank(pending_id):

    receipt_no = request.form.get("receipt_no", "").strip().upper()
    receipt_date = request.form.get("receipt_date") or date.today()

    if not receipt_no:
        return "必须填写收条号码", 400

    existing = db_query("""
        select id
        from finance_records
        where receipt_no = %s
        limit 1
    """, (receipt_no,), fetchone=True)

    if existing:
        return "这个收条号码已经用过了，请检查是否重复输入", 400

    p = db_query("""
        select *
        from bank_pending_records
        where id = %s and status = 'pending'
    """, (pending_id,), fetchone=True)

    if not p:
        return redirect(url_for("finance.bank_pending"))

    member = None
    paid_until_date = None
    month_from = ""
    month_to = ""
    month_from_db = None
    month_to_db = None
    month_count = None

    if p["category"] == "月费" and p["member_id"]:

        member = db_query("""
            select *
            from members
            where member_id = %s
            limit 1
        """, (p["member_id"],), fetchone=True)

        paid = db_query("""
            select max(end_month) as paid_until
            from member_payments
            where member_id = %s
        """, (p["member_id"],), fetchone=True)

        paid_until_date = paid["paid_until"] if paid else None

        month_from = next_month_ym(paid_until_date)

        month_count = max(1, round(float(p["amount"] or 0) / 50))
        month_to = add_months_ym(month_from, month_count - 1)

        month_from_db = month_from + "-01"
        month_to_db = month_to + "-01"

    db_query("""
        insert into finance_records
        (
            record_type,
            fund_account,
            record_date,
            receipt_date,
            category,
            receipt_no,
            member_id,
            name,
            amount,
            payment_method,
            bank_name,
            bank_ref,
            month_from,
            month_to,
            remarks
        )
        values
        (
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            '银行过账',
            %s, %s, %s, %s, %s
        )
    """, (
        "income",
        get_fund_account(p["category"], "income"),
        p["payment_date"],
        receipt_date,
        p["category"],
        receipt_no,
        p["member_id"],
        p["name"],
        p["amount"],
        p["bank_name"],
        p["bank_ref"],
        month_from,
        month_to,
        p["remarks"]
    ))

    if p["category"] == "月费" and p["member_id"]:

        db_query("""
            insert into member_payments
            (
                payment_date,
                receipt_date,
                member_id,
                name,
                receipt_no,
                amount,
                start_month,
                end_month,
                month_count
            )
            values
            (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s
            )
        """, (
            p["payment_date"],
            receipt_date,
            p["member_id"],
            p["name"],
            receipt_no,
            p["amount"],
            month_from_db,
            month_to_db,
            month_count
        ))

    db_query("""
        update bank_pending_records
        set status = 'confirmed'
        where id = %s
    """, (pending_id,))

    return redirect(url_for("finance.records"))

@finance_bp.route("/records/<int:record_id>/cancel", methods=["POST"])
def cancel_record(record_id):

    record = db_query("""
        select receipt_no, status
        from finance_records
        where id = %s
    """, (record_id,), fetchone=True)

    if not record:
        return redirect(url_for("finance.records"))

    if record["status"] == "cancelled":
        return redirect(url_for("finance.records"))

    cancel_reason = request.form.get("cancel_reason", "").strip()

    if not cancel_reason:
        cancel_reason = "未填写原因"

    db_query("""
        update finance_records
        set
            status = 'cancelled',
            cancel_reason = %s
        where id = %s
    """, (cancel_reason, record_id))

    if record["receipt_no"]:
        db_query("""
            delete from member_payments
            where receipt_no = %s
        """, (record["receipt_no"],))

    return redirect(url_for("finance.records"))


@finance_bp.route("/records")
def records():

    q = request.args.get("q", "").strip()

    if q:
        keyword = f"%{q}%"

        rows = db_query("""
            select *
            from finance_records
            where
                receipt_no ilike %s
                or member_id ilike %s
                or name ilike %s
                or category ilike %s
                or payment_method ilike %s
                or bank_ref ilike %s
            order by record_date desc, id desc
            limit 300
        """, (
            keyword,
            keyword,
            keyword,
            keyword,
            keyword,
            keyword
        ), fetchall=True)

    else:
        rows = db_query("""
            select *
            from finance_records
            order by record_date desc, id desc
            limit 300
        """, fetchall=True)

    summary = db_query("""
        select
            sum(
                case
                    when coalesce(status, 'confirmed') <> 'cancelled'
                    then 1
                    else 0
                end
            ) as active_count,

            sum(
                case
                    when status = 'cancelled'
                    then 1
                    else 0
                end
            ) as cancelled_count
        from finance_records
    """, fetchone=True)

    active_count = int(summary["active_count"] or 0)
    cancelled_count = int(summary["cancelled_count"] or 0)

    shown_count = len(rows)

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>财政记录</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .records-page {
                max-width: 1500px;
            }

            .records-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;

                box-shadow:
                    0 12px 30px
                    rgba(37, 99, 235, 0.18);
            }

            .records-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .records-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .search-grid {
                display: grid;
                grid-template-columns: 1fr auto auto;
                gap: 10px;
                align-items: center;
            }

            .search-grid .form-input {
                margin: 0;
            }

            .records-summary {
                display: grid;
                grid-template-columns:
                    repeat(3, minmax(0, 1fr));
                gap: 16px;
                margin-bottom: 20px;
            }

            .records-summary .summary-box {
                text-align: center;
                min-height: 115px;

                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
            }

            .summary-icon {
                font-size: 29px;
                margin-bottom: 5px;
            }

            .summary-label {
                color: #64748b;
                font-size: 16px;
                margin-bottom: 6px;
            }

            .summary-value {
                font-size: 27px;
                font-weight: 800;
                color: #0f172a;
            }

            .summary-value.good {
                color: #15803d;
            }

            .summary-value.cancelled {
                color: #b91c1c;
            }

            .table-topbar {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 12px;
                margin-bottom: 14px;
                flex-wrap: wrap;
            }

            .result-text {
                color: #64748b;
                font-size: 16px;
            }

            .record-table {
                min-width: 1450px;
            }

            .record-table th {
                white-space: nowrap;
            }

            .record-table td {
                vertical-align: middle;
            }

            .receipt-no {
                color: #1d4ed8;
                font-weight: 700;
                white-space: nowrap;
            }

            .member-id {
                font-weight: 700;
                white-space: nowrap;
            }

            .record-name {
                font-weight: 700;
                color: #1e293b;
                min-width: 90px;
            }

            .money {
                font-weight: 800;
                color: #15803d;
                text-align: right;
                white-space: nowrap;
            }

            .record-date {
                white-space: nowrap;
            }

            .month-range {
                white-space: nowrap;
            }

            .remarks-cell {
                min-width: 150px;
                max-width: 260px;
                white-space: normal;
                line-height: 1.5;
            }

            .status-badge {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                padding: 7px 11px;
                border-radius: 999px;
                font-size: 14px;
                font-weight: 800;
                white-space: nowrap;
            }

            .status-confirmed {
                background: #dcfce7;
                color: #166534;
            }

            .status-pending {
                background: #fef3c7;
                color: #92400e;
            }

            .status-cancelled {
                background: #fee2e2;
                color: #991b1b;
            }

            .cancelled-row {
                background: #f8fafc;
                color: #94a3b8;
            }

            .cancelled-row td:not(.action-cell) {
                text-decoration: line-through;
            }

            .cancelled-row .status-badge {
                text-decoration: none;
            }

            .cancel-form {
                display: grid;
                grid-template-columns: minmax(120px, 1fr) auto;
                gap: 8px;
                align-items: center;
                min-width: 240px;
            }

            .cancel-form .form-input {
                margin: 0;
                min-height: 40px;
                padding: 8px 10px;
                font-size: 14px;
            }

            .cancel-form .btn-tool {
                min-height: 40px;
                padding: 8px 12px;
                font-size: 14px;
                white-space: nowrap;
            }

            .empty-state-custom {
                text-align: center;
                padding: 55px 20px;
                color: #64748b;
            }

            .empty-state-custom .empty-icon {
                font-size: 48px;
                margin-bottom: 10px;
            }

            .empty-state-custom h3 {
                color: #334155;
                margin: 0 0 8px;
            }

            .bottom-actions {
                margin-top: 20px;
            }

            @media (max-width: 800px) {

                .records-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .records-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .records-header h1 {
                    font-size: 26px;
                }

                .search-grid {
                    grid-template-columns: 1fr;
                }

                .search-grid .btn-tool {
                    width: 100%;
                }

                .records-summary {
                    grid-template-columns: 1fr;
                }

                .records-summary .summary-box {
                    min-height: 100px;
                }

                .bottom-actions .btn-tool {
                    width: 100%;
                }
            }
        </style>
    </head>

    <body>

    <div class="page records-page">

        <div class="records-header">

            <h1>📚 财政记录</h1>

            <p>
                查询所有收入、支出、银行过账及已作废记录。
            </p>

        </div>

        <div class="card">

            <div class="section-title">
                🔍 搜索财政记录
            </div>

            <form method="get">

                <div class="search-grid">

                    <input
                        class="form-input"
                        name="q"
                        value="{{ q }}"
                        placeholder="搜索收条、编号、姓名、项目、付款方式或 Reference"
                        autocomplete="off"
                    >

                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                    >
                        🔍 搜索
                    </button>

                    {% if q %}

                    <a
                        class="btn-tool btn-secondary"
                        href="{{ url_for('finance.records') }}"
                    >
                        ✕ 清除
                    </a>

                    {% endif %}

                </div>

            </form>

        </div>

        <div class="records-summary">

            <div class="summary-box">

                <div class="summary-icon">
                    ✅
                </div>

                <div class="summary-label">
                    正常记录
                </div>

                <div class="summary-value good">
                    {{ active_count }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    ❌
                </div>

                <div class="summary-label">
                    作废记录
                </div>

                <div class="summary-value cancelled">
                    {{ cancelled_count }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    📋
                </div>

                <div class="summary-label">
                    当前显示
                </div>

                <div class="summary-value">
                    {{ shown_count }}
                </div>

            </div>

        </div>

        <div class="card">

            <div class="table-topbar">

                <div class="section-title" style="margin-bottom:0;">
                    🧾 记录明细
                </div>

                <div class="result-text">

                    {% if q %}
                        搜索：
                        <strong>{{ q }}</strong>
                        ，找到
                        <strong>{{ shown_count }}</strong>
                        笔记录
                    {% else %}
                        显示最近
                        <strong>{{ shown_count }}</strong>
                        笔记录
                    {% endif %}

                </div>

            </div>

            {% if rows %}

            <div class="table-responsive">

                <table class="record-table">

                    <thead>
                        <tr>
                            <th>付款日期</th>
                            <th>开收条日期</th>
                            <th>项目</th>
                            <th>收条</th>
                            <th>编号</th>
                            <th>姓名</th>
                            <th>金额</th>
                            <th>方式</th>
                            <th>Reference</th>
                            <th>月份</th>
                            <th>备注</th>
                            <th>状态</th>
                            <th>操作</th>
                        </tr>
                    </thead>

                    <tbody>

                        {% for r in rows %}

                        <tr
                            {% if r.status == 'cancelled' %}
                                class="cancelled-row"
                            {% endif %}
                        >

                            <td class="record-date">
                                {{ r.record_date }}
                            </td>

                            <td class="record-date">
                                {{ r.receipt_date or "-" }}
                            </td>

                            <td>
                                {{ r.category or "-" }}
                            </td>

                            <td>
                                <span class="receipt-no">
                                    {{ r.receipt_no or "-" }}
                                </span>
                            </td>

                            <td>
                                <span class="member-id">
                                    {{ r.member_id or "-" }}
                                </span>
                            </td>

                            <td>
                                <span class="record-name">
                                    {{ r.name or "-" }}
                                </span>
                            </td>

                            <td class="money">
                                RM {{ "%.2f"|format(r.amount or 0) }}
                            </td>

                            <td>
                                {{ r.payment_method or "-" }}
                            </td>

                            <td>
                                {{ r.bank_ref or "-" }}
                            </td>

                            <td class="month-range">

                                {% if r.month_from or r.month_to %}

                                    {{ r.month_from or "-" }}
                                    至
                                    {{ r.month_to or "-" }}

                                {% else %}

                                    -

                                {% endif %}

                            </td>

                            <td class="remarks-cell">
                                {{ r.remarks or "-" }}
                            </td>

                            <td>

                                {% if r.status == 'cancelled' %}

                                    <span
                                        class="
                                            status-badge
                                            status-cancelled
                                        "
                                    >
                                        ❌ 已作废
                                    </span>

                                {% elif r.status == 'confirmed'
                                    or not r.status %}

                                    <span
                                        class="
                                            status-badge
                                            status-confirmed
                                        "
                                    >
                                        ✅ 已入账
                                    </span>

                                {% else %}

                                    <span
                                        class="
                                            status-badge
                                            status-pending
                                        "
                                    >
                                        ⏳ 待确认
                                    </span>

                                {% endif %}

                            </td>

                            <td class="action-cell">

                                {% if r.status != 'cancelled' %}

                                    <form
                                        class="cancel-form"
                                        method="post"
                                        action="{{ url_for(
                                            'finance.cancel_record',
                                            record_id=r.id
                                        ) }}"
                                        onsubmit="
                                            return confirm(
                                                '确定要作废这笔记录吗？作废后不会计入财政统计。'
                                            );
                                        "
                                    >

                                        <input
                                            class="form-input"
                                            name="cancel_reason"
                                            placeholder="填写作废原因"
                                            autocomplete="off"
                                            required
                                        >

                                        <button
                                            class="btn-tool btn-warning"
                                            type="submit"
                                        >
                                            作废
                                        </button>

                                    </form>

                                {% else %}

                                    <span style="color:#94a3b8;">
                                        已作废
                                    </span>

                                {% endif %}

                            </td>

                        </tr>

                        {% endfor %}

                    </tbody>

                </table>

            </div>

            {% else %}

            <div class="empty-state-custom">

                <div class="empty-icon">
                    🔎
                </div>

                <h3>没有找到财政记录</h3>

                <p>
                    请检查收条编号、会员编号、姓名、项目或
                    Reference 是否正确。
                </p>

            </div>

            {% endif %}

        </div>

        <div class="bottom-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_home') }}"
            >
                ← 返回财政首页
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        rows=rows,
        active_count=active_count,
        cancelled_count=cancelled_count,
        shown_count=shown_count,
        q=q
    )

@finance_bp.route("/dashboard")
def dashboard():

    ym = request.args.get(
        "ym",
        date.today().strftime("%Y-%m")
    )

    daily_income = db_query("""
        select
            category,
            coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'income'
        group by category
        order by category
    """, (ym,), fetchall=True)

    daily_expense = db_query("""
        select
            category,
            coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'expense'
        group by category
        order by category
    """, (ym,), fetchall=True)

    hq_income = db_query("""
        select
            category,
            coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '总会户口'
          and record_type = 'income'
        group by category
        order by category
    """, (ym,), fetchall=True)

    daily_income_total = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'income'
    """, (ym,), fetchone=True)

    daily_expense_total = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '观音堂日常户口'
          and record_type = 'expense'
    """, (ym,), fetchone=True)

    hq_income_total = db_query("""
        select coalesce(sum(amount), 0) as total
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
          and coalesce(status, 'confirmed') <> 'cancelled'
          and fund_account = '总会户口'
          and record_type = 'income'
    """, (ym,), fetchone=True)

    try:
        opening_balance = float(
            request.args.get("opening_balance") or 0
        )
    except (TypeError, ValueError):
        opening_balance = 0.0

    daily_income_value = float(
        daily_income_total["total"] or 0
    )

    daily_expense_value = float(
        daily_expense_total["total"] or 0
    )

    hq_income_value = float(
        hq_income_total["total"] or 0
    )

    daily_balance = (
        opening_balance
        + daily_income_value
        - daily_expense_value
    )

    monthly_net = (
        daily_income_value
        - daily_expense_value
    )

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>
        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>财政统计 Dashboard</title>

        <link
            rel="stylesheet"
            href="{{ url_for('static', filename='css/toolbox.css') }}"
        >

        <style>
            .dashboard-page {
                max-width: 1180px;
            }

            .dashboard-header {
                background:
                    linear-gradient(
                        135deg,
                        #2563eb,
                        #1d4ed8
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;
                box-shadow:
                    0 12px 30px
                    rgba(37, 99, 235, 0.18);
            }

            .dashboard-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .dashboard-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .filter-grid {
                display: grid;
                grid-template-columns:
                    minmax(180px, 1fr)
                    minmax(220px, 1fr)
                    auto;
                gap: 14px;
                align-items: end;
            }

            .dashboard-summary {
                display: grid;
                grid-template-columns:
                    repeat(4, minmax(0, 1fr));
                gap: 16px;
                margin-bottom: 20px;
            }

            .dashboard-summary .summary-box {
                min-height: 130px;
                display: flex;
                flex-direction: column;
                justify-content: center;
                text-align: center;
            }

            .summary-icon {
                font-size: 30px;
                margin-bottom: 6px;
            }

            .summary-label {
                color: #64748b;
                font-size: 16px;
                margin-bottom: 7px;
            }

            .summary-value {
                color: #0f172a;
                font-size: 25px;
                font-weight: 800;
                white-space: nowrap;
            }

            .summary-positive {
                color: #15803d;
            }

            .summary-negative {
                color: #b91c1c;
            }

            .account-grid {
                display: grid;
                grid-template-columns:
                    minmax(0, 1fr)
                    minmax(0, 1fr);
                gap: 20px;
                align-items: start;
            }

            .account-title {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 12px;
                margin-bottom: 14px;
            }

            .account-title h2 {
                margin: 0;
                font-size: 23px;
                color: #1e293b;
            }

            .account-badge {
                display: inline-flex;
                align-items: center;
                padding: 7px 12px;
                border-radius: 999px;
                background: #dbeafe;
                color: #1d4ed8;
                font-size: 14px;
                font-weight: 700;
                white-space: nowrap;
            }

            .account-badge.hq {
                background: #ede9fe;
                color: #6d28d9;
            }

            .money {
                text-align: right;
                font-weight: 700;
                white-space: nowrap;
            }

            .money-income {
                color: #15803d;
            }

            .money-expense {
                color: #b91c1c;
            }

            .total-row {
                background: #f8fafc;
                font-weight: 800;
            }

            .balance-row {
                background: #eff6ff;
                font-weight: 800;
            }

            .dashboard-actions {
                display: flex;
                justify-content: space-between;
                gap: 12px;
                flex-wrap: wrap;
                margin-top: 20px;
            }

            .empty-row {
                text-align: center;
                padding: 28px 12px;
                color: #64748b;
            }

            .section-gap {
                margin-top: 20px;
            }

            @media (max-width: 900px) {

                .dashboard-summary {
                    grid-template-columns:
                        repeat(2, minmax(0, 1fr));
                }

                .account-grid {
                    grid-template-columns: 1fr;
                }
            }

            @media (max-width: 650px) {

                .dashboard-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .dashboard-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .dashboard-header h1 {
                    font-size: 26px;
                }

                .filter-grid {
                    grid-template-columns: 1fr;
                }

                .filter-grid .btn-tool {
                    width: 100%;
                }

                .dashboard-summary {
                    grid-template-columns: 1fr;
                }

                .dashboard-summary .summary-box {
                    min-height: 110px;
                }

                .dashboard-actions {
                    flex-direction: column;
                }

                .dashboard-actions .btn-tool {
                    width: 100%;
                }

                .record-table {
                    min-width: 620px;
                }
            }
        </style>
    </head>

    <body>

    <div class="page dashboard-page">

        <div class="dashboard-header">

            <h1>📊 财政统计 Dashboard</h1>

            <p>
                查看每月收入、支出、户口结余及总会户口收入。
            </p>

        </div>

        <div class="card">

            <div class="section-title">
                🔎 查询条件
            </div>

            <form method="get">

                <div class="filter-grid">

                    <div class="form-group">

                        <label class="form-label">
                            月份
                        </label>

                        <input
                            class="form-input"
                            type="month"
                            name="ym"
                            value="{{ ym }}"
                            required
                        >

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            上月结余 RM
                        </label>

                        <input
                            class="form-input"
                            type="number"
                            name="opening_balance"
                            value="{{ opening_balance }}"
                            step="0.01"
                            min="0"
                            placeholder="例如 15000.00"
                        >

                    </div>

                    <button
                        class="btn-tool btn-primary"
                        type="submit"
                    >
                        🔍 查看统计
                    </button>

                </div>

            </form>

        </div>

        <div class="dashboard-summary">

            <div class="summary-box">

                <div class="summary-icon">
                    💰
                </div>

                <div class="summary-label">
                    日常户口收入
                </div>

                <div class="summary-value summary-positive">
                    RM {{ "%.2f"|format(daily_income_value) }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    💸
                </div>

                <div class="summary-label">
                    日常户口支出
                </div>

                <div class="summary-value summary-negative">
                    RM {{ "%.2f"|format(daily_expense_value) }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    📈
                </div>

                <div class="summary-label">
                    本月收支差额
                </div>

                <div class="
                    summary-value
                    {% if monthly_net >= 0 %}
                        summary-positive
                    {% else %}
                        summary-negative
                    {% endif %}
                ">
                    RM {{ "%.2f"|format(monthly_net) }}
                </div>

            </div>

            <div class="summary-box">

                <div class="summary-icon">
                    🏦
                </div>

                <div class="summary-label">
                    日常户口结余
                </div>

                <div class="
                    summary-value
                    {% if daily_balance >= 0 %}
                        summary-positive
                    {% else %}
                        summary-negative
                    {% endif %}
                ">
                    RM {{ "%.2f"|format(daily_balance) }}
                </div>

            </div>

        </div>

        <div class="account-grid">

            <div class="card">

                <div class="account-title">

                    <h2>
                        🏠 观音堂日常户口
                    </h2>

                    <span class="account-badge">
                        收入与支出
                    </span>

                </div>

                <div class="section-title">
                    📥 收入明细
                </div>

                <div class="table-responsive">

                    <table class="record-table">

                        <thead>
                            <tr>
                                <th>收入项目</th>
                                <th>金额 RM</th>
                            </tr>
                        </thead>

                        <tbody>

                            {% for r in daily_income %}

                            <tr>
                                <td>
                                    {{ r.category }}
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(r.total) }}
                                </td>
                            </tr>

                            {% else %}

                            <tr>
                                <td
                                    colspan="2"
                                    class="empty-row"
                                >
                                    本月暂无收入记录
                                </td>
                            </tr>

                            {% endfor %}

                            <tr class="total-row">
                                <td>
                                    日常户口总收入
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(daily_income_value) }}
                                </td>
                            </tr>

                        </tbody>

                    </table>

                </div>

                <div class="section-title section-gap">
                    📤 支出明细
                </div>

                <div class="table-responsive">

                    <table class="record-table">

                        <thead>
                            <tr>
                                <th>支出项目</th>
                                <th>金额 RM</th>
                            </tr>
                        </thead>

                        <tbody>

                            {% for r in daily_expense %}

                            <tr>
                                <td>
                                    {{ r.category }}
                                </td>

                                <td class="money money-expense">
                                    {{ "%.2f"|format(r.total) }}
                                </td>
                            </tr>

                            {% else %}

                            <tr>
                                <td
                                    colspan="2"
                                    class="empty-row"
                                >
                                    本月暂无支出记录
                                </td>
                            </tr>

                            {% endfor %}

                            <tr class="total-row">
                                <td>
                                    日常户口总支出
                                </td>

                                <td class="money money-expense">
                                    {{ "%.2f"|format(daily_expense_value) }}
                                </td>
                            </tr>

                            <tr class="balance-row">
                                <td>
                                    上月结余
                                </td>

                                <td class="money">
                                    {{ "%.2f"|format(opening_balance) }}
                                </td>
                            </tr>

                            <tr class="balance-row">
                                <td>
                                    本月结余
                                </td>

                                <td class="
                                    money
                                    {% if daily_balance >= 0 %}
                                        money-income
                                    {% else %}
                                        money-expense
                                    {% endif %}
                                ">
                                    {{ "%.2f"|format(daily_balance) }}
                                </td>
                            </tr>

                        </tbody>

                    </table>

                </div>

            </div>

            <div class="card">

                <div class="account-title">

                    <h2>
                        🏛️ 总会户口
                    </h2>

                    <span class="account-badge hq">
                        收入统计
                    </span>

                </div>

                <div class="section-title">
                    📥 收入明细
                </div>

                <div class="table-responsive">

                    <table class="record-table">

                        <thead>
                            <tr>
                                <th>收入项目</th>
                                <th>金额 RM</th>
                            </tr>
                        </thead>

                        <tbody>

                            {% for r in hq_income %}

                            <tr>
                                <td>
                                    {{ r.category }}
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(r.total) }}
                                </td>
                            </tr>

                            {% else %}

                            <tr>
                                <td
                                    colspan="2"
                                    class="empty-row"
                                >
                                    本月暂无总会户口收入
                                </td>
                            </tr>

                            {% endfor %}

                            <tr class="total-row">
                                <td>
                                    总会户口总收入
                                </td>

                                <td class="money money-income">
                                    {{ "%.2f"|format(hq_income_value) }}
                                </td>
                            </tr>

                        </tbody>

                    </table>

                </div>

            </div>

        </div>

        <div class="dashboard-actions">

            <a
                class="btn-tool btn-secondary"
                href="{{ url_for('finance.finance_home') }}"
            >
                ← 返回财政首页
            </a>

            <a
                class="btn-tool btn-success"
                href="{{ url_for(
                    'finance.export_monthly_report',
                    ym=ym
                ) }}"
            >
                📥 下载专业版 Excel 月报
            </a>

        </div>

    </div>

    </body>
    </html>
    """,
        ym=ym,
        opening_balance=opening_balance,
        daily_income=daily_income,
        daily_expense=daily_expense,
        hq_income=hq_income,
        daily_income_value=daily_income_value,
        daily_expense_value=daily_expense_value,
        hq_income_value=hq_income_value,
        daily_balance=daily_balance,
        monthly_net=monthly_net
    )


@finance_bp.route(
    "/expense/<category>",
    methods=["GET", "POST"]
)
@finance_bp.route(
    "/expense/<category>",
    methods=["GET", "POST"]
)
def expense(category):

    message = ""

    suggested_payment_voucher_no = (
        get_next_payment_voucher()
    )

    form_data = {
        "payment_voucher_no": suggested_payment_voucher_no,
        "record_date": date.today().isoformat(),
        "amount": "",
        "payment_method": "现金",
        "remarks": "",
    }

    if request.method == "POST":

        payment_voucher_no = request.form.get(
            "payment_voucher_no",
            ""
        ).strip().upper()

        record_date = (
            request.form.get("record_date")
            or date.today().isoformat()
        )

        amount_raw = request.form.get(
            "amount",
            ""
        ).strip()

        amount = money(amount_raw)

        payment_method = request.form.get(
            "payment_method",
            "现金"
        ).strip()

        remarks = request.form.get(
            "remarks",
            ""
        ).strip()

        form_data = {
            "payment_voucher_no": payment_voucher_no,
            "record_date": str(record_date),
            "amount": amount_raw,
            "payment_method": payment_method,
            "remarks": remarks,
        }

        existing_voucher = None

        if payment_voucher_no:

            existing_voucher = db_query("""
                select id
                from finance_records
                where payment_voucher_no = %s
                limit 1
            """, (
                payment_voucher_no,
            ), fetchone=True)

        if not payment_voucher_no:

            message = "请填写 Payment Voucher 编号。"

        elif not re.match(
            r"^PV\d+$",
            payment_voucher_no
        ):

            message = (
                "Payment Voucher 格式错误，"
                "例如：PV000001。"
            )

        elif existing_voucher:

            message = (
                "这个 Payment Voucher 编号已经存在，"
                "请检查是否重复。"
            )

        elif amount <= 0:

            message = "请输入正确的支出金额。"

        elif not remarks:

            message = "请填写支出用途或说明。"

        else:

            db_query("""
                insert into finance_records
                (
                    record_type,
                    payment_voucher_no,
                    record_date,
                    category,
                    amount,
                    payment_method,
                    fund_account,
                    remarks
                )
                values
                (
                    'expense',
                    %s, %s, %s, %s, %s, %s, %s
                )
            """, (
                payment_voucher_no,
                record_date,
                category,
                amount,
                payment_method,
                "观音堂日常户口",
                remarks
            ))

            return redirect(
                url_for("finance.records")
            )

    return render_template_string("""
    <!doctype html>
    <html lang="zh">
    <head>

        <meta charset="utf-8">

        <meta
            name="viewport"
            content="width=device-width, initial-scale=1"
        >

        <title>{{ category }}支出记录</title>

        <link
            rel="stylesheet"
            href="{{ url_for(
                'static',
                filename='css/toolbox.css'
            ) }}"
        >

        <style>

            .expense-page {
                max-width: 820px;
            }

            .expense-header {
                background:
                    linear-gradient(
                        135deg,
                        #dc2626,
                        #b91c1c
                    );

                color: white;
                padding: 28px;
                border-radius: 22px;
                margin-bottom: 20px;

                box-shadow:
                    0 12px 30px
                    rgba(185, 28, 28, 0.18);
            }

            .expense-header h1 {
                margin: 0 0 8px;
                font-size: 30px;
            }

            .expense-header p {
                margin: 0;
                opacity: 0.92;
                line-height: 1.6;
            }

            .account-note {
                display: flex;
                align-items: flex-start;
                gap: 12px;
                background: #fff7ed;
                border: 1px solid #fed7aa;
                color: #9a3412;
                border-radius: 14px;
                padding: 14px 16px;
                margin-bottom: 20px;
                line-height: 1.6;
            }

            .account-note-icon {
                font-size: 24px;
                line-height: 1;
            }

            .expense-form-grid {
                display: grid;
                grid-template-columns:
                    repeat(2, minmax(0, 1fr));
                gap: 16px;
            }

            .expense-form-grid .full-width {
                grid-column: 1 / -1;
            }

            .voucher-box {
                background: #eff6ff;
                border: 1px solid #bfdbfe;
                border-radius: 14px;
                padding: 16px;
                margin-bottom: 18px;
            }

            .voucher-input {
                font-weight: 800;
                letter-spacing: 1px;
                color: #1d4ed8;
            }

            .amount-input-wrap {
                position: relative;
            }

            .amount-prefix {
                position: absolute;
                left: 14px;
                top: 50%;
                transform: translateY(-50%);
                color: #64748b;
                font-weight: 700;
                pointer-events: none;
            }

            .amount-input-wrap .form-input {
                padding-left: 48px;
            }

            .expense-actions {
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 12px;
                flex-wrap: wrap;
                margin-top: 22px;
            }

            .expense-actions-right {
                display: flex;
                gap: 10px;
                flex-wrap: wrap;
            }

            .required-mark {
                color: #dc2626;
            }

            .form-help {
                color: #64748b;
                font-size: 14px;
                line-height: 1.5;
                margin-top: 6px;
            }

            @media (max-width: 700px) {

                .expense-page {
                    padding-left: 12px;
                    padding-right: 12px;
                }

                .expense-header {
                    padding: 22px 18px;
                    border-radius: 18px;
                }

                .expense-header h1 {
                    font-size: 26px;
                }

                .expense-form-grid {
                    grid-template-columns: 1fr;
                }

                .expense-form-grid .full-width {
                    grid-column: auto;
                }

                .expense-actions {
                    flex-direction: column-reverse;
                    align-items: stretch;
                }

                .expense-actions-right {
                    flex-direction: column;
                }

                .expense-actions .btn-tool {
                    width: 100%;
                }

            }

        </style>

    </head>

    <body>

    <div class="page expense-page">

        <div class="expense-header">

            <h1>💸 {{ category }}</h1>

            <p>
                填写本次支出资料，保存后将自动列入财政记录与月报。
            </p>

        </div>

        {% if message %}

            <div class="alert alert-danger">
                ⚠️ {{ message }}
            </div>

        {% endif %}

        <div class="account-note">

            <div class="account-note-icon">
                🏦
            </div>

            <div>
                <strong>支出户口：</strong>
                观音堂日常户口

                <br>

                所有支出记录统一归入观音堂日常户口。
            </div>

        </div>

        <div class="card">

            <div class="section-title">
                🧾 支出资料
            </div>

            <form method="post">

                <div class="voucher-box">

                    <div class="form-group">

                        <label class="form-label">
                            Payment Voucher No.
                            <span class="required-mark">*</span>
                        </label>

                        <input
                            class="form-input voucher-input"
                            name="payment_voucher_no"
                            value="{{ form_data.payment_voucher_no }}"
                            placeholder="例如 PV000001"
                            autocomplete="off"
                            required
                        >

                        <div class="form-help">
                            系统建议下一张：
                            <strong>
                                {{ suggested_payment_voucher_no }}
                            </strong>
                            。如实体 Payment Voucher 编号不同，
                            财政可以直接修改。
                        </div>

                    </div>

                </div>

                <div class="expense-form-grid">

                    <div class="form-group">

                        <label class="form-label">
                            支出日期
                            <span class="required-mark">*</span>
                        </label>

                        <input
                            class="form-input"
                            name="record_date"
                            type="date"
                            value="{{ form_data.record_date }}"
                            required
                        >

                    </div>

                    <div class="form-group">

                        <label class="form-label">
                            付款方式
                            <span class="required-mark">*</span>
                        </label>

                        <select
                            class="form-input"
                            name="payment_method"
                            required
                        >

                            {% for method in [
                                '现金',
                                '银行过账',
                                '支票'
                            ] %}

                                <option
                                    value="{{ method }}"
                                    {% if
                                        form_data.payment_method
                                        == method
                                    %}
                                        selected
                                    {% endif %}
                                >
                                    {{ method }}
                                </option>

                            {% endfor %}

                        </select>

                    </div>

                    <div class="form-group full-width">

                        <label class="form-label">
                            支出金额
                            <span class="required-mark">*</span>
                        </label>

                        <div class="amount-input-wrap">

                            <span class="amount-prefix">
                                RM
                            </span>

                            <input
                                class="form-input"
                                name="amount"
                                type="number"
                                step="0.01"
                                min="0.01"
                                value="{{ form_data.amount }}"
                                placeholder="0.00"
                                inputmode="decimal"
                                required
                            >

                        </div>

                    </div>

                    <div class="form-group full-width">

                        <label class="form-label">
                            支出用途／备注
                            <span class="required-mark">*</span>
                        </label>

                        <textarea
                            class="form-input"
                            name="remarks"
                            rows="5"
                            placeholder="例如：购买清洁用品、维修冷气、支付水电费"
                            required
                        >{{ form_data.remarks }}</textarea>

                        <div class="form-help">
                            请清楚填写支出用途，方便日后查询及核对月报。
                        </div>

                    </div>

                </div>

                <div class="expense-actions">

                    <a
                        class="btn-tool btn-secondary"
                        href="{{ url_for(
                            'finance.finance_expense_menu'
                        ) }}"
                    >
                        ← 返回支出项目
                    </a>

                    <div class="expense-actions-right">

                        <a
                            class="btn-tool btn-secondary"
                            href="{{ url_for(
                                'finance.records'
                            ) }}"
                        >
                            📚 查看财政记录
                        </a>

                        <button
                            class="btn-tool btn-danger"
                            type="submit"
                            onclick="
                                return confirm(
                                    '确定保存这笔支出记录？'
                                );
                            "
                        >
                            💾 保存支出
                        </button>

                    </div>

                </div>

            </form>

        </div>

    </div>

    </body>
    </html>
    """,
        category=category,
        message=message,
        form_data=form_data,
        suggested_payment_voucher_no=(
            suggested_payment_voucher_no
        )
    )

@finance_bp.route("/export_monthly_report")
def export_monthly_report():
    ym = request.args.get("ym", date.today().strftime("%Y-%m"))

    rows = db_query("""
        select
            record_date,
            receipt_date,
            receipt_no,
            category,
            record_type,
            fund_account,
            name,
            phone,
            amount,
            remarks as note
        from finance_records
        where to_char(record_date, 'YYYY-MM') = %s
        and coalesce(status,'confirmed') <> 'cancelled'
        order by record_date, id
    """, (ym,), fetchall=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "财政摘要"
    ws_income = wb.create_sheet("收入明细")
    ws_expense = wb.create_sheet("支出明细")

    # =========================
    # 样式
    # =========================
    green = "1F5E20"
    light_green = "EAF4E4"
    blue = "0B3A78"
    light_blue = "EAF2FF"
    gold = "F8E7B8"
    dark_text = "1F2937"
    red = "D00000"

    thin_green = Side(style="thin", color="8DBA7C")
    thin_blue = Side(style="thin", color="7FA6D9")
    thin_gold = Side(style="thin", color="C9962C")
    thin_gray = Side(style="thin", color="CCCCCC")

    border_green = Border(left=thin_green, right=thin_green, top=thin_green, bottom=thin_green)
    border_blue = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)
    border_gold = Border(left=thin_gold, right=thin_gold, top=thin_gold, bottom=thin_gold)
    border_gray = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    money_fmt = '"RM"#,##0.00'

    def money(v):
        return float(v or 0)

    def set_range_border(ws, cell_range, border):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border

    def fill_range(ws, cell_range, fill):
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = fill

    def center_range(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def format_money_cell(cell):
        cell.number_format = money_fmt
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 22)

    # =========================
    # 数据整理
    # =========================
    summary = {
        "观音堂日常户口": {"income": 0, "expense": 0},
        "总会户口": {"income": 0, "expense": 0},
    }

    income_rows = []
    expense_rows = []

    for r in rows:
        fund = r["fund_account"] or "未分类"
        record_type = r["record_type"] or "income"
        amount = money(r["amount"])

        if fund not in summary:
            summary[fund] = {"income": 0, "expense": 0}

        summary[fund][record_type] += amount

        if record_type == "income":
            income_rows.append(r)
        else:
            expense_rows.append(r)

    daily_income = summary.get("观音堂日常户口", {}).get("income", 0)
    daily_expense = summary.get("观音堂日常户口", {}).get("expense", 0)
    daily_balance = daily_income - daily_expense

    hq_income = summary.get("总会户口", {}).get("income", 0)
    hq_expense = summary.get("总会户口", {}).get("expense", 0)
    hq_balance = hq_income - hq_expense

    total_income = sum(v["income"] for v in summary.values())
    total_expense = sum(v["expense"] for v in summary.values())
    total_balance = total_income - total_expense

    # =========================
    # Sheet 1 财政摘要：图片同款
    # =========================
    ws.sheet_view.showGridLines = False

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 12

    ws.merge_cells("A1:O4")
    ws["A1"] = f"观音堂财政月报\n{ym}"
    ws["A1"].font = Font(bold=True, size=24, color=dark_text)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].fill = PatternFill("solid", fgColor="FFF8E8")
    set_range_border(ws, "A1:O4", border_gold)

    # 左表：观音堂日常户口
    ws.merge_cells("A5:G5")
    ws["A5"] = "观音堂日常户口"
    ws["A5"].fill = PatternFill("solid", fgColor=green)
    ws["A5"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A5"].alignment = Alignment(horizontal="center")

    daily_table = [
        ["项目", "收入（RM）", "支出（RM）", "本月结余（RM）"],
        ["收入", daily_income, "-", daily_income],
        ["支出", "-", daily_expense, -daily_expense],
        ["本月结余", "-", "-", daily_balance],
    ]

    start_row = 6
    for i, row in enumerate(daily_table, start_row):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        ws.merge_cells(start_row=i, start_column=5, end_row=i, end_column=5)
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)

        ws.cell(i, 1).value = row[0]
        ws.cell(i, 3).value = row[1]
        ws.cell(i, 5).value = row[2]
        ws.cell(i, 6).value = row[3]

        for c in [1, 3, 5, 6]:
            ws.cell(i, c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(i, c).border = border_green

        if i == 6:
            fill_range(ws, f"A{i}:G{i}", PatternFill("solid", fgColor=light_green))
            for c in [1, 3, 5, 6]:
                ws.cell(i, c).font = Font(bold=True)
        elif i == 9:
            fill_range(ws, f"A{i}:G{i}", PatternFill("solid", fgColor="DDEED2"))
            ws.cell(i, 6).font = Font(bold=True, size=14, color=green)

    for cell_ref in ["C7", "E8", "F7", "F8", "F9"]:
        ws[cell_ref].number_format = money_fmt

    # 右表：总会户口
    ws.merge_cells("I5:O5")
    ws["I5"] = "总会户口"
    ws["I5"].fill = PatternFill("solid", fgColor=blue)
    ws["I5"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["I5"].alignment = Alignment(horizontal="center")

    hq_table = [
        ["项目", "收入（RM）", "支出（RM）", "本月结余（RM）"],
        ["收入", hq_income, "-", hq_income],
        ["支出", "-", hq_expense, -hq_expense],
        ["本月结余", "-", "-", hq_balance],
    ]

    for i, row in enumerate(hq_table, start_row):
        ws.merge_cells(start_row=i, start_column=9, end_row=i, end_column=10)
        ws.merge_cells(start_row=i, start_column=11, end_row=i, end_column=12)
        ws.merge_cells(start_row=i, start_column=13, end_row=i, end_column=13)
        ws.merge_cells(start_row=i, start_column=14, end_row=i, end_column=15)

        ws.cell(i, 9).value = row[0]
        ws.cell(i, 11).value = row[1]
        ws.cell(i, 13).value = row[2]
        ws.cell(i, 14).value = row[3]

        for c in [9, 11, 13, 14]:
            ws.cell(i, c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(i, c).border = border_blue

        if i == 6:
            fill_range(ws, f"I{i}:O{i}", PatternFill("solid", fgColor=light_blue))
            for c in [9, 11, 13, 14]:
                ws.cell(i, c).font = Font(bold=True, color=blue)
        elif i == 9:
            fill_range(ws, f"I{i}:O{i}", PatternFill("solid", fgColor="DCEAFF"))
            ws.cell(i, 14).font = Font(bold=True, size=14, color=blue)

    for cell_ref in ["K7", "M8", "N7", "N8", "N9"]:
        ws[cell_ref].number_format = money_fmt

    # 总计横条
    ws.merge_cells("A11:C12")
    ws["A11"] = "总计"
    ws["A11"].font = Font(bold=True, size=16, color="7A5200")
    ws["A11"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A11"].fill = PatternFill("solid", fgColor="FFF2CC")

    ws.merge_cells("D11:F12")
    ws["D11"] = f"总收入（RM）\n{total_income}"
    ws["D11"].font = Font(bold=True, size=12, color=green)
    ws["D11"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D11"].fill = PatternFill("solid", fgColor="FFF8E8")

    ws.merge_cells("G11:I12")
    ws["G11"] = f"总支出（RM）\n{total_expense}"
    ws["G11"].font = Font(bold=True, size=12, color=red)
    ws["G11"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G11"].fill = PatternFill("solid", fgColor="FFF8E8")

    ws.merge_cells("J11:O12")
    ws["J11"] = f"总体结余（RM）\n{total_balance}"
    ws["J11"].font = Font(bold=True, size=12, color=blue)
    ws["J11"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["J11"].fill = PatternFill("solid", fgColor="FFF8E8")

    set_range_border(ws, "A11:O12", border_gold)

    # 收入明细表
    ws.merge_cells("A14:G14")
    ws["A14"] = "收入明细"
    ws["A14"].fill = PatternFill("solid", fgColor=green)
    ws["A14"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A14"].alignment = Alignment(horizontal="center")

    income_headers = [
        "付款日期",
        "开收条日期",
        "收条号码",
        "分类",
        "户口",
        "姓名",
        "金额（RM）"
    ]
    for col, h in enumerate(income_headers, 1):
        c = ws.cell(15, col)
        c.value = h
        c.fill = PatternFill("solid", fgColor=light_green)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = border_green

    rno = 16
    for r in income_rows[:8]:
        vals = [
            r["record_date"],
            r["receipt_date"],
            r["receipt_no"],
            r["category"],
            r["fund_account"],
            r["name"],
            money(r["amount"]),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(rno, col)
            c.value = v
            c.border = border_green
            c.alignment = Alignment(horizontal="center")
            if col == 7:
                c.number_format = money_fmt
        rno += 1

    while rno <= 23:
        for col in range(1, 8):
            ws.cell(rno, col).border = border_green
            ws.cell(rno, col).fill = PatternFill("solid", fgColor="FAFCF7")
        rno += 1

    ws.merge_cells("A24:F24")
    ws["A24"] = "收入总计"
    ws["A24"].fill = PatternFill("solid", fgColor="DDEED2")
    ws["A24"].font = Font(bold=True, color=green)
    ws["A24"].alignment = Alignment(horizontal="right")
    ws["G24"] = total_income
    ws["G24"].number_format = money_fmt
    ws["G24"].fill = PatternFill("solid", fgColor="DDEED2")
    ws["G24"].font = Font(bold=True, color=green)
    ws["G24"].alignment = Alignment(horizontal="center")
    set_range_border(ws, "A24:G24", border_green)

    # 支出明细表
    ws.merge_cells("I14:O14")
    ws["I14"] = "支出明细"
    ws["I14"].fill = PatternFill("solid", fgColor=blue)
    ws["I14"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["I14"].alignment = Alignment(horizontal="center")

    expense_headers = [
        "付款日期",
        "开收条日期",
        "收条号码",
        "分类",
        "户口",
        "姓名",
        "金额（RM）"
    ]
    for col, h in enumerate(expense_headers, 9):
        c = ws.cell(15, col)
        c.value = h
        c.fill = PatternFill("solid", fgColor=light_blue)
        c.font = Font(bold=True, color=blue)
        c.alignment = Alignment(horizontal="center")
        c.border = border_blue

    rno = 16
    for r in expense_rows[:8]:
        vals = [
            r["record_date"],
            r["receipt_date"],
            r["receipt_no"],
            r["category"],
            r["fund_account"],
            r["name"],
            money(r["amount"]),
        ]
        for col, v in enumerate(vals, 9):
            c = ws.cell(rno, col)
            c.value = v
            c.border = border_blue
            c.alignment = Alignment(horizontal="center")
            if col == 15:
                c.number_format = money_fmt
        rno += 1

    while rno <= 23:
        for col in range(9, 16):
            ws.cell(rno, col).border = border_blue
            ws.cell(rno, col).fill = PatternFill("solid", fgColor="F8FBFF")
        rno += 1

    ws.merge_cells("I24:N24")
    ws["I24"] = "支出总计"
    ws["I24"].fill = PatternFill("solid", fgColor="DCEAFF")
    ws["I24"].font = Font(bold=True, color=blue)
    ws["I24"].alignment = Alignment(horizontal="right")
    ws["O24"] = total_expense
    ws["O24"].number_format = money_fmt
    ws["O24"].fill = PatternFill("solid", fgColor="DCEAFF")
    ws["O24"].font = Font(bold=True, color=red)
    ws["O24"].alignment = Alignment(horizontal="center")
    set_range_border(ws, "I24:O24", border_blue)

    ws.freeze_panes = "A14"

    # =========================
    # Sheet 2 / Sheet 3 明细
    # =========================
    def write_detail_sheet(ws2, title, data_rows, main_color, light_color):
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells("A1:I2")
        ws2["A1"] = f"{title} - {ym}"
        ws2["A1"].font = Font(bold=True, size=18, color="FFFFFF")
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2["A1"].fill = PatternFill("solid", fgColor=main_color)

        headers = [
            "付款日期",
            "开收条日期",
            "收条号码",
            "分类",
            "户口",
            "姓名",
            "金额（RM）",
            "备注",
            "类型"
        ]

        for col, h in enumerate(headers, 1):
            c = ws2.cell(4, col)
            c.value = h
            c.fill = PatternFill("solid", fgColor=light_color)
            c.font = Font(bold=True, color=dark_text)
            c.alignment = Alignment(horizontal="center")
            c.border = border_gray

        row_no = 5
        for r in data_rows:
            vals = [
                r["record_date"],
                r["receipt_date"],
                r["receipt_no"],
                r["category"],
                r["fund_account"],
                r["name"],
                money(r["amount"]),
                r["note"],
                ...
            ]

            for col, v in enumerate(vals, 1):
                c = ws2.cell(row_no, col)
                c.value = v
                c.border = border_gray
                c.alignment = Alignment(horizontal="center")
                if col == 7:
                    c.number_format = money_fmt

            row_no += 1

        total_row = row_no + 1
        ws2.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
        ws2.cell(total_row, 1).value = "总计"
        ws2.cell(total_row, 1).font = Font(bold=True)
        ws2.cell(total_row, 1).alignment = Alignment(horizontal="right")
        ws2.cell(total_row, 7).value = sum(money(r["amount"]) for r in data_rows)
        ws2.cell(total_row, 7).number_format = money_fmt
        ws2.cell(total_row, 7).font = Font(bold=True, color=main_color)

        fill_range(ws2, f"A{total_row}:I{total_row}", PatternFill("solid", fgColor=light_color))
        set_range_border(ws2, f"A{total_row}:I{total_row}", border_gray)

        ws2.freeze_panes = "A5"
        auto_width(ws2)

    write_detail_sheet(ws_income, "收入明细", income_rows, green, light_green)
    write_detail_sheet(ws_expense, "支出明细", expense_rows, blue, light_blue)

    # =========================
    # 输出
    # =========================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"观音堂财政月报_{ym}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def get_next_payment_voucher():

    row = db_query("""
        select payment_voucher_no
        from finance_records
        where payment_voucher_no is not null
          and payment_voucher_no <> ''
        order by payment_voucher_no desc
        limit 1
    """, fetchone=True)

    if not row:
        return "PV000001"

    m = re.search(r"(\d+)$", row["payment_voucher_no"])

    if not m:
        return "PV000001"

    return f"PV{int(m.group(1))+1:06d}"

